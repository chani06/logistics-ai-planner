# ดึงข้อมูลจาก Google Sheets ด้วย gspread
import gspread
from google.oauth2.service_account import Credentials

# กำหนด scope สำหรับ Google Sheets API
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']

# URL หรือ ID ของ Google Sheet
SPREADSHEET_ID = '12DmIfECwVpsWfl8rl2r1A_LB4_5XMrmnmwlPUHKNU-o'
# ชื่อชีตหรือ gid (เช่น 'Sheet1' หรือใช้ worksheet_by_id)
WORKSHEET_GID = 876257177

# โหลด credentials (optional - graceful fallback)
gc = None
sh = None
SHEETS_AVAILABLE = False

try:
    creds = Credentials.from_service_account_file('credentials.json', scopes=SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SPREADSHEET_ID)
    SHEETS_AVAILABLE = True
except FileNotFoundError:
    # credentials.json not found - will show setup message later
    pass
except Exception as e:
    # Other errors - will show message
    pass


# Streamlit Web App สำหรับโหลดข้อมูล Google Sheets และบันทึก Excel
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import os

# Import OR-Tools Vehicle Routing Optimization (ถ้ามี)
try:
    from ortools_vrp import predict_trips_ortools
    ORTOOLS_AVAILABLE = True
except ImportError:
    ORTOOLS_AVAILABLE = False
    predict_trips_ortools = None

# Import Vehicle Logic (ข้อจำกัดรถ, Buffer, Drop Limits)
try:
    from vehicle_logic import (
        load_vehicle_restrictions_from_excel,
        get_buffer_for_trip,
        get_punthai_drop_limit,
        check_branch_vehicle_compatibility,
        get_max_vehicle_for_branch,
        get_max_vehicle_for_trip,
        filter_vehicles_by_region,
        suggest_truck,
        calculate_utilization,
        is_punthai_only,
        is_central_region,
        PUNTHAI_BUFFER,
        MAXMART_BUFFER
    )
    VEHICLE_LOGIC_AVAILABLE = True
    
    # โหลด Vehicle Restrictions จากไฟล์
    VEHICLE_RESTRICTIONS = load_vehicle_restrictions_from_excel()
    if VEHICLE_RESTRICTIONS:
        print(f"✅ โหลด Vehicle Restrictions: {len(VEHICLE_RESTRICTIONS)} สาขา")
except ImportError:
    VEHICLE_LOGIC_AVAILABLE = False
    VEHICLE_RESTRICTIONS = {}

st.title('🚚 ระบบจัดเที่ยว - Route Optimizer')

# Show Google Sheets connection status
if SHEETS_AVAILABLE:
    st.success("✅ เชื่อมต่อ Google Sheets แล้ว")
else:
    st.warning("⚠️ Google Sheets ยังไม่ได้ตั้งค่า - สามารถอัปโหลดไฟล์ Excel ได้เลย")
st.write('**โหลดข้อมูลสาขา** → **จัดเที่ยวแบบ Optimization** → **ดาวน์โหลดผลลัพธ์**')

import json
from datetime import datetime, time

def sync_branch_data_from_sheets():
    """
    ดึงข้อมูลจาก Google Sheets และ sync กับ JSON file
    ใช้รหัสสาขา (Code/Plan Code) เป็น key หลัก
    
    Returns:
        DataFrame หรือ None ถ้าล้มเหลว
    """
    global SHEETS_AVAILABLE, sh
    
    json_file = 'branch_data.json'
    
    # โหลดข้อมูลเก่าจาก JSON
    existing_data = {}
    if os.path.exists(json_file):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)
        except Exception as e:
            print(f"⚠️ ไม่สามารถอ่าน JSON: {e}")
    
    # ถ้าไม่มี Google Sheets ให้ใช้ข้อมูลเก่า
    if not SHEETS_AVAILABLE or sh is None:
        if existing_data:
            print(f"📦 ใช้ข้อมูลจาก JSON ({len(existing_data)} สาขา)")
            # แปลง dict กลับเป็น DataFrame
            df = pd.DataFrame.from_dict(existing_data, orient='index')
            df.reset_index(drop=True, inplace=True)
            return df
        else:
            print("ℹ️ ไม่มีข้อมูลใน JSON และไม่ได้เชื่อมต่อ Google Sheets")
            print("💡 สร้างข้อมูลตัวอย่างด้วย: python create_sample_data.py")
            return None
    
    # ดึงข้อมูลจาก Google Sheets
    try:
        print("🔄 กำลังดึงข้อมูลจาก Google Sheets...")
        worksheet = None
        for ws in sh.worksheets():
            if ws.id == WORKSHEET_GID:
                worksheet = ws
                break
        
        if worksheet is None:
            print(f"❌ ไม่พบ Worksheet GID {WORKSHEET_GID}")
            return None
        
        # ดึงข้อมูลทั้งหมด
        data = worksheet.get_all_values()
        
        if len(data) < 2:
            print("❌ ไม่มีข้อมูลใน Sheet")
            return None
        
        # แปลงเป็น DataFrame
        df_new = pd.DataFrame(data[1:], columns=data[0])
        
        # หา column ที่เป็นรหัสสาขา
        code_col = None
        for col in ['Code', 'Plan Code', 'รหัสสาขา', 'สาขา']:
            if col in df_new.columns:
                code_col = col
                break
        
        if code_col is None:
            print("❌ ไม่พบคอลัมน์รหัสสาขา")
            return None
        
        # นับข้อมูลใหม่
        new_count = 0
        updated_count = 0
        
        # เพิ่ม DC วังน้อย ในฐานข้อมูล
        dc_wangnoi = {
            '8nvDC011': {
                code_col: '8nvDC011',
                'สาขา': 'DC วังน้อย',
                'รายละเอียด': 'PROJECT-บ.พีทีจี เอ็นเนอยี จำกัด (มหาชน) (DCวังน้อย)',
                'ตำบล': 'พยอม',
                'อำเภอ': 'วังน้อย',
                'จังหวัด': 'พระนครศรีอยุธยา',
                'ละ': 14.1793943,
                'ลอง': 100.6481489,
                'MaxTruckType': '6W'
            }
        }
        
        # เพิ่ม DC วังน้อยถ้ายังไม่มี
        if '8nvDC011' not in existing_data:
            existing_data['8nvDC011'] = dc_wangnoi['8nvDC011']
            new_count += 1
        
        # อัปเดตข้อมูล
        for idx, row in df_new.iterrows():
            code = str(row[code_col]).strip().upper()
            if not code or code == '':
                continue
            
            # แปลง row เป็น dict
            row_dict = row.to_dict()
            
            if code in existing_data:
                # ข้อมูลเก่า - เช็คว่ามีการเปลี่ยนแปลงจริงหรือไม่
                if existing_data[code] != row_dict:
                    existing_data[code] = row_dict
                    updated_count += 1
                # ถ้าข้อมูลเหมือนเดิม ไม่นับเป็น update
            else:
                # ข้อมูลใหม่ - เพิ่ม
                existing_data[code] = row_dict
                new_count += 1
        
        # บันทึกกลับเป็น JSON
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=2)
        
        print(f"✅ Sync เสร็จสิ้น: {new_count} สาขาใหม่, {updated_count} สาขาอัปเดต, รวม {len(existing_data)} สาขา")
        
        # แปลงกลับเป็น DataFrame
        df = pd.DataFrame.from_dict(existing_data, orient='index')
        df.reset_index(drop=True, inplace=True)
        
        return df
        
    except Exception as e:
        print(f"❌ Error: {e}")
        # ถ้าเกิด error ให้ใช้ข้อมูลเก่า
        if existing_data:
            print(f"📦 ใช้ข้อมูลเก่าจาก JSON")
            df = pd.DataFrame.from_dict(existing_data, orient='index')
            return df
        return None


def load_sheet_data():
    """โหลดข้อมูลจาก Google Sheets (deprecated - ใช้ sync_branch_data_from_sheets แทน)"""
    global SHEETS_AVAILABLE
    
    if not SHEETS_AVAILABLE or gc is None or sh is None:
        st.error("❌ ไม่ได้เชื่อมต่อ Google Sheets - ไม่พบไฟล์ credentials.json")
        st.info("""
        📝 **วิธีตั้งค่า Google Sheets Integration:**
        
        1. ไปที่ [Google Cloud Console](https://console.cloud.google.com/)
        2. สร้าง Service Account สำหรับ Sheets API
        3. ดาวน์โหลด JSON key file
        4. บันทึกไฟล์เป็น `credentials.json` ในโฟลเดอร์ app
        5. Restart Streamlit app
        
        **หรือ:** อัปโหลดไฟล์ Excel โดยตรงแทน (ดูแท็บ "📦 จัดเที่ยว")
        """)
        return None
    
    try:
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
        SPREADSHEET_ID = '12DmIfECwVpsWfl8rl2r1A_LB4_5XMrmnmwlPUHKNU-o'
        WORKSHEET_GID = 876257177
        
        worksheet = None
        for ws in sh.worksheets():
            if ws.id == WORKSHEET_GID:
                worksheet = ws
                break
        
        if worksheet is None:
            st.error(f"❌ ไม่พบ Worksheet ที่มี GID {WORKSHEET_GID}")
            return None
        
        data = worksheet.get_all_values()
        return data
    except Exception as e:
        st.error(f"❌ เกิดข้อผิดพลาดในการเชื่อมต่อ Google Sheets: {e}")
        return None

# =============================
# ฟังก์ชันเลือกขนาดรถที่เหมาะสมตาม branch/zone
# =============================
def get_allowed_vehicle_for_branch(branch_code, zone, restrictions):
    allowed = restrictions.get(str(branch_code).strip(), ['4W', 'JB', '6W'])
    if zone == 'CENTRAL' and '6W' in allowed:
        allowed = [v for v in allowed if v != '6W']
    for v in ['6W', 'JB', '4W']:
        if v in allowed:
            return v
    return allowed[0]

# =============================
# ตัวอย่างการใช้งาน (comment ไว้)
# =============================
# restrictions = load_vehicle_restrictions('Dc/Auto planning (1).xlsx', 'Info')
# vehicle = get_allowed_vehicle_for_branch(branch_code, zone, restrictions)
"""
Logistics Planner 
"""

import streamlit as st
import pandas as pd
import numpy as np
import pickle
import os
import glob
from datetime import datetime, time
import io
from math import radians, sin, cos, sqrt, atan2

# Auto-refresh component
try:
    from streamlit_autorefresh import st_autorefresh
    AUTOREFRESH_AVAILABLE = True
except ImportError:
    AUTOREFRESH_AVAILABLE = False
    st.warning("⚠️ ติดตั้ง streamlit-autorefresh: pip install streamlit-autorefresh")

# ==========================================
# CONFIG
# ==========================================
MODEL_PATH = 'models/decision_tree_model.pkl'

# ขีดจำกัดรถแต่ละประเภท (มาตรฐาน)
LIMITS = {
    '4W': {'max_w': 2500, 'max_c': 5.0, 'max_drops': 12},   # ไม่เกิน 12 จุด, Cube ≤ 5
    'JB': {'max_w': 3500, 'max_c': 7.0, 'max_drops': 12},   # ไม่เกิน 12 จุด, Cube ≤ 7
    '6W': {'max_w': 6000, 'max_c': 20.0, 'max_drops': 999}  # ไม่จำกัดจุด, Cube ต้องเต็ม, Weight ≤ 6000
}

# 🔒 ขีดจำกัดสำหรับ Punthai ล้วน (ห้ามเกิน 100%)
PUNTHAI_LIMITS = {
    '4W': {'max_w': 2500, 'max_c': 5.0, 'max_drops': 5},   # Punthai ล้วน 4W: สูงสุด 5 สาขา
    'JB': {'max_w': 3500, 'max_c': 7.0, 'max_drops': 10},  # Punthai ล้วน JB: สูงสุด 10 สาขา
    '6W': {'max_w': 6000, 'max_c': 20.0, 'max_drops': 999}
}

# 🎯 Minimum utilization ต่อประเภทรถ (สำหรับ balancing)
MIN_UTIL = {
    '4W': 80,   # 4W ต้องใช้อย่างน้อย 70%
    'JB': 80,   # JB ต้องใช้อย่างน้อย 80%
    '6W': 90    # 6W ต้องใช้อย่างน้อย 90%
}

# Buffer สำหรับการใช้รถ (ตาม BU)
BUFFER = 1.0  # Default buffer
PUNTHAI_BUFFER = 1.0  # 🅿️ Punthai ล้วน: ห้ามเกิน 100%
MAXMART_BUFFER = 1.10  # 🅼 Maxmart/ผสม: เกินได้ 10%

# จำนวนสาขาต่อทริป - ใช้กับ 4W/JB เท่านั้น (6W ไม่จำกัด)
MAX_BRANCHES_PER_TRIP = 12  # สูงสุด 12 สาขาต่อทริปสำหรับ 4W/JB (6W ไม่จำกัด)

# Performance Config
MAX_DETOUR_KM = 12  # ลดจาก 15km เป็น 12km เพื่อประมวลผลเร็วขึ้น
MAX_MERGE_ITERATIONS = 25  # จำกัดรอบการรวมทริป (ลดจาก 50 เพื่อเร็วขึ้น)

# ==========================================
# REGION ORDER CONFIG (Far-to-Near Sorting)
# ==========================================
# ลำดับการจัด: เหนือ → อีสาน → ใต้ → ตะวันออก → กลาง
REGION_ORDER = {
    'เหนือ': 1, 'NORTH': 1,
    'อีสาน': 2, 'NE': 2,
    'ใต้': 3, 'SOUTH': 3,
    'ตะวันออก': 4, 'EAST': 4,
    'ตะวันตก': 5, 'WEST': 5,
    'กลาง': 6, 'CENTRAL': 6,
    'ไม่ระบุ': 99
}

# ภาคกลาง: ห้ามใช้ 6W (เฉพาะ 4W, JB)
CENTRAL_REGIONS = ['กลาง', 'CENTRAL']
CENTRAL_ALLOWED_VEHICLES = ['4W', 'JB']  # NO 6W in Central

# รายการสาขาที่ไม่ต้องการจัดส่ง (ตัดออก)
EXCLUDE_BRANCHES = ['DC011', 'PTDC', 'PTG DISTRIBUTION CENTER']

# รายชื่อที่ต้องตัดออก (ใช้ตรวจสอบชื่อ)
EXCLUDE_NAMES = ['Distribution Center', 'PTG Distribution', 'บ.พีทีจี เอ็นเนอยี']

# พิกัด DC วังน้อย (จุดกลาง)
DC_WANG_NOI_LAT = 14.179394
DC_WANG_NOI_LON = 100.648149

# ระยะทางที่ต้องใช้รถ 6W (กม.)
DISTANCE_REQUIRE_6W = 100  # ถ้าห่างจาก DC เกิน 100 กม. ต้องใช้ 6W

# ==========================================
# ZONE/REGION CONFIG - รหัสภาคและจังหวัด
# ==========================================
# รหัสภาค: 1=กลาง, 2=ตะวันออก, 3=ตะวันตก, 4=เหนือ, 5=อีสาน, 6=ใต้
REGION_CODE = {
    # ภาคกลาง (รหัส 1)
    'กรุงเทพมหานคร': '10', 'กรุงเทพฯ': '10',
    'นนทบุรี': '11',
    'ปทุมธานี': '12',
    'พระนครศรีอยุธยา': '13', 'อยุธยา': '13',
    'สระบุรี': '14',
    'ลพบุรี': '15',
    'สิงห์บุรี': '16',
    'อ่างทอง': '17',
    'ชัยนาท': '18',
    'นครปฐม': '19',
    'สมุทรปราการ': '1A',
    'สมุทรสาคร': '1B',
    'สมุทรสงคราม': '1C',
    
    # ภาคตะวันออก (รหัส 2)
    'ชลบุรี': '20',
    'ระยอง': '21',
    'จันทบุรี': '22',
    'ตราด': '23',
    'ฉะเชิงเทรา': '24',
    'ปราจีนบุรี': '25',
    'สระแก้ว': '26',
    'นครนายก': '27',
    
    # ภาคตะวันตก (รหัส 3)
    'ราชบุรี': '30',
    'กาญจนบุรี': '31',
    'สุพรรณบุรี': '32',
    'เพชรบุรี': '33',
    'ประจวบคีรีขันธ์': '34',
    
    # ภาคเหนือ (รหัส 4) - ไกล ใช้ 6W เป็นหลัก
    'นครสวรรค์': '40',
    'อุทัยธานี': '41',
    'กำแพงเพชร': '42',
    'ตาก': '43',
    'สุโขทัย': '44',
    'พิษณุโลก': '45',
    'พิจิตร': '46',
    'เพชรบูรณ์': '47',
    'อุตรดิตถ์': '48',
    'แพร่': '49',
    'น่าน': '4A',
    'พะเยา': '4B',
    'เชียงราย': '4C',
    'เชียงใหม่': '4D',
    'แม่ฮ่องสอน': '4E',
    'ลำพูน': '4F',
    'ลำปาง': '4G',
    
    # ภาคตะวันออกเฉียงเหนือ/อีสาน (รหัส 5)
    'นครราชสีมา': '50', 'โคราช': '50',
    'บุรีรัมย์': '51',
    'สุรินทร์': '52',
    'ศรีสะเกษ': '53',
    'อุบลราชธานี': '54',
    'ยโสธร': '55',
    'ชัยภูมิ': '56',
    'อำนาจเจริญ': '57',
    'หนองบัวลำภู': '58',
    'ขอนแก่น': '59',
    'อุดรธานี': '5A',
    'เลย': '5B',
    'หนองคาย': '5C',
    'มหาสารคาม': '5D',
    'ร้อยเอ็ด': '5E',
    'กาฬสินธุ์': '5F',
    'สกลนคร': '5G',
    'นครพนม': '5H',
    'มุกดาหาร': '5I',
    'บึงกาฬ': '5J',
    
    # ภาคใต้ (รหัส 6) - ไกลมาก ใช้ 6W
    'ชุมพร': '60',
    'ระนอง': '61',
    'สุราษฎร์ธานี': '62',
    'พังงา': '63',
    'กระบี่': '64',
    'ภูเก็ต': '65',
    'นครศรีธรรมราช': '66',
    'ตรัง': '67',
    'พัทลุง': '68',
    'สงขลา': '69',
    'สตูล': '6A',
    'ปัตตานี': '6B',
    'ยะลา': '6C',
    'นราธิวาส': '6D',
}

# ภาคที่ต้องใช้ 6W เป็นหลัก (ไกลจาก DC)
REGIONS_REQUIRE_6W = ['4', '5', '6']  # เหนือ, อีสาน, ใต้

# ชื่อภาค
REGION_NAMES = {
    '1': 'กลาง',
    '2': 'ตะวันออก',
    '3': 'ตะวันตก',
    '4': 'เหนือ',
    '5': 'อีสาน',
    '6': 'ใต้',
    '9': 'ไม่ระบุ'
}

# ==========================================
# HELPER: ZONE/REGION FUNCTIONS
# ==========================================
def get_region_code(province):
    """ดึงรหัสภาค/โซนจากจังหวัด"""
    if not province or str(province).strip() == '' or str(province) == 'nan':
        return '99'  # ไม่ระบุ
    province = str(province).strip()
    return REGION_CODE.get(province, '99')

def get_region_name(province):
    """ดึงชื่อภาคจากจังหวัด"""
    code = get_region_code(province)
    if code == '99':
        return 'ไม่ระบุ'
    region_prefix = code[0]
    return REGION_NAMES.get(region_prefix, 'ไม่ระบุ')

def get_recommended_vehicle_by_region(province, distance_from_dc=None):
    """แนะนำรถตามภาค/ระยะทาง"""
    code = get_region_code(province)
    region_prefix = code[0] if code != '99' else '9'
    
    # ภาคเหนือ, อีสาน, ใต้ → ใช้ 6W
    if region_prefix in REGIONS_REQUIRE_6W:
        return '6W'
    
    # ถ้ามีระยะทาง และเกิน threshold → ใช้ 6W
    if distance_from_dc and distance_from_dc > DISTANCE_REQUIRE_6W:
        return '6W'
    
    # ภาคกลาง, ตะวันออก, ตะวันตก → ใช้ 4W/JB ได้
    return 'JB'  # default เป็น JB

def sort_branches_by_region_route(branches_df, master_data=None):
    """
    จัดเรียงสาขาตามภาค → จังหวัด → อำเภอ → ตำบล → Route
    เพื่อให้ทริปเรียงติดกันไม่กระโดด
    """
    if branches_df.empty:
        return branches_df
    
    df = branches_df.copy()
    
    # เพิ่มคอลัมน์สำหรับ sort
    df['_region_code'] = df['Province'].apply(get_region_code) if 'Province' in df.columns else '99'
    df['_province'] = df['Province'].fillna('') if 'Province' in df.columns else ''
    df['_district'] = df['District'].fillna('') if 'District' in df.columns else ''
    df['_subdistrict'] = df['Subdistrict'].fillna('') if 'Subdistrict' in df.columns else ''
    
    # แยก Route number
    if 'Route' in df.columns:
        df['_route_num'] = df['Route'].apply(lambda x: int(str(x).replace('CD', '')) if pd.notna(x) and str(x).startswith('CD') else 99999)
    else:
        df['_route_num'] = 99999
    
    # Sort
    df = df.sort_values(by=['_region_code', '_province', '_district', '_subdistrict', '_route_num'])
    
    # ลบคอลัมน์ชั่วคราว
    df = df.drop(columns=['_region_code', '_province', '_district', '_subdistrict', '_route_num'])
    
    return df.reset_index(drop=True)

def check_trip_route_spread(trip_df):
    """
    ตรวจสอบว่าทริปมี Route กระจายมากไหม
    คืนค่า: (route_range, is_spread, provinces)
    """
    if trip_df.empty or 'Route' not in trip_df.columns:
        return 0, False, []
    
    routes = trip_df['Route'].dropna().unique()
    route_nums = []
    for r in routes:
        if pd.notna(r) and str(r).startswith('CD'):
            try:
                route_nums.append(int(str(r).replace('CD', '')))
            except:
                pass
    
    if len(route_nums) < 2:
        return 0, False, trip_df['Province'].dropna().unique().tolist() if 'Province' in trip_df.columns else []
    
    route_range = max(route_nums) - min(route_nums)
    is_spread = route_range > 4000  # ถ้ามากกว่า 4000 ถือว่ากระจาย
    
    provinces = trip_df['Province'].dropna().unique().tolist() if 'Province' in trip_df.columns else []
    
    return route_range, is_spread, provinces

def validate_trip_vehicle(trip_df, assigned_vehicle):
    """
    ตรวจสอบว่ารถที่จัดให้ทริปเหมาะสมกับภาค/ระยะทางหรือไม่
    คืนค่า: (is_valid, recommended_vehicle, reason)
    """
    if trip_df.empty:
        return True, assigned_vehicle, ''
    
    provinces = trip_df['Province'].dropna().unique() if 'Province' in trip_df.columns else []
    
    # หาภาคที่ไกลที่สุดในทริป
    farthest_region = '1'  # default กลาง
    for prov in provinces:
        code = get_region_code(prov)
        region = code[0] if code != '99' else '1'
        if region > farthest_region:
            farthest_region = region
    
    # ตรวจสอบ
    if farthest_region in REGIONS_REQUIRE_6W:
        # ภาคไกล ควรใช้ 6W
        if assigned_vehicle in ['4W', 'JB']:
            return False, '6W', f'ภาค{REGION_NAMES.get(farthest_region, "ไกล")} ควรใช้ 6W'
    
    return True, assigned_vehicle, ''

# ==========================================
# LOAD MASTER DATA
# ==========================================
@st.cache_data(ttl=7200)  # Cache 2 ชั่วโมง (เร็วขึ้น)
def load_master_data():
    """
    โหลดข้อมูล Master สถานที่ส่ง
    ลำดับความสำคัญ: 1. JSON (sync จาก Sheets), 2. Excel files
    """
    # 1. ลองโหลดจาก JSON ที่ sync จาก Google Sheets
    json_file = 'branch_data.json'
    if os.path.exists(json_file):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                branch_data = json.load(f)
            
            if branch_data:
                print(f"📦 โหลด Master จาก JSON: {len(branch_data)} สาขา")
                df_master = pd.DataFrame.from_dict(branch_data, orient='index')
                df_master.reset_index(drop=True, inplace=True)
                
                # ตรวจสอบและจัด column ให้ตรงกับที่ต้องการ
                required_cols = ['Plan Code', 'ตำบล', 'อำเภอ', 'จังหวัด', 'ละติจูด', 'ลองติจูด']
                
                # Mapping column names
                column_mapping = {}
                for col in df_master.columns:
                    col_clean = str(col).strip()
                    if 'plan code' in col_clean.lower() or 'รหัส' in col_clean or col_clean.upper() == 'CODE':
                        column_mapping[col] = 'Plan Code'
                    elif 'ตำบล' in col_clean or 'subdistrict' in col_clean.lower():
                        column_mapping[col] = 'ตำบล'
                    elif 'อำเภอ' in col_clean or 'district' in col_clean.lower():
                        column_mapping[col] = 'อำเภอ'
                    elif 'จังหวัด' in col_clean or 'province' in col_clean.lower():
                        column_mapping[col] = 'จังหวัด'
                    elif 'ละติจูด' in col_clean or 'latitude' in col_clean.lower() or col_clean.upper() == 'LAT':
                        column_mapping[col] = 'ละติจูด'
                    elif 'ลองติจูด' in col_clean or 'ลองจิจูด' in col_clean or 'longitude' in col_clean.lower() or col_clean.upper() in ['LON', 'LONG', 'LNG']:
                        column_mapping[col] = 'ลองติจูด'
                
                if column_mapping:
                    df_master = df_master.rename(columns=column_mapping)
                
                # สร้าง Plan Code ถ้าไม่มี (ใช้ Code แทน)
                if 'Plan Code' not in df_master.columns:
                    if 'Code' in df_master.columns:
                        df_master['Plan Code'] = df_master['Code']
                    elif 'รหัสสาขา' in df_master.columns:
                        df_master['Plan Code'] = df_master['รหัสสาขา']
                
                # ทำความสะอาด Plan Code
                if 'Plan Code' in df_master.columns:
                    df_master['Plan Code'] = df_master['Plan Code'].astype(str).str.strip().str.upper()
                    df_master = df_master[df_master['Plan Code'] != '']
                
                # เลือกเฉพาะคอลัมน์ที่มี
                available_cols = [col for col in required_cols if col in df_master.columns]
                if available_cols:
                    return df_master[available_cols]
                
                return df_master
                
        except Exception as e:
            print(f"⚠️ ไม่สามารถโหลด JSON: {e}")
    
    # 2. ถ้าไม่มี JSON ให้ใช้ไฟล์ Excel (Fallback)
    try:
        usecols = ['Plan Code', 'ตำบล', 'อำเภอ', 'จังหวัด', 'ละติจูด', 'ลองติจูด']
        possible_files = ['Dc/สถานที่ส่ง.xlsx', 'Dc/Master สถานที่ส่ง.xlsx']
        df_master = pd.DataFrame()
        for file_path in possible_files:
            try:
                df_master = pd.read_excel(file_path, usecols=usecols)
                print(f"📦 โหลด Master จาก Excel: {file_path}")
                break
            except:
                continue
        if df_master.empty:
            return pd.DataFrame()
        
        # ทำความสะอาด Plan Code (vectorized)
        if 'Plan Code' in df_master.columns:
            df_master['Plan Code'] = df_master['Plan Code'].astype(str).str.strip().str.upper()
        
        df_master = df_master[df_master['Plan Code'] != '']
        return df_master
    except FileNotFoundError:
        return pd.DataFrame()
    except Exception as e:
        try:
            st.warning(f"ไม่สามารถโหลดไฟล์ Master: {e} (จะใช้ข้อมูลจากไฟล์อัปโหลดแทน)")
        except:
            pass
        return pd.DataFrame()

# โหลด Master Data
MASTER_DATA = load_master_data()

# ==========================================
# CLEAN NAME FUNCTION (สำหรับทำ Join_Key)
# ==========================================
def clean_name(text):
    """
    ทำความสะอาดชื่อ: ลบ prefix จ./อ./ต. และ trim whitespace
    ใช้สำหรับสร้าง Join_Key เพื่อเทียบกับ Master Data
    """
    if pd.isna(text) or text is None:
        return ''
    text = str(text)
    # ลบ prefix ภาษาไทย
    text = text.replace('จ. ', '').replace('จ.', '')
    text = text.replace('อ. ', '').replace('อ.', '')
    text = text.replace('ต. ', '').replace('ต.', '')
    # ลบ prefix ภาษาอังกฤษ (ถ้ามี)
    text = text.replace('Tambon ', '').replace('Amphoe ', '').replace('Changwat ', '')
    return text.strip()

def normalize_province_name(province):
    """
    แปลงชื่อจังหวัดให้เป็นมาตรฐาน (แก้ปัญหาชื่อเพี้ยน)
    """
    if pd.isna(province) or province is None:
        return ''
    province = clean_name(province)
    # Mapping ชื่อที่พบบ่อย
    province_mapping = {
        'พระนครศรีอยุธยา': 'อยุธยา',
        'กรุงเทพฯ': 'กรุงเทพมหานคร',
        'กทม': 'กรุงเทพมหานคร',
        'กทม.': 'กรุงเทพมหานคร',
        'โคราช': 'นครราชสีมา',
    }
    return province_mapping.get(province, province)

def load_master_dist_data():
    """
    โหลดไฟล์ Master Dist.xlsx สำหรับ:
    1. ระยะทางระดับตำบล
    2. Sum_Code (Sort_Code) สำหรับเรียงลำดับตามภูมิศาสตร์
    
    หลักการ: ใช้ Join_Key (จังหวัด_อำเภอ_ตำบล) เป็นตัวเชื่อม
    เพื่อดึง Sum_Code มาใช้ในการ Sort
    """
    try:
        file_path = 'Dc/Master Dist.xlsx'
        df = pd.read_excel(file_path)
        
        # สร้าง lookup dict - สอง key: Sum_Code และ Join_Key (จังหวัด_อำเภอ_ตำบล)
        dist_lookup = {}   # key = Sum_Code
        name_lookup = {}   # key = Join_Key (จังหวัด_อำเภอ_ตำบล)
        
        for _, row in df.iterrows():
            sum_code = str(row.get('Sum_Code', '')).strip()
            
            # ข้อมูลสำคัญ: เพิ่ม sum_code (Sort_Code) เข้าไปด้วย!
            data = {
                'sum_code': sum_code,  # 🔑 กุญแจสำคัญสำหรับ Sort!
                'region': row.get('Region', ''),
                'region_code': row.get('Region_Code', ''),
                'province': row.get('Province', ''),
                'prov_code': row.get('Prov_Code', ''),
                'district': row.get('District', ''),
                'dist_code': row.get('Dist_Code', ''),
                'subdistrict': row.get('Subdistrict', ''),
                'subdist_code': row.get('Subdist_Code', ''),
                'dist_from_dc_km': float(row.get('Dist_from_DC_km', 9999)) if pd.notna(row.get('Dist_from_DC_km')) else 9999,
                'prov_dist_km': float(row.get('Prov_Dist_km', 0)) if pd.notna(row.get('Prov_Dist_km')) else 0,
                'dist_subdist_km': float(row.get('Dist_Subdist_km', 0)) if pd.notna(row.get('Dist_Subdist_km')) else 0,
            }
            
            # Key 1: Sum_Code (สำหรับ lookup โดยตรง)
            if sum_code:
                dist_lookup[sum_code] = data
            
            # Key 2: Join_Key (จังหวัด_อำเภอ_ตำบล) - หัวใจของ Lookup!
            prov_raw = str(row.get('Province', ''))
            dist_raw = str(row.get('District', ''))
            subdist_raw = str(row.get('Subdistrict', ''))
            
            # Clean name สำหรับ Join
            prov_clean = clean_name(prov_raw)
            dist_clean = clean_name(dist_raw)
            subdist_clean = clean_name(subdist_raw)
            
            # Join_Key แบบ clean (มาตรฐาน)
            join_key = f"{prov_clean}_{dist_clean}_{subdist_clean}"
            if join_key and join_key != '__':
                name_lookup[join_key] = data
            
            # Join_Key แบบ normalized province (เผื่อชื่อเพี้ยน)
            prov_normalized = normalize_province_name(prov_raw)
            if prov_normalized != prov_clean:
                alt_key = f"{prov_normalized}_{dist_clean}_{subdist_clean}"
                if alt_key and alt_key != '__':
                    name_lookup[alt_key] = data
            
            # Join_Key แบบมี prefix (เผื่อข้อมูลมี prefix)
            raw_key = f"{prov_raw.strip()}_{dist_raw.strip()}_{subdist_raw.strip()}"
            if raw_key and raw_key != '__' and raw_key not in name_lookup:
                name_lookup[raw_key] = data
        
        return {'by_code': dist_lookup, 'by_name': name_lookup}
    except Exception as e:
        return {'by_code': {}, 'by_name': {}}

# โหลด Master Dist Data
MASTER_DIST_DATA = load_master_dist_data()

# ==========================================
# PUNTHAI/MAXMART BUFFER FUNCTIONS
# ==========================================
def is_punthai_only(trip_data):
    """
    ตรวจสอบว่าทริปนี้เป็น Punthai ล้วน, Maxmart ล้วน หรือผสม
    
    Returns:
        'punthai_only': ถ้าทั้งหมดเป็น Punthai (BU = 211 หรือชื่อมี PUNTHAI)
        'maxmart_only': ถ้าทั้งหมดเป็น Maxmart (BU = 200 หรือชื่อมี MAXMART)
        'mixed': ถ้ามีทั้ง Punthai และ Maxmart
        'other': ถ้าไม่มีข้อมูล BU
    """
    if trip_data is None or len(trip_data) == 0:
        return 'other'
    
    punthai_count = 0
    maxmart_count = 0
    total_count = len(trip_data)
    
    for _, row in trip_data.iterrows():
        bu = row.get('BU', None)
        name = str(row.get('Name', '')).upper()
        
        # เช็ค Punthai: BU = 211 หรือชื่อมี PUNTHAI
        if bu == 211 or bu == '211' or 'PUNTHAI' in name or 'PUN-' in name:
            punthai_count += 1
        # เช็ค Maxmart: BU = 200 หรือชื่อมี MAXMART/MAX MART
        elif bu == 200 or bu == '200' or 'MAXMART' in name or 'MAX MART' in name:
            maxmart_count += 1
    
    if punthai_count == total_count:
        return 'punthai_only'
    elif maxmart_count == total_count:
        return 'maxmart_only'
    elif punthai_count > 0 or maxmart_count > 0:
        return 'mixed'
    else:
        return 'other'

def get_buffer_for_trip(trip_data):
    """
    ดึง Buffer ที่เหมาะสมตาม BU ของทริป
    
    Rules:
    - Punthai ล้วน: BUFFER = 1.0 (ห้ามเกิน 100%)
    - Maxmart ล้วน/ผสม: BUFFER = 1.10 (เกินได้ 10%)
    
    Returns:
        float: buffer multiplier (1.0 หรือ 1.10)
    """
    trip_type = is_punthai_only(trip_data)
    
    if trip_type == 'punthai_only':
        return PUNTHAI_BUFFER  
    elif trip_type in ['maxmart_only', 'mixed']:
        return MAXMART_BUFFER  
    else:
        return BUFFER  

def get_punthai_drop_limit(trip_data, vehicle_type):
    """
    ดึงจำกัดจำนวน Drop สำหรับ Punthai ล้วน
    
    Rules:
    - Punthai ล้วน + 4W: สูงสุด 5 สาขา
    - Punthai ล้วน + JB: สูงสุด 7 drop
    - อื่นๆ: ไม่จำกัด (999)
    
    Returns:
        int: max drops allowed
    """
    trip_type = is_punthai_only(trip_data)
    
    if trip_type == 'punthai_only':
        return PUNTHAI_LIMITS.get(vehicle_type, {}).get('max_drops', 999)
    else:
        return 999  # ไม่จำกัด

@st.cache_data(ttl=3600)  # Cache 1 ชั่วโมง
def load_booking_history_restrictions():
    """โหลดประวัติการจัดส่งจาก Booking History - ข้อมูลจริง 3,053 booking (Optimized)"""
    try:
        # ลองหาไฟล์ Booking History (อาจมีชื่อหลายแบบ)
        possible_files = [
            'Dc/ประวัติงานจัดส่ง DC วังน้อย(1).xlsx',
            'Dc/ประวัติงานจัดส่ง DC วังน้อย.xlsx',
            'branch_vehicle_restrictions_from_booking.xlsx'
        ]
        
        file_path = None
        for path in possible_files:
            if os.path.exists(path):
                file_path = path
                break
        
        if not file_path:
            # ถ้าไม่มีไฟล์ ใช้ข้อมูลที่เคยเรียนรู้ (fallback)
            return load_learned_restrictions_fallback()
        
        df = pd.read_excel(file_path)
        
        # แปลงประเภทรถ
        vehicle_mapping = {
            '4 ล้อ จัมโบ้ ตู้ทึบ': 'JB',
            '6 ล้อ ตู้ทึบ': '6W',
            '4 ล้อ ตู้ทึบ': '4W'
        }
        df['Vehicle_Type'] = df['ประเภทรถ'].map(vehicle_mapping)
        
        # วิเคราะห์ความสัมพันธ์สาขา-รถ (ไม่ใช้ Booking No)
        branch_vehicle_history = {}
        
        # จัดกลุ่มตามสาขา
        for branch_code in df['รหัสสาขา'].dropna().unique():
            branch_data = df[df['รหัสสาขา'] == branch_code]
            vehicle_types = branch_data['Vehicle_Type'].dropna().unique()
            if len(vehicle_types) > 0:
                # เก็บประวัติรถที่ใช้กับสาขานี้
                branch_vehicle_history[branch_code] = list(branch_data['Vehicle_Type'].dropna())
        
        # สร้าง restrictions
        branch_restrictions = {}
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for branch_code, vehicle_list in branch_vehicle_history.items():
            vehicles_used = set(vehicle_list)
            vehicle_counts = pd.Series(vehicle_list).value_counts().to_dict()
            
            if len(vehicles_used) == 1:
                # STRICT - ใช้รถเดียว
                vehicle = list(vehicles_used)[0]
                branch_restrictions[str(branch_code)] = {
                    'max_vehicle': vehicle,
                    'allowed': [vehicle],
                    'total_bookings': len(vehicle_list),
                    'restriction_type': 'STRICT'
                }
            else:
                # FLEXIBLE - ใช้ได้หลายประเภท
                max_vehicle = max(vehicles_used, key=lambda v: vehicle_sizes.get(v, 0))
                branch_restrictions[str(branch_code)] = {
                    'max_vehicle': max_vehicle,
                    'allowed': list(vehicles_used),
                    'total_bookings': len(vehicle_list),
                    'restriction_type': 'FLEXIBLE'
                }
        
        stats = {
            'total_branches': len(branch_restrictions),
            'strict': len([b for b, r in branch_restrictions.items() if r['restriction_type'] == 'STRICT']),
            'flexible': len([b for b, r in branch_restrictions.items() if r['restriction_type'] == 'FLEXIBLE']),
            'total_bookings': len(df)
        }
        
        return {
            'branch_restrictions': branch_restrictions,
            'stats': stats
        }
    except Exception as e:
        # ถ้าเกิด error ใช้ข้อมูลที่เคยเรียนรู้แทน
        return load_learned_restrictions_fallback()

def load_learned_restrictions_fallback():
    """
    ข้อมูลที่เรียนรู้จาก Booking History (backup)
    ใช้เมื่อไม่สามารถโหลดไฟล์ได้
    
    จากการวิเคราะห์ 3,053 bookings, 2,790 สาขา:
    - JB: รถกลาง (ใช้มากที่สุด 54.7%)
    - 6W: รถใหญ่ (30.1%)
    - 4W: รถเล็ก (0.2%)
    
    กลยุทธ์: ถ้าไม่มีข้อมูล default เป็น JB (รถกลาง ใช้ได้กับสาขาส่วนใหญ่)
    """
    return {
        'branch_restrictions': {},
        'stats': {
            'total_branches': 0,
            'strict': 0,
            'flexible': 0,
            'total_bookings': 0,
            'fallback': True,
            'message': 'ใช้ Punthai เป็นหลัก (ไม่พบไฟล์ Booking History)'
        }
    }

@st.cache_data(ttl=3600)  # Cache 1 ชั่วโมง
def load_punthai_reference():
    """โหลดไฟล์ Punthai Maxmart เพื่อเรียนรู้หลักการจัดทริป (Location patterns - Optimized)"""
    try:
        file_path = 'Dc/แผนงาน Punthai Maxmart รอบสั่ง 24หยิบ 25พฤศจิกายน 2568 To.เฟิ(1) - สำเนา.xlsx'
        df = pd.read_excel(file_path, sheet_name='2.Punthai', header=1)
        
        # กรองเฉพาะแถวที่มี Trip และไม่ใช่ DC/Distribution Center
        df_clean = df[df['Trip'].notna()].copy()
        df_clean = df_clean[~df_clean['BranchCode'].isin(['DC011', 'PTDC', 'PTG Distribution Center'])].copy()
        
        # Extract vehicle type from Trip no (เช่น 4W009 → 4W)
        df_clean['Vehicle_Type'] = df_clean['Trip no'].apply(
            lambda x: str(x)[:2] if pd.notna(x) else 'Unknown'
        )
        
        # Merge กับ Master เพื่อได้ข้อมูลตำบล/อำเภอ/จังหวัด
        try:
            df_master = pd.read_excel('Dc/Master สถานที่ส่ง.xlsx')
            df_clean = df_clean.merge(
                df_master[['Plan Code', 'ตำบล', 'อำเภอ', 'จังหวัด']],
                left_on='BranchCode',
                right_on='Plan Code',
                how='left'
            )
        except:
            pass
        
        # เรียนรู้ข้อจำกัดรถจาก Punthai (แผน) - สำหรับสาขาที่ไม่มีใน Booking
        punthai_restrictions = {}
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for branch_code in df_clean['BranchCode'].unique():
            branch_data = df_clean[df_clean['BranchCode'] == branch_code]
            vehicles_used = set(branch_data['Vehicle_Type'].dropna().tolist())
            vehicles_used = {v for v in vehicles_used if v in ['4W', 'JB', '6W']}
            
            if vehicles_used:
                if len(vehicles_used) == 1:
                    vehicle = list(vehicles_used)[0]
                    punthai_restrictions[str(branch_code)] = {
                        'max_vehicle': vehicle,
                        'allowed': [vehicle],
                        'source': 'PUNTHAI'
                    }
                else:
                    max_vehicle = max(vehicles_used, key=lambda v: vehicle_sizes.get(v, 0))
                    punthai_restrictions[str(branch_code)] = {
                        'max_vehicle': max_vehicle,
                        'allowed': list(vehicles_used),
                        'source': 'PUNTHAI'
                    }
        
        # สร้าง dictionary: Trip → ข้อมูล (location patterns)
        trip_patterns = {}
        location_stats = {
            'same_province': 0,
            'mixed_province': 0,
            'avg_branches': 0
        }
        
        for trip_num in df_clean['Trip'].unique():
            trip_data = df_clean[df_clean['Trip'] == trip_num]
            
            # Get location info
            provinces = set(trip_data['จังหวัด'].dropna().tolist()) if 'จังหวัด' in trip_data.columns else set()
            
            # Count same vs mixed province
            if len(provinces) == 1:
                location_stats['same_province'] += 1
            elif len(provinces) > 1:
                location_stats['mixed_province'] += 1
            
            trip_patterns[int(trip_num)] = {
                'branches': len(trip_data),
                'codes': trip_data['BranchCode'].tolist(),
                'weight': trip_data['TOTALWGT'].sum() if 'TOTALWGT' in trip_data.columns else 0,
                'cube': trip_data['TOTALCUBE'].sum() if 'TOTALCUBE' in trip_data.columns else 0,
                'provinces': list(provinces),
                'same_province': len(provinces) == 1
            }
        
        # Calculate stats
        if trip_patterns:
            location_stats['avg_branches'] = sum(t['branches'] for t in trip_patterns.values()) / len(trip_patterns)
            total = location_stats['same_province'] + location_stats['mixed_province']
            location_stats['same_province_pct'] = (location_stats['same_province'] / total * 100) if total > 0 else 0
        
        return {
            'patterns': trip_patterns, 
            'stats': location_stats,
            'punthai_restrictions': punthai_restrictions
        }
    except:
        return {'patterns': {}, 'stats': {}, 'punthai_restrictions': {}}

# โหลด Booking History (ข้อจำกัดรถ)
BOOKING_RESTRICTIONS = load_booking_history_restrictions()

# โหลด Punthai Reference (location patterns)
PUNTHAI_PATTERNS = load_punthai_reference()

# ==========================================
# HELPER FUNCTIONS
# ==========================================
def normalize(val):
    """ทำให้รหัสสาขาเป็นมาตรฐาน"""
    return str(val).strip().upper().replace(" ", "").replace(".0", "")

def calculate_distance(lat1, lon1, lat2, lon2):
    """คำนวณระยะทางระหว่างสองจุด (กม.) - Haversine formula"""
    if lat1 == 0 or lon1 == 0 or lat2 == 0 or lon2 == 0:
        return 0
    import math
    lat1_rad, lon1_rad = math.radians(lat1), math.radians(lon1)
    lat2_rad, lon2_rad = math.radians(lat2), math.radians(lon2)
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    return 6371 * c

def calculate_distance_from_dc(lat, lon):
    """คำนวณระยะทางจาก DC วังน้อย (กม.)"""
    return calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)

def calculate_bearing(lat1, lon1, lat2, lon2):
    """
    คำนวณทิศทาง (bearing) จากจุด 1 ไปจุด 2
    Returns: มุมเป็นองศา (0-360) โดย 0 = เหนือ, 90 = ตะวันออก, 180 = ใต้, 270 = ตะวันตก
    """
    if lat1 == 0 or lon1 == 0 or lat2 == 0 or lon2 == 0:
        return None
    import math
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    dlon_rad = math.radians(lon2 - lon1)
    
    x = math.sin(dlon_rad) * math.cos(lat2_rad)
    y = math.cos(lat1_rad) * math.sin(lat2_rad) - math.sin(lat1_rad) * math.cos(lat2_rad) * math.cos(dlon_rad)
    
    bearing_rad = math.atan2(x, y)
    bearing_deg = math.degrees(bearing_rad)
    
    # แปลงเป็น 0-360
    return (bearing_deg + 360) % 360

def is_opposite_direction(bearing1, bearing2, threshold=120):
    """
    เช็คว่าสองทิศทางตรงกันข้ามหรือไม่
    Args:
        bearing1, bearing2: ทิศทางเป็นองศา (0-360)
        threshold: เกณฑ์แตกต่างขั้นต่ำที่ถือว่าตรงข้าม (องศา)
    Returns:
        True ถ้าทิศทางตรงข้ามกัน
    """
    if bearing1 is None or bearing2 is None:
        return False
    
    # คำนวณความแตกต่างของมุม
    diff = abs(bearing1 - bearing2)
    if diff > 180:
        diff = 360 - diff
    
    # ถ้าแตกต่างมากกว่า threshold ถือว่าตรงข้าม
    return diff >= threshold

def check_branch_vehicle_compatibility(branch_code, vehicle_type):
    """ตรวจสอบว่าสาขานี้ใช้รถประเภทนี้ได้ไหม (รวม Booking + Punthai)"""
    branch_code_str = str(branch_code).strip()
    
    # 1. ลองหาจาก Booking History ก่อน (ข้อมูลจริง)
    booking_restrictions = BOOKING_RESTRICTIONS.get('branch_restrictions', {})
    if branch_code_str in booking_restrictions:
        allowed = booking_restrictions[branch_code_str].get('allowed', [])
        return vehicle_type in allowed
    
    # 2. ถ้าไม่มี ลองหาจาก Punthai (แผน)
    punthai_restrictions = PUNTHAI_PATTERNS.get('punthai_restrictions', {})
    if branch_code_str in punthai_restrictions:
        allowed = punthai_restrictions[branch_code_str].get('allowed', [])
        return vehicle_type in allowed
    
    # 3. ถ้าไม่มีข้อมูล = ยืดหยุ่น
    return True

def get_max_vehicle_for_branch(branch_code):
    """ดึงรถใหญ่สุดที่สาขานี้รองรับ (รวม Booking History + Punthai)"""
    branch_code_str = str(branch_code).strip()
    
    # 1. ลองหาจาก Booking History ก่อน (ข้อมูลจริง - ความเชื่อมั่นสูง)
    booking_restrictions = BOOKING_RESTRICTIONS.get('branch_restrictions', {})
    if branch_code_str in booking_restrictions:
        return booking_restrictions[branch_code_str].get('max_vehicle', '6W')
    
    # 2. ถ้าไม่มี ลองหาจาก Punthai (แผน - สำรอง)
    punthai_restrictions = PUNTHAI_PATTERNS.get('punthai_restrictions', {})
    if branch_code_str in punthai_restrictions:
        return punthai_restrictions[branch_code_str].get('max_vehicle', '6W')
    
    # 3. ถ้าไม่มีทั้งสองแหล่ง = ใช้รถใหญ่ได้
    return '6W'

def get_max_vehicle_for_trip(trip_codes):
    """
    หารถใหญ่สุดที่ทริปนี้ใช้ได้ (เช็คข้อจำกัดของทุกสาขาในทริป)
    
    Args:
        trip_codes: set ของ branch codes ในทริป
    
    Returns:
        str: '4W', 'JB', หรือ '6W'
    """
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    max_allowed = '6W'  # เริ่มจากใหญ่สุด แล้วจำกัดตามข้อจำกัดสาขา
    min_priority = 3  # ค่าใหญ่สุดคือไม่มีข้อจำกัด
    
    for code in trip_codes:
        branch_max = get_max_vehicle_for_branch(code)
        priority = vehicle_priority.get(branch_max, 3)
        
        # 🔒 เลือกรถที่เล็กที่สุด (ข้อจำกัดมากที่สุด) จากทุกสาขาในทริป
        if priority < min_priority:
            min_priority = priority
            max_allowed = branch_max
    
    return max_allowed

def get_required_vehicle_by_distance(branch_code):
    """ตรวจสอบว่าสาขาต้องใช้รถอะไรตามระยะทางจาก DC"""
    # ดึงพิกัดจาก Master
    if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
        master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == branch_code]
        if len(master_row) > 0:
            lat = master_row.iloc[0].get('ละติจูด', 0)
            lon = master_row.iloc[0].get('ลองติจูด', 0)
            distance = calculate_distance_from_dc(lat, lon)
            
            # ถ้าห่างจาก DC เกินกำหนด → ต้องใช้ 6W
            if distance > DISTANCE_REQUIRE_6W:
                return '6W', distance
    
    return None, 0

def can_fit_truck(total_weight, total_cube, truck_type):
    """เช็คว่าน้ำหนัก/คิวใส่รถได้หรือไม่"""
    limits = LIMITS[truck_type]
    max_w = limits['max_w'] * BUFFER
    max_c = limits['max_c'] * BUFFER
    return total_weight <= max_w and total_cube <= max_c

def suggest_truck(total_weight, total_cube, max_allowed='6W', trip_codes=None):
    """
    แนะนำรถที่เหมาะสม โดยเลือกรถที่:
    1. ใส่ของได้พอดี (ไม่เกินขีดจำกัด 105%)
    2. ใช้งานได้ใกล้ 100% มากที่สุด (เป้าหมาย: 90-100%)
    3. เคารพข้อจำกัดของสาขา (ถ้าสาขาใช้แค่ 4W = ต้องใช้ 4W เท่านั้น)
    """
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    max_size = vehicle_sizes.get(max_allowed, 3)
    
    # ตรวจสอบข้อจำกัดของสาขาทั้งหมดในกลุ่ม
    branch_max_vehicle = '4W'  # 🔒 เริ่มต้นที่ 4W (เล็กสุด) แล้วขยายเมื่อจำเป็น
    if trip_codes is not None and len(trip_codes) > 0:
        for code in trip_codes:
            branch_max = get_max_vehicle_for_branch(code)
            # หารถที่เล็กที่สุดที่ต้องใช้
            if vehicle_sizes.get(branch_max, 3) < vehicle_sizes.get(branch_max_vehicle, 3):
                branch_max_vehicle = branch_max
        
        # จำกัด max_allowed ตามข้อจำกัดของสาขา
        if vehicle_sizes.get(branch_max_vehicle, 3) < max_size:
            max_allowed = branch_max_vehicle
            max_size = vehicle_sizes.get(max_allowed, 3)
    
    best_truck = None
    best_utilization = 0
    best_distance_from_100 = 999  # ระยะห่างจาก 100%
    
    for truck in ['4W', 'JB', '6W']:
        truck_size = vehicle_sizes.get(truck, 0)
        # ถ้ารถใหญ่กว่าที่อนุญาต ข้ามไป
        if truck_size > max_size:
            continue
        if can_fit_truck(total_weight, total_cube, truck):
            # คำนวณ % การใช้รถ
            limits = LIMITS[truck]
            w_util = (total_weight / limits['max_w']) * 100
            c_util = (total_cube / limits['max_c']) * 100
            utilization = max(w_util, c_util)
            
            # คำนวณระยะห่างจาก 100%
            distance_from_100 = abs(100 - utilization)
            
            # เลือกรถที่ใกล้ 100% ที่สุด (90-105% เป็นเป้าหมาย)
            # ถ้าใช้งานใกล้เคียงกัน เลือกรถที่ใช้งานสูงกว่า
            if best_truck is None:
                best_truck = truck
                best_utilization = utilization
                best_distance_from_100 = distance_from_100
            else:
                # ถ้าอยู่ในช่วง 90-105% เลือกที่ใกล้ 100% ที่สุด
                if 90 <= utilization <= 105:
                    if distance_from_100 < best_distance_from_100 or best_utilization < 90:
                        best_truck = truck
                        best_utilization = utilization
                        best_distance_from_100 = distance_from_100
                # ถ้าทั้งคู่ไม่อยู่ในช่วง เลือกที่ใช้งานสูงกว่า
                elif utilization > best_utilization:
                    best_truck = truck
                    best_utilization = utilization
                    best_distance_from_100 = distance_from_100
    
    if best_truck:
        return best_truck
    
    # ถ้าไม่มีรถที่เหมาะสม ใช้รถใหญ่สุดที่อนุญาต
    return max_allowed if max_allowed in LIMITS else '6W+'

def calculate_optimal_vehicle_split(total_weight, total_cube, max_allowed='6W', branch_count=0):
    """
    🚛 คำนวณการแบ่งรถที่เหมาะสม
    
    เงื่อนไข:
    - 4W: ≤12 จุด, Cube ≤ 5
    - JB: ≤12 จุด, Cube ≤ 8  
    - 6W: ไม่จำกัดจุด, Cube ต้องเต็ม ≥100%
    
    ลำดับการเลือก:
    1. 4W (ถ้า cube ≤ 5)
    2. JB (ถ้า cube ≤ 8)
    3. JB + 4W (แยก 2 คัน, 75%-95% ต่อคัน)
    4. JB + JB (แยก 2 คัน, 75%-95% ต่อคัน)
    5. 6W + JB (แยก 2 คัน, 75%-95% ต่อคัน)
    6. 4W + 4W (แยก 2 คัน, 75%-95% ต่อคัน)
    7. 6W (cube ต้อง ≥100%)
    
    Returns: (vehicle_type, split_needed, split_config)
    """
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    max_priority = vehicle_priority.get(max_allowed, 3)
    
    # คำนวณ utilization สำหรับแต่ละรถ (ใช้ Cube เป็นหลัก)
    cube_util_4w = (total_cube / LIMITS['4W']['max_c']) * 100  # max 5 cube
    cube_util_jb = (total_cube / LIMITS['JB']['max_c']) * 100  # max 8 cube
    cube_util_6w = (total_cube / LIMITS['6W']['max_c']) * 100  # max 20 cube
    
    weight_util_4w = (total_weight / LIMITS['4W']['max_w']) * 100
    weight_util_jb = (total_weight / LIMITS['JB']['max_w']) * 100
    weight_util_6w = (total_weight / LIMITS['6W']['max_w']) * 100
    
    # 🎯 เป้าหมาย: Utilization 75%-95% สำหรับการแยก, 95%-105% สำหรับคันเดียว
    SPLIT_MIN = 75   # ขั้นต่ำสำหรับแต่ละคันเมื่อแยก
    SPLIT_MAX = 95   # สูงสุดสำหรับแต่ละคันเมื่อแยก
    SINGLE_MIN = 95  # ขั้นต่ำสำหรับคันเดียว
    SINGLE_MAX = 105 # สูงสุดสำหรับคันเดียว
    
    # ตรวจสอบจำนวนสาขา (4W/JB ไม่เกิน 12 จุด)
    branch_ok_for_small = branch_count <= 12 or branch_count == 0
    
    # 1. ลอง 4W ก่อน (ถ้า cube ≤ 5 และ ≤12 จุด)
    if max_priority >= 1 and total_cube <= 5.0 and branch_ok_for_small:
        if cube_util_4w <= 105 and weight_util_4w <= 105:
            return ('4W', False, None)
    
    # 2. ลอง JB (ถ้า cube ≤ 8 และ ≤12 จุด)
    if max_priority >= 2 and total_cube <= 8.0 and branch_ok_for_small:
        if cube_util_jb <= 105 and weight_util_jb <= 105:
            return ('JB', False, None)
    
    # 3. ถ้ารถเดียวไม่พอ ต้องแยก (cube > 8 หรือ จุด > 12)
    need_split = total_cube > 8.0 or not branch_ok_for_small
    
    if need_split:
        # 🔄 ลองแบบต่างๆ ตามลำดับ - เป้าหมาย 75%-95% ต่อคัน
        
        # JB + 4W (JB 8 cube + 4W 5 cube = 13 cube max)
        if max_priority >= 2 and total_cube <= 13.0:
            # แบ่ง: JB รับ cube มากกว่า, 4W รับส่วนที่เหลือ
            jb_cube = min(total_cube * 0.6, 8.0)  # JB รับ 60% แต่ไม่เกิน 8
            four_w_cube = total_cube - jb_cube
            
            jb_util = (jb_cube / LIMITS['JB']['max_c']) * 100
            four_w_util = (four_w_cube / LIMITS['4W']['max_c']) * 100
            
            if SPLIT_MIN <= jb_util <= SPLIT_MAX and SPLIT_MIN <= four_w_util <= SPLIT_MAX:
                return ('JB', True, {'split': ['JB', '4W'], 'ratio': [jb_cube/total_cube, four_w_cube/total_cube]})
        
        # JB + JB (JB 8 + JB 8 = 16 cube max)
        if max_priority >= 2 and total_cube <= 16.0:
            jb_util_half = (total_cube / 2 / LIMITS['JB']['max_c']) * 100
            if SPLIT_MIN <= jb_util_half <= SPLIT_MAX:
                return ('JB', True, {'split': ['JB', 'JB'], 'ratio': [0.5, 0.5]})
        
        # 6W + JB (6W 20 + JB 8 = 28 cube max)
        if max_priority >= 3 and total_cube <= 28.0:
            # แบ่ง: 6W รับส่วนใหญ่
            six_w_cube = min(total_cube * 0.7, 20.0)
            jb_cube = total_cube - six_w_cube
            
            six_w_util = (six_w_cube / LIMITS['6W']['max_c']) * 100
            jb_util = (jb_cube / LIMITS['JB']['max_c']) * 100
            
            if six_w_util >= 75 and SPLIT_MIN <= jb_util <= SPLIT_MAX:
                return ('6W', True, {'split': ['6W', 'JB'], 'ratio': [six_w_cube/total_cube, jb_cube/total_cube]})
        
        # 4W + 4W (4W 5 + 4W 5 = 10 cube max) - สำหรับสาขาที่จำกัด 4W
        if max_priority == 1 and total_cube <= 10.0:
            four_w_util_half = (total_cube / 2 / LIMITS['4W']['max_c']) * 100
            if SPLIT_MIN <= four_w_util_half <= SPLIT_MAX:
                return ('4W', True, {'split': ['4W', '4W'], 'ratio': [0.5, 0.5]})
    
    # 4. 6W (ไม่จำกัดจุด แต่ cube ต้อง ≥100%)
    if max_priority >= 3:
        if cube_util_6w >= 100:
            return ('6W', False, None)
        elif cube_util_6w >= 80:
            # 6W ไม่เต็ม (80-99%) → ยังพอรับได้
            return ('6W', False, None)
        else:
            # 6W ว่างมาก (<80%) → ลดเป็น JB ถ้าได้
            if total_cube <= 8.0 and branch_ok_for_small and max_priority >= 2:
                return ('JB', False, None)
            # ถ้า JB ไม่ได้ ลดเป็น 4W
            if total_cube <= 5.0 and branch_ok_for_small:
                return ('4W', False, None)
    
    # Default: ใช้ max_allowed
    return (max_allowed, False, None)

def can_branch_use_vehicle(code, vehicle_type, branch_vehicles):
    """
    เช็คว่าสาขาสามารถใช้รถประเภทนี้ได้หรือไม่
    - ถ้าไม่มีประวัติ = ใช้ได้ทุกประเภท
    - ถ้ามีประวัติใช้รถใหญ่ = ใช้รถเล็กกว่าได้
    - ถ้ามีประวัติใช้แค่รถเล็ก (เช่น 4W) = ใช้รถใหญ่ไม่ได้ (รถใหญ่เข้าไม่ได้)
    """
    if not branch_vehicles or code not in branch_vehicles:
        return True  # ไม่มีประวัติ = ใช้ได้ทุกประเภท
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return True  # ไม่มีข้อมูลรถ = ใช้ได้ทุกประเภท
    
    # ถ้าเคยใช้รถประเภทนี้ = ใช้ได้
    if vehicle_type in vehicle_history:
        return True
    
    # เช็คขนาดรถ (6W > JB > 4W)
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    requested_size = vehicle_sizes.get(vehicle_type, 0)
    
    # หารถที่ใหญ่ที่สุดที่สาขาเคยใช้
    max_used_size = max(vehicle_sizes.get(v, 0) for v in vehicle_history)
    
    # ถ้าขอใช้รถเล็กกว่าหรือเท่ากับที่เคยใช้ = ใช้ได้
    # ถ้าขอใช้รถใหญ่กว่าที่เคยใช้ = ใช้ไม่ได้ (รถใหญ่อาจเข้าไม่ได้)
    return requested_size <= max_used_size

def get_max_vehicle_for_branch_old(code, branch_vehicles):
    """[OLD] ดึงประเภทรถที่ใหญ่ที่สุดที่สาขาเคยใช้ (จำกัดไม่ให้ใช้รถใหญ่กว่านี้)"""
    if not branch_vehicles or code not in branch_vehicles:
        return '6W'  # ไม่มีประวัติ = ใช้ได้ถึง 6W
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return '6W'
    
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    max_vehicle = max(vehicle_history.keys(), key=lambda v: vehicle_sizes.get(v, 0))
    return max_vehicle

def get_most_used_vehicle_for_branch(code, branch_vehicles):
    """ดึงประเภทรถที่สาขาใช้บ่อยที่สุด"""
    if not branch_vehicles or code not in branch_vehicles:
        return None
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return None
    
    return max(vehicle_history, key=vehicle_history.get)

def is_similar_name(name1, name2):
    """เช็คว่าชื่อสาขาคล้ายกันหรือไม่ - รองรับทั้งไทยและอังกฤษ + ดูคำสำคัญ"""
    def extract_keywords(name):
        """ดึงคำสำคัญจากชื่อสาขา"""
        if pd.isna(name) or name is None:
            return set(), "", ""
        s = str(name).strip().upper()
        
        # คำสำคัญที่ต้องการจับคู่ (ไทย + อังกฤษ)
        keywords = set()
        
        # เช็คคำสำคัญแบบ exact match
        important_words = [
            'ฟิวเจอร์', 'FUTURE', 'รังสิต', 'RANGSIT',
            'เซ็นทรัล', 'CENTRAL', 'เทสโก้', 'TESCO', 'โลตัส', 'LOTUS',
            'บิ๊กซี', 'BIGC', 'แม็คโคร', 'MAKRO', 'โฮมโปร', 'HOMEPRO',
            'ซีคอน', 'SEACON', 'เมกา', 'MEGA', 'พาราไดซ์', 'PARADISE',
            'เทอร์มินอล', 'TERMINAL', 'สยามพารากอน', 'SIAM', 'PARAGON'
        ]
        
        for word in important_words:
            if word in s:
                keywords.add(word)
        
        # ลบ prefix/suffix ที่พบบ่อย
        prefixes = ['PTC-MRT-', 'FC PTF ', 'PTC-', 'PTC ', 'PUN-', 'PTF ', 
                   'MAXMART', 'CW', 'FC', 'NW', 'MI', 'PI']
        for prefix in prefixes:
            if s.startswith(prefix):
                s = s[len(prefix):].strip()
                break
        
        # ลบตัวอักษรเดี่ยวที่ขึ้นต้น (M, P, N) ถ้าตามด้วยตัวเลข
        import re
        if re.match(r'^[MPN]\d', s):
            s = s[1:]
        
        # แยกภาษาไทยและอังกฤษ
        thai_chars = ''.join([c for c in s if '\u0e01' <= c <= '\u0e5b'])
        eng_chars = ''.join([c for c in s if c.isalpha() and c.isascii()])
        
        return keywords, thai_chars, eng_chars
    
        # fallback: ไม่มีคำนำหน้า ให้เดาว่าเป็นคำหลัง + หรือคำแรกที่ไม่ใช่รหัส/อำเภอ/จังหวัด
        parts = address.split()
        # ข้ามรหัส plus code ถ้ามี
        idx = 0
        if parts and ('+' in parts[0] or re.match(r'^[A-Z0-9]{4,}', parts[0])):
            idx = 1
        # หาตำบลเป็นคำแรกที่ไม่ใช่รหัส/อำเภอ/จังหวัด/ตัวเลข
        for i in range(idx, len(parts)):
            if not re.match(r'^(อำเภอ|จังหวัด|\d{5}|[A-Z0-9]{4,}|ประเทศไทย)$', parts[i]):
                tambon = parts[i]
                break
    keywords1, thai1, eng1 = extract_keywords(name1) if 'name1' in locals() else (set(), '', '')
    keywords2, thai2, eng2 = extract_keywords(name2) if 'name2' in locals() else (set(), '', '')

    # 🔥 ลำดับแรก: เช็คคำสำคัญก่อน (เช่น ฟิวเจอร์+รังสิต)
    if keywords1 and keywords2:
        # ถ้ามีคำสำคัญร่วมกัน >= 2 คำ → ถือว่าเหมือนกัน
        common_keywords = keywords1 & keywords2
        if len(common_keywords) >= 2:
            return True
        # ถ้ามีคำสำคัญร่วมกัน 1 คำ แต่เป็นคำเฉพาะ → ถือว่าเหมือนกัน
        if len(common_keywords) >= 1:
            # เช็คว่ามีคำที่เป็นชื่อสถานที่เฉพาะ
            specific_places = {'รังสิต', 'RANGSIT', 'เซ็นทรัล', 'CENTRAL', 'ซีคอน', 'SEACON'}
            if common_keywords & specific_places:
                # ต้องมีอีก 1 คำ หรือ ชื่อคล้ายกัน
                if len(common_keywords) >= 2 or (thai1 and thai2 and len(thai1) >= 4 and thai1[:4] in thai2):
                    return True

    # ต้องมีความยาวพอสมควร
    if len(thai1) < 3 and len(eng1) < 3:
        return False
    if len(thai2) < 3 and len(eng2) < 3:
        return False

    # เช็คภาษาไทย
    if thai1 and thai2:
        shorter_thai = min(thai1, thai2, key=len)
        longer_thai = max(thai1, thai2, key=len)
        if len(shorter_thai) >= 3 and shorter_thai in longer_thai:
            return True
        # ความคล้าย 80%+
        if len(shorter_thai) >= 5:
            common = sum(1 for c in shorter_thai if c in longer_thai)
            if common / len(shorter_thai) >= 0.8:
                return True

    # เช็คภาษาอังกฤษ
    if eng1 and eng2:
        shorter_eng = min(eng1, eng2, key=len)
        longer_eng = max(eng1, eng2, key=len)
        if len(shorter_eng) >= 3 and shorter_eng in longer_eng:
            return True
        # ความคล้าย 80%+
        if len(shorter_eng) >= 5:
            common = sum(1 for c in shorter_eng if c in longer_eng)
            if common / len(shorter_eng) >= 0.8:
                return True

    # เช็คคำสำคัญระหว่างไทย-อังกฤษ (เช่น Future = ฟิวเจอร์, Rangsit = รังสิต)
    thai_eng_map = {
        'RANGSIT': 'รังสิต',
        'FUTURE': 'ฟิวเจอร',
        'PARK': 'ปารค',
        'TRIANGLE': 'ไตรแองเกิล',
    }

    for eng_word, thai_word in thai_eng_map.items():
        # ตรวจสอบว่ามีคำนี้ในชื่อทั้งสองฝั่ง (ไทย-อังกฤษ หรือ อังกฤษ-อังกฤษ หรือ ไทย-ไทย)
        has_eng_in_1 = eng_word in eng1
        has_eng_in_2 = eng_word in eng2
        has_thai_in_1 = thai_word in thai1
        has_thai_in_2 = thai_word in thai2
        # ถ้าทั้งสองมีคำเดียวกัน (ไม่ว่าจะไทยหรืออังกฤษ) = คล้ายกัน
        if (has_eng_in_1 and has_eng_in_2) or (has_thai_in_1 and has_thai_in_2):
            return True
        # ถ้าข้ามภาษา (อังกฤษ-ไทย)
        if (has_eng_in_1 and has_thai_in_2) or (has_eng_in_2 and has_thai_in_1):
            return True
    
    return False

def haversine_distance(lat1, lon1, lat2, lon2):
    """
    คำนวณระยะทางระหว่างจุดสองจุดบนพื้นโลก (km)
    ใช้สูตร Haversine
    """
    from math import radians, sin, cos, sqrt, atan2
    
    # แปลงองศาเป็น radians
    lat1_rad = radians(lat1)
    lon1_rad = radians(lon1)
    lat2_rad = radians(lat2)
    lon2_rad = radians(lon2)
    
    # ความต่างของพิกัด
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    
    # สูตร Haversine
    a = sin(dlat/2)**2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    
    # รัศมีโลก (km)
    R = 6371.0
    distance = R * c
    
    return distance

def get_region_type(province):
    """
    กำหนดประเภทพื้นที่และรถที่เหมาะสม
    
    Returns:
        str: 'nearby' (ใกล้ - ใช้ 4W/JB), 'far' (ไกล - ใช้ 6W), 
             'very_far' (ไกลมาก - ต้อง 6W เท่านั้น), 'unknown'
    """
    if pd.isna(province):
        return 'unknown'
    
    prov = str(province).strip()
    
    # 🚛 พื้นที่ไกลมากๆ (ภาคเหนือตอนบน + ภาคใต้ลึก) → ต้องใช้ 6W เท่านั้น
    very_far_provinces = [
        # ภาคเหนือตอนบน (ไกลจาก DC วังน้อย ~500-700 กม.)
        'เชียงใหม่', 'เชียงราย', 'แม่ฮ่องสอน', 'น่าน', 'พะเยา',
        # ภาคใต้ลึก (ไกลจาก DC วังน้อย ~700-1000 กม.)
        'สงขลา', 'ปัตตานี', 'ยะลา', 'นราธิวาส', 'พัทลุง', 'ตรัง', 'สตูล'
    ]
    
    for very_far in very_far_provinces:
        if very_far in prov:
            return 'very_far'
    
    # กรุงเทพ + ปริมณฑล + ภาคกลาง = ใกล้ → ใช้ 4W/JB
    nearby_provinces = [
        'กรุงเทพมหานคร', 'กรุงเทพ',
        'นครปฐม', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ', 'สมุทรสาคร',
        'ชัยนาท', 'พระนครศรีอยุธยา', 'ลพบุรี', 'สระบุรี', 'สิงห์บุรี', 'อ่างทอง', 'อยุธยา',
        'สมุทรสงคราม', 'สุพรรณบุรี', 'นครนายก'
    ]
    
    for nearby in nearby_provinces:
        if nearby in prov:
            return 'nearby'
    
    # จังหวัดอื่นๆ = ไกล → ใช้ 6W
    return 'far'

def is_nearby_province(prov1, prov2):
    """เช็คว่าจังหวัดใกล้เคียงกันหรือไม่ (จากไฟล์ประวัติ)"""
    if pd.isna(prov1) or pd.isna(prov2):
        return False
    
    if prov1 == prov2:
        return True
    
    # จัดกลุ่มจังหวัดตามภาคย่อย (จากไฟล์ประวัติ)
    province_groups = {
        'กรุงเทพ': ['กรุงเทพมหานคร', 'กรุงเทพ'],
        'ปริมณฑล': ['นครปฐม', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ', 'สมุทรสาคร'],
        'กลางตอนบน': ['ชัยนาท', 'พระนครศรีอยุธยา', 'ลพบุรี', 'สระบุรี', 'สิงห์บุรี', 'อ่างทอง', 'อยุธยา'],
        'กลางตอนล่าง': ['สมุทรสงคราม', 'สุพรรณบุรี'],
        'ภาคตะวันตก': ['กาญจนบุรี', 'ประจวบคีรีขันธ์', 'ราชบุรี', 'เพชรบุรี'],
        'ภาคตะวันออก': ['จันทบุรี', 'ชลบุรี', 'ตราด', 'นครนายก', 'ปราจีนบุรี', 'ระยอง', 'สระแก้ว', 'ฉะเชิงเทรา'],
        'อีสานเหนือ': ['นครพนม', 'บึงกาฬ', 'มุกดาหาร', 'สกลนคร', 'หนองคาย', 'หนองบัวลำภู', 'อุดรธานี', 'เลย'],
        'อีสานกลาง': ['กาฬสินธุ์', 'ขอนแก่น', 'ชัยภูมิ', 'มหาสารคาม', 'ร้อยเอ็ด'],
        'อีสานใต้': ['นครราชสีมา', 'โคราช', 'บุรีรัมย์', 'ยโสธร', 'ศรีสะเกษ', 'สุรินทร์', 'อำนาจเจริญ', 'อุบลราชธานี'],
        'เหนือตอนบน': ['น่าน', 'พะเยา', 'ลำปาง', 'ลำพูน', 'เชียงราย', 'เชียงใหม่', 'แพร่', 'แม่ฮ่องสอน'],
        'เหนือตอนล่าง': ['กำแพงเพชร', 'ตาก', 'นครสวรรค์', 'พิจิตร', 'พิษณุโลก', 'สุโขทัย', 'อุตรดิตถ์', 'อุทัยธานี', 'เพชรบูรณ์'],
        'ใต้ฝั่งอันดามัน': ['กระบี่', 'ตรัง', 'พังงา', 'ภูเก็ต', 'ระนอง', 'สตูล'],
        'ใต้ฝั่งอ่าวไทย': ['ชุมพร', 'นครศรีธรรมราช', 'พัทลุง', 'ยะลา', 'สงขลา', 'สุราษฎร์ธานี', 'ปัตตานี', 'นราธิวาส']
    }
    
    # หาว่าจังหวัดทั้ง 2 อยู่กลุ่มเดียวกันหรือไม่
    for group, provinces in province_groups.items():
        in_group_1 = any(p in str(prov1) for p in provinces)
        in_group_2 = any(p in str(prov2) for p in provinces)
        
        if in_group_1 and in_group_2:
            return True
    
    return False

def load_model():
    """โหลดโมเดลที่เทรนไว้"""
    if not os.path.exists(MODEL_PATH):
        return None
    
    try:
        with open(MODEL_PATH, 'rb') as f:
            model_data = pickle.load(f)
        return model_data
    except Exception as e:
        st.error(f"❌ Error loading model: {e}")
        return None

def create_pair_features(code1, code2, branch_info):
    """สร้าง features สำหรับคู่สาขา"""
    import math
    
    info1 = branch_info[code1]
    info2 = branch_info[code2]
    
    # คำนวณความต่างของน้ำหนักและคิว
    weight_diff = abs(info1['avg_weight'] - info2['avg_weight'])
    cube_diff = abs(info1['avg_cube'] - info2['avg_cube'])
    weight_sum = info1['avg_weight'] + info2['avg_weight']
    cube_sum = info1['avg_cube'] + info2['avg_cube']
    
    # จังหวัดเดียวกันหรือไม่
    same_province = 1 if info1['province'] == info2['province'] else 0
    
    # คำนวณระยะทางจากพิกัด
    distance_km = 0.0
    if info1['latitude'] != 0 and info2['latitude'] != 0:
        lat1, lon1 = math.radians(info1['latitude']), math.radians(info1['longitude'])
        lat2, lon2 = math.radians(info2['latitude']), math.radians(info2['longitude'])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        distance_km = 6371 * c
    
    # ความถี่
    freq_product = info1['total_trips'] * info2['total_trips']
    freq_diff = abs(info1['total_trips'] - info2['total_trips'])
    
    # Ratio
    weight_ratio = (info1['avg_weight'] / info2['avg_weight']) if info2['avg_weight'] > 0 else 0
    cube_ratio = (info1['avg_cube'] / info2['avg_cube']) if info2['avg_cube'] > 0 else 0
    
    # ข้อจำกัดรถ
    over_4w = 1 if (weight_sum > 2500 or cube_sum > 5.0) else 0
    over_jb = 1 if (weight_sum > 3500 or cube_sum > 8.0) else 0
    over_6w = 1 if (weight_sum > 5800 or cube_sum > 22.0) else 0
    
    return {
        'weight_sum': weight_sum,
        'cube_sum': cube_sum,
        'weight_diff': weight_diff,
        'cube_diff': cube_diff,
        'same_province': same_province,
        'distance_km': distance_km,
        'avg_weight_1': info1['avg_weight'],
        'avg_weight_2': info2['avg_weight'],
        'avg_cube_1': info1['avg_cube'],
        'avg_cube_2': info2['avg_cube'],
        'freq_product': freq_product,
        'freq_diff': freq_diff,
        'weight_ratio': weight_ratio,
        'cube_ratio': cube_ratio,
        'over_4w': over_4w,
        'over_jb': over_jb,
        'over_6w': over_6w
    }

def load_excel(file_content, sheet_name=None):
    """โหลด Excel"""
    try:
        xls = pd.ExcelFile(io.BytesIO(file_content))
        
        target_sheet = None
        if sheet_name and sheet_name in xls.sheet_names:
            target_sheet = sheet_name
        else:
            for s in xls.sheet_names:
                if 'punthai' in s.lower() or '2.' in s.lower():
                    target_sheet = s
                    break
        
        if not target_sheet:
            target_sheet = xls.sheet_names[0]
        
        # หา header row
        df_temp = pd.read_excel(xls, sheet_name=target_sheet, header=None)
        header_row = 0
        
        for i in range(min(10, len(df_temp))):
            row_values = df_temp.iloc[i].astype(str).str.upper()
            match_count = sum([
                'BRANCH' in ' '.join(row_values),
                'TRIP' in ' '.join(row_values),
                'รหัสสาขา' in ' '.join(df_temp.iloc[i].astype(str))
            ])
            if match_count >= 2:
                header_row = i
                break
        
        df = pd.read_excel(xls, sheet_name=target_sheet, header=header_row)
        df = df.loc[:, ~df.columns.duplicated()]
        
        return df
    except Exception as e:
        st.error(f"❌ Error: {e}")
        return None

def process_dataframe(df):
    """แปลงคอลัมน์เป็นรูปแบบมาตรฐาน - รองรับข้อมูลจาก Google Sheets"""
    if df is None or df.empty:
        return None
    
    rename_map = {}
    
    # ตรวจสอบจากชื่อคอลัมน์ทั้งหมดก่อน (สำหรับ Google Sheets)
    for col in df.columns:
        col_clean = str(col).strip()
        col_upper = col_clean.upper().replace(' ', '').replace('_', '')
        
        # รหัสสาขา
        if any(keyword in col_clean.lower() for keyword in ['plan code', 'branch code', 'รหัสสาขา', 'รหัส wms', 'code']):
            if 'Code' not in rename_map.values():  # เอาตัวแรกที่เจอ
                rename_map[col] = 'Code'
        
        # ชื่อสาขา
        elif any(keyword in col_clean.lower() for keyword in ['branch name', 'ชื่อสาขา', 'สาขา', 'branch']) and 'code' not in col_clean.lower():
            if 'Name' not in rename_map.values():
                rename_map[col] = 'Name'
        
        # น้ำหนัก
        elif any(keyword in col_upper for keyword in ['TOTALWGT', 'WEIGHT', 'WGT']) or 'น้ำหนัก' in col_clean:
            if 'Weight' not in rename_map.values():
                rename_map[col] = 'Weight'
        
        # คิว/ปริมาตร
        elif any(keyword in col_upper for keyword in ['TOTALCUBE', 'CUBE', 'VOLUME']) or 'คิว' in col_clean or 'ปริมาตร' in col_clean:
            if 'Cube' not in rename_map.values():
                rename_map[col] = 'Cube'
        
        # พิกัด
        elif 'latitude' in col_clean.lower() or col_clean == 'ละติจูด' or col_upper == 'LAT':
            rename_map[col] = 'Latitude'
        elif 'longitude' in col_clean.lower() or 'ลองติจูด' in col_clean or 'ลองจิจูด' in col_clean or col_upper in ['LON', 'LONG', 'LNG']:
            rename_map[col] = 'Longitude'
        
        # ที่อยู่
        elif 'จังหวัด' in col_clean or 'province' in col_clean.lower():
            if 'Province' not in rename_map.values():
                rename_map[col] = 'Province'
        elif 'อำเภอ' in col_clean or 'district' in col_clean.lower() and 'sub' not in col_clean.lower():
            if 'District' not in rename_map.values():
                rename_map[col] = 'District'
        elif 'ตำบล' in col_clean or 'subdistrict' in col_clean.lower() or 'sub district' in col_clean.lower():
            if 'Subdistrict' not in rename_map.values():
                rename_map[col] = 'Subdistrict'
        
        # ทริป
        elif col_upper in ['TRIPNO', 'TRIP_NO'] or col_clean == 'Trip no':
            rename_map[col] = 'TripNo'
        elif col_upper == 'TRIP' or 'ทริป' in col_clean or 'เที่ยว' in col_clean:
            rename_map[col] = 'Trip'
        elif 'BOOKING' in col_upper:
            rename_map[col] = 'Booking'
    
    # ถ้าไม่เจอจากชื่อคอลัมน์ ให้ลองใช้ลำดับ (สำหรับไฟล์ Auto Plan)
    if 'Code' not in rename_map.values() and len(df.columns) >= 8:
        col_list = list(df.columns)
        if len(col_list) > 2 and col_list[2] not in rename_map:
            rename_map[col_list[2]] = 'Code'
        if len(col_list) > 4 and col_list[4] not in rename_map:
            rename_map[col_list[4]] = 'Name'
        if len(col_list) > 5 and col_list[5] not in rename_map:
            rename_map[col_list[5]] = 'Cube'
        if len(col_list) > 6 and col_list[6] not in rename_map:
            rename_map[col_list[6]] = 'Weight'
    
    # ใช้ mapping
    df = df.rename(columns=rename_map)
    
    # ลบคอลัมน์ซ้ำ
    df = df.loc[:, ~df.columns.duplicated()]
    
    # ทำความสะอาด Code
    if 'Code' in df.columns:
        df['Code'] = df['Code'].apply(normalize)
        
        # ตัดสาขาที่ไม่ต้องการออก
        df = df[~df['Code'].isin(EXCLUDE_BRANCHES)]
        
        # ตัดสาขาที่ชื่อมี keyword ที่ไม่ต้องการ
        if 'Name' in df.columns:
            exclude_pattern = '|'.join(EXCLUDE_NAMES)
            df = df[~df['Name'].str.contains(exclude_pattern, case=False, na=False)]
    
    # แปลง Weight และ Cube เป็นตัวเลข
    for col in ['Weight', 'Cube']:
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    
    # เพิ่มจังหวัดจาก Master ถ้ายังไม่มี
    if 'Province' not in df.columns or df['Province'].isna().all():
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns and 'Code' in df.columns:
            # สร้าง mapping จาก Master
            province_map = {}
            for _, row in MASTER_DATA.iterrows():
                code = row.get('Plan Code', '')
                province = row.get('จังหวัด', '')
                if code and province:
                    province_map[code] = province
            
            # ฟังก์ชันค้นหาจังหวัดจากชื่อสาขา
            def find_province_by_name(code, name):
                # ลองหาจาก code ก่อน
                if code in province_map:
                    return province_map[code]
                
                # ถ้าไม่เจอ ลองค้นหาจากชื่อสาขา
                if not name or pd.isna(name):
                    return None
                
                # แยกคำสำคัญจากชื่อ (เอาคำแรกที่ไม่ใช่ prefix)
                keywords = str(name).replace('MAX MART-', '').replace('PUNTHAI-', '').replace('LUBE', '').strip()
                if not keywords:
                    return None
                
                # ค้นหาในชื่อสาขาของ Master
                for _, master_row in MASTER_DATA.iterrows():
                    master_name = str(master_row.get('สาขา', ''))
                    # ถ้าชื่อคล้ายกัน (มีคำสำคัญเหมือนกัน)
                    if keywords[:10] in master_name or master_name[:10] in keywords:
                        province = master_row.get('จังหวัด', '')
                        if province:
                            return province
                
                return None
            
            # ใส่จังหวัดให้แต่ละสาขา
            if 'Name' in df.columns:
                df['Province'] = df.apply(lambda row: find_province_by_name(row['Code'], row.get('Name', '')), axis=1)
            else:
                df['Province'] = df['Code'].map(province_map)
    
    return df.reset_index(drop=True)

# ============================================================================
# 🎯 MAIN APPLICATION
# ============================================================================

def main():
    st.set_page_config(
        page_title="ระบบจัดเที่ยว",
        page_icon="🚚",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # 🔄 Sync ข้อมูลจาก Google Sheets ทุกครั้งที่โหลดเว็บ (ไม่แสดง error ถ้าไม่มี)
    try:
        synced_df = sync_branch_data_from_sheets()
        if synced_df is not None and not synced_df.empty:
            st.success(f"✅ โหลดข้อมูล Master สำเร็จ: {len(synced_df)} สาขา", icon="📊")
    except Exception as e:
        # ไม่แสดง error ถ้าเป็นแค่ไม่มี credentials
        pass
    
    # ==========================================
    # Step 6: DISTRICT CLUSTERING ALLOCATION (OPTIMIZED)
    # 🔥 จัดทริปโดยเน้นสาขาที่มีข้อจำกัดรถก่อน (4W → JB → 6W)
    # ==========================================
    trip_counter = 1
    df['Trip'] = 0
    
    # 🚀 CACHE: Pre-compute branch constraints และ BU type
    branch_max_vehicle_cache = {}
    branch_bu_cache = {}
    branch_priority_cache = {}  # เพิ่ม: cache สำหรับ priority ตามข้อจำกัดรถ
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    
    for _, row in df.iterrows():
        code = row['Code']
        max_vehicle = row.get('_max_vehicle', '6W')
        branch_max_vehicle_cache[code] = max_vehicle
        bu = str(row.get('BU', '')).upper()
        branch_bu_cache[code] = bu in ['211', 'PUNTHAI']
        # สาขาที่มีข้อจำกัดรถ (4W, JB) จะมี priority สูงกว่า
        branch_priority_cache[code] = vehicle_priority.get(max_vehicle, 3)
    
    # 🚀 Pre-compute limits with buffer (CACHED VERSION)
    limits_cache = {}  # Cache: (max_vehicle, is_punthai) -> limits
    
    def get_max_limits(allowed_vehicles, is_punthai):
        """หา capacity สูงสุดที่ใช้ได้ (with caching)"""
        buffer_mult = punthai_buffer if is_punthai else maxmart_buffer
        max_vehicle = '6W' if '6W' in allowed_vehicles else ('JB' if 'JB' in allowed_vehicles else '4W')
        
        # Check cache first
        cache_key = (max_vehicle, is_punthai)
        if cache_key in limits_cache:
            return limits_cache[cache_key]
        
        limits_to_use = PUNTHAI_LIMITS if is_punthai else LIMITS
        lim = limits_to_use.get(max_vehicle, LIMITS['6W'])
        result = {
            'max_w': lim.get('max_w', 6000) * buffer_mult,
            'max_c': lim.get('max_c', 20.0) * buffer_mult,
            'max_d': lim.get('max_drops', 12)
        }
        
        # Store in cache
        limits_cache[cache_key] = result
        return result
    
    # Helper function: เลือกรถที่เหมาะสม (Optimized - v2)
    def select_vehicle_for_load(weight, cube, drops, is_punthai, allowed_vehicles, global_limiting_factor):
        """
        เลือกรถที่เล็กที่สุดที่รับโหลดได้ (ต้องไม่เกิน Buffer)
        🔥 ใช้ global_limiting_factor จากการวิเคราะห์ทั้งระบบ
        🚀 Optimized: ใช้ cache และลดการคำนวณ
        
        Parameters:
            global_limiting_factor: 'weight' หรือ 'cube' จากการวิเคราะห์ข้อมูลทั้งหมด
        
        Returns: (vehicle, limiting_factor)
            vehicle: ประเภทรถ ('4W', 'JB', '6W')
            limiting_factor: จะเป็นค่าเดียวกับ global_limiting_factor เสมอ
        """
        buffer_mult = punthai_buffer if is_punthai else maxmart_buffer
        limits_to_use = PUNTHAI_LIMITS if is_punthai else LIMITS
        
        # Pre-compute limits for all vehicles (ทำครั้งเดียว)
        for v in ['4W', 'JB', '6W']:
            if v not in allowed_vehicles:
                continue
            lim = limits_to_use[v]
            # เช็คว่าไม่เกินขีดจำกัดทั้ง 3 มิติ (inline calculation)
            if (weight <= lim['max_w'] * buffer_mult and 
                cube <= lim['max_c'] * buffer_mult and 
                drops <= lim.get('max_drops', 12)):
                # 🎯 ใช้ global_limiting_factor ที่กำหนดไว้แล้ว
                return v, global_limiting_factor
        
        return None, None
    
    # 🔥 NEW: Helper function - เช็คว่าทริปมีน้ำหนักขั้นต่ำตามมาตรฐานหรือไม่
    def meets_minimum_standard(weight, cube, drops, allowed_vehicles, limiting_factor=None):
        """
        เช็คว่าทริปมีน้ำหนัก/คิวขั้นต่ำตามมาตรฐานหรือไม่
        🔥 ถ้าตัดที่น้ำหนัก → วัดที่น้ำหนัก
        🔥 ถ้าตัดที่คิว → วัดที่คิว
        
        เกณฑ์มาตรฐาน:
        - Weight: ≥ 70% ของขีดจำกัด
        - Cube: ≥ 70% ของขีดจำกัด  
        """
        if not allowed_vehicles:
            return True
        
        # หารถที่เล็กที่สุดที่ใช้ได้
        min_vehicle = '4W' if '4W' in allowed_vehicles else ('JB' if 'JB' in allowed_vehicles else '6W')
        lim = LIMITS[min_vehicle]
        
        # ถ้าไม่ระบุ limiting_factor ให้หาเอง (จาก weight กับ cube)
        if limiting_factor is None:
            weight_pct = (weight / lim['max_w']) * 100 if lim['max_w'] > 0 else 0
            cube_pct = (cube / lim['max_c']) * 100 if lim['max_c'] > 0 else 0
            
            if weight_pct >= cube_pct:
                limiting_factor = 'weight'
            else:
                limiting_factor = 'cube'
        
        # 🎯 วัดมาตรฐานตาม limiting_factor เท่านั้น!
        if limiting_factor == 'weight':
            min_weight = lim['max_w'] * 0.70  # ≥ 70%
            return weight >= min_weight
        elif limiting_factor == 'cube':
            min_cube = lim['max_c'] * 0.70  # ≥ 70%
            return cube >= min_cube
        
        # Fallback (ไม่ควรเกิด)
        return True
    
    # Helper function: เช็คว่าเป็น Punthai ล้วนหรือไม่ (Optimized - ใช้ cache)
    def is_all_punthai_codes(codes):
        if not codes:
            return False
        return all(branch_bu_cache.get(c, False) for c in codes)
    
    # Helper function: หา allowed vehicles จาก codes (Optimized)
    def get_allowed_from_codes(codes, base_allowed):
        """หา allowed vehicles โดยรวม branch constraints"""
        result = set(base_allowed)
        for code in codes:
            branch_max = branch_max_vehicle_cache.get(code, '6W')
            if branch_max == 'JB':
                result.discard('6W')
            elif branch_max == '4W':
                result.discard('6W')
                result.discard('JB')
        return list(result)
    
    # 🔥 NEW: Helper function - คำนวณ centroid ของทริปจากพิกัดจริง
    def calculate_trip_centroid(codes):
        """คำนวณจุดกึ่งกลางของทริปจากพิกัด lat/lon"""
        if not codes:
            return None, None
        trip_branches = df[df['Code'].isin(codes)]
        valid_coords = trip_branches[(trip_branches['_lat'] != 0) & (trip_branches['_lon'] != 0)]
        if valid_coords.empty:
            return None, None
        return valid_coords['_lat'].mean(), valid_coords['_lon'].mean()
    
    # 🔥 NEW: Helper function - เช็คว่าสาขาใกล้ทริปปัจจุบันหรือไม่ (+ เช็คทิศทาง)
    def is_branch_near_trip(branch_code, trip_codes, max_distance_km=80):
        """
        เช็คว่าสาขาอยู่ใกล้กับทริปปัจจุบันหรือไม่ และอยู่ในทิศทางเดียวกัน
        Args:
            branch_code: รหัสสาขาที่จะเช็ค
            trip_codes: รหัสสาขาในทริปปัจจุบัน
            max_distance_km: ระยะห่างสูงสุดที่ยอมรับได้ (กม.)
        Returns:
            True ถ้าใกล้พอและทิศทางเหมาะสม, False ถ้าไกลเกินไปหรือทิศทางตรงข้าม
        """
        if not trip_codes:
            return True  # ทริปว่าง รับได้เลย
        
        # หาพิกัดสาขาที่จะเพิ่ม
        branch_data = df[df['Code'] == branch_code]
        if branch_data.empty:
            return True
        
        branch_lat = branch_data['_lat'].iloc[0]
        branch_lon = branch_data['_lon'].iloc[0]
        branch_province = branch_data['_province'].iloc[0]
        
        if branch_lat == 0 or branch_lon == 0:
            return True  # ไม่มีพิกัด ให้ผ่าน
        
        # คำนวณ centroid ของทริป
        trip_lat, trip_lon = calculate_trip_centroid(trip_codes)
        if trip_lat is None or trip_lon is None:
            return True  # ทริปไม่มีพิกัด ให้ผ่าน
        
        # 1. เช็คระยะห่าง
        distance = haversine_distance(trip_lat, trip_lon, branch_lat, branch_lon)
        if distance > max_distance_km:
            return False  # ไกลเกินไป
        
        # 2. 🔥 เช็คทิศทาง (ป้องกันทิศทางตรงข้าม)
        # หาจังหวัดหลักของทริป (จังหวัดที่มีสาขามากที่สุด)
        trip_provinces = df[df['Code'].isin(trip_codes)]['_province'].value_counts()
        if trip_provinces.empty:
            return True
        
        main_province = trip_provinces.index[0]
        
        # ถ้าสาขาใหม่อยู่จังหวัดเดียวกับจังหวัดหลัก → OK
        if branch_province == main_province:
            return True
        
        # ถ้าต่างจังหวัด → เช็คทิศทาง
        # หา centroid ของจังหวัดหลัก
        main_prov_branches = df[(df['Code'].isin(trip_codes)) & (df['_province'] == main_province)]
        valid_main = main_prov_branches[(main_prov_branches['_lat'] != 0) & (main_prov_branches['_lon'] != 0)]
        
        if valid_main.empty:
            return True
        
        main_prov_lat = valid_main['_lat'].mean()
        main_prov_lon = valid_main['_lon'].mean()
        
        # คำนวณทิศทางจากจังหวัดหลักไปยังสาขาแต่ละสาขาในทริป
        trip_bearings = []
        for code in trip_codes:
            code_data = df[df['Code'] == code]
            if code_data.empty:
                continue
            code_lat = code_data['_lat'].iloc[0]
            code_lon = code_data['_lon'].iloc[0]
            if code_lat != 0 and code_lon != 0:
                bearing = calculate_bearing(main_prov_lat, main_prov_lon, code_lat, code_lon)
                if bearing is not None:
                    trip_bearings.append(bearing)
        
        if not trip_bearings:
            return True  # ไม่มีข้อมูลทิศทาง ให้ผ่าน
        
        # ทิศทางเฉลี่ยของทริป
        avg_bearing = sum(trip_bearings) / len(trip_bearings)
        
        # ทิศทางของสาขาใหม่จากจังหวัดหลัก
        new_bearing = calculate_bearing(main_prov_lat, main_prov_lon, branch_lat, branch_lon)
        if new_bearing is None:
            return True
        
        # เช็คว่าทิศทางตรงข้ามหรือไม่
        if is_opposite_direction(avg_bearing, new_bearing, threshold=100):
            return False  # ทิศทางตรงข้าม ไม่รวมทริป
        
        return True  # ผ่านทั้งระยะทางและทิศทาง
    
    # Current trip state
    current_trip = {
        'codes': [], 'weight': 0, 'cube': 0, 'drops': 0,
        'region': None, 'allowed_vehicles': ['4W', 'JB', '6W'],
        'province': None,  # 🔥 เพิ่มติดตามจังหวัด
        'district': None, 'is_punthai': False,
        'limiting_factor': None  # 🔥 เก็บว่าตัดทริปด้วยอะไร (weight/cube/drops)
    }
    
    overflow_queue = []  # Queue สำหรับ stores ที่ overflow
    
    def finalize_current_trip():
        """ปิดทริปปัจจุบันและบันทึก"""
        nonlocal trip_counter
        if current_trip['codes']:
            for c in current_trip['codes']:
                df.loc[df['Code'] == c, 'Trip'] = trip_counter
    
    def split_until_fits(allowed_vehicles, region):
        """แยก stores ออกจาก current_trip จนกว่าจะพอดีรถ (ไม่เกิน buffer) - OPTIMIZED"""
        nonlocal trip_counter, overflow_queue
        
        while True:
            # ใช้ cached is_punthai
            is_punthai = current_trip['is_punthai']
            limits = get_max_limits(current_trip['allowed_vehicles'], is_punthai)
            
            # เช็คว่าเกินหรือไม่
            if (current_trip['weight'] <= limits['max_w'] and 
                current_trip['cube'] <= limits['max_c'] and 
                current_trip['drops'] <= limits['max_d']):
                break
            
            if len(current_trip['codes']) <= 1:
                break
            
            # ตัด store สุดท้ายออก
            overflow_code = current_trip['codes'].pop()
            overflow_weight = df.loc[df['Code'] == overflow_code, 'Weight'].iloc[0]
            overflow_cube = df.loc[df['Code'] == overflow_code, 'Cube'].iloc[0]
            current_trip['weight'] -= overflow_weight
            current_trip['cube'] -= overflow_cube
            current_trip['drops'] -= 1
            
            # Update is_punthai และ allowed_vehicles
            current_trip['is_punthai'] = is_all_punthai_codes(current_trip['codes'])
            current_trip['allowed_vehicles'] = get_allowed_from_codes(current_trip['codes'], allowed_vehicles)
            
            overflow_queue.append({
                'code': overflow_code,
                'weight': overflow_weight,
                'cube': overflow_cube,
                'region': region,
                'allowed_vehicles': allowed_vehicles
            })
    
    def process_overflow_queue():
        """ประมวลผล overflow queue - สร้างทริปใหม่สำหรับ stores ที่ล้น - OPTIMIZED v2"""
        nonlocal trip_counter, current_trip, overflow_queue
        
        while overflow_queue:
            item = overflow_queue.pop(0)
            code = item['code']
            weight = item['weight']
            cube = item['cube']
            region = item['region']
            allowed_vehicles = item['allowed_vehicles']
            
            # 🚀 Skip if already assigned (ป้องกัน double assignment)
            if df.loc[df['Code'] == code, 'Trip'].iloc[0] != 0:
                continue
            
            # ลองเพิ่มเข้า current_trip
            if current_trip['codes']:
                test_codes = current_trip['codes'] + [code]
                test_weight = current_trip['weight'] + weight
                test_cube = current_trip['cube'] + cube
                test_drops = current_trip['drops'] + 1
                test_punthai = is_all_punthai_codes(test_codes)
                test_allowed = get_allowed_from_codes(test_codes, allowed_vehicles)
                
                vehicle, limiting_factor = select_vehicle_for_load(test_weight, test_cube, test_drops, test_punthai, test_allowed, GLOBAL_LIMITING_FACTOR)
                
                if vehicle:
                    # พอดี! เพิ่มเข้า
                    current_trip['codes'].append(code)
                    current_trip['weight'] = test_weight
                    current_trip['cube'] = test_cube
                    current_trip['drops'] = test_drops
                    current_trip['is_punthai'] = test_punthai
                    current_trip['allowed_vehicles'] = test_allowed
                    current_trip['limiting_factor'] = limiting_factor  # 🔥 เก็บว่าตัดด้วยอะไร
                    
                    # Double check
                    split_until_fits(allowed_vehicles, region)
                else:
                    # ไม่พอดี → ปิดทริปเก่า, เริ่มใหม่
                    finalize_current_trip()
                    trip_counter += 1
                    new_allowed = get_allowed_from_codes([code], allowed_vehicles)
                    current_trip = {
                        'codes': [code],
                        'weight': weight,
                        'cube': cube,
                        'drops': 1,
                        'region': region,
                        'allowed_vehicles': new_allowed,
                        'district': None,
                        'is_punthai': branch_bu_cache.get(code, False)
                    }
            else:
                # ทริปว่าง
                new_allowed = get_allowed_from_codes([code], allowed_vehicles)
                current_trip = {
                    'codes': [code],
                    'weight': weight,
                    'cube': cube,
                    'drops': 1,
                    'region': region,
                    'allowed_vehicles': new_allowed,
                    'district': None,
                    'is_punthai': branch_bu_cache.get(code, False)
                }
    
    # ==========================================
    # GROUP BY DISTRICT BUCKETS - OPTIMIZED WITH PROXIMITY GROUPING
    # 🔥 ใช้พิกัดจริง (lat/lon) เพื่อเช็คความใกล้กันก่อนรวมทริป
    # ==========================================
    # Pre-group data for faster iteration
    district_groups = df.groupby(['_region_name', '_province', '_district'], sort=False)
    
    # ตัวแปรเก็บข้อมูลจังหวัดล่าสุด
    prev_province = None
    
    for (region, province, district), district_df in district_groups:
        # ข้อมูล District (vectorized - no dict conversion)
        district_codes = district_df['Code'].tolist()
        district_weight = district_df['Weight'].sum()
        district_cube = district_df['Cube'].sum()
        district_drops = len(district_codes)
        
        # หารถที่ใช้ได้ตามภาค
        allowed_vehicles = ['4W', 'JB', '6W']
        if region in CENTRAL_REGIONS:
            allowed_vehicles = CENTRAL_ALLOWED_VEHICLES.copy()
        
        # ==========================================
        # 🔥 Rule 0 (UPDATED): Province-Complete Mode - จัดจังหวัดเดียวกันให้เสร็จก่อน
        # ไม่แยกจังหวัดออกจากกัน - ตัดทริปทุกครั้งที่เปลี่ยนจังหวัด
        # ==========================================
        should_close_trip = False
        
        if current_trip['region'] and current_trip['region'] != region:
            # 1. เปลี่ยนภาค → ปิดทริปแน่นอน
            should_close_trip = True
        elif prev_province and prev_province != province:
            # 2. 🔥 เปลี่ยนจังหวัด → ปิดทริปทันที (ไม่ผสมจังหวัด)
            should_close_trip = True
        
        if should_close_trip:
            process_overflow_queue()
            
            # 🔥 เช็คเกณฑ์ขั้นต่ำก่อนปิดทริป (ใช้ limiting_factor ที่ตัดจริง)
            if current_trip['codes']:
                if not meets_minimum_standard(current_trip['weight'], current_trip['cube'], 
                                              current_trip['drops'], current_trip['allowed_vehicles'],
                                              current_trip['limiting_factor']):
                    # ทริปไม่ถึงเกณฑ์ขั้นต่ำ - ลองรวมต่อก่อน (ยกเว้นเปลี่ยนภาค)
                    if current_trip['region'] == region:
                        should_close_trip = False
            
            if should_close_trip:
                finalize_current_trip()
                trip_counter += 1
                current_trip = {
                    'codes': [], 'weight': 0, 'cube': 0, 'drops': 0,
                    'region': None, 'allowed_vehicles': allowed_vehicles,
                    'district': None, 'province': None,
                    'is_punthai': False, 'limiting_factor': None
                }
        
        # ==========================================
        # Rule 1: ลองใส่ทั้ง District - OPTIMIZED WITH PROXIMITY CHECK
        # 🔥 เพิ่ม: เช็คระยะห่างจากพิกัดก่อนรวมทริป
        # ==========================================
        if current_trip['codes']:
            # 🔥 เช็คว่าอำเภอนี้อยู่ใกล้กับทริปปัจจุบันหรือไม่
            # ใช้สาขาแรกของอำเภอเป็นตัวแทน
            sample_branch = district_codes[0]
            is_near = is_branch_near_trip(sample_branch, current_trip['codes'], max_distance_km=80)
            
            if not is_near:
                # อำเภอนี้ไกลเกินไป → ปิดทริปเก่า เริ่มทริปใหม่
                finalize_current_trip()
                trip_counter += 1
                
                new_allowed = get_allowed_from_codes(district_codes, allowed_vehicles)
                new_punthai = is_all_punthai_codes(district_codes)
                
                current_trip = {
                    'codes': district_codes.copy(),
                    'weight': district_weight,
                    'cube': district_cube,
                    'drops': district_drops,
                    'region': region,
                    'allowed_vehicles': new_allowed,
                    'province': province,
                    'district': district,
                    'is_punthai': new_punthai
                }
                
                split_until_fits(allowed_vehicles, region)
            else:
                # อำเภอนี้ใกล้พอ → ลองรวมทริป
                test_codes = current_trip['codes'] + district_codes
                test_weight = current_trip['weight'] + district_weight
                test_cube = current_trip['cube'] + district_cube
                test_drops = current_trip['drops'] + district_drops
                test_punthai = is_all_punthai_codes(test_codes)
                test_allowed = get_allowed_from_codes(test_codes, allowed_vehicles)
                
                vehicle, limiting_factor = select_vehicle_for_load(test_weight, test_cube, test_drops, test_punthai, test_allowed, GLOBAL_LIMITING_FACTOR)
                
                if vehicle:
                    # District พอดี!
                    current_trip['codes'].extend(district_codes)
                    current_trip['weight'] = test_weight
                    current_trip['cube'] = test_cube
                    current_trip['drops'] = test_drops
                    current_trip['allowed_vehicles'] = test_allowed
                    current_trip['region'] = region
                    current_trip['province'] = province  # 🔥 เก็บจังหวัดปัจจุบัน
                    current_trip['district'] = district
                    current_trip['is_punthai'] = test_punthai
                    current_trip['limiting_factor'] = limiting_factor  # 🔥 เก็บว่าตัดด้วยอะไร
                    
                    # Double check
                    split_until_fits(test_allowed, region)
                else:
                    # District ไม่พอดี → ปิดทริปเก่า
                    finalize_current_trip()
                    trip_counter += 1
                    
                    new_allowed = get_allowed_from_codes(district_codes, allowed_vehicles)
                    new_punthai = is_all_punthai_codes(district_codes)
                    
                    current_trip = {
                        'codes': district_codes.copy(),
                        'weight': district_weight,
                        'cube': district_cube,
                        'drops': district_drops,
                        'region': region,
                        'allowed_vehicles': new_allowed,
                        'province': province,  # 🔥 เก็บจังหวัดปัจจุบัน
                        'district': district,
                        'is_punthai': new_punthai
                    }
                    
                    # ==========================================
                    # Rule 2: ถ้า District ใหญ่เกินรถ → Split ทันที!
                    # ==========================================
                    split_until_fits(allowed_vehicles, region)
        else:
            # ทริปว่าง - หา allowed_vehicles รวม branch constraints (ใช้ cache)
            new_allowed = get_allowed_from_codes(district_codes, allowed_vehicles)
            new_punthai = is_all_punthai_codes(district_codes)
            
            current_trip = {
                'codes': district_codes.copy(),
                'weight': district_weight,
                'cube': district_cube,
                'drops': district_drops,
                'region': region,
                'allowed_vehicles': new_allowed,
                'province': province,  # 🔥 เก็บจังหวัดปัจจุบัน
                'district': district,
                'is_punthai': new_punthai
            }
            
            # ==========================================
            # Rule 2: ถ้า District ใหญ่เกินรถ → Split ทันที!
            # ==========================================
            split_until_fits(new_allowed, region)
        
        # 🔥 อัพเดตตัวแปรติดตามจังหวัด
        prev_province = province
    
    # ==========================================
    # Final: Process remaining overflow และปิดทริปสุดท้าย (รอบ 1)
    # ==========================================
    process_overflow_queue()
    finalize_current_trip()

    # ==========================================
    # 🔥 PHASE 2: CLEANUP MODE - รวมเศษจากทุกจังหวัด
    # รวมสาขาที่เหลือจากแต่ละจังหวัดที่ยังไม่ได้ทริป (Trip = 0)
    # โดยรวมตามความใกล้กันของตำบล/อำเภอข้ามจังหวัด
    # ==========================================
    remaining_df = df[df['Trip'] == 0].copy()
    
    if not remaining_df.empty:
        # Log cleanup phase
        print(f"🔄 Cleanup Phase: {remaining_df['_province'].nunique()} provinces, {len(remaining_df)} branches remaining")
        
        # เรียงตามภาค → ระยะทาง (ไกลมาใกล้) เพื่อรวมพื้นที่ใกล้กัน
        remaining_df = remaining_df.sort_values(
            ['_vehicle_priority', '_region_order', '_distance_from_dc'],
            ascending=[True, True, False]
        )
        
        # Reset current trip
        current_trip = {
            'codes': [], 'weight': 0, 'cube': 0, 'drops': 0,
            'region': None, 'allowed_vehicles': ['4W', 'JB', '6W'],
            'province': None, 'district': None, 'is_punthai': False
        }
        
        prev_region = None
        
        for idx, row in remaining_df.iterrows():
            code = row['Code']
            weight = row['Weight']
            cube = row['Cube']
            region = row['_region_name']
            
            # หารถที่ใช้ได้ตามภาค
            allowed_vehicles = ['4W', 'JB', '6W']
            if region in CENTRAL_REGIONS:
                allowed_vehicles = CENTRAL_ALLOWED_VEHICLES.copy()
            
            # 🔥 ตัดทริปเมื่อเปลี่ยนภาค (cleanup mode ยังไม่ผสมข้ามภาค)
            if prev_region and prev_region != region:
                finalize_current_trip()
                trip_counter += 1
                current_trip = {
                    'codes': [], 'weight': 0, 'cube': 0, 'drops': 0,
                    'region': None, 'allowed_vehicles': allowed_vehicles,
                    'province': None, 'district': None, 'is_punthai': False
                }
            
            # ลองเพิ่มเข้าทริปปัจจุบัน
            if current_trip['codes']:
                test_codes = current_trip['codes'] + [code]
                test_weight = current_trip['weight'] + weight
                test_cube = current_trip['cube'] + cube
                test_drops = current_trip['drops'] + 1
                test_punthai = is_all_punthai_codes(test_codes)
                test_allowed = get_allowed_from_codes(test_codes, allowed_vehicles)
                
                vehicle, limiting_factor = select_vehicle_for_load(test_weight, test_cube, test_drops, test_punthai, test_allowed, GLOBAL_LIMITING_FACTOR)
                
                if vehicle:
                    # พอดี! เพิ่มเข้า
                    current_trip['codes'].append(code)
                    current_trip['weight'] = test_weight
                    current_trip['cube'] = test_cube
                    current_trip['drops'] = test_drops
                    current_trip['region'] = region
                    current_trip['is_punthai'] = test_punthai
                    current_trip['allowed_vehicles'] = test_allowed
                    current_trip['limiting_factor'] = limiting_factor  # 🔥 เก็บว่าตัดด้วยอะไร
                else:
                    # ไม่พอดี → ปิดทริปเก่า, เริ่มใหม่
                    finalize_current_trip()
                    trip_counter += 1
                    new_allowed = get_allowed_from_codes([code], allowed_vehicles)
                    current_trip = {
                        'codes': [code],
                        'weight': weight,
                        'cube': cube,
                        'drops': 1,
                        'region': region,
                        'allowed_vehicles': new_allowed,
                        'province': None,
                        'district': None,
                        'is_punthai': branch_bu_cache.get(code, False)
                    }
            else:
                # ทริปว่าง
                new_allowed = get_allowed_from_codes([code], allowed_vehicles)
                current_trip = {
                    'codes': [code],
                    'weight': weight,
                    'cube': cube,
                    'drops': 1,
                    'region': region,
                    'allowed_vehicles': new_allowed,
                    'province': None,
                    'district': None,
                    'is_punthai': branch_bu_cache.get(code, False)
                }
            
            prev_region = region
        
        # ปิดทริปสุดท้าย
        finalize_current_trip()

    # ==========================================
    # Step 7: สร้าง Summary + Central Rule + Punthai Drop Limits
    # ==========================================
    summary_data = []
    
    for trip_num in sorted(df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = df[df['Trip'] == trip_num]
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        trip_codes = trip_data['Code'].unique()
        trip_drops = len(trip_codes)
        
        # หาภาคของทริป (ใช้ภาคแรก)
        trip_region = trip_data['_region_name'].iloc[0] if '_region_name' in trip_data.columns else 'ไม่ระบุ'
        
        # หารถที่เหมาะสม (รวม Central Rule)
        max_vehicles = [get_max_vehicle_for_branch(c) for c in trip_codes]
        min_max_size = min(vehicle_priority.get(v, 3) for v in max_vehicles)
        max_allowed_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(min_max_size, '6W')
        
        # 🚫 Central Region Rule: ห้าม 6W
        if trip_region in CENTRAL_REGIONS and max_allowed_vehicle == '6W':
            max_allowed_vehicle = 'JB'  # ลดเป็น JB
        
        # ตรวจ BU ของทริป
        is_punthai_only_trip = True
        for _, r in trip_data.iterrows():
            bu = str(r.get('BU', '')).upper()
            if bu not in ['211', 'PUNTHAI']:
                is_punthai_only_trip = False
                break
        
        buffer = punthai_buffer if is_punthai_only_trip else maxmart_buffer
        buffer_pct = int(buffer * 100)
        buffer_label = f"🅿️ {buffer_pct}%" if is_punthai_only_trip else f"🅼 {buffer_pct}%"
        trip_type = 'punthai' if is_punthai_only_trip else 'maxmart'
        
        # เลือกรถที่พอดีที่สุด
        suggested = max_allowed_vehicle
        source = "📋 จำกัดสาขา" if min_max_size < 3 else "🤖 อัตโนมัติ"
        
        # 🔒 Punthai Drop Limit Check
        if is_punthai_only_trip:
            punthai_drop_limit = PUNTHAI_LIMITS.get(suggested, {}).get('max_drops', 999)
            if trip_drops > punthai_drop_limit:
                # ต้องเพิ่มขนาดรถเพื่อรองรับ drops
                if suggested == '4W' and trip_drops <= PUNTHAI_LIMITS['JB']['max_drops']:
                    suggested = 'JB'
                    source += " → JB (Drop Limit)"
                elif suggested == 'JB' or trip_drops > PUNTHAI_LIMITS['JB']['max_drops']:
                    # ถ้า Central ห้าม 6W → WARNING
                    if trip_region not in CENTRAL_REGIONS:
                        suggested = '6W'
                        source += " → 6W (Drop Limit)"
                    else:
                        source += " ⚠️ Drops เกิน!"
        
        # คำนวณ utilization
        max_util_threshold = buffer * 100  # 100% หรือ 110% ตาม BU
        if suggested in LIMITS:
            w_util = (total_w / LIMITS[suggested]['max_w']) * 100
            c_util = (total_c / LIMITS[suggested]['max_c']) * 100
            max_util = max(w_util, c_util)
            
            # ถ้าเกิน threshold ตาม BU ต้องเพิ่มขนาดรถ
            if max_util > max_util_threshold:
                if suggested == '4W' and min_max_size >= 2:
                    jb_util = max((total_w / LIMITS['JB']['max_w']), (total_c / LIMITS['JB']['max_c'])) * 100
                    if jb_util <= max_util_threshold:
                        suggested = 'JB'
                        source += " → JB"
                        w_util = (total_w / LIMITS['JB']['max_w']) * 100
                        c_util = (total_c / LIMITS['JB']['max_c']) * 100
                    elif min_max_size >= 3:
                        suggested = '6W'
                        source += " → 6W"
                        w_util = (total_w / LIMITS['6W']['max_w']) * 100
                        c_util = (total_c / LIMITS['6W']['max_c']) * 100
                elif suggested == 'JB' and min_max_size >= 3:
                    suggested = '6W'
                    source += " → 6W"
                    w_util = (total_w / LIMITS['6W']['max_w']) * 100
                    c_util = (total_c / LIMITS['6W']['max_c']) * 100
        else:
            w_util = c_util = 0
        
        # คำนวณระยะทางรวม
        total_distance = 0
        branch_coords = []
        for code in trip_codes:
            loc = location_map.get(str(code).upper(), {})
            if loc.get('lat') and loc.get('lon'):
                branch_coords.append((loc['lat'], loc['lon']))
        
        if branch_coords:
            # DC → สาขาแรก
            total_distance += haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, branch_coords[0][0], branch_coords[0][1])
            # สาขา → สาขา
            for i in range(len(branch_coords) - 1):
                total_distance += haversine_distance(branch_coords[i][0], branch_coords[i][1], branch_coords[i+1][0], branch_coords[i+1][1])
            # สาขาสุดท้าย → DC
            total_distance += haversine_distance(branch_coords[-1][0], branch_coords[-1][1], DC_WANG_NOI_LAT, DC_WANG_NOI_LON)
        
        summary_data.append({
            'Trip': trip_num,
            'Branches': len(trip_codes),
            'Weight': total_w,
            'Cube': total_c,
            'Truck': f"{suggested} {source}",
            'BU_Type': trip_type,
            'Buffer': buffer_label,
            'Weight_Use%': w_util,
            'Cube_Use%': c_util,
            'Total_Distance': round(total_distance, 1)
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    # ==========================================
    # Step 8: เพิ่มคอลัมน์เสริม
    # ==========================================
    # เพิ่มคอลัมน์รถ
    trip_truck_map = {}
    for _, row in summary_df.iterrows():
        trip_truck_map[row['Trip']] = row['Truck']
    df['Truck'] = df['Trip'].map(trip_truck_map)
    
    # เพิ่มคอลัมน์ Region
    df['Region'] = df['_region_name']
    
    # เพิ่มคอลัมน์ Province (ถ้ายังไม่มี)
    if 'Province' not in df.columns:
        df['Province'] = df['_province']
    
    # เพิ่มคอลัมน์ระยะทางจาก DC
    df['Distance_from_DC'] = df['_distance_from_dc'].round(1)
    
    # เพิ่มคอลัมน์เช็ครถ
    df['VehicleCheck'] = '✅ ใช้ได้'
    
    # ==========================================
    # Step 9: เรียงทริปใหม่ให้ทริปติดกัน (สำหรับ export)
    
    return df.reset_index(drop=True)

# ============================================================================
# 🎯 MAIN APPLICATION
# ============================================================================

def main():
    st.set_page_config(
        page_title="ระบบจัดเที่ยว",
        page_icon="🚚",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # 🔄 Sync ข้อมูลจาก Google Sheets ทุกครั้งที่โหลดเว็บ (ไม่แสดง error ถ้าไม่มี)
    try:
        synced_df = sync_branch_data_from_sheets()
        if synced_df is not None and not synced_df.empty:
            st.success(f"✅ โหลดข้อมูล Master สำเร็จ: {len(synced_df)} สาขา", icon="📊")
    except Exception as e:
        # ไม่แสดง error ถ้าเป็นแค่ไม่มี credentials
        pass
    
    # 🔄 Auto-refresh ทุกเที่ยงคืน (ล้างแคช)
    if AUTOREFRESH_AVAILABLE:
        now = datetime.now()
        # คำนวณเวลาถึงเที่ยงคืน (00:00:00)
        midnight = datetime.combine(now.date(), time(0, 0, 0))
        
        # ถ้ายังไม่ถึงเที่ยงคืน เอาเที่ยงคืนวันถัดไป
        if now < midnight:
            next_midnight = midnight
        else:
            from datetime import timedelta
            next_midnight = midnight + timedelta(days=1)
        
        # คำนวณเวลาที่เหลือ (วินาที)
        seconds_until_midnight = int((next_midnight - now).total_seconds())
        
        # Refresh ทุกเที่ยงคืน
        if seconds_until_midnight > 0:
            # เช็คในช่วง 5 นาทีก่อนเที่ยงคืน (หลัง 23:55)
            if seconds_until_midnight <= 300:  # 5 minutes
                st.info(f"🔄 ระบบจะ Refresh อัตโนมัติใน {seconds_until_midnight // 60} นาที")
                st_autorefresh(interval=seconds_until_midnight * 1000, key="midnight_refresh")
            else:
                # ตรวจสอบทุก 1 ชั่วโมง
                st_autorefresh(interval=3600000, limit=24, key="hourly_check")
    
    # Header
    col1, col2 = st.columns([3, 1])
    with col2:
        st.markdown("###  ")  # Spacing
        if st.button("🔄 Refresh"):
            st.rerun()
    
    # Sidebar - จัดการ Master Data
    st.sidebar.header("⚙️ ตั้งค่า")
    
    with st.sidebar.expander("📊 จัดการข้อมูล Master (ข้อมูลสาขา)", expanded=False):
        st.markdown("""
        🔄 **อัปโหลด Master Data ใหม่**
        - อัปโหลดไฟล์ Excel เพื่ออัปเดต branch_data.json
        - รองรับคอลัมน์: Plan Code, จังหวัด, อำเภอ, ตำบล, Route, Distance, Lat, Lon
        """)
        
        # อัปโหลด Master
        master_file = st.file_uploader(
            "อัปโหลด Master Data (Excel)",
            type=["xlsx", "xls"],
            help="อัปโหลดไฟล์ Excel ที่มีข้อมูลสาขาทั้งหมด",
            key="master_uploader"
        )
        
        if master_file:
            if st.button("⬆️ อัปเดต Master Data", type="primary"):
                try:
                    with st.spinner("⏳ กำลังอัปเดตข้อมูล..."):
                        # อ่านไฟล์ Excel
                        df_master = pd.read_excel(master_file)
                        
                        # หาคอลัมน์ Code
                        code_col = None
                        for col in ['Plan Code', 'Code', 'รหัสสาขา', 'สาขา']:
                            if col in df_master.columns:
                                code_col = col
                                break
                        
                        if code_col is None:
                            st.error("❌ ไม่พบคอลัมน์รหัสสาขา (Plan Code, Code, รหัสสาขา, สาขา)")
                        else:
                            # แปลงเป็น JSON
                            master_dict = {}
                            for _, row in df_master.iterrows():
                                code = str(row[code_col]).strip().upper()
                                if code and code != '':
                                    master_dict[code] = row.to_dict()
                            
                            # บันทึกเป็น JSON
                            json_file = 'branch_data.json'
                            with open(json_file, 'w', encoding='utf-8') as f:
                                json.dump(master_dict, f, ensure_ascii=False, indent=2)
                            
                            st.success(f"✅ อัปเดต Master Data สำเร็จ: **{len(master_dict):,}** สาขา")
                            st.info("🔄 กรุณา Refresh หน้าเพื่อโหลดข้อมูลใหม่")
                            
                except Exception as e:
                    st.error(f"❌ เกิดข้อผิดพลาด: {e}")
        
        # แสดงข้อมูล Master ปัจจุบัน
        if st.checkbox("👁️ ดูข้อมูล Master ปัจจุบัน"):
            json_file = 'branch_data.json'
            if os.path.exists(json_file):
                with open(json_file, 'r', encoding='utf-8') as f:
                    master_data = json.load(f)
                st.info(f"📋 ข้อมูล Master ทั้งหมด: **{len(master_data):,}** สาขา")
                
                # แสดง sample 5 สาขาแรก
                sample_codes = list(master_data.keys())[:5]
                st.markdown("**ตัวอย่าง 5 สาขาแรก:**")
                for code in sample_codes:
                    st.text(f"- {code}")
            else:
                st.warning("⚠️ ยังไม่มีไฟล์ branch_data.json")
    
    # โหลดโมเดล
    model_data = load_model()
    if not model_data:
        st.error("❌ ไม่สามารถโหลดโมเดลได้")
        st.stop()
    
    # Show Punthai learning stats
    if PUNTHAI_PATTERNS and 'stats' in PUNTHAI_PATTERNS and PUNTHAI_PATTERNS['stats']:
        stats = PUNTHAI_PATTERNS['stats']
        with st.expander("📊 สถิติที่เรียนรู้จากไฟล์ Punthai Maxmart", expanded=False):
            col_a, col_b, col_c = st.columns(3)
            with col_a:
                st.metric("เฉลี่ยสาขา/ทริป", f"{stats.get('avg_branches', 0):.1f}")
            with col_b:
                st.metric("ทริปจังหวัดเดียว", f"{stats.get('same_province_pct', 0):.1f}%")
            with col_c:
                total_trips = stats.get('same_province', 0) + stats.get('mixed_province', 0)
                st.metric("จำนวนทริปอ้างอิง", total_trips)
    
    st.markdown("---")
    
    # โหลดโมเดล
    model_data = load_model()
    
    if not model_data:
        st.error("❌ ไม่พบข้อมูลโมเดล กรุณาเทรนโมเดลก่อนใช้งาน")
        st.info("💡 รันคำสั่ง: `python test_model.py`")
        st.stop()
    
    # ==========================================
    # แสดงสถานะ Master Data และ Sync
    # ==========================================
    st.markdown("### 📊 ข้อมูล Master (ฐานข้อมูลสาขา)")
    
    json_file = 'branch_data.json'
    col_m1, col_m2, col_m3 = st.columns([2, 1, 1])
    
    with col_m1:
        if os.path.exists(json_file):
            # อ่านข้อมูล JSON และเวลา
            try:
                file_stat = os.stat(json_file)
                last_modified = datetime.fromtimestamp(file_stat.st_mtime)
                time_diff = datetime.now() - last_modified
                
                with open(json_file, 'r', encoding='utf-8') as f:
                    master_data = json.load(f)
                    master_branch_count = len(master_data)
                    has_dc = '8nvDC011' in master_data
                
                # แสดงสถานะ
                if time_diff.total_seconds() < 300:  # < 5 นาที
                    status_icon = "🟢"
                    status_text = "อัปเดตล่าสุด"
                elif time_diff.total_seconds() < 3600:  # < 1 ชม.
                    status_icon = "🟡"
                    status_text = "อัปเดตเมื่อสักครู่"
                else:
                    status_icon = "🔴"
                    status_text = "ควร Sync ใหม่"
                
                st.success(f"{status_icon} **{master_branch_count:,} สาขา** (รวม DC วังน้อย: {'✅' if has_dc else '❌'})")
                st.caption(f"📅 {status_text}: {last_modified.strftime('%Y-%m-%d %H:%M:%S')}")
                
            except Exception as e:
                st.warning(f"⚠️ อ่านไฟล์ JSON ไม่สำเร็จ: {e}")
        else:
            st.warning("⚠️ ไม่พบไฟล์ branch_data.json - กรุณา Sync ข้อมูลจาก Google Sheets")
    
    with col_m2:
        # ปุ่ม Sync จาก Google Sheets
        if st.button("🔄 Sync จาก Sheets", use_container_width=True, type="primary"):
            if not SHEETS_AVAILABLE or gc is None:
                st.error("❌ ไม่พบ credentials.json")
                st.info("💡 วางไฟล์ credentials.json ในโฟลเดอร์ app แล้ว Refresh")
            else:
                with st.spinner("⏳ กำลัง Sync จาก Google Sheets..."):
                    try:
                        df_synced = sync_branch_data_from_sheets()
                        if df_synced is not None and not df_synced.empty:
                            st.success(f"✅ Sync สำเร็จ: {len(df_synced)} สาขา")
                            st.rerun()
                        else:
                            st.error("❌ Sync ไม่สำเร็จ")
                    except Exception as e:
                        st.error(f"❌ เกิดข้อผิดพลาด: {e}")
    
    with col_m3:
        # ปุ่มดูข้อมูล Master
        if st.button("👁️ ดูข้อมูล", use_container_width=True):
            if os.path.exists(json_file):
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        master_data = json.load(f)
                    
                    st.markdown("---")
                    st.markdown("#### 📋 ข้อมูล Master สาขา")
                    
                    # แสดงสรุป
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("จำนวนสาขา", f"{len(master_data):,}")
                    with col2:
                        dc_count = sum(1 for code in master_data.keys() if 'DC' in code.upper())
                        st.metric("DC", f"{dc_count}")
                    with col3:
                        provinces = set()
                        for branch in master_data.values():
                            prov = branch.get('จังหวัด', branch.get('Province', ''))
                            if prov:
                                provinces.add(prov)
                        st.metric("จังหวัด", len(provinces))
                    
                    # แสดงตัวอย่าง 10 สาขาแรก
                    st.markdown("**ตัวอย่าง 10 สาขาแรก:**")
                    sample_codes = list(master_data.keys())[:10]
                    sample_data = []
                    for code in sample_codes:
                        branch = master_data[code]
                        sample_data.append({
                            'รหัส': code,
                            'สาขา': branch.get('สาขา', branch.get('Branch Name', '')),
                            'จังหวัด': branch.get('จังหวัด', branch.get('Province', '')),
                            'อำเภอ': branch.get('อำเภอ', branch.get('District', ''))
                        })
                    st.dataframe(pd.DataFrame(sample_data), use_container_width=True, hide_index=True)
                    st.markdown("---")
                    
                except Exception as e:
                    st.error(f"❌ ไม่สามารถแสดงข้อมูลได้: {e}")
            else:
                st.warning("⚠️ ไม่พบไฟล์ branch_data.json")
    
    st.markdown("---")
    
    # อัปโหลดไฟล์สำหรับวิเคราะห์
    st.markdown("### 📁 อัปโหลดไฟล์รายการออเดอร์")
    st.markdown("""
    📝 **ข้อมูลสาขา (Master Data)** ถูก Sync จาก **Google Sheets** → ใช้ปุ่ม 🔄 ด้านบนเพื่อ Sync  
    🔄 Auto Sync ทุกครั้งที่โหลดเว็บ - รวม DC วังน้อย และข้อมูลละ/ลอง
    """)
    
    uploaded_file = st.file_uploader(
        "เลือกไฟล์ Excel (.xlsx)", 
        type=['xlsx'],
        help="อัปโหลดไฟล์ Excel ที่มีรายการ Booking No, สาขา, น้ำหนัก, คิว ฯลฯ"
    )
    
    df = None
    
    if uploaded_file:
        # เก็บไฟล์ต้นฉบับไว้ใน session_state เพื่อใช้ตอน export
        uploaded_file_content = uploaded_file.read()
        st.session_state['original_file_content'] = uploaded_file_content
        
        with st.spinner("⏳ กำลังอ่านข้อมูล..."):
            df = load_excel(uploaded_file_content)
            df = process_dataframe(df)
    
    # ถ้ามีข้อมูลแล้ว แสดงผล
    if df is not None and 'Code' in df.columns:
        st.success(f"✅ อ่านข้อมูลสำเร็จ: **{len(df):,}** รายการ")
        
        # แสดงข้อมูลพื้นฐาน
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📍 จำนวนสาขา", f"{df['Code'].nunique():,}")
        with col2:
            st.metric("⚖️ น้ำหนักรวม", f"{df['Weight'].sum():,.0f} kg")
        with col3:
            st.metric("📦 คิวรวม", f"{df['Cube'].sum():.1f} m³")
        with col4:
            provinces = df['Province'].nunique() if 'Province' in df.columns else 0
            st.metric("🗺️ จังหวัด", f"{provinces}")
        
        # แสดงตัวอย่างข้อมูล
        with st.expander("🔍 ดูข้อมูลตัวอย่าง"):
            st.dataframe(df.head(10), use_container_width=True)
        
        # ==========================================
        # เติมข้อมูลพื้นที่จาก Master (ทำในหลังบ้าน)
        # ==========================================
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
            # สร้าง dict สำหรับค้นหาเร็ว
            master_lookup = {}
            for _, row in MASTER_DATA.iterrows():
                code = str(row['Plan Code']).strip().upper()
                master_lookup[code] = {
                    'province': row.get('จังหวัด', ''),
                    'district': row.get('อำเภอ', ''),
                    'subdistrict': row.get('ตำบล', ''),
                    'lat': row.get('ละติจูด', 0),
                    'lon': row.get('ลองติจูด', 0)
                }
            
            # เติมข้อมูลที่ขาด
            filled_count = 0
            for idx, row in df.iterrows():
                code = str(row['Code']).strip().upper()
                if code in master_lookup:
                    master_info = master_lookup[code]
                    # เติม Province ถ้าว่าง
                    if 'Province' not in df.columns or pd.isna(df.loc[idx, 'Province']) or df.loc[idx, 'Province'] == '' or df.loc[idx, 'Province'] == 'UNKNOWN':
                        if master_info['province']:
                            df.loc[idx, 'Province'] = master_info['province']
                            filled_count += 1
                    # เติม District ถ้าว่าง
                    if 'District' not in df.columns:
                        df['District'] = ''
                    if pd.isna(df.loc[idx, 'District']) or df.loc[idx, 'District'] == '':
                        if master_info['district']:
                            df.loc[idx, 'District'] = master_info['district']
                    # เติม Subdistrict ถ้าว่าง
                    if 'Subdistrict' not in df.columns:
                        df['Subdistrict'] = ''
                    if pd.isna(df.loc[idx, 'Subdistrict']) or df.loc[idx, 'Subdistrict'] == '':
                        if master_info['subdistrict']:
                            df.loc[idx, 'Subdistrict'] = master_info['subdistrict']
            
            if filled_count > 0:
                st.info(f"📍 เติมข้อมูลพื้นที่จาก Master แล้ว {filled_count} รายการ")
        
        # ตรวจสอบว่ายังมีข้อมูลที่ขาดหรือไม่ (แสดงแค่จำนวน)
        if 'Province' in df.columns:
            missing_count = len(df[(df['Province'].isna()) | (df['Province'] == '') | (df['Province'] == 'UNKNOWN')])
            if missing_count > 0:
                st.warning(f"⚠️ ยังมี {missing_count} สาขาที่ไม่พบข้อมูลพื้นที่ใน Master")
        
        st.markdown("---")
        
        # แท็บหลัก
        tab1, tab2 = st.tabs([
            "📦 จัดเที่ยว (ตามน้ำหนัก)", 
            "🗺️ จัดกลุ่มตามภาค"
        ])
            
        # ==========================================
        # แท็บ 1: จัดเที่ยว (ตามน้ำหนัก)
        # ==========================================
        with tab1:
            # เพิ่ม Region ถ้ายังไม่มี
            if 'Region' not in df.columns and 'Province' in df.columns:
                df['Region'] = df['Province'].apply(get_region_name)
            
            # ==========================================
            # ตัวเลือกการตั้งค่า
            # ==========================================
            st.markdown("#### ⚙️ ตั้งค่าการจัดทริป")
            
            # กรอก Buffer แยกตามประเภท
            col_buf1, col_buf2 = st.columns(2)
            
            with col_buf1:
                punthai_buffer = st.number_input(
                    "🅿️ Punthai Buffer %",
                    min_value=80,
                    max_value=120,
                    value=100,
                    step=5
                )
            
            with col_buf2:
                maxmart_buffer = st.number_input(
                    "🅼 Maxmart/ผสม Buffer %",
                    min_value=80,
                    max_value=150,
                    value=110,
                    step=5
                )
            
            # แปลงเป็น buffer value
            punthai_buffer_value = punthai_buffer / 100.0
            maxmart_buffer_value = maxmart_buffer / 100.0
            
            st.markdown("---")
            
            # ปุ่มจัดทริป
            if st.button("🚀 เริ่มจัดเที่ยว (Google OR-Tools)", type="primary", use_container_width=True):
                with st.spinner("⏳ กำลังประมวลผล..."):
                    # จัดเรียงตามภาค/จังหวัด/อำเภอ/ตำบล/Route (ในฟังก์ชัน predict_trips)
                            df_to_process = df.copy()
                            
                            # ใช้ OR-Tools สำหรับ optimization
                            if ORTOOLS_AVAILABLE:
                                try:
                                    st.info("🤖 กำลังใช้ Google OR-Tools optimization...")
                                    result_df, summary = predict_trips_ortools(
                                        df_to_process,
                                        buffer_punthai=punthai_buffer_value,
                                        buffer_maxmart=maxmart_buffer_value,
                                        master_data=MASTER_DATA if not MASTER_DATA.empty else None,
                                        max_trips=80,
                                        time_limit=50,
                                        restrictions=VEHICLE_RESTRICTIONS if VEHICLE_LOGIC_AVAILABLE else None
                                    )
                                    if len(summary) == 0:
                                        st.warning("⚠️ OR-Tools ไม่สามารถหาผลลัพธ์ได้")
                                except Exception as e:
                                    st.error(f"❌ Error with OR-Tools: {e}")
                            else:
                                st.error("❌ OR-Tools ไม่พร้อมใช้งาน - ติดตั้งด้วย: pip install ortools")
                                st.stop()
                            
                            # ตรวจสอบสาขาที่ไม่ได้จัดทริป (Trip = 0)
                            unassigned_count = len(result_df[result_df['Trip'] == 0])
                            if unassigned_count > 0:
                                st.warning(f"⚠️ มี {unassigned_count} สาขาที่ไม่ได้จัดทริป (Trip = 0)")
                            
                            # กรองเฉพาะสาขาที่จัดทริปแล้ว สำหรับการแสดงผล
                            assigned_df = result_df[result_df['Trip'] > 0].copy()
                            
                            st.balloons()
                            st.success(f"✅ **จัดทริปเสร็จสมบูรณ์!** รวม **{len(summary)}** ทริป ({len(assigned_df)} สาขา)")
                            
                            st.markdown("---")
                            
                            # สถิติโดยรวม
                            st.markdown("### 📊 สรุปผลการจัดทริป")
                            
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("🚚 จำนวนทริป", len(summary))
                            with col2:
                                st.metric("📍 จำนวนสาขา", len(assigned_df))
                            with col3:
                                avg_branches = len(assigned_df) / max(1, assigned_df['Trip'].nunique())
                                st.metric("📊 เฉลี่ยสาขา/ทริป", f"{avg_branches:.1f}")
                            with col4:
                                # เช็คว่า summary เป็น DataFrame หรือ dict
                                if isinstance(summary, pd.DataFrame):
                                    # เช็คว่ามีคอลัมน์ Cube_Use% หรือไม่
                                    if 'Cube_Use%' in summary.columns and len(summary) > 0:
                                        avg_util = summary['Cube_Use%'].mean()
                                    elif 'Cube' in summary.columns and 'Vehicle' in summary.columns:
                                        # คำนวณเอง
                                        vehicle_limits = {'4W': 5.0, 'JB': 12.0, '6W': 20.0}
                                        summary['Cube_Use%'] = summary.apply(
                                            lambda row: (row['Cube'] / vehicle_limits.get(row['Vehicle'], 20.0)) * 100 
                                            if row['Vehicle'] in vehicle_limits else 0,
                                            axis=1
                                        )
                                        avg_util = summary['Cube_Use%'].mean()
                                    else:
                                        avg_util = 0
                                else:
                                    # ถ้า summary เป็น dict ให้แปลงเป็น DataFrame
                                    summary = pd.DataFrame([summary]) if isinstance(summary, dict) else pd.DataFrame()
                                    avg_util = 0
                                st.metric("📈 การใช้รถเฉลี่ย", f"{avg_util:.0f}%")
                            
                            st.markdown("---")
                            
                            # ตารางสรุปแต่ละทริป
                            st.markdown("### 🚛 รายละเอียดแต่ละทริป")
                            
                            # เช็คว่า summary เป็น DataFrame
                            if not isinstance(summary, pd.DataFrame):
                                if isinstance(summary, dict):
                                    summary = pd.DataFrame([summary])
                                else:
                                    summary = pd.DataFrame()
                            
                            if not summary.empty:
                                # เตรียม format dict โดยเช็คว่ามีคอลัมน์หรือไม่
                                format_dict = {}
                                if 'Weight' in summary.columns:
                                    format_dict['Weight'] = '{:.2f}'
                                if 'Cube' in summary.columns:
                                    format_dict['Cube'] = '{:.2f}'
                                if 'Weight_Use%' in summary.columns:
                                    format_dict['Weight_Use%'] = '{:.1f}%'
                                if 'Cube_Use%' in summary.columns:
                                    format_dict['Cube_Use%'] = '{:.1f}%'
                                if 'Total_Distance' in summary.columns:
                                    format_dict['Total_Distance'] = '{:.1f} km'
                                
                                # เตรียม subset สำหรับ gradient
                                gradient_cols = [col for col in ['Weight_Use%', 'Cube_Use%'] if col in summary.columns]
                                
                                if format_dict and gradient_cols:
                                    st.dataframe(
                                        summary.style.format(format_dict).background_gradient(
                                            subset=gradient_cols,
                                            cmap='RdYlGn',
                                            vmin=0,
                                            vmax=100
                                        ),
                                        use_container_width=True,
                                        height=400
                                    )
                                else:
                                    # ถ้าไม่มีคอลัมน์ที่ต้องการ format แสดงแบบปกติ
                                    st.dataframe(summary, use_container_width=True, height=400)
                            else:
                                st.info("ไม่มีข้อมูลสรุปทริป")
                            
                            # ตารางรายละเอียดทั้งหมด (มีคอลัมน์รถและระยะทาง)
                            with st.expander("📋 ดูรายละเอียดรายสาขา (เรียงตามน้ำหนัก)"):
                                # จัดเรียงคอลัมน์ที่สำคัญ
                                display_cols = ['Trip', 'Code', 'Name']
                                if 'Province' in result_df.columns:
                                    display_cols.append('Province')
                                if 'Region' in result_df.columns:
                                    display_cols.append('Region')
                                
                                # เพิ่มคอลัมน์ที่มีอยู่จริง
                                for col in ['Max_Distance_in_Trip', 'Weight', 'Cube', 'Truck', 'VehicleCheck']:
                                    if col in result_df.columns:
                                        display_cols.append(col)
                                
                                # กรองคอลัมน์ที่มีอยู่จริง
                                display_cols = [col for col in display_cols if col in result_df.columns]
                                display_df = result_df[display_cols].copy()
                                
                                # ตั้งชื่อคอลัมน์ภาษาไทย
                                col_names = {'Trip': 'ทริป', 'Code': 'รหัส', 'Name': 'ชื่อสาขา', 'Province': 'จังหวัด', 
                                           'Region': 'ภาค', 'Max_Distance_in_Trip': 'ระยะทาง Max(km)', 
                                           'Weight': 'น้ำหนัก(kg)', 'Cube': 'คิว(m³)', 'Truck': 'รถ', 'VehicleCheck': 'ตรวจสอบรถ'}
                                display_df.columns = [col_names.get(c, c) for c in display_cols]
                                
                                # จัดรูปแบบคอลัมน์ระยะทาง
                                st.dataframe(
                                    display_df.style.format({
                                        'ระยะทาง(km)': '{:.1f}',
                                        'น้ำหนัก(kg)': '{:.2f}',
                                        'คิว(m³)': '{:.2f}'
                                    }),
                                    use_container_width=True, 
                                    height=400
                                )
                            
                            # แสดงสาขาที่มีคำเตือน (ถ้ามีคอลัมน์ VehicleCheck)
                            if 'VehicleCheck' in result_df.columns:
                                warning_branches = result_df[result_df['VehicleCheck'].str.contains('⚠️', na=False)]
                                if len(warning_branches) > 0:
                                    with st.expander(f"⚠️ สาขาที่ใช้รถต่างจากปกติ ({len(warning_branches)} สาขา)"):
                                        st.warning("สาขาเหล่านี้ปกติใช้รถประเภทอื่น แต่ถูกจัดให้ใช้รถประเภทที่ต่างออกไป")
                                        display_cols_warn = ['Trip', 'Code', 'Name', 'Truck', 'VehicleCheck']
                                        # กรองคอลัมน์ที่มีอยู่จริง
                                        display_cols_warn = [col for col in display_cols_warn if col in warning_branches.columns]
                                        display_warn_df = warning_branches[display_cols_warn].copy()
                                        col_names_warn = {'Trip': 'ทริป', 'Code': 'รหัส', 'Name': 'ชื่อสาขา', 'Truck': 'รถที่จัด', 'VehicleCheck': 'ประวัติการใช้รถ'}
                                        display_warn_df.columns = [col_names_warn.get(c, c) for c in display_cols_warn]
                                        st.dataframe(display_warn_df, use_container_width=True)
                            
                            st.markdown("---")
                            
                            # ดาวน์โหลด - เขียนทับชีต 2.Punthai ในไฟล์ต้นฉบับ พร้อมสลับสีเหลืองโทนส้ม-ขาว
                            from openpyxl import load_workbook
                            from openpyxl.styles import PatternFill, Font, Border, Side
                            
                            output = io.BytesIO()
                            
                            # สร้าง location_map จาก MASTER_DATA
                            location_map = {}
                            if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                                for _, row in MASTER_DATA.iterrows():
                                    code = str(row.get('Plan Code', '')).strip().upper()
                                    if code:
                                        location_map[code] = {
                                            'ตำบล': row.get('ตำบล', ''),
                                            'อำเภอ': row.get('อำเภอ', ''),
                                            'จังหวัด': row.get('จังหวัด', ''),
                                            'Route': row.get('Reference', '')
                                        }
                            
                            # สร้าง Trip_No map
                            trip_no_map = {}
                            vehicle_counts = {'4W': 0, '4WJ': 0, '6W': 0}
                            
                            # เรียง trip ตาม Zone Order + Province Max Dist + District Max Dist (เหมือนตอนจัดทริป)
                            ZONE_ORDER_EXPORT = {'NORTH': 1, 'NE': 2, 'SOUTH': 3, 'EAST': 4, 'WEST': 5, 'CENTRAL': 6}
                            trip_sort_keys = {}
                            
                            for trip_num in result_df['Trip'].unique():
                                if trip_num == 0:
                                    continue
                                trip_data = result_df[result_df['Trip'] == trip_num]
                                
                                # หา Region Order
                                region = trip_data['Region'].iloc[0] if 'Region' in trip_data.columns else 'ไม่ระบุ'
                                region_order = ZONE_ORDER_EXPORT.get(region, 99)
                                
                                # หา Province Max Distance และ District Max Distance
                                prov_max_dist = 0
                                dist_max_dist = 0
                                
                                for code in trip_data['Code'].unique():
                                    loc = location_map.get(str(code).upper(), {})
                                    # ดึงระยะทางจาก MASTER_DATA
                                    if not MASTER_DATA.empty:
                                        master_row = MASTER_DATA[MASTER_DATA['Plan Code'].astype(str).str.upper() == str(code).upper()]
                                        if len(master_row) > 0:
                                            dist_km = master_row.iloc[0].get('Distance from DC (km)', 0)
                                            if pd.notna(dist_km):
                                                prov_max_dist = max(prov_max_dist, float(dist_km))
                                                dist_max_dist = max(dist_max_dist, float(dist_km))
                                
                                # Sort key: Region Order (Asc), Prov Max Dist (Desc), Dist Max Dist (Desc)
                                # ใช้ค่าลบเพื่อให้ sort Desc
                                trip_sort_keys[trip_num] = (region_order, -prov_max_dist, -dist_max_dist)
                            
                            # Sort: Zone Order → Province Max Dist (ไกลก่อน) → District Max Dist (ไกลก่อน)
                            sorted_trips = sorted(
                                [t for t in result_df['Trip'].unique() if t != 0],
                                key=lambda t: trip_sort_keys.get(t, (99, 0, 0))
                            )
                            
                            for trip_num in sorted_trips:
                                trip_summary = summary[summary['Trip'] == trip_num]
                                if len(trip_summary) > 0:
                                    truck_info = trip_summary.iloc[0]['Truck']
                                    vehicle_type = truck_info.split()[0] if truck_info else '6W'
                                    # JB ใช้ prefix 4WJ
                                    if vehicle_type == 'JB':
                                        vehicle_type = '4WJ'
                                    vehicle_counts[vehicle_type] = vehicle_counts.get(vehicle_type, 0) + 1
                                    trip_no = f"{vehicle_type}{vehicle_counts[vehicle_type]:03d}"
                                    trip_no_map[trip_num] = trip_no
                            
                            try:
                                # โหลด workbook ต้นฉบับ
                                wb = load_workbook(io.BytesIO(st.session_state.get('original_file_content', b'')))
                                
                                # หาชีตเป้าหมาย (2.Punthai)
                                target_sheet = None
                                for sheet_name in wb.sheetnames:
                                    if 'punthai' in sheet_name.lower() or '2.' in sheet_name.lower():
                                        target_sheet = sheet_name
                                        break
                                
                                if not target_sheet:
                                    target_sheet = '2.Punthai'
                                    if target_sheet not in wb.sheetnames:
                                        wb.create_sheet(target_sheet)
                                
                                ws = wb[target_sheet]
                                
                                # หา header row
                                header_row = 1
                                for row_idx in range(1, min(5, ws.max_row + 1)):
                                    for col_idx in range(1, min(15, ws.max_column + 1)):
                                        cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                                        if 'รหัสสาขา' in cell_val or 'Trip' in cell_val.upper():
                                            header_row = row_idx
                                            break
                                
                                # ลบข้อมูลเก่า
                                if ws.max_row > header_row:
                                    ws.delete_rows(header_row + 1, ws.max_row - header_row)
                                
                                # เขียน header ใหม่
                                new_headers = ['Sep.', 'BU', 'รหัสสาขา', 'รหัส WMS', 'สาขา', 'ตำบล', 'อำเภอ', 'จังหวัด', 'Route',
                                              'Total Cube', 'Total Wgt', 'Original QTY', 'Trip', 'Trip no']
                                for col_idx, header_val in enumerate(new_headers, 1):
                                    ws.cell(row=header_row, column=col_idx, value=header_val)
                                
                                # สีเหลืองโทนส้ม-ขาว (สลับ 2 สี)
                                yellow_orange = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                                white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                                thin_border = Border(
                                    left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin')
                                )
                                red_font = Font(color='FF0000', bold=True)
                                
                                # หาทริปที่ไม่ผ่านเกณฑ์
                                failed_trips = set()
                                vehicle_limits = {'4W': {'max_w': 2500, 'max_c': 5.0}, 'JB': {'max_w': 3500, 'max_c': 7.0}, '6W': {'max_w': 6000, 'max_c': 20.0}}
                                for t in result_df['Trip'].unique():
                                    if t == 0:
                                        continue
                                    trip_data = result_df[result_df['Trip'] == t]
                                    trip_cube = trip_data['Cube'].sum()
                                    trip_weight = trip_data['Weight'].sum()
                                    trip_no = trip_no_map.get(t, '6W001')
                                    veh_type = 'JB' if trip_no.startswith('4WJ') else ('4W' if trip_no.startswith('4W') else '6W')
                                    limits = vehicle_limits.get(veh_type, vehicle_limits['6W'])
                                    max_util = max((trip_cube / limits['max_c']) * 100, (trip_weight / limits['max_w']) * 100)
                                    if max_util > 105 or max_util < 50:
                                        failed_trips.add(t)
                                
                                # เขียนข้อมูล
                                current_trip = None
                                use_yellow = True
                                row_num = header_row + 1
                                sep_num = 1
                                
                                for trip_num in sorted_trips:
                                    trip_data = result_df[result_df['Trip'] == trip_num].copy()
                                    
                                    # Sort ตาม ตำบล → อำเภอ → จังหวัด
                                    trip_data['_sort_sub'] = trip_data['Code'].apply(lambda c: location_map.get(str(c).upper(), {}).get('ตำบล', ''))
                                    trip_data['_sort_dist'] = trip_data['Code'].apply(lambda c: location_map.get(str(c).upper(), {}).get('อำเภอ', ''))
                                    trip_data['_sort_prov'] = trip_data['Code'].apply(lambda c: location_map.get(str(c).upper(), {}).get('จังหวัด', ''))
                                    trip_data = trip_data.sort_values(['_sort_prov', '_sort_dist', '_sort_sub', 'Code'])
                                    
                                    trip_no = trip_no_map.get(trip_num, '')
                                    
                                    # สลับสีเมื่อเปลี่ยนทริป
                                    if current_trip != trip_num:
                                        current_trip = trip_num
                                        use_yellow = not use_yellow
                                    
                                    fill = yellow_orange if use_yellow else white_fill
                                    
                                    for _, row in trip_data.iterrows():
                                        branch_code = row.get('Code', '')
                                        loc = location_map.get(str(branch_code).upper(), {})
                                        
                                        data = [
                                            sep_num,
                                            row.get('BU', 211),
                                            branch_code,
                                            branch_code,
                                            row.get('Name', ''),
                                            loc.get('ตำบล', ''),
                                            loc.get('อำเภอ', ''),
                                            loc.get('จังหวัด', ''),
                                            loc.get('Route', ''),
                                            round(row.get('Cube', 0), 2) if pd.notna(row.get('Cube')) else 0,
                                            round(row.get('Weight', 0), 2) if pd.notna(row.get('Weight')) else 0,
                                            row.get('OriginalQty', 0) if pd.notna(row.get('OriginalQty')) else 0,
                                            int(trip_num),
                                            trip_no,
                                        ]
                                        
                                        for col_idx, value in enumerate(data, 1):
                                            cell = ws.cell(row=row_num, column=col_idx, value=value)
                                            cell.fill = fill
                                            cell.border = thin_border
                                            if trip_num in failed_trips:
                                                cell.font = red_font
                                        
                                        row_num += 1
                                        sep_num += 1
                                
                                wb.save(output)
                                output.seek(0)
                                
                            except Exception as e:
                                st.warning(f"⚠️ ไม่สามารถเขียนทับไฟล์ต้นฉบับได้: {e} - ใช้รูปแบบมาตรฐานแทน")
                                # Fallback: สร้างไฟล์ใหม่ด้วย xlsxwriter
                                output = io.BytesIO()
                                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                    export_df = result_df.copy()
                                    export_df['Trip_No'] = export_df['Trip'].map(lambda x: trip_no_map.get(x, ''))
                                    export_df.to_excel(writer, sheet_name='รายละเอียดทริป', index=False)
                                    summary.to_excel(writer, sheet_name='สรุปทริป', index=False)
                            
                            col1, col2, col3 = st.columns([1, 2, 1])
                            with col2:
                                st.download_button(
                                    label="📥 ดาวน์โหลดผลลัพธ์ (Excel)",
                                    data=output.getvalue(),
                            file_name=f"ผลจัดทริป_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
        
        # ==========================================
        # แท็บ 2: จัดกลุ่มสาขาตามภาค (ไม่สนน้ำหนัก)
        # ==========================================
        with tab2:
            df_region = df.copy()
            
            # จัดกลุ่มตามภาค
            branch_info = model_data.get('branch_info', {})
            trip_pairs = model_data.get('trip_pairs', set())
            
            # สร้างข้อมูลภาคสำหรับแต่ละสาขา (จากไฟล์ประวัติ)
            region_groups = {
                'ภาคกลาง-กรุงเทพชั้นใน': ['กรุงเทพมหานคร'],
                'ภาคกลาง-กรุงเทพชั้นกลาง': ['กรุงเทพมหานคร'],
                'ภาคกลาง-กรุงเทพชั้นนอก': ['กรุงเทพมหานคร'],
                'ภาคกลาง-ปริมณฑล': ['นครปฐม', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ', 'สมุทรสาคร'],
                'ภาคกลาง-กลางตอนบน': ['ชัยนาท', 'พระนครศรีอยุธยา', 'ลพบุรี', 'สระบุรี', 'สิงห์บุรี', 'อ่างทอง', 'อยุธยา'],
                'ภาคกลาง-กลางตอนล่าง': ['สมุทรสงคราม', 'สุพรรณบุรี'],
                        'ภาคตะวันตก': ['กาญจนบุรี', 'ประจวบคีรีขันธ์', 'ราชบุรี', 'เพชรบุรี'],
                        'ภาคตะวันออก': ['จันทบุรี', 'ชลบุรี', 'ตราด', 'นครนายก', 'ปราจีนบุรี', 'ระยอง', 'สระแก้ว', 'ฉะเชิงเทรา'],
                        'ภาคอีสาน-อีสานเหนือ': ['นครพนม', 'บึงกาฬ', 'มุกดาหาร', 'สกลนคร', 'หนองคาย', 'หนองบัวลำภู', 'อุดรธานี', 'เลย'],
                        'ภาคอีสาน-อีสานกลาง': ['กาฬสินธุ์', 'ขอนแก่น', 'ชัยภูมิ', 'มหาสารคาม', 'ร้อยเอ็ด'],
                        'ภาคอีสาน-อีสานใต้': ['นครราชสีมา', 'โคราช', 'บุรีรัมย์', 'ยโสธร', 'ศรีสะเกษ', 'สุรินทร์', 'อำนาจเจริญ', 'อุบลราชธานี'],
                        'ภาคเหนือ-เหนือตอนบน': ['น่าน', 'พะเยา', 'ลำปาง', 'ลำพูน', 'เชียงราย', 'เชียงใหม่', 'แพร่', 'แม่ฮ่องสอน'],
                        'ภาคเหนือ-เหนือตอนล่าง': ['กำแพงเพชร', 'ตาก', 'นครสวรรค์', 'พิจิตร', 'พิษณุโลก', 'สุโขทัย', 'อุตรดิตถ์', 'อุทัยธานี', 'เพชรบูรณ์'],
                        'ภาคใต้-ใต้ฝั่งอันดามัน': ['กระบี่', 'ตรัง', 'พังงา', 'ภูเก็ต', 'ระนอง', 'สตูล'],
                        'ภาคใต้-ใต้ฝั่งอ่าวไทย': ['ชุมพร', 'นครศรีธรรมราช', 'พัทลุง', 'ยะลา', 'สงขลา', 'สุราษฎร์ธานี', 'ปัตตานี', 'นราธิวาส']
            }
  
            
            # เพิ่มคอลัมน์ภาค - ดึงจังหวัดจาก Master ถ้าไม่มี
            if 'Province' not in df_region.columns or df_region['Province'].isna().any():
                # ดึงจังหวัดจาก Master
                if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                    province_map = {}
                    for _, row in MASTER_DATA.iterrows():
                        code = row.get('Plan Code', '')
                        province = row.get('จังหวัด', '')
                        if code and province:
                            province_map[code] = province
                    
                    # ใส่จังหวัดให้แต่ละสาขา
                    if 'Province' not in df_region.columns:
                        df_region['Province'] = df_region['Code'].map(province_map)
                    else:
                        # เติมเฉพาะที่เป็น NaN
                        df_region['Province'] = df_region.apply(
                            lambda row: province_map.get(row['Code'], row.get('Province', 'UNKNOWN')) 
                            if pd.isna(row.get('Province')) else row['Province'],
                            axis=1
                        )
                
                # Define region mapping
                REGION_PROVINCES = {
                    'NORTH': ['เชียงใหม่', 'เชียงราย', 'ลำปาง', 'ลำพูน', 'แม่ฮ่องสอน', 'น่าน', 'พะเยา', 'แพร่', 'อุตรดิตถ์'],
                    'NORTHEAST': ['ขอนแก่น', 'อุดรธานี', 'นครราชสีมา', 'อุบลราชธานี', 'สกลนคร', 'ร้อยเอ็ด', 'มหาสารคาม', 'กาฬสินธุ์', 'เลย', 'หนองคาย', 'หนองบัวลำภู', 'ชัยภูมิ', 'ยโสธร', 'มุกดาหาร', 'นครพนม', 'ศรีสะเกษ', 'สุรินทร์', 'บุรีรัมย์', 'อำนาจเจริญ', 'บึงกาฬ'],
                    'CENTRAL': ['กรุงเทพมหานคร', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ', 'พระนครศรีอยุธยา', 'อ่างทอง', 'ลพบุรี', 'สิงห์บุรี', 'ชัยนาท', 'สระบุรี', 'ฉะเชิงเทรา', 'นครนายก', 'ปราจีนบุรี', 'สมุทรสาคร', 'สมุทรสงคราม', 'นครปฐม', 'กาญจนบุรี', 'ราชบุรี', 'สุพรรณบุรี', 'เพชรบุรี', 'ประจวบคีรีขันธ์'],
                    'EAST': ['ชลบุรี', 'ระยอง', 'จันทบุรี', 'ตราด', 'สระแก้ว'],
                    'SOUTH': ['สุราษฎร์ธานี', 'นครศรีธรรมราช', 'ภูเก็ต', 'กระบี่', 'พังงา', 'ระนอง', 'ชุมพร', 'สงขลา', 'ตรัง', 'พัทลุง', 'ปัตตานี', 'ยะลา', 'นราธิวาส', 'สตูล']
                }
                
                # Map province to region
                def get_region_from_province(prov):
                    for region, provinces in REGION_PROVINCES.items():
                        if prov in provinces:
                            return region
                    return 'OTHER'
                
                df_region['Region'] = df_region['Province'].apply(get_region_from_province)
                
                # หากลุ่มสาขา (ตามความใกล้ทางภูมิศาสตร์)
                def find_paired_branches(code, code_province, df_data):
                    paired = set()
                    
                    # หาสาขาที่อยู่ใกล้กัน
                    code_rows = df_data[df_data['Code'] == code]
                    if len(code_rows) == 0:
                        return paired
                    
                    # หาสาขาอื่นในจังหวัดเดียวกัน
                    same_province = df_data[df_data['Province'] == code_province]
                    for _, other_row in same_province.iterrows():
                        other_code = other_row['Code']
                        if other_code != code:
                            paired.add(other_code)
                    
                    return paired
                    
                    all_codes_set = set(df_region['Code'].unique())
                    
                    # สร้างกลุ่มสาขาแบบ Union-Find (ตามลำดับ: ตำบล → อำเภอ → จังหวัด)
                    # Step 1: เริ่มจากแต่ละสาขาเป็นกลุ่มๆ พร้อมข้อมูล Master
                    initial_groups = {}
                    for code in all_codes_set:
                        # ดึงข้อมูลจาก Master
                        location = {}
                        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                            if len(master_row) > 0:
                                master_row = master_row.iloc[0]
                                location = {
                                    'subdistrict': master_row.get('ตำบล', ''),
                                    'district': master_row.get('อำเภอ', ''),
                                    'province': master_row.get('จังหวัด', 'UNKNOWN'),
                                    'lat': master_row.get('ละติจูด', 0),
                                    'lon': master_row.get('ลองติจูด', 0)
                                }
                        
                        # ถ้าไม่มีใน Master ลองดึงจากไฟล์อัปโหลด
                        if not location or location.get('province', 'UNKNOWN') == 'UNKNOWN':
                            c_row = df_region[df_region['Code'] == code].iloc[0] if len(df_region[df_region['Code'] == code]) > 0 else None
                            if c_row is not None:
                                location = {
                                    'subdistrict': '',
                                    'district': '',
                                    'province': c_row.get('Province', 'UNKNOWN'),
                                    'lat': 0,
                                    'lon': 0
                                }
                        
                        if location:
                            initial_groups[(code,)] = {code: location}
                    
                    # ใช้ initial_groups แทน booking_groups
                    booking_groups = initial_groups
                    
                    # Step 2: รวมกลุ่มตามลำดับ ตำบล → อำเภอ → จังหวัด
                    def groups_can_merge(locs1, locs2):
                        """เช็คว่า 2 กลุ่มควรรวมกันไหม (ตามลำดับความละเอียด)"""
                        # 1. เช็คตำบลเดียวกัน (ต้องมีข้อมูลตำบล)
                        subdistricts1 = set(loc.get('subdistrict', '') for loc in locs1.values() if loc.get('subdistrict', ''))
                        subdistricts2 = set(loc.get('subdistrict', '') for loc in locs2.values() if loc.get('subdistrict', ''))
                        if subdistricts1 and subdistricts2 and (subdistricts1 & subdistricts2):
                            return True, 'ตำบล'
                        
                        # 2. เช็คอำเภอเดียวกัน (ต้องมีข้อมูลอำเภอและจังหวัดเดียวกัน)
                        districts1 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs1.values() if loc.get('district', '')}
                        districts2 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs2.values() if loc.get('district', '')}
                        if districts1 and districts2:
                            # เช็คว่ามีอำเภอและจังหวัดตรงกัน
                            for d1, p1 in districts1:
                                for d2, p2 in districts2:
                                    if d1 == d2 and p1 == p2 and p1:
                                        return True, 'อำเภอ'
                        
                        # 3. เช็คจังหวัดเดียวกัน
                        provinces1 = set(loc.get('province', '') for loc in locs1.values() if loc.get('province', ''))
                        provinces2 = set(loc.get('province', '') for loc in locs2.values() if loc.get('province', ''))
                        if provinces1 & provinces2:
                            return True, 'จังหวัด'
                        
                        return False, None
                    
                    merged_groups = []
                    used_groups = set()
                    
                    for group1, locs1 in booking_groups.items():
                        if group1 in used_groups:
                            continue
                        
                        merged_codes = set(group1)
                        merged_locs = locs1.copy()
                        used_groups.add(group1)
                        
                        # หากลุ่มอื่นที่ใกล้เคียง
                        changed = True
                        while changed:
                            changed = False
                            for group2, locs2 in booking_groups.items():
                                if group2 in used_groups:
                                    continue
                                can_merge, level = groups_can_merge(merged_locs, locs2)
                                if can_merge:
                                    merged_codes |= set(group2)
                                    merged_locs.update(locs2)
                                    used_groups.add(group2)
                                    changed = True
                        
                        merged_groups.append({
                            'codes': merged_codes,
                            'locations': merged_locs
                        })
                    
                    # Step 3: แปลงเป็น groups format
                    groups = []
                    for mg in merged_groups:
                        rep_code = list(mg['codes'])[0]
                        rep_row = df_region[df_region['Code'] == rep_code].iloc[0]
                        # กรองเฉพาะจังหวัดที่ไม่ใช่ UNKNOWN และไม่เป็น NaN
                        provinces = set(
                            str(loc.get('province', '')).strip() 
                            for loc in mg['locations'].values() 
                            if loc.get('province') and str(loc.get('province', '')).strip() not in ['UNKNOWN', 'nan', '']
                        )
                        
                        # ถ้าไม่มีจังหวัดเลย ใส่ "ไม่ระบุ"
                        province_str = ', '.join(sorted(provinces)) if provinces else 'ไม่ระบุ'
                        
                        groups.append({
                            'codes': mg['codes'],
                            'region': rep_row.get('Region', 'ไม่ระบุ'),
                            'province': province_str
                        })
                    
                    # แสดงสถิติ
                    st.markdown("---")
                    st.markdown("### 📊 สรุปการจัดกลุ่ม")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("📍 จำนวนสาขา", df_region['Code'].nunique())
                    with col2:
                        st.metric("🗂️ จำนวนกลุ่ม", len(groups))
                    with col3:
                        regions_count = df_region['Region'].nunique()
                        st.metric("🗺️ จำนวนภาค", regions_count)
                    
                    # แสดงตามภาค
                    st.markdown("---")
                    st.markdown("### 🗺️ สาขาแยกตามภาค")
                    
                    region_summary = df_region.groupby('Region').agg({
                        'Code': 'nunique',
                        'Weight': 'sum',
                        'Cube': 'sum'
                    }).reset_index()
                    region_summary.columns = ['ภาค', 'จำนวนสาขา', 'น้ำหนักรวม', 'คิวรวม']
                    st.dataframe(region_summary, width='stretch')
                    
                    # แสดงรายละเอียดแต่ละภาค
                    for region in sorted(df_region['Region'].unique()):
                        region_data = df_region[df_region['Region'] == region]
                        with st.expander(f"📍 {region} ({region_data['Code'].nunique()} สาขา)"):
                            display_cols = ['Code', 'Name', 'Province', 'Weight', 'Cube']
                            display_cols = [c for c in display_cols if c in region_data.columns]
                            
                            region_display = region_data[display_cols].drop_duplicates('Code')
                            col_names = {'Code': 'รหัส', 'Name': 'ชื่อสาขา', 'Province': 'จังหวัด', 'Weight': 'น้ำหนัก', 'Cube': 'คิว'}
                            region_display.columns = [col_names.get(c, c) for c in display_cols]
                            st.dataframe(region_display, use_container_width=True)
                    
                    # แสดงกลุ่มสาขาที่เคยไปด้วยกัน
                    st.markdown("---")
                    st.markdown("### 🔗 กลุ่มสาขาที่เคยไปด้วยกัน (จากประวัติ)")
                    
                    paired_groups = [g for g in groups if len(g['codes']) > 1]
                    if paired_groups:
                        for i, group in enumerate(paired_groups, 1):
                            codes_list = list(group['codes'])
                            names = []
                            for c in codes_list:
                                name_row = df_region[df_region['Code'] == c]
                                if len(name_row) > 0 and 'Name' in name_row.columns:
                                    names.append(f"{c} ({name_row['Name'].iloc[0]})")
                                else:
                                    names.append(c)
                            
                            st.write(f"**กลุ่ม {i}** - {group['region']}: {', '.join(names)}")
                    else:
                        st.info("ไม่พบกลุ่มสาขาที่เคยไปด้วยกันในรายการนี้")
                    
                    # ดาวน์โหลด
                    st.markdown("---")
                    output_region = io.BytesIO()
                    with pd.ExcelWriter(output_region, engine='xlsxwriter') as writer:
                        df_region.to_excel(writer, sheet_name='สาขาทั้งหมด', index=False)
                        region_summary.to_excel(writer, sheet_name='สรุปตามภาค', index=False)
                    
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        st.download_button(
                            label="📥 ดาวน์โหลดข้อมูลจัดกลุ่ม (Excel)",
                            data=output_region.getvalue(),
                            file_name=f"จัดกลุ่มสาขา_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

if __name__ == "__main__":
    main()
