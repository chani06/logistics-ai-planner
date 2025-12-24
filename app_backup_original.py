"""
Logistics Planner 
"""

import streamlit as st
import pandas as pd
import numpy as np
import pickle
import os
import glob
from datetime import datetime, time
import io
from math import radians, sin, cos, sqrt, atan2

# Auto-refresh component
try:
    from streamlit_autorefresh import st_autorefresh
    AUTOREFRESH_AVAILABLE = True
except ImportError:
    AUTOREFRESH_AVAILABLE = False
    st.warning("ΓÜá∩╕Å α╕òα╕┤α╕öα╕òα╕▒α╣ëα╕ç streamlit-autorefresh: pip install streamlit-autorefresh")

# ==========================================
# CONFIG
# ==========================================
MODEL_PATH = 'models/decision_tree_model.pkl'

# α╕éα╕╡α╕öα╕êα╕│α╕üα╕▒α╕öα╕úα╕ûα╣üα╕òα╣êα╕Ñα╕░α╕¢α╕úα╕░α╣Çα╕áα╕ù (α╕íα╕▓α╕òα╕úα╕Éα╕▓α╕Ö)
LIMITS = {
    '4W': {'max_w': 2500, 'max_c': 5.0, 'max_drops': 12},   # α╣äα╕íα╣êα╣Çα╕üα╕┤α╕Ö 12 α╕êα╕╕α╕ö, Cube Γëñ 5
    'JB': {'max_w': 3500, 'max_c': 7.0, 'max_drops': 12},   # α╣äα╕íα╣êα╣Çα╕üα╕┤α╕Ö 12 α╕êα╕╕α╕ö, Cube Γëñ 7
    '6W': {'max_w': 6000, 'max_c': 20.0, 'max_drops': 999}  # α╣äα╕íα╣êα╕êα╕│α╕üα╕▒α╕öα╕êα╕╕α╕ö, Cube α╕òα╣ëα╕¡α╕çα╣Çα╕òα╣çα╕í, Weight Γëñ 6000
}

# ≡ƒöÆ α╕éα╕╡α╕öα╕êα╕│α╕üα╕▒α╕öα╕¬α╕│α╕½α╕úα╕▒α╕Ü Punthai α╕Ñα╣ëα╕ºα╕Ö (α╕½α╣ëα╕▓α╕íα╣Çα╕üα╕┤α╕Ö 100%)
PUNTHAI_LIMITS = {
    '4W': {'max_w': 2500, 'max_c': 5.0, 'max_drops': 5},   # Punthai α╕Ñα╣ëα╕ºα╕Ö 4W: α╕¬α╕╣α╕çα╕¬α╕╕α╕ö 5 α╕¬α╕▓α╕éα╕▓
    'JB': {'max_w': 3500, 'max_c': 7.0, 'max_drops': 10},  # Punthai α╕Ñα╣ëα╕ºα╕Ö JB: α╕¬α╕╣α╕çα╕¬α╕╕α╕ö 10 α╕¬α╕▓α╕éα╕▓
    '6W': {'max_w': 6000, 'max_c': 20.0, 'max_drops': 999}
}

# ≡ƒÄ» Minimum utilization α╕òα╣êα╕¡α╕¢α╕úα╕░α╣Çα╕áα╕ùα╕úα╕û (α╕¬α╕│α╕½α╕úα╕▒α╕Ü balancing)
MIN_UTIL = {
    '4W': 70,   # 4W α╕òα╣ëα╕¡α╕çα╣âα╕èα╣ëα╕¡α╕óα╣êα╕▓α╕çα╕Öα╣ëα╕¡α╕ó 70%
    'JB': 80,   # JB α╕òα╣ëα╕¡α╕çα╣âα╕èα╣ëα╕¡α╕óα╣êα╕▓α╕çα╕Öα╣ëα╕¡α╕ó 80%
    '6W': 90    # 6W α╕òα╣ëα╕¡α╕çα╣âα╕èα╣ëα╕¡α╕óα╣êα╕▓α╕çα╕Öα╣ëα╕¡α╕ó 90%
}

# Buffer α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕üα╕▓α╕úα╣âα╕èα╣ëα╕úα╕û (α╕òα╕▓α╕í BU)
BUFFER = 1.0  # Default buffer
PUNTHAI_BUFFER = 1.0  # ≡ƒà┐∩╕Å Punthai α╕Ñα╣ëα╕ºα╕Ö: α╕½α╣ëα╕▓α╕íα╣Çα╕üα╕┤α╕Ö 100%
MAXMART_BUFFER = 1.10  # ≡ƒà╝ Maxmart/α╕£α╕¬α╕í: α╣Çα╕üα╕┤α╕Öα╣äα╕öα╣ë 10%

# α╕êα╕│α╕Öα╕ºα╕Öα╕¬α╕▓α╕éα╕▓α╕òα╣êα╕¡α╕ùα╕úα╕┤α╕¢ - α╣âα╕èα╣ëα╕üα╕▒α╕Ü 4W/JB α╣Çα╕ùα╣êα╕▓α╕Öα╕▒α╣ëα╕Ö (6W α╣äα╕íα╣êα╕êα╕│α╕üα╕▒α╕ö)
MAX_BRANCHES_PER_TRIP = 12  # α╕¬α╕╣α╕çα╕¬α╕╕α╕ö 12 α╕¬α╕▓α╕éα╕▓α╕òα╣êα╕¡α╕ùα╕úα╕┤α╕¢α╕¬α╕│α╕½α╕úα╕▒α╕Ü 4W/JB (6W α╣äα╕íα╣êα╕êα╕│α╕üα╕▒α╕ö)

# Performance Config
MAX_DETOUR_KM = 12  # α╕Ñα╕öα╕êα╕▓α╕ü 15km α╣Çα╕¢α╣çα╕Ö 12km α╣Çα╕₧α╕╖α╣êα╕¡α╕¢α╕úα╕░α╕íα╕ºα╕Ñα╕£α╕Ñα╣Çα╕úα╣çα╕ºα╕éα╕╢α╣ëα╕Ö
MAX_MERGE_ITERATIONS = 25  # α╕êα╕│α╕üα╕▒α╕öα╕úα╕¡α╕Üα╕üα╕▓α╕úα╕úα╕ºα╕íα╕ùα╕úα╕┤α╕¢ (α╕Ñα╕öα╕êα╕▓α╕ü 50 α╣Çα╕₧α╕╖α╣êα╕¡α╣Çα╕úα╣çα╕ºα╕éα╕╢α╣ëα╕Ö)

# ==========================================
# REGION ORDER CONFIG (Far-to-Near Sorting)
# ==========================================
# α╕Ñα╕│α╕öα╕▒α╕Üα╕üα╕▓α╕úα╕êα╕▒α╕ö: α╣Çα╕½α╕Öα╕╖α╕¡ ΓåÆ α╕¡α╕╡α╕¬α╕▓α╕Ö ΓåÆ α╣âα╕òα╣ë ΓåÆ α╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü ΓåÆ α╕üα╕Ñα╕▓α╕ç
REGION_ORDER = {
    'α╣Çα╕½α╕Öα╕╖α╕¡': 1, 'NORTH': 1,
    'α╕¡α╕╡α╕¬α╕▓α╕Ö': 2, 'NE': 2,
    'α╣âα╕òα╣ë': 3, 'SOUTH': 3,
    'α╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü': 4, 'EAST': 4,
    'α╕òα╕░α╕ºα╕▒α╕Öα╕òα╕ü': 5, 'WEST': 5,
    'α╕üα╕Ñα╕▓α╕ç': 6, 'CENTRAL': 6,
    'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕': 99
}

# α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç: α╕½α╣ëα╕▓α╕íα╣âα╕èα╣ë 6W (α╣Çα╕ëα╕₧α╕▓α╕░ 4W, JB)
CENTRAL_REGIONS = ['α╕üα╕Ñα╕▓α╕ç', 'CENTRAL']
CENTRAL_ALLOWED_VEHICLES = ['4W', 'JB']  # NO 6W in Central

# α╕úα╕▓α╕óα╕üα╕▓α╕úα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣äα╕íα╣êα╕òα╣ëα╕¡α╕çα╕üα╕▓α╕úα╕êα╕▒α╕öα╕¬α╣êα╕ç (α╕òα╕▒α╕öα╕¡α╕¡α╕ü)
EXCLUDE_BRANCHES = ['DC011', 'PTDC', 'PTG DISTRIBUTION CENTER']

# α╕úα╕▓α╕óα╕èα╕╖α╣êα╕¡α╕ùα╕╡α╣êα╕òα╣ëα╕¡α╕çα╕òα╕▒α╕öα╕¡α╕¡α╕ü (α╣âα╕èα╣ëα╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕èα╕╖α╣êα╕¡)
EXCLUDE_NAMES = ['Distribution Center', 'PTG Distribution', 'α╕Ü.α╕₧α╕╡α╕ùα╕╡α╕êα╕╡ α╣Çα╕¡α╣çα╕Öα╣Çα╕Öα╕¡α╕óα╕╡']

# α╕₧α╕┤α╕üα╕▒α╕ö DC α╕ºα╕▒α╕çα╕Öα╣ëα╕¡α╕ó (α╕êα╕╕α╕öα╕üα╕Ñα╕▓α╕ç)
DC_WANG_NOI_LAT = 14.179394
DC_WANG_NOI_LON = 100.648149

# α╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕ùα╕╡α╣êα╕òα╣ëα╕¡α╕çα╣âα╕èα╣ëα╕úα╕û 6W (α╕üα╕í.)
DISTANCE_REQUIRE_6W = 100  # α╕ûα╣ëα╕▓α╕½α╣êα╕▓α╕çα╕êα╕▓α╕ü DC α╣Çα╕üα╕┤α╕Ö 100 α╕üα╕í. α╕òα╣ëα╕¡α╕çα╣âα╕èα╣ë 6W

# ==========================================
# ZONE/REGION CONFIG - α╕úα╕½α╕▒α╕¬α╕áα╕▓α╕äα╣üα╕Ñα╕░α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö
# ==========================================
# α╕úα╕½α╕▒α╕¬α╕áα╕▓α╕ä: 1=α╕üα╕Ñα╕▓α╕ç, 2=α╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü, 3=α╕òα╕░α╕ºα╕▒α╕Öα╕òα╕ü, 4=α╣Çα╕½α╕Öα╕╖α╕¡, 5=α╕¡α╕╡α╕¬α╕▓α╕Ö, 6=α╣âα╕òα╣ë
REGION_CODE = {
    # α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç (α╕úα╕½α╕▒α╕¬ 1)
    'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕íα╕½α╕▓α╕Öα╕äα╕ú': '10', 'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕»': '10',
    'α╕Öα╕Öα╕ùα╕Üα╕╕α╕úα╕╡': '11',
    'α╕¢α╕ùα╕╕α╕íα╕ÿα╕▓α╕Öα╕╡': '12',
    'α╕₧α╕úα╕░α╕Öα╕äα╕úα╕¿α╕úα╕╡α╕¡α╕óα╕╕α╕ÿα╕óα╕▓': '13', 'α╕¡α╕óα╕╕α╕ÿα╕óα╕▓': '13',
    'α╕¬α╕úα╕░α╕Üα╕╕α╕úα╕╡': '14',
    'α╕Ñα╕₧α╕Üα╕╕α╕úα╕╡': '15',
    'α╕¬α╕┤α╕çα╕½α╣îα╕Üα╕╕α╕úα╕╡': '16',
    'α╕¡α╣êα╕▓α╕çα╕ùα╕¡α╕ç': '17',
    'α╕èα╕▒α╕óα╕Öα╕▓α╕ù': '18',
    'α╕Öα╕äα╕úα╕¢α╕Éα╕í': '19',
    'α╕¬α╕íα╕╕α╕ùα╕úα╕¢α╕úα╕▓α╕üα╕▓α╕ú': '1A',
    'α╕¬α╕íα╕╕α╕ùα╕úα╕¬α╕▓α╕äα╕ú': '1B',
    'α╕¬α╕íα╕╕α╕ùα╕úα╕¬α╕çα╕äα╕úα╕▓α╕í': '1C',
    
    # α╕áα╕▓α╕äα╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü (α╕úα╕½α╕▒α╕¬ 2)
    'α╕èα╕Ñα╕Üα╕╕α╕úα╕╡': '20',
    'α╕úα╕░α╕óα╕¡α╕ç': '21',
    'α╕êα╕▒α╕Öα╕ùα╕Üα╕╕α╕úα╕╡': '22',
    'α╕òα╕úα╕▓α╕ö': '23',
    'α╕ëα╕░α╣Çα╕èα╕┤α╕çα╣Çα╕ùα╕úα╕▓': '24',
    'α╕¢α╕úα╕▓α╕êα╕╡α╕Öα╕Üα╕╕α╕úα╕╡': '25',
    'α╕¬α╕úα╕░α╣üα╕üα╣ëα╕º': '26',
    'α╕Öα╕äα╕úα╕Öα╕▓α╕óα╕ü': '27',
    
    # α╕áα╕▓α╕äα╕òα╕░α╕ºα╕▒α╕Öα╕òα╕ü (α╕úα╕½α╕▒α╕¬ 3)
    'α╕úα╕▓α╕èα╕Üα╕╕α╕úα╕╡': '30',
    'α╕üα╕▓α╕ìα╕êα╕Öα╕Üα╕╕α╕úα╕╡': '31',
    'α╕¬α╕╕α╕₧α╕úα╕úα╕ôα╕Üα╕╕α╕úα╕╡': '32',
    'α╣Çα╕₧α╕èα╕úα╕Üα╕╕α╕úα╕╡': '33',
    'α╕¢α╕úα╕░α╕êα╕ºα╕Üα╕äα╕╡α╕úα╕╡α╕éα╕▒α╕Öα╕ÿα╣î': '34',
    
    # α╕áα╕▓α╕äα╣Çα╕½α╕Öα╕╖α╕¡ (α╕úα╕½α╕▒α╕¬ 4) - α╣äα╕üα╕Ñ α╣âα╕èα╣ë 6W α╣Çα╕¢α╣çα╕Öα╕½α╕Ñα╕▒α╕ü
    'α╕Öα╕äα╕úα╕¬α╕ºα╕úα╕úα╕äα╣î': '40',
    'α╕¡α╕╕α╕ùα╕▒α╕óα╕ÿα╕▓α╕Öα╕╡': '41',
    'α╕üα╕│α╣üα╕₧α╕çα╣Çα╕₧α╕èα╕ú': '42',
    'α╕òα╕▓α╕ü': '43',
    'α╕¬α╕╕α╣éα╕éα╕ùα╕▒α╕ó': '44',
    'α╕₧α╕┤α╕⌐α╕ôα╕╕α╣éα╕Ñα╕ü': '45',
    'α╕₧α╕┤α╕êα╕┤α╕òα╕ú': '46',
    'α╣Çα╕₧α╕èα╕úα╕Üα╕╣α╕úα╕ôα╣î': '47',
    'α╕¡α╕╕α╕òα╕úα╕öα╕┤α╕òα╕ûα╣î': '48',
    'α╣üα╕₧α╕úα╣ê': '49',
    'α╕Öα╣êα╕▓α╕Ö': '4A',
    'α╕₧α╕░α╣Çα╕óα╕▓': '4B',
    'α╣Çα╕èα╕╡α╕óα╕çα╕úα╕▓α╕ó': '4C',
    'α╣Çα╕èα╕╡α╕óα╕çα╣âα╕½α╕íα╣ê': '4D',
    'α╣üα╕íα╣êα╕«α╣êα╕¡α╕çα╕¬α╕¡α╕Ö': '4E',
    'α╕Ñα╕│α╕₧α╕╣α╕Ö': '4F',
    'α╕Ñα╕│α╕¢α╕▓α╕ç': '4G',
    
    # α╕áα╕▓α╕äα╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕üα╣Çα╕ëα╕╡α╕óα╕çα╣Çα╕½α╕Öα╕╖α╕¡/α╕¡α╕╡α╕¬α╕▓α╕Ö (α╕úα╕½α╕▒α╕¬ 5)
    'α╕Öα╕äα╕úα╕úα╕▓α╕èα╕¬α╕╡α╕íα╕▓': '50', 'α╣éα╕äα╕úα╕▓α╕è': '50',
    'α╕Üα╕╕α╕úα╕╡α╕úα╕▒α╕íα╕óα╣î': '51',
    'α╕¬α╕╕α╕úα╕┤α╕Öα╕ùα╕úα╣î': '52',
    'α╕¿α╕úα╕╡α╕¬α╕░α╣Çα╕üα╕⌐': '53',
    'α╕¡α╕╕α╕Üα╕Ñα╕úα╕▓α╕èα╕ÿα╕▓α╕Öα╕╡': '54',
    'α╕óα╣éα╕¬α╕ÿα╕ú': '55',
    'α╕èα╕▒α╕óα╕áα╕╣α╕íα╕┤': '56',
    'α╕¡α╕│α╕Öα╕▓α╕êα╣Çα╕êα╕úα╕┤α╕ì': '57',
    'α╕½α╕Öα╕¡α╕çα╕Üα╕▒α╕ºα╕Ñα╕│α╕áα╕╣': '58',
    'α╕éα╕¡α╕Öα╣üα╕üα╣êα╕Ö': '59',
    'α╕¡α╕╕α╕öα╕úα╕ÿα╕▓α╕Öα╕╡': '5A',
    'α╣Çα╕Ñα╕ó': '5B',
    'α╕½α╕Öα╕¡α╕çα╕äα╕▓α╕ó': '5C',
    'α╕íα╕½α╕▓α╕¬α╕▓α╕úα╕äα╕▓α╕í': '5D',
    'α╕úα╣ëα╕¡α╕óα╣Çα╕¡α╣çα╕ö': '5E',
    'α╕üα╕▓α╕¼α╕¬α╕┤α╕Öα╕ÿα╕╕α╣î': '5F',
    'α╕¬α╕üα╕Ñα╕Öα╕äα╕ú': '5G',
    'α╕Öα╕äα╕úα╕₧α╕Öα╕í': '5H',
    'α╕íα╕╕α╕üα╕öα╕▓α╕½α╕▓α╕ú': '5I',
    'α╕Üα╕╢α╕çα╕üα╕▓α╕¼': '5J',
    
    # α╕áα╕▓α╕äα╣âα╕òα╣ë (α╕úα╕½α╕▒α╕¬ 6) - α╣äα╕üα╕Ñα╕íα╕▓α╕ü α╣âα╕èα╣ë 6W
    'α╕èα╕╕α╕íα╕₧α╕ú': '60',
    'α╕úα╕░α╕Öα╕¡α╕ç': '61',
    'α╕¬α╕╕α╕úα╕▓α╕⌐α╕Äα╕úα╣îα╕ÿα╕▓α╕Öα╕╡': '62',
    'α╕₧α╕▒α╕çα╕çα╕▓': '63',
    'α╕üα╕úα╕░α╕Üα╕╡α╣ê': '64',
    'α╕áα╕╣α╣Çα╕üα╣çα╕ò': '65',
    'α╕Öα╕äα╕úα╕¿α╕úα╕╡α╕ÿα╕úα╕úα╕íα╕úα╕▓α╕è': '66',
    'α╕òα╕úα╕▒α╕ç': '67',
    'α╕₧α╕▒α╕ùα╕Ñα╕╕α╕ç': '68',
    'α╕¬α╕çα╕éα╕Ñα╕▓': '69',
    'α╕¬α╕òα╕╣α╕Ñ': '6A',
    'α╕¢α╕▒α╕òα╕òα╕▓α╕Öα╕╡': '6B',
    'α╕óα╕░α╕Ñα╕▓': '6C',
    'α╕Öα╕úα╕▓α╕ÿα╕┤α╕ºα╕▓α╕¬': '6D',
}

# α╕áα╕▓α╕äα╕ùα╕╡α╣êα╕òα╣ëα╕¡α╕çα╣âα╕èα╣ë 6W α╣Çα╕¢α╣çα╕Öα╕½α╕Ñα╕▒α╕ü (α╣äα╕üα╕Ñα╕êα╕▓α╕ü DC)
REGIONS_REQUIRE_6W = ['4', '5', '6']  # α╣Çα╕½α╕Öα╕╖α╕¡, α╕¡α╕╡α╕¬α╕▓α╕Ö, α╣âα╕òα╣ë

# α╕èα╕╖α╣êα╕¡α╕áα╕▓α╕ä
REGION_NAMES = {
    '1': 'α╕üα╕Ñα╕▓α╕ç',
    '2': 'α╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü',
    '3': 'α╕òα╕░α╕ºα╕▒α╕Öα╕òα╕ü',
    '4': 'α╣Çα╕½α╕Öα╕╖α╕¡',
    '5': 'α╕¡α╕╡α╕¬α╕▓α╕Ö',
    '6': 'α╣âα╕òα╣ë',
    '9': 'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕'
}

# ==========================================
# HELPER: ZONE/REGION FUNCTIONS
# ==========================================
def get_region_code(province):
    """α╕öα╕╢α╕çα╕úα╕½α╕▒α╕¬α╕áα╕▓α╕ä/α╣éα╕ïα╕Öα╕êα╕▓α╕üα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö"""
    if not province or str(province).strip() == '' or str(province) == 'nan':
        return '99'  # α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕
    province = str(province).strip()
    return REGION_CODE.get(province, '99')

def get_region_name(province):
    """α╕öα╕╢α╕çα╕èα╕╖α╣êα╕¡α╕áα╕▓α╕äα╕êα╕▓α╕üα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö"""
    code = get_region_code(province)
    if code == '99':
        return 'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕'
    region_prefix = code[0]
    return REGION_NAMES.get(region_prefix, 'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕')

def get_recommended_vehicle_by_region(province, distance_from_dc=None):
    """α╣üα╕Öα╕░α╕Öα╕│α╕úα╕ûα╕òα╕▓α╕íα╕áα╕▓α╕ä/α╕úα╕░α╕óα╕░α╕ùα╕▓α╕ç"""
    code = get_region_code(province)
    region_prefix = code[0] if code != '99' else '9'
    
    # α╕áα╕▓α╕äα╣Çα╕½α╕Öα╕╖α╕¡, α╕¡α╕╡α╕¬α╕▓α╕Ö, α╣âα╕òα╣ë ΓåÆ α╣âα╕èα╣ë 6W
    if region_prefix in REGIONS_REQUIRE_6W:
        return '6W'
    
    # α╕ûα╣ëα╕▓α╕íα╕╡α╕úα╕░α╕óα╕░α╕ùα╕▓α╕ç α╣üα╕Ñα╕░α╣Çα╕üα╕┤α╕Ö threshold ΓåÆ α╣âα╕èα╣ë 6W
    if distance_from_dc and distance_from_dc > DISTANCE_REQUIRE_6W:
        return '6W'
    
    # α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç, α╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü, α╕òα╕░α╕ºα╕▒α╕Öα╕òα╕ü ΓåÆ α╣âα╕èα╣ë 4W/JB α╣äα╕öα╣ë
    return 'JB'  # default α╣Çα╕¢α╣çα╕Ö JB

def sort_branches_by_region_route(branches_df, master_data=None):
    """
    α╕êα╕▒α╕öα╣Çα╕úα╕╡α╕óα╕çα╕¬α╕▓α╕éα╕▓α╕òα╕▓α╕íα╕áα╕▓α╕ä ΓåÆ α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö ΓåÆ α╕¡α╕│α╣Çα╕áα╕¡ ΓåÆ α╕òα╕│α╕Üα╕Ñ ΓåÆ Route
    α╣Çα╕₧α╕╖α╣êα╕¡α╣âα╕½α╣ëα╕ùα╕úα╕┤α╕¢α╣Çα╕úα╕╡α╕óα╕çα╕òα╕┤α╕öα╕üα╕▒α╕Öα╣äα╕íα╣êα╕üα╕úα╕░α╣éα╕öα╕ö
    """
    if branches_df.empty:
        return branches_df
    
    df = branches_df.copy()
    
    # α╣Çα╕₧α╕┤α╣êα╕íα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕¬α╕│α╕½α╕úα╕▒α╕Ü sort
    df['_region_code'] = df['Province'].apply(get_region_code) if 'Province' in df.columns else '99'
    df['_province'] = df['Province'].fillna('') if 'Province' in df.columns else ''
    df['_district'] = df['District'].fillna('') if 'District' in df.columns else ''
    df['_subdistrict'] = df['Subdistrict'].fillna('') if 'Subdistrict' in df.columns else ''
    
    # α╣üα╕óα╕ü Route number
    if 'Route' in df.columns:
        df['_route_num'] = df['Route'].apply(lambda x: int(str(x).replace('CD', '')) if pd.notna(x) and str(x).startswith('CD') else 99999)
    else:
        df['_route_num'] = 99999
    
    # Sort
    df = df.sort_values(by=['_region_code', '_province', '_district', '_subdistrict', '_route_num'])
    
    # α╕Ñα╕Üα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕èα╕▒α╣êα╕ºα╕äα╕úα╕▓α╕º
    df = df.drop(columns=['_region_code', '_province', '_district', '_subdistrict', '_route_num'])
    
    return df.reset_index(drop=True)

def check_trip_route_spread(trip_df):
    """
    α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕ºα╣êα╕▓α╕ùα╕úα╕┤α╕¢α╕íα╕╡ Route α╕üα╕úα╕░α╕êα╕▓α╕óα╕íα╕▓α╕üα╣äα╕½α╕í
    α╕äα╕╖α╕Öα╕äα╣êα╕▓: (route_range, is_spread, provinces)
    """
    if trip_df.empty or 'Route' not in trip_df.columns:
        return 0, False, []
    
    routes = trip_df['Route'].dropna().unique()
    route_nums = []
    for r in routes:
        if pd.notna(r) and str(r).startswith('CD'):
            try:
                route_nums.append(int(str(r).replace('CD', '')))
            except:
                pass
    
    if len(route_nums) < 2:
        return 0, False, trip_df['Province'].dropna().unique().tolist() if 'Province' in trip_df.columns else []
    
    route_range = max(route_nums) - min(route_nums)
    is_spread = route_range > 4000  # α╕ûα╣ëα╕▓α╕íα╕▓α╕üα╕üα╕ºα╣êα╕▓ 4000 α╕ûα╕╖α╕¡α╕ºα╣êα╕▓α╕üα╕úα╕░α╕êα╕▓α╕ó
    
    provinces = trip_df['Province'].dropna().unique().tolist() if 'Province' in trip_df.columns else []
    
    return route_range, is_spread, provinces

def validate_trip_vehicle(trip_df, assigned_vehicle):
    """
    α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕ºα╣êα╕▓α╕úα╕ûα╕ùα╕╡α╣êα╕êα╕▒α╕öα╣âα╕½α╣ëα╕ùα╕úα╕┤α╕¢α╣Çα╕½α╕íα╕▓α╕░α╕¬α╕íα╕üα╕▒α╕Üα╕áα╕▓α╕ä/α╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê
    α╕äα╕╖α╕Öα╕äα╣êα╕▓: (is_valid, recommended_vehicle, reason)
    """
    if trip_df.empty:
        return True, assigned_vehicle, ''
    
    provinces = trip_df['Province'].dropna().unique() if 'Province' in trip_df.columns else []
    
    # α╕½α╕▓α╕áα╕▓α╕äα╕ùα╕╡α╣êα╣äα╕üα╕Ñα╕ùα╕╡α╣êα╕¬α╕╕α╕öα╣âα╕Öα╕ùα╕úα╕┤α╕¢
    farthest_region = '1'  # default α╕üα╕Ñα╕▓α╕ç
    for prov in provinces:
        code = get_region_code(prov)
        region = code[0] if code != '99' else '1'
        if region > farthest_region:
            farthest_region = region
    
    # α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Ü
    if farthest_region in REGIONS_REQUIRE_6W:
        # α╕áα╕▓α╕äα╣äα╕üα╕Ñ α╕äα╕ºα╕úα╣âα╕èα╣ë 6W
        if assigned_vehicle in ['4W', 'JB']:
            return False, '6W', f'α╕áα╕▓α╕ä{REGION_NAMES.get(farthest_region, "α╣äα╕üα╕Ñ")} α╕äα╕ºα╕úα╣âα╕èα╣ë 6W'
    
    return True, assigned_vehicle, ''

# ==========================================
# LOAD MASTER DATA
# ==========================================
@st.cache_data(ttl=7200)  # Cache 2 α╕èα╕▒α╣êα╕ºα╣éα╕íα╕ç (α╣Çα╕úα╣çα╕ºα╕éα╕╢α╣ëα╕Ö)
def load_master_data():
    """α╣éα╕½α╕Ñα╕öα╣äα╕ƒα╕Ñα╣î Master α╕¬α╕ûα╕▓α╕Öα╕ùα╕╡α╣êα╕¬α╣êα╕ç (Optimized)"""
    try:
        # α╣éα╕½α╕Ñα╕öα╣Çα╕ëα╕₧α╕▓α╕░α╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕ùα╕╡α╣êα╕êα╕│α╣Çα╕¢α╣çα╕Ö
        usecols = ['Plan Code', 'α╕òα╕│α╕Üα╕Ñ', 'α╕¡α╕│α╣Çα╕áα╕¡', 'α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', 'α╕Ñα╕░α╕òα╕┤α╕êα╕╣α╕ö', 'α╕Ñα╕¡α╕çα╕òα╕┤α╕êα╕╣α╕ö']
        # α╕Ñα╕¡α╕çα╕½α╕▓α╣äα╕ƒα╕Ñα╣îα╕ùα╕╡α╣êα╕íα╕╡α╕¡α╕óα╕╣α╣êα╕êα╕úα╕┤α╕ç
        possible_files = ['Dc/α╕¬α╕ûα╕▓α╕Öα╕ùα╕╡α╣êα╕¬α╣êα╕ç.xlsx', 'Dc/Master α╕¬α╕ûα╕▓α╕Öα╕ùα╕╡α╣êα╕¬α╣êα╕ç.xlsx']
        df_master = pd.DataFrame()
        for file_path in possible_files:
            try:
                df_master = pd.read_excel(file_path, usecols=usecols)
                break
            except:
                continue
        if df_master.empty:
            return pd.DataFrame()
        # α╕ùα╕│α╕äα╕ºα╕▓α╕íα╕¬α╕░α╕¡α╕▓α╕ö Plan Code (vectorized)
        if 'Plan Code' in df_master.columns:
            df_master['Plan Code'] = df_master['Plan Code'].astype(str).str.strip().str.upper()
        # α╕¬α╕úα╣ëα╕▓α╕ç dict α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕äα╣ëα╕Öα╕½α╕▓α╣Çα╕úα╣çα╕º
        df_master = df_master[df_master['Plan Code'] != '']
        return df_master
    except FileNotFoundError:
        return pd.DataFrame()
    except Exception as e:
        try:
            st.warning(f"α╣äα╕íα╣êα╕¬α╕▓α╕íα╕▓α╕úα╕ûα╣éα╕½α╕Ñα╕öα╣äα╕ƒα╕Ñα╣î Master: {e} (α╕êα╕░α╣âα╕èα╣ëα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕êα╕▓α╕üα╣äα╕ƒα╕Ñα╣îα╕¡α╕▒α╕¢α╣éα╕½α╕Ñα╕öα╣üα╕ùα╕Ö)")
        except:
            pass
        return pd.DataFrame()

# α╣éα╕½α╕Ñα╕ö Master Data
MASTER_DATA = load_master_data()

# ==========================================
# CLEAN NAME FUNCTION (α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕ùα╕│ Join_Key)
# ==========================================
def clean_name(text):
    """
    α╕ùα╕│α╕äα╕ºα╕▓α╕íα╕¬α╕░α╕¡α╕▓α╕öα╕èα╕╖α╣êα╕¡: α╕Ñα╕Ü prefix α╕ê./α╕¡./α╕ò. α╣üα╕Ñα╕░ trim whitespace
    α╣âα╕èα╣ëα╕¬α╕│α╕½α╕úα╕▒α╕Üα╕¬α╕úα╣ëα╕▓α╕ç Join_Key α╣Çα╕₧α╕╖α╣êα╕¡α╣Çα╕ùα╕╡α╕óα╕Üα╕üα╕▒α╕Ü Master Data
    """
    if pd.isna(text) or text is None:
        return ''
    text = str(text)
    # α╕Ñα╕Ü prefix α╕áα╕▓α╕⌐α╕▓α╣äα╕ùα╕ó
    text = text.replace('α╕ê. ', '').replace('α╕ê.', '')
    text = text.replace('α╕¡. ', '').replace('α╕¡.', '')
    text = text.replace('α╕ò. ', '').replace('α╕ò.', '')
    # α╕Ñα╕Ü prefix α╕áα╕▓α╕⌐α╕▓α╕¡α╕▒α╕çα╕üα╕ñα╕⌐ (α╕ûα╣ëα╕▓α╕íα╕╡)
    text = text.replace('Tambon ', '').replace('Amphoe ', '').replace('Changwat ', '')
    return text.strip()

def normalize_province_name(province):
    """
    α╣üα╕¢α╕Ñα╕çα╕èα╕╖α╣êα╕¡α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣âα╕½α╣ëα╣Çα╕¢α╣çα╕Öα╕íα╕▓α╕òα╕úα╕Éα╕▓α╕Ö (α╣üα╕üα╣ëα╕¢α╕▒α╕ìα╕½α╕▓α╕èα╕╖α╣êα╕¡α╣Çα╕₧α╕╡α╣ëα╕óα╕Ö)
    """
    if pd.isna(province) or province is None:
        return ''
    province = clean_name(province)
    # Mapping α╕èα╕╖α╣êα╕¡α╕ùα╕╡α╣êα╕₧α╕Üα╕Üα╣êα╕¡α╕ó
    province_mapping = {
        'α╕₧α╕úα╕░α╕Öα╕äα╕úα╕¿α╕úα╕╡α╕¡α╕óα╕╕α╕ÿα╕óα╕▓': 'α╕¡α╕óα╕╕α╕ÿα╕óα╕▓',
        'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕»': 'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕íα╕½α╕▓α╕Öα╕äα╕ú',
        'α╕üα╕ùα╕í': 'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕íα╕½α╕▓α╕Öα╕äα╕ú',
        'α╕üα╕ùα╕í.': 'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕íα╕½α╕▓α╕Öα╕äα╕ú',
        'α╣éα╕äα╕úα╕▓α╕è': 'α╕Öα╕äα╕úα╕úα╕▓α╕èα╕¬α╕╡α╕íα╕▓',
    }
    return province_mapping.get(province, province)

def load_master_dist_data():
    """
    α╣éα╕½α╕Ñα╕öα╣äα╕ƒα╕Ñα╣î Master Dist.xlsx α╕¬α╕│α╕½α╕úα╕▒α╕Ü:
    1. α╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕úα╕░α╕öα╕▒α╕Üα╕òα╕│α╕Üα╕Ñ
    2. Sum_Code (Sort_Code) α╕¬α╕│α╕½α╕úα╕▒α╕Üα╣Çα╕úα╕╡α╕óα╕çα╕Ñα╕│α╕öα╕▒α╕Üα╕òα╕▓α╕íα╕áα╕╣α╕íα╕┤α╕¿α╕▓α╕¬α╕òα╕úα╣î
    
    α╕½α╕Ñα╕▒α╕üα╕üα╕▓α╕ú: α╣âα╕èα╣ë Join_Key (α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö_α╕¡α╕│α╣Çα╕áα╕¡_α╕òα╕│α╕Üα╕Ñ) α╣Çα╕¢α╣çα╕Öα╕òα╕▒α╕ºα╣Çα╕èα╕╖α╣êα╕¡α╕í
    α╣Çα╕₧α╕╖α╣êα╕¡α╕öα╕╢α╕ç Sum_Code α╕íα╕▓α╣âα╕èα╣ëα╣âα╕Öα╕üα╕▓α╕ú Sort
    """
    try:
        file_path = 'Dc/Master Dist.xlsx'
        df = pd.read_excel(file_path)
        
        # α╕¬α╕úα╣ëα╕▓α╕ç lookup dict - α╕¬α╕¡α╕ç key: Sum_Code α╣üα╕Ñα╕░ Join_Key (α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö_α╕¡α╕│α╣Çα╕áα╕¡_α╕òα╕│α╕Üα╕Ñ)
        dist_lookup = {}   # key = Sum_Code
        name_lookup = {}   # key = Join_Key (α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö_α╕¡α╕│α╣Çα╕áα╕¡_α╕òα╕│α╕Üα╕Ñ)
        
        for _, row in df.iterrows():
            sum_code = str(row.get('Sum_Code', '')).strip()
            
            # α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕¬α╕│α╕äα╕▒α╕ì: α╣Çα╕₧α╕┤α╣êα╕í sum_code (Sort_Code) α╣Çα╕éα╣ëα╕▓α╣äα╕¢α╕öα╣ëα╕ºα╕ó!
            data = {
                'sum_code': sum_code,  # ≡ƒöæ α╕üα╕╕α╕ìα╣üα╕êα╕¬α╕│α╕äα╕▒α╕ìα╕¬α╕│α╕½α╕úα╕▒α╕Ü Sort!
                'region': row.get('Region', ''),
                'region_code': row.get('Region_Code', ''),
                'province': row.get('Province', ''),
                'prov_code': row.get('Prov_Code', ''),
                'district': row.get('District', ''),
                'dist_code': row.get('Dist_Code', ''),
                'subdistrict': row.get('Subdistrict', ''),
                'subdist_code': row.get('Subdist_Code', ''),
                'dist_from_dc_km': float(row.get('Dist_from_DC_km', 9999)) if pd.notna(row.get('Dist_from_DC_km')) else 9999,
                'prov_dist_km': float(row.get('Prov_Dist_km', 0)) if pd.notna(row.get('Prov_Dist_km')) else 0,
                'dist_subdist_km': float(row.get('Dist_Subdist_km', 0)) if pd.notna(row.get('Dist_Subdist_km')) else 0,
            }
            
            # Key 1: Sum_Code (α╕¬α╕│α╕½α╕úα╕▒α╕Ü lookup α╣éα╕öα╕óα╕òα╕úα╕ç)
            if sum_code:
                dist_lookup[sum_code] = data
            
            # Key 2: Join_Key (α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö_α╕¡α╕│α╣Çα╕áα╕¡_α╕òα╕│α╕Üα╕Ñ) - α╕½α╕▒α╕ºα╣âα╕êα╕éα╕¡α╕ç Lookup!
            prov_raw = str(row.get('Province', ''))
            dist_raw = str(row.get('District', ''))
            subdist_raw = str(row.get('Subdistrict', ''))
            
            # Clean name α╕¬α╕│α╕½α╕úα╕▒α╕Ü Join
            prov_clean = clean_name(prov_raw)
            dist_clean = clean_name(dist_raw)
            subdist_clean = clean_name(subdist_raw)
            
            # Join_Key α╣üα╕Üα╕Ü clean (α╕íα╕▓α╕òα╕úα╕Éα╕▓α╕Ö)
            join_key = f"{prov_clean}_{dist_clean}_{subdist_clean}"
            if join_key and join_key != '__':
                name_lookup[join_key] = data
            
            # Join_Key α╣üα╕Üα╕Ü normalized province (α╣Çα╕£α╕╖α╣êα╕¡α╕èα╕╖α╣êα╕¡α╣Çα╕₧α╕╡α╣ëα╕óα╕Ö)
            prov_normalized = normalize_province_name(prov_raw)
            if prov_normalized != prov_clean:
                alt_key = f"{prov_normalized}_{dist_clean}_{subdist_clean}"
                if alt_key and alt_key != '__':
                    name_lookup[alt_key] = data
            
            # Join_Key α╣üα╕Üα╕Üα╕íα╕╡ prefix (α╣Çα╕£α╕╖α╣êα╕¡α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕íα╕╡ prefix)
            raw_key = f"{prov_raw.strip()}_{dist_raw.strip()}_{subdist_raw.strip()}"
            if raw_key and raw_key != '__' and raw_key not in name_lookup:
                name_lookup[raw_key] = data
        
        return {'by_code': dist_lookup, 'by_name': name_lookup}
    except Exception as e:
        return {'by_code': {}, 'by_name': {}}

# α╣éα╕½α╕Ñα╕ö Master Dist Data
MASTER_DIST_DATA = load_master_dist_data()

# ==========================================
# PUNTHAI/MAXMART BUFFER FUNCTIONS
# ==========================================
def is_punthai_only(trip_data):
    """
    α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕ºα╣êα╕▓α╕ùα╕úα╕┤α╕¢α╕Öα╕╡α╣ëα╣Çα╕¢α╣çα╕Ö Punthai α╕Ñα╣ëα╕ºα╕Ö, Maxmart α╕Ñα╣ëα╕ºα╕Ö α╕½α╕úα╕╖α╕¡α╕£α╕¬α╕í
    
    Returns:
        'punthai_only': α╕ûα╣ëα╕▓α╕ùα╕▒α╣ëα╕çα╕½α╕íα╕öα╣Çα╕¢α╣çα╕Ö Punthai (BU = 211 α╕½α╕úα╕╖α╕¡α╕èα╕╖α╣êα╕¡α╕íα╕╡ PUNTHAI)
        'maxmart_only': α╕ûα╣ëα╕▓α╕ùα╕▒α╣ëα╕çα╕½α╕íα╕öα╣Çα╕¢α╣çα╕Ö Maxmart (BU = 200 α╕½α╕úα╕╖α╕¡α╕èα╕╖α╣êα╕¡α╕íα╕╡ MAXMART)
        'mixed': α╕ûα╣ëα╕▓α╕íα╕╡α╕ùα╕▒α╣ëα╕ç Punthai α╣üα╕Ñα╕░ Maxmart
        'other': α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡α╕éα╣ëα╕¡α╕íα╕╣α╕Ñ BU
    """
    if trip_data is None or len(trip_data) == 0:
        return 'other'
    
    punthai_count = 0
    maxmart_count = 0
    total_count = len(trip_data)
    
    for _, row in trip_data.iterrows():
        bu = row.get('BU', None)
        name = str(row.get('Name', '')).upper()
        
        # α╣Çα╕èα╣çα╕ä Punthai: BU = 211 α╕½α╕úα╕╖α╕¡α╕èα╕╖α╣êα╕¡α╕íα╕╡ PUNTHAI
        if bu == 211 or bu == '211' or 'PUNTHAI' in name or 'PUN-' in name:
            punthai_count += 1
        # α╣Çα╕èα╣çα╕ä Maxmart: BU = 200 α╕½α╕úα╕╖α╕¡α╕èα╕╖α╣êα╕¡α╕íα╕╡ MAXMART/MAX MART
        elif bu == 200 or bu == '200' or 'MAXMART' in name or 'MAX MART' in name:
            maxmart_count += 1
    
    if punthai_count == total_count:
        return 'punthai_only'
    elif maxmart_count == total_count:
        return 'maxmart_only'
    elif punthai_count > 0 or maxmart_count > 0:
        return 'mixed'
    else:
        return 'other'

def get_buffer_for_trip(trip_data):
    """
    α╕öα╕╢α╕ç Buffer α╕ùα╕╡α╣êα╣Çα╕½α╕íα╕▓α╕░α╕¬α╕íα╕òα╕▓α╕í BU α╕éα╕¡α╕çα╕ùα╕úα╕┤α╕¢
    
    Rules:
    - Punthai α╕Ñα╣ëα╕ºα╕Ö: BUFFER = 1.0 (α╕½α╣ëα╕▓α╕íα╣Çα╕üα╕┤α╕Ö 100%)
    - Maxmart α╕Ñα╣ëα╕ºα╕Ö/α╕£α╕¬α╕í: BUFFER = 1.10 (α╣Çα╕üα╕┤α╕Öα╣äα╕öα╣ë 10%)
    
    Returns:
        float: buffer multiplier (1.0 α╕½α╕úα╕╖α╕¡ 1.10)
    """
    trip_type = is_punthai_only(trip_data)
    
    if trip_type == 'punthai_only':
        return PUNTHAI_BUFFER  # 1.0 - α╕½α╣ëα╕▓α╕íα╣Çα╕üα╕┤α╕Ö 100%
    elif trip_type in ['maxmart_only', 'mixed']:
        return MAXMART_BUFFER  # 1.10 - α╣Çα╕üα╕┤α╕Öα╣äα╕öα╣ë 10%
    else:
        return BUFFER  # default 1.0

def get_punthai_drop_limit(trip_data, vehicle_type):
    """
    α╕öα╕╢α╕çα╕êα╕│α╕üα╕▒α╕öα╕êα╕│α╕Öα╕ºα╕Ö Drop α╕¬α╕│α╕½α╕úα╕▒α╕Ü Punthai α╕Ñα╣ëα╕ºα╕Ö
    
    Rules:
    - Punthai α╕Ñα╣ëα╕ºα╕Ö + 4W: α╕¬α╕╣α╕çα╕¬α╕╕α╕ö 5 α╕¬α╕▓α╕éα╕▓
    - Punthai α╕Ñα╣ëα╕ºα╕Ö + JB: α╕¬α╕╣α╕çα╕¬α╕╕α╕ö 7 drop
    - α╕¡α╕╖α╣êα╕Öα╣å: α╣äα╕íα╣êα╕êα╕│α╕üα╕▒α╕ö (999)
    
    Returns:
        int: max drops allowed
    """
    trip_type = is_punthai_only(trip_data)
    
    if trip_type == 'punthai_only':
        return PUNTHAI_LIMITS.get(vehicle_type, {}).get('max_drops', 999)
    else:
        return 999  # α╣äα╕íα╣êα╕êα╕│α╕üα╕▒α╕ö

@st.cache_data(ttl=3600)  # Cache 1 α╕èα╕▒α╣êα╕ºα╣éα╕íα╕ç
def load_booking_history_restrictions():
    """α╣éα╕½α╕Ñα╕öα╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤α╕üα╕▓α╕úα╕êα╕▒α╕öα╕¬α╣êα╕çα╕êα╕▓α╕ü Booking History - α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕êα╕úα╕┤α╕ç 3,053 booking (Optimized)"""
    try:
        # α╕Ñα╕¡α╕çα╕½α╕▓α╣äα╕ƒα╕Ñα╣î Booking History (α╕¡α╕▓α╕êα╕íα╕╡α╕èα╕╖α╣êα╕¡α╕½α╕Ñα╕▓α╕óα╣üα╕Üα╕Ü)
        possible_files = [
            'Dc/α╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤α╕çα╕▓α╕Öα╕êα╕▒α╕öα╕¬α╣êα╕ç DC α╕ºα╕▒α╕çα╕Öα╣ëα╕¡α╕ó(1).xlsx',
            'Dc/α╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤α╕çα╕▓α╕Öα╕êα╕▒α╕öα╕¬α╣êα╕ç DC α╕ºα╕▒α╕çα╕Öα╣ëα╕¡α╕ó.xlsx',
            'branch_vehicle_restrictions_from_booking.xlsx'
        ]
        
        file_path = None
        for path in possible_files:
            if os.path.exists(path):
                file_path = path
                break
        
        if not file_path:
            # α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡α╣äα╕ƒα╕Ñα╣î α╣âα╕èα╣ëα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕ùα╕╡α╣êα╣Çα╕äα╕óα╣Çα╕úα╕╡α╕óα╕Öα╕úα╕╣α╣ë (fallback)
            return load_learned_restrictions_fallback()
        
        df = pd.read_excel(file_path)
        
        # α╣üα╕¢α╕Ñα╕çα╕¢α╕úα╕░α╣Çα╕áα╕ùα╕úα╕û
        vehicle_mapping = {
            '4 α╕Ñα╣ëα╕¡ α╕êα╕▒α╕íα╣éα╕Üα╣ë α╕òα╕╣α╣ëα╕ùα╕╢α╕Ü': 'JB',
            '6 α╕Ñα╣ëα╕¡ α╕òα╕╣α╣ëα╕ùα╕╢α╕Ü': '6W',
            '4 α╕Ñα╣ëα╕¡ α╕òα╕╣α╣ëα╕ùα╕╢α╕Ü': '4W'
        }
        df['Vehicle_Type'] = df['α╕¢α╕úα╕░α╣Çα╕áα╕ùα╕úα╕û'].map(vehicle_mapping)
        
        # α╕ºα╕┤α╣Çα╕äα╕úα╕▓α╕░α╕½α╣îα╕äα╕ºα╕▓α╕íα╕¬α╕▒α╕íα╕₧α╕▒α╕Öα╕ÿα╣îα╕¬α╕▓α╕éα╕▓-α╕úα╕û
        branch_vehicle_history = {}
        booking_groups = df.groupby('Booking No')
        
        for booking_no, booking_data in booking_groups:
            vehicle_types = booking_data['Vehicle_Type'].dropna().unique()
            if len(vehicle_types) > 0:
                vehicle = booking_data['Vehicle_Type'].mode()[0] if len(booking_data['Vehicle_Type'].mode()) > 0 else vehicle_types[0]
                for branch_code in booking_data['α╕úα╕½α╕▒α╕¬α╕¬α╕▓α╕éα╕▓'].dropna().unique():
                    if branch_code not in branch_vehicle_history:
                        branch_vehicle_history[branch_code] = []
                    branch_vehicle_history[branch_code].append(vehicle)
        
        # α╕¬α╕úα╣ëα╕▓α╕ç restrictions
        branch_restrictions = {}
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for branch_code, vehicle_list in branch_vehicle_history.items():
            vehicles_used = set(vehicle_list)
            vehicle_counts = pd.Series(vehicle_list).value_counts().to_dict()
            
            if len(vehicles_used) == 1:
                # STRICT - α╣âα╕èα╣ëα╕úα╕ûα╣Çα╕öα╕╡α╕óα╕º
                vehicle = list(vehicles_used)[0]
                branch_restrictions[str(branch_code)] = {
                    'max_vehicle': vehicle,
                    'allowed': [vehicle],
                    'total_bookings': len(vehicle_list),
                    'restriction_type': 'STRICT'
                }
            else:
                # FLEXIBLE - α╣âα╕èα╣ëα╣äα╕öα╣ëα╕½α╕Ñα╕▓α╕óα╕¢α╕úα╕░α╣Çα╕áα╕ù
                max_vehicle = max(vehicles_used, key=lambda v: vehicle_sizes.get(v, 0))
                branch_restrictions[str(branch_code)] = {
                    'max_vehicle': max_vehicle,
                    'allowed': list(vehicles_used),
                    'total_bookings': len(vehicle_list),
                    'restriction_type': 'FLEXIBLE'
                }
        
        stats = {
            'total_branches': len(branch_restrictions),
            'strict': len([b for b, r in branch_restrictions.items() if r['restriction_type'] == 'STRICT']),
            'flexible': len([b for b, r in branch_restrictions.items() if r['restriction_type'] == 'FLEXIBLE']),
            'total_bookings': len(booking_groups)
        }
        
        return {
            'branch_restrictions': branch_restrictions,
            'stats': stats
        }
    except Exception as e:
        # α╕ûα╣ëα╕▓α╣Çα╕üα╕┤α╕ö error α╣âα╕èα╣ëα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕ùα╕╡α╣êα╣Çα╕äα╕óα╣Çα╕úα╕╡α╕óα╕Öα╕úα╕╣α╣ëα╣üα╕ùα╕Ö
        return load_learned_restrictions_fallback()

def load_learned_restrictions_fallback():
    """
    α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕ùα╕╡α╣êα╣Çα╕úα╕╡α╕óα╕Öα╕úα╕╣α╣ëα╕êα╕▓α╕ü Booking History (backup)
    α╣âα╕èα╣ëα╣Çα╕íα╕╖α╣êα╕¡α╣äα╕íα╣êα╕¬α╕▓α╕íα╕▓α╕úα╕ûα╣éα╕½α╕Ñα╕öα╣äα╕ƒα╕Ñα╣îα╣äα╕öα╣ë
    
    α╕êα╕▓α╕üα╕üα╕▓α╕úα╕ºα╕┤α╣Çα╕äα╕úα╕▓α╕░α╕½α╣î 3,053 bookings, 2,790 α╕¬α╕▓α╕éα╕▓:
    - JB: α╕úα╕ûα╕üα╕Ñα╕▓α╕ç (α╣âα╕èα╣ëα╕íα╕▓α╕üα╕ùα╕╡α╣êα╕¬α╕╕α╕ö 54.7%)
    - 6W: α╕úα╕ûα╣âα╕½α╕ìα╣ê (30.1%)
    - 4W: α╕úα╕ûα╣Çα╕Ñα╣çα╕ü (0.2%)
    
    α╕üα╕Ñα╕óα╕╕α╕ùα╕ÿα╣î: α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡α╕éα╣ëα╕¡α╕íα╕╣α╕Ñ default α╣Çα╕¢α╣çα╕Ö JB (α╕úα╕ûα╕üα╕Ñα╕▓α╕ç α╣âα╕èα╣ëα╣äα╕öα╣ëα╕üα╕▒α╕Üα╕¬α╕▓α╕éα╕▓α╕¬α╣êα╕ºα╕Öα╣âα╕½α╕ìα╣ê)
    """
    return {
        'branch_restrictions': {},
        'stats': {
            'total_branches': 0,
            'strict': 0,
            'flexible': 0,
            'total_bookings': 0,
            'fallback': True,
            'message': 'α╣âα╕èα╣ë Punthai α╣Çα╕¢α╣çα╕Öα╕½α╕Ñα╕▒α╕ü (α╣äα╕íα╣êα╕₧α╕Üα╣äα╕ƒα╕Ñα╣î Booking History)'
        }
    }

@st.cache_data(ttl=3600)  # Cache 1 α╕èα╕▒α╣êα╕ºα╣éα╕íα╕ç
def load_punthai_reference():
    """α╣éα╕½α╕Ñα╕öα╣äα╕ƒα╕Ñα╣î Punthai Maxmart α╣Çα╕₧α╕╖α╣êα╕¡α╣Çα╕úα╕╡α╕óα╕Öα╕úα╕╣α╣ëα╕½α╕Ñα╕▒α╕üα╕üα╕▓α╕úα╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢ (Location patterns - Optimized)"""
    try:
        file_path = 'Dc/α╣üα╕£α╕Öα╕çα╕▓α╕Ö Punthai Maxmart α╕úα╕¡α╕Üα╕¬α╕▒α╣êα╕ç 24α╕½α╕óα╕┤α╕Ü 25α╕₧α╕ñα╕¿α╕êα╕┤α╕üα╕▓α╕óα╕Ö 2568 To.α╣Çα╕ƒα╕┤(1) - α╕¬α╕│α╣Çα╕Öα╕▓.xlsx'
        df = pd.read_excel(file_path, sheet_name='2.Punthai', header=1)
        
        # α╕üα╕úα╕¡α╕çα╣Çα╕ëα╕₧α╕▓α╕░α╣üα╕ûα╕ºα╕ùα╕╡α╣êα╕íα╕╡ Trip α╣üα╕Ñα╕░α╣äα╕íα╣êα╣âα╕èα╣ê DC/Distribution Center
        df_clean = df[df['Trip'].notna()].copy()
        df_clean = df_clean[~df_clean['BranchCode'].isin(['DC011', 'PTDC', 'PTG Distribution Center'])].copy()
        
        # Extract vehicle type from Trip no (α╣Çα╕èα╣êα╕Ö 4W009 ΓåÆ 4W)
        df_clean['Vehicle_Type'] = df_clean['Trip no'].apply(
            lambda x: str(x)[:2] if pd.notna(x) else 'Unknown'
        )
        
        # Merge α╕üα╕▒α╕Ü Master α╣Çα╕₧α╕╖α╣êα╕¡α╣äα╕öα╣ëα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕òα╕│α╕Üα╕Ñ/α╕¡α╕│α╣Çα╕áα╕¡/α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö
        try:
            df_master = pd.read_excel('Dc/Master α╕¬α╕ûα╕▓α╕Öα╕ùα╕╡α╣êα╕¬α╣êα╕ç.xlsx')
            df_clean = df_clean.merge(
                df_master[['Plan Code', 'α╕òα╕│α╕Üα╕Ñ', 'α╕¡α╕│α╣Çα╕áα╕¡', 'α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö']],
                left_on='BranchCode',
                right_on='Plan Code',
                how='left'
            )
        except:
            pass
        
        # α╣Çα╕úα╕╡α╕óα╕Öα╕úα╕╣α╣ëα╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕úα╕ûα╕êα╕▓α╕ü Punthai (α╣üα╕£α╕Ö) - α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣äα╕íα╣êα╕íα╕╡α╣âα╕Ö Booking
        punthai_restrictions = {}
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for branch_code in df_clean['BranchCode'].unique():
            branch_data = df_clean[df_clean['BranchCode'] == branch_code]
            vehicles_used = set(branch_data['Vehicle_Type'].dropna().tolist())
            vehicles_used = {v for v in vehicles_used if v in ['4W', 'JB', '6W']}
            
            if vehicles_used:
                if len(vehicles_used) == 1:
                    vehicle = list(vehicles_used)[0]
                    punthai_restrictions[str(branch_code)] = {
                        'max_vehicle': vehicle,
                        'allowed': [vehicle],
                        'source': 'PUNTHAI'
                    }
                else:
                    max_vehicle = max(vehicles_used, key=lambda v: vehicle_sizes.get(v, 0))
                    punthai_restrictions[str(branch_code)] = {
                        'max_vehicle': max_vehicle,
                        'allowed': list(vehicles_used),
                        'source': 'PUNTHAI'
                    }
        
        # α╕¬α╕úα╣ëα╕▓α╕ç dictionary: Trip ΓåÆ α╕éα╣ëα╕¡α╕íα╕╣α╕Ñ (location patterns)
        trip_patterns = {}
        location_stats = {
            'same_province': 0,
            'mixed_province': 0,
            'avg_branches': 0
        }
        
        for trip_num in df_clean['Trip'].unique():
            trip_data = df_clean[df_clean['Trip'] == trip_num]
            
            # Get location info
            provinces = set(trip_data['α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö'].dropna().tolist()) if 'α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö' in trip_data.columns else set()
            
            # Count same vs mixed province
            if len(provinces) == 1:
                location_stats['same_province'] += 1
            elif len(provinces) > 1:
                location_stats['mixed_province'] += 1
            
            trip_patterns[int(trip_num)] = {
                'branches': len(trip_data),
                'codes': trip_data['BranchCode'].tolist(),
                'weight': trip_data['TOTALWGT'].sum() if 'TOTALWGT' in trip_data.columns else 0,
                'cube': trip_data['TOTALCUBE'].sum() if 'TOTALCUBE' in trip_data.columns else 0,
                'provinces': list(provinces),
                'same_province': len(provinces) == 1
            }
        
        # Calculate stats
        if trip_patterns:
            location_stats['avg_branches'] = sum(t['branches'] for t in trip_patterns.values()) / len(trip_patterns)
            total = location_stats['same_province'] + location_stats['mixed_province']
            location_stats['same_province_pct'] = (location_stats['same_province'] / total * 100) if total > 0 else 0
        
        return {
            'patterns': trip_patterns, 
            'stats': location_stats,
            'punthai_restrictions': punthai_restrictions
        }
    except:
        return {'patterns': {}, 'stats': {}, 'punthai_restrictions': {}}

# α╣éα╕½α╕Ñα╕ö Booking History (α╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕úα╕û)
BOOKING_RESTRICTIONS = load_booking_history_restrictions()

# α╣éα╕½α╕Ñα╕ö Punthai Reference (location patterns)
PUNTHAI_PATTERNS = load_punthai_reference()

# ==========================================
# HELPER FUNCTIONS
# ==========================================
def normalize(val):
    """α╕ùα╕│α╣âα╕½α╣ëα╕úα╕½α╕▒α╕¬α╕¬α╕▓α╕éα╕▓α╣Çα╕¢α╣çα╕Öα╕íα╕▓α╕òα╕úα╕Éα╕▓α╕Ö"""
    return str(val).strip().upper().replace(" ", "").replace(".0", "")

def calculate_distance(lat1, lon1, lat2, lon2):
    """α╕äα╕│α╕Öα╕ºα╕ôα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕úα╕░α╕½α╕ºα╣êα╕▓α╕çα╕¬α╕¡α╕çα╕êα╕╕α╕ö (α╕üα╕í.) - Haversine formula"""
    if lat1 == 0 or lon1 == 0 or lat2 == 0 or lon2 == 0:
        return 0
    import math
    lat1_rad, lon1_rad = math.radians(lat1), math.radians(lon1)
    lat2_rad, lon2_rad = math.radians(lat2), math.radians(lon2)
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    return 6371 * c

def calculate_distance_from_dc(lat, lon):
    """α╕äα╕│α╕Öα╕ºα╕ôα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕êα╕▓α╕ü DC α╕ºα╕▒α╕çα╕Öα╣ëα╕¡α╕ó (α╕üα╕í.)"""
    return calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)

def check_branch_vehicle_compatibility(branch_code, vehicle_type):
    """α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕ºα╣êα╕▓α╕¬α╕▓α╕éα╕▓α╕Öα╕╡α╣ëα╣âα╕èα╣ëα╕úα╕ûα╕¢α╕úα╕░α╣Çα╕áα╕ùα╕Öα╕╡α╣ëα╣äα╕öα╣ëα╣äα╕½α╕í (α╕úα╕ºα╕í Booking + Punthai)"""
    branch_code_str = str(branch_code).strip()
    
    # 1. α╕Ñα╕¡α╕çα╕½α╕▓α╕êα╕▓α╕ü Booking History α╕üα╣êα╕¡α╕Ö (α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕êα╕úα╕┤α╕ç)
    booking_restrictions = BOOKING_RESTRICTIONS.get('branch_restrictions', {})
    if branch_code_str in booking_restrictions:
        allowed = booking_restrictions[branch_code_str].get('allowed', [])
        return vehicle_type in allowed
    
    # 2. α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡ α╕Ñα╕¡α╕çα╕½α╕▓α╕êα╕▓α╕ü Punthai (α╣üα╕£α╕Ö)
    punthai_restrictions = PUNTHAI_PATTERNS.get('punthai_restrictions', {})
    if branch_code_str in punthai_restrictions:
        allowed = punthai_restrictions[branch_code_str].get('allowed', [])
        return vehicle_type in allowed
    
    # 3. α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡α╕éα╣ëα╕¡α╕íα╕╣α╕Ñ = α╕óα╕╖α╕öα╕½α╕óα╕╕α╣êα╕Ö
    return True

def get_max_vehicle_for_branch(branch_code):
    """α╕öα╕╢α╕çα╕úα╕ûα╣âα╕½α╕ìα╣êα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╕¬α╕▓α╕éα╕▓α╕Öα╕╡α╣ëα╕úα╕¡α╕çα╕úα╕▒α╕Ü (α╕úα╕ºα╕í Booking History + Punthai)"""
    branch_code_str = str(branch_code).strip()
    
    # 1. α╕Ñα╕¡α╕çα╕½α╕▓α╕êα╕▓α╕ü Booking History α╕üα╣êα╕¡α╕Ö (α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕êα╕úα╕┤α╕ç - α╕äα╕ºα╕▓α╕íα╣Çα╕èα╕╖α╣êα╕¡α╕íα╕▒α╣êα╕Öα╕¬α╕╣α╕ç)
    booking_restrictions = BOOKING_RESTRICTIONS.get('branch_restrictions', {})
    if branch_code_str in booking_restrictions:
        return booking_restrictions[branch_code_str].get('max_vehicle', '6W')
    
    # 2. α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡ α╕Ñα╕¡α╕çα╕½α╕▓α╕êα╕▓α╕ü Punthai (α╣üα╕£α╕Ö - α╕¬α╕│α╕úα╕¡α╕ç)
    punthai_restrictions = PUNTHAI_PATTERNS.get('punthai_restrictions', {})
    if branch_code_str in punthai_restrictions:
        return punthai_restrictions[branch_code_str].get('max_vehicle', '6W')
    
    # 3. α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡α╕ùα╕▒α╣ëα╕çα╕¬α╕¡α╕çα╣üα╕½α╕Ñα╣êα╕ç = α╣âα╕èα╣ëα╕úα╕ûα╣âα╕½α╕ìα╣êα╣äα╕öα╣ë
    return '6W'

def get_max_vehicle_for_trip(trip_codes):
    """
    α╕½α╕▓α╕úα╕ûα╣âα╕½α╕ìα╣êα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╕ùα╕úα╕┤α╕¢α╕Öα╕╡α╣ëα╣âα╕èα╣ëα╣äα╕öα╣ë (α╣Çα╕èα╣çα╕äα╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕éα╕¡α╕çα╕ùα╕╕α╕üα╕¬α╕▓α╕éα╕▓α╣âα╕Öα╕ùα╕úα╕┤α╕¢)
    
    Args:
        trip_codes: set α╕éα╕¡α╕ç branch codes α╣âα╕Öα╕ùα╕úα╕┤α╕¢
    
    Returns:
        str: '4W', 'JB', α╕½α╕úα╕╖α╕¡ '6W'
    """
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    max_allowed = '6W'  # α╣Çα╕úα╕┤α╣êα╕íα╕êα╕▓α╕üα╣âα╕½α╕ìα╣êα╕¬α╕╕α╕ö α╣üα╕Ñα╣ëα╕ºα╕êα╕│α╕üα╕▒α╕öα╕òα╕▓α╕íα╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕¬α╕▓α╕éα╕▓
    min_priority = 3  # α╕äα╣êα╕▓α╣âα╕½α╕ìα╣êα╕¬α╕╕α╕öα╕äα╕╖α╕¡α╣äα╕íα╣êα╕íα╕╡α╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕ö
    
    for code in trip_codes:
        branch_max = get_max_vehicle_for_branch(code)
        priority = vehicle_priority.get(branch_max, 3)
        
        # ≡ƒöÆ α╣Çα╕Ñα╕╖α╕¡α╕üα╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕Ñα╣çα╕üα╕ùα╕╡α╣êα╕¬α╕╕α╕ö (α╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕íα╕▓α╕üα╕ùα╕╡α╣êα╕¬α╕╕α╕ö) α╕êα╕▓α╕üα╕ùα╕╕α╕üα╕¬α╕▓α╕éα╕▓α╣âα╕Öα╕ùα╕úα╕┤α╕¢
        if priority < min_priority:
            min_priority = priority
            max_allowed = branch_max
    
    return max_allowed

def get_required_vehicle_by_distance(branch_code):
    """α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕ºα╣êα╕▓α╕¬α╕▓α╕éα╕▓α╕òα╣ëα╕¡α╕çα╣âα╕èα╣ëα╕úα╕ûα╕¡α╕░α╣äα╕úα╕òα╕▓α╕íα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕êα╕▓α╕ü DC"""
    # α╕öα╕╢α╕çα╕₧α╕┤α╕üα╕▒α╕öα╕êα╕▓α╕ü Master
    if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
        master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == branch_code]
        if len(master_row) > 0:
            lat = master_row.iloc[0].get('α╕Ñα╕░α╕òα╕┤α╕êα╕╣α╕ö', 0)
            lon = master_row.iloc[0].get('α╕Ñα╕¡α╕çα╕òα╕┤α╕êα╕╣α╕ö', 0)
            distance = calculate_distance_from_dc(lat, lon)
            
            # α╕ûα╣ëα╕▓α╕½α╣êα╕▓α╕çα╕êα╕▓α╕ü DC α╣Çα╕üα╕┤α╕Öα╕üα╕│α╕½α╕Öα╕ö ΓåÆ α╕òα╣ëα╕¡α╕çα╣âα╕èα╣ë 6W
            if distance > DISTANCE_REQUIRE_6W:
                return '6W', distance
    
    return None, 0

def can_fit_truck(total_weight, total_cube, truck_type):
    """α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓α╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü/α╕äα╕┤α╕ºα╣âα╕¬α╣êα╕úα╕ûα╣äα╕öα╣ëα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê"""
    limits = LIMITS[truck_type]
    max_w = limits['max_w'] * BUFFER
    max_c = limits['max_c'] * BUFFER
    return total_weight <= max_w and total_cube <= max_c

def suggest_truck(total_weight, total_cube, max_allowed='6W', trip_codes=None):
    """
    α╣üα╕Öα╕░α╕Öα╕│α╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕½α╕íα╕▓α╕░α╕¬α╕í α╣éα╕öα╕óα╣Çα╕Ñα╕╖α╕¡α╕üα╕úα╕ûα╕ùα╕╡α╣ê:
    1. α╣âα╕¬α╣êα╕éα╕¡α╕çα╣äα╕öα╣ëα╕₧α╕¡α╕öα╕╡ (α╣äα╕íα╣êα╣Çα╕üα╕┤α╕Öα╕éα╕╡α╕öα╕êα╕│α╕üα╕▒α╕ö 105%)
    2. α╣âα╕èα╣ëα╕çα╕▓α╕Öα╣äα╕öα╣ëα╣âα╕üα╕Ñα╣ë 100% α╕íα╕▓α╕üα╕ùα╕╡α╣êα╕¬α╕╕α╕ö (α╣Çα╕¢α╣ëα╕▓α╕½α╕íα╕▓α╕ó: 90-100%)
    3. α╣Çα╕äα╕▓α╕úα╕₧α╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕éα╕¡α╕çα╕¬α╕▓α╕éα╕▓ (α╕ûα╣ëα╕▓α╕¬α╕▓α╕éα╕▓α╣âα╕èα╣ëα╣üα╕äα╣ê 4W = α╕òα╣ëα╕¡α╕çα╣âα╕èα╣ë 4W α╣Çα╕ùα╣êα╕▓α╕Öα╕▒α╣ëα╕Ö)
    """
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    max_size = vehicle_sizes.get(max_allowed, 3)
    
    # α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕éα╕¡α╕çα╕¬α╕▓α╕éα╕▓α╕ùα╕▒α╣ëα╕çα╕½α╕íα╕öα╣âα╕Öα╕üα╕Ñα╕╕α╣êα╕í
    branch_max_vehicle = '4W'  # ≡ƒöÆ α╣Çα╕úα╕┤α╣êα╕íα╕òα╣ëα╕Öα╕ùα╕╡α╣ê 4W (α╣Çα╕Ñα╣çα╕üα╕¬α╕╕α╕ö) α╣üα╕Ñα╣ëα╕ºα╕éα╕óα╕▓α╕óα╣Çα╕íα╕╖α╣êα╕¡α╕êα╕│α╣Çα╕¢α╣çα╕Ö
    if trip_codes is not None and len(trip_codes) > 0:
        for code in trip_codes:
            branch_max = get_max_vehicle_for_branch(code)
            # α╕½α╕▓α╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕Ñα╣çα╕üα╕ùα╕╡α╣êα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╕òα╣ëα╕¡α╕çα╣âα╕èα╣ë
            if vehicle_sizes.get(branch_max, 3) < vehicle_sizes.get(branch_max_vehicle, 3):
                branch_max_vehicle = branch_max
        
        # α╕êα╕│α╕üα╕▒α╕ö max_allowed α╕òα╕▓α╕íα╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕éα╕¡α╕çα╕¬α╕▓α╕éα╕▓
        if vehicle_sizes.get(branch_max_vehicle, 3) < max_size:
            max_allowed = branch_max_vehicle
            max_size = vehicle_sizes.get(max_allowed, 3)
    
    best_truck = None
    best_utilization = 0
    best_distance_from_100 = 999  # α╕úα╕░α╕óα╕░α╕½α╣êα╕▓α╕çα╕êα╕▓α╕ü 100%
    
    for truck in ['4W', 'JB', '6W']:
        truck_size = vehicle_sizes.get(truck, 0)
        # α╕ûα╣ëα╕▓α╕úα╕ûα╣âα╕½α╕ìα╣êα╕üα╕ºα╣êα╕▓α╕ùα╕╡α╣êα╕¡α╕Öα╕╕α╕ìα╕▓α╕ò α╕éα╣ëα╕▓α╕íα╣äα╕¢
        if truck_size > max_size:
            continue
        if can_fit_truck(total_weight, total_cube, truck):
            # α╕äα╕│α╕Öα╕ºα╕ô % α╕üα╕▓α╕úα╣âα╕èα╣ëα╕úα╕û
            limits = LIMITS[truck]
            w_util = (total_weight / limits['max_w']) * 100
            c_util = (total_cube / limits['max_c']) * 100
            utilization = max(w_util, c_util)
            
            # α╕äα╕│α╕Öα╕ºα╕ôα╕úα╕░α╕óα╕░α╕½α╣êα╕▓α╕çα╕êα╕▓α╕ü 100%
            distance_from_100 = abs(100 - utilization)
            
            # α╣Çα╕Ñα╕╖α╕¡α╕üα╕úα╕ûα╕ùα╕╡α╣êα╣âα╕üα╕Ñα╣ë 100% α╕ùα╕╡α╣êα╕¬α╕╕α╕ö (90-105% α╣Çα╕¢α╣çα╕Öα╣Çα╕¢α╣ëα╕▓α╕½α╕íα╕▓α╕ó)
            # α╕ûα╣ëα╕▓α╣âα╕èα╣ëα╕çα╕▓α╕Öα╣âα╕üα╕Ñα╣ëα╣Çα╕äα╕╡α╕óα╕çα╕üα╕▒α╕Ö α╣Çα╕Ñα╕╖α╕¡α╕üα╕úα╕ûα╕ùα╕╡α╣êα╣âα╕èα╣ëα╕çα╕▓α╕Öα╕¬α╕╣α╕çα╕üα╕ºα╣êα╕▓
            if best_truck is None:
                best_truck = truck
                best_utilization = utilization
                best_distance_from_100 = distance_from_100
            else:
                # α╕ûα╣ëα╕▓α╕¡α╕óα╕╣α╣êα╣âα╕Öα╕èα╣êα╕ºα╕ç 90-105% α╣Çα╕Ñα╕╖α╕¡α╕üα╕ùα╕╡α╣êα╣âα╕üα╕Ñα╣ë 100% α╕ùα╕╡α╣êα╕¬α╕╕α╕ö
                if 90 <= utilization <= 105:
                    if distance_from_100 < best_distance_from_100 or best_utilization < 90:
                        best_truck = truck
                        best_utilization = utilization
                        best_distance_from_100 = distance_from_100
                # α╕ûα╣ëα╕▓α╕ùα╕▒α╣ëα╕çα╕äα╕╣α╣êα╣äα╕íα╣êα╕¡α╕óα╕╣α╣êα╣âα╕Öα╕èα╣êα╕ºα╕ç α╣Çα╕Ñα╕╖α╕¡α╕üα╕ùα╕╡α╣êα╣âα╕èα╣ëα╕çα╕▓α╕Öα╕¬α╕╣α╕çα╕üα╕ºα╣êα╕▓
                elif utilization > best_utilization:
                    best_truck = truck
                    best_utilization = utilization
                    best_distance_from_100 = distance_from_100
    
    if best_truck:
        return best_truck
    
    # α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡α╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕½α╕íα╕▓α╕░α╕¬α╕í α╣âα╕èα╣ëα╕úα╕ûα╣âα╕½α╕ìα╣êα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╕¡α╕Öα╕╕α╕ìα╕▓α╕ò
    return max_allowed if max_allowed in LIMITS else '6W+'

def calculate_optimal_vehicle_split(total_weight, total_cube, max_allowed='6W', branch_count=0):
    """
    ≡ƒÜ¢ α╕äα╕│α╕Öα╕ºα╕ôα╕üα╕▓α╕úα╣üα╕Üα╣êα╕çα╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕½α╕íα╕▓α╕░α╕¬α╕í
    
    α╣Çα╕çα╕╖α╣êα╕¡α╕Öα╣äα╕é:
    - 4W: Γëñ12 α╕êα╕╕α╕ö, Cube Γëñ 5
    - JB: Γëñ12 α╕êα╕╕α╕ö, Cube Γëñ 8  
    - 6W: α╣äα╕íα╣êα╕êα╕│α╕üα╕▒α╕öα╕êα╕╕α╕ö, Cube α╕òα╣ëα╕¡α╕çα╣Çα╕òα╣çα╕í ΓëÑ100%
    
    α╕Ñα╕│α╕öα╕▒α╕Üα╕üα╕▓α╕úα╣Çα╕Ñα╕╖α╕¡α╕ü:
    1. 4W (α╕ûα╣ëα╕▓ cube Γëñ 5)
    2. JB (α╕ûα╣ëα╕▓ cube Γëñ 8)
    3. JB + 4W (α╣üα╕óα╕ü 2 α╕äα╕▒α╕Ö, 75%-95% α╕òα╣êα╕¡α╕äα╕▒α╕Ö)
    4. JB + JB (α╣üα╕óα╕ü 2 α╕äα╕▒α╕Ö, 75%-95% α╕òα╣êα╕¡α╕äα╕▒α╕Ö)
    5. 6W + JB (α╣üα╕óα╕ü 2 α╕äα╕▒α╕Ö, 75%-95% α╕òα╣êα╕¡α╕äα╕▒α╕Ö)
    6. 4W + 4W (α╣üα╕óα╕ü 2 α╕äα╕▒α╕Ö, 75%-95% α╕òα╣êα╕¡α╕äα╕▒α╕Ö)
    7. 6W (cube α╕òα╣ëα╕¡α╕ç ΓëÑ100%)
    
    Returns: (vehicle_type, split_needed, split_config)
    """
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    max_priority = vehicle_priority.get(max_allowed, 3)
    
    # α╕äα╕│α╕Öα╕ºα╕ô utilization α╕¬α╕│α╕½α╕úα╕▒α╕Üα╣üα╕òα╣êα╕Ñα╕░α╕úα╕û (α╣âα╕èα╣ë Cube α╣Çα╕¢α╣çα╕Öα╕½α╕Ñα╕▒α╕ü)
    cube_util_4w = (total_cube / LIMITS['4W']['max_c']) * 100  # max 5 cube
    cube_util_jb = (total_cube / LIMITS['JB']['max_c']) * 100  # max 8 cube
    cube_util_6w = (total_cube / LIMITS['6W']['max_c']) * 100  # max 20 cube
    
    weight_util_4w = (total_weight / LIMITS['4W']['max_w']) * 100
    weight_util_jb = (total_weight / LIMITS['JB']['max_w']) * 100
    weight_util_6w = (total_weight / LIMITS['6W']['max_w']) * 100
    
    # ≡ƒÄ» α╣Çα╕¢α╣ëα╕▓α╕½α╕íα╕▓α╕ó: Utilization 75%-95% α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕üα╕▓α╕úα╣üα╕óα╕ü, 95%-105% α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕äα╕▒α╕Öα╣Çα╕öα╕╡α╕óα╕º
    SPLIT_MIN = 75   # α╕éα╕▒α╣ëα╕Öα╕òα╣êα╕│α╕¬α╕│α╕½α╕úα╕▒α╕Üα╣üα╕òα╣êα╕Ñα╕░α╕äα╕▒α╕Öα╣Çα╕íα╕╖α╣êα╕¡α╣üα╕óα╕ü
    SPLIT_MAX = 95   # α╕¬α╕╣α╕çα╕¬α╕╕α╕öα╕¬α╕│α╕½α╕úα╕▒α╕Üα╣üα╕òα╣êα╕Ñα╕░α╕äα╕▒α╕Öα╣Çα╕íα╕╖α╣êα╕¡α╣üα╕óα╕ü
    SINGLE_MIN = 95  # α╕éα╕▒α╣ëα╕Öα╕òα╣êα╕│α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕äα╕▒α╕Öα╣Çα╕öα╕╡α╕óα╕º
    SINGLE_MAX = 105 # α╕¬α╕╣α╕çα╕¬α╕╕α╕öα╕¬α╕│α╕½α╕úα╕▒α╕Üα╕äα╕▒α╕Öα╣Çα╕öα╕╡α╕óα╕º
    
    # α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕êα╕│α╕Öα╕ºα╕Öα╕¬α╕▓α╕éα╕▓ (4W/JB α╣äα╕íα╣êα╣Çα╕üα╕┤α╕Ö 12 α╕êα╕╕α╕ö)
    branch_ok_for_small = branch_count <= 12 or branch_count == 0
    
    # 1. α╕Ñα╕¡α╕ç 4W α╕üα╣êα╕¡α╕Ö (α╕ûα╣ëα╕▓ cube Γëñ 5 α╣üα╕Ñα╕░ Γëñ12 α╕êα╕╕α╕ö)
    if max_priority >= 1 and total_cube <= 5.0 and branch_ok_for_small:
        if cube_util_4w <= 105 and weight_util_4w <= 105:
            return ('4W', False, None)
    
    # 2. α╕Ñα╕¡α╕ç JB (α╕ûα╣ëα╕▓ cube Γëñ 8 α╣üα╕Ñα╕░ Γëñ12 α╕êα╕╕α╕ö)
    if max_priority >= 2 and total_cube <= 8.0 and branch_ok_for_small:
        if cube_util_jb <= 105 and weight_util_jb <= 105:
            return ('JB', False, None)
    
    # 3. α╕ûα╣ëα╕▓α╕úα╕ûα╣Çα╕öα╕╡α╕óα╕ºα╣äα╕íα╣êα╕₧α╕¡ α╕òα╣ëα╕¡α╕çα╣üα╕óα╕ü (cube > 8 α╕½α╕úα╕╖α╕¡ α╕êα╕╕α╕ö > 12)
    need_split = total_cube > 8.0 or not branch_ok_for_small
    
    if need_split:
        # ≡ƒöä α╕Ñα╕¡α╕çα╣üα╕Üα╕Üα╕òα╣êα╕▓α╕çα╣å α╕òα╕▓α╕íα╕Ñα╕│α╕öα╕▒α╕Ü - α╣Çα╕¢α╣ëα╕▓α╕½α╕íα╕▓α╕ó 75%-95% α╕òα╣êα╕¡α╕äα╕▒α╕Ö
        
        # JB + 4W (JB 8 cube + 4W 5 cube = 13 cube max)
        if max_priority >= 2 and total_cube <= 13.0:
            # α╣üα╕Üα╣êα╕ç: JB α╕úα╕▒α╕Ü cube α╕íα╕▓α╕üα╕üα╕ºα╣êα╕▓, 4W α╕úα╕▒α╕Üα╕¬α╣êα╕ºα╕Öα╕ùα╕╡α╣êα╣Çα╕½α╕Ñα╕╖α╕¡
            jb_cube = min(total_cube * 0.6, 8.0)  # JB α╕úα╕▒α╕Ü 60% α╣üα╕òα╣êα╣äα╕íα╣êα╣Çα╕üα╕┤α╕Ö 8
            four_w_cube = total_cube - jb_cube
            
            jb_util = (jb_cube / LIMITS['JB']['max_c']) * 100
            four_w_util = (four_w_cube / LIMITS['4W']['max_c']) * 100
            
            if SPLIT_MIN <= jb_util <= SPLIT_MAX and SPLIT_MIN <= four_w_util <= SPLIT_MAX:
                return ('JB', True, {'split': ['JB', '4W'], 'ratio': [jb_cube/total_cube, four_w_cube/total_cube]})
        
        # JB + JB (JB 8 + JB 8 = 16 cube max)
        if max_priority >= 2 and total_cube <= 16.0:
            jb_util_half = (total_cube / 2 / LIMITS['JB']['max_c']) * 100
            if SPLIT_MIN <= jb_util_half <= SPLIT_MAX:
                return ('JB', True, {'split': ['JB', 'JB'], 'ratio': [0.5, 0.5]})
        
        # 6W + JB (6W 20 + JB 8 = 28 cube max)
        if max_priority >= 3 and total_cube <= 28.0:
            # α╣üα╕Üα╣êα╕ç: 6W α╕úα╕▒α╕Üα╕¬α╣êα╕ºα╕Öα╣âα╕½α╕ìα╣ê
            six_w_cube = min(total_cube * 0.7, 20.0)
            jb_cube = total_cube - six_w_cube
            
            six_w_util = (six_w_cube / LIMITS['6W']['max_c']) * 100
            jb_util = (jb_cube / LIMITS['JB']['max_c']) * 100
            
            if six_w_util >= 75 and SPLIT_MIN <= jb_util <= SPLIT_MAX:
                return ('6W', True, {'split': ['6W', 'JB'], 'ratio': [six_w_cube/total_cube, jb_cube/total_cube]})
        
        # 4W + 4W (4W 5 + 4W 5 = 10 cube max) - α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╕êα╕│α╕üα╕▒α╕ö 4W
        if max_priority == 1 and total_cube <= 10.0:
            four_w_util_half = (total_cube / 2 / LIMITS['4W']['max_c']) * 100
            if SPLIT_MIN <= four_w_util_half <= SPLIT_MAX:
                return ('4W', True, {'split': ['4W', '4W'], 'ratio': [0.5, 0.5]})
    
    # 4. 6W (α╣äα╕íα╣êα╕êα╕│α╕üα╕▒α╕öα╕êα╕╕α╕ö α╣üα╕òα╣ê cube α╕òα╣ëα╕¡α╕ç ΓëÑ100%)
    if max_priority >= 3:
        if cube_util_6w >= 100:
            return ('6W', False, None)
        elif cube_util_6w >= 80:
            # 6W α╣äα╕íα╣êα╣Çα╕òα╣çα╕í (80-99%) ΓåÆ α╕óα╕▒α╕çα╕₧α╕¡α╕úα╕▒α╕Üα╣äα╕öα╣ë
            return ('6W', False, None)
        else:
            # 6W α╕ºα╣êα╕▓α╕çα╕íα╕▓α╕ü (<80%) ΓåÆ α╕Ñα╕öα╣Çα╕¢α╣çα╕Ö JB α╕ûα╣ëα╕▓α╣äα╕öα╣ë
            if total_cube <= 8.0 and branch_ok_for_small and max_priority >= 2:
                return ('JB', False, None)
            # α╕ûα╣ëα╕▓ JB α╣äα╕íα╣êα╣äα╕öα╣ë α╕Ñα╕öα╣Çα╕¢α╣çα╕Ö 4W
            if total_cube <= 5.0 and branch_ok_for_small:
                return ('4W', False, None)
    
    # Default: α╣âα╕èα╣ë max_allowed
    return (max_allowed, False, None)

def can_branch_use_vehicle(code, vehicle_type, branch_vehicles):
    """
    α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓α╕¬α╕▓α╕éα╕▓α╕¬α╕▓α╕íα╕▓α╕úα╕ûα╣âα╕èα╣ëα╕úα╕ûα╕¢α╕úα╕░α╣Çα╕áα╕ùα╕Öα╕╡α╣ëα╣äα╕öα╣ëα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê
    - α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡α╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤ = α╣âα╕èα╣ëα╣äα╕öα╣ëα╕ùα╕╕α╕üα╕¢α╕úα╕░α╣Çα╕áα╕ù
    - α╕ûα╣ëα╕▓α╕íα╕╡α╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤α╣âα╕èα╣ëα╕úα╕ûα╣âα╕½α╕ìα╣ê = α╣âα╕èα╣ëα╕úα╕ûα╣Çα╕Ñα╣çα╕üα╕üα╕ºα╣êα╕▓α╣äα╕öα╣ë
    - α╕ûα╣ëα╕▓α╕íα╕╡α╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤α╣âα╕èα╣ëα╣üα╕äα╣êα╕úα╕ûα╣Çα╕Ñα╣çα╕ü (α╣Çα╕èα╣êα╕Ö 4W) = α╣âα╕èα╣ëα╕úα╕ûα╣âα╕½α╕ìα╣êα╣äα╕íα╣êα╣äα╕öα╣ë (α╕úα╕ûα╣âα╕½α╕ìα╣êα╣Çα╕éα╣ëα╕▓α╣äα╕íα╣êα╣äα╕öα╣ë)
    """
    if not branch_vehicles or code not in branch_vehicles:
        return True  # α╣äα╕íα╣êα╕íα╕╡α╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤ = α╣âα╕èα╣ëα╣äα╕öα╣ëα╕ùα╕╕α╕üα╕¢α╕úα╕░α╣Çα╕áα╕ù
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return True  # α╣äα╕íα╣êα╕íα╕╡α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕úα╕û = α╣âα╕èα╣ëα╣äα╕öα╣ëα╕ùα╕╕α╕üα╕¢α╕úα╕░α╣Çα╕áα╕ù
    
    # α╕ûα╣ëα╕▓α╣Çα╕äα╕óα╣âα╕èα╣ëα╕úα╕ûα╕¢α╕úα╕░α╣Çα╕áα╕ùα╕Öα╕╡α╣ë = α╣âα╕èα╣ëα╣äα╕öα╣ë
    if vehicle_type in vehicle_history:
        return True
    
    # α╣Çα╕èα╣çα╕äα╕éα╕Öα╕▓α╕öα╕úα╕û (6W > JB > 4W)
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    requested_size = vehicle_sizes.get(vehicle_type, 0)
    
    # α╕½α╕▓α╕úα╕ûα╕ùα╕╡α╣êα╣âα╕½α╕ìα╣êα╕ùα╕╡α╣êα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╕¬α╕▓α╕éα╕▓α╣Çα╕äα╕óα╣âα╕èα╣ë
    max_used_size = max(vehicle_sizes.get(v, 0) for v in vehicle_history)
    
    # α╕ûα╣ëα╕▓α╕éα╕¡α╣âα╕èα╣ëα╕úα╕ûα╣Çα╕Ñα╣çα╕üα╕üα╕ºα╣êα╕▓α╕½α╕úα╕╖α╕¡α╣Çα╕ùα╣êα╕▓α╕üα╕▒α╕Üα╕ùα╕╡α╣êα╣Çα╕äα╕óα╣âα╕èα╣ë = α╣âα╕èα╣ëα╣äα╕öα╣ë
    # α╕ûα╣ëα╕▓α╕éα╕¡α╣âα╕èα╣ëα╕úα╕ûα╣âα╕½α╕ìα╣êα╕üα╕ºα╣êα╕▓α╕ùα╕╡α╣êα╣Çα╕äα╕óα╣âα╕èα╣ë = α╣âα╕èα╣ëα╣äα╕íα╣êα╣äα╕öα╣ë (α╕úα╕ûα╣âα╕½α╕ìα╣êα╕¡α╕▓α╕êα╣Çα╕éα╣ëα╕▓α╣äα╕íα╣êα╣äα╕öα╣ë)
    return requested_size <= max_used_size

def get_max_vehicle_for_branch_old(code, branch_vehicles):
    """[OLD] α╕öα╕╢α╕çα╕¢α╕úα╕░α╣Çα╕áα╕ùα╕úα╕ûα╕ùα╕╡α╣êα╣âα╕½α╕ìα╣êα╕ùα╕╡α╣êα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╕¬α╕▓α╕éα╕▓α╣Çα╕äα╕óα╣âα╕èα╣ë (α╕êα╕│α╕üα╕▒α╕öα╣äα╕íα╣êα╣âα╕½α╣ëα╣âα╕èα╣ëα╕úα╕ûα╣âα╕½α╕ìα╣êα╕üα╕ºα╣êα╕▓α╕Öα╕╡α╣ë)"""
    if not branch_vehicles or code not in branch_vehicles:
        return '6W'  # α╣äα╕íα╣êα╕íα╕╡α╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤ = α╣âα╕èα╣ëα╣äα╕öα╣ëα╕ûα╕╢α╕ç 6W
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return '6W'
    
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    max_vehicle = max(vehicle_history.keys(), key=lambda v: vehicle_sizes.get(v, 0))
    return max_vehicle

def get_most_used_vehicle_for_branch(code, branch_vehicles):
    """α╕öα╕╢α╕çα╕¢α╕úα╕░α╣Çα╕áα╕ùα╕úα╕ûα╕ùα╕╡α╣êα╕¬α╕▓α╕éα╕▓α╣âα╕èα╣ëα╕Üα╣êα╕¡α╕óα╕ùα╕╡α╣êα╕¬α╕╕α╕ö"""
    if not branch_vehicles or code not in branch_vehicles:
        return None
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return None
    
    return max(vehicle_history, key=vehicle_history.get)

def is_similar_name(name1, name2):
    """α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓α╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓α╕äα╕Ñα╣ëα╕▓α╕óα╕üα╕▒α╕Öα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê - α╕úα╕¡α╕çα╕úα╕▒α╕Üα╕ùα╕▒α╣ëα╕çα╣äα╕ùα╕óα╣üα╕Ñα╕░α╕¡α╕▒α╕çα╕üα╕ñα╕⌐ + α╕öα╕╣α╕äα╕│α╕¬α╕│α╕äα╕▒α╕ì"""
    def extract_keywords(name):
        """α╕öα╕╢α╕çα╕äα╕│α╕¬α╕│α╕äα╕▒α╕ìα╕êα╕▓α╕üα╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓"""
        if pd.isna(name) or name is None:
            return set(), "", ""
        s = str(name).strip().upper()
        
        # α╕äα╕│α╕¬α╕│α╕äα╕▒α╕ìα╕ùα╕╡α╣êα╕òα╣ëα╕¡α╕çα╕üα╕▓α╕úα╕êα╕▒α╕Üα╕äα╕╣α╣ê (α╣äα╕ùα╕ó + α╕¡α╕▒α╕çα╕üα╕ñα╕⌐)
        keywords = set()
        
        # α╣Çα╕èα╣çα╕äα╕äα╕│α╕¬α╕│α╕äα╕▒α╕ìα╣üα╕Üα╕Ü exact match
        important_words = [
            'α╕ƒα╕┤α╕ºα╣Çα╕êα╕¡α╕úα╣î', 'FUTURE', 'α╕úα╕▒α╕çα╕¬α╕┤α╕ò', 'RANGSIT',
            'α╣Çα╕ïα╣çα╕Öα╕ùα╕úα╕▒α╕Ñ', 'CENTRAL', 'α╣Çα╕ùα╕¬α╣éα╕üα╣ë', 'TESCO', 'α╣éα╕Ñα╕òα╕▒α╕¬', 'LOTUS',
            'α╕Üα╕┤α╣èα╕üα╕ïα╕╡', 'BIGC', 'α╣üα╕íα╣çα╕äα╣éα╕äα╕ú', 'MAKRO', 'α╣éα╕«α╕íα╣éα╕¢α╕ú', 'HOMEPRO',
            'α╕ïα╕╡α╕äα╕¡α╕Ö', 'SEACON', 'α╣Çα╕íα╕üα╕▓', 'MEGA', 'α╕₧α╕▓α╕úα╕▓α╣äα╕öα╕ïα╣î', 'PARADISE',
            'α╣Çα╕ùα╕¡α╕úα╣îα╕íα╕┤α╕Öα╕¡α╕Ñ', 'TERMINAL', 'α╕¬α╕óα╕▓α╕íα╕₧α╕▓α╕úα╕▓α╕üα╕¡α╕Ö', 'SIAM', 'PARAGON'
        ]
        
        for word in important_words:
            if word in s:
                keywords.add(word)
        
        # α╕Ñα╕Ü prefix/suffix α╕ùα╕╡α╣êα╕₧α╕Üα╕Üα╣êα╕¡α╕ó
        prefixes = ['PTC-MRT-', 'FC PTF ', 'PTC-', 'PTC ', 'PUN-', 'PTF ', 
                   'MAXMART', 'CW', 'FC', 'NW', 'MI', 'PI']
        for prefix in prefixes:
            if s.startswith(prefix):
                s = s[len(prefix):].strip()
                break
        
        # α╕Ñα╕Üα╕òα╕▒α╕ºα╕¡α╕▒α╕üα╕⌐α╕úα╣Çα╕öα╕╡α╣êα╕óα╕ºα╕ùα╕╡α╣êα╕éα╕╢α╣ëα╕Öα╕òα╣ëα╕Ö (M, P, N) α╕ûα╣ëα╕▓α╕òα╕▓α╕íα╕öα╣ëα╕ºα╕óα╕òα╕▒α╕ºα╣Çα╕Ñα╕é
        import re
        if re.match(r'^[MPN]\d', s):
            s = s[1:]
        
        # α╣üα╕óα╕üα╕áα╕▓α╕⌐α╕▓α╣äα╕ùα╕óα╣üα╕Ñα╕░α╕¡α╕▒α╕çα╕üα╕ñα╕⌐
        thai_chars = ''.join([c for c in s if '\u0e01' <= c <= '\u0e5b'])
        eng_chars = ''.join([c for c in s if c.isalpha() and c.isascii()])
        
        return keywords, thai_chars, eng_chars
    
    keywords1, thai1, eng1 = extract_keywords(name1)
    keywords2, thai2, eng2 = extract_keywords(name2)
    
    # ≡ƒöÑ α╕Ñα╕│α╕öα╕▒α╕Üα╣üα╕úα╕ü: α╣Çα╕èα╣çα╕äα╕äα╕│α╕¬α╕│α╕äα╕▒α╕ìα╕üα╣êα╕¡α╕Ö (α╣Çα╕èα╣êα╕Ö α╕ƒα╕┤α╕ºα╣Çα╕êα╕¡α╕úα╣î+α╕úα╕▒α╕çα╕¬α╕┤α╕ò)
    if keywords1 and keywords2:
        # α╕ûα╣ëα╕▓α╕íα╕╡α╕äα╕│α╕¬α╕│α╕äα╕▒α╕ìα╕úα╣êα╕ºα╕íα╕üα╕▒α╕Ö >= 2 α╕äα╕│ ΓåÆ α╕ûα╕╖α╕¡α╕ºα╣êα╕▓α╣Çα╕½α╕íα╕╖α╕¡α╕Öα╕üα╕▒α╕Ö
        common_keywords = keywords1 & keywords2
        if len(common_keywords) >= 2:
            return True
        # α╕ûα╣ëα╕▓α╕íα╕╡α╕äα╕│α╕¬α╕│α╕äα╕▒α╕ìα╕úα╣êα╕ºα╕íα╕üα╕▒α╕Ö 1 α╕äα╕│ α╣üα╕òα╣êα╣Çα╕¢α╣çα╕Öα╕äα╕│α╣Çα╕ëα╕₧α╕▓α╕░ ΓåÆ α╕ûα╕╖α╕¡α╕ºα╣êα╕▓α╣Çα╕½α╕íα╕╖α╕¡α╕Öα╕üα╕▒α╕Ö
        if len(common_keywords) >= 1:
            # α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓α╕íα╕╡α╕äα╕│α╕ùα╕╡α╣êα╣Çα╕¢α╣çα╕Öα╕èα╕╖α╣êα╕¡α╕¬α╕ûα╕▓α╕Öα╕ùα╕╡α╣êα╣Çα╕ëα╕₧α╕▓α╕░
            specific_places = {'α╕úα╕▒α╕çα╕¬α╕┤α╕ò', 'RANGSIT', 'α╣Çα╕ïα╣çα╕Öα╕ùα╕úα╕▒α╕Ñ', 'CENTRAL', 'α╕ïα╕╡α╕äα╕¡α╕Ö', 'SEACON'}
            if common_keywords & specific_places:
                # α╕òα╣ëα╕¡α╕çα╕íα╕╡α╕¡α╕╡α╕ü 1 α╕äα╕│ α╕½α╕úα╕╖α╕¡ α╕èα╕╖α╣êα╕¡α╕äα╕Ñα╣ëα╕▓α╕óα╕üα╕▒α╕Ö
                if len(common_keywords) >= 2 or (thai1 and thai2 and len(thai1) >= 4 and thai1[:4] in thai2):
                    return True
    
    # α╕òα╣ëα╕¡α╕çα╕íα╕╡α╕äα╕ºα╕▓α╕íα╕óα╕▓α╕ºα╕₧α╕¡α╕¬α╕íα╕äα╕ºα╕ú
    if len(thai1) < 3 and len(eng1) < 3:
        return False
    if len(thai2) < 3 and len(eng2) < 3:
        return False
    
    # α╣Çα╕èα╣çα╕äα╕áα╕▓α╕⌐α╕▓α╣äα╕ùα╕ó
    if thai1 and thai2:
        shorter_thai = min(thai1, thai2, key=len)
        longer_thai = max(thai1, thai2, key=len)
        if len(shorter_thai) >= 3 and shorter_thai in longer_thai:
            return True
        # α╕äα╕ºα╕▓α╕íα╕äα╕Ñα╣ëα╕▓α╕ó 80%+
        if len(shorter_thai) >= 5:
            common = sum(1 for c in shorter_thai if c in longer_thai)
            if common / len(shorter_thai) >= 0.8:
                return True
    
    # α╣Çα╕èα╣çα╕äα╕áα╕▓α╕⌐α╕▓α╕¡α╕▒α╕çα╕üα╕ñα╕⌐
    if eng1 and eng2:
        shorter_eng = min(eng1, eng2, key=len)
        longer_eng = max(eng1, eng2, key=len)
        if len(shorter_eng) >= 3 and shorter_eng in longer_eng:
            return True
        # α╕äα╕ºα╕▓α╕íα╕äα╕Ñα╣ëα╕▓α╕ó 80%+
        if len(shorter_eng) >= 5:
            common = sum(1 for c in shorter_eng if c in longer_eng)
            if common / len(shorter_eng) >= 0.8:
                return True
    
    # α╣Çα╕èα╣çα╕äα╕äα╕│α╕¬α╕│α╕äα╕▒α╕ìα╕úα╕░α╕½α╕ºα╣êα╕▓α╕çα╣äα╕ùα╕ó-α╕¡α╕▒α╕çα╕üα╕ñα╕⌐ (α╣Çα╕èα╣êα╕Ö Future = α╕ƒα╕┤α╕ºα╣Çα╕êα╕¡α╕úα╣î, Rangsit = α╕úα╕▒α╕çα╕¬α╕┤α╕ò)
    thai_eng_map = {
        'RANGSIT': 'α╕úα╕▒α╕çα╕¬α╕┤α╕ò',
        'FUTURE': 'α╕ƒα╕┤α╕ºα╣Çα╕êα╕¡α╕ú',
        'PARK': 'α╕¢α╕▓α╕úα╕ä',
        'TRIANGLE': 'α╣äα╕òα╕úα╣üα╕¡α╕çα╣Çα╕üα╕┤α╕Ñ',
    }
    
    for eng_word, thai_word in thai_eng_map.items():
        # α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕ºα╣êα╕▓α╕íα╕╡α╕äα╕│α╕Öα╕╡α╣ëα╣âα╕Öα╕èα╕╖α╣êα╕¡α╕ùα╕▒α╣ëα╕çα╕¬α╕¡α╕çα╕¥α╕▒α╣êα╕ç (α╣äα╕ùα╕ó-α╕¡α╕▒α╕çα╕üα╕ñα╕⌐ α╕½α╕úα╕╖α╕¡ α╕¡α╕▒α╕çα╕üα╕ñα╕⌐-α╕¡α╕▒α╕çα╕üα╕ñα╕⌐ α╕½α╕úα╕╖α╕¡ α╣äα╕ùα╕ó-α╣äα╕ùα╕ó)
        has_eng_in_1 = eng_word in eng1
        has_eng_in_2 = eng_word in eng2
        has_thai_in_1 = thai_word in thai1
        has_thai_in_2 = thai_word in thai2
        
        # α╕ûα╣ëα╕▓α╕ùα╕▒α╣ëα╕çα╕¬α╕¡α╕çα╕íα╕╡α╕äα╕│α╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Ö (α╣äα╕íα╣êα╕ºα╣êα╕▓α╕êα╕░α╣äα╕ùα╕óα╕½α╕úα╕╖α╕¡α╕¡α╕▒α╕çα╕üα╕ñα╕⌐) = α╕äα╕Ñα╣ëα╕▓α╕óα╕üα╕▒α╕Ö
        if (has_eng_in_1 and has_eng_in_2) or (has_thai_in_1 and has_thai_in_2):
            return True
        # α╕ûα╣ëα╕▓α╕éα╣ëα╕▓α╕íα╕áα╕▓α╕⌐α╕▓ (α╕¡α╕▒α╕çα╕üα╕ñα╕⌐-α╣äα╕ùα╕ó)
        if (has_eng_in_1 and has_thai_in_2) or (has_eng_in_2 and has_thai_in_1):
            return True
    
    return False

def haversine_distance(lat1, lon1, lat2, lon2):
    """
    α╕äα╕│α╕Öα╕ºα╕ôα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕úα╕░α╕½α╕ºα╣êα╕▓α╕çα╕êα╕╕α╕öα╕¬α╕¡α╕çα╕êα╕╕α╕öα╕Üα╕Öα╕₧α╕╖α╣ëα╕Öα╣éα╕Ñα╕ü (km)
    α╣âα╕èα╣ëα╕¬α╕╣α╕òα╕ú Haversine
    """
    from math import radians, sin, cos, sqrt, atan2
    
    # α╣üα╕¢α╕Ñα╕çα╕¡α╕çα╕¿α╕▓α╣Çα╕¢α╣çα╕Ö radians
    lat1_rad = radians(lat1)
    lon1_rad = radians(lon1)
    lat2_rad = radians(lat2)
    lon2_rad = radians(lon2)
    
    # α╕äα╕ºα╕▓α╕íα╕òα╣êα╕▓α╕çα╕éα╕¡α╕çα╕₧α╕┤α╕üα╕▒α╕ö
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    
    # α╕¬α╕╣α╕òα╕ú Haversine
    a = sin(dlat/2)**2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    
    # α╕úα╕▒α╕¿α╕íα╕╡α╣éα╕Ñα╕ü (km)
    R = 6371.0
    distance = R * c
    
    return distance

def get_region_type(province):
    """
    α╕üα╕│α╕½α╕Öα╕öα╕¢α╕úα╕░α╣Çα╕áα╕ùα╕₧α╕╖α╣ëα╕Öα╕ùα╕╡α╣êα╣üα╕Ñα╕░α╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕½α╕íα╕▓α╕░α╕¬α╕í
    
    Returns:
        str: 'nearby' (α╣âα╕üα╕Ñα╣ë - α╣âα╕èα╣ë 4W/JB), 'far' (α╣äα╕üα╕Ñ - α╣âα╕èα╣ë 6W), 
             'very_far' (α╣äα╕üα╕Ñα╕íα╕▓α╕ü - α╕òα╣ëα╕¡α╕ç 6W α╣Çα╕ùα╣êα╕▓α╕Öα╕▒α╣ëα╕Ö), 'unknown'
    """
    if pd.isna(province):
        return 'unknown'
    
    prov = str(province).strip()
    
    # ≡ƒÜ¢ α╕₧α╕╖α╣ëα╕Öα╕ùα╕╡α╣êα╣äα╕üα╕Ñα╕íα╕▓α╕üα╣å (α╕áα╕▓α╕äα╣Çα╕½α╕Öα╕╖α╕¡α╕òα╕¡α╕Öα╕Üα╕Ö + α╕áα╕▓α╕äα╣âα╕òα╣ëα╕Ñα╕╢α╕ü) ΓåÆ α╕òα╣ëα╕¡α╕çα╣âα╕èα╣ë 6W α╣Çα╕ùα╣êα╕▓α╕Öα╕▒α╣ëα╕Ö
    very_far_provinces = [
        # α╕áα╕▓α╕äα╣Çα╕½α╕Öα╕╖α╕¡α╕òα╕¡α╕Öα╕Üα╕Ö (α╣äα╕üα╕Ñα╕êα╕▓α╕ü DC α╕ºα╕▒α╕çα╕Öα╣ëα╕¡α╕ó ~500-700 α╕üα╕í.)
        'α╣Çα╕èα╕╡α╕óα╕çα╣âα╕½α╕íα╣ê', 'α╣Çα╕èα╕╡α╕óα╕çα╕úα╕▓α╕ó', 'α╣üα╕íα╣êα╕«α╣êα╕¡α╕çα╕¬α╕¡α╕Ö', 'α╕Öα╣êα╕▓α╕Ö', 'α╕₧α╕░α╣Çα╕óα╕▓',
        # α╕áα╕▓α╕äα╣âα╕òα╣ëα╕Ñα╕╢α╕ü (α╣äα╕üα╕Ñα╕êα╕▓α╕ü DC α╕ºα╕▒α╕çα╕Öα╣ëα╕¡α╕ó ~700-1000 α╕üα╕í.)
        'α╕¬α╕çα╕éα╕Ñα╕▓', 'α╕¢α╕▒α╕òα╕òα╕▓α╕Öα╕╡', 'α╕óα╕░α╕Ñα╕▓', 'α╕Öα╕úα╕▓α╕ÿα╕┤α╕ºα╕▓α╕¬', 'α╕₧α╕▒α╕ùα╕Ñα╕╕α╕ç', 'α╕òα╕úα╕▒α╕ç', 'α╕¬α╕òα╕╣α╕Ñ'
    ]
    
    for very_far in very_far_provinces:
        if very_far in prov:
            return 'very_far'
    
    # α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧ + α╕¢α╕úα╕┤α╕íα╕ôα╕æα╕Ñ + α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç = α╣âα╕üα╕Ñα╣ë ΓåÆ α╣âα╕èα╣ë 4W/JB
    nearby_provinces = [
        'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕íα╕½α╕▓α╕Öα╕äα╕ú', 'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧',
        'α╕Öα╕äα╕úα╕¢α╕Éα╕í', 'α╕Öα╕Öα╕ùα╕Üα╕╕α╕úα╕╡', 'α╕¢α╕ùα╕╕α╕íα╕ÿα╕▓α╕Öα╕╡', 'α╕¬α╕íα╕╕α╕ùα╕úα╕¢α╕úα╕▓α╕üα╕▓α╕ú', 'α╕¬α╕íα╕╕α╕ùα╕úα╕¬α╕▓α╕äα╕ú',
        'α╕èα╕▒α╕óα╕Öα╕▓α╕ù', 'α╕₧α╕úα╕░α╕Öα╕äα╕úα╕¿α╕úα╕╡α╕¡α╕óα╕╕α╕ÿα╕óα╕▓', 'α╕Ñα╕₧α╕Üα╕╕α╕úα╕╡', 'α╕¬α╕úα╕░α╕Üα╕╕α╕úα╕╡', 'α╕¬α╕┤α╕çα╕½α╣îα╕Üα╕╕α╕úα╕╡', 'α╕¡α╣êα╕▓α╕çα╕ùα╕¡α╕ç', 'α╕¡α╕óα╕╕α╕ÿα╕óα╕▓',
        'α╕¬α╕íα╕╕α╕ùα╕úα╕¬α╕çα╕äα╕úα╕▓α╕í', 'α╕¬α╕╕α╕₧α╕úα╕úα╕ôα╕Üα╕╕α╕úα╕╡', 'α╕Öα╕äα╕úα╕Öα╕▓α╕óα╕ü'
    ]
    
    for nearby in nearby_provinces:
        if nearby in prov:
            return 'nearby'
    
    # α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╕¡α╕╖α╣êα╕Öα╣å = α╣äα╕üα╕Ñ ΓåÆ α╣âα╕èα╣ë 6W
    return 'far'

def is_nearby_province(prov1, prov2):
    """α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣âα╕üα╕Ñα╣ëα╣Çα╕äα╕╡α╕óα╕çα╕üα╕▒α╕Öα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê (α╕êα╕▓α╕üα╣äα╕ƒα╕Ñα╣îα╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤)"""
    if pd.isna(prov1) or pd.isna(prov2):
        return False
    
    if prov1 == prov2:
        return True
    
    # α╕êα╕▒α╕öα╕üα╕Ñα╕╕α╣êα╕íα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╕òα╕▓α╕íα╕áα╕▓α╕äα╕óα╣êα╕¡α╕ó (α╕êα╕▓α╕üα╣äα╕ƒα╕Ñα╣îα╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤)
    province_groups = {
        'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧': ['α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕íα╕½α╕▓α╕Öα╕äα╕ú', 'α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧'],
        'α╕¢α╕úα╕┤α╕íα╕ôα╕æα╕Ñ': ['α╕Öα╕äα╕úα╕¢α╕Éα╕í', 'α╕Öα╕Öα╕ùα╕Üα╕╕α╕úα╕╡', 'α╕¢α╕ùα╕╕α╕íα╕ÿα╕▓α╕Öα╕╡', 'α╕¬α╕íα╕╕α╕ùα╕úα╕¢α╕úα╕▓α╕üα╕▓α╕ú', 'α╕¬α╕íα╕╕α╕ùα╕úα╕¬α╕▓α╕äα╕ú'],
        'α╕üα╕Ñα╕▓α╕çα╕òα╕¡α╕Öα╕Üα╕Ö': ['α╕èα╕▒α╕óα╕Öα╕▓α╕ù', 'α╕₧α╕úα╕░α╕Öα╕äα╕úα╕¿α╕úα╕╡α╕¡α╕óα╕╕α╕ÿα╕óα╕▓', 'α╕Ñα╕₧α╕Üα╕╕α╕úα╕╡', 'α╕¬α╕úα╕░α╕Üα╕╕α╕úα╕╡', 'α╕¬α╕┤α╕çα╕½α╣îα╕Üα╕╕α╕úα╕╡', 'α╕¡α╣êα╕▓α╕çα╕ùα╕¡α╕ç', 'α╕¡α╕óα╕╕α╕ÿα╕óα╕▓'],
        'α╕üα╕Ñα╕▓α╕çα╕òα╕¡α╕Öα╕Ñα╣êα╕▓α╕ç': ['α╕¬α╕íα╕╕α╕ùα╕úα╕¬α╕çα╕äα╕úα╕▓α╕í', 'α╕¬α╕╕α╕₧α╕úα╕úα╕ôα╕Üα╕╕α╕úα╕╡'],
        'α╕áα╕▓α╕äα╕òα╕░α╕ºα╕▒α╕Öα╕òα╕ü': ['α╕üα╕▓α╕ìα╕êα╕Öα╕Üα╕╕α╕úα╕╡', 'α╕¢α╕úα╕░α╕êα╕ºα╕Üα╕äα╕╡α╕úα╕╡α╕éα╕▒α╕Öα╕ÿα╣î', 'α╕úα╕▓α╕èα╕Üα╕╕α╕úα╕╡', 'α╣Çα╕₧α╕èα╕úα╕Üα╕╕α╕úα╕╡'],
        'α╕áα╕▓α╕äα╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü': ['α╕êα╕▒α╕Öα╕ùα╕Üα╕╕α╕úα╕╡', 'α╕èα╕Ñα╕Üα╕╕α╕úα╕╡', 'α╕òα╕úα╕▓α╕ö', 'α╕Öα╕äα╕úα╕Öα╕▓α╕óα╕ü', 'α╕¢α╕úα╕▓α╕êα╕╡α╕Öα╕Üα╕╕α╕úα╕╡', 'α╕úα╕░α╕óα╕¡α╕ç', 'α╕¬α╕úα╕░α╣üα╕üα╣ëα╕º', 'α╕ëα╕░α╣Çα╕èα╕┤α╕çα╣Çα╕ùα╕úα╕▓'],
        'α╕¡α╕╡α╕¬α╕▓α╕Öα╣Çα╕½α╕Öα╕╖α╕¡': ['α╕Öα╕äα╕úα╕₧α╕Öα╕í', 'α╕Üα╕╢α╕çα╕üα╕▓α╕¼', 'α╕íα╕╕α╕üα╕öα╕▓α╕½α╕▓α╕ú', 'α╕¬α╕üα╕Ñα╕Öα╕äα╕ú', 'α╕½α╕Öα╕¡α╕çα╕äα╕▓α╕ó', 'α╕½α╕Öα╕¡α╕çα╕Üα╕▒α╕ºα╕Ñα╕│α╕áα╕╣', 'α╕¡α╕╕α╕öα╕úα╕ÿα╕▓α╕Öα╕╡', 'α╣Çα╕Ñα╕ó'],
        'α╕¡α╕╡α╕¬α╕▓α╕Öα╕üα╕Ñα╕▓α╕ç': ['α╕üα╕▓α╕¼α╕¬α╕┤α╕Öα╕ÿα╕╕α╣î', 'α╕éα╕¡α╕Öα╣üα╕üα╣êα╕Ö', 'α╕èα╕▒α╕óα╕áα╕╣α╕íα╕┤', 'α╕íα╕½α╕▓α╕¬α╕▓α╕úα╕äα╕▓α╕í', 'α╕úα╣ëα╕¡α╕óα╣Çα╕¡α╣çα╕ö'],
        'α╕¡α╕╡α╕¬α╕▓α╕Öα╣âα╕òα╣ë': ['α╕Öα╕äα╕úα╕úα╕▓α╕èα╕¬α╕╡α╕íα╕▓', 'α╣éα╕äα╕úα╕▓α╕è', 'α╕Üα╕╕α╕úα╕╡α╕úα╕▒α╕íα╕óα╣î', 'α╕óα╣éα╕¬α╕ÿα╕ú', 'α╕¿α╕úα╕╡α╕¬α╕░α╣Çα╕üα╕⌐', 'α╕¬α╕╕α╕úα╕┤α╕Öα╕ùα╕úα╣î', 'α╕¡α╕│α╕Öα╕▓α╕êα╣Çα╕êα╕úα╕┤α╕ì', 'α╕¡α╕╕α╕Üα╕Ñα╕úα╕▓α╕èα╕ÿα╕▓α╕Öα╕╡'],
        'α╣Çα╕½α╕Öα╕╖α╕¡α╕òα╕¡α╕Öα╕Üα╕Ö': ['α╕Öα╣êα╕▓α╕Ö', 'α╕₧α╕░α╣Çα╕óα╕▓', 'α╕Ñα╕│α╕¢α╕▓α╕ç', 'α╕Ñα╕│α╕₧α╕╣α╕Ö', 'α╣Çα╕èα╕╡α╕óα╕çα╕úα╕▓α╕ó', 'α╣Çα╕èα╕╡α╕óα╕çα╣âα╕½α╕íα╣ê', 'α╣üα╕₧α╕úα╣ê', 'α╣üα╕íα╣êα╕«α╣êα╕¡α╕çα╕¬α╕¡α╕Ö'],
        'α╣Çα╕½α╕Öα╕╖α╕¡α╕òα╕¡α╕Öα╕Ñα╣êα╕▓α╕ç': ['α╕üα╕│α╣üα╕₧α╕çα╣Çα╕₧α╕èα╕ú', 'α╕òα╕▓α╕ü', 'α╕Öα╕äα╕úα╕¬α╕ºα╕úα╕úα╕äα╣î', 'α╕₧α╕┤α╕êα╕┤α╕òα╕ú', 'α╕₧α╕┤α╕⌐α╕ôα╕╕α╣éα╕Ñα╕ü', 'α╕¬α╕╕α╣éα╕éα╕ùα╕▒α╕ó', 'α╕¡α╕╕α╕òα╕úα╕öα╕┤α╕òα╕ûα╣î', 'α╕¡α╕╕α╕ùα╕▒α╕óα╕ÿα╕▓α╕Öα╕╡', 'α╣Çα╕₧α╕èα╕úα╕Üα╕╣α╕úα╕ôα╣î'],
        'α╣âα╕òα╣ëα╕¥α╕▒α╣êα╕çα╕¡α╕▒α╕Öα╕öα╕▓α╕íα╕▒α╕Ö': ['α╕üα╕úα╕░α╕Üα╕╡α╣ê', 'α╕òα╕úα╕▒α╕ç', 'α╕₧α╕▒α╕çα╕çα╕▓', 'α╕áα╕╣α╣Çα╕üα╣çα╕ò', 'α╕úα╕░α╕Öα╕¡α╕ç', 'α╕¬α╕òα╕╣α╕Ñ'],
        'α╣âα╕òα╣ëα╕¥α╕▒α╣êα╕çα╕¡α╣êα╕▓α╕ºα╣äα╕ùα╕ó': ['α╕èα╕╕α╕íα╕₧α╕ú', 'α╕Öα╕äα╕úα╕¿α╕úα╕╡α╕ÿα╕úα╕úα╕íα╕úα╕▓α╕è', 'α╕₧α╕▒α╕ùα╕Ñα╕╕α╕ç', 'α╕óα╕░α╕Ñα╕▓', 'α╕¬α╕çα╕éα╕Ñα╕▓', 'α╕¬α╕╕α╕úα╕▓α╕⌐α╕Äα╕úα╣îα╕ÿα╕▓α╕Öα╕╡', 'α╕¢α╕▒α╕òα╕òα╕▓α╕Öα╕╡', 'α╕Öα╕úα╕▓α╕ÿα╕┤α╕ºα╕▓α╕¬']
    }
    
    # α╕½α╕▓α╕ºα╣êα╕▓α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╕ùα╕▒α╣ëα╕ç 2 α╕¡α╕óα╕╣α╣êα╕üα╕Ñα╕╕α╣êα╕íα╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Öα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê
    for group, provinces in province_groups.items():
        in_group_1 = any(p in str(prov1) for p in provinces)
        in_group_2 = any(p in str(prov2) for p in provinces)
        
        if in_group_1 and in_group_2:
            return True
    
    return False

def load_model():
    """α╣éα╕½α╕Ñα╕öα╣éα╕íα╣Çα╕öα╕Ñα╕ùα╕╡α╣êα╣Çα╕ùα╕úα╕Öα╣äα╕ºα╣ë"""
    if not os.path.exists(MODEL_PATH):
        return None
    
    try:
        with open(MODEL_PATH, 'rb') as f:
            model_data = pickle.load(f)
        return model_data
    except Exception as e:
        st.error(f"Γ¥î Error loading model: {e}")
        return None

def create_pair_features(code1, code2, branch_info):
    """α╕¬α╕úα╣ëα╕▓α╕ç features α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕äα╕╣α╣êα╕¬α╕▓α╕éα╕▓"""
    import math
    
    info1 = branch_info[code1]
    info2 = branch_info[code2]
    
    # α╕äα╕│α╕Öα╕ºα╕ôα╕äα╕ºα╕▓α╕íα╕òα╣êα╕▓α╕çα╕éα╕¡α╕çα╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕üα╣üα╕Ñα╕░α╕äα╕┤α╕º
    weight_diff = abs(info1['avg_weight'] - info2['avg_weight'])
    cube_diff = abs(info1['avg_cube'] - info2['avg_cube'])
    weight_sum = info1['avg_weight'] + info2['avg_weight']
    cube_sum = info1['avg_cube'] + info2['avg_cube']
    
    # α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Öα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê
    same_province = 1 if info1['province'] == info2['province'] else 0
    
    # α╕äα╕│α╕Öα╕ºα╕ôα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕êα╕▓α╕üα╕₧α╕┤α╕üα╕▒α╕ö
    distance_km = 0.0
    if info1['latitude'] != 0 and info2['latitude'] != 0:
        lat1, lon1 = math.radians(info1['latitude']), math.radians(info1['longitude'])
        lat2, lon2 = math.radians(info2['latitude']), math.radians(info2['longitude'])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        distance_km = 6371 * c
    
    # α╕äα╕ºα╕▓α╕íα╕ûα╕╡α╣ê
    freq_product = info1['total_trips'] * info2['total_trips']
    freq_diff = abs(info1['total_trips'] - info2['total_trips'])
    
    # Ratio
    weight_ratio = (info1['avg_weight'] / info2['avg_weight']) if info2['avg_weight'] > 0 else 0
    cube_ratio = (info1['avg_cube'] / info2['avg_cube']) if info2['avg_cube'] > 0 else 0
    
    # α╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕úα╕û
    over_4w = 1 if (weight_sum > 2500 or cube_sum > 5.0) else 0
    over_jb = 1 if (weight_sum > 3500 or cube_sum > 8.0) else 0
    over_6w = 1 if (weight_sum > 5800 or cube_sum > 22.0) else 0
    
    return {
        'weight_sum': weight_sum,
        'cube_sum': cube_sum,
        'weight_diff': weight_diff,
        'cube_diff': cube_diff,
        'same_province': same_province,
        'distance_km': distance_km,
        'avg_weight_1': info1['avg_weight'],
        'avg_weight_2': info2['avg_weight'],
        'avg_cube_1': info1['avg_cube'],
        'avg_cube_2': info2['avg_cube'],
        'freq_product': freq_product,
        'freq_diff': freq_diff,
        'weight_ratio': weight_ratio,
        'cube_ratio': cube_ratio,
        'over_4w': over_4w,
        'over_jb': over_jb,
        'over_6w': over_6w
    }

def load_excel(file_content, sheet_name=None):
    """α╣éα╕½α╕Ñα╕ö Excel"""
    try:
        xls = pd.ExcelFile(io.BytesIO(file_content))
        
        target_sheet = None
        if sheet_name and sheet_name in xls.sheet_names:
            target_sheet = sheet_name
        else:
            for s in xls.sheet_names:
                if 'punthai' in s.lower() or '2.' in s.lower():
                    target_sheet = s
                    break
        
        if not target_sheet:
            target_sheet = xls.sheet_names[0]
        
        # α╕½α╕▓ header row
        df_temp = pd.read_excel(xls, sheet_name=target_sheet, header=None)
        header_row = 0
        
        for i in range(min(10, len(df_temp))):
            row_values = df_temp.iloc[i].astype(str).str.upper()
            match_count = sum([
                'BRANCH' in ' '.join(row_values),
                'TRIP' in ' '.join(row_values),
                'α╕úα╕½α╕▒α╕¬α╕¬α╕▓α╕éα╕▓' in ' '.join(df_temp.iloc[i].astype(str))
            ])
            if match_count >= 2:
                header_row = i
                break
        
        df = pd.read_excel(xls, sheet_name=target_sheet, header=header_row)
        df = df.loc[:, ~df.columns.duplicated()]
        
        return df
    except Exception as e:
        st.error(f"Γ¥î Error: {e}")
        return None

def process_dataframe(df):
    """α╣üα╕¢α╕Ñα╕çα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╣Çα╕¢α╣çα╕Öα╕úα╕╣α╕¢α╣üα╕Üα╕Üα╕íα╕▓α╕òα╕úα╕Éα╕▓α╕Ö"""
    if df is None:
        return None
    
    rename_map = {}
    
    # α╕ûα╣ëα╕▓α╕íα╕╡α╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕Öα╣ëα╕¡α╕óα╕üα╕ºα╣êα╕▓ 15 = α╣âα╕èα╣ëα╕Ñα╕│α╕öα╕▒α╕Üα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣î
    # α╕Ñα╕│α╕öα╕▒α╕Üα╕íα╕▓α╕òα╕úα╕Éα╕▓α╕Ö: Sep, BU, α╕úα╕½α╕▒α╕¬α╕¬α╕▓α╕éα╕▓, α╕úα╕½α╕▒α╕¬ WMS, α╕¬α╕▓α╕éα╕▓, Total Cube, Total Wgt, α╕êα╕│α╕Öα╕ºα╕Öα╕èα╕┤α╣ëα╕Ö, Trip, Trip no, ...
    if len(df.columns) >= 8:
        col_list = list(df.columns)
        # α╕Ñα╕│α╕öα╕▒α╕Ü 2 = α╕úα╕½α╕▒α╕¬α╕¬α╕▓α╕éα╕▓
        if len(col_list) > 2:
            rename_map[col_list[2]] = 'Code'
        # α╕Ñα╕│α╕öα╕▒α╕Ü 4 = α╕¬α╕▓α╕éα╕▓/α╕èα╕╖α╣êα╕¡
        if len(col_list) > 4:
            rename_map[col_list[4]] = 'Name'
        # α╕Ñα╕│α╕öα╕▒α╕Ü 5 = Total Cube
        if len(col_list) > 5:
            rename_map[col_list[5]] = 'Cube'
        # α╕Ñα╕│α╕öα╕▒α╕Ü 6 = Total Wgt
        if len(col_list) > 6:
            rename_map[col_list[6]] = 'Weight'
        # α╕Ñα╕│α╕öα╕▒α╕Ü 8 = Trip
        if len(col_list) > 8:
            rename_map[col_list[8]] = 'Trip'
        # α╕Ñα╕│α╕öα╕▒α╕Ü 9 = Trip no
        if len(col_list) > 9:
            rename_map[col_list[9]] = 'TripNo'
    
    # α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╣Çα╕₧α╕┤α╣êα╕íα╣Çα╕òα╕┤α╕íα╕êα╕▓α╕üα╕èα╕╖α╣êα╕¡α╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣î
    for col in df.columns:
        if col in rename_map:
            continue
        col_clean = str(col).strip()
        col_upper = col_clean.upper().replace(' ', '').replace('_', '')
        
        if col_clean == 'BranchCode' or 'α╕úα╕½α╕▒α╕¬α╕¬α╕▓α╕éα╕▓' in col_clean or col_clean == 'α╕úα╕½α╕▒α╕¬ WMS' or 'BRANCH_CODE' in col_upper:
            rename_map[col] = 'Code'
        elif col_clean == 'Branch' or 'α╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓' in col_clean or col_clean == 'α╕¬α╕▓α╕éα╕▓' or 'BRANCH' in col_upper:
            rename_map[col] = 'Name'
        elif 'TOTALWGT' in col_upper or 'α╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü' in col_clean or 'WGT' in col_upper or 'WEIGHT' in col_upper:
            rename_map[col] = 'Weight'
        elif 'TOTALCUBE' in col_upper or 'α╕äα╕┤α╕º' in col_clean or 'CUBE' in col_upper:
            rename_map[col] = 'Cube'
        elif 'latitude' in col_clean.lower() or col_clean == 'α╕Ñα╕░α╕òα╕┤α╕êα╕╣α╕ö' or 'LAT' in col_upper:
            rename_map[col] = 'Latitude'
        elif 'longitude' in col_clean.lower() or col_clean == 'α╕Ñα╕¡α╕çα╕òα╕┤α╕êα╕╣α╕ö' or 'LONG' in col_upper or 'LNG' in col_upper:
            rename_map[col] = 'Longitude'
        elif 'α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö' in col_clean or 'PROVINCE' in col_upper:
            rename_map[col] = 'Province'
        elif col_upper in ['TRIPNO', 'TRIP_NO'] or col_clean == 'Trip no':
            rename_map[col] = 'TripNo'
        elif col_upper == 'TRIP' or 'α╕ùα╕úα╕┤α╕¢' in col_clean or 'α╣Çα╕ùα╕╡α╣êα╕óα╕º' in col_clean:
            rename_map[col] = 'Trip'
        elif 'BOOKING' in col_upper:
            rename_map[col] = 'Booking'
    
    df = df.rename(columns=rename_map)
    
    # α╕Ñα╕Üα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕ïα╣ëα╕│
    df = df.loc[:, ~df.columns.duplicated()]
    
    if 'Code' in df.columns:
        df['Code'] = df['Code'].apply(normalize)
        
        # α╕òα╕▒α╕öα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣äα╕íα╣êα╕òα╣ëα╕¡α╕çα╕üα╕▓α╕úα╕¡α╕¡α╕ü (α╕úα╕½α╕▒α╕¬)
        df = df[~df['Code'].isin(EXCLUDE_BRANCHES)]
        
        # α╕òα╕▒α╕öα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╕èα╕╖α╣êα╕¡α╕íα╕╡ keyword α╕ùα╕╡α╣êα╣äα╕íα╣êα╕òα╣ëα╕¡α╕çα╕üα╕▓α╕ú
        if 'Name' in df.columns:
            exclude_pattern = '|'.join(EXCLUDE_NAMES)
            df = df[~df['Name'].str.contains(exclude_pattern, case=False, na=False)]
    
    for col in ['Weight', 'Cube']:
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    
    # α╣Çα╕₧α╕┤α╣êα╕íα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╕êα╕▓α╕ü Master α╕ûα╣ëα╕▓α╕óα╕▒α╕çα╣äα╕íα╣êα╕íα╕╡
    if 'Province' not in df.columns or df['Province'].isna().all():
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns and 'Code' in df.columns:
            # α╕¬α╕úα╣ëα╕▓α╕ç mapping α╕êα╕▓α╕ü Master
            province_map = {}
            for _, row in MASTER_DATA.iterrows():
                code = row.get('Plan Code', '')
                province = row.get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', '')
                if code and province:
                    province_map[code] = province
            
            # α╕ƒα╕▒α╕çα╕üα╣îα╕èα╕▒α╕Öα╕äα╣ëα╕Öα╕½α╕▓α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╕êα╕▓α╕üα╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓
            def find_province_by_name(code, name):
                # α╕Ñα╕¡α╕çα╕½α╕▓α╕êα╕▓α╕ü code α╕üα╣êα╕¡α╕Ö
                if code in province_map:
                    return province_map[code]
                
                # α╕ûα╣ëα╕▓α╣äα╕íα╣êα╣Çα╕êα╕¡ α╕Ñα╕¡α╕çα╕äα╣ëα╕Öα╕½α╕▓α╕êα╕▓α╕üα╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓
                if not name or pd.isna(name):
                    return None
                
                # α╣üα╕óα╕üα╕äα╕│α╕¬α╕│α╕äα╕▒α╕ìα╕êα╕▓α╕üα╕èα╕╖α╣êα╕¡ (α╣Çα╕¡α╕▓α╕äα╕│α╣üα╕úα╕üα╕ùα╕╡α╣êα╣äα╕íα╣êα╣âα╕èα╣ê prefix)
                keywords = str(name).replace('MAX MART-', '').replace('PUNTHAI-', '').replace('LUBE', '').strip()
                if not keywords:
                    return None
                
                # α╕äα╣ëα╕Öα╕½α╕▓α╣âα╕Öα╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓α╕éα╕¡α╕ç Master
                for _, master_row in MASTER_DATA.iterrows():
                    master_name = str(master_row.get('α╕¬α╕▓α╕éα╕▓', ''))
                    # α╕ûα╣ëα╕▓α╕èα╕╖α╣êα╕¡α╕äα╕Ñα╣ëα╕▓α╕óα╕üα╕▒α╕Ö (α╕íα╕╡α╕äα╕│α╕¬α╕│α╕äα╕▒α╕ìα╣Çα╕½α╕íα╕╖α╕¡α╕Öα╕üα╕▒α╕Ö)
                    if keywords[:10] in master_name or master_name[:10] in keywords:
                        province = master_row.get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', '')
                        if province:
                            return province
                
                return None
            
            # α╣âα╕¬α╣êα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣âα╕½α╣ëα╣üα╕òα╣êα╕Ñα╕░α╕¬α╕▓α╕éα╕▓
            if 'Name' in df.columns:
                df['Province'] = df.apply(lambda row: find_province_by_name(row['Code'], row.get('Name', '')), axis=1)
            else:
                df['Province'] = df['Code'].map(province_map)
    
    return df.reset_index(drop=True)

def predict_trips(test_df, model_data, punthai_buffer=1.0, maxmart_buffer=1.10):
    """
    α╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢α╣üα╕Üα╕Üα╣âα╕½α╕íα╣ê - α╣Çα╕úα╕╡α╕óα╕Üα╕çα╣êα╕▓α╕óα╣üα╕Ñα╕░α╕íα╕╡α╕¢α╕úα╕░α╕¬α╕┤α╕ùα╕ÿα╕┤α╕áα╕▓α╕₧
    
    α╕½α╕Ñα╕▒α╕üα╕üα╕▓α╕ú:
    1. α╣Çα╕úα╕╡α╕óα╕çα╕òα╕▓α╕í: α╕áα╕▓α╕ä ΓåÆ α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö ΓåÆ α╕¡α╕│α╣Çα╕áα╕¡ ΓåÆ α╕òα╕│α╕Üα╕Ñ ΓåÆ Route (α╣âα╕èα╣ëα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕êα╕▓α╕ü Master Dist.xlsx α╣äα╕íα╣êα╣âα╕èα╣êα╕òα╕▒α╕ºα╕¡α╕▒α╕üα╕⌐α╕ú)
    2. α╕êα╕▒α╕Üα╕üα╕Ñα╕╕α╣êα╕í Route α╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Ö α╕úα╕ºα╕íα╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕üα╣äα╕ºα╣ëα╕öα╣ëα╕ºα╕óα╕üα╕▒α╕Ö
    3. α╣Çα╕úα╕╡α╕óα╕çα╕êα╕▓α╕üα╣äα╕üα╕Ñα╕íα╕▓α╣âα╕üα╕Ñα╣ë (α╕êα╕▓α╕ü DC)
    4. α╕òα╕▒α╕öα╣Çα╕¢α╣çα╕Öα╕ùα╕úα╕┤α╕¢α╕òα╕▓α╕íα╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü/α╕äα╕┤α╕ºα╕éα╕¡α╕çα╕úα╕ûα╣üα╕òα╣êα╕Ñα╕░α╕¢α╕úα╕░α╣Çα╕áα╕ù
    5. α╣âα╕èα╣ë BUFFER α╕òα╕▓α╕í BU (α╕òα╕úα╕ºα╕êα╕êα╕▓α╕üα╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓)
    
    Args:
        test_df: DataFrame α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╕êα╕░α╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢
        model_data: α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╣éα╕íα╣Çα╕öα╕Ñ (branch_vehicles, etc.)
        punthai_buffer: Buffer α╕¬α╕│α╕½α╕úα╕▒α╕Ü Punthai (α╣Çα╕èα╣êα╕Ö 1.0 = 100%)
        maxmart_buffer: Buffer α╕¬α╕│α╕½α╕úα╕▒α╕Ü Maxmart/α╕£α╕¬α╕í (α╣Çα╕èα╣êα╕Ö 1.10 = 110%)
    """
    branch_vehicles = model_data.get('branch_vehicles', {})
    
    # ==========================================
    # Step 1: α╣Çα╕òα╕úα╕╡α╕óα╕í Master Dist Lookup (Join_Key ΓåÆ Sort_Code)
    # α╕½α╕Ñα╕▒α╕üα╕üα╕▓α╕ú: α╣âα╕èα╣ë Join_Key (α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö_α╕¡α╕│α╣Çα╕áα╕¡_α╕òα╕│α╕Üα╕Ñ) α╣Çα╕¢α╣çα╕Öα╕òα╕▒α╕ºα╣Çα╕èα╕╖α╣êα╕¡α╕í
    # α╣Çα╕₧α╕╖α╣êα╕¡α╕öα╕╢α╕ç Sum_Code (Sort_Code) α╕íα╕▓α╣âα╕èα╣ëα╣âα╕Öα╕üα╕▓α╕úα╣Çα╕úα╕╡α╕óα╕çα╕Ñα╕│α╕öα╕▒α╕Ü
    # ==========================================
    subdistrict_dist_lookup = {}  # {Join_Key: {sum_code, dist_from_dc, ...}}
    if MASTER_DIST_DATA and 'by_name' in MASTER_DIST_DATA:
        subdistrict_dist_lookup = MASTER_DIST_DATA['by_name']
    
    # α╕¬α╕úα╣ëα╕▓α╕ç location_map α╕êα╕▓α╕ü MASTER_DATA (α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕¬α╕▓α╕éα╕▓)
    location_map = {}  # {code: {province, district, subdistrict, route, sum_code, ...}}
    
    if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
        for _, row in MASTER_DATA.iterrows():
            code = str(row.get('Plan Code', '')).strip().upper()
            if not code:
                continue
            
            province = str(row.get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', '')).strip() if pd.notna(row.get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö')) else ''
            district = str(row.get('α╕¡α╕│α╣Çα╕áα╕¡', '')).strip() if pd.notna(row.get('α╕¡α╕│α╣Çα╕áα╕¡')) else ''
            subdistrict = str(row.get('α╕òα╕│α╕Üα╕Ñ', '')).strip() if pd.notna(row.get('α╕òα╕│α╕Üα╕Ñ')) else ''
            route = str(row.get('Reference', '')).strip() if pd.notna(row.get('Reference')) else ''
            lat = float(row.get('α╕Ñα╕░α╕òα╕┤α╕êα╕╣α╕ö', 0)) if pd.notna(row.get('α╕Ñα╕░α╕òα╕┤α╕êα╕╣α╕ö')) else 0
            lon = float(row.get('α╕Ñα╕¡α╕çα╕òα╕┤α╕êα╕╣α╕ö', 0)) if pd.notna(row.get('α╕Ñα╕¡α╕çα╕òα╕┤α╕êα╕╣α╕ö')) else 0
            
            # ≡ƒöæ α╕¬α╕úα╣ëα╕▓α╕ç Join_Key α╣Çα╕₧α╕╖α╣êα╕¡α╣Çα╕ùα╕╡α╕óα╕Üα╕üα╕▒α╕Ü Master Dist (VLOOKUP)
            prov_clean = clean_name(province)
            dist_clean = clean_name(district)
            subdist_clean = clean_name(subdistrict)
            join_key = f"{prov_clean}_{dist_clean}_{subdist_clean}"
            
            # α╕Ñα╕¡α╕çα╕½α╕Ñα╕▓α╕ó key α╣Çα╕£α╕╖α╣êα╕¡α╕èα╕╖α╣êα╕¡α╣äα╕íα╣êα╕òα╕úα╕ç
            dist_data = subdistrict_dist_lookup.get(join_key, {})
            if not dist_data:
                # α╕Ñα╕¡α╕ç normalize α╕èα╕╖α╣êα╕¡α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö
                prov_normalized = normalize_province_name(province)
                alt_key = f"{prov_normalized}_{dist_clean}_{subdist_clean}"
                dist_data = subdistrict_dist_lookup.get(alt_key, {})
            
            # α╕öα╕╢α╕çα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕êα╕▓α╕ü Master Dist (α╕ûα╣ëα╕▓α╕íα╕╡)
            if dist_data:
                sum_code = dist_data.get('sum_code', '')  # ≡ƒÄ» Sort_Code α╕½α╕Ñα╕▒α╕ü!
                dist_from_dc = dist_data.get('dist_from_dc_km', 9999)
                region_code = dist_data.get('region_code', '')
                prov_code = dist_data.get('prov_code', '')
                dist_code_val = dist_data.get('dist_code', '')
                subdist_code = dist_data.get('subdist_code', '')
            else:
                # Fallback: α╕¬α╕úα╣ëα╕▓α╕ç sort_code α╕êα╕▓α╕ü region code α╣üα╕Ñα╕░α╕äα╕│α╕Öα╕ºα╕ôα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕êα╕▓α╕ü lat/lon
                region_code = get_region_code(province)
                sum_code = f"R99P999D9999S99999"  # Default α╕¬α╕│α╕½α╕úα╕▒α╕Üα╣äα╕íα╣êα╕₧α╕Ü
                dist_from_dc = 9999
                if lat and lon:
                    dist_from_dc = haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
                prov_code = 'P999'
                dist_code_val = 'D9999'
                subdist_code = 'S99999'
            
            region_name = get_region_name(province)
            
            location_map[code] = {
                'province': province,
                'district': district,
                'subdistrict': subdistrict,
                'route': route,
                'lat': lat,
                'lon': lon,
                'join_key': join_key,  # ≡ƒöæ Join_Key α╕ùα╕╡α╣êα╣âα╕èα╣ë lookup
                'sum_code': sum_code,  # ≡ƒÄ» Sort_Code α╕½α╕Ñα╕▒α╕ü (α╕êα╕▓α╕ü Master Dist)
                'distance_from_dc': dist_from_dc,
                'region_code': region_code,
                'prov_code': prov_code,
                'dist_code': dist_code_val,
                'subdist_code': subdist_code,
                'region_name': region_name
            }
    
    # ==========================================
    # Step 2: α╣Çα╕₧α╕┤α╣êα╕íα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕₧α╕╖α╣ëα╕Öα╕ùα╕╡α╣êα╣âα╕½α╣ëα╣üα╕òα╣êα╕Ñα╕░α╕¬α╕▓α╕éα╕▓ (pd.merge α╣üα╕Üα╕Ü manual)
    # ==========================================
    df = test_df.copy()
    
    def get_location_info(code):
        code_upper = str(code).strip().upper()
        return location_map.get(code_upper, {
            'province': '', 'district': '', 'subdistrict': '', 'route': '',
            'lat': 0, 'lon': 0, 'join_key': '', 
            'sum_code': 'R99P999D9999S99999',  # Default sort_code
            'distance_from_dc': 9999,
            'region_code': 'R99', 'prov_code': 'P999', 'dist_code': 'D9999', 'subdist_code': 'S99999',
            'region_name': 'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕'
        })
    
    # α╣Çα╕₧α╕┤α╣êα╕íα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕₧α╕╖α╣ëα╕Öα╕ùα╕╡α╣ê (α╕úα╕ºα╕í sum_code α╕¬α╕│α╕½α╕úα╕▒α╕Ü sort)
    df['_sum_code'] = df['Code'].apply(lambda c: get_location_info(c)['sum_code'])  # ≡ƒÄ» Sort_Code!
    df['_join_key'] = df['Code'].apply(lambda c: get_location_info(c)['join_key'])
    df['_region_code'] = df['Code'].apply(lambda c: get_location_info(c)['region_code'])
    df['_region_name'] = df['Code'].apply(lambda c: get_location_info(c)['region_name'])
    df['_prov_code'] = df['Code'].apply(lambda c: get_location_info(c)['prov_code'])
    df['_dist_code'] = df['Code'].apply(lambda c: get_location_info(c)['dist_code'])
    df['_subdist_code'] = df['Code'].apply(lambda c: get_location_info(c)['subdist_code'])
    df['_province'] = df['Code'].apply(lambda c: get_location_info(c)['province'])
    df['_district'] = df['Code'].apply(lambda c: get_location_info(c)['district'])
    df['_subdistrict'] = df['Code'].apply(lambda c: get_location_info(c)['subdistrict'])
    df['_route'] = df['Code'].apply(lambda c: get_location_info(c)['route'])
    df['_distance_from_dc'] = df['Code'].apply(lambda c: get_location_info(c)['distance_from_dc'])
    
    # ==========================================
    # Step 3: α╣Çα╕úα╕╡α╕óα╕çα╕Ñα╕│α╕öα╕▒α╕Üα╣üα╕Üα╕Ü Hierarchical (Region > Province Max Dist > District Max Dist > Distance)
    # ≡ƒÄ» α╕½α╕▒α╕ºα╣âα╕êα╕¬α╕│α╕äα╕▒α╕ì: α╣Çα╕úα╕╡α╕óα╕çα╕òα╕▓α╕í Region Order α╕üα╣êα╕¡α╕Ö (α╣äα╕üα╕Ñα╕íα╕▓α╣âα╕üα╕Ñα╣ë)
    # ==========================================
    
    # α╣Çα╕₧α╕┤α╣êα╕í Region Order α╕¬α╕│α╕½α╕úα╕▒α╕Ü sorting
    df['_region_order'] = df['_region_name'].map(REGION_ORDER).fillna(99)
    
    # α╕äα╕│α╕Öα╕ºα╕ô Province Max Distance (α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣äα╕½α╕Öα╕íα╕╡α╕êα╕╕α╕öα╣äα╕üα╕Ñα╕¬α╕╕α╕öα╕íα╕▓α╕üα╣êα╕¡α╕Ö)
    prov_max_dist = df.groupby('_province')['_distance_from_dc'].max().reset_index()
    prov_max_dist.columns = ['_province', '_prov_max_dist']
    df = df.merge(prov_max_dist, on='_province', how='left')
    df['_prov_max_dist'] = df['_prov_max_dist'].fillna(9999)
    
    # α╕äα╕│α╕Öα╕ºα╕ô District Max Distance (α╕¡α╕│α╣Çα╕áα╕¡α╣äα╕½α╕Öα╕íα╕╡α╕êα╕╕α╕öα╣äα╕üα╕Ñα╕¬α╕╕α╕öα╕íα╕▓α╕üα╣êα╕¡α╕Ö)
    dist_max_dist = df.groupby(['_province', '_district'])['_distance_from_dc'].max().reset_index()
    dist_max_dist.columns = ['_province', '_district', '_dist_max_dist']
    df = df.merge(dist_max_dist, on=['_province', '_district'], how='left')
    df['_dist_max_dist'] = df['_dist_max_dist'].fillna(9999)
    
    # Sort: Region Order (Asc) ΓåÆ Prov Max Dist (Desc) ΓåÆ Dist Max Dist (Desc) ΓåÆ Sum_Code ΓåÆ Distance (Desc)
    df = df.sort_values(
        ['_region_order', '_prov_max_dist', '_dist_max_dist', '_sum_code', '_route', '_distance_from_dc'],
        ascending=[True, False, False, True, True, False]  # Region/Province/District α╣äα╕üα╕Ñα╕íα╕▓α╕üα╣êα╕¡α╕Ö
    ).reset_index(drop=True)
    
    # ==========================================
    # Step 4: α╕êα╕▒α╕Üα╕üα╕Ñα╕╕α╣êα╕í Route α╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Ö α╕úα╕ºα╕íα╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü
    # ==========================================
    # α╕¬α╕úα╣ëα╕▓α╕ç grouping key α╕êα╕▓α╕ü route (α╕ûα╣ëα╕▓α╕íα╕╡) α╕½α╕úα╕╖α╕¡ α╕òα╕│α╕Üα╕Ñ+α╕¡α╕│α╣Çα╕áα╕¡+α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö
    def get_group_key(row):
        route = row['_route']
        if route and route.strip():
            return f"R_{route}"
        # α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡ route α╣âα╕èα╣ë α╕úα╕½α╕▒α╕¬α╕òα╕│α╕Üα╕Ñ (α╣Çα╕úα╕╡α╕óα╕çα╕òα╕▓α╕íα╕úα╕░α╕óα╕░α╕ùα╕▓α╕ç)
        return f"L_{row['_subdist_code']}_{row['_dist_code']}_{row['_prov_code']}"
    
    df['_group_key'] = df.apply(get_group_key, axis=1)
    
    # ==========================================
    # Step 5: α╕½α╕▓α╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕½α╕íα╕▓α╕░α╕¬α╕íα╕êα╕▓α╕üα╕éα╣ëα╕¡α╕êα╕│α╕üα╕▒α╕öα╕¬α╕▓α╕éα╕▓ + Central Region Rule
    # ==========================================
    def get_max_vehicle_for_code(code):
        """α╕½α╕▓α╕úα╕ûα╕ùα╕╡α╣êα╣âα╕½α╕ìα╣êα╕ùα╕╡α╣êα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╕¬α╕▓α╕éα╕▓α╕¬α╕▓α╕íα╕▓α╕úα╕ûα╣âα╕èα╣ëα╣äα╕öα╣ë"""
        max_vehicle = get_max_vehicle_for_branch(code)
        return max_vehicle
    
    def get_allowed_vehicles_for_region(region_name):
        """α╕½α╕▓α╕úα╕ûα╕ùα╕╡α╣êα╣âα╕èα╣ëα╣äα╕öα╣ëα╕òα╕▓α╕íα╕áα╕▓α╕ä (Central α╕½α╣ëα╕▓α╕í 6W)"""
        if region_name in CENTRAL_REGIONS:
            return CENTRAL_ALLOWED_VEHICLES  # ['4W', 'JB']
        return ['4W', 'JB', '6W']  # All vehicles
    
    df['_max_vehicle'] = df['Code'].apply(get_max_vehicle_for_code)
    df['_region_allowed_vehicles'] = df['_region_name'].apply(get_allowed_vehicles_for_region)
    
    # ==========================================
    # Step 6: DISTRICT CLUSTERING ALLOCATION (FIXED)
    # α╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢α╕òα╕▓α╕í District Buckets α╕₧α╕úα╣ëα╕¡α╕í Split α╣Çα╕íα╕╖α╣êα╕¡α╣Çα╕üα╕┤α╕Ö
    # ==========================================
    trip_counter = 1
    df['Trip'] = 0
    
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    
    # Helper function: α╣Çα╕Ñα╕╖α╕¡α╕üα╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕½α╕íα╕▓α╕░α╕¬α╕í
    def select_vehicle_for_load(weight, cube, drops, is_punthai, allowed_vehicles):
        """α╣Çα╕Ñα╕╖α╕¡α╕üα╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕Ñα╣çα╕üα╕ùα╕╡α╣êα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╕úα╕▒α╕Üα╣éα╕½α╕Ñα╕öα╣äα╕öα╣ë (α╕òα╣ëα╕¡α╕çα╣äα╕íα╣êα╣Çα╕üα╕┤α╕Ö Buffer)"""
        limits_to_use = PUNTHAI_LIMITS if is_punthai else LIMITS
        buffer_mult = punthai_buffer if is_punthai else maxmart_buffer
        
        for v in ['4W', 'JB', '6W']:
            if v not in allowed_vehicles:
                continue
            lim = limits_to_use.get(v, LIMITS['6W'])
            max_w = lim.get('max_w', lim.get('max_weight', 6000)) * buffer_mult
            max_c = lim.get('max_c', lim.get('max_cube', 20.0)) * buffer_mult
            max_d = lim.get('max_drops', 12)
            
            if weight <= max_w and cube <= max_c and drops <= max_d:
                return v
        
        # α╣äα╕íα╣êα╕íα╕╡α╕úα╕ûα╕úα╕¡α╕çα╕úα╕▒α╕Üα╣äα╕öα╣ëα╕áα╕▓α╕óα╣âα╕Ö buffer ΓåÆ return None (α╕òα╣ëα╕¡α╕ç split)
        return None
    
    # Helper function: α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓α╣Çα╕¢α╣çα╕Ö Punthai α╕Ñα╣ëα╕ºα╕Öα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê
    def is_all_punthai(rows):
        if not rows:
            return False
        return all(str(r.get('BU', '')).upper() in ['211', 'PUNTHAI'] for r in rows)
    
    # Current trip state
    current_trip = {
        'codes': [], 'weight': 0, 'cube': 0, 'drops': 0,
        'rows': [], 'region': None, 'allowed_vehicles': ['4W', 'JB', '6W'],
        'district': None
    }
    
    overflow_queue = []  # Queue α╕¬α╕│α╕½α╕úα╕▒α╕Ü stores α╕ùα╕╡α╣ê overflow
    
    def finalize_current_trip():
        """α╕¢α╕┤α╕öα╕ùα╕úα╕┤α╕¢α╕¢α╕▒α╕êα╕êα╕╕α╕Üα╕▒α╕Öα╣üα╕Ñα╕░α╕Üα╕▒α╕Öα╕ùα╕╢α╕ü"""
        nonlocal trip_counter
        if current_trip['codes']:
            for c in current_trip['codes']:
                df.loc[df['Code'] == c, 'Trip'] = trip_counter
    
    def split_until_fits(allowed_vehicles, region):
        """α╣üα╕óα╕ü stores α╕¡α╕¡α╕üα╕êα╕▓α╕ü current_trip α╕êα╕Öα╕üα╕ºα╣êα╕▓α╕êα╕░α╕₧α╕¡α╕öα╕╡α╕úα╕û (α╣äα╕íα╣êα╣Çα╕üα╕┤α╕Ö buffer)"""
        nonlocal trip_counter, overflow_queue
        
        split_count = 0
        
        while True:
            # α╕½α╕▓ capacity α╕¬α╕╣α╕çα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╣âα╕èα╣ëα╣äα╕öα╣ë (α╕úα╕ºα╕í branch constraints)
            is_punthai = is_all_punthai(current_trip['rows'])
            buffer_mult = punthai_buffer if is_punthai else maxmart_buffer
            
            # α╕½α╕▓ allowed_vehicles α╕ùα╕╡α╣êα╣üα╕ùα╣ëα╕êα╕úα╕┤α╕ç (α╕úα╕ºα╕í branch constraints α╕éα╕¡α╕çα╕ùα╕úα╕┤α╕¢α╕¢α╕▒α╕êα╕êα╕╕α╕Üα╕▒α╕Ö)
            effective_allowed = current_trip['allowed_vehicles'].copy() if current_trip['allowed_vehicles'] else allowed_vehicles.copy()
            
            # α╕½α╕▓α╕úα╕ûα╕ùα╕╡α╣êα╣âα╕½α╕ìα╣êα╕ùα╕╡α╣êα╕¬α╕╕α╕öα╕ùα╕╡α╣êα╕¡α╕Öα╕╕α╕ìα╕▓α╕ò
            max_vehicle = '6W' if '6W' in effective_allowed else ('JB' if 'JB' in effective_allowed else '4W')
            limits_to_use = PUNTHAI_LIMITS if is_punthai else LIMITS
            lim = limits_to_use.get(max_vehicle, LIMITS['6W'])
            max_w = lim.get('max_w', 6000) * buffer_mult
            max_c = lim.get('max_c', 20.0) * buffer_mult
            max_d = lim.get('max_drops', 12)
            
            # α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓α╣Çα╕üα╕┤α╕Öα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê (Weight α╕½α╕úα╕╖α╕¡ Cube α╕½α╕úα╕╖α╕¡ Drops)
            weight_over = current_trip['weight'] > max_w
            cube_over = current_trip['cube'] > max_c
            drops_over = current_trip['drops'] > max_d
            
            if not weight_over and not cube_over and not drops_over:
                # α╕₧α╕¡α╕öα╕╡α╣üα╕Ñα╣ëα╕º α╣äα╕íα╣êα╕òα╣ëα╕¡α╕ç split
                break
            
            if len(current_trip['codes']) <= 1:
                # α╣Çα╕½α╕Ñα╕╖α╕¡ 1 store ΓåÆ α╕òα╣ëα╕¡α╕çα╕óα╕¡α╕íα╕úα╕▒α╕Ü (α╣üα╕òα╣êα╕êα╕░α╕ûα╕╣α╕ü flag α╣âα╕Ö summary)
                break
            
            # ≡ƒöÑ α╕òα╕▒α╕öα╕ùα╕▒α╕Öα╕ùα╕╡: α╣Çα╕¡α╕▓ store α╕¬α╕╕α╕öα╕ùα╣ëα╕▓α╕óα╕¡α╕¡α╕ü
            overflow_code = current_trip['codes'].pop()
            overflow_row = current_trip['rows'].pop()
            current_trip['weight'] -= overflow_row['Weight']
            current_trip['cube'] -= overflow_row['Cube']
            current_trip['drops'] -= 1
            split_count += 1
            
            # α╕¡α╕▒α╕₧α╣Çα╕öα╕ù allowed_vehicles α╕½α╕Ñα╕▒α╕çα╣Çα╕¡α╕▓ store α╕¡α╕¡α╕ü
            new_allowed = allowed_vehicles.copy()
            for code in current_trip['codes']:
                branch_max_v = get_max_vehicle_for_code(code)
                if branch_max_v == 'JB' and '6W' in new_allowed:
                    new_allowed.remove('6W')
                elif branch_max_v == '4W':
                    if '6W' in new_allowed:
                        new_allowed.remove('6W')
                    if 'JB' in new_allowed:
                        new_allowed.remove('JB')
            current_trip['allowed_vehicles'] = new_allowed
            
            overflow_queue.append({
                'code': overflow_code,
                'row': overflow_row,
                'region': region,
                'allowed_vehicles': allowed_vehicles
            })
    
    def process_overflow_queue():
        """α╕¢α╕úα╕░α╕íα╕ºα╕Ñα╕£α╕Ñ overflow queue - α╕¬α╕úα╣ëα╕▓α╕çα╕ùα╕úα╕┤α╕¢α╣âα╕½α╕íα╣êα╕¬α╕│α╕½α╕úα╕▒α╕Ü stores α╕ùα╕╡α╣êα╕Ñα╣ëα╕Ö"""
        nonlocal trip_counter, current_trip, overflow_queue
        
        while overflow_queue:
            item = overflow_queue.pop(0)
            code = item['code']
            row = item['row']
            region = item['region']
            allowed_vehicles = item['allowed_vehicles']
            
            # α╕Ñα╕¡α╕çα╣Çα╕₧α╕┤α╣êα╕íα╣Çα╕éα╣ëα╕▓ current_trip
            if current_trip['codes']:
                test_weight = current_trip['weight'] + row['Weight']
                test_cube = current_trip['cube'] + row['Cube']
                test_drops = current_trip['drops'] + 1
                test_rows = current_trip['rows'] + [row]
                test_punthai = is_all_punthai(test_rows)
                
                vehicle = select_vehicle_for_load(test_weight, test_cube, test_drops, test_punthai, allowed_vehicles)
                
                if vehicle:
                    # α╕₧α╕¡α╕öα╕╡! α╣Çα╕₧α╕┤α╣êα╕íα╣Çα╕éα╣ëα╕▓
                    current_trip['codes'].append(code)
                    current_trip['weight'] = test_weight
                    current_trip['cube'] = test_cube
                    current_trip['drops'] = test_drops
                    current_trip['rows'].append(row)
                    
                    # ≡ƒöÆ DOUBLE CHECK: Split α╕ûα╣ëα╕▓α╕óα╕▒α╕çα╣Çα╕üα╕┤α╕Öα╕¡α╕óα╕╣α╣ê
                    split_until_fits(allowed_vehicles, region)
                else:
                    # α╣äα╕íα╣êα╕₧α╕¡α╕öα╕╡ ΓåÆ α╕¢α╕┤α╕öα╕ùα╕úα╕┤α╕¢α╣Çα╕üα╣êα╕▓, α╣Çα╕úα╕┤α╣êα╕íα╣âα╕½α╕íα╣ê
                    finalize_current_trip()
                    trip_counter += 1
                    current_trip = {
                        'codes': [code],
                        'weight': row['Weight'],
                        'cube': row['Cube'],
                        'drops': 1,
                        'rows': [row],
                        'region': region,
                        'allowed_vehicles': allowed_vehicles,
                        'district': None
                    }
            else:
                # α╕ùα╕úα╕┤α╕¢α╕ºα╣êα╕▓α╕ç
                current_trip = {
                    'codes': [code],
                    'weight': row['Weight'],
                    'cube': row['Cube'],
                    'drops': 1,
                    'rows': [row],
                    'region': region,
                    'allowed_vehicles': allowed_vehicles,
                    'district': None
                }
    
    # ==========================================
    # GROUP BY DISTRICT BUCKETS
    # ==========================================
    district_groups = df.groupby(['_region_name', '_province', '_district'], sort=False)
    
    for (region, province, district), district_df in district_groups:
        # α╕éα╣ëα╕¡α╕íα╕╣α╕Ñ District
        district_codes = district_df['Code'].tolist()
        district_weight = district_df['Weight'].sum()
        district_cube = district_df['Cube'].sum()
        district_drops = len(district_codes)
        district_rows = district_df.to_dict('records')
        
        # α╕½α╕▓α╕úα╕ûα╕ùα╕╡α╣êα╣âα╕èα╣ëα╣äα╕öα╣ëα╕òα╕▓α╕íα╕áα╕▓α╕ä
        allowed_vehicles = ['4W', 'JB', '6W']
        if region in CENTRAL_REGIONS:
            allowed_vehicles = CENTRAL_ALLOWED_VEHICLES
        
        # ==========================================
        # Rule 0: Region Change ΓåÆ α╕¢α╕┤α╕öα╕ùα╕úα╕┤α╕¢α╣Çα╕üα╣êα╕▓ + process overflow
        # ==========================================
        if current_trip['region'] and current_trip['region'] != region:
            # Process overflow α╕éα╕¡α╕ç region α╣Çα╕üα╣êα╕▓α╕üα╣êα╕¡α╕Ö
            process_overflow_queue()
            finalize_current_trip()
            trip_counter += 1
            current_trip = {
                'codes': [], 'weight': 0, 'cube': 0, 'drops': 0,
                'rows': [], 'region': None, 'allowed_vehicles': allowed_vehicles,
                'district': None
            }
        
        # ==========================================
        # Rule 1: α╕Ñα╕¡α╕çα╣âα╕¬α╣êα╕ùα╕▒α╣ëα╕ç District
        # ==========================================
        if current_trip['codes']:
            test_weight = current_trip['weight'] + district_weight
            test_cube = current_trip['cube'] + district_cube
            test_drops = current_trip['drops'] + district_drops
            test_rows = current_trip['rows'] + district_rows
            test_punthai = is_all_punthai(test_rows)
            
            # α╕úα╕ºα╕í allowed_vehicles α╕êα╕▓α╕ü region + branch constraints
            test_allowed = list(set(current_trip['allowed_vehicles']) & set(allowed_vehicles))
            
            # α╣Çα╕₧α╕┤α╣êα╕í: α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Ü branch constraints α╕éα╕¡α╕ç District α╣âα╕½α╕íα╣ê
            for code in district_codes:
                branch_max_v = get_max_vehicle_for_code(code)
                # α╕ûα╣ëα╕▓α╕¬α╕▓α╕éα╕▓α╕½α╣ëα╕▓α╕í 6W α╕òα╣ëα╕¡α╕çα╣Çα╕¡α╕▓ 6W α╕¡α╕¡α╕ü
                if branch_max_v == 'JB' and '6W' in test_allowed:
                    test_allowed.remove('6W')
                elif branch_max_v == '4W':
                    if '6W' in test_allowed:
                        test_allowed.remove('6W')
                    if 'JB' in test_allowed:
                        test_allowed.remove('JB')
            
            vehicle = select_vehicle_for_load(test_weight, test_cube, test_drops, test_punthai, test_allowed)
            
            if vehicle:
                # District α╕₧α╕¡α╕öα╕╡! α╣Çα╕₧α╕┤α╣êα╕íα╣Çα╕éα╣ëα╕▓α╕ùα╕úα╕┤α╕¢
                current_trip['codes'].extend(district_codes)
                current_trip['weight'] = test_weight
                current_trip['cube'] = test_cube
                current_trip['drops'] = test_drops
                current_trip['rows'].extend(district_rows)
                current_trip['allowed_vehicles'] = test_allowed
                current_trip['region'] = region
                
                # ≡ƒöÆ DOUBLE CHECK: Split α╕ûα╣ëα╕▓α╕óα╕▒α╕çα╣Çα╕üα╕┤α╕Öα╕¡α╕óα╕╣α╣ê
                split_until_fits(test_allowed, region)
            else:
                # District α╣äα╕íα╣êα╕₧α╕¡α╕öα╕╡ ΓåÆ α╕¢α╕┤α╕öα╕ùα╕úα╕┤α╕¢α╣Çα╕üα╣êα╕▓
                finalize_current_trip()
                trip_counter += 1
                
                # α╕½α╕▓ allowed_vehicles α╕¬α╕│α╕½α╕úα╕▒α╕Ü District α╣âα╕½α╕íα╣ê (α╕úα╕ºα╕í branch constraints)
                new_allowed = allowed_vehicles.copy()
                for code in district_codes:
                    branch_max_v = get_max_vehicle_for_code(code)
                    if branch_max_v == 'JB' and '6W' in new_allowed:
                        new_allowed.remove('6W')
                    elif branch_max_v == '4W':
                        if '6W' in new_allowed:
                            new_allowed.remove('6W')
                        if 'JB' in new_allowed:
                            new_allowed.remove('JB')
                
                # α╣Çα╕úα╕┤α╣êα╕íα╕ùα╕úα╕┤α╕¢α╣âα╕½α╕íα╣êα╕üα╕▒α╕Ü District α╕Öα╕╡α╣ë
                current_trip = {
                    'codes': district_codes.copy(),
                    'weight': district_weight,
                    'cube': district_cube,
                    'drops': district_drops,
                    'rows': district_rows.copy(),
                    'region': region,
                    'allowed_vehicles': new_allowed,
                    'district': district
                }
                
                # ==========================================
                # Rule 2: α╕ûα╣ëα╕▓ District α╣âα╕½α╕ìα╣êα╣Çα╕üα╕┤α╕Öα╕úα╕û ΓåÆ Split α╕ùα╕▒α╕Öα╕ùα╕╡!
                # ==========================================
                split_until_fits(allowed_vehicles, region)
        else:
            # α╕ùα╕úα╕┤α╕¢α╕ºα╣êα╕▓α╕ç - α╕½α╕▓ allowed_vehicles α╕úα╕ºα╕í branch constraints
            new_allowed = allowed_vehicles.copy()
            for code in district_codes:
                branch_max_v = get_max_vehicle_for_code(code)
                if branch_max_v == 'JB' and '6W' in new_allowed:
                    new_allowed.remove('6W')
                elif branch_max_v == '4W':
                    if '6W' in new_allowed:
                        new_allowed.remove('6W')
                    if 'JB' in new_allowed:
                        new_allowed.remove('JB')
            
            current_trip = {
                'codes': district_codes.copy(),
                'weight': district_weight,
                'cube': district_cube,
                'drops': district_drops,
                'rows': district_rows.copy(),
                'region': region,
                'allowed_vehicles': new_allowed,
                'district': district
            }
            
            # ==========================================
            # Rule 2: α╕ûα╣ëα╕▓ District α╣âα╕½α╕ìα╣êα╣Çα╕üα╕┤α╕Öα╕úα╕û ΓåÆ Split α╕ùα╕▒α╕Öα╕ùα╕╡!
            # ==========================================
            split_until_fits(new_allowed, region)
    
    # ==========================================
    # Final: Process remaining overflow α╣üα╕Ñα╕░α╕¢α╕┤α╕öα╕ùα╕úα╕┤α╕¢α╕¬α╕╕α╕öα╕ùα╣ëα╕▓α╕ó
    # ==========================================
    process_overflow_queue()
    finalize_current_trip()

    # ==========================================
    # Step 7: α╕¬α╕úα╣ëα╕▓α╕ç Summary + Central Rule + Punthai Drop Limits
    # ==========================================
    summary_data = []
    
    for trip_num in sorted(df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = df[df['Trip'] == trip_num]
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        trip_codes = trip_data['Code'].unique()
        trip_drops = len(trip_codes)
        
        # α╕½α╕▓α╕áα╕▓α╕äα╕éα╕¡α╕çα╕ùα╕úα╕┤α╕¢ (α╣âα╕èα╣ëα╕áα╕▓α╕äα╣üα╕úα╕ü)
        trip_region = trip_data['_region_name'].iloc[0] if '_region_name' in trip_data.columns else 'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕'
        
        # α╕½α╕▓α╕úα╕ûα╕ùα╕╡α╣êα╣Çα╕½α╕íα╕▓α╕░α╕¬α╕í (α╕úα╕ºα╕í Central Rule)
        max_vehicles = [get_max_vehicle_for_branch(c) for c in trip_codes]
        min_max_size = min(vehicle_priority.get(v, 3) for v in max_vehicles)
        max_allowed_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(min_max_size, '6W')
        
        # ≡ƒÜ½ Central Region Rule: α╕½α╣ëα╕▓α╕í 6W
        if trip_region in CENTRAL_REGIONS and max_allowed_vehicle == '6W':
            max_allowed_vehicle = 'JB'  # α╕Ñα╕öα╣Çα╕¢α╣çα╕Ö JB
        
        # α╕òα╕úα╕ºα╕ê BU α╕éα╕¡α╕çα╕ùα╕úα╕┤α╕¢
        is_punthai_only_trip = True
        for _, r in trip_data.iterrows():
            bu = str(r.get('BU', '')).upper()
            if bu not in ['211', 'PUNTHAI']:
                is_punthai_only_trip = False
                break
        
        buffer = punthai_buffer if is_punthai_only_trip else maxmart_buffer
        buffer_pct = int(buffer * 100)
        buffer_label = f"≡ƒà┐∩╕Å {buffer_pct}%" if is_punthai_only_trip else f"≡ƒà╝ {buffer_pct}%"
        trip_type = 'punthai' if is_punthai_only_trip else 'maxmart'
        
        # α╣Çα╕Ñα╕╖α╕¡α╕üα╕úα╕ûα╕ùα╕╡α╣êα╕₧α╕¡α╕öα╕╡α╕ùα╕╡α╣êα╕¬α╕╕α╕ö
        suggested = max_allowed_vehicle
        source = "≡ƒôï α╕êα╕│α╕üα╕▒α╕öα╕¬α╕▓α╕éα╕▓" if min_max_size < 3 else "≡ƒñû α╕¡α╕▒α╕òα╣éα╕Öα╕íα╕▒α╕òα╕┤"
        
        # ≡ƒöÆ Punthai Drop Limit Check
        if is_punthai_only_trip:
            punthai_drop_limit = PUNTHAI_LIMITS.get(suggested, {}).get('max_drops', 999)
            if trip_drops > punthai_drop_limit:
                # α╕òα╣ëα╕¡α╕çα╣Çα╕₧α╕┤α╣êα╕íα╕éα╕Öα╕▓α╕öα╕úα╕ûα╣Çα╕₧α╕╖α╣êα╕¡α╕úα╕¡α╕çα╕úα╕▒α╕Ü drops
                if suggested == '4W' and trip_drops <= PUNTHAI_LIMITS['JB']['max_drops']:
                    suggested = 'JB'
                    source += " ΓåÆ JB (Drop Limit)"
                elif suggested == 'JB' or trip_drops > PUNTHAI_LIMITS['JB']['max_drops']:
                    # α╕ûα╣ëα╕▓ Central α╕½α╣ëα╕▓α╕í 6W ΓåÆ WARNING
                    if trip_region not in CENTRAL_REGIONS:
                        suggested = '6W'
                        source += " ΓåÆ 6W (Drop Limit)"
                    else:
                        source += " ΓÜá∩╕Å Drops α╣Çα╕üα╕┤α╕Ö!"
        
        # α╕äα╕│α╕Öα╕ºα╕ô utilization
        max_util_threshold = buffer * 100  # 100% α╕½α╕úα╕╖α╕¡ 110% α╕òα╕▓α╕í BU
        if suggested in LIMITS:
            w_util = (total_w / LIMITS[suggested]['max_w']) * 100
            c_util = (total_c / LIMITS[suggested]['max_c']) * 100
            max_util = max(w_util, c_util)
            
            # α╕ûα╣ëα╕▓α╣Çα╕üα╕┤α╕Ö threshold α╕òα╕▓α╕í BU α╕òα╣ëα╕¡α╕çα╣Çα╕₧α╕┤α╣êα╕íα╕éα╕Öα╕▓α╕öα╕úα╕û
            if max_util > max_util_threshold:
                if suggested == '4W' and min_max_size >= 2:
                    jb_util = max((total_w / LIMITS['JB']['max_w']), (total_c / LIMITS['JB']['max_c'])) * 100
                    if jb_util <= max_util_threshold:
                        suggested = 'JB'
                        source += " ΓåÆ JB"
                        w_util = (total_w / LIMITS['JB']['max_w']) * 100
                        c_util = (total_c / LIMITS['JB']['max_c']) * 100
                    elif min_max_size >= 3:
                        suggested = '6W'
                        source += " ΓåÆ 6W"
                        w_util = (total_w / LIMITS['6W']['max_w']) * 100
                        c_util = (total_c / LIMITS['6W']['max_c']) * 100
                elif suggested == 'JB' and min_max_size >= 3:
                    suggested = '6W'
                    source += " ΓåÆ 6W"
                    w_util = (total_w / LIMITS['6W']['max_w']) * 100
                    c_util = (total_c / LIMITS['6W']['max_c']) * 100
        else:
            w_util = c_util = 0
        
        # α╕äα╕│α╕Öα╕ºα╕ôα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕úα╕ºα╕í
        total_distance = 0
        branch_coords = []
        for code in trip_codes:
            loc = location_map.get(str(code).upper(), {})
            if loc.get('lat') and loc.get('lon'):
                branch_coords.append((loc['lat'], loc['lon']))
        
        if branch_coords:
            # DC ΓåÆ α╕¬α╕▓α╕éα╕▓α╣üα╕úα╕ü
            total_distance += haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, branch_coords[0][0], branch_coords[0][1])
            # α╕¬α╕▓α╕éα╕▓ ΓåÆ α╕¬α╕▓α╕éα╕▓
            for i in range(len(branch_coords) - 1):
                total_distance += haversine_distance(branch_coords[i][0], branch_coords[i][1], branch_coords[i+1][0], branch_coords[i+1][1])
            # α╕¬α╕▓α╕éα╕▓α╕¬α╕╕α╕öα╕ùα╣ëα╕▓α╕ó ΓåÆ DC
            total_distance += haversine_distance(branch_coords[-1][0], branch_coords[-1][1], DC_WANG_NOI_LAT, DC_WANG_NOI_LON)
        
        summary_data.append({
            'Trip': trip_num,
            'Branches': len(trip_codes),
            'Weight': total_w,
            'Cube': total_c,
            'Truck': f"{suggested} {source}",
            'BU_Type': trip_type,
            'Buffer': buffer_label,
            'Weight_Use%': w_util,
            'Cube_Use%': c_util,
            'Total_Distance': round(total_distance, 1)
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    # ==========================================
    # Step 8: α╣Çα╕₧α╕┤α╣êα╕íα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╣Çα╕¬α╕úα╕┤α╕í
    # ==========================================
    # α╣Çα╕₧α╕┤α╣êα╕íα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕úα╕û
    trip_truck_map = {}
    for _, row in summary_df.iterrows():
        trip_truck_map[row['Trip']] = row['Truck']
    df['Truck'] = df['Trip'].map(trip_truck_map)
    
    # α╣Çα╕₧α╕┤α╣êα╕íα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣î Region
    df['Region'] = df['_region_name']
    
    # α╣Çα╕₧α╕┤α╣êα╕íα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣î Province (α╕ûα╣ëα╕▓α╕óα╕▒α╕çα╣äα╕íα╣êα╕íα╕╡)
    if 'Province' not in df.columns:
        df['Province'] = df['_province']
    
    # α╣Çα╕₧α╕┤α╣êα╕íα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕êα╕▓α╕ü DC
    df['Distance_from_DC'] = df['_distance_from_dc'].round(1)
    
    # α╣Çα╕₧α╕┤α╣êα╕íα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╣Çα╕èα╣çα╕äα╕úα╕û
    df['VehicleCheck'] = 'Γ£à α╣âα╕èα╣ëα╣äα╕öα╣ë'
    
    # ==========================================
    # Step 9: α╣Çα╕úα╕╡α╕óα╕çα╕ùα╕úα╕┤α╕¢α╣âα╕½α╕íα╣êα╣âα╕½α╣ëα╕ùα╕úα╕┤α╕¢α╕òα╕┤α╕öα╕üα╕▒α╕Ö (α╕¬α╕│α╕½α╕úα╕▒α╕Ü export)
    # ==========================================
    df = df.sort_values(['Trip', '_distance_from_dc'], ascending=[True, False]).reset_index(drop=True)
    
    # α╕Ñα╕Üα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕èα╕▒α╣êα╕ºα╕äα╕úα╕▓α╕º
    cols_to_drop = ['_region_code', '_region_name', '_prov_code', '_dist_code', '_subdist_code', '_province', '_district', '_subdistrict', '_route', '_distance_from_dc', '_group_key', '_max_vehicle', '_region_order', '_prov_max_dist', '_dist_max_dist', '_region_allowed_vehicles']
    df = df.drop(columns=[c for c in cols_to_drop if c in df.columns], errors='ignore')
    
    return df, summary_df
def main():
    st.set_page_config(
        page_title="α╕úα╕░α╕Üα╕Üα╕êα╕▒α╕öα╣Çα╕ùα╕╡α╣êα╕óα╕º",
        page_icon="≡ƒÜÜ",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # ≡ƒöä Auto-refresh α╕ùα╕╕α╕üα╣Çα╕ùα╕╡α╣êα╕óα╕çα╕äα╕╖α╕Ö (α╕Ñα╣ëα╕▓α╕çα╣üα╕äα╕è)
    if AUTOREFRESH_AVAILABLE:
        now = datetime.now()
        # α╕äα╕│α╕Öα╕ºα╕ôα╣Çα╕ºα╕Ñα╕▓α╕ûα╕╢α╕çα╣Çα╕ùα╕╡α╣êα╕óα╕çα╕äα╕╖α╕Ö (00:00:00)
        midnight = datetime.combine(now.date(), time(0, 0, 0))
        
        # α╕ûα╣ëα╕▓α╕óα╕▒α╕çα╣äα╕íα╣êα╕ûα╕╢α╕çα╣Çα╕ùα╕╡α╣êα╕óα╕çα╕äα╕╖α╕Ö α╣Çα╕¡α╕▓α╣Çα╕ùα╕╡α╣êα╕óα╕çα╕äα╕╖α╕Öα╕ºα╕▒α╕Öα╕ûα╕▒α╕öα╣äα╕¢
        if now < midnight:
            next_midnight = midnight
        else:
            from datetime import timedelta
            next_midnight = midnight + timedelta(days=1)
        
        # α╕äα╕│α╕Öα╕ºα╕ôα╣Çα╕ºα╕Ñα╕▓α╕ùα╕╡α╣êα╣Çα╕½α╕Ñα╕╖α╕¡ (α╕ºα╕┤α╕Öα╕▓α╕ùα╕╡)
        seconds_until_midnight = int((next_midnight - now).total_seconds())
        
        # Refresh α╕ùα╕╕α╕üα╣Çα╕ùα╕╡α╣êα╕óα╕çα╕äα╕╖α╕Ö
        if seconds_until_midnight > 0:
            # α╣Çα╕èα╣çα╕äα╣âα╕Öα╕èα╣êα╕ºα╕ç 5 α╕Öα╕▓α╕ùα╕╡α╕üα╣êα╕¡α╕Öα╣Çα╕ùα╕╡α╣êα╕óα╕çα╕äα╕╖α╕Ö (α╕½α╕Ñα╕▒α╕ç 23:55)
            if seconds_until_midnight <= 300:  # 5 minutes
                st.info(f"≡ƒöä α╕úα╕░α╕Üα╕Üα╕êα╕░ Refresh α╕¡α╕▒α╕òα╣éα╕Öα╕íα╕▒α╕òα╕┤α╣âα╕Ö {seconds_until_midnight // 60} α╕Öα╕▓α╕ùα╕╡")
                st_autorefresh(interval=seconds_until_midnight * 1000, key="midnight_refresh")
            else:
                # α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕ùα╕╕α╕ü 1 α╕èα╕▒α╣êα╕ºα╣éα╕íα╕ç
                st_autorefresh(interval=3600000, limit=24, key="hourly_check")
    
    # Header
    col1, col2 = st.columns([3, 1])
    with col1:
        st.title("≡ƒÜÜ α╕úα╕░α╕Üα╕Üα╕êα╕▒α╕öα╣Çα╕ùα╕╡α╣êα╕óα╕º")
    with col2:
        st.image("https://raw.githubusercontent.com/twitter/twemoji/master/assets/svg/1f69a.svg", width=100)
    
    # Show Punthai learning stats
    if PUNTHAI_PATTERNS and 'stats' in PUNTHAI_PATTERNS and PUNTHAI_PATTERNS['stats']:
        stats = PUNTHAI_PATTERNS['stats']
        with st.expander("≡ƒôè α╕¬α╕ûα╕┤α╕òα╕┤α╕ùα╕╡α╣êα╣Çα╕úα╕╡α╕óα╕Öα╕úα╕╣α╣ëα╕êα╕▓α╕üα╣äα╕ƒα╕Ñα╣î Punthai Maxmart", expanded=False):
            col_a, col_b, col_c = st.columns(3)
            with col_a:
                st.metric("α╣Çα╕ëα╕Ñα╕╡α╣êα╕óα╕¬α╕▓α╕éα╕▓/α╕ùα╕úα╕┤α╕¢", f"{stats.get('avg_branches', 0):.1f}")
            with col_b:
                st.metric("α╕ùα╕úα╕┤α╕¢α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣Çα╕öα╕╡α╕óα╕º", f"{stats.get('same_province_pct', 0):.1f}%")
            with col_c:
                total_trips = stats.get('same_province', 0) + stats.get('mixed_province', 0)
                st.metric("α╕êα╕│α╕Öα╕ºα╕Öα╕ùα╕úα╕┤α╕¢α╕¡α╣ëα╕▓α╕çα╕¡α╕┤α╕ç", total_trips)
    
    st.markdown("---")
    
    # α╣éα╕½α╕Ñα╕öα╣éα╕íα╣Çα╕öα╕Ñ
    model_data = load_model()
    
    if not model_data:
        st.error("Γ¥î α╣äα╕íα╣êα╕₧α╕Üα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╣éα╕íα╣Çα╕öα╕Ñ α╕üα╕úα╕╕α╕ôα╕▓α╣Çα╕ùα╕úα╕Öα╣éα╕íα╣Çα╕öα╕Ñα╕üα╣êα╕¡α╕Öα╣âα╕èα╣ëα╕çα╕▓α╕Ö")
        st.info("≡ƒÆí α╕úα╕▒α╕Öα╕äα╕│α╕¬α╕▒α╣êα╕ç: `python test_model.py`")
        st.stop()
    
    # α╕¡α╕▒α╕¢α╣éα╕½α╕Ñα╕öα╣äα╕ƒα╕Ñα╣îα╕äα╕úα╕▒α╣ëα╕çα╣Çα╕öα╕╡α╕óα╕º
    st.markdown("### ≡ƒôé α╕¡α╕▒α╕¢α╣éα╕½α╕Ñα╕öα╣äα╕ƒα╕Ñα╣îα╕úα╕▓α╕óα╕üα╕▓α╕úα╕¡α╕¡α╣Çα╕öα╕¡α╕úα╣î")
    uploaded_file = st.file_uploader(
        "α╣Çα╕Ñα╕╖α╕¡α╕üα╣äα╕ƒα╕Ñα╣î Excel (.xlsx)", 
        type=['xlsx'],
        help="α╕¡α╕▒α╕¢α╣éα╕½α╕Ñα╕öα╣äα╕ƒα╕Ñα╣î Excel α╕ùα╕╡α╣êα╕íα╕╡α╕úα╕▓α╕óα╕üα╕▓α╕úα╕¬α╕▓α╕éα╕▓α╣üα╕Ñα╕░α╕¡α╕¡α╣Çα╕öα╕¡α╕úα╣î"
    )
    
    if uploaded_file:
        # α╣Çα╕üα╣çα╕Üα╣äα╕ƒα╕Ñα╣îα╕òα╣ëα╕Öα╕ëα╕Üα╕▒α╕Üα╣äα╕ºα╣ëα╣âα╕Ö session_state α╣Çα╕₧α╕╖α╣êα╕¡α╣âα╕èα╣ëα╕òα╕¡α╕Ö export
        uploaded_file_content = uploaded_file.read()
        st.session_state['original_file_content'] = uploaded_file_content
        
        with st.spinner("ΓÅ│ α╕üα╕│α╕Ñα╕▒α╕çα╕¡α╣êα╕▓α╕Öα╕éα╣ëα╕¡α╕íα╕╣α╕Ñ..."):
            df = load_excel(uploaded_file_content)
            df = process_dataframe(df)
            
            if df is not None and 'Code' in df.columns:
                st.success(f"Γ£à α╕¡α╣êα╕▓α╕Öα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕¬α╕│α╣Çα╕úα╣çα╕ê: **{len(df):,}** α╕úα╕▓α╕óα╕üα╕▓α╕ú")
                
                # α╣üα╕¬α╕öα╕çα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕₧α╕╖α╣ëα╕Öα╕Éα╕▓α╕Ö
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("≡ƒôì α╕êα╕│α╕Öα╕ºα╕Öα╕¬α╕▓α╕éα╕▓", f"{df['Code'].nunique():,}")
                with col2:
                    st.metric("ΓÜû∩╕Å α╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕üα╕úα╕ºα╕í", f"{df['Weight'].sum():,.0f} kg")
                with col3:
                    st.metric("≡ƒôª α╕äα╕┤α╕ºα╕úα╕ºα╕í", f"{df['Cube'].sum():.1f} m┬│")
                with col4:
                    provinces = df['Province'].nunique() if 'Province' in df.columns else 0
                    st.metric("≡ƒù║∩╕Å α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö", f"{provinces}")
                
                # α╣üα╕¬α╕öα╕çα╕òα╕▒α╕ºα╕¡α╕óα╣êα╕▓α╕çα╕éα╣ëα╕¡α╕íα╕╣α╕Ñ
                with st.expander("≡ƒöì α╕öα╕╣α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕òα╕▒α╕ºα╕¡α╕óα╣êα╕▓α╕ç"):
                    st.dataframe(df.head(10), use_container_width=True)
                
                # ==========================================
                # α╣Çα╕òα╕┤α╕íα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕₧α╕╖α╣ëα╕Öα╕ùα╕╡α╣êα╕êα╕▓α╕ü Master (α╕ùα╕│α╣âα╕Öα╕½α╕Ñα╕▒α╕çα╕Üα╣ëα╕▓α╕Ö)
                # ==========================================
                if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                    # α╕¬α╕úα╣ëα╕▓α╕ç dict α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕äα╣ëα╕Öα╕½α╕▓α╣Çα╕úα╣çα╕º
                    master_lookup = {}
                    for _, row in MASTER_DATA.iterrows():
                        code = str(row['Plan Code']).strip().upper()
                        master_lookup[code] = {
                            'province': row.get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', ''),
                            'district': row.get('α╕¡α╕│α╣Çα╕áα╕¡', ''),
                            'subdistrict': row.get('α╕òα╕│α╕Üα╕Ñ', ''),
                            'lat': row.get('α╕Ñα╕░α╕òα╕┤α╕êα╕╣α╕ö', 0),
                            'lon': row.get('α╕Ñα╕¡α╕çα╕òα╕┤α╕êα╕╣α╕ö', 0)
                        }
                    
                    # α╣Çα╕òα╕┤α╕íα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕ùα╕╡α╣êα╕éα╕▓α╕ö
                    filled_count = 0
                    for idx, row in df.iterrows():
                        code = str(row['Code']).strip().upper()
                        if code in master_lookup:
                            master_info = master_lookup[code]
                            # α╣Çα╕òα╕┤α╕í Province α╕ûα╣ëα╕▓α╕ºα╣êα╕▓α╕ç
                            if 'Province' not in df.columns or pd.isna(df.loc[idx, 'Province']) or df.loc[idx, 'Province'] == '' or df.loc[idx, 'Province'] == 'UNKNOWN':
                                if master_info['province']:
                                    df.loc[idx, 'Province'] = master_info['province']
                                    filled_count += 1
                            # α╣Çα╕òα╕┤α╕í District α╕ûα╣ëα╕▓α╕ºα╣êα╕▓α╕ç
                            if 'District' not in df.columns:
                                df['District'] = ''
                            if pd.isna(df.loc[idx, 'District']) or df.loc[idx, 'District'] == '':
                                if master_info['district']:
                                    df.loc[idx, 'District'] = master_info['district']
                            # α╣Çα╕òα╕┤α╕í Subdistrict α╕ûα╣ëα╕▓α╕ºα╣êα╕▓α╕ç
                            if 'Subdistrict' not in df.columns:
                                df['Subdistrict'] = ''
                            if pd.isna(df.loc[idx, 'Subdistrict']) or df.loc[idx, 'Subdistrict'] == '':
                                if master_info['subdistrict']:
                                    df.loc[idx, 'Subdistrict'] = master_info['subdistrict']
                    
                    if filled_count > 0:
                        st.info(f"≡ƒôì α╣Çα╕òα╕┤α╕íα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕₧α╕╖α╣ëα╕Öα╕ùα╕╡α╣êα╕êα╕▓α╕ü Master α╣üα╕Ñα╣ëα╕º {filled_count} α╕úα╕▓α╕óα╕üα╕▓α╕ú")
                
                # α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕ºα╣êα╕▓α╕óα╕▒α╕çα╕íα╕╡α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕ùα╕╡α╣êα╕éα╕▓α╕öα╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê (α╣üα╕¬α╕öα╕çα╣üα╕äα╣êα╕êα╕│α╕Öα╕ºα╕Ö)
                if 'Province' in df.columns:
                    missing_count = len(df[(df['Province'].isna()) | (df['Province'] == '') | (df['Province'] == 'UNKNOWN')])
                    if missing_count > 0:
                        st.warning(f"ΓÜá∩╕Å α╕óα╕▒α╕çα╕íα╕╡ {missing_count} α╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣äα╕íα╣êα╕₧α╕Üα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕₧α╕╖α╣ëα╕Öα╕ùα╕╡α╣êα╣âα╕Ö Master")
                
                st.markdown("---")
                
                # α╣üα╕ùα╣çα╕Üα╕½α╕Ñα╕▒α╕ü
                tab1, tab2 = st.tabs([
                    "≡ƒôª α╕êα╕▒α╕öα╣Çα╕ùα╕╡α╣êα╕óα╕º (α╕òα╕▓α╕íα╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü)", 
                    "≡ƒù║∩╕Å α╕êα╕▒α╕öα╕üα╕Ñα╕╕α╣êα╕íα╕òα╕▓α╕íα╕áα╕▓α╕ä"
                ])
                    
                # ==========================================
                # α╣üα╕ùα╣çα╕Ü 1: α╕êα╕▒α╕öα╣Çα╕ùα╕╡α╣êα╕óα╕º (α╕òα╕▓α╕íα╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü)
                # ==========================================
                with tab1:
                    # α╣Çα╕₧α╕┤α╣êα╕í Region α╕ûα╣ëα╕▓α╕óα╕▒α╕çα╣äα╕íα╣êα╕íα╕╡
                    if 'Region' not in df.columns and 'Province' in df.columns:
                        df['Region'] = df['Province'].apply(get_region_name)
                    
                    # ==========================================
                    # α╕òα╕▒α╕ºα╣Çα╕Ñα╕╖α╕¡α╕üα╕üα╕▓α╕úα╕òα╕▒α╣ëα╕çα╕äα╣êα╕▓
                    # ==========================================
                    st.markdown("#### ΓÜÖ∩╕Å α╕òα╕▒α╣ëα╕çα╕äα╣êα╕▓α╕üα╕▓α╕úα╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢")
                    
                    # α╕üα╕úα╕¡α╕ü Buffer α╣üα╕óα╕üα╕òα╕▓α╕íα╕¢α╕úα╕░α╣Çα╕áα╕ù
                    col_buf1, col_buf2 = st.columns(2)
                    
                    with col_buf1:
                        punthai_buffer = st.number_input(
                            "≡ƒà┐∩╕Å Punthai Buffer %",
                            min_value=80,
                            max_value=120,
                            value=100,
                            step=5
                        )
                    
                    with col_buf2:
                        maxmart_buffer = st.number_input(
                            "≡ƒà╝ Maxmart/α╕£α╕¬α╕í Buffer %",
                            min_value=80,
                            max_value=150,
                            value=110,
                            step=5
                        )
                    
                    # α╣üα╕¢α╕Ñα╕çα╣Çα╕¢α╣çα╕Ö buffer value
                    punthai_buffer_value = punthai_buffer / 100.0
                    maxmart_buffer_value = maxmart_buffer / 100.0
                    
                    st.markdown("---")
                    
                    # α╕¢α╕╕α╣êα╕íα╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢
                    if st.button("≡ƒÜÇ α╣Çα╕úα╕┤α╣êα╕íα╕êα╕▒α╕öα╣Çα╕ùα╕╡α╣êα╕óα╕º", type="primary", use_container_width=True):
                        with st.spinner("ΓÅ│ α╕üα╕│α╕Ñα╕▒α╕çα╕¢α╕úα╕░α╕íα╕ºα╕Ñα╕£α╕Ñ..."):
                            # α╕êα╕▒α╕öα╣Çα╕úα╕╡α╕óα╕çα╕òα╕▓α╕íα╕áα╕▓α╕ä/α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö/α╕¡α╕│α╣Çα╕áα╕¡/α╕òα╕│α╕Üα╕Ñ/Route (α╣âα╕Öα╕ƒα╕▒α╕çα╕üα╣îα╕èα╕▒α╕Ö predict_trips)
                            df_to_process = df.copy()
                            
                            # α╕¬α╣êα╕ç buffer α╣üα╕óα╕üα╕òα╕▓α╕í BU
                            result_df, summary = predict_trips(
                                df_to_process, 
                                model_data, 
                                punthai_buffer=punthai_buffer_value,
                                maxmart_buffer=maxmart_buffer_value
                            )
                            
                            # α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣äα╕íα╣êα╣äα╕öα╣ëα╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢ (Trip = 0)
                            unassigned_count = len(result_df[result_df['Trip'] == 0])
                            if unassigned_count > 0:
                                st.warning(f"ΓÜá∩╕Å α╕íα╕╡ {unassigned_count} α╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣äα╕íα╣êα╣äα╕öα╣ëα╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢ (Trip = 0)")
                            
                            # α╕üα╕úα╕¡α╕çα╣Çα╕ëα╕₧α╕▓α╕░α╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢α╣üα╕Ñα╣ëα╕º α╕¬α╕│α╕½α╕úα╕▒α╕Üα╕üα╕▓α╕úα╣üα╕¬α╕öα╕çα╕£α╕Ñ
                            assigned_df = result_df[result_df['Trip'] > 0].copy()
                            
                            st.balloons()
                            st.success(f"Γ£à **α╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢α╣Çα╕¬α╕úα╣çα╕êα╕¬α╕íα╕Üα╕╣α╕úα╕ôα╣î!** α╕úα╕ºα╕í **{len(summary)}** α╕ùα╕úα╕┤α╕¢ ({len(assigned_df)} α╕¬α╕▓α╕éα╕▓)")
                            
                            st.markdown("---")
                            
                            # α╕¬α╕ûα╕┤α╕òα╕┤α╣éα╕öα╕óα╕úα╕ºα╕í
                            st.markdown("### ≡ƒôè α╕¬α╕úα╕╕α╕¢α╕£α╕Ñα╕üα╕▓α╕úα╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢")
                            
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("≡ƒÜÜ α╕êα╕│α╕Öα╕ºα╕Öα╕ùα╕úα╕┤α╕¢", len(summary))
                            with col2:
                                st.metric("≡ƒôì α╕êα╕│α╕Öα╕ºα╕Öα╕¬α╕▓α╕éα╕▓", len(assigned_df))
                            with col3:
                                avg_branches = len(assigned_df) / max(1, assigned_df['Trip'].nunique())
                                st.metric("≡ƒôè α╣Çα╕ëα╕Ñα╕╡α╣êα╕óα╕¬α╕▓α╕éα╕▓/α╕ùα╕úα╕┤α╕¢", f"{avg_branches:.1f}")
                            with col4:
                                avg_util = summary['Cube_Use%'].mean() if len(summary) > 0 else 0
                                st.metric("≡ƒôê α╕üα╕▓α╕úα╣âα╕èα╣ëα╕úα╕ûα╣Çα╕ëα╕Ñα╕╡α╣êα╕ó", f"{avg_util:.0f}%")
                            
                            st.markdown("---")
                            
                            # α╕òα╕▓α╕úα╕▓α╕çα╕¬α╕úα╕╕α╕¢α╣üα╕òα╣êα╕Ñα╕░α╕ùα╕úα╕┤α╕¢
                            st.markdown("### ≡ƒÜ¢ α╕úα╕▓α╕óα╕Ñα╕░α╣Çα╕¡α╕╡α╕óα╕öα╣üα╕òα╣êα╕Ñα╕░α╕ùα╕úα╕┤α╕¢")
                            st.dataframe(
                                summary.style.format({
                                    'Weight': '{:.2f}',
                                    'Cube': '{:.2f}',
                                    'Weight_Use%': '{:.1f}%',
                                    'Cube_Use%': '{:.1f}%',
                                    'Total_Distance': '{:.1f} km'
                                }).background_gradient(
                                    subset=['Weight_Use%', 'Cube_Use%'],
                                    cmap='RdYlGn',
                                    vmin=0,
                                    vmax=100
                                ),
                                use_container_width=True,
                                height=400
                            )
                            
                            # α╕òα╕▓α╕úα╕▓α╕çα╕úα╕▓α╕óα╕Ñα╕░α╣Çα╕¡α╕╡α╕óα╕öα╕ùα╕▒α╣ëα╕çα╕½α╕íα╕ö (α╕íα╕╡α╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕úα╕ûα╣üα╕Ñα╕░α╕úα╕░α╕óα╕░α╕ùα╕▓α╕ç)
                            with st.expander("≡ƒôï α╕öα╕╣α╕úα╕▓α╕óα╕Ñα╕░α╣Çα╕¡α╕╡α╕óα╕öα╕úα╕▓α╕óα╕¬α╕▓α╕éα╕▓ (α╣Çα╕úα╕╡α╕óα╕çα╕òα╕▓α╕íα╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü)"):
                                # α╕êα╕▒α╕öα╣Çα╕úα╕╡α╕óα╕çα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕ùα╕╡α╣êα╕¬α╕│α╕äα╕▒α╕ì
                                display_cols = ['Trip', 'Code', 'Name']
                                if 'Province' in result_df.columns:
                                    display_cols.append('Province')
                                if 'Region' in result_df.columns:
                                    display_cols.append('Region')
                                display_cols.extend(['Max_Distance_in_Trip', 'Weight', 'Cube', 'Truck', 'VehicleCheck'])
                                
                                # α╕üα╕úα╕¡α╕çα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕ùα╕╡α╣êα╕íα╕╡α╕¡α╕óα╕╣α╣êα╕êα╕úα╕┤α╕ç
                                display_cols = [col for col in display_cols if col in result_df.columns]
                                display_df = result_df[display_cols].copy()
                                
                                # α╕òα╕▒α╣ëα╕çα╕èα╕╖α╣êα╕¡α╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕áα╕▓α╕⌐α╕▓α╣äα╕ùα╕ó
                                col_names = {'Trip': 'α╕ùα╕úα╕┤α╕¢', 'Code': 'α╕úα╕½α╕▒α╕¬', 'Name': 'α╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓', 'Province': 'α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', 
                                           'Region': 'α╕áα╕▓α╕ä', 'Max_Distance_in_Trip': 'α╕úα╕░α╕óα╕░α╕ùα╕▓α╕ç Max(km)', 
                                           'Weight': 'α╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü(kg)', 'Cube': 'α╕äα╕┤α╕º(m┬│)', 'Truck': 'α╕úα╕û', 'VehicleCheck': 'α╕òα╕úα╕ºα╕êα╕¬α╕¡α╕Üα╕úα╕û'}
                                display_df.columns = [col_names.get(c, c) for c in display_cols]
                                
                                # α╕êα╕▒α╕öα╕úα╕╣α╕¢α╣üα╕Üα╕Üα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕úα╕░α╕óα╕░α╕ùα╕▓α╕ç
                                st.dataframe(
                                    display_df.style.format({
                                        'α╕úα╕░α╕óα╕░α╕ùα╕▓α╕ç(km)': '{:.1f}',
                                        'α╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü(kg)': '{:.2f}',
                                        'α╕äα╕┤α╕º(m┬│)': '{:.2f}'
                                    }),
                                    use_container_width=True, 
                                    height=400
                                )
                            
                            # α╣üα╕¬α╕öα╕çα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╕íα╕╡α╕äα╕│α╣Çα╕òα╕╖α╕¡α╕Ö
                            warning_branches = result_df[result_df['VehicleCheck'].str.contains('ΓÜá∩╕Å', na=False)]
                            if len(warning_branches) > 0:
                                with st.expander(f"ΓÜá∩╕Å α╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣âα╕èα╣ëα╕úα╕ûα╕òα╣êα╕▓α╕çα╕êα╕▓α╕üα╕¢α╕üα╕òα╕┤ ({len(warning_branches)} α╕¬α╕▓α╕éα╕▓)"):
                                    st.warning("α╕¬α╕▓α╕éα╕▓α╣Çα╕½α╕Ñα╣êα╕▓α╕Öα╕╡α╣ëα╕¢α╕üα╕òα╕┤α╣âα╕èα╣ëα╕úα╕ûα╕¢α╕úα╕░α╣Çα╕áα╕ùα╕¡α╕╖α╣êα╕Ö α╣üα╕òα╣êα╕ûα╕╣α╕üα╕êα╕▒α╕öα╣âα╕½α╣ëα╣âα╕èα╣ëα╕úα╕ûα╕¢α╕úα╕░α╣Çα╕áα╕ùα╕ùα╕╡α╣êα╕òα╣êα╕▓α╕çα╕¡α╕¡α╕üα╣äα╕¢")
                                    display_cols_warn = ['Trip', 'Code', 'Name', 'Truck', 'VehicleCheck']
                                    display_warn_df = warning_branches[display_cols_warn].copy()
                                    display_warn_df.columns = ['α╕ùα╕úα╕┤α╕¢', 'α╕úα╕½α╕▒α╕¬', 'α╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓', 'α╕úα╕ûα╕ùα╕╡α╣êα╕êα╕▒α╕ö', 'α╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤α╕üα╕▓α╕úα╣âα╕èα╣ëα╕úα╕û']
                                    st.dataframe(display_warn_df, use_container_width=True)
                            
                            st.markdown("---")
                            
                            # α╕öα╕▓α╕ºα╕Öα╣îα╣éα╕½α╕Ñα╕ö - α╣Çα╕éα╕╡α╕óα╕Öα╕ùα╕▒α╕Üα╕èα╕╡α╕ò 2.Punthai α╣âα╕Öα╣äα╕ƒα╕Ñα╣îα╕òα╣ëα╕Öα╕ëα╕Üα╕▒α╕Ü α╕₧α╕úα╣ëα╕¡α╕íα╕¬α╕Ñα╕▒α╕Üα╕¬α╕╡α╣Çα╕½α╕Ñα╕╖α╕¡α╕çα╣éα╕ùα╕Öα╕¬α╣ëα╕í-α╕éα╕▓α╕º
                            from openpyxl import load_workbook
                            from openpyxl.styles import PatternFill, Font, Border, Side
                            
                            output = io.BytesIO()
                            
                            # α╕¬α╕úα╣ëα╕▓α╕ç location_map α╕êα╕▓α╕ü MASTER_DATA
                            location_map = {}
                            if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                                for _, row in MASTER_DATA.iterrows():
                                    code = str(row.get('Plan Code', '')).strip().upper()
                                    if code:
                                        location_map[code] = {
                                            'α╕òα╕│α╕Üα╕Ñ': row.get('α╕òα╕│α╕Üα╕Ñ', ''),
                                            'α╕¡α╕│α╣Çα╕áα╕¡': row.get('α╕¡α╕│α╣Çα╕áα╕¡', ''),
                                            'α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö': row.get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', ''),
                                            'Route': row.get('Reference', '')
                                        }
                            
                            # α╕¬α╕úα╣ëα╕▓α╕ç Trip_No map
                            trip_no_map = {}
                            vehicle_counts = {'4W': 0, '4WJ': 0, '6W': 0}
                            
                            # α╣Çα╕úα╕╡α╕óα╕ç trip α╕òα╕▓α╕í Zone Order + Province Max Dist + District Max Dist (α╣Çα╕½α╕íα╕╖α╕¡α╕Öα╕òα╕¡α╕Öα╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢)
                            ZONE_ORDER_EXPORT = {'NORTH': 1, 'NE': 2, 'SOUTH': 3, 'EAST': 4, 'WEST': 5, 'CENTRAL': 6}
                            trip_sort_keys = {}
                            
                            for trip_num in result_df['Trip'].unique():
                                if trip_num == 0:
                                    continue
                                trip_data = result_df[result_df['Trip'] == trip_num]
                                
                                # α╕½α╕▓ Region Order
                                region = trip_data['Region'].iloc[0] if 'Region' in trip_data.columns else 'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕'
                                region_order = ZONE_ORDER_EXPORT.get(region, 99)
                                
                                # α╕½α╕▓ Province Max Distance α╣üα╕Ñα╕░ District Max Distance
                                prov_max_dist = 0
                                dist_max_dist = 0
                                
                                for code in trip_data['Code'].unique():
                                    loc = location_map.get(str(code).upper(), {})
                                    # α╕öα╕╢α╕çα╕úα╕░α╕óα╕░α╕ùα╕▓α╕çα╕êα╕▓α╕ü MASTER_DATA
                                    if not MASTER_DATA.empty:
                                        master_row = MASTER_DATA[MASTER_DATA['Plan Code'].astype(str).str.upper() == str(code).upper()]
                                        if len(master_row) > 0:
                                            dist_km = master_row.iloc[0].get('Distance from DC (km)', 0)
                                            if pd.notna(dist_km):
                                                prov_max_dist = max(prov_max_dist, float(dist_km))
                                                dist_max_dist = max(dist_max_dist, float(dist_km))
                                
                                # Sort key: Region Order (Asc), Prov Max Dist (Desc), Dist Max Dist (Desc)
                                # α╣âα╕èα╣ëα╕äα╣êα╕▓α╕Ñα╕Üα╣Çα╕₧α╕╖α╣êα╕¡α╣âα╕½α╣ë sort Desc
                                trip_sort_keys[trip_num] = (region_order, -prov_max_dist, -dist_max_dist)
                            
                            # Sort: Zone Order ΓåÆ Province Max Dist (α╣äα╕üα╕Ñα╕üα╣êα╕¡α╕Ö) ΓåÆ District Max Dist (α╣äα╕üα╕Ñα╕üα╣êα╕¡α╕Ö)
                            sorted_trips = sorted(
                                [t for t in result_df['Trip'].unique() if t != 0],
                                key=lambda t: trip_sort_keys.get(t, (99, 0, 0))
                            )
                            
                            for trip_num in sorted_trips:
                                trip_summary = summary[summary['Trip'] == trip_num]
                                if len(trip_summary) > 0:
                                    truck_info = trip_summary.iloc[0]['Truck']
                                    vehicle_type = truck_info.split()[0] if truck_info else '6W'
                                    # JB α╣âα╕èα╣ë prefix 4WJ
                                    if vehicle_type == 'JB':
                                        vehicle_type = '4WJ'
                                    vehicle_counts[vehicle_type] = vehicle_counts.get(vehicle_type, 0) + 1
                                    trip_no = f"{vehicle_type}{vehicle_counts[vehicle_type]:03d}"
                                    trip_no_map[trip_num] = trip_no
                            
                            try:
                                # α╣éα╕½α╕Ñα╕ö workbook α╕òα╣ëα╕Öα╕ëα╕Üα╕▒α╕Ü
                                wb = load_workbook(io.BytesIO(st.session_state.get('original_file_content', b'')))
                                
                                # α╕½α╕▓α╕èα╕╡α╕òα╣Çα╕¢α╣ëα╕▓α╕½α╕íα╕▓α╕ó (2.Punthai)
                                target_sheet = None
                                for sheet_name in wb.sheetnames:
                                    if 'punthai' in sheet_name.lower() or '2.' in sheet_name.lower():
                                        target_sheet = sheet_name
                                        break
                                
                                if not target_sheet:
                                    target_sheet = '2.Punthai'
                                    if target_sheet not in wb.sheetnames:
                                        wb.create_sheet(target_sheet)
                                
                                ws = wb[target_sheet]
                                
                                # α╕½α╕▓ header row
                                header_row = 1
                                for row_idx in range(1, min(5, ws.max_row + 1)):
                                    for col_idx in range(1, min(15, ws.max_column + 1)):
                                        cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                                        if 'α╕úα╕½α╕▒α╕¬α╕¬α╕▓α╕éα╕▓' in cell_val or 'Trip' in cell_val.upper():
                                            header_row = row_idx
                                            break
                                
                                # α╕Ñα╕Üα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╣Çα╕üα╣êα╕▓
                                if ws.max_row > header_row:
                                    ws.delete_rows(header_row + 1, ws.max_row - header_row)
                                
                                # α╣Çα╕éα╕╡α╕óα╕Ö header α╣âα╕½α╕íα╣ê
                                new_headers = ['Sep.', 'BU', 'α╕úα╕½α╕▒α╕¬α╕¬α╕▓α╕éα╕▓', 'α╕úα╕½α╕▒α╕¬ WMS', 'α╕¬α╕▓α╕éα╕▓', 'α╕òα╕│α╕Üα╕Ñ', 'α╕¡α╕│α╣Çα╕áα╕¡', 'α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', 'Route',
                                              'Total Cube', 'Total Wgt', 'Original QTY', 'Trip', 'Trip no']
                                for col_idx, header_val in enumerate(new_headers, 1):
                                    ws.cell(row=header_row, column=col_idx, value=header_val)
                                
                                # α╕¬α╕╡α╣Çα╕½α╕Ñα╕╖α╕¡α╕çα╣éα╕ùα╕Öα╕¬α╣ëα╕í-α╕éα╕▓α╕º (α╕¬α╕Ñα╕▒α╕Ü 2 α╕¬α╕╡)
                                yellow_orange = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                                white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                                thin_border = Border(
                                    left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin')
                                )
                                red_font = Font(color='FF0000', bold=True)
                                
                                # α╕½α╕▓α╕ùα╕úα╕┤α╕¢α╕ùα╕╡α╣êα╣äα╕íα╣êα╕£α╣êα╕▓α╕Öα╣Çα╕üα╕ôα╕æα╣î
                                failed_trips = set()
                                vehicle_limits = {'4W': {'max_w': 2500, 'max_c': 5.0}, 'JB': {'max_w': 3500, 'max_c': 7.0}, '6W': {'max_w': 6000, 'max_c': 20.0}}
                                for t in result_df['Trip'].unique():
                                    if t == 0:
                                        continue
                                    trip_data = result_df[result_df['Trip'] == t]
                                    trip_cube = trip_data['Cube'].sum()
                                    trip_weight = trip_data['Weight'].sum()
                                    trip_no = trip_no_map.get(t, '6W001')
                                    veh_type = 'JB' if trip_no.startswith('4WJ') else ('4W' if trip_no.startswith('4W') else '6W')
                                    limits = vehicle_limits.get(veh_type, vehicle_limits['6W'])
                                    max_util = max((trip_cube / limits['max_c']) * 100, (trip_weight / limits['max_w']) * 100)
                                    if max_util > 105 or max_util < 50:
                                        failed_trips.add(t)
                                
                                # α╣Çα╕éα╕╡α╕óα╕Öα╕éα╣ëα╕¡α╕íα╕╣α╕Ñ
                                current_trip = None
                                use_yellow = True
                                row_num = header_row + 1
                                sep_num = 1
                                
                                for trip_num in sorted_trips:
                                    trip_data = result_df[result_df['Trip'] == trip_num].copy()
                                    
                                    # Sort α╕òα╕▓α╕í α╕òα╕│α╕Üα╕Ñ ΓåÆ α╕¡α╕│α╣Çα╕áα╕¡ ΓåÆ α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö
                                    trip_data['_sort_sub'] = trip_data['Code'].apply(lambda c: location_map.get(str(c).upper(), {}).get('α╕òα╕│α╕Üα╕Ñ', ''))
                                    trip_data['_sort_dist'] = trip_data['Code'].apply(lambda c: location_map.get(str(c).upper(), {}).get('α╕¡α╕│α╣Çα╕áα╕¡', ''))
                                    trip_data['_sort_prov'] = trip_data['Code'].apply(lambda c: location_map.get(str(c).upper(), {}).get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', ''))
                                    trip_data = trip_data.sort_values(['_sort_prov', '_sort_dist', '_sort_sub', 'Code'])
                                    
                                    trip_no = trip_no_map.get(trip_num, '')
                                    
                                    # α╕¬α╕Ñα╕▒α╕Üα╕¬α╕╡α╣Çα╕íα╕╖α╣êα╕¡α╣Çα╕¢α╕Ñα╕╡α╣êα╕óα╕Öα╕ùα╕úα╕┤α╕¢
                                    if current_trip != trip_num:
                                        current_trip = trip_num
                                        use_yellow = not use_yellow
                                    
                                    fill = yellow_orange if use_yellow else white_fill
                                    
                                    for _, row in trip_data.iterrows():
                                        branch_code = row.get('Code', '')
                                        loc = location_map.get(str(branch_code).upper(), {})
                                        
                                        data = [
                                            sep_num,
                                            row.get('BU', 211),
                                            branch_code,
                                            branch_code,
                                            row.get('Name', ''),
                                            loc.get('α╕òα╕│α╕Üα╕Ñ', ''),
                                            loc.get('α╕¡α╕│α╣Çα╕áα╕¡', ''),
                                            loc.get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', ''),
                                            loc.get('Route', ''),
                                            round(row.get('Cube', 0), 2) if pd.notna(row.get('Cube')) else 0,
                                            round(row.get('Weight', 0), 2) if pd.notna(row.get('Weight')) else 0,
                                            row.get('OriginalQty', 0) if pd.notna(row.get('OriginalQty')) else 0,
                                            int(trip_num),
                                            trip_no,
                                        ]
                                        
                                        for col_idx, value in enumerate(data, 1):
                                            cell = ws.cell(row=row_num, column=col_idx, value=value)
                                            cell.fill = fill
                                            cell.border = thin_border
                                            if trip_num in failed_trips:
                                                cell.font = red_font
                                        
                                        row_num += 1
                                        sep_num += 1
                                
                                wb.save(output)
                                output.seek(0)
                                
                            except Exception as e:
                                st.warning(f"ΓÜá∩╕Å α╣äα╕íα╣êα╕¬α╕▓α╕íα╕▓α╕úα╕ûα╣Çα╕éα╕╡α╕óα╕Öα╕ùα╕▒α╕Üα╣äα╕ƒα╕Ñα╣îα╕òα╣ëα╕Öα╕ëα╕Üα╕▒α╕Üα╣äα╕öα╣ë: {e} - α╣âα╕èα╣ëα╕úα╕╣α╕¢α╣üα╕Üα╕Üα╕íα╕▓α╕òα╕úα╕Éα╕▓α╕Öα╣üα╕ùα╕Ö")
                                # Fallback: α╕¬α╕úα╣ëα╕▓α╕çα╣äα╕ƒα╕Ñα╣îα╣âα╕½α╕íα╣êα╕öα╣ëα╕ºα╕ó xlsxwriter
                                output = io.BytesIO()
                                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                    export_df = result_df.copy()
                                    export_df['Trip_No'] = export_df['Trip'].map(lambda x: trip_no_map.get(x, ''))
                                    export_df.to_excel(writer, sheet_name='α╕úα╕▓α╕óα╕Ñα╕░α╣Çα╕¡α╕╡α╕óα╕öα╕ùα╕úα╕┤α╕¢', index=False)
                                    summary.to_excel(writer, sheet_name='α╕¬α╕úα╕╕α╕¢α╕ùα╕úα╕┤α╕¢', index=False)
                            
                            col1, col2, col3 = st.columns([1, 2, 1])
                            with col2:
                                st.download_button(
                                    label="≡ƒôÑ α╕öα╕▓α╕ºα╕Öα╣îα╣éα╕½α╕Ñα╕öα╕£α╕Ñα╕Ñα╕▒α╕₧α╕ÿα╣î (Excel)",
                                    data=output.getvalue(),
                                    file_name=f"α╕£α╕Ñα╕êα╕▒α╕öα╕ùα╕úα╕┤α╕¢_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                
                # ==========================================
                # α╣üα╕ùα╣çα╕Ü 2: α╕êα╕▒α╕öα╕üα╕Ñα╕╕α╣êα╕íα╕¬α╕▓α╕éα╕▓α╕òα╕▓α╕íα╕áα╕▓α╕ä (α╣äα╕íα╣êα╕¬α╕Öα╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü)
                # ==========================================
                with tab2:
                    df_region = df.copy()
                    
                    # α╕êα╕▒α╕öα╕üα╕Ñα╕╕α╣êα╕íα╕òα╕▓α╕íα╕áα╕▓α╕ä
                    branch_info = model_data.get('branch_info', {})
                    trip_pairs = model_data.get('trip_pairs', set())
                    
                    # α╕¬α╕úα╣ëα╕▓α╕çα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕áα╕▓α╕äα╕¬α╕│α╕½α╕úα╕▒α╕Üα╣üα╕òα╣êα╕Ñα╕░α╕¬α╕▓α╕éα╕▓ (α╕êα╕▓α╕üα╣äα╕ƒα╕Ñα╣îα╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤)
                    region_groups = {
                        'α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç-α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕èα╕▒α╣ëα╕Öα╣âα╕Ö': ['α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕íα╕½α╕▓α╕Öα╕äα╕ú'],
                        'α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç-α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕èα╕▒α╣ëα╕Öα╕üα╕Ñα╕▓α╕ç': ['α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕íα╕½α╕▓α╕Öα╕äα╕ú'],
                        'α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç-α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕èα╕▒α╣ëα╕Öα╕Öα╕¡α╕ü': ['α╕üα╕úα╕╕α╕çα╣Çα╕ùα╕₧α╕íα╕½α╕▓α╕Öα╕äα╕ú'],
                        'α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç-α╕¢α╕úα╕┤α╕íα╕ôα╕æα╕Ñ': ['α╕Öα╕äα╕úα╕¢α╕Éα╕í', 'α╕Öα╕Öα╕ùα╕Üα╕╕α╕úα╕╡', 'α╕¢α╕ùα╕╕α╕íα╕ÿα╕▓α╕Öα╕╡', 'α╕¬α╕íα╕╕α╕ùα╕úα╕¢α╕úα╕▓α╕üα╕▓α╕ú', 'α╕¬α╕íα╕╕α╕ùα╕úα╕¬α╕▓α╕äα╕ú'],
                        'α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç-α╕üα╕Ñα╕▓α╕çα╕òα╕¡α╕Öα╕Üα╕Ö': ['α╕èα╕▒α╕óα╕Öα╕▓α╕ù', 'α╕₧α╕úα╕░α╕Öα╕äα╕úα╕¿α╕úα╕╡α╕¡α╕óα╕╕α╕ÿα╕óα╕▓', 'α╕Ñα╕₧α╕Üα╕╕α╕úα╕╡', 'α╕¬α╕úα╕░α╕Üα╕╕α╕úα╕╡', 'α╕¬α╕┤α╕çα╕½α╣îα╕Üα╕╕α╕úα╕╡', 'α╕¡α╣êα╕▓α╕çα╕ùα╕¡α╕ç', 'α╕¡α╕óα╕╕α╕ÿα╕óα╕▓'],
                        'α╕áα╕▓α╕äα╕üα╕Ñα╕▓α╕ç-α╕üα╕Ñα╕▓α╕çα╕òα╕¡α╕Öα╕Ñα╣êα╕▓α╕ç': ['α╕¬α╕íα╕╕α╕ùα╕úα╕¬α╕çα╕äα╕úα╕▓α╕í', 'α╕¬α╕╕α╕₧α╕úα╕úα╕ôα╕Üα╕╕α╕úα╕╡'],
                        'α╕áα╕▓α╕äα╕òα╕░α╕ºα╕▒α╕Öα╕òα╕ü': ['α╕üα╕▓α╕ìα╕êα╕Öα╕Üα╕╕α╕úα╕╡', 'α╕¢α╕úα╕░α╕êα╕ºα╕Üα╕äα╕╡α╕úα╕╡α╕éα╕▒α╕Öα╕ÿα╣î', 'α╕úα╕▓α╕èα╕Üα╕╕α╕úα╕╡', 'α╣Çα╕₧α╕èα╕úα╕Üα╕╕α╕úα╕╡'],
                        'α╕áα╕▓α╕äα╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü': ['α╕êα╕▒α╕Öα╕ùα╕Üα╕╕α╕úα╕╡', 'α╕èα╕Ñα╕Üα╕╕α╕úα╕╡', 'α╕òα╕úα╕▓α╕ö', 'α╕Öα╕äα╕úα╕Öα╕▓α╕óα╕ü', 'α╕¢α╕úα╕▓α╕êα╕╡α╕Öα╕Üα╕╕α╕úα╕╡', 'α╕úα╕░α╕óα╕¡α╕ç', 'α╕¬α╕úα╕░α╣üα╕üα╣ëα╕º', 'α╕ëα╕░α╣Çα╕èα╕┤α╕çα╣Çα╕ùα╕úα╕▓'],
                        'α╕áα╕▓α╕äα╕¡α╕╡α╕¬α╕▓α╕Ö-α╕¡α╕╡α╕¬α╕▓α╕Öα╣Çα╕½α╕Öα╕╖α╕¡': ['α╕Öα╕äα╕úα╕₧α╕Öα╕í', 'α╕Üα╕╢α╕çα╕üα╕▓α╕¼', 'α╕íα╕╕α╕üα╕öα╕▓α╕½α╕▓α╕ú', 'α╕¬α╕üα╕Ñα╕Öα╕äα╕ú', 'α╕½α╕Öα╕¡α╕çα╕äα╕▓α╕ó', 'α╕½α╕Öα╕¡α╕çα╕Üα╕▒α╕ºα╕Ñα╕│α╕áα╕╣', 'α╕¡α╕╕α╕öα╕úα╕ÿα╕▓α╕Öα╕╡', 'α╣Çα╕Ñα╕ó'],
                        'α╕áα╕▓α╕äα╕¡α╕╡α╕¬α╕▓α╕Ö-α╕¡α╕╡α╕¬α╕▓α╕Öα╕üα╕Ñα╕▓α╕ç': ['α╕üα╕▓α╕¼α╕¬α╕┤α╕Öα╕ÿα╕╕α╣î', 'α╕éα╕¡α╕Öα╣üα╕üα╣êα╕Ö', 'α╕èα╕▒α╕óα╕áα╕╣α╕íα╕┤', 'α╕íα╕½α╕▓α╕¬α╕▓α╕úα╕äα╕▓α╕í', 'α╕úα╣ëα╕¡α╕óα╣Çα╕¡α╣çα╕ö'],
                        'α╕áα╕▓α╕äα╕¡α╕╡α╕¬α╕▓α╕Ö-α╕¡α╕╡α╕¬α╕▓α╕Öα╣âα╕òα╣ë': ['α╕Öα╕äα╕úα╕úα╕▓α╕èα╕¬α╕╡α╕íα╕▓', 'α╣éα╕äα╕úα╕▓α╕è', 'α╕Üα╕╕α╕úα╕╡α╕úα╕▒α╕íα╕óα╣î', 'α╕óα╣éα╕¬α╕ÿα╕ú', 'α╕¿α╕úα╕╡α╕¬α╕░α╣Çα╕üα╕⌐', 'α╕¬α╕╕α╕úα╕┤α╕Öα╕ùα╕úα╣î', 'α╕¡α╕│α╕Öα╕▓α╕êα╣Çα╕êα╕úα╕┤α╕ì', 'α╕¡α╕╕α╕Üα╕Ñα╕úα╕▓α╕èα╕ÿα╕▓α╕Öα╕╡'],
                        'α╕áα╕▓α╕äα╣Çα╕½α╕Öα╕╖α╕¡-α╣Çα╕½α╕Öα╕╖α╕¡α╕òα╕¡α╕Öα╕Üα╕Ö': ['α╕Öα╣êα╕▓α╕Ö', 'α╕₧α╕░α╣Çα╕óα╕▓', 'α╕Ñα╕│α╕¢α╕▓α╕ç', 'α╕Ñα╕│α╕₧α╕╣α╕Ö', 'α╣Çα╕èα╕╡α╕óα╕çα╕úα╕▓α╕ó', 'α╣Çα╕èα╕╡α╕óα╕çα╣âα╕½α╕íα╣ê', 'α╣üα╕₧α╕úα╣ê', 'α╣üα╕íα╣êα╕«α╣êα╕¡α╕çα╕¬α╕¡α╕Ö'],
                        'α╕áα╕▓α╕äα╣Çα╕½α╕Öα╕╖α╕¡-α╣Çα╕½α╕Öα╕╖α╕¡α╕òα╕¡α╕Öα╕Ñα╣êα╕▓α╕ç': ['α╕üα╕│α╣üα╕₧α╕çα╣Çα╕₧α╕èα╕ú', 'α╕òα╕▓α╕ü', 'α╕Öα╕äα╕úα╕¬α╕ºα╕úα╕úα╕äα╣î', 'α╕₧α╕┤α╕êα╕┤α╕òα╕ú', 'α╕₧α╕┤α╕⌐α╕ôα╕╕α╣éα╕Ñα╕ü', 'α╕¬α╕╕α╣éα╕éα╕ùα╕▒α╕ó', 'α╕¡α╕╕α╕òα╕úα╕öα╕┤α╕òα╕ûα╣î', 'α╕¡α╕╕α╕ùα╕▒α╕óα╕ÿα╕▓α╕Öα╕╡', 'α╣Çα╕₧α╕èα╕úα╕Üα╕╣α╕úα╕ôα╣î'],
                        'α╕áα╕▓α╕äα╣âα╕òα╣ë-α╣âα╕òα╣ëα╕¥α╕▒α╣êα╕çα╕¡α╕▒α╕Öα╕öα╕▓α╕íα╕▒α╕Ö': ['α╕üα╕úα╕░α╕Üα╕╡α╣ê', 'α╕òα╕úα╕▒α╕ç', 'α╕₧α╕▒α╕çα╕çα╕▓', 'α╕áα╕╣α╣Çα╕üα╣çα╕ò', 'α╕úα╕░α╕Öα╕¡α╕ç', 'α╕¬α╕òα╕╣α╕Ñ'],
                        'α╕áα╕▓α╕äα╣âα╕òα╣ë-α╣âα╕òα╣ëα╕¥α╕▒α╣êα╕çα╕¡α╣êα╕▓α╕ºα╣äα╕ùα╕ó': ['α╕èα╕╕α╕íα╕₧α╕ú', 'α╕Öα╕äα╕úα╕¿α╕úα╕╡α╕ÿα╕úα╕úα╕íα╕úα╕▓α╕è', 'α╕₧α╕▒α╕ùα╕Ñα╕╕α╕ç', 'α╕óα╕░α╕Ñα╕▓', 'α╕¬α╕çα╕éα╕Ñα╕▓', 'α╕¬α╕╕α╕úα╕▓α╕⌐α╕Äα╕úα╣îα╕ÿα╕▓α╕Öα╕╡', 'α╕¢α╕▒α╕òα╕òα╕▓α╕Öα╕╡', 'α╕Öα╕úα╕▓α╕ÿα╕┤α╕ºα╕▓α╕¬']
                    }
                    
                    def get_region(province):
                        if pd.isna(province) or not province or str(province).strip() in ['', 'nan', 'UNKNOWN']:
                            return 'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕'
                        
                        # ≡ƒÜ¿ Override: α╕ëα╕░α╣Çα╕èα╕┤α╕çα╣Çα╕ùα╕úα╕▓ ΓåÆ α╕áα╕▓α╕äα╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü (α╣äα╕íα╣êα╣âα╕èα╣êα╕¢α╕úα╕┤α╕íα╕ôα╕æα╕Ñ)
                        if 'α╕ëα╕░α╣Çα╕èα╕┤α╕çα╣Çα╕ùα╕úα╕▓' in str(province):
                            return 'α╕áα╕▓α╕äα╕òα╕░α╕ºα╕▒α╕Öα╕¡α╕¡α╕ü'
                        
                        for region, provinces in region_groups.items():
                            if any(p in str(province) for p in provinces):
                                return region
                        return 'α╕¡α╕╖α╣êα╕Öα╣å'
                    
                    # α╣Çα╕₧α╕┤α╣êα╕íα╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣îα╕áα╕▓α╕ä - α╕öα╕╢α╕çα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╕êα╕▓α╕ü Master α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡
                    if 'Province' not in df_region.columns or df_region['Province'].isna().any():
                        # α╕öα╕╢α╕çα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╕êα╕▓α╕ü Master
                        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                            province_map = {}
                            for _, row in MASTER_DATA.iterrows():
                                code = row.get('Plan Code', '')
                                province = row.get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', '')
                                if code and province:
                                    province_map[code] = province
                            
                            # α╣âα╕¬α╣êα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣âα╕½α╣ëα╣üα╕òα╣êα╕Ñα╕░α╕¬α╕▓α╕éα╕▓
                            if 'Province' not in df_region.columns:
                                df_region['Province'] = df_region['Code'].map(province_map)
                            else:
                                # α╣Çα╕òα╕┤α╕íα╣Çα╕ëα╕₧α╕▓α╕░α╕ùα╕╡α╣êα╣Çα╕¢α╣çα╕Ö NaN
                                df_region['Province'] = df_region.apply(
                                    lambda row: province_map.get(row['Code'], row.get('Province', 'UNKNOWN')) 
                                    if pd.isna(row.get('Province')) else row['Province'],
                                    axis=1
                                )
                    
                    df_region['Region'] = df_region['Province'].apply(get_region)
                    
                    # α╕½α╕▓α╕üα╕Ñα╕╕α╣êα╕íα╕¬α╕▓α╕éα╕▓ (α╣âα╕èα╣ë Booking No. α╣Çα╕¢α╣çα╕Öα╕½α╕Ñα╕▒α╕ü)
                    def find_paired_branches(code, code_province, df_data):
                        paired = set()
                        
                        # α╕½α╕▓ Booking No. α╕éα╕¡α╕çα╕¬α╕▓α╕éα╕▓α╕Öα╕╡α╣ë
                        code_rows = df_data[df_data['Code'] == code]
                        if len(code_rows) == 0:
                            return paired
                        
                        # α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓α╕íα╕╡α╕äα╕¡α╕Ñα╕▒α╕íα╕Öα╣î Booking α╕½α╕úα╕╖α╕¡α╣äα╕íα╣ê
                        if 'Booking' not in df_data.columns and 'Trip' not in df_data.columns:
                            return paired
                        
                        booking_col = 'Booking' if 'Booking' in df_data.columns else 'Trip'
                        code_bookings = set(code_rows[booking_col].dropna().astype(str))
                        
                        if not code_bookings:
                            return paired
                        
                        # α╕½α╕▓α╕¬α╕▓α╕éα╕▓α╕¡α╕╖α╣êα╕Öα╕ùα╕╡α╣êα╕¡α╕óα╕╣α╣ê Booking α╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Ö (α╣äα╕íα╣êα╕¬α╕Öα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö)
                        for booking in code_bookings:
                            if booking == 'nan' or not booking.strip():
                                continue
                            
                            same_booking = df_data[df_data[booking_col].astype(str) == booking]
                            for _, other_row in same_booking.iterrows():
                                other_code = other_row['Code']
                                
                                # α╣Çα╕çα╕╖α╣êα╕¡α╕Öα╣äα╕é: Booking α╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Ö = α╕úα╕ºα╕íα╕üα╕Ñα╕╕α╣êα╕í (α╣äα╕íα╣êα╕¬α╕Öα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö)
                                if other_code != code:
                                    paired.add(other_code)
                        
                        return paired
                    
                    all_codes_set = set(df_region['Code'].unique())
                    
                    # α╕¬α╕úα╣ëα╕▓α╕çα╕üα╕Ñα╕╕α╣êα╕íα╕¬α╕▓α╕éα╕▓α╣üα╕Üα╕Ü Union-Find (α╕òα╕▓α╕íα╕Ñα╕│α╕öα╕▒α╕Ü: α╕òα╕│α╕Üα╕Ñ ΓåÆ α╕¡α╕│α╣Çα╕áα╕¡ ΓåÆ α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö)
                    # Step 1: α╣Çα╕úα╕┤α╣êα╕íα╕êα╕▓α╕üα╣üα╕òα╣êα╕Ñα╕░α╕¬α╕▓α╕éα╕▓α╣Çα╕¢α╣çα╕Öα╕üα╕Ñα╕╕α╣êα╕íα╣å α╕₧α╕úα╣ëα╕¡α╕íα╕éα╣ëα╕¡α╕íα╕╣α╕Ñ Master
                    initial_groups = {}
                    for code in all_codes_set:
                        # α╕öα╕╢α╕çα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕êα╕▓α╕ü Master
                        location = {}
                        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                            if len(master_row) > 0:
                                master_row = master_row.iloc[0]
                                location = {
                                    'subdistrict': master_row.get('α╕òα╕│α╕Üα╕Ñ', ''),
                                    'district': master_row.get('α╕¡α╕│α╣Çα╕áα╕¡', ''),
                                    'province': master_row.get('α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', 'UNKNOWN'),
                                    'lat': master_row.get('α╕Ñα╕░α╕òα╕┤α╕êα╕╣α╕ö', 0),
                                    'lon': master_row.get('α╕Ñα╕¡α╕çα╕òα╕┤α╕êα╕╣α╕ö', 0)
                                }
                        
                        # α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡α╣âα╕Ö Master α╕Ñα╕¡α╕çα╕öα╕╢α╕çα╕êα╕▓α╕üα╣äα╕ƒα╕Ñα╣îα╕¡α╕▒α╕¢α╣éα╕½α╕Ñα╕ö
                        if not location or location.get('province', 'UNKNOWN') == 'UNKNOWN':
                            c_row = df_region[df_region['Code'] == code].iloc[0] if len(df_region[df_region['Code'] == code]) > 0 else None
                            if c_row is not None:
                                location = {
                                    'subdistrict': '',
                                    'district': '',
                                    'province': c_row.get('Province', 'UNKNOWN'),
                                    'lat': 0,
                                    'lon': 0
                                }
                        
                        if location:
                            initial_groups[(code,)] = {code: location}
                    
                    # α╣âα╕èα╣ë initial_groups α╣üα╕ùα╕Ö booking_groups
                    booking_groups = initial_groups
                    
                    # Step 2: α╕úα╕ºα╕íα╕üα╕Ñα╕╕α╣êα╕íα╕òα╕▓α╕íα╕Ñα╕│α╕öα╕▒α╕Ü α╕òα╕│α╕Üα╕Ñ ΓåÆ α╕¡α╕│α╣Çα╕áα╕¡ ΓåÆ α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö
                    def groups_can_merge(locs1, locs2):
                        """α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓ 2 α╕üα╕Ñα╕╕α╣êα╕íα╕äα╕ºα╕úα╕úα╕ºα╕íα╕üα╕▒α╕Öα╣äα╕½α╕í (α╕òα╕▓α╕íα╕Ñα╕│α╕öα╕▒α╕Üα╕äα╕ºα╕▓α╕íα╕Ñα╕░α╣Çα╕¡α╕╡α╕óα╕ö)"""
                        # 1. α╣Çα╕èα╣çα╕äα╕òα╕│α╕Üα╕Ñα╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Ö (α╕òα╣ëα╕¡α╕çα╕íα╕╡α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕òα╕│α╕Üα╕Ñ)
                        subdistricts1 = set(loc.get('subdistrict', '') for loc in locs1.values() if loc.get('subdistrict', ''))
                        subdistricts2 = set(loc.get('subdistrict', '') for loc in locs2.values() if loc.get('subdistrict', ''))
                        if subdistricts1 and subdistricts2 and (subdistricts1 & subdistricts2):
                            return True, 'α╕òα╕│α╕Üα╕Ñ'
                        
                        # 2. α╣Çα╕èα╣çα╕äα╕¡α╕│α╣Çα╕áα╕¡α╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Ö (α╕òα╣ëα╕¡α╕çα╕íα╕╡α╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕¡α╕│α╣Çα╕áα╕¡α╣üα╕Ñα╕░α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Ö)
                        districts1 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs1.values() if loc.get('district', '')}
                        districts2 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs2.values() if loc.get('district', '')}
                        if districts1 and districts2:
                            # α╣Çα╕èα╣çα╕äα╕ºα╣êα╕▓α╕íα╕╡α╕¡α╕│α╣Çα╕áα╕¡α╣üα╕Ñα╕░α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╕òα╕úα╕çα╕üα╕▒α╕Ö
                            for d1, p1 in districts1:
                                for d2, p2 in districts2:
                                    if d1 == d2 and p1 == p2 and p1:
                                        return True, 'α╕¡α╕│α╣Çα╕áα╕¡'
                        
                        # 3. α╣Çα╕èα╣çα╕äα╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣Çα╕öα╕╡α╕óα╕ºα╕üα╕▒α╕Ö
                        provinces1 = set(loc.get('province', '') for loc in locs1.values() if loc.get('province', ''))
                        provinces2 = set(loc.get('province', '') for loc in locs2.values() if loc.get('province', ''))
                        if provinces1 & provinces2:
                            return True, 'α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö'
                        
                        return False, None
                    
                    merged_groups = []
                    used_groups = set()
                    
                    for group1, locs1 in booking_groups.items():
                        if group1 in used_groups:
                            continue
                        
                        merged_codes = set(group1)
                        merged_locs = locs1.copy()
                        used_groups.add(group1)
                        
                        # α╕½α╕▓α╕üα╕Ñα╕╕α╣êα╕íα╕¡α╕╖α╣êα╕Öα╕ùα╕╡α╣êα╣âα╕üα╕Ñα╣ëα╣Çα╕äα╕╡α╕óα╕ç
                        changed = True
                        while changed:
                            changed = False
                            for group2, locs2 in booking_groups.items():
                                if group2 in used_groups:
                                    continue
                                can_merge, level = groups_can_merge(merged_locs, locs2)
                                if can_merge:
                                    merged_codes |= set(group2)
                                    merged_locs.update(locs2)
                                    used_groups.add(group2)
                                    changed = True
                        
                        merged_groups.append({
                            'codes': merged_codes,
                            'locations': merged_locs
                        })
                    
                    # Step 3: α╣üα╕¢α╕Ñα╕çα╣Çα╕¢α╣çα╕Ö groups format
                    groups = []
                    for mg in merged_groups:
                        rep_code = list(mg['codes'])[0]
                        rep_row = df_region[df_region['Code'] == rep_code].iloc[0]
                        # α╕üα╕úα╕¡α╕çα╣Çα╕ëα╕₧α╕▓α╕░α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╕ùα╕╡α╣êα╣äα╕íα╣êα╣âα╕èα╣ê UNKNOWN α╣üα╕Ñα╕░α╣äα╕íα╣êα╣Çα╕¢α╣çα╕Ö NaN
                        provinces = set(
                            str(loc.get('province', '')).strip() 
                            for loc in mg['locations'].values() 
                            if loc.get('province') and str(loc.get('province', '')).strip() not in ['UNKNOWN', 'nan', '']
                        )
                        
                        # α╕ûα╣ëα╕▓α╣äα╕íα╣êα╕íα╕╡α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕öα╣Çα╕Ñα╕ó α╣âα╕¬α╣ê "α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕"
                        province_str = ', '.join(sorted(provinces)) if provinces else 'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕'
                        
                        groups.append({
                            'codes': mg['codes'],
                            'region': rep_row.get('Region', 'α╣äα╕íα╣êα╕úα╕░α╕Üα╕╕'),
                            'province': province_str
                        })
                    
                    # α╣üα╕¬α╕öα╕çα╕¬α╕ûα╕┤α╕òα╕┤
                    st.markdown("---")
                    st.markdown("### ≡ƒôè α╕¬α╕úα╕╕α╕¢α╕üα╕▓α╕úα╕êα╕▒α╕öα╕üα╕Ñα╕╕α╣êα╕í")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("≡ƒôì α╕êα╕│α╕Öα╕ºα╕Öα╕¬α╕▓α╕éα╕▓", df_region['Code'].nunique())
                    with col2:
                        st.metric("≡ƒùé∩╕Å α╕êα╕│α╕Öα╕ºα╕Öα╕üα╕Ñα╕╕α╣êα╕í", len(groups))
                    with col3:
                        regions_count = df_region['Region'].nunique()
                        st.metric("≡ƒù║∩╕Å α╕êα╕│α╕Öα╕ºα╕Öα╕áα╕▓α╕ä", regions_count)
                    
                    # α╣üα╕¬α╕öα╕çα╕òα╕▓α╕íα╕áα╕▓α╕ä
                    st.markdown("---")
                    st.markdown("### ≡ƒù║∩╕Å α╕¬α╕▓α╕éα╕▓α╣üα╕óα╕üα╕òα╕▓α╕íα╕áα╕▓α╕ä")
                    
                    region_summary = df_region.groupby('Region').agg({
                        'Code': 'nunique',
                        'Weight': 'sum',
                        'Cube': 'sum'
                    }).reset_index()
                    region_summary.columns = ['α╕áα╕▓α╕ä', 'α╕êα╕│α╕Öα╕ºα╕Öα╕¬α╕▓α╕éα╕▓', 'α╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕üα╕úα╕ºα╕í', 'α╕äα╕┤α╕ºα╕úα╕ºα╕í']
                    st.dataframe(region_summary, use_container_width=True)
                    
                    # α╣üα╕¬α╕öα╕çα╕úα╕▓α╕óα╕Ñα╕░α╣Çα╕¡α╕╡α╕óα╕öα╣üα╕òα╣êα╕Ñα╕░α╕áα╕▓α╕ä
                    for region in sorted(df_region['Region'].unique()):
                        region_data = df_region[df_region['Region'] == region]
                        with st.expander(f"≡ƒôì {region} ({region_data['Code'].nunique()} α╕¬α╕▓α╕éα╕▓)"):
                            display_cols = ['Code', 'Name', 'Province', 'Weight', 'Cube']
                            display_cols = [c for c in display_cols if c in region_data.columns]
                            
                            region_display = region_data[display_cols].drop_duplicates('Code')
                            col_names = {'Code': 'α╕úα╕½α╕▒α╕¬', 'Name': 'α╕èα╕╖α╣êα╕¡α╕¬α╕▓α╕éα╕▓', 'Province': 'α╕êα╕▒α╕çα╕½α╕ºα╕▒α╕ö', 'Weight': 'α╕Öα╣ëα╕│α╕½α╕Öα╕▒α╕ü', 'Cube': 'α╕äα╕┤α╕º'}
                            region_display.columns = [col_names.get(c, c) for c in display_cols]
                            st.dataframe(region_display, use_container_width=True)
                    
                    # α╣üα╕¬α╕öα╕çα╕üα╕Ñα╕╕α╣êα╕íα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣Çα╕äα╕óα╣äα╕¢α╕öα╣ëα╕ºα╕óα╕üα╕▒α╕Ö
                    st.markdown("---")
                    st.markdown("### ≡ƒöù α╕üα╕Ñα╕╕α╣êα╕íα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣Çα╕äα╕óα╣äα╕¢α╕öα╣ëα╕ºα╕óα╕üα╕▒α╕Ö (α╕êα╕▓α╕üα╕¢α╕úα╕░α╕ºα╕▒α╕òα╕┤)")
                    
                    paired_groups = [g for g in groups if len(g['codes']) > 1]
                    if paired_groups:
                        for i, group in enumerate(paired_groups, 1):
                            codes_list = list(group['codes'])
                            names = []
                            for c in codes_list:
                                name_row = df_region[df_region['Code'] == c]
                                if len(name_row) > 0 and 'Name' in name_row.columns:
                                    names.append(f"{c} ({name_row['Name'].iloc[0]})")
                                else:
                                    names.append(c)
                            
                            st.write(f"**α╕üα╕Ñα╕╕α╣êα╕í {i}** - {group['region']}: {', '.join(names)}")
                    else:
                        st.info("α╣äα╕íα╣êα╕₧α╕Üα╕üα╕Ñα╕╕α╣êα╕íα╕¬α╕▓α╕éα╕▓α╕ùα╕╡α╣êα╣Çα╕äα╕óα╣äα╕¢α╕öα╣ëα╕ºα╕óα╕üα╕▒α╕Öα╣âα╕Öα╕úα╕▓α╕óα╕üα╕▓α╕úα╕Öα╕╡α╣ë")
                    
                    # α╕öα╕▓α╕ºα╕Öα╣îα╣éα╕½α╕Ñα╕ö
                    st.markdown("---")
                    output_region = io.BytesIO()
                    with pd.ExcelWriter(output_region, engine='xlsxwriter') as writer:
                        df_region.to_excel(writer, sheet_name='α╕¬α╕▓α╕éα╕▓α╕ùα╕▒α╣ëα╕çα╕½α╕íα╕ö', index=False)
                        region_summary.to_excel(writer, sheet_name='α╕¬α╕úα╕╕α╕¢α╕òα╕▓α╕íα╕áα╕▓α╕ä', index=False)
                    
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        st.download_button(
                            label="≡ƒôÑ α╕öα╕▓α╕ºα╕Öα╣îα╣éα╕½α╕Ñα╕öα╕éα╣ëα╕¡α╕íα╕╣α╕Ñα╕êα╕▒α╕öα╕üα╕Ñα╕╕α╣êα╕í (Excel)",
                            data=output_region.getvalue(),
                            file_name=f"α╕êα╕▒α╕öα╕üα╕Ñα╕╕α╣êα╕íα╕¬α╕▓α╕éα╕▓_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

if __name__ == "__main__":
    main()
