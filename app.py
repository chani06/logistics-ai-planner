# Logistics Planner
# Version: 2025-12-26-v3.4

import streamlit as st
import pandas as pd
import numpy as np
import pickle
import os
import glob
from datetime import datetime, time as datetime_time, timedelta
import io
from math import radians, sin, cos, sqrt, atan2
import json
import time as time_module
import sys
import re

# ฟังก์ชัน safe print สำหรับ Windows console
def safe_print(*args, **kwargs):
    """Print with fallback encoding for Windows console + append to UI log buffer"""
    text = ' '.join(str(arg) for arg in args)
    # ── dump to debug log file ──
    try:
        import os as _os2
        _dbg = _os2.path.join(_os2.path.dirname(__file__), 'trip_debug.log')
        with open(_dbg, 'a', encoding='utf-8') as _df:
            _df.write(text + '\n')
    except Exception:
        pass
    # ── append to session_state log buffer ──
    try:
        import streamlit as _st
        if hasattr(_st, 'session_state'):
            if '_ui_log' not in _st.session_state:
                _st.session_state['_ui_log'] = []
            _st.session_state['_ui_log'].append(text)
    except Exception:
        pass
    try:
        # ใช้ buffer แทน stdout โดยตรงเพื่อหลีกเลี่ยง Streamlit wrapper
        output = sys.stdout.buffer if hasattr(sys.stdout, 'buffer') else sys.stdout
        if hasattr(output, 'write'):
            if isinstance(output, (type(sys.stdout.buffer), type(sys.stderr.buffer))):
                output.write((text + '\n').encode('utf-8', errors='replace'))
            else:
                output.write(text + '\n')
            output.flush() if hasattr(output, 'flush') else None
    except Exception:
        try:
            # Fallback: ใช้ ASCII
            sys.__stdout__.write(text.encode('ascii', 'replace').decode('ascii') + '\n')
            sys.__stdout__.flush()
        except Exception:
            pass

def safe_int_trip(val, default=0):
    """แปลง Trip value เป็น int อย่างปลอดภัย — รับมือ NaN/None/float"""
    try:
        if val is None:
            return default
        import math
        if isinstance(val, float) and math.isnan(val):
            return default
        return int(val)
    except (ValueError, TypeError):
        return default

def safe_qty(val, default=0):
    """แปลง OriginalQty / priority เป็น int อย่างปลอดภัย — NaN → default"""
    try:
        if val is None:
            return default
        fval = float(val)
        import math
        if math.isnan(fval):
            return default
        return int(fval)
    except (ValueError, TypeError):
        return default

def safe_join(values, sep=', '):
    """Join mixed-type values safely, skipping None/NaN."""
    if values is None:
        return ''
    cleaned = []
    for v in values:
        try:
            if v is None or bool(pd.isna(v)):
                continue
        except Exception:
            pass
        cleaned.append(str(v))
    return sep.join(cleaned)

# ── AI TRIP LEARNING SYSTEM ──────────────────────────────────────────────────
# บันทึก/โหลดประวัติการจัดทริป เพื่อเรียนรู้ว่าสาขาไหนมักอยู่ทริปเดียวกัน

TRIP_HISTORY_FILE = os.path.join(os.path.dirname(__file__), 'trip_history.json')
_TRIP_HISTORY_CACHE = None   # in-process cache (cleared each save)

def load_trip_history() -> dict:
    """โหลด pair_freq dict: {"CODE_A|CODE_B": count}  (sorted keys)"""
    global _TRIP_HISTORY_CACHE
    if _TRIP_HISTORY_CACHE is not None:
        return _TRIP_HISTORY_CACHE
    if os.path.exists(TRIP_HISTORY_FILE):
        try:
            with open(TRIP_HISTORY_FILE, encoding='utf-8') as f:
                data = json.load(f)
            _TRIP_HISTORY_CACHE = data.get('pair_freq', {})
            return _TRIP_HISTORY_CACHE
        except Exception:
            pass
    return {}

def save_trip_history(assigned_df) -> int:
    """
    เรียกหลัง export — บันทึกว่าสาขาไหนอยู่ทริปเดียวกัน
    คืนค่าจำนวน pairs ที่บันทึกเพิ่มใน session นี้
    """
    global _TRIP_HISTORY_CACHE
    # โหลดข้อมูลเดิม
    if os.path.exists(TRIP_HISTORY_FILE):
        try:
            with open(TRIP_HISTORY_FILE, encoding='utf-8') as f:
                data = json.load(f)
        except Exception:
            data = {}
    else:
        data = {}

    pair_freq   = data.get('pair_freq', {})
    sessions    = data.get('sessions', [])
    total_saved = 0

    # สร้าง group: trip_id → list of codes
    trip_groups = {}
    for _, row in assigned_df.iterrows():
        tid  = str(safe_int_trip(row.get('Trip', 0)))
        code = str(row.get('Code', '')).strip().upper()
        if code and tid != '0':
            trip_groups.setdefault(tid, []).append(code)

    session_pairs = []
    import itertools
    for tid, codes in trip_groups.items():
        codes = sorted(set(codes))
        for a, b in itertools.combinations(codes, 2):
            key = f"{a}|{b}"
            pair_freq[key] = pair_freq.get(key, 0) + 1
            session_pairs.append(key)
            total_saved += 1

    sessions.append({
        'date':       datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),
        'trips':      len(trip_groups),
        'pair_count': total_saved,
    })
    # เก็บเฉพาะ 500 sessions ล่าสุด
    if len(sessions) > 500:
        sessions = sessions[-500:]

    data = {'pair_freq': pair_freq, 'sessions': sessions,
            'total_sessions': len(sessions)}

    with open(TRIP_HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    _TRIP_HISTORY_CACHE = pair_freq   # refresh in-process cache
    return total_saved

def get_trip_learning_stats() -> dict:
    """สรุปสถิติการเรียนรู้เพื่อแสดงใน UI"""
    if not os.path.exists(TRIP_HISTORY_FILE):
        return {'sessions': 0, 'unique_pairs': 0, 'top_pairs': []}
    try:
        with open(TRIP_HISTORY_FILE, encoding='utf-8') as f:
            data = json.load(f)
        pf  = data.get('pair_freq', {})
        top = sorted(pf.items(), key=lambda x: -x[1])[:10]
        return {
            'sessions':     len(data.get('sessions', [])),
            'unique_pairs': len(pf),
            'top_pairs':    top,
        }
    except Exception:
        return {'sessions': 0, 'unique_pairs': 0, 'top_pairs': []}

# Map visualization
try:
    import folium
    from folium import plugins
    from streamlit_folium import folium_static  # ใช้ folium_static แทน st_folium เพื่อไม่ให้โหลดซ้ำ
    import requests
    FOLIUM_AVAILABLE = True
except ImportError:
    FOLIUM_AVAILABLE = False

# Google Sheets Integration
try:
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    
    # ตรวจสอบว่ามีไฟล์ credentials.json หรือ Streamlit secrets
    credentials_file = 'credentials.json'
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    
    creds = None
    
    # 1️⃣ ลองใช้ Streamlit Secrets ก่อน (สำหรับ Streamlit Cloud)
    try:
        if hasattr(st, 'secrets') and 'gcp_service_account' in st.secrets:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(
                dict(st.secrets['gcp_service_account']), 
                scope
            )
            safe_print("✅ ใช้ credentials จาก Streamlit Secrets")
    except Exception as e:
        safe_print(f"⚠️ Streamlit Secrets ไม่พร้อมใช้งาน: {e}")
    
    # 2️⃣ ถ้าไม่มี secrets ให้ใช้ไฟล์ local
    if creds is None:
        if os.path.exists(credentials_file):
            try:
                creds = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
                safe_print(f"✅ ใช้ credentials จาก {credentials_file}")
            except Exception as e:
                safe_print(f"⚠️ ไม่สามารถอ่านไฟล์ {credentials_file}: {e}")
        else:
            safe_print(f"⚠️ ไม่พบ {credentials_file} และไม่มี Streamlit Secrets")
            safe_print(f"💡 ดูวิธีตั้งค่าได้ที่: CREDENTIALS_SETUP.md")
    
    # เชื่อมต่อ Google Sheets (ใส่ socket timeout ป้องกัน hang ตอน startup)
    if creds:
        import socket as _socket
        _prev_timeout = _socket.getdefaulttimeout()
        _socket.setdefaulttimeout(10)  # 10 วินาที max สำหรับ network calls
        try:
            gc = gspread.authorize(creds)
            SPREADSHEET_ID = '12DmIfECwVpsWfl8rl2r1A_LB4_5XMrmnmwlPUHKNU-o'
            sh = gc.open_by_key(SPREADSHEET_ID)
            SHEETS_AVAILABLE = True
            safe_print("✅ เชื่อมต่อ Google Sheets สำเร็จ")
        except Exception as e:
            safe_print(f"⚠️ Google Sheets Error: {e}")
            safe_print(f"💡 ตรวจสอบ credentials หรือดูคู่มือที่ CREDENTIALS_SETUP.md")
            SHEETS_AVAILABLE = False
            gc = None
            sh = None
        finally:
            _socket.setdefaulttimeout(_prev_timeout)  # restore
    else:
        SHEETS_AVAILABLE = False
        gc = None
        sh = None
        
except ImportError:
    safe_print("⚠️ ไม่พบ gspread library - ติดตั้งด้วย: pip install gspread oauth2client")
    SHEETS_AVAILABLE = False
    gc = None
    sh = None

# Auto-refresh component
try:
    from streamlit_autorefresh import st_autorefresh
    AUTOREFRESH_AVAILABLE = True
except ImportError:
    AUTOREFRESH_AVAILABLE = False
    # แสดง warning เฉพาะใน local dev (ไม่แสดงใน deployment)
    if os.environ.get('ENVIRONMENT') != 'production':
        pass  # ไม่แสดง warning - ใช้ manual refresh แทน

# ==========================================
# CACHE SYSTEM - ป้องกันการโหลดซ้ำและเพิ่มความเร็ว
# ==========================================
USE_CACHE = True  # เปิดใช้งาน cache system
DISTANCE_CACHE_FILE = 'distance_cache.json'
ROUTE_CACHE_FILE = 'route_cache.json'

# โหลด cache จากไฟล์
@st.cache_data(show_spinner=False)
def load_distance_cache():
    """โหลด distance cache จากไฟล์"""
    if os.path.exists(DISTANCE_CACHE_FILE):
        try:
            with open(DISTANCE_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

# นับ entry ใหม่ที่ยังไม่ได้ save (dirty counter)
_DIST_CACHE_DIRTY = 0
_DIST_CACHE_SAVE_BATCH = 50  # บันทึกทุก N entries ใหม่

_ROUTE_CACHE_DIRTY = 0
_ROUTE_CACHE_SAVE_BATCH = 10  # route แต่ละ entry ใหญ่กว่า → batch เล็กกว่า

def save_distance_cache(cache_dict, force=False):
    """บันทึก distance cache ลงไฟล์ — เฉพาะเมื่อมีการเปลี่ยนแปลงเท่านั้น"""
    global _DIST_CACHE_DIRTY
    if not force and _DIST_CACHE_DIRTY == 0:
        return  # ไม่มีการเปลี่ยนแปลง ไม่ต้อง save
    try:
        with open(DISTANCE_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache_dict, f, ensure_ascii=False, indent=2)
        _DIST_CACHE_DIRTY = 0
    except Exception as e:
        safe_print(f"⚠️ ไม่สามารถบันทึก distance cache: {e}")

def load_route_cache():
    """โหลด route cache จากไฟล์"""
    if os.path.exists(ROUTE_CACHE_FILE):
        try:
            with open(ROUTE_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_route_cache(cache_dict, force=False):
    """บันทึก route cache ลงไฟล์ — เฉพาะเมื่อมีการเปลี่ยนแปลง (dirty counter)"""
    global _ROUTE_CACHE_DIRTY
    if not force and _ROUTE_CACHE_DIRTY == 0:
        return
    try:
        with open(ROUTE_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(dict(cache_dict), f, ensure_ascii=False, separators=(',', ':'))
        _ROUTE_CACHE_DIRTY = 0
    except Exception as e:
        safe_print(f"⚠️ ไม่สามารถบันทึก route cache: {e}")

# โหลด cache ตอนเริ่มต้น
if USE_CACHE:
    DISTANCE_CACHE = load_distance_cache()
    ROUTE_CACHE_DATA = load_route_cache()
    
    # แยกประเภท cache
    dc_distances = sum(1 for k in DISTANCE_CACHE.keys() if k.startswith('14.1') or k.startswith('14.2'))
    branch_distances = len(DISTANCE_CACHE) - dc_distances
    
    safe_print(f"✅ โหลด distance_cache.json: {len(DISTANCE_CACHE):,} รายการ")
    if dc_distances > 0 or branch_distances > 0:
        safe_print(f"   - DC→สาขา: ~{dc_distances:,} รายการ")
        safe_print(f"   - สาขา↔สาขา: ~{branch_distances:,} รายการ")
    safe_print(f"✅ โหลด route_cache.json: {len(ROUTE_CACHE_DATA):,} เส้นทาง")
else:
    DISTANCE_CACHE = {}
    ROUTE_CACHE_DATA = {}

# ==========================================
# GOOGLE SHEETS SYNC FUNCTION
# ==========================================
def sync_branch_data_from_sheets():
    """
    ดึงข้อมูลจาก Google Sheets และ sync กับ JSON file
    ใช้รหัสสาขา (Code/Plan Code) เป็น key หลัก
    
    Returns:
        DataFrame หรือ None ถ้าล้มเหลว
    """
    global SHEETS_AVAILABLE, sh, gc

    # ── พยายาม reconnect ถ้า sh หรือ SHEETS_AVAILABLE เป็น False/None ──
    if not SHEETS_AVAILABLE or sh is None:
        try:
            _creds = None
            _scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
            if hasattr(st, 'secrets') and 'gcp_service_account' in st.secrets:
                from oauth2client.service_account import ServiceAccountCredentials as _SAC
                _creds = _SAC.from_json_keyfile_dict(dict(st.secrets['gcp_service_account']), _scope)
            elif os.path.exists('credentials.json'):
                from oauth2client.service_account import ServiceAccountCredentials as _SAC
                _creds = _SAC.from_json_keyfile_name('credentials.json', _scope)
            if _creds:
                import socket as _s2
                _pt = _s2.getdefaulttimeout()
                _s2.setdefaulttimeout(15)
                try:
                    import gspread as _gs
                    gc = _gs.authorize(_creds)
                    sh = gc.open_by_key('12DmIfECwVpsWfl8rl2r1A_LB4_5XMrmnmwlPUHKNU-o')
                    SHEETS_AVAILABLE = True
                    safe_print("✅ Reconnect Google Sheets สำเร็จ")
                finally:
                    _s2.setdefaulttimeout(_pt)
        except Exception as _re:
            safe_print(f"⚠️ Reconnect ล้มเหลว: {_re}")

    json_file = 'branch_data.json'
    
    # โหลดข้อมูลเก่าจาก JSON
    existing_data = {}
    if os.path.exists(json_file):
        try:
            import re as _re_j
            with open(json_file, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            # normalize inner dict keys (column names อาจมี space/newline)
            for k, v in raw.items():
                if isinstance(v, dict):
                    existing_data[k] = {_re_j.sub(r'[\s\n\r\t]+', '', str(ck)): cv for ck, cv in v.items()}
                else:
                    existing_data[k] = v
        except Exception as e:
            safe_print(f"⚠️ ไม่สามารถอ่าน JSON: {e}")
    
    # ถ้าไม่มี Google Sheets ให้ใช้ข้อมูลเก่า
    if not SHEETS_AVAILABLE or sh is None:
        if existing_data:
            safe_print(f"⚠️ Google Sheets ไม่พร้อม - ใช้ข้อมูลจาก JSON ({len(existing_data)} สาขา)")
            df = pd.DataFrame.from_dict(existing_data, orient='index')
            # ตรวจสอบว่ามีคอลัมน์ Plan Code หรือไม่
            # (JSON keys ถูก normalize แล้ว → 'Plan Code' กลายเป็น 'PlanCode')
            if 'Plan Code' in df.columns:
                df.reset_index(drop=True, inplace=True)
            elif 'PlanCode' in df.columns:
                df.reset_index(drop=True, inplace=True)
                df.rename(columns={'PlanCode': 'Plan Code'}, inplace=True)
            else:
                df.reset_index(inplace=True)
                df.rename(columns={'index': 'Plan Code'}, inplace=True)
            return df
        else:
            safe_print("❌ ไม่พบข้อมูล: ไม่มี Google Sheets และไม่มี JSON cache")
            return pd.DataFrame()  # Return empty DataFrame แทน None
    
    try:
        # ดึงข้อมูลจาก Sheets (GID: 876257177)
        worksheet = None
        for ws in sh.worksheets():
            if ws.id == 876257177:
                worksheet = ws
                break
        
        if worksheet is None:
            worksheet = sh.get_worksheet(0)
        
        # ดึงข้อมูลทั้งหมด
        data = worksheet.get_all_values()
        if not data or len(data) < 2:
            return None
        
        # สร้าง DataFrame — normalize headers ก่อนเพื่อให้ column names สะอาด
        import re as _re_hdr
        headers = [_re_hdr.sub(r'[\s\n\r\t]+', '', str(h)) for h in data[0]]
        # คืน 'Plan Code' ให้ถูกต้อง (หลัง normalize 'Plan Code' → 'PlanCode')
        headers = ['Plan Code' if h == 'PlanCode' else h for h in headers]
        df_new = pd.DataFrame(data[1:], columns=headers)
        
        # หา column รหัสสาขา
        code_col = None
        for col in ['Code', 'Plan Code', 'รหัสสาขา', 'สาขา']:
            if col in df_new.columns:
                code_col = col
                break
        
        if not code_col:
            safe_print("❌ ไม่พบคอลัมน์รหัสสาขา")
            return None
        
        # นับข้อมูลใหม่
        new_count = 0
        updated_count = 0
        
        # อัปเดตข้อมูลจาก Google Sheets (รวม DC วังน้อยที่มีอยู่ใน Sheets)
        for idx, row in df_new.iterrows():
            code = str(row[code_col]).strip().upper()
            if not code or code == '':
                continue
            
            # แปลง row เป็น dict — normalize column keys ให้ตรงกับ JSON (ลบ space/newline)
            import re as _re_rd
            row_dict = {_re_rd.sub(r'[\s\n\r\t]+', '', str(ck)): cv for ck, cv in row.to_dict().items()}
            
            if code in existing_data:
                # ข้อมูลเก่า - เช็คว่ามีการเปลี่ยนแปลงจริงหรือไม่
                if existing_data[code] != row_dict:
                    existing_data[code] = row_dict
                    updated_count += 1
                # ถ้าข้อมูลเหมือนเดิม ไม่นับเป็น update
            else:
                # ข้อมูลใหม่ - เพิ่ม
                existing_data[code] = row_dict
                new_count += 1
        
        # บันทึกกลับเป็น JSON
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=2)

        safe_print(f"✅ Sync เสร็จสิ้น: {new_count} สาขาใหม่, {updated_count} สาขาอัปเดต, รวม {len(existing_data)} สาขา")

        # 🔄 Auto-rebuild branch_groups.json เฉพาะเมื่อมีสาขาใหม่/อัพเดท
        if new_count == 0 and updated_count == 0:
            safe_print("⏩ branch_groups.json ไม่ต้อง rebuild (ไม่มีการเปลี่ยนแปลง)")
        else:
            try:
                import math as _math

                def _hav_m(la1, lo1, la2, lo2):
                    R = 6371.0
                    d1 = _math.radians(la2 - la1); d2 = _math.radians(lo2 - lo1)
                    a = _math.sin(d1/2)**2 + _math.cos(_math.radians(la1)) * _math.cos(_math.radians(la2)) * _math.sin(d2/2)**2
                    return R * 2 * _math.asin(_math.sqrt(max(0.0, min(1.0, a))))

                # ดึงพิกัด
                _coords_bg: dict = {}
                for _bc, _bi in existing_data.items():
                    _cu = str(_bc).strip().upper()
                    _lat_v = _lon_v = 0.0
                    for _lk in ('ละติจูด', 'Latitude', 'ละ'):
                        _v = _bi.get(_lk)
                        if _v and str(_v).strip() not in ('', '0', 'None'):
                            try: _lat_v = float(str(_v).replace(',', '')); break
                            except: pass
                    for _lk in ('ลองติจูด', 'Longitude', 'ลอง'):
                        _v = _bi.get(_lk)
                        if _v and str(_v).strip() not in ('', '0', 'None'):
                            try: _lon_v = float(str(_v).replace(',', '')); break
                            except: pass
                    if _lat_v > 0 and _lon_v > 0:
                        _coords_bg[_cu] = (_lat_v, _lon_v)

                # Spatial grid (cell ≈ 500m) — ขยายจาก 200m → 500m ให้ตรงกับ precompute_branch_data.py
                _COLOC = 0.5
                _CELL = 0.005  # ~550m per cell
                _grid_bg: dict = {}
                for _c, (_la, _lo) in _coords_bg.items():
                    _key = (int(_la / _CELL), int(_lo / _CELL))
                    _grid_bg.setdefault(_key, []).append(_c)

                _par = {c: c for c in _coords_bg}
                def _find(x):
                    while _par[x] != x:
                        _par[x] = _par[_par[x]]; x = _par[x]
                    return x
                def _union(x, y): _par[_find(x)] = _find(y)

                # สร้าง province map สำหรับตรวจจังหวัดเดียวกัน
                _prov_bg = {}
                for _bc2, _bi2 in existing_data.items():
                    _cu2 = str(_bc2).strip().upper()
                    _prov_bg[_cu2] = str(_bi2.get('จังหวัด', '') or '').strip()

                for (_glat, _glon), _cell_codes in _grid_bg.items():
                    _nbrs = []
                    for _dg in range(-1, 2):
                        for _dh in range(-1, 2):
                            _nbrs.extend(_grid_bg.get((_glat + _dg, _glon + _dh), []))
                    for _ci in _cell_codes:
                        _lai, _loi = _coords_bg[_ci]
                        _pi = _prov_bg.get(_ci, '')
                        for _cj in _nbrs:
                            if _cj <= _ci: continue
                            _pj = _prov_bg.get(_cj, '')
                            # ต้องจังหวัดเดียวกัน (ถ้ามีข้อมูล)
                            if _pi and _pj and _pi != _pj:
                                continue
                            _laj, _loj = _coords_bg[_cj]
                            _d_m = _hav_m(_lai, _loi, _laj, _loj)
                            # ≤200m → union เสมอ (GPS drift ชายแดน, data ผิด)
                            # ≤500m → ต้องจังหวัดเดียวกันก่อน
                            if _d_m <= 0.2 or (_d_m <= _COLOC and not (_pi and _pj and _pi != _pj)):
                                _union(_ci, _cj)

                # Build groups
                _gmap: dict = {}
                for _c in _coords_bg:
                    _gmap.setdefault(_find(_c), []).append(_c)

                _new_groups: dict = {}; _new_b2g: dict = {}; _gi = 1
                for _root, _members in _gmap.items():
                    if len(_members) < 2: continue
                    _gk = f'G{_gi:04d}'; _gi += 1
                    _new_groups[_gk] = sorted(_members)
                    for _m in _members: _new_b2g[_m] = _gk

                # Merge กับ old groups
                try:
                    with open('branch_groups.json', 'r', encoding='utf-8') as _f:
                        _old_bg = json.load(_f)
                    for _ogk, _ogm in _old_bg.get('groups', {}).items():
                        _old_mems = [str(_m).strip().upper() for _m in _ogm]
                        _tgt = next((_new_b2g[_om] for _om in _old_mems if _om in _new_b2g), None)
                        if _tgt:
                            for _om in _old_mems:
                                if _om not in _new_b2g:
                                    _new_groups[_tgt].append(_om); _new_b2g[_om] = _tgt
                            _new_groups[_tgt] = sorted(set(_new_groups[_tgt]))
                        else:
                            _gk2 = f'G{_gi:04d}'; _gi += 1
                            _new_groups[_gk2] = sorted(set(_old_mems))
                            for _om in _old_mems: _new_b2g[_om] = _gk2
                except: pass

                with open('branch_groups.json', 'w', encoding='utf-8') as _f:
                    json.dump({'groups': _new_groups, 'branch_to_group': _new_b2g},
                              _f, ensure_ascii=False, separators=(',', ':'))
                safe_print(f"✅ Auto-rebuild branch_groups.json: {len(_new_groups)} groups, {len(_new_b2g)} branches")
                # 🔄 Reload global ทันที
                global BRANCH_GROUPS, BRANCH_TO_GROUP
                BRANCH_GROUPS   = {k: [str(c).strip().upper() for c in v]
                                   for k, v in _new_groups.items()}
                BRANCH_TO_GROUP = {str(k).strip().upper(): v
                                   for k, v in _new_b2g.items()}
                safe_print(f"🔄 Reloaded BRANCH_GROUPS: {len(BRANCH_GROUPS)} groups")
            except Exception as _bge:
                safe_print(f"⚠️ Auto-rebuild branch_groups error: {_bge}")
        
        # แปลงกลับเป็น DataFrame (ใช้ index เป็น Plan Code)
        df = pd.DataFrame.from_dict(existing_data, orient='index')
        
        # ป้องกัน duplicate column: PlanCode (inner dict) + index (branch code) จะซ้ำกัน
        # ให้ drop index เสมอ แล้วใช้ PlanCode column แทน
        if 'PlanCode' in df.columns:
            df.reset_index(drop=True, inplace=True)
            df.rename(columns={'PlanCode': 'Plan Code'}, inplace=True)
        elif 'Plan Code' in df.columns:
            df.reset_index(drop=True, inplace=True)
        else:
            df.reset_index(inplace=True)
            df.rename(columns={'index': 'Plan Code'}, inplace=True)
        
        return df
        
    except Exception as e:
        safe_print(f"❌ Error: {e}")
        # ถ้าเกิด error ให้ใช้ข้อมูลเก่า
        if existing_data:
            safe_print(f"📦 ใช้ข้อมูลเก่าจาก JSON")
            df = pd.DataFrame.from_dict(existing_data, orient='index')
            if 'PlanCode' in df.columns:
                df.reset_index(drop=True, inplace=True)
                df.rename(columns={'PlanCode': 'Plan Code'}, inplace=True)
            elif 'Plan Code' in df.columns:
                df.reset_index(drop=True, inplace=True)
            else:
                df.reset_index(inplace=True)
                df.rename(columns={'index': 'Plan Code'}, inplace=True)
            return df
        return None

# ==========================================
# CONFIG
# ==========================================
MODEL_PATH = 'models/decision_tree_model.pkl'

# ขีดจำกัดรถแต่ละประเภท (มาตรฐาน)
LIMITS = {
    '4W': {'max_w': 2000, 'max_c': 7.0,  'max_drops': 999},   # Cube ≤ 7m³, W ≤ 2000 kg
    'JB': {'max_w': 3000, 'max_c': 10.0, 'max_drops': 999},   # Cube ≤ 10m³, W ≤ 3000 kg
    '6W': {'max_w': 5500, 'max_c': 25.0, 'max_drops': 999},   # Cube ≤ 25m³, W ≤ 5500 kg
}

# 🔒 ขีดจำกัดสำหรับ Punthai ล้วน — เฉพาะน้ำหนัก/คิว (ไม่จำกัดจุด)
PUNTHAI_LIMITS = {
    '4W': {'max_w': 2000, 'max_c': 7.0,  'max_drops': 999},
    'JB': {'max_w': 3000, 'max_c': 10.0, 'max_drops': 999},
    '6W': {'max_w': 5500, 'max_c': 25.0, 'max_drops': 999},
}

#  Geographic Clustering Config
MAX_DISTRICT_DISTANCE_KM = 30  # คนละจังหวัด: ห่างกันเกิน 30km ไม่ควรรวมทริป (จังหวัดเดียวกันสามารถ 80km)

# Utilization Config (ใช้ buffer จากหน้าเว็บเท่านั้น ไม่ fix ตายตัว)
MIN_VEHICLE_UTILIZATION = 1.0  # เป้าหมาย: รถต้องเต็ม 100% (แสดง warning ถ้าต่ำกว่า)

# ==========================================
# REGION ORDER CONFIG (Far-to-Near Sorting)
# ==========================================
# ลำดับการจัด: เหนือ → อีสาน → ใต้ → ตะวันออก → กลาง
REGION_ORDER = {
    'เหนือ': 1, 'NORTH': 1,
    'อีสาน': 2, 'NE': 2,
    'ใต้': 3, 'SOUTH': 3,
    'ตะวันออก': 4, 'EAST': 4,
    'ตะวันตก': 5, 'WEST': 5,
    'กลาง': 6, 'CENTRAL': 6,
    'ไม่ระบุ': 99
}

# จังหวัดที่ควรอยู่ในทริปเดียวกัน (bypass zone check) — แต่ละ frozenset = กลุ่มที่รวมได้
PROVINCE_PAIR_GROUPS: list = [
    frozenset(['ฉะเชิงเทรา', 'ชลบุรี']),          # ตะวันออก: สาย 304/331 ผ่านกัน
    frozenset(['ชลบุรี', 'ระยอง']),               # EEC ชายฝั่งตะวันออก
    frozenset(['ฉะเชิงเทรา', 'ปราจีนบุรี']),       # สาย 304 เดียวกัน
    frozenset(['จันทบุรี', 'ตราด']),               # ปลายตะวันออก
]

def _provinces_are_paired(prov_a: str, prov_b: str) -> bool:
    """คืน True ถ้า 2 จังหวัดอยู่ใน PROVINCE_PAIR_GROUPS เดียวกัน"""
    if not prov_a or not prov_b or prov_a == prov_b:
        return False
    for _grp in PROVINCE_PAIR_GROUPS:
        if prov_a in _grp and prov_b in _grp:
            return True
    return False

# ════════════════════════════════════════════════════════════════
# PROVINCE ZONE MAP — ตรงกับ zone_viewer.py (ใช้จัดทริป ป้องกันกระโดด)
# ════════════════════════════════════════════════════════════════
PROVINCE_ZONE_MAP: dict = {
    "กรุงเทพมหานคร": "__BKK__", "กรุงเทพฯ": "__BKK__",
    "กทม": "__BKK__", "กทม.": "__BKK__",
    # ปริมณฑล
    "นนทบุรี":"ปริมณฑล_นนทบุรี","ปทุมธานี":"ปริมณฑล_ปทุมธานี",
    "สมุทรปราการ":"ปริมณฑล_สมุทรปราการ","นครปฐม":"ปริมณฑล_นครปฐม",
    "สมุทรสาคร":"ปริมณฑล_สมุทรสาคร","สมุทรสงคราม":"ปริมณฑล_สมุทรสงคราม",
    "พระนครศรีอยุธยา":"ปริมณฑล_อยุธยา","สระบุรี":"ปริมณฑล_สระบุรี",
    "อ่างทอง":"ปริมณฑล_อ่างทอง","สิงห์บุรี":"ปริมณฑล_สิงห์บุรี",
    "ชัยนาท":"ปริมณฑล_ชัยนาท","ลพบุรี":"ปริมณฑล_ลพบุรี",
    # ภาคเหนือ
    "นครสวรรค์":"เหนือ_นครสวรรค์","อุทัยธานี":"เหนือ_อุทัยธานี",
    "กำแพงเพชร":"เหนือ_กำแพงเพชร","ตาก":"เหนือ_ตาก",
    "สุโขทัย":"เหนือ_สุโขทัย","พิษณุโลก":"เหนือ_พิษณุโลก",
    "พิจิตร":"เหนือ_พิจิตร","เพชรบูรณ์":"เหนือ_เพชรบูรณ์",
    "อุตรดิตถ์":"เหนือ_อุตรดิตถ์","แพร่":"เหนือ_แพร่",
    "น่าน":"เหนือ_น่าน","พะเยา":"เหนือ_พะเยา",
    "เชียงราย":"เหนือ_เชียงราย","เชียงใหม่":"เหนือ_เชียงใหม่",
    "ลำพูน":"เหนือ_ลำพูน","ลำปาง":"เหนือ_ลำปาง",
    "แม่ฮ่องสอน":"เหนือ_แม่ฮ่องสอน",
    # ภาคอีสาน
    "หนองบัวลำภู":"อีสาน_หนองบัวลำภู","อุดรธานี":"อีสาน_อุดรธานี",
    "หนองคาย":"อีสาน_หนองคาย","บึงกาฬ":"อีสาน_บึงกาฬ",
    "เลย":"อีสาน_เลย","สกลนคร":"อีสาน_สกลนคร",
    "นครพนม":"อีสาน_นครพนม","มุกดาหาร":"อีสาน_มุกดาหาร",
    "ชัยภูมิ":"อีสาน_ชัยภูมิ","ขอนแก่น":"อีสาน_ขอนแก่น",
    "กาฬสินธุ์":"อีสาน_กาฬสินธุ์","มหาสารคาม":"อีสาน_มหาสารคาม",
    "ร้อยเอ็ด":"อีสาน_ร้อยเอ็ด","นครราชสีมา":"อีสาน_นครราชสีมา",
    "บุรีรัมย์":"อีสาน_บุรีรัมย์","สุรินทร์":"อีสาน_สุรินทร์",
    "ศรีสะเกษ":"อีสาน_ศรีสะเกษ","อุบลราชธานี":"อีสาน_อุบลราชธานี",
    "ยโสธร":"อีสาน_ยโสธร","อำนาจเจริญ":"อีสาน_อำนาจเจริญ",
    # ภาคตะวันออก
    "ฉะเชิงเทรา":"ตะวันออก_ฉะเชิงเทรา","นครนายก":"ตะวันออก_นครนายก",
    "ปราจีนบุรี":"ตะวันออก_ปราจีนบุรี","สระแก้ว":"ตะวันออก_สระแก้ว",
    "ชลบุรี":"ตะวันออก_ชลบุรี","ระยอง":"ตะวันออก_ระยอง",
    "จันทบุรี":"ตะวันออก_จันทบุรี","ตราด":"ตะวันออก_ตราด",
    # ภาคตะวันตก
    "กาญจนบุรี":"ตะวันตก_กาญจนบุรี","ราชบุรี":"ตะวันตก_ราชบุรี",
    "สุพรรณบุรี":"ตะวันตก_สุพรรณบุรี","เพชรบุรี":"ตะวันตก_เพชรบุรี",
    "ประจวบคีรีขันธ์":"ตะวันตก_ประจวบคีรีขันธ์",
    # ภาคใต้
    "ชุมพร":"ใต้_ชุมพร","ระนอง":"ใต้_ระนอง",
    "สุราษฎร์ธานี":"ใต้_สุราษฎร์ธานี","นครศรีธรรมราช":"ใต้_นครศรีธรรมราช",
    "พังงา":"ใต้_พังงา","กระบี่":"ใต้_กระบี่","ภูเก็ต":"ใต้_ภูเก็ต",
    "ตรัง":"ใต้_ตรัง","พัทลุง":"ใต้_พัทลุง","สตูล":"ใต้_สตูล",
    "สงขลา":"ใต้_สงขลา","ปัตตานี":"ใต้_ปัตตานี",
    "ยะลา":"ใต้_ยะลา","นราธิวาส":"ใต้_นราธิวาส",
}

# รายการสาขาที่ไม่ต้องการจัดส่ง (ตัดออก)
EXCLUDE_BRANCHES = ['DC011', 'PTDC', 'PTG DISTRIBUTION CENTER']

# รายชื่อที่ต้องตัดออก (ใช้ตรวจสอบชื่อ)
EXCLUDE_NAMES = ['Distribution Center', 'PTG Distribution', 'บ.พีทีจี เอ็นเนอยี']

# พิกัด DC วังน้อย (จุดกลาง)
DC_WANG_NOI_LAT = 14.179394
DC_WANG_NOI_LON = 100.648149

# ==========================================
# 🚛 HIGHWAY-BASED LOGISTICS ROUTES & ZONES
# ==========================================
# หลักการ: "อย่าลากเส้นตรง ให้ลากตามถนน"
# ยึดเลขทางหลวงแผ่นดินเป็นเกณฑ์หลักในการจัดกลุ่ม ไม่ใช่เขตจังหวัด

HIGHWAY_ROUTES = {
    # สาย 1 (พหลโยธิน): ภาคเหนือตอนบน
    'ROUTE_1_พหลโยธิน': {
        'highway': '1',
        'description': 'กทม → สระบุรี → นครสวรรค์ → ตาก → ลำปาง → เชียงใหม่ → เชียงราย',
        'provinces': ['สระบุรี', 'ลพบุรี', 'นครสวรรค์', 'กำแพงเพชร', 'ตาก', 'ลำปาง', 'ลำพูน', 'เชียงใหม่', 'เชียงราย'],
        'branches': ['พหลโยธิน', 'เอเชีย'],
    },
    # สาย 11 (เอเชียสายเก่า): พิษณุโลก-แพร่-น่าน
    'ROUTE_11_เอเชียสายเก่า': {
        'highway': '11',
        'description': 'นครสวรรค์ → พิจิตร → พิษณุโลก → อุตรดิตถ์ → แพร่',
        'provinces': ['นครสวรรค์', 'พิจิตร', 'พิษณุโลก', 'อุตรดิตถ์', 'แพร่'],
    },
    # สาย 101: แพร่-น่าน
    'ROUTE_101_แพร่น่าน': {
        'highway': '101',
        'description': 'แพร่ → น่าน (หุบเขา)',
        'provinces': ['แพร่', 'น่าน'],
    },
    # สาย 32 (สายเอเชีย): ภาคเหนือตอนล่าง
    'ROUTE_32_สายเอเชีย': {
        'highway': '32',
        'description': 'กทม → อยุธยา → อ่างทอง → สิงห์บุรี → ชัยนาท → นครสวรรค์',
        'provinces': ['พระนครศรีอยุธยา', 'อ่างทอง', 'สิงห์บุรี', 'ชัยนาท', 'นครสวรรค์'],
    },
    # สาย 2 (มิตรภาพ): อีสานเหนือ
    'ROUTE_2_มิตรภาพ': {
        'highway': '2',
        'description': 'สระบุรี → นครราชสีมา → ขอนแก่น → อุดรธานี → หนองคาย',
        'provinces': ['นครราชสีมา', 'ขอนแก่น', 'อุดรธานี', 'หนองคาย', 'เลย', 'หนองบัวลำภู', 'สกลนคร', 'นครพนม', 'มุกดาหาร', 'กาฬสินธุ์', 'มหาสารคาม', 'ร้อยเอ็ด'],
    },
    # สาย 24 (เดชอุดม): อีสานใต้
    'ROUTE_24_เดชอุดม': {
        'highway': '24',
        'description': 'นครราชสีมา → บุรีรัมย์ → สุรินทร์ → อุบลราชธานี',
        'provinces': ['บุรีรัมย์', 'สุรินทร์', 'ศรีสะเกษ', 'อุบลราชธานี', 'ยโสธร', 'อำนาจเจริญ'],
    },
    # สาย 304: ปราจีนบุรี-โคราช
    'ROUTE_304_ปราจีนโคราช': {
        'highway': '304',
        'description': 'ชลบุรี → ปราจีนบุรี → นครราชสีมา',
        'provinces': ['ปราจีนบุรี', 'นครราชสีมา'],
    },
    # สาย 4 (เพชรเกษม): ภาคใต้
    'ROUTE_4_เพชรเกษม': {
        'highway': '4',
        'description': 'กทม → เพชรบุรี → ประจวบฯ → ชุมพร → สุราษฎร์ → นครศรีฯ → สงขลา',
        'provinces': ['เพชรบุรี', 'ประจวบคีรีขันธ์', 'ชุมพร', 'ระนอง', 'สุราษฎร์ธานี', 'นครศรีธรรมราช', 'พัทลุง', 'สงขลา', 'ปัตตานี', 'ยะลา', 'นราธิวาส'],
    },
    # สาย 401/402: อันดามัน
    'ROUTE_401_อันดามัน': {
        'highway': '401/402',
        'description': 'สุราษฎร์ → กระบี่ → ภูเก็ต',
        'provinces': ['กระบี่', 'พังงา', 'ภูเก็ต', 'ตรัง', 'สตูล'],
    },
    # สาย 3 (สุขุมวิท): ภาคตะวันออก
    'ROUTE_3_สุขุมวิท': {
        'highway': '3',
        'description': 'กทม → ชลบุรี → ระยอง → จันทบุรี → ตราด',
        'provinces': ['ชลบุรี', 'ระยอง', 'จันทบุรี', 'ตราด'],
    },
    # สาย 331/344: EEC
    'ROUTE_331_EEC': {
        'highway': '331/344',
        'description': 'ชลบุรี-ระยอง (เขตเศรษฐกิจพิเศษ)',
        'provinces': ['ชลบุรี', 'ระยอง'],
    },
    # สาย 9 (กาญจนาภิเษก): ปริมณฑลด้านเหนือ
    'ROUTE_9_กาญจนาภิเษก': {
        'highway': '9',
        'description': 'รอบนอกกรุงเทพ นนทบุรี-ปทุมธานี-สมุทรปราการ',
        'provinces': ['กรุงเทพมหานคร', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ'],
    },
    # สาย 35 (บรมราชชนนี): ปริมณฑลด้านตะวันตก
    'ROUTE_35_บรมราชชนนี': {
        'highway': '35',
        'description': 'กทม → นนทบุรี → นครปฐม → สมุทรสาคร',
        'provinces': ['กรุงเทพมหานคร', 'นนทบุรี', 'นครปฐม', 'สมุทรสาคร'],
    },
    # สาย 305: สระบุรี-ฉะเชิงเทรา
    'ROUTE_305_สระบุรีฉะเชิงเทรา': {
        'highway': '305',
        'description': 'สระบุรี → นครนายก → ฉะเชิงเทรา',
        'provinces': ['สระบุรี', 'นครนายก', 'ฉะเชิงเทรา', 'ปราจีนบุรี'],
    },
    # สาย 340: สุพรรณบุรี-ชัยนาท
    'ROUTE_340_สุพรรณ': {
        'highway': '340',
        'description': 'สุพรรณบุรี → ชัยนาท → อุทัยธานี',
        'provinces': ['สุพรรณบุรี', 'ชัยนาท', 'อุทัยธานี'],
    },
    # สาย 321: ราชบุรี-กาญจนบุรี
    'ROUTE_321_ราชบุรีกาญจน': {
        'highway': '321',
        'description': 'ราชบุรี → กาญจนบุรี → สังขละบุรี',
        'provinces': ['ราชบุรี', 'กาญจนบุรี'],
    },
}

# ===== NO CROSS-ZONE RULES (ห้ามข้ามโซน) =====
# หลักการ: จังหวัดที่ไปคนละทาง/ต้องข้ามภูเขา
NO_CROSS_ZONE_PAIRS = [
    ('เพชรบูรณ์', 'ชัยภูมิ'),
    ('เพชรบูรณ์', 'เลย'),
    ('ตาก', 'สุโขทัย'),
    ('ใต้ฝั่งอันดามัน', 'ใต้ฝั่งอ่าวไทย'),
    ('กระบี่', 'สุราษฎร์ธานี'),
    # ฝั่งตะวันตกของสมุทรปราการ (เมือง/พระประแดง) ≠ ตะวันออก (ชลบุรี/ฉะเชิงเทรา)
    # บางบ่อ/บางเสาธง แยกออกไปอยู่ ZONE_EAST_สมุทรปราการ แล้ว จึงไม่ต้องปิดกั้นทั้งจังหวัด
]
# ✅ น่าน-แพร่-พะเยา รวมกันได้ (เส้นทาง: DC → แพร่ (สาย 11) → น่าน (สาย 101) หรือ → พะเยา (สาย 1))
# ✅ แพร่-อุตรดิตถ์ รวมได้ (สาย 11 เดียวกัน)

# 🎯 ROUTE GROUPS: กลุ่มจังหวัดที่ควรรวมกัน (ไปทางเดียวกัน)
ROUTE_GROUPS = {
    'ROUTE_สาย1_พะเยา': {
        'provinces': ['พะเยา', 'เชียงราย', 'ลำปาง'],
        'description': 'สาย 1 พหลโยธิน ไปเชียงราย-พะเยา',
        'next_routes': [],  # พะเยาไม่ควรรวมกับจังหวัดอื่นนอกสาย 1
    },
    'ROUTE_สาย11_แพร่น่าน': {
        'provinces': ['แพร่', 'น่าน'],
        'description': 'สาย 11/101 แพร่-น่าน (ต้องผ่านแพร่)',
        'next_routes': ['ROUTE_สาย11_อุตรดิตถ์'],  # รวมกับอุตรดิตถ์ได้
    },
    'ROUTE_สาย11_อุตรดิตถ์': {
        'provinces': ['อุตรดิตถ์', 'สุโขทัย'],
        'description': 'สาย 11 อุตรดิตถ์-สุโขทัย',
        'next_routes': ['ROUTE_สาย11_แพร่น่าน', 'ROUTE_สาย11_พิษณุโลก'],
    },
    'ROUTE_สาย11_พิษณุโลก': {
        'provinces': ['พิษณุโลก', 'พิจิตร'],
        'description': 'สาย 11 พิษณุโลก-พิจิตร',
        'next_routes': ['ROUTE_สาย11_อุตรดิตถ์', 'ROUTE_สาย32_นครสวรรค์'],
    },
}

LOGISTICS_ZONES = {
    # ============ ภาคเหนือ - สาย 1/11/101 ============
    # --- เชียงราย (sub-zones) ---
    'ZONE_CRI1_เชียงราย_เมือง': {
        'provinces': ['เชียงราย'],
        'districts': ['เมืองเชียงราย', 'แม่ลาว', 'เวียงชัย', 'เวียงเชียงรุ้ง'],
        'highway': '1',
        'priority': 4,
        'distance_from_dc_km': 730,
        'description': 'เชียงรายในเมือง สาย 1'
    },
    'ZONE_CRI2_เชียงราย_เหนือ': {
        'provinces': ['เชียงราย'],
        'districts': ['แม่สาย', 'แม่จัน', 'เชียงแสน', 'เชียงของ', 'เทิง', 'ขุนตาล'],
        'highway': '1',
        'priority': 3,
        'distance_from_dc_km': 760,
        'description': 'เชียงรายเหนือ ชายแดนพม่า-ลาว'
    },
    'ZONE_CRI3_เชียงราย_ใต้': {
        'provinces': ['เชียงราย'],
        'districts': ['พาน', 'แม่สรวย', 'เวียงป่าเป้า', 'ป่าแดด', 'พญาเม็งราย'],
        'highway': '1',
        'priority': 5,
        'distance_from_dc_km': 720,
        'description': 'เชียงรายใต้'
    },
    # --- เชียงใหม่ (sub-zones) ---
    'ZONE_CNX1_เชียงใหม่_เมือง': {
        'provinces': ['เชียงใหม่'],
        'districts': ['เมืองเชียงใหม่', 'สารภี', 'หางดง', 'สันป่าตอง', 'จอมทอง', 'ดอยหล่อ', 'แม่วาง'],
        'highway': '11',
        'priority': 8,
        'distance_from_dc_km': 650,
        'description': 'เชียงใหม่ในเมือง+ทิศใต้'
    },
    'ZONE_CNX2_เชียงใหม่_รอบเมือง': {
        'provinces': ['เชียงใหม่'],
        'districts': ['สันทราย', 'สันกำแพง', 'ดอยสะเก็ด', 'แม่ริม', 'แม่แตง', 'สะเมิง'],
        'highway': '11',
        'priority': 7,
        'distance_from_dc_km': 660,
        'description': 'เชียงใหม่รอบเมือง ทิศตะวันออก-เหนือ'
    },
    'ZONE_CNX3_เชียงใหม่_เหนือ': {
        'provinces': ['เชียงใหม่'],
        'districts': ['ฝาง', 'เชียงดาว', 'แม่อาย', 'พร้าว', 'ไชยปราการ', 'เวียงแหง'],
        'highway': '107',
        'priority': 6,
        'distance_from_dc_km': 700,
        'description': 'เชียงใหม่เหนือ ฝาง-เชียงดาว'
    },
    'ZONE_A_พะเยา': {
        'provinces': ['พะเยา'],
        'districts': ['เมืองพะเยา', 'แม่ใจ', 'เชียงคำ', 'เชียงม่วน', 'ดอกคำใต้', 'ปง', 'จุน', 'ภูซาง', 'ภูกามยาว'],
        'highway': '1',
        'priority': 1,
        'distance_from_dc_km': 680,
        'description': 'โซนเหนือสุด สาย 1 ปลายทาง'
    },
    'ZONE_B_น่าน': {
        'provinces': ['น่าน'],
        'districts': ['เมืองน่าน', 'ท่าวังผา', 'ปัว', 'เชียงกลาง', 'ทุ่งช้าง', 'บ่อเกลือ', 'เวียงสา', 'นาน้อย', 'นาหมื่น', 'แม่จริม', 'บ้านหลวง', 'สันติสุข', 'ภูเพียง', 'สองแคว', 'เฉลิมพระเกียรติ'],
        'highway': '101',
        'priority': 2,
        'distance_from_dc_km': 620,
        'description': 'หุบเขา ต้องผ่านแพร่ (สาย 101)'
    },
    'ZONE_C_แพร่': {
        'provinces': ['แพร่'],
        'districts': ['เมืองแพร่', 'สูงเม่น', 'เด่นชัย', 'ร้องกวาง', 'สอง', 'ลอง', 'วังชิ้น', 'หนองม่วงไข่'],
        'highway': '11',
        'priority': 3,
        'distance_from_dc_km': 540,
        'description': 'Gateway เหนือตอนบน สาย 11 → น่าน/พะเยา'
    },
    'ZONE_D_อุตรดิตถ์': {
        'provinces': ['อุตรดิตถ์', 'สุโขทัย'],
        'districts': ['เมืองอุตรดิตถ์', 'ตรอน', 'ท่าปลา', 'น้ำปาด', 'ฟากท่า', 'บ้านโคก', 'พิชัย', 'ลับแล', 'ทองแสนขัน',
                      'เมืองสุโขทัย', 'กงไกรลาศ', 'คีรีมาศ', 'ศรีสำโรง', 'สวรรคโลก', 'ศรีนคร', 'บ้านด่านลานหอย', 'ทุ่งเสลี่ยม', 'ศรีสัชนาลัย'],
        'highway': '11',
        'priority': 4,
        'distance_from_dc_km': 450,
        'description': 'หน้าด่านก่อนเข้าแพร่ สาย 11'
    },
    'ZONE_E1_พิษณุโลก_ในเมือง': {
        'provinces': ['พิษณุโลก'],
        'districts': ['เมืองพิษณุโลก'],
        'subdistricts': ['วัดจันทร์', 'ในเมือง', 'บ้านคลอง', 'หัวรอ', 'บึงพระ', 'ท่าทอง', 'บ้านกร่าง'],
        'highway': '11',
        'priority': 5,
        'distance_from_dc_km': 380,
        'description': 'Hub ใหญ่ โซนในเมือง+ตลาด'
    },
    'ZONE_E2_พิษณุโลก_มหาวิทยาลัย': {
        'provinces': ['พิษณุโลก'],
        'districts': ['เมืองพิษณุโลก'],
        'subdistricts': ['ท่าโพธิ์', 'อรัญญิก', 'แม่กา', 'สมอแข', 'บ้านป่า'],
        'highway': '11',
        'priority': 6,
        'distance_from_dc_km': 385,
        'description': 'โซน ม.นเรศวร'
    },
    'ZONE_E3_พิษณุโลก_ตะวันออก': {
        'provinces': ['พิษณุโลก', 'เพชรบูรณ์'],
        'districts': ['วังทอง', 'พรหมพิราม', 'เนินมะปราง', 'บางระกำ', 'ชาติตระการ', 'นครไทย',
                      'หล่มสัก', 'หล่มเก่า', 'เขาค้อ'],
        'highway': '12',
        'priority': 7,
        'distance_from_dc_km': 400,
        'description': 'โซนตะวันออก สาย 12 ไปเขาค้อ'
    },
    'ZONE_F1_พิจิตร_สายหลัก': {
        'provinces': ['พิจิตร'],
        'districts': ['เมืองพิจิตร', 'สากเหล็ก', 'สามง่าม', 'วังทรายพูน'],
        'highway': '11',
        'priority': 8,
        'distance_from_dc_km': 330,
        'description': 'พิจิตรสายหลัก สาย 11'
    },
    'ZONE_F2_พิจิตร_ตะวันออก': {
        'provinces': ['พิจิตร'],
        'districts': ['ตะพานหิน', 'ทับคล้อ', 'ดงเจริญ', 'บางมูลนาก'],
        'highway': '113',
        'priority': 9,
        'distance_from_dc_km': 340,
        'description': 'พิจิตรตะวันออก สาย 113'
    },
    'ZONE_F3_พิจิตร_สาย117': {
        'provinces': ['พิจิตร'],
        'districts': ['โพธิ์ประทับช้าง', 'บึงนาราง', 'วชิรบารมี', 'โพทะเล'],
        'highway': '117',
        'priority': 10,
        'distance_from_dc_km': 320,
        'description': 'พิจิตรสาย 117'
    },
    'ZONE_G_นครสวรรค์': {
        'provinces': ['นครสวรรค์'],
        'districts': ['เมืองนครสวรรค์', 'หนองบัว', 'ท่าตะโก', 'ไพศาลี', 'ตาคลี', 'บรรพตพิสัย', 'ชุมตาบง', 'ลาดยาว', 'ตากฟ้า', 'พยุหะคีรี', 'โกรกพระ', 'เก้าเลี้ยว', 'ชุมแสง', 'แม่วงก์', 'แม่เปิน'],
        'highway': '1/32',
        'priority': 11,
        'distance_from_dc_km': 240,
        'description': 'ประตูเหนือ สาย 1/32'
    },
    # ============ ภาคอีสาน - สาย 2/24 (เพิ่มโซนย่อย) ============
    # --- โคราช (sub-zones ใหม่ละเอียด) ---
    'ZONE_KRM1_โคราช_เมือง': {
        'provinces': ['นครราชสีมา'],
        'districts': ['เมืองนครราชสีมา', 'ปักธงชัย'],
        'highway': '2',
        'priority': 35,
        'distance_from_dc_km': 250,
        'description': 'โคราชในเมือง ประตูอีสาน'
    },
    'ZONE_KRM2_โคราช_ตะวันตก': {
        'provinces': ['นครราชสีมา'],
        'districts': ['สีคิ้ว', 'สูงเนิน', 'ด่านขุนทด', 'เทพารักษ์'],
        'highway': '2',
        'priority': 34,
        'distance_from_dc_km': 270,
        'description': 'โคราชตะวันตก สาย 2'
    },
    'ZONE_KRM3_โคราช_ใต้': {
        'provinces': ['นครราชสีมา'],
        'districts': ['ปากช่อง', 'วังน้ำเขียว', 'เฉลิมพระเกียรติ'],
        'highway': '2/304',
        'priority': 33,
        'distance_from_dc_km': 280,
        'description': 'โคราชใต้ ปากช่อง-เขาใหญ่'
    },
    'ZONE_KRM4_โคราช_ตะวันออก': {
        'provinces': ['นครราชสีมา'],
        'districts': ['โชคชัย', 'ครบุรี', 'โนนแดง', 'แก้งสนามนาง'],
        'highway': '304',
        'priority': 36,
        'distance_from_dc_km': 260,
        'description': 'โคราชตะวันออก'
    },
    'ZONE_KRM5_โคราช_เหนือ': {
        'provinces': ['นครราชสีมา'],
        'districts': ['โนนสูง', 'พิมาย', 'บัวใหญ่', 'ประทาย', 'ชุมพวง', 'ห้วยแถลง', 'บ้านเหลื่อม'],
        'highway': '2',
        'priority': 37,
        'distance_from_dc_km': 255,
        'description': 'โคราชเหนือ สาย 2 มิตรภาพ'
    },
    'ZONE_H1_โคราช_เมือง': {
        'provinces': ['นครราชสีมา'],
        'districts': ['เมืองนครราชสีมา', 'ปักธงชัย'],
        'highway': '2',
        'priority': 12,
        'distance_from_dc_km': 260,
        'description': 'โคราชในเมือง ประตูอีสาน'
    },
    'ZONE_H2_โคราช_ตะวันออก': {
        'provinces': ['นครราชสีมา'],
        'districts': ['บัวใหญ่', 'ครบุรี', 'สีคิ้ว', 'สูงเนิน', 'โนนสูง', 'โนนแดง', 'ด่านขุนทด'],
        'highway': '2/304',
        'priority': 12,
        'distance_from_dc_km': 280,
        'description': 'โคราชตะวันออก-เขาใหญ่'
    },
    'ZONE_H3_โคราช_เหนือ': {
        'provinces': ['นครราชสีมา'],
        'districts': ['พิมาย', 'ห้วยแถลง', 'บ้านเหลื่อม', 'โชคชัย', 'แก้งสนามนาง', 'เทพารักษ์'],
        'highway': '2',
        'priority': 12,
        'distance_from_dc_km': 270,
        'description': 'โคราชเหนือ-เส้นมิตรภาพ'
    },
    'ZONE_H4_โคราช_ใต้': {
        'provinces': ['นครราชสีมา'],
        'districts': ['ปากช่อง', 'วังน้ำเขียว', 'เฉลิมพระเกียรติ', 'คง', 'ชุมพวง'],
        'highway': '2/304',
        'priority': 12,
        'distance_from_dc_km': 290,
        'description': 'โคราชใต้-เส้น304'
    },
    # --- ขอนแก่น (sub-zones ใหม่ละเอียด) ---
    'ZONE_KKN1_ขอนแก่น_เมือง': {
        'provinces': ['ขอนแก่น'],
        'districts': ['เมืองขอนแก่น', 'พระยืน', 'ซำสูง'],
        'highway': '2',
        'priority': 28,
        'distance_from_dc_km': 380,
        'description': 'ขอนแก่นในเมือง Hub อีสานกลาง'
    },
    'ZONE_KKN2_ขอนแก่น_รอบนอกเหนือ': {
        'provinces': ['ขอนแก่น'],
        'districts': ['น้ำพอง', 'อุบลรัตน์', 'หนองเรือ', 'ชุมแพ', 'ภูผาม่าน', 'ภูเวียง'],
        'highway': '2',
        'priority': 27,
        'distance_from_dc_km': 400,
        'description': 'ขอนแก่นรอบนอกเหนือ ชุมแพ-น้ำพอง'
    },
    'ZONE_KKN3_ขอนแก่น_ใต้': {
        'provinces': ['ขอนแก่น'],
        'districts': ['บ้านไผ่', 'บ้านฝาง', 'ชนบท', 'พล', 'แวงใหญ่', 'แวงน้อย', 'มัญจาคีรี', 'เปือยน้อย', 'กระนวน', 'หนองสองห้อง', 'โนนศิลา'],
        'highway': '2',
        'priority': 29,
        'distance_from_dc_km': 420,
        'description': 'ขอนแก่นใต้ บ้านไผ่-พล'
    },
    'ZONE_I1_ขอนแก่น_เมือง': {
        'provinces': ['ขอนแก่น'],
        'districts': ['เมืองขอนแก่น', 'น้ำพอง', 'อุบลรัตน์', 'บ้านไผ่'],
        'highway': '2',
        'priority': 13,
        'distance_from_dc_km': 450,
        'description': 'ขอนแก่นในเมือง Hub อีสานกลาง'
    },
    'ZONE_I2_ขอนแก่น_ใต้': {
        'provinces': ['ขอนแก่น'],
        'districts': ['บ้านฝาง', 'ชนบท', 'พล', 'แวงใหญ่', 'แวงน้อย', 'มัญจาคีรี'],
        'highway': '2',
        'priority': 13,
        'distance_from_dc_km': 470,
        'description': 'ขอนแก่นใต้'
    },
    'ZONE_I3_ขอนแก่น_เหนือ': {
        'provinces': ['ขอนแก่น'],
        'districts': ['กระนวน', 'ซำสูง', 'เปือยน้อย', 'พระยืน', 'ภูผาม่าน', 'หนองสองห้อง', 'หนองเรือ'],
        'highway': '2',
        'priority': 13,
        'distance_from_dc_km': 460,
        'description': 'ขอนแก่นเหนือ'
    },
    'ZONE_I4_มหาสารคาม': {
        'provinces': ['มหาสารคาม'],
        'districts': ['เมืองมหาสารคาม', 'กันทรวิชัย', 'แกดำ', 'โกสุมพิสัย', 'ชื่นชม', 'นาเชือก', 'นาดูน', 'บรบือ', 'พยัคฆภูมิพิสัย', 'วาปีปทุม', 'ยางสีสุราช'],
        'highway': '2',
        'priority': 13,
        'distance_from_dc_km': 480,
        'description': 'มหาสารคาม'
    },
    'ZONE_I5_ร้อยเอ็ด': {
        'provinces': ['ร้อยเอ็ด'],
        'districts': ['เมืองร้อยเอ็ด', 'เกษตรวิสัย', 'ปทุมรัตต์', 'ธวัชบุรี', 'พนมไพร', 'โพนทอง', 'เมืองสรวง', 'เสลภูมิ', 'สุวรรณภูมิ', 'อาจสามารถ'],
        'highway': '2/214',
        'priority': 13,
        'distance_from_dc_km': 510,
        'description': 'ร้อยเอ็ด'
    },
    'ZONE_I6_กาฬสินธุ์': {
        'provinces': ['กาฬสินธุ์'],
        'districts': ['เมืองกาฬสินธุ์', 'กมลาไสย', 'กุฉินารายณ์', 'เขาวง', 'คำม่วง', 'ดอนจาน', 'ท่าคันโท', 'นาคู', 'นามน', 'ยางตลาด', 'ฆ้องชัย', 'ร่องคำ', 'สหัสขันธ์', 'สมเด็จ', 'สามชัย', 'หนองกุงศรี', 'ห้วยเม็ก', 'ห้วยผึ้ง'],
        'highway': '213',
        'priority': 13,
        'distance_from_dc_km': 520,
        'description': 'กาฬสินธุ์'
    },
    # --- อุดรธานี (sub-zones) ---
    'ZONE_UDR1_อุดร_เมือง': {
        'provinces': ['อุดรธานี'],
        'districts': ['เมืองอุดรธานี', 'กุดจับ', 'หนองวัวซอ'],
        'highway': '2',
        'priority': 22,
        'distance_from_dc_km': 470,
        'description': 'อุดรธานีในเมือง'
    },
    'ZONE_UDR2_อุดร_รอบนอก': {
        'provinces': ['อุดรธานี'],
        'districts': ['กุมภวาปี', 'หนองหาน', 'เพ็ญ', 'บ้านดุง', 'บ้านผือ', 'วังสามหมอ', 'กู่แก้ว', 'ไชยวาน', 'ประจักษ์ศิลปาคม'],
        'highway': '2',
        'priority': 21,
        'distance_from_dc_km': 490,
        'description': 'อุดรธานีรอบนอก กุมภวาปี-หนองหาน'
    },
    # --- บุรีรัมย์ (5 sub-zones ตามภูมิศาสตร์จริง) ---
    'ZONE_BRM1_บุรีรัมย์_เมือง': {
        'provinces': ['บุรีรัมย์'],
        'districts': ['เมืองบุรีรัมย์'],
        'highway': '24',
        'priority': 26,
        'distance_from_dc_km': 360,
        'description': 'บุรีรัมย์ใจกลางเมือง'
    },
    'ZONE_BRM2_บุรีรัมย์_เหนือ': {
        'provinces': ['บุรีรัมย์'],
        'districts': ['ลำปลายมาศ', 'คูเมือง', 'หนองหงส์', 'ชำนิ', 'บ้านด่าน'],
        'highway': '24',
        'priority': 25,
        'distance_from_dc_km': 370,
        'description': 'บุรีรัมย์เหนือ ลำปลายมาศ-คูเมือง'
    },
    'ZONE_BRM3_บุรีรัมย์_ตะวันออก': {
        'provinces': ['บุรีรัมย์'],
        'districts': ['นางรอง', 'หนองกี่', 'ละหานทราย', 'ปะคำ', 'โนนดินแดง', 'เฉลิมพระเกียรติ'],
        'highway': '24/348',
        'priority': 24,
        'distance_from_dc_km': 400,
        'description': 'บุรีรัมย์ตะวันออก นางรอง-หนองกี่'
    },
    'ZONE_BRM4_บุรีรัมย์_ใต้': {
        'provinces': ['บุรีรัมย์'],
        'districts': ['ประโคนชัย', 'บ้านกรวด', 'กระสัง', 'พลับพลาชัย'],
        'highway': '24/219',
        'priority': 23,
        'distance_from_dc_km': 390,
        'description': 'บุรีรัมย์ใต้ ประโคนชัย-บ้านกรวด ชายแดนกัมพูชา'
    },
    'ZONE_BRM5_บุรีรัมย์_ตะวันตก': {
        'provinces': ['บุรีรัมย์'],
        'districts': ['สตึก', 'พุทไธสง', 'บ้านใหม่ไชยพจน์', 'แคนดง', 'นาโพธิ์', 'พลับพลาชัย'],
        'highway': '226',
        'priority': 25,
        'distance_from_dc_km': 380,
        'description': 'บุรีรัมย์ตะวันตก สตึก-พุทไธสง ฝั่งแม่น้ำมูล'
    },
    # --- อุบลราชธานี (sub-zones) ---
    'ZONE_UBL1_อุบล_เมือง': {
        'provinces': ['อุบลราชธานี'],
        'districts': ['เมืองอุบลราชธานี', 'วารินชำราบ', 'สว่างวีระวงศ์', 'เขื่องใน'],
        'highway': '24',
        'priority': 23,
        'distance_from_dc_km': 500,
        'description': 'อุบลราชธานีในเมือง'
    },
    'ZONE_UBL2_อุบล_รอบนอก': {
        'provinces': ['อุบลราชธานี'],
        'districts': ['ตระการพืชผล', 'เดชอุดม', 'พิบูลมังสาหาร', 'บุณฑริก', 'น้ำยืน', 'ศรีเมืองใหม่', 'ม่วงสามสิบ', 'น้ำขุ่น'],
        'highway': '24',
        'priority': 22,
        'distance_from_dc_km': 520,
        'description': 'อุบลราชธานีรอบนอก ตระการ-เดชอุดม'
    },
    'ZONE_J_อุดร': {
        'provinces': ['อุดรธานี', 'หนองคาย', 'หนองบัวลำภู', 'เลย', 'สกลนคร', 'นครพนม', 'บึงกาฬ'],
        'highway': '2',
        'priority': 14,
        'distance_from_dc_km': 560,
        'description': 'อีสานเหนือ สาย 2 ปลายทาง'
    },
    'ZONE_K_อีสานใต้': {
        'provinces': ['บุรีรัมย์', 'สุรินทร์', 'ศรีสะเกษ', 'อุบลราชธานี', 'ยโสธร', 'อำนาจเจริญ', 'มุกดาหาร'],
        'highway': '24',
        'priority': 15,
        'distance_from_dc_km': 500,
        'description': 'อีสานใต้ สาย 24'
    },
    # ============ ภาคตะวันออก - สาย 3 ============
    # --- ฉะเชิงเทรา (sub-zones) ---
    'ZONE_CCT1_ฉะเชิงเทรา_เมือง': {
        'provinces': ['ฉะเชิงเทรา'],
        'districts': ['เมืองฉะเชิงเทรา', 'บางน้ำเปรี้ยว', 'บ้านโพธิ์'],
        'highway': '304',
        'priority': 70,
        'distance_from_dc_km': 70,
        'description': 'ฉะเชิงเทราในเมือง สาย 304'
    },
    'ZONE_CCT2_ฉะเชิงเทรา_ตะวันออก': {
        'provinces': ['ฉะเชิงเทรา'],
        'districts': ['บางปะกง', 'พนมสารคาม', 'บางคล้า', 'แปลงยาว'],
        'highway': '304/331',
        'priority': 69,
        'distance_from_dc_km': 80,
        'description': 'ฉะเชิงเทราตะวันออก สาย 331'
    },
    # --- ชลบุรี (sub-zones) ---
    'ZONE_CBI1_ชลบุรี_เหนือ': {
        'provinces': ['ชลบุรี'],
        'districts': ['บ้านบึง', 'พนัสนิคม', 'บ่อทอง', 'หนองใหญ่', 'เกาะจันทร์'],
        'highway': '304/331',
        'priority': 67,
        'distance_from_dc_km': 90,
        'description': 'ชลบุรีเหนือในแผ่นดิน'
    },
    'ZONE_CBI2_ชลบุรี_เมือง': {
        'provinces': ['ชลบุรี'],
        'districts': ['เมืองชลบุรี', 'พานทอง'],
        'highway': '3',
        'priority': 66,
        'distance_from_dc_km': 110,
        'description': 'ชลบุรีในเมือง สาย 3'
    },
    'ZONE_CBI3_ชลบุรี_ศรีราชา': {
        'provinces': ['ชลบุรี'],
        'districts': ['ศรีราชา'],
        'highway': '3',
        'priority': 65,
        'distance_from_dc_km': 120,
        'description': 'ศรีราชา EEC สาย 3'
    },
    'ZONE_CBI4_ชลบุรี_พัทยา': {
        'provinces': ['ชลบุรี'],
        'districts': ['บางละมุง', 'สัตหีบ'],
        'highway': '3',
        'priority': 64,
        'distance_from_dc_km': 130,
        'description': 'พัทยา-สัตหีบ ชายฝั่ง'
    },
    # --- ระยอง (sub-zones) ---
    'ZONE_RYG1_ระยอง_เมือง': {
        'provinces': ['ระยอง'],
        'districts': ['เมืองระยอง', 'บ้านฉาง', 'นิคมพัฒนา', 'ปลวกแดง', 'บ้านค่าย'],
        'highway': '3',
        'priority': 63,
        'distance_from_dc_km': 140,
        'description': 'ระยองในเมือง+นิคมอุตสาหกรรม'
    },
    'ZONE_RYG2_ระยอง_ตะวันออก': {
        'provinces': ['ระยอง'],
        'districts': ['แกลง', 'วังจันทร์', 'เขาชะเมา'],
        'highway': '36',
        'priority': 62,
        'distance_from_dc_km': 160,
        'description': 'ระยองตะวันออก แกลง-วังจันทร์'
    },
    'ZONE_L_ชลบุรีระยอง': {
        'provinces': ['ชลบุรี', 'ระยอง'],
        # เฉพาะชลบุรีชายฝั่ง (เมือง/ศรีราชา/บ้านบึง/พัทยา/สัตหีบ) + ระยอง ทางสาย 3 (Bang Na-Trat)
        'highway': '3',
        'priority': 16,
        'distance_from_dc_km': 120,
        'description': 'ชลบุรีชายฝั่ง+ระยอง สาย 3 (Bang Na→Trat) EEC'
    },
    # ชลบุรีเหนือ/ในแผ่นดิน: พนัสนิคม/บ่อทอง/หนองใหญ่ เข้าถึงผ่านสาย 304→31 จากฉะเชิงเทรา
    'ZONE_L1_ชลบุรีเหนือ': {
        'provinces': ['ชลบุรี'],
        'districts': ['พนัสนิคม', 'บ่อทอง', 'หนองใหญ่', 'เกาะจันทร์'],
        'highway': '304/331',
        'priority': 16.5,
        'distance_from_dc_km': 90,
        'description': 'ชลบุรีเหนือในแผ่นดิน สาย 304→331 ผ่านฉะเชิงเทรา'
    },
    # สมุทรปราการ ฝั่งตะวันออก: บางบ่อ-บางเสาธง อยู่บนสาย 3 เส้นทางเดียวกับชลบุรี
    'ZONE_EAST_สมุทรปราการ': {
        'provinces': ['สมุทรปราการ'],
        'districts': ['บางบ่อ', 'บางเสาธง'],
        'highway': '3/331',
        'priority': 28,
        'distance_from_dc_km': 55,
        'description': 'สมุทรปราการตะวันออก บางบ่อ-บางเสาธง สาย 3 ต่อชลบุรี'
    },
    'ZONE_M_จันทบุรีตราด': {
        'provinces': ['จันทบุรี', 'ตราด'],
        'highway': '3',
        'priority': 17,
        'distance_from_dc_km': 300,
        'description': 'ตะวันออกไกล สาย 3 ปลายทาง'
    },
    # ============ ภาคใต้ - สาย 4 ============
    'ZONE_N_ใต้ตอนบน': {
        'provinces': ['เพชรบุรี', 'ประจวบคีรีขันธ์', 'ระนอง'],
        'highway': '4',
        'priority': 18,
        'distance_from_dc_km': 400,
        'description': 'ใต้ตอนบน สาย 4 (ไม่รวมชุมพร)'
    },
    # ชุมพร - แยกโซนย่อย
    'ZONE_N1_ชุมพรเหนือ': {
        'provinces': ['ชุมพร'],
        'districts': ['ปะทิว', 'สวี', 'ละแม', 'เมืองชุมพร'],
        'highway': '4',
        'priority': 18.1,
        'distance_from_dc_km': 420,
        'description': 'ชุมพรเหนือ (ปะทิว-สวี-เมือง)'
    },
    'ZONE_N2_ชุมพรใต้': {
        'provinces': ['ชุมพร'],
        'districts': ['ทุ่งตะโก', 'พะโต๊ะ', 'หลังสวน', 'ท่าแซะ'],
        'highway': '4',
        'priority': 18.2,
        'distance_from_dc_km': 450,
        'description': 'ชุมพรใต้ (ทุ่งตะโก-หลังสวน)'
    },
    'ZONE_N3_ชุมพรกลาง': {
        'provinces': ['ชุมพร'],
        'districts': ['บางสะพานน้อย', 'ทับสะแก', 'บางสะพาน'],
        'highway': '4',
        'priority': 18.3,
        'distance_from_dc_km': 440,
        'description': 'ชุมพรกลาง (บางสะพานน้อย-ทับสะแก-บางสะพาน)'
    },
    'ZONE_O_ใต้อ่าวไทย': {
        'provinces': ['สุราษฎร์ธานี', 'นครศรีธรรมราช', 'พัทลุง', 'สงขลา', 'ปัตตานี', 'ยะลา', 'นราธิวาส'],
        'highway': '4',
        'priority': 19,
        'distance_from_dc_km': 700,
        'description': 'ใต้ฝั่งอ่าวไทย สาย 4'
    },
    'ZONE_P_ใต้อันดามัน': {
        'provinces': ['กระบี่', 'พังงา', 'ภูเก็ต', 'ตรัง', 'สตูล'],
        'highway': '401/402',
        'priority': 20,
        'distance_from_dc_km': 850,
        'description': 'ใต้ฝั่งอันดามัน สาย 401/402'
    },
    # ============ ปริมณฑล (แบ่งโซนละเอียด) ============
    # --- พระนครศรีอยุธยา (sub-zones) ---
    'ZONE_AYA1_อยุธยา_ใกล้DC': {
        'provinces': ['พระนครศรีอยุธยา'],
        'districts': ['วังน้อย', 'บางปะอิน', 'อุทัย', 'บางบาล', 'นครหลวง'],
        'highway': '1/32',
        'priority': 88,
        'distance_from_dc_km': 30,
        'description': 'อยุธยาใกล้ DC วังน้อย-บางปะอิน'
    },
    'ZONE_AYA2_อยุธยา_เมือง': {
        'provinces': ['พระนครศรีอยุธยา'],
        'districts': ['พระนครศรีอยุธยา', 'บางไทร', 'บางซ้าย', 'ผักไห่', 'บางปะหัน', 'ท่าเรือ', 'ภาชี'],
        'highway': '1/32',
        'priority': 87,
        'distance_from_dc_km': 45,
        'description': 'อยุธยาในเมือง-รอบนอก'
    },
    # --- ปทุมธานี (sub-zones) ---
    'ZONE_PTM1_ปทุมธานี_คลองหลวง': {
        'provinces': ['ปทุมธานี'],
        'districts': ['คลองหลวง'],
        'highway': '1/305',
        'priority': 82,
        'distance_from_dc_km': 50,
        'description': 'ปทุมธานี คลองหลวง ม.เกษตร-รังสิต'
    },
    'ZONE_PTM2_ปทุมธานี_ลำลูกกา': {
        'provinces': ['ปทุมธานี'],
        'districts': ['ลำลูกกา', 'ธัญบุรี'],
        'highway': '1/305',
        'priority': 82,
        'distance_from_dc_km': 55,
        'description': 'ปทุมธานี ลำลูกกา-ธัญบุรี'
    },
    'ZONE_PTM3_ปทุมธานี_เมือง': {
        'provinces': ['ปทุมธานี'],
        'districts': ['เมืองปทุมธานี', 'สามโคก', 'ลาดหลุมแก้ว'],
        'highway': '1/306',
        'priority': 83,
        'distance_from_dc_km': 45,
        'description': 'ปทุมธานี เมือง-สามโคก-ลาดหลุมแก้ว'
    },
    # --- นนทบุรี (sub-zones) ---
    'ZONE_NBI1_นนทบุรี_บางบัวทอง': {
        'provinces': ['นนทบุรี'],
        'districts': ['บางบัวทอง', 'บางใหญ่', 'ไทรน้อย'],
        'highway': '9/346',
        'priority': 84,
        'distance_from_dc_km': 60,
        'description': 'นนทบุรี บางบัวทอง-บางใหญ่-ไทรน้อย'
    },
    'ZONE_NBI2_นนทบุรี_ปากเกร็ด': {
        'provinces': ['นนทบุรี'],
        'districts': ['ปากเกร็ด', 'บางกรวย'],
        'highway': '9/302',
        'priority': 85,
        'distance_from_dc_km': 65,
        'description': 'นนทบุรี ปากเกร็ด-บางกรวย'
    },
    'ZONE_NBI3_นนทบุรี_เมือง': {
        'provinces': ['นนทบุรี'],
        'districts': ['เมืองนนทบุรี'],
        'highway': '9/302',
        'priority': 85,
        'distance_from_dc_km': 65,
        'description': 'นนทบุรี เมือง'
    },
    # --- สมุทรปราการ (sub-zones ฝั่งตะวันตก/เมือง) ---
    'ZONE_SPK1_สมุทรปราการ_เมือง': {
        'provinces': ['สมุทรปราการ'],
        'districts': ['เมืองสมุทรปราการ', 'พระประแดง', 'พระสมุทรเจดีย์'],
        'highway': '3/34',
        'priority': 86,
        'distance_from_dc_km': 60,
        'description': 'สมุทรปราการ เมือง-พระประแดง'
    },
    'ZONE_SPK2_สมุทรปราการ_บางพลี': {
        'provinces': ['สมุทรปราการ'],
        'districts': ['บางพลี'],
        'highway': '3',
        'priority': 87,
        'distance_from_dc_km': 55,
        'description': 'สมุทรปราการ บางพลี'
    },
    # --- นครปฐม (sub-zones) ---
    'ZONE_NPT1_นครปฐม_เมือง': {
        'provinces': ['นครปฐม'],
        'districts': ['เมืองนครปฐม', 'พุทธมณฑล'],
        'highway': '4/35',
        'priority': 75,
        'distance_from_dc_km': 80,
        'description': 'นครปฐมในเมือง-พุทธมณฑล'
    },
    'ZONE_NPT2_นครปฐม_สามพราน': {
        'provinces': ['นครปฐม'],
        'districts': ['สามพราน', 'นครชัยศรี'],
        'highway': '4',
        'priority': 75,
        'distance_from_dc_km': 75,
        'description': 'นครปฐม สามพราน-นครชัยศรี'
    },
    'ZONE_NPT3_นครปฐม_กำแพงแสน': {
        'provinces': ['นครปฐม'],
        'districts': ['กำแพงแสน', 'ดอนตูม', 'บางเลน'],
        'highway': '4',
        'priority': 74,
        'distance_from_dc_km': 90,
        'description': 'นครปฐม กำแพงแสน-บางเลน'
    },
    # --- สมุทรสาคร (sub-zones) ---
    'ZONE_SSK1_สมุทรสาคร_เมือง': {
        'provinces': ['สมุทรสาคร'],
        'districts': ['เมืองสมุทรสาคร'],
        'highway': '35',
        'priority': 76,
        'distance_from_dc_km': 75,
        'description': 'สมุทรสาคร มหาชัย'
    },
    'ZONE_SSK2_สมุทรสาคร_กระทุ่มแบน': {
        'provinces': ['สมุทรสาคร'],
        'districts': ['กระทุ่มแบน', 'บ้านแพ้ว'],
        'highway': '35',
        'priority': 76,
        'distance_from_dc_km': 80,
        'description': 'สมุทรสาคร กระทุ่มแบน-บ้านแพ้ว'
    },
    # ── BKK sub-zones (แยกย่อยตามทิศ เขตใกล้กันอยู่โซนเดียวกัน) ──────────
    # ── BKK per-เขต zones: แต่ละเขต = 1 โซน → ทริปหมดเขตก่อนค่อยข้าม ──
    # กลุ่มเหนือ
    'ZONE_BKK_ดุสิต':           {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ดุสิต'],           'highway': '1/9', 'priority': 94, 'distance_from_dc_km': 55},
    'ZONE_BKK_บางซื่อ':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางซื่อ'],         'highway': '1/9', 'priority': 94, 'distance_from_dc_km': 55},
    'ZONE_BKK_จตุจักร':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['จตุจักร'],         'highway': '1/9', 'priority': 94, 'distance_from_dc_km': 52},
    'ZONE_BKK_บางเขน':          {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางเขน'],          'highway': '1/9', 'priority': 95, 'distance_from_dc_km': 48},
    'ZONE_BKK_ลาดพร้าว':        {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ลาดพร้าว'],        'highway': '1/9', 'priority': 95, 'distance_from_dc_km': 46},
    'ZONE_BKK_วังทองหลาง':      {'provinces': ['กรุงเทพมหานคร'], 'districts': ['วังทองหลาง'],      'highway': '1/9', 'priority': 95, 'distance_from_dc_km': 47},
    'ZONE_BKK_คันนายาว':        {'provinces': ['กรุงเทพมหานคร'], 'districts': ['คันนายาว'],        'highway': '1/9', 'priority': 95, 'distance_from_dc_km': 43},
    'ZONE_BKK_หลักสี่':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['หลักสี่'],         'highway': '1',   'priority': 93, 'distance_from_dc_km': 40},
    'ZONE_BKK_ดอนเมือง':        {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ดอนเมือง'],        'highway': '1',   'priority': 93, 'distance_from_dc_km': 35},
    'ZONE_BKK_สายไหม':          {'provinces': ['กรุงเทพมหานคร'], 'districts': ['สายไหม'],          'highway': '1',   'priority': 93, 'distance_from_dc_km': 38},
    # กลุ่มกลาง
    'ZONE_BKK_พระนคร':          {'provinces': ['กรุงเทพมหานคร'], 'districts': ['พระนคร'],          'highway': 'CBD', 'priority': 96, 'distance_from_dc_km': 62},
    'ZONE_BKK_ป้อมปราบศัตรูพ่าย':{'provinces': ['กรุงเทพมหานคร'], 'districts': ['ป้อมปราบศัตรูพ่าย'],'highway': 'CBD','priority': 96, 'distance_from_dc_km': 62},
    'ZONE_BKK_สัมพันธวงศ์':     {'provinces': ['กรุงเทพมหานคร'], 'districts': ['สัมพันธวงศ์'],     'highway': 'CBD', 'priority': 96, 'distance_from_dc_km': 60},
    'ZONE_BKK_ราชเทวี':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ราชเทวี'],         'highway': 'CBD', 'priority': 96, 'distance_from_dc_km': 58},
    'ZONE_BKK_พญาไท':           {'provinces': ['กรุงเทพมหานคร'], 'districts': ['พญาไท'],           'highway': 'CBD', 'priority': 96, 'distance_from_dc_km': 57},
    'ZONE_BKK_ดินแดง':          {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ดินแดง'],          'highway': 'CBD', 'priority': 96, 'distance_from_dc_km': 55},
    'ZONE_BKK_ห้วยขวาง':        {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ห้วยขวาง'],        'highway': 'CBD', 'priority': 96, 'distance_from_dc_km': 53},
    'ZONE_BKK_ปทุมวัน':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ปทุมวัน'],         'highway': 'CBD', 'priority': 97, 'distance_from_dc_km': 60},
    'ZONE_BKK_บางรัก':          {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางรัก'],          'highway': 'CBD', 'priority': 97, 'distance_from_dc_km': 62},
    # กลุ่มตะวันออก
    'ZONE_BKK_วัฒนา':           {'provinces': ['กรุงเทพมหานคร'], 'districts': ['วัฒนา'],           'highway': '3/9', 'priority': 97, 'distance_from_dc_km': 57},
    'ZONE_BKK_คลองเตย':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['คลองเตย'],         'highway': '3/9', 'priority': 97, 'distance_from_dc_km': 58},
    'ZONE_BKK_พระโขนง':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['พระโขนง'],         'highway': '3',   'priority': 97, 'distance_from_dc_km': 55},
    'ZONE_BKK_บางกะปิ':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางกะปิ'],         'highway': '3/9', 'priority': 97, 'distance_from_dc_km': 52},
    'ZONE_BKK_บึงกุ่ม':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บึงกุ่ม'],         'highway': '3/9', 'priority': 97, 'distance_from_dc_km': 48},
    'ZONE_BKK_สะพานสูง':        {'provinces': ['กรุงเทพมหานคร'], 'districts': ['สะพานสูง'],        'highway': '3/9', 'priority': 97, 'distance_from_dc_km': 46},
    'ZONE_BKK_บางนา':           {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางนา'],           'highway': '3',   'priority': 96, 'distance_from_dc_km': 53},
    'ZONE_BKK_สวนหลวง':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['สวนหลวง'],         'highway': '3',   'priority': 96, 'distance_from_dc_km': 50},
    'ZONE_BKK_ประเวศ':           {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ประเวศ'],           'highway': '3',   'priority': 96, 'distance_from_dc_km': 48},
    'ZONE_BKK_ลาดกระบัง':       {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ลาดกระบัง'],       'highway': '3/9', 'priority': 95, 'distance_from_dc_km': 42},
    'ZONE_BKK_มีนบุรี':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['มีนบุรี'],         'highway': '3/9', 'priority': 95, 'distance_from_dc_km': 40},
    'ZONE_BKK_คลองสามวา':       {'provinces': ['กรุงเทพมหานคร'], 'districts': ['คลองสามวา'],       'highway': '3/9', 'priority': 95, 'distance_from_dc_km': 38},
    'ZONE_BKK_หนองจอก':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['หนองจอก'],         'highway': '3/9', 'priority': 95, 'distance_from_dc_km': 36},
    # กลุ่มใต้
    'ZONE_BKK_ยานนาวา':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ยานนาวา'],         'highway': '35',  'priority': 98, 'distance_from_dc_km': 62},
    'ZONE_BKK_สาทร':            {'provinces': ['กรุงเทพมหานคร'], 'districts': ['สาทร'],            'highway': '35',  'priority': 98, 'distance_from_dc_km': 63},
    'ZONE_BKK_บางคอแหลม':       {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางคอแหลม'],       'highway': '35',  'priority': 98, 'distance_from_dc_km': 62},
    'ZONE_BKK_ราษฎร์บูรณะ':     {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ราษฎร์บูรณะ'],     'highway': '35',  'priority': 98, 'distance_from_dc_km': 60},
    'ZONE_BKK_ทุ่งครุ':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ทุ่งครุ'],         'highway': '35',  'priority': 98, 'distance_from_dc_km': 58},
    'ZONE_BKK_จอมทอง':          {'provinces': ['กรุงเทพมหานคร'], 'districts': ['จอมทอง'],          'highway': '35',  'priority': 98, 'distance_from_dc_km': 60},
    'ZONE_BKK_บางบอน':          {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางบอน'],          'highway': '35',  'priority': 98, 'distance_from_dc_km': 57},
    'ZONE_BKK_บางขุนเทียน':     {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางขุนเทียน'],     'highway': '35',  'priority': 98, 'distance_from_dc_km': 55},
    # กลุ่มตะวันตก/ธนบุรี
    'ZONE_BKK_บางกอกน้อย':      {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางกอกน้อย'],      'highway': '35/4','priority': 99, 'distance_from_dc_km': 60},
    'ZONE_BKK_บางกอกใหญ่':      {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางกอกใหญ่'],      'highway': '35/4','priority': 99, 'distance_from_dc_km': 62},
    'ZONE_BKK_ธนบุรี':          {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ธนบุรี'],          'highway': '35/4','priority': 99, 'distance_from_dc_km': 62},
    'ZONE_BKK_คลองสาน':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['คลองสาน'],         'highway': '35/4','priority': 99, 'distance_from_dc_km': 62},
    'ZONE_BKK_ตลิ่งชัน':        {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ตลิ่งชัน'],        'highway': '35/4','priority': 99, 'distance_from_dc_km': 58},
    'ZONE_BKK_ทวีวัฒนา':        {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ทวีวัฒนา'],        'highway': '4',   'priority': 99, 'distance_from_dc_km': 56},
    'ZONE_BKK_ภาษีเจริญ':       {'provinces': ['กรุงเทพมหานคร'], 'districts': ['ภาษีเจริญ'],       'highway': '35/4','priority': 99, 'distance_from_dc_km': 58},
    'ZONE_BKK_หนองแขม':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['หนองแขม'],         'highway': '4',   'priority': 99, 'distance_from_dc_km': 55},
    'ZONE_BKK_บางแค':           {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางแค'],           'highway': '4',   'priority': 99, 'distance_from_dc_km': 56},
    'ZONE_BKK_บางพลัด':         {'provinces': ['กรุงเทพมหานคร'], 'districts': ['บางพลัด'],         'highway': '35/4','priority': 99, 'distance_from_dc_km': 60},
    # Fallback กรุงเทพ กรณีเขตไม่ match
    'ZONE_NEARBY_นนทบุรี': {
        'provinces': ['นนทบุรี'],
        'districts': ['เมืองนนทบุรี', 'บางกรวย', 'บางใหญ่', 'บางบัวทอง', 'ไทรน้อย', 'ปากเกร็ด'],
        'highway': '9/35',
        'priority': 99,
        'distance_from_dc_km': 35,
        'description': 'นนทบุรี ใกล้กทม'
    },
    'ZONE_NEARBY_ปทุมธานี': {
        'provinces': ['ปทุมธานี'],
        'districts': ['เมืองปทุมธานี', 'คลองหลวง', 'ธัญบุรี', 'หนองเสือ', 'ลาดหลุมแก้ว', 'ลำลูกกา', 'สามโคก'],
        'highway': '1/9/305',
        'priority': 99,
        'distance_from_dc_km': 25,
        'description': 'ปทุมธานี ใกล้ DC วังน้อย'
    },
    'ZONE_NEARBY_สมุทรปราการ': {
        'provinces': ['สมุทรปราการ'],
        'districts': ['เมืองสมุทรปราการ', 'บางพลี', 'พระประแดง', 'พระสมุทรเจดีย์'],
        'highway': '3/9/34',
        'priority': 99,
        'distance_from_dc_km': 40,
        'description': 'สมุทรปราการตะวันตก เมือง-พระประแดง-บางพลี (ฝั่งกทม)'
    },
    'ZONE_NEARBY_สมุทรสาคร': {
        'provinces': ['สมุทรสาคร'],
        'districts': ['เมืองสมุทรสาคร', 'กระทุ่มแบน', 'บ้านแพ้ว'],
        'highway': '35',
        'priority': 99,
        'distance_from_dc_km': 50,
        'description': 'สมุทรสาคร มหาชัย'
    },
    'ZONE_NEARBY_นครปฐม': {
        'provinces': ['นครปฐม'],
        'districts': ['เมืองนครปฐม', 'กำแพงแสน', 'นครชัยศรี', 'ดอนตูม', 'บางเลน', 'สามพราน', 'พุทธมณฑล'],
        'highway': '35/4',
        'priority': 99,
        'distance_from_dc_km': 55,
        'description': 'นครปฐม ม.เกษตร-สามพราน'
    },
    'ZONE_NEARBY_อยุธยา': {
        'provinces': ['พระนครศรีอยุธยา'],
        'districts': ['เมืองพระนครศรีอยุธยา', 'ท่าเรือ', 'นครหลวง', 'บางไทร', 'บางปะหัน', 'บางซ้าย', 'ผักไห่', 'ภาชี', 'ลาดบัวหลวง', 'วังน้อย', 'เสนา', 'บางปะอิน', 'อุทัย'],
        'highway': '1/32',
        'priority': 99,
        'distance_from_dc_km': 20,
        'description': 'อยุธยา-วังน้อย (DC อยู่ที่นี่!)'
    },
    # ============ โซนเพิ่มเติม - จังหวัดใกล้เคียง DC ============
    'ZONE_F4_กำแพงเพชร': {
        'provinces': ['กำแพงเพชร'],
        'highway': '1',
        'priority': 10.5,
        'distance_from_dc_km': 340,
        'description': 'กำแพงเพชร สาย 1 ภาคเหนือตอนล่าง'
    },
    'ZONE_F4_สุโขทัย': {
        'provinces': ['สุโขทัย'],
        'highway': '1',
        'priority': 11,
        'distance_from_dc_km': 400,
        'description': 'สุโขทัย สาย 1 ภาคเหนือตอนล่าง'
    },
    'ZONE_TAK_ตาก': {
        'provinces': ['ตาก'],
        'highway': '1/105',
        'priority': 10,
        'distance_from_dc_km': 420,
        'description': 'ตาก-แม่สอด สาย 1/105 เหนือตอนล่างฝั่งตะวันตก'
    },
    'ZONE_NEARBY_สิงห์บุรี': {
        'provinces': ['สิงห์บุรี'],
        'highway': '1/32',
        'priority': 99,
        'distance_from_dc_km': 100,
        'description': 'สิงห์บุรี สาย 1/32 ภาคกลางตอนบน'
    },
    'ZONE_NEARBY_อ่างทอง': {
        'provinces': ['อ่างทอง'],
        'highway': '1/32',
        'priority': 99,
        'distance_from_dc_km': 80,
        'description': 'อ่างทอง สาย 1/32 ภาคกลางตอนบน'
    },
    'ZONE_NEARBY_ชัยนาท': {
        'provinces': ['ชัยนาท'],
        'highway': '1/32',
        'priority': 99,
        'distance_from_dc_km': 150,
        'description': 'ชัยนาท สาย 1/32 ภาคกลางตอนบน'
    },
    'ZONE_NEARBY_ลพบุรี': {
        'provinces': ['ลพบุรี'],
        'highway': '1/21',
        'priority': 99,
        'distance_from_dc_km': 140,
        'description': 'ลพบุรี สาย 1/21 ภาคกลางตอนบน'
    },
    'ZONE_NEARBY_สระบุรี': {
        'provinces': ['สระบุรี'],
        'highway': '1/2',
        'priority': 99,
        'distance_from_dc_km': 80,
        'description': 'สระบุรี สาย 1/2 ประตูอีสาน-เหนือ'
    },
    # Fallback กรุงเทพ กรณีเขตไม่ match
    'ZONE_BKK_ทั่วไป': {
        'provinces': ['กรุงเทพมหานคร'],
        'highway': 'กทม',
        'priority': 99,
        'distance_from_dc_km': 50,
        'description': 'กรุงเทพทั่วไป (fallback)'
    },
    # ============ โซนปริมณฑล ============
    'ZONE_CENTRAL_นนทบุรี': {
        'provinces': ['นนทบุรี'],
        'highway': 'กทม',
        'priority': 99,
        'distance_from_dc_km': 40,
        'description': 'นนทบุรี ปริมณฑล'
    },
    'ZONE_CENTRAL_ปทุมธานี': {
        'provinces': ['ปทุมธานี'],
        'highway': 'กทม',
        'priority': 99,
        'distance_from_dc_km': 30,
        'description': 'ปทุมธานี ใกล้ DC'
    },
    'ZONE_CENTRAL_สมุทรปราการ': {
        'provinces': ['สมุทรปราการ'],
        'highway': 'กทม',
        'priority': 99,
        'distance_from_dc_km': 60,
        'description': 'สมุทรปราการ ปริมณฑล'
    },
    'ZONE_CENTRAL_นครปฐม': {
        'provinces': ['นครปฐม'],
        'highway': '35/4',
        'priority': 99,
        'distance_from_dc_km': 55,
        'description': 'นครปฐม สาย 35/4'
    },
    'ZONE_CENTRAL_สมุทรสาคร': {
        'provinces': ['สมุทรสาคร'],
        'highway': '35',
        'priority': 99,
        'distance_from_dc_km': 70,
        'description': 'สมุทรสาคร สาย 35'
    },
    'ZONE_CENTRAL_สมุทรสงคราม': {
        'provinces': ['สมุทรสงคราม'],
        'highway': '35',
        'priority': 99,
        'distance_from_dc_km': 90,
        'description': 'สมุทรสงคราม สาย 35'
    },
    # ============ โซนเพิ่มเติม - ภาคกลางตอนบน ============
    'ZONE_F4_นครสวรรค์': {
        'provinces': ['นครสวรรค์'],
        'highway': '1',
        'priority': 10,
        'distance_from_dc_km': 240,
        'description': 'นครสวรรค์ สาย 1 ภาคกลางตอนบน'
    },
    'ZONE_CENTRAL_อยุธยา': {
        'provinces': ['พระนครศรีอยุธยา'],
        'highway': '1/32',
        'priority': 99,
        'distance_from_dc_km': 25,
        'description': 'อยุธยา ใกล้ DC'
    },
    # ============ โซนเพิ่มเติม - ภาคตะวันออก ============
    'ZONE_EAST_นครนายก': {
        'provinces': ['นครนายก'],
        'highway': '305',
        'priority': 90,
        'distance_from_dc_km': 100,
        'description': 'นครนายก สาย 305'
    },
    'ZONE_EAST_ฉะเชิงเทรา': {
        'provinces': ['ฉะเชิงเทรา'],
        'highway': '304/331',  # เส้นทางเดียวกับชลบุรีเหนือ
        'priority': 17,  # อยู่ถัดชลบุรีเหนือ (สามารถรวมทริปได้ผ่าน highway 304/331)
        'distance_from_dc_km': 80,
        'description': 'ฉะเชิงเทรา สาย 304/331 (รวมทริปกับชลบุรีเหนือได้)'
    },
    'ZONE_EAST_ปราจีนบุรี': {
        'provinces': ['ปราจีนบุรี'],
        'highway': '304',
        'priority': 85,
        'distance_from_dc_km': 130,
        'description': 'ปราจีนบุรี สาย 304'
    },
    'ZONE_EAST_สระแก้ว': {
        'provinces': ['สระแก้ว'],
        'highway': '33',
        'priority': 80,
        'distance_from_dc_km': 220,
        'description': 'สระแก้ว สาย 33'
    },
    # ============ โซนเพิ่มเติม - ภาคตะวันตก ============
    'ZONE_WEST_ราชบุรี': {
        'provinces': ['ราชบุรี'],
        'highway': '4',
        'priority': 85,
        'distance_from_dc_km': 100,
        'description': 'ราชบุรี สาย 4'
    },
    'ZONE_WEST_กาญจนบุรี': {
        'provinces': ['กาญจนบุรี'],
        'highway': '323',
        'priority': 80,
        'distance_from_dc_km': 150,
        'description': 'กาญจนบุรี สาย 323'
    },
    'ZONE_WEST_สุพรรณบุรี': {
        'provinces': ['สุพรรณบุรี'],
        'highway': '340',
        'priority': 85,
        'distance_from_dc_km': 110,
        'description': 'สุพรรณบุรี สาย 340'
    },
    # ============ โซนเพิ่มเติม - ภาคเหนือ ============
    'ZONE_NORTH_พะเยา': {
        'provinces': ['พะเยา'],
        'highway': '1',
        'priority': 1,
        'distance_from_dc_km': 680,
        'description': 'พะเยา สาย 1'
    },
    'ZONE_NORTH_น่าน': {
        'provinces': ['น่าน'],
        'highway': '101',
        'priority': 2,
        'distance_from_dc_km': 620,
        'description': 'น่าน สาย 101'
    },
    'ZONE_NORTH_แพร่': {
        'provinces': ['แพร่'],
        'highway': '11',
        'priority': 3,
        'distance_from_dc_km': 540,
        'description': 'แพร่ สาย 11'
    },
    'ZONE_NORTH_อุตรดิตถ์': {
        'provinces': ['อุตรดิตถ์'],
        'highway': '11',
        'priority': 4,
        'distance_from_dc_km': 450,
        'description': 'อุตรดิตถ์ สาย 11'
    },
    'ZONE_F4_พิษณุโลก': {
        'provinces': ['พิษณุโลก'],
        'highway': '12',
        'priority': 12,
        'distance_from_dc_km': 380,
        'description': 'พิษณุโลก สาย 12'
    },
    'ZONE_F4_พิจิตร': {
        'provinces': ['พิจิตร'],
        'highway': '1',
        'priority': 11,
        'distance_from_dc_km': 330,
        'description': 'พิจิตร สาย 1'
    },
    'ZONE_F4_เพชรบูรณ์': {
        'provinces': ['เพชรบูรณ์'],
        'highway': '21',
        'priority': 15,
        'distance_from_dc_km': 350,
        'description': 'เพชรบูรณ์ สาย 21'
    },
    'ZONE_F4_ตาก': {
        'provinces': ['ตาก'],
        'highway': '1',
        'priority': 13,
        'distance_from_dc_km': 430,
        'description': 'ตาก สาย 1'
    },
    'ZONE_F4_อุทัยธานี': {
        'provinces': ['อุทัยธานี'],
        'highway': '333',
        'priority': 14,
        'distance_from_dc_km': 230,
        'description': 'อุทัยธานี สาย 333'
    },
    'ZONE_NORTH_เชียงใหม่': {
        'provinces': ['เชียงใหม่'],
        'highway': '11',
        'priority': 5,
        'distance_from_dc_km': 700,
        'description': 'เชียงใหม่ สาย 11'
    },
    'ZONE_NORTH_เชียงราย': {
        'provinces': ['เชียงราย'],
        'highway': '1',
        'priority': 3,
        'distance_from_dc_km': 780,
        'description': 'เชียงราย สาย 1'
    },
    'ZONE_NORTH_ลำพูน': {
        'provinces': ['ลำพูน'],
        'highway': '11',
        'priority': 6,
        'distance_from_dc_km': 680,
        'description': 'ลำพูน สาย 11'
    },
    'ZONE_NORTH_ลำปาง': {
        'provinces': ['ลำปาง'],
        'highway': '11',
        'priority': 7,
        'distance_from_dc_km': 600,
        'description': 'ลำปาง สาย 11'
    },
    'ZONE_NORTH_แม่ฮ่องสอน': {
        'provinces': ['แม่ฮ่องสอน'],
        'highway': '108',
        'priority': 2,
        'distance_from_dc_km': 850,
        'description': 'แม่ฮ่องสอน สาย 108 (ไกลสุด)'
    },
    # ============ โซนเพิ่มเติม - ภาคอีสาน ============
    'ZONE_ISAN_นครราชสีมา': {
        'provinces': ['นครราชสีมา'],
        'highway': '2',
        'priority': 50,
        'distance_from_dc_km': 260,
        'description': 'นครราชสีมา สาย 2 (มิตรภาพ)'
    },
    'ZONE_ISAN_ขอนแก่น': {
        'provinces': ['ขอนแก่น'],
        'highway': '2',
        'priority': 45,
        'distance_from_dc_km': 450,
        'description': 'ขอนแก่น สาย 2 (มิตรภาพ)'
    },
    'ZONE_ISAN_ชัยภูมิ': {
        'provinces': ['ชัยภูมิ'],
        'highway': '201',
        'priority': 48,
        'distance_from_dc_km': 340,
        'description': 'ชัยภูมิ สาย 201'
    },
    'ZONE_ISAN_กาฬสินธุ์': {
        'provinces': ['กาฬสินธุ์'],
        'highway': '12',
        'priority': 40,
        'distance_from_dc_km': 500,
        'description': 'กาฬสินธุ์ สาย 12'
    },
    'ZONE_ISAN_มหาสารคาม': {
        'provinces': ['มหาสารคาม'],
        'highway': '2',
        'priority': 42,
        'distance_from_dc_km': 470,
        'description': 'มหาสารคาม สาย 2'
    },
    'ZONE_ISAN_ร้อยเอ็ด': {
        'provinces': ['ร้อยเอ็ด'],
        'highway': '23',
        'priority': 38,
        'distance_from_dc_km': 520,
        'description': 'ร้อยเอ็ด สาย 23'
    },
    # ============ โซนเพิ่มเติม - ภาคใต้ตอนบน ============
    'ZONE_SOUTH_ชุมพร': {
        'provinces': ['ชุมพร'],
        'highway': '4',
        'priority': 60,
        'distance_from_dc_km': 470,
        'description': 'ชุมพร สาย 4 (ประตูใต้)'
    },
    # ============ พระนครศรีอยุธยา (sub-zones เพิ่มเติม) ============
    'ZONE_AYA3_อยุธยา_เหนือ': {
        'provinces': ['พระนครศรีอยุธยา'],
        'districts': ['บางปะหัน', 'ภาชี', 'ท่าเรือ'],
        'highway': '1',
        'priority': 87,
        'distance_from_dc_km': 55,
        'description': 'อยุธยาเหนือ'
    },
    # ============ สระบุรี (sub-zones) ============
    'ZONE_SRB1_สระบุรี_ใต้': {
        'provinces': ['สระบุรี'],
        'districts': ['หนองแค', 'บ้านหมอ', 'เสาไห้'],
        'highway': '1',
        'priority': 80,
        'distance_from_dc_km': 60,
        'description': 'สระบุรีใต้'
    },
    'ZONE_SRB2_สระบุรี_เมือง': {
        'provinces': ['สระบุรี'],
        'districts': ['เมืองสระบุรี', 'เฉลิมพระเกียรติ', 'วิหารแดง', 'พระพุทธบาท'],
        'highway': '1',
        'priority': 79,
        'distance_from_dc_km': 75,
        'description': 'สระบุรีเมือง'
    },
    'ZONE_SRB3_สระบุรี_เหนือ': {
        'provinces': ['สระบุรี'],
        'districts': ['แก่งคอย', 'มวกเหล็ก', 'วังม่วง'],
        'highway': '1',
        'priority': 78,
        'distance_from_dc_km': 90,
        'description': 'สระบุรีเหนือ'
    },
    # ============ ลพบุรี (sub-zones) ============
    'ZONE_LRI1_ลพบุรี_เมือง': {
        'provinces': ['ลพบุรี'],
        'districts': ['เมืองลพบุรี', 'ท่าวุ้ง', 'บ้านหมี่', 'โคกสำโรง'],
        'highway': '1',
        'priority': 77,
        'distance_from_dc_km': 80,
        'description': 'ลพบุรีเมือง'
    },
    'ZONE_LRI2_ลพบุรี_ใต้': {
        'provinces': ['ลพบุรี'],
        'districts': ['พัฒนานิคม', 'หนองม่วง', 'สระโบสถ์'],
        'highway': '1',
        'priority': 76,
        'distance_from_dc_km': 90,
        'description': 'ลพบุรีใต้'
    },
    'ZONE_LRI3_ลพบุรี_เหนือ': {
        'provinces': ['ลพบุรี'],
        'districts': ['ชัยบาดาล', 'ลำสนธิ', 'ท่าหลวง'],
        'highway': '1',
        'priority': 75,
        'distance_from_dc_km': 110,
        'description': 'ลพบุรีเหนือ'
    },
    # ============ ชัยนาท (sub-zones) ============
    'ZONE_CNT1_ชัยนาท_เมือง': {
        'provinces': ['ชัยนาท'],
        'districts': ['เมืองชัยนาท', 'มโนรมย์', 'สรรพยา'],
        'highway': '1',
        'priority': 78,
        'distance_from_dc_km': 85,
        'description': 'ชัยนาทเมือง'
    },
    'ZONE_CNT2_ชัยนาท_เหนือ': {
        'provinces': ['ชัยนาท'],
        'districts': ['สรรคบุรี', 'หันคา', 'วัดสิงห์', 'เนินขาม'],
        'highway': '1',
        'priority': 77,
        'distance_from_dc_km': 95,
        'description': 'ชัยนาทเหนือ'
    },
    # ============ สิงห์บุรี ============
    'ZONE_SBR1_สิงห์บุรี': {
        'provinces': ['สิงห์บุรี'],
        'districts': ['เมืองสิงห์บุรี', 'อินทร์บุรี', 'พรหมบุรี', 'บางระจัน', 'ท่าช้าง', 'ค่ายบางระจัน'],
        'highway': '1',
        'priority': 79,
        'distance_from_dc_km': 90,
        'description': 'สิงห์บุรีทุกอำเภอ'
    },
    # ============ อ่างทอง ============
    'ZONE_ATG1_อ่างทอง': {
        'provinces': ['อ่างทอง'],
        'districts': ['เมืองอ่างทอง', 'วิเศษชัยชาญ', 'ไชโย', 'ป่าโมก', 'โพธิ์ทอง', 'สามโก้'],
        'highway': '1',
        'priority': 80,
        'distance_from_dc_km': 80,
        'description': 'อ่างทองทุกอำเภอ'
    },
    # ============ นครนายก ============
    'ZONE_NNK1_นครนายก': {
        'provinces': ['นครนายก'],
        'districts': ['เมืองนครนายก', 'บ้านนา', 'องครักษ์', 'ปากพลี'],
        'highway': '33',
        'priority': 72,
        'distance_from_dc_km': 85,
        'description': 'นครนายกทุกอำเภอ'
    },
    # ============ ปราจีนบุรี (sub-zones) ============
    'ZONE_PRB1_ปราจีนบุรี_เมือง': {
        'provinces': ['ปราจีนบุรี'],
        'districts': ['เมืองปราจีนบุรี', 'บ้านสร้าง', 'ศรีมโหสถ'],
        'highway': '33',
        'priority': 71,
        'distance_from_dc_km': 90,
        'description': 'ปราจีนบุรีเมือง'
    },
    'ZONE_PRB2_ปราจีนบุรี_ตะวันออก': {
        'provinces': ['ปราจีนบุรี'],
        'districts': ['กบินทร์บุรี', 'ศรีมหาโพธิ', 'ประจันตคาม', 'นาดี'],
        'highway': '33',
        'priority': 70,
        'distance_from_dc_km': 110,
        'description': 'ปราจีนบุรีตะวันออก'
    },
    # ============ สระแก้ว (sub-zones) ============
    'ZONE_SKW1_สระแก้ว_เมือง': {
        'provinces': ['สระแก้ว'],
        'districts': ['เมืองสระแก้ว', 'วังน้ำเย็น', 'เขาฉกรรจ์', 'วังสมบูรณ์'],
        'highway': '33',
        'priority': 68,
        'distance_from_dc_km': 130,
        'description': 'สระแก้วเมือง'
    },
    'ZONE_SKW2_สระแก้ว_ตะวันออก': {
        'provinces': ['สระแก้ว'],
        'districts': ['วัฒนานคร', 'อรัญประเทศ'],
        'highway': '33',
        'priority': 67,
        'distance_from_dc_km': 150,
        'description': 'สระแก้วตะวันออก ชายแดน'
    },
    # ============ ราชบุรี (sub-zones) ============
    'ZONE_RBR1_ราชบุรี_เมือง': {
        'provinces': ['ราชบุรี'],
        'districts': ['เมืองราชบุรี', 'ดำเนินสะดวก', 'บ้านคา'],
        'highway': '4',
        'priority': 73,
        'distance_from_dc_km': 90,
        'description': 'ราชบุรีเมือง'
    },
    'ZONE_RBR2_ราชบุรี_เหนือ': {
        'provinces': ['ราชบุรี'],
        'districts': ['บ้านโป่ง', 'โพธาราม', 'สวนผึ้ง', 'จอมบึง'],
        'highway': '4',
        'priority': 72,
        'distance_from_dc_km': 100,
        'description': 'ราชบุรีเหนือ'
    },
    'ZONE_RBR3_ราชบุรี_ใต้': {
        'provinces': ['ราชบุรี'],
        'districts': ['ปากท่อ', 'บางแพ'],
        'highway': '4',
        'priority': 71,
        'distance_from_dc_km': 110,
        'description': 'ราชบุรีใต้'
    },
    # ============ สุพรรณบุรี (sub-zones) ============
    'ZONE_SPB1_สุพรรณบุรี_เมือง': {
        'provinces': ['สุพรรณบุรี'],
        'districts': ['เมืองสุพรรณบุรี', 'บางปลาม้า', 'ศรีประจันต์', 'ดอนเจดีย์'],
        'highway': '340',
        'priority': 74,
        'distance_from_dc_km': 100,
        'description': 'สุพรรณบุรีเมือง'
    },
    'ZONE_SPB2_สุพรรณบุรี_ใต้': {
        'provinces': ['สุพรรณบุรี'],
        'districts': ['สองพี่น้อง', 'สามชุก', 'อู่ทอง', 'เดิมบางนางบวช'],
        'highway': '340',
        'priority': 73,
        'distance_from_dc_km': 115,
        'description': 'สุพรรณบุรีใต้'
    },
    'ZONE_SPB3_สุพรรณบุรี_เหนือ': {
        'provinces': ['สุพรรณบุรี'],
        'districts': ['ด่านช้าง', 'หนองหญ้าไซ'],
        'highway': '340',
        'priority': 72,
        'distance_from_dc_km': 130,
        'description': 'สุพรรณบุรีเหนือ'
    },
    # ============ กาญจนบุรี (sub-zones) ============
    'ZONE_KBR1_กาญจนบุรี_ใต้': {
        'provinces': ['กาญจนบุรี'],
        'districts': ['เมืองกาญจนบุรี', 'ท่าม่วง', 'พนมทวน'],
        'highway': '323',
        'priority': 69,
        'distance_from_dc_km': 120,
        'description': 'กาญจนบุรีใต้'
    },
    'ZONE_KBR2_กาญจนบุรี_ตะวันออก': {
        'provinces': ['กาญจนบุรี'],
        'districts': ['ท่ามะกา', 'บ่อพลอย', 'เลาขวัญ'],
        'highway': '323',
        'priority': 68,
        'distance_from_dc_km': 130,
        'description': 'กาญจนบุรีตะวันออก'
    },
    'ZONE_KBR3_กาญจนบุรี_เหนือ': {
        'provinces': ['กาญจนบุรี'],
        'districts': ['ไทรโยค', 'สังขละบุรี'],
        'highway': '323',
        'priority': 67,
        'distance_from_dc_km': 160,
        'description': 'กาญจนบุรีเหนือ ชายแดน'
    },
    # ============ สมุทรสงคราม ============
    'ZONE_SSG1_สมุทรสงคราม': {
        'provinces': ['สมุทรสงคราม'],
        'districts': ['เมืองสมุทรสงคราม', 'อัมพวา', 'บางคนที'],
        'highway': '35',
        'priority': 77,
        'distance_from_dc_km': 85,
        'description': 'สมุทรสงครามทุกอำเภอ'
    },
    # ============ เพชรบุรี (sub-zones) ============
    'ZONE_PBI1_เพชรบุรี_เมือง': {
        'provinces': ['เพชรบุรี'],
        'districts': ['เมืองเพชรบุรี', 'เขาย้อย', 'บ้านลาด', 'บ้านแหลม'],
        'highway': '4',
        'priority': 66,
        'distance_from_dc_km': 140,
        'description': 'เพชรบุรีเมือง'
    },
    'ZONE_PBI2_เพชรบุรี_ชายฝั่ง': {
        'provinces': ['เพชรบุรี'],
        'districts': ['ชะอำ', 'ท่ายาง', 'แก่งกระจาน'],
        'highway': '4',
        'priority': 65,
        'distance_from_dc_km': 155,
        'description': 'เพชรบุรีชายฝั่ง'
    },
    # ============ อุทัยธานี ============
    'ZONE_UTN1_อุทัยธานี': {
        'provinces': ['อุทัยธานี'],
        'districts': ['เมืองอุทัยธานี', 'หนองฉาง', 'ทัพทัน', 'ลานสัก', 'บ้านไร่', 'สว่างอารมณ์', 'หนองขาหย่าง'],
        'highway': '333',
        'priority': 55,
        'distance_from_dc_km': 180,
        'description': 'อุทัยธานีทุกอำเภอ'
    },
    # ============ นครสวรรค์ (sub-zones) ============
    'ZONE_NSW1_นครสวรรค์_เมือง': {
        'provinces': ['นครสวรรค์'],
        'districts': ['เมืองนครสวรรค์', 'โกรกพระ', 'พยุหะคีรี', 'เก้าเลี้ยว'],
        'highway': '1',
        'priority': 50,
        'distance_from_dc_km': 230,
        'description': 'นครสวรรค์เมือง'
    },
    'ZONE_NSW2_นครสวรรค์_ตะวันออก': {
        'provinces': ['นครสวรรค์'],
        'districts': ['ตาคลี', 'ตากฟ้า', 'ไพศาลี', 'หนองบัว'],
        'highway': '1',
        'priority': 49,
        'distance_from_dc_km': 250,
        'description': 'นครสวรรค์ตะวันออก'
    },
    'ZONE_NSW3_นครสวรรค์_เหนือ': {
        'provinces': ['นครสวรรค์'],
        'districts': ['บรรพตพิสัย', 'ลาดยาว', 'ชุมแสง', 'ชุมตาบง', 'แม่วงก์', 'แม่เปิน'],
        'highway': '1',
        'priority': 48,
        'distance_from_dc_km': 260,
        'description': 'นครสวรรค์เหนือ'
    },
    # ============ กำแพงเพชร (sub-zones) ============
    'ZONE_KPT1_กำแพงเพชร_เมือง': {
        'provinces': ['กำแพงเพชร'],
        'districts': ['เมืองกำแพงเพชร', 'ลานกระบือ', 'พรานกระต่าย', 'ไทรงาม'],
        'highway': '1',
        'priority': 40,
        'distance_from_dc_km': 290,
        'description': 'กำแพงเพชรเมือง'
    },
    'ZONE_KPT2_กำแพงเพชร_ตะวันตก': {
        'provinces': ['กำแพงเพชร'],
        'districts': ['คลองขลุง', 'ขาณุวรลักษบุรี', 'บึงสามัคคี', 'คลองลาน'],
        'highway': '1',
        'priority': 39,
        'distance_from_dc_km': 310,
        'description': 'กำแพงเพชรตะวันตก'
    },
    # ============ ตาก (sub-zones) ============
    'ZONE_TAK1_ตาก_เมือง': {
        'provinces': ['ตาก'],
        'districts': ['เมืองตาก', 'บ้านตาก', 'สามเงา'],
        'highway': '1',
        'priority': 38,
        'distance_from_dc_km': 330,
        'description': 'ตากเมือง'
    },
    'ZONE_TAK2_ตาก_แม่สอด': {
        'provinces': ['ตาก'],
        'districts': ['แม่สอด', 'แม่ระมาด', 'พบพระ'],
        'highway': '105',
        'priority': 37,
        'distance_from_dc_km': 380,
        'description': 'ตากแม่สอด ชายแดนพม่า'
    },
    # ============ สุโขทัย (sub-zones) ============
    'ZONE_STT1_สุโขทัย_ใต้': {
        'provinces': ['สุโขทัย'],
        'districts': ['เมืองสุโขทัย', 'กงไกรลาศ', 'ศรีสำโรง', 'บ้านด่านลานหอย', 'คีรีมาศ'],
        'highway': '101',
        'priority': 41,
        'distance_from_dc_km': 380,
        'description': 'สุโขทัยใต้'
    },
    'ZONE_STT2_สุโขทัย_เหนือ': {
        'provinces': ['สุโขทัย'],
        'districts': ['สวรรคโลก', 'ศรีสัชนาลัย', 'ทุ่งเสลี่ยม', 'ศรีนคร'],
        'highway': '101',
        'priority': 40,
        'distance_from_dc_km': 400,
        'description': 'สุโขทัยเหนือ'
    },
    # ============ ลำพูน (sub-zones) ============
    'ZONE_LPN1_ลำพูน_เมือง': {
        'provinces': ['ลำพูน'],
        'districts': ['เมืองลำพูน', 'ป่าซาง', 'บ้านธิ', 'เวียงหนองล่อง'],
        'highway': '11',
        'priority': 12,
        'distance_from_dc_km': 640,
        'description': 'ลำพูนเมือง'
    },
    'ZONE_LPN2_ลำพูน_ใต้': {
        'provinces': ['ลำพูน'],
        'districts': ['ลี้', 'บ้านโฮ่ง', 'ทุ่งหัวช้าง'],
        'highway': '106',
        'priority': 11,
        'distance_from_dc_km': 680,
        'description': 'ลำพูนใต้'
    },
    # ============ ลำปาง (sub-zones) ============
    'ZONE_LPG1_ลำปาง_เมือง': {
        'provinces': ['ลำปาง'],
        'districts': ['เมืองลำปาง', 'ห้างฉัตร', 'เกาะคา', 'แม่ทะ'],
        'highway': '11',
        'priority': 13,
        'distance_from_dc_km': 580,
        'description': 'ลำปางเมือง'
    },
    'ZONE_LPG2_ลำปาง_เหนือ': {
        'provinces': ['ลำปาง'],
        'districts': ['งาว', 'แจ้ห่ม', 'วังเหนือ'],
        'highway': '11',
        'priority': 12,
        'distance_from_dc_km': 620,
        'description': 'ลำปางเหนือ'
    },
    'ZONE_LPG3_ลำปาง_ใต้': {
        'provinces': ['ลำปาง'],
        'districts': ['เถิน', 'สบปราบ', 'แม่พริก'],
        'highway': '11',
        'priority': 11,
        'distance_from_dc_km': 600,
        'description': 'ลำปางใต้'
    },
    # ============ แพร่ (sub-zones) ============
    'ZONE_PRE1_แพร่_เมือง': {
        'provinces': ['แพร่'],
        'districts': ['เมืองแพร่', 'สูงเม่น', 'ร้องกวาง'],
        'highway': '11',
        'priority': 14,
        'distance_from_dc_km': 540,
        'description': 'แพร่เมือง'
    },
    'ZONE_PRE2_แพร่_เหนือ': {
        'provinces': ['แพร่'],
        'districts': ['สอง', 'เด่นชัย', 'วังชิ้น', 'ลอง', 'หนองม่วงไข่'],
        'highway': '11',
        'priority': 13,
        'distance_from_dc_km': 560,
        'description': 'แพร่เหนือ'
    },
    # ============ น่าน ============
    'ZONE_NAN1_น่าน': {
        'provinces': ['น่าน'],
        'districts': ['เมืองน่าน', 'เวียงสา', 'ท่าวังผา', 'ภูเพียง', 'สันติสุข', 'เชียงกลาง'],
        'highway': '101',
        'priority': 15,
        'distance_from_dc_km': 620,
        'description': 'น่านทุกอำเภอ'
    },
    # ============ พะเยา (sub-zones) ============
    'ZONE_PYO1_พะเยา_เมือง': {
        'provinces': ['พะเยา'],
        'districts': ['เมืองพะเยา', 'แม่ใจ', 'ดอกคำใต้', 'ภูกามยาว'],
        'highway': '1',
        'priority': 16,
        'distance_from_dc_km': 680,
        'description': 'พะเยาเมือง'
    },
    'ZONE_PYO2_พะเยา_เหนือ': {
        'provinces': ['พะเยา'],
        'districts': ['เชียงคำ', 'ภูซาง', 'เชียงม่วน', 'ปง', 'จุน'],
        'highway': '1',
        'priority': 15,
        'distance_from_dc_km': 720,
        'description': 'พะเยาเหนือ'
    },
    # ============ ชัยภูมิ (sub-zones) ============
    'ZONE_CPM1_ชัยภูมิ_เมือง': {
        'provinces': ['ชัยภูมิ'],
        'districts': ['เมืองชัยภูมิ', 'บ้านเขว้า', 'คอนสวรรค์'],
        'highway': '2',
        'priority': 32,
        'distance_from_dc_km': 270,
        'description': 'ชัยภูมิเมือง'
    },
    'ZONE_CPM2_ชัยภูมิ_เหนือ': {
        'provinces': ['ชัยภูมิ'],
        'districts': ['ภูเขียว', 'แก้งคร้อ', 'หนองบัวแดง', 'เกษตรสมบูรณ์', 'บ้านแท่น', 'ภักดีชุมพล'],
        'highway': '2',
        'priority': 31,
        'distance_from_dc_km': 290,
        'description': 'ชัยภูมิเหนือ'
    },
    'ZONE_CPM3_ชัยภูมิ_ใต้': {
        'provinces': ['ชัยภูมิ'],
        'districts': ['จัตุรัส', 'บำเหน็จณรงค์', 'เทพสถิต'],
        'highway': '2',
        'priority': 30,
        'distance_from_dc_km': 300,
        'description': 'ชัยภูมิใต้'
    },
    # ============ สุรินทร์ (sub-zones) ============
    'ZONE_SRN1_สุรินทร์_เมือง': {
        'provinces': ['สุรินทร์'],
        'districts': ['เมืองสุรินทร์', 'ศีขรภูมิ', 'สำโรงทาบ'],
        'highway': '226',
        'priority': 26,
        'distance_from_dc_km': 400,
        'description': 'สุรินทร์เมือง'
    },
    'ZONE_SRN2_สุรินทร์_ใต้': {
        'provinces': ['สุรินทร์'],
        'districts': ['ปราสาท', 'สังขะ', 'กาบเชิง'],
        'highway': '226',
        'priority': 25,
        'distance_from_dc_km': 420,
        'description': 'สุรินทร์ใต้'
    },
    'ZONE_SRN3_สุรินทร์_เหนือ': {
        'provinces': ['สุรินทร์'],
        'districts': ['ท่าตูม', 'ชุมพลบุรี', 'รัตนบุรี', 'สนม', 'จอมพระ', 'ลำดวน'],
        'highway': '226',
        'priority': 24,
        'distance_from_dc_km': 410,
        'description': 'สุรินทร์เหนือ'
    },
    # ============ มหาสารคาม (sub-zones) ============
    'ZONE_MSK1_มหาสารคาม_เมือง': {
        'provinces': ['มหาสารคาม'],
        'districts': ['เมืองมหาสารคาม', 'กันทรวิชัย', 'โกสุมพิสัย'],
        'highway': '2',
        'priority': 30,
        'distance_from_dc_km': 440,
        'description': 'มหาสารคามเมือง'
    },
    'ZONE_MSK2_มหาสารคาม_ใต้': {
        'provinces': ['มหาสารคาม'],
        'districts': ['พยัคฆภูมิพิสัย', 'วาปีปทุม', 'นาดูน', 'นาเชือก', 'ยางสีสุราช'],
        'highway': '2',
        'priority': 29,
        'distance_from_dc_km': 460,
        'description': 'มหาสารคามใต้'
    },
    'ZONE_MSK3_มหาสารคาม_เหนือ': {
        'provinces': ['มหาสารคาม'],
        'districts': ['เชียงยืน', 'บรบือ', 'แกดำ', 'กุดรัง'],
        'highway': '2',
        'priority': 29,
        'distance_from_dc_km': 450,
        'description': 'มหาสารคามเหนือ'
    },
    # ============ ร้อยเอ็ด (sub-zones) ============
    'ZONE_ROI1_ร้อยเอ็ด_เมือง': {
        'provinces': ['ร้อยเอ็ด'],
        'districts': ['เมืองร้อยเอ็ด', 'ธวัชบุรี', 'จังหาร', 'จตุรพักตรพิมาน'],
        'highway': '23',
        'priority': 28,
        'distance_from_dc_km': 490,
        'description': 'ร้อยเอ็ดเมือง'
    },
    'ZONE_ROI2_ร้อยเอ็ด_ตะวันตก': {
        'provinces': ['ร้อยเอ็ด'],
        'districts': ['พนมไพร', 'เกษตรวิสัย', 'ปทุมรัตต์', 'โพนทอง'],
        'highway': '23',
        'priority': 27,
        'distance_from_dc_km': 510,
        'description': 'ร้อยเอ็ดตะวันตก'
    },
    'ZONE_ROI3_ร้อยเอ็ด_ตะวันออก': {
        'provinces': ['ร้อยเอ็ด'],
        'districts': ['สุวรรณภูมิ', 'เสลภูมิ', 'อาจสามารถ', 'โพธิ์ชัย', 'ทุ่งเขาหลวง'],
        'highway': '23',
        'priority': 26,
        'distance_from_dc_km': 520,
        'description': 'ร้อยเอ็ดตะวันออก'
    },
    # ============ ยโสธร (sub-zones) ============
    'ZONE_YST1_ยโสธร_เมือง': {
        'provinces': ['ยโสธร'],
        'districts': ['เมืองยโสธร', 'คำเขื่อนแก้ว', 'มหาชนะชัย', 'ค้อวัง'],
        'highway': '23',
        'priority': 24,
        'distance_from_dc_km': 510,
        'description': 'ยโสธรเมือง'
    },
    'ZONE_YST2_ยโสธร_เหนือ': {
        'provinces': ['ยโสธร'],
        'districts': ['เลิงนกทา', 'กุดชุม', 'ไทยเจริญ', 'ป่าติ้ว'],
        'highway': '23',
        'priority': 23,
        'distance_from_dc_km': 530,
        'description': 'ยโสธรเหนือ'
    },
    # ============ ศรีสะเกษ (sub-zones) ============
    'ZONE_SSK1_ศรีสะเกษ_เมือง': {
        'provinces': ['ศรีสะเกษ'],
        'districts': ['เมืองศรีสะเกษ', 'กันทรารมย์', 'ยางชุมน้อย', 'ราษีไศล', 'โนนคูณ'],
        'highway': '226',
        'priority': 24,
        'distance_from_dc_km': 480,
        'description': 'ศรีสะเกษเมือง'
    },
    'ZONE_SSK2_ศรีสะเกษ_ใต้': {
        'provinces': ['ศรีสะเกษ'],
        'districts': ['ขุขันธ์', 'ปรางค์กู่', 'ไพรบึง', 'ภูสิงห์'],
        'highway': '226',
        'priority': 23,
        'distance_from_dc_km': 500,
        'description': 'ศรีสะเกษใต้'
    },
    'ZONE_SSK3_ศรีสะเกษ_เหนือ': {
        'provinces': ['ศรีสะเกษ'],
        'districts': ['กันทรลักษ์', 'ขุนหาญ', 'อุทุมพรพิสัย', 'ห้วยทับทัน', 'ศรีรัตนะ'],
        'highway': '226',
        'priority': 22,
        'distance_from_dc_km': 510,
        'description': 'ศรีสะเกษเหนือ'
    },
    # ============ สกลนคร (sub-zones) ============
    'ZONE_SNK1_สกลนคร_เมือง': {
        'provinces': ['สกลนคร'],
        'districts': ['เมืองสกลนคร', 'โพนนาแก้ว', 'กุสุมาลย์', 'โคกศรีสุพรรณ'],
        'highway': '22',
        'priority': 20,
        'distance_from_dc_km': 550,
        'description': 'สกลนครเมือง'
    },
    'ZONE_SNK2_สกลนคร_รอบนอก': {
        'provinces': ['สกลนคร'],
        'districts': ['สว่างแดนดิน', 'วานรนิวาส', 'พรรณานิคม', 'วาริชภูมิ', 'บ้านม่วง', 'คำตากล้า'],
        'highway': '22',
        'priority': 19,
        'distance_from_dc_km': 580,
        'description': 'สกลนครรอบนอก'
    },
    # ============ อำนาจเจริญ ============
    'ZONE_AJN1_อำนาจเจริญ': {
        'provinces': ['อำนาจเจริญ'],
        'districts': ['เมืองอำนาจเจริญ', 'หัวตะพาน', 'ชานุมาน', 'พนา', 'เสนางคนิคม', 'ลืออำนาจ'],
        'highway': '212',
        'priority': 22,
        'distance_from_dc_km': 520,
        'description': 'อำนาจเจริญทุกอำเภอ'
    },
    # ============ มุกดาหาร ============
    'ZONE_MKH1_มุกดาหาร': {
        'provinces': ['มุกดาหาร'],
        'districts': ['เมืองมุกดาหาร', 'คำชะอี', 'ดงหลวง', 'หนองสูง', 'หว้านใหญ่', 'ดอนตาล'],
        'highway': '212',
        'priority': 20,
        'distance_from_dc_km': 540,
        'description': 'มุกดาหารทุกอำเภอ'
    },
    # ============ นครพนม ============
    'ZONE_NPN1_นครพนม': {
        'provinces': ['นครพนม'],
        'districts': ['เมืองนครพนม', 'ธาตุพนม', 'ท่าอุเทน', 'ปลาปาก', 'นาแก', 'ศรีสงคราม', 'นาหว้า', 'วังยาง', 'บ้านแพง', 'นาทม'],
        'highway': '22',
        'priority': 19,
        'distance_from_dc_km': 570,
        'description': 'นครพนมทุกอำเภอ'
    },
    # ============ บึงกาฬ (sub-zones) ============
    'ZONE_BKN1_บึงกาฬ_เมือง': {
        'provinces': ['บึงกาฬ'],
        'districts': ['เมืองบึงกาฬ', 'ปากคาด', 'พรเจริญ', 'บุ่งคล้า'],
        'highway': '212',
        'priority': 18,
        'distance_from_dc_km': 590,
        'description': 'บึงกาฬเมือง'
    },
    'ZONE_BKN2_บึงกาฬ_ตะวันออก': {
        'provinces': ['บึงกาฬ'],
        'districts': ['เซกา', 'บึงโขงหลง', 'โซ่พิสัย', 'ศรีวิไล'],
        'highway': '212',
        'priority': 17,
        'distance_from_dc_km': 610,
        'description': 'บึงกาฬตะวันออก'
    },
    # ============ หนองคาย ============
    'ZONE_NKI1_หนองคาย': {
        'provinces': ['หนองคาย'],
        'districts': ['เมืองหนองคาย', 'ศรีเชียงใหม่', 'ท่าบ่อ', 'โพนพิสัย', 'สังคม', 'เฝ้าไร่', 'โพธิ์ตาก', 'รัตนวาปี'],
        'highway': '2',
        'priority': 18,
        'distance_from_dc_km': 600,
        'description': 'หนองคายทุกอำเภอ'
    },
    # ============ หนองบัวลำภู ============
    'ZONE_NBL1_หนองบัวลำภู': {
        'provinces': ['หนองบัวลำภู'],
        'districts': ['เมืองหนองบัวลำภู', 'โนนสัง', 'ศรีบุญเรือง', 'นากลาง', 'สุวรรณคูหา', 'นาวัง'],
        'highway': '210',
        'priority': 22,
        'distance_from_dc_km': 490,
        'description': 'หนองบัวลำภูทุกอำเภอ'
    },
    # ============ เลย ============
    'ZONE_LEI1_เลย': {
        'provinces': ['เลย'],
        'districts': ['เมืองเลย', 'นาด้วง', 'ท่าลี่', 'เชียงคาน', 'ปากชม', 'ด่านซ้าย', 'วังสะพุง', 'ภูกระดึง', 'ภูเรือ', 'ภูหลวง'],
        'highway': '21',
        'priority': 21,
        'distance_from_dc_km': 520,
        'description': 'เลยทุกอำเภอ'
    },
    # ============ กาฬสินธุ์ (sub-zones) ============
    'ZONE_KSN1_กาฬสินธุ์_ใต้': {
        'provinces': ['กาฬสินธุ์'],
        'districts': ['เมืองกาฬสินธุ์', 'กมลาไสย', 'ร่องคำ', 'ฆ้องชัย', 'ยางตลาด'],
        'highway': '12',
        'priority': 27,
        'distance_from_dc_km': 510,
        'description': 'กาฬสินธุ์ใต้'
    },
    'ZONE_KSN2_กาฬสินธุ์_เหนือ': {
        'provinces': ['กาฬสินธุ์'],
        'districts': ['สมเด็จ', 'สหัสขันธ์', 'คำม่วง', 'กุฉินารายณ์', 'เขาวง', 'นาคู', 'สามชัย'],
        'highway': '12',
        'priority': 26,
        'distance_from_dc_km': 530,
        'description': 'กาฬสินธุ์เหนือ'
    },
    'ZONE_KSN3_กาฬสินธุ์_ตะวันตก': {
        'provinces': ['กาฬสินธุ์'],
        'districts': ['ท่าคันโท', 'ห้วยเม็ก', 'ห้วยผึ้ง', 'นามน', 'ดอนจาน'],
        'highway': '12',
        'priority': 25,
        'distance_from_dc_km': 520,
        'description': 'กาฬสินธุ์ตะวันตก'
    },
    # ============ จันทบุรี (sub-zones) ============
    'ZONE_CTB1_จันทบุรี_เมือง': {
        'provinces': ['จันทบุรี'],
        'districts': ['เมืองจันทบุรี', 'ท่าใหม่', 'มะขาม', 'แหลมสิงห์', 'นายายอาม'],
        'highway': '3',
        'priority': 55,
        'distance_from_dc_km': 280,
        'description': 'จันทบุรีเมือง'
    },
    'ZONE_CTB2_จันทบุรี_ตะวันออก': {
        'provinces': ['จันทบุรี'],
        'districts': ['สอยดาว', 'ขลุง', 'โป่งน้ำร้อน', 'เขาคิชฌกูฏ'],
        'highway': '317',
        'priority': 54,
        'distance_from_dc_km': 310,
        'description': 'จันทบุรีตะวันออก'
    },
    # ============ ตราด ============
    'ZONE_TRT1_ตราด': {
        'provinces': ['ตราด'],
        'districts': ['เมืองตราด', 'คลองใหญ่', 'เขาสมิง', 'บ่อไร่', 'แหลมงอบ'],
        'highway': '3',
        'priority': 53,
        'distance_from_dc_km': 350,
        'description': 'ตราดทุกอำเภอ'
    },
    # ============ ประจวบคีรีขันธ์ (sub-zones) ============
    'ZONE_PKK1_ประจวบ_เหนือ': {
        'provinces': ['ประจวบคีรีขันธ์'],
        'districts': ['หัวหิน', 'ปราณบุรี', 'สามร้อยยอด', 'กุยบุรี'],
        'highway': '4',
        'priority': 63,
        'distance_from_dc_km': 200,
        'description': 'ประจวบเหนือ หัวหิน'
    },
    'ZONE_PKK2_ประจวบ_เมือง': {
        'provinces': ['ประจวบคีรีขันธ์'],
        'districts': ['เมืองประจวบคีรีขันธ์'],
        'highway': '4',
        'priority': 62,
        'distance_from_dc_km': 230,
        'description': 'ประจวบเมือง'
    },
    'ZONE_PKK3_ประจวบ_ใต้': {
        'provinces': ['ประจวบคีรีขันธ์'],
        'districts': ['ทับสะแก', 'บางสะพานน้อย', 'บางสะพาน'],
        'highway': '4',
        'priority': 61,
        'distance_from_dc_km': 260,
        'description': 'ประจวบใต้'
    },
    # ============ สุราษฎร์ธานี (sub-zones) ============
    'ZONE_STN1_สุราษฎร์_เมือง': {
        'provinces': ['สุราษฎร์ธานี'],
        'districts': ['เมืองสุราษฎร์ธานี', 'พุนพิน', 'กาญจนดิษฐ์'],
        'highway': '4',
        'priority': 56,
        'distance_from_dc_km': 500,
        'description': 'สุราษฎร์ธานีเมือง'
    },
    'ZONE_STN2_สุราษฎร์_เหนือ': {
        'provinces': ['สุราษฎร์ธานี'],
        'districts': ['ไชยา', 'ท่าชนะ', 'ท่าฉาง', 'คีรีรัฐนิคม'],
        'highway': '4',
        'priority': 55,
        'distance_from_dc_km': 480,
        'description': 'สุราษฎร์ธานีเหนือ'
    },
    'ZONE_STN3_สุราษฎร์_ใต้': {
        'provinces': ['สุราษฎร์ธานี'],
        'districts': ['เวียงสระ', 'บ้านนาสาร', 'บ้านนาเดิม', 'เคียนซา', 'บ้านตาขุน'],
        'highway': '4',
        'priority': 54,
        'distance_from_dc_km': 520,
        'description': 'สุราษฎร์ธานีใต้'
    },
    'ZONE_STN4_เกาะสมุย': {
        'provinces': ['สุราษฎร์ธานี'],
        'districts': ['เกาะสมุย', 'เกาะพงัน'],
        'highway': '4',
        'priority': 53,
        'distance_from_dc_km': 550,
        'description': 'เกาะสมุย เกาะพงัน'
    },
    # ============ นครศรีธรรมราช (sub-zones) ============
    'ZONE_NRT1_นครศรีฯ_เมือง': {
        'provinces': ['นครศรีธรรมราช'],
        'districts': ['เมืองนครศรีธรรมราช', 'พรหมคีรี', 'ท่าศาลา', 'นบพิตำ'],
        'highway': '4',
        'priority': 52,
        'distance_from_dc_km': 560,
        'description': 'นครศรีธรรมราชเมือง'
    },
    'ZONE_NRT2_นครศรีฯ_เหนือ': {
        'provinces': ['นครศรีธรรมราช'],
        'districts': ['ทุ่งสง', 'ฉวาง', 'พิปูน', 'ช้างกลาง', 'นาบอน'],
        'highway': '4',
        'priority': 51,
        'distance_from_dc_km': 580,
        'description': 'นครศรีธรรมราชเหนือ'
    },
    'ZONE_NRT3_นครศรีฯ_ใต้': {
        'provinces': ['นครศรีธรรมราช'],
        'districts': ['ชะอวด', 'ร่อนพิบูลย์', 'หัวไทร', 'เฉลิมพระเกียรติ', 'จุฬาภรณ์', 'บางขัน'],
        'highway': '4',
        'priority': 50,
        'distance_from_dc_km': 600,
        'description': 'นครศรีธรรมราชใต้'
    },
    # ============ สงขลา (sub-zones) ============
    'ZONE_SKL1_สงขลา_หาดใหญ่': {
        'provinces': ['สงขลา'],
        'districts': ['หาดใหญ่', 'สะเดา', 'คลองหอยโข่ง', 'นาหม่อม', 'บางกล่ำ'],
        'highway': '4',
        'priority': 47,
        'distance_from_dc_km': 680,
        'description': 'สงขลาหาดใหญ่'
    },
    'ZONE_SKL2_สงขลา_เมือง': {
        'provinces': ['สงขลา'],
        'districts': ['เมืองสงขลา', 'สิงหนคร', 'ควนเนียง'],
        'highway': '4',
        'priority': 46,
        'distance_from_dc_km': 700,
        'description': 'สงขลาเมือง'
    },
    'ZONE_SKL3_สงขลา_เหนือ': {
        'provinces': ['สงขลา'],
        'districts': ['ระโนด', 'สทิงพระ', 'กระแสสินธุ์', 'รัตภูมิ', 'นาทวี', 'เทพา', 'จะนะ'],
        'highway': '4',
        'priority': 45,
        'distance_from_dc_km': 660,
        'description': 'สงขลาเหนือ'
    },
    # ============ พัทลุง (sub-zones) ============
    'ZONE_PTL1_พัทลุง_เมือง': {
        'provinces': ['พัทลุง'],
        'districts': ['เมืองพัทลุง', 'ควนขนุน', 'เขาชัยสน', 'บางแก้ว', 'ศรีนครินทร์'],
        'highway': '4',
        'priority': 48,
        'distance_from_dc_km': 640,
        'description': 'พัทลุงเมือง'
    },
    'ZONE_PTL2_พัทลุง_รอบนอก': {
        'provinces': ['พัทลุง'],
        'districts': ['ป่าบอน', 'ตะโหมด', 'ป่าพะยอม', 'ปากพะยูน', 'กงหรา'],
        'highway': '4',
        'priority': 47,
        'distance_from_dc_km': 660,
        'description': 'พัทลุงรอบนอก'
    },
    # ============ กระบี่ (sub-zones) ============
    'ZONE_KBI1_กระบี่_เมือง': {
        'provinces': ['กระบี่'],
        'districts': ['เมืองกระบี่', 'อ่าวลึก', 'ปลายพระยา'],
        'highway': '4',
        'priority': 44,
        'distance_from_dc_km': 750,
        'description': 'กระบี่เมือง'
    },
    'ZONE_KBI2_กระบี่_เกาะ': {
        'provinces': ['กระบี่'],
        'districts': ['เกาะลันตา', 'คลองท่อม', 'ลำทับ'],
        'highway': '4',
        'priority': 43,
        'distance_from_dc_km': 780,
        'description': 'กระบี่เกาะลันตา'
    },
    # ============ ตรัง ============
    'ZONE_TRG1_ตรัง': {
        'provinces': ['ตรัง'],
        'districts': ['เมืองตรัง', 'กันตัง', 'ปะเหลียน', 'ห้วยยอด', 'วังวิเศษ', 'รัษฎา', 'สิเกา', 'ย่านตาขาว', 'นาโยง', 'หาดสำราญ'],
        'highway': '4',
        'priority': 46,
        'distance_from_dc_km': 720,
        'description': 'ตรังทุกอำเภอ'
    },
    # ============ พังงา ============
    'ZONE_PNG1_พังงา': {
        'provinces': ['พังงา'],
        'districts': ['เมืองพังงา', 'ตะกั่วทุ่ง', 'ตะกั่วป่า', 'คุระบุรี', 'กะปง', 'ทับปุด', 'ท้ายเหมือง'],
        'highway': '4',
        'priority': 44,
        'distance_from_dc_km': 730,
        'description': 'พังงาทุกอำเภอ'
    },
    # ============ ภูเก็ต ============
    'ZONE_PKT1_ภูเก็ต': {
        'provinces': ['ภูเก็ต'],
        'districts': ['เมืองภูเก็ต', 'ถลาง', 'กะทู้'],
        'highway': '402',
        'priority': 43,
        'distance_from_dc_km': 850,
        'description': 'ภูเก็ตทุกอำเภอ'
    },
    # ============ ระนอง ============
    'ZONE_RNG1_ระนอง': {
        'provinces': ['ระนอง'],
        'districts': ['เมืองระนอง', 'กระบุรี', 'สุขสำราญ', 'ละอุ่น', 'กะเปอร์'],
        'highway': '4',
        'priority': 42,
        'distance_from_dc_km': 700,
        'description': 'ระนองทุกอำเภอ'
    },
    # ============ สตูล ============
    'ZONE_STL1_สตูล': {
        'provinces': ['สตูล'],
        'districts': ['เมืองสตูล', 'ควนโดน', 'ควนกาหลง', 'ท่าแพ', 'ละงู', 'ทุ่งหว้า', 'มะนัง'],
        'highway': '4',
        'priority': 44,
        'distance_from_dc_km': 800,
        'description': 'สตูลทุกอำเภอ'
    },

    # ============ ยะลา (sub-zones) ============
    'ZONE_YLA1_ยะลา_เมือง': {
        'provinces': ['ยะลา'],
        'districts': ['เมืองยะลา', 'ยะหา', 'กาบัง'],
        'highway': '410',
        'priority': 46,
        'distance_from_dc_km': 1100,
        'description': 'ยะลาเมืองและตะวันตก'
    },
    'ZONE_YLA2_ยะลา_รามัน': {
        'provinces': ['ยะลา'],
        'districts': ['รามัน', 'กรงปินัง', 'บันนังสตา'],
        'highway': '410',
        'priority': 45,
        'distance_from_dc_km': 1120,
        'description': 'ยะลากลาง-ตะวันออก'
    },
    'ZONE_YLA3_ยะลา_เบตง': {
        'provinces': ['ยะลา'],
        'districts': ['เบตง', 'ธารโต'],
        'highway': '410',
        'priority': 44,
        'distance_from_dc_km': 1200,
        'description': 'ยะลาใต้-เบตง'
    },

    # ============ ปัตตานี (sub-zones) ============
    'ZONE_PTN1_ปัตตานี_เมือง': {
        'provinces': ['ปัตตานี'],
        'districts': ['เมืองปัตตานี', 'หนองจิก', 'โคกโพธิ์', 'แม่ลาน'],
        'highway': '42',
        'priority': 46,
        'distance_from_dc_km': 1050,
        'description': 'ปัตตานีเมืองและใกล้เคียง'
    },
    'ZONE_PTN2_ปัตตานี_กลาง': {
        'provinces': ['ปัตตานี'],
        'districts': ['ยะรัง', 'ยะหริ่ง', 'มายอ', 'กะพ้อ'],
        'highway': '42',
        'priority': 45,
        'distance_from_dc_km': 1070,
        'description': 'ปัตตานีกลาง'
    },
    'ZONE_PTN3_ปัตตานี_ตะวันออก': {
        'provinces': ['ปัตตานี'],
        'districts': ['สายบุรี', 'ปะนาเระ', 'ทุ่งยางแดง', 'ไม้แก่น'],
        'highway': '42',
        'priority': 44,
        'distance_from_dc_km': 1090,
        'description': 'ปัตตานีตะวันออก'
    },

    # ============ นราธิวาส (sub-zones) ============
    'ZONE_NWT1_นราธิวาส_เมือง': {
        'provinces': ['นราธิวาส'],
        'districts': ['เมืองนราธิวาส', 'ตากใบ', 'บาเจาะ'],
        'highway': '42',
        'priority': 46,
        'distance_from_dc_km': 1150,
        'description': 'นราธิวาสเมืองและชายทะเล'
    },
    'ZONE_NWT2_นราธิวาส_กลาง': {
        'provinces': ['นราธิวาส'],
        'districts': ['รือเสาะ', 'ระแงะ', 'ยี่งอ', 'จะแนะ'],
        'highway': '42',
        'priority': 45,
        'distance_from_dc_km': 1170,
        'description': 'นราธิวาสกลาง'
    },
    'ZONE_NWT3_นราธิวาส_เหนือ': {
        'provinces': ['นราธิวาส'],
        'districts': ['สุไหงโก-ลก', 'สุไหงปาดี', 'แว้ง', 'สุคิริน'],
        'highway': '42',
        'priority': 44,
        'distance_from_dc_km': 1200,
        'description': 'นราธิวาสเหนือ-ชายแดนมาเลเซีย'
    },

    # ============ แม่ฮ่องสอน (sub-zones) ============
    'ZONE_MHS1_แม่ฮ่องสอน_เหนือ': {
        'provinces': ['แม่ฮ่องสอน'],
        'districts': ['เมืองแม่ฮ่องสอน', 'ปาย', 'ขุนยวม'],
        'highway': '108',
        'priority': 22,
        'distance_from_dc_km': 870,
        'description': 'แม่ฮ่องสอนเหนือ-เมือง-ปาย'
    },
    'ZONE_MHS2_แม่ฮ่องสอน_ใต้': {
        'provinces': ['แม่ฮ่องสอน'],
        'districts': ['แม่สะเรียง', 'แม่ลาน้อย', 'สบเมย'],
        'highway': '108',
        'priority': 21,
        'distance_from_dc_km': 900,
        'description': 'แม่ฮ่องสอนใต้-แม่สะเรียง'
    },
}

# ==========================================
# ZONE/REGION CONFIG - รหัสภาคและจังหวัด
# ==========================================
# รหัสภาค: 1=กลาง, 2=ตะวันออก, 3=ตะวันตก, 4=เหนือ, 5=อีสาน, 6=ใต้
REGION_CODE = {
    # ภาคกลาง (รหัส 1)
    'กรุงเทพมหานคร': '10', 'กรุงเทพฯ': '10',
    'นนทบุรี': '11',
    'ปทุมธานี': '12',
    'พระนครศรีอยุธยา': '13', 'อยุธยา': '13',
    'สระบุรี': '14',
    'ลพบุรี': '15',
    'สิงห์บุรี': '16',
    'อ่างทอง': '17',
    'ชัยนาท': '18',
    'นครปฐม': '19',
    'สมุทรปราการ': '1A',
    'สมุทรสาคร': '1B',
    'สมุทรสงคราม': '1C',
    
    # ภาคตะวันออก (รหัส 2)
    'ชลบุรี': '20',
    'ระยอง': '21',
    'จันทบุรี': '22',
    'ตราด': '23',
    'ฉะเชิงเทรา': '24',
    'ปราจีนบุรี': '25',
    'สระแก้ว': '26',
    'นครนายก': '27',
    
    # ภาคตะวันตก (รหัส 3)
    'ราชบุรี': '30',
    'กาญจนบุรี': '31',
    'สุพรรณบุรี': '32',
    'เพชรบุรี': '33',
    'ประจวบคีรีขันธ์': '34',
    
    # ภาคเหนือ (รหัส 4)
    'นครสวรรค์': '40',
    'อุทัยธานี': '41',
    'กำแพงเพชร': '42',
    'ตาก': '43',
    'สุโขทัย': '44',
    'พิษณุโลก': '45',
    'พิจิตร': '46',
    'เพชรบูรณ์': '47',
    'อุตรดิตถ์': '48',
    'แพร่': '49',
    'น่าน': '4A',
    'พะเยา': '4B',
    'เชียงราย': '4C',
    'เชียงใหม่': '4D',
    'แม่ฮ่องสอน': '4E',
    'ลำพูน': '4F',
    'ลำปาง': '4G',
    
    # ภาคตะวันออกเฉียงเหนือ/อีสาน (รหัส 5)
    'นครราชสีมา': '50', 'โคราช': '50',
    'บุรีรัมย์': '51',
    'สุรินทร์': '52',
    'ศรีสะเกษ': '53',
    'อุบลราชธานี': '54',
    'ยโสธร': '55',
    'ชัยภูมิ': '56',
    'อำนาจเจริญ': '57',
    'หนองบัวลำภู': '58',
    'ขอนแก่น': '59',
    'อุดรธานี': '5A',
    'เลย': '5B',
    'หนองคาย': '5C',
    'มหาสารคาม': '5D',
    'ร้อยเอ็ด': '5E',
    'กาฬสินธุ์': '5F',
    'สกลนคร': '5G',
    'นครพนม': '5H',
    'มุกดาหาร': '5I',
    'บึงกาฬ': '5J',
    
    # ภาคใต้ (รหัส 6)
    'ชุมพร': '60',
    'ระนอง': '61',
    'สุราษฎร์ธานี': '62',
    'พังงา': '63',
    'กระบี่': '64',
    'ภูเก็ต': '65',
    'นครศรีธรรมราช': '66',
    'ตรัง': '67',
    'พัทลุง': '68',
    'สงขลา': '69',
    'สตูล': '6A',
    'ปัตตานี': '6B',
    'ยะลา': '6C',
    'นราธิวาส': '6D',
}

# ชื่อภาค
REGION_NAMES = {
    '1': 'กลาง',
    '2': 'ตะวันออก',
    '3': 'ตะวันตก',
    '4': 'เหนือ',
    '5': 'อีสาน',
    '6': 'ใต้',
    '9': 'ไม่ระบุ'
}

# ==========================================
# HELPER: ZONE/REGION FUNCTIONS
# ==========================================
def get_region_code(province):
    """ดึงรหัสภาค/โซนจากจังหวัด"""
    if not province or str(province).strip() == '' or str(province) == 'nan':
        return '99'  # ไม่ระบุ
    province = clean_name(str(province).strip())  # ลบ จ./อ./ต. prefix
    # normalize aliases (พระนครศรีอยุธยา → อยุธยา ฯลฯ)
    _alias = {
        'พระนครศรีอยุธยา': 'อยุธยา',
        'กรุงเทพฯ': 'กรุงเทพมหานคร',
        'กทม': 'กรุงเทพมหานคร',
        'กทม.': 'กรุงเทพมหานคร',
        'โคราช': 'นครราชสีมา',
    }
    province = _alias.get(province, province)
    return REGION_CODE.get(province, '99')

_region_name_cache: dict = {}
def get_region_name(province):
    """ดึงชื่อภาคจากจังหวัด (cached)"""
    _k = str(province)
    if _k in _region_name_cache:
        return _region_name_cache[_k]
    code = get_region_code(province)
    if code == '99':
        _region_name_cache[_k] = 'ไม่ระบุ'
        return 'ไม่ระบุ'
    region_prefix = code[0]
    _v = REGION_NAMES.get(region_prefix, 'ไม่ระบุ')
    _region_name_cache[_k] = _v
    return _v

def _hav_vec(lat0: float, lon0: float, lats, lons) -> np.ndarray:
    """Vectorized haversine×1.35 จาก 1 จุด → numpy array ของจุดปลาย (km)"""
    R = 6371.0 * 1.35
    phi0 = radians(lat0)
    phi1 = np.radians(lats)
    dphi = np.radians(lats - lat0)
    dlam = np.radians(lons - lon0)
    a = np.sin(dphi / 2) ** 2 + cos(phi0) * np.cos(phi1) * np.sin(dlam / 2) ** 2
    return R * 2 * np.arctan2(np.sqrt(a), np.sqrt(1.0 - a))

# ==========================================
# LOGISTICS ZONE FUNCTIONS
# ==========================================

# 🏙️ Bangkok geographic center (สำหรับแบ่ง sub-zone)
_BKK_CENTER_LAT = 13.7563
_BKK_CENTER_LON = 100.5018
_BKK_CENTER_RADIUS_KM = 4.5  # รัศมี BKK_CENTER

# ชื่อ sub-zone กรุงเทพ (8 ทิศ + กลาง)
BKK_SUBZONE_NAMES = {
    'BKK_CENTER': 'กรุงเทพ - ใจกลาง (Silom/Sathorn/Siam)',
    'BKK_N':      'กรุงเทพ - เหนือ (ดอนเมือง/ลาดยาว/หลักสี่)',
    'BKK_NE':     'กรุงเทพ - ตะวันออกเฉียงเหนือ (ลาดพร้าว/มีนบุรี)',
    'BKK_E':      'กรุงเทพ - ตะวันออก (วังทองหลาง/ลาดกระบัง)',
    'BKK_SE':     'กรุงเทพ - ตะวันออกเฉียงใต้ (พระโขนง/บางนา)',
    'BKK_S':      'กรุงเทพ - ใต้ (ราษฎร์บูรณะ/บางขุนเทียน)',
    'BKK_SW':     'กรุงเทพ - ตะวันตกเฉียงใต้ (ธนบุรี/หนองแขม)',
    'BKK_W':      'กรุงเทพ - ตะวันตก (ตลิ่งชัน/บางแค)',
    'BKK_NW':     'กรุงเทพ - ตะวันตกเฉียงเหนือ (บางพลัด/บางซื่อ)',
}

def get_bkk_sub_zone(lat, lon):
    """
    จัดกรุงเทพแบ่ง sub-zone จากทิศทาง + ระยะจากใจกลาง
    Returns: 'BKK_CENTER' | 'BKK_N' | 'BKK_NE' | ... | 'BKK_NW'
    """
    if not lat or not lon or lat == 0 or lon == 0:
        return 'BKK_CENTER'
    dist = haversine_distance(_BKK_CENTER_LAT, _BKK_CENTER_LON, lat, lon, use_osrm_cache=False)
    if dist <= _BKK_CENTER_RADIUS_KM:
        return 'BKK_CENTER'
    bearing = calculate_bearing(_BKK_CENTER_LAT, _BKK_CENTER_LON, lat, lon)
    # 8 sectors (45° each), starting from North
    if bearing < 22.5 or bearing >= 337.5:
        return 'BKK_N'
    elif bearing < 67.5:
        return 'BKK_NE'
    elif bearing < 112.5:
        return 'BKK_E'
    elif bearing < 157.5:
        return 'BKK_SE'
    elif bearing < 202.5:
        return 'BKK_S'
    elif bearing < 247.5:
        return 'BKK_SW'
    elif bearing < 292.5:
        return 'BKK_W'
    else:
        return 'BKK_NW'


def get_prov_zone(province: str, district: str = '') -> str:
    """
    ดึงโซนจัดส่งระดับจังหวัด (ระบบเดียวกับ zone_viewer.py)
    BKK  → BKK_{เขต}    |    จังหวัดอื่น → {ภาค}_{จังหวัด}
    ใช้เป็น primary key สำหรับจัดทริป ป้องกันกระโดดข้ามภาค/จังหวัด
    """
    if not province:
        return 'ไม่ระบุ'
    prov = str(province).strip()
    _alias = {'กรุงเทพฯ': 'กรุงเทพมหานคร', 'กทม': 'กรุงเทพมหานคร',
              'กทม.': 'กรุงเทพมหานคร', 'โคราช': 'นครราชสีมา'}
    prov = _alias.get(prov, prov)
    rz = PROVINCE_ZONE_MAP.get(prov)
    if rz == '__BKK__':
        dist = str(district).strip() if district else ''
        return f'BKK_{dist}' if dist else 'BKK_ไม่ระบุ'
    return rz if rz else f'ไม่ระบุ_{prov}'


def classify_all_branch_zones(master_df=None):
    """
    จัดทุกสาขาใน MASTER_DATA เข้าโซนจัดส่ง (ล้วน geographic — ไม่คำนึง weight/cube)

    Returns:
        dict: {branch_code: zone_name}  (e.g. 'BKK_N', 'ZONE_A_พะเยา', ...)
        dict: zone_summary {zone_name: {'count': N, 'branches': [...]}}
    """
    if master_df is None:
        master_df = MASTER_DATA
    if master_df is None or master_df.empty:
        return {}, {}

    branch_zone_map = {}
    zone_summary = {}

    for _, row in master_df.iterrows():
        code = str(row.get('Plan Code', '')).strip().upper()
        if not code:
            continue

        province = str(row.get('จังหวัด', '') or '').strip()
        district  = str(row.get('อำเภอ', '')  or '').strip()
        subdistrict = str(row.get('ตำบล', '') or '').strip()

        # ──── กรุงเทพมหานคร: แบ่ง sub-zone ────
        _prov_alias = {'กรุงเทพฯ': 'กรุงเทพมหานคร', 'กทม': 'กรุงเทพมหานคร', 'กทม.': 'กรุงเทพมหานคร'}
        _prov_clean = _prov_alias.get(province, province)
        if _prov_clean == 'กรุงเทพมหานคร':
            lat_val = row.get('ละติจูด', 0) or 0
            lon_val = row.get('ลองติจูด', 0) or 0
            try:
                lat_val = float(lat_val)
                lon_val = float(lon_val)
            except (ValueError, TypeError):
                lat_val = lon_val = 0
            zone = get_bkk_sub_zone(lat_val, lon_val)
        else:
            # ──── จังหวัดอื่น: ใช้ LOGISTICS_ZONES ────
            zone = get_logistics_zone(_prov_clean, district, subdistrict)
            if not zone:
                zone = f'UNCLASSIFIED_{_prov_clean}' if _prov_clean else 'UNCLASSIFIED'

        branch_zone_map[code] = zone

        # สะสมสถิติ
        if zone not in zone_summary:
            zone_summary[zone] = {'count': 0, 'branches': [], 'province': _prov_clean}
        zone_summary[zone]['count'] += 1
        zone_summary[zone]['branches'].append(code)

    return branch_zone_map, zone_summary


def _build_zone_color_map(zone_summary):
    """
    สร้าง color map: {zone_name: '#rrggbb'}
    - BKK_* → 9 สีทิศ
    - โซนจังหวัดอื่น → จัดกลุ่มตามภาค แล้วใช้ palette ต่อเนื่อง
    """
    bkk_fixed = {
        'BKK_CENTER': '#C62828',
        'BKK_N':      '#1565C0',
        'BKK_NE':     '#0097A7',
        'BKK_E':      '#2E7D32',
        'BKK_SE':     '#F9A825',
        'BKK_S':      '#E65100',
        'BKK_SW':     '#6A1B9A',
        'BKK_W':      '#AD1457',
        'BKK_NW':     '#4527A0',
    }
    # Region-grouped palettes for province zones
    _region_palettes = {
        'เหนือ':       ['#0D47A1','#1565C0','#1976D2','#1E88E5','#2196F3','#42A5F5','#64B5F6','#90CAF9'],
        'อีสาน':       ['#1B5E20','#2E7D32','#388E3C','#43A047','#4CAF50','#66BB6A','#81C784','#A5D6A7'],
        'ใต้':         ['#006064','#00838F','#00ACC1','#00BCD4','#26C6DA','#4DD0E1','#80DEEA','#B2EBF2'],
        'ตะวันออก':    ['#E65100','#EF6C00','#F57C00','#FB8C00','#FF9800','#FFA726','#FFB74D','#FFCC80'],
        'กลาง':        ['#4A148C','#6A1B9A','#7B1FA2','#8E24AA','#9C27B0','#AB47BC','#BA68C8','#CE93D8'],
        'ตะวันตก':     ['#BF360C','#D84315','#E64A19','#F4511E','#FF5722','#FF7043','#FF8A65','#FFAB91'],
    }
    _fallback_colors = [
        '#607D8B','#78909C','#90A4AE','#B0BEC5',
        '#795548','#8D6E63','#A1887F','#BCAAA4',
    ]

    color_map = {}
    color_map.update(bkk_fixed)

    # group province zones by region
    from collections import defaultdict as _dd
    region_zones = _dd(list)
    for zk, zv in zone_summary.items():
        if zk.startswith('BKK_') or zk.startswith('UNCLASSIFIED'):
            continue
        prov = zv.get('province', '')
        region = get_region_name(prov) if prov else 'ไม่ระบุ'
        region_zones[region].append(zk)

    for region, zlist in region_zones.items():
        palette = _region_palettes.get(region, _fallback_colors)
        for i, zk in enumerate(sorted(zlist)):
            color_map[zk] = palette[i % len(palette)]

    # UNCLASSIFIED → grey
    for zk in zone_summary:
        if zk.startswith('UNCLASSIFIED'):
            color_map[zk] = '#9E9E9E'

    return color_map


def _build_zone_folium_map(master_df, branch_zone_map, color_map):
    """
    สร้าง Folium map แสดงสาขาทุกสาขาระบายสีตามโซน พร้อม Label โซน
    """
    if not FOLIUM_AVAILABLE:
        return None

    # Thailand center
    m = folium.Map(location=[13.0, 101.5], zoom_start=6,
                   tiles='CartoDB positron', control_scale=True)

    # Build lat/lon lookup from master_df
    lat_col = 'ละติจูด' if 'ละติจูด' in master_df.columns else None
    lon_col = 'ลองติจูด' if 'ลองติจูด' in master_df.columns else None
    name_col = 'สาขา' if 'สาขา' in master_df.columns else None
    code_col = 'Plan Code' if 'Plan Code' in master_df.columns else None

    if not (lat_col and lon_col and code_col):
        return m

    # Build coord dict  {code: (lat, lon, name)}
    _coords = {}
    for _, row in master_df.iterrows():
        code = str(row.get(code_col, '')).strip().upper()
        if not code:
            continue
        try:
            lat = float(row.get(lat_col, 0) or 0)
            lon = float(row.get(lon_col, 0) or 0)
        except (ValueError, TypeError):
            lat = lon = 0
        name = str(row.get(name_col, '') or '') if name_col else ''
        _coords[code] = (lat, lon, name)

    # Group branches by zone
    from collections import defaultdict as _ddict
    zone_branches = _ddict(list)
    for code, zone in branch_zone_map.items():
        code_upper = str(code).strip().upper()
        if code_upper in _coords:
            zone_branches[zone].append((code_upper, *_coords[code_upper]))

    # Create a FeatureGroup per zone + compute centroid for label
    zone_centroids = {}  # {zone: (lat, lon, count)}
    for zone, branches in sorted(zone_branches.items()):
        color = color_map.get(zone, '#9E9E9E')
        fg = folium.FeatureGroup(name=f"{zone} ({len(branches)})", show=True)

        lats, lons = [], []
        for code, lat, lon, name in branches:
            if lat == 0 or lon == 0:
                continue
            lats.append(lat)
            lons.append(lon)
            tooltip_html = f"<b>{code}</b><br>{name}<br><i>{zone}</i>"
            folium.CircleMarker(
                location=[lat, lon],
                radius=4,
                color=color,
                fill=True,
                fill_color=color,
                fill_opacity=0.75,
                weight=1,
                tooltip=folium.Tooltip(tooltip_html, sticky=False),
            ).add_to(fg)

        fg.add_to(m)

        if lats:
            zone_centroids[zone] = (sum(lats)/len(lats), sum(lons)/len(lons), len(lats), color)

    # Add zone label markers at centroids
    label_fg = folium.FeatureGroup(name="🏷️ Zone Labels", show=True)
    for zone, (clat, clon, cnt, color) in zone_centroids.items():
        short_label = zone.replace('ZONE_', '').replace('_', ' ')
        icon_html = (
            f'<div style="background:{color};color:#fff;border-radius:6px;'
            f'padding:3px 7px;font-size:11px;font-weight:700;white-space:nowrap;'
            f'border:1.5px solid rgba(0,0,0,.3);box-shadow:1px 1px 3px rgba(0,0,0,.25);'
            f'opacity:.92;">{short_label}</div>'
        )
        folium.Marker(
            location=[clat, clon],
            icon=folium.DivIcon(html=icon_html, icon_size=(120, 28), icon_anchor=(60, 14)),
            tooltip=f"{zone} — {cnt} สาขา",
        ).add_to(label_fg)
    label_fg.add_to(m)

    folium.LayerControl(collapsed=False, position='topright').add_to(m)
    return m


def _build_zone_excel(master_df, branch_zone_map, zone_summary, color_map):
    """
    สร้าง Excel หลายชีต:
    - สาขาทั้งหมด_โซน
    - สรุปโซน
    - กรุงเทพ_SubZone
    - หนึ่งชีตต่อภาค (จังหวัดโซน)
    """
    import io as _io
    output = _io.BytesIO()

    lat_col = 'ละติจูด' if 'ละติจูด' in master_df.columns else None
    lon_col = 'ลองติจูด' if 'ลองติจูด' in master_df.columns else None
    name_col = 'สาขา' if 'สาขา' in master_df.columns else None

    # Build main dataframe
    rows = []
    for _, row in master_df.iterrows():
        code = str(row.get('Plan Code', '')).strip().upper()
        if not code:
            continue
        zone = branch_zone_map.get(code, 'UNCLASSIFIED')
        color = color_map.get(zone, '#9E9E9E')
        name = str(row.get(name_col, '') or '') if name_col else ''
        prov = str(row.get('จังหวัด', '') or '')
        dist = str(row.get('อำเภอ', '') or '')
        subdist = str(row.get('ตำบล', '') or '')
        lat = ''
        lon = ''
        if lat_col:
            try: lat = float(row.get(lat_col, 0) or 0)
            except: lat = ''
        if lon_col:
            try: lon = float(row.get(lon_col, 0) or 0)
            except: lon = ''
        region = get_region_name(prov) if prov else ''
        zone_label = BKK_SUBZONE_NAMES.get(zone, zone)
        rows.append({
            'Plan Code': code,
            'ชื่อสาขา': name,
            'จังหวัด': prov,
            'อำเภอ': dist,
            'ตำบล': subdist,
            'ภาค': region,
            'Zone': zone,
            'Zone_Description': zone_label,
            '_hex': color,
            'ละติจูด': lat,
            'ลองติจูด': lon,
        })
    main_df = pd.DataFrame(rows)

    # Summary dataframe
    sum_rows = []
    for zone, zv in sorted(zone_summary.items(), key=lambda x: (-x[1]['count'], x[0])):
        prov = zv.get('province', '')
        region = get_region_name(prov) if prov else ''
        zone_desc = BKK_SUBZONE_NAMES.get(zone, zone)
        sum_rows.append({
            'Zone': zone,
            'คำอธิบาย': zone_desc,
            'จังหวัด': prov,
            'ภาค': region,
            'จำนวนสาขา': zv['count'],
            'สีโซน': color_map.get(zone, '#9E9E9E'),
        })
    sum_df = pd.DataFrame(sum_rows)

    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        wb = writer.book

        # ─── Helper: write df to sheet with header format + zone color ───
        def _write_sheet(df, sheet_name, freeze=True, color_col=None):
            export_df = df.drop(columns=[c for c in ['_hex'] if c in df.columns], errors='ignore')
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            # Header format
            hdr_fmt = wb.add_format({'bold': True, 'bg_color': '#1B5E20',
                                      'font_color': '#FFFFFF', 'border': 1,
                                      'align': 'center', 'valign': 'vcenter'})
            for col_idx, col_name in enumerate(export_df.columns):
                ws.write(0, col_idx, col_name, hdr_fmt)
                ws.set_column(col_idx, col_idx, max(12, min(40, len(str(col_name)) + 4)))
            if freeze:
                ws.freeze_panes(1, 0)
            # Color rows by zone
            if color_col and color_col in df.columns and '_hex' in df.columns:
                col_idx = list(export_df.columns).index(color_col)
                _fmt_cache = {}
                for row_idx, (_, row) in enumerate(df.iterrows(), start=1):
                    hex_c = str(row.get('_hex', '#FFFFFF')).replace('#', '')
                    if hex_c not in _fmt_cache:
                        _fmt_cache[hex_c] = wb.add_format({
                            'bg_color': f'#{hex_c}', 'font_color': '#FFFFFF',
                            'bold': True, 'border': 1, 'align': 'center'
                        })
                    ws.write(row_idx, col_idx, row.get(color_col, ''), _fmt_cache[hex_c])

        # Sheet 1: ทุกสาขา
        _write_sheet(main_df, 'สาขาทั้งหมด_โซน', color_col='Zone')

        # Sheet 2: สรุปโซน
        _write_sheet(sum_df, 'สรุปโซนทั้งหมด', color_col='Zone')

        # Sheet 3: กรุงเทพ sub-zone detail
        bkk_df = main_df[main_df['Zone'].str.startswith('BKK_', na=False)].copy()
        if not bkk_df.empty:
            _write_sheet(bkk_df, 'กรุงเทพ_SubZone', color_col='Zone')

        # Sheet 4-N: province zones by region
        from collections import defaultdict as _ddef
        region_map_ex = _ddef(list)
        for _, row in main_df.iterrows():
            z = row.get('Zone', '')
            if z.startswith('BKK_') or z.startswith('UNCLASSIFIED'):
                continue
            region_map_ex[row.get('ภาค', 'ไม่ระบุ')].append(row)
        for region_name, region_rows in sorted(region_map_ex.items()):
            sheet_df = pd.DataFrame(region_rows)
            safe_name = f"โซน_{region_name}"[:31]
            _write_sheet(sheet_df, safe_name, color_col='Zone')

        # Sheet last: UNCLASSIFIED
        unc_df = main_df[main_df['Zone'].str.startswith('UNCLASSIFIED', na=False)].copy()
        if not unc_df.empty:
            _write_sheet(unc_df, 'ไม่ระบุโซน')

    output.seek(0)
    return output.getvalue()


def get_logistics_zone(province, district='', subdistrict=''):
    """
    หาโซนโลจิสติกส์จาก จังหวัด/อำเภอ/ตำบล
    หลักการ: ทุกจังหวัด = per-district zone เพื่อให้ทริปหมดอำเภอก่อนข้าม
    Returns: 'ZONE_{จังหวัด}_{อำเภอ}' หรือ 'ZONE_{จังหวัด}' ถ้าไม่มีอำเภอ
    """
    if not province or str(province).strip() == '':
        return None

    province = str(province).strip()
    district  = str(district).strip()  if district  else ''
    subdistrict = str(subdistrict).strip() if subdistrict else ''

    # ถ้ามีทั้ง province + district → คืน per-district zone เสมอ
    if district:
        return f'ZONE_{province}_{district}'

    # ไม่มี district → คืน province-level zone
    return f'ZONE_{province}'

def get_zone_priority(zone_name):
    """Priority ของโซน: ZONE_{จังหวัด}_{อำเภอ} → ดึงจาก _distance_from_dc (ผ่าน LOGISTICS_ZONES fallback)"""
    if not zone_name:
        return 999
    if zone_name in LOGISTICS_ZONES:
        return LOGISTICS_ZONES[zone_name].get('priority', 50)
    # per-district format: ZONE_{province}_{district} → ใช้ priority 50 (default)
    return 50

def zone_is_specific(zone_name: str) -> bool:
    """True ถ้า zone เป็น per-district (มี 3 ส่วน: ZONE_{จังหวัด}_{อำเภอ}) หรือ LOGISTICS_ZONES ที่มี districts"""
    if not zone_name:
        return False
    # LOGISTICS_ZONES ที่มี districts กำหนด
    zi = LOGISTICS_ZONES.get(zone_name, {})
    if zi.get('districts'):
        return True
    # per-district format: ZONE_{จังหวัด}_{อำเภอ} = 3 parts เสมอ
    parts = zone_name.split('_')
    return len(parts) >= 3 and parts[0] == 'ZONE' and bool(parts[2])

def get_zone_highway(zone_name):
    """
    ดึงทางหลวงหลักของโซน
    
    Returns:
        str: เช่น 'สาย 1 (พหลโยธิน)', 'สาย 2 (มิตรภาพ)'
    """
    if not zone_name:
        return ''
    if zone_name in LOGISTICS_ZONES:
        return LOGISTICS_ZONES[zone_name].get('highway', '')
    # zone_viewer format — derive highway from zone prefix
    _ZV_HW = {
        'เหนือ': '1/11', 'อีสาน': '2/24', 'ใต้': '4',
        'ตะวันออก': '3', 'ตะวันตก': '32/4', 'ปริมณฑล': '',
    }
    prefix = zone_name.split('_')[0] if '_' in zone_name else zone_name
    return _ZV_HW.get(prefix, '')

def can_combine_zones_by_highway(zone1, zone2):
    """
    เช็คว่า 2 โซนอยู่บนทางหลวงเดียวกันหรือไม่
    (ถ้าใช่ → สามารถรวมทริปได้)
    
    Returns:
        bool: True ถ้าอยู่ทางเดียวกัน
    """
    if not zone1 or not zone2:
        return False
    
    highway1 = get_zone_highway(zone1)
    highway2 = get_zone_highway(zone2)
    
    if not highway1 or not highway2:
        return False
    
    # เช็คว่าทางหลวงมีส่วนร่วมกัน (set intersection) รองรับ '304' == '304/331'
    return bool(set(highway1.split('/')) & set(highway2.split('/')))

def is_cross_zone_violation(province1, province2):
    """
    เช็คว่าจังหวัดทั้ง 2 อยู่ใน NO_CROSS_ZONE_PAIRS หรือไม่
    (พยายามหลีกเลี่ยง - Soft Rule)
    
    Returns:
        bool: True ถ้าควรหลีกเลี่ยงการรวมโซน
    """
    if not province1 or not province2:
        return False
    
    prov1 = str(province1).strip()
    prov2 = str(province2).strip()
    
    # เช็คทั้ง 2 ทาง
    return (prov1, prov2) in NO_CROSS_ZONE_PAIRS or (prov2, prov1) in NO_CROSS_ZONE_PAIRS

def are_provinces_on_same_route(province1, province2):
    """
    เช็คว่าจังหวัดทั้ง 2 อยู่ใน ROUTE เดียวกันหรือไม่
    
    Returns:
        bool: True ถ้าอยู่ route เดียวกัน (ควรรวมกัน)
    """
    if not province1 or not province2:
        return False
    
    prov1 = str(province1).strip()
    prov2 = str(province2).strip()
    
    # หา route ของแต่ละจังหวัด
    route1 = None
    route2 = None
    
    for route_name, route_info in ROUTE_GROUPS.items():
        if prov1 in route_info['provinces']:
            route1 = route_name
        if prov2 in route_info['provinces']:
            route2 = route_name
    
    # ถ้าอยู่ route เดียวกัน
    if route1 and route2:
        if route1 == route2:
            return True
        # เช็ค next_routes
        if route2 in ROUTE_GROUPS.get(route1, {}).get('next_routes', []):
            return True
        if route1 in ROUTE_GROUPS.get(route2, {}).get('next_routes', []):
            return True
    
    return False

def calculate_district_centroid(district_df):
    """คำนวณจุดกลางของอำเภอจากพิกัดสาขา"""
    valid_coords = district_df[district_df['_lat'] > 0]
    if valid_coords.empty:
        return None, None
    return valid_coords['_lat'].mean(), valid_coords['_lon'].mean()

def check_geographic_proximity(district1_df, district2_df, max_distance_km=MAX_DISTRICT_DISTANCE_KM):
    """ตรวจสอบว่า 2 อำเภอใกล้กันพอที่จะอยู่ทริปเดียวกันได้หรือไม่"""
    # ตรวจสอบจังหวัด
    prov1 = district1_df['_province'].iloc[0] if not district1_df.empty else ''
    prov2 = district2_df['_province'].iloc[0] if not district2_df.empty else ''
    
    # 🚨 ถ้าคนละจังหวัด → เช็ค Logistics Zone ก่อน
    if prov1 and prov2 and prov1 != prov2:
        # แช็คว่าอยู่ Logistics Zone เดียวกันหรือไม่
        zone1 = get_logistics_zone(prov1, '', '')
        zone2 = get_logistics_zone(prov2, '', '')
        
        if zone1 and zone2:
            if zone1 == zone2:
                # ✅ Zone เดียวกัน → ไม่จำกัดระยะทาง (เลือกใกล้ที่สุดในโซน)
                return True
            else:
                # คนละ Zone → เช็คว่าอยู่บนทางหลวงเดียวกันหรือไม่
                if not can_combine_zones_by_highway(zone1, zone2):
                    return False  # คนละทางหลวง → ห้ามรวม
    
    # คำนวณระยะห่างระหว่าง centroids
    lat1, lon1 = calculate_district_centroid(district1_df)
    lat2, lon2 = calculate_district_centroid(district2_df)
    
    if lat1 is None or lat2 is None:
        return True  # ไม่มีพิกัด ให้ผ่าน
    
    distance = haversine_distance(lat1, lon1, lat2, lon2, use_osrm_cache=False)
    
    if prov1 and prov2 and prov1 == prov2:
        # ✅ จังหวัดเดียวกัน → ใช้ threshold กว้างกว่า (60km)
        return distance <= (max_distance_km * 2.0)  # 30km * 2.0 = 60km
    else:
        # ⚠️ คนละจังหวัด + คนละ Zone → ใช้ threshold เข้มงวด (30km)
        return distance <= max_distance_km

def sort_branches_by_region_route(branches_df, master_data=None):
    """
    จัดเรียงสาขาตามภาค → จังหวัด → อำเภอ → ตำบล → Route
    เพื่อให้ทริปเรียงติดกันไม่กระโดด
    """
    if branches_df.empty:
        return branches_df
    
    df = branches_df.copy()
    
    # หาชื่อคอลัมน์จังหวัด (รองรับทั้ง Province และ จังหวัด)
    province_col = 'Province' if 'Province' in df.columns else 'จังหวัด' if 'จังหวัด' in df.columns else None
    
    # เพิ่มคอลัมน์สำหรับ sort
    df['_region_code'] = df[province_col].apply(get_region_code) if province_col else '99'
    df['_province'] = df[province_col].fillna('') if province_col else ''
    df['_district'] = df['District'].fillna('') if 'District' in df.columns else ''
    df['_subdistrict'] = df['Subdistrict'].fillna('') if 'Subdistrict' in df.columns else ''
    
    # แยก Route number
    if 'Route' in df.columns:
        df['_route_num'] = df['Route'].apply(lambda x: int(str(x).replace('CD', '')) if pd.notna(x) and str(x).startswith('CD') else 99999)
    else:
        df['_route_num'] = 99999
    
    # Sort
    df = df.sort_values(by=['_region_code', '_province', '_district', '_subdistrict', '_route_num'])
    
    # ลบคอลัมน์ชั่วคราว
    df = df.drop(columns=['_region_code', '_province', '_district', '_subdistrict', '_route_num'])
    
    return df.reset_index(drop=True)

def check_trip_route_spread(trip_df):
    """
    ตรวจสอบว่าทริปมี Route กระจายมากไหม
    คืนค่า: (route_range, is_spread, provinces)
    """
    if trip_df.empty or 'Route' not in trip_df.columns:
        return 0, False, []
    
    routes = trip_df['Route'].dropna().unique()
    route_nums = []
    for r in routes:
        if pd.notna(r) and str(r).startswith('CD'):
            try:
                route_nums.append(int(str(r).replace('CD', '')))
            except:
                pass
    
    if len(route_nums) < 2:
        return 0, False, trip_df['Province'].dropna().unique().tolist() if 'Province' in trip_df.columns else []
    
    route_range = max(route_nums) - min(route_nums)
    is_spread = route_range > 4000  # ถ้ามากกว่า 4000 ถือว่ากระจาย
    
    provinces = trip_df['Province'].dropna().unique().tolist() if 'Province' in trip_df.columns else []
    
    return route_range, is_spread, provinces

# ==========================================
# LOAD MASTER DATA
# ==========================================
@st.cache_data(ttl=300, show_spinner=False)  # Cache 5 นาที (real-time เมื่อ Sheets เปลี่ยน)
def load_master_data():
    """โหลด Master Data จาก Google Sheets หรือ JSON (auto-sync)"""
    try:
        # ใช้ข้อมูลจาก Google Sheets ที่ sync มาแล้ว
        df_from_sheets = sync_branch_data_from_sheets()
        
        if df_from_sheets is None or df_from_sheets.empty:
            safe_print("⚠️ ไม่สามารถโหลดข้อมูล - ตรวจสอบ Google Sheets หรือ branch_data.json")
            return pd.DataFrame()
        
        # ตรวจสอบคอลัมน์ที่จำเป็น
        required_cols = ['Plan Code']
        missing = [c for c in required_cols if c not in df_from_sheets.columns]
        if missing:
            safe_print(f"⚠️ ขาดคอลัมน์: {missing}")
        
        # แปลงชื่อคอลัมน์ที่อาจต่างกัน
        col_mapping = {
            'ละ': 'ละติจูด',
            'ลอง': 'ลองติจูด'
        }
        df_from_sheets = df_from_sheets.rename(columns=col_mapping)
        
        # ──────────────────────────────────────────────────────────
        # 🔑 Normalize column names: ลบ space, newline, tab ที่ Sheets ใส่มา
        # (เช่น 'Max       x\nTruckType' → 'MaxTruckType')
        # ──────────────────────────────────────────────────────────
        import re as _re_col
        df_from_sheets.columns = [
            _re_col.sub(r'[\s\n\r\t]+', '', str(c)) for c in df_from_sheets.columns
        ]

        # ทำความสะอาด Plan Code
        if 'Plan Code' in df_from_sheets.columns:
            df_from_sheets['Plan Code'] = df_from_sheets['Plan Code'].astype(str).str.strip().str.upper()
            df_from_sheets = df_from_sheets[df_from_sheets['Plan Code'] != '']
        elif 'PlanCode' in df_from_sheets.columns:   # หลัง normalize อาจกลายเป็น PlanCode
            df_from_sheets.rename(columns={'PlanCode': 'Plan Code'}, inplace=True)
            df_from_sheets['Plan Code'] = df_from_sheets['Plan Code'].astype(str).str.strip().str.upper()
            df_from_sheets = df_from_sheets[df_from_sheets['Plan Code'] != '']
        
        safe_print(f"✅ โหลด MASTER_DATA: {len(df_from_sheets)} สาขา")
        
        # 🔍 Debug: แสดงคอลัมน์ทั้งหมดที่มี
        safe_print(f"📋 คอลัมน์ทั้งหมด ({len(df_from_sheets.columns)}): {safe_join(df_from_sheets.columns.tolist())}")
        
        # 🔍 Debug: แสดงคอลัมน์ที่เกี่ยวข้องกับรถ (ค้นหาแบบยืดหยุ่น)
        vehicle_cols = [
            'MaxTruckType', 'Max Truck Type', 'MaxVehicle', 'Max Vehicle', 
            'รถสูงสุด', 'Max_Truck_Type', 'max_truck', 'MaxTruck',
            'ข้อจำกัดรถ', 'Truck', 'truck_type', 'TruckType',
            'ประเภทรถ', 'Vehicle', 'vehicle_type', 'VehicleType'
        ]
        found_vehicle_cols = [col for col in vehicle_cols if col in df_from_sheets.columns]
        if found_vehicle_cols:
            safe_print(f"✅ พบคอลัมน์ข้อจำกัดรถ: {safe_join(found_vehicle_cols)}")
            # แสดงสถิติข้อจำกัดรถ
            for col in found_vehicle_cols:
                vehicle_counts = df_from_sheets[col].value_counts(dropna=False)
                safe_print(f"   - {col}: {dict(vehicle_counts)}")
        else:
            safe_print(f"⚠️ ไม่พบคอลัมน์ข้อจำกัดรถ!")
            # ค้นหาคอลัมน์ที่อาจเกี่ยวข้อง
            for col in df_from_sheets.columns:
                if 'truck' in col.lower() or 'vehicle' in col.lower() or 'รถ' in col:
                    safe_print(f"   💡 คอลัมน์ที่อาจเกี่ยวข้อง: '{col}'")
        
        return df_from_sheets
        
    except Exception as e:
        safe_print(f"❌ Error loading MASTER_DATA: {e}")
        return pd.DataFrame()

# โหลด Master Data จาก Google Sheets
MASTER_DATA = load_master_data()

# ──────────────────────────────────────────────────────────────────
# 🔑 MASTER_DATA_DICT  — Plan Code (upper) เป็น PK → O(1) lookup
# ── สร้างทุกครั้งที่ MASTER_DATA โหลดใหม่ ──
# ──────────────────────────────────────────────────────────────────
def _build_master_dict(md: 'pd.DataFrame') -> dict:
    """สร้าง dict {plan_code_upper: row_dict} จาก MASTER_DATA"""
    if md is None or md.empty or 'Plan Code' not in md.columns:
        return {}
    result = {}
    truck_cols_priority = [
        'MaxTruckType', 'Max Truck Type', 'MaxVehicle', 'Max Vehicle',
        'รถสูงสุด', 'Max_Truck_Type', 'max_truck', 'MaxTruck',
        'ข้อจำกัดรถ', 'Truck', 'truck_type', 'TruckType',
        'ประเภทรถ', 'Vehicle', 'vehicle_type', 'VehicleType'
    ]
    found_truck_col = next((c for c in truck_cols_priority if c in md.columns), None)
    for _, row in md.iterrows():
        code = str(row.get('Plan Code', '')).strip().upper()
        if not code:
            continue
        entry = {
            'max_truck': '6W',   # default
            '_row': row.to_dict(),
        }
        if found_truck_col:
            raw = str(row.get(found_truck_col, '')).strip().upper()
            if raw in ('4W', '4 W', '4-W'):
                entry['max_truck'] = '4W'
            elif raw in ('JB', 'J B', 'J-B', '4WJ', '4WJ '):
                entry['max_truck'] = 'JB'
            elif raw in ('6W', '6 W', '6-W'):
                entry['max_truck'] = '6W'
        result[code] = entry
    safe_print(f"🔑 MASTER_DATA_DICT: {len(result)} สาขา (PK=Plan Code) | truck col='{found_truck_col}'")
    return result

MASTER_DATA_DICT: dict = _build_master_dict(MASTER_DATA)

# ══════════════════════════════════════════════════════════════════════════════
# 🗺️ BRANCH_ZONES_CACHE — โหลดจาก branch_zones.json ที่ zone_viewer.py สร้าง
# Format: {branch_code_upper: zone_string}  เช่น "NY00" → "เหนือ_เชียงใหม่_เมือง"
# ══════════════════════════════════════════════════════════════════════════════
def _load_branch_zones() -> dict:
    """โหลด branch_zones.json ที่ zone_viewer.py export ไว้"""
    _path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'branch_zones.json')
    if not os.path.exists(_path):
        safe_print("⚠️ ไม่พบ branch_zones.json — ใช้ LOGISTICS_ZONES แทน (รัน zone_viewer.py ก่อน)")
        return {}
    try:
        with open(_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        safe_print(f"🗺️ โหลด branch_zones.json: {len(data):,} สาขา")
        return {str(k).upper(): v for k, v in data.items()}
    except Exception as e:
        safe_print(f"⚠️ โหลด branch_zones.json ล้มเหลว: {e}")
        return {}

BRANCH_ZONES_CACHE: dict = _load_branch_zones()

# ==========================================
# 🔄 BRANCH GROUPING (จุดส่งเดียวกัน ≤200 เมตร)
# โหลดจาก branch_groups.json (สร้างโดย precompute_branch_data.py ด้วย haversine ≤500m + ตำบล/อำเภอ/จังหวัดเดียวกัน)
# ==========================================
@st.cache_resource(show_spinner=False)
def load_branch_groups():
    """
    โหลด branch_groups.json ที่สร้างด้วย haversine ≤500m + ตำบล/อำเภอ/จังหวัดเดียวกัน
    Return: (groups_dict, branch_to_group_dict)
    """
    try:
        with open('branch_groups.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        groups = data.get('groups', {})  # {group_id: [codes]}
        branch_to_group = data.get('branch_to_group', {})  # {code: group_id}

        # Auto-rebuild branch_to_group ถ้าว่าง (backward compat กับ JSON เก่า)
        if not branch_to_group and groups:
            branch_to_group = {}
            for _gid, _codes in groups.items():
                for _c in _codes:
                    branch_to_group[str(_c).strip().upper()] = _gid
            safe_print(f"🔧 Rebuild branch_to_group: {len(branch_to_group)} entries จาก {len(groups)} groups")

        # Normalize keys เป็น UPPER เสมอ
        branch_to_group = {str(k).strip().upper(): v for k, v in branch_to_group.items()}

        safe_print(f"✅ โหลด branch_groups.json: {len(groups)} กลุ่ม, {len(branch_to_group)} สาขา")
        return groups, branch_to_group
    except FileNotFoundError:
        safe_print("⚠️ ไม่พบ branch_groups.json - รัน python precompute_branch_data.py ก่อน")
        return {}, {}
    except Exception as e:
        safe_print(f"⚠️ โหลด branch_groups.json ไม่สำเร็จ: {e}")
        return {}, {}

# โหลด branch groups
BRANCH_GROUPS, BRANCH_TO_GROUP = load_branch_groups()

def get_group_branches(code: str) -> list:
    """
    ดึงสาขาทั้งหมดในกลุ่มเดียวกัน (จุดส่งเดียวกัน ≤200 เมตร)
    ถ้าไม่มีกลุ่ม return [code] (สาขาเดียว)
    """
    code_upper = str(code).strip().upper()
    group_id = BRANCH_TO_GROUP.get(code_upper)
    if group_id:
        return BRANCH_GROUPS.get(group_id, [code_upper])
    return [code_upper]

def is_same_group(code1: str, code2: str) -> bool:
    """เช็คว่า 2 สาขาอยู่กลุ่มเดียวกันหรือไม่"""
    c1 = str(code1).strip().upper()
    c2 = str(code2).strip().upper()
    g1 = BRANCH_TO_GROUP.get(c1)
    g2 = BRANCH_TO_GROUP.get(c2)
    return g1 and g2 and g1 == g2

# ==========================================
# 🚀 BRANCH CLUSTERS & SPATIAL DATA (Pre-computed)
# โหลดจาก branch_clusters.json (สร้างโดย precompute_branch_data.py)
# ==========================================
@st.cache_data(show_spinner=False)
def load_branch_clusters():
    """
    โหลด branch_clusters.json ที่มี:
    - branch_info: พิกัด, ระยะห่างจาก DC, ทิศทาง, cluster
    - nearby_branches: สาขาใกล้เคียง (< 15km)
    - clusters: กลุ่มตามระยะทาง, ทิศทาง, จังหวัด, อำเภอ
    """
    try:
        with open('branch_clusters.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        branch_info = {str(k).strip().upper(): v for k, v in data.get('branch_info', {}).items()}
        nearby_branches = {str(k).strip().upper(): v for k, v in data.get('nearby_branches', {}).items()}
        clusters = data.get('clusters', {})
        
        safe_print(f"✅ โหลด branch_clusters.json:")
        safe_print(f"   - {len(branch_info)} สาขามีข้อมูล spatial")
        safe_print(f"   - {len(nearby_branches)} สาขามี nearby branches")
        if clusters:
            safe_print(f"   - Distance clusters: {len(clusters.get('distance', {}))}")
            safe_print(f"   - Direction clusters: {len(clusters.get('direction', {}))}")
            safe_print(f"   - Province clusters: {len(clusters.get('province', {}))}")
            safe_print(f"   - District clusters: {len(clusters.get('district', {}))}")
        
        return branch_info, nearby_branches, clusters
    except FileNotFoundError:
        safe_print("⚠️ ไม่พบ branch_clusters.json - รัน python precompute_branch_data.py ก่อน")
        return {}, {}, {}
    except Exception as e:
        safe_print(f"⚠️ โหลด branch_clusters.json ไม่สำเร็จ: {e}")
        return {}, {}, {}

# โหลด branch clusters
BRANCH_INFO, NEARBY_BRANCHES, BRANCH_CLUSTERS = load_branch_clusters()

# ==========================================
# �️ PRE-SEED DISTANCE CACHE จาก branch_clusters.json
# inject ระยะทาง pre-computed ทุกคู่ใน NEARBY_BRANCHES → DISTANCE_CACHE
# เพื่อให้ hot-path haversine_distance(use_osrm_cache=False) ได้ cache hit เสมอ
# (รันใน background thread ไม่บล็อก UI)
# ==========================================
def _preseed_distance_cache_from_clusters():
    """
    อ่านระยะทางจาก branch_clusters.json (pre-computed OSRM distances)
    และ inject เข้า DISTANCE_CACHE ทุกคู่ที่ยังไม่มี
    จากนั้น OSRM live สำหรับคู่ที่ไม่มี pre-computed distance (batch ทีละ 20 คู่)
    """
    if not USE_CACHE or not BRANCH_INFO or not NEARBY_BRANCHES:
        return
    import threading
    import time as _time

    def _run():
        global _DIST_CACHE_DIRTY
        _injected = 0
        _to_fetch: list = []  # [(key, lat1, lon1, lat2, lon2)]

        for code, nearby_list in NEARBY_BRANCHES.items():
            code_up = str(code).strip().upper()
            info1 = BRANCH_INFO.get(code_up, {})
            lat1 = info1.get('lat', 0)
            lon1 = info1.get('lon', 0)
            if not lat1 or not lon1:
                continue
            for item in nearby_list:
                if isinstance(item, dict):
                    nb_code = str(item.get('code', '')).strip().upper()
                    pre_dist = item.get('distance', None)
                else:
                    nb_code = str(item).strip().upper()
                    pre_dist = None
                if not nb_code:
                    continue
                info2 = BRANCH_INFO.get(nb_code, {})
                lat2 = info2.get('lat', 0)
                lon2 = info2.get('lon', 0)
                if not lat2 or not lon2:
                    continue
                ck  = f"{lat1:.4f},{lon1:.4f}_{lat2:.4f},{lon2:.4f}"
                ckr = f"{lat2:.4f},{lon2:.4f}_{lat1:.4f},{lon1:.4f}"
                if ck in DISTANCE_CACHE or ckr in DISTANCE_CACHE:
                    continue  # มีแล้ว
                if pre_dist and pre_dist > 0:
                    # inject จาก pre-computed ทันที
                    DISTANCE_CACHE[ck] = round(pre_dist, 2)
                    _DIST_CACHE_DIRTY += 1
                    _injected += 1
                else:
                    # ไม่มี pre_dist → inject haversine×1.35 ทันที (ไม่เรียก OSRM)
                    from math import radians, sin, cos, sqrt, atan2
                    R = 6371.0
                    phi1, phi2 = radians(lat1), radians(lat2)
                    dphi = radians(lat2 - lat1)
                    dlambda = radians(lon2 - lon1)
                    a = sin(dphi/2)**2 + cos(phi1)*cos(phi2)*sin(dlambda/2)**2
                    hav_dist = round(R * 2 * atan2(sqrt(a), sqrt(1-a)) * 1.35, 2)
                    DISTANCE_CACHE[ck] = hav_dist
                    _DIST_CACHE_DIRTY += 1
                    _injected += 1

        # inject DC → ทุกสาขาใน BRANCH_INFO (haversine×1.35, zero-network)
        _dc_injected = 0
        _dc_lat = DC_WANG_NOI_LAT
        _dc_lon = DC_WANG_NOI_LON
        from math import radians as _r, sin as _s, cos as _c, sqrt as _sq, atan2 as _at
        for _bc_code, _bc_info in BRANCH_INFO.items():
            _blat = _bc_info.get('lat', 0)
            _blon = _bc_info.get('lon', 0)
            if not _blat or not _blon:
                continue
            _ck_dc  = f"{_dc_lat:.4f},{_dc_lon:.4f}_{_blat:.4f},{_blon:.4f}"
            _ck_dcr = f"{_blat:.4f},{_blon:.4f}_{_dc_lat:.4f},{_dc_lon:.4f}"
            if _ck_dc in DISTANCE_CACHE or _ck_dcr in DISTANCE_CACHE:
                continue
            _phi1, _phi2 = _r(_dc_lat), _r(_blat)
            _dphi = _r(_blat - _dc_lat)
            _dlam = _r(_blon - _dc_lon)
            _a = _s(_dphi/2)**2 + _c(_phi1)*_c(_phi2)*_s(_dlam/2)**2
            _hav = round(6371.0 * 2 * _at(_sq(_a), _sq(1-_a)) * 1.35, 2)
            DISTANCE_CACHE[_ck_dc] = _hav
            _DIST_CACHE_DIRTY += 1
            _dc_injected += 1

        # save ถ้ามีของใหม่
        if _DIST_CACHE_DIRTY > 0:
            save_distance_cache(DISTANCE_CACHE, force=True)
        safe_print(f"🗄️ Pre-seed cache: nearby {_injected} + DC→สาขา {_dc_injected} คู่ ({len(DISTANCE_CACHE):,} total)")

    t = threading.Thread(target=_run, daemon=True, name="preseed-cache")
    t.start()

_preseed_distance_cache_from_clusters()

# ==========================================
# �🚀 PRE-COMPUTE: Distance Matrix & Nearby Branches
# ใช้ข้อมูลจาก branch_clusters.json แทนการคำนวณใหม่
# ==========================================
@st.cache_data(ttl=3600)  # Cache 1 ชั่วโมง
def precompute_branch_distances(master_df):
    """
    โหลดข้อมูล pre-computed จาก branch_clusters.json
    หรือคำนวณใหม่ถ้าไม่มีไฟล์
    """
    # ถ้ามีข้อมูล pre-computed ให้ใช้เลย
    if BRANCH_INFO and NEARBY_BRANCHES:
        safe_print("✅ ใช้ข้อมูล pre-computed จาก branch_clusters.json")
        
        # แปลง branch_info เป็น branch_coords (ใช้ uppercase key ทั้งหมด)
        branch_coords = {}
        for code, info in BRANCH_INFO.items():
            if 'lat' in info and 'lon' in info:
                branch_coords[str(code).strip().upper()] = (info['lat'], info['lon'])
        
        # แปลง nearby_branches เป็นรูปแบบที่ต้องการ
        nearby_dict = {}
        for code, nearby_list in NEARBY_BRANCHES.items():
            code_upper = str(code).strip().upper()
            # nearby_list มีได้สองรูปแบบ:
            #   - [code1, code2, ...]  (รูปแบบเก่า)
            #   - [{"code": code1, "distance": d1}, ...]  (รูปแบบใหม่จาก precompute_branch_data.py)
            nearby_with_dist = []
            if code_upper in branch_coords:
                lat1, lon1 = branch_coords[code_upper]
                for item in nearby_list:
                    # รองรับทั้งสองรูปแบบ
                    if isinstance(item, dict):
                        nearby_code = str(item.get('code', '')).strip().upper()
                        pre_dist = item.get('distance', None)
                    else:
                        nearby_code = str(item).strip().upper()
                        pre_dist = None

                    if not nearby_code or nearby_code not in branch_coords:
                        continue

                    lat2, lon2 = branch_coords[nearby_code]
                    if pre_dist is not None and pre_dist > 0:
                        # ใช้ค่า pre-computed จาก branch_clusters.json (OSRM road dist)
                        dist = pre_dist
                    else:
                        # Inline: ตรวจ DISTANCE_CACHE ก่อน → fallback haversine × 1.35
                        _ck  = f"{lat1:.4f},{lon1:.4f}_{lat2:.4f},{lon2:.4f}"
                        _ckr = f"{lat2:.4f},{lon2:.4f}_{lat1:.4f},{lon1:.4f}"
                        if USE_CACHE and _ck in DISTANCE_CACHE:
                            dist = DISTANCE_CACHE[_ck]
                        elif USE_CACHE and _ckr in DISTANCE_CACHE:
                            dist = DISTANCE_CACHE[_ckr]
                        else:
                            from math import radians, sin, cos, sqrt, atan2
                            _phi1, _phi2 = radians(lat1), radians(lat2)
                            _a = sin((radians(lat2-lat1))/2)**2 + cos(_phi1)*cos(_phi2)*sin((radians(lon2-lon1))/2)**2
                            dist = round(6371.0 * 2 * atan2(sqrt(_a), sqrt(1-_a)) * 1.35, 2)
                    nearby_with_dist.append((nearby_code, round(dist, 2)))
            nearby_dict[code_upper] = sorted(nearby_with_dist, key=lambda x: x[1])
        
        # สร้าง same_area_branches จาก clusters (ใช้ uppercase key)
        same_area_branches = {}
        if BRANCH_CLUSTERS and 'district' in BRANCH_CLUSTERS:
            district_clusters = BRANCH_CLUSTERS['district']
            for code, info in BRANCH_INFO.items():
                code_up = str(code).strip().upper()
                district_id = info.get('district_cluster')
                if district_id and district_id in district_clusters:
                    same_area_branches[code_up] = [str(c).strip().upper() for c in district_clusters[district_id] if str(c).strip().upper() != code_up]
                else:
                    same_area_branches[code_up] = []
        
        safe_print(f"   ✅ {len(branch_coords)} สาขามีพิกัด, {len(nearby_dict)} สาขามี nearby")
        return branch_coords, nearby_dict, same_area_branches
    
    # ถ้าไม่มี pre-computed ให้คำนวณใหม่
    safe_print("⚠️ ไม่มี pre-computed data - คำนวณใหม่...")
    
    if master_df.empty:
        return {}, {}, {}
    
    # ดึงพิกัดสาขาทั้งหมด
    branch_coords = {}
    for _, row in master_df.iterrows():
        code = str(row.get('Plan Code', '')).strip().upper()
        lat = row.get('ละติจูด') or row.get('Latitude') or row.get('ละ', 0)
        lon = row.get('ลองติจูด') or row.get('Longitude') or row.get('ลอง', 0)
        if code and lat and lon:
            try:
                lat_float = float(lat)
                lon_float = float(lon)
                if lat_float > 0 and lon_float > 0:
                    branch_coords[code] = (lat_float, lon_float)
            except (ValueError, TypeError):
                pass
    
    safe_print(f"   📍 พบ {len(branch_coords)} สาขาที่มีพิกัด")
    
    # สร้าง nearby_branches (คำนวณแบบเร็ว)
    nearby_branches = {}
    same_area_branches = {}
    
    codes = list(branch_coords.keys())
    for code in codes:
        nearby_branches[code] = []
        same_area_branches[code] = []
    
    safe_print(f"   ✅ เตรียมข้อมูลเบื้องต้นเสร็จสิ้น")
    
    return branch_coords, nearby_branches, same_area_branches

# Pre-compute distances
BRANCH_COORDS, NEARBY_BRANCHES, SAME_AREA_BRANCHES = precompute_branch_distances(MASTER_DATA)

# ==========================================
# CLEAN NAME FUNCTION (สำหรับทำ Join_Key)
# ==========================================
def clean_name(text):
    """
    ทำความสะอาดชื่อ: ลบ prefix จ./อ./ต. และ trim whitespace
    ใช้สำหรับสร้าง Join_Key เพื่อเทียบกับ Master Data
    """
    if pd.isna(text) or text is None:
        return ''
    text = str(text)
    # ลบ prefix ภาษาไทย
    text = text.replace('จ. ', '').replace('จ.', '')
    text = text.replace('อ. ', '').replace('อ.', '')
    text = text.replace('ต. ', '').replace('ต.', '')
    # ลบ prefix ภาษาอังกฤษ (ถ้ามี)
    text = text.replace('Tambon ', '').replace('Amphoe ', '').replace('Changwat ', '')
    return text.strip()

def normalize_province_name(province):
    """
    แปลงชื่อจังหวัดให้เป็นมาตรฐาน (แก้ปัญหาชื่อเพี้ยน)
    """
    if pd.isna(province) or province is None:
        return ''
    province = clean_name(province)
    # Mapping ชื่อที่พบบ่อย
    province_mapping = {
        'พระนครศรีอยุธยา': 'อยุธยา',
        'กรุงเทพฯ': 'กรุงเทพมหานคร',
        'กทม': 'กรุงเทพมหานคร',
        'กทม.': 'กรุงเทพมหานคร',
        'โคราช': 'นครราชสีมา',
    }
    return province_mapping.get(province, province)

def normalize(val):
    """ทำให้รหัสสาขาเป็นมาตรฐาน"""
    return str(val).strip().upper().replace(" ", "").replace(".0", "")

# ==========================================
# PUNTHAI/MAXMART BUFFER FUNCTIONS (REMOVED - ใช้โลจิกใหม่แล้ว)
# ==========================================
# HELPER FUNCTIONS
# ==========================================

def get_max_vehicle_for_branch(branch_code, test_df=None, debug=False):
    """ดึงรถใหญ่สุดที่สาขานี้รองรับ
    ใช้ MASTER_DATA_DICT (PK=Plan Code) เพื่อ O(1) lookup
    """
    branch_code_str = str(branch_code).strip().upper()

    # ── 1. Fast path: ค้นจาก MASTER_DATA_DICT โดยตรง ──
    if MASTER_DATA_DICT:
        entry = MASTER_DATA_DICT.get(branch_code_str)

        # ── 2. ถ้าไม่พบ ลอง prefix-strip fallback ──
        if entry is None:
            prefixes = ['PUN-', 'MAX-', 'MM-', 'PT-']
            code_clean = branch_code_str
            for p in prefixes:
                if code_clean.startswith(p):
                    code_clean = code_clean[len(p):]
                    break
            if code_clean != branch_code_str:
                entry = MASTER_DATA_DICT.get(code_clean)
            # ลอง match แบบ strip prefix จากฝั่ง master
            if entry is None:
                for mk, mv in MASTER_DATA_DICT.items():
                    mk_clean = mk
                    for p in prefixes:
                        if mk_clean.startswith(p):
                            mk_clean = mk_clean[len(p):]
                            break
                    if mk_clean == code_clean:
                        entry = mv
                        break

        if entry is not None:
            return entry['max_truck']   # '4W' / 'JB' / '6W'

    # ── 3. Legacy fallback: scan DataFrame (ถ้า dict ยังไม่ build) ──
    if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
        master_codes = MASTER_DATA['Plan Code'].str.strip().str.upper()
        branch_row = MASTER_DATA[master_codes == branch_code_str]
        if not branch_row.empty:
            possible_cols = [
                'MaxTruckType', 'Max Truck Type', 'MaxVehicle', 'Max Vehicle',
                'รถสูงสุด', 'Max_Truck_Type', 'max_truck', 'MaxTruck',
                'ข้อจำกัดรถ', 'Truck', 'truck_type', 'TruckType',
                'ประเภทรถ', 'Vehicle', 'vehicle_type', 'VehicleType'
            ]
            for col in possible_cols:
                if col in branch_row.columns and pd.notna(branch_row.iloc[0][col]):
                    raw = str(branch_row.iloc[0][col]).strip().upper()
                    if raw in ('4W', '4 W', '4-W'):
                        return '4W'
                    elif raw in ('JB', 'J B', 'J-B', '4WJ', '4WJ '):
                        return 'JB'
                    elif raw in ('6W', '6 W', '6-W'):
                        return '6W'

    # Default: ไม่มีข้อจำกัด = ใช้รถใหญ่ได้
    return '6W'

def get_max_vehicle_for_trip(trip_codes):
    """
    หารถใหญ่สุดที่ทริปนี้ใช้ได้ (เช็คข้อจำกัดของทุกสาขาในทริป)
    
    Args:
        trip_codes: set ของ branch codes ในทริป
    
    Returns:
        str: '4W', 'JB', หรือ '6W'
    """
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    max_allowed = '6W'  # เริ่มจากใหญ่สุด แล้วจำกัดตามข้อจำกัดสาขา
    min_priority = 3  # ค่าใหญ่สุดคือไม่มีข้อจำกัด
    
    for code in trip_codes:
        branch_max = get_max_vehicle_for_branch(code)
        priority = vehicle_priority.get(branch_max, 3)
        
        # 🔒 เลือกรถที่เล็กที่สุด (ข้อจำกัดมากที่สุด) จากทุกสาขาในทริป
        if priority < min_priority:
            min_priority = priority
            max_allowed = branch_max
    
    return max_allowed

def get_route_osrm(pickup_lat, pickup_lon, dropoff_lat, dropoff_lon, max_retries=1):
    """
    ขอเส้นทางจริงจาก OSRM API (วิ่งตามถนน)
    ตรวจ ROUTE_CACHE_DATA ก่อนเสมอ — เรียก API เฉพาะตอนที่ยังไม่เคย cache ไว้
    """
    if not FOLIUM_AVAILABLE:
        return [[pickup_lat, pickup_lon], [dropoff_lat, dropoff_lon]]

    # ตรวจ cache ก่อน
    cache_key = f"{pickup_lat:.4f},{pickup_lon:.4f}|{dropoff_lat:.4f},{dropoff_lon:.4f}"
    if USE_CACHE and cache_key in ROUTE_CACHE_DATA:
        cached = ROUTE_CACHE_DATA[cache_key]
        if isinstance(cached, dict):
            return cached.get('coords', [[pickup_lat, pickup_lon], [dropoff_lat, dropoff_lon]])
        return cached  # backward compat (list)

    # OSRM Public Server (lon, lat format!)
    loc = f"{pickup_lon},{pickup_lat};{dropoff_lon},{dropoff_lat}"
    url = f"http://router.project-osrm.org/route/v1/driving/{loc}?overview=full&geometries=geojson"

    for attempt in range(max_retries):
        try:
            r = requests.get(url, timeout=4)
            res = r.json()

            if "routes" in res and len(res["routes"]) > 0:
                coords = res["routes"][0]["geometry"]["coordinates"]
                route_coords = [[lat, lon] for lon, lat in coords]
                # บันทึก cache ทันที
                if USE_CACHE:
                    ROUTE_CACHE_DATA[cache_key] = {'coords': route_coords, 'distance': 0}
                    global _ROUTE_CACHE_DIRTY
                    _ROUTE_CACHE_DIRTY += 1
                    if _ROUTE_CACHE_DIRTY >= _ROUTE_CACHE_SAVE_BATCH:
                        save_route_cache(ROUTE_CACHE_DATA)
                return route_coords
            else:
                return [[pickup_lat, pickup_lon], [dropoff_lat, dropoff_lon]]
        except Exception:
            return [[pickup_lat, pickup_lon], [dropoff_lat, dropoff_lon]]

    return [[pickup_lat, pickup_lon], [dropoff_lat, dropoff_lon]]


def get_multi_point_route_osrm(waypoints, max_retries=2):
    """
    ขอเส้นทางจริงจาก OSRM API สำหรับหลายจุด พร้อม cache
    
    Args:
        waypoints: list ของ [lat, lon] เช่น [[14.1, 100.6], [14.2, 100.7], ...]
        max_retries: จำนวนครั้งที่ลองใหม่
    
    Returns:
        tuple: (route_coords, distance_km) - พิกัดเส้นทาง และระยะทางรวม
    """
    if not FOLIUM_AVAILABLE or len(waypoints) < 2:
        return waypoints, 0
    
    # สร้าง cache key
    cache_key = "|".join([f"{lat:.4f},{lon:.4f}" for lat, lon in waypoints])
    
    # ตรวจสอบ cache ก่อน
    if USE_CACHE and cache_key in ROUTE_CACHE_DATA:
        cached = ROUTE_CACHE_DATA[cache_key]
        return cached['coords'], cached['distance']
    
    # OSRM รับพิกัดแบบ lon,lat (ไม่ใช่ lat,lon!)
    coords_str = ";".join([f"{lon},{lat}" for lat, lon in waypoints])
    url = f"http://router.project-osrm.org/route/v1/driving/{coords_str}?overview=full&geometries=geojson"
    
    for attempt in range(max_retries):
        try:
            r = requests.get(url, timeout=10)
            res = r.json()
            
            if "routes" in res and len(res["routes"]) > 0:
                route = res["routes"][0]
                # แปลง GeoJSON coordinates (lon, lat) เป็น (lat, lon)
                coords = route["geometry"]["coordinates"]
                route_coords = [[lat, lon] for lon, lat in coords]
                # ระยะทางจาก OSRM เป็นเมตร
                distance_km = route.get("distance", 0) / 1000
                
                # บันทึก cache ทันทีทุกครั้งที่ได้ข้อมูลใหม่
                if USE_CACHE:
                    ROUTE_CACHE_DATA[cache_key] = {
                        'coords': route_coords,
                        'distance': distance_km
                    }
                    global _ROUTE_CACHE_DIRTY
                    _ROUTE_CACHE_DIRTY += 1
                    if _ROUTE_CACHE_DIRTY >= _ROUTE_CACHE_SAVE_BATCH:
                        save_route_cache(ROUTE_CACHE_DATA)
                
                return route_coords, distance_km
            else:
                return waypoints, 0
        except Exception as e:
            if attempt < max_retries - 1:
                time_module.sleep(0.3)
                continue
            return waypoints, 0
    
    return waypoints, 0

def calculate_bearing(lat1, lon1, lat2, lon2):
    """
    คำนวณทิศทาง (bearing) จากจุด 1 ไปจุด 2 เป็นองศา (0-360)
    0 = เหนือ, 90 = ตะวันออก, 180 = ใต้, 270 = ตะวันตก
    """
    from math import radians, sin, cos, atan2, degrees
    
    lat1_rad = radians(lat1)
    lat2_rad = radians(lat2)
    dlon = radians(lon2 - lon1)
    
    x = sin(dlon) * cos(lat2_rad)
    y = cos(lat1_rad) * sin(lat2_rad) - sin(lat1_rad) * cos(lat2_rad) * cos(dlon)
    
    bearing = atan2(x, y)
    bearing = degrees(bearing)
    bearing = (bearing + 360) % 360  # Normalize to 0-360
    
    return bearing

def get_bearing_zone(bearing):
    """
    แบ่งทิศทางเป็น 8 โซน (ทุก 45 องศา)
    0-1 = N, 2-3 = NE, 4-5 = E, 6-7 = SE, 8-9 = S, 10-11 = SW, 12-13 = W, 14-15 = NW
    """
    # แบ่งเป็น 16 โซน (ทุก 22.5 องศา) เพื่อจัดกลุ่มสาขาที่อยู่ทิศเดียวกัน
    zone = int((bearing + 11.25) / 22.5) % 16
    return zone

def get_osrm_distance_live(lat1, lon1, lat2, lon2):
    """
    เรียก OSRM Table API เพื่อดึงระยะทางถนนจริงระหว่างสองจุด (km)
    คืนค่า float km ถ้าสำเร็จ หรือ None ถ้าล้มเหลว
    """
    try:
        url = (
            f"http://router.project-osrm.org/table/v1/driving/"
            f"{lon1},{lat1};{lon2},{lat2}?annotations=distance"
        )
        r = requests.get(url, timeout=6)
        data = r.json()
        if data.get("code") == "Ok":
            dist_m = data["distances"][0][1]  # จากจุด 0 → จุด 1
            if dist_m and dist_m > 0:
                return dist_m / 1000.0
    except Exception:
        pass
    return None


def _osrm_table_batch(points):
    """
    ระยะ branch-to-branch (km) ทุกคู่ — zero latency
    cache hit → ระยะถนนจริง, miss → haversine×1.35 ทันที (ไม่ยิง network)
    """
    import math as _m
    n = len(points)
    result = {}

    def _hav_km(la1, lo1, la2, lo2):
        R = 6371.0
        p1, p2 = _m.radians(la1), _m.radians(la2)
        a = (_m.sin(_m.radians(la2-la1)/2)**2 +
             _m.cos(p1)*_m.cos(p2)*_m.sin(_m.radians(lo2-lo1)/2)**2)
        return 2*R*_m.atan2(_m.sqrt(max(0,a)), _m.sqrt(max(0,1-a))) * 1.35

    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            la1, lo1 = points[i]
            la2, lo2 = points[j]
            ck   = f"{la1:.4f},{lo1:.4f}_{la2:.4f},{lo2:.4f}"
            ck_r = f"{la2:.4f},{lo2:.4f}_{la1:.4f},{lo1:.4f}"
            if USE_CACHE and ck in DISTANCE_CACHE:
                result[(i, j)] = DISTANCE_CACHE[ck]
            elif USE_CACHE and ck_r in DISTANCE_CACHE:
                result[(i, j)] = DISTANCE_CACHE[ck_r]
            else:
                result[(i, j)] = _hav_km(la1, lo1, la2, lo2)

    return result


def haversine_distance(lat1, lon1, lat2, lon2, use_osrm_cache=True):
    """
    คืนค่าระยะทางถนน (km)
    ลำดับ:
      1. DISTANCE_CACHE → คืนค่าระยะทางถนนจริงทันที (เร็วที่สุด, ทั้งสองโหมด)
      2. Cache miss + use_osrm_cache=True  → OSRM live (6s), cache ผล, fallback haversine×1.35
      3. Cache miss + use_osrm_cache=False → haversine×1.35 ทันที (zero-latency, hot-path)
    """
    # 1. ตรวจ DISTANCE_CACHE ก่อนเสมอ (ทั้งสองโหมดได้ระยะทางจริงถ้ามีแคช)
    cache_key = f"{lat1:.4f},{lon1:.4f}_{lat2:.4f},{lon2:.4f}"
    cache_key_reverse = f"{lat2:.4f},{lon2:.4f}_{lat1:.4f},{lon1:.4f}"

    if USE_CACHE:
        if cache_key in DISTANCE_CACHE:
            return DISTANCE_CACHE[cache_key]
        if cache_key_reverse in DISTANCE_CACHE:
            return DISTANCE_CACHE[cache_key_reverse]

    # 2a. Cache miss + hot-path → haversine×1.35 ทันที (ไม่ network เด็ดขาด)
    if not use_osrm_cache:
        R = 6371.0
        phi1, phi2 = radians(lat1), radians(lat2)
        dphi = radians(lat2 - lat1)
        dlambda = radians(lon2 - lon1)
        a = sin(dphi/2)**2 + cos(phi1)*cos(phi2)*sin(dlambda/2)**2
        return round(R * 2 * atan2(sqrt(a), sqrt(1-a)) * 1.35, 2)

    # 2b. Cache miss + precision → OSRM live, cache ไว้
    try:
        _url = (
            f"http://router.project-osrm.org/table/v1/driving/"
            f"{lon1},{lat1};{lon2},{lat2}?annotations=distance"
        )
        _r = requests.get(_url, timeout=6)
        _data = _r.json()
        if _data.get("code") == "Ok":
            _dist_m = _data["distances"][0][1]
            if _dist_m and _dist_m > 0:
                dist_km = round(_dist_m / 1000.0, 2)
                if USE_CACHE:
                    DISTANCE_CACHE[cache_key] = dist_km
                    global _DIST_CACHE_DIRTY
                    _DIST_CACHE_DIRTY += 1
                    if _DIST_CACHE_DIRTY >= _DIST_CACHE_SAVE_BATCH:
                        save_distance_cache(DISTANCE_CACHE)
                return dist_km
    except Exception:
        pass

    # 3. OSRM ล้มเหลว/timeout → haversine×1.35 (fallback)
    R = 6371.0
    phi1, phi2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dlambda = radians(lon2 - lon1)
    a = sin(dphi/2)**2 + cos(phi1)*cos(phi2)*sin(dlambda/2)**2
    return round(R * 2 * atan2(sqrt(a), sqrt(1-a)) * 1.35, 2)

def load_model():
    """โหลดโมเดลที่เทรนไว้"""
    if not os.path.exists(MODEL_PATH):
        return None
    
    try:
        import warnings as _warnings
        with _warnings.catch_warnings(record=True) as _caught:
            _warnings.simplefilter("always")
            with open(MODEL_PATH, 'rb') as f:
                model_data = pickle.load(f)
        # suppress model version mismatch warning in UI
        return model_data
    except Exception as e:
        st.error(f"❌ Error loading model: {e}")
        return None

@st.cache_data(show_spinner=False)
def _extract_all_info(file_content):
    """
    เปิด openpyxl ครั้งเดียว แล้วดึง header info + style info + DC row ในรอบเดียว
    คืน (header_list, style_dict, dc_dict)
    """
    import openpyxl

    header_list = []
    style_dict  = {'row_height': 15.0, 'font_name': 'Angsana New', 'font_size': 14.0}
    dc_dict     = {}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
        ws = next((wb[sn] for sn in wb.sheetnames
                   if 'punthai' in sn.lower() or '2.' in sn.lower()), wb.active)

        # หา header row
        hrow = 1
        headers = []
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)), start=1):
            vals_upper = ' '.join(str(c.value or '').upper() for c in row)
            if sum(kw in vals_upper for kw in ('BRANCH', 'TRIP', 'รหัสสาขา', 'BU')) >= 2:
                hrow = ri
                headers = [str(c.value) if c.value is not None else '' for c in row]
                break

        # ── 1. header colors ──
        for cell in ws[hrow]:
            if cell.column > ws.max_column:
                break
            name  = str(cell.value) if cell.value is not None else ''
            color = '#D9D9D9'
            try:
                fill = cell.fill
                if fill and fill.fill_type == 'solid' and fill.fgColor:
                    fg = fill.fgColor
                    if fg.type == 'rgb' and fg.rgb and len(fg.rgb) >= 6:
                        rgb_hex = fg.rgb[-6:]
                        if rgb_hex.upper() not in ('FFFFFF', '000000'):
                            color = '#' + rgb_hex
            except Exception:
                pass
            header_list.append((name, color))

        # ── 2. style (row height + font จากแถวข้อมูลแรก) ──
        data_row_idx = hrow + 1
        rd = ws.row_dimensions.get(data_row_idx)
        if rd and rd.height:
            style_dict['row_height'] = float(rd.height)
        for cell in ws[data_row_idx]:
            if cell.value is not None:
                try:
                    if cell.font:
                        if cell.font.name:  style_dict['font_name'] = cell.font.name
                        if cell.font.size:  style_dict['font_size'] = float(cell.font.size)
                except Exception:
                    pass
                break

        # ── 3. DC row ──
        if not headers:
            headers = [str(c.value) if c.value is not None else ''
                       for c in next(ws.iter_rows(min_row=hrow, max_row=hrow))]
        _DC_CODES = {'DC011', 'PTDC', 'PTG DISTRIBUTION CENTER'}
        for row in ws.iter_rows(min_row=hrow + 1, max_row=ws.max_row):
            vals = [c.value for c in row]
            if any(str(v or '').strip().upper() in _DC_CODES for v in vals[:6]):
                dc_dict = {headers[i]: ('' if vals[i] is None else vals[i])
                           for i in range(min(len(headers), len(vals)))}
                break

        wb.close()
    except Exception as ex:
        safe_print(f"⚠️ _extract_all_info: {ex}")

    return header_list, style_dict, dc_dict


def _extract_header_info(file_content):
    return _extract_all_info(file_content)[0]

def _extract_style_info(file_content):
    return _extract_all_info(file_content)[1]

def _extract_dc_row_info(file_content):
    return _extract_all_info(file_content)[2]


@st.cache_data(show_spinner=False)
def load_excel(file_content, sheet_name=None):
    """โหลด Excel — อ่านไฟล์ครั้งเดียว, cache ตาม content hash"""
    try:
        xls = pd.ExcelFile(io.BytesIO(file_content))

        target_sheet = None
        if sheet_name and sheet_name in xls.sheet_names:
            target_sheet = sheet_name
        else:
            for s in xls.sheet_names:
                if 'punthai' in s.lower() or '2.' in s.lower():
                    target_sheet = s
                    break
        if not target_sheet:
            target_sheet = xls.sheet_names[0]

        # อ่านครั้งเดียว header=None แล้วหา header row จาก df ที่โหลดแล้ว
        df_raw = pd.read_excel(xls, sheet_name=target_sheet, header=None)
        header_row = 1
        for i in range(min(10, len(df_raw))):
            row_joined = ' '.join(str(v) for v in df_raw.iloc[i])
            row_upper  = row_joined.upper()
            if sum(['BRANCH' in row_upper, 'TRIP' in row_upper,
                    'รหัสสาขา' in row_joined, 'จำนวนชิ้น' in row_joined,
                    'น้ำหนัก' in row_joined or 'น้ําหนัก' in row_joined,
                    'คิว' in row_joined, 'WMS' in row_upper, 'SEP' in row_upper]) >= 2:
                header_row = i
                break

        # ใช้ df_raw แทนการอ่านซ้ำ
        df = df_raw.iloc[header_row + 1:].copy()
        df.columns = df_raw.iloc[header_row].tolist()
        df = df.reset_index(drop=True)
        df = df.loc[:, ~df.columns.duplicated()]
        return df
    except Exception as e:
        import traceback as _tb
        safe_print(f"❌ load_excel: {e}\n{_tb.format_exc()}")
        return None

def process_dataframe(df):
    """แปลงคอลัมน์เป็นรูปแบบมาตรฐาน"""
    if df is None:
        return None
    
    rename_map = {}
    
    # ใช้ลำดับตำแหน่งคอลัมน์ตามไฟล์ test.xlsx sheet 2.Punthai
    # 0:Sep, 1:BU, 2:BranchCode, 3:รหัสWMS, 4:Branch, 5:TOTALCUBE, 6:TOTALWGT, 7:OriginalQTY, ...
    col_list = list(df.columns)
    
    # ลำดับ 1 = BU
    if len(col_list) > 1:
        rename_map[col_list[1]] = 'BU'
    # ลำดับ 2 = รหัสสาขา (BranchCode)
    if len(col_list) > 2:
        rename_map[col_list[2]] = 'Code'
    # ลำดับ 3 = รหัส WMS
    if len(col_list) > 3:
        rename_map[col_list[3]] = 'WMSCode'
    # ลำดับ 4 = สาขา/ชื่อ (Branch)
    if len(col_list) > 4:
        rename_map[col_list[4]] = 'Name'
    # ลำดับ 5 = TOTALCUBE
    if len(col_list) > 5:
        rename_map[col_list[5]] = 'Cube'
    # ลำดับ 6 = TOTALWGT
    if len(col_list) > 6:
        rename_map[col_list[6]] = 'Weight'
    # ลำดับ 7 = OriginalQTY
    if len(col_list) > 7:
        rename_map[col_list[7]] = 'OriginalQty'
    # ลำดับ 15 = latitude
    if len(col_list) > 15:
        rename_map[col_list[15]] = 'Latitude'
    # ลำดับ 16 = longitude
    if len(col_list) > 16:
        rename_map[col_list[16]] = 'Longitude'
    
    # ตรวจสอบเพิ่มเติมจากชื่อคอลัมน์ (สำรองถ้าไฟล์มีคอลัมน์น้อยหรือโครงสร้างต่าง)
    for col in df.columns:
        if col in rename_map.values():  # ถ้า map แล้วข้าม
            continue
        if col in rename_map:  # ถ้าเป็น key ใน map แล้วข้าม
            continue
        col_clean = str(col).strip()
        col_upper = col_clean.upper().replace(' ', '').replace('_', '')
        
        # BU
        if col_upper == 'BU' or col_clean == 'BU':
            rename_map[col] = 'BU'
        # Code
        elif col_clean == 'BranchCode' or 'รหัสสาขา' in col_clean or col_clean ==  'BRANCH_CODE' in col_upper or 'CODE' in col_upper:
            if 'Weight' not in col_upper and 'Cube' not in col_upper:  # ป้องกันไม่ให้จับ WeightCode
                rename_map[col] = 'Code'
        # Name
        elif col_clean == 'Branch' or 'ชื่อสาขา' in col_clean or col_clean == 'สาขา' or ('BRANCH' in col_upper and 'CODE' not in col_upper):
            rename_map[col] = 'Name'
        # ตำบล
        elif 'ตำบล' in col_clean or 'SUBDISTRICT' in col_upper or 'TAMBON' in col_upper:
            rename_map[col] = 'Subdistrict'
        # อำเภอ
        elif 'อำเภอ' in col_clean or ('DISTRICT' in col_upper and 'SUB' not in col_upper) or 'AMPHOE' in col_upper or 'AMPHUR' in col_upper:
            rename_map[col] = 'District'
        # จังหวัด
        elif 'จังหวัด' in col_clean or 'PROVINCE' in col_upper or 'CHANGWAT' in col_upper:
            rename_map[col] = 'Province'
        # Weight - ตรวจสอบหลายรูปแบบ
        elif ('น้ำหนัก' in col_clean or 
              'WEIGHT' in col_upper or 
              'WGT' in col_upper or 
              'TOTALWGT' in col_upper or
              'น้ําหนัก' in col_clean or  # รองรับ ำ ที่พิมพ์ผิด
              col_upper in ['WEIGHT', 'WGT', 'TOTALWEIGHT']):
            rename_map[col] = 'Weight'
        # Cube - ตรวจสอบหลายรูปแบบ
        elif ('คิว' in col_clean or 
              'CUBE' in col_upper or 
              'TOTALCUBE' in col_upper or
              'CBM' in col_upper or
              col_upper in ['CUBE', 'CBM', 'TOTALCUBE']):
            rename_map[col] = 'Cube'
        # Latitude
        elif 'latitude' in col_clean.lower() or col_clean == 'ละติจูด' or 'LAT' == col_upper or col_upper == 'LATITUDE':
            rename_map[col] = 'Latitude'
        # Longitude
        elif 'longitude' in col_clean.lower() or col_clean == 'ลองติจูด' or col_upper in ['LONG', 'LNG', 'LON', 'LONGITUDE']:
            rename_map[col] = 'Longitude'
        # Trip
        elif col_upper in ['TRIPNO', 'TRIP_NO', 'TRIPNUMBER'] or col_clean == 'Trip no':
            rename_map[col] = 'TripNo'
        elif col_upper == 'TRIP' or 'ทริป' in col_clean or 'เที่ยว' in col_clean:
            rename_map[col] = 'Trip'
        # WMSCode
        elif 'WMS' in col_upper or col_upper in ['WMSCODE', 'BRANCHCODEWMS', 'CODEWMS']:
            rename_map[col] = 'WMSCode'
        # OriginalQty
        elif col_upper in ['ORIGINALQTY', 'ORIGINALQUANTITY', 'ORIGINALQTY', 'ORIG_QTY', 'ORIGQTY'] or 'ORIGINALQ' in col_upper:
            rename_map[col] = 'OriginalQty'
        # Booking
        elif 'BOOKING' in col_upper:
            rename_map[col] = 'Booking'
    
    df = df.rename(columns=rename_map)

    # บันทึก rename_map สำหรับ export (original col name → internal name)
    try:
        st.session_state['_col_rename_map'] = dict(rename_map)
    except Exception:
        pass

    # ลบคอลัมน์ซ้ำ
    df = df.loc[:, ~df.columns.duplicated()]
    
    if 'Code' in df.columns:
        df['Code'] = df['Code'].apply(normalize)
        
        # ตัดสาขาที่ไม่ต้องการออก (รหัส)
        df = df[~df['Code'].isin(EXCLUDE_BRANCHES)]
        
        # ตัดสาขาที่ชื่อมี keyword ที่ไม่ต้องการ
        if 'Name' in df.columns:
            _excl_kws = [re.escape(str(n)) for n in EXCLUDE_NAMES
                         if n is not None and not (isinstance(n, float) and pd.isna(n))]
            if _excl_kws:
                exclude_pattern = '|'.join(_excl_kws)
                df = df[~df['Name'].str.contains(exclude_pattern, case=False, na=False)]
    
    for col in ['Weight', 'Cube', 'OriginalQty']:
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    
    # เพิ่มจังหวัดจาก Master ถ้ายังไม่มี (รองรับทั้ง Province และ จังหวัด)
    province_col = 'Province' if 'Province' in df.columns else 'จังหวัด' if 'จังหวัด' in df.columns else None
    
    if not province_col or df[province_col].isna().all():
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns and 'Code' in df.columns:
            # สร้าง mapping จาก Master
            province_map = {}
            for _, row in MASTER_DATA.iterrows():
                code = row.get('Plan Code', '')
                province = row.get('จังหวัด', '')
                if code and province:
                    province_map[code] = province
            
            # ฟังก์ชันค้นหาจังหวัดจากชื่อสาขา
            def find_province_by_name(code, name):
                # ลองหาจาก code ก่อน
                if code in province_map:
                    return province_map[code]
                
                # ถ้าไม่เจอ ลองค้นหาจากชื่อสาขา
                if not name or pd.isna(name):
                    return ''
                
                # แยกคำสำคัญจากชื่อ (เอาคำแรกที่ไม่ใช่ prefix)
                keywords = str(name).replace('MAX MART-', '').replace('PUNTHAI-', '').replace('LUBE', '').strip()
                if not keywords:
                    return ''
                
                # ค้นหาในชื่อสาขาของ Master (ใช้ vectorized แทน iterrows)
                name_lookup = {}
                if not MASTER_DATA.empty and 'สาขา' in MASTER_DATA.columns and 'จังหวัด' in MASTER_DATA.columns:
                    for rec in MASTER_DATA[['สาขา', 'จังหวัด']].dropna().to_dict('records'):
                        name_lookup[str(rec['สาขา'])[:10]] = str(rec.get('จังหวัด', ''))
                for prefix, prov in name_lookup.items():
                    if keywords[:10] == prefix or prefix in keywords:
                        return prov if prov else ''
                return ''
            
            # ใส่จังหวัดให้แต่ละสาขา (สร้างคอลัมน์ Province ถ้ายังไม่มี)
            target_col = 'Province' if 'Province' in df.columns else 'จังหวัด'
            if 'Name' in df.columns:
                df[target_col] = df.apply(lambda row: find_province_by_name(row['Code'], row.get('Name', '')), axis=1).fillna('')
            else:
                df[target_col] = df['Code'].map(province_map).fillna('')
            
            # สร้าง Province ถ้ายังไม่มี (เพื่อ backward compatibility)
            if 'Province' not in df.columns and 'จังหวัด' in df.columns:
                df['Province'] = df['จังหวัด']
    
    return df.reset_index(drop=True)

def predict_trips(test_df, model_data, punthai_buffer=1.0, maxmart_buffer=1.10, fleet_limits=None, max_qty_per_trip=0):
    import traceback as _tb_mod
    try:
        return _predict_trips_inner(test_df, model_data, punthai_buffer, maxmart_buffer, fleet_limits, max_qty_per_trip)
    except Exception as _e:
        import traceback as _tb2
        _msg = _tb2.format_exc()
        safe_print(f"❌ predict_trips ERROR:\n{_msg}")
        raise

def _predict_trips_inner(test_df, model_data, punthai_buffer=1.0, maxmart_buffer=1.10, fleet_limits=None, max_qty_per_trip=0):
    """
    จัดทริปแบบใหม่ - เรียบง่ายและมีประสิทธิภาพ
    
    หลักการ:
    1. เรียงตาม: ภาค → จังหวัด → อำเภอ → ตำบล → Route (ใช้ระยะทางจากพิกัดจริง)
    2. จับกลุ่ม Route เดียวกัน รวมน้ำหนักไว้ด้วยกัน
    3. เรียงจากไกลมาใกล้ (จาก DC)
    4. ตัดเป็นทริปตามน้ำหนัก/คิวของรถแต่ละประเภท
    5. ใช้ BUFFER ตาม BU (ตรวจจากชื่อสาขา)
    
    Args:
        test_df: DataFrame ข้อมูลสาขาที่จะจัดทริป
        model_data: ข้อมูลโมเดล (branch_vehicles, etc.)
        punthai_buffer: Buffer สำหรับ Punthai (เช่น 1.0 = 100%)
        maxmart_buffer: Buffer สำหรับ Maxmart/ผสม (เช่น 1.10 = 110%)
    """
    branch_vehicles = model_data.get('branch_vehicles', {})

    # pre-build MASTER_DATA dict สำหรับ O(1) lookup แทน O(N) scan ในลูป
    _master_dict: dict = {}
    if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
        for _, _mr in MASTER_DATA.iterrows():
            _mc = str(_mr.get('Plan Code', '')).strip().upper()
            if _mc:
                _master_dict[_mc] = _mr

    # ==========================================
    # Step 1: สร้าง location_map จากข้อมูล MASTER_DATA (Google Sheets) + พิกัด
    # ==========================================
    location_map = {}  # {code: {province, district, subdistrict, route, lat, lon, distance_from_dc, region_name}}
    
    for _, row in test_df.iterrows():
        code = str(row.get('Code', '')).strip().upper()
        if not code:
            continue
        
        province = ''
        district = ''
        subdistrict = ''
        route = ''
        
        # ── Priority 1: MASTER DATA (source of truth) ──
        _mrow = _master_dict.get(code)
        if _mrow is not None:
            province    = str(_mrow.get('จังหวัด', '')).strip() if pd.notna(_mrow.get('จังหวัด')) else ''
            district    = str(_mrow.get('อำเภอ', '')).strip()   if pd.notna(_mrow.get('อำเภอ'))   else ''
            subdistrict = str(_mrow.get('ตำบล', '')).strip()    if pd.notna(_mrow.get('ตำบล'))    else ''
            route       = str(_mrow.get('Route', '')).strip()   if pd.notna(_mrow.get('Route'))   else ''

        # ── Priority 2: Excel upload (fallback เฉพาะที่ Master ว่าง) ──
        if not province:    province    = str(row.get('Province', row.get('จังหวัด', ''))).strip() if pd.notna(row.get('Province', row.get('จังหวัด', ''))) else ''
        if not district:    district    = str(row.get('District', row.get('อำเภอ', ''))).strip()   if pd.notna(row.get('District', row.get('อำเภอ', '')))   else ''
        if not subdistrict: subdistrict = str(row.get('Subdistrict', row.get('ตำบล', ''))).strip() if pd.notna(row.get('Subdistrict', row.get('ตำบล', ''))) else ''
        if not route:       route       = str(row.get('Route', '')).strip()                        if pd.notna(row.get('Route', ''))                        else ''

        # normalize province alias
        _prov_alias = {'พระนครศรีอยุธยา':'อยุธยา','กรุงเทพฯ':'กรุงเทพมหานคร',
                       'กทม':'กรุงเทพมหานคร','กทม.':'กรุงเทพมหานคร','โคราช':'นครราชสีมา'}
        province    = _prov_alias.get(clean_name(province), clean_name(province))
        district    = clean_name(district)
        subdistrict = clean_name(subdistrict)

        # ── พิกัด Priority 1: Master (O(1) dict lookup) ──
        lat = 0; lon = 0
        if _mrow is not None:
            try: lat = float(_mrow.get('ละติจูด', 0) or 0)
            except: lat = 0
            try: lon = float(_mrow.get('ลองติจูด', 0) or 0)
            except: lon = 0

        # ── พิกัด Priority 2: Excel upload ──
        if lat == 0 or lon == 0:
            for lat_col in ['Latitude', 'latitude', 'ละติจูด', 'lat']:
                if lat_col in row and pd.notna(row[lat_col]):
                    try: lat = float(row[lat_col]); break
                    except: pass
            for lon_col in ['Longitude', 'longitude', 'ลองจิจูด', 'ลองติจูด', 'lon', 'long']:
                if lon_col in row and pd.notna(row[lon_col]):
                    try: lon = float(row[lon_col]); break
                    except: pass
        
        # ระยะทางจาก DC (haversine จากพิกัด)
        dist_from_dc = haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon, use_osrm_cache=False) if (lat and lon) else 9999

        # 🗺️ GPS fallback: ถ้าไม่มีจังหวัด แต่มีพิกัด → หาสาขาใกล้สุดใน BRANCH_INFO
        if not province and lat and lon and BRANCH_INFO:
            _best_d, _best_info = 9999, None
            for _bi in BRANCH_INFO.values():
                _blat = _bi.get('lat', 0); _blon = _bi.get('lon', 0)
                _bprov = _bi.get('province', '')
                if not _bprov or not _blat or not _blon:
                    continue
                _d = abs(lat - _blat) + abs(lon - _blon)  # fast Manhattan pre-filter
                if _d < _best_d:
                    _best_d = _d
                    _best_info = _bi
            if _best_info:
                province    = _best_info.get('province', '')
                district    = district    or _best_info.get('district', '')
                subdistrict = subdistrict or _best_info.get('subdistrict', '')

        location_map[code] = {
            'province': province,
            'district': district,
            'subdistrict': subdistrict,
            'route': route,
            'lat': lat,
            'lon': lon,
            'distance_from_dc': dist_from_dc,
            'region_name': get_region_name(province)
        }
    
    # ==========================================
    # Step 2: เพิ่มข้อมูลพื้นที่ให้แต่ละสาขา (pd.merge แบบ manual)
    # ==========================================
    df = test_df.copy()
    
    def get_location_info(code):
        code_upper = str(code).strip().upper()
        return location_map.get(code_upper, {
            'province': '', 'district': '', 'subdistrict': '', 'route': '',
            'lat': 0, 'lon': 0,
            'distance_from_dc': 9999,
            'region_name': 'ไม่ระบุ'
        })
    
    # เพิ่มคอลัมน์ข้อมูลพื้นที่
    _loc_cache = {str(c).strip().upper(): get_location_info(c) for c in df['Code'].unique()}
    df['_region_name']      = df['Code'].map(lambda c: _loc_cache.get(str(c).strip().upper(), {}).get('region_name', 'ไม่ระบุ'))
    # _loc_cache มาจาก location_map ซึ่ง Master-first แล้ว — ใช้ตรงได้เลย
    df['_province']    = df['Code'].map(lambda c: _loc_cache.get(str(c).strip().upper(), {}).get('province', ''))
    df['_district']    = df['Code'].map(lambda c: _loc_cache.get(str(c).strip().upper(), {}).get('district', ''))
    df['_subdistrict'] = df['Code'].map(lambda c: _loc_cache.get(str(c).strip().upper(), {}).get('subdistrict', ''))
    df['_route']            = df['Code'].map(lambda c: _loc_cache.get(str(c).strip().upper(), {}).get('route', ''))
    df['_distance_from_dc'] = df['Code'].map(lambda c: _loc_cache.get(str(c).strip().upper(), {}).get('distance_from_dc', 9999))
    df['_lat']              = df['Code'].map(lambda c: _loc_cache.get(str(c).strip().upper(), {}).get('lat', 0))
    df['_lon']              = df['Code'].map(lambda c: _loc_cache.get(str(c).strip().upper(), {}).get('lon', 0))
    
    # 🎯 คำนวณ Bearing (ทิศทาง) จาก DC เพื่อจัดกลุ่มสาขาที่อยู่ทิศเดียวกัน
    DC_LAT = 14.117451
    DC_LON = 100.633408
    
    def calc_bearing(row):
        if row['_lat'] > 0 and row['_lon'] > 0:
            return calculate_bearing(DC_LAT, DC_LON, row['_lat'], row['_lon'])
        return 0
    
    df['_bearing_from_dc'] = df.apply(calc_bearing, axis=1)
    df['_bearing_zone'] = df['_bearing_from_dc'].apply(get_bearing_zone)
    
    # 🚨 เพิ่ม Logistics Zone สำหรับ routing ตามทางหลวง
    # 🗺️ ลำดับความสำคัญ:
    #   1. BRANCH_ZONES_CACHE (จาก zone_viewer.py) — ถ้ามี branch_zones.json
    #   2. PROVINCE_ZONE_MAP fallback — ใช้ logic เดียวกับ zone_viewer.py
    def _get_zone_for_row(row):
        """คืน per-district zone: ZONE_{จังหวัด}_{อำเภอ} เสมอ"""
        prov = str(row.get('_province', '') or '').strip()
        dist = str(row.get('_district',  '') or '').strip()
        for _alias, _full in [("กรุงเทพฯ","กรุงเทพมหานคร"),("กทม","กรุงเทพมหานคร"),("กทม.","กรุงเทพมหานคร"),("โคราช","นครราชสีมา")]:
            if prov == _alias: prov = _full; break
        return get_logistics_zone(prov, dist) or f'ZONE_{prov}' if prov else 'ZONE_ไม่ระบุ'

    df['_logistics_zone'] = df.apply(_get_zone_for_row, axis=1)
    df['_zone_priority'] = df['_logistics_zone'].apply(get_zone_priority)
    df['_zone_highway'] = df['_logistics_zone'].apply(get_zone_highway)
    # 🎯 Province Zone (zone_viewer.py system) — ใช้ป้องกันกระโดดข้ามจังหวัด
    df['_prov_zone'] = df.apply(
        lambda row: get_prov_zone(row['_province'], row['_district']), axis=1
    )
    
    # ==========================================
    # Step 3: เรียงลำดับแบบ Hierarchical (Zone Priority > Region > Province Max Dist > District Max Dist > Distance)
    # 🎯 หัวใจสำคัญ: เรียงตาม Region Order ก่อน (ไกลมาใกล้)
    # ==========================================
    
    # เพิ่ม Region Order สำหรับ sorting
    df['_region_order'] = df['_region_name'].map(REGION_ORDER).fillna(99)
    
    # คำนวณ Province Max Distance (จังหวัดไหนมีจุดไกลสุดมาก่อน)
    prov_max_dist = df.groupby('_province')['_distance_from_dc'].max().reset_index()
    prov_max_dist.columns = ['_province', '_prov_max_dist']
    df = df.merge(prov_max_dist, on='_province', how='left')
    df['_prov_max_dist'] = df['_prov_max_dist'].fillna(9999)
    
    # คำนวณ District Max Distance (อำเภอไหนมีจุดไกลสุดมาก่อน)
    dist_max_dist = df.groupby(['_province', '_district'])['_distance_from_dc'].max().reset_index()
    dist_max_dist.columns = ['_province', '_district', '_dist_max_dist']
    df = df.merge(dist_max_dist, on=['_province', '_district'], how='left')
    df['_dist_max_dist'] = df['_dist_max_dist'].fillna(9999)
    
    # คำนวณ Subdistrict Max Distance (ตำบลไหนมีจุดไกลสุด — แทน sum_code)
    subdist_max_dist = df.groupby(['_province', '_district', '_subdistrict'])['_distance_from_dc'].max().reset_index()
    subdist_max_dist.columns = ['_province', '_district', '_subdistrict', '_subdist_max_dist']
    df = df.merge(subdist_max_dist, on=['_province', '_district', '_subdistrict'], how='left')
    df['_subdist_max_dist'] = df['_subdist_max_dist'].fillna(9999)
    
    # 🎯 Sort: LIFO (Last In First Out) - ไกลส่งก่อน, ใกล้ส่งทีหลัง
    # Zone → Region → Province Max Dist → District Max Dist → Subdistrict Max Dist → Route → Distance
    df = df.sort_values(
        ['_zone_priority', '_region_order', '_prov_max_dist', '_dist_max_dist', '_subdist_max_dist', '_route', '_distance_from_dc'],
        ascending=[True, True, False, False, False, True, False]
    ).reset_index(drop=True)
    
    # ==========================================
    # Step 4: จับกลุ่ม Route เดียวกัน รวมน้ำหนัก
    # ==========================================
    # สร้าง grouping key จาก route (ถ้ามี) หรือ ตำบล+อำเภอ+จังหวัด
    def get_group_key(row):
        route = row['_route']
        if route and route.strip():
            return f"R_{route}"
        # ถ้าไม่มี route ใช้ รหัสตำบล (เรียงตามระยะทาง)
        return f"L_{row['_subdistrict']}_{row['_district']}_{row['_province']}"
    
    df['_group_key'] = df.apply(get_group_key, axis=1)
    
    # ==========================================
    # Step 5: หารถที่เหมาะสมจากข้อจำกัดสาขา + Central Region Rule
    # ==========================================
    def get_max_vehicle_for_code(code):
        """หารถที่ใหญ่ที่สุดที่สาขาสามารถใช้ได้ - อ่านจาก Sheets"""
        max_vehicle = get_max_vehicle_for_branch(code, test_df=test_df)
        return max_vehicle
    
    def get_allowed_vehicles_for_region(region_name):
        """หารถที่ใช้ได้ (อิงตาม Master data เท่านั้น)"""
        return ['4W', 'JB', '6W']  # All vehicles - restrictions from Master data only
    
    df['_max_vehicle'] = df['Code'].apply(get_max_vehicle_for_code)
    df['_region_allowed_vehicles'] = df['_region_name'].apply(get_allowed_vehicles_for_region)
    
    # 🎯 สร้าง Vehicle Priority: สาขา 4W = 1 (จัดก่อน), JB = 2, 6W = 3 (จัดทีหลัง)
    vehicle_priority_map = {'4W': 1, 'JB': 2, '6W': 3}
    df['_vehicle_priority'] = df['_max_vehicle'].map(vehicle_priority_map).fillna(3)
    
    # 🎯 Sort: ใช้ PROVINCE ZONE เป็นหลัก (ระบบเดียวกับ zone_viewer.py)
    # หลักการ: สาขาในจังหวัด/ภาคเดียวกันอยู่ติดกัน ป้องกันกระโดดข้ามจังหวัด
    # 1. Province Zone (BKK_เขต / ภาค_จังหวัด) — สาขากลุ่มเดียวกันอยู่ติดกัน
    # 2. ระยะทางจาก DC (ไกลมาใกล้) - LIFO
    # 3. จังหวัด/อำเภอ/ตำบล - จัดกลุ่มในพื้นที่เดียวกัน
    df = df.sort_values(
        [
            '_prov_zone',           # 1. 🗺️ Province Zone — BKK_เขต หรือ ภาค_จังหวัด
            '_distance_from_dc',    # 2. ระยะทางไกลก่อน (LIFO)
            '_province',            # 3. จังหวัดเดียวกัน
            '_district',            # 4. อำเภอเดียวกัน
            '_subdistrict',         # 5. ตำบลเดียวกัน
            '_vehicle_priority'     # 6. ข้อจำกัดรถ (secondary)
        ],
        ascending=[
            True,   # Province Zone เรียง A-Z (จัดกลุ่มจังหวัดเดียวกัน)
            False,  # ไกลมาใกล้ (LIFO)
            True,   # จังหวัดเรียง A-Z
            True,   # อำเภอเรียง A-Z
            True,   # ตำบลเรียง A-Z
            True    # ข้อจำกัดมากก่อน
        ]
    ).reset_index(drop=True)
    
    safe_print(f"📊 DEBUG: Province zones = {df['_prov_zone'].unique().tolist()}")
    
    # ==========================================
    # Step 6: DISTRICT CLUSTERING ALLOCATION (OPTIMIZED)
    # จัดทริปตาม District Buckets พร้อม Split เมื่อเกิน
    # ==========================================
    trip_counter = 1
    df['Trip'] = 0
    
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    
    # 🚀 CACHE: Pre-compute branch constraints และ BU type จาก Excel upload
    branch_max_vehicle_cache = {}
    branch_bu_cache = {}
    for _, row in df.iterrows():
        code = row['Code']
        branch_max_vehicle_cache[code] = row['_max_vehicle']
        
        # 📊 ดึง BU จากไฟล์ Excel ที่ upload มา (คอลัมน์ BU)
        bu = row.get('BU', None)  # อ่านค่าจริงจาก Excel
        
        # เช็คว่าเป็น Punthai หรือไม่
        is_punthai = False
        if bu is not None:
            bu_str = str(bu).strip().upper()
            # เช็ค BU = 211 หรือ 'PUNTHAI'
            is_punthai = bu_str in ['211', 'PUNTHAI']
        else:
            # ถ้าไม่มีคอลัมน์ BU → ลองเช็คจากชื่อสาขา (fallback)
            name = str(row.get('Name', '')).upper()
            is_punthai = 'PUNTHAI' in name or 'PUN-' in name
        
        branch_bu_cache[code] = is_punthai
    
    # 🚀 Pre-compute limits with buffer
    def get_max_limits(allowed_vehicles, is_punthai):
        """หา capacity สูงสุดที่ใช้ได้"""
        buffer_mult = punthai_buffer if is_punthai else maxmart_buffer
        max_vehicle = '6W' if '6W' in allowed_vehicles else ('JB' if 'JB' in allowed_vehicles else '4W')
        limits_to_use = PUNTHAI_LIMITS if is_punthai else LIMITS
        lim = limits_to_use.get(max_vehicle, LIMITS['6W'])
        return {
            'max_w': lim.get('max_w', 6000),           # น้ำหนัก = hard limit ไม่ใช้ buffer
            'max_c': lim.get('max_c', 20.0) * buffer_mult,  # คิ้ว = ยืดได้ตาม buffer
            'max_d': lim.get('max_drops', 12)
        }
    
    # Helper function: เลือกรถที่เหมาะสม (STRICT - บังคับข้อจำกัด)
    def select_vehicle_for_load(weight, cube, drops, is_punthai, allowed_vehicles, strict_constraint=True):
        """
        เลือกรถที่เหมาะสมตามโหลดและข้อจำกัด
        
        Logic: ใช้ buffer จากหน้าเว็บ (punthai_buffer, maxmart_buffer)
        - Punthai: buffer = 100% (ห้ามเกิน)
        - Maxmart: buffer = 110% (เกินได้ 10%)
        - strict_constraint=True: ห้ามใช้รถที่ใหญ่กว่า allowed_vehicles (บังคับข้อจำกัด)
        """
        buffer_mult = punthai_buffer if is_punthai else maxmart_buffer
        limits_to_use = PUNTHAI_LIMITS if is_punthai else LIMITS
        
        # 🚨 STRICT MODE: ใช้เฉพาะรถที่อนุญาตเท่านั้น ห้ามเกิน
        vehicle_rank = {'4W': 1, 'JB': 2, '6W': 3}
        max_allowed_rank = max([vehicle_rank[v] for v in allowed_vehicles if v in vehicle_rank], default=3)
        
        # เรียงจากเล็กไปใหญ่ แต่ห้ามเกิน max_allowed
        vehicle_order = ['4W', 'JB', '6W']
        
        for v in vehicle_order:
            # ข้ามรถที่ใหญ่กว่าที่อนุญาต
            if strict_constraint and vehicle_rank.get(v, 3) > max_allowed_rank:
                continue
            
            if v not in allowed_vehicles:
                continue
                
            lim = limits_to_use[v]
            
            # น้ำหนัก = hard limit, คิ้ว = ยืดได้ตาม buffer
            if (weight <= lim['max_w'] and
                cube <= lim['max_c'] * buffer_mult and
                drops <= lim.get('max_drops', 12)):
                return v
        
        return None
    
    # Helper function: เช็ค Geographic Spread ภายในทริป
    def check_intra_trip_spread(trip_codes_list):
        """ตรวจสอบว่าสาขาในทริปไม่กระจายทางภูมิศาสตร์เกินไป (ห้ามคนละทิศ)"""
        if len(trip_codes_list) < 2:
            return True  # 1 สาขา = OK
        
        trip_df = df[df['Code'].isin(trip_codes_list)]
        if trip_df.empty:
            return True
        
        # คำนวณ centroid ของทริป
        trip_lat_mean = trip_df['_lat'].mean()
        trip_lon_mean = trip_df['_lon'].mean()
        
        # เช็คว่าทุกสาขาห่างจาก centroid ไม่เกิน 80km
        max_dist_from_center = 0
        for _, row in trip_df.iterrows():
            if row['_lat'] > 0 and row['_lon'] > 0:
                dist = haversine_distance(trip_lat_mean, trip_lon_mean, row['_lat'], row['_lon'], use_osrm_cache=False)
                max_dist_from_center = max(max_dist_from_center, dist)
        
        # ถ้า spread เกิน 80km ถือว่ากระจายเกินไป (คนละทิศ)
        return max_dist_from_center <= 80
    
    # Helper function: เช็คว่าเป็น Punthai ล้วนหรือไม่ (Optimized - ใช้ cache)
    def is_all_punthai_codes(codes):
        if not codes:
            return False
        return all(branch_bu_cache.get(c, False) for c in codes)
    
    # Helper function: หา allowed vehicles จาก codes (Optimized)
    def get_allowed_from_codes(codes, base_allowed):
        """หา allowed vehicles โดยรวม branch constraints"""
        result = set(base_allowed)
        for code in codes:
            branch_max = branch_max_vehicle_cache.get(code, '6W')
            if branch_max == 'JB':
                result.discard('6W')
            elif branch_max == '4W':
                result.discard('6W')
                result.discard('JB')
        return list(result)
    
    # Step 6.4: 🎯 ZONE-STRICT GREEDY - จัดทริปแบบแยกโซน + ห้ามข้ามโซน
    # หลักการ: ใช้ LOGISTICS_ZONES + NO_CROSS_ZONE_PAIRS
    # ==========================================
    safe_print("🎯 กำลังจัดทริปใหม่แบบ Zone-Strict (LOGISTICS_ZONES + NO_CROSS_ZONE_PAIRS)...")

    # ─── Runtime Nearby Groups (≤500m) ──────────────────────────────────────
    # สาขาที่ต้องอยู่ด้วยกัน = ≤500m เท่านั้น (ตรงกับ branch_groups.json)
    _NEARBY_GROUP_KM = 0.5   # รัศมีรวมกลุ่ม 500m
    _df_codes_upper = {str(c).strip().upper() for c in df['Code'].tolist()}

    # สร้าง _rt_same_loc จาก NEARBY_BRANCHES (pre-computed) + haversine fallback
    _rt_same_loc: dict = {}   # code_upper → [code_upper, ...]

    # Pass 1: ใช้ NEARBY_BRANCHES (เร็ว)
    for _nc in _df_codes_upper:
        if _nc in NEARBY_BRANCHES:
            _nearby_in_run = [
                nb_code for nb_code, nb_dist in NEARBY_BRANCHES[_nc]
                if nb_dist <= _NEARBY_GROUP_KM and nb_code in _df_codes_upper
            ]
            if _nearby_in_run:
                _rt_same_loc[_nc] = [_nc] + _nearby_in_run

    # Pass 2: สาขาที่ไม่อยู่ใน NEARBY_BRANCHES → ใช้ haversine จาก df พิกัด
    _df_coord_map = {}
    for _, _sr in df.iterrows():
        _slat = float(_sr.get('_lat', 0) or 0)
        _slon = float(_sr.get('_lon', 0) or 0)
        if _slat > 0 and _slon > 0:
            _df_coord_map[str(_sr['Code']).strip().upper()] = (_slat, _slon)

    for _nc in _df_codes_upper:
        if _nc not in _rt_same_loc and _nc in _df_coord_map:
            _nlat, _nlon = _df_coord_map[_nc]
            _nearby_hv = [
                _oc for _oc, (_olat, _olon) in _df_coord_map.items()
                if _oc != _nc and haversine_distance(_nlat, _nlon, _olat, _olon, use_osrm_cache=False) <= _NEARBY_GROUP_KM
            ]
            if _nearby_hv:
                _rt_same_loc[_nc] = [_nc] + _nearby_hv

    # Pass 3: Symmetry — ถ้า Y ∈ list[X] แต่ X ∉ list[Y] → เพิ่ม X เข้าไป
    # แก้กรณี Pass 1 (NEARBY_BRANCHES) เจอ X→Y แต่ Y อยู่ใน Pass 1 โดยไม่เจอ X
    # ป้องกันสาขาที่เดียวกันถูกแยกทริป เพราะ lookup ทิศทางเดียว
    for _nc in list(_rt_same_loc.keys()):
        for _nb in list(_rt_same_loc[_nc]):
            if _nb != _nc and _nb in _df_codes_upper:
                if _nb not in _rt_same_loc:
                    _rt_same_loc[_nb] = [_nb, _nc]
                elif _nc not in _rt_same_loc[_nb]:
                    _rt_same_loc[_nb].append(_nc)

    _rt_grp_count = sum(1 for v in _rt_same_loc.values() if len(v) > 1)
    safe_print(f"📍 Runtime nearby group (≤{_NEARBY_GROUP_KM:.0f}km): {_rt_grp_count} สาขามีเพื่อนร่วมทริป จาก {len(_df_codes_upper)} สาขา")

    # Pass 4: runtime co-location ≤200m — สาขาที่พิกัดเดียวกัน/ใกล้กันมาก (สำหรับสาขาใหม่ที่ไม่อยู่ใน branch_groups.json)
    # ใช้ haversine จาก _df_coord_map เพื่อหาคู่ที่แน่ใจว่าเป็น "จุดส่งเดียวกัน"
    _COLOC_STRICT_KM = 0.2   # 200m
    _coloc_200m: dict = {}   # code_upper → [other_code_upper, ...]
    if len(_df_coord_map) <= 3000:   # guard: ป้องกัน O(n²) ช้าสำหรับ run ใหญ่มาก
        _coord_items = list(_df_coord_map.items())
        for _i, (_nc, (_nlat, _nlon)) in enumerate(_coord_items):
            if _nlat <= 0 or _nlon <= 0:
                continue
            _near_strict = []
            for _oc, (_olat, _olon) in _coord_items:
                if _oc == _nc or _olat <= 0 or _olon <= 0:
                    continue
                if abs(_olat - _nlat) > 0.003:   # lat bound ≈300m → skip haversine
                    continue
                if haversine_distance(_nlat, _nlon, _olat, _olon, use_osrm_cache=False) <= _COLOC_STRICT_KM:
                    _near_strict.append(_oc)
            if _near_strict:
                _coloc_200m[_nc] = _near_strict
        _coloc_rt_count = sum(1 for v in _coloc_200m.values() if v)
        safe_print(f"📍 Runtime co-location (≤{_COLOC_STRICT_KM*1000:.0f}m): {_coloc_rt_count} สาขามีคู่พิกัดเดียวกัน")
    else:
        safe_print(f"⚠️ Skip Pass 4 co-location (run ใหญ่ {len(_df_coord_map)} สาขา) → ใช้ precomputed groups")

    # Pass 5: same-ตำบล/อำเภอ ≤5km — สาขาในพื้นที่เดียวกันและใกล้กันต้องไปทริปเดียวกัน
    # สร้าง _code_area_map ก่อน (fast lookup)
    _COLOC_AREA_KM = 5.0
    _code_area_map: dict = {}   # code_upper → (subdistrict, district)
    for _, _car in df.iterrows():
        _ccu = str(_car['Code']).strip().upper()
        _code_area_map[_ccu] = (
            str(_car.get('_subdistrict', '') or ''),
            str(_car.get('_district',    '') or ''),
        )
    # จัดกลุ่มตาม (subdistrict, district) เพื่อ lookup เร็ว
    _area_to_codes: dict = {}   # (sub, dis) → [code_upper]
    for _ccu, (_csub, _cdis) in _code_area_map.items():
        if not _csub and not _cdis:
            continue
        _akey = (_csub, _cdis)
        _area_to_codes.setdefault(_akey, []).append(_ccu)
    # สร้าง _coloc_subdist
    _coloc_subdist: dict = {}
    for _nc, (_nlat, _nlon) in _df_coord_map.items():
        if _nlat <= 0 or _nlon <= 0:
            continue
        _nsub, _ndis = _code_area_map.get(_nc, ('', ''))
        if not _nsub and not _ndis:
            continue
        _akey = (_nsub, _ndis)
        _near_area = []
        for _oc in _area_to_codes.get(_akey, []):
            if _oc == _nc:
                continue
            _olat, _olon = _df_coord_map.get(_oc, (0, 0))
            if _olat <= 0 or _olon <= 0:
                continue
            if abs(_olat - _nlat) > 0.06:   # lat bound ≈6km
                continue
            if haversine_distance(_nlat, _nlon, _olat, _olon, use_osrm_cache=False) <= _COLOC_AREA_KM:
                _near_area.append(_oc)
        if _near_area:
            _coloc_subdist[_nc] = _near_area
    _subdist_count = sum(1 for v in _coloc_subdist.values() if v)
    safe_print(f"📍 Same-ตำบล/อำเภอ (≤{_COLOC_AREA_KM:.0f}km): {_subdist_count} สาขามีเพื่อนตำบลเดียวกัน")

    def get_group_branches_rt(code: str) -> list:
        """รวม precomputed group (≤200m) + runtime co-location (≤200m) + runtime nearby (≤10km)"""
        code_upper = str(code).strip().upper()
        # precomputed group (จาก branch_groups.json)
        grp = list(get_group_branches(code_upper))
        grp_upper = {str(c).strip().upper() for c in grp}
        # runtime co-location ≤200m (สาขาใหม่ที่ไม่อยู่ใน precomputed)
        for _cl_c in _coloc_200m.get(code_upper, []):
            if _cl_c not in grp_upper and _cl_c in _df_codes_upper:
                grp.append(_cl_c)
                grp_upper.add(_cl_c)
        # สาขาในรัศมี 10km (runtime trip grouping)
        for _rt_c in _rt_same_loc.get(code_upper, []):
            if _rt_c not in grp_upper:
                grp.append(_rt_c)
                grp_upper.add(_rt_c)
        return grp

    # 🧠 AI LEARNING: โหลด pair_freq สำหรับ affinity boost
    _ai_pair_freq = load_trip_history()
    _ai_active = bool(_ai_pair_freq)
    if _ai_active:
        safe_print(f"🧠 AI Learning: โหลด {len(_ai_pair_freq):,} pair records")

    # _logistics_zone, _zone_priority, _zone_highway คำนวณไว้แล้วตั้งแต่ Step 2

    # 2️⃣ เรียงลำดับตาม zone priority (priority ต่ำ = ไกล DC = จัดก่อน)
    df = df.sort_values(['_zone_priority', '_distance_from_dc'], ascending=[True, False]).reset_index(drop=True)
    
    # 🚗 คำนวณ MaxVehicle ของแต่ละสาขา (4W=1, JB=2, 6W=3)
    vehicle_rank = {'4W': 1, 'JB': 2, '6W': 3}
    df['_max_vehicle'] = df['Code'].apply(lambda c: get_max_vehicle_for_branch(c))
    df['_vehicle_rank'] = df['_max_vehicle'].map(vehicle_rank).fillna(3).astype(int)
    
    # รีเซ็ตทริปทั้งหมด
    df['Trip'] = 0
    trip_counter = 1

    # ─── INDEX: code_upper → df index (O(1) lookup แทน df.loc ตลอด algorithm) ───
    _df_idx: dict = {str(r['Code']).strip().upper(): idx for idx, r in df[['Code']].iterrows()}

    # ═══════════════════════════════════════════════════════════════════
    # 🔒 ABSOLUTE GROUP LOCK — กลุ่มสาขาจาก branch_groups.json
    #    ห้ามแยกทริปในทุกกรณี ไม่มีข้อยกเว้น
    #
    #    assign_to_trip(code, trip_num):
    #      → assign code + สมาชิกกลุ่มทั้งหมดเข้า trip_num ทันที
    #      → ลบออกจาก unassigned ทุกคน
    # ═══════════════════════════════════════════════════════════════════
    def _assign_with_group(code: str, trip_num: int,
                           trip_codes_list: list,
                           extra_w: list, extra_c: list, extra_qty: list):
        """Assign code + group members ทั้งหมดเข้า trip_num ทันที ไม่มีข้อยกเว้น"""
        _todo = [str(code).strip().upper()]
        _done: set = set()
        while _todo:
            _cu = _todo.pop()
            if _cu in _done:
                continue
            _done.add(_cu)
            _idx = _df_idx.get(_cu)
            if _idx is None:
                continue
            _actual = df.at[_idx, 'Code']
            _cur_trip = safe_int_trip(df.at[_idx, 'Trip'])
            if _cur_trip == trip_num:
                pass  # อยู่แล้ว — แต่ยังต้องตรวจ group members
            else:
                df.at[_idx, 'Trip'] = trip_num
                if _actual not in trip_codes_list:
                    trip_codes_list.append(_actual)
                    extra_w.append(float(df.at[_idx, 'Weight'] or 0))
                    extra_c.append(float(df.at[_idx, 'Cube'] or 0))
                    extra_qty.append(safe_qty(df.at[_idx, 'OriginalQty']))
                # ลบออกจาก unassigned
                if _actual in unassigned:
                    unassigned.discard(_actual)
                else:
                    for _u in list(unassigned):
                        if str(_u).strip().upper() == _cu:
                            unassigned.discard(_u)
                            break
            # เพิ่ม group members ทั้งหมดเข้า todo
            _gid = BRANCH_TO_GROUP.get(_cu)
            if _gid:
                for _sib in BRANCH_GROUPS.get(_gid, []):
                    _sib_up = str(_sib).strip().upper()
                    if _sib_up not in _done:
                        _todo.append(_sib_up)

    # สร้าง set ของสาขาที่ยังไม่ได้จัด
    unassigned = set(df['Code'].tolist())

    # 🔒 PRE-CLUSTER: สาขาที่มีกลุ่ม branch_groups → ยึดไว้ด้วยกันก่อน planning
    # เลือก "anchor" 1 ตัวต่อกลุ่ม, ซ่อน "followers" ออกจาก unassigned
    # เมื่อ anchor ถูกเพิ่มเข้าทริป → followers ตามเข้าทันที (ไม่มีข้อยกเว้น)
    _anchor_followers: dict = {}  # anchor_code_upper → [follower_actual_code, ...]
    _follower_of: dict = {}       # follower_code_upper → anchor_code_upper
    _seen_groups: set = set()

    # pre-build code→actual_code lookup เพื่อความเร็ว
    _code_actual: dict = {str(r['Code']).strip().upper(): r['Code'] for _, r in df.iterrows()}

    for _pu in list(unassigned):
        _pu_up = str(_pu).strip().upper()
        _gid = BRANCH_TO_GROUP.get(_pu_up)
        if not _gid or _gid in _seen_groups:
            continue
        _seen_groups.add(_gid)
        _grp_all = [str(c).strip().upper() for c in BRANCH_GROUPS.get(_gid, [])]
        # กรองเฉพาะที่อยู่ใน order (df) และยัง unassigned
        _grp_in_run = [c for c in _grp_all if c in _code_actual and
                       any(str(u).strip().upper() == c for u in unassigned)]
        if len(_grp_in_run) <= 1:
            continue
        # anchor = member แรกใน df (ตาม index df)
        _anchor_up = None
        for _, _dr in df.iterrows():
            _dcu = str(_dr['Code']).strip().upper()
            if _dcu in _grp_in_run:
                _anchor_up = _dcu
                break
        if not _anchor_up:
            _anchor_up = _grp_in_run[0]
        _followers = [c for c in _grp_in_run if c != _anchor_up]
        _anchor_followers[_anchor_up] = [_code_actual[f] for f in _followers]
        for _f_up in _followers:
            _follower_of[_f_up] = _anchor_up
            # ซ่อน follower ออกจาก unassigned
            for _u in list(unassigned):
                if str(_u).strip().upper() == _f_up:
                    unassigned.remove(_u)
                    break

    _pre_cluster_count = sum(len(v) for v in _anchor_followers.values())
    safe_print(f"🔒 Pre-cluster: {len(_anchor_followers)} anchors, {_pre_cluster_count} followers ถูกล็อคไว้")

    def _add_followers(anchor_code, trip_codes):
        """เพิ่ม followers ของ anchor เข้าทริปทันที และลบออกจาก unassigned เพื่อไม่ให้เป็น seed ทริปใหม่"""
        _a_up = str(anchor_code).strip().upper()
        _flist = _anchor_followers.get(_a_up, [])
        _added_w = _added_c = 0.0; _added_qty = 0
        for _fact in _flist:
            _f_up = str(_fact).strip().upper()
            if any(str(tc).strip().upper() == _f_up for tc in trip_codes):
                continue
            _f_rows = df[df['Code'].apply(lambda x: str(x).strip().upper() == _f_up)]
            if _f_rows.empty:
                continue
            trip_codes.append(_fact)
            # ลบ follower ออกจาก unassigned ป้องกันถูกเลือกเป็น seed ทริปใหม่
            if _fact in unassigned:
                unassigned.remove(_fact)
            else:
                for _u in list(unassigned):
                    if str(_u).strip().upper() == _f_up:
                        unassigned.remove(_u)
                        break
            _added_w   += float(_f_rows.iloc[0].get('Weight', 0) or 0)
            _added_c   += float(_f_rows.iloc[0].get('Cube', 0) or 0)
            _added_qty += safe_qty(_f_rows.iloc[0].get('OriginalQty', 0))
            safe_print(f"      🔒 LOCK: {_fact} (follower of {anchor_code})")
        return _added_w, _added_c, _added_qty


    # ==========================================
    # 3 OR-Tools VRP (100%) -- Dynamic Routing
    # ==========================================
    safe_print('🤖 OR-Tools VRP: โยนพิกัดทั้งหมดรอบเดียว...')
    try:
        from ortools_vrp import solve_vrp_by_province
        _ortools_ok = True
    except Exception as _ort_err:
        safe_print(f'   ⚠️ import ortools_vrp ล้มเหลว: {_ort_err}')
        _ortools_ok = False

    if _ortools_ok:
        # ดึง buffer ที่ถูกต้อง
        _ort_buf = punthai_buffer if all(
            str(r.get('BU','')).upper() in ['211','PUNTHAI']
            for _, r in df.iterrows()
        ) else maxmart_buffer

        # ดึง max_vehicles จาก fleet_limits (ถ้ามี)
        _ort_max_veh = None
        if fleet_limits:
            _ort_max_veh = {
                '4W': fleet_limits.get('4W', 30),
                'JB': fleet_limits.get('JB', 20),
                '6W': fleet_limits.get('6W', 30),
            }

        _ort_result = solve_vrp_by_province(
            df.copy(),
            dc_lat=DC_WANG_NOI_LAT,
            dc_lon=DC_WANG_NOI_LON,
            buffer=_ort_buf,
            time_limit_sec=15,
            max_vehicles_per_type=_ort_max_veh,
        )

        # copy Trip กลับเข้า df
        df['Trip']  = _ort_result['Trip'].values
        df['Truck'] = _ort_result['Truck'].values

        # นับ trip_counter ต่อจาก OR-Tools
        _ort_max_trip = int(df['Trip'].max()) if df['Trip'].max() > 0 else 0
        trip_counter = _ort_max_trip + 1

        safe_print(f'   ✅ OR-Tools: {_ort_max_trip} ทริป ({len(df[df["Trip"]>0])} สาขา)')
    else:
        safe_print('   ⚠️ fallback: ไม่มีทริป (OR-Tools ใช้ไม่ได้)')
        trip_counter = 1

    safe_print(f'🎯 จัดทริปเสร็จ: {trip_counter - 1} ทริป')


    # ==========================================
    # Step 6.4.4: 🔋 FILL-UP PASS — เติมรถที่ยังไม่เต็มด้วยสาขาที่เหลือ
    # สาขาที่ยังไม่ได้จัด (unassigned) → ลองเพิ่มเข้าทริปที่ util < 70%
    # เฉพาะสาขาที่อยู่ในภาค/จังหวัดเดียวกัน และใกล้ทริปนั้น ≤ 60km
    # ==========================================
    # ==========================================
    # Step 6.4.4: 🔋 GROUP-AWARE BEST-FIT FILL-UP
    # หลักการ: เต็ม | ไปด้วยกัน | พอดี
    #   1) รวม unassigned branch เป็น "unit" ตาม branch_groups (group = 1 unit)
    #   2) เรียง unit ตาม weight มากสุด (best-fit decreasing)
    #   3) แต่ละ unit หาทริปที่ "พอดีที่สุด" (space เหลือน้อยสุดที่ยังจุได้) ภาคเดียวกัน
    #   4) วนซ้ำจนไม่มี unit เหลือ หรือไม่มีทริปรับ
    # ==========================================
    safe_print("🔋 Group-Aware Best-Fit Fill-Up...")
    _FILLUP_MAX_KM = 150.0
    _fu_code_idx   = {str(r['Code']).strip().upper(): idx
                      for idx, r in df[['Code']].iterrows()}
    _fillup_added  = 0

    def _fu_get_trip_state(trip_num):
        """คืน (codes, w, c, max_w, max_c, buf, lims, region, prov, coords) ของทริป"""
        _rows  = df[df['Trip'] == trip_num]
        if _rows.empty: return None
        _codes = _rows['Code'].tolist()
        _w     = float(_rows['Weight'].sum())
        _c     = float(_rows['Cube'].sum())
        _is_pt = all(branch_bu_cache.get(str(x).strip().upper(), False) for x in _codes)
        _buf   = punthai_buffer if _is_pt else maxmart_buffer
        _lims  = PUNTHAI_LIMITS if _is_pt else LIMITS
        _allow = get_allowed_from_codes(_codes, ['4W','JB','6W'])
        _veh   = next((v for v in ['4W','JB','6W'] if v in (_allow or ['6W'])), '6W')
        _lim   = _lims.get(_veh, _lims['6W'])
        _mw    = _lim['max_w'] * _buf
        _mc    = _lim['max_c'] * _buf
        _reg   = str(_rows.iloc[0].get('_region_name','') or '')
        # ใช้ dominant province (mode) ไม่ใช่แค่ iloc[0]
        _prov_series = _rows['_province'].dropna() if '_province' in _rows.columns else pd.Series(dtype=str)
        _prov  = str(_prov_series.mode().iloc[0]) if not _prov_series.empty else ''
        _coords= [(float(r['_lat'] or 0), float(r['_lon'] or 0))
                  for _, r in _rows.iterrows() if float(r.get('_lat',0) or 0) > 0]
        return {'codes':_codes,'w':_w,'c':_c,'max_w':_mw,'max_c':_mc,
                'buf':_buf,'lims':_lims,'allow':_allow,'veh':_veh,
                'region':_reg,'prov':_prov,'coords':_coords}

    _fu_rounds = 0
    while _fu_rounds < 8:
        _fu_rounds += 1
        _unassigned_df = df[df['Trip'] == 0].copy()
        if _unassigned_df.empty: break

        # ── สร้าง code→trip map และ units ──
        _code_trip_map = {str(r['Code']).strip().upper(): safe_int_trip(r['Trip'])
                          for _, r in df[['Code','Trip']].iterrows()}
        _seen_unit: set = set()
        # (total_w, total_c, [codes], clat, clon, reg, prov, forced_trip)
        # forced_trip != 0 = unit ต้องไปทริปนี้เท่านั้น (เพราะ member อื่นอยู่ที่นั่นแล้ว)
        _units: list   = []
        _una_codes_up  = {str(c).strip().upper() for c in _unassigned_df['Code']}

        for _, _ur in _unassigned_df.iterrows():
            _uc_up = str(_ur['Code']).strip().upper()
            if _uc_up in _seen_unit: continue

            _gid = BRANCH_TO_GROUP.get(_uc_up)
            if _gid:
                _gmems_up = [str(c).strip().upper() for c in BRANCH_GROUPS.get(_gid,[])]
                # แยก: members ที่ยัง unassigned vs ที่ถูก assign แล้ว
                _unit_codes_up  = [c for c in _gmems_up if c in _una_codes_up]
                _assigned_trips = {_code_trip_map[c] for c in _gmems_up
                                   if c in _code_trip_map and _code_trip_map[c] > 0}
                # ถ้า member อื่นอยู่ทริปไหนแล้ว → forced ไปทริปนั้น
                _forced_trip = next(iter(_assigned_trips)) if len(_assigned_trips) == 1 else 0
                if len(_assigned_trips) > 1:
                    # group แตกอยู่แล้ว → ข้ามไปให้ FINAL LOCK จัดการ
                    for _x in _unit_codes_up: _seen_unit.add(_x)
                    continue
            else:
                _unit_codes_up = [_uc_up]
                _forced_trip   = 0

            if not _unit_codes_up: continue
            for _x in _unit_codes_up: _seen_unit.add(_x)

            _unit_rows = _unassigned_df[
                _unassigned_df['Code'].apply(lambda x: str(x).strip().upper() in set(_unit_codes_up))
            ]
            _unit_w = float(_unit_rows['Weight'].sum())
            _unit_c = float(_unit_rows['Cube'].sum())
            _lats   = _unit_rows['_lat'].fillna(0).to_numpy(dtype=float)
            _lons   = _unit_rows['_lon'].fillna(0).to_numpy(dtype=float)
            _valid  = (_lats > 0) & (_lons > 0)
            _clat   = float(_lats[_valid].mean()) if _valid.any() else 0.0
            _clon   = float(_lons[_valid].mean()) if _valid.any() else 0.0
            _reg    = str(_unit_rows.iloc[0].get('_region_name','') or '')
            _prov   = str(_unit_rows.iloc[0].get('_province','') or '')
            _dist   = str(_unit_rows.iloc[0].get('_district','') or '')
            _sub    = str(_unit_rows.iloc[0].get('_subdistrict','') or '')
            _zone   = str(_unit_rows.iloc[0].get('_logistics_zone','') or '')
            _actual_codes = _unit_rows['Code'].tolist()
            _units.append((_unit_w, _unit_c, _actual_codes, _clat, _clon,
                           _reg, _prov, _dist, _sub, _zone, _forced_trip))

        if not _units: break
        # เรียง unit ตาม weight มากสุดก่อน (best-fit decreasing)
        _units.sort(key=lambda x: -x[0])

        # ── สร้าง trip state cache (เพิ่ม subdistricts/districts/zones) ──
        _trip_states = {}
        for _ft in df[df['Trip'] > 0]['Trip'].unique():
            _ts = _fu_get_trip_state(_ft)
            if not _ts: continue
            _ft_rows2 = df[df['Trip'] == _ft]
            _ts['subdists'] = set(_ft_rows2['_subdistrict'].dropna().unique()) if '_subdistrict' in _ft_rows2 else set()
            _ts['dists']    = set(_ft_rows2['_district'].dropna().unique())    if '_district'    in _ft_rows2 else set()
            _ts['zones']    = set(_ft_rows2['_logistics_zone'].dropna().unique()) if '_logistics_zone' in _ft_rows2 else set()
            _trip_states[_ft] = _ts

        _added_this_round = 0
        for _uw, _uc, _ucodes, _ulat, _ulon, _ureg, _uprov, _udist, _usub, _uzone, _forced_trip in _units:
            # ถ้า forced_trip → ต้องไปทริปนั้นเท่านั้น (group member อื่นอยู่ที่นั่นแล้ว)
            if _forced_trip > 0:
                _ts = _trip_states.get(_forced_trip)
                if _ts:
                    _tw = _ts['w'] + _uw; _tc = _ts['c'] + _uc
                    _vok = None
                    for _v in ['4W','JB','6W']:
                        if _v not in (_ts['allow'] or []): continue
                        _lim2 = _ts['lims'].get(_v, _ts['lims']['6W'])
                        if _tw <= _lim2['max_w']*_ts['buf'] and _tc <= _lim2['max_c']*_ts['buf']:
                            _vok = _v; break
                    if _vok:
                        _best_t = _forced_trip
                    else:
                        # ไม่พอดี — สร้างทริปใหม่สำหรับ group นี้ (ห้ามยัดเกิน capacity)
                        safe_print(f"   ⚠️ Forced-group overflow: unit {_ucodes} → สร้างทริปใหม่แทน Trip {_forced_trip}")
                        _best_t = None
                else:
                    _best_t = None
            else:
                # Zone-first Best-fit: zone เดียวกันก่อน แล้วค่อย gap น้อยสุด
                _best_t    = None
                _best_gap  = 9e9
                _best_zprio= 9
                _BKK       = 'กรุงเทพมหานคร'
                for _ft, _ts in _trip_states.items():
                    if (_ureg and _ts['region'] and
                            _ureg not in ('','ไม่ระบุ') and _ts['region'] not in ('','ไม่ระบุ') and
                            _ureg != _ts['region']):
                        continue
                    if ((_uprov == _BKK and _ts['prov'] != _BKK and _ts['prov']) or
                            (_ts['prov'] == _BKK and _uprov and _uprov != _BKK)):
                        continue
                    # ห้าม fill-up ข้ามจังหวัด — hard rule ก่อน distance
                    if _uprov and _ts['prov'] and _uprov != _ts['prov']:
                        continue
                    if _ulat > 0 and _ts['coords']:
                        _min_d = min(haversine_distance(_ulat,_ulon,lt,ln,use_osrm_cache=False)
                                     for lt,ln in _ts['coords'] if lt > 0)
                        _radius = 10.0 if (_uprov==_BKK or _ts['prov']==_BKK) else _FILLUP_MAX_KM
                        if _min_d > _radius: continue
                    _tw = _ts['w'] + _uw; _tc = _ts['c'] + _uc
                    _vok = None
                    for _v in ['4W','JB','6W']:
                        if _v not in (_ts['allow'] or []): continue
                        _lim2 = _ts['lims'].get(_v, _ts['lims']['6W'])
                        if _tw <= _lim2['max_w']*_ts['buf'] and _tc <= _lim2['max_c']*_ts['buf']:
                            _vok = _v; break
                    if not _vok: continue
                    # priority: 0=zone เดียวกัน (สูงสุด), 1=ตำบล, 2=อำเภอ, 3=จังหวัด, 4=อื่น
                    if _uzone and _ts.get('zones') and _uzone in _ts['zones']:   _zprio = 0
                    elif _usub  and _usub  in _ts.get('subdists', set()):         _zprio = 1
                    elif _udist and _udist in _ts.get('dists',    set()):         _zprio = 2
                    elif _uprov and _uprov == _ts['prov']:                        _zprio = 3
                    else:                                                          _zprio = 4
                    _gap = _ts['max_w'] - _tw
                    if (_zprio, _gap) < (_best_zprio, _best_gap):
                        _best_zprio = _zprio; _best_gap = _gap; _best_t = _ft

            if _best_t is None: continue

            # ✅ เพิ่ม unit ทั้งกลุ่มเข้า trip
            _ts = _trip_states[_best_t]
            for _uc_code in _ucodes:
                _uc_idx = _fu_code_idx.get(str(_uc_code).strip().upper())
                if _uc_idx is not None:
                    df.at[_uc_idx, 'Trip'] = _best_t
                else:
                    df.loc[df['Code'] == _uc_code, 'Trip'] = _best_t
                _ts['codes'].append(_uc_code)
            _ts['w'] += _uw; _ts['c'] += _uc
            _fillup_added += len(_ucodes); _added_this_round += len(_ucodes)
            _util_now = max(_ts['w']/_ts['max_w'], _ts['c']/_ts['max_c']) if _ts['max_w'] > 0 else 0
            safe_print(f"   🔋 Fill-up trip {_best_t}: +{_ucodes} util={_util_now*100:.0f}%")

        safe_print(f"   🔋 รอบ {_fu_rounds}: เพิ่ม {_added_this_round} สาขา")
        if _added_this_round == 0: break

    safe_print(f"🔋 Group-Aware Fill-Up: {_fu_rounds} รอบ, รวม {_fillup_added} สาขา")

    # ==========================================
    # Step 6.4.4b: 📍 SAME-COORDINATE FORCE MERGE
    # สาขาพิกัดเดียวกัน (≤50m) ในต่างทริป → รวมทริปเข้าด้วยกัน (ยอมเกิน capacity)
    # รวมถึงสาขาชื่อเดียวกันที่อยู่ห่างกัน ≤50m
    # ==========================================
    _SAME_COORD_KM = 0.05   # 50 เมตร
    safe_print("📍 SAME-COORDINATE FORCE MERGE: ตรวจสาขาพิกัดเดียวกันต่างทริป...")
    _samecoord_merged = 0
    _sc_changed = True
    while _sc_changed:
        _sc_changed = False
        # สร้าง {trip: [(code, lat, lon, name), ...]}
        _trip_coord_map2: dict = {}
        for _, _scr2 in df[df['Trip'] > 0].iterrows():
            _sc_t2 = safe_int_trip(_scr2['Trip'])
            _sc_lat2 = float(_scr2.get('_lat', 0) or 0)
            _sc_lon2 = float(_scr2.get('_lon', 0) or 0)
            _sc_name2 = str(_scr2.get('Name', '') or '').strip()
            if _sc_lat2 > 0 and _sc_lon2 > 0:
                if _sc_t2 not in _trip_coord_map2:
                    _trip_coord_map2[_sc_t2] = []
                _trip_coord_map2[_sc_t2].append((str(_scr2['Code']), _sc_lat2, _sc_lon2, _sc_name2))
        _trip_list2 = sorted(_trip_coord_map2.keys())
        for _i2, _ta2 in enumerate(_trip_list2):
            if _sc_changed: break
            for _tb2 in _trip_list2[_i2+1:]:
                if _sc_changed: break
                for (_ca2, _lat_a2, _lon_a2, _nm_a2) in _trip_coord_map2.get(_ta2, []):
                    if _sc_changed: break
                    for (_cb2, _lat_b2, _lon_b2, _nm_b2) in _trip_coord_map2.get(_tb2, []):
                        _d_sc2 = haversine_distance(_lat_a2, _lon_a2, _lat_b2, _lon_b2, use_osrm_cache=False)
                        # พิกัดใกล้กัน ≤50m หรือ ชื่อสาขาเดียวกัน + ≤200m
                        _name_match = (_nm_a2 and _nm_b2 and _nm_a2 == _nm_b2 and _d_sc2 <= 0.2)
                        if _d_sc2 <= _SAME_COORD_KM or _name_match:
                            _len_a2 = len(_trip_coord_map2.get(_ta2, []))
                            _len_b2 = len(_trip_coord_map2.get(_tb2, []))
                            _base_sc2  = _ta2 if _len_a2 >= _len_b2 else _tb2
                            _other_sc2 = _tb2 if _base_sc2 == _ta2 else _ta2
                            df.loc[df['Trip'] == _other_sc2, 'Trip'] = _base_sc2
                            _samecoord_merged += 1
                            safe_print(f"   📍 รวม trip {_other_sc2} → {_base_sc2} ({_ca2}↔{_cb2} ห่าง {_d_sc2*1000:.0f}m name={_name_match})")
                            _sc_changed = True
                            break
    if _samecoord_merged:
        safe_print(f"✅ same-coord merge: รวม {_samecoord_merged} ทริป")

    # ถ้าทริปใด มีแค่ 1 unique Code และสาขาอยู่ห่างจาก solo-trip อื่น ≤ 500m → รวม
    # ยอมเกิน limit ได้เพราะสาขาอยู่จุดเดียวกัน (ต้องส่งพร้อมกัน)
    # ==========================================
    _SOLO_MERGE_KM = 0.5   # 500 เมตร
    safe_print(f"🔀 ตรวจสอบ same-location solo trips (≤{_SOLO_MERGE_KM*1000:.0f}m)...")
    _sc_merged = 0
    _sc_iters = 0
    while _sc_iters < 100:
        _sc_iters += 1
        _changed = False
        # สร้าง {trip_num: list of rows}
        _solo_info: dict = {}   # {trip_num: (code, lat, lon, region, province, district, subdistrict)}
        for _tnum_sc in df[df['Trip'] > 0]['Trip'].unique():
            _tdf = df[df['Trip'] == _tnum_sc]
            _ucodes = _tdf['Code'].astype(str).str.strip().str.upper().unique().tolist()
            if len(_ucodes) == 1:   # solo trip
                _r0 = _tdf.iloc[0]
                _lat0  = float(_r0.get('_lat', 0) or 0)
                _lon0  = float(_r0.get('_lon', 0) or 0)
                _reg0  = str(_r0.get('_region_name', '') or '')
                _prov0 = str(_r0.get('_province', '') or '')
                _dist0 = str(_r0.get('_district', '') or '')
                _sub0  = str(_r0.get('_subdistrict', '') or '')
                _solo_info[_tnum_sc] = (_ucodes[0], _lat0, _lon0, _reg0, _prov0, _dist0, _sub0)
        if len(_solo_info) < 2:
            break
        _solo_list = sorted(_solo_info.items())   # [(trip_num, (...)), ...]
        _merged_this_round: set = set()
        for _i, (_ta, (_ca, _lata, _lona, _rega, _pova, _disa, _suba)) in enumerate(_solo_list):
            if _ta in _merged_this_round:
                continue
            for _tb, (_cb, _latb, _lonb, _regb, _povb, _disb, _subb) in _solo_list[_i+1:]:
                if _tb in _merged_this_round:
                    continue
                # ตรวจ ตำบล อำเภอ จังหวัด ภาค ต้องตรงกัน (ถ้ามีค่า)
                def _ne(a, b): return a and b and a not in ('', 'ไม่ระบุ') and b not in ('', 'ไม่ระบุ') and a != b
                if _ne(_rega, _regb): continue
                if _ne(_pova, _povb): continue
                if _ne(_disa, _disb): continue
                if _ne(_suba, _subb): continue
                # ตรวจระยะ
                if _lata <= 0 or _lona <= 0 or _latb <= 0 or _lonb <= 0:
                    continue
                _d = haversine_distance(_lata, _lona, _latb, _lonb, use_osrm_cache=False)
                if _d <= _SOLO_MERGE_KM:
                    # รวม trip_num ใหญ่ → เล็ก
                    _base_t2  = min(_ta, _tb)
                    _other_t2 = max(_ta, _tb)
                    df.loc[df['Trip'] == _other_t2, 'Trip'] = _base_t2
                    _merged_this_round.add(_other_t2)
                    _sc_merged += 1
                    safe_print(f"   🔀 รวม trip {_other_t2}({_cb}) → {_base_t2}({_ca}) ห่าง {_d*1000:.0f}m [{_suba}/{_disa}/{_pova}]")
                    _changed = True
                    break   # หาพาร์ทเนอร์ให้ _ta ได้แล้ว → ไปตัวถัดไป
        if not _changed:
            break
    if _sc_merged:
        safe_print(f"✅ same-location merge: รวม {_sc_merged} ทริป")


    # ==========================================
    # Step 6.6: 🔄 BRANCH-LEVEL MERGE - ดึงสาขาจากทริปถัดไปมาเติมทริปปัจจุบัน
    # หลักการ: เริ่มจากทริปไกลสุด ถ้ายังไม่เต็ม ดึงสาขาที่ใกล้จากทริปถัดไปมาทีละสาขา
    # ==========================================
    safe_print("🔄 กำลังเติมทริปที่ไม่เต็ม buffer ด้วยสาขาใกล้เคียง...")
    
    def get_trip_capacity(trip_num):
        """คำนวณความจุที่เหลือของทริป"""
        trip_data = df[df['Trip'] == trip_num]
        if len(trip_data) == 0:
            return None
        
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        codes = trip_data['Code'].tolist()
        
        # เช็ค BU
        is_punthai = all(branch_bu_cache.get(c, False) for c in codes)
        buffer = punthai_buffer if is_punthai else maxmart_buffer
        
        # หารถที่รับ constraint ได้
        max_vehicles = [branch_max_vehicle_cache.get(c, '6W') for c in codes]
        min_priority = min(vehicle_priority.get(v, 3) for v in max_vehicles)
        allowed_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(min_priority, '6W')
        
        limits = PUNTHAI_LIMITS if is_punthai else LIMITS
        max_w = limits[allowed_vehicle]['max_w']                # hard limit
        max_c = limits[allowed_vehicle]['max_c'] * buffer       # คิ้ว ยืดได้
        max_drops = limits[allowed_vehicle]['max_drops']

        # รวบรวม zone/highway เพื่อตรวจ compatibility ใน merge loop
        _provinces = set(trip_data['_province'].dropna().unique()) if '_province' in trip_data.columns else set()
        _zones     = set(trip_data['_logistics_zone'].dropna().unique()) if '_logistics_zone' in trip_data.columns else set()
        _hws: set  = set()
        if '_zone_highway' in trip_data.columns:
            for _hw in trip_data['_zone_highway'].dropna().unique():
                _hws.update(str(_hw).split('/'))
        # 🔒 รวบรวม regions ของทริป (ใช้กรองไม่ให้ข้ามภาค)
        _regions: set = set()
        for _p in _provinces:
            _r = get_region_name(str(_p)) if _p else 'ไม่ระบุ'
            if _r and _r != 'ไม่ระบุ':
                _regions.add(_r)
        # fallback: ดึงจาก _region_name column
        if not _regions and '_region_name' in trip_data.columns:
            for _rn in trip_data['_region_name'].dropna().unique():
                if _rn and _rn != 'ไม่ระบุ':
                    _regions.add(_rn)
        _sd_col65 = next((c for c in ['Subdistrict', '_subdistrict', 'ตำบล'] if c in trip_data.columns), None)
        _subds65 = set(trip_data[_sd_col65].dropna().unique()) if _sd_col65 else set()
        return {
            'weight': total_w,
            'cube': total_c,
            'codes': codes,
            'drops': len(codes),
            'max_w': max_w,
            'max_c': max_c,
            'max_drops': max_drops,
            'is_punthai': is_punthai,
            'allowed_vehicle': allowed_vehicle,
            'min_priority': min_priority,
            'centroid_lat': trip_data['_lat'].mean(),
            'centroid_lon': trip_data['_lon'].mean(),
            'provinces': _provinces,
            'logistics_zones': _zones,
            'highways': _hws,
            'regions': _regions,
            'subdistricts': _subds65,
            'branch_pts': [(float(r['_lat']), float(r['_lon']))
                           for _, r in trip_data.iterrows()
                           if float(r.get('_lat', 0) or 0) > 0 and float(r.get('_lon', 0) or 0) > 0],
        }
    
    def can_add_branch_to_trip(branch_row, trip_capacity):
        """เช็คว่าสามารถเพิ่มสาขานี้เข้าทริปได้หรือไม่"""
        branch_code = branch_row['Code']
        branch_w = branch_row['Weight']
        branch_c = branch_row['Cube']
        branch_vehicle = branch_max_vehicle_cache.get(branch_code, '6W')
        branch_priority = vehicle_priority.get(branch_vehicle, 3)

        # 🚫 เช็ค vehicle constraint: effective vehicle = min(trip, branch)
        # ถ้าเพิ่มสาขานี้แล้วต้อง downgrade รถ → ตรวจว่าโหลดรวมยังพอดีรถเล็กลงไหม
        effective_priority = min(trip_capacity['min_priority'], branch_priority)
        effective_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(effective_priority, '6W')
        is_punthai = trip_capacity.get('is_punthai', False)
        eff_limits = (PUNTHAI_LIMITS if is_punthai else LIMITS)[effective_vehicle]
        # น้ำหนัก = hard limit (ห้ามเกิน), คิ้ว = ยืดได้นิดหน่อย
        w_buffer = 1.0    # น้ำหนักห้ามเกิน
        c_buffer = 1.009  # คิ้วยืดได้ < 101%

        # เช็คน้ำหนัก/ปริมาตร/drops กับรถที่ effective จริงๆ
        new_w = trip_capacity['weight'] + branch_w
        new_c = trip_capacity['cube'] + branch_c
        new_drops = trip_capacity['drops'] + 1

        if new_w > eff_limits['max_w'] * w_buffer:
            return False, f"น้ำหนักเกิน ({effective_vehicle})"
        if new_c > eff_limits['max_c'] * c_buffer:
            return False, f"ปริมาตรเกิน ({effective_vehicle})"
        if new_drops > eff_limits['max_drops']:
            return False, f"Drop เกิน ({effective_vehicle})"

        return True, "OK"
    
    def get_nearby_branches(branch_row, all_branches_df, max_dist_km=6.0):
        """หาสาขาที่อยู่ใกล้กัน (ตำบลเดียวกัน หรือ ห่างกัน < 6 km)"""
        branch_lat = branch_row['_lat']
        branch_lon = branch_row['_lon']
        branch_subdistrict = branch_row.get('_subdistrict', '')
        branch_code = branch_row['Code']
        
        nearby_codes = []
        
        for _, other_row in all_branches_df.iterrows():
            other_code = other_row['Code']
            if other_code == branch_code:
                continue
            
            # 1. ตำบลเดียวกัน → ต้องมาด้วยกัน
            if other_row.get('_subdistrict', '') == branch_subdistrict and branch_subdistrict:
                nearby_codes.append(other_code)
                continue
            
            # 2. ห่างกัน < 6 km → ต้องมาด้วยกัน
            other_lat = other_row['_lat']
            other_lon = other_row['_lon']
            if other_lat > 0 and other_lon > 0 and branch_lat > 0 and branch_lon > 0:
                dist = haversine_distance(branch_lat, branch_lon, other_lat, other_lon, use_osrm_cache=False)
                if dist <= max_dist_km:
                    nearby_codes.append(other_code)
        
        return nearby_codes
    
    # วนลูปทริปจากไกลสุด (1) ไปใกล้สุด
    all_trips = sorted(df[df['Trip'] > 0]['Trip'].unique())
    moved_branches = 0
    
    for i, current_trip in enumerate(all_trips[:-1]):  # ไม่รวมทริปสุดท้าย
        trip_cap = get_trip_capacity(current_trip)
        if not trip_cap:
            continue
        
        # เช็คว่าทริปนี้ยังมีที่เหลือไหม
        w_util = trip_cap['weight'] / trip_cap['max_w']
        c_util = trip_cap['cube'] / trip_cap['max_c']
        
        if max(w_util, c_util) >= 0.97:  # ถ้าเต็มแล้วไม่ต้องเติม
            continue
        
        # หาสาขาจากทริปถัดไปทันที (i+1 เท่านั้น ไม่ข้ามทริป)
        for next_trip in all_trips[i+1:i+2]:
            next_trip_data = df[df['Trip'] == next_trip].copy()
            if len(next_trip_data) == 0:
                continue
            
            # คำนวณระยะห่างของแต่ละสาขาใน next_trip จาก centroid ของ current_trip (vectorized)
            _ntd_lats = next_trip_data['_lat'].fillna(0).to_numpy(dtype=float)
            _ntd_lons = next_trip_data['_lon'].fillna(0).to_numpy(dtype=float)
            _ntd_valid = (_ntd_lats > 0) & (_ntd_lons > 0)
            _ntd_dists = _hav_vec(trip_cap['centroid_lat'], trip_cap['centroid_lon'], _ntd_lats, _ntd_lons)
            _ntd_dists[~_ntd_valid] = 999.0
            next_trip_data = next_trip_data.copy()
            next_trip_data['_dist_to_current'] = _ntd_dists
            
            # เรียงตามระยะใกล้สุดก่อน
            next_trip_data = next_trip_data.sort_values('_dist_to_current')
            
            # เก็บสาขาที่ย้ายแล้วเพื่อไม่ให้ซ้ำ
            already_moved = set()
            
            # ดึงสาขาที่ใกล้และเข้ากันได้
            for _, branch_row in next_trip_data.iterrows():
                branch_code = branch_row['Code']
                
                # ข้ามถ้าย้ายไปแล้ว
                if branch_code in already_moved:
                    continue
                
                dist_to_trip = branch_row['_dist_to_current']

                # ขยาย merge radius ตาม utilization + province + BKK strict
                _b_prov = branch_row.get('_province', '')
                _mg_util = max(trip_cap['weight'] / trip_cap['max_w'], trip_cap['cube'] / trip_cap['max_c'])
                _mg_same_prov = bool(trip_cap.get('provinces', set()) & {_b_prov}) if _b_prov else False
                _BKK_MG = 'กรุงเทพมหานคร'
                _trip_has_bkk = _BKK_MG in trip_cap.get('provinces', set())
                _b_is_bkk    = (_b_prov == _BKK_MG)
                if _trip_has_bkk or _b_is_bkk:
                    _mg_dist_limit = 5.0    # กรุงเทพ: เขตเดียวกัน ≤5km
                elif _mg_same_prov:
                    _mg_dist_limit = 20.0   # จังหวัดเดียวกัน ≤20km
                else:
                    _mg_dist_limit = 10.0   # ต่างจังหวัด: ใกล้ชิดเท่านั้น ≤10km
                if dist_to_trip > _mg_dist_limit:
                    continue
                
                # อัพเดต trip_cap เพราะอาจมีการเพิ่มสาขาแล้ว
                trip_cap = get_trip_capacity(current_trip)
                if not trip_cap:
                    break
                
                # เช็คว่าเต็มหรือยัง
                w_util = trip_cap['weight'] / trip_cap['max_w']
                c_util = trip_cap['cube'] / trip_cap['max_c']
                if max(w_util, c_util) >= 0.98:
                    break  # เต็มแล้ว หยุด
                
                # 🚫 Zone + Region compatibility: ห้ามรวมสาขาคนละทิศ/highway/ภาค
                # (_b_prov already assigned above for distance check)
                _b_zone = branch_row.get('_logistics_zone', '')
                _b_hw   = branch_row.get('_zone_highway', '')
                _b_hws  = set(str(_b_hw).split('/')) if _b_hw else set()
                # 🔒 ตรวจภาคก่อน — คำนวณจาก province โดยตรง
                _b_region = get_region_name(str(_b_prov)) if _b_prov else ''
                _trip_regions = trip_cap.get('regions', set())
                _region_ok = (
                    not _b_region or _b_region == 'ไม่ระบุ' or   # candidate ไม่ทราบภาค
                    (not _trip_regions and (not _b_region or _b_region == 'ไม่ระบุ')) or  # ทั้งคู่ไม่ทราบ → zone guard
                    (len(_trip_regions) == 1 and _b_region in _trip_regions) or  # ทริปมีภาคเดียว ตรงกัน
                    (len(_trip_regions) > 1 and _b_region in _trip_regions)   # ทริปมีหลายภาค (ผิดปกติ — zone guard ดูแล)
                )
                if not _region_ok:
                    safe_print(f"      🚫 merge skip {branch_code} ภาคต่างกัน ({_b_region}/{_b_prov} ≠ {_trip_regions})")
                    continue
                # 🛑 Distance fallback: สาขาไม่รู้จังหวัด + ทริปมีภาคที่รู้ → เช็คระยะ (>15km block)
                if (not _b_region or _b_region == 'ไม่ระบุ') and _trip_regions:
                    _bx_lat = float(branch_row.get('_lat', 0) or 0)
                    _bx_lon = float(branch_row.get('_lon', 0) or 0)
                    _tx_lat = float(trip_cap.get('centroid_lat', 0) or 0)
                    _tx_lon = float(trip_cap.get('centroid_lon', 0) or 0)
                    if _bx_lat and _tx_lat:
                        _dp6 = radians(_bx_lat - _tx_lat); _dl6 = radians(_bx_lon - _tx_lon)
                        _a6 = sin(_dp6/2)**2 + cos(radians(_tx_lat))*cos(radians(_bx_lat))*sin(_dl6/2)**2
                        _dist_mg = 2*6371*atan2(sqrt(_a6), sqrt(1-_a6))
                        if _dist_mg > 15.0:
                            safe_print(f"      🛑 MERGE DIST GUARD: ตัด {branch_code} ห่าง {_dist_mg:.1f}km (ไม่รู้จังหวัด ภาคทริป={_trip_regions})")
                            continue
                # 🔒 กรุงเทพฯ isolation (Step 6.6 merge): ห้ามกรุงเทพฯ ปนกับจังหวัดอื่น
                _BKK = 'กรุงเทพมหานคร'
                _trip_provs_mg = trip_cap.get('provinces', set())
                # fallback: ถ้า provinces ว่าง → ดูจาก zone name (BKK_ prefix)
                _trip_zones_mg_bkk = trip_cap.get('logistics_zones', set())
                _trip_is_bkk = (
                    _BKK in _trip_provs_mg or
                    any(str(z).startswith('BKK_') for z in _trip_zones_mg_bkk)
                )
                _branch_is_bkk = (_b_prov == _BKK)
                if _trip_is_bkk != _branch_is_bkk and (_b_prov or _trip_is_bkk):
                    safe_print(f"      🚫 BKK isolation merge: ตัด {branch_code} ({_b_prov}) ≠ trip BKK={_trip_is_bkk}")
                    continue
                # 🔒 ZONE_NEARBY strict (Step 6.6 merge): ห้ามรวม ZONE_NEARBY ต่างจังหวัด
                _trip_zones_mg = trip_cap.get('logistics_zones', set())
                _is_trip_nearby_mg = any(str(z).startswith('ZONE_NEARBY_') for z in _trip_zones_mg)
                _is_branch_nearby_mg = str(_b_zone or '').startswith('ZONE_NEARBY_')
                if _is_trip_nearby_mg or _is_branch_nearby_mg:
                    if _b_prov and _trip_provs_mg and _b_prov not in _trip_provs_mg:
                        safe_print(f"      🚫 NEARBY strict merge: ตัด {branch_code} ({_b_prov}/{_b_zone}) ≠ trip provinces {_trip_provs_mg}")
                        continue
                _trip_provs_mg = trip_cap.get('provinces', set())
                _trip_zones_mg2 = trip_cap.get('logistics_zones', set())
                # ❌ ห้ามข้ามโซนย่อยเด็ดขาด: ถ้ามี zone ใดก็ตาม ต้องตรงกัน
                _branch_has_zone = bool(_b_zone and str(_b_zone).strip())
                _trip_has_zone   = bool(_trip_zones_mg2)
                if _trip_has_zone and _branch_has_zone and _b_zone not in _trip_zones_mg2:
                    continue  # ต่าง sub-zone → ห้าม merge
                if _trip_has_zone and not _branch_has_zone:
                    continue  # trip มี zone แต่ branch ไม่มี → ห้าม merge
                _zone_ok = (
                    _b_prov in _trip_provs_mg or
                    _b_zone in _trip_zones_mg2
                    # ห้าม merge ข้ามจังหวัด แม้จะเป็น paired province
                )
                if not _zone_ok:
                    safe_print(f"      🚫 merge skip {branch_code} ({_b_prov}/{_b_zone}) ≠ trip zone {_trip_provs_mg}")
                    continue

                # เช็คว่าเพิ่มสาขานี้ได้ไหม
                can_add, reason = can_add_branch_to_trip(branch_row, trip_cap)
                
                if can_add:
                    # ✅ ย้ายสาขานี้มาทริปปัจจุบัน
                    df.loc[df['Code'] == branch_code, 'Trip'] = current_trip
                    already_moved.add(branch_code)
                    moved_branches += 1
                    safe_print(f"   ✅ ย้าย {branch_code} จาก Trip {next_trip} → Trip {current_trip} (ห่าง {dist_to_trip:.1f} km)")
                    
                    # 🔗 หาสาขาใกล้เคียง (ตำบลเดียวกัน หรือ ห่าง < 6 km) แล้วย้ายมาด้วย
                    nearby_codes = get_nearby_branches(branch_row, next_trip_data[~next_trip_data['Code'].isin(already_moved)])
                    
                    for nearby_code in nearby_codes:
                        if nearby_code in already_moved:
                            continue
                        
                        # อัพเดต trip_cap อีกครั้ง
                        trip_cap = get_trip_capacity(current_trip)
                        if not trip_cap:
                            break
                        
                        # เช็คว่าเต็มหรือยัง
                        w_util = trip_cap['weight'] / trip_cap['max_w']
                        c_util = trip_cap['cube'] / trip_cap['max_c']
                        if max(w_util, c_util) >= 0.95:
                            break
                        
                        nearby_row = next_trip_data[next_trip_data['Code'] == nearby_code]
                        if len(nearby_row) == 0:
                            continue
                        nearby_row = nearby_row.iloc[0]
                        
                        # zone + region check สำหรับ nearby สาขาด้วย
                        _nb_prov = nearby_row.get('_province', '')
                        _nb_zone = nearby_row.get('_logistics_zone', '')
                        _nb_hw   = nearby_row.get('_zone_highway', '')
                        _nb_hws  = set(str(_nb_hw).split('/')) if _nb_hw else set()
                        # 🔒 region check ก่อน
                        _nb_region = get_region_name(str(_nb_prov)) if _nb_prov else ''
                        _trip_regions_nb = trip_cap.get('regions', set())
                        _nb_region_ok = (
                            not _nb_region or _nb_region == 'ไม่ระบุ' or   # candidate ไม่ทราบภาค
                            (not _trip_regions_nb and (not _nb_region or _nb_region == 'ไม่ระบุ')) or  # ทั้งคู่ไม่ทราบ
                            (len(_trip_regions_nb) == 1 and _nb_region in _trip_regions_nb) or  # ทริปมีภาคเดียว
                            (len(_trip_regions_nb) > 1 and _nb_region in _trip_regions_nb)  # ทริปมีหลายภาค
                        )
                        if not _nb_region_ok:
                            continue
                        # 🛑 Nearby distance fallback: สาขาไม่รู้จังหวัด + ทริปมีภาค → >15km block
                        if (not _nb_region or _nb_region == 'ไม่ระบุ') and _trip_regions_nb:
                            _nbx_lat = float(nearby_row.get('_lat', 0) or 0)
                            _nbx_lon = float(nearby_row.get('_lon', 0) or 0)
                            _txn_lat = float(trip_cap.get('centroid_lat', 0) or 0)
                            _txn_lon = float(trip_cap.get('centroid_lon', 0) or 0)
                            if _nbx_lat and _txn_lat:
                                _dp7 = radians(_nbx_lat - _txn_lat); _dl7 = radians(_nbx_lon - _txn_lon)
                                _a7 = sin(_dp7/2)**2 + cos(radians(_txn_lat))*cos(radians(_nbx_lat))*sin(_dl7/2)**2
                                _dist_mg7 = 2*6371*atan2(sqrt(_a7), sqrt(1-_a7))
                                if _dist_mg7 > 15.0:
                                    continue
                        # 🔒 กรุงเทพฯ isolation (nearby merge)
                        _BKK = 'กรุงเทพมหานคร'
                        _nb_trip_provs = trip_cap.get('provinces', set())
                        _nb_trip_zones_bkk = trip_cap.get('logistics_zones', set())
                        _nb_trip_is_bkk = (
                            _BKK in _nb_trip_provs or
                            any(str(z).startswith('BKK_') for z in _nb_trip_zones_bkk)
                        )
                        _nb_branch_is_bkk = (_nb_prov == _BKK)
                        if _nb_trip_is_bkk != _nb_branch_is_bkk and (_nb_prov or _nb_trip_is_bkk):
                            continue
                        _nb_zone_ok = (
                            _nb_prov in trip_cap.get('provinces', set()) or
                            _nb_zone in trip_cap.get('logistics_zones', set()) or
                            bool(trip_cap.get('highways', set()) & _nb_hws)
                        )
                        if not _nb_zone_ok:
                            continue
                        can_add_nearby, _ = can_add_branch_to_trip(nearby_row, trip_cap)
                        if can_add_nearby:
                            df.loc[df['Code'] == nearby_code, 'Trip'] = current_trip
                            already_moved.add(nearby_code)
                            moved_branches += 1
                            safe_print(f"   🔗 ย้ายด้วย {nearby_code} (ใกล้กัน/ตำบลเดียวกัน)")
        
        # หลังจากเติมเสร็จ เช็คอีกครั้ง
        trip_cap = get_trip_capacity(current_trip)
        if trip_cap:
            w_util = trip_cap['weight'] / trip_cap['max_w']
            c_util = trip_cap['cube'] / trip_cap['max_c']
            safe_print(f"   📊 Trip {current_trip}: {max(w_util, c_util)*100:.1f}% ({len(trip_cap['codes'])} สาขา)")
    
    if moved_branches > 0:
        safe_print(f"🔄 ย้ายสาขาเสร็จ: ย้าย {moved_branches} สาขา")
        
        # ลบทริปที่ว่างเปล่า
        empty_trips = [t for t in df['Trip'].unique() if t > 0 and len(df[df['Trip'] == t]) == 0]
        
        # Renumber ทริปใหม่หลังย้าย
        remaining_trips = sorted(df[df['Trip'] > 0]['Trip'].unique())
        trip_renumber = {old: new for new, old in enumerate(remaining_trips, start=1)}
        df['Trip'] = df['Trip'].map(lambda x: trip_renumber.get(x, x) if x > 0 else x)

    # ── Pre-seed DISTANCE_CACHE ด้วย OSRM road distance ทุกสาขา (1 batch call) ──
    # ยิงก่อน consolidation เริ่ม — loop ทั้งหมดหลังนี้ lookup cache ทันที ไม่มี network
    try:
        _pre_lats = df['_lat'].fillna(0).to_numpy(dtype=float)
        _pre_lons = df['_lon'].fillna(0).to_numpy(dtype=float)
        _pre_valid = (_pre_lats > 0) & (_pre_lons > 0)
        _pre_pts = list(zip(_pre_lats[_pre_valid].tolist(), _pre_lons[_pre_valid].tolist()))
        # dedup พิกัด (round 4 ตำแหน่ง)
        _pre_pts_u = list({(round(la,4), round(lo,4)) for la, lo in _pre_pts})
        if len(_pre_pts_u) >= 2:
            safe_print(f"📡 Pre-compute road distances: {len(_pre_pts_u)} จุด ({len(_pre_pts_u)**2} คู่)...")
            coords_str = ";".join([f"{lo},{la}" for la, lo in _pre_pts_u])
            _pre_url = (f"http://router.project-osrm.org/table/v1/driving/"
                        f"{coords_str}?annotations=distance")
            _pre_r = requests.get(_pre_url, timeout=30)
            _pre_data = _pre_r.json()
            if _pre_data.get("code") == "Ok":
                _pre_mat = _pre_data["distances"]
                _seeded = 0
                for _pi, (_la1, _lo1) in enumerate(_pre_pts_u):
                    for _pj, (_la2, _lo2) in enumerate(_pre_pts_u):
                        if _pi == _pj: continue
                        _dm = _pre_mat[_pi][_pj]
                        if _dm and _dm > 0:
                            _ck = f"{_la1:.4f},{_lo1:.4f}_{_la2:.4f},{_lo2:.4f}"
                            if _ck not in DISTANCE_CACHE:
                                DISTANCE_CACHE[_ck] = _dm / 1000.0
                                _seeded += 1
                safe_print(f"   ✅ Seeded {_seeded} road distances → consolidation ใช้ถนนจริง")
    except Exception as _pre_ex:
        safe_print(f"   ⚠️ Pre-compute skipped: {_pre_ex} (fallback haversine×1.35)")

    # ==========================================
    # Step 6.65: 🔗 AGGRESSIVE CONSOLIDATION — รวมทริปที่ยังว่างอยู่
    # หลักการ: "จะตัดใหม่ต้องเต็มก่อน" — รวม 2 ทริปที่ util ต่ำเข้าด้วยกัน
    # ถ้าน้ำหนัก+ปริมาตร+drops รวมกันแล้วยังพอดีรถ
    # ==========================================
    MIN_CONSOLIDATION_UTIL = 0.98  # มาตรฐาน 98% — รวมทริปที่ยังไม่ถึง 98%
    _consol_rounds = 0
    _consol_total = 0
    # Pass 1: same-province เท่านั้น (เติมตากด้วยตากก่อน)
    # Pass 2: cross-province สำหรับทริปที่ยังว่าง (<60%)
    _same_prov_only = True   # เริ่ม pass 1
    _pass1_done = False
    while _consol_rounds < 60:
        _consol_rounds += 1
        _trips_now = sorted(df[df['Trip'] > 0]['Trip'].unique())

        # Build capacity info for all trips
        _caps_cs = {}
        for _t_cs in _trips_now:
            _c_cs = get_trip_capacity(_t_cs)
            if _c_cs:
                _caps_cs[_t_cs] = _c_cs

        # ── Pre-compute OSRM road distance ทุกคู่สาขา ครั้งเดียวต่อ round ──
        _all_pts_665: list = []
        _pt_trip_665: list = []
        for _t6, _c6 in _caps_cs.items():
            for _pt6 in _c6.get('branch_pts', []):
                _all_pts_665.append(_pt6)
                _pt_trip_665.append(_t6)
        _road_mat_665 = _osrm_table_batch(_all_pts_665) if len(_all_pts_665) >= 2 else {}

        def _min_road_665(ta, tb):
            best = float('inf')
            for i, ti in enumerate(_pt_trip_665):
                if ti != ta: continue
                for j, tj in enumerate(_pt_trip_665):
                    if tj != tb: continue
                    d = _road_mat_665.get((i, j), float('inf'))
                    if d < best: best = d
            if best == float('inf'):
                ca, cb = _caps_cs.get(ta, {}), _caps_cs.get(tb, {})
                return haversine_distance(
                    float(ca.get('centroid_lat',0) or 0), float(ca.get('centroid_lon',0) or 0),
                    float(cb.get('centroid_lat',0) or 0), float(cb.get('centroid_lon',0) or 0),
                    use_osrm_cache=False)
            return best

        # Find under-utilized trips (sorted: lowest util first)
        _under_cs = sorted(
            [t for t, c in _caps_cs.items()
             if max(c['weight'] / c['max_w'], c['cube'] / c['max_c']) < MIN_CONSOLIDATION_UTIL],
            key=lambda t: max(_caps_cs[t]['weight'] / _caps_cs[t]['max_w'],
                              _caps_cs[t]['cube'] / _caps_cs[t]['max_c'])
        )
        if not _under_cs:
            break

        _merged_cs = False
        for _ta_cs in _under_cs:
            if _ta_cs not in _caps_cs:
                continue
            _ca_cs = _caps_cs[_ta_cs]

            # เรียง candidate trips: same-province ก่อน → cross-province ทีหลัง
            # ทำให้ ตากเต็มก่อนค่อย spill ไปสุโขทัย
            _pa_cs_sort = _ca_cs.get('provinces', set())
            def _tb_priority(t):
                _pb = _caps_cs[t].get('provinces', set()) if t in _caps_cs else set()
                _same_p = bool(_pa_cs_sort & _pb)
                return (0 if _same_p else 1, t)  # same-province ก่อน
            for _tb_cs in sorted(_caps_cs.keys(), key=_tb_priority):
                if _tb_cs == _ta_cs or _tb_cs not in _caps_cs:
                    continue
                _cb_cs = _caps_cs[_tb_cs]

                _pa_cs = _ca_cs.get('provinces', set())
                _pb_cs = _cb_cs.get('provinces', set())
                _za_cs = _ca_cs.get('logistics_zones', set())
                _zb_cs = _cb_cs.get('logistics_zones', set())
                _ra_cs = _ca_cs.get('regions', set())
                _rb_cs = _cb_cs.get('regions', set())

                # Zone family: ใช้ 2 prefix แรก (เช่น ZONE_BKK, ZONE_H1, ZONE_K)
                def _zfam_cs(z): parts = str(z).split('_'); return '_'.join(parts[:2]) if len(parts) >= 2 else str(z)
                _za_fam_cs = {_zfam_cs(z) for z in _za_cs if z}
                _zb_fam_cs = {_zfam_cs(z) for z in _zb_cs if z}

                # คำนวณ util ของ trip A (ทริปที่กำลังจะรับ merge)
                _util_a = max(_ca_cs['weight'] / _ca_cs['max_w'],
                              _ca_cs['cube'] / _ca_cs['max_c']) if _ca_cs['max_w'] > 0 else 0
                _is_very_low = _util_a < 0.20  # ทริปที่ต่ำมาก (<20%) → ผ่อน zone ได้

                # Must share province OR zone OR zone family — หรือ min road dist ≤30km ภาคเดียวกัน
                _cs_share_area = ((_pa_cs & _pb_cs) or (_za_cs & _zb_cs) or (_za_fam_cs & _zb_fam_cs))
                if not _cs_share_area:
                    _cd = _min_road_665(_ta_cs, _tb_cs)
                    _same_r = bool(_ra_cs & _rb_cs) if (_ra_cs and _rb_cs) else True
                    if _cd <= 30.0 and _same_r:
                        _cs_share_area = True
                    elif not _same_prov_only and _util_a < 0.60 and _same_r and _cd <= 80.0:
                        _cs_share_area = True
                if not _cs_share_area:
                    continue
                # Must share region (ภาค) — strict: if EITHER side has known region, both must match
                if _ra_cs and _rb_cs and not (_ra_cs & _rb_cs):
                    continue
                if _ra_cs and not _rb_cs and _pa_cs and _pb_cs:
                    # Trip B has no region info but has provinces → compute region from provinces
                    _rb_cs_calc = {get_region_name(str(p)) for p in _pb_cs if p}
                    _rb_cs_calc.discard('ไม่ระบุ'); _rb_cs_calc.discard('')
                    if _rb_cs_calc and not (_ra_cs & _rb_cs_calc):
                        continue
                if _rb_cs and not _ra_cs and _pa_cs and _pb_cs:
                    _ra_cs_calc = {get_region_name(str(p)) for p in _pa_cs if p}
                    _ra_cs_calc.discard('ไม่ระบุ'); _ra_cs_calc.discard('')
                    if _ra_cs_calc and not (_rb_cs & _ra_cs_calc):
                        continue
                # BKK isolation — normalize aliases
                _BKK_cs = 'กรุงเทพมหานคร'
                _BKK_ALIASES_cs = {'กรุงเทพฯ', 'กทม', 'กทม.', 'Bangkok'}
                _pa_has_bkk = bool(_pa_cs & ({_BKK_cs} | _BKK_ALIASES_cs))
                _pb_has_bkk = bool(_pb_cs & ({_BKK_cs} | _BKK_ALIASES_cs))
                if _pa_has_bkk != _pb_has_bkk:
                    continue
                # 📐 Province strict — ห้าม consolidate ข้ามจังหวัดทุกกรณี
                if _pa_cs and _pb_cs:
                    _prov_overlap = _pa_cs & _pb_cs
                    if not _prov_overlap:
                        continue  # คนละจังหวัด → ห้าม consolidate

                # 📐 Subdistrict check — ข้ามตำบล: ใช้ระยะถนนจริง สาขาต่อสาขา ≤ 5km
                _sda_cs = _ca_cs.get('subdistricts', set())
                _sdb_cs = _cb_cs.get('subdistricts', set())
                if _sda_cs and _sdb_cs and not (_sda_cs & _sdb_cs):
                    if _min_road_665(_ta_cs, _tb_cs) > 5.0:
                        continue
                # Zone isolation by centroid distance: จังหวัดเดียวกัน ≤80km, คู่จังหวัด ≤120km
                if not (_pa_cs & _pb_cs):  # คู่จังหวัด (ผ่าน PROVINCE_PAIR_GROUPS มาแล้ว)
                    _ca_lat_cs = float(_ca_cs.get('centroid_lat', 0) or 0)
                    _ca_lon_cs = float(_ca_cs.get('centroid_lon', 0) or 0)
                    _cb_lat_cs = float(_cb_cs.get('centroid_lat', 0) or 0)
                    _cb_lon_cs = float(_cb_cs.get('centroid_lon', 0) or 0)
                    if _ca_lat_cs and _cb_lat_cs:
                        _dp_cs2 = radians(_cb_lat_cs - _ca_lat_cs)
                        _dl_cs2 = radians(_cb_lon_cs - _ca_lon_cs)
                        _aa_cs2 = sin(_dp_cs2/2)**2 + cos(radians(_ca_lat_cs))*cos(radians(_cb_lat_cs))*sin(_dl_cs2/2)**2
                        _cdist_cs = 2*6371*atan2(sqrt(_aa_cs2), sqrt(1-_aa_cs2))
                        if _cdist_cs > 120.0:
                            continue
                # ZONE_NEARBY: only same province
                _a_nb_cs = any(str(z).startswith('ZONE_NEARBY_') for z in _za_cs)
                _b_nb_cs = any(str(z).startswith('ZONE_NEARBY_') for z in _zb_cs)
                if (_a_nb_cs or _b_nb_cs) and not (_pa_cs & _pb_cs):
                    continue

                # Check combined load fits in a truck
                _cw_cs = _ca_cs['weight'] + _cb_cs['weight']
                _cc_cs = _ca_cs['cube'] + _cb_cs['cube']
                _cd_cs = _ca_cs['drops'] + _cb_cs['drops']
                _call_cs = get_allowed_from_codes(
                    _ca_cs['codes'] + _cb_cs['codes'], ['4W', 'JB', '6W'])
                if not _call_cs:
                    continue
                _cpunthai_cs = all(branch_bu_cache.get(c, False)
                                   for c in _ca_cs['codes'] + _cb_cs['codes'])
                _cbuf_cs = 1.009  # คิ้วยืดได้ < 101%, น้ำหนัก hard limit
                _clims_cs = PUNTHAI_LIMITS if _cpunthai_cs else LIMITS

                _fits_veh_cs = None
                for _fv_cs in ['4W', 'JB', '6W']:
                    if _fv_cs not in _call_cs:
                        continue
                    _fl_cs = _clims_cs[_fv_cs]
                    if (_cw_cs <= _fl_cs['max_w'] and
                            _cc_cs <= _fl_cs['max_c'] * _cbuf_cs and
                            _cd_cs <= _fl_cs.get('max_drops', 999)):
                        _fits_veh_cs = _fv_cs
                        break

                if not _fits_veh_cs:
                    continue

                # ✅ Merge _tb_cs into _ta_cs
                _new_util = max(_cw_cs / (_clims_cs[_fits_veh_cs]['max_w'] * _cbuf_cs),
                                _cc_cs / (_clims_cs[_fits_veh_cs]['max_c'] * _cbuf_cs))
                df.loc[df['Trip'] == _tb_cs, 'Trip'] = _ta_cs
                safe_print(f"   🔗 Consolidate Trip {_tb_cs} → Trip {_ta_cs} "
                           f"[{_fits_veh_cs}] {_cd_cs} drops {_cw_cs:.0f}kg "
                           f"→ {_new_util*100:.0f}%")
                _caps_cs[_ta_cs] = get_trip_capacity(_ta_cs)
                del _caps_cs[_tb_cs]
                _consol_total += 1
                _merged_cs = True
                break

            if _merged_cs:
                break

        if not _merged_cs:
            if _same_prov_only and not _pass1_done:
                # Pass 1 หมดแล้ว → เริ่ม Pass 2 (cross-province สำหรับทริปว่าง <60%)
                _same_prov_only = False
                _pass1_done = True
                safe_print("🔗 Consolidation pass 2: cross-province สำหรับทริปที่ว่าง (<60%)")
            else:
                break

    if _consol_total > 0:
        safe_print(f"🔗 Consolidation done: merged {_consol_total} trips")
        # Renumber after consolidation
        _remaining_cs = sorted(df[df['Trip'] > 0]['Trip'].unique())
        _renumber_cs = {old: new for new, old in enumerate(_remaining_cs, start=1)}
        df['Trip'] = df['Trip'].map(lambda x: _renumber_cs.get(x, x) if x > 0 else x)
    else:
        safe_print("🔗 Consolidation: no further merges possible")

    # ── Split ทริปเกิน capacity (หลัง Step 6.65) ───────────────────────────
    _split_next_tid = (df[df['Trip'] > 0]['Trip'].max() + 1) if df[df['Trip'] > 0].shape[0] else 1
    import math as _math_hav
    def _hav_app(a, b, c, d):
        R = 6371000.0
        p1, p2 = _math_hav.radians(a), _math_hav.radians(c)
        x = (_math_hav.sin(_math_hav.radians(c-a)/2)**2 +
             _math_hav.cos(p1)*_math_hav.cos(p2)*_math_hav.sin(_math_hav.radians(d-b)/2)**2)
        return 2*R*_math_hav.atan2(_math_hav.sqrt(max(0,x)), _math_hav.sqrt(max(0,1-x)))
    def _app_split_over(df, next_tid):
        """ตรวจทริปเกิน capacity → reset + repack sequential"""
        _vp = {'4W': 1, 'JB': 2, '6W': 3}
        for _pass in range(5):
            _found = False
            for _tid_sp in sorted(df.loc[df['Trip']>0,'Trip'].unique()):
                _m = df['Trip'] == _tid_sp
                _td_sp = df[_m]
                _codes_sp = [str(c).strip().upper() for c in _td_sp['Code'].tolist()]
                _is_pt_sp = all(branch_bu_cache.get(c, False) for c in _codes_sp)
                _lims_sp = PUNTHAI_LIMITS if _is_pt_sp else LIMITS
                # หารถที่เข้มงวดสุด (priority ต่ำสุด = รถเล็กสุดที่ constraint กำหนด)
                _min_pri = min(
                    _vp.get(str(branch_max_vehicle_cache.get(c, '6W')), 3)
                    for c in _codes_sp
                ) if _codes_sp else 3
                _mv_sp = {1: '4W', 2: 'JB', 3: '6W'}.get(_min_pri, '6W')
                _lim_sp = _lims_sp[_mv_sp]
                _tw_sp = float(_td_sp['Weight'].sum())
                _tc_sp = float(_td_sp['Cube'].sum())
                if _tw_sp <= _lim_sp['max_w'] and _tc_sp <= _lim_sp['max_c']:
                    continue
                # เกิน → reset แล้ว repack ตาม location
                _rows_sp = _td_sp.index.tolist()
                df.loc[_rows_sp, 'Trip'] = 0
                df.loc[_rows_sp, 'Truck'] = ''
                # เรียงตาม lat/lon nearest-neighbor จาก DC
                _rem = list(_rows_sp)
                _cur_lat, _cur_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
                _ordered = []
                while _rem:
                    _nxt = min(_rem, key=lambda i: _hav_app(
                        _cur_lat, _cur_lon,
                        float(df.at[i,'_lat']) if '_lat' in df.columns else 0,
                        float(df.at[i,'_lon']) if '_lon' in df.columns else 0))
                    _ordered.append(_nxt)
                    _rem.remove(_nxt)
                    _cur_lat = float(df.at[_nxt,'_lat']) if '_lat' in df.columns else _cur_lat
                    _cur_lon = float(df.at[_nxt,'_lon']) if '_lon' in df.columns else _cur_lon
                # bin-fill
                _cur_w = _cur_c = 0.0
                _cur_tid = next_tid
                next_tid += 1
                for _ri in _ordered:
                    _rw = float(df.at[_ri,'Weight']) if 'Weight' in df.columns else 0
                    _rc = float(df.at[_ri,'Cube']) if 'Cube' in df.columns else 0
                    if _cur_w + _rw > _lim_sp['max_w'] or _cur_c + _rc > _lim_sp['max_c']:
                        _cur_tid = next_tid
                        next_tid += 1
                        _cur_w = _cur_c = 0.0
                    df.at[_ri,'Trip'] = _cur_tid
                    df.at[_ri,'Truck'] = _mv_sp
                    _cur_w += _rw; _cur_c += _rc
                _found = True
            if not _found:
                break
        return df, next_tid
    df, _split_next_tid = _app_split_over(df, _split_next_tid)

    # ==========================================
    # Step 6.66: 🚀 FINAL UNASSIGNED SWEEP — บังคับ assign สาขาที่เหลือทั้งหมด (ห้ามทิ้ง)
    # เป้า: ทุกสาขาต้องได้รับ Trip ≥ 1 และ utilization รวม ≥ 98%
    # ==========================================
    _sweep_unassigned = df[df['Trip'] == 0].copy()
    if not _sweep_unassigned.empty:
        safe_print(f"🚀 Final sweep: {len(_sweep_unassigned)} สาขายังไม่ได้จัด → บังคับ assign...")
        _sweep_added = 0
        _sweep_new_trips = 0
        _sweep_trips_now = sorted(df[df['Trip'] > 0]['Trip'].unique())

        # สร้าง coord map ของทุกทริป (centroid)
        _sw_trip_coords: dict = {}  # trip → [(lat, lon), ...]
        _sw_trip_caps: dict   = {}  # trip → get_trip_capacity() — cache ไว้ ไม่เรียกซ้ำ
        for _sw_t in _sweep_trips_now:
            _sw_rows = df[df['Trip'] == _sw_t]
            _sw_lats = _sw_rows['_lat'].fillna(0).to_numpy(dtype=float)
            _sw_lons = _sw_rows['_lon'].fillna(0).to_numpy(dtype=float)
            _sw_valid = (_sw_lats > 0) & (_sw_lons > 0)
            _sw_trip_coords[_sw_t] = list(zip(_sw_lats[_sw_valid], _sw_lons[_sw_valid]))
            _sw_trip_caps[_sw_t]   = get_trip_capacity(_sw_t)

        # dict: code_upper → row index ใน df — ป้องกัน df.loc ใน loop
        _sw_code_idx = {str(r['Code']).strip().upper(): idx for idx, r in df[['Code', 'Trip']].iterrows()}

        for _, _sw_row in _sweep_unassigned.iterrows():
            _sw_code = _sw_row['Code']
            _sw_idx  = _sw_code_idx.get(str(_sw_code).strip().upper())
            # ตรวจซ้ำ — อาจถูก assign แล้วโดย force-sync/repair ก่อนหน้า
            if _sw_idx is not None and safe_int_trip(df.at[_sw_idx, 'Trip']) > 0:
                continue
            _sw_lat = float(_sw_row.get('_lat', 0) or 0)
            _sw_lon = float(_sw_row.get('_lon', 0) or 0)
            _sw_w   = float(_sw_row.get('Weight', 0) or 0)
            _sw_c   = float(_sw_row.get('Cube', 0) or 0)
            _sw_prov   = str(_sw_row.get('_province', '') or '')
            _sw_region = get_region_name(_sw_prov) if _sw_prov else ''

            # หาทริปที่ใกล้สุด + ยังรับน้ำหนักได้ + ภาคเดียวกัน
            _sw_best_trip = None
            _sw_best_dist = 9999.0
            for _sw_t, _sw_coords in _sw_trip_coords.items():
                if not _sw_coords or _sw_lat <= 0 or _sw_lon <= 0:
                    continue
                _sw_cap = _sw_trip_caps.get(_sw_t)  # ใช้ cache แทน get_trip_capacity()
                if not _sw_cap:
                    continue
                # ตรวจภาค
                _sw_t_regions = _sw_cap.get('regions', set())
                if (_sw_region and _sw_t_regions and
                        _sw_region not in ('', 'ไม่ระบุ') and
                        not any(r == _sw_region for r in _sw_t_regions)):
                    continue
                # ตรวจ BKK isolation
                _sw_t_provs = _sw_cap.get('provinces', set())
                _BKK_SW = 'กรุงเทพมหานคร'
                if ((_sw_prov == _BKK_SW and _BKK_SW not in _sw_t_provs and _sw_t_provs) or
                        (_BKK_SW in _sw_t_provs and _sw_prov and _sw_prov != _BKK_SW)):
                    continue
                # ตรวจว่ายังจุได้
                _sw_test_w = _sw_cap['weight'] + _sw_w
                _sw_test_c = _sw_cap['cube'] + _sw_c
                _sw_veh = _sw_cap.get('vehicle', '6W')
                _sw_lims = PUNTHAI_LIMITS if _sw_cap.get('is_punthai') else LIMITS
                _sw_buf  = 1.009  # อนุญาตเกิน 100% แต่ < 101% สำหรับ final sweep
                _sw_lim  = _sw_lims.get(_sw_veh, _sw_lims['6W'])
                if (_sw_test_w > _sw_lim['max_w'] * _sw_buf or
                        _sw_test_c > _sw_lim['max_c'] * _sw_buf):
                    continue
                # หาระยะใกล้สุด
                _sw_min_d = min(
                    haversine_distance(_sw_lat, _sw_lon, lt, ln, use_osrm_cache=False)
                    for lt, ln in _sw_coords if lt > 0
                ) if _sw_lat > 0 else 9999.0
                if _sw_min_d < _sw_best_dist:
                    _sw_best_dist = _sw_min_d
                    _sw_best_trip = _sw_t

            if _sw_best_trip is not None:
                if _sw_idx is not None:
                    df.at[_sw_idx, 'Trip'] = _sw_best_trip  # O(1) แทน df.loc O(n)
                else:
                    df.loc[df['Code'] == _sw_code, 'Trip'] = _sw_best_trip
                _sw_trip_coords[_sw_best_trip].append((_sw_lat, _sw_lon))
                # อัปเดต cap cache หลัง assign
                _sw_trip_caps[_sw_best_trip] = get_trip_capacity(_sw_best_trip)
                _sweep_added += 1
                safe_print(f"   🚀 Sweep: {_sw_code} → Trip {_sw_best_trip} ({_sw_best_dist:.1f}km)")
            else:
                _max_trip_sw = safe_int_trip(df['Trip'].max()) + 1 if pd.notna(df['Trip'].max()) and df['Trip'].max() > 0 else 1
                if _sw_idx is not None:
                    df.at[_sw_idx, 'Trip'] = _max_trip_sw
                else:
                    df.loc[df['Code'] == _sw_code, 'Trip'] = _max_trip_sw
                _sw_trip_coords[_max_trip_sw] = [(_sw_lat, _sw_lon)] if _sw_lat > 0 else []
                _sw_trip_caps[_max_trip_sw] = get_trip_capacity(_max_trip_sw)
                _sweep_new_trips += 1
                safe_print(f"   🆕 Sweep new trip {_max_trip_sw}: {_sw_code} (ไม่มีทริปรับ)")

        safe_print(f"🚀 Final sweep: เพิ่ม {_sweep_added} สาขาเข้าทริปเดิม, เปิดทริปใหม่ {_sweep_new_trips} ทริป")
    else:
        safe_print("✅ Final sweep: ไม่มีสาขาค้าง — สาขาทั้งหมดได้รับทริปแล้ว")

    # ==========================================
    # Step 6.7: 🔍 REGION AUDIT — ตรวจและแยกทริปที่มีการปนภาค
    # ==========================================
    safe_print("🔍 ตรวจสอบการปนภาคใน trips...")
    _audit_fixed = 0
    _max_trip_now = df[df['Trip'] > 0]['Trip'].max() if len(df[df['Trip'] > 0]) > 0 else 0
    for _aud_trip in sorted(df[df['Trip'] > 0]['Trip'].unique()):
        _aud_data = df[df['Trip'] == _aud_trip]
        _aud_regions = {}
        for _, _aud_row in _aud_data.iterrows():
            _ap = str(_aud_row.get('_province', '') or '')
            _ar = get_region_name(_ap) if _ap else ''
            # fallback: ใช้ _region_name column ถ้า _province ว่าง
            if (not _ar or _ar == 'ไม่ระบุ'):
                _ar = str(_aud_row.get('_region_name', '') or '')
            if _ar and _ar != 'ไม่ระบุ':
                _aud_regions[_ar] = _aud_regions.get(_ar, 0) + 1
        if len(_aud_regions) <= 1:
            continue  # ไม่มีการปนภาค
        # พบการปนภาค → แยกสาขา minority ออกเป็นทริปใหม่
        # ใช้ region ที่มีสาขามากที่สุดเป็น dominant (ถ้าเท่ากัน ใช้ตามตำแหน่งใน sort)
        _dominant = max(_aud_regions, key=lambda k: (_aud_regions[k], ['เหนือ','อีสาน','ตะวันออก','กลาง','ตะวันตก','ใต้'].index(k) if k in ['เหนือ','อีสาน','ตะวันออก','กลาง','ตะวันตก','ใต้'] else 99))
        _minority_codes = []
        for _, _aud_row in _aud_data.iterrows():
            _ap2 = str(_aud_row.get('_province', '') or '')
            _ar2 = get_region_name(_ap2) if _ap2 else ''
            if (not _ar2 or _ar2 == 'ไม่ระบุ'):
                _ar2 = str(_aud_row.get('_region_name', '') or '')
            if _ar2 and _ar2 != 'ไม่ระบุ' and _ar2 != _dominant:
                _minority_codes.append(_aud_row['Code'])
        if _minority_codes:
            # 🔒 GROUP GUARD: ถ้า minority code มี group member อยู่ใน majority → ห้ามแยก
            _minority_safe = []
            for _mc in _minority_codes:
                _mc_up = str(_mc).strip().upper()
                _mc_gid = BRANCH_TO_GROUP.get(_mc_up)
                if _mc_gid:
                    _mc_sibs = {str(s).strip().upper() for s in BRANCH_GROUPS.get(_mc_gid, [])}
                    _majority_codes_up = {str(c).strip().upper() for c in _aud_data['Code']
                                          if c not in _minority_codes}
                    if _mc_sibs & _majority_codes_up:
                        safe_print(f"   🔒 AUDIT GUARD: ข้าม {_mc} — group member อยู่ใน majority ห้ามแยก")
                        continue
                _minority_safe.append(_mc)
            if _minority_safe:
                _max_trip_now += 1
                df.loc[df['Code'].isin(_minority_safe), 'Trip'] = _max_trip_now
                safe_print(f"   ⚠️ AUDIT: Trip {_aud_trip} ปนภาค {_aud_regions} → แยก {_minority_safe} → Trip ใหม่ {_max_trip_now}")
                _audit_fixed += 1
    if _audit_fixed > 0:
        safe_print(f"   🔧 AUDIT: แก้ไขการปนภาค {_audit_fixed} ทริป")
        # Renumber หลัง audit
        _aud_remaining = sorted(df[df['Trip'] > 0]['Trip'].unique())
        _aud_remap = {old: new for new, old in enumerate(_aud_remaining, start=1)}
        df['Trip'] = df['Trip'].map(lambda x: _aud_remap.get(x, x) if x > 0 else x)
    else:
        safe_print("   ✅ ไม่พบการปนภาค")

    # ==========================================
    # Step 6.8: 🔗 POST-AUDIT CONSOLIDATION — รวมเศษทริปที่เกิดจากการ audit แตก
    # เพราะ Step 6.7 อาจแยกทริปแล้วทิ้ง fragment เล็กๆ ไว้ ต้องรวมกลับ
    # ==========================================
    _pa_total = 0
    _pa_rounds = 0
    while _pa_rounds < 20:
        _pa_rounds += 1
        _pa_caps = {}
        for _t_pa in sorted(df[df['Trip'] > 0]['Trip'].unique()):
            _c_pa = get_trip_capacity(_t_pa)
            if _c_pa:
                _pa_caps[_t_pa] = _c_pa
        _pa_under = sorted(
            [t for t, c in _pa_caps.items()
             if max(c['weight'] / c['max_w'], c['cube'] / c['max_c']) < MIN_CONSOLIDATION_UTIL],
            key=lambda t: max(_pa_caps[t]['weight'] / _pa_caps[t]['max_w'],
                              _pa_caps[t]['cube'] / _pa_caps[t]['max_c'])
        )
        if not _pa_under:
            break
        _pa_merged = False
        for _ta_pa in _pa_under:
            if _ta_pa not in _pa_caps:
                continue
            _ca_pa = _pa_caps[_ta_pa]
            for _tb_pa in sorted(_pa_caps.keys()):
                if _tb_pa == _ta_pa or _tb_pa not in _pa_caps:
                    continue
                _cb_pa = _pa_caps[_tb_pa]
                _pa_a = _ca_pa.get('provinces', set())
                _pb_pa = _cb_pa.get('provinces', set())
                _za_pa = _ca_pa.get('logistics_zones', set())
                _zb_pa = _cb_pa.get('logistics_zones', set())
                _ra_pa = _ca_pa.get('regions', set())
                _rb_pa = _cb_pa.get('regions', set())
                def _zfam_pa(z): _p = str(z).split('_'); return '_'.join(_p[:2]) if len(_p) >= 2 else str(z)
                _za_fam_pa = {_zfam_pa(z) for z in _za_pa if z}
                _zb_fam_pa = {_zfam_pa(z) for z in _zb_pa if z}
                if not ((_pa_a & _pb_pa) or (_za_pa & _zb_pa) or (_za_fam_pa & _zb_fam_pa)):
                    continue
                if _ra_pa and _rb_pa and not (_ra_pa & _rb_pa):
                    continue
                if _ra_pa and not _rb_pa and _pa_a and _pb_pa:
                    _rb_pa_calc = {get_region_name(str(p)) for p in _pb_pa if p}
                    _rb_pa_calc.discard('ไม่ระบุ'); _rb_pa_calc.discard('')
                    if _rb_pa_calc and not (_ra_pa & _rb_pa_calc):
                        continue
                if _rb_pa and not _ra_pa and _pa_a and _pb_pa:
                    _ra_pa_calc = {get_region_name(str(p)) for p in _pa_a if p}
                    _ra_pa_calc.discard('ไม่ระบุ'); _ra_pa_calc.discard('')
                    if _ra_pa_calc and not (_rb_pa & _ra_pa_calc):
                        continue
                _BKK_pa = 'กรุงเทพมหานคร'
                _BKK_ALIASES_pa = {'กรุงเทพฯ', 'กทม', 'กทม.', 'Bangkok'}
                _pa_has_bkk_pa = bool(_pa_a & ({_BKK_pa} | _BKK_ALIASES_pa))
                _pb_has_bkk_pa = bool(_pb_pa & ({_BKK_pa} | _BKK_ALIASES_pa))
                if _pa_has_bkk_pa != _pb_has_bkk_pa:
                    continue
                # 📐 Zone isolation by centroid distance (post-audit consolidation)
                if not (_pa_a & _pb_pa):  # ต่างจังหวัด: max 120km centroid
                    _ca_lat_pa = float(_ca_pa.get('centroid_lat', 0) or 0)
                    _ca_lon_pa = float(_ca_pa.get('centroid_lon', 0) or 0)
                    _cb_lat_pa = float(_cb_pa.get('centroid_lat', 0) or 0)
                    _cb_lon_pa = float(_cb_pa.get('centroid_lon', 0) or 0)
                    if _ca_lat_pa and _cb_lat_pa:
                        _dp_pa2 = radians(_cb_lat_pa - _ca_lat_pa)
                        _dl_pa2 = radians(_cb_lon_pa - _ca_lon_pa)
                        _aa_pa2 = sin(_dp_pa2/2)**2 + cos(radians(_ca_lat_pa))*cos(radians(_cb_lat_pa))*sin(_dl_pa2/2)**2
                        _cdist_pa = 2*6371*atan2(sqrt(_aa_pa2), sqrt(1-_aa_pa2))
                        if _cdist_pa > 120.0:
                            continue
                _a_nb_pa = any(str(z).startswith('ZONE_NEARBY_') for z in _za_pa)
                _b_nb_pa = any(str(z).startswith('ZONE_NEARBY_') for z in _zb_pa)
                if (_a_nb_pa or _b_nb_pa) and not (_pa_a & _pb_pa):
                    continue
                _cw_pa = _ca_pa['weight'] + _cb_pa['weight']
                _cc_pa = _ca_pa['cube'] + _cb_pa['cube']
                _cd_pa = _ca_pa['drops'] + _cb_pa['drops']
                _call_pa = get_allowed_from_codes(
                    _ca_pa['codes'] + _cb_pa['codes'], ['4W', 'JB', '6W'])
                if not _call_pa:
                    continue
                _cpun_pa = all(branch_bu_cache.get(c, False)
                               for c in _ca_pa['codes'] + _cb_pa['codes'])
                _cbuf_pa = punthai_buffer if _cpun_pa else maxmart_buffer
                _clim_pa = PUNTHAI_LIMITS if _cpun_pa else LIMITS
                _fveh_pa = None
                for _fv_pa in ['4W', 'JB', '6W']:
                    if _fv_pa not in _call_pa:
                        continue
                    _fl_pa = _clim_pa[_fv_pa]
                    if (_cw_pa <= _fl_pa['max_w'] and
                            _cc_pa <= _fl_pa['max_c'] * _cbuf_pa and
                            _cd_pa <= _fl_pa.get('max_drops', 999)):
                        _fveh_pa = _fv_pa
                        break
                if not _fveh_pa:
                    continue
                df.loc[df['Trip'] == _tb_pa, 'Trip'] = _ta_pa
                _nutil = max(_cw_pa / _clim_pa[_fveh_pa]['max_w'],
                             _cc_pa / (_clim_pa[_fveh_pa]['max_c'] * _cbuf_pa))
                safe_print(f"   🔗 Post-audit merge Trip {_tb_pa} → Trip {_ta_pa} "
                           f"[{_fveh_pa}] {_cd_pa}drops {_cw_pa:.0f}kg → {_nutil*100:.0f}%")
                _pa_caps[_ta_pa] = get_trip_capacity(_ta_pa)
                del _pa_caps[_tb_pa]
                _pa_total += 1
                _pa_merged = True
                break
            if _pa_merged:
                break
        if not _pa_merged:
            break
    if _pa_total > 0:
        safe_print(f"🔗 Post-audit consolidation: merged {_pa_total} trips")
        _pa_rem = sorted(df[df['Trip'] > 0]['Trip'].unique())
        _pa_ren = {old: new for new, old in enumerate(_pa_rem, start=1)}
        df['Trip'] = df['Trip'].map(lambda x: _pa_ren.get(x, x) if x > 0 else x)
    else:
        safe_print("🔗 Post-audit consolidation: nothing to merge")

    # ==========================================
    # Step 6.9: 📊 UTILIZATION AUDIT — รายงาน utilization และสาขาค้าง
    # ==========================================
    _unassigned_final = df[df['Trip'] == 0]
    _total_branches = len(df)
    _assigned_branches = len(df[df['Trip'] > 0])
    _assign_rate = _assigned_branches / _total_branches * 100 if _total_branches > 0 else 0
    safe_print(f"\n📊 Utilization Audit:")
    safe_print(f"   สาขาทั้งหมด: {_total_branches}, assigned: {_assigned_branches} ({_assign_rate:.1f}%)")
    if not _unassigned_final.empty:
        safe_print(f"   ⚠️ ยังไม่ได้จัด {len(_unassigned_final)} สาขา: {_unassigned_final['Code'].tolist()}")
    else:
        safe_print(f"   ✅ ทุกสาขาได้รับทริปแล้ว")

    _trip_utils = []
    for _ut in df[df['Trip'] > 0]['Trip'].unique():
        _ut_rows = df[df['Trip'] == _ut]
        _ut_cap = get_trip_capacity(_ut)
        if _ut_cap and _ut_cap['max_w'] > 0:
            _ut_util = max(
                _ut_cap['weight'] / _ut_cap['max_w'],
                _ut_cap['cube'] / _ut_cap['max_c'] if _ut_cap['max_c'] > 0 else 0
            )
            _trip_utils.append(_ut_util)
    if _trip_utils:
        _avg_util = sum(_trip_utils) / len(_trip_utils) * 100
        _below_98 = sum(1 for u in _trip_utils if u < 0.98)
        safe_print(f"   Avg utilization: {_avg_util:.1f}% | ทริปที่ < 98%: {_below_98}/{len(_trip_utils)}")

    # ── Final Group Integrity Check (O(n) ด้วย dict lookup แทน iterrows+apply)
    _split_groups = 0
    _orphan_fixed = 0
    # สร้าง code_upper → (index, trip) dict ครั้งเดียว
    _code_idx_trip: dict = {
        str(r['Code']).strip().upper(): (idx, safe_int_trip(r['Trip']))
        for idx, r in df[['Code', 'Trip']].iterrows()
    }
    _seen_gids_final: set = set()
    for _fgid, _fgmembers in BRANCH_GROUPS.items():
        if _fgid in _seen_gids_final:
            continue
        _seen_gids_final.add(_fgid)
        _mem_ups = [str(c).strip().upper() for c in _fgmembers]
        _mem_data = [(up, *_code_idx_trip[up]) for up in _mem_ups if up in _code_idx_trip]
        if len(_mem_data) < 2:
            continue
        _trips_set = {t for _, _, t in _mem_data if t > 0}
        _orphan_idxs = [idx for _, idx, t in _mem_data if t == 0]
        if len(_trips_set) > 1:
            from collections import Counter as _Ctr
            _dominant_trip = _Ctr(t for _, _, t in _mem_data if t > 0).most_common(1)[0][0]
            # เช็ค capacity ของ dominant trip ก่อน merge
            _dom_cap = get_trip_capacity(_dominant_trip)
            _merge_ok = True
            if _dom_cap:
                _extra_codes = [up for up, idx, t in _mem_data if t != _dominant_trip]
                _extra_w = sum(float(df.at[idx, 'Weight'] or 0) for up, idx, t in _mem_data if t != _dominant_trip)
                _extra_c = sum(float(df.at[idx, 'Cube'] or 0) for up, idx, t in _mem_data if t != _dominant_trip)
                _new_w = _dom_cap['weight'] + _extra_w
                _new_c = _dom_cap['cube'] + _extra_c
                _dom_lim = (PUNTHAI_LIMITS if _dom_cap.get('is_punthai') else LIMITS).get(_dom_cap.get('allowed_vehicle', '6W'), LIMITS['6W'])
                # group lock เด็ดขาด — ไม่มี capacity cap, Step 7.5 จะ split ทีหลัง
                _ = _dom_lim  # suppress unused warning
            if _merge_ok:
                for _, idx, _ in _mem_data:
                    df.at[idx, 'Trip'] = _dominant_trip
                _split_groups += 1
                safe_print(f"   🔗 Final fix: group {_fgid} แตกทริป {_trips_set} → Trip {_dominant_trip}")
            else:
                # เกิน 150% → ให้ minor trips อยู่ทริปแรกของตัวเอง (group lock ยังคงได้ แต่แยกทริป)
                _minor_trips = _trips_set - {_dominant_trip}
                for _mt in _minor_trips:
                    _mt_codes = [up for up, idx, t in _mem_data if t == _mt]
                    safe_print(f"   ⚠️ Final fix SKIP: group {_fgid} overflow >150% → Trip {_mt} คงอยู่ (group lock ยังได้แต่แยกทริป)")
                _split_groups += 1
        elif _orphan_idxs and _trips_set:
            _target = next(iter(_trips_set))
            for idx in _orphan_idxs:
                df.at[idx, 'Trip'] = _target
            _orphan_fixed += len(_orphan_idxs)
            safe_print(f"   🔗 Final fix: {len(_orphan_idxs)} orphans → Trip {_target}")
    if _split_groups or _orphan_fixed:
        safe_print(f"   ⚠️ Final group fix: {_split_groups} groups แตก, {_orphan_fixed} orphans ถูกรวม")
    else:
        safe_print(f"   ✅ Group integrity: ทุก group อยู่ทริปเดียวกัน")

    # ==========================================
    # Step 7: สร้าง Summary + Central Rule + Punthai Drop Limits
    # ==========================================
    summary_data = []

    # 🚛 Fleet Constraint: ติดตามจำนวนรถแต่ละประเภทที่ใช้ไป
    _fleet_limits = fleet_limits or {'4W': 999, 'JB': 999, '6W': 999}
    fleet_used = {'4W': 0, 'JB': 0, '6W': 0}
    _fleet_rank = {1: '4W', 2: 'JB', 3: '6W'}
    _rank_fleet = {'4W': 1, 'JB': 2, '6W': 3}

    for trip_num in sorted(df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = df[df['Trip'] == trip_num]
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        trip_codes = trip_data['Code'].unique()
        trip_drops = len(trip_codes)
        
        # หาภาคของทริป (ใช้ภาคแรก)
        trip_region = trip_data['_region_name'].iloc[0] if '_region_name' in trip_data.columns else 'ไม่ระบุ'
        
        # หารถที่เหมาะสม (รวม Central Rule)
        max_vehicles = [get_max_vehicle_for_branch(c) for c in trip_codes]
        min_max_size = min(vehicle_priority.get(v, 3) for v in max_vehicles)
        max_allowed_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(min_max_size, '6W')
        
        # ตรวจ BU ของทริป
        is_punthai_only_trip = True
        for _, r in trip_data.iterrows():
            bu = str(r.get('BU', '')).upper()
            if bu not in ['211', 'PUNTHAI']:
                is_punthai_only_trip = False
                break
        
        buffer = punthai_buffer if is_punthai_only_trip else maxmart_buffer
        buffer_pct = int(buffer * 100)
        buffer_label = f"🅿️ {buffer_pct}%" if is_punthai_only_trip else f"🅼 {buffer_pct}%"
        trip_type = 'punthai' if is_punthai_only_trip else 'maxmart'
        
        # 🎯 เลือกรถตามภาค + ข้อจำกัดสาขา
        # เหนือ/ใต้ → ใช้รถใหญ่สุดที่อนุญาตเสมอ (ไม่ downgrade — เส้นทางไกล)
        # ภาคอื่น  → เล็กสุดที่รับโหลดได้ (ประหยัดรถ)
        limits_to_check = PUNTHAI_LIMITS if is_punthai_only_trip else LIMITS
        is_long_haul = str(trip_region) in ('เหนือ', 'ใต้')
        suggested = max_allowed_vehicle  # fallback = รถใหญ่สุดที่อนุญาต
        source = "📋 จำกัดสาขา" if min_max_size < 3 else "🤖 อัตโนมัติ"
        if is_long_haul:
            # เหนือ/ใต้: ใช้ max_allowed_vehicle ตรงๆ ไม่ลอง downgrade
            if min_max_size >= 3:
                source = "🚛 ไกล (เหนือ/ใต้)"
            else:
                source = "📋 จำกัดสาขา (เหนือ/ใต้)"
        else:
            # ภาคอื่น: ลอง 4W → JB → 6W เลือกเล็กสุดที่รับโหลดได้
            for _veh in ['4W', 'JB', '6W']:
                _vr = vehicle_priority.get(_veh, 3)
                if _vr > min_max_size:
                    break  # ห้ามเกินข้อจำกัดสาขา
                _lim = limits_to_check[_veh]
                if (total_w <= _lim['max_w'] * buffer and
                        total_c <= _lim['max_c'] * buffer and
                        trip_drops <= _lim['max_drops']):
                    suggested = _veh
                    if _vr < min_max_size:
                        source = "🔽 Downgrade (ขนาดพอดี)"
                    break  # เล็กสุดที่รับโหลดได้
        
        # 🔒 Punthai Drop Limit Check
        if is_punthai_only_trip:
            punthai_drop_limit = PUNTHAI_LIMITS.get(suggested, {}).get('max_drops', 999)
            if trip_drops > punthai_drop_limit:
                # ต้องเพิ่มขนาดรถเพื่อรองรับ drops - แต่ห้ามเกินข้อจำกัดสาขา!
                if suggested == '4W' and trip_drops <= PUNTHAI_LIMITS['JB']['max_drops']:
                    # เช็คว่าสาขาอนุญาต JB ไหม
                    if min_max_size >= 2:  # JB หรือ 6W
                        suggested = 'JB'
                        source += " → JB (Drop Limit)"
                    else:
                        # สาขาจำกัดแค่ 4W - ไม่สามารถ upgrade ได้!
                        source += " ⚠️ Drop เกิน (แต่สาขาจำกัด 4W)"
                elif suggested == 'JB' or trip_drops > PUNTHAI_LIMITS['JB']['max_drops']:
                    # เช็คว่าสาขาอนุญาต 6W ไหม
                    if min_max_size >= 3:  # 6W
                        suggested = '6W'
                        source += " → 6W (Drop Limit)"
                    else:
                        # 🚫 สาขาจำกัดไม่เกิน JB - ห้ามใช้ 6W!
                        suggested = max_allowed_vehicle  # ใช้รถตามข้อจำกัดสาขา (JB หรือ 4W)
                        source += f" ⚠️ Drop เกิน (แต่สาขาจำกัด {max_allowed_vehicle})"
        
        # คำนวณ utilization - ใช้ limits ตาม BU type
        max_util_threshold = buffer * 100  # 100% หรือ 110% ตาม BU
        limits_for_util = PUNTHAI_LIMITS if is_punthai_only_trip else LIMITS
        if suggested in limits_for_util:
            w_util = (total_w / limits_for_util[suggested]['max_w']) * 100
            c_util = (total_c / limits_for_util[suggested]['max_c']) * 100
            max_util = max(w_util, c_util)
            
            # ถ้าเกิน threshold ตาม BU ต้องเพิ่มขนาดรถ
            if max_util > max_util_threshold:
                # 🚫 ห้าม upgrade เกินข้อจำกัดสาขา!
                if suggested == '4W' and min_max_size >= 2:
                    jb_util = max((total_w / limits_for_util['JB']['max_w']), (total_c / limits_for_util['JB']['max_c'])) * 100
                    if jb_util <= max_util_threshold:
                        suggested = 'JB'
                        source += " → JB"
                        w_util = (total_w / limits_for_util['JB']['max_w']) * 100
                        c_util = (total_c / limits_for_util['JB']['max_c']) * 100
                    elif min_max_size >= 3:  # สาขาอนุญาต 6W
                        suggested = '6W'
                        source += " → 6W"
                        w_util = (total_w / limits_for_util['6W']['max_w']) * 100
                        c_util = (total_c / limits_for_util['6W']['max_c']) * 100
                    else:
                        # 🚫 ไม่สามารถ upgrade ได้ (สาขาจำกัด JB) → ยังคงใช้ JB (จะเกิน buffer)
                        suggested = 'JB'
                        source += " ⚠️ เกินแต่สาขาจำกัด"
                        w_util = (total_w / limits_for_util['JB']['max_w']) * 100
                        c_util = (total_c / limits_for_util['JB']['max_c']) * 100
                elif suggested == 'JB' and min_max_size >= 3:  # สาขาอนุญาต 6W
                    suggested = '6W'
                    source += " → 6W"
                    w_util = (total_w / limits_for_util['6W']['max_w']) * 100
                    c_util = (total_c / limits_for_util['6W']['max_c']) * 100
                elif suggested == 'JB' and min_max_size < 3:
                    # 🚫 ไม่สามารถ upgrade เป็น 6W ได้ (สาขาจำกัด JB)
                    source += " ⚠️ เกินแต่สาขาจำกัด"
                elif suggested == '4W' and min_max_size < 2:
                    # 🚫 ไม่สามารถ upgrade ได้ (สาขาจำกัด 4W)
                    source += " ⚠️ เกินแต่สาขาจำกัด"
        else:
            w_util = c_util = 0
        
        # คำนวณระยะทางรวม - ใช้พิกัดจาก DataFrame โดยตรง
        total_distance = 0
        branch_coords = []
        for code in trip_codes:
            # ดึงพิกัดจาก df (มีคอลัมน์ _lat, _lon)
            branch_data = df[df['Code'] == code]
            if not branch_data.empty:
                lat = branch_data.iloc[0].get('_lat', 0)
                lon = branch_data.iloc[0].get('_lon', 0)
                if lat > 0 and lon > 0:
                    branch_coords.append((lat, lon))

        if branch_coords:
            # คำนวณระยะทางรวม: ลอง ROUTE_CACHE ก่อน; ถ้าไม่มีใช้ haversine×1.35 ประมาณ (ไม่เรียก network)
            _wp_td = [[DC_WANG_NOI_LAT, DC_WANG_NOI_LON]] + [[la, lo] for la, lo in branch_coords] + [[DC_WANG_NOI_LAT, DC_WANG_NOI_LON]]
            _ck_td = "|".join([f"{la:.4f},{lo:.4f}" for la, lo in _wp_td])
            if USE_CACHE and _ck_td in ROUTE_CACHE_DATA:
                _rc_td = ROUTE_CACHE_DATA[_ck_td]
                total_distance = _rc_td.get('distance', 0)
            else:
                # ประมาณจาก haversine×1.35 (zero-network) DC→b1→b2→...→DC
                _pts = _wp_td
                total_distance = sum(
                    haversine_distance(_pts[_pi][0], _pts[_pi][1], _pts[_pi+1][0], _pts[_pi+1][1], use_osrm_cache=False)
                    for _pi in range(len(_pts) - 1)
                )

        # 🚛 Fleet Constraint: ถ้าโควต้ารถประเภทนี้เต็ม → ลอง upgrade ไปรถใหญ่กว่า
        _sv = suggested  # บันทึกรถเดิม
        _sv_rank = _rank_fleet.get(suggested, 3)
        _upgraded_by_fleet = False
        while fleet_used.get(suggested, 0) >= _fleet_limits.get(suggested, 999):
            _next_rank = _sv_rank + 1
            if _next_rank > 3:
                # ไม่มีรถให้ upgrade → ใช้รถเดิม + เตือน
                source += " ⚠️ เกินโควต้า"
                break
            _next_veh = _fleet_rank.get(_next_rank, '6W')
            # เช็คว่าสาขาอนุญาตรถใหญ่กว่าไหม
            if _next_rank <= min_max_size:
                suggested = _next_veh
                _sv_rank = _next_rank
                _upgraded_by_fleet = True
                safe_print(f"      🚛 Fleet upgrade: Trip {trip_num} {_sv}→{suggested} (โควต้า {_sv} เต็ม {fleet_used.get(_sv,0)}/{_fleet_limits.get(_sv,999)})")
            else:
                # สาขาจำกัดไม่ให้ใช้รถใหญ่กว่า → ยังคงใช้รถเดิม + เตือน
                source += " ⚠️ เกินโควต้า"
                break
        if _upgraded_by_fleet:
            source += f" ↑ Fleet({_sv}→{suggested})"
            # คำนวณ utilization ใหม่ด้วยรถที่ upgrade
            if suggested in limits_for_util:
                w_util = (total_w / limits_for_util[suggested]['max_w']) * 100
                c_util = (total_c / limits_for_util[suggested]['max_c']) * 100
        fleet_used[suggested] = fleet_used.get(suggested, 0) + 1

        summary_data.append({
            'Trip': trip_num,
            'Branches': len(trip_codes),
            'Weight': total_w,
            'Cube': total_c,
            'Truck': f"{suggested} {source}",
            'BU_Type': trip_type,
            'Buffer': buffer_label,
            'Weight_Use%': w_util,
            'Cube_Use%': c_util,
            'Total_Distance': round(total_distance, 1)
        })
    
    # ==========================================
    # 🚨 Step 7.5: ตัดสาขาออกถ้าเกิน buffer หรือรถผิดประเภท (Strict Enforcement)
    # ==========================================
    safe_print("\n📋 Step 7.5: ตรวจสอบและตัดสาขาที่เกิน Buffer + ข้อจำกัดรถ...")
    overflow_branches = []
    # เพิ่มคอลัมน์ _overflow_reason เพื่อ track สาขาที่เกิน buffer
    if '_overflow_reason' not in df.columns:
        df['_overflow_reason'] = ''
    
    for i, trip_summary in enumerate(summary_data):
        trip_num = trip_summary['Trip']
        buffer_pct = float(trip_summary['Buffer'].replace('🅿️ ', '').replace('🅼 ', '').replace('%', ''))
        
        # ดึงข้อมูลทริป
        trip_data = df[df['Trip'] == trip_num].copy()
        if trip_data.empty:
            continue
            
        trip_codes = trip_data['Code'].tolist()
        
        # 🚗 หารถที่ถูกต้องตามข้อจำกัดสาขา (รถเล็กสุดที่รับโหลดได้)
        max_vehicles = [get_max_vehicle_for_branch(c) for c in trip_codes]
        vehicle_priority_map = {'4W': 1, 'JB': 2, '6W': 3}
        min_max_size = min(vehicle_priority_map.get(v, 3) for v in max_vehicles)
        max_allowed_v = {1: '4W', 2: 'JB', 3: '6W'}.get(min_max_size, '6W')
        
        # ดึง limits ตาม BU
        bu_type = trip_summary['BU_Type']
        is_punthai = (bu_type == 'punthai')
        buffer = punthai_buffer if is_punthai else maxmart_buffer
        limits = PUNTHAI_LIMITS if is_punthai else LIMITS
        
        # คำนวณน้ำหนัก/คิวก่อน (ใช้เลือกรถ)
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        
        # 🎯 เลือกรถตามภาค + ข้อจำกัดสาขา (เหนือ/ใต้ → รถใหญ่สุดที่อนุญาต)
        trip_region_75 = trip_data['_region_name'].iloc[0] if '_region_name' in trip_data.columns else 'ไม่ระบุ'
        is_long_haul_75 = str(trip_region_75) in ('เหนือ', 'ใต้')
        correct_vehicle = max_allowed_v  # fallback = รถใหญ่สุดที่อนุญาต
        if not is_long_haul_75:
            # ภาคอื่น: เล็กสุดที่รับโหลดได้
            for _veh in ['4W', 'JB', '6W']:
                _vr = vehicle_priority_map.get(_veh, 3)
                if _vr > min_max_size:
                    break  # ห้ามเกินข้อจำกัดสาขา
                _lim = limits[_veh]
                if (total_w <= _lim['max_w'] and                  # weight: hard limit, no buffer
                        total_c <= _lim['max_c'] * buffer and       # cube: ยืดได้ตาม buffer
                        len(trip_codes) <= _lim['max_drops']):
                    correct_vehicle = _veh
                    break  # เล็กสุดที่รับโหลดได้
        # เหนือ/ใต้: correct_vehicle = max_allowed_v แล้ว (ไม่ downgrade)
        
        max_w = limits[correct_vehicle]['max_w']           # น้ำหนัก = hard limit ห้ามเกิน
        max_c = limits[correct_vehicle]['max_c'] * buffer  # คิ้ว = ยืดได้ตาม buffer
        max_drops = limits[correct_vehicle]['max_drops']
        
        w_util = (total_w / limits[correct_vehicle]['max_w']) * 100
        c_util = (total_c / limits[correct_vehicle]['max_c']) * 100
        max_util = max(w_util, c_util)
        
        # อัพเดต summary ด้วยรถที่ถูกต้อง
        if is_long_haul_75:
            truck_source = "🚛 ไกล (เหนือ/ใต้)" if min_max_size >= 3 else "📋 จำกัดสาขา (เหนือ/ใต้)"
        elif min_max_size < 3:
            truck_source = "📋 จำกัดสาขา"
        elif vehicle_priority_map.get(correct_vehicle, 3) < min_max_size:
            truck_source = "🔽 Downgrade (ขนาดพอดี)"
        else:
            truck_source = "🤖 อัตโนมัติ"
        summary_data[i]['Truck'] = f"{correct_vehicle} {truck_source}"
        summary_data[i]['Weight_Use%'] = w_util
        summary_data[i]['Cube_Use%'] = c_util
        
        # ตรวจสอบว่าเกิน buffer หรือ drops หรือไม่
        is_over_buffer = total_w > max_w or total_c > max_c
        is_over_drops = len(trip_codes) > max_drops
        
        if is_over_buffer or is_over_drops:
            reason = "เกิน buffer" if is_over_buffer else f"เกิน drops ({len(trip_codes)}>{max_drops})"
            safe_print(f"   ✂️ Trip {trip_num} {reason}: {max_util:.1f}% → แยก 2 คัน")

            # 🚨 ถ้ามีแค่ 1 สาขา แต่เกิน buffer → ยอมรับ (ไม่สามารถแยกได้)
            if len(trip_data) <= 1:
                safe_print(f"      ⚠️ 1 สาขาแต่เกิน buffer → รับไว้ในทริปเดิม")
                continue

            # ── แยก 2 คันตรงๆ (ไม่ผ่าน overflow limbo) ──────────────────────
            # เรียงสาขา: ใกล้ DC ก่อน (สาขาใกล้อยู่คันแรก, ไกลไปคันสอง)
            trip_data_sorted = trip_data.sort_values('_distance_from_dc', ascending=True)

            truck_str = correct_vehicle
            if truck_str not in limits:
                continue
            _sp_max_w = limits[truck_str]['max_w']
            _sp_max_c = limits[truck_str]['max_c'] * buffer
            _sp_max_d = limits[truck_str]['max_drops']

            # สร้าง unit list ที่รักษา group lock (ตัดทีละ group ไม่ใช่ทีละสาขา)
            _sp_units = []   # [(group_codes, w, c)]
            _sp_seen  = set()
            for _, _sp_row in trip_data_sorted.iterrows():
                _sp_c  = _sp_row['Code']
                _sp_cu = str(_sp_c).strip().upper()
                if _sp_cu in _sp_seen:
                    continue
                _gid = BRANCH_TO_GROUP.get(_sp_cu)
                if _gid:
                    _grp = [c for c in BRANCH_GROUPS.get(_gid, [])
                            if str(c).strip().upper() in {str(x).strip().upper() for x in trip_data['Code']}
                            and str(c).strip().upper() not in _sp_seen]
                else:
                    _grp = [_sp_c]
                if not _grp:
                    _grp = [_sp_c]  # fallback: สาขานี้เป็น unit เดี่ยว (ป้องกันหาย)
                _grp_w = sum(float(trip_data[trip_data['Code'].apply(lambda x: str(x).strip().upper()==str(g).strip().upper())]['Weight'].sum()) for g in _grp)
                _grp_c = sum(float(trip_data[trip_data['Code'].apply(lambda x: str(x).strip().upper()==str(g).strip().upper())]['Cube'].sum())   for g in _grp)
                _sp_units.append((_grp, _grp_w, _grp_c))
                for g in _grp:
                    _sp_seen.add(str(g).strip().upper())

            # safety: สาขาที่ไม่อยู่ใน units (edge case) → เพิ่มเป็น unit เดี่ยว
            _all_in_units = {str(c).strip().upper() for _grp, _, _ in _sp_units for c in _grp}
            for _, _sr in trip_data.iterrows():
                _sc_chk = str(_sr['Code']).strip().upper()
                if _sc_chk not in _all_in_units:
                    _sr_w = float(_sr.get('Weight', 0) or 0)
                    _sr_c = float(_sr.get('Cube', 0) or 0)
                    _sp_units.append(([_sr['Code']], _sr_w, _sr_c))

            # แบ่ง: คันแรก = trip_num (เติมจนเต็ม), คันสอง = new trip
            _trip1_codes = []; _trip1_w = 0; _trip1_c = 0; _trip1_d = 0
            _trip2_codes = []
            for _grp, _gw, _gc2 in _sp_units:
                if (_trip1_w + _gw <= _sp_max_w and
                        _trip1_c + _gc2 <= _sp_max_c and
                        _trip1_d + len(_grp) <= _sp_max_d):
                    _trip1_codes.extend(_grp)
                    _trip1_w += _gw; _trip1_c += _gc2; _trip1_d += len(_grp)
                else:
                    _trip2_codes.extend(_grp)

            # ถ้าคันแรกว่างเปล่า (unit แรกก็เกินแล้ว) → บังคับใส่อย่างน้อย 1 unit
            if not _trip1_codes and _sp_units:
                _grp0, _gw0, _gc0 = _sp_units[0]
                _trip1_codes.extend(_grp0)
                _trip1_upper0 = {str(c).strip().upper() for c in _grp0}
                _trip2_codes = [c for c in _trip2_codes if str(c).strip().upper() not in _trip1_upper0]

            # Assign คันแรก (ทริปเดิม) — ไม่แตะสาขาที่ไม่อยู่ใน trip_data
            _trip1_upper = {str(c).strip().upper() for c in _trip1_codes}
            _trip2_upper = {str(c).strip().upper() for c in _trip2_codes}
            for _, _sp_row in trip_data.iterrows():
                _sp_code = _sp_row['Code']
                _sp_code_up = str(_sp_code).strip().upper()
                if _sp_code_up not in _trip1_upper and _sp_code_up not in _trip2_upper:
                    # ตกหล่น → ใส่คันสองเสมอ
                    _trip2_codes.append(_sp_code)
                    _trip2_upper.add(_sp_code_up)
                if _sp_code_up not in _trip1_upper:
                    df.loc[df['Code'] == _sp_code, 'Trip'] = 0  # ชั่วคราว (จะ assign ทริปสอง)

            # Assign คันสอง (ทริปใหม่)
            if _trip2_codes:
                _new_split_trip = int(df['Trip'].max()) + 1
                _trip2_upper = {str(c).strip().upper() for c in _trip2_codes}
                for _sp2_code in _trip2_codes:
                    df.loc[df['Code'].apply(lambda x: str(x).strip().upper()==str(_sp2_code).strip().upper()), 'Trip'] = _new_split_trip

                # vehicle สำหรับคันสอง
                _t2_data = df[df['Trip'] == _new_split_trip]
                _t2_w = _t2_data['Weight'].sum(); _t2_c = _t2_data['Cube'].sum()
                _t2_is_pt = all(str(df[df['Code']==c]['BU'].values[0] if len(df[df['Code']==c])>0 else '').upper() in ['211','PUNTHAI'] for c in _trip2_codes)
                _t2_lims  = PUNTHAI_LIMITS if _t2_is_pt else LIMITS
                _t2_buf   = punthai_buffer if _t2_is_pt else maxmart_buffer
                _t2_buf_lbl = f"🅿️ {int(_t2_buf*100)}%" if _t2_is_pt else f"🅼 {int(_t2_buf*100)}%"
                _t2_veh   = correct_vehicle
                for _v2 in ['4W','JB','6W']:
                    _l2 = _t2_lims[_v2]
                    if _t2_w <= _l2['max_w'] and _t2_c <= _l2['max_c']*_t2_buf:
                        _t2_veh = _v2; break
                summary_data.append({
                    'Trip': _new_split_trip,
                    'Branches': len(_t2_data),
                    'Weight': _t2_w,
                    'Cube': _t2_c,
                    'Truck': f'{_t2_veh} ✂️ แยก',
                    'BU_Type': 'punthai' if _t2_is_pt else 'mixed',
                    'Buffer': _t2_buf_lbl,
                    'Weight_Use%': (_t2_w / _t2_lims[_t2_veh]['max_w']) * 100,
                    'Cube_Use%': (_t2_c / _t2_lims[_t2_veh]['max_c']) * 100,
                    'Total_Distance': 0
                })
                safe_print(f"      ✂️ คันสอง: Trip {_new_split_trip} ({len(_t2_data)} สาขา, {_t2_w:.0f}kg, {_t2_veh})")

            # อัปเดต summary คันแรก
            _t1_data = df[df['Trip'] == trip_num]
            _t1_w = _t1_data['Weight'].sum(); _t1_c = _t1_data['Cube'].sum()
            summary_data[i]['Branches'] = len(_t1_data)
            summary_data[i]['Weight']   = _t1_w
            summary_data[i]['Cube']     = _t1_c
            summary_data[i]['Weight_Use%'] = (_t1_w / limits[truck_str]['max_w']) * 100
            summary_data[i]['Cube_Use%']   = (_t1_c / limits[truck_str]['max_c']) * 100
            summary_data[i]['Truck'] = f'{truck_str} ✂️ แยก'
            safe_print(f"      ✂️ คันแรก: Trip {trip_num} ({len(_t1_data)} สาขา, {_t1_w:.0f}kg, {truck_str})")
    
    summary_df = pd.DataFrame(summary_data)
    
    # ==========================================
    # Step 8: เพิ่มคอลัมน์เสริม
    # ==========================================
    # เพิ่มคอลัมน์รถ
    trip_truck_map = {}
    for _, row in summary_df.iterrows():
        trip_truck_map[row['Trip']] = row['Truck']
    df['Truck'] = df['Trip'].map(trip_truck_map)
    
    # เพิ่มคอลัมน์ Region
    df['Region'] = df['_region_name']
    
    # เพิ่มคอลัมน์ Province/District/Subdistrict (ถ้ายังไม่มี)
    if 'Province' not in df.columns:
        df['Province'] = df['_province']
    if 'District' not in df.columns:
        df['District'] = df['_district']
    if 'Subdistrict' not in df.columns:
        df['Subdistrict'] = df['_subdistrict']
    
    # เพิ่มคอลัมน์ระยะทางจาก DC
    df['Distance_from_DC'] = df['_distance_from_dc'].round(1)
    
    # เพิ่มคอลัมน์ MaxVehicle constraint
    df['MaxVehicle'] = df['_max_vehicle']
    
    # 🚨 เพิ่มคอลัมน์เช็ครถ - ตรวจสอบว่ารถที่จัดตรงกับข้อจำกัดหรือไม่
    def check_vehicle_compliance(row):
        """ตรวจสอบว่ารถที่จัดไปตรงกับข้อจำกัดหรือไม่"""
        if row['Trip'] == 0:
            return '⚠️ ไม่ได้จัด'
        
        max_allowed = row['_max_vehicle']
        truck_assigned = str(row.get('Truck', '')).split()[0] if pd.notna(row.get('Truck')) else ''
        
        # แปลง JB เป็น 4WJ ถ้าจำเป็น
        if truck_assigned == '4WJ':
            truck_assigned = 'JB'
        
        # Vehicle hierarchy: 4W < JB < 6W
        vehicle_rank = {'4W': 1, 'JB': 2, '6W': 3}
        
        if max_allowed not in vehicle_rank or truck_assigned not in vehicle_rank:
            return '✅ ใช้ได้'
        
        # ตรวจสอบว่ารถที่จัดเล็กกว่าหรือเท่ากับรถที่อนุญาต
        if vehicle_rank[truck_assigned] <= vehicle_rank[max_allowed]:
            return '✅ ใช้ได้'
        else:
            return f'❌ เกินข้อจำกัด (Max: {max_allowed}, ใช้: {truck_assigned})'
    
    df['VehicleCheck'] = df.apply(check_vehicle_compliance, axis=1)
    
    # ==========================================
    # 🚨 Step 8.5: บังคับแก้ไขสาขาที่เกินข้อจำกัดรถ (Enforce Vehicle Constraints)
    # ==========================================
    safe_print("\n📋 Step 8.5: บังคับข้อจำกัดรถ...")
    vehicle_violations = df[df['VehicleCheck'].str.contains('❌', na=False)]
    
    if len(vehicle_violations) > 0:
        safe_print(f"   ⚠️ พบ {len(vehicle_violations)} สาขาที่ใช้รถเกินข้อจำกัด")
        
        # แยกสาขาที่เกินข้อจำกัดออกมาจัดทริปใหม่
        for _, viol_row in vehicle_violations.iterrows():
            viol_code = viol_row['Code']
            viol_trip = viol_row['Trip']
            max_allowed = viol_row['_max_vehicle']
            
            # หาสาขาอื่นในทริปเดียวกันที่มีข้อจำกัดเดียวกันหรือน้อยกว่า
            same_trip = df[df['Trip'] == viol_trip]
            
            # ตรวจสอบว่าสาขาอื่นในทริปมีข้อจำกัดอย่างไร
            vehicle_rank = {'4W': 1, 'JB': 2, '6W': 3}
            max_allowed_rank = vehicle_rank.get(max_allowed, 3)
            
            # หาสาขาที่ทำให้ต้องใช้รถใหญ่ (น้ำหนัก/คิวมาก หรือ max vehicle ใหญ่กว่า)
            other_branches = same_trip[same_trip['Code'] != viol_code]
            
            if len(other_branches) > 0:
                # ตรวจสอบว่าสาขาอื่นมีข้อจำกัดใหญ่กว่าหรือไม่
                other_max_vehicles = other_branches['_max_vehicle'].apply(lambda x: vehicle_rank.get(x, 3))
                min_other_rank = other_max_vehicles.min()
                
                if min_other_rank > max_allowed_rank:
                    # สาขาอื่นมีข้อจำกัดใหญ่กว่า → ย้ายสาขานี้ออก
                    df.loc[df['Code'] == viol_code, 'Trip'] = 0  # ย้ายออกไปจัดใหม่
                    safe_print(f"      🔄 ย้าย {viol_code} ออกจาก Trip {viol_trip} (Max: {max_allowed})")
    
    # จัดทริปใหม่สำหรับสาขาที่ถูกย้ายออก
    unassigned_violations = df[df['Trip'] == 0]
    if len(unassigned_violations) > 0:
        safe_print(f"   📦 จัดทริปใหม่สำหรับ {len(unassigned_violations)} สาขา...")
        max_trip = df[df['Trip'] > 0]['Trip'].max() if len(df[df['Trip'] > 0]) > 0 else 0
        
        # จัดกลุ่มตาม max_vehicle
        for max_veh in ['4W', 'JB', '6W']:
            veh_branches = unassigned_violations[unassigned_violations['_max_vehicle'] == max_veh]
            if len(veh_branches) == 0:
                continue
            
            # สร้างทริปใหม่สำหรับสาขาที่มี max_vehicle เดียวกัน
            new_trip = max_trip + 1
            
            # เช็คว่าเป็น Punthai หรือไม่
            is_punthai = False
            if 'BU' in veh_branches.columns and len(veh_branches) > 0:
                bu_val = str(veh_branches['BU'].iloc[0]).upper()
                is_punthai = bu_val in ['211', 'PUNTHAI']
            limits = PUNTHAI_LIMITS if is_punthai else LIMITS
            _veh_buffer = punthai_buffer if is_punthai else maxmart_buffer
            
            current_w = 0
            current_c = 0
            current_drops = 0
            max_w = limits[max_veh]['max_w'] * _veh_buffer
            max_c = limits[max_veh]['max_c'] * _veh_buffer
            max_d = limits[max_veh]['max_drops']
            
            for _, br in veh_branches.iterrows():
                br_w = br['Weight']
                br_c = br['Cube']
                
                if current_w + br_w > max_w or current_c + br_c > max_c or current_drops >= max_d:
                    # ปิดทริปปัจจุบัน เริ่มทริปใหม่
                    new_trip += 1
                    current_w = 0
                    current_c = 0
                    current_drops = 0
                
                df.loc[df['Code'] == br['Code'], 'Trip'] = new_trip
                current_w += br_w
                current_c += br_c
                current_drops += 1
            
            max_trip = new_trip
            safe_print(f"      ✅ จัด {len(veh_branches)} สาขา {max_veh} เสร็จ")
        
        # อัพเดต Truck และ VehicleCheck หลังจัดใหม่
        for trip_num in df[df['Trip'] > 0]['Trip'].unique():
            trip_codes = df[df['Trip'] == trip_num]['Code'].tolist()
            max_vehicles = [get_max_vehicle_for_branch(c) for c in trip_codes]
            _vp_local = {'4W': 1, 'JB': 2, '6W': 3}  # local copy — ไม่ shadow outer vehicle_priority
            min_rank = min(_vp_local.get(v, 3) for v in max_vehicles)
            suggested = {1: '4W', 2: 'JB', 3: '6W'}.get(min_rank, '6W')
            df.loc[df['Trip'] == trip_num, 'Truck'] = f"{suggested} 📋 จัดใหม่"
        
        df['VehicleCheck'] = df.apply(check_vehicle_compliance, axis=1)

    # ==========================================
    # Step 8.8: 🔒 FINAL REGION & BKK ISOLATION AUDIT
    # รันหลังทุก step เพื่อรับประกันไม่มีทริปที่ปนภาค/ปนกรุงเทพฯ
    # ==========================================
    safe_print("\n🔒 Step 8.8: Final Region & BKK Isolation Audit...")
    _BKK_PROV = 'กรุงเทพมหานคร'
    _final_audit_fixed = 0
    _fa_max_trip = df[df['Trip'] > 0]['Trip'].max() if len(df[df['Trip'] > 0]) > 0 else 0

    for _fa_trip in sorted(df[df['Trip'] > 0]['Trip'].unique()):
        _fa_data = df[df['Trip'] == _fa_trip]
        _fa_provs = [str(r.get('_province', '') or '') for _, r in _fa_data.iterrows()]
        _fa_provs_clean = [p for p in _fa_provs if p and p != 'nan']

        # 1️⃣ BKK Isolation: กรุงเทพฯ ห้ามปนกับจังหวัดอื่น
        _fa_has_bkk = _BKK_PROV in _fa_provs_clean
        _fa_has_non_bkk = any(p != _BKK_PROV for p in _fa_provs_clean)
        if _fa_has_bkk and _fa_has_non_bkk:
            # แยกสาขาที่ไม่ใช่กรุงเทพฯ ออก
            _fa_split_codes = [
                r['Code'] for _, r in _fa_data.iterrows()
                if str(r.get('_province', '') or '') != _BKK_PROV
            ]
            if _fa_split_codes:
                _fa_max_trip += 1
                df.loc[df['Code'].isin(_fa_split_codes), 'Trip'] = _fa_max_trip
                safe_print(f"   🔒 BKK AUDIT: Trip {_fa_trip} → แยก {len(_fa_split_codes)} สาขา non-BKK → Trip {_fa_max_trip}")
                _final_audit_fixed += 1
            continue  # ตรวจข้ออื่นบนข้อมูลใหม่ในรอบถัดไป

        # 2️⃣ Region Mixing: ห้ามปนภาค
        _fa_regions: dict = {}
        for _, _far in _fa_data.iterrows():
            _fap = str(_far.get('_province', '') or '')
            _fareg = get_region_name(_fap) if _fap and _fap != 'nan' else ''
            if not _fareg or _fareg == 'ไม่ระบุ':
                _fareg = str(_far.get('_region_name', '') or '')
            if _fareg and _fareg != 'ไม่ระบุ':
                _fa_regions[_fareg] = _fa_regions.get(_fareg, 0) + 1
        if len(_fa_regions) <= 1:
            continue  # clean — no mixing

        # พบการปนภาค → dominant = ภาคที่มีสาขามากสุด
        _fa_region_order = ['เหนือ', 'อีสาน', 'ตะวันออก', 'กลาง', 'ตะวันตก', 'ใต้']
        _fa_dominant = max(
            _fa_regions,
            key=lambda k: (_fa_regions[k], -(_fa_region_order.index(k) if k in _fa_region_order else 99))
        )
        _fa_minority_codes = []
        for _, _far2 in _fa_data.iterrows():
            _fap2 = str(_far2.get('_province', '') or '')
            _fareg2 = get_region_name(_fap2) if _fap2 and _fap2 != 'nan' else ''
            if not _fareg2 or _fareg2 == 'ไม่ระบุ':
                _fareg2 = str(_far2.get('_region_name', '') or '')
            if _fareg2 and _fareg2 != 'ไม่ระบุ' and _fareg2 != _fa_dominant:
                _fa_minority_codes.append(_far2['Code'])
        if _fa_minority_codes:
            # 🔒 GROUP GUARD: ห้ามแยก code ที่มี group member อยู่ใน majority
            _fa_majority_up = {str(r['Code']).strip().upper() for _, r in _fa_data.iterrows()
                               if r['Code'] not in _fa_minority_codes}
            _fa_minority_safe = []
            for _fmc in _fa_minority_codes:
                _fmc_up = str(_fmc).strip().upper()
                _fmc_gid = BRANCH_TO_GROUP.get(_fmc_up)
                if _fmc_gid:
                    _fmc_sibs = {str(s).strip().upper() for s in BRANCH_GROUPS.get(_fmc_gid, [])}
                    if _fmc_sibs & _fa_majority_up:
                        safe_print(f"   🔒 FA AUDIT GUARD: ข้าม {_fmc} — group member อยู่ใน majority")
                        continue
                _fa_minority_safe.append(_fmc)
            _fa_minority_codes = _fa_minority_safe
        if _fa_minority_codes:
            _fa_max_trip += 1
            df.loc[df['Code'].isin(_fa_minority_codes), 'Trip'] = _fa_max_trip
            safe_print(f"   🔒 REGION AUDIT: Trip {_fa_trip} ปนภาค {_fa_regions} → แยก {_fa_minority_codes} → Trip {_fa_max_trip}")
            _final_audit_fixed += 1

    if _final_audit_fixed > 0:
        safe_print(f"   ✅ Final Audit: แก้ไข {_final_audit_fixed} ทริป")
        # Renumber trips after final audit
        _fa_rem = sorted(df[df['Trip'] > 0]['Trip'].unique())
        _fa_ren = {old: new for new, old in enumerate(_fa_rem, start=1)}
        df['Trip'] = df['Trip'].map(lambda x: _fa_ren.get(x, x) if x > 0 else x)
        # อัพเดต Truck mapping หลัง renumber
        try:
            for _fa_t in df[df['Trip'] > 0]['Trip'].unique():
                if _fa_t not in trip_truck_map:
                    _fa_codes = df[df['Trip'] == _fa_t]['Code'].tolist()
                    _fa_vp = {'4W': 1, 'JB': 2, '6W': 3}
                    _fa_max_veh_list = [branch_max_vehicle_cache.get(str(c).strip().upper(), '6W') for c in _fa_codes]
                    _fa_min_rank = min(_fa_vp.get(v, 3) for v in _fa_max_veh_list)
                    _fa_truck = {1: '4W', 2: 'JB', 3: '6W'}.get(_fa_min_rank, '6W')
                    df.loc[df['Trip'] == _fa_t, 'Truck'] = f"{_fa_truck} ✂️ audit-split"
        except Exception:
            pass
        # อัพเดต trip_truck_map ใหม่
        trip_truck_map = {}
        for _fa_t2 in df[df['Trip'] > 0]['Trip'].unique():
            _fa_td = df[df['Trip'] == _fa_t2]
            if not _fa_td.empty:
                _raw_trk = str(_fa_td.iloc[0].get('Truck', '6W') or '6W').split()[0]
                trip_truck_map[_fa_t2] = _raw_trk
        df['Truck'] = df['Trip'].map(trip_truck_map)
    else:
        safe_print("   ✅ Final Audit: ไม่พบการปนภาค/BKK")

    # ==========================================
    # Step 8.9: Catch-all — สาขาที่ยังไม่ได้จัดทริป (Trip=0)
    # รองรับ Z*, LUBE, SUPPLY, USE, สาขาไม่มีพิกัด ฯลฯ
    # ==========================================
    _catchall_remaining = df[df['Trip'] == 0].copy()
    if len(_catchall_remaining) > 0:
        safe_print(f"\n⚠️  Step 8.9: พบ {len(_catchall_remaining)} สาขายังไม่ได้จัดทริป → จัดทริปเดี่ยว...")
        _ca_max_trip = safe_int_trip(df[df['Trip'] > 0]['Trip'].max()) if len(df[df['Trip'] > 0]) > 0 else 0
        for _, _ca_row in _catchall_remaining.iterrows():
            _ca_max_trip += 1
            _ca_code = _ca_row['Code']
            df.loc[df['Code'] == _ca_code, 'Trip'] = _ca_max_trip
            _ca_veh = branch_max_vehicle_cache.get(str(_ca_code).strip().upper(), '6W')
            df.loc[df['Code'] == _ca_code, 'Truck'] = f"{_ca_veh} ⚙️ จัดเดี่ยว"
            safe_print(f"   ➕ {_ca_code} → Trip {_ca_max_trip} ({_ca_veh})")
        safe_print(f"   ✅ Step 8.9: จัดทริปเพิ่ม {len(_catchall_remaining)} สาขา")

    # ==========================================
    # Step 8.95: REPACK OVERWEIGHT TRIPS (Fill-first bin-packing)
    # ==========================================
    safe_print("\n📦 Step 8.95: ตรวจสอบและจัดใหม่ทริปที่เกิน weight/cube limit...")
    
    # ตรวจสอบ trip ไหนเกิน limit
    _repack_trips_list = []
    for _ov_trip in sorted(df[df['Trip'] > 0]['Trip'].unique()):
        _ov_df = df[df['Trip'] == _ov_trip].copy()
        _ov_codes = _ov_df['Code'].tolist()
        if len(_ov_codes) < 2:
            continue
        
        # หา vehicle limit
        _ov_is_pt = all(branch_bu_cache.get(str(c).strip().upper(), False) for c in _ov_codes)
        _ov_lims = PUNTHAI_LIMITS if _ov_is_pt else LIMITS
        _ov_allowed = get_allowed_from_codes(_ov_codes, ['4W', 'JB', '6W'])
        _ov_veh = next((v for v in ['6W', 'JB', '4W'] if v in _ov_allowed), '6W')
        _ov_max_w = _ov_lims[_ov_veh]['max_w']
        _ov_max_c = _ov_lims[_ov_veh]['max_c']
        _ov_total_w = _ov_df['Weight'].sum()
        _ov_total_c = _ov_df['Cube'].sum()
        
        if _ov_total_w <= _ov_max_w and _ov_total_c <= _ov_max_c:
            continue  # ไม่เกิน ข้ามไป
        
        _repack_trips_list.append((_ov_trip, _ov_df, _ov_veh, _ov_max_w, _ov_max_c))
    
    # Repack trips ที่เกิน limit ด้วย bin-packing (fill-first strategy)
    _repack_count = 0
    _orig_max_trip_repack = safe_int_trip(df['Trip'].max())
    _next_trip_id_repack = _orig_max_trip_repack + 1
    for _ov_trip, _ov_df, _ov_veh, _ov_max_w, _ov_max_c in _repack_trips_list:
        # เรียงสาขาตาม weight มากสุดก่อน (First-Fit Decreasing heuristic)
        _ov_sorted = _ov_df.sort_values('Weight', ascending=False).reset_index(drop=True)
        _branches_to_assign = list(_ov_sorted['Code'])
        
        # Bin-packing: Fill current trip then create new trips
        _current_trip_id = _ov_trip
        _current_w = 0
        _current_c = 0
        _current_items = []
        _repack_map = {}  # map branch code -> new trip id
        
        for _branch_code in _branches_to_assign:
            _branch_row = _ov_sorted[_ov_sorted['Code'] == _branch_code].iloc[0]
            _rw = float(_branch_row.get('Weight', 0) or 0)
            _rc = float(_branch_row.get('Cube', 0) or 0)
            
            # ถ้าเพิ่มสาขานี้เข้าไปจะเกิน limit → สร้างทริปใหม่
            if _current_items and (_current_w + _rw > _ov_max_w or _current_c + _rc > _ov_max_c):
                # บันทึก mapping สำหรับ items ในทริปปัจจุบัน
                for _item_code in _current_items:
                    _repack_map[_item_code] = _current_trip_id
                # เรียมตัวแปรสำหรับทริปใหม่
                _current_trip_id = _next_trip_id_repack
                _next_trip_id_repack += 1
                _current_w = 0
                _current_c = 0
                _current_items = []
            
            # เพิ่มสาขาเข้าทริปปัจจุบัน
            _current_items.append(_branch_code)
            _current_w += _rw
            _current_c += _rc
        
        # บันทึก items ที่เหลือในทริปปัจจุบัน
        for _item_code in _current_items:
            _repack_map[_item_code] = _current_trip_id
        
        # Apply mapping to dataframe (อัพเดต Trip ของสาขาที่ย้ายไป)
        _moved_count = 0
        for _code, _new_trip_id in _repack_map.items():
            if _new_trip_id != _ov_trip:
                df.loc[df['Code'] == _code, 'Trip'] = _new_trip_id
                _moved_count += 1
                _repack_count += 1
        
        # Log ผล repacking
        _groups = {}
        for _code, _tid in _repack_map.items():
            if _tid not in _groups:
                _groups[_tid] = 0
            _groups[_tid] += 1
        
        _trip_list = ', '.join([f"{_tid}({_groups[_tid]})" for _tid in sorted(_groups.keys())])
        if _moved_count > 0:
            safe_print(f"   📦 Trip {_ov_trip} → {_trip_list} [{_ov_veh}] (bin-packing)")
    
    if _repack_count > 0:
        safe_print(f"   ✅ Step 8.95: จัดใหม่ {_repack_count} สาขา → เต็มทริปแรกก่อน (fill-first)")
    else:
        safe_print(f"   ✅ Step 8.95: ทุกทริปอยู่ในลิมิต (ไม่ต้องจัดใหม่)")

    # ==========================================
    # Step 8.96: Merge spill trips เข้าทริปที่มี capacity เหลือ
    # (เศษจาก repack → รวมกับจังหวัดเดิม หรือจังหวัดถัดไปในลำดับ)
    # ==========================================
    _spill_trip_ids = set(range(_orig_max_trip_repack + 1, _next_trip_id_repack))
    if _spill_trip_ids:
        safe_print(f"\n🔗 Step 8.96: รวม {len(_spill_trip_ids)} spill trips เข้าทริปใกล้เคียง...")
        _sp_prov_col = '_province' if '_province' in df.columns else ('Province' if 'Province' in df.columns else None)
        _merged_spills = 0

        # รวบรวม metadata ของทุก original trip (ไม่ใช่ spill)
        _orig_trip_meta: dict = {}  # trip_id → {prov, w, c, max_w, max_c, veh, is_pt}
        for _ot in sorted(df[(df['Trip'] > 0) & (~df['Trip'].isin(_spill_trip_ids))]['Trip'].unique()):
            _odf = df[df['Trip'] == _ot]
            _oprov = ''
            if _sp_prov_col:
                _opvc = _odf[_sp_prov_col].dropna().value_counts()
                if len(_opvc): _oprov = str(_opvc.index[0])
            _ois_pt = all(branch_bu_cache.get(str(c).strip().upper(), False) for c in _odf['Code'])
            _olims = PUNTHAI_LIMITS if _ois_pt else LIMITS
            _oallowed = get_allowed_from_codes(_odf['Code'].tolist(), ['4W', 'JB', '6W'])
            _oveh = next((v for v in ['6W', 'JB', '4W'] if v in _oallowed), '6W')
            _obuf = punthai_buffer if _ois_pt else maxmart_buffer
            _orig_trip_meta[_ot] = {
                'prov': _oprov, 'is_pt': _ois_pt,
                'w': float(_odf['Weight'].sum()), 'c': float(_odf['Cube'].sum()),
                'max_w': _olims[_oveh]['max_w'],
                'max_c': _olims[_oveh]['max_c'] * _obuf,
                'veh': _oveh,
            }

        for _sp_tid in sorted(_spill_trip_ids):
            _sp_df = df[df['Trip'] == _sp_tid].copy()
            if _sp_df.empty:
                continue
            _sp_prov = ''
            if _sp_prov_col:
                _spvc = _sp_df[_sp_prov_col].dropna().value_counts()
                if len(_spvc): _sp_prov = str(_spvc.index[0])
            _sp_w = float(_sp_df['Weight'].sum())
            _sp_c = float(_sp_df['Cube'].sum())
            _sp_is_pt = all(branch_bu_cache.get(str(c).strip().upper(), False) for c in _sp_df['Code'])
            _sp_allowed = get_allowed_from_codes(_sp_df['Code'].tolist(), ['4W', 'JB', '6W'])
            _sp_veh = next((v for v in ['6W', 'JB', '4W'] if v in _sp_allowed), '6W')

            # ผ่านสองรอบ: รอบ 1 = จังหวัดเดิม, รอบ 2 = จังหวัดอื่นในภาคเดียวกัน
            _best_cand = None
            _sp_region = get_region_name(_sp_prov)
            for _pass in (1, 2):
                for _cot, _cmeta in _orig_trip_meta.items():
                    if _pass == 1 and _cmeta['prov'] != _sp_prov:
                        continue
                    if _pass == 2 and _cmeta['prov'] == _sp_prov:
                        continue
                    if _pass == 2 and get_region_name(_cmeta['prov']) != _sp_region:
                        continue
                    if _cmeta['veh'] != _sp_veh:
                        continue
                    if _cmeta['w'] + _sp_w <= _cmeta['max_w'] and _cmeta['c'] + _sp_c <= _cmeta['max_c']:
                        _best_cand = _cot
                        break
                if _best_cand is not None:
                    break

            if _best_cand is not None:
                df.loc[df['Trip'] == _sp_tid, 'Trip'] = _best_cand
                # อัพเดต meta ของ target trip
                _orig_trip_meta[_best_cand]['w'] += _sp_w
                _orig_trip_meta[_best_cand]['c'] += _sp_c
                _merged_spills += 1
                safe_print(f"   🔗 Spill {_sp_tid} ({_sp_prov} {_sp_w:.0f}kg) → Trip {_best_cand} ({_orig_trip_meta[_best_cand]['prov']})")
            else:
                safe_print(f"   ⚠️ Spill {_sp_tid} ({_sp_prov} {_sp_w:.0f}kg) → ไม่มีทริปรองรับ คงเป็นทริปแยก")

        safe_print(f"   ✅ Step 8.96: รวม {_merged_spills}/{len(_spill_trip_ids)} spill trips สำเร็จ")

    # ==========================================
    # Step 8.97: รวมทริปใกล้กันที่ยังว่างอยู่ (Consolidate nearby underutilized trips)
    # เป้าหมาย: ลดรถว่างโดยรวมทริปเล็กๆ ใกล้กันเข้าด้วยกัน
    # เงื่อนไข: จังหวัดเดียวกัน + ระยะ centroid ≤ 40km + รวมแล้วไม่เกิน capacity
    # ==========================================
    safe_print("\n🔗 Step 8.97: รวมทริปใกล้กันที่รถว่าง...")
    _CONSOL_MAX_KM   = 20.0   # ระยะ centroid สูงสุดระหว่างทริปที่จะรวม
    _CONSOL_MIN_UTIL = 0.98   # มาตรฐาน 98% — รวมทริปที่ยังไม่ถึง 98%
    _consol_merged   = 0

    def _trip_meta_consol(trip_id):
        """คืน metadata ของทริป: w, c, prov, centroid lat/lon, max_w, max_c, allow, is_pt, buf"""
        _td = df[df['Trip'] == trip_id]
        if _td.empty:
            return None
        _codes = _td['Code'].tolist()
        _is_pt = all(branch_bu_cache.get(str(c).strip().upper(), False) for c in _codes)
        _lims  = PUNTHAI_LIMITS if _is_pt else LIMITS
        _buf   = punthai_buffer if _is_pt else maxmart_buffer
        _allow = get_allowed_from_codes(_codes, ['4W', 'JB', '6W'])
        _veh   = next((v for v in ['6W', 'JB', '4W'] if v in _allow), '6W')
        _mw    = _lims[_veh]['max_w']
        _mc    = _lims[_veh]['max_c']
        _w     = float(_td['Weight'].sum())
        _c     = float(_td['Cube'].sum())
        _prov_vc = _td['_province'].dropna().value_counts()
        _prov  = str(_prov_vc.index[0]) if len(_prov_vc) else ''
        _sd_col = next((c for c in ['Subdistrict', '_subdistrict', 'ตำบล'] if c in _td.columns), None)
        _sd_vc  = _td[_sd_col].dropna().value_counts() if _sd_col else None
        _subd   = str(_sd_vc.index[0]) if (_sd_vc is not None and len(_sd_vc)) else ''
        _lats  = _td['_lat'].dropna().to_numpy(dtype=float)
        _lons  = _td['_lon'].dropna().to_numpy(dtype=float)
        _valid = (_lats > 0) & (_lons > 0)
        _clat  = float(_lats[_valid].mean()) if _valid.any() else 0.0
        _clon  = float(_lons[_valid].mean()) if _valid.any() else 0.0
        _pts   = list(zip(_lats[_valid].tolist(), _lons[_valid].tolist()))
        return {'w': _w, 'c': _c, 'prov': _prov, 'subd': _subd,
                'clat': _clat, 'clon': _clon, 'pts': _pts,
                'max_w': _mw, 'max_c': _mc, 'buf': _buf, 'allow': _allow,
                'veh': _veh, 'is_pt': _is_pt, 'codes': _codes}

    # วนซ้ำจนกว่าไม่มีการรวมอีก (chain merges)
    for _cs_round in range(10):
        _cs_changed = False
        _all_trips_cs = sorted(df[df['Trip'] > 0]['Trip'].unique())
        # สร้าง metadata ทุกทริปครั้งเดียว
        _meta_cs = {t: _trip_meta_consol(t) for t in _all_trips_cs}
        _meta_cs = {t: m for t, m in _meta_cs.items() if m}

        # ── Pre-compute OSRM road distance ทุกคู่สาขาจริง ครั้งเดียวต่อ round ──
        # รวมพิกัดสาขาทุกตัวจากทุกทริป พร้อม index กลับหาทริป
        _all_pts_cs: list = []   # [(lat,lon), ...]
        _pt_trip_cs: list = []   # trip id ของแต่ละ point
        for _t, _m in _meta_cs.items():
            for _pt in _m.get('pts', []):
                _all_pts_cs.append(_pt)
                _pt_trip_cs.append(_t)
        _road_matrix_cs = _osrm_table_batch(_all_pts_cs) if len(_all_pts_cs) >= 2 else {}

        def _min_road_dist_trips(ta, tb):
            """ระยะถนนสั้นสุด (km) ระหว่างสาขาจริงของ ta กับ tb"""
            best = float('inf')
            for i, ti in enumerate(_pt_trip_cs):
                if ti != ta: continue
                for j, tj in enumerate(_pt_trip_cs):
                    if tj != tb: continue
                    d = _road_matrix_cs.get((i, j), float('inf'))
                    if d < best:
                        best = d
            return best if best < float('inf') else haversine_distance(
                _meta_cs[ta]['clat'], _meta_cs[ta]['clon'],
                _meta_cs[tb]['clat'], _meta_cs[tb]['clon'], use_osrm_cache=False)

        # เรียงทริปตาม utilization น้อยสุดก่อน (รถว่างมากสุดก่อน)
        _trips_by_util = sorted(
            _meta_cs.items(),
            key=lambda x: max(x[1]['w'] / x[1]['max_w'], x[1]['c'] / x[1]['max_c']) if x[1]['max_w'] > 0 else 1
        )
        _merged_set_cs: set = set()

        for _ta, _ma in _trips_by_util:
            if _ta in _merged_set_cs:
                continue
            _util_a = max(_ma['w'] / _ma['max_w'], _ma['c'] / _ma['max_c']) if _ma['max_w'] > 0 else 1
            if _util_a >= _CONSOL_MIN_UTIL:
                continue
            if not _ma['prov']:
                continue

            # เรียง candidates ตามระยะถนนจริง → nearest first
            _candidates_cs = []
            for _tb, _mb in _trips_by_util:
                if _tb == _ta or _tb in _merged_set_cs:
                    continue
                if _mb['prov'] != _ma['prov']:
                    continue
                _util_b = max(_mb['w'] / _mb['max_w'], _mb['c'] / _mb['max_c']) if _mb['max_w'] > 0 else 1
                if _util_b >= _CONSOL_MIN_UTIL:
                    continue

                _d_cs = _min_road_dist_trips(_ta, _tb)
                if _d_cs > _CONSOL_MAX_KM:
                    continue

                _comb_w = _ma['w'] + _mb['w']
                _comb_c = _ma['c'] + _mb['c']
                _comb_allow = [v for v in ['4W', 'JB', '6W']
                               if v in _ma['allow'] and v in _mb['allow']]
                if not _comb_allow:
                    continue
                _is_pt_comb = _ma['is_pt'] and _mb['is_pt']
                _lims_comb  = PUNTHAI_LIMITS if _is_pt_comb else LIMITS
                _fit_veh    = next((v for v in ['4W', 'JB', '6W']
                                    if v in _comb_allow and
                                    _comb_w <= _lims_comb[v]['max_w'] and
                                    _comb_c <= _lims_comb[v]['max_c']), None)
                if not _fit_veh:
                    continue

                _u_comb = max(_comb_w / _lims_comb[_fit_veh]['max_w'],
                              _comb_c / _lims_comb[_fit_veh]['max_c'])
                _candidates_cs.append((_d_cs, -_u_comb, _tb))

            if not _candidates_cs:
                continue
            _candidates_cs.sort()  # ใกล้สุดตามถนนจริงก่อน → util สูงก่อน
            _best_tb = _candidates_cs[0][2]
            _best_combined_util = -_candidates_cs[0][1]

            # รวมทริป _best_tb เข้า _ta
            df.loc[df['Trip'] == _best_tb, 'Trip'] = _ta
            # อัป Truck ของ _ta ให้ตรงกับ vehicle ที่จะใช้หลัง merge
            _mb_merged = _meta_cs.get(_best_tb)
            if _mb_merged:
                _comb_w_up = _ma['w'] + _mb_merged['w']
                _comb_c_up = _ma['c'] + _mb_merged['c']
                _is_pt_up  = _ma['is_pt'] and _mb_merged['is_pt']
                _lims_up   = PUNTHAI_LIMITS if _is_pt_up else LIMITS
                _allow_up  = [v for v in ['4W', 'JB', '6W']
                               if v in _ma['allow'] and v in _mb_merged['allow']]
                _truck_up  = next((v for v in ['4W', 'JB', '6W']
                                   if v in _allow_up and
                                   _comb_w_up <= _lims_up[v]['max_w'] and
                                   _comb_c_up <= _lims_up[v]['max_c']), '6W')
                df.loc[df['Trip'] == _ta, 'Truck'] = _truck_up
            _merged_set_cs.add(_best_tb)
            _cs_changed = True
            _consol_merged += 1
            safe_print(f"   🔗 Consolidate: Trip {_best_tb} ({_meta_cs[_best_tb]['prov']}) → Trip {_ta} "
                       f"(util {_best_combined_util*100:.0f}%)")

        if not _cs_changed:
            break

    safe_print(f"   ✅ Step 8.97: รวม {_consol_merged} คู่ทริป (ลดรถว่าง)")

    # ── Split ทริปเกิน capacity (หลัง Step 8.97) ───────────────────────────
    df, _split_next_tid = _app_split_over(df, _split_next_tid)

    # ==========================================
    # Step 9: เรียงทริปใหม่ตามภาค → จังหวัด → ระยะทาง
    # ==========================================
    safe_print("\n📋 Step 9: เรียงทริปแบบ chain ต่อเนื่อง (average distance) → จังหวัด → ตำบล...")

    # คำนวณ avg distance + max distance + dominant province + dominant subdistrict ของแต่ละทริป
    trip_max_distances: dict = {}
    _trip_avg_dist9:    dict = {}   # ← ใช้เรียง chain ต่อเนื่อง
    _trip_dom_prov9:    dict = {}
    _trip_dom_subdist9: dict = {}
    _prov_col9_name    = '_province' if '_province' in df.columns else ('Province' if 'Province' in df.columns else None)
    _subdist_col9_name = next((c for c in ['Subdistrict', '_subdistrict', 'ตำบล'] if c in df.columns), None)

    for trip_num in df[df['Trip'] > 0]['Trip'].unique():
        trip_data = df[df['Trip'] == trip_num]
        if '_distance_from_dc' in trip_data.columns:
            _dists = trip_data['_distance_from_dc'].dropna()
            max_dist = _dists.max() if len(_dists) else 0
            avg_dist = _dists.mean() if len(_dists) else 0
        else:
            max_dist = avg_dist = 0
        trip_max_distances[trip_num] = float(max_dist) if pd.notna(max_dist) else 0.0
        _trip_avg_dist9[trip_num]    = float(avg_dist) if pd.notna(avg_dist) else 0.0

        # dominant province
        _dom_prov9 = ''
        if _prov_col9_name:
            _vc9 = trip_data[_prov_col9_name].dropna().value_counts()
            if len(_vc9):
                _dom_prov9 = str(_vc9.index[0])
        _trip_dom_prov9[trip_num] = _dom_prov9

        # dominant subdistrict
        _dom_sd9 = ''
        if _subdist_col9_name:
            _vcs9 = trip_data[_subdist_col9_name].dropna().value_counts()
            if len(_vcs9):
                _dom_sd9 = str(_vcs9.index[0])
        _trip_dom_subdist9[trip_num] = _dom_sd9

    # dominant district
    _dist_col9_name = next((c for c in ['District', '_district', 'อำเภอ'] if c in df.columns), None)
    _trip_dom_dist9: dict = {}
    for trip_num in df[df['Trip'] > 0]['Trip'].unique():
        _dom_d9 = ''
        if _dist_col9_name:
            _vcd9 = df[df['Trip'] == trip_num][_dist_col9_name].dropna().value_counts()
            if len(_vcd9):
                _dom_d9 = str(_vcd9.index[0])
        _trip_dom_dist9[trip_num] = _dom_d9

    # province → avg dist (centroid) และ → region_order
    _prov_avg_dist9: dict = {}
    _prov_region_ord9: dict = {}
    for _tn, _dp in _trip_dom_prov9.items():
        _d = _trip_avg_dist9.get(_tn, 0)
        if _dp not in _prov_avg_dist9:
            _prov_avg_dist9[_dp] = _d
        else:
            _prov_avg_dist9[_dp] = (_prov_avg_dist9[_dp] + _d) / 2
        if _dp and _dp not in _prov_region_ord9:
            _prov_region_ord9[_dp] = REGION_ORDER.get(get_region_name(str(_dp)), 99)

    # district → avg dist (เรียงอำเภอตามระยะ ไม่ใช่ alphabetical)
    _dist_avg_dist9: dict = {}
    for _tn, _dd in _trip_dom_dist9.items():
        if not _dd:
            continue
        _d = _trip_avg_dist9.get(_tn, 0)
        if _dd not in _dist_avg_dist9:
            _dist_avg_dist9[_dd] = _d
        else:
            _dist_avg_dist9[_dd] = (_dist_avg_dist9[_dd] + _d) / 2

    # sort key: ภาค → จังหวัด (ไกลก่อน) → อำเภอ (ไกลก่อน) → ตำบล (ไกลก่อน)
    # ทุกระดับเรียงตาม avg distance (ไม่ใช่ชื่อ) → ทริปต่อเนื่องกัน ไม่กระโดด
    trip_sort9_keys = {
        trip_num: (
            _prov_region_ord9.get(_trip_dom_prov9.get(trip_num, ''), 99),    # ภาค
            -_prov_avg_dist9.get(_trip_dom_prov9.get(trip_num, ''), 0),      # จังหวัด ไกลก่อน
            _trip_dom_prov9.get(trip_num, ''),                                # กลุ่มจังหวัดเดียวกัน
            -_dist_avg_dist9.get(_trip_dom_dist9.get(trip_num, ''), 0),      # อำเภอ ไกลก่อน (ระยะ)
            -_trip_avg_dist9[trip_num],                                       # ภายในอำเภอ: ไกลก่อน
            _trip_dom_subdist9.get(trip_num, ''),                             # ตำบล
        )
        for trip_num in trip_max_distances
    }

    sorted_trips = sorted(trip_max_distances.keys(), key=lambda x: trip_sort9_keys.get(x, (99, 0, '', '', 0, '')))
    
    # สร้าง mapping ใหม่
    trip_renumber = {old_trip: new_trip for new_trip, old_trip in enumerate(sorted_trips, 1)}
    df['Trip'] = df['Trip'].map(lambda x: trip_renumber.get(x, 0) if x > 0 else 0)
    
    # อัพเดต summary_df ใหม่ทั้งหมดหลัง renumber (ให้ข้อมูลตรงกับ df)
    summary_data_new = []
    for trip_num in sorted(df[df['Trip'] > 0]['Trip'].unique()):
        trip_data = df[df['Trip'] == trip_num]
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        trip_codes_list = trip_data['Code'].tolist()
        max_dist = trip_data['_distance_from_dc'].max() if '_distance_from_dc' in trip_data.columns else 0
        
        # หารถจากคอลัมน์ Truck ใน df
        truck = trip_data['Truck'].iloc[0] if 'Truck' in trip_data.columns and len(trip_data) > 0 else '6W'
        truck_str = str(truck).split()[0] if pd.notna(truck) else '6W'
        
        # หา BU type
        is_punthai = all(str(r.get('BU', '')).upper() in ['211', 'PUNTHAI'] for _, r in trip_data.iterrows())
        limits = PUNTHAI_LIMITS if is_punthai else LIMITS
        
        max_w = limits.get(truck_str, limits['6W'])['max_w']
        max_c = limits.get(truck_str, limits['6W'])['max_c']
        
        summary_data_new.append({
            'Trip': trip_num,
            'Branches': len(trip_codes_list),
            'Weight': total_w,
            'Cube': total_c,
            'Truck': truck,
            'BU_Type': 'punthai' if is_punthai else 'maxmart',
            'Buffer': f"🅿️ {int(punthai_buffer*100)}%" if is_punthai else f"🅼 {int(maxmart_buffer*100)}%",
            'Weight_Use%': (total_w / max_w) * 100,
            'Cube_Use%': (total_c / max_c) * 100,
            'Total_Distance': max_dist if pd.notna(max_dist) else 0
        })
    
    summary_df = pd.DataFrame(summary_data_new)
    summary_df = summary_df.sort_values('Trip').reset_index(drop=True)
    
    if sorted_trips:
        safe_print(f"   ✅ เรียงใหม่: {len(sorted_trips)} ทริป (Trip 1 = ไกลสุด {trip_max_distances[sorted_trips[0]]:.0f} km)")
    else:
        safe_print("   ✅ เรียงใหม่: 0 ทริป")
    
    # 📋 เรียงลำดับสาขาภายในทริป: จังหวัด (ไกลก่อน, ตามระยะทาง) → อำเภอ (จัดกลุ่ม) → ระยะทาง DC (ไกลก่อน)
    if '_distance_from_dc' in df.columns:
        _dist_col_s9 = '_distance_from_dc'
    elif 'Distance_from_DC' in df.columns:
        _dist_col_s9 = 'Distance_from_DC'
    else:
        _dist_col_s9 = None

    _subdist_col_s9  = next((c for c in ['Subdistrict', '_subdistrict', 'ตำบล']  if c in df.columns), None)
    _dist_grp_col_s9 = next((c for c in ['District',    '_district',   'อำเภอ']  if c in df.columns), None)
    if _dist_col_s9 and '_province' in df.columns:
        # groupby (Trip, province) → max dist สำหรับจัดกลุ่มจังหวัด
        df['_prov_maxd_s9'] = df.groupby(['Trip', '_province'])[_dist_col_s9].transform('max').fillna(0)
        # groupby (Trip, province, district) → max dist สำหรับจัดกลุ่มอำเภอภายในจังหวัด
        if _dist_grp_col_s9:
            df['_dist_maxd_s9'] = df.groupby(['Trip', '_province', _dist_grp_col_s9])[_dist_col_s9].transform('max').fillna(0)
        if _dist_grp_col_s9 and _subdist_col_s9:
            df['_subd_maxd_s9'] = df.groupby(['Trip', '_province', _dist_grp_col_s9, _subdist_col_s9])[_dist_col_s9].transform('max').fillna(0)
        # เรียง: Trip → จังหวัด(ไกล) → จังหวัด → อำเภอ(ไกล) → อำเภอ → ตำบล(ไกล) → ตำบล → ระยะทาง(ไกล)
        _sort_cols_s9 = ['Trip', '_prov_maxd_s9', '_province']
        _sort_asc_s9  = [True, False, True]
        if _dist_grp_col_s9:
            _sort_cols_s9 += ['_dist_maxd_s9', _dist_grp_col_s9]
            _sort_asc_s9  += [False, True]
        if _subdist_col_s9:
            if _dist_grp_col_s9:
                _sort_cols_s9 += ['_subd_maxd_s9', _subdist_col_s9]
                _sort_asc_s9  += [False, True]
            else:
                _sort_cols_s9.append(_subdist_col_s9); _sort_asc_s9.append(True)
        _sort_cols_s9.append(_dist_col_s9); _sort_asc_s9.append(False)
        df = df.sort_values(_sort_cols_s9, ascending=_sort_asc_s9).reset_index(drop=True)
        df = df.drop(columns=['_prov_maxd_s9', '_dist_maxd_s9', '_subd_maxd_s9'], errors='ignore')
    elif _dist_col_s9:
        df = df.sort_values(['Trip', _dist_col_s9], ascending=[True, False]).reset_index(drop=True)
    else:
        df = df.sort_values(['Trip'], ascending=[True]).reset_index(drop=True)
    
    # ลบคอลัมน์ชั่วคราว
    cols_to_drop = ['_region_name', '_route', '_group_key', '_region_order', '_prov_max_dist', '_dist_max_dist', '_subdist_max_dist', '_region_allowed_vehicles', '_vehicle_priority']
    df = df.drop(columns=[c for c in cols_to_drop if c in df.columns], errors='ignore')

    # ═══════════════════════════════════════════════════════════════════
    # 🔒 ABSOLUTE FINAL GROUP LOCK — ใช้ BRANCH_GROUPS global (โหลดตอน startup แล้ว)
    _fl_groups_fresh = BRANCH_GROUPS
    safe_print(f"🔒 FINAL LOCK: {len(_fl_groups_fresh)} กลุ่ม")

    # สร้าง code_upper → (index, trip) จาก df ปัจจุบัน
    _fl_code_idx: dict = {}
    _fl_code_trip: dict = {}
    for _fl_i, _fl_r in df[['Code', 'Trip']].iterrows():
        _fl_cu = str(_fl_r['Code']).strip().upper()
        _fl_code_idx[_fl_cu]  = _fl_i
        _fl_code_trip[_fl_cu] = safe_int_trip(_fl_r['Trip'])

    from collections import Counter as _FLC

    # วนซ้ำจนกว่า stable (ไม่มีการเปลี่ยนแปลง) — ป้องกัน chain dependency
    _total_fixed = 0
    for _pass in range(10):   # max 10 rounds กันวนไม่จบ
        _pass_fixed = 0
        for _fgid, _fgmems in _fl_groups_fresh.items():
            _mups = [str(c).strip().upper() for c in _fgmems
                     if str(c).strip().upper() in _fl_code_trip]
            if len(_mups) < 2:
                continue
            _trips_now = [_fl_code_trip[m] for m in _mups]
            # รวมทุก member — ทั้ง Trip>0 และ Trip=0 (orphan ก็ต้องดึงเข้ากลุ่ม)
            _assigned   = [(m, t) for m, t in zip(_mups, _trips_now) if t > 0]
            _orphans_fl = [(m, t) for m, t in zip(_mups, _trips_now) if t == 0]
            if len(_assigned) < 1:
                continue
            _trips_assigned = [t for _, t in _assigned]
            if len(set(_trips_assigned)) == 1 and not _orphans_fl:
                continue  # ✓ ทุก member อยู่ทริปเดียวกันแล้ว
            # เลือกทริปที่มี member มากที่สุด
            _target_t = _FLC(_trips_assigned).most_common(1)[0][0]
            # วน: member ที่อยู่ต่างทริป + orphan (Trip=0) ทั้งหมด
            for _mu, _mt in _assigned + _orphans_fl:
                if _mt == _target_t:
                    continue
                _fl_i = _fl_code_idx.get(_mu)
                # FINAL GROUP LOCK = ไม่มีข้อยกเว้น ยกเว้นข้ามจังหวัดเท่านั้น
                _raw_prov_fl = df.at[_fl_i, '_province'] if _fl_i is not None else None
                _mu_prov_fl = '' if (not _raw_prov_fl or str(_raw_prov_fl).strip().lower() in ('nan', 'none', '')) else str(_raw_prov_fl).strip()
                _tgt_rows = df[df['Trip'] == _target_t]
                _tgt_provs = set(_tgt_rows['_province'].dropna().astype(str).str.strip().unique()) if not _tgt_rows.empty else set()
                _tgt_provs = {p for p in _tgt_provs if p and p.lower() not in ('nan', 'none', '')}
                # Province-strict เท่านั้น (zone/capacity ไม่บล็อค — group lock เด็ดขาด)
                if _mu_prov_fl and _tgt_provs and _mu_prov_fl not in _tgt_provs:
                    safe_print(f"   ⚠️ FINAL LOCK skip: {_mu} จังหวัด {_mu_prov_fl} ≠ target trip {_target_t} ({_tgt_provs})")
                    continue
                # zone/capacity: ยอมเกิน (group lock สำคัญกว่า) — Step 7.5 จะ split ทีหลังถ้าจำเป็น
                if _fl_i is not None:
                    df.at[_fl_i, 'Trip'] = _target_t
                else:
                    df.loc[df['Code'].str.strip().str.upper() == _mu, 'Trip'] = _target_t
                _fl_code_trip[_mu] = _target_t
                _pass_fixed += 1
                safe_print(f"   🔒 FINAL LOCK pass{_pass+1}: {_mu} Trip {_mt}→{_target_t} (group {_fgid})")
        _total_fixed += _pass_fixed
        if _pass_fixed == 0:
            break  # stable แล้ว

    if _total_fixed:
        safe_print(f"🔒 FINAL GROUP LOCK: แก้ {_total_fixed} สาขาที่แยกกลุ่ม ({_pass+1} passes)")
    else:
        safe_print("🔒 FINAL GROUP LOCK: ✅ ทุก group อยู่ทริปเดียวกัน")

    # ── FINAL RE-NUMBER: sort key เดียวกับ Step 9 + Export ──
    # (-province_max_dist, province_name, subdistrict, -trip_max_dist)
    _fn_prov_col   = '_province'      if '_province'      in df.columns else None
    _fn_dist_col   = '_distance_from_dc' if '_distance_from_dc' in df.columns else None
    _fn_subdist_col = next((c for c in ['Subdistrict', '_subdistrict', 'ตำบล'] if c in df.columns), None)

    # pass 1: trip → avg dist + max dist + dominant province + dominant subdistrict
    _fn_trip_pmx:    dict = {}
    _fn_trip_avg:    dict = {}
    _fn_trip_prov:   dict = {}
    _fn_trip_dist:   dict = {}
    _fn_trip_subdist: dict = {}
    _fn_dist_col2 = next((c for c in ['_district', 'District', 'อำเภอ'] if c in df.columns), None)
    for _tn in df[df['Trip'] > 0]['Trip'].unique():
        _td = df[df['Trip'] == _tn]
        if _fn_dist_col:
            _dv = _td[_fn_dist_col].dropna()
            _mx_v = _dv.max() if len(_dv) else 0
            _av_v = _dv.mean() if len(_dv) else 0
        else:
            _mx_v = _av_v = 0
        _fn_trip_pmx[_tn] = float(_mx_v) if pd.notna(_mx_v) else 0.0
        _fn_trip_avg[_tn] = float(_av_v) if pd.notna(_av_v) else 0.0
        if _fn_prov_col:
            _vc = _td[_fn_prov_col].dropna().value_counts()
            _fn_trip_prov[_tn] = str(_vc.index[0]) if len(_vc) else ''
        else:
            _fn_trip_prov[_tn] = ''
        if _fn_dist_col2:
            _vcd = _td[_fn_dist_col2].dropna().value_counts()
            _fn_trip_dist[_tn] = str(_vcd.index[0]) if len(_vcd) else ''
        else:
            _fn_trip_dist[_tn] = ''
        if _fn_subdist_col:
            _vcs = _td[_fn_subdist_col].dropna().value_counts()
            _fn_trip_subdist[_tn] = str(_vcs.index[0]) if len(_vcs) else ''
        else:
            _fn_trip_subdist[_tn] = ''

    # สร้าง global max dist ในระดับ จังหวัด / อำเภอ / ตำบล
    # เพื่อให้ทริปเรียงต่อเนื่องโดยไม่กระโดดข้ามระดับ
    _fn_prov_gmax:   dict = {}   # province → max dist
    _fn_prov_rord:   dict = {}   # province → region order
    _fn_dist_gmax:   dict = {}   # (prov, dist) → max dist
    _fn_subdist_gmax: dict = {}  # (prov, dist, subdist) → max dist
    for _tn, _dp in _fn_trip_prov.items():
        _d   = _fn_trip_pmx.get(_tn, 0)
        _dd  = _fn_trip_dist.get(_tn, '')
        _ds  = _fn_trip_subdist.get(_tn, '')
        if _dp not in _fn_prov_gmax or _d > _fn_prov_gmax[_dp]:
            _fn_prov_gmax[_dp] = _d
        if _dp and _dp not in _fn_prov_rord:
            _fn_prov_rord[_dp] = REGION_ORDER.get(get_region_name(str(_dp)), 99)
        _dkey  = (_dp, _dd)
        _sdkey = (_dp, _dd, _ds)
        if _dkey not in _fn_dist_gmax or _d > _fn_dist_gmax[_dkey]:
            _fn_dist_gmax[_dkey] = _d
        if _sdkey not in _fn_subdist_gmax or _d > _fn_subdist_gmax[_sdkey]:
            _fn_subdist_gmax[_sdkey] = _d

    # sort key: ภาค → จังหวัด(ไกลก่อน) → จังหวัด
    #           → อำเภอ(ไกลก่อน) → อำเภอ
    #           → ตำบล(ไกลก่อน) → ตำบล
    #           → avg_dist (เรียงทริปภายในตำบลเดียวกัน)
    _fn_sort_keys = {
        _tn: (
            _fn_prov_rord.get(_fn_trip_prov.get(_tn, ''), 99),
            -_fn_prov_gmax.get(_fn_trip_prov.get(_tn, ''), 0),
            _fn_trip_prov.get(_tn, ''),
            -_fn_dist_gmax.get((_fn_trip_prov.get(_tn,''), _fn_trip_dist.get(_tn,'')), 0),
            _fn_trip_dist.get(_tn, ''),
            -_fn_subdist_gmax.get((_fn_trip_prov.get(_tn,''), _fn_trip_dist.get(_tn,''), _fn_trip_subdist.get(_tn,'')), 0),
            _fn_trip_subdist.get(_tn, ''),
            -_fn_trip_avg.get(_tn, 0),
        )
        for _tn in _fn_trip_pmx
    }
    _final_sorted = sorted(_fn_sort_keys.items(), key=lambda x: x[1])
    _final_map = {old: new for new, (old, _) in enumerate(_final_sorted, 1)}
    df['Trip'] = df['Trip'].map(lambda x: _final_map.get(x, x) if x > 0 else x)
    # sync summary_df ให้ตรงกับ FINAL RE-NUMBER (ป้องกัน vehicle lookup ผิดทริป)
    summary_df['Trip'] = summary_df['Trip'].map(
        lambda x: _final_map.get(int(x), int(x)) if pd.notna(x) and int(x) > 0 else 0
    )
    summary_df = summary_df.sort_values('Trip').reset_index(drop=True)
    safe_print(f"🔢 FINAL RE-NUMBER: เรียงเลขทริปใหม่ {len(_final_sorted)} ทริป (ภาค→จังหวัด→avg dist)")

    # ── Blanket NaN guard: แปลง NaN ใน Trip column เป็น 0 ก่อน return ──
    df['Trip'] = df['Trip'].fillna(0).astype(int)

    return df, summary_df, fleet_used
def main():
    st.set_page_config(
        page_title="ระบบจัดเที่ยว",
        page_icon="🚚",
        layout="wide",
        initial_sidebar_state="collapsed"
    )

    # ── Global white/green theme CSS ──────────────────────────────────────
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sarabun:wght@300;400;500;600;700;800&family=Inter:wght@300;400;500;600;700;800;900&display=swap');

/* ═══════════════════════════════════════════════════════════════
   RESET & BASE — FORCE LIGHT MODE
═══════════════════════════════════════════════════════════════ */
*, *::before, *::after { box-sizing: border-box; }
html, body {
    font-family: 'Sarabun', 'Inter', system-ui, sans-serif !important;
    background: #f0faf4 !important;
    color: #0f172a !important;
    font-size: 16px !important;
}
[class*="css"] {
    font-family: 'Sarabun', 'Inter', system-ui, sans-serif !important;
}
/* Nuke every possible dark container */
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stAppViewBlockContainer"],
[data-testid="stMain"],
[data-testid="stMainBlockContainer"],
[data-testid="stVerticalBlock"],
[data-testid="stHorizontalBlock"],
[data-testid="stHeader"],
[data-testid="stToolbar"],
[data-testid="stDecoration"],
[data-testid="stBottomBlockContainer"],
section[tabindex="0"],
.main, section.main, .block-container,
div[data-layout="wide"] {
    background-color: #f0faf4 !important;
    color: #0f172a !important;
}
.stApp {
    background: #f0faf4 !important;
    background-image:
        radial-gradient(ellipse 70% 40% at 50% -5%, rgba(16,185,129,0.10), transparent),
        radial-gradient(ellipse 50% 35% at 90% 70%, rgba(6,182,212,0.05), transparent) !important;
}
.main .block-container, [data-testid="stMainBlockContainer"] {
    padding-top: 0 !important;
    padding-bottom: 4rem !important;
    max-width: 1480px !important;
    background-color: transparent !important;
}

/* ═══════════════════════════════════════════════════════════════
   TOP NAV BAR
═══════════════════════════════════════════════════════════════ */
.app-navbar {
    background: #ffffff;
    border-bottom: 3px solid #10b981;
    padding: 0 2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    height: 64px;
    margin: -1rem -1rem 0 -1rem;
    position: sticky;
    top: 0;
    z-index: 999;
    box-shadow: 0 2px 16px rgba(16,185,129,0.10);
}
.app-navbar-brand {
    display: flex; align-items: center; gap: 14px;
    color: #064e3b !important;
    font-size: 1.15rem;
    font-weight: 800;
    letter-spacing: -0.3px;
    text-decoration: none;
}
.app-navbar-brand .brand-icon {
    width: 40px; height: 40px;
    background: linear-gradient(135deg, #059669 0%, #10b981 100%);
    border-radius: 12px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.3rem;
    box-shadow: 0 4px 12px rgba(16,185,129,0.3);
    flex-shrink: 0;
}
.app-navbar-brand .brand-name { line-height: 1.1; }
.app-navbar-brand .brand-name small { display: block; font-size: 0.65rem; font-weight: 500; color: #6b7280; letter-spacing: 0.5px; }
.app-navbar-status {
    display: flex; align-items: center; gap: 10px;
    font-size: 0.8rem; color: #047857;
    font-weight: 500;
}
.nav-badge {
    display: inline-flex; align-items: center; gap: 6px;
    background: #f0fdf4;
    border: 1px solid #bbf7d0;
    padding: 4px 12px; border-radius: 20px;
    font-size: 0.76rem; font-weight: 600;
    color: #065f46; white-space: nowrap;
}
.nav-sync-btn {
    display: inline-flex; align-items: center; gap: 6px;
    background: #059669; color: #ffffff !important;
    border: none; border-radius: 20px;
    padding: 6px 16px; font-size: 0.78rem; font-weight: 700;
    cursor: pointer; transition: background .2s, box-shadow .2s;
    box-shadow: 0 2px 8px rgba(5,150,105,0.3);
    text-decoration: none;
}
.nav-sync-btn:hover { background: #047857; box-shadow: 0 4px 12px rgba(5,150,105,0.4); }
/* push stButton inside navbar area */
.navbar-sync-wrapper { position: relative; margin: -64px 0 0 0; height: 64px; display: flex; align-items: center; justify-content: flex-end; padding-right: 2rem; pointer-events: none; }
.navbar-sync-wrapper > * { pointer-events: all; }
.nav-dot { width: 8px; height: 8px; border-radius: 50%; background: #4ade80; display: inline-block; box-shadow: 0 0 6px rgba(74,222,128,0.9); animation: pulse-dot 2s infinite; flex-shrink: 0; }
.nav-dot.offline { background: #fbbf24; box-shadow: 0 0 6px rgba(251,191,36,0.8); animation: none; }
.nav-badge-warn { background: #fffbeb !important; border-color: #fde68a !important; color: #92400e !important; }
@keyframes pulse-dot { 0%, 100% { opacity: 1; } 50% { opacity: 0.4; } }
/* ── Page title ── */
.page-section-title {
    font-size: 1.4rem; font-weight: 800; color: #064e3b;
    margin: 1.2rem 0 0.8rem;
    display: flex; align-items: center; gap: 10px;
}
.page-section-title small { font-size: 0.82rem; font-weight: 500; color: #6b7280; }
/* ── KPI cards ── */
.kpi-card {
    background: #ffffff;
    border: 1.5px solid #d1fae5;
    border-radius: 16px;
    padding: 1rem 1.25rem;
    text-align: center;
    box-shadow: 0 2px 8px rgba(5,150,105,0.06);
    transition: border-color .2s, box-shadow .2s;
}
.kpi-card:hover { border-color: #6ee7b7; box-shadow: 0 4px 16px rgba(16,185,129,0.12); }
.kpi-value { font-size: 1.55rem; font-weight: 800; color: #059669; line-height: 1.1; }
.kpi-label { font-size: 0.72rem; color: #6b7280; font-weight: 500; margin-top: 4px; }

/* ═══════════════════════════════════════════════════════════════
   SECTION CARD & DIVIDER
═══════════════════════════════════════════════════════════════ */
.glass-card {
    background: #ffffff;
    border: 1.5px solid #d1fae5;
    border-radius: 20px;
    padding: 1.75rem;
    margin-bottom: 1.25rem;
    box-shadow: 0 2px 8px rgba(5,150,105,0.06), 0 1px 3px rgba(15,23,42,0.05);
    transition: border-color .2s, box-shadow .2s;
}
.glass-card:hover {
    border-color: #6ee7b7;
    box-shadow: 0 6px 24px rgba(16,185,129,0.12);
}
.divider-label {
    display: flex; align-items: center; gap: 12px;
    color: #059669; font-size: 0.78rem; font-weight: 800;
    text-transform: uppercase; letter-spacing: 0.1em;
    margin: 2rem 0 1.25rem 0;
}
.divider-label::before, .divider-label::after {
    content: ''; flex: 1; height: 1.5px;
    background: linear-gradient(90deg, #d1fae5, transparent);
}
.divider-label::before { background: linear-gradient(90deg, transparent, #d1fae5); }

/* ═══════════════════════════════════════════════════════════════
   TYPOGRAPHY
═══════════════════════════════════════════════════════════════ */
h1 { color: #064e3b !important; font-weight: 900 !important; font-size: 2rem !important; }
h2 { color: #065f46 !important; font-weight: 800 !important; font-size: 1.5rem !important; }
h3 { color: #047857 !important; font-weight: 700 !important; font-size: 1.2rem !important; }
h4 { color: #0f172a !important; font-weight: 700 !important; font-size: 1.05rem !important; }
h5, h6 { color: #1e293b !important; font-weight: 600 !important; }
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3,
[data-testid="stMarkdownContainer"] h4 {
    color: #064e3b !important;
}
[data-testid="stMarkdownContainer"] p {
    color: #1e293b !important;
    font-size: 0.95rem !important;
    line-height: 1.65 !important;
}
[data-testid="stMarkdownContainer"] li {
    color: #374151 !important;
    font-size: 0.95rem !important;
}
/* All text elements */
label, span, div, p, td, th, li {
    color: #0f172a;
}

/* ═══════════════════════════════════════════════════════════════
   METRIC CARDS
═══════════════════════════════════════════════════════════════ */
[data-testid="metric-container"] {
    background: #ffffff !important;
    border: 1.5px solid #d1fae5 !important;
    border-radius: 18px !important;
    padding: 20px 24px !important;
    box-shadow: 0 2px 8px rgba(5,150,105,0.07) !important;
    transition: all .22s cubic-bezier(.4,0,.2,1) !important;
    position: relative !important;
    overflow: hidden !important;
}
[data-testid="metric-container"]::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 4px;
    background: linear-gradient(90deg, #059669, #10b981, #34d399);
    border-radius: 18px 18px 0 0;
}
[data-testid="metric-container"]:hover {
    border-color: #6ee7b7 !important;
    box-shadow: 0 8px 28px rgba(16,185,129,0.15) !important;
    transform: translateY(-3px) !important;
}
[data-testid="stMetricLabel"] {
    color: #059669 !important;
    font-size: 0.78rem !important;
    font-weight: 800 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
}
[data-testid="stMetricValue"] {
    color: #064e3b !important;
    font-weight: 900 !important;
    font-size: 2rem !important;
    letter-spacing: -1px !important;
}
[data-testid="stMetricDelta"] { font-weight: 700 !important; font-size: 0.85rem !important; }

/* ═══════════════════════════════════════════════════════════════
   BUTTONS
═══════════════════════════════════════════════════════════════ */
.stButton > button {
    background: #ffffff !important;
    color: #374151 !important;
    border: 1.5px solid #d1d5db !important;
    border-radius: 12px !important;
    font-weight: 700 !important;
    font-family: inherit !important;
    font-size: 0.95rem !important;
    padding: 0.6rem 1.3rem !important;
    transition: all .18s cubic-bezier(.4,0,.2,1) !important;
    box-shadow: 0 1px 3px rgba(15,23,42,0.08) !important;
}
.stButton > button:hover {
    background: #f0fdf4 !important;
    border-color: #10b981 !important;
    color: #065f46 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 14px rgba(16,185,129,0.15) !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
    color: #ffffff !important;
    border: none !important;
    font-weight: 800 !important;
    font-size: 1rem !important;
    padding: 0.7rem 1.75rem !important;
    box-shadow: 0 4px 20px rgba(5,150,105,0.35) !important;
}
.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #047857 0%, #059669 100%) !important;
    box-shadow: 0 8px 28px rgba(5,150,105,0.45) !important;
    transform: translateY(-2px) !important;
}

/* ═══════════════════════════════════════════════════════════════
   TABS
═══════════════════════════════════════════════════════════════ */
.stTabs [data-baseweb="tab-list"] {
    gap: 3px;
    background: #ecfdf5;
    border-radius: 14px;
    padding: 5px;
    border: 1.5px solid #d1fae5;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px !important;
    font-weight: 700 !important;
    color: #6b7280 !important;
    padding: 10px 26px !important;
    font-size: 0.92rem !important;
    transition: all .18s !important;
    border: none !important;
    background: transparent !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #059669, #10b981) !important;
    color: #ffffff !important;
    box-shadow: 0 2px 12px rgba(5,150,105,0.35) !important;
    font-weight: 800 !important;
}
[data-baseweb="tab-highlight"] { display: none !important; }
[data-baseweb="tab-border"] { display: none !important; }

/* ═══════════════════════════════════════════════════════════════
   EXPANDERS
═══════════════════════════════════════════════════════════════ */
details[data-testid="stExpander"] {
    background: #ffffff !important;
    border: 1.5px solid #e2e8f0 !important;
    border-radius: 14px !important;
    margin-bottom: 10px;
    box-shadow: 0 1px 4px rgba(15,23,42,0.05);
    transition: all .2s;
}
details[data-testid="stExpander"][open] {
    border-color: #6ee7b7 !important;
    box-shadow: 0 3px 14px rgba(16,185,129,0.10) !important;
}
details[data-testid="stExpander"] summary {
    font-weight: 700 !important;
    color: #1e293b !important;
    padding: 15px 20px !important;
    border-radius: 14px !important;
    font-size: 0.95rem !important;
    cursor: pointer;
}
details[data-testid="stExpander"] summary:hover { color: #059669 !important; }

/* ═══════════════════════════════════════════════════════════════
   HERO UPLOAD CARD
═══════════════════════════════════════════════════════════════ */
.hero-upload-card {
    background: linear-gradient(135deg, #ecfdf5 0%, #f0fdf4 60%, #ffffff 100%);
    border: 2px solid #a7f3d0;
    border-radius: 24px;
    padding: 2rem 2.5rem 1.5rem;
    margin: 1rem 0 1.5rem;
    box-shadow: 0 4px 24px rgba(16,185,129,0.08);
}
.hero-upload-title {
    font-size: 1.3rem; font-weight: 800; color: #064e3b;
    margin-bottom: 0.25rem;
}
.hero-upload-sub {
    font-size: 0.88rem; color: #6b7280; margin-bottom: 1.2rem;
}
/* FILE UPLOADER */
[data-testid="stFileUploader"] {
    background: #ffffff !important;
    border: 2px dashed #6ee7b7 !important;
    border-radius: 16px !important;
    padding: 1.2rem !important;
    transition: all .25s;
}
[data-testid="stFileUploader"]:hover {
    border-color: #10b981 !important;
    background: #f0fdf4 !important;
    box-shadow: 0 0 0 4px rgba(16,185,129,0.1) !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] { color: #6b7280 !important; }
[data-testid="stFileUploaderDropzoneInstructions"] small { color: #9ca3af !important; }

/* ═══════════════════════════════════════════════════════════════
   INPUTS
═══════════════════════════════════════════════════════════════ */
[data-testid="stNumberInput"] > div,
[data-testid="stTextInput"] > div > div {
    background: #ffffff !important;
    border: 1.5px solid #d1d5db !important;
    border-radius: 12px !important;
}
[data-testid="stNumberInput"] input,
[data-testid="stTextInput"] input {
    background: transparent !important;
    border: none !important;
    color: #0f172a !important;
    font-weight: 700 !important;
    font-size: 1.05rem !important;
    caret-color: #10b981;
}
[data-testid="stNumberInput"] > div:focus-within,
[data-testid="stTextInput"] > div > div:focus-within {
    border-color: #10b981 !important;
    box-shadow: 0 0 0 3px rgba(16,185,129,0.12) !important;
    background: #f0fdf4 !important;
}
[data-testid="stSelectbox"] > div > div {
    background: #ffffff !important;
    border: 1.5px solid #d1d5db !important;
    border-radius: 12px !important;
    color: #0f172a !important;
}
/* Input labels */
[data-testid="stNumberInput"] label,
[data-testid="stTextInput"] label,
[data-testid="stSelectbox"] label,
.stSlider label {
    color: #059669 !important;
    font-size: 0.82rem !important;
    font-weight: 800 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
}

/* ═══════════════════════════════════════════════════════════════
   DATAFRAME
═══════════════════════════════════════════════════════════════ */
[data-testid="stDataFrame"] {
    border-radius: 14px !important;
    overflow: hidden !important;
    border: 1.5px solid #d1fae5 !important;
    box-shadow: 0 2px 8px rgba(5,150,105,0.06) !important;
}
[data-testid="stDataFrame"] thead th {
    background: #ecfdf5 !important;
    color: #065f46 !important;
    font-weight: 800 !important;
    font-size: 0.78rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
    border-bottom: 1.5px solid #a7f3d0 !important;
}
[data-testid="stDataFrame"] tbody tr:hover td { background: #f0fdf4 !important; }

/* ═══════════════════════════════════════════════════════════════
   ALERTS — LIGHT MODE COLORS
═══════════════════════════════════════════════════════════════ */
[data-testid="stAlert"] {
    border-radius: 14px !important;
    font-weight: 600 !important;
    font-size: 0.92rem !important;
}
[data-testid="stAlert"][data-type="success"],
div.stAlert[data-baseweb="notification"].st-success {
    background: #f0fdf4 !important;
    border: 1.5px solid #a7f3d0 !important;
    border-left: 4px solid #059669 !important;
    color: #065f46 !important;
}
[data-testid="stAlert"][data-type="warning"] {
    background: #fffbeb !important;
    border: 1.5px solid #fde68a !important;
    border-left: 4px solid #d97706 !important;
    color: #92400e !important;
}
[data-testid="stAlert"][data-type="error"] {
    background: #fef2f2 !important;
    border: 1.5px solid #fecaca !important;
    border-left: 4px solid #ef4444 !important;
    color: #991b1b !important;
}
[data-testid="stAlert"][data-type="info"] {
    background: #eff6ff !important;
    border: 1.5px solid #bfdbfe !important;
    border-left: 4px solid #3b82f6 !important;
    color: #1e40af !important;
}

/* ═══════════════════════════════════════════════════════════════
   DOWNLOAD BUTTON
═══════════════════════════════════════════════════════════════ */
[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #059669, #10b981) !important;
    color: #ffffff !important;
    border: none !important;
    font-weight: 800 !important;
    font-size: 0.95rem !important;
    border-radius: 12px !important;
    box-shadow: 0 3px 10px rgba(5,150,105,0.28) !important;
}
[data-testid="stDownloadButton"] > button:hover {
    filter: brightness(1.07) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 7px 22px rgba(5,150,105,0.38) !important;
}

/* ═══════════════════════════════════════════════════════════════
   PROGRESS & SPINNER
═══════════════════════════════════════════════════════════════ */
[data-testid="stProgress"] > div {
    background: #d1fae5 !important;
    border-radius: 99px !important;
    overflow: hidden;
}
[data-testid="stProgress"] > div > div {
    background: linear-gradient(90deg, #059669, #10b981, #34d399) !important;
    border-radius: 99px !important;
    box-shadow: 0 0 8px rgba(16,185,129,0.4);
}
[data-testid="stSpinner"] > div { border-top-color: #10b981 !important; }

/* ═══════════════════════════════════════════════════════════════
   SCROLLBAR
═══════════════════════════════════════════════════════════════ */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f0fdf4; border-radius: 3px; }
::-webkit-scrollbar-thumb { background: #a7f3d0; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #6ee7b7; }

/* ═══════════════════════════════════════════════════════════════
   MISC
═══════════════════════════════════════════════════════════════ */
[data-testid="stCheckbox"] label { font-weight: 700 !important; color: #1e293b !important; font-size: 0.95rem !important; }
.stCaption { color: #6b7280 !important; font-size: 0.78rem !important; }
hr { border: none !important; border-top: 1.5px solid #d1fae5 !important; margin: 1.5rem 0 !important; }
[data-testid="stRadio"] label { font-weight: 600 !important; color: #1e293b !important; font-size: 0.95rem !important; }

/* ═══════════════════════════════════════════════════════════════
   HIDE STREAMLIT BRANDING
═══════════════════════════════════════════════════════════════ */
#MainMenu, footer, header { visibility: hidden; }
[data-testid="stAppViewContainer"] > section > div:first-child { padding-top: 0 !important; }
.stDeployButton { display: none; }
</style>
""", unsafe_allow_html=True)


    # ป้องกันการโหลดโซนซ้ำ - ใช้ session state
    if 'zones_loaded' not in st.session_state:
        st.session_state.zones_loaded = False
    if 'cache_stats_shown' not in st.session_state:
        st.session_state.cache_stats_shown = False
    
    # 🔄 Auto-refresh (Optional - ไม่กระทบการใช้งานหลักถ้าไม่มี)
    # ใช้สำหรับ refresh cache ทุกเที่ยงคืน (เฉพาะ local dev)
    if AUTOREFRESH_AVAILABLE:
        try:
            now = datetime.now()
            # คำนวณเวลาถึงเที่ยงคืน (00:00:00)
            midnight = datetime.combine(now.date(), datetime_time(0, 0, 0))
            
            # ถ้ายังไม่ถึงเที่ยงคืน เอาเที่ยงคืนวันถัดไป
            if now < midnight:
                next_midnight = midnight
            else:
                next_midnight = midnight + timedelta(days=1)
            
            # คำนวณเวลาที่เหลือ (วินาที)
            seconds_until_midnight = int((next_midnight - now).total_seconds())
            
            # Refresh ทุกเที่ยงคืน (เฉพาะถ้ามี autorefresh)
            if seconds_until_midnight > 0:
                # เช็คในช่วง 5 นาทีก่อนเที่ยงคืน (หลัง 23:55)
                if seconds_until_midnight <= 300:  # 5 minutes
                    st.info(f"🔄 ระบบจะ Refresh อัตโนมัติใน {seconds_until_midnight // 60} นาที")
                    st_autorefresh(interval=seconds_until_midnight * 1000, key="midnight_refresh")
                else:
                    # ตรวจสอบทุก 1 ชั่วโมง
                    st_autorefresh(interval=3600000, limit=24, key="hourly_check")
        except Exception as e:
            # ถ้า autorefresh มีปัญหา → ไม่แสดง error (ฟีเจอร์เสริมเท่านั้น)
            pass
    
    # ── Top Navigation Bar ──────────────────────────────────────────────────
    # check precompute status
    _precompute_running = False
    if st.session_state.get('_precompute_pid'):
        import subprocess as _sp2
        _pid = st.session_state['_precompute_pid']
        try:
            if os.name == 'nt':
                _r = _sp2.run(['tasklist', '/FI', f'PID eq {_pid}', '/NH', '/FO', 'CSV'],
                              capture_output=True, text=True)
                _precompute_running = str(_pid) in _r.stdout
            else:
                os.kill(_pid, 0)
                _precompute_running = True
        except Exception:
            _precompute_running = False
        if not _precompute_running:
            st.session_state.pop('_precompute_pid', None)

    # ── Top Navbar ──────────────────────────────────────────────────────────
    _sheets_ok    = SHEETS_AVAILABLE
    _dot_color    = '#4ade80' if _sheets_ok else '#fbbf24'
    _sheets_label = 'Google Sheets' if _sheets_ok else 'ออฟไลน์'
    _sheets_sub   = 'เชื่อมต่อแล้ว' if _sheets_ok else 'ออฟไลน์'
    _branch_count = len(MASTER_DATA) if not MASTER_DATA.empty else 0
    _today_str    = datetime.now().strftime('%d %b %Y')
    _rebuild_badge = '<span class="nav-badge nav-badge-warn">⏳ กำลัง Rebuild...</span>' if _precompute_running else ''
    _dot_span = f'<span style="width:8px;height:8px;border-radius:50%;background:{_dot_color};display:inline-block;flex-shrink:0;margin-right:4px"></span>'
    _badge1 = f'<span class="nav-badge">{_dot_span} {_sheets_label}: {_sheets_sub}</span>'
    _badge2 = f'<span class="nav-badge">📍 {_branch_count:,} สาขา</span>'
    _navbar_html = f'<div class="app-navbar"><div class="app-navbar-brand"><div class="brand-icon">🚚</div><div class="brand-name">Route <strong>Optimizer</strong><small>ระบบจัดเที่ยวอัจฉริยะ</small></div></div><div class="app-navbar-status">{_rebuild_badge}{_badge1}{_badge2}</div></div>'
    st.markdown(_navbar_html, unsafe_allow_html=True)

    # ── Page title ─────────────────────────────────────────────────────────
    _cache_dist = len(DISTANCE_CACHE)
    _cache_rt   = len(ROUTE_CACHE_DATA)
    st.markdown(f'<div class="page-section-title"><span>📦</span> จัดเส้นทางจัดส่ง <small>· {_today_str}</small></div>', unsafe_allow_html=True)

    # ── KPI row: 3 stats + sync button ──────────────────────────────────────
    _kc1, _kc2, _kc3, _kc4 = st.columns(4)
    with _kc1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{_branch_count:,}</div><div class="kpi-label">🏪 สาขาในระบบ</div></div>', unsafe_allow_html=True)
    with _kc2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{_cache_dist:,}</div><div class="kpi-label">📍 ระยะทางแคช</div></div>', unsafe_allow_html=True)
    with _kc3:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{_cache_rt:,}</div><div class="kpi-label">🛣️ เส้นทางแคช</div></div>', unsafe_allow_html=True)
    with _kc4:
        st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
        if st.button("🔄 ซิงค์ข้อมูล", use_container_width=True, type="primary",
                     help="ดึงข้อมูลจาก Google Sheets + rebuild distances"):
            with st.spinner("⏳ กำลังดึงข้อมูล..."):
                try:
                    st.cache_data.clear()
                    for _k in ['trip_result', 'trip_summary', '_imap_html', '_imap_key',
                                'trip_result_excel', '_imap_build_time']:
                        st.session_state.pop(_k, None)
                    import subprocess as _sp, sys as _sys
                    _precompute_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'precompute_branch_data.py')
                    if os.path.exists(_precompute_script):
                        _proc = _sp.Popen(
                            [_sys.executable, _precompute_script],
                            cwd=os.path.dirname(os.path.abspath(__file__)),
                            stdout=_sp.DEVNULL, stderr=_sp.DEVNULL,
                            creationflags=_sp.CREATE_NO_WINDOW if os.name == 'nt' else 0
                        )
                        st.session_state['_precompute_pid'] = _proc.pid
                except Exception as _re:
                    st.error(f"❌ {_re}")
                finally:
                    st.rerun()
        if _precompute_running:
            st.caption("⏳ กำลัง rebuild...")

    # โหลดโมเดล
    model_data = load_model()
    if not model_data:
        st.error("❌ ไม่พบข้อมูลโมเดล กรุณาเทรนโมเดลก่อนใช้งาน")
        st.stop()

    # ── Upload card ─────────────────────────────────────────────────────────
    st.markdown("""
<div class="hero-upload-card">
  <div class="hero-upload-title">📂 นำเข้าไฟล์ออเดอร์</div>
  <div class="hero-upload-sub">รองรับไฟล์ Excel (.xlsx) ที่มีรายการสาขาและน้ำหนัก / คิว</div>
</div>""", unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "upload", type=['xlsx'],
        help="อัปโหลดไฟล์ Excel ที่มีรายการสาขาและออเดอร์",
        label_visibility="collapsed"
    )
    
    if uploaded_file:
        # เก็บไฟล์ต้นฉบับไว้ใน session_state เพื่อใช้ตอน export
        uploaded_file_content = uploaded_file.read()
        # เคลียร์ผลลัพธ์เก่าถ้าไฟล์เปลี่ยน
        _prev_file_id = st.session_state.get('_uploaded_file_id')
        _curr_file_id = (uploaded_file.name, uploaded_file.size)
        if _prev_file_id != _curr_file_id:
            for _k in ('trip_result', 'trip_summary', 'fleet_used', 'fleet_limits', 'trip_buffers', '_trip_result_fresh', '_trip_elapsed'):
                st.session_state.pop(_k, None)
            st.session_state['_uploaded_file_id'] = _curr_file_id
        st.session_state['original_file_content'] = uploaded_file_content
        # เคลียร์ log buffer ก่อนโหลด
        st.session_state['_ui_log'] = []
        
        if _prev_file_id != _curr_file_id:
            with st.spinner("⏳ กำลังอ่านข้อมูล..."):
                # เรียก _extract_all_info ก่อน (openpyxl) — cache ผลไว้
                _orig_hdr, _orig_style, _orig_dc_raw = _extract_all_info(uploaded_file_content)
                # load_excel ใช้ pandas — ทั้งคู่ cache ด้วย content hash ไม่อ่านซ้ำ
                df = load_excel(uploaded_file_content)
                if _orig_hdr:
                    st.session_state['_orig_headers'] = _orig_hdr
                st.session_state['_orig_style_info'] = _orig_style
                if _orig_dc_raw:
                    st.session_state['_orig_dc_row_raw'] = _orig_dc_raw
                df = process_dataframe(df)
                st.session_state['_df_processed'] = df
            # rerun เพื่อให้ UI แสดงสะอาด ไม่ซ้อน spinner กับ content
            st.rerun()
        else:
            df = st.session_state.get('_df_processed')
        
        if df is not None and 'Code' in df.columns:
            total_rows = len(df)
            unique_codes = df['Code'].nunique()
            duplicate_count = total_rows - unique_codes

            # ── สรุปไฟล์ (compact) ──
            _c1, _c2, _c3, _c4 = st.columns(4)
            _c1.metric("📍 สาขา", f"{total_rows:,}")
            _c2.metric("⚖️ น้ำหนัก", f"{df['Weight'].sum():,.0f} kg")
            _c3.metric("📦 คิว", f"{df['Cube'].sum():.1f} m³")
            _prov_n = df['Province'].nunique() if 'Province' in df.columns else 0
            _c4.metric("🗺️ จังหวัด", f"{_prov_n}")

            # ==========================================
            # เติมข้อมูลพื้นที่จาก Master (vectorized)
            # ==========================================
            _filled_count = 0
            _missing_count = 0
            if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                _m = MASTER_DATA[['Plan Code', 'จังหวัด', 'อำเภอ', 'ตำบล']].copy()
                _m['_code'] = _m['Plan Code'].astype(str).str.strip().str.upper()
                _m = _m.rename(columns={'จังหวัด': '_m_prov', 'อำเภอ': '_m_dist', 'ตำบล': '_m_subdist'})
                df['_code'] = df['Code'].astype(str).str.strip().str.upper()
                df = df.merge(_m[['_code', '_m_prov', '_m_dist', '_m_subdist']].drop_duplicates('_code'),
                              on='_code', how='left')
                need_prov = df['Province'].isna() | (df['Province'] == '') | (df['Province'] == 'UNKNOWN') if 'Province' in df.columns else pd.Series([True]*len(df))
                _filled_count = int(need_prov.sum())
                if 'Province' not in df.columns:
                    df['Province'] = df['_m_prov']
                else:
                    df.loc[need_prov, 'Province'] = df.loc[need_prov, '_m_prov']
                for col_upload, col_master in [('District', '_m_dist'), ('Subdistrict', '_m_subdist')]:
                    if col_upload not in df.columns:
                        df[col_upload] = df[col_master].fillna('')
                    else:
                        need = df[col_upload].isna() | (df[col_upload] == '')
                        df.loc[need, col_upload] = df.loc[need, col_master]
                df = df.drop(columns=['_code', '_m_prov', '_m_dist', '_m_subdist'], errors='ignore')

            if 'Province' in df.columns:
                _missing_df = df[(df['Province'].isna()) | (df['Province'] == '') | (df['Province'] == 'UNKNOWN')]
                _missing_count = len(_missing_df)

            # ── รวม warnings ทั้งหมดเป็น expander เดียว ──
            _warn_items = []
            if duplicate_count > 0:
                _warn_items.append(f"⚠️ Code ซ้ำ {duplicate_count} รายการ")
            if _filled_count > 0:
                _warn_items.append(f"📍 เติมข้อมูลพื้นที่ {_filled_count} รายการ")
            if _missing_count > 0:
                _warn_items.append(f"❓ ไม่พบใน Master {_missing_count} รายการ")
            if _warn_items:
                with st.expander("ℹ️ " + " · ".join(_warn_items), expanded=False):
                    if duplicate_count > 0:
                        st.caption("**Code ซ้ำ**")
                        dup_codes = df[df.duplicated(subset=['Code'], keep=False)].groupby('Code').size().reset_index(name='ซ้ำ')
                        st.dataframe(dup_codes[dup_codes['ซ้ำ'] > 1], hide_index=True, height=150)
                    if _missing_count > 0:
                        st.caption("**สาขาที่ไม่พบข้อมูลพื้นที่**")
                        _show_cols = [c for c in ['Code', 'Name', 'Province', 'District'] if c in _missing_df.columns]
                        st.dataframe(_missing_df[_show_cols].reset_index(drop=True), hide_index=True, height=150)

            st.markdown('<div class="divider-label">⚙️ การจัดการ</div>', unsafe_allow_html=True)

            # แท็บหลัก
            tab1, tab2, tab3 = st.tabs([
                "📦 จัดเที่ยว",
                "🗺️ จัดกลุ่มตามภาค",
                "🏙️ โซนจัดส่ง"
            ])
                
            # ==========================================
            # แท็บ 1: จัดเที่ยว (ตามน้ำหนัก)
            # ==========================================
            with tab1:
                # เพิ่ม Region (cache ใน session เพื่อไม่ recompute ทุก rerun)
                if 'Region' not in df.columns and 'Province' in df.columns:
                    _region_key = f"_region_{_curr_file_id}"
                    if _region_key not in st.session_state:
                        _region_map = df['Province'].map(lambda p: get_region_name(str(p)) if p else 'ไม่ระบุ')
                        st.session_state[_region_key] = _region_map.tolist()
                    df['Region'] = st.session_state[_region_key]
                
                # ==========================================
                # ── vehicle restrictions (คำนวณครั้งเดียว cache ใน session) ──
                _vr_key = f"_vrestrict_{_curr_file_id}"
                if _vr_key not in st.session_state:
                    _vr = {code: get_max_vehicle_for_branch(code) for code in df['Code']}
                    _unmatched = []
                    if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                        _mset = set(MASTER_DATA['Plan Code'].str.strip().str.upper())
                        _unmatched = [str(c).strip().upper() for c in df['Code']
                                      if str(c).strip().upper() not in _mset]
                    st.session_state[_vr_key] = (_vr, _unmatched)
                vehicle_restrictions, unmatched_codes = st.session_state[_vr_key]

                # ── แสดง warnings แบบ compact ──
                _rc = pd.Series(vehicle_restrictions).value_counts()
                _4w, _jb = _rc.get('4W', 0), _rc.get('JB', 0)
                if unmatched_codes:
                    with st.expander(f"⚠️ {len(unmatched_codes)} สาขาไม่พบใน Master (ใช้ 6W default)"):
                        st.dataframe(df[df['Code'].isin(unmatched_codes[:20])][['Code','Name']].rename(
                            columns={'Code':'รหัส','Name':'ชื่อ'}), hide_index=True)
                if _4w > 0 or _jb > 0:
                    with st.expander(f"📋 ข้อจำกัดรถ: 4W={_4w}, JB={_jb}"):
                        _r_df = df[df['Code'].isin([k for k,v in vehicle_restrictions.items() if v in ['4W','JB']])]
                        _r_df = _r_df[['Code','Name']].copy()
                        _r_df['รถสูงสุด'] = _r_df['Code'].map(vehicle_restrictions)
                        st.dataframe(_r_df.sort_values('รถสูงสุด'), hide_index=True, height=200)

                # ── Form: settings + ปุ่มจัดทริป (ไม่ rerun ทุก input) ──
                with st.form("trip_form"):
                    _fc1, _fc2 = st.columns(2)
                    with _fc1:
                        punthai_buffer = st.number_input("🅿️ Punthai Buffer %", 80, 120, 100, 5)
                    with _fc2:
                        maxmart_buffer = st.number_input("🅼 Maxmart Buffer %", 80, 150, 100, 5)

                    _ff1, _ff2, _ff3, _ff4 = st.columns(4)
                    with _ff1:
                        fleet_4w = st.number_input("🚗 4W", 0, 99, 0, 1, help="0=ไม่จำกัด")
                    with _ff2:
                        fleet_jb = st.number_input("🚚 JB", 0, 99, 0, 1, help="0=ไม่จำกัด")
                    with _ff3:
                        fleet_6w = st.number_input("🚛 6W", 0, 99, 0, 1, help="0=ไม่จำกัด")
                    with _ff4:
                        max_qty_per_trip = st.number_input("📦 QTY/ทริป", 0, step=100, help="0=ไม่จำกัด")

                    _td1, _td2 = st.columns(2)
                    with _td1:
                        st.time_input("🕐 เริ่มโหลด", datetime.strptime("00:00","%H:%M").time(),
                                      key="load_start_time", step=1800)
                    with _td2:
                        st.date_input("📅 วันที่โหลด", datetime.now().date(),
                                      key="load_date_input", format="DD/MM/YYYY")

                    _submitted = st.form_submit_button("🚀 เริ่มจัดเที่ยว", type="primary", use_container_width=True)

                punthai_buffer_value = punthai_buffer / 100.0
                maxmart_buffer_value = maxmart_buffer / 100.0
                fleet_limits_input = {
                    '4W': int(fleet_4w) if fleet_4w > 0 else 999,
                    'JB': int(fleet_jb) if fleet_jb > 0 else 999,
                    '6W': int(fleet_6w) if fleet_6w > 0 else 999,
                }

                if _submitted:
                    # เคลียร์ log เก่า
                    st.session_state['_ui_log'] = []
                    st.session_state['_trip_log'] = []
                    # สร้าง status container แบบ popup
                    with st.status("🚀 กำลังประมวลผล...", expanded=True) as status:
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        log_area = st.empty()

                        status_text.write("⏳ กำลังเตรียมข้อมูล...")
                        progress_bar.progress(10)

                        df_to_process = df.copy()
                        progress_bar.progress(20)

                        import time as time_module, threading as _threading
                        start_time = time_module.time()

                        # ── รัน predict_trips ใน thread แยก เพื่อให้ log แสดง live ──
                        _result_box = {'result': None, 'error': None, 'done': False}

                        def _run_predict():
                            try:
                                _result_box['result'] = predict_trips(
                                    df_to_process, model_data,
                                    punthai_buffer=punthai_buffer_value,
                                    maxmart_buffer=maxmart_buffer_value,
                                    fleet_limits=fleet_limits_input,
                                    max_qty_per_trip=int(max_qty_per_trip)
                                )
                            except Exception as _ex:
                                _result_box['error'] = _ex
                            finally:
                                _result_box['done'] = True

                        _t = _threading.Thread(target=_run_predict, daemon=True)
                        _t.start()
                        # รอโดยตรง (ไม่ loop) — st.status แสดง spinner อัตโนมัติ
                        _t.join()

                        if _result_box['error']:
                            status.update(label=f"❌ เกิดข้อผิดพลาด", state="error", expanded=True)
                            st.error(f"❌ {_result_box['error']}")
                            import traceback as _tb2
                            st.code(_tb2.format_exc(), language='text')
                            st.stop()

                        result_df, summary, fleet_used = _result_box['result']
                        elapsed_time = time_module.time() - start_time

                        # snapshot log หลังเสร็จ
                        _collected_log = list(st.session_state.get('_ui_log', []))
                        st.session_state['_trip_log'] = _collected_log
                        if _collected_log:
                            log_area.code('\n'.join(_collected_log[-30:]), language=None)
                            # dump log ลงไฟล์เพื่อ debug
                            try:
                                import os as _os
                                _log_path = _os.path.join(_os.path.dirname(__file__), 'trip_debug.log')
                                with open(_log_path, 'w', encoding='utf-8') as _lf:
                                    _lf.write('\n'.join(_collected_log))
                            except Exception:
                                pass

                        # 💾 เก็บผลลัพธ์ใน session_state
                        st.session_state['trip_result'] = result_df
                        st.session_state['trip_summary'] = summary
                        st.session_state['fleet_used'] = fleet_used
                        st.session_state['fleet_limits'] = fleet_limits_input
                        st.session_state['trip_buffers'] = {
                            'punthai': punthai_buffer_value,
                            'maxmart': maxmart_buffer_value
                        }
                        st.session_state['_trip_result_fresh'] = True
                        st.session_state['_trip_elapsed'] = elapsed_time

                        progress_bar.progress(100)
                        status_text.write(f"✅ จัดทริปเสร็จสิ้น! (ใช้เวลา {elapsed_time:.1f} วินาที)")
                        status.update(label=f"✅ ประมวลผลเสร็จสมบูรณ์! ({elapsed_time:.1f}s)", state="complete", expanded=False)
                        st.session_state['_trip_just_done'] = True

                        # 🗺️ Pre-cache เส้นทาง OSRM ทุกทริปใน background thread
                        # เพื่อให้แผนที่แสดงเส้นจริงทันทีโดยไม่ต้องรอ API ขณะ render
                        def _precache_routes(df_snapshot):
                            import threading
                            _dc_lat = DC_WANG_NOI_LAT
                            _dc_lon = DC_WANG_NOI_LON
                            _trip_ids = sorted(df_snapshot[df_snapshot['Trip'] > 0]['Trip'].unique())
                            _cached_count = 0
                            for _tid in _trip_ids:
                                _tdata = df_snapshot[df_snapshot['Trip'] == _tid]
                                _pts = []
                                for _, _r in _tdata.iterrows():
                                    _la = float(_r.get('_lat', 0) or 0)
                                    _lo = float(_r.get('_lon', 0) or 0)
                                    if _la > 0 and _lo > 0:
                                        _pts.append([_la, _lo])
                                if not _pts:
                                    continue
                                _wp = [[_dc_lat, _dc_lon]] + _pts + [[_dc_lat, _dc_lon]]
                                _ck = "|".join([f"{la:.4f},{lo:.4f}" for la, lo in _wp])
                                if USE_CACHE and _ck in ROUTE_CACHE_DATA:
                                    continue  # มีแล้ว ข้าม
                                get_multi_point_route_osrm(_wp)  # จะ cache ผลอัตโนมัติ
                                _cached_count += 1
                            if _cached_count > 0:
                                save_route_cache(ROUTE_CACHE_DATA, force=True)
                                safe_print(f"🗺️ Pre-cached {_cached_count} trip routes → route_cache.json")

                        import threading as _thr
                        _rt = _thr.Thread(
                            target=_precache_routes,
                            args=(result_df.copy(),),
                            daemon=True,
                            name="precache-routes"
                        )
                        _rt.start()

                # 📊 แสดงผลลัพธ์ถ้ามีข้อมูลใน session_state
                if 'trip_result' in st.session_state and 'trip_summary' in st.session_state:
                    result_df = st.session_state['trip_result']
                    summary = st.session_state['trip_summary']
                    # guard: ล้าง NaN ใน Trip ที่อาจค้างมาจากการคำนวณ
                    if 'Trip' in result_df.columns:
                        result_df['Trip'] = result_df['Trip'].fillna(0).astype(int)

                    # ── แสดง log การจัดทริป ──
                    _trip_log = st.session_state.get('_trip_log', [])
                    if _trip_log:
                        _elapsed = st.session_state.get('_trip_elapsed', 0)
                        with st.expander(f"📋 Log การจัดทริป ({len(_trip_log)} บรรทัด · {_elapsed:.1f}s)", expanded=False):
                            st.code('\n'.join(_trip_log), language=None)

                    # ── เรียงลำดับ result_df: ทริป → ระยะทางจาก DC (ไกลก่อน) ──
                    _rd_sort_cols = ['Trip']
                    _rd_sort_asc  = [True]
                    if '_distance_from_dc' in result_df.columns:
                        _rd_sort_cols.append('_distance_from_dc')
                        _rd_sort_asc.append(False)   # ไกลก่อนภายในทริปเดียวกัน
                    result_df = result_df.sort_values(
                        _rd_sort_cols,
                        ascending=_rd_sort_asc,
                        na_position='last'
                    ).reset_index(drop=True)
                    st.session_state['trip_result'] = result_df

                    # ตรวจสอบสาขาที่ไม่ได้จัดทริป (Trip = 0)
                    unassigned_count = len(result_df[result_df['Trip'] == 0])
                    if unassigned_count > 0:
                        st.warning(f"⚠️ มี {unassigned_count} สาขาที่ไม่ได้จัดทริป (Trip = 0)")
                        with st.expander(f"🔍 ดูรายละเอียดสาขาที่ไม่ได้จัดทริป ({unassigned_count} สาขา)"):
                            _ua_cols = [c for c in ['Code', 'Name', 'BU', 'Weight', 'Cube', 'Province'] if c in result_df.columns]
                            st.dataframe(result_df[result_df['Trip'] == 0][_ua_cols].reset_index(drop=True), hide_index=True)

                    
                    # กรองเฉพาะสาขาที่จัดทริปแล้ว สำหรับการแสดงผล
                    assigned_df = result_df[result_df['Trip'] > 0].copy()
                    
                    # แสดง balloons เฉพาะครั้งแรกที่ผลลัพธ์ใหม่ (ไม่เป็นทุก rerender)
                    if st.session_state.get('_trip_result_fresh', False):
                        st.balloons()
                        st.session_state['_trip_result_fresh'] = False
                    st.success(f"✅ **จัดทริปเสร็จสมบูรณ์!** รวม **{len(summary)}** ทริป ({len(assigned_df)} สาขา)")
                    
                    st.markdown('<div class="divider-label">📊 สรุปผลการจัดทริป</div>', unsafe_allow_html=True)
                    
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("🚚 จำนวนทริป", len(summary))
                    with col2:
                        st.metric("📍 จำนวนสาขา", len(assigned_df))
                    with col3:
                        avg_branches = len(assigned_df) / max(1, assigned_df['Trip'].nunique())
                        st.metric("📊 เฉลี่ยสาขา/ทริป", f"{avg_branches:.1f}")
                    with col4:
                        avg_util = summary['Cube_Use%'].mean() if len(summary) > 0 else 0
                        st.metric("📈 การใช้รถเฉลี่ย", f"{avg_util:.0f}%")

                    # 🚛 Fleet Usage Summary
                    _fu = st.session_state.get('fleet_used', {})
                    _fl = st.session_state.get('fleet_limits', {})
                    _any_limit = any(v < 999 for v in _fl.values()) if _fl else False
                    if _fu:
                        st.markdown("---")
                        st.markdown('<div class="divider-label">🚛 การใช้รถ</div>', unsafe_allow_html=True)
                        _fc1, _fc2, _fc3 = st.columns(3)
                        for _col, _vtype, _icon in [(_fc1, '4W', '🚗'), (_fc2, 'JB', '🚚'), (_fc3, '6W', '🚛')]:
                            _used = _fu.get(_vtype, 0)
                            _limit = _fl.get(_vtype, 999) if _fl else 999
                            _limit_str = str(_limit) if _limit < 999 else '∞'
                            _delta = f"/{_limit_str} คัน"
                            _over = _limit < 999 and _used > _limit
                            with _col:
                                st.metric(f"{_icon} {_vtype}", f"{_used} ทริป", delta=_delta,
                                          delta_color="inverse" if _over else "normal")
                                if _over:
                                    st.warning(f"⚠️ เกินโควต้า {_vtype}: ใช้ {_used}/{_limit}")
                    elif _any_limit:
                        st.info("ℹ️ ตั้งโควต้ารถไว้แล้ว — จัดทริปใหม่เพื่อดูผล")
                    
                    # ⏱️ timing inline (ไม่เป็น expander)
                    _trip_elapsed = st.session_state.get('_trip_elapsed', 0)
                    _map_elapsed  = st.session_state.get('_imap_build_time', None)
                    if _trip_elapsed or _map_elapsed is not None:
                        _t_parts = []
                        if _trip_elapsed: _t_parts.append(f"จัดทริป {_trip_elapsed:.1f}s")
                        if _map_elapsed is not None: _t_parts.append(f"แผนที่ {_map_elapsed:.1f}s")
                        st.caption("⏱️ " + " · ".join(_t_parts))

                    st.markdown('<div class="divider-label">🚛 รายละเอียดแต่ละทริป</div>', unsafe_allow_html=True)

                    # styled dataframe (ใช้แบบเดียว)
                    format_dict = {}
                    gradient_cols = []
                    for _col, _fmt in [('Weight','{:.0f}'),('Cube','{:.2f}'),
                                       ('Weight_Use%','{:.1f}%'),('Cube_Use%','{:.1f}%'),
                                       ('Total_Distance','{:.1f} km')]:
                        if _col in summary.columns:
                            format_dict[_col] = _fmt
                            if _col.endswith('_Use%'): gradient_cols.append(_col)
                    try:
                        _sdf = summary.style.format(format_dict)
                        if gradient_cols:
                            _sdf = _sdf.background_gradient(subset=gradient_cols, cmap='RdYlGn', vmin=0, vmax=100)
                        st.dataframe(_sdf, use_container_width=True, height=min(400, 60+len(summary)*40))
                    except Exception:
                        st.dataframe(summary, use_container_width=True, height=400)

                    with st.expander("📋 ดูรายละเอียดรายสาขา (เรียงตามทริป → จังหวัด → อำเภอ)"):
                        # จัดเรียงคอลัมน์ที่สำคัญ
                        display_cols = ['Trip', 'Code', 'Name']
                        # จังหวัด
                        if '_province' in result_df.columns:
                            display_cols.append('_province')
                        elif 'Province' in result_df.columns:
                            display_cols.append('Province')
                        # อำเภอ
                        if '_district' in result_df.columns:
                            display_cols.append('_district')
                        elif 'District' in result_df.columns:
                            display_cols.append('District')
                        # ตำบล
                        if '_subdistrict' in result_df.columns:
                            display_cols.append('_subdistrict')
                        if 'Region' in result_df.columns:
                            display_cols.append('Region')
                        display_cols.extend(['Max_Distance_in_Trip', 'Weight', 'Cube', 'Truck', 'VehicleCheck'])
                        
                        # กรองคอลัมน์ที่มีอยู่จริง
                        display_cols = [col for col in dict.fromkeys(display_cols) if col in result_df.columns]
                        display_df = result_df[display_cols].copy()
                        
                        # ตั้งชื่อคอลัมน์ภาษาไทย
                        col_names = {'Trip': 'ทริป', 'Code': 'รหัส', 'Name': 'ชื่อสาขา',
                                   'Province': 'จังหวัด', '_province': 'จังหวัด',
                                   'District': 'อำเภอ', '_district': 'อำเภอ',
                                   '_subdistrict': 'ตำบล',
                                   'Region': 'ภาค', 'Max_Distance_in_Trip': 'ระยะทาง Max(km)', 
                                   'Weight': 'น้ำหนัก(kg)', 'Cube': 'คิว(m³)', 'Truck': 'รถ', 'VehicleCheck': 'ตรวจสอบรถ'}
                        display_df.columns = [col_names.get(c, c) for c in display_cols]
                        
                        # จัดรูปแบบคอลัมน์ระยะทาง
                        _fmt_disp = {}
                        if 'ระยะทาง Max(km)' in display_df.columns: _fmt_disp['ระยะทาง Max(km)'] = '{:.1f}'
                        if 'น้ำหนัก(kg)' in display_df.columns: _fmt_disp['น้ำหนัก(kg)'] = '{:.2f}'
                        if 'คิว(m³)' in display_df.columns: _fmt_disp['คิว(m³)'] = '{:.2f}'
                        st.dataframe(
                            display_df.style.format(_fmt_disp) if _fmt_disp else display_df,
                            width="stretch", 
                            height=500
                        )
                    
                    # แสดงสาขาที่มีคำเตือน - รวมทั้ง ⚠️ และ ❌
                    warning_branches = result_df[result_df['VehicleCheck'].str.contains('⚠️|❌', na=False, regex=True)]
                    if len(warning_branches) > 0:
                        # นับจำนวนแต่ละประเภท
                        error_count = len(result_df[result_df['VehicleCheck'].str.contains('❌', na=False)])
                        warning_count = len(result_df[result_df['VehicleCheck'].str.contains('⚠️', na=False)])
                        
                        with st.expander(f"🚨 สาขาที่มีปัญหา ({len(warning_branches)} สาขา: ❌ {error_count} ข้อจำกัด, ⚠️ {warning_count} อื่นๆ)", expanded=(error_count > 0)):
                            if error_count > 0:
                                st.error(f"❌ มี {error_count} สาขาที่ใช้รถเกินข้อจำกัดจาก Master Data!")
                            if warning_count > 0:
                                st.warning(f"⚠️ มี {warning_count} สาขาที่มีคำเตือนอื่นๆ")
                            
                            display_cols_warn = ['Trip', 'Code', 'Name', 'MaxVehicle', 'Truck', 'VehicleCheck']
                            display_warn_df = warning_branches[display_cols_warn].copy()
                            display_warn_df.columns = ['ทริป', 'รหัส', 'ชื่อสาขา', 'รถ Max', 'รถที่จัด', 'สถานะ']
                            st.dataframe(display_warn_df, width="stretch")
                    
                    # ── 📥 Excel build (cached) — สร้างครั้งเดียว ไม่ rebuild ถ้า result ไม่เปลี่ยน ──
                    import hashlib as _hl_xl
                    # อ่าน load_start_time (datetime.time จาก st.time_input)
                    _t_val = st.session_state.get('load_start_time')
                    if hasattr(_t_val, 'hour'):
                        _load_start_min = _t_val.hour * 60 + _t_val.minute
                    else:
                        try:
                            _lh2, _lm3 = map(int, str(_t_val or '00:00').split(':'))
                            _load_start_min = _lh2 * 60 + _lm3
                        except Exception:
                            _load_start_min = 0
                    # อ่าน load_date_input (datetime.date จาก st.date_input)
                    _d_val = st.session_state.get('load_date_input')
                    if hasattr(_d_val, 'strftime'):
                        _load_date_sig = _d_val.strftime('%d/%m/%Y')
                    else:
                        _load_date_sig = str(_d_val or datetime.now().strftime('%d/%m/%Y'))
                    _xl_sig = f"v9|{len(result_df)}|{safe_int_trip(result_df['Trip'].max())}|{sorted(result_df['Trip'].unique().tolist())}|{_load_start_min}|{_load_date_sig}"
                    _xl_key = _hl_xl.md5(_xl_sig.encode()).hexdigest()[:12]

                    if st.session_state.get('_excel_key') != _xl_key:
                        with st.spinner("📊 กำลังสร้างไฟล์ Excel..."):
                            import xlsxwriter as _xlw

                            # ── 1. location_map → vectorized dict map ──
                            _loc_sp = {}; _loc_sd = {}; _loc_sv = {}; _loc_rt = {}
                            if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                                _lm_cols = [c for c in ['Plan Code','ตำบล','อำเภอ','จังหวัด','Reference'] if c in MASTER_DATA.columns]
                                _lm = MASTER_DATA[_lm_cols].copy()
                                _lm['_k'] = _lm['Plan Code'].astype(str).str.strip().str.upper()
                                for _r in _lm.to_dict('records'):
                                    _k = _r['_k']
                                    if _k:
                                        _loc_sp[_k] = str(_r.get('ตำบล', '') or '')
                                        _loc_sd[_k] = str(_r.get('อำเภอ', '') or '')
                                        _loc_sv[_k] = str(_r.get('จังหวัด', '') or '')
                                        _loc_rt[_k] = str(_r.get('Reference', '') or '')

                            # ── 2. Pre-join ทั้งหมดแบบ vectorized ──
                            _rd = result_df[result_df['Trip'] != 0].copy()
                            _rd['_key_u'] = _rd['Code'].astype(str).str.strip().str.upper()
                            _rd['_sp'] = _rd['_key_u'].map(_loc_sp).fillna('')
                            _rd['_sd'] = _rd['_key_u'].map(_loc_sd).fillna('')
                            _rd['_sv'] = _rd['_key_u'].map(_loc_sv).fillna('')
                            _rd['_rt'] = _rd['_key_u'].map(_loc_rt).fillna('')

                            # fallback → ใช้ planning columns ถ้า MASTER_DATA ว่าง
                            _prov_col  = '_province'    if '_province'    in _rd.columns else ('Province'    if 'Province'    in _rd.columns else None)
                            _dist_col2 = '_district'    if '_district'    in _rd.columns else ('District'    if 'District'    in _rd.columns else None)
                            _subd_col2 = '_subdistrict' if '_subdistrict' in _rd.columns else ('Subdistrict' if 'Subdistrict' in _rd.columns else None)
                            _empty = ''
                            _rd['_sv_eff'] = _rd['_sv'].where(_rd['_sv'].str.strip() != '', _rd[_prov_col]  if _prov_col  else _empty)
                            _rd['_sd_eff'] = _rd['_sd'].where(_rd['_sd'].str.strip() != '', _rd[_dist_col2] if _dist_col2 else _empty)
                            _rd['_sp_eff'] = _rd['_sp'].where(_rd['_sp'].str.strip() != '', _rd[_subd_col2] if _subd_col2 else _empty)

                            # ── 3. pre-build province→rorder dict ครั้งเดียว ──
                            _prov_rorder: dict = {}
                            for _p in _rd['_sv_eff'].unique():
                                _prov_rorder[_p] = REGION_ORDER.get(get_region_name(str(_p)), 99)
                            _rd['_rorder'] = _rd['_sv_eff'].map(_prov_rorder).fillna(99).astype(int)

                            # ── 4. distance from result_df directly (ไม่ต้อง MASTER_DATA loop) ──
                            _dist_src_col = '_distance_from_dc' if '_distance_from_dc' in _rd.columns else None

                            # ── 5. sort keys: จังหวัด (ไกลสุดก่อน) → ระยะทริป (ไกลสุดก่อน) ──
                            trip_no_map = {}
                            vehicle_counts = {'4W': 0, 'JB': 0, '6W': 0}
                            _prov_col_rd = '_province' if '_province' in _rd.columns else ('Province' if 'Province' in _rd.columns else None)

                            # pass 1: หา avg dist + max dist + dominant province + subdistrict
                            _trip_pmx:     dict = {}
                            _trip_avg:     dict = {}
                            _trip_prov:    dict = {}
                            _trip_subdist: dict = {}
                            _subdist_col_rd = next((c for c in ['Subdistrict', '_subdistrict', 'ตำบล'] if c in _rd.columns), None)
                            for _tnum2, _tg2 in _rd.groupby('Trip', sort=False):
                                if _tnum2 == 0: continue
                                if _dist_src_col:
                                    _dv2 = _tg2[_dist_src_col].dropna()
                                    _mx2 = _dv2.max() if len(_dv2) else 0
                                    _av2 = _dv2.mean() if len(_dv2) else 0
                                else:
                                    _mx2 = _av2 = 0
                                _trip_pmx[_tnum2] = float(_mx2) if pd.notna(_mx2) else 0.0
                                _trip_avg[_tnum2] = float(_av2) if pd.notna(_av2) else 0.0
                                if _prov_col_rd:
                                    _vc2 = _tg2[_prov_col_rd].dropna().value_counts()
                                    _trip_prov[_tnum2] = str(_vc2.index[0]) if len(_vc2) else ''
                                else:
                                    _trip_prov[_tnum2] = ''
                                if _subdist_col_rd:
                                    _vcs2 = _tg2[_subdist_col_rd].dropna().value_counts()
                                    _trip_subdist[_tnum2] = str(_vcs2.index[0]) if len(_vcs2) else ''
                                else:
                                    _trip_subdist[_tnum2] = ''

                            # province → global max dist + region_order
                            _prov_gmax_rd: dict = {}
                            _prov_rord_rd: dict = {}
                            for _tn2x, _dp2x in _trip_prov.items():
                                _d2x = _trip_pmx.get(_tn2x, 0)
                                if _dp2x not in _prov_gmax_rd or _d2x > _prov_gmax_rd[_dp2x]:
                                    _prov_gmax_rd[_dp2x] = _d2x
                                if _dp2x and _dp2x not in _prov_rord_rd:
                                    _prov_rord_rd[_dp2x] = REGION_ORDER.get(get_region_name(str(_dp2x)), 99)

                            # ใช้ trip number ตามลำดับ (FINAL RE-NUMBER จัดไว้ถูกแล้ว)
                            # ไม่ re-sort อีกเพื่อป้องกัน export สลับกับ algorithm
                            trip_sort_keys = {_tn3: (_tn3,) for _tn3 in _trip_pmx}

                            sorted_trips = sorted(
                                [t for t in result_df['Trip'].unique() if t > 0]
                            )

                            trip_vehicle_map = {}   # _tnum → '4W'/'JB'/'6W'
                            _global_trip_seq = 0
                            for _tnum in sorted_trips:
                                _global_trip_seq += 1
                                _ts = summary[summary['Trip'] == _tnum]
                                _vt = '6W'
                                if len(_ts) > 0:
                                    _vi = _ts.iloc[0]['Truck']
                                    _vt = _vi.split()[0] if _vi else '6W'
                                    if _vt in ('4WJ',): _vt = 'JB'
                                    if _vt not in vehicle_counts: _vt = '6W'
                                    vehicle_counts[_vt] = vehicle_counts.get(_vt, 0) + 1
                                trip_vehicle_map[_tnum] = _vt
                                trip_no_map[_tnum] = f"{_vt}{_global_trip_seq:03d}"
                            # บันทึก trip_no_map ทันทีหลังสร้าง เพื่อให้แผนที่ใช้ได้ก่อน export
                            st.session_state['_trip_no_map'] = dict(trip_no_map)
                            # ── 5.5 helper functions สำหรับ load schedule (ใช้ใน section 7.5) ──
                            _PT_RATE_PURE = 25000   # P ล้วน: 25,000 ชิ้น/ชม.
                            _MM_RATE_PURE = 35000   # M ล้วน: 35,000 ชิ้น/ชม.
                            _MIX_RATE     = 40000   # P+M คละ: 15,000+25,000 = 40,000 ชิ้น/ชม. (2 ช่องพร้อมกัน)
                            _BREAK_STARTS = {7*60, 13*60, 19*60, 1*60}
                            _BREAK_DUR    = 60
                            _HUB_STARTS   = {8*60, 10*60, 12*60, 14*60}
                            _HUB_DUR      = 60
                            _DOORS_6W     = [2, 9]
                            _DOORS_SMALL  = [d for d in range(1, 10) if d not in _DOORS_6W]

                            def _skip_blocked(t_min):
                                for _ in range(20):
                                    t_mod = t_min % 1440
                                    changed = False
                                    for _bs in _BREAK_STARTS:
                                        if _bs <= t_mod < _bs + _BREAK_DUR:
                                            t_min += (_bs + _BREAK_DUR) - t_mod
                                            changed = True
                                            break
                                    if changed: continue
                                    for _hs in _HUB_STARTS:
                                        if _hs <= t_mod < _hs + _HUB_DUR:
                                            t_min += (_hs + _HUB_DUR) - t_mod
                                            changed = True
                                            break
                                    if not changed:
                                        break
                                return t_min

                            def _fmt_time(t_min):
                                t_mod = int(t_min) % 1440
                                return f"{t_mod//60:02d}:{t_mod%60:02d}"

                            def _fmt_date(t_min, base_date_str):
                                try:
                                    from datetime import datetime as _dt, timedelta as _td
                                    _bd = _dt.strptime(base_date_str.strip(), '%d/%m/%Y')
                                    _days = int(t_min) // 1440
                                    return (_bd + _td(days=_days)).strftime('%d/%m/%Y')
                                except Exception:
                                    return base_date_str

                            _trip_order_map = {t: i for i, t in enumerate(sorted_trips)}
                            _rd['_trip_order'] = _rd['Trip'].map(_trip_order_map).fillna(9999)

                            # เรียงแถว: trip → จังหวัด(ไกล) → จังหวัด → อำเภอ(ไกล) → อำเภอ → ตำบล(ไกล) → ตำบล → ระยะทาง(ไกล)
                            if _dist_src_col:
                                _rd['_prov_maxd_rd2'] = (
                                    _rd.groupby(['_trip_order', '_sv_eff'])[_dist_src_col]
                                    .transform('max').fillna(0)
                                )
                                _rd['_dist_maxd_rd2'] = (
                                    _rd.groupby(['_trip_order', '_sv_eff', '_sd_eff'])[_dist_src_col]
                                    .transform('max').fillna(0)
                                )
                                _rd['_subd_maxd_rd2'] = (
                                    _rd.groupby(['_trip_order', '_sv_eff', '_sd_eff', '_sp_eff'])[_dist_src_col]
                                    .transform('max').fillna(0)
                                )
                                _sc_rd2 = ['_trip_order',
                                           '_prov_maxd_rd2', '_sv_eff',
                                           '_dist_maxd_rd2', '_sd_eff',
                                           '_subd_maxd_rd2', '_sp_eff',
                                           _dist_src_col]
                                _sa_rd2 = [True, False, True, False, True, False, True, False]
                                _rd = _rd.sort_values(_sc_rd2, ascending=_sa_rd2)
                                _rd = _rd.drop(columns=['_prov_maxd_rd2', '_dist_maxd_rd2', '_subd_maxd_rd2'], errors='ignore')
                            else:
                                _rd = _rd.sort_values(['_trip_order', '_sv_eff', '_sd_eff', '_sp_eff'], ascending=[True, True, True, True])

                            # ── 7. pre-group rows ──
                            _trip_rows: dict = {}
                            for _rec in _rd.to_dict('records'):
                                _t_val = _rec.get('Trip', 0)
                                try:
                                    _t_key = int(_t_val) if _t_val is not None and str(_t_val) not in ('nan', 'None', '') else 0
                                except (ValueError, TypeError):
                                    _t_key = 0
                                _trip_rows.setdefault(_t_key, []).append(_rec)

                            # ── 7.5 คำนวณ load schedule (เวลาโหลด/วันที่/ประตู) ──
                            # อ่านวันที่ (datetime.date หรือ string)
                            _d_val2 = st.session_state.get('load_date_input')
                            if hasattr(_d_val2, 'strftime'):
                                _base_date = _d_val2.strftime('%d/%m/%Y')
                            else:
                                _base_date = str(_d_val2 or datetime.now().strftime('%d/%m/%Y'))
                            _cur_min = _load_start_min
                            _cur_min = _skip_blocked(_cur_min)

                            trip_load_date: dict = {}
                            trip_load_time: dict = {}
                            trip_door:      dict = {}
                            _door_small_idx = 0
                            _6w_door_idx    = 0
                            _hour_6w_count: dict = {}

                            # ── batch scheduling: บวกจำนวนชิ้นสะสม แล้วใส่เวลา ──
                            # ทุกทริปใน batch เดียวกัน → ได้เวลาเดียวกัน
                            # เมื่อ qty รวม > max_qty_sched → flush (คำนวณ duration จาก qty รวม) แล้วขึ้น batch ใหม่
                            # max_qty_sched = 0 → แต่ละทริปเป็น batch ตัวเอง (ใช้ qty ทริปนั้นคำนวณ)
                            _max_qty_sched = int(max_qty_per_trip) if max_qty_per_trip and int(max_qty_per_trip) > 0 else 0

                            def _batch_rate_from_flags(has_pt, has_mm):
                                """เลือกอัตราโหลดตามประเภทสินค้าใน batch"""
                                if has_pt and has_mm:
                                    return _MIX_RATE       # P+M คละ: 40,000 ชิ้น/ชม.
                                elif has_pt:
                                    return _PT_RATE_PURE   # P ล้วน:  25,000 ชิ้น/ชม.
                                else:
                                    return _MM_RATE_PURE   # M ล้วน:  35,000 ชิ้น/ชม.

                            _batch_qty   = 0.0   # qty สะสม batch ปัจจุบัน
                            _batch_start = _cur_min
                            _batch_has_pt = False  # มี P ใน batch ไหม
                            _batch_has_mm = False  # มี M ใน batch ไหม

                            # limit ต่อชั่วโมง: ถ้าผู้ใช้กำหนด max_qty_per_trip → ใช้ค่านั้น
                            # ถ้าไม่กำหนด → ใช้อัตราตาม BU (PT=25k, MM=35k, Mix=40k)
                            _user_qty_limit = int(max_qty_per_trip) if max_qty_per_trip and int(max_qty_per_trip) > 0 else 0

                            def _get_batch_limit(has_pt, has_mm):
                                if _user_qty_limit > 0:
                                    return _user_qty_limit
                                return _batch_rate_from_flags(has_pt, has_mm)

                            for _tnum in sorted_trips:
                                _vt_s = trip_vehicle_map.get(_tnum, '6W')
                                _rows_s = _trip_rows.get(_tnum, [])
                                _is_pt_s = all(str(_r.get('BU', '')).upper() in ('211', 'PUNTHAI') for _r in _rows_s)
                                _is_mm_s = not _is_pt_s and all(str(_r.get('BU', '')).upper() not in ('211', 'PUNTHAI') for _r in _rows_s)
                                _trip_has_pt = _is_pt_s or (not _is_mm_s)  # มี P component
                                _trip_has_mm = _is_mm_s or (not _is_pt_s)  # มี M component
                                _trip_qty_s = sum(float(_r.get('OriginalQty', 1) or 1) for _r in _rows_s)
                                if _trip_qty_s <= 0: _trip_qty_s = len(_rows_s) * 10

                                # ตรวจว่า batch เต็มไหม (อิง BU rate หรือค่าผู้ใช้)
                                _limit_now = _get_batch_limit(
                                    _batch_has_pt or _trip_has_pt,
                                    _batch_has_mm or _trip_has_mm
                                )
                                if _batch_qty > 0 and _batch_qty + _trip_qty_s > _limit_now:
                                    # flush: เลื่อน 1 ชั่วโมง แล้วขึ้น batch ใหม่
                                    _batch_start = _skip_blocked(_batch_start + 60)
                                    _batch_qty    = 0.0
                                    _batch_has_pt = False
                                    _batch_has_mm = False

                                # ใส่เวลา batch ปัจจุบัน
                                trip_load_date[_tnum] = _fmt_date(_batch_start, _base_date)
                                trip_load_time[_tnum] = _fmt_time(_batch_start)

                                # สะสม qty + flags ของ batch
                                _batch_qty   += _trip_qty_s
                                _batch_has_pt = _batch_has_pt or _is_pt_s
                                _batch_has_mm = _batch_has_mm or _is_mm_s

                                # ประตู (คละรถทุกประเภทแชร์ประตูตามขนาดรถ)
                                _hour_bucket = int(_batch_start) // 60
                                if _vt_s == '6W':
                                    _door = _DOORS_6W[_6w_door_idx % len(_DOORS_6W)]
                                    _6w_door_idx += 1
                                    _hour_6w_count[_hour_bucket] = _hour_6w_count.get(_hour_bucket, 0) + 1
                                else:
                                    _door = _DOORS_SMALL[_door_small_idx % len(_DOORS_SMALL)]
                                    _door_small_idx += 1
                                trip_door[_tnum] = _door

                            # ── 8. failed_trips — สีแดงเมื่อ util ต่ำกว่า 98% ──
                            failed_trips = set()
                            for _t in sorted_trips:
                                _rows_t = _trip_rows.get(_t, [])
                                if not _rows_t: continue
                                _is_pt = all(str(_r.get('BU', '')).upper() in ('211', 'PUNTHAI') for _r in _rows_t)
                                _vt2 = trip_vehicle_map.get(_t, '6W')
                                _lim = (PUNTHAI_LIMITS if _is_pt else LIMITS).get(_vt2, LIMITS['6W'])
                                _tw  = sum(float(_r.get('Weight', 0) or 0) for _r in _rows_t)
                                _tc  = sum(float(_r.get('Cube',   0) or 0) for _r in _rows_t)
                                if (_tw / _lim['max_w'] * 100) < 98 and (_tc / _lim['max_c'] * 100) < 98:
                                    failed_trips.add(_t)

                            # ── 9. xlsxwriter write ──
                            _output = io.BytesIO()
                            try:
                                _wb_xl = _xlw.Workbook(_output, {'in_memory': True, 'constant_memory': True})
                                _ws_xl = _wb_xl.add_worksheet('2.Punthai')

                                # ── ดึงสไตล์ต้นฉบับ (font/row height) ──
                                _ostyle = st.session_state.get('_orig_style_info', {})
                                _ofont  = _ostyle.get('font_name', 'Angsana New')
                                _ofsize = _ostyle.get('font_size', 14.0)
                                _orh    = _ostyle.get('row_height', 15.0)
                                def _f(d):
                                    """เพิ่ม font ต้นฉบับเข้า format dict"""
                                    return {**d, 'font_name': _ofont, 'font_size': _ofsize}

                                _hdr_fmt = _wb_xl.add_format(_f({'bold':True,'border':1,'bg_color':'#D9D9D9','align':'center'}))
                                _yfmt    = _wb_xl.add_format(_f({'bg_color':'#FFE699','border':1,'font_color':'#000000'}))
                                _wfmt    = _wb_xl.add_format(_f({'bg_color':'#FFFFFF','border':1,'font_color':'#000000'}))
                                _yfmt_r  = _wb_xl.add_format(_f({'bg_color':'#FFE699','border':1,'font_color':'#FF0000'}))
                                _wfmt_r  = _wb_xl.add_format(_f({'bg_color':'#FFFFFF','border':1,'font_color':'#FF0000'}))
                                _ynfmt   = _wb_xl.add_format(_f({'bg_color':'#FFE699','border':1,'num_format':'#,##0.00','font_color':'#000000'}))
                                _wnfmt   = _wb_xl.add_format(_f({'bg_color':'#FFFFFF','border':1,'num_format':'#,##0.00','font_color':'#000000'}))
                                _ynfmt_r = _wb_xl.add_format(_f({'bg_color':'#FFE699','border':1,'num_format':'#,##0.00','font_color':'#FF0000'}))
                                _wnfmt_r = _wb_xl.add_format(_f({'bg_color':'#FFFFFF','border':1,'num_format':'#,##0.00','font_color':'#FF0000'}))
                                # ── แดง: ทริปที่ util < 98% ──
                                _LOW_BG  = '#FFCCCC'  # พื้นแดงอ่อน
                                _LOW_FG  = '#CC0000'  # ตัวอักษรแดงเข้ม
                                _rfmt    = _wb_xl.add_format(_f({'bg_color':_LOW_BG,'border':1,'font_color':_LOW_FG}))
                                _rnfmt   = _wb_xl.add_format(_f({'bg_color':_LOW_BG,'border':1,'num_format':'#,##0.00','font_color':_LOW_FG}))
                                # DC summary row formats
                                _dc_fmt  = _wb_xl.add_format(_f({'bg_color':'#BDD7EE','border':1,'bold':True}))
                                _dc_nfmt = _wb_xl.add_format(_f({'bg_color':'#BDD7EE','border':1,'bold':True,'num_format':'#,##0.00'}))
                                # helper: 0-indexed col → Excel letter (A, B, ..., Z, AA, ...)
                                def _col_letter(n):
                                    s = ''
                                    n += 1
                                    while n > 0:
                                        n, r = divmod(n - 1, 26)
                                        s = chr(65 + r) + s
                                    return s

                                # ── column plan จากหัวคอลัมน์ต้นฉบับ ──
                                # คอลัมน์ที่ระบบเพิ่มเองหรือ internal — ห้ามซ้ำใน extra_export_cols
                                _FIXED_EXPORT_COLS = {
                                    'BU','Code','WMSCode','Name','Cube','Weight','OriginalQty','Trip',
                                    'Truck','MaxVehicle','VehicleCheck','Region','Distance_from_DC',
                                    'Province','District','Subdistrict','TripNo','Booking',
                                    'Latitude','Longitude','Max_Distance_in_Trip',
                                    'Route','Reference',   # ไม่ให้ extra_cols เพิ่ม Route จาก _rd
                                }
                                _extra_export_cols = [
                                    c for c in _rd.columns
                                    if isinstance(c, str) and c.strip()
                                    and c not in _FIXED_EXPORT_COLS
                                    and not c.startswith('_')
                                    and c.lower() not in {'หมายเหตุ', 'remark'}
                                ]

                                # โหลด original header info
                                _orig_hdr_info  = st.session_state.get('_orig_headers', [])   # [(orig_name, color), ...]
                                _rename_map_sv  = st.session_state.get('_col_rename_map', {}) # orig → internal
                                # DC row: map original col names → internal keys
                                _dc_row_raw = st.session_state.get('_orig_dc_row_raw', {})
                                _dc_row = {_rename_map_sv.get(k, k): v for k, v in _dc_row_raw.items()}

                                # format cache สำหรับสีหัวคอลัมน์
                                _hdr_fmt_cache = {}
                                def _get_hdr_fmt(c):
                                    if c not in _hdr_fmt_cache:
                                        _hdr_fmt_cache[c] = _wb_xl.add_format(_f({'bold':True,'border':1,'bg_color':c,'align':'center','font_color':'#000000'}))
                                    return _hdr_fmt_cache[c]

                                # default widths ต่อ internal name
                                _DEF_W = {'__SEP__':6,'BU':6,'Code':12,'WMSCode':12,'Name':30,
                                          'Cube':11,'Weight':11,'OriginalQty':12,
                                          '__SUB_GEN__':14,'__DIS_GEN__':14,'__PROV_GEN__':16,'__PROV_BLANK__':16,
                                          '__TRIP__':6,'__TRIPNO__':10,
                                          '__LOAD_DATE__':14,'__LOAD_TIME__':14,'__DOOR__':8,'__REMARK__':28}

                                # สร้าง column plan: [(display_name, color, internal_key), ...]
                                _GEO_COLOR  = '#E2EFDA'
                                _TRIP_COLOR = '#BDD7EE'
                                # คอลัมน์ geo จากต้นฉบับ → ข้าม (จะแทรกหลัง Name แทน)
                                _SKIP_IKEYS = {'Route','Reference','_rt',
                                               'Subdistrict','ตำบล','District','อำเภอ','Province','จังหวัด',
                                               '__TRIP__','__TRIPNO__',
                                               '__PROV_BLANK__','__SUB_GEN__','__DIS_GEN__','__PROV_GEN__',
                                               '__LOAD_DATE__','__LOAD_TIME__','__DOOR__',
                                               'remark','Remark','REMARK','หมายเหตุ'}
                                # Trip / Trip no / วันที่โหลด / เวลาโหลด / ประตู → แทนด้วย generated ในตำแหน่งเดิม
                                _SCH_COLOR  = '#FCE4D6'   # สีส้มอ่อน — วันที่/เวลา/ประตู
                                _REPLACE_MAP = {
                                    'Sep.':            ('__SEP__',       '#D9D9D9'),
                                    'Sep':             ('__SEP__',       '#D9D9D9'),
                                    'SEP':             ('__SEP__',       '#D9D9D9'),
                                    'sep.':            ('__SEP__',       '#D9D9D9'),
                                    'ลำดับ':          ('__SEP__',       '#D9D9D9'),
                                    'Trip':            ('__TRIP__',      _TRIP_COLOR),
                                    'T':               ('__TRIP__',      _TRIP_COLOR),
                                    'Trip no':         ('__TRIPNO__',    _TRIP_COLOR),
                                    'Trip No':         ('__TRIPNO__',    _TRIP_COLOR),
                                    'TripNo':          ('__TRIPNO__',    _TRIP_COLOR),
                                    'TripNumber':      ('__TRIPNO__',    _TRIP_COLOR),
                                    'วันที่โหลด':    ('__LOAD_DATE__', _SCH_COLOR),
                                    'LoadDate':        ('__LOAD_DATE__', _SCH_COLOR),
                                    'เวลาโหลด(ประมาณ)': ('__LOAD_TIME__', _SCH_COLOR),
                                    'เวลาโหลด':     ('__LOAD_TIME__', _SCH_COLOR),
                                    'LoadTime':        ('__LOAD_TIME__', _SCH_COLOR),
                                    'ประตู':          ('__DOOR__',      _SCH_COLOR),
                                    'Door':            ('__DOOR__',      _SCH_COLOR),
                                }
                                _col_plan = []
                                _orig_has_trip    = False
                                _orig_has_tripno  = False
                                _orig_has_loaddate= False
                                _orig_has_loadtime= False
                                _orig_has_door    = False
                                _orig_has_sep     = False
                                _geo_inserted    = False   # แทรก geo หลัง Name แค่ครั้งเดียว
                                if _orig_hdr_info:
                                    for _oname, _oclr in _orig_hdr_info:
                                        _ikey = _rename_map_sv.get(_oname, _oname)
                                        if _ikey in _SKIP_IKEYS:
                                            continue
                                        if _ikey in _REPLACE_MAP:
                                            _gen_ikey, _gen_clr = _REPLACE_MAP[_ikey]
                                            _col_plan.append((_oname, _gen_clr, _gen_ikey))
                                            if _gen_ikey == '__SEP__':
                                                _orig_has_sep = True
                                            elif _gen_ikey == '__TRIP__':
                                                _orig_has_trip = True
                                            elif _gen_ikey == '__TRIPNO__':
                                                _orig_has_tripno = True
                                            elif _gen_ikey == '__LOAD_DATE__':
                                                _orig_has_loaddate = True
                                            elif _gen_ikey == '__LOAD_TIME__':
                                                _orig_has_loadtime = True
                                            elif _gen_ikey == '__DOOR__':
                                                _orig_has_door = True
                                        else:
                                            _col_plan.append((_oname, _oclr, _ikey))
                                            # แทรก ตำบล/อำเภอ/จังหวัด(generated) หลัง Name
                                            if _ikey == 'Name' and not _geo_inserted:
                                                _geo_inserted = True
                                                _col_plan.append(('ตำบล',    _GEO_COLOR, '__SUB_GEN__'))
                                                _col_plan.append(('อำเภอ',   _GEO_COLOR, '__DIS_GEN__'))
                                                _col_plan.append(('จังหวัด', _GEO_COLOR, '__PROV_GEN__'))
                                else:
                                    # fallback ถ้าไม่มีข้อมูลต้นฉบับ
                                    _col_plan = [
                                        ('Sep.',        '#D9D9D9', '__SEP__'),
                                        ('BU',          '#D9D9D9', 'BU'),
                                        ('รหัสสาขา',    '#D9D9D9', 'Code'),
                                        ('รหัส WMS',    '#D9D9D9', 'WMSCode'),
                                        ('สาขา',        '#D9D9D9', 'Name'),
                                        ('ตำบล',        _GEO_COLOR, '__SUB_GEN__'),
                                        ('อำเภอ',       _GEO_COLOR, '__DIS_GEN__'),
                                        ('จังหวัด',     _GEO_COLOR, '__PROV_GEN__'),
                                        ('Total Cube',  '#D9D9D9', 'Cube'),
                                        ('Total Wgt',   '#D9D9D9', 'Weight'),
                                        ('Original QTY','#D9D9D9', 'OriginalQty'),
                                    ]

                                # ต่อด้วย extra cols จากต้นฉบับที่ยังไม่อยู่ใน plan
                                _plan_ikeys  = {k for _, _, k in _col_plan}
                                _plan_dnames = {n for n, _, _ in _col_plan}
                                for _ec in _extra_export_cols:
                                    if _ec not in _plan_ikeys and _ec not in _plan_dnames:
                                        _col_plan.append((_ec, '#D9D9D9', _ec))

                                # เพิ่ม T + Trip no ท้ายสุด เฉพาะกรณีต้นฉบับไม่มี
                                if not _orig_has_sep:
                                    _col_plan.insert(0, ('Sep.', '#D9D9D9', '__SEP__'))
                                if not _orig_has_trip:
                                    _col_plan.append(('T',           _TRIP_COLOR, '__TRIP__'))
                                if not _orig_has_tripno:
                                    _col_plan.append(('Trip no',     _TRIP_COLOR, '__TRIPNO__'))
                                # วันที่โหลด / เวลาโหลด / ประตู — ต่อท้าย ถ้าต้นฉบับไม่มี
                                if not _orig_has_loaddate:
                                    _col_plan.append(('วันที่โหลด',       _SCH_COLOR, '__LOAD_DATE__'))
                                if not _orig_has_loadtime:
                                    _col_plan.append(('เวลาโหลด(ประมาณ)', _SCH_COLOR, '__LOAD_TIME__'))
                                if not _orig_has_door:
                                    _col_plan.append(('ประตู',            _SCH_COLOR, '__DOOR__'))
                                # หมายเหตุ: แทรกทันทีหลัง ประตู (__DOOR__) — ถ้าไม่มีประตู → ต่อท้ายก่อนจังหวัด
                                _door_idx = next((i for i, (_, _, k) in enumerate(_col_plan) if k == '__DOOR__'), None)
                                if _door_idx is not None:
                                    _col_plan.insert(_door_idx + 1, ('หมายเหตุ', _SCH_COLOR, '__REMARK__'))
                                else:
                                    _col_plan.append(('หมายเหตุ', _SCH_COLOR, '__REMARK__'))
                                # จังหวัด (ว่าง) — ท้ายสุดเสมอ
                                _col_plan.append(('จังหวัด', _GEO_COLOR, '__PROV_BLANK__'))

                                # ── Set column widths FIRST (must precede any write() with constant_memory=True) ──
                                for _ci, (_, _, _ikey) in enumerate(_col_plan):
                                    _cw = _DEF_W.get(_ikey, max(12, len(str(_ikey))+2))
                                    _ws_xl.set_column(_ci, _ci, _cw)

                                # ── Title row (row 0): แผนงาน + วันที่ (อิงจาก load_date_input) ──
                                _title_fmt  = _wb_xl.add_format(_f({'bold':True,'font_size':_ofsize,'align':'left'}))
                                _title_rfmt = _wb_xl.add_format(_f({'bold':True,'font_size':_ofsize,'num_format':'#,##0','font_color':'#FF0000','align':'right'}))
                                try:
                                    from datetime import datetime as _dt2, timedelta as _td2
                                    _bd2 = _dt2.strptime(_base_date, '%d/%m/%Y')
                                    _ord_d = _bd2.strftime('%d/%m/%y')
                                    _pik_d = (_bd2 + _td2(days=1)).strftime('%d/%m/%y')
                                except Exception:
                                    _ord_d = _base_date; _pik_d = _base_date
                                _title_text = f"แผนงานรอบสั่งวันที่ {_ord_d} รอบหยิบวันที่ {_pik_d}"
                                _ws_xl.write(0, 1, _title_text, _title_fmt)
                                # SUM จำนวนชิ้นทั้งหมด (สีแดง) ในหัวแถว
                                _qty_ci = next((i for i, (_, _, k) in enumerate(_col_plan) if k == 'OriginalQty'), None)
                                if _qty_ci is not None:
                                    _qty_ltr = _col_letter(_qty_ci)
                                    _ws_xl.write_formula(0, _qty_ci, f'=SUM({_qty_ltr}3:{_qty_ltr}9999)', _title_rfmt, 0)
                                _ws_xl.set_row(0, 20)

                                # ── Header row (row 1) ──
                                for _ci, (_dname, _dcolor, _ikey) in enumerate(_col_plan):
                                    _ws_xl.write(1, _ci, _dname, _get_hdr_fmt(_dcolor))
                                _ws_xl.set_row(1, 18)

                                # format สำหรับแถวว่างท้ายทริป
                                _sep_fmt = _wb_xl.add_format({'border':0})

                                # ── คำนวณ constraint ที่ถึงก่อนต่อทริป (น้ำหนัก/คิว) ──
                                trip_remark = {}
                                for _t_rm in sorted_trips:
                                    _rows_rm = _trip_rows.get(_t_rm, [])
                                    _vt_rm = trip_vehicle_map.get(_t_rm, '6W')
                                    _is_pt_rm = all(str(_r.get('BU','')).strip() in ('211','PUNTHAI') for _r in _rows_rm) if _rows_rm else False
                                    _lim_rm = (PUNTHAI_LIMITS if _is_pt_rm else LIMITS).get(_vt_rm, LIMITS['6W'])
                                    _tw_rm = sum(float(_r.get('Weight',0) or 0) for _r in _rows_rm)
                                    _tc_rm = sum(float(_r.get('Cube',0) or 0) for _r in _rows_rm)
                                    _wu_rm = _tw_rm / _lim_rm['max_w'] if _lim_rm['max_w'] > 0 else 0
                                    _cu_rm = _tc_rm / _lim_rm['max_c'] if _lim_rm['max_c'] > 0 else 0
                                    if _wu_rm >= _cu_rm:
                                        trip_remark[_t_rm] = f"น้ำหนัก {int(_tw_rm):,}/{int(_lim_rm['max_w']):,}kg ({_wu_rm*100:.0f}%)"
                                    else:
                                        trip_remark[_t_rm] = f"คิว {_tc_rm:.2f}/{_lim_rm['max_c']:.1f}m\u00b3 ({_cu_rm*100:.0f}%)"

                                # export ตาม trip number ที่ FINAL RE-NUMBER จัดไว้
                                # trip_no_map ถูก build ครั้งเดียวตอน display (ไม่ rebuild เพื่อป้องกัน label สลับ)
                                export_sorted_trips = sorted(sorted_trips)

                                # ── pre-compute util% ต่อทริป (เทียบกับ max ไม่รวม buffer เหมือน remark column) ──
                                _trip_util_map: dict = {}
                                for _tu in sorted_trips:
                                    _ru = _trip_rows.get(_tu, [])
                                    _vtu = trip_vehicle_map.get(_tu, '6W')
                                    _is_pt_u = all(str(_r.get('BU','')).strip() in ('211','PUNTHAI') for _r in _ru) if _ru else False
                                    _lim_u = (PUNTHAI_LIMITS if _is_pt_u else LIMITS).get(_vtu, LIMITS['6W'])
                                    _tw_u = sum(float(_r.get('Weight',0) or 0) for _r in _ru)
                                    _tc_u = sum(float(_r.get('Cube',0) or 0) for _r in _ru)
                                    _wu_u = _tw_u / _lim_u['max_w'] if _lim_u['max_w'] > 0 else 0
                                    _cu_u = _tc_u / _lim_u['max_c'] if _lim_u['max_c'] > 0 else 0
                                    _trip_util_map[_tu] = max(_wu_u, _cu_u)

                                use_yellow = True
                                _row_xl = 2
                                _row_seq = 1   # Sep. sequential ต่อแถว (รวม DC row)

                                # Pre-compute remark text ต่อ trip (static — ไม่ใช้ SUMIF ป้องกันไฟล์ค้าง)
                                _trip_remark_text: dict = {}
                                for _tr_t in export_sorted_trips:
                                    _tr_rows = _trip_rows.get(_tr_t, [])
                                    _tr_veh  = trip_vehicle_map.get(_tr_t, '6W')
                                    _tr_is_pt = all(str(_r.get('BU','')).strip() in ('211','PUNTHAI') for _r in _tr_rows) if _tr_rows else False
                                    _tr_lim  = (PUNTHAI_LIMITS if _tr_is_pt else LIMITS).get(_tr_veh, LIMITS['6W'])
                                    _tr_w    = sum(float(_r.get('Weight', 0) or 0) for _r in _tr_rows)
                                    _tr_c    = sum(float(_r.get('Cube',   0) or 0) for _r in _tr_rows)
                                    _tr_mw   = _tr_lim['max_w']
                                    _tr_mc   = _tr_lim['max_c']
                                    _tr_uw   = _tr_w / _tr_mw if _tr_mw > 0 else 0
                                    _tr_uc   = _tr_c / _tr_mc if _tr_mc > 0 else 0
                                    if _tr_uw >= _tr_uc:
                                        _trip_remark_text[_tr_t] = f"น้ำหนัก {_tr_w:,.0f}/{_tr_mw:,.0f}kg ({_tr_uw:.0%})"
                                    else:
                                        _trip_remark_text[_tr_t] = f"คิว {_tr_c:,.2f}/{_tr_mc:,.0f}m³ ({_tr_uc:.0%})"
                                _tn_col = None  # ไม่ใช้ SUMIF formula อีกต่อไป

                                _xl_trip_seq = 0
                                for _tnum in export_sorted_trips:
                                    _xl_trip_seq += 1
                                    _rows = _trip_rows.get(_tnum, [])
                                    _tno  = trip_no_map.get(_tnum, '')
                                    _is_low_util = _tnum in failed_trips
                                    if _is_low_util:
                                        _tf = _yfmt_r if use_yellow else _wfmt_r
                                        _nf = _ynfmt_r if use_yellow else _wnfmt_r
                                    else:
                                        _tf = _yfmt if use_yellow else _wfmt
                                        _nf = _ynfmt if use_yellow else _wnfmt
                                    _tnum_int = _xl_trip_seq  # sequential 1,2,3… ตาม export order
                                    _tno_str  = str(_tno)
                                    _first_row_of_trip = True
                                    _trip_start_row = _row_xl  # เก็บแถวเริ่มต้น (0-indexed) สำหรับสูตร SUM
                                    for _rec in _rows:
                                        for _ci, (_, _, _ikey) in enumerate(_col_plan):
                                            _is_num = _ikey in ('Cube', 'Weight')
                                            _dfmt = _nf if _is_num else _tf
                                            if _ikey == '__SEP__':
                                                _val = _row_seq  # sequential ทุกแถว
                                            elif _ikey == '__PROV_BLANK__':
                                                _val = ''   # ว่างไว้ให้ user กรอกเอง
                                            elif _ikey == '__TRIP__':
                                                if _tn_col:
                                                    _xl_r2   = _row_xl + 1
                                                    _tn_r2   = f'${_tn_col}{_xl_r2}'
                                                    _tn_rng2 = f'${_tn_col}$3:${_tn_col}{_xl_r2}'
                                                    _sp2 = f'IFERROR(SUMPRODUCT((1/COUNTIF({_tn_rng2},{_tn_rng2}))*({_tn_rng2}<>"")),1)'
                                                    _ws_xl.write_formula(_row_xl, _ci,
                                                        f'=IF({_tn_r2}=""," ",TEXT({_sp2},"000"))',
                                                        _dfmt, str(_tnum_int))
                                                    continue
                                                _val = _tnum_int
                                            elif _ikey == '__TRIPNO__':
                                                _val = _tno_str
                                            elif _ikey == 'BU':
                                                _val = _rec.get('BU', 211)
                                            elif _ikey == 'Code':
                                                _val = str(_rec.get('Code', ''))
                                            elif _ikey == 'WMSCode':
                                                _val = str(_rec.get('WMSCode', _rec.get('Code', '')))
                                            elif _ikey == 'Name':
                                                _val = str(_rec.get('Name', ''))
                                            elif _ikey == 'Cube':
                                                _val = round(float(_rec.get('Cube', 0) or 0), 2)
                                            elif _ikey == 'Weight':
                                                _val = round(float(_rec.get('Weight', 0) or 0), 2)
                                            elif _ikey == 'OriginalQty':
                                                _val = safe_qty(_rec.get('OriginalQty', 0))
                                            elif _ikey == '__SUB_GEN__':
                                                _val = str(_rec.get('_sp_eff','') or _rec.get('_sp','') or _rec.get('Subdistrict','') or _rec.get('ตำบล',''))
                                            elif _ikey == '__DIS_GEN__':
                                                _val = str(_rec.get('_sd_eff','') or _rec.get('_sd','') or _rec.get('District','') or _rec.get('อำเภอ',''))
                                            elif _ikey == '__PROV_GEN__':
                                                _val = str(_rec.get('_sv_eff','') or _rec.get('_sv','') or _rec.get('Province','') or _rec.get('จังหวัด',''))
                                            elif _ikey == '__PROV_BLANK__':
                                                _val = ''   # จังหวัดจากต้นฉบับ — ว่างไว้ให้กรอกเอง
                                            elif _ikey == '__LOAD_DATE__':
                                                _val = trip_load_date.get(_tnum, '')  # ทุกแถว
                                            elif _ikey == '__LOAD_TIME__':
                                                _val = trip_load_time.get(_tnum, '')  # ทุกแถว
                                            elif _ikey == '__DOOR__':
                                                _val = trip_door.get(_tnum, '')  # ทุกแถว
                                            elif _ikey == '__REMARK__':
                                                _val = _trip_remark_text.get(_tnum, '') if _first_row_of_trip else ''
                                            else:
                                                # คอลัมน์จากต้นฉบับ — ใช้ค่าตรงๆ จากไฟล์
                                                _val = _rec.get(_ikey, '')
                                                if _val is None or (isinstance(_val, float) and _val != _val):
                                                    _val = ''
                                            _ws_xl.write(_row_xl, _ci, _val, _dfmt)
                                        _ws_xl.set_row(_row_xl, _orh)
                                        _row_xl += 1
                                        _row_seq += 1
                                        _first_row_of_trip = False

                                    # ── จุดสุดท้าย: DC return row (ท้ายทริป) ──
                                    for _ci, (_, _, _ikey) in enumerate(_col_plan):
                                        if _ikey == '__SEP__':
                                            _dcval = _row_seq
                                        elif _ikey == 'BU':
                                            _dcval = 'PROJECT'
                                        elif _ikey == 'Code':
                                            _dcval = 'DC011'
                                        elif _ikey == 'WMSCode':
                                            _dcval = 'DC011'
                                        elif _ikey == 'Name':
                                            _dcval = 'บ.พีทีจี เอ็นเนอยี จำกัด (มหาชน) (DCวังน้อย)'
                                        elif _ikey in ('Cube', 'Weight'):
                                            _dcval = 0.0
                                        elif _ikey == 'OriginalQty':
                                            _dcval = 0
                                        elif _ikey == '__LOAD_TIME__':
                                            _dcval = trip_load_time.get(_tnum, '')
                                        elif _ikey == '__DOOR__':
                                            _dcval = trip_door.get(_tnum, '')
                                        elif _ikey == '__TRIP__':
                                            if _tn_col:
                                                _xl_r2   = _row_xl + 1
                                                _tn_r2   = f'${_tn_col}{_xl_r2}'
                                                _tn_rng2 = f'${_tn_col}$3:${_tn_col}{_xl_r2}'
                                                _sp2 = f'IFERROR(SUMPRODUCT((1/COUNTIF({_tn_rng2},{_tn_rng2}))*({_tn_rng2}<>"")),1)'
                                                _ws_xl.write_formula(_row_xl, _ci,
                                                    f'=IF({_tn_r2}=""," ",TEXT({_sp2},"000"))',
                                                    _tf, str(_tnum_int))
                                                continue
                                            _dcval = _tnum_int
                                        elif _ikey == '__TRIPNO__':
                                            _dcval = _tno_str
                                        elif _ikey == '__LOAD_DATE__':
                                            _dcval = trip_load_date.get(_tnum, '')
                                        else:
                                            _dcval = ''
                                        _is_dc_num = _ikey in ('Cube', 'Weight')
                                        _ws_xl.write(_row_xl, _ci, _dcval, _nf if _is_dc_num else _tf)
                                    _ws_xl.set_row(_row_xl, _orh)
                                    _row_xl += 1
                                    _row_seq += 1
                                    use_yellow = not use_yellow

                                # Conditional format: auto highlight + font color (Excel 2016 compatible)
                                if _tn_col and _wt_col and _cb_col and _row_xl > 2:
                                    _cf_last_r = _row_xl - 1
                                    _cf_ncols  = len(_col_plan)
                                    _anc3         = f'${_tn_col}3'
                                    _anc_exp_cur  = f'${_tn_col}$3:${_tn_col}3'
                                    _anc_exp_prev = f'${_tn_col}$2:${_tn_col}2'
                                    _anc_rng      = f'${_tn_col}$3:${_tn_col}$9999'
                                    _wt_rng_cf    = f'${_wt_col}$3:${_wt_col}$9999'
                                    _cb_rng_cf    = f'${_cb_col}$3:${_cb_col}$9999'
                                    _vw3 = f'IF(LEFT({_anc3},2)="6W",5500,IF(LEFT({_anc3},2)="JB",3000,2000))'
                                    _vc3 = f'IF(LEFT({_anc3},2)="6W",25,IF(LEFT({_anc3},2)="JB",10,7))'
                                    # นับจำนวน transition (แถวปัจจุบัน≠แถวบน) — ไม่มี division → Excel 2016 safe
                                    _odd_grp = (
                                        f'ISODD(SUMPRODUCT(({_anc_exp_cur}<>{_anc_exp_prev})'
                                        f'*({_anc_exp_cur}<>"")))'
                                    )
                                    _util_lt = (
                                        f'MAX(SUMIF({_anc_rng},{_anc3},{_wt_rng_cf})/{_vw3},'
                                        f'SUMIF({_anc_rng},{_anc3},{_cb_rng_cf})/{_vc3})<0.98'
                                    )
                                    _util_ge = (
                                        f'MAX(SUMIF({_anc_rng},{_anc3},{_wt_rng_cf})/{_vw3},'
                                        f'SUMIF({_anc_rng},{_anc3},{_cb_rng_cf})/{_vc3})>=0.98'
                                    )
                                    # 4 rules: (odd/even trip group) × (util ต่ำ/สูง)
                                    _cf_rules = [
                                        (f'=AND({_anc3}<>"",{_odd_grp},{_util_lt})',
                                         _wb_xl.add_format(_f({'bg_color':'#FFE699','border':1,'font_color':'#CC0000'}))),
                                        (f'=AND({_anc3}<>"",NOT({_odd_grp}),{_util_lt})',
                                         _wb_xl.add_format(_f({'bg_color':'#FFFFFF','border':1,'font_color':'#CC0000'}))),
                                        (f'=AND({_anc3}<>"",{_odd_grp},{_util_ge})',
                                         _wb_xl.add_format(_f({'bg_color':'#FFE699','border':1,'font_color':'#000000'}))),
                                        (f'=AND({_anc3}<>"",NOT({_odd_grp}),{_util_ge})',
                                         _wb_xl.add_format(_f({'bg_color':'#FFFFFF','border':1,'font_color':'#000000'}))),
                                    ]
                                    for _cf_fml, _cf_fmt in _cf_rules:
                                        _ws_xl.conditional_format(2, 0, _cf_last_r, _cf_ncols - 1,
                                            {'type': 'formula', 'criteria': _cf_fml, 'format': _cf_fmt})
                                _wb_xl.close()
                                _output.seek(0)

                            except Exception as _xe:
                                st.warning(f"⚠️ xlsxwriter error: {_xe} — fallback to basic")
                                _output = io.BytesIO()
                                with pd.ExcelWriter(_output, engine='xlsxwriter') as _writer:
                                    _exp = _rd.drop(columns=[c for c in ['_key_u','_sp','_sd','_sv','_rt','_trip_order','_sp_eff','_sd_eff','_sv_eff','_rorder'] if c in _rd.columns], errors='ignore').copy()
                                    _exp['Trip_No'] = _exp['Trip'].map(lambda x: trip_no_map.get(x, ''))
                                    _exp.to_excel(_writer, sheet_name='รายละเอียดทริป', index=False)
                                    summary.to_excel(_writer, sheet_name='สรุปทริป', index=False)

                            st.session_state['_excel_bytes']  = _output.getvalue()
                            st.session_state['_excel_key']    = _xl_key
                            st.session_state['_trip_no_map']  = trip_no_map

                    # trip_no_map ต้องพร้อมสำหรับแผนที่ด้านล่าง
                    trip_no_map = st.session_state.get('_trip_no_map', {})

                    _btn_col1, _btn_col2 = st.columns([1, 3])
                    with _btn_col1:
                        if st.button("🔢 รันเลขทริปใหม่",
                                     help="สร้าง Excel ใหม่ รันเลขทริป 001,002,... ตามลำดับปัจจุบัน",
                                     use_container_width=True):
                            st.session_state.pop('_excel_key', None)
                            st.rerun()
                    with _btn_col2:
                        st.download_button(
                            label="📥 ดาวน์โหลดผลลัพธ์ (Excel)",
                            data=st.session_state.get('_excel_bytes', b''),
                            file_name=f"ผลจัดทริป_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True
                        )

                    st.markdown("---")
                    
                    # 🗺️ แผนที่เส้นทาง (Interactive - Leaflet.js)
                    with st.expander("🗺️ แผนที่เส้นทาง (Interactive)", expanded=True):
                        try:
                            import importlib as _imp, trip_map_interactive as _tmi
                            import time as _time_mod
                            import streamlit.components.v1 as _cmp2
                            import hashlib as _hl
                            _build_imap = _tmi.build_interactive_map_html

                            _imap_sig = f"v30|{len(assigned_df)}|{safe_int_trip(assigned_df['Trip'].max())}|{sorted(assigned_df['Trip'].unique().tolist())}"
                            _imap_key = _hl.md5(_imap_sig.encode()).hexdigest()[:12]

                            if st.session_state.get('_imap_key') != _imap_key:
                                with st.spinner("🗺️ กำลังสร้างแผนที่..."):
                                    _t_map = _time_mod.time()
                                    _imap_html = _build_imap(
                                        result_df=assigned_df,
                                        summary_df=summary,
                                        limits=LIMITS,
                                        punthai_limits=PUNTHAI_LIMITS,
                                        trip_no_map=trip_no_map,
                                        dc_lat=14.1459, dc_lon=100.6873,
                                        route_cache=ROUTE_CACHE_DATA,
                                    )
                                    st.session_state['_imap_html'] = _imap_html
                                    st.session_state['_imap_key'] = _imap_key
                                    st.session_state['_imap_build_time'] = _time_mod.time() - _t_map

                            _htm = st.session_state.get('_imap_html', '')
                            if not _htm:
                                st.warning("⚠️ ยังไม่มีข้อมูลแผนที่ กดจัดเที่ยวก่อน")
                            else:
                                # Sanitize surrogates
                                try:
                                    _htm.encode('utf-8')
                                except UnicodeEncodeError:
                                    _htm = _htm.encode('utf-8', errors='replace').decode('utf-8')
                                    st.session_state['_imap_html'] = _htm
                                import re as _re
                                _nb = len(_re.findall(r'"code":', _htm))
                                _build_t = st.session_state.get('_imap_build_time', 0)
                                st.caption(f"🗺️ HTML: {len(_htm)//1024} KB · {_nb} สาขา · build: {_build_t:.1f}s")
                                _cmp2.html(_htm, height=860, scrolling=False)
                        except Exception as _e:
                            import traceback as _tb
                            st.error(f"❌ Interactive map error: {_e}")
                            st.code(_tb.format_exc(), language='text')
                            st.info(f"📋 columns: {list(assigned_df.columns)} | rows: {len(assigned_df)} | trips: {sorted(assigned_df['Trip'].unique().tolist())}")

                    # ── FOLIUM FALLBACK (ใช้เมื่อ interactive map error) ──
                    if 'FOLIUM_AVAILABLE' in dir() and FOLIUM_AVAILABLE and locals().get('_FOLIUM_FALLBACK_', False):
                        with st.expander("🗺️ แผนที่เส้นทางแต่ละทริป (Fallback)", expanded=True):
                            # ตัวกรอง
                            col_filter1, col_filter2, col_filter3 = st.columns([1, 1, 1])
                            
                            with col_filter1:
                                # กรองตามเลขทริป - เรียงจากไกลมาใกล้
                                trip_distances = {}
                                for t in assigned_df['Trip'].unique():
                                    if t > 0 and '_distance_from_dc' in assigned_df.columns:
                                        max_dist = assigned_df[assigned_df['Trip'] == t]['_distance_from_dc'].max()
                                        trip_distances[t] = max_dist if pd.notna(max_dist) else 0
                                sorted_trips = sorted(trip_distances.keys(), key=lambda x: trip_distances.get(x, 0), reverse=True)
                                trip_options = ['ทั้งหมด'] + [f"Trip {t} ({trip_distances.get(t, 0):.0f}km)" for t in sorted_trips]
                                selected_trip = st.selectbox("🚚 เลือกทริป (ไกล→ใกล้)", trip_options, key="map_trip_filter")
                            
                            with col_filter2:
                                # กรองตามประเภทรถ
                                truck_types = ['ทั้งหมด']
                                if 'Truck' in assigned_df.columns:
                                    unique_trucks = assigned_df['Truck'].dropna().unique()
                                    truck_types.extend(sorted(set([t.split()[0] for t in unique_trucks if t])))
                                selected_truck = st.selectbox("🚛 ประเภทรถ", truck_types, key="map_truck_filter")
                            
                            with col_filter3:
                                # เลือกแสดงเส้นทาง
                                show_route = st.checkbox("🛣️ แสดงเส้นทาง", value=True, key="map_show_route")
                            
                            # กรองข้อมูล
                            map_df = assigned_df.copy()
                            if selected_trip != 'ทั้งหมด':
                                trip_num = int(selected_trip.split()[1])
                                map_df = map_df[map_df['Trip'] == trip_num]
                            if selected_truck != 'ทั้งหมด':
                                map_df = map_df[map_df['Truck'].str.startswith(selected_truck, na=False)]
                            
                            if len(map_df) == 0:
                                st.warning("⚠️ ไม่มีข้อมูลตามเงื่อนไขที่เลือก")
                            else:
                                # ตรวจสอบว่ามีพิกัด
                                if '_lat' in map_df.columns and '_lon' in map_df.columns:
                                    valid_coords = map_df[(map_df['_lat'] > 0) & (map_df['_lon'] > 0)]
                                    
                                    if len(valid_coords) == 0:
                                        st.warning("⚠️ ไม่มีข้อมูลพิกัดสำหรับแสดงแผนที่")
                                    else:
                                        # สร้างแผนที่พร้อม progress
                                        # Map cache - ตรวจสอบว่ามีแผนที่ cached หรือไม่
                                        map_cache_key = f'map|{selected_trip}|{selected_truck}|{show_route}|{len(valid_coords)}'
                                        _map_is_cached = (st.session_state.get('_map_cache_key') == map_cache_key and '_map_html' in st.session_state)
                                        if not _map_is_cached:
                                            with st.spinner("🗺️ กำลังสร้างแผนที่..."):
                                                # DC Wang Noi coordinates
                                                DC_LAT, DC_LON = 14.1459, 100.6873
                                            
                                                # หาจุดกึ่งกลาง
                                                center_lat = valid_coords['_lat'].mean()
                                                center_lon = valid_coords['_lon'].mean()
                                            
                                                # สร้างแผนที่
                                                m = folium.Map(
                                                    location=[center_lat, center_lon],
                                                    zoom_start=8,
                                                    tiles='OpenStreetMap',
                                                    prefer_canvas=True  # เร็วขึ้น
                                                )
                                            
                                                # เพิ่มปุ่ม Fullscreen
                                                plugins.Fullscreen(
                                                    position='topleft',
                                                    title='เต็มจอ',
                                                    title_cancel='ออกจากโหมดเต็มจอ',
                                                    force_separate_button=True
                                                ).add_to(m)
                                            
                                                # เพิ่ม DC Marker
                                                folium.Marker(
                                                    location=[DC_LAT, DC_LON],
                                                    popup="<b>🏭 DC Wang Noi</b>",
                                                    tooltip="DC Wang Noi",
                                                    icon=folium.Icon(color='black', icon='home', prefix='fa')
                                                ).add_to(m)
                                            
                                                # สี palette สำหรับแต่ละทริป - 50 สีไม่ซ้ำกัน
                                                colors = [
                                                    '#e41a1c', '#377eb8', '#4daf4a', '#984ea3', '#ff7f00',  # 1-5
                                                    '#a65628', '#f781bf', '#1b9e77', '#d95f02', '#7570b3',  # 6-10
                                                    '#e7298a', '#66a61e', '#e6ab02', '#a6761d', '#666666',  # 11-15
                                                    '#1f78b4', '#33a02c', '#fb9a99', '#fdbf6f', '#cab2d6',  # 16-20
                                                    '#b15928', '#8dd3c7', '#ffffb3', '#bebada', '#fb8072',  # 21-25
                                                    '#80b1d3', '#fdb462', '#b3de69', '#fccde5', '#d9d9d9',  # 26-30
                                                    '#bc80bd', '#ccebc5', '#ffed6f', '#e31a1c', '#1b7837',  # 31-35
                                                    '#762a83', '#e66101', '#5e3c99', '#d53e4f', '#3288bd',  # 36-40
                                                    '#f46d43', '#fdae61', '#fee08b', '#66c2a5', '#3d9970',  # 41-45
                                                    '#001f3f', '#39cccc', '#85144b', '#ff4136', '#2ecc40'   # 46-50
                                                ]
                                            
                                                # เรียงทริปตามเลข (Trip 1, 2, 3...) เพราะ renumber แล้ว
                                                trip_max_dist = {}
                                                for trip_id in valid_coords['Trip'].unique():
                                                    if '_distance_from_dc' in valid_coords.columns:
                                                        max_d = valid_coords[valid_coords['Trip'] == trip_id]['_distance_from_dc'].max()
                                                        trip_max_dist[trip_id] = max_d if pd.notna(max_d) else 0
                                                    else:
                                                        trip_max_dist[trip_id] = 0
                                                sorted_trip_ids = sorted(trip_max_dist.keys(), key=lambda x: trip_max_dist[x], reverse=True)
                                            
                                                # ฟังก์ชันเรียงสาขาแบบ Nearest Neighbor (ป้องกันกระโดด)
                                                def optimize_branch_order(trip_df, dc_lat, dc_lon):
                                                    """เรียงสาขาให้ต่อเนื่องกัน: DC → สาขาไกลสุด → nearest → nearest → ... → DC"""
                                                    if len(trip_df) <= 1:
                                                        return trip_df
                                                
                                                    df = trip_df.copy()
                                                    # เริ่มจากสาขาไกลสุด
                                                    ordered = []
                                                    remaining = df.to_dict('records')
                                                
                                                    # หาสาขาไกลสุดเป็นจุดเริ่มต้น
                                                    farthest_idx = max(range(len(remaining)), key=lambda i: remaining[i]['_distance_from_dc'])
                                                    current = remaining.pop(farthest_idx)
                                                    ordered.append(current)
                                                
                                                    # Nearest neighbor: หาสาขาใกล้สุดกับสาขาปัจจุบัน
                                                    while remaining:
                                                        current_lat, current_lon = current['_lat'], current['_lon']
                                                        nearest_idx = 0
                                                        nearest_dist = float('inf')
                                                    
                                                        for i, branch in enumerate(remaining):
                                                            # ใช้ cache distance
                                                            dist_key = f"{current_lat:.4f},{current_lon:.4f}_{branch['_lat']:.4f},{branch['_lon']:.4f}"
                                                            dist_key_rev = f"{branch['_lat']:.4f},{branch['_lon']:.4f}_{current_lat:.4f},{current_lon:.4f}"
                                                        
                                                            if dist_key in DISTANCE_CACHE:
                                                                dist = DISTANCE_CACHE[dist_key]
                                                            elif dist_key_rev in DISTANCE_CACHE:
                                                                dist = DISTANCE_CACHE[dist_key_rev]
                                                            else:
                                                                # คำนวณ haversine
                                                                dlat = radians(branch['_lat'] - current_lat)
                                                                dlon = radians(branch['_lon'] - current_lon)
                                                                a = sin(dlat/2)**2 + cos(radians(current_lat)) * cos(radians(branch['_lat'])) * sin(dlon/2)**2
                                                                c = 2 * atan2(sqrt(a), sqrt(1-a))
                                                                dist = 6371 * c
                                                        
                                                            if dist < nearest_dist:
                                                                nearest_dist = dist
                                                                nearest_idx = i
                                                    
                                                        current = remaining.pop(nearest_idx)
                                                        ordered.append(current)
                                                
                                                    return pd.DataFrame(ordered)
                                            
                                                # สร้าง Feature Groups สำหรับ Layer Control
                                                trip_groups = {}
                                            
                                                # วนลูปแต่ละทริป
                                                for idx, trip_id in enumerate(sorted_trip_ids):
                                                    trip_data = valid_coords[valid_coords['Trip'] == trip_id].copy()
                                                    # เรียงสาขาแบบ Nearest Neighbor (ป้องกันกระโดด)
                                                    trip_data = optimize_branch_order(trip_data, DC_LAT, DC_LON)
                                                
                                                    trip_color = colors[idx % len(colors)]
                                                    max_dist = trip_max_dist.get(trip_id, 0)
                                                
                                                    # ดึงชื่อรถจาก summary
                                                    truck_info = summary[summary['Trip'] == trip_id]['Truck'].iloc[0] if trip_id in summary['Trip'].values else 'N/A'
                                                
                                                    # สร้าง Feature Group สำหรับทริปนี้
                                                    fg = folium.FeatureGroup(name=f"Trip {trip_id} ({max_dist:.0f}km) - {truck_info}")
                                                    trip_groups[trip_id] = fg
                                                
                                                    # เก็บพิกัดสาขา
                                                    points = []
                                                    point_names = []
                                                    point_distances = []
                                                
                                                    for _, row in trip_data.iterrows():
                                                        if row['_lat'] > 0 and row['_lon'] > 0:
                                                            points.append([row['_lat'], row['_lon']])
                                                            point_names.append(f"{row.get('Name', row.get('Code', 'Unknown'))}")
                                                            point_distances.append(row.get('_distance_from_dc', 0))
                                                
                                                    if len(points) == 0:
                                                        continue
                                                
                                                    # 🛣️ ดึงเส้นทางจริงจาก OSRM (DC → สาขา1 → สาขา2 → ... → DC)
                                                    waypoints = [[DC_LAT, DC_LON]] + points + [[DC_LAT, DC_LON]]

                                                    # ใช้ ROUTE_CACHE_DATA (global file-backed) โดยตรงผ่าน get_multi_point_route_osrm
                                                    real_route_coords, total_trip_distance = get_multi_point_route_osrm(waypoints)

                                                    # fallback: ถ้า OSRM ล้มเหลว (coords = waypoints เหมือนกัน) → คำนวณ distance จาก cache
                                                    if total_trip_distance == 0 or real_route_coords == waypoints:
                                                        _fb_dist = 0
                                                        for _wi in range(len(waypoints) - 1):
                                                            lat1, lon1 = waypoints[_wi]
                                                            lat2, lon2 = waypoints[_wi + 1]
                                                            _fb_dist += haversine_distance(lat1, lon1, lat2, lon2, use_osrm_cache=False)
                                                        total_trip_distance = _fb_dist
                                                        # เส้นเชื่อม DC → waypoints แบบ straight-segment (fallback)
                                                        real_route_coords = waypoints

                                                    # 🏭 หมุด DC ออกเดิน (ลำดับ 0)
                                                    _dc_start_label = f'<div style="background-color:{trip_color};color:#fff;border-radius:12px;min-width:50px;height:24px;text-align:center;line-height:24px;font-weight:bold;font-size:10px;border:2px solid #000;box-shadow:2px 2px 6px rgba(0,0,0,0.5);padding:0 4px;">T{trip_id}(0)</div>'
                                                    folium.Marker(
                                                        location=[DC_LAT, DC_LON],
                                                        popup=folium.Popup(f"<b>🏭 DC ออกเดิน — Trip {trip_id}</b><br>ลำดับ: 0<br>ระยะทางรวม: {total_trip_distance:.1f} km", max_width=250),
                                                        tooltip=f"Trip {trip_id} - 0. DC (ออกเดิน)",
                                                        icon=folium.DivIcon(html=_dc_start_label)
                                                    ).add_to(fg)
                                                
                                                    # ปักหมุดแต่ละจุด
                                                    for i, (point, name, dist) in enumerate(zip(points, point_names, point_distances)):
                                                        # 🎯 แสดง T{trip}({ลำดับ}) บนหมุด เช่น T1(1), T1(2)
                                                        trip_label = f'<div style="background-color:{trip_color};color:#fff;border-radius:12px;min-width:50px;height:24px;text-align:center;line-height:24px;font-weight:bold;font-size:10px;border:2px solid #000;box-shadow:2px 2px 6px rgba(0,0,0,0.5);padding:0 4px;">T{trip_id}({i+1})</div>'
                                                    
                                                        popup_html = f"""
                                                        <div style="font-family:Arial;min-width:200px;">
                                                            <h4 style="margin:0;color:{trip_color};">🚚 Trip {trip_id}</h4>
                                                            <hr style="margin:5px 0;">
                                                            <b>ลำดับ:</b> {i+1}/{len(points)}<br>
                                                            <b>สาขา:</b> {name}<br>
                                                            <b>ห่างจาก DC:</b> {dist:.1f} km<br>
                                                            <b>รถ:</b> {truck_info}<br>
                                                            <hr style="margin:5px 0;">
                                                            <b>📏 ระยะทางรวมทริป:</b> {total_trip_distance:.1f} km<br>
                                                            <b>📍 จำนวนจุด:</b> {len(points)} สาขา
                                                        </div>
                                                        """
                                                    
                                                        folium.Marker(
                                                            location=point,
                                                            popup=folium.Popup(popup_html, max_width=300),
                                                            tooltip=f"Trip {trip_id} - {i+1}. {name} ({dist:.1f}km)",
                                                            icon=folium.DivIcon(html=trip_label)
                                                        ).add_to(fg)

                                                    # 🏭 หมุด DC กลับ (ลำดับสุดท้าย = len(points)+1)
                                                    _dc_return_label = f'<div style="background-color:{trip_color};color:#fff;border-radius:12px;min-width:50px;height:24px;text-align:center;line-height:24px;font-weight:bold;font-size:10px;border:2px solid #000;box-shadow:2px 2px 6px rgba(0,0,0,0.5);padding:0 4px;">T{trip_id}({len(points)+1})</div>'
                                                    folium.Marker(
                                                        location=[DC_LAT, DC_LON],
                                                        popup=folium.Popup(f"<b>🏭 DC กลับ — Trip {trip_id}</b><br>ลำดับ: {len(points)+1}<br>ระยะทางรวม: {total_trip_distance:.1f} km", max_width=250),
                                                        tooltip=f"Trip {trip_id} - {len(points)+1}. DC (กลับ)",
                                                        icon=folium.DivIcon(html=_dc_return_label)
                                                    ).add_to(fg)
                                                
                                                    # วาดเส้นทางจริง DC → สาขา → DC (ถ้าเปิด)
                                                    if show_route and len(points) >= 1:
                                                        # ใช้เส้นทางจริงจาก OSRM
                                                        folium.PolyLine(
                                                            locations=real_route_coords,
                                                            weight=4,
                                                            color=trip_color,
                                                            opacity=0.8,
                                                            popup=f"Trip {trip_id}: {total_trip_distance:.1f} km (เส้นทางจริง)",
                                                            tooltip=f"🛣️ Trip {trip_id} - ระยะทาง {total_trip_distance:.1f} km"
                                                        ).add_to(fg)
                                                
                                                    fg.add_to(m)
                                            
                                                # เพิ่ม Layer Control สำหรับเปิด/ปิดแต่ละทริป
                                                folium.LayerControl(collapsed=False).add_to(m)
                                        
                                            # แสดงแผนที่
                                            folium_static(m, width=1200, height=700)

                                            # บันทึก Map Cache
                                            st.session_state['_map_cache_key'] = map_cache_key
                                            st.session_state['_map_html'] = m._repr_html_()

                                        # แสดงแผนที่จาก cache เมื่อ cached
                                        if _map_is_cached and '_map_html' in st.session_state:
                                            import streamlit.components.v1 as _cmp
                                            _cmp.html(st.session_state['_map_html'], height=720, scrolling=False)
                                        
                                        # สรุประยะทางแต่ละทริป - แสดงข้อมูลละเอียด
                                        st.markdown("#### 📏 ระยะทางแต่ละทริป (เรียงจากไกล→ใกล้)")
                                        
                                        # สร้าง DataFrame สำหรับแสดงตาราง
                                        trip_details = []
                                        for trip_id in sorted_trip_ids:
                                            trip_data = valid_coords[valid_coords['Trip'] == trip_id].copy()
                                            if len(trip_data) == 0:
                                                continue
                                            
                                            # เรียงตามระยะทางจาก DC (ไกล → ใกล้)
                                            trip_data = trip_data.sort_values('_distance_from_dc', ascending=False).reset_index(drop=True)
                                            
                                            # ดึงข้อมูลรถจาก summary
                                            truck_info = summary[summary['Trip'] == trip_id]['Truck'].iloc[0] if trip_id in summary['Trip'].values else 'N/A'
                                            truck_type = truck_info.split()[0] if truck_info else 'N/A'
                                            
                                            # คำนวณน้ำหนักและคิวรวม
                                            total_weight = trip_data['Weight'].sum()
                                            total_cube = trip_data['Cube'].sum()
                                            
                                            # สาขาไกลสุดจาก DC
                                            max_dist_from_dc = trip_data['_distance_from_dc'].max()
                                            
                                            # คำนวณระยะทางรวม (DC → สาขา1 → สาขา2 → ... → สาขาสุดท้าย) - ไม่รวมกลับ DC
                                            points = []
                                            for _, row in trip_data.iterrows():
                                                if row['_lat'] > 0 and row['_lon'] > 0:
                                                    points.append([row['_lat'], row['_lon']])
                                            
                                            route_distance = 0  # DC → สาขา1 → ... → สาขาสุดท้าย
                                            inter_branch_distance = 0  # สาขา → สาขา (ไม่รวม DC)
                                            
                                            if len(points) > 0:
                                                # ใช้ OSRM multi-point route เพื่อระยะทางถนนจริง
                                                _wp = [[DC_LAT, DC_LON]] + points
                                                _cache_k = "|".join([f"{la:.4f},{lo:.4f}" for la, lo in _wp])
                                                if USE_CACHE and _cache_k in ROUTE_CACHE_DATA:
                                                    _rc = ROUTE_CACHE_DATA[_cache_k]
                                                    route_distance = _rc.get('distance', 0)
                                                else:
                                                    _, route_distance = get_multi_point_route_osrm(_wp)
                                                # inter_branch = OSRM ไม่รวม DC leg แรก → ประมาณจาก DISTANCE_CACHE
                                                for j in range(len(points) - 1):
                                                    seg = haversine_distance(points[j][0], points[j][1], points[j+1][0], points[j+1][1])
                                                    if seg < 9999:
                                                        inter_branch_distance += seg
                                            
                                            trip_details.append({
                                                'ทริป': trip_id,
                                                'รถ': truck_type,
                                                'สาขา': len(trip_data),
                                                'น้ำหนัก (kg)': f"{total_weight:,.0f}",
                                                'คิว (m³)': f"{total_cube:.1f}",
                                                'ไกลสุดจาก DC': f"{max_dist_from_dc:.1f} km",
                                                'ระยะทางรวม': f"{route_distance:.1f} km",
                                                'ระหว่างสาขา': f"{inter_branch_distance:.1f} km"
                                            })
                                        
                                        # แสดงตาราง (หลังวนครบทุก trip แล้ว)
                                        if trip_details:
                                            trip_df = pd.DataFrame(trip_details)
                                            st.dataframe(
                                                trip_df,
                                                width="stretch",
                                                hide_index=True,
                                                column_config={
                                                    'ทริป': st.column_config.NumberColumn('🚚 ทริป', width='small'),
                                                    'รถ': st.column_config.TextColumn('🚛 รถ', width='small'),
                                                    'สาขา': st.column_config.NumberColumn('📍 สาขา', width='small'),
                                                    'น้ำหนัก (kg)': st.column_config.TextColumn('⚖️ น้ำหนัก', width='small'),
                                                    'คิว (m³)': st.column_config.TextColumn('📦 คิว', width='small'),
                                                    'ไกลสุดจาก DC': st.column_config.TextColumn('🎯 ไกลสุด', width='small'),
                                                    'ระยะทางรวม': st.column_config.TextColumn('📏 รวม (DC→สาขา)', width='medium'),
                                                    'ระหว่างสาขา': st.column_config.TextColumn('↔️ ระหว่างสาขา', width='small')
                                                }
                                            )
                                            
                                            # สรุปรวม
                                            total_route = sum(float(d['ระยะทางรวม'].replace(' km', '').replace(',', '')) for d in trip_details)
                                            total_inter = sum(float(d['ระหว่างสาขา'].replace(' km', '').replace(',', '')) for d in trip_details)
                                            st.caption(f"📊 **รวมทั้งหมด:** {len(trip_details)} ทริป | ระยะทางรวม: {total_route:,.1f} km | ระหว่างสาขา: {total_inter:,.1f} km")
                                        
                                        st.caption(f"📍 แสดง {len(valid_coords)} สาขาใน {len(sorted_trip_ids)} ทริป | 💡 คลิกปุ่มมุมซ้ายบนเพื่อเต็มจอ | ใช้ Layer Control ด้านขวาเพื่อเปิด/ปิดทริป")
                                else:
                                    st.warning("⚠️ ไม่มีข้อมูลพิกัดในผลลัพธ์ (ต้องมีคอลัมน์ _lat และ _lon)")
                    
            # ==========================================
            # แท็บ 2: จัดกลุ่มสาขาตามภาค (ไม่สนน้ำหนัก)
            # ==========================================
            with tab2:
                df_region = df.copy()
                
                # จัดกลุ่มตามภาค
                branch_info = model_data.get('branch_info', {})
                trip_pairs = model_data.get('trip_pairs', set())
                
                # สร้างข้อมูลภาคสำหรับแต่ละสาขา (จากไฟล์ประวัติ)
                region_groups = {
                    'ภาคกลาง-กรุงเทพชั้นใน': ['กรุงเทพมหานคร'],
                    'ภาคกลาง-กรุงเทพชั้นกลาง': ['กรุงเทพมหานคร'],
                    'ภาคกลาง-กรุงเทพชั้นนอก': ['กรุงเทพมหานคร'],
                    'ภาคกลาง-ปริมณฑล': ['นครปฐม', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ', 'สมุทรสาคร'],
                    'ภาคกลาง-กลางตอนบน': ['ชัยนาท', 'พระนครศรีอยุธยา', 'ลพบุรี', 'สระบุรี', 'สิงห์บุรี', 'อ่างทอง', 'อยุธยา'],
                    'ภาคกลาง-กลางตอนล่าง': ['สมุทรสงคราม', 'สุพรรณบุรี'],
                    'ภาคตะวันตก': ['กาญจนบุรี', 'ประจวบคีรีขันธ์', 'ราชบุรี', 'เพชรบุรี'],
                    'ภาคตะวันออก': ['จันทบุรี', 'ชลบุรี', 'ตราด', 'นครนายก', 'ปราจีนบุรี', 'ระยอง', 'สระแก้ว', 'ฉะเชิงเทรา'],
                    'ภาคอีสาน-อีสานเหนือ': ['นครพนม', 'บึงกาฬ', 'มุกดาหาร', 'สกลนคร', 'หนองคาย', 'หนองบัวลำภู', 'อุดรธานี', 'เลย'],
                    'ภาคอีสาน-อีสานกลาง': ['กาฬสินธุ์', 'ขอนแก่น', 'ชัยภูมิ', 'มหาสารคาม', 'ร้อยเอ็ด'],
                    'ภาคอีสาน-อีสานใต้': ['นครราชสีมา', 'โคราช', 'บุรีรัมย์', 'ยโสธร', 'ศรีสะเกษ', 'สุรินทร์', 'อำนาจเจริญ', 'อุบลราชธานี'],
                    'ภาคเหนือ-เหนือตอนบน': ['น่าน', 'พะเยา', 'ลำปาง', 'ลำพูน', 'เชียงราย', 'เชียงใหม่', 'แพร่', 'แม่ฮ่องสอน'],
                    'ภาคเหนือ-เหนือตอนล่าง': ['กำแพงเพชร', 'ตาก', 'นครสวรรค์', 'พิจิตร', 'พิษณุโลก', 'สุโขทัย', 'อุตรดิตถ์', 'อุทัยธานี', 'เพชรบูรณ์'],
                    'ภาคใต้-ใต้ฝั่งอันดามัน': ['กระบี่', 'ตรัง', 'พังงา', 'ภูเก็ต', 'ระนอง', 'สตูล'],
                    'ภาคใต้-ใต้ฝั่งอ่าวไทย': ['ชุมพร', 'นครศรีธรรมราช', 'พัทลุง', 'ยะลา', 'สงขลา', 'สุราษฎร์ธานี', 'ปัตตานี', 'นราธิวาส']
                }
                
                def get_region(province):
                    if pd.isna(province) or not province or str(province).strip() in ['', 'nan', 'UNKNOWN']:
                        return 'ไม่ระบุ'
                    
                    # 🚨 Override: ฉะเชิงเทรา → ภาคตะวันออก (ไม่ใช่ปริมณฑล)
                    if 'ฉะเชิงเทรา' in str(province):
                        return 'ภาคตะวันออก'
                    
                    for region, provinces in region_groups.items():
                        if any(p in str(province) for p in provinces):
                            return region
                    return 'อื่นๆ'
                
                # เพิ่มคอลัมน์ภาค - ดึงจังหวัดจาก Master ถ้าไม่มี
                # รองรับทั้งชื่อคอลัมน์ภาษาอังกฤษ (Province) และไทย (จังหวัด)
                province_col = None
                if 'Province' in df_region.columns:
                    province_col = 'Province'
                elif 'จังหวัด' in df_region.columns:
                    province_col = 'จังหวัด'
                
                # ถ้าไม่มีคอลัมน์จังหวัดเลย หรือมีแต่เป็น NaN ทั้งหมด → ดึงจาก MASTER_DATA
                # Vectorized: สร้าง province_map ครั้งเดียวแทน iterrows
                _prov_map = {}
                if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns and 'จังหวัด' in MASTER_DATA.columns:
                    _pm = MASTER_DATA[['Plan Code', 'จังหวัด']].dropna(subset=['Plan Code'])
                    _prov_map = dict(zip(_pm['Plan Code'].astype(str).str.strip(), _pm['จังหวัด'].fillna('')))
                
                if not province_col or df_region[province_col].isna().all():
                    df_region['จังหวัด'] = df_region['Code'].astype(str).map(_prov_map).fillna('UNKNOWN')
                    province_col = 'จังหวัด'
                elif province_col and df_region[province_col].isna().any():
                    # เติมเฉพาะที่เป็น NaN
                    _missing = df_region[province_col].isna()
                    df_region.loc[_missing, province_col] = df_region.loc[_missing, 'Code'].astype(str).map(_prov_map).fillna('UNKNOWN')
                
                # ตรวจสอบอีกครั้งว่ามีคอลัมน์จังหวัดแล้ว
                if not province_col or province_col not in df_region.columns:
                    st.error("❌ ไม่พบข้อมูลจังหวัด กรุณาตรวจสอบไฟล์ข้อมูล")
                    return
                
                df_region['Region'] = df_region[province_col].apply(get_region)
                
                # หากลุ่มสาขา (ใช้ Booking No. เป็นหลัก)
                def find_paired_branches(code, code_province, df_data):
                    paired = set()
                    
                    # หา Booking No. ของสาขานี้
                    code_rows = df_data[df_data['Code'] == code]
                    if len(code_rows) == 0:
                        return paired
                    
                    # เช็คว่ามีคอลัมน์ Booking หรือไม่
                    if 'Booking' not in df_data.columns and 'Trip' not in df_data.columns:
                        return paired
                    
                    booking_col = 'Booking' if 'Booking' in df_data.columns else 'Trip'
                    code_bookings = set(code_rows[booking_col].dropna().astype(str))
                    
                    if not code_bookings:
                        return paired
                    
                    # หาสาขาอื่นที่อยู่ Booking เดียวกัน (ไม่สนจังหวัด)
                    for booking in code_bookings:
                        if booking == 'nan' or not booking.strip():
                            continue
                        
                        same_booking = df_data[df_data[booking_col].astype(str) == booking]
                        for _, other_row in same_booking.iterrows():
                            other_code = other_row['Code']
                            
                            # เงื่อนไข: Booking เดียวกัน = รวมกลุ่ม (ไม่สนจังหวัด)
                            if other_code != code:
                                paired.add(other_code)
                    
                    return paired
                
                all_codes_set = set(df_region['Code'].unique())
                
                # สร้างกลุ่มสาขาแบบ Union-Find (ตามลำดับ: ตำบล → อำเภอ → จังหวัด)
                # Step 1: เริ่มจากแต่ละสาขาเป็นกลุ่มๆ พร้อมข้อมูล Master
                initial_groups = {}
                for code in all_codes_set:
                    # ดึงข้อมูลจาก Master
                    location = {}
                    if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                        master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                        if len(master_row) > 0:
                            master_row = master_row.iloc[0]
                            location = {
                                'subdistrict': master_row.get('ตำบล', ''),
                                'district': master_row.get('อำเภอ', ''),
                                'province': master_row.get('จังหวัด', 'UNKNOWN'),
                                'lat': master_row.get('ละติจูด', 0),
                                'lon': master_row.get('ลองติจูด', 0)
                            }
                    
                    # ถ้าไม่มีใน Master ลองดึงจากไฟล์อัปโหลด
                    if not location or location.get('province', 'UNKNOWN') == 'UNKNOWN':
                        c_row = df_region[df_region['Code'] == code].iloc[0] if len(df_region[df_region['Code'] == code]) > 0 else None
                        if c_row is not None:
                            location = {
                                'subdistrict': '',
                                'district': '',
                                'province': c_row.get('Province', 'UNKNOWN'),
                                'lat': 0,
                                'lon': 0
                            }
                    
                    if location:
                        initial_groups[(code,)] = {code: location}
                
                # ใช้ initial_groups แทน booking_groups
                booking_groups = initial_groups
                
                # Step 2: รวมกลุ่มตามลำดับ ตำบล → อำเภอ → จังหวัด
                def groups_can_merge(locs1, locs2):
                    """เช็คว่า 2 กลุ่มควรรวมกันไหม (ตามลำดับความละเอียด)"""
                    # 1. เช็คตำบลเดียวกัน (ต้องมีข้อมูลตำบล)
                    subdistricts1 = set(loc.get('subdistrict', '') for loc in locs1.values() if loc.get('subdistrict', ''))
                    subdistricts2 = set(loc.get('subdistrict', '') for loc in locs2.values() if loc.get('subdistrict', ''))
                    if subdistricts1 and subdistricts2 and (subdistricts1 & subdistricts2):
                        return True, 'ตำบล'
                    
                    # 2. เช็คอำเภอเดียวกัน (ต้องมีข้อมูลอำเภอและจังหวัดเดียวกัน)
                    districts1 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs1.values() if loc.get('district', '')}
                    districts2 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs2.values() if loc.get('district', '')}
                    if districts1 and districts2:
                        # เช็คว่ามีอำเภอและจังหวัดตรงกัน
                        for d1, p1 in districts1:
                            for d2, p2 in districts2:
                                if d1 == d2 and p1 == p2 and p1:
                                    return True, 'อำเภอ'
                    
                    # 3. เช็คจังหวัดเดียวกัน
                    provinces1 = set(loc.get('province', '') for loc in locs1.values() if loc.get('province', ''))
                    provinces2 = set(loc.get('province', '') for loc in locs2.values() if loc.get('province', ''))
                    if provinces1 & provinces2:
                        return True, 'จังหวัด'
                    
                    return False, None
                
                merged_groups = []
                used_groups = set()
                
                for group1, locs1 in booking_groups.items():
                    if group1 in used_groups:
                        continue
                    
                    merged_codes = set(group1)
                    merged_locs = locs1.copy()
                    used_groups.add(group1)
                    
                    # หากลุ่มอื่นที่ใกล้เคียง
                    changed = True
                    while changed:
                        changed = False
                        for group2, locs2 in booking_groups.items():
                            if group2 in used_groups:
                                continue
                            can_merge, level = groups_can_merge(merged_locs, locs2)
                            if can_merge:
                                merged_codes |= set(group2)
                                merged_locs.update(locs2)
                                used_groups.add(group2)
                                changed = True
                    
                    merged_groups.append({
                        'codes': merged_codes,
                        'locations': merged_locs
                    })
                
                # Step 3: แปลงเป็น groups format
                groups = []
                for mg in merged_groups:
                    rep_code = list(mg['codes'])[0]
                    rep_row = df_region[df_region['Code'] == rep_code].iloc[0]
                    # กรองเฉพาะจังหวัดที่ไม่ใช่ UNKNOWN และไม่เป็น NaN
                    provinces = set(
                        str(loc.get('province', '')).strip() 
                        for loc in mg['locations'].values() 
                        if loc.get('province') and str(loc.get('province', '')).strip() not in ['UNKNOWN', 'nan', '']
                    )
                    
                    # ถ้าไม่มีจังหวัดเลย ใส่ "ไม่ระบุ"
                    province_str = ', '.join(sorted(provinces)) if provinces else 'ไม่ระบุ'
                    
                    groups.append({
                        'codes': mg['codes'],
                        'region': str(rep_row.get('Region') or 'ไม่ระบุ'),
                        'province': province_str
                    })
                
                # แสดงสถิติ
                st.markdown("---")
                st.markdown("### 📊 สรุปการจัดกลุ่ม")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("📍 จำนวนสาขา", df_region['Code'].nunique())
                with col2:
                    st.metric("🗂️ จำนวนกลุ่ม", len(groups))
                with col3:
                    regions_count = df_region['Region'].nunique()
                    st.metric("🗺️ จำนวนภาค", regions_count)
                
                # แสดงตามภาค
                st.markdown("---")
                st.markdown('<div class="divider-label">🗺️ สาขาแยกตามภาค</div>', unsafe_allow_html=True)
                
                region_summary = df_region.groupby('Region').agg({
                    'Code': 'nunique',
                    'Weight': 'sum',
                    'Cube': 'sum'
                }).reset_index()
                region_summary.columns = ['ภาค', 'จำนวนสาขา', 'น้ำหนักรวม', 'คิวรวม']
                st.dataframe(region_summary, width="stretch")
                
                # แสดงรายละเอียดแต่ละภาค
                for region in sorted(df_region['Region'].unique()):
                    region_data = df_region[df_region['Region'] == region]
                    with st.expander(f"📍 {region} ({region_data['Code'].nunique()} สาขา)"):
                        display_cols = ['Code', 'Name', 'Province', 'Weight', 'Cube']
                        display_cols = [c for c in display_cols if c in region_data.columns]
                        
                        region_display = region_data[display_cols].drop_duplicates('Code')
                        col_names = {'Code': 'รหัส', 'Name': 'ชื่อสาขา', 'Province': 'จังหวัด', 'Weight': 'น้ำหนัก', 'Cube': 'คิว'}
                        region_display.columns = [col_names.get(c, c) for c in display_cols]
                        st.dataframe(region_display, width="stretch")
                
                # แสดงกลุ่มสาขาที่เคยไปด้วยกัน
                st.markdown("---")
                st.markdown('<div class="divider-label">🔗 กลุ่มสาขาที่เคยไปด้วยกัน</div>', unsafe_allow_html=True)
                
                paired_groups = [g for g in groups if len(g['codes']) > 1]
                if paired_groups:
                    for i, group in enumerate(paired_groups, 1):
                        codes_list = list(group['codes'])
                        names = []
                        for c in codes_list:
                            name_row = df_region[df_region['Code'] == c]
                            if len(name_row) > 0 and 'Name' in name_row.columns:
                                _nm = name_row['Name'].iloc[0]
                                _nm_str = str(_nm) if (_nm is not None and not (isinstance(_nm, float) and pd.isna(_nm))) else ''
                                names.append(f"{c}" + (f" ({_nm_str})" if _nm_str else ''))
                            else:
                                names.append(str(c))
                        
                        region_label = group['region'] or 'ไม่ระบุ'
                        st.write(f"**กลุ่ม {i}** — {region_label}: {', '.join(names)}")
                else:
                    st.info("ไม่พบกลุ่มสาขาที่เคยไปด้วยกันในรายการนี้")
                
                # ดาวน์โหลด
                st.markdown("---")
                output_region = io.BytesIO()
                with pd.ExcelWriter(output_region, engine='xlsxwriter') as writer:
                    df_region.to_excel(writer, sheet_name='สาขาทั้งหมด', index=False)
                    region_summary.to_excel(writer, sheet_name='สรุปตามภาค', index=False)
                
                st.download_button(
                    label="📥 ดาวน์โหลดข้อมูลจัดกลุ่ม (Excel)",
                    data=output_region.getvalue(),
                    file_name=f"จัดกลุ่มสาขา_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    width="stretch"
                )

            # ==========================================
            # แท็บ 3: โซนจัดส่งสาขา (geographic zone classification)
            # ==========================================
            with tab3:
                st.markdown('<div class="divider-label">🏙️ โซนจัดส่งของสาขา</div>', unsafe_allow_html=True)
                st.markdown(
                    "จัดทุกสาขาเข้าโซนจัดส่งตาม **ที่ตั้งทางภูมิศาสตร์** (ไม่คำนึงถึง Weight/Cube)  \n"
                    "กรุงเทพฯ แบ่ง **9 sub-zone** (ใจกลาง + 8 ทิศ)  •  จังหวัดอื่นอิง **LOGISTICS_ZONES** (เส้นทางถนน)"
                )

                if st.button("🔍 จำแนกโซนสาขาทั้งหมด", type="primary", key="btn_classify_zones"):
                    with st.spinner("กำลังจำแนกโซน..."):
                        _bz_map, _bz_summary = classify_all_branch_zones(MASTER_DATA)
                        st.session_state['branch_zone_map'] = _bz_map
                        st.session_state['branch_zone_summary'] = _bz_summary
                    st.success(f"✅ จำแนกเสร็จ: {len(_bz_map):,} สาขา / {len(_bz_summary)} โซน")

                _bz_map = st.session_state.get('branch_zone_map', {})
                _bz_summary = st.session_state.get('branch_zone_summary', {})

                if _bz_summary:
                    _bz_colors = _build_zone_color_map(_bz_summary)

                    bkk_zones  = {k: v for k, v in _bz_summary.items() if k.startswith('BKK_')}
                    prov_zones = {k: v for k, v in _bz_summary.items()
                                  if not k.startswith('BKK_') and not k.startswith('UNCLASSIFIED')}
                    unc_zones  = {k: v for k, v in _bz_summary.items() if k.startswith('UNCLASSIFIED')}

                    # ── Metrics ──
                    _zm1, _zm2, _zm3, _zm4 = st.columns(4)
                    _zm1.metric("📍 สาขาทั้งหมด", f"{len(_bz_map):,}")
                    _zm2.metric("🗺️ โซนทั้งหมด", f"{len(_bz_summary)}")
                    _zm3.metric("🏙️ BKK Sub-zones", f"{len(bkk_zones)}")
                    _unc_total = sum(v['count'] for v in unc_zones.values())
                    _zm4.metric("❓ ไม่ระบุโซน", f"{_unc_total}")

                    # ── Map ──
                    st.markdown("---")
                    st.markdown("#### 🗺️ แผนที่แสดงโซนจัดส่ง (สีต่างกันต่างโซน — คลิก layer ด้านขวาเพื่อเปิด/ปิดโซน)")

                    if FOLIUM_AVAILABLE:
                        with st.spinner("กำลังสร้างแผนที่..."):
                            _zone_fmap = _build_zone_folium_map(MASTER_DATA, _bz_map, _bz_colors)
                        if _zone_fmap:
                            folium_static(_zone_fmap, width=1100, height=680)
                        else:
                            st.warning("ไม่สามารถสร้างแผนที่: ตรวจสอบว่า MASTER_DATA มีคอลัมน์ ละติจูด/ลองติจูด")
                    else:
                        st.warning("⚠️ ต้องติดตั้ง folium และ streamlit-folium เพื่อดูแผนที่")

                    # ── Zone Legend ──
                    st.markdown("---")
                    st.markdown("#### 🎨 Legend — สีและโซน")

                    _leg_tabs = st.tabs(["🏙️ กรุงเทพฯ (Sub-zones)", "🗺️ จังหวัด (Logistics Zones)", "❓ ไม่ระบุโซน"])

                    with _leg_tabs[0]:
                        if bkk_zones:
                            _bkk_order = ['BKK_CENTER','BKK_N','BKK_NE','BKK_E','BKK_SE',
                                          'BKK_S','BKK_SW','BKK_W','BKK_NW']
                            _bkk_cols = st.columns(3)
                            for _bi, _bk in enumerate([z for z in _bkk_order if z in bkk_zones]):
                                _bv = bkk_zones[_bk]
                                _bc = _bz_colors.get(_bk, '#888')
                                _desc = BKK_SUBZONE_NAMES.get(_bk, _bk)
                                with _bkk_cols[_bi % 3]:
                                    st.markdown(
                                        f'<div style="background:{_bc};color:#fff;border-radius:8px;'
                                        f'padding:10px 14px;margin:4px 0;font-size:13px;">'
                                        f'<b>{_bk}</b><br><span style="font-size:11px;">{_desc}</span>'
                                        f'<br><b>{_bv["count"]} สาขา</b></div>',
                                        unsafe_allow_html=True
                                    )
                        else:
                            st.info("ไม่พบข้อมูลกรุงเทพฯ")

                    with _leg_tabs[1]:
                        if prov_zones:
                            # Group by region for display
                            from collections import defaultdict as _ddleg
                            _leg_by_region = _ddleg(list)
                            for _zk, _zv in prov_zones.items():
                                _pv = _zv.get('province', '')
                                _rv = get_region_name(_pv) if _pv else 'ไม่ระบุ'
                                _leg_by_region[_rv].append((_zk, _zv))

                            for _rv in ['เหนือ','อีสาน','ใต้','ตะวันออก','กลาง','ตะวันตก','ไม่ระบุ']:
                                if _rv not in _leg_by_region:
                                    continue
                                _rlist = sorted(_leg_by_region[_rv], key=lambda x: -x[1]['count'])
                                with st.expander(f"📍 ภาค{_rv} ({len(_rlist)} โซน)", expanded=False):
                                    _rcols = st.columns(4)
                                    for _ri, (_zk, _zv) in enumerate(_rlist):
                                        _zc = _bz_colors.get(_zk, '#888')
                                        _zlabel = _zk.replace('ZONE_', '').replace('_', ' ')
                                        with _rcols[_ri % 4]:
                                            st.markdown(
                                                f'<div style="background:{_zc};color:#fff;border-radius:6px;'
                                                f'padding:6px 10px;margin:3px 0;font-size:12px;">'
                                                f'<b>{_zlabel}</b><br>{_zv["count"]} สาขา</div>',
                                                unsafe_allow_html=True
                                            )
                        else:
                            st.info("ไม่พบโซนจังหวัด")

                    with _leg_tabs[2]:
                        if unc_zones:
                            _unc_rows = [{'โซน (จังหวัดที่ไม่พบใน LOGISTICS_ZONES)': k,
                                          'จำนวนสาขา': v['count']} for k, v in unc_zones.items()]
                            st.dataframe(pd.DataFrame(_unc_rows), hide_index=True, use_container_width=True)
                        else:
                            st.success("✅ ทุกสาขามีโซนครบถ้วน")

                    # ── Zone Summary Table ──
                    st.markdown("---")
                    st.markdown("#### 📋 ตารางสรุปโซนทั้งหมด")
                    _sum_rows = []
                    for _zk, _zv in sorted(_bz_summary.items(), key=lambda x: (-x[1]['count'], x[0])):
                        _pv = _zv.get('province', '')
                        _rv = get_region_name(_pv) if _pv else ''
                        _zdesc = BKK_SUBZONE_NAMES.get(_zk, _zk.replace('ZONE_','').replace('_',' '))
                        _zc = _bz_colors.get(_zk, '#9E9E9E')
                        _sum_rows.append({
                            '🎨': f'<div style="background:{_zc};width:18px;height:18px;border-radius:4px;"></div>',
                            'Zone': _zk,
                            'คำอธิบาย': _zdesc,
                            'จังหวัด': _pv,
                            'ภาค': _rv,
                            'จำนวนสาขา': _zv['count'],
                        })
                    _sum_df = pd.DataFrame(_sum_rows)
                    st.dataframe(
                        _sum_df.drop(columns=['🎨']),
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            'จำนวนสาขา': st.column_config.ProgressColumn(
                                'จำนวนสาขา', format='%d สาขา',
                                min_value=0,
                                max_value=int(_sum_df['จำนวนสาขา'].max()) if len(_sum_df) > 0 else 1
                            )
                        }
                    )

                    # ── Downloads ──
                    st.markdown("---")
                    st.markdown("#### 📥 ดาวน์โหลดข้อมูล")
                    _dl1, _dl2 = st.columns(2)

                    with _dl1:
                        # Excel multi-sheet
                        with st.spinner("เตรียมไฟล์ Excel..."):
                            _excel_bytes = _build_zone_excel(MASTER_DATA, _bz_map, _bz_summary, _bz_colors)
                        st.download_button(
                            label="📊 ดาวน์โหลด Excel แยกโซน (หลายชีต)",
                            data=_excel_bytes,
                            file_name=f"branch_zones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            key="dl_zone_excel"
                        )

                    with _dl2:
                        # CSV quick export
                        _bz_csv_rows = []
                        _nm_map2 = {}
                        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns and 'สาขา' in MASTER_DATA.columns:
                            _nm_map2 = dict(zip(
                                MASTER_DATA['Plan Code'].astype(str).str.strip(),
                                MASTER_DATA['สาขา'].fillna('')))
                        for _code, _zone in sorted(_bz_map.items()):
                            _bz_csv_rows.append({
                                'Plan Code': _code,
                                'ชื่อสาขา': _nm_map2.get(_code, ''),
                                'Zone': _zone,
                                'Zone_Description': BKK_SUBZONE_NAMES.get(_zone, _zone),
                            })
                        _bz_csv = pd.DataFrame(_bz_csv_rows).to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            label="📄 ดาวน์โหลด CSV (ไฟล์เดียว)",
                            data=_bz_csv,
                            file_name=f"branch_zones_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv",
                            key="dl_zone_csv"
                        )
                else:
                    st.info("กด **🔍 จำแนกโซนสาขาทั้งหมด** เพื่อเริ่มต้น")

if __name__ == "__main__":
    try:
        main()
    finally:
        # บันทึก cache ก่อนปิดโปรแกรม (เฉพาะเมื่อมีการเปลี่ยนแปลง)
        if USE_CACHE:
            save_distance_cache(DISTANCE_CACHE, force=True)
            save_route_cache(ROUTE_CACHE_DATA, force=True)
            safe_print(f"💾 บันทึก cache: {len(DISTANCE_CACHE)} ระยะทาง, {len(ROUTE_CACHE_DATA)} เส้นทาง")

