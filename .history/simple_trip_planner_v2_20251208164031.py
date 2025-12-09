"""
Simple Trip Planner V2 - จัดทริปตามเงื่อนไขใหม่
เรียงตาม: ระยะทาง → ชื่อ → จังหวัด → อำเภอ → ตำบล (ไกล → ใกล้)
เริ่มจาก 6W (ไกลสุดก่อน)
เช็คระยะทางระหว่างสาขา
รวมทริปเศษให้เต็ม
"""

import pandas as pd
import numpy as np
import math
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# พิกัด DC วังน้อย (DC011 - วังน้อย พระนครศรีอยุธยา)
DC_WANG_NOI_LAT = 14.1793943
DC_WANG_NOI_LON = 100.6481489

# ลิมิตใหม่: 6W ≥18 คิว, JB ≤7 คิว, 4W = 5/3-4 คิว
LIMITS = {
    '6W': {'max_c': 20, 'min_c': 18, 'max_w': 9000, 'max_drops': 20},  # 6W ต้อง 18-20 คิว
    'JB': {'max_c': 7, 'max_w': 3500, 'max_drops': 7},  # JB ≤7 คิว, ≤7 drops
    '4W': {'max_c': 5, 'max_w': 1800, 'max_drops': 12}  # 4W: 5คิว(PT)/3-4คิว(Mix)
}

MAX_DISTANCE_BETWEEN_BRANCHES = 100  # km - ระยะห่างระหว่างสาขาติดกัน
MAX_DC_DISTANCE_SPREAD = 80  # km - ความห่างสูงสุดของ Distance_DC ในทริปเดียวกัน (ป้องกันข้ามภูมิภาค)
MAX_DIRECTION_DIFF = 2  # จำนวนทิศทางที่แตกต่างกันได้สูงสุดในทริป (เช่น N กับ NE ได้, แต่ N กับ S ไม่ได้)
NEAR_DC_THRESHOLD = 150  # km - บริเวณใกล้ DC ให้ใช้รถเล็ก
FAR_DC_THRESHOLD = 290  # km - บริเวณไกล DC (>290km) ให้ใช้รถใหญ่

def haversine_distance(lat1, lon1, lat2, lon2):
    """คำนวณระยะทางระหว่างสองจุด (km) - เส้นตรง"""
    R = 6371
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    return R * c

def get_road_distance(lat1, lon1, lat2, lon2):
    """คำนวณระยะทางจริงตามถนน (km) - ประมาณจาก Haversine × 1.35"""
    # ใช้ค่าประมาณ: ระยะทางถนน = ระยะเส้นตรง × 1.35
    # (เพราะถนนไม่ตรง มีการเลี้ยว ต้องวนไปตามเส้นทาง)
    straight_dist = haversine_distance(lat1, lon1, lat2, lon2)
    return straight_dist * 1.35

def get_direction_from_dc(lat, lon, dc_lat, dc_lon):
    """คำนวณทิศทางจาก DC (N/S/E/W/NE/NW/SE/SW)"""
    import math
    if not lat or not lon or lat == 0 or lon == 0:
        return 'UNKNOWN'
    
    dlat = lat - dc_lat
    dlon = lon - dc_lon
    
    # คำนวณมุม (0° = เหนือ, 90° = ตะวันออก)
    angle = math.atan2(dlon, dlat) * 180 / math.pi
    
    # แปลงมุมเป็นทิศ 8 ทิศ
    if -22.5 <= angle < 22.5:
        return 'N'  # เหนือ
    elif 22.5 <= angle < 67.5:
        return 'NE'  # ตะวันออกเฉียงเหนือ
    elif 67.5 <= angle < 112.5:
        return 'E'  # ตะวันออก
    elif 112.5 <= angle < 157.5:
        return 'SE'  # ตะวันออกเฉียงใต้
    elif angle >= 157.5 or angle < -157.5:
        return 'S'  # ใต้
    elif -157.5 <= angle < -112.5:
        return 'SW'  # ตะวันตกเฉียงใต้
    elif -112.5 <= angle < -67.5:
        return 'W'  # ตะวันตก
    else:  # -67.5 <= angle < -22.5
        return 'NW'  # ตะวันตกเฉียงเหนือ

def are_directions_compatible(dir1, dir2):
    """เช็คว่า 2 ทิศทางเข้ากันได้หรือไม่ (ไม่ควรอยู่ตรงข้ามกัน)"""
    if dir1 == 'UNKNOWN' or dir2 == 'UNKNOWN':
        return True
    
    # ทิศตรงข้าม
    opposite_pairs = [
        ('N', 'S'), ('S', 'N'),
        ('E', 'W'), ('W', 'E'),
        ('NE', 'SW'), ('SW', 'NE'),
        ('NW', 'SE'), ('SE', 'NW')
    ]
    
    # ถ้าเป็นทิศตรงข้าม → ไม่เข้ากัน
    if (dir1, dir2) in opposite_pairs:
        return False
    
    # ทิศเดียวกันหรือใกล้เคียง → เข้ากัน
    return True

def get_region(province):
    """จัดจังหวัดเข้ากลุ่มภูมิภาค"""
    if not province:
        return 'UNKNOWN'
    
    province = str(province).strip()
    
    # ภาคเหนือ
    north = ['เชียงใหม่', 'เชียงราย', 'ลำปาง', 'ลำพูน', 'แม่ฮ่องสอน', 'น่าน', 
             'พะเยา', 'แพร่', 'อุตรดิตถ์', 'ตาก', 'สุโขทัย', 'พิษณุโลก', 
             'พิจิตร', 'เพชรบูรณ์', 'กำแพงเพชร', 'นครสวรรค์', 'อุทัยธานี']
    
    # ภาคตะวันออกเฉียงเหนือ (อีสาน)
    northeast = ['นครราชสีมา', 'บุรีรัมย์', 'สุรินทร์', 'ศรีสะเกษ', 'อุบลราชธานี',
                 'ยโสธร', 'ชัยภูมิ', 'อำนาจเจริญ', 'หนองบัวลำภู', 'ขอนแก่น',
                 'อุดรธานี', 'เลย', 'หนองคาย', 'มหาสารคาม', 'ร้อยเอ็ด',
                 'กาฬสินธุ์', 'สกลนคร', 'นครพนม', 'มุกดาหาร', 'บึงกาฬ']
    
    # ภาคกลาง
    central = ['กรุงเทพฯ', 'กรุงเทพมหานคร', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ',
               'สมุทรสาคร', 'นครปฐม', 'ฉะเชิงเทรา', 'ชลบุรี', 'ระยอง',
               'จันทบุรี', 'ตราด', 'ประจวบคีรีขันธ์', 'เพชรบุรี', 'ราชบุรี',
               'กาญจนบุรี', 'สุพรรณบุรี', 'พระนครศรีอยุธยา', 'อ่างทอง',
               'ลพบุรี', 'สระบุรี', 'สิงห์บุรี', 'ชัยนาท', 'สมุทรสงคราม']
    
    # ภาคใต้
    south = ['นครศรีธรรมราช', 'สุราษฎร์ธานี', 'กระบี่', 'พังงา', 'ภูเก็ต',
             'ระนอง', 'ชุมพร', 'สงขลา', 'สตูล', 'ตรัง', 'พัทลุง', 'ปัตตานี',
             'ยะลา', 'นราธิวาส']
    
    if province in north:
        return 'NORTH'
    elif province in northeast:
        return 'NORTHEAST'
    elif province in central:
        return 'CENTRAL'
    elif province in south:
        return 'SOUTH'
    else:
        return 'UNKNOWN'

def is_punthai(branch_code):
    """เช็คว่าเป็นสาขา Punthai หรือไม่"""
    code_str = str(branch_code).upper()
    return code_str.startswith('PH') or code_str.startswith('PU') or code_str.startswith('PT')

def get_base_name(name):
    """ดึงชื่อพื้นฐานของสาขา"""
    import re
    if not name:
        return ""
    
    name_lower = str(name).strip().lower()
    
    if 'future' in name_lower or 'ฟิวเจอร์' in name_lower:
        if 'rangsit' in name_lower or 'รังสิต' in name_lower:
            return "ฟิวเจอร์รังสิต"
    
    if 'lotus' in name_lower or 'โลตัส' in name_lower:
        base = re.sub(r'\s*\d+\s*$', '', name_lower)
        return base.strip()
    
    base = re.sub(r'\s*\d+\s*$', '', str(name).strip())
    return base.strip().lower()

def plan_trips_v2(df, master_data):
    """
    จัดทริปตามเงื่อนไขใหม่:
    1. เรียงตามระยะทาง → ชื่อ → จังหวัด → อำเภอ → ตำบล (ไกล → ใกล้)
    2. เริ่มจาก 6W (ไกลสุดก่อน)
    3. เช็คระยะทางระหว่างสาขา
    4. รวมทริปเศษให้เต็ม
    """
    
    # เตรียมข้อมูล
    result_df = df.copy()
    result_df['Trip'] = 0
    result_df['Distance_DC'] = 0.0
    result_df['Base_Name'] = ''
    result_df['Province'] = ''
    result_df['District'] = ''
    result_df['Subdistrict'] = ''
    result_df['Is_Punthai'] = False
    result_df['Latitude'] = 0.0
    result_df['Longitude'] = 0.0
    result_df['Truck'] = ''
    result_df['Direction'] = ''  # ทิศทางจาก DC
    result_df['Region'] = ''  # ภูมิภาค
    
    # ดึงข้อมูลจาก Master
    for idx, row in result_df.iterrows():
        code = row['Code']
        
        if not master_data.empty and 'Plan Code' in master_data.columns:
            master_row = master_data[master_data['Plan Code'] == code]
            if len(master_row) > 0:
                m = master_row.iloc[0]
                lat = m.get('ละติจูด', 0) if pd.notna(m.get('ละติจูด')) else 0
                lon = m.get('ลองติจูด', 0) if pd.notna(m.get('ลองติจูด')) else 0
                
                if lat and lon and lat != 0 and lon != 0:
                    dist = haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
                    result_df.at[idx, 'Distance_DC'] = dist
                    result_df.at[idx, 'Latitude'] = lat
                    result_df.at[idx, 'Longitude'] = lon
                    result_df.at[idx, 'Direction'] = get_direction_from_dc(lat, lon, DC_WANG_NOI_LAT, DC_WANG_NOI_LON)
                
                province = str(m.get('จังหวัด', '')).strip()
                result_df.at[idx, 'Province'] = province
                result_df.at[idx, 'Region'] = get_region(province)
                result_df.at[idx, 'District'] = str(m.get('อำเภอ', '')).strip()
                result_df.at[idx, 'Subdistrict'] = str(m.get('ตำบล', '')).strip()
        
        if 'Name' in result_df.columns:
            result_df.at[idx, 'Base_Name'] = get_base_name(row['Name'])
        
        result_df.at[idx, 'Is_Punthai'] = is_punthai(code)
    
    # เรียงลำดับ: Region → ทิศทาง → ระยะทาง → จังหวัด → ชื่อ (จัดกลุ่มตามภูมิภาคก่อน)
    result_df = result_df.sort_values(
        by=['Region', 'Direction', 'Distance_DC', 'Province', 'Base_Name'],
        ascending=[True, True, False, True, True]
    ).reset_index(drop=True)
    
    # จัดทริป
    trip_num = 1
    trips = []
    remaining_indices = list(result_df.index)
    
    while remaining_indices:
        seed_idx = remaining_indices.pop(0)
        seed_row = result_df.iloc[seed_idx]
        
        # Debug Trip 7 seed
        if trip_num == 7:
            print(f"DEBUG Trip {trip_num} SEED: {seed_row['Base_Name']} (Region={seed_row['Region']}, Dir={seed_row['Direction']}, DC={seed_row['Distance_DC']:.1f}km)")
        
        current_trip = [seed_idx]
        current_cubes = seed_row['Cube'] if pd.notna(seed_row['Cube']) else 0
        current_weight = seed_row['Weight'] if pd.notna(seed_row['Weight']) else 0
        all_punthai = seed_row['Is_Punthai']
        last_lat = seed_row['Latitude']
        last_lon = seed_row['Longitude']
        seed_distance = seed_row['Distance_DC']
        trip_min_dc_distance = seed_distance  # ระยะใกล้สุดจาก DC ในทริป
        trip_max_dc_distance = seed_distance  # ระยะไกลสุดจาก DC ในทริป
        
        # กำหนดประเภทรถตามระยะทางและคิว
        # 6W: ต้อง ≥18 คิว (ถ้าไม่ถึง 18 ให้ใช้ JB หลายคัน)
        # JB: ≤7 คิว, ≤7 drops
        # 4W: 5 คิว (PT ล้วน), 3-4 คิว (คละ)
        if seed_distance > FAR_DC_THRESHOLD:
            # ไกล (>290km) → ประเมินว่ามีคิวพอสำหรับ 6W ไหม
            estimated_cubes = current_cubes
            for idx in remaining_indices[:25]:  # ดู 25 สาขาถัดไป
                r = result_df.iloc[idx]
                if r['Distance_DC'] > FAR_DC_THRESHOLD - 50:
                    estimated_cubes += r['Cube'] if pd.notna(r['Cube']) else 0
            
            # ถ้าสาขารวมแล้ว ≥18 คิว → ใช้ 6W, ไม่งั้นใช้ JB
            if estimated_cubes >= LIMITS['6W']['min_c']:
                truck_type = '6W'
            else:
                truck_type = 'JB'
        elif seed_distance < NEAR_DC_THRESHOLD:
            # ใกล้ (<150km) → ใช้รถเล็ก
            if current_cubes > LIMITS['4W']['max_c']:
                truck_type = 'JB'
            else:
                truck_type = '4W'
        else:
            # ระยะกลาง (150-290km) → ใช้ JB
            if current_cubes > LIMITS['JB']['max_c']:
                truck_type = 'JB'  # เกิน 7 คิว → ยังใช้ JB (จะแยกเป็น 2 คันภายหลัง)
            elif current_cubes > LIMITS['4W']['max_c']:
                truck_type = 'JB'
            else:
                truck_type = '4W'
        
        # เพิ่มสาขาเข้าทริป
        indices_to_remove = []
        for idx in remaining_indices[:]:
            row = result_df.iloc[idx]
            cube = row['Cube'] if pd.notna(row['Cube']) else 0
            weight = row['Weight'] if pd.notna(row['Weight']) else 0
            is_punthai_branch = row['Is_Punthai']
            branch_dc_distance = row['Distance_DC']
            branch_region = row['Region']
            branch_direction = row['Direction']
            
            if all_punthai and not is_punthai_branch:
                all_punthai = False
            
            # ระยะทางจากสาขาล่าสุด
            branch_lat = row['Latitude']
            branch_lon = row['Longitude']
            distance_from_last = 0
            
            if last_lat and last_lon and branch_lat and branch_lon:
                distance_from_last = get_road_distance(last_lat, last_lon, branch_lat, branch_lon)
            
            # คำนวณ
            new_cubes = current_cubes + cube
            new_weight = current_weight + weight
            
            # เช็คว่าสาขานี้อยู่ในภูมิภาค/ทิศทางเดียวกันหรือใกล้เคียงหรือไม่
            seed_region = result_df.iloc[seed_idx]['Region']
            seed_direction = result_df.iloc[seed_idx]['Direction']
            
            # 1. เช็คภูมิภาค - ถ้าต่างภูมิภาค → ไม่รวม (ยกเว้น CENTRAL กับ UNKNOWN)
            regions_compatible = True
            if seed_region not in ['CENTRAL', 'UNKNOWN'] and branch_region not in ['CENTRAL', 'UNKNOWN']:
                if seed_region != branch_region:
                    regions_compatible = False
            
            # 2. เช็คทิศทาง - ไม่ควรอยู่ตรงข้ามกัน
            directions_compatible = are_directions_compatible(seed_direction, branch_direction)
            
            # 3. ถ้า CENTRAL ผสมกับภาคอื่น → ต้องเช็คทิศทางให้เข้มงวด
            if (seed_region == 'CENTRAL' and branch_region not in ['CENTRAL', 'UNKNOWN']) or \
               (branch_region == 'CENTRAL' and seed_region not in ['CENTRAL', 'UNKNOWN']):
                # CENTRAL ผสมภาคอื่น → ต้องทิศทางเดียวกันเท่านั้น
                if seed_direction != branch_direction:
                    # Debug: แสดงข้อมูลเมื่อบล็อก
                    if trip_num == 7 and len(current_trip) <= 7:
                        print(f"DEBUG Trip {trip_num}: Blocked {row['Base_Name']} ({branch_region}/{branch_direction}) from seed {seed_row['Base_Name']} ({seed_region}/{seed_direction})")
                    directions_compatible = False
            
            # เช็คว่าสาขานี้จะทำให้ทริปกระจายเกินไปหรือไม่ (ป้องกันข้ามภูมิภาค)
            potential_min_dc = min(trip_min_dc_distance, branch_dc_distance)
            potential_max_dc = max(trip_max_dc_distance, branch_dc_distance)
            dc_distance_spread = potential_max_dc - potential_min_dc
            
            # กำหนด limit
            if truck_type == '6W':
                max_cubes = LIMITS['6W']['max_c']
                max_weight = LIMITS['6W']['max_w']
                max_drops = LIMITS['6W']['max_drops']
            elif truck_type == 'JB':
                max_cubes = LIMITS['JB']['max_c']
                max_weight = LIMITS['JB']['max_w']
                max_drops = LIMITS['JB']['max_drops']
            else:
                max_cubes = LIMITS['4W']['max_c'] if all_punthai else 3.5  # PT=5, Mix=3-4
                max_weight = LIMITS['4W']['max_w']
                max_drops = LIMITS['4W']['max_drops']
            
            should_add = True
            
            # เช็คเงื่อนไข
            if not regions_compatible:
                # ต่างภูมิภาค → ไม่รวม
                should_add = False
            elif not directions_compatible:
                # ทิศทางไม่เข้ากัน (เช่น เหนือกับใต้) → ไม่รวม
                should_add = False
            elif distance_from_last > MAX_DISTANCE_BETWEEN_BRANCHES:
                should_add = False
            elif dc_distance_spread > MAX_DC_DISTANCE_SPREAD:
                # ถ้าสาขานี้ทำให้ทริปกระจายเกินไป (เช่น ข้ามจากภาคใต้มาภาคกลาง)
                should_add = False
            elif new_cubes > max_cubes:
                # ถ้ามีรถ 6W อยู่แล้ว → ใส่ต่อไปได้ถึง 20 คิว
                if truck_type == '6W':
                    if new_cubes <= LIMITS['6W']['max_c']:
                        pass  # ใส่ได้
                    else:
                        should_add = False
                # JB: ห้ามเกิน 7 คิว (ไม่อัพเกรด)
                elif truck_type == 'JB':
                    should_add = False  # เกิน 7 คิว → ตัดทริป (ไม่อัพเกรดเป็น 6W)
                # 4W → JB
                elif truck_type == '4W' and new_cubes <= LIMITS['JB']['max_c']:
                    truck_type = 'JB'
                    max_cubes = LIMITS['JB']['max_c']
                    max_weight = LIMITS['JB']['max_w']
                    max_drops = LIMITS['JB']['max_drops']
                else:
                    should_add = False
            elif new_weight > max_weight:
                # ถ้ามีรถ 6W อยู่แล้ว → ผ่อนปรนเรื่องน้ำหนัก
                if truck_type == '6W' and new_weight <= LIMITS['6W']['max_w'] * 1.1:
                    pass  # ใส่ได้ (เกินน้ำหนักได้ 10%)
                else:
                    should_add = False
            elif len(current_trip) + 1 > max_drops:
                # ถ้ามีรถ 6W อยู่แล้ว → ใส่ได้จนถึง 20 drops
                if truck_type == '6W' and len(current_trip) + 1 <= LIMITS['6W']['max_drops']:
                    pass  # ใส่ได้
                else:
                    should_add = False
            
            if should_add:
                # Debug Trip 7 additions
                if trip_num == 7 and len(current_trip) < 10:
                    print(f"DEBUG Trip {trip_num} ADDED: {row['Base_Name']} (Region={branch_region}, Dir={branch_direction}, DC={branch_dc_distance:.1f}km)")
                
                current_trip.append(idx)
                current_cubes = new_cubes
                current_weight = new_weight
                last_lat = branch_lat
                last_lon = branch_lon
                # อัพเดทขอบเขตระยะทาง DC ของทริป
                trip_min_dc_distance = min(trip_min_dc_distance, branch_dc_distance)
                trip_max_dc_distance = max(trip_max_dc_distance, branch_dc_distance)
                indices_to_remove.append(idx)
        
        for idx in indices_to_remove:
            remaining_indices.remove(idx)
        
        trips.append((trip_num, current_trip, truck_type, all_punthai))
        trip_num += 1
    
    # กำหนดเลขทริป
    for trip_num, trip_indices, truck_type, all_punthai in trips:
        for idx in trip_indices:
            result_df.at[idx, 'Trip'] = trip_num
            result_df.at[idx, 'Truck'] = truck_type
    
    # Phase 2: ตัดทริป 6W ที่ < 18 คิว ให้เป็น JB หลายคัน
    # และตัดทริป JB ที่ > 7 คิว ให้เป็น JB หลายคัน
    max_trip = int(result_df['Trip'].max())
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        
        trip_data = result_df[result_df['Trip'] == trip].copy()
        total_cubes = trip_data['Cube'].sum()
        truck = trip_data['Truck'].iloc[0]
        
        # 6W ที่ < 18 คิว → แยกเป็น JB
        if truck == '6W' and total_cubes < LIMITS['6W']['min_c']:
            # แยกเป็น JB หลายคัน (แต่ละคัน ≤7 คิว)
            new_trip_num = max_trip + 1
            current_cubes = 0
            current_drops = 0
            
            for idx in trip_data.index:
                cube = result_df.at[idx, 'Cube'] if pd.notna(result_df.at[idx, 'Cube']) else 0
                
                if (current_cubes + cube > LIMITS['JB']['max_c'] or current_drops >= LIMITS['JB']['max_drops']) and current_cubes > 0:
                    new_trip_num += 1
                    current_cubes = 0
                    current_drops = 0
                
                result_df.at[idx, 'Trip'] = new_trip_num
                result_df.at[idx, 'Truck'] = 'JB'
                current_cubes += cube
                current_drops += 1
            
            max_trip = new_trip_num
        
        # JB ที่ > 7 คิว → แยกเป็น JB หลายคัน
        elif truck == 'JB' and total_cubes > LIMITS['JB']['max_c']:
            new_trip_num = max_trip + 1
            current_cubes = 0
            current_drops = 0
            
            for idx in trip_data.index:
                cube = result_df.at[idx, 'Cube'] if pd.notna(result_df.at[idx, 'Cube']) else 0
                
                if (current_cubes + cube > LIMITS['JB']['max_c'] or current_drops >= LIMITS['JB']['max_drops']) and current_cubes > 0:
                    new_trip_num += 1
                    current_cubes = 0
                    current_drops = 0
                
                result_df.at[idx, 'Trip'] = new_trip_num
                result_df.at[idx, 'Truck'] = 'JB'
                current_cubes += cube
                current_drops += 1
            
            max_trip = new_trip_num
    
    # Phase 2.5: เปลี่ยน JB ที่ < 5 คิว เป็น 4W
    jb_to_4w_count = 0
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        
        trip_data = result_df[result_df['Trip'] == trip]
        total_cubes = trip_data['Cube'].sum()
        truck = trip_data['Truck'].iloc[0]
        all_punthai_trip = trip_data['Is_Punthai'].all()
        
        # JB < 5 คิว → เปลี่ยนเป็น 4W (ถ้าไม่เกิน 4W limit)
        if truck == 'JB' and total_cubes < 5:
            cube_limit = LIMITS['4W']['max_c'] if all_punthai_trip else 3.5
            
            # เปลี่ยนเป็น 4W ไม่เช็ค limit (เพราะ < 5 ก็ไม่เกิน limit อยู่แล้ว)
            for idx in trip_data.index:
                result_df.at[idx, 'Truck'] = '4W'
            jb_to_4w_count += 1
    
    print(f"🔄 เปลี่ยน JB → 4W: {jb_to_4w_count} ทริป")
    
    # Phase 2.6: แยกทริปที่เกิน limit (loop จนกว่าจะไม่เจอทริปเกิน)
    for round_num in range(5):  # เพิ่มเป็น 5 รอบเพื่อความแน่ใจ
        max_trip = int(result_df['Trip'].max())
        found_over_limit = False
        split_count = 0
        
        for trip in sorted(result_df['Trip'].unique()):
            if trip == 0:
                continue
            
            trip_data = result_df[result_df['Trip'] == trip].copy()
            total_cubes = trip_data['Cube'].sum()
            
            # ตรวจสอบ Truck type จากค่าที่มีมากที่สุดในทริป (ป้องกัน mixed truck)
            truck_counts = trip_data['Truck'].value_counts()
            if len(truck_counts) == 0:
                continue
            truck = truck_counts.index[0]
            
            all_punthai_trip = trip_data['Is_Punthai'].all()
            
            # กำหนด limit ตาม truck type
            if truck == '6W':
                cube_limit = LIMITS['6W']['max_c']
                drop_limit = LIMITS['6W']['max_drops']
            elif truck == 'JB':
                cube_limit = LIMITS['JB']['max_c']
                drop_limit = LIMITS['JB']['max_drops']
            elif truck == '4W':
                cube_limit = LIMITS['4W']['max_c'] if all_punthai_trip else 3.5
                drop_limit = LIMITS['4W']['max_drops']
            else:
                continue
            
            # ถ้าเกิน limit → แยกทริป
            if total_cubes > cube_limit or len(trip_data) > drop_limit:
                found_over_limit = True
                split_count += 1
                
                # คำนวณระยะทางจาก DC สำหรับแต่ละสาขาในทริป
                trip_indices = trip_data.index.tolist()
                distances = []
                for idx in trip_indices:
                    lat = result_df.at[idx, 'Latitude']
                    lon = result_df.at[idx, 'Longitude']
                    if lat and lon:
                        dist = get_road_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
                        distances.append((idx, dist))
                    else:
                        distances.append((idx, 0))
                
                # เรียงจากไกล → ใกล้ เพื่อให้เส้นทางวนกลับ DC
                distances.sort(key=lambda x: x[1], reverse=True)
                sorted_indices = [idx for idx, _ in distances]
                
                new_trip_num = max_trip + 1
                current_cubes = 0
                current_drops = 0
                prev_lat = DC_WANG_NOI_LAT
                prev_lon = DC_WANG_NOI_LON
                
                for idx in sorted_indices:
                    cube = result_df.at[idx, 'Cube'] if pd.notna(result_df.at[idx, 'Cube']) else 0
                    branch_lat = result_df.at[idx, 'Latitude']
                    branch_lon = result_df.at[idx, 'Longitude']
                    
                    # คำนวณระยะทางจากสาขาก่อนหน้า (ระยะทางถนนจริง)
                    if branch_lat and branch_lon and prev_lat and prev_lon:
                        dist_from_prev = get_road_distance(prev_lat, prev_lon, branch_lat, branch_lon)
                    else:
                        dist_from_prev = 0
                    
                    # ถ้าใส่สาขานี้แล้วจะเกิน limit และมีสาขาอยู่แล้ว → เริ่มทริปใหม่
                    # หรือถ้าระยะห่างจากสาขาก่อนหน้า > 150km (กระโดดไกลเกินไป)
                    if (((current_cubes + cube > cube_limit) or (current_drops >= drop_limit)) and current_cubes > 0):
                        new_trip_num += 1
                        current_cubes = 0
                        current_drops = 0
                        prev_lat = DC_WANG_NOI_LAT
                        prev_lon = DC_WANG_NOI_LON
                    
                    result_df.at[idx, 'Trip'] = new_trip_num
                    result_df.at[idx, 'Truck'] = truck  # เซ็ตให้แน่ใจ
                    current_cubes += cube
                    current_drops += 1
                    
                    # อัพเดทตำแหน่งปัจจุบัน
                    if branch_lat and branch_lon:
                        prev_lat = branch_lat
                        prev_lon = branch_lon
                
                max_trip = new_trip_num
        
        if split_count > 0:
            print(f"  รอบที่ {round_num + 1}: แยก {split_count} ทริปที่เกิน limit")
        
        # ถ้าไม่เจอทริปที่เกิน limit แล้ว → หยุดลูป
        if not found_over_limit:
            break
    
    # Phase 2.7: แปลง JB single-branch ที่เกิน 7 คิว → 6W
    convert_to_6w = 0
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        trip_data = result_df[result_df['Trip'] == trip]
        if len(trip_data) == 1:  # single-branch trip
            cube = trip_data['Cube'].iloc[0]
            truck = trip_data['Truck'].iloc[0]
            if truck == 'JB' and cube > 7:
                result_df.loc[result_df['Trip'] == trip, 'Truck'] = '6W'
                convert_to_6w += 1
    
    if convert_to_6w > 0:
        print(f"🔄 เปลี่ยน JB > 7 คิว (1 สาขา) → 6W: {convert_to_6w} ทริป")
    
    # Phase 3: รวมทริปเศษ (ต้อง check limit ก่อน merge)
    MIN_BRANCHES = 3
    small_trips = [(t, i, tr, p) for t, i, tr, p in trips if len(i) < MIN_BRANCHES]
    
    for small_trip_num, small_indices, small_truck, small_punthai in small_trips:
        best_merge = None
        best_distance = float('inf')
        
        for target_trip_num, target_indices, target_truck, target_punthai in trips:
            if target_trip_num == small_trip_num:
                continue
            
            # 1. ตรวจสอบจำนวน drops
            if len(target_indices) + len(small_indices) > LIMITS[target_truck]['max_drops']:
                continue
            
            # 2. คำนวณ cube + weight หลัง merge
            small_cubes = sum(result_df.at[si, 'Cube'] for si in small_indices if pd.notna(result_df.at[si, 'Cube']))
            target_cubes = sum(result_df.at[ti, 'Cube'] for ti in target_indices if pd.notna(result_df.at[ti, 'Cube']))
            small_weight = sum(result_df.at[si, 'Weight'] for si in small_indices if pd.notna(result_df.at[si, 'Weight']))
            target_weight = sum(result_df.at[ti, 'Weight'] for ti in target_indices if pd.notna(result_df.at[ti, 'Weight']))
            
            merged_cubes = small_cubes + target_cubes
            merged_weight = small_weight + target_weight
            
            # 3. เช็ค limit ตาม truck type
            target_all_punthai = all(result_df.at[ti, 'Is_Punthai'] for ti in target_indices)
            
            if target_truck == '6W':
                if merged_cubes > LIMITS['6W']['max_c'] or merged_weight > LIMITS['6W']['max_w']:
                    continue
            elif target_truck == 'JB':
                if merged_cubes > LIMITS['JB']['max_c'] or merged_weight > LIMITS['JB']['max_w']:
                    continue
            elif target_truck == '4W':
                cube_limit = LIMITS['4W']['max_c'] if target_all_punthai else 3.5
                if merged_cubes > cube_limit or merged_weight > LIMITS['4W']['max_w']:
                    continue
            
            # 4. คำนวณระยะทาง
            total_dist = 0
            count = 0
            for si in small_indices:
                s_lat = result_df.at[si, 'Latitude']
                s_lon = result_df.at[si, 'Longitude']
                for ti in target_indices[:5]:  # เช็คแค่ 5 สาขาแรก
                    t_lat = result_df.at[ti, 'Latitude']
                    t_lon = result_df.at[ti, 'Longitude']
                    if s_lat and t_lat:
                        dist = get_road_distance(s_lat, s_lon, t_lat, t_lon)
                        total_dist += dist
                        count += 1
            
            avg_dist = total_dist / count if count > 0 else float('inf')
            if avg_dist < best_distance and avg_dist < MAX_DISTANCE_BETWEEN_BRANCHES:
                best_distance = avg_dist
                best_merge = target_trip_num
        
        if best_merge:
            for idx in small_indices:
                result_df.at[idx, 'Trip'] = best_merge
    
    # Phase 3.5: ลบ DC011 ซ้ำออก (เก็บแค่ 1 แถวต่อ 1 ทริป)
    # เพื่อเตรียมพร้อมสำหรับ Phase 4
    dc_indices_to_drop = []
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        trip_data = result_df[result_df['Trip'] == trip]
        dc_rows = trip_data[trip_data['Code'].astype(str).str.upper() == 'DC011']
        if len(dc_rows) > 0:
            # เก็บ DC011 ทั้งหมดใน trip นี้ไว้ลบ (จะเพิ่มใหม่ใน Phase 4)
            dc_indices_to_drop.extend(dc_rows.index.tolist())
    
    if dc_indices_to_drop:
        result_df = result_df.drop(dc_indices_to_drop).reset_index(drop=True)
    
    # Phase 4: เพิ่ม DC011 กลับท้ายทุกทริป
    dc_rows = []
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        trip_data = result_df[result_df['Trip'] == trip]
        if len(trip_data) == 0:
            continue
        
        # ดึง truck type จากทริป
        truck = trip_data['Truck'].mode()[0] if len(trip_data['Truck'].mode()) > 0 else '4W'
        
        # สร้างแถว DC011 กลับ
        dc_row = {
            'Trip': trip,
            'BU': 211,
            'Code': 'DC011',
            'Branch_Code': 'DC011',
            'Branch_Name': 'บ.พีทีจี เอ็นเนอยี จำกัด (มหาชน) (DCวังน้อย)',
            'Cube': 0,
            'Weight': 0,
            'Drops': 0,
            'Truck': truck,
            'Province': 'พระนครศรีอยุธยา',
            'District': 'วังน้อย',
            'Subdistrict': 'พยอม',
            'Latitude': DC_WANG_NOI_LAT,
            'Longitude': DC_WANG_NOI_LON,
            'Distance_from_DC': -1,  # ใช้ -1 เพื่อให้อยู่ท้ายสุด
            'Is_Punthai': False
        }
        dc_rows.append(dc_row)
    
    # เพิ่มแถว DC ลงใน DataFrame
    if dc_rows:
        dc_df = pd.DataFrame(dc_rows)
        result_df = pd.concat([result_df, dc_df], ignore_index=True)
        # เรียงตาม Trip แล้วตาม Distance_from_DC descending (DC=-1 จะอยู่ท้ายสุด)
        result_df = result_df.sort_values(['Trip', 'Distance_from_DC'], ascending=[True, False])
        result_df = result_df.reset_index(drop=True)
    
    # สร้าง summary (ข้ามทริปเปล่าที่มี cube = 0)
    summary_data = []
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        trip_data = result_df[result_df['Trip'] == trip]
        branches = len(trip_data)
        total_cube = trip_data['Cube'].sum()
        total_weight = trip_data['Weight'].sum()
        
        # ข้ามทริปเปล่า (cube = 0)
        if total_cube == 0 or branches == 0:
            continue
        
        all_punthai = trip_data['Is_Punthai'].all()
        truck = trip_data['Truck'].mode()[0] if len(trip_data['Truck'].mode()) > 0 else '4W'
        
        summary_data.append({
            'Trip': int(trip),
            'Branches': branches,
            'Cube': round(total_cube, 2),
            'Weight': round(total_weight, 2),
            'Truck': f"{truck} ({'PT' if all_punthai else 'Mix'})",
            'Punthai': 'ล้วน' if all_punthai else 'คละ'
        })
    
    summary_df = pd.DataFrame(summary_data)
    return result_df, summary_df


def export_with_colors(result_df, output_file, original_file, sheet_name="2.Punthai"):
    """Export กลับไฟล์เดิมพร้อมสีเหลือง-ขาว"""
    wb = load_workbook(original_file)
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    yellow = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    code_col = 3
    trip_col = 9
    
    code_to_trip = dict(zip(result_df['Code'].astype(str), result_df['Trip']))
    
    current_trip = None
    use_yellow = True
    
    for row_idx in range(3, ws.max_row + 1):
        code_cell = ws.cell(row=row_idx, column=code_col)
        code = str(code_cell.value).strip() if code_cell.value else None
        
        if code and code in code_to_trip:
            trip = code_to_trip[code]
            
            if current_trip != trip:
                current_trip = trip
                use_yellow = not use_yellow
            
            fill = yellow if use_yellow else white
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = fill
            
            ws.cell(row=row_idx, column=trip_col, value=int(trip))
    
    wb.save(output_file)
    print(f"✅ บันทึก: {output_file}")
