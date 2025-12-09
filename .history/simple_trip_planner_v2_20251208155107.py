"""
Simple Trip Planner V2 - จัดทริปตามเงื่อนไขใหม่
เรียงตาม: ระยะทาง → ชื่อ → จังหวัด → อำเภอ → ตำบล (ไกล → ใกล้)
เริ่มจาก 6W (ไกลสุดก่อน)
เช็คระยะทางระหว่างสาขา
รวมทริปเศษให้เต็ม
"""

import pandas as pd
import numpy as np
import math
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# พิกัด DC วังน้อย
DC_WANG_NOI_LAT = 14.2682524
DC_WANG_NOI_LON = 100.8434858

# ลิมิตใหม่: 6W ≥18 คิว, JB ≤7 คิว, 4W = 5/3-4 คิว
LIMITS = {
    '6W': {'max_c': 20, 'min_c': 18, 'max_w': 9000, 'max_drops': 20},  # 6W ต้อง 18-20 คิว
    'JB': {'max_c': 7, 'max_w': 3500, 'max_drops': 7},  # JB ≤7 คิว, ≤7 drops
    '4W': {'max_c': 5, 'max_w': 1800, 'max_drops': 12}  # 4W: 5คิว(PT)/3-4คิว(Mix)
}

MAX_DISTANCE_BETWEEN_BRANCHES = 100  # km
NEAR_DC_THRESHOLD = 150  # km - บริเวณใกล้ DC ให้ใช้รถเล็ก
FAR_DC_THRESHOLD = 290  # km - บริเวณไกล DC (>290km) ให้ใช้รถใหญ่

def haversine_distance(lat1, lon1, lat2, lon2):
    """คำนวณระยะทางระหว่างสองจุด (km)"""
    R = 6371
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    return R * c

def is_punthai(branch_code):
    """เช็คว่าเป็นสาขา Punthai หรือไม่"""
    code_str = str(branch_code).upper()
    return code_str.startswith('PH') or code_str.startswith('PU') or code_str.startswith('PT')

def get_base_name(name):
    """ดึงชื่อพื้นฐานของสาขา"""
    import re
    if not name:
        return ""
    
    name_lower = str(name).strip().lower()
    
    if 'future' in name_lower or 'ฟิวเจอร์' in name_lower:
        if 'rangsit' in name_lower or 'รังสิต' in name_lower:
            return "ฟิวเจอร์รังสิต"
    
    if 'lotus' in name_lower or 'โลตัส' in name_lower:
        base = re.sub(r'\s*\d+\s*$', '', name_lower)
        return base.strip()
    
    base = re.sub(r'\s*\d+\s*$', '', str(name).strip())
    return base.strip().lower()

def plan_trips_v2(df, master_data):
    """
    จัดทริปตามเงื่อนไขใหม่:
    1. เรียงตามระยะทาง → ชื่อ → จังหวัด → อำเภอ → ตำบล (ไกล → ใกล้)
    2. เริ่มจาก 6W (ไกลสุดก่อน)
    3. เช็คระยะทางระหว่างสาขา
    4. รวมทริปเศษให้เต็ม
    """
    
    # เตรียมข้อมูล
    result_df = df.copy()
    result_df['Trip'] = 0
    result_df['Distance_DC'] = 0.0
    result_df['Base_Name'] = ''
    result_df['Province'] = ''
    result_df['District'] = ''
    result_df['Subdistrict'] = ''
    result_df['Is_Punthai'] = False
    result_df['Latitude'] = 0.0
    result_df['Longitude'] = 0.0
    result_df['Truck'] = ''
    
    # ดึงข้อมูลจาก Master
    for idx, row in result_df.iterrows():
        code = row['Code']
        
        if not master_data.empty and 'Plan Code' in master_data.columns:
            master_row = master_data[master_data['Plan Code'] == code]
            if len(master_row) > 0:
                m = master_row.iloc[0]
                lat = m.get('ละติจูด', 0) if pd.notna(m.get('ละติจูด')) else 0
                lon = m.get('ลองติจูด', 0) if pd.notna(m.get('ลองติจูด')) else 0
                
                if lat and lon and lat != 0 and lon != 0:
                    dist = haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
                    result_df.at[idx, 'Distance_DC'] = dist
                    result_df.at[idx, 'Latitude'] = lat
                    result_df.at[idx, 'Longitude'] = lon
                
                result_df.at[idx, 'Province'] = str(m.get('จังหวัด', '')).strip()
                result_df.at[idx, 'District'] = str(m.get('อำเภอ', '')).strip()
                result_df.at[idx, 'Subdistrict'] = str(m.get('ตำบล', '')).strip()
        
        if 'Name' in result_df.columns:
            result_df.at[idx, 'Base_Name'] = get_base_name(row['Name'])
        
        result_df.at[idx, 'Is_Punthai'] = is_punthai(code)
    
    # เรียงลำดับ: ไกล → ใกล้
    result_df = result_df.sort_values(
        by=['Distance_DC', 'Base_Name', 'Province', 'District', 'Subdistrict'],
        ascending=[False, True, True, True, True]
    ).reset_index(drop=True)
    
    # จัดทริป
    trip_num = 1
    trips = []
    remaining_indices = list(result_df.index)
    
    while remaining_indices:
        seed_idx = remaining_indices.pop(0)
        seed_row = result_df.iloc[seed_idx]
        
        current_trip = [seed_idx]
        current_cubes = seed_row['Cube'] if pd.notna(seed_row['Cube']) else 0
        current_weight = seed_row['Weight'] if pd.notna(seed_row['Weight']) else 0
        all_punthai = seed_row['Is_Punthai']
        last_lat = seed_row['Latitude']
        last_lon = seed_row['Longitude']
        seed_distance = seed_row['Distance_DC']
        
        # กำหนดประเภทรถตามระยะทางและคิว
        # 6W: ต้อง ≥18 คิว (ถ้าไม่ถึง 18 ให้ใช้ JB หลายคัน)
        # JB: ≤7 คิว, ≤7 drops
        # 4W: 5 คิว (PT ล้วน), 3-4 คิว (คละ)
        if seed_distance > FAR_DC_THRESHOLD:
            # ไกล (>290km) → ประเมินว่ามีคิวพอสำหรับ 6W ไหม
            estimated_cubes = current_cubes
            for idx in remaining_indices[:25]:  # ดู 25 สาขาถัดไป
                r = result_df.iloc[idx]
                if r['Distance_DC'] > FAR_DC_THRESHOLD - 50:
                    estimated_cubes += r['Cube'] if pd.notna(r['Cube']) else 0
            
            # ถ้าสาขารวมแล้ว ≥18 คิว → ใช้ 6W, ไม่งั้นใช้ JB
            if estimated_cubes >= LIMITS['6W']['min_c']:
                truck_type = '6W'
            else:
                truck_type = 'JB'
        elif seed_distance < NEAR_DC_THRESHOLD:
            # ใกล้ (<150km) → ใช้รถเล็ก
            if current_cubes > LIMITS['4W']['max_c']:
                truck_type = 'JB'
            else:
                truck_type = '4W'
        else:
            # ระยะกลาง (150-290km) → ใช้ JB
            if current_cubes > LIMITS['JB']['max_c']:
                truck_type = 'JB'  # เกิน 7 คิว → ยังใช้ JB (จะแยกเป็น 2 คันภายหลัง)
            elif current_cubes > LIMITS['4W']['max_c']:
                truck_type = 'JB'
            else:
                truck_type = '4W'
        
        # เพิ่มสาขาเข้าทริป
        indices_to_remove = []
        for idx in remaining_indices[:]:
            row = result_df.iloc[idx]
            cube = row['Cube'] if pd.notna(row['Cube']) else 0
            weight = row['Weight'] if pd.notna(row['Weight']) else 0
            is_punthai_branch = row['Is_Punthai']
            
            if all_punthai and not is_punthai_branch:
                all_punthai = False
            
            # ระยะทางจากสาขาล่าสุด
            branch_lat = row['Latitude']
            branch_lon = row['Longitude']
            distance_from_last = 0
            
            if last_lat and last_lon and branch_lat and branch_lon:
                distance_from_last = haversine_distance(last_lat, last_lon, branch_lat, branch_lon)
            
            # คำนวณ
            new_cubes = current_cubes + cube
            new_weight = current_weight + weight
            
            # กำหนด limit
            if truck_type == '6W':
                max_cubes = LIMITS['6W']['max_c']
                max_weight = LIMITS['6W']['max_w']
                max_drops = LIMITS['6W']['max_drops']
            elif truck_type == 'JB':
                max_cubes = LIMITS['JB']['max_c']
                max_weight = LIMITS['JB']['max_w']
                max_drops = LIMITS['JB']['max_drops']
            else:
                max_cubes = LIMITS['4W']['max_c'] if all_punthai else 3.5  # PT=5, Mix=3-4
                max_weight = LIMITS['4W']['max_w']
                max_drops = LIMITS['4W']['max_drops']
            
            should_add = True
            
            # เช็คเงื่อนไข
            if distance_from_last > MAX_DISTANCE_BETWEEN_BRANCHES:
                should_add = False
            elif new_cubes > max_cubes:
                # ถ้ามีรถ 6W อยู่แล้ว → ใส่ต่อไปได้ถึง 20 คิว
                if truck_type == '6W':
                    if new_cubes <= LIMITS['6W']['max_c']:
                        pass  # ใส่ได้
                    else:
                        should_add = False
                # JB: ห้ามเกิน 7 คิว (ไม่อัพเกรด)
                elif truck_type == 'JB':
                    should_add = False  # เกิน 7 คิว → ตัดทริป (ไม่อัพเกรดเป็น 6W)
                # 4W → JB
                elif truck_type == '4W' and new_cubes <= LIMITS['JB']['max_c']:
                    truck_type = 'JB'
                    max_cubes = LIMITS['JB']['max_c']
                    max_weight = LIMITS['JB']['max_w']
                    max_drops = LIMITS['JB']['max_drops']
                else:
                    should_add = False
            elif new_weight > max_weight:
                # ถ้ามีรถ 6W อยู่แล้ว → ผ่อนปรนเรื่องน้ำหนัก
                if truck_type == '6W' and new_weight <= LIMITS['6W']['max_w'] * 1.1:
                    pass  # ใส่ได้ (เกินน้ำหนักได้ 10%)
                else:
                    should_add = False
            elif len(current_trip) + 1 > max_drops:
                # ถ้ามีรถ 6W อยู่แล้ว → ใส่ได้จนถึง 20 drops
                if truck_type == '6W' and len(current_trip) + 1 <= LIMITS['6W']['max_drops']:
                    pass  # ใส่ได้
                else:
                    should_add = False
            
            if should_add:
                current_trip.append(idx)
                current_cubes = new_cubes
                current_weight = new_weight
                last_lat = branch_lat
                last_lon = branch_lon
                indices_to_remove.append(idx)
        
        for idx in indices_to_remove:
            remaining_indices.remove(idx)
        
        trips.append((trip_num, current_trip, truck_type, all_punthai))
        trip_num += 1
    
    # กำหนดเลขทริป
    for trip_num, trip_indices, truck_type, all_punthai in trips:
        for idx in trip_indices:
            result_df.at[idx, 'Trip'] = trip_num
            result_df.at[idx, 'Truck'] = truck_type
    
    # Phase 2: ตัดทริป 6W ที่ < 18 คิว ให้เป็น JB หลายคัน
    # และตัดทริป JB ที่ > 7 คิว ให้เป็น JB หลายคัน
    max_trip = int(result_df['Trip'].max())
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        
        trip_data = result_df[result_df['Trip'] == trip].copy()
        total_cubes = trip_data['Cube'].sum()
        truck = trip_data['Truck'].iloc[0]
        
        # 6W ที่ < 18 คิว → แยกเป็น JB
        if truck == '6W' and total_cubes < LIMITS['6W']['min_c']:
            # แยกเป็น JB หลายคัน (แต่ละคัน ≤7 คิว)
            new_trip_num = max_trip + 1
            current_cubes = 0
            current_drops = 0
            
            for idx in trip_data.index:
                cube = result_df.at[idx, 'Cube'] if pd.notna(result_df.at[idx, 'Cube']) else 0
                
                if (current_cubes + cube > LIMITS['JB']['max_c'] or current_drops >= LIMITS['JB']['max_drops']) and current_cubes > 0:
                    new_trip_num += 1
                    current_cubes = 0
                    current_drops = 0
                
                result_df.at[idx, 'Trip'] = new_trip_num
                result_df.at[idx, 'Truck'] = 'JB'
                current_cubes += cube
                current_drops += 1
            
            max_trip = new_trip_num
        
        # JB ที่ > 7 คิว → แยกเป็น JB หลายคัน
        elif truck == 'JB' and total_cubes > LIMITS['JB']['max_c']:
            new_trip_num = max_trip + 1
            current_cubes = 0
            current_drops = 0
            
            for idx in trip_data.index:
                cube = result_df.at[idx, 'Cube'] if pd.notna(result_df.at[idx, 'Cube']) else 0
                
                if (current_cubes + cube > LIMITS['JB']['max_c'] or current_drops >= LIMITS['JB']['max_drops']) and current_cubes > 0:
                    new_trip_num += 1
                    current_cubes = 0
                    current_drops = 0
                
                result_df.at[idx, 'Trip'] = new_trip_num
                result_df.at[idx, 'Truck'] = 'JB'
                current_cubes += cube
                current_drops += 1
            
            max_trip = new_trip_num
    
    # Phase 2.5: เปลี่ยน JB ที่ < 5 คิว เป็น 4W
    jb_to_4w_count = 0
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        
        trip_data = result_df[result_df['Trip'] == trip]
        total_cubes = trip_data['Cube'].sum()
        truck = trip_data['Truck'].iloc[0]
        all_punthai_trip = trip_data['Is_Punthai'].all()
        
        # JB < 5 คิว → เปลี่ยนเป็น 4W (ถ้าไม่เกิน 4W limit)
        if truck == 'JB' and total_cubes < 5:
            cube_limit = LIMITS['4W']['max_c'] if all_punthai_trip else 3.5
            
            # เปลี่ยนเป็น 4W ไม่เช็ค limit (เพราะ < 5 ก็ไม่เกิน limit อยู่แล้ว)
            for idx in trip_data.index:
                result_df.at[idx, 'Truck'] = '4W'
            jb_to_4w_count += 1
    
    print(f"🔄 เปลี่ยน JB → 4W: {jb_to_4w_count} ทริป")
    
    # Phase 2.6: แยกทริปที่เกิน limit (loop จนกว่าจะไม่เจอทริปเกิน)
    for round_num in range(5):  # เพิ่มเป็น 5 รอบเพื่อความแน่ใจ
        max_trip = int(result_df['Trip'].max())
        found_over_limit = False
        split_count = 0
        
        for trip in sorted(result_df['Trip'].unique()):
            if trip == 0:
                continue
            
            trip_data = result_df[result_df['Trip'] == trip].copy()
            total_cubes = trip_data['Cube'].sum()
            
            # ตรวจสอบ Truck type จากค่าที่มีมากที่สุดในทริป (ป้องกัน mixed truck)
            truck_counts = trip_data['Truck'].value_counts()
            if len(truck_counts) == 0:
                continue
            truck = truck_counts.index[0]
            
            all_punthai_trip = trip_data['Is_Punthai'].all()
            
            # กำหนด limit ตาม truck type
            if truck == '6W':
                cube_limit = LIMITS['6W']['max_c']
                drop_limit = LIMITS['6W']['max_drops']
            elif truck == 'JB':
                cube_limit = LIMITS['JB']['max_c']
                drop_limit = LIMITS['JB']['max_drops']
            elif truck == '4W':
                cube_limit = LIMITS['4W']['max_c'] if all_punthai_trip else 3.5
                drop_limit = LIMITS['4W']['max_drops']
            else:
                continue
            
            # ถ้าเกิน limit → แยกทริป
            if total_cubes > cube_limit or len(trip_data) > drop_limit:
                found_over_limit = True
                split_count += 1
                
                # คำนวณระยะทางจาก DC สำหรับแต่ละสาขาในทริป
                trip_indices = trip_data.index.tolist()
                distances = []
                for idx in trip_indices:
                    lat = result_df.at[idx, 'Latitude']
                    lon = result_df.at[idx, 'Longitude']
                    if lat and lon:
                        dist = haversine_distance(dc_lat, dc_lon, lat, lon)
                        distances.append((idx, dist))
                    else:
                        distances.append((idx, 0))
                
                # เรียงจากไกล → ใกล้ เพื่อให้เส้นทางวนกลับ DC
                distances.sort(key=lambda x: x[1], reverse=True)
                sorted_indices = [idx for idx, _ in distances]
                
                new_trip_num = max_trip + 1
                current_cubes = 0
                current_drops = 0
                prev_lat = dc_lat
                prev_lon = dc_lon
                
                for idx in trip_data.index:
                    cube = result_df.at[idx, 'Cube'] if pd.notna(result_df.at[idx, 'Cube']) else 0
                    branch_lat = result_df.at[idx, 'Latitude']
                    branch_lon = result_df.at[idx, 'Longitude']
                    
                    # คำนวณระยะทางจากสาขาก่อนหน้า
                    if branch_lat and branch_lon and prev_lat and prev_lon:
                        dist_from_prev = haversine_distance(prev_lat, prev_lon, branch_lat, branch_lon)
                    else:
                        dist_from_prev = 0
                    
                    # ถ้าใส่สาขานี้แล้วจะเกิน limit และมีสาขาอยู่แล้ว → เริ่มทริปใหม่
                    # หรือถ้าระยะห่างจากสาขาก่อนหน้า > 150km (กระโดดไกลเกินไป)
                    if (((current_cubes + cube > cube_limit) or (current_drops >= drop_limit)) and current_cubes > 0):
                        new_trip_num += 1
                        current_cubes = 0
                        current_drops = 0
                        prev_lat = dc_lat
                        prev_lon = dc_lon
                    
                    result_df.at[idx, 'Trip'] = new_trip_num
                    result_df.at[idx, 'Truck'] = truck  # เซ็ตให้แน่ใจ
                    current_cubes += cube
                    current_drops += 1
                    
                    # อัพเดทตำแหน่งปัจจุบัน
                    if branch_lat and branch_lon:
                        prev_lat = branch_lat
                        prev_lon = branch_lon
                
                max_trip = new_trip_num
        
        if split_count > 0:
            print(f"  รอบที่ {round_num + 1}: แยก {split_count} ทริปที่เกิน limit")
        
        # ถ้าไม่เจอทริปที่เกิน limit แล้ว → หยุดลูป
        if not found_over_limit:
            break
    
    # Phase 2.7: แปลง JB single-branch ที่เกิน 7 คิว → 6W
    convert_to_6w = 0
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        trip_data = result_df[result_df['Trip'] == trip]
        if len(trip_data) == 1:  # single-branch trip
            cube = trip_data['Cube'].iloc[0]
            truck = trip_data['Truck'].iloc[0]
            if truck == 'JB' and cube > 7:
                result_df.loc[result_df['Trip'] == trip, 'Truck'] = '6W'
                convert_to_6w += 1
    
    if convert_to_6w > 0:
        print(f"🔄 เปลี่ยน JB > 7 คิว (1 สาขา) → 6W: {convert_to_6w} ทริป")
    
    # Phase 3: รวมทริปเศษ (ต้อง check limit ก่อน merge)
    MIN_BRANCHES = 3
    small_trips = [(t, i, tr, p) for t, i, tr, p in trips if len(i) < MIN_BRANCHES]
    
    for small_trip_num, small_indices, small_truck, small_punthai in small_trips:
        best_merge = None
        best_distance = float('inf')
        
        for target_trip_num, target_indices, target_truck, target_punthai in trips:
            if target_trip_num == small_trip_num:
                continue
            
            # 1. ตรวจสอบจำนวน drops
            if len(target_indices) + len(small_indices) > LIMITS[target_truck]['max_drops']:
                continue
            
            # 2. คำนวณ cube + weight หลัง merge
            small_cubes = sum(result_df.at[si, 'Cube'] for si in small_indices if pd.notna(result_df.at[si, 'Cube']))
            target_cubes = sum(result_df.at[ti, 'Cube'] for ti in target_indices if pd.notna(result_df.at[ti, 'Cube']))
            small_weight = sum(result_df.at[si, 'Weight'] for si in small_indices if pd.notna(result_df.at[si, 'Weight']))
            target_weight = sum(result_df.at[ti, 'Weight'] for ti in target_indices if pd.notna(result_df.at[ti, 'Weight']))
            
            merged_cubes = small_cubes + target_cubes
            merged_weight = small_weight + target_weight
            
            # 3. เช็ค limit ตาม truck type
            target_all_punthai = all(result_df.at[ti, 'Is_Punthai'] for ti in target_indices)
            
            if target_truck == '6W':
                if merged_cubes > LIMITS['6W']['max_c'] or merged_weight > LIMITS['6W']['max_w']:
                    continue
            elif target_truck == 'JB':
                if merged_cubes > LIMITS['JB']['max_c'] or merged_weight > LIMITS['JB']['max_w']:
                    continue
            elif target_truck == '4W':
                cube_limit = LIMITS['4W']['max_c'] if target_all_punthai else 3.5
                if merged_cubes > cube_limit or merged_weight > LIMITS['4W']['max_w']:
                    continue
            
            # 4. คำนวณระยะทาง
            total_dist = 0
            count = 0
            for si in small_indices:
                s_lat = result_df.at[si, 'Latitude']
                s_lon = result_df.at[si, 'Longitude']
                for ti in target_indices[:5]:  # เช็คแค่ 5 สาขาแรก
                    t_lat = result_df.at[ti, 'Latitude']
                    t_lon = result_df.at[ti, 'Longitude']
                    if s_lat and t_lat:
                        dist = haversine_distance(s_lat, s_lon, t_lat, t_lon)
                        total_dist += dist
                        count += 1
            
            avg_dist = total_dist / count if count > 0 else float('inf')
            if avg_dist < best_distance and avg_dist < MAX_DISTANCE_BETWEEN_BRANCHES:
                best_distance = avg_dist
                best_merge = target_trip_num
        
        if best_merge:
            for idx in small_indices:
                result_df.at[idx, 'Trip'] = best_merge
    
    # สร้าง summary (ข้ามทริปเปล่าที่มี cube = 0)
    summary_data = []
    for trip in sorted(result_df['Trip'].unique()):
        if trip == 0:
            continue
        trip_data = result_df[result_df['Trip'] == trip]
        branches = len(trip_data)
        total_cube = trip_data['Cube'].sum()
        total_weight = trip_data['Weight'].sum()
        
        # ข้ามทริปเปล่า (cube = 0)
        if total_cube == 0 or branches == 0:
            continue
        
        all_punthai = trip_data['Is_Punthai'].all()
        truck = trip_data['Truck'].mode()[0] if len(trip_data['Truck'].mode()) > 0 else '4W'
        
        summary_data.append({
            'Trip': int(trip),
            'Branches': branches,
            'Cube': round(total_cube, 2),
            'Weight': round(total_weight, 2),
            'Truck': f"{truck} ({'PT' if all_punthai else 'Mix'})",
            'Punthai': 'ล้วน' if all_punthai else 'คละ'
        })
    
    summary_df = pd.DataFrame(summary_data)
    return result_df, summary_df


def export_with_colors(result_df, output_file, original_file, sheet_name="2.Punthai"):
    """Export กลับไฟล์เดิมพร้อมสีเหลือง-ขาว"""
    wb = load_workbook(original_file)
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    yellow = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    code_col = 3
    trip_col = 9
    
    code_to_trip = dict(zip(result_df['Code'].astype(str), result_df['Trip']))
    
    current_trip = None
    use_yellow = True
    
    for row_idx in range(3, ws.max_row + 1):
        code_cell = ws.cell(row=row_idx, column=code_col)
        code = str(code_cell.value).strip() if code_cell.value else None
        
        if code and code in code_to_trip:
            trip = code_to_trip[code]
            
            if current_trip != trip:
                current_trip = trip
                use_yellow = not use_yellow
            
            fill = yellow if use_yellow else white
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = fill
            
            ws.cell(row=row_idx, column=trip_col, value=int(trip))
    
    wb.save(output_file)
    print(f"✅ บันทึก: {output_file}")
