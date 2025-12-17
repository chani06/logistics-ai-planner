"""
Logistics Planner 
"""

import streamlit as st
import pandas as pd
import numpy as np
import pickle
import os
import glob
from datetime import datetime, time
import io

# Fuzzy String Matching
try:
    from rapidfuzz import fuzz, process
    FUZZY_AVAILABLE = True
except ImportError:
    FUZZY_AVAILABLE = False
    # Fallback: ‡πÉ‡∏ä‡πâ difflib
    from difflib import SequenceMatcher

# Auto-refresh component
try:
    from streamlit_autorefresh import st_autorefresh
    AUTOREFRESH_AVAILABLE = True
except ImportError:
    AUTOREFRESH_AVAILABLE = False
    st.warning("‚ö†Ô∏è ‡∏ï‡∏¥‡∏î‡∏ï‡∏±‡πâ‡∏á streamlit-autorefresh: pip install streamlit-autorefresh")

# ==========================================
# CONFIG
# ==========================================
MODEL_PATH = 'models/decision_tree_model.pkl'

# ‡∏Ç‡∏µ‡∏î‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó
LIMITS = {
    '4W': {'max_w': 2500, 'max_c': 5.0},   # ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 12 ‡∏à‡∏∏‡∏î, Cube ‚â§ 5 (Punthai ‡∏•‡πâ‡∏ß‡∏ô)
    'JB': {'max_w': 3500, 'max_c': 7.0},   # ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 12 ‡∏à‡∏∏‡∏î, Cube ‚â§ 7
    '6W': {'max_w': 6000, 'max_c': 20.0}   # ‡πÑ‡∏°‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏à‡∏∏‡∏î, Cube ‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏ï‡πá‡∏°, Weight ‚â§ 6000
}

# üîí ‡∏Ç‡∏µ‡∏î‡∏à‡∏≥‡∏Å‡∏±‡∏î 4W ‡∏ï‡∏≤‡∏° BU (Punthai ‡∏•‡πâ‡∏ß‡∏ô vs ‡∏ú‡∏™‡∏°)
# - Punthai ‡∏•‡πâ‡∏ß‡∏ô: 4W max_c = 5.0
# - ‡∏ú‡∏™‡∏° (Punthai + ‡∏≠‡∏∑‡πà‡∏ô): 4W max_c = 4.0 (‡∏ñ‡πâ‡∏≤ cube 3-4 ‡πÉ‡∏ä‡πâ 4W, ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô ‚Üí JB)
LIMITS_4W_PUNTHAI_ONLY = 5.0   # Punthai ‡∏•‡πâ‡∏ß‡∏ô
LIMITS_4W_MIXED = 4.0          # ‡∏ú‡∏™‡∏° BU

# üîí ‡∏Ç‡∏µ‡∏î‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö Punthai ‡∏•‡πâ‡∏ß‡∏ô
# - JB (Jumbo): ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 7 drop, Cube ‚â§ 8
# - 4W: ‡∏ñ‡πâ‡∏≤ Cube > 5 ‚Üí ‡∏ï‡∏±‡∏î‡πÄ‡∏õ‡πá‡∏ô JB
# - ‡∏ñ‡πâ‡∏≤‡∏ú‡∏™‡∏° (Punthai + ‡∏≠‡∏∑‡πà‡∏ô): ‡∏ñ‡πâ‡∏≤ Cube 3-4 ‚Üí ‡πÉ‡∏ä‡πâ 4W ‡πÑ‡∏î‡πâ, ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô ‚Üí JB
PUNTHAI_LIMITS = {
    '4W': {'max_w': 2500, 'max_c': 5.0, 'max_drops': 12},  # Punthai ‡∏•‡πâ‡∏ß‡∏ô: ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 5 cube ‚Üí ‡πÉ‡∏ä‡πâ JB
    'JB': {'max_w': 3500, 'max_c': 7.0, 'max_drops': 7},   # Punthai ‡∏•‡πâ‡∏ß‡∏ô: ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 7 drop
    '6W': {'max_w': 6000, 'max_c': 20.0, 'max_drops': 999}
}

# üö® ‡∏´‡πâ‡∏≤‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 100% - ‡πÑ‡∏°‡πà‡∏°‡∏µ Buffer
BUFFER = 1.0

# üö® ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
MAX_DISTANCE_IN_TRIP = 50  # km - ‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡∏´‡πâ‡∏≤‡∏°‡∏´‡πà‡∏≤‡∏á‡∏Å‡∏±‡∏ô‡πÄ‡∏Å‡∏¥‡∏ô 50km

# üéØ Minimum utilization ‡∏ï‡πà‡∏≠‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ (‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö balancing)
MIN_UTIL = {
    '4W': 70,   # 4W ‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏ô‡πâ‡∏≠‡∏¢ 70%
    'JB': 80,   # JB ‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏ô‡πâ‡∏≠‡∏¢ 80%
    '6W': 90    # 6W ‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏ô‡πâ‡∏≠‡∏¢ 90%
}

# ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡πà‡∏≠‡∏ó‡∏£‡∏¥‡∏õ - ‡πÉ‡∏ä‡πâ‡∏Å‡∏±‡∏ö 4W/JB ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô (6W ‡πÑ‡∏°‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î)
MAX_BRANCHES_PER_TRIP = 12  # ‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î 12 ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡πà‡∏≠‡∏ó‡∏£‡∏¥‡∏õ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö 4W/JB (6W ‡πÑ‡∏°‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î)
TARGET_BRANCHES_PER_TRIP = 12  # ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢ 12 ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡πà‡∏≠‡∏ó‡∏£‡∏¥‡∏õ

# Performance Config - Optimized for < 1 minute
MAX_DETOUR_KM = 10  # ‡∏•‡∏î‡∏à‡∏≤‡∏Å 12km ‡πÄ‡∏õ‡πá‡∏ô 10km
MAX_MERGE_ITERATIONS = 5  # ‡∏•‡∏î‡∏à‡∏≤‡∏Å 10 ‡πÄ‡∏õ‡πá‡∏ô 5 ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÄ‡∏£‡πá‡∏ß‡∏Ç‡∏∂‡πâ‡∏ô
MAX_REBALANCE_ITERATIONS = 3  # ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏Å‡∏≤‡∏£ rebalance
MAX_PROCESSING_TIME = 55  # ‡∏ß‡∏¥‡∏ô‡∏≤‡∏ó‡∏µ - ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢ < 1 minute
EARLY_STOP_UTIL = 95  # ‡∏´‡∏¢‡∏∏‡∏î‡∏ñ‡πâ‡∏≤‡πÑ‡∏î‡πâ utilization >= 95%
MAX_REBALANCE_ITERATIONS = 3  # ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏Å‡∏≤‡∏£ rebalance
EARLY_STOP_THRESHOLD = 0.95  # ‡∏´‡∏¢‡∏∏‡∏î‡∏ñ‡πâ‡∏≤‡πÑ‡∏î‡πâ utilization >= 95%

# ‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏™‡πà‡∏á (‡∏ï‡∏±‡∏î‡∏≠‡∏≠‡∏Å)
EXCLUDE_BRANCHES = ['DC011', 'PTDC', 'PTG DISTRIBUTION CENTER']

# ‡∏£‡∏≤‡∏¢‡∏ä‡∏∑‡πà‡∏≠‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏ï‡∏±‡∏î‡∏≠‡∏≠‡∏Å (‡πÉ‡∏ä‡πâ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ä‡∏∑‡πà‡∏≠)
EXCLUDE_NAMES = ['Distribution Center', 'PTG Distribution', '‡∏ö.‡∏û‡∏µ‡∏ó‡∏µ‡∏à‡∏µ ‡πÄ‡∏≠‡πá‡∏ô‡πÄ‡∏ô‡∏≠‡∏¢‡∏µ']

# üîí ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå Auto Plan (‡∏ä‡∏µ‡∏ï info) - MaxTruckType
# ‡∏à‡∏∞‡∏ñ‡∏π‡∏Å populate ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡∏ó‡∏µ‡πà‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î
AUTO_PLAN_TRUCK_LIMITS = {}  # {branch_code: max_truck_type} ‡πÄ‡∏ä‡πà‡∏ô {'11005514': '4W', 'G015': 'JB'}

# ‡∏û‡∏¥‡∏Å‡∏±‡∏î DC ‡∏ß‡∏±‡∏á‡∏ô‡πâ‡∏≠‡∏¢ (‡∏à‡∏∏‡∏î‡∏Å‡∏•‡∏≤‡∏á)
DC_WANG_NOI_LAT = 14.179394
DC_WANG_NOI_LON = 100.648149

# ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ‡∏£‡∏ñ 6W (‡∏Å‡∏°.)
DISTANCE_REQUIRE_6W = 100  # ‡∏ñ‡πâ‡∏≤‡∏´‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å DC ‡πÄ‡∏Å‡∏¥‡∏ô 100 ‡∏Å‡∏°. ‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ 6W

# ==========================================
# LOAD MASTER DATA
# ==========================================
@st.cache_data(ttl=7200)  # Cache 2 ‡∏ä‡∏±‡πà‡∏ß‡πÇ‡∏°‡∏á (‡πÄ‡∏£‡πá‡∏ß‡∏Ç‡∏∂‡πâ‡∏ô)
def load_master_data():
    """‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Master ‡∏™‡∏ñ‡∏≤‡∏ô‡∏ó‡∏µ‡πà‡∏™‡πà‡∏á (Optimized)"""
    try:
        # ‡πÇ‡∏´‡∏•‡∏î‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ó‡∏µ‡πà‡∏à‡∏≥‡πÄ‡∏õ‡πá‡∏ô
        usecols = ['Plan Code', '‡∏ï‡∏≥‡∏ö‡∏•', '‡∏≠‡∏≥‡πÄ‡∏†‡∏≠', '‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', '‡∏•‡∏∞‡∏ï‡∏¥‡∏à‡∏π‡∏î', '‡∏•‡∏≠‡∏á‡∏ï‡∏¥‡∏à‡∏π‡∏î']
        df_master = pd.read_excel('Dc/Master ‡∏™‡∏ñ‡∏≤‡∏ô‡∏ó‡∏µ‡πà‡∏™‡πà‡∏á.xlsx', usecols=usecols)
        # ‡∏ó‡∏≥‡∏Ñ‡∏ß‡∏≤‡∏°‡∏™‡∏∞‡∏≠‡∏≤‡∏î Plan Code (vectorized)
        if 'Plan Code' in df_master.columns:
            df_master['Plan Code'] = df_master['Plan Code'].astype(str).str.strip().str.upper()
        # ‡∏™‡∏£‡πâ‡∏≤‡∏á dict ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏Ñ‡πâ‡∏ô‡∏´‡∏≤‡πÄ‡∏£‡πá‡∏ß
        df_master = df_master[df_master['Plan Code'] != '']
        return df_master
    except FileNotFoundError:
        return pd.DataFrame()
    except Exception as e:
        st.warning(f"‡πÑ‡∏°‡πà‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Master: {e} (‡∏à‡∏∞‡πÉ‡∏ä‡πâ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÅ‡∏ó‡∏ô)")
        return pd.DataFrame()

# ‡πÇ‡∏´‡∏•‡∏î Master Data
MASTER_DATA = load_master_data()

@st.cache_data(ttl=3600)  # Cache 1 ‡∏ä‡∏±‡πà‡∏ß‡πÇ‡∏°‡∏á
def load_booking_history_restrictions():
    """‡πÇ‡∏´‡∏•‡∏î‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏™‡πà‡∏á‡∏à‡∏≤‡∏Å Booking History - ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏à‡∏£‡∏¥‡∏á 3,053 booking (Optimized)"""
    try:
        # ‡∏•‡∏≠‡∏á‡∏´‡∏≤‡πÑ‡∏ü‡∏•‡πå Booking History (‡∏≠‡∏≤‡∏à‡∏°‡∏µ‡∏ä‡∏∑‡πà‡∏≠‡∏´‡∏•‡∏≤‡∏¢‡πÅ‡∏ö‡∏ö)
        possible_files = [
            'Dc/‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥‡∏á‡∏≤‡∏ô‡∏à‡∏±‡∏î‡∏™‡πà‡∏á DC ‡∏ß‡∏±‡∏á‡∏ô‡πâ‡∏≠‡∏¢(1).xlsx',
            'Dc/‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥‡∏á‡∏≤‡∏ô‡∏à‡∏±‡∏î‡∏™‡πà‡∏á DC ‡∏ß‡∏±‡∏á‡∏ô‡πâ‡∏≠‡∏¢.xlsx',
            'branch_vehicle_restrictions_from_booking.xlsx'
        ]
        
        file_path = None
        for path in possible_files:
            if os.path.exists(path):
                file_path = path
                break
        
        if not file_path:
            # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡πÑ‡∏ü‡∏•‡πå ‡πÉ‡∏ä‡πâ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ó‡∏µ‡πà‡πÄ‡∏Ñ‡∏¢‡πÄ‡∏£‡∏µ‡∏¢‡∏ô‡∏£‡∏π‡πâ (fallback)
            return load_learned_restrictions_fallback()
        
        df = pd.read_excel(file_path)
        
        # ‡πÅ‡∏õ‡∏•‡∏á‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ
        vehicle_mapping = {
            '4 ‡∏•‡πâ‡∏≠ ‡∏à‡∏±‡∏°‡πÇ‡∏ö‡πâ ‡∏ï‡∏π‡πâ‡∏ó‡∏∂‡∏ö': 'JB',
            '6 ‡∏•‡πâ‡∏≠ ‡∏ï‡∏π‡πâ‡∏ó‡∏∂‡∏ö': '6W',
            '4 ‡∏•‡πâ‡∏≠ ‡∏ï‡∏π‡πâ‡∏ó‡∏∂‡∏ö': '4W'
        }
        df['Vehicle_Type'] = df['‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ'].map(vehicle_mapping)
        
        # ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå‡∏Ñ‡∏ß‡∏≤‡∏°‡∏™‡∏±‡∏°‡∏û‡∏±‡∏ô‡∏ò‡πå‡∏™‡∏≤‡∏Ç‡∏≤-‡∏£‡∏ñ
        branch_vehicle_history = {}
        booking_groups = df.groupby('Booking No')
        
        for booking_no, booking_data in booking_groups:
            vehicle_types = booking_data['Vehicle_Type'].dropna().unique()
            if len(vehicle_types) > 0:
                vehicle = booking_data['Vehicle_Type'].mode()[0] if len(booking_data['Vehicle_Type'].mode()) > 0 else vehicle_types[0]
                for branch_code in booking_data['‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤'].dropna().unique():
                    if branch_code not in branch_vehicle_history:
                        branch_vehicle_history[branch_code] = []
                    branch_vehicle_history[branch_code].append(vehicle)
        
        # ‡∏™‡∏£‡πâ‡∏≤‡∏á restrictions
        branch_restrictions = {}
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for branch_code, vehicle_list in branch_vehicle_history.items():
            vehicles_used = set(vehicle_list)
            vehicle_counts = pd.Series(vehicle_list).value_counts().to_dict()
            
            if len(vehicles_used) == 1:
                # STRICT - ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß
                vehicle = list(vehicles_used)[0]
                branch_restrictions[str(branch_code)] = {
                    'max_vehicle': vehicle,
                    'allowed': [vehicle],
                    'total_bookings': len(vehicle_list),
                    'restriction_type': 'STRICT'
                }
            else:
                # FLEXIBLE - ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡∏´‡∏•‡∏≤‡∏¢‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó
                max_vehicle = max(vehicles_used, key=lambda v: vehicle_sizes.get(v, 0))
                branch_restrictions[str(branch_code)] = {
                    'max_vehicle': max_vehicle,
                    'allowed': list(vehicles_used),
                    'total_bookings': len(vehicle_list),
                    'restriction_type': 'FLEXIBLE'
                }
        
        stats = {
            'total_branches': len(branch_restrictions),
            'strict': len([b for b, r in branch_restrictions.items() if r['restriction_type'] == 'STRICT']),
            'flexible': len([b for b, r in branch_restrictions.items() if r['restriction_type'] == 'FLEXIBLE']),
            'total_bookings': len(booking_groups)
        }
        
        return {
            'branch_restrictions': branch_restrictions,
            'stats': stats
        }
    except Exception as e:
        # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏î error ‡πÉ‡∏ä‡πâ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ó‡∏µ‡πà‡πÄ‡∏Ñ‡∏¢‡πÄ‡∏£‡∏µ‡∏¢‡∏ô‡∏£‡∏π‡πâ‡πÅ‡∏ó‡∏ô
        return load_learned_restrictions_fallback()

def load_learned_restrictions_fallback():
    """
    ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ó‡∏µ‡πà‡πÄ‡∏£‡∏µ‡∏¢‡∏ô‡∏£‡∏π‡πâ‡∏à‡∏≤‡∏Å Booking History (backup)
    ‡πÉ‡∏ä‡πâ‡πÄ‡∏°‡∏∑‡πà‡∏≠‡πÑ‡∏°‡πà‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡πÑ‡∏î‡πâ
    
    ‡∏à‡∏≤‡∏Å‡∏Å‡∏≤‡∏£‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå 3,053 bookings, 2,790 ‡∏™‡∏≤‡∏Ç‡∏≤:
    - JB: ‡∏£‡∏ñ‡∏Å‡∏•‡∏≤‡∏á (‡πÉ‡∏ä‡πâ‡∏°‡∏≤‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î 54.7%)
    - 6W: ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà (30.1%)
    - 4W: ‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å (0.2%)
    
    ‡∏Å‡∏•‡∏¢‡∏∏‡∏ó‡∏ò‡πå: ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• default ‡πÄ‡∏õ‡πá‡∏ô JB (‡∏£‡∏ñ‡∏Å‡∏•‡∏≤‡∏á ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡∏Å‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏™‡πà‡∏ß‡∏ô‡πÉ‡∏´‡∏ç‡πà)
    """
    return {
        'branch_restrictions': {},
        'stats': {
            'total_branches': 0,
            'strict': 0,
            'flexible': 0,
            'total_bookings': 0,
            'fallback': True,
            'message': '‡πÉ‡∏ä‡πâ Punthai ‡πÄ‡∏õ‡πá‡∏ô‡∏´‡∏•‡∏±‡∏Å (‡πÑ‡∏°‡πà‡∏û‡∏ö‡πÑ‡∏ü‡∏•‡πå Booking History)'
        }
    }

@st.cache_data(ttl=3600)  # Cache 1 ‡∏ä‡∏±‡πà‡∏ß‡πÇ‡∏°‡∏á
def load_punthai_reference():
    """‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Punthai Maxmart ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÄ‡∏£‡∏µ‡∏¢‡∏ô‡∏£‡∏π‡πâ‡∏´‡∏•‡∏±‡∏Å‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ (Location patterns - Optimized)"""
    try:
        file_path = 'Dc/‡πÅ‡∏ú‡∏ô‡∏á‡∏≤‡∏ô Punthai Maxmart ‡∏£‡∏≠‡∏ö‡∏™‡∏±‡πà‡∏á 24‡∏´‡∏¢‡∏¥‡∏ö 25‡∏û‡∏§‡∏®‡∏à‡∏¥‡∏Å‡∏≤‡∏¢‡∏ô 2568 To.‡πÄ‡∏ü‡∏¥(1) - ‡∏™‡∏≥‡πÄ‡∏ô‡∏≤.xlsx'
        df = pd.read_excel(file_path, sheet_name='2.Punthai', header=1)
        
        # ‡∏Å‡∏£‡∏≠‡∏á‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡πÅ‡∏ñ‡∏ß‡∏ó‡∏µ‡πà‡∏°‡∏µ Trip ‡πÅ‡∏•‡∏∞‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πà DC/Distribution Center
        df_clean = df[df['Trip'].notna()].copy()
        df_clean = df_clean[~df_clean['BranchCode'].isin(['DC011', 'PTDC', 'PTG Distribution Center'])].copy()
        
        # Extract vehicle type from Trip no (‡πÄ‡∏ä‡πà‡∏ô 4W009 ‚Üí 4W)
        df_clean['Vehicle_Type'] = df_clean['Trip no'].apply(
            lambda x: str(x)[:2] if pd.notna(x) else 'Unknown'
        )
        
        # Merge ‡∏Å‡∏±‡∏ö Master ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÑ‡∏î‡πâ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ï‡∏≥‡∏ö‡∏•/‡∏≠‡∏≥‡πÄ‡∏†‡∏≠/‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î
        try:
            df_master = pd.read_excel('Dc/Master ‡∏™‡∏ñ‡∏≤‡∏ô‡∏ó‡∏µ‡πà‡∏™‡πà‡∏á.xlsx')
            df_clean = df_clean.merge(
                df_master[['Plan Code', '‡∏ï‡∏≥‡∏ö‡∏•', '‡∏≠‡∏≥‡πÄ‡∏†‡∏≠', '‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î']],
                left_on='BranchCode',
                right_on='Plan Code',
                how='left'
            )
        except:
            pass
        
        # ‡πÄ‡∏£‡∏µ‡∏¢‡∏ô‡∏£‡∏π‡πâ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡∏à‡∏≤‡∏Å Punthai (‡πÅ‡∏ú‡∏ô) - ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏°‡∏µ‡πÉ‡∏ô Booking
        punthai_restrictions = {}
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for branch_code in df_clean['BranchCode'].unique():
            branch_data = df_clean[df_clean['BranchCode'] == branch_code]
            vehicles_used = set(branch_data['Vehicle_Type'].dropna().tolist())
            vehicles_used = {v for v in vehicles_used if v in ['4W', 'JB', '6W']}
            
            if vehicles_used:
                if len(vehicles_used) == 1:
                    vehicle = list(vehicles_used)[0]
                    punthai_restrictions[str(branch_code)] = {
                        'max_vehicle': vehicle,
                        'allowed': [vehicle],
                        'source': 'PUNTHAI'
                    }
                else:
                    max_vehicle = max(vehicles_used, key=lambda v: vehicle_sizes.get(v, 0))
                    punthai_restrictions[str(branch_code)] = {
                        'max_vehicle': max_vehicle,
                        'allowed': list(vehicles_used),
                        'source': 'PUNTHAI'
                    }
        
        # ‡∏™‡∏£‡πâ‡∏≤‡∏á dictionary: Trip ‚Üí ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• (location patterns)
        trip_patterns = {}
        location_stats = {
            'same_province': 0,
            'mixed_province': 0,
            'avg_branches': 0
        }
        
        for trip_num in df_clean['Trip'].unique():
            trip_data = df_clean[df_clean['Trip'] == trip_num]
            
            # Get location info
            provinces = set(trip_data['‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î'].dropna().tolist()) if '‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î' in trip_data.columns else set()
            
            # Count same vs mixed province
            if len(provinces) == 1:
                location_stats['same_province'] += 1
            elif len(provinces) > 1:
                location_stats['mixed_province'] += 1
            
            trip_patterns[int(trip_num)] = {
                'branches': len(trip_data),
                'codes': trip_data['BranchCode'].tolist(),
                'weight': trip_data['TOTALWGT'].sum() if 'TOTALWGT' in trip_data.columns else 0,
                'cube': trip_data['TOTALCUBE'].sum() if 'TOTALCUBE' in trip_data.columns else 0,
                'provinces': list(provinces),
                'same_province': len(provinces) == 1
            }
        
        # Calculate stats
        if trip_patterns:
            location_stats['avg_branches'] = sum(t['branches'] for t in trip_patterns.values()) / len(trip_patterns)
            total = location_stats['same_province'] + location_stats['mixed_province']
            location_stats['same_province_pct'] = (location_stats['same_province'] / total * 100) if total > 0 else 0
        
        return {
            'patterns': trip_patterns, 
            'stats': location_stats,
            'punthai_restrictions': punthai_restrictions
        }
    except:
        return {'patterns': {}, 'stats': {}, 'punthai_restrictions': {}}

# ‡πÇ‡∏´‡∏•‡∏î Booking History (‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ)
BOOKING_RESTRICTIONS = load_booking_history_restrictions()

# ‡πÇ‡∏´‡∏•‡∏î Punthai Reference (location patterns)
PUNTHAI_PATTERNS = load_punthai_reference()

# ==========================================
# HELPER FUNCTIONS
# ==========================================
def normalize(val):
    """‡∏ó‡∏≥‡πÉ‡∏´‡πâ‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏õ‡πá‡∏ô‡∏°‡∏≤‡∏ï‡∏£‡∏ê‡∏≤‡∏ô"""
    return str(val).strip().upper().replace(" ", "").replace(".0", "")

def calculate_distance(lat1, lon1, lat2, lon2):
    """‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏£‡∏∞‡∏´‡∏ß‡πà‡∏≤‡∏á‡∏™‡∏≠‡∏á‡∏à‡∏∏‡∏î (‡∏Å‡∏°.) - Haversine formula"""
    if lat1 == 0 or lon1 == 0 or lat2 == 0 or lon2 == 0:
        return 0
    import math
    lat1_rad, lon1_rad = math.radians(lat1), math.radians(lon1)
    lat2_rad, lon2_rad = math.radians(lat2), math.radians(lon2)
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    return 6371 * c

def calculate_distance_from_dc(lat, lon):
    """‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å DC ‡∏ß‡∏±‡∏á‡∏ô‡πâ‡∏≠‡∏¢ (‡∏Å‡∏°.)"""
    return calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)

def check_branch_vehicle_compatibility(branch_code, vehicle_type):
    """‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏ô‡∏µ‡πâ‡πÑ‡∏î‡πâ‡πÑ‡∏´‡∏° (‡∏£‡∏ß‡∏° Booking + Punthai)"""
    branch_code_str = str(branch_code).strip()
    
    # 1. ‡∏•‡∏≠‡∏á‡∏´‡∏≤‡∏à‡∏≤‡∏Å Booking History ‡∏Å‡πà‡∏≠‡∏ô (‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏à‡∏£‡∏¥‡∏á)
    booking_restrictions = BOOKING_RESTRICTIONS.get('branch_restrictions', {})
    if branch_code_str in booking_restrictions:
        allowed = booking_restrictions[branch_code_str].get('allowed', [])
        return vehicle_type in allowed
    
    # 2. ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ ‡∏•‡∏≠‡∏á‡∏´‡∏≤‡∏à‡∏≤‡∏Å Punthai (‡πÅ‡∏ú‡∏ô)
    punthai_restrictions = PUNTHAI_PATTERNS.get('punthai_restrictions', {})
    if branch_code_str in punthai_restrictions:
        allowed = punthai_restrictions[branch_code_str].get('allowed', [])
        return vehicle_type in allowed
    
    # 3. ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• = ‡∏¢‡∏∑‡∏î‡∏´‡∏¢‡∏∏‡πà‡∏ô
    return True

def is_punthai_only(trip_data):
    """
    ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ô‡∏µ‡πâ‡πÄ‡∏õ‡πá‡∏ô Punthai ‡∏•‡πâ‡∏ß‡∏ô‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
    
    Returns:
        'punthai_only': ‡∏ñ‡πâ‡∏≤‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î‡πÄ‡∏õ‡πá‡∏ô Punthai (BU = 211 ‡∏´‡∏£‡∏∑‡∏≠‡∏ä‡∏∑‡πà‡∏≠‡∏°‡∏µ PUNTHAI)
        'mixed': ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏ó‡∏±‡πâ‡∏á Punthai ‡πÅ‡∏•‡∏∞‡∏≠‡∏∑‡πà‡∏ô
        'other': ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ Punthai ‡πÄ‡∏•‡∏¢
    """
    if trip_data is None or len(trip_data) == 0:
        return 'other'
    
    punthai_count = 0
    total_count = len(trip_data)
    
    for _, row in trip_data.iterrows():
        bu = row.get('BU', None)
        name = str(row.get('Name', '')).upper()
        
        # ‡πÄ‡∏ä‡πá‡∏Ñ BU = 211 ‡∏´‡∏£‡∏∑‡∏≠‡∏ä‡∏∑‡πà‡∏≠‡∏°‡∏µ PUNTHAI
        if bu == 211 or bu == '211' or 'PUNTHAI' in name:
            punthai_count += 1
    
    if punthai_count == total_count:
        return 'punthai_only'
    elif punthai_count > 0:
        return 'mixed'
    else:
        return 'other'

def get_punthai_vehicle_limits(trip_data, total_cube, branch_count):
    """
    ‡∏î‡∏∂‡∏á‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö Punthai
    
    ‡∏Å‡∏é:
    - Punthai ‡∏•‡πâ‡∏ß‡∏ô + JB: ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 7 drop
    - Punthai ‡∏•‡πâ‡∏ß‡∏ô + 4W: ‡∏ñ‡πâ‡∏≤ Cube > 5 ‚Üí ‡∏ï‡∏±‡∏î‡πÄ‡∏õ‡πá‡∏ô JB
    - ‡∏ú‡∏™‡∏° (Punthai + ‡∏≠‡∏∑‡πà‡∏ô): ‡∏ñ‡πâ‡∏≤ Cube 3-4 ‚Üí 6W ‡πÑ‡∏î‡πâ, ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô ‚Üí ‡∏ï‡∏±‡∏î‡πÄ‡∏õ‡πá‡∏ô 4W ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
    
    Returns:
        dict: {'max_vehicle': '4W'/'JB'/'6W', 'max_drops': int, 'should_split': bool}
    """
    punthai_type = is_punthai_only(trip_data)
    
    if punthai_type == 'punthai_only':
        # Punthai ‡∏•‡πâ‡∏ß‡∏ô
        if total_cube > 5.0:
            # Cube ‡πÄ‡∏Å‡∏¥‡∏ô 5 ‚Üí ‡πÉ‡∏ä‡πâ JB (‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πâ 4W)
            if branch_count > 7:
                # ‡πÄ‡∏Å‡∏¥‡∏ô 7 drop ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å
                return {'max_vehicle': 'JB', 'max_drops': 7, 'should_split': True}
            else:
                return {'max_vehicle': 'JB', 'max_drops': 7, 'should_split': False}
        else:
            # Cube ‚â§ 5 ‚Üí ‡πÉ‡∏ä‡πâ 4W ‡πÑ‡∏î‡πâ
            return {'max_vehicle': '4W', 'max_drops': 12, 'should_split': False}
    
    elif punthai_type == 'mixed':
        # ‡∏ú‡∏™‡∏° (Punthai + ‡∏≠‡∏∑‡πà‡∏ô): 4W max_c = 4.0
        if total_cube <= 4.0:
            # Cube ‚â§ 4 ‚Üí ‡πÉ‡∏ä‡πâ 4W ‡πÑ‡∏î‡πâ
            return {'max_vehicle': '4W', 'max_drops': 12, 'should_split': False}
        elif total_cube <= 7.0:
            # Cube 4-7 ‚Üí ‡πÉ‡∏ä‡πâ JB
            return {'max_vehicle': 'JB', 'max_drops': 12, 'should_split': False}
        else:
            # Cube > 7 ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å
            return {'max_vehicle': 'JB', 'max_drops': 12, 'should_split': True}
    
    else:
        # ‡πÑ‡∏°‡πà‡∏°‡∏µ Punthai ‚Üí ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏û‡∏¥‡πÄ‡∏®‡∏©
        return {'max_vehicle': '6W', 'max_drops': 999, 'should_split': False}

def get_max_vehicle_for_branch(branch_code):
    """‡∏î‡∏∂‡∏á‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ‡∏£‡∏≠‡∏á‡∏£‡∏±‡∏ö (Auto Plan > Booking History > Punthai)"""
    branch_code_str = str(branch_code).strip()
    
    # üîí 1. ‡∏•‡∏≠‡∏á‡∏´‡∏≤‡∏à‡∏≤‡∏Å Auto Plan (‡∏ä‡∏µ‡∏ï info - MaxTruckType) ‡∏Å‡πà‡∏≠‡∏ô (‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î!)
    if branch_code_str in AUTO_PLAN_TRUCK_LIMITS:
        return AUTO_PLAN_TRUCK_LIMITS[branch_code_str]
    
    # 2. ‡∏•‡∏≠‡∏á‡∏´‡∏≤‡∏à‡∏≤‡∏Å Booking History (‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏à‡∏£‡∏¥‡∏á - ‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏ä‡∏∑‡πà‡∏≠‡∏°‡∏±‡πà‡∏ô‡∏™‡∏π‡∏á)
    booking_restrictions = BOOKING_RESTRICTIONS.get('branch_restrictions', {})
    if branch_code_str in booking_restrictions:
        return booking_restrictions[branch_code_str].get('max_vehicle', '6W')
    
    # 3. ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ ‡∏•‡∏≠‡∏á‡∏´‡∏≤‡∏à‡∏≤‡∏Å Punthai (‡πÅ‡∏ú‡∏ô - ‡∏™‡∏≥‡∏£‡∏≠‡∏á)
    punthai_restrictions = PUNTHAI_PATTERNS.get('punthai_restrictions', {})
    if branch_code_str in punthai_restrictions:
        return punthai_restrictions[branch_code_str].get('max_vehicle', '6W')
    
    # 4. ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î = ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡πÑ‡∏î‡πâ
    return '6W'

def get_max_vehicle_for_trip(trip_codes):
    """
    ‡∏´‡∏≤‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏ó‡∏£‡∏¥‡∏õ‡∏ô‡∏µ‡πâ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ (‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏Ç‡∏≠‡∏á‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ)
    
    Args:
        trip_codes: set ‡∏Ç‡∏≠‡∏á branch codes ‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
    
    Returns:
        str: '4W', 'JB', ‡∏´‡∏£‡∏∑‡∏≠ '6W'
    """
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    max_allowed = '6W'  # ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å‡πÉ‡∏´‡∏ç‡πà‡∏™‡∏∏‡∏î ‡πÅ‡∏•‡πâ‡∏ß‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏ï‡∏≤‡∏°‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤
    min_priority = 3  # ‡∏Ñ‡πà‡∏≤‡πÉ‡∏´‡∏ç‡πà‡∏™‡∏∏‡∏î‡∏Ñ‡∏∑‡∏≠‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î
    
    for code in trip_codes:
        branch_max = get_max_vehicle_for_branch(code)
        priority = vehicle_priority.get(branch_max, 3)
        
        # üîí ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏•‡πá‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î (‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏°‡∏≤‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î) ‡∏à‡∏≤‡∏Å‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
        if priority < min_priority:
            min_priority = priority
            max_allowed = branch_max
    
    return max_allowed

def get_required_vehicle_by_distance(branch_code):
    """‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏≠‡∏∞‡πÑ‡∏£‡∏ï‡∏≤‡∏°‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å DC"""
    # ‡∏î‡∏∂‡∏á‡∏û‡∏¥‡∏Å‡∏±‡∏î‡∏à‡∏≤‡∏Å Master
    if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
        master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == branch_code]
        if len(master_row) > 0:
            lat = master_row.iloc[0].get('‡∏•‡∏∞‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
            lon = master_row.iloc[0].get('‡∏•‡∏≠‡∏á‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
            distance = calculate_distance_from_dc(lat, lon)
            
            # ‡∏ñ‡πâ‡∏≤‡∏´‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å DC ‡πÄ‡∏Å‡∏¥‡∏ô‡∏Å‡∏≥‡∏´‡∏ô‡∏î ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ 6W
            if distance > DISTANCE_REQUIRE_6W:
                return '6W', distance
    
    return None, 0

def can_fit_truck(total_weight, total_cube, truck_type):
    """‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å/‡∏Ñ‡∏¥‡∏ß‡πÉ‡∏™‡πà‡∏£‡∏ñ‡πÑ‡∏î‡πâ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà"""
    limits = LIMITS[truck_type]
    max_w = limits['max_w'] * BUFFER
    max_c = limits['max_c'] * BUFFER
    return total_weight <= max_w and total_cube <= max_c

def suggest_truck(total_weight, total_cube, max_allowed='6W', trip_codes=None):
    """
    ‡πÅ‡∏ô‡∏∞‡∏ô‡∏≥‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏° ‡πÇ‡∏î‡∏¢‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ‡∏ó‡∏µ‡πà:
    1. ‡πÉ‡∏™‡πà‡∏Ç‡∏≠‡∏á‡πÑ‡∏î‡πâ‡∏û‡∏≠‡∏î‡∏µ (‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô‡∏Ç‡∏µ‡∏î‡∏à‡∏≥‡∏Å‡∏±‡∏î 105%)
    2. ‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô‡πÑ‡∏î‡πâ‡πÉ‡∏Å‡∏•‡πâ 100% ‡∏°‡∏≤‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î (‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢: 90-100%)
    3. ‡πÄ‡∏Ñ‡∏≤‡∏£‡∏û‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏Ç‡∏≠‡∏á‡∏™‡∏≤‡∏Ç‡∏≤ (‡∏ñ‡πâ‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ä‡πâ‡πÅ‡∏Ñ‡πà 4W = ‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ 4W ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô)
    """
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    max_size = vehicle_sizes.get(max_allowed, 3)
    
    # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏Ç‡∏≠‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î‡πÉ‡∏ô‡∏Å‡∏•‡∏∏‡πà‡∏°
    branch_max_vehicle = '4W'  # üîí ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏ï‡πâ‡∏ô‡∏ó‡∏µ‡πà 4W (‡πÄ‡∏•‡πá‡∏Å‡∏™‡∏∏‡∏î) ‡πÅ‡∏•‡πâ‡∏ß‡∏Ç‡∏¢‡∏≤‡∏¢‡πÄ‡∏°‡∏∑‡πà‡∏≠‡∏à‡∏≥‡πÄ‡∏õ‡πá‡∏ô
    if trip_codes is not None and len(trip_codes) > 0:
        for code in trip_codes:
            branch_max = get_max_vehicle_for_branch(code)
            # ‡∏´‡∏≤‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏•‡πá‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ
            if vehicle_sizes.get(branch_max, 3) < vehicle_sizes.get(branch_max_vehicle, 3):
                branch_max_vehicle = branch_max
        
        # ‡∏à‡∏≥‡∏Å‡∏±‡∏î max_allowed ‡∏ï‡∏≤‡∏°‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏Ç‡∏≠‡∏á‡∏™‡∏≤‡∏Ç‡∏≤
        if vehicle_sizes.get(branch_max_vehicle, 3) < max_size:
            max_allowed = branch_max_vehicle
            max_size = vehicle_sizes.get(max_allowed, 3)
    
    best_truck = None
    best_utilization = 0
    best_distance_from_100 = 999  # ‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å 100%
    
    for truck in ['4W', 'JB', '6W']:
        truck_size = vehicle_sizes.get(truck, 0)
        # ‡∏ñ‡πâ‡∏≤‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏Å‡∏ß‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï ‡∏Ç‡πâ‡∏≤‡∏°‡πÑ‡∏õ
        if truck_size > max_size:
            continue
        if can_fit_truck(total_weight, total_cube, truck):
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì % ‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏£‡∏ñ
            limits = LIMITS[truck]
            w_util = (total_weight / limits['max_w']) * 100
            c_util = (total_cube / limits['max_c']) * 100
            utilization = max(w_util, c_util)
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å 100%
            distance_from_100 = abs(100 - utilization)
            
            # ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ 100% ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î (90-105% ‡πÄ‡∏õ‡πá‡∏ô‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢)
            # ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á‡∏Å‡∏±‡∏ô ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô‡∏™‡∏π‡∏á‡∏Å‡∏ß‡πà‡∏≤
            if best_truck is None:
                best_truck = truck
                best_utilization = utilization
                best_distance_from_100 = distance_from_100
            else:
                # ‡∏ñ‡πâ‡∏≤‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏ô‡∏ä‡πà‡∏ß‡∏á 90-100% ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ 100% ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
                if 90 <= utilization <= 100:
                    if distance_from_100 < best_distance_from_100 or best_utilization < 90:
                        best_truck = truck
                        best_utilization = utilization
                        best_distance_from_100 = distance_from_100
                # ‡∏ñ‡πâ‡∏≤‡∏ó‡∏±‡πâ‡∏á‡∏Ñ‡∏π‡πà‡πÑ‡∏°‡πà‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏ô‡∏ä‡πà‡∏ß‡∏á ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô‡∏™‡∏π‡∏á‡∏Å‡∏ß‡πà‡∏≤
                elif utilization > best_utilization:
                    best_truck = truck
                    best_utilization = utilization
                    best_distance_from_100 = distance_from_100
    
    if best_truck:
        return best_truck
    
    # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏° ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï
    return max_allowed if max_allowed in LIMITS else '6W+'

def calculate_optimal_vehicle_split(total_weight, total_cube, max_allowed='6W', branch_count=0):
    """
    üöõ ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏Å‡∏≤‡∏£‡πÅ‡∏ö‡πà‡∏á‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°
    
    ‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç:
    - 4W: ‚â§12 ‡∏à‡∏∏‡∏î, Cube ‚â§ 5
    - JB: ‚â§12 ‡∏à‡∏∏‡∏î, Cube ‚â§ 8  
    - 6W: ‡πÑ‡∏°‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏à‡∏∏‡∏î, Cube ‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏ï‡πá‡∏° ‚â•100%
    
    ‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏Å‡∏≤‡∏£‡πÄ‡∏•‡∏∑‡∏≠‡∏Å:
    1. 4W (‡∏ñ‡πâ‡∏≤ cube ‚â§ 5)
    2. JB (‡∏ñ‡πâ‡∏≤ cube ‚â§ 8)
    3. JB + 4W (‡πÅ‡∏¢‡∏Å 2 ‡∏Ñ‡∏±‡∏ô, 75%-95% ‡∏ï‡πà‡∏≠‡∏Ñ‡∏±‡∏ô)
    4. JB + JB (‡πÅ‡∏¢‡∏Å 2 ‡∏Ñ‡∏±‡∏ô, 75%-95% ‡∏ï‡πà‡∏≠‡∏Ñ‡∏±‡∏ô)
    5. 6W + JB (‡πÅ‡∏¢‡∏Å 2 ‡∏Ñ‡∏±‡∏ô, 75%-95% ‡∏ï‡πà‡∏≠‡∏Ñ‡∏±‡∏ô)
    6. 4W + 4W (‡πÅ‡∏¢‡∏Å 2 ‡∏Ñ‡∏±‡∏ô, 75%-95% ‡∏ï‡πà‡∏≠‡∏Ñ‡∏±‡∏ô)
    7. 6W (cube ‡∏ï‡πâ‡∏≠‡∏á ‚â•100%)
    
    Returns: (vehicle_type, split_needed, split_config)
    """
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    max_priority = vehicle_priority.get(max_allowed, 3)
    
    # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì utilization ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏£‡∏ñ (‡πÉ‡∏ä‡πâ Cube ‡πÄ‡∏õ‡πá‡∏ô‡∏´‡∏•‡∏±‡∏Å)
    cube_util_4w = (total_cube / LIMITS['4W']['max_c']) * 100  # max 5 cube
    cube_util_jb = (total_cube / LIMITS['JB']['max_c']) * 100  # max 7 cube
    cube_util_6w = (total_cube / LIMITS['6W']['max_c']) * 100  # max 20 cube
    
    weight_util_4w = (total_weight / LIMITS['4W']['max_w']) * 100
    weight_util_jb = (total_weight / LIMITS['JB']['max_w']) * 100
    weight_util_6w = (total_weight / LIMITS['6W']['max_w']) * 100
    
    # üéØ ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢: Utilization 75%-95% ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏Å‡∏≤‡∏£‡πÅ‡∏¢‡∏Å, 95%-100% ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏Ñ‡∏±‡∏ô‡πÄ‡∏î‡∏µ‡∏¢‡∏ß
    SPLIT_MIN = 75   # ‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏Ñ‡∏±‡∏ô‡πÄ‡∏°‡∏∑‡πà‡∏≠‡πÅ‡∏¢‡∏Å
    SPLIT_MAX = 95   # ‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏Ñ‡∏±‡∏ô‡πÄ‡∏°‡∏∑‡πà‡∏≠‡πÅ‡∏¢‡∏Å
    SINGLE_MIN = 95  # ‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏Ñ‡∏±‡∏ô‡πÄ‡∏î‡∏µ‡∏¢‡∏ß
    SINGLE_MAX = 100 # ‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏Ñ‡∏±‡∏ô‡πÄ‡∏î‡∏µ‡∏¢‡∏ß (‡∏´‡πâ‡∏≤‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 100%)
    
    # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤ (4W/JB ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 12 ‡∏à‡∏∏‡∏î)
    branch_ok_for_small = branch_count <= 12 or branch_count == 0
    
    # 1. ‡∏•‡∏≠‡∏á 4W ‡∏Å‡πà‡∏≠‡∏ô (‡∏ñ‡πâ‡∏≤ cube ‚â§ 5 ‡πÅ‡∏•‡∏∞ ‚â§12 ‡∏à‡∏∏‡∏î)
    if max_priority >= 1 and total_cube <= 5.0 and branch_ok_for_small:
        if cube_util_4w <= 100 and weight_util_4w <= 100:
            return ('4W', False, None)
    
    # 2. ‡∏•‡∏≠‡∏á JB (‡∏ñ‡πâ‡∏≤ cube ‚â§ 7 ‡πÅ‡∏•‡∏∞ ‚â§12 ‡∏à‡∏∏‡∏î)
    if max_priority >= 2 and total_cube <= 7.0 and branch_ok_for_small:
        if cube_util_jb <= 100 and weight_util_jb <= 100:
            return ('JB', False, None)
    
    # 3. ‡∏ñ‡πâ‡∏≤‡∏£‡∏ñ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡πÑ‡∏°‡πà‡∏û‡∏≠ ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å (cube > 7 ‡∏´‡∏£‡∏∑‡∏≠ ‡∏à‡∏∏‡∏î > 12)
    need_split = total_cube > 7.0 or not branch_ok_for_small
    
    if need_split:
        # üîÑ ‡∏•‡∏≠‡∏á‡πÅ‡∏ö‡∏ö‡∏ï‡πà‡∏≤‡∏á‡πÜ ‡∏ï‡∏≤‡∏°‡∏•‡∏≥‡∏î‡∏±‡∏ö
        
        # üîí ‡∏Å‡∏é‡πÉ‡∏´‡∏°‡πà: cube 7-14 ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB+JB ‡∏´‡∏£‡∏∑‡∏≠ JB+4W
        if total_cube > 7.0 and total_cube <= 14.0:
            if total_cube <= 12.0:
                # JB + 4W (JB 7 + 4W 5 = 12 max)
                return ('JB', True, {'split': ['JB', '4W'], 'reason': f'Cube {total_cube:.1f} ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB+4W'})
            else:
                # JB + JB (7 + 7 = 14 max)
                return ('JB', True, {'split': ['JB', 'JB'], 'reason': f'Cube {total_cube:.1f} ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB+JB'})
        
        # üîí ‡∏Å‡∏é‡πÉ‡∏´‡∏°‡πà: cube 14-18 ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB+JB+4W ‡∏´‡∏£‡∏∑‡∏≠ 6W ‡∏ñ‡πâ‡∏≤‡∏à‡∏≥‡πÄ‡∏õ‡πá‡∏ô
        if total_cube > 14.0 and total_cube < 18.0:
            # ‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á 6W (‡∏ï‡πâ‡∏≠‡∏á ‚â•18) ‡πÅ‡∏ï‡πà‡πÄ‡∏Å‡∏¥‡∏ô JB+JB ‚Üí ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB+JB+4W ‡∏´‡∏£‡∏∑‡∏≠‡∏¢‡∏≠‡∏°‡πÉ‡∏ä‡πâ 6W
            if max_priority >= 3:
                # ‡∏¢‡∏≠‡∏°‡πÉ‡∏ä‡πâ 6W ‡πÅ‡∏°‡πâ‡∏ß‡πà‡∏≤‡∏à‡∏∞‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á 18 cube (‡∏î‡∏µ‡∏Å‡∏ß‡πà‡∏≤‡πÅ‡∏¢‡∏Å 3 ‡∏Ñ‡∏±‡∏ô)
                return ('6W', False, {'reason': f'Cube {total_cube:.1f} ‡πÉ‡∏ä‡πâ 6W (‡∏£‡∏≠‡∏£‡∏ß‡∏°‡πÄ‡∏û‡∏¥‡πà‡∏°)'})
            else:
                return ('JB', True, {'split': ['JB', 'JB'], 'reason': f'Cube {total_cube:.1f} ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB+JB'})
        
        # 6W + JB (‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö cube > 20)
        if max_priority >= 3 and total_cube > 20.0 and total_cube <= 27.0:
            return ('6W', True, {'split': ['6W', 'JB'], 'reason': f'Cube {total_cube:.1f} ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô 6W+JB'})
        
        # 4W + 4W (4W 5 + 4W 5 = 10 cube max) - ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î 4W
        if max_priority == 1 and total_cube <= 10.0:
            four_w_util_half = (total_cube / 2 / LIMITS['4W']['max_c']) * 100
            if SPLIT_MIN <= four_w_util_half <= SPLIT_MAX:
                return ('4W', True, {'split': ['4W', '4W'], 'ratio': [0.5, 0.5]})
    
    # 4. 6W (‡πÑ‡∏°‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏à‡∏∏‡∏î ‡πÅ‡∏ï‡πà‡∏ï‡πâ‡∏≠‡∏á‡πÑ‡∏î‡πâ‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥ 18 cube (90%) ‡πÅ‡∏•‡∏∞‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 20 cube)
    if max_priority >= 3:
        # üîí 6W ‡πÄ‡∏Å‡∏¥‡∏ô 20 cube ‚Üí ‡∏ï‡∏±‡∏î‡∏™‡πà‡∏ß‡∏ô‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ 4W
        if total_cube > 20.0:
            return ('6W', True, {'split': ['6W', '4W'], 'reason': '‡πÄ‡∏Å‡∏¥‡∏ô 20 cube ‡∏ï‡∏±‡∏î‡∏™‡πà‡∏ß‡∏ô‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ 4W'})
        
        # üîí 6W ‡∏ï‡πâ‡∏≠‡∏á‡πÑ‡∏î‡πâ‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥ 18 cube (90%) ‡∏ñ‡πâ‡∏≤‡∏ï‡πà‡∏≥‡∏Å‡∏ß‡πà‡∏≤ ‚Üí ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB
        if total_cube >= 18.0:
            return ('6W', False, None)
        elif total_cube >= 7.0 and total_cube < 18.0:
            # 6W ‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á 18 cube ‚Üí ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB ‡πÅ‡∏ó‡∏ô (‡∏ñ‡πâ‡∏≤‡πÑ‡∏î‡πâ)
            if max_priority >= 2 and branch_ok_for_small:
                # ‡∏•‡∏≠‡∏á‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB + JB ‡∏´‡∏£‡∏∑‡∏≠ JB + 4W
                if total_cube <= 14.0:  # JB + JB = 14 cube max
                    return ('JB', True, {'split': ['JB', 'JB'], 'reason': '6W ‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á 18 cube ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB'})
                else:
                    return ('JB', True, {'split': ['JB', '4W'], 'reason': '6W ‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á 18 cube ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB+4W'})
            else:
                # ‡∏ñ‡πâ‡∏≤‡πÅ‡∏¢‡∏Å JB ‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ ‚Üí ‡∏¢‡∏≠‡∏°‡πÉ‡∏ä‡πâ 6W ‡πÅ‡∏°‡πâ‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á 18 cube
                return ('6W', False, None)
        else:
            # 6W ‡∏ß‡πà‡∏≤‡∏á‡∏°‡∏≤‡∏Å (<7 cube) ‚Üí ‡∏•‡∏î‡πÄ‡∏õ‡πá‡∏ô JB ‡∏´‡∏£‡∏∑‡∏≠ 4W
            if total_cube <= 7.0 and branch_ok_for_small and max_priority >= 2:
                return ('JB', False, None)
            if total_cube <= 5.0 and branch_ok_for_small:
                return ('4W', False, None)
    
    # Default: ‡πÉ‡∏ä‡πâ max_allowed
    return (max_allowed, False, None)

def can_branch_use_vehicle(code, vehicle_type, branch_vehicles):
    """
    ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏ô‡∏µ‡πâ‡πÑ‡∏î‡πâ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
    - ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥ = ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡∏ó‡∏∏‡∏Å‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó
    - ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà = ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å‡∏Å‡∏ß‡πà‡∏≤‡πÑ‡∏î‡πâ
    - ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥‡πÉ‡∏ä‡πâ‡πÅ‡∏Ñ‡πà‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å (‡πÄ‡∏ä‡πà‡∏ô 4W) = ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ (‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡πÄ‡∏Ç‡πâ‡∏≤‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ)
    """
    if not branch_vehicles or code not in branch_vehicles:
        return True  # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥ = ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡∏ó‡∏∏‡∏Å‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return True  # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏£‡∏ñ = ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡∏ó‡∏∏‡∏Å‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó
    
    # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏ô‡∏µ‡πâ = ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ
    if vehicle_type in vehicle_history:
        return True
    
    # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ç‡∏ô‡∏≤‡∏î‡∏£‡∏ñ (6W > JB > 4W)
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    requested_size = vehicle_sizes.get(vehicle_type, 0)
    
    # ‡∏´‡∏≤‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÉ‡∏´‡∏ç‡πà‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ
    max_used_size = max(vehicle_sizes.get(v, 0) for v in vehicle_history)
    
    # ‡∏ñ‡πâ‡∏≤‡∏Ç‡∏≠‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å‡∏Å‡∏ß‡πà‡∏≤‡∏´‡∏£‡∏∑‡∏≠‡πÄ‡∏ó‡πà‡∏≤‡∏Å‡∏±‡∏ö‡∏ó‡∏µ‡πà‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ = ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ
    # ‡∏ñ‡πâ‡∏≤‡∏Ç‡∏≠‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏Å‡∏ß‡πà‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ = ‡πÉ‡∏ä‡πâ‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ (‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏≠‡∏≤‡∏à‡πÄ‡∏Ç‡πâ‡∏≤‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ)
    return requested_size <= max_used_size

def get_max_vehicle_for_branch_old(code, branch_vehicles):
    """[OLD] ‡∏î‡∏∂‡∏á‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÉ‡∏´‡∏ç‡πà‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ (‡∏à‡∏≥‡∏Å‡∏±‡∏î‡πÑ‡∏°‡πà‡πÉ‡∏´‡πâ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏Å‡∏ß‡πà‡∏≤‡∏ô‡∏µ‡πâ)"""
    if not branch_vehicles or code not in branch_vehicles:
        return '6W'  # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥ = ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡∏ñ‡∏∂‡∏á 6W
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return '6W'
    
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    max_vehicle = max(vehicle_history.keys(), key=lambda v: vehicle_sizes.get(v, 0))
    return max_vehicle

def get_most_used_vehicle_for_branch(code, branch_vehicles):
    """‡∏î‡∏∂‡∏á‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ä‡πâ‡∏ö‡πà‡∏≠‡∏¢‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î"""
    if not branch_vehicles or code not in branch_vehicles:
        return None
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return None
    
    return max(vehicle_history, key=vehicle_history.get)

def is_similar_name(name1, name2, similarity_threshold=85):
    """‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà - ‡πÉ‡∏ä‡πâ Fuzzy Matching + ‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç
    
    Args:
        name1: ‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà 1
        name2: ‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà 2
        similarity_threshold: ‡∏Ñ‡∏ß‡∏≤‡∏°‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥ (0-100, default=85)
    
    Returns:
        True ‡∏ñ‡πâ‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô
    """
    def extract_keywords(name):
        """‡∏î‡∏∂‡∏á‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤"""
        if pd.isna(name) or name is None:
            return set(), "", ""
        s = str(name).strip().upper()
        
        # ‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏ö‡∏Ñ‡∏π‡πà (‡πÑ‡∏ó‡∏¢ + ‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©) - ‡πÄ‡∏û‡∏¥‡πà‡∏°‡πÄ‡∏ï‡∏¥‡∏°‡∏Ñ‡∏≥‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏Ç‡∏≠‡∏á‡∏™‡∏≤‡∏Ç‡∏≤
        keywords = set()
        
        # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡πÅ‡∏ö‡∏ö exact match
        important_words = [
            # ‡∏™‡∏ñ‡∏≤‡∏ô‡∏ó‡∏µ‡πà‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç
            '‡∏ü‡∏¥‡∏ß‡πÄ‡∏à‡∏≠‡∏£‡πå', 'FUTURE', '‡∏£‡∏±‡∏á‡∏™‡∏¥‡∏ï', 'RANGSIT', '‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á', 'KHLONGLUANG',
            '‡πÄ‡∏ã‡πá‡∏ô‡∏ó‡∏£‡∏±‡∏•', 'CENTRAL', '‡πÄ‡∏ó‡∏™‡πÇ‡∏Å‡πâ', 'TESCO', '‡πÇ‡∏•‡∏ï‡∏±‡∏™', 'LOTUS',
            '‡∏ö‡∏¥‡πä‡∏Å‡∏ã‡∏µ', 'BIGC', '‡πÅ‡∏°‡πá‡∏Ñ‡πÇ‡∏Ñ‡∏£', 'MAKRO', '‡πÇ‡∏Æ‡∏°‡πÇ‡∏õ‡∏£', 'HOMEPRO',
            '‡∏ã‡∏µ‡∏Ñ‡∏≠‡∏ô', 'SEACON', '‡πÄ‡∏°‡∏Å‡∏≤', 'MEGA', '‡∏û‡∏≤‡∏£‡∏≤‡πÑ‡∏î‡∏ã‡πå', 'PARADISE',
            '‡πÄ‡∏ó‡∏≠‡∏£‡πå‡∏°‡∏¥‡∏ô‡∏≠‡∏•', 'TERMINAL', '‡∏™‡∏¢‡∏≤‡∏°‡∏û‡∏≤‡∏£‡∏≤‡∏Å‡∏≠‡∏ô', 'SIAM', 'PARAGON',
            # ‡∏´‡∏°‡∏≤‡∏¢‡πÄ‡∏•‡∏Ç‡∏™‡∏≤‡∏Ç‡∏≤ (‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á 3, 4, 8, 10 ‡∏Ø‡∏•‡∏Ø)
            '‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á3', '‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á4', '‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á8', '‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á10',
        ]
        
        for word in important_words:
            if word in s:
                keywords.add(word)
        
        # ‡∏î‡∏∂‡∏á‡∏ä‡∏∑‡πà‡∏≠‡∏û‡∏∑‡πâ‡∏ô‡∏ê‡∏≤‡∏ô (‡πÄ‡∏ä‡πà‡∏ô "‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á" ‡∏à‡∏≤‡∏Å "‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á3")
        import re
        base_match = re.search(r'([‡∏Å-‡πôA-Z]+)\s*\d+', s)
        if base_match:
            base_name = base_match.group(1).strip()
            if len(base_name) >= 3:
                keywords.add(base_name)
        
        # Pattern 3: ‡∏î‡∏∂‡∏á‡∏ä‡∏∑‡πà‡∏≠‡πÉ‡∏ô‡∏ß‡∏á‡πÄ‡∏•‡πá‡∏ö (‡πÄ‡∏ä‡πà‡∏ô "‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á4(‡∏ñ.‡∏û‡∏´‡∏•‡πÇ‡∏¢‡∏ò‡∏¥‡∏ô ‡∏Å‡∏°.34)" ‚Üí "‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á4")
        paren_match = re.search(r'^([^(]+)', s)
        if paren_match:
            main_name = paren_match.group(1).strip()
            if len(main_name) >= 3 and main_name != s:
                keywords.add(main_name)  # "‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á4"
        
        # ‡∏•‡∏ö prefix/suffix ‡∏ó‡∏µ‡πà‡∏û‡∏ö‡∏ö‡πà‡∏≠‡∏¢
        prefixes = ['PTC-MRT-', 'FC PTF ', 'PTC-', 'PTC ', 'PUN-', 'PTF ', 'FC ', 
                   'MAXMART', 'CW', 'NW', 'MI', 'PI', 'MH', 'ME', 'SE', 'SG', 'SH', 'MG']
        clean_s = s
        for prefix in prefixes:
            if clean_s.startswith(prefix):
                clean_s = clean_s[len(prefix):].strip()
                break
        
        # ‡∏•‡∏ö‡∏ï‡∏±‡∏ß‡∏≠‡∏±‡∏Å‡∏©‡∏£‡πÄ‡∏î‡∏µ‡πà‡∏¢‡∏ß‡∏ó‡∏µ‡πà‡∏Ç‡∏∂‡πâ‡∏ô‡∏ï‡πâ‡∏ô (M, P, N, S) ‡∏ñ‡πâ‡∏≤‡∏ï‡∏≤‡∏°‡∏î‡πâ‡∏ß‡∏¢‡∏ï‡∏±‡∏ß‡πÄ‡∏•‡∏Ç
        if re.match(r'^[MPNS]\d', clean_s):
            clean_s = clean_s[1:]
        
        # ‡πÅ‡∏¢‡∏Å‡∏†‡∏≤‡∏©‡∏≤‡πÑ‡∏ó‡∏¢‡πÅ‡∏•‡∏∞‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©
        thai_chars = ''.join([c for c in s if '\u0e01' <= c <= '\u0e5b'])
        eng_chars = ''.join([c for c in s if c.isalpha() and c.isascii()])
        
        return keywords, thai_chars, eng_chars, clean_s
    
    keywords1, thai1, eng1, clean1 = extract_keywords(name1)
    keywords2, thai2, eng2, clean2 = extract_keywords(name2)
    
    # üî• ‡∏•‡∏≥‡∏î‡∏±‡∏ö‡πÅ‡∏£‡∏Å: ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏Å‡πà‡∏≠‡∏ô (‡πÄ‡∏ä‡πà‡∏ô ‡∏ü‡∏¥‡∏ß‡πÄ‡∏à‡∏≠‡∏£‡πå+‡∏£‡∏±‡∏á‡∏™‡∏¥‡∏ï, ‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á)
    if keywords1 and keywords2:
        common_keywords = keywords1 & keywords2
        
        # ‚úÖ Case 1: ‡∏ä‡∏∑‡πà‡∏≠‡∏û‡∏∑‡πâ‡∏ô‡∏ê‡∏≤‡∏ô‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô (‡πÄ‡∏ä‡πà‡∏ô "‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á" ‡πÉ‡∏ô ‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á3, ‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á4, ‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á6)
        base_names = {k for k in common_keywords if len(k) >= 3 and not k.isdigit()}
        if base_names:
            # ‡∏ñ‡∏∑‡∏≠‡∏ß‡πà‡∏≤‡πÄ‡∏õ‡πá‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô (‡πÄ‡∏ä‡πà‡∏ô ‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á 3, 4, 6, 8, 10)
            return True
        
        # ‚úÖ Case 1.5: ‡πÄ‡∏ä‡πá‡∏Ñ partial match ‡∏Ç‡∏≠‡∏á‡∏ä‡∏∑‡πà‡∏≠‡∏û‡∏∑‡πâ‡∏ô‡∏ê‡∏≤‡∏ô (‡πÄ‡∏ä‡πà‡∏ô "KHLONG" ‡πÉ‡∏ô keywords1, "LUANG" ‡πÉ‡∏ô keywords1 + "KHLONG" ‡πÉ‡∏ô keywords2)
        # ‚Üí ‡∏ñ‡∏∑‡∏≠‡∏ß‡πà‡∏≤‡πÄ‡∏õ‡πá‡∏ô "KHLONG LUANG" ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô (‡∏£‡∏≠‡∏á‡∏£‡∏±‡∏ö‡∏†‡∏≤‡∏©‡∏≤‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©)
        for k1 in keywords1:
            for k2 in keywords2:
                # ‡∏ñ‡πâ‡∏≤ k1 ‡πÄ‡∏õ‡πá‡∏ô substring ‡∏Ç‡∏≠‡∏á k2 ‡∏´‡∏£‡∏∑‡∏≠‡∏ï‡∏£‡∏á‡∏Å‡∏±‡∏ô‡∏Ç‡πâ‡∏≤‡∏°
                if len(k1) >= 4 and len(k2) >= 4:
                    if k1 in k2 or k2 in k1:
                        return True
        
        # ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏£‡πà‡∏ß‡∏°‡∏Å‡∏±‡∏ô >= 2 ‡∏Ñ‡∏≥ ‚Üí ‡∏ñ‡∏∑‡∏≠‡∏ß‡πà‡∏≤‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô‡∏Å‡∏±‡∏ô
        if len(common_keywords) >= 2:
            return True
        
        # ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏£‡πà‡∏ß‡∏°‡∏Å‡∏±‡∏ô 1 ‡∏Ñ‡∏≥ ‡πÅ‡∏ï‡πà‡πÄ‡∏õ‡πá‡∏ô‡∏Ñ‡∏≥‡πÄ‡∏â‡∏û‡∏≤‡∏∞ ‚Üí ‡∏ñ‡∏∑‡∏≠‡∏ß‡πà‡∏≤‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô‡∏Å‡∏±‡∏ô
        if len(common_keywords) >= 1:
            specific_places = {'‡∏£‡∏±‡∏á‡∏™‡∏¥‡∏ï', 'RANGSIT', '‡πÄ‡∏ã‡πá‡∏ô‡∏ó‡∏£‡∏±‡∏•', 'CENTRAL', '‡∏ã‡∏µ‡∏Ñ‡∏≠‡∏ô', 'SEACON', '‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á', 'KHLONGLUANG', '‡∏ï‡∏•‡∏≤‡∏î‡πÑ‡∏ó', 'TALADTHAI'}
            if common_keywords & specific_places:
                if len(common_keywords) >= 2 or (thai1 and thai2 and len(thai1) >= 4 and thai1[:4] in thai2):
                    return True
    
    # üéØ Fuzzy Matching - ‡πÉ‡∏ä‡πâ rapidfuzz ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ ‡∏´‡∏£‡∏∑‡∏≠ difflib ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ
    if FUZZY_AVAILABLE:
        # ‡πÉ‡∏ä‡πâ rapidfuzz (‡πÄ‡∏£‡πá‡∏ß‡πÅ‡∏•‡∏∞‡πÅ‡∏°‡πà‡∏ô‡∏¢‡∏≥‡∏Å‡∏ß‡πà‡∏≤)
        ratio = fuzz.token_sort_ratio(clean1, clean2)
        if ratio >= similarity_threshold:
            return True
        
        # ‡πÄ‡∏ä‡πá‡∏Ñ partial ratio ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏ä‡∏∑‡πà‡∏≠‡∏ó‡∏µ‡πà‡πÄ‡∏õ‡πá‡∏ô substring
        partial_ratio = fuzz.partial_ratio(clean1, clean2)
        if partial_ratio >= 90:  # ‡∏Ñ‡∏ß‡∏≤‡∏°‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏™‡∏π‡∏á‡∏°‡∏≤‡∏Å
            return True
    else:
        # Fallback: ‡πÉ‡∏ä‡πâ difflib
        ratio = SequenceMatcher(None, clean1, clean2).ratio() * 100
        if ratio >= similarity_threshold:
            return True
    
    # ‡∏ï‡πâ‡∏≠‡∏á‡∏°‡∏µ‡∏Ñ‡∏ß‡∏≤‡∏°‡∏¢‡∏≤‡∏ß‡∏û‡∏≠‡∏™‡∏°‡∏Ñ‡∏ß‡∏£
    if len(thai1) < 3 and len(eng1) < 3:
        return False
    if len(thai2) < 3 and len(eng2) < 3:
        return False
    
    # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏†‡∏≤‡∏©‡∏≤‡πÑ‡∏ó‡∏¢
    if thai1 and thai2:
        shorter_thai = min(thai1, thai2, key=len)
        longer_thai = max(thai1, thai2, key=len)
        if len(shorter_thai) >= 3 and shorter_thai in longer_thai:
            return True
        # ‡∏Ñ‡∏ß‡∏≤‡∏°‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢ 80%+
        if len(shorter_thai) >= 5:
            common = sum(1 for c in shorter_thai if c in longer_thai)
            if common / len(shorter_thai) >= 0.8:
                return True
    
    # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏†‡∏≤‡∏©‡∏≤‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©
    if eng1 and eng2:
        shorter_eng = min(eng1, eng2, key=len)
        longer_eng = max(eng1, eng2, key=len)
        if len(shorter_eng) >= 3 and shorter_eng in longer_eng:
            return True
        # ‡∏Ñ‡∏ß‡∏≤‡∏°‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢ 80%+
        if len(shorter_eng) >= 5:
            common = sum(1 for c in shorter_eng if c in longer_eng)
            if common / len(shorter_eng) >= 0.8:
                return True
    
    # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏£‡∏∞‡∏´‡∏ß‡πà‡∏≤‡∏á‡πÑ‡∏ó‡∏¢-‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏© (‡πÄ‡∏ä‡πà‡∏ô Future = ‡∏ü‡∏¥‡∏ß‡πÄ‡∏à‡∏≠‡∏£‡πå, Rangsit = ‡∏£‡∏±‡∏á‡∏™‡∏¥‡∏ï)
    thai_eng_map = {
        'RANGSIT': '‡∏£‡∏±‡∏á‡∏™‡∏¥‡∏ï',
        'FUTURE': '‡∏ü‡∏¥‡∏ß‡πÄ‡∏à‡∏≠‡∏£',
        'PARK': '‡∏õ‡∏≤‡∏£‡∏Ñ',
        'TRIANGLE': '‡πÑ‡∏ï‡∏£‡πÅ‡∏≠‡∏á‡πÄ‡∏Å‡∏¥‡∏•',
    }
    
    for eng_word, thai_word in thai_eng_map.items():
        # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏°‡∏µ‡∏Ñ‡∏≥‡∏ô‡∏µ‡πâ‡πÉ‡∏ô‡∏ä‡∏∑‡πà‡∏≠‡∏ó‡∏±‡πâ‡∏á‡∏™‡∏≠‡∏á‡∏ù‡∏±‡πà‡∏á (‡πÑ‡∏ó‡∏¢-‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏© ‡∏´‡∏£‡∏∑‡∏≠ ‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©-‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏© ‡∏´‡∏£‡∏∑‡∏≠ ‡πÑ‡∏ó‡∏¢-‡πÑ‡∏ó‡∏¢)
        has_eng_in_1 = eng_word in eng1
        has_eng_in_2 = eng_word in eng2
        has_thai_in_1 = thai_word in thai1
        has_thai_in_2 = thai_word in thai2
        
        # ‡∏ñ‡πâ‡∏≤‡∏ó‡∏±‡πâ‡∏á‡∏™‡∏≠‡∏á‡∏°‡∏µ‡∏Ñ‡∏≥‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô (‡πÑ‡∏°‡πà‡∏ß‡πà‡∏≤‡∏à‡∏∞‡πÑ‡∏ó‡∏¢‡∏´‡∏£‡∏∑‡∏≠‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©) = ‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô
        if (has_eng_in_1 and has_eng_in_2) or (has_thai_in_1 and has_thai_in_2):
            return True
        # ‡∏ñ‡πâ‡∏≤‡∏Ç‡πâ‡∏≤‡∏°‡∏†‡∏≤‡∏©‡∏≤ (‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©-‡πÑ‡∏ó‡∏¢)
        if (has_eng_in_1 and has_thai_in_2) or (has_eng_in_2 and has_thai_in_1):
            return True
    
    return False

def haversine_distance(lat1, lon1, lat2, lon2):
    """
    ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏£‡∏∞‡∏´‡∏ß‡πà‡∏≤‡∏á‡∏à‡∏∏‡∏î‡∏™‡∏≠‡∏á‡∏à‡∏∏‡∏î‡∏ö‡∏ô‡∏û‡∏∑‡πâ‡∏ô‡πÇ‡∏•‡∏Å (km)
    ‡πÉ‡∏ä‡πâ‡∏™‡∏π‡∏ï‡∏£ Haversine
    """
    from math import radians, sin, cos, sqrt, atan2
    
    # ‡πÅ‡∏õ‡∏•‡∏á‡∏≠‡∏á‡∏®‡∏≤‡πÄ‡∏õ‡πá‡∏ô radians
    lat1_rad = radians(lat1)
    lon1_rad = radians(lon1)
    lat2_rad = radians(lat2)
    lon2_rad = radians(lon2)
    
    # ‡∏Ñ‡∏ß‡∏≤‡∏°‡∏ï‡πà‡∏≤‡∏á‡∏Ç‡∏≠‡∏á‡∏û‡∏¥‡∏Å‡∏±‡∏î
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    
    # ‡∏™‡∏π‡∏ï‡∏£ Haversine
    a = sin(dlat/2)**2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    
    # ‡∏£‡∏±‡∏®‡∏°‡∏µ‡πÇ‡∏•‡∏Å (km)
    R = 6371.0
    distance = R * c
    
    return distance

def get_region_type(province):
    """
    ‡∏Å‡∏≥‡∏´‡∏ô‡∏î‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏û‡∏∑‡πâ‡∏ô‡∏ó‡∏µ‡πà‡πÅ‡∏•‡∏∞‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°
    
    Returns:
        str: 'nearby' (‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û+‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•+‡∏†‡∏≤‡∏Ñ‡∏Å‡∏•‡∏≤‡∏á - ‡∏´‡πâ‡∏≤‡∏° 6W ‡πÉ‡∏ä‡πâ 4W/JB ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô),
             'north' (‡∏†‡∏≤‡∏Ñ‡πÄ‡∏´‡∏ô‡∏∑‡∏≠ - ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏ä‡πâ 6W),
             'south' (‡∏†‡∏≤‡∏Ñ‡πÉ‡∏ï‡πâ - ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏ä‡πâ 6W),
             'far' (‡∏†‡∏π‡∏°‡∏¥‡∏†‡∏≤‡∏Ñ‡∏≠‡∏∑‡πà‡∏ô - ‡∏¢‡∏∑‡∏î‡∏´‡∏¢‡∏∏‡πà‡∏ô‡∏ï‡∏≤‡∏° utilization),
             'unknown'
    """
    if pd.isna(province):
        return 'unknown'
    
    prov = str(province).strip()
    
    # üö´ ‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û + ‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏• + ‡∏†‡∏≤‡∏Ñ‡∏Å‡∏•‡∏≤‡∏á ‚Üí ‡∏´‡πâ‡∏≤‡∏°‡πÉ‡∏ä‡πâ 6W ‡πÄ‡∏î‡πá‡∏î‡∏Ç‡∏≤‡∏î! ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡πÅ‡∏Ñ‡πà 4W/JB
    nearby_provinces = [
        '‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏°‡∏´‡∏≤‡∏ô‡∏Ñ‡∏£', '‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û',
        '‡∏ô‡∏Ñ‡∏£‡∏õ‡∏ê‡∏°', '‡∏ô‡∏ô‡∏ó‡∏ö‡∏∏‡∏£‡∏µ', '‡∏õ‡∏ó‡∏∏‡∏°‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏õ‡∏£‡∏≤‡∏Å‡∏≤‡∏£', '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏™‡∏≤‡∏Ñ‡∏£',
        '‡∏ä‡∏±‡∏¢‡∏ô‡∏≤‡∏ó', '‡∏û‡∏£‡∏∞‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤', '‡∏•‡∏û‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏£‡∏∞‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏¥‡∏á‡∏´‡πå‡∏ö‡∏∏‡∏£‡∏µ', '‡∏≠‡πà‡∏≤‡∏á‡∏ó‡∏≠‡∏á', '‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤',
        '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏™‡∏á‡∏Ñ‡∏£‡∏≤‡∏°', '‡∏™‡∏∏‡∏û‡∏£‡∏£‡∏ì‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ô‡∏Ñ‡∏£‡∏ô‡∏≤‡∏¢‡∏Å'
    ]
    
    for nearby in nearby_provinces:
        if nearby in prov:
            return 'nearby'
    
    # üöõ ‡∏†‡∏≤‡∏Ñ‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î (18 ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î) ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏ä‡πâ 6W
    north_provinces = [
        # ‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ï‡∏≠‡∏ô‡∏ö‡∏ô
        '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡πÉ‡∏´‡∏°‡πà', '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡∏£‡∏≤‡∏¢', '‡πÅ‡∏°‡πà‡∏Æ‡πà‡∏≠‡∏á‡∏™‡∏≠‡∏ô', '‡∏ô‡πà‡∏≤‡∏ô', '‡∏û‡∏∞‡πÄ‡∏¢‡∏≤', '‡∏•‡∏≥‡∏õ‡∏≤‡∏á', '‡∏•‡∏≥‡∏û‡∏π‡∏ô', '‡πÅ‡∏û‡∏£‡πà',
        # ‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ï‡∏≠‡∏ô‡∏•‡πà‡∏≤‡∏á
        '‡∏Å‡∏≥‡πÅ‡∏û‡∏á‡πÄ‡∏û‡∏ä‡∏£', '‡∏ï‡∏≤‡∏Å', '‡∏ô‡∏Ñ‡∏£‡∏™‡∏ß‡∏£‡∏£‡∏Ñ‡πå', '‡∏û‡∏¥‡∏à‡∏¥‡∏ï‡∏£', '‡∏û‡∏¥‡∏©‡∏ì‡∏∏‡πÇ‡∏•‡∏Å', '‡∏™‡∏∏‡πÇ‡∏Ç‡∏ó‡∏±‡∏¢', 
        '‡∏≠‡∏∏‡∏ï‡∏£‡∏î‡∏¥‡∏ï‡∏ñ‡πå', '‡∏≠‡∏∏‡∏ó‡∏±‡∏¢‡∏ò‡∏≤‡∏ô‡∏µ', '‡πÄ‡∏û‡∏ä‡∏£‡∏ö‡∏π‡∏£‡∏ì‡πå', '‡∏ä‡∏±‡∏¢‡∏†‡∏π‡∏°‡∏¥'
    ]
    
    for north in north_provinces:
        if north in prov:
            return 'north'
    
    # üöõ ‡∏†‡∏≤‡∏Ñ‡πÉ‡∏ï‡πâ‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î (14 ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î) ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏ä‡πâ 6W
    south_provinces = [
        # ‡∏ù‡∏±‡πà‡∏á‡∏≠‡∏±‡∏ô‡∏î‡∏≤‡∏°‡∏±‡∏ô
        '‡∏Å‡∏£‡∏∞‡∏ö‡∏µ‡πà', '‡∏ï‡∏£‡∏±‡∏á', '‡∏û‡∏±‡∏á‡∏á‡∏≤', '‡∏†‡∏π‡πÄ‡∏Å‡πá‡∏ï', '‡∏£‡∏∞‡∏ô‡∏≠‡∏á', '‡∏™‡∏ï‡∏π‡∏•',
        # ‡∏ù‡∏±‡πà‡∏á‡∏≠‡πà‡∏≤‡∏ß‡πÑ‡∏ó‡∏¢
        '‡∏ä‡∏∏‡∏°‡∏û‡∏£', '‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏ò‡∏£‡∏£‡∏°‡∏£‡∏≤‡∏ä', '‡∏û‡∏±‡∏ó‡∏•‡∏∏‡∏á', '‡∏¢‡∏∞‡∏•‡∏≤', '‡∏™‡∏á‡∏Ç‡∏•‡∏≤', 
        '‡∏™‡∏∏‡∏£‡∏≤‡∏©‡∏é‡∏£‡πå‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏õ‡∏±‡∏ï‡∏ï‡∏≤‡∏ô‡∏µ', '‡∏ô‡∏£‡∏≤‡∏ò‡∏¥‡∏ß‡∏≤‡∏™'
    ]
    
    for south in south_provinces:
        if south in prov:
            return 'south'
    
    # ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏≠‡∏∑‡πà‡∏ô‡πÜ (‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏≠‡∏≠‡∏Å, ‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏ï‡∏Å, ‡∏≠‡∏µ‡∏™‡∏≤‡∏ô) = ‡∏¢‡∏∑‡∏î‡∏´‡∏¢‡∏∏‡πà‡∏ô‡∏ï‡∏≤‡∏° utilization
    return 'far'

def is_nearby_province(prov1, prov2):
    """‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á‡∏Å‡∏±‡∏ô‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà (‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥)"""
    if pd.isna(prov1) or pd.isna(prov2):
        return False
    
    if prov1 == prov2:
        return True
    
    # ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏ï‡∏≤‡∏°‡∏†‡∏≤‡∏Ñ‡∏¢‡πà‡∏≠‡∏¢ (‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥)
    province_groups = {
        '‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û': ['‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏°‡∏´‡∏≤‡∏ô‡∏Ñ‡∏£', '‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û'],
        '‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•': ['‡∏ô‡∏Ñ‡∏£‡∏õ‡∏ê‡∏°', '‡∏ô‡∏ô‡∏ó‡∏ö‡∏∏‡∏£‡∏µ', '‡∏õ‡∏ó‡∏∏‡∏°‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏õ‡∏£‡∏≤‡∏Å‡∏≤‡∏£', '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏™‡∏≤‡∏Ñ‡∏£'],
        '‡∏Å‡∏•‡∏≤‡∏á‡∏ï‡∏≠‡∏ô‡∏ö‡∏ô': ['‡∏ä‡∏±‡∏¢‡∏ô‡∏≤‡∏ó', '‡∏û‡∏£‡∏∞‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤', '‡∏•‡∏û‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏£‡∏∞‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏¥‡∏á‡∏´‡πå‡∏ö‡∏∏‡∏£‡∏µ', '‡∏≠‡πà‡∏≤‡∏á‡∏ó‡∏≠‡∏á', '‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤'],
        '‡∏Å‡∏•‡∏≤‡∏á‡∏ï‡∏≠‡∏ô‡∏•‡πà‡∏≤‡∏á': ['‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏™‡∏á‡∏Ñ‡∏£‡∏≤‡∏°', '‡∏™‡∏∏‡∏û‡∏£‡∏£‡∏ì‡∏ö‡∏∏‡∏£‡∏µ'],
        '‡∏†‡∏≤‡∏Ñ‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏ï‡∏Å': ['‡∏Å‡∏≤‡∏ç‡∏à‡∏ô‡∏ö‡∏∏‡∏£‡∏µ', '‡∏õ‡∏£‡∏∞‡∏à‡∏ß‡∏ö‡∏Ñ‡∏µ‡∏£‡∏µ‡∏Ç‡∏±‡∏ô‡∏ò‡πå', '‡∏£‡∏≤‡∏ä‡∏ö‡∏∏‡∏£‡∏µ', '‡πÄ‡∏û‡∏ä‡∏£‡∏ö‡∏∏‡∏£‡∏µ'],
        '‡∏†‡∏≤‡∏Ñ‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏≠‡∏≠‡∏Å': ['‡∏à‡∏±‡∏ô‡∏ó‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ä‡∏•‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ï‡∏£‡∏≤‡∏î', '‡∏ô‡∏Ñ‡∏£‡∏ô‡∏≤‡∏¢‡∏Å', '‡∏õ‡∏£‡∏≤‡∏à‡∏µ‡∏ô‡∏ö‡∏∏‡∏£‡∏µ', '‡∏£‡∏∞‡∏¢‡∏≠‡∏á', '‡∏™‡∏£‡∏∞‡πÅ‡∏Å‡πâ‡∏ß', '‡∏â‡∏∞‡πÄ‡∏ä‡∏¥‡∏á‡πÄ‡∏ó‡∏£‡∏≤'],
        '‡∏≠‡∏µ‡∏™‡∏≤‡∏ô‡πÄ‡∏´‡∏ô‡∏∑‡∏≠': ['‡∏ô‡∏Ñ‡∏£‡∏û‡∏ô‡∏°', '‡∏ö‡∏∂‡∏á‡∏Å‡∏≤‡∏¨', '‡∏°‡∏∏‡∏Å‡∏î‡∏≤‡∏´‡∏≤‡∏£', '‡∏™‡∏Å‡∏•‡∏ô‡∏Ñ‡∏£', '‡∏´‡∏ô‡∏≠‡∏á‡∏Ñ‡∏≤‡∏¢', '‡∏´‡∏ô‡∏≠‡∏á‡∏ö‡∏±‡∏ß‡∏•‡∏≥‡∏†‡∏π', '‡∏≠‡∏∏‡∏î‡∏£‡∏ò‡∏≤‡∏ô‡∏µ', '‡πÄ‡∏•‡∏¢'],
        '‡∏≠‡∏µ‡∏™‡∏≤‡∏ô‡∏Å‡∏•‡∏≤‡∏á': ['‡∏Å‡∏≤‡∏¨‡∏™‡∏¥‡∏ô‡∏ò‡∏∏‡πå', '‡∏Ç‡∏≠‡∏ô‡πÅ‡∏Å‡πà‡∏ô', '‡∏ä‡∏±‡∏¢‡∏†‡∏π‡∏°‡∏¥', '‡∏°‡∏´‡∏≤‡∏™‡∏≤‡∏£‡∏Ñ‡∏≤‡∏°', '‡∏£‡πâ‡∏≠‡∏¢‡πÄ‡∏≠‡πá‡∏î'],
        '‡∏≠‡∏µ‡∏™‡∏≤‡∏ô‡πÉ‡∏ï‡πâ': ['‡∏ô‡∏Ñ‡∏£‡∏£‡∏≤‡∏ä‡∏™‡∏µ‡∏°‡∏≤', '‡πÇ‡∏Ñ‡∏£‡∏≤‡∏ä', '‡∏ö‡∏∏‡∏£‡∏µ‡∏£‡∏±‡∏°‡∏¢‡πå', '‡∏¢‡πÇ‡∏™‡∏ò‡∏£', '‡∏®‡∏£‡∏µ‡∏™‡∏∞‡πÄ‡∏Å‡∏©', '‡∏™‡∏∏‡∏£‡∏¥‡∏ô‡∏ó‡∏£‡πå', '‡∏≠‡∏≥‡∏ô‡∏≤‡∏à‡πÄ‡∏à‡∏£‡∏¥‡∏ç', '‡∏≠‡∏∏‡∏ö‡∏•‡∏£‡∏≤‡∏ä‡∏ò‡∏≤‡∏ô‡∏µ'],
        '‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ï‡∏≠‡∏ô‡∏ö‡∏ô': ['‡∏ô‡πà‡∏≤‡∏ô', '‡∏û‡∏∞‡πÄ‡∏¢‡∏≤', '‡∏•‡∏≥‡∏õ‡∏≤‡∏á', '‡∏•‡∏≥‡∏û‡∏π‡∏ô', '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡∏£‡∏≤‡∏¢', '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡πÉ‡∏´‡∏°‡πà', '‡πÅ‡∏û‡∏£‡πà', '‡πÅ‡∏°‡πà‡∏Æ‡πà‡∏≠‡∏á‡∏™‡∏≠‡∏ô'],
        '‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ï‡∏≠‡∏ô‡∏•‡πà‡∏≤‡∏á': ['‡∏Å‡∏≥‡πÅ‡∏û‡∏á‡πÄ‡∏û‡∏ä‡∏£', '‡∏ï‡∏≤‡∏Å', '‡∏ô‡∏Ñ‡∏£‡∏™‡∏ß‡∏£‡∏£‡∏Ñ‡πå', '‡∏û‡∏¥‡∏à‡∏¥‡∏ï‡∏£', '‡∏û‡∏¥‡∏©‡∏ì‡∏∏‡πÇ‡∏•‡∏Å', '‡∏™‡∏∏‡πÇ‡∏Ç‡∏ó‡∏±‡∏¢', '‡∏≠‡∏∏‡∏ï‡∏£‡∏î‡∏¥‡∏ï‡∏ñ‡πå', '‡∏≠‡∏∏‡∏ó‡∏±‡∏¢‡∏ò‡∏≤‡∏ô‡∏µ', '‡πÄ‡∏û‡∏ä‡∏£‡∏ö‡∏π‡∏£‡∏ì‡πå'],
        '‡πÉ‡∏ï‡πâ‡∏ù‡∏±‡πà‡∏á‡∏≠‡∏±‡∏ô‡∏î‡∏≤‡∏°‡∏±‡∏ô': ['‡∏Å‡∏£‡∏∞‡∏ö‡∏µ‡πà', '‡∏ï‡∏£‡∏±‡∏á', '‡∏û‡∏±‡∏á‡∏á‡∏≤', '‡∏†‡∏π‡πÄ‡∏Å‡πá‡∏ï', '‡∏£‡∏∞‡∏ô‡∏≠‡∏á', '‡∏™‡∏ï‡∏π‡∏•'],
        '‡πÉ‡∏ï‡πâ‡∏ù‡∏±‡πà‡∏á‡∏≠‡πà‡∏≤‡∏ß‡πÑ‡∏ó‡∏¢': ['‡∏ä‡∏∏‡∏°‡∏û‡∏£', '‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏ò‡∏£‡∏£‡∏°‡∏£‡∏≤‡∏ä', '‡∏û‡∏±‡∏ó‡∏•‡∏∏‡∏á', '‡∏¢‡∏∞‡∏•‡∏≤', '‡∏™‡∏á‡∏Ç‡∏•‡∏≤', '‡∏™‡∏∏‡∏£‡∏≤‡∏©‡∏é‡∏£‡πå‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏õ‡∏±‡∏ï‡∏ï‡∏≤‡∏ô‡∏µ', '‡∏ô‡∏£‡∏≤‡∏ò‡∏¥‡∏ß‡∏≤‡∏™']
    }
    
    # ‡∏´‡∏≤‡∏ß‡πà‡∏≤‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏ó‡∏±‡πâ‡∏á 2 ‡∏≠‡∏¢‡∏π‡πà‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
    for group, provinces in province_groups.items():
        in_group_1 = any(p in str(prov1) for p in provinces)
        in_group_2 = any(p in str(prov2) for p in provinces)
        
        if in_group_1 and in_group_2:
            return True
    
    return False

def load_model():
    """‡πÇ‡∏´‡∏•‡∏î‡πÇ‡∏°‡πÄ‡∏î‡∏•‡∏ó‡∏µ‡πà‡πÄ‡∏ó‡∏£‡∏ô‡πÑ‡∏ß‡πâ"""
    if not os.path.exists(MODEL_PATH):
        return None
    
    try:
        with open(MODEL_PATH, 'rb') as f:
            model_data = pickle.load(f)
        return model_data
    except Exception as e:
        st.error(f"‚ùå Error loading model: {e}")
        return None

def create_pair_features(code1, code2, branch_info):
    """‡∏™‡∏£‡πâ‡∏≤‡∏á features ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏Ñ‡∏π‡πà‡∏™‡∏≤‡∏Ç‡∏≤"""
    import math
    
    info1 = branch_info[code1]
    info2 = branch_info[code2]
    
    # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏Ñ‡∏ß‡∏≤‡∏°‡∏ï‡πà‡∏≤‡∏á‡∏Ç‡∏≠‡∏á‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å‡πÅ‡∏•‡∏∞‡∏Ñ‡∏¥‡∏ß
    weight_diff = abs(info1['avg_weight'] - info2['avg_weight'])
    cube_diff = abs(info1['avg_cube'] - info2['avg_cube'])
    weight_sum = info1['avg_weight'] + info2['avg_weight']
    cube_sum = info1['avg_cube'] + info2['avg_cube']
    
    # ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
    same_province = 1 if info1['province'] == info2['province'] else 0
    
    # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å‡∏û‡∏¥‡∏Å‡∏±‡∏î
    distance_km = 0.0
    if info1['latitude'] != 0 and info2['latitude'] != 0:
        lat1, lon1 = math.radians(info1['latitude']), math.radians(info1['longitude'])
        lat2, lon2 = math.radians(info2['latitude']), math.radians(info2['longitude'])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        distance_km = 6371 * c
    
    # ‡∏Ñ‡∏ß‡∏≤‡∏°‡∏ñ‡∏µ‡πà
    freq_product = info1['total_trips'] * info2['total_trips']
    freq_diff = abs(info1['total_trips'] - info2['total_trips'])
    
    # Ratio
    weight_ratio = (info1['avg_weight'] / info2['avg_weight']) if info2['avg_weight'] > 0 else 0
    cube_ratio = (info1['avg_cube'] / info2['avg_cube']) if info2['avg_cube'] > 0 else 0
    
    # ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ
    over_4w = 1 if (weight_sum > 2500 or cube_sum > 5.0) else 0
    over_jb = 1 if (weight_sum > 3500 or cube_sum > 7.0) else 0
    over_6w = 1 if (weight_sum > 5800 or cube_sum > 22.0) else 0
    
    return {
        'weight_sum': weight_sum,
        'cube_sum': cube_sum,
        'weight_diff': weight_diff,
        'cube_diff': cube_diff,
        'same_province': same_province,
        'distance_km': distance_km,
        'avg_weight_1': info1['avg_weight'],
        'avg_weight_2': info2['avg_weight'],
        'avg_cube_1': info1['avg_cube'],
        'avg_cube_2': info2['avg_cube'],
        'freq_product': freq_product,
        'freq_diff': freq_diff,
        'weight_ratio': weight_ratio,
        'cube_ratio': cube_ratio,
        'over_4w': over_4w,
        'over_jb': over_jb,
        'over_6w': over_6w
    }

def load_info_sheet_truck_limits(xls):
    """
    üîí ‡∏≠‡πà‡∏≤‡∏ô‡∏ä‡∏µ‡∏ï info ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå Auto Plan ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏î‡∏∂‡∏á MaxTruckType
    
    ‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£:
    - Location Code: ‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤
    - MaxTruckType: ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ‡∏ö‡∏£‡∏£‡∏ó‡∏∏‡∏Å‡πÉ‡∏´‡∏ç‡πà‡∏™‡∏∏‡∏î (‡πÄ‡∏ä‡πà‡∏ô 4W, 6W, 10W)
    """
    global AUTO_PLAN_TRUCK_LIMITS
    
    try:
        # ‡∏´‡∏≤‡∏ä‡∏µ‡∏ï info
        info_sheet = None
        for s in xls.sheet_names:
            if 'info' in s.lower():
                info_sheet = s
                break
        
        if not info_sheet:
            return  # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ä‡∏µ‡∏ï info
        
        # ‡∏≠‡πà‡∏≤‡∏ô‡∏ä‡∏µ‡∏ï info
        df_info = pd.read_excel(xls, sheet_name=info_sheet)
        
        # ‡∏´‡∏≤‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå Location Code ‡πÅ‡∏•‡∏∞ MaxTruckType
        location_col = None
        truck_col = None
        
        for col in df_info.columns:
            col_str = str(col).lower()
            if 'location' in col_str and 'code' in col_str:
                location_col = col
            elif 'maxtruck' in col_str or 'max truck' in col_str:
                truck_col = col
        
        if location_col and truck_col:
            # ‡∏™‡∏£‡πâ‡∏≤‡∏á mapping
            AUTO_PLAN_TRUCK_LIMITS.clear()
            
            for _, row in df_info.iterrows():
                code = str(row[location_col]).strip()
                truck_type = str(row[truck_col]).strip().upper()
                
                if code and truck_type and code != 'nan' and truck_type != 'NAN':
                    # üîí ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏°‡∏µ 6W ‡πÉ‡∏ô‡πÄ‡∏ã‡∏•‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà (‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ = ‡∏´‡πâ‡∏≤‡∏° 6W)
                    # ‡∏ï‡∏±‡∏ß‡∏≠‡∏¢‡πà‡∏≤‡∏á: "6W" = ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡∏ó‡∏∏‡∏Å‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó, "4WJB" = ‡∏´‡πâ‡∏≤‡∏° 6W (‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡πÅ‡∏Ñ‡πà 4W, JB)
                    has_6w = '6W' in truck_type or '10W' in truck_type
                    has_jb = 'JB' in truck_type or 'JUMBO' in truck_type
                    has_4w = '4W' in truck_type
                    
                    # ‡∏Å‡∏≥‡∏´‡∏ô‡∏î max vehicle ‡∏ï‡∏≤‡∏°‡∏ó‡∏µ‡πà‡∏£‡∏∞‡∏ö‡∏∏‡πÉ‡∏ô‡πÑ‡∏ü‡∏•‡πå
                    if has_6w:
                        # ‡∏°‡∏µ 6W ‡∏´‡∏£‡∏∑‡∏≠ 10W = ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡∏ó‡∏∏‡∏Å‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó
                        normalized_truck = '6W'
                    elif has_jb or truck_type == '4WJB':
                        # ‡∏°‡∏µ JB ‡∏´‡∏£‡∏∑‡∏≠ "4WJB" ‡πÅ‡∏ï‡πà‡πÑ‡∏°‡πà‡∏°‡∏µ 6W = ‡∏´‡πâ‡∏≤‡∏° 6W (‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡πÅ‡∏Ñ‡πà 4W, JB)
                        normalized_truck = 'JB'
                    elif has_4w:
                        # ‡∏°‡∏µ‡πÅ‡∏Ñ‡πà 4W = ‡∏´‡πâ‡∏≤‡∏° 6W ‡πÅ‡∏•‡∏∞ JB
                        normalized_truck = '4W'
                    else:
                        # ‡πÑ‡∏°‡πà‡∏£‡∏∞‡∏ö‡∏∏ = ‡∏´‡πâ‡∏≤‡∏° 6W (default ‡πÄ‡∏õ‡πá‡∏ô JB)
                        normalized_truck = 'JB'
                    
                    AUTO_PLAN_TRUCK_LIMITS[code] = normalized_truck
            
            if AUTO_PLAN_TRUCK_LIMITS:
                # ‡∏ô‡∏±‡∏ö‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó
                count_4w = sum(1 for v in AUTO_PLAN_TRUCK_LIMITS.values() if v == '4W')
                count_jb = sum(1 for v in AUTO_PLAN_TRUCK_LIMITS.values() if v == 'JB')
                count_6w = sum(1 for v in AUTO_PLAN_TRUCK_LIMITS.values() if v == '6W')
                st.info(f"üìã ‡πÇ‡∏´‡∏•‡∏î‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡∏à‡∏≤‡∏Å‡∏ä‡∏µ‡∏ï info: {len(AUTO_PLAN_TRUCK_LIMITS)} ‡∏™‡∏≤‡∏Ç‡∏≤ (4W: {count_4w}, JB: {count_jb}, 6W: {count_6w})")
                
    except Exception as e:
        # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ä‡∏µ‡∏ï info ‡∏´‡∏£‡∏∑‡∏≠‡∏≠‡πà‡∏≤‡∏ô‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ ‚Üí ‡πÑ‡∏°‡πà‡πÄ‡∏õ‡πá‡∏ô‡πÑ‡∏£ ‡πÉ‡∏ä‡πâ default
        pass

def load_excel(file_content, sheet_name=None):
    """‡πÇ‡∏´‡∏•‡∏î Excel ‡πÅ‡∏•‡∏∞‡∏≠‡πà‡∏≤‡∏ô‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡∏à‡∏≤‡∏Å‡∏ä‡∏µ‡∏ï info"""
    global AUTO_PLAN_TRUCK_LIMITS
    
    try:
        xls = pd.ExcelFile(io.BytesIO(file_content))
        
        # üîí ‡∏•‡∏≠‡∏á‡∏≠‡πà‡∏≤‡∏ô‡∏ä‡∏µ‡∏ï info ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏î‡∏∂‡∏á MaxTruckType
        load_info_sheet_truck_limits(xls)
        
        target_sheet = None
        if sheet_name and sheet_name in xls.sheet_names:
            target_sheet = sheet_name
        else:
            for s in xls.sheet_names:
                if 'punthai' in s.lower() or '2.' in s.lower():
                    target_sheet = s
                    break
        
        if not target_sheet:
            target_sheet = xls.sheet_names[0]
        
        # ‡∏´‡∏≤ header row
        df_temp = pd.read_excel(xls, sheet_name=target_sheet, header=None)
        header_row = 0
        
        for i in range(min(10, len(df_temp))):
            row_values = df_temp.iloc[i].astype(str).str.upper()
            match_count = sum([
                'BRANCH' in ' '.join(row_values),
                'TRIP' in ' '.join(row_values),
                '‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤' in ' '.join(df_temp.iloc[i].astype(str))
            ])
            if match_count >= 2:
                header_row = i
                break
        
        df = pd.read_excel(xls, sheet_name=target_sheet, header=header_row)
        df = df.loc[:, ~df.columns.duplicated()]
        
        return df
    except Exception as e:
        st.error(f"‚ùå Error: {e}")
        return None

def process_dataframe(df):
    """‡πÅ‡∏õ‡∏•‡∏á‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡πÄ‡∏õ‡πá‡∏ô‡∏£‡∏π‡∏õ‡πÅ‡∏ö‡∏ö‡∏°‡∏≤‡∏ï‡∏£‡∏ê‡∏≤‡∏ô"""
    if df is None:
        return None
    
    rename_map = {}
    
    # ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ô‡πâ‡∏≠‡∏¢‡∏Å‡∏ß‡πà‡∏≤ 15 = ‡πÉ‡∏ä‡πâ‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå
    # ‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏°‡∏≤‡∏ï‡∏£‡∏ê‡∏≤‡∏ô: Sep, BU, ‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤, ‡∏£‡∏´‡∏±‡∏™ WMS, ‡∏™‡∏≤‡∏Ç‡∏≤, Total Cube, Total Wgt, ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏ä‡∏¥‡πâ‡∏ô, Trip, Trip no, ...
    if len(df.columns) >= 8:
        col_list = list(df.columns)
        # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 1 = BU
        if len(col_list) > 1:
            rename_map[col_list[1]] = 'BU'
        # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 2 = ‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤
        if len(col_list) > 2:
            rename_map[col_list[2]] = 'Code'
        # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 4 = ‡∏™‡∏≤‡∏Ç‡∏≤/‡∏ä‡∏∑‡πà‡∏≠
        if len(col_list) > 4:
            rename_map[col_list[4]] = 'Name'
        # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 5 = Total Cube
        if len(col_list) > 5:
            rename_map[col_list[5]] = 'Cube'
        # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 6 = Total Wgt
        if len(col_list) > 6:
            rename_map[col_list[6]] = 'Weight'
        # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 7 = ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏ä‡∏¥‡πâ‡∏ô (Original QTY)
        if len(col_list) > 7:
            rename_map[col_list[7]] = 'OriginalQty'
        # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 8 = Trip
        if len(col_list) > 8:
            rename_map[col_list[8]] = 'Trip'
        # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 9 = Trip no
        if len(col_list) > 9:
            rename_map[col_list[9]] = 'TripNo'
    
    # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡πÄ‡∏û‡∏¥‡πà‡∏°‡πÄ‡∏ï‡∏¥‡∏°‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå
    for col in df.columns:
        if col in rename_map:
            continue
        col_clean = str(col).strip()
        col_upper = col_clean.upper().replace(' ', '').replace('_', '')
        
        if col_clean == 'BranchCode' or '‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤' in col_clean or col_clean == '‡∏£‡∏´‡∏±‡∏™ WMS' or 'BRANCH_CODE' in col_upper:
            rename_map[col] = 'Code'
        elif col_clean == 'Branch' or '‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤' in col_clean or col_clean == '‡∏™‡∏≤‡∏Ç‡∏≤' or 'BRANCH' in col_upper:
            rename_map[col] = 'Name'
        elif 'TOTALWGT' in col_upper or '‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å' in col_clean or 'WGT' in col_upper or 'WEIGHT' in col_upper:
            rename_map[col] = 'Weight'
        elif 'TOTALCUBE' in col_upper or '‡∏Ñ‡∏¥‡∏ß' in col_clean or 'CUBE' in col_upper:
            rename_map[col] = 'Cube'
        elif 'latitude' in col_clean.lower() or col_clean == '‡∏•‡∏∞‡∏ï‡∏¥‡∏à‡∏π‡∏î' or 'LAT' in col_upper:
            rename_map[col] = 'Latitude'
        elif 'longitude' in col_clean.lower() or col_clean == '‡∏•‡∏≠‡∏á‡∏ï‡∏¥‡∏à‡∏π‡∏î' or 'LONG' in col_upper or 'LNG' in col_upper:
            rename_map[col] = 'Longitude'
        elif '‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î' in col_clean or 'PROVINCE' in col_upper:
            rename_map[col] = 'Province'
        elif col_upper in ['TRIPNO', 'TRIP_NO'] or col_clean == 'Trip no':
            rename_map[col] = 'TripNo'
        elif col_upper == 'TRIP' or '‡∏ó‡∏£‡∏¥‡∏õ' in col_clean or '‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏ß' in col_clean:
            rename_map[col] = 'Trip'
        elif 'BOOKING' in col_upper:
            rename_map[col] = 'Booking'
    
    df = df.rename(columns=rename_map)
    
    # ‡∏•‡∏ö‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ã‡πâ‡∏≥
    df = df.loc[:, ~df.columns.duplicated()]
    
    if 'Code' in df.columns:
        df['Code'] = df['Code'].apply(normalize)
        
        # ‡∏ï‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£‡∏≠‡∏≠‡∏Å (‡∏£‡∏´‡∏±‡∏™)
        df = df[~df['Code'].isin(EXCLUDE_BRANCHES)]
        
        # ‡∏ï‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏ä‡∏∑‡πà‡∏≠‡∏°‡∏µ keyword ‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏Å‡∏≤‡∏£
        if 'Name' in df.columns:
            exclude_pattern = '|'.join(EXCLUDE_NAMES)
            df = df[~df['Name'].str.contains(exclude_pattern, case=False, na=False)]
    
    for col in ['Weight', 'Cube']:
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    
    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏à‡∏≤‡∏Å Master ‡∏ñ‡πâ‡∏≤‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡∏°‡∏µ
    if 'Province' not in df.columns or df['Province'].isna().all():
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns and 'Code' in df.columns:
            # ‡∏™‡∏£‡πâ‡∏≤‡∏á mapping ‡∏à‡∏≤‡∏Å Master
            province_map = {}
            for _, row in MASTER_DATA.iterrows():
                code = row.get('Plan Code', '')
                province = row.get('‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', '')
                if code and province:
                    province_map[code] = province
            
            # ‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô‡∏Ñ‡πâ‡∏ô‡∏´‡∏≤‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤
            def find_province_by_name(code, name):
                # ‡∏•‡∏≠‡∏á‡∏´‡∏≤‡∏à‡∏≤‡∏Å code ‡∏Å‡πà‡∏≠‡∏ô
                if code in province_map:
                    return province_map[code]
                
                # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡πÄ‡∏à‡∏≠ ‡∏•‡∏≠‡∏á‡∏Ñ‡πâ‡∏ô‡∏´‡∏≤‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤
                if not name or pd.isna(name):
                    return None
                
                # ‡πÅ‡∏¢‡∏Å‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠ (‡πÄ‡∏≠‡∏≤‡∏Ñ‡∏≥‡πÅ‡∏£‡∏Å‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πà prefix)
                keywords = str(name).replace('MAX MART-', '').replace('PUNTHAI-', '').replace('LUBE', '').strip()
                if not keywords:
                    return None
                
                # ‡∏Ñ‡πâ‡∏ô‡∏´‡∏≤‡πÉ‡∏ô‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤‡∏Ç‡∏≠‡∏á Master
                for _, master_row in MASTER_DATA.iterrows():
                    master_name = str(master_row.get('‡∏™‡∏≤‡∏Ç‡∏≤', ''))
                    # ‡∏ñ‡πâ‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô (‡∏°‡∏µ‡∏Ñ‡∏≥‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô‡∏Å‡∏±‡∏ô)
                    if keywords[:10] in master_name or master_name[:10] in keywords:
                        province = master_row.get('‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', '')
                        if province:
                            return province
                
                return None
            
            # ‡πÉ‡∏™‡πà‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÉ‡∏´‡πâ‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏™‡∏≤‡∏Ç‡∏≤
            if 'Name' in df.columns:
                df['Province'] = df.apply(lambda row: find_province_by_name(row['Code'], row.get('Name', '')), axis=1)
            else:
                df['Province'] = df['Code'].map(province_map)
    
    return df.reset_index(drop=True)

def predict_trips(test_df, model_data):
    """
    ‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ‡∏î‡πâ‡∏ß‡∏¢‡∏£‡∏∞‡∏ö‡∏ö‡∏≠‡∏±‡∏à‡∏â‡∏£‡∏¥‡∏¢‡∏∞ ‡πÇ‡∏î‡∏¢‡πÉ‡∏ä‡πâ‡∏Å‡∏é‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏Ñ‡∏ß‡∏≤‡∏°‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç:
    
    ‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö (‡∏ï‡πâ‡∏≠‡∏á‡∏ú‡πà‡∏≤‡∏ô‡∏Å‡πà‡∏≠‡∏ô):
    0. ‚úÖ ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡πà‡∏≠‡∏ó‡∏£‡∏¥‡∏õ (‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î 12 ‡∏™‡∏≤‡∏Ç‡∏≤)
    0. ‚úÖ ‡πÄ‡∏ä‡πá‡∏Ñ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á‡∏Å‡∏±‡∏ô (‡∏ï‡πâ‡∏≠‡∏á‡∏≠‡∏¢‡∏π‡πà‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô)
    0. ‚úÖ ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 10 ‡∏™‡∏≤‡∏Ç‡∏≤ ‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏õ‡πá‡∏ô‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
    0. ‚úÖ ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏ô‡∏µ‡πâ‡πÑ‡∏î‡πâ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà‡∏à‡∏≤‡∏Å‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥
    
    ‡∏Å‡∏é‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏ö‡∏Ñ‡∏π‡πà (‡∏´‡∏•‡∏±‡∏á‡∏ú‡πà‡∏≤‡∏ô‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö):
    1. ‚úÖ ‡πÄ‡∏Ñ‡∏¢‡πÑ‡∏õ‡∏î‡πâ‡∏ß‡∏¢‡∏Å‡∏±‡∏ô‡πÉ‡∏ô‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥ (trip_pairs) + ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÅ‡∏ö‡∏ö‡πÄ‡∏î‡∏¥‡∏°
    2. ‚úÖ ‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô (‡πÄ‡∏ä‡πà‡∏ô ‡∏ô‡∏Ñ‡∏£‡∏£‡∏≤‡∏ä‡∏™‡∏µ‡∏°‡∏≤1, ‡∏ô‡∏Ñ‡∏£‡∏£‡∏≤‡∏ä‡∏™‡∏µ‡∏°‡∏≤2)
    3. ‚úÖ AI ‡∏ó‡∏≥‡∏ô‡∏≤‡∏¢‡∏à‡∏≤‡∏Å Decision Tree Model
    4. ‚úÖ ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å/‡∏Ñ‡∏¥‡∏ß ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô‡∏Ç‡∏µ‡∏î‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ
    """
    # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤ model_data ‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏Ñ‡∏£‡∏ö‡∏ñ‡πâ‡∏ß‡∏ô
    if not model_data or not isinstance(model_data, dict):
        st.error("‚ùå ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÇ‡∏°‡πÄ‡∏î‡∏•‡πÑ‡∏°‡πà‡∏ñ‡∏π‡∏Å‡∏ï‡πâ‡∏≠‡∏á ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡πÄ‡∏ó‡∏£‡∏ô‡πÇ‡∏°‡πÄ‡∏î‡∏•‡∏Å‡πà‡∏≠‡∏ô")
        return test_df, []
    
    model = model_data.get('model')
    trip_pairs = model_data.get('trip_pairs', set()).copy()  # ‡∏Ñ‡∏±‡∏î‡∏•‡∏≠‡∏Å‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÑ‡∏°‡πà‡πÉ‡∏´‡πâ‡∏Å‡∏£‡∏∞‡∏ó‡∏ö‡∏ï‡πâ‡∏ô‡∏â‡∏ö‡∏±‡∏ö
    branch_info = model_data.get('branch_info', {})
    trip_vehicles = model_data.get('trip_vehicles', {}).copy()
    branch_vehicles = model_data.get('branch_vehicles', {})
    
    # ‚ö° ‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï‡πÉ‡∏´‡πâ‡∏ó‡∏≥‡∏á‡∏≤‡∏ô‡πÇ‡∏î‡∏¢‡πÑ‡∏°‡πà‡∏°‡∏µ model (‡πÉ‡∏ä‡πâ‡∏Å‡∏é‡πÅ‡∏•‡∏∞‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô)
    # if model is None:
    #     st.error("‚ùå ‡πÑ‡∏°‡πà‡∏û‡∏ö‡πÇ‡∏°‡πÄ‡∏î‡∏• ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡πÄ‡∏ó‡∏£‡∏ô‡πÇ‡∏°‡πÄ‡∏î‡∏•‡∏Å‡πà‡∏≠‡∏ô")
    #     return test_df, []
    
    # ‚òÖ ‡∏ñ‡πâ‡∏≤‡πÑ‡∏ü‡∏•‡πå‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡∏°‡∏µ‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå Trip ‡πÉ‡∏´‡πâ‡πÉ‡∏ä‡πâ‡πÄ‡∏õ‡πá‡∏ô‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏≠‡πâ‡∏≤‡∏á‡∏≠‡∏¥‡∏á‡∏´‡∏•‡∏±‡∏Å
    # ‡πÄ‡∏û‡∏£‡∏≤‡∏∞‡πÄ‡∏õ‡πá‡∏ô‡πÅ‡∏ú‡∏ô‡∏á‡∏≤‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡∏à‡∏£‡∏¥‡∏á‡∏°‡∏≤‡πÅ‡∏•‡πâ‡∏ß
    file_trip_vehicles = {}  # ‡πÄ‡∏Å‡πá‡∏ö‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡πÅ‡∏ú‡∏ô‡∏á‡∏≤‡∏ô
    use_file_trips = False  # ‡πÉ‡∏ä‡πâ‡∏ó‡∏£‡∏¥‡∏õ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡πÇ‡∏î‡∏¢‡∏ï‡∏£‡∏á
    
    if 'Trip' in test_df.columns and test_df['Trip'].notna().any():
        use_file_trips = True
        st.info(f"üìã ‡∏û‡∏ö‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏ô‡πÑ‡∏ü‡∏•‡πå - ‡πÉ‡∏ä‡πâ‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡πÇ‡∏î‡∏¢‡∏ï‡∏£‡∏á")
        
        # ‡∏™‡∏£‡πâ‡∏≤‡∏á trip_pairs ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡πÅ‡∏ú‡∏ô‡∏á‡∏≤‡∏ô
        for trip_id, group in test_df.groupby('Trip'):
            if pd.isna(trip_id):
                continue
            codes = group['Code'].unique().tolist()
            
            # ‡∏î‡∏∂‡∏á‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ‡∏à‡∏≤‡∏Å TripNo (‡πÄ‡∏ä‡πà‡∏ô 4W009 -> 4W, JB014 -> JB)
            if 'TripNo' in group.columns:
                trip_no = group['TripNo'].iloc[0]
                if pd.notna(trip_no):
                    trip_no_str = str(trip_no).strip()
                    if trip_no_str.startswith('4W'):
                        vehicle_type = '4W'
                    elif trip_no_str.startswith('JB'):
                        vehicle_type = 'JB'
                    elif trip_no_str.startswith('6W'):
                        vehicle_type = '6W'
                    else:
                        vehicle_type = None
                    
                    # ‡πÄ‡∏Å‡πá‡∏ö‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏Ñ‡∏π‡πà‡∏™‡∏≤‡∏Ç‡∏≤
                    if vehicle_type:
                        for i in range(len(codes)):
                            for j in range(i+1, len(codes)):
                                pair = tuple(sorted([codes[i], codes[j]]))
                                file_trip_vehicles[pair] = vehicle_type
            
            # ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏Ñ‡∏π‡πà‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
            for i in range(len(codes)):
                for j in range(i+1, len(codes)):
                    pair = tuple(sorted([codes[i], codes[j]]))
                    trip_pairs.add(pair)  # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡πÄ‡∏Ç‡πâ‡∏≤ trip_pairs
    
    # ‡∏£‡∏ß‡∏° file_trip_vehicles ‡πÄ‡∏Ç‡πâ‡∏≤‡∏Å‡∏±‡∏ö trip_vehicles (‡πÉ‡∏´‡πâ‡πÑ‡∏ü‡∏•‡πå‡∏°‡∏µ‡∏Ñ‡∏ß‡∏≤‡∏°‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏Å‡∏ß‡πà‡∏≤)
    for pair, vehicle in file_trip_vehicles.items():
        trip_vehicles[pair] = {'most_used': vehicle, 'vehicle': vehicle}
    
    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏´‡∏°‡πà
    for code in test_df['Code'].unique():
        if code not in branch_info:
            code_data = test_df[test_df['Code'] == code]
            branch_info[code] = {
                'avg_weight': code_data['Weight'].mean(),
                'avg_cube': code_data['Cube'].mean(),
                'total_trips': 1,
                'province': code_data['Province'].iloc[0] if 'Province' in code_data.columns else 'UNKNOWN',
                'latitude': 0.0,
                'longitude': 0.0
            }
    
    # ‚òÖ‚òÖ‚òÖ ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå Trip ‡πÉ‡∏ô‡πÑ‡∏ü‡∏•‡πå ‡πÉ‡∏ä‡πâ‡πÇ‡∏î‡∏¢‡∏ï‡∏£‡∏á‡πÄ‡∏•‡∏¢ ‚òÖ‚òÖ‚òÖ
    if use_file_trips:
        # ‡πÉ‡∏ä‡πâ Trip ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡πÇ‡∏î‡∏¢‡∏ï‡∏£‡∏á
        test_df_result = test_df.copy()
        
        # ‡∏î‡∏∂‡∏á‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ‡∏à‡∏≤‡∏Å TripNo
        trip_truck_map_file = {}
        if 'TripNo' in test_df.columns:
            for trip_id in test_df['Trip'].dropna().unique():
                trip_data = test_df[test_df['Trip'] == trip_id]
                if 'TripNo' in trip_data.columns and len(trip_data) > 0:
                    trip_no = trip_data['TripNo'].iloc[0]
                    if pd.notna(trip_no):
                        trip_no_str = str(trip_no).strip()
                        if trip_no_str.startswith('4W'):
                            trip_truck_map_file[trip_id] = '4W'
                        elif trip_no_str.startswith('JB'):
                            trip_truck_map_file[trip_id] = 'JB'
                        elif trip_no_str.startswith('6W'):
                            trip_truck_map_file[trip_id] = '6W'
        
        # ‡∏™‡∏£‡πâ‡∏≤‡∏á summary
        summary_data = []
        for trip_num in sorted(test_df_result['Trip'].dropna().unique()):
            trip_data = test_df_result[test_df_result['Trip'] == trip_num]
            total_w = trip_data['Weight'].sum()
            total_c = trip_data['Cube'].sum()
            
            # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏Ç‡∏≠‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
            trip_codes = trip_data['Code'].unique()
            max_vehicles = []
            for c in trip_codes:
                max_vehicles.append(get_max_vehicle_for_branch(c))
            
            vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
            min_max_size = min(vehicle_sizes.get(v, 3) for v in max_vehicles)
            max_allowed_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(min_max_size, '6W')
            

            # üö® STRICT: Branch constraint (üîí) > History (üìú) > AI (ü§ñ)
            # 1. Branch constraint (never allow 6W if any branch restricts to 4W/JB)
            if min_max_size < 3:  # 1=4W, 2=JB
                # Only allow 4W/JB, never 6W
                allowed = ['JB', '4W'] if min_max_size == 2 else ['4W']
            else:
                allowed = ['JB', '4W', '6W']

            if trip_num in trip_truck_map_file:
                suggested = trip_truck_map_file[trip_num]
                # If suggested vehicle is not allowed, override to strictest allowed
                if suggested not in allowed:
                    suggested = allowed[0]
                    source = f"üìã ‡πÑ‡∏ü‡∏•‡πå ‚Üí {suggested} (üîí ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤)"
                else:
                    source = "üìã ‡πÑ‡∏ü‡∏•‡πå"
            else:
                # AI suggestion, but must respect allowed
                ai_suggested = suggest_truck(total_w, total_c, max_allowed_vehicle, trip_codes)
                if ai_suggested not in allowed:
                    suggested = allowed[0]
                    source = f"ü§ñ AI ‚Üí {suggested} (üîí ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤)"
                else:
                    suggested = ai_suggested
                    source = "ü§ñ AI"

            # Double check: If strict constraint, never allow 6W even if utilization >105%
            if min_max_size < 3:
                # Only JB or 4W allowed, never 6W
                if suggested == '6W':
                    # fallback to JB if possible, else 4W
                    if 'JB' in allowed:
                        suggested = 'JB'
                        source = source + " (üîí ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤)"
                    else:
                        suggested = '4W'
                        source = source + " (üîí ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤)"

            # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÉ‡∏™‡πà‡∏Ç‡∏≠‡∏á‡πÑ‡∏î‡πâ‡∏à‡∏£‡∏¥‡∏á‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà (‡∏´‡πâ‡∏≤‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 100%)
            if suggested in LIMITS:
                w_util = (total_w / LIMITS[suggested]['max_w']) * 100
                c_util = (total_c / LIMITS[suggested]['max_c']) * 100
                max_util = max(w_util, c_util)

                # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 100% ‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ç‡∏ô‡∏≤‡∏î‡∏£‡∏ñ
                if max_util > 100:
                    # ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤ ‡∏´‡πâ‡∏≤‡∏°‡∏Ç‡∏¢‡∏≤‡∏¢‡πÄ‡∏õ‡πá‡∏ô 6W
                    if min_max_size < 3:
                        # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö JB ‡∏´‡∏£‡∏∑‡∏≠ 4W ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
                        if 'JB' in allowed and suggested == '4W':
                            jb_w_util = (total_w / LIMITS['JB']['max_w']) * 100
                            jb_c_util = (total_c / LIMITS['JB']['max_c']) * 100
                            if max(jb_w_util, jb_c_util) <= 100:
                                suggested = 'JB'
                                source = source + " ‚Üí JB"
                                w_util, c_util = jb_w_util, jb_c_util
                        # ‡∏ñ‡πâ‡∏≤ JB ‡∏Å‡πá‡∏¢‡∏±‡∏á‡πÄ‡∏Å‡∏¥‡∏ô ‡πÉ‡∏´‡πâ‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô‡∏ß‡πà‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô ‡πÑ‡∏°‡πà‡∏Ç‡∏¢‡∏≤‡∏¢‡πÄ‡∏õ‡πá‡∏ô 6W
                        elif suggested == 'JB':
                            source = source + " (üö´ ‡πÄ‡∏Å‡∏¥‡∏ô‡∏Ç‡∏ô‡∏≤‡∏î‡πÅ‡∏ï‡πà‡∏´‡πâ‡∏≤‡∏°‡πÉ‡∏ä‡πâ 6W)"
                    else:
                        # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤ ‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡∏Ç‡∏¢‡∏≤‡∏¢‡πÄ‡∏õ‡πá‡∏ô 6W ‡πÑ‡∏î‡πâ
                        if suggested == '4W' and 'JB' in LIMITS:
                            jb_w_util = (total_w / LIMITS['JB']['max_w']) * 100
                            jb_c_util = (total_c / LIMITS['JB']['max_c']) * 100
                            if max(jb_w_util, jb_c_util) <= 100:
                                suggested = 'JB'
                                source = source + " ‚Üí JB"
                                w_util, c_util = jb_w_util, jb_c_util
                            else:
                                suggested = '6W'
                                source = source + " ‚Üí 6W"
                                w_util = (total_w / LIMITS['6W']['max_w']) * 100
                                c_util = (total_c / LIMITS['6W']['max_c']) * 100
                        elif suggested == 'JB' or suggested == '4W':
                            suggested = '6W'
                            source = source + " ‚Üí 6W"
                            w_util = (total_w / LIMITS['6W']['max_w']) * 100
                            c_util = (total_c / LIMITS['6W']['max_c']) * 100
            else:
                w_util = c_util = 0
            
            # ‚ö° Skip distance calculation completely for speed optimization
            trip_codes = trip_data['Code'].unique()
            total_distance = 0  # Skip all distance calculations
            
            summary_data.append({
                'Trip': int(trip_num),
                'Branches': len(trip_data['Code'].unique()),
                'Weight': total_w,
                'Cube': total_c,
                'Truck': f"{suggested} {source}",
                'Weight_Use%': w_util,
                'Cube_Use%': c_util,
                'Total_Distance': total_distance
            })
        

        summary_df = pd.DataFrame(summary_data)

        # üö® Double Check: No trip uses a vehicle larger than allowed by any branch
        for idx, row in summary_df.iterrows():
            trip_num = row['Trip']
            trip_codes = test_df_result[test_df_result['Trip'] == trip_num]['Code'].unique()
            max_allowed = get_max_vehicle_for_trip(trip_codes)
            vehicle_type = row['Truck'].split()[0]
            vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
            if vehicle_sizes.get(vehicle_type, 3) > vehicle_sizes.get(max_allowed, 3):
                # Override to strictest allowed
                summary_df.at[idx, 'Truck'] = f"{max_allowed} üîí ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤"

        # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏£‡∏ñ
        trip_truck_display = {}
        for _, row in summary_df.iterrows():
            trip_truck_display[row['Trip']] = row['Truck']

        test_df_result['Truck'] = test_df_result['Trip'].map(trip_truck_display)
        # üîí Final enforcement: Never allow 6W if any branch restricts to 4W/JB
        test_df_result = enforce_vehicle_constraints(test_df_result)
        
        # Mark VehicleCheck if strict constraint enforced
        def vehicle_check_str(row):
            truck = row['Truck']
            if 'üîí' in truck or '‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤' in truck:
                return 'üîí ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤'
            return '‚úÖ ‡πÉ‡∏ä‡πâ‡∏ï‡∏≤‡∏°‡πÑ‡∏ü‡∏•‡πå'
        test_df_result['VehicleCheck'] = test_df_result.apply(vehicle_check_str, axis=1)

        return test_df_result, summary_df
    
    # üîí Final enforcement of vehicle constraints
    def enforce_vehicle_constraints(test_df):
        """‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡∏Ç‡∏±‡πâ‡∏ô‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢ - ‡πÑ‡∏°‡πà‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï 6W ‡∏´‡∏≤‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏à‡∏≥‡∏Å‡∏±‡∏î 4W/JB ‡∏´‡∏£‡∏∑‡∏≠‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏ô‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•"""
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for trip_num in test_df['Trip'].unique():
            trip_data = test_df[test_df['Trip'] == trip_num]
            trip_codes = trip_data['Code'].unique()
            
            # üîí ‡πÄ‡∏ä‡πá‡∏Ñ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î - ‡∏´‡πâ‡∏≤‡∏° 6W ‡πÉ‡∏ô‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•!
            provinces = set()
            for code in trip_codes:
                prov = get_province(code) if 'get_province' in dir() else None
                if not prov and not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                    master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                    if len(master_row) > 0:
                        prov = master_row.iloc[0].get('‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', '')
                if prov and prov != 'UNKNOWN':
                    provinces.add(prov)
            
            all_nearby = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
            
            # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏ó‡∏µ‡πà‡πÄ‡∏Ç‡πâ‡∏°‡∏á‡∏ß‡∏î‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
            max_vehicles = []
            for code in trip_codes:
                max_vehicle = get_max_vehicle_for_branch(code)
                max_vehicles.append(max_vehicle)
            
            min_max_size = min(vehicle_sizes.get(v, 3) for v in max_vehicles)
            
            # üîí ‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏• = ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö JB ‡∏´‡∏£‡∏∑‡∏≠‡πÄ‡∏•‡πá‡∏Å‡∏Å‡∏ß‡πà‡∏≤ (‡∏´‡πâ‡∏≤‡∏° 6W)
            if all_nearby and min_max_size == 3:
                min_max_size = 2  # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏•‡∏á‡∏°‡∏≤‡πÄ‡∏õ‡πá‡∏ô JB
            
            # ‡∏´‡∏≤‡∏Å‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏î‡∏à‡∏≥‡∏Å‡∏±‡∏î 4W/JB ‡∏´‡∏£‡∏∑‡∏≠‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏ô‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏• ‚Üí ‡∏´‡πâ‡∏≤‡∏° 6W
            if min_max_size < 3:
                # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡πÄ‡∏õ‡πá‡∏ô JB ‡∏´‡∏£‡∏∑‡∏≠ 4W
                allowed_vehicle = 'JB' if min_max_size >= 2 else '4W'
                current_truck = test_df.loc[test_df['Trip'] == trip_num, 'Truck'].iloc[0] if len(test_df[test_df['Trip'] == trip_num]) > 0 else ''
                if '6W' in str(current_truck):
                    test_df.loc[test_df['Trip'] == trip_num, 'Truck'] = f'{allowed_vehicle} üîí {"‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•" if all_nearby else "‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤"}'
        
        return test_df
    
    # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå Trip ‡πÉ‡∏´‡πâ‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
    
    # üó∫Ô∏è ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡∏≤‡∏°‡∏û‡∏¥‡∏Å‡∏±‡∏î‡∏Å‡πà‡∏≠‡∏ô (Spatial Clustering) + ‡∏à‡∏±‡∏ö‡∏Ñ‡∏π‡πà‡∏™‡∏≤‡∏Ç‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô
    def create_distance_based_clusters(codes, max_distance_km=25):
        """‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ô (‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô max_distance_km) + ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏£‡∏ß‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô"""
        # ‚ö° Speed: Skip clustering if too few codes
        if len(codes) < 10:
            return [codes]  # Return all as one cluster
        
        # üî• Phase 0: ‡∏à‡∏±‡∏ö‡∏Ñ‡∏π‡πà‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô (‡πÄ‡∏ä‡πà‡∏ô ‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á 3,4,8,10) ‡πÉ‡∏´‡πâ‡∏≠‡∏¢‡∏π‡πà‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡πÄ‡∏™‡∏°‡∏≠
        similar_groups = []  # ‡πÄ‡∏Å‡πá‡∏ö‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô
        grouped_codes = set()  # ‡πÄ‡∏Å‡πá‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏ñ‡∏π‡∏Å‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÅ‡∏•‡πâ‡∏ß
        
        # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ó‡∏∏‡∏Å‡∏Ñ‡∏π‡πà‡∏™‡∏≤‡∏Ç‡∏≤
        for i, code1 in enumerate(codes):
            if code1 in grouped_codes:
                continue
            
            # ‡∏´‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤
            name1 = test_df[test_df['Code'] == code1]['Name'].iloc[0] if 'Name' in test_df.columns and len(test_df[test_df['Code'] == code1]) > 0 else ''
            
            # ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô
            similar_group = [code1]
            for j, code2 in enumerate(codes):
                if i >= j or code2 in grouped_codes:
                    continue
                
                name2 = test_df[test_df['Code'] == code2]['Name'].iloc[0] if 'Name' in test_df.columns and len(test_df[test_df['Code'] == code2]) > 0 else ''
                
                # ‡∏ñ‡πâ‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô (‡πÄ‡∏ä‡πà‡∏ô "‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á" ‡πÉ‡∏ô "‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á 3", "‡∏Ñ‡∏•‡∏≠‡∏á‡∏´‡∏•‡∏ß‡∏á 4")
                if is_similar_name(name1, name2, similarity_threshold=75):  # ‡∏•‡∏î‡πÄ‡∏Å‡∏ì‡∏ë‡πå‡πÄ‡∏´‡∏•‡∏∑‡∏≠ 75% ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏à‡∏±‡∏ö‡πÑ‡∏î‡πâ‡∏°‡∏≤‡∏Å‡∏Ç‡∏∂‡πâ‡∏ô
                    # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á - ‡∏¢‡∏≠‡∏°‡πÉ‡∏´‡πâ‡πÑ‡∏Å‡∏•‡πÑ‡∏î‡πâ‡∏ñ‡∏∂‡∏á 80km (‡πÄ‡∏û‡∏£‡∏≤‡∏∞‡∏™‡∏≤‡∏Ç‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡∏≠‡∏≤‡∏à‡∏Å‡∏£‡∏∞‡∏à‡∏≤‡∏¢)
                    lat1, lon1 = get_lat_lon_from_master(code1)
                    lat2, lon2 = get_lat_lon_from_master(code2)
                    
                    if lat1 and lat2:
                        dist = haversine_distance(lat1, lon1, lat2, lon2)
                        if dist < 80:  # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å 50km ‚Üí 80km
                            similar_group.append(code2)
                            grouped_codes.add(code2)
                    else:
                        # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏û‡∏¥‡∏Å‡∏±‡∏î ‚Üí ‡∏£‡∏ß‡∏°‡πÄ‡∏Ç‡πâ‡∏≤‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÄ‡∏•‡∏¢ (‡∏ñ‡∏∑‡∏≠‡∏ß‡πà‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô)
                        similar_group.append(code2)
                        grouped_codes.add(code2)
            
            if len(similar_group) > 1:
                # ‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô ‚Üí ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏Å‡∏•‡∏∏‡πà‡∏°
                similar_groups.append(similar_group)
                grouped_codes.add(code1)
        
        # ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏•‡∏∑‡∏≠ (‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô)
        remaining_codes = [c for c in codes if c not in grouped_codes]
        
        clusters = []
        remaining = remaining_codes.copy()
        
        while remaining:
            # ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÉ‡∏´‡∏°‡πà
            seed = remaining.pop(0)
            cluster = [seed]
            seed_lat, seed_lon = get_lat_lon_from_master(seed)
            
            if seed_lat is None:
                # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏û‡∏¥‡∏Å‡∏±‡∏î ‚Üí ‡πÉ‡∏™‡πà‡∏Ñ‡∏•‡∏±‡∏™‡πÄ‡∏ï‡∏≠‡∏£‡πå‡πÄ‡∏î‡∏µ‡πà‡∏¢‡∏ß
                clusters.append(cluster)
                continue
            
            # ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ö seed
            to_remove = []
            for code in remaining[:]:
                lat, lon = get_lat_lon_from_master(code)
                if lat and lon:
                    dist = haversine_distance(seed_lat, seed_lon, lat, lon)
                    if dist <= max_distance_km:
                        cluster.append(code)
                        to_remove.append(code)
            
            # ‡∏•‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏û‡∏¥‡πà‡∏°‡πÑ‡∏õ‡πÅ‡∏•‡πâ‡∏ß
            for code in to_remove:
                if code in remaining:
                    remaining.remove(code)
            
            clusters.append(cluster)
        
        # üî• ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô‡πÄ‡∏Ç‡πâ‡∏≤‡πÑ‡∏õ (‡∏à‡∏∞‡∏≠‡∏¢‡∏π‡πà‡∏Ç‡πâ‡∏≤‡∏á‡∏´‡∏ô‡πâ‡∏≤‡∏™‡∏∏‡∏î - ‡∏™‡πà‡∏á‡∏Å‡πà‡∏≠‡∏ô)
        all_clusters = similar_groups + clusters
        
        return all_clusters
    
    def get_lat_lon_from_master(code):
        """‡∏î‡∏∂‡∏á‡∏û‡∏¥‡∏Å‡∏±‡∏î‡∏à‡∏≤‡∏Å Master Data"""
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
            if len(master_row) > 0:
                lat = master_row.iloc[0].get('‡∏•‡∏∞‡∏ï‡∏¥‡∏à‡∏π‡∏î', None)
                lon = master_row.iloc[0].get('‡∏•‡∏≠‡∏á‡∏ï‡∏¥‡∏à‡∏π‡∏î', None)
                if pd.notna(lat) and pd.notna(lon) and lat != 0 and lon != 0:
                    try:
                        return float(lat), float(lon)
                    except:
                        pass
        return None, None
    
    def build_route_nearest_neighbor(codes):
        """‡∏™‡∏£‡πâ‡∏≤‡∏á‡πÄ‡∏™‡πâ‡∏ô‡∏ó‡∏≤‡∏á‡πÇ‡∏î‡∏¢‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏ñ‡∏±‡∏î‡πÑ‡∏õ (Nearest Neighbor)"""
        if len(codes) <= 1:
            return codes
        
        # ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å DC
        route = []
        remaining = codes.copy()
        current_lat, current_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
        
        while remaining:
            # ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏à‡∏≤‡∏Å‡∏ï‡∏≥‡πÅ‡∏´‡∏ô‡πà‡∏á‡∏õ‡∏±‡∏à‡∏à‡∏∏‡∏ö‡∏±‡∏ô
            min_dist = float('inf')
            nearest_code = None
            
            for code in remaining:
                lat, lon = get_lat_lon_from_master(code)
                if lat and lon:
                    dist = haversine_distance(current_lat, current_lon, lat, lon)
                    if dist < min_dist:
                        min_dist = dist
                        nearest_code = code
            
            if nearest_code:
                route.append(nearest_code)
                remaining.remove(nearest_code)
                current_lat, current_lon = get_lat_lon_from_master(nearest_code)
                if current_lat is None:
                    current_lat, current_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
            else:
                # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏û‡∏¥‡∏Å‡∏±‡∏î ‚Üí ‡πÉ‡∏™‡πà‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏•‡∏∑‡∏≠‡∏ï‡∏≤‡∏°‡∏•‡∏≥‡∏î‡∏±‡∏ö
                route.extend(remaining)
                break
        
        return route
    
    all_codes = test_df['Code'].unique().tolist()
    assigned_trips = {}
    trip_counter = 1
    trip_recommended_vehicles = {}  # ‡πÄ‡∏Å‡πá‡∏ö‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÅ‡∏ô‡∏∞‡∏ô‡∏≥‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏ó‡∏£‡∏¥‡∏õ
    
    total_codes = len(all_codes)
    processed = 0
    
    # ‚è±Ô∏è Timer ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö early stopping
    import time
    start_time = time.time()
    MAX_PROCESSING_TIME = 20  # ‡∏ß‡∏¥‡∏ô‡∏≤‡∏ó‡∏µ - ‡∏•‡∏î‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÉ‡∏´‡πâ‡πÄ‡∏£‡πá‡∏ß‡∏Ç‡∏∂‡πâ‡∏ô (target: 30 ‡∏ß‡∏¥‡∏ô‡∏≤‡∏ó‡∏µ ‡∏£‡∏ß‡∏°)
    
    # üöÄ Cache ‡∏û‡∏¥‡∏Å‡∏±‡∏î‡πÅ‡∏•‡∏∞‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏•‡πà‡∏ß‡∏á‡∏´‡∏ô‡πâ‡∏≤ (‡∏õ‡∏£‡∏∞‡∏´‡∏¢‡∏±‡∏î‡πÄ‡∏ß‡∏•‡∏≤ 70%)
    coord_cache = {}
    province_cache = {}
    
    # üÜï ‡∏£‡∏≤‡∏¢‡∏ä‡∏∑‡πà‡∏≠ 77 ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏Ç‡∏≠‡∏á‡πÑ‡∏ó‡∏¢‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏î‡∏∂‡∏á‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤
    THAI_PROVINCES = [
        '‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û', '‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏°‡∏´‡∏≤‡∏ô‡∏Ñ‡∏£', '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏õ‡∏£‡∏≤‡∏Å‡∏≤‡∏£', '‡∏ô‡∏ô‡∏ó‡∏ö‡∏∏‡∏£‡∏µ', '‡∏õ‡∏ó‡∏∏‡∏°‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏ô‡∏Ñ‡∏£‡∏õ‡∏ê‡∏°', '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏™‡∏≤‡∏Ñ‡∏£',
        '‡∏û‡∏£‡∏∞‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤', '‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤', '‡∏≠‡πà‡∏≤‡∏á‡∏ó‡∏≠‡∏á', '‡∏•‡∏û‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏¥‡∏á‡∏´‡πå‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ä‡∏±‡∏¢‡∏ô‡∏≤‡∏ó', '‡∏™‡∏£‡∏∞‡∏ö‡∏∏‡∏£‡∏µ',
        '‡∏ä‡∏•‡∏ö‡∏∏‡∏£‡∏µ', '‡∏£‡∏∞‡∏¢‡∏≠‡∏á', '‡∏à‡∏±‡∏ô‡∏ó‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ï‡∏£‡∏≤‡∏î', '‡∏â‡∏∞‡πÄ‡∏ä‡∏¥‡∏á‡πÄ‡∏ó‡∏£‡∏≤', '‡∏õ‡∏£‡∏≤‡∏à‡∏µ‡∏ô‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ô‡∏Ñ‡∏£‡∏ô‡∏≤‡∏¢‡∏Å', '‡∏™‡∏£‡∏∞‡πÅ‡∏Å‡πâ‡∏ß',
        '‡∏ô‡∏Ñ‡∏£‡∏£‡∏≤‡∏ä‡∏™‡∏µ‡∏°‡∏≤', '‡πÇ‡∏Ñ‡∏£‡∏≤‡∏ä', '‡∏ö‡∏∏‡∏£‡∏µ‡∏£‡∏±‡∏°‡∏¢‡πå', '‡∏™‡∏∏‡∏£‡∏¥‡∏ô‡∏ó‡∏£‡πå', '‡∏®‡∏£‡∏µ‡∏™‡∏∞‡πÄ‡∏Å‡∏©', '‡∏≠‡∏∏‡∏ö‡∏•‡∏£‡∏≤‡∏ä‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏¢‡πÇ‡∏™‡∏ò‡∏£', '‡∏ä‡∏±‡∏¢‡∏†‡∏π‡∏°‡∏¥',
        '‡∏≠‡∏≥‡∏ô‡∏≤‡∏à‡πÄ‡∏à‡∏£‡∏¥‡∏ç', '‡∏´‡∏ô‡∏≠‡∏á‡∏ö‡∏±‡∏ß‡∏•‡∏≥‡∏†‡∏π', '‡∏Ç‡∏≠‡∏ô‡πÅ‡∏Å‡πà‡∏ô', '‡∏≠‡∏∏‡∏î‡∏£‡∏ò‡∏≤‡∏ô‡∏µ', '‡πÄ‡∏•‡∏¢', '‡∏´‡∏ô‡∏≠‡∏á‡∏Ñ‡∏≤‡∏¢', '‡∏°‡∏´‡∏≤‡∏™‡∏≤‡∏£‡∏Ñ‡∏≤‡∏°',
        '‡∏£‡πâ‡∏≠‡∏¢‡πÄ‡∏≠‡πá‡∏î', '‡∏Å‡∏≤‡∏¨‡∏™‡∏¥‡∏ô‡∏ò‡∏∏‡πå', '‡∏™‡∏Å‡∏•‡∏ô‡∏Ñ‡∏£', '‡∏ô‡∏Ñ‡∏£‡∏û‡∏ô‡∏°', '‡∏°‡∏∏‡∏Å‡∏î‡∏≤‡∏´‡∏≤‡∏£', '‡∏ö‡∏∂‡∏á‡∏Å‡∏≤‡∏¨',
        '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡πÉ‡∏´‡∏°‡πà', '‡∏•‡∏≥‡∏û‡∏π‡∏ô', '‡∏•‡∏≥‡∏õ‡∏≤‡∏á', '‡∏≠‡∏∏‡∏ï‡∏£‡∏î‡∏¥‡∏ï‡∏ñ‡πå', '‡πÅ‡∏û‡∏£‡πà', '‡∏ô‡πà‡∏≤‡∏ô', '‡∏û‡∏∞‡πÄ‡∏¢‡∏≤', '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡∏£‡∏≤‡∏¢', '‡πÅ‡∏°‡πà‡∏Æ‡πà‡∏≠‡∏á‡∏™‡∏≠‡∏ô',
        '‡∏ô‡∏Ñ‡∏£‡∏™‡∏ß‡∏£‡∏£‡∏Ñ‡πå', '‡∏≠‡∏∏‡∏ó‡∏±‡∏¢‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏Å‡∏≥‡πÅ‡∏û‡∏á‡πÄ‡∏û‡∏ä‡∏£', '‡∏ï‡∏≤‡∏Å', '‡∏™‡∏∏‡πÇ‡∏Ç‡∏ó‡∏±‡∏¢', '‡∏û‡∏¥‡∏©‡∏ì‡∏∏‡πÇ‡∏•‡∏Å', '‡∏û‡∏¥‡∏à‡∏¥‡∏ï‡∏£', '‡πÄ‡∏û‡∏ä‡∏£‡∏ö‡∏π‡∏£‡∏ì‡πå',
        '‡∏£‡∏≤‡∏ä‡∏ö‡∏∏‡∏£‡∏µ', '‡∏Å‡∏≤‡∏ç‡∏à‡∏ô‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏∏‡∏û‡∏£‡∏£‡∏ì‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ô‡∏Ñ‡∏£‡∏õ‡∏ê‡∏°', '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏™‡∏á‡∏Ñ‡∏£‡∏≤‡∏°', '‡πÄ‡∏û‡∏ä‡∏£‡∏ö‡∏∏‡∏£‡∏µ', '‡∏õ‡∏£‡∏∞‡∏à‡∏ß‡∏ö‡∏Ñ‡∏µ‡∏£‡∏µ‡∏Ç‡∏±‡∏ô‡∏ò‡πå',
        '‡∏ä‡∏∏‡∏°‡∏û‡∏£', '‡∏£‡∏∞‡∏ô‡∏≠‡∏á', '‡∏™‡∏∏‡∏£‡∏≤‡∏©‡∏é‡∏£‡πå‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏û‡∏±‡∏á‡∏á‡∏≤', '‡∏Å‡∏£‡∏∞‡∏ö‡∏µ‡πà', '‡∏†‡∏π‡πÄ‡∏Å‡πá‡∏ï', '‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏ò‡∏£‡∏£‡∏°‡∏£‡∏≤‡∏ä', '‡∏ï‡∏£‡∏±‡∏á',
        '‡∏û‡∏±‡∏ó‡∏•‡∏∏‡∏á', '‡∏™‡∏á‡∏Ç‡∏•‡∏≤', '‡∏™‡∏ï‡∏π‡∏•', '‡∏õ‡∏±‡∏ï‡∏ï‡∏≤‡∏ô‡∏µ', '‡∏¢‡∏∞‡∏•‡∏≤', '‡∏ô‡∏£‡∏≤‡∏ò‡∏¥‡∏ß‡∏≤‡∏™'
    ]
    
    def extract_province_from_name(branch_name):
        """‡∏î‡∏∂‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤ ‡πÄ‡∏ä‡πà‡∏ô '‡∏û‡∏¥‡∏©‡∏ì‡∏∏‡πÇ‡∏•‡∏Å5' -> '‡∏û‡∏¥‡∏©‡∏ì‡∏∏‡πÇ‡∏•‡∏Å'"""
        if not branch_name:
            return None
        name = str(branch_name).strip()
        for province in THAI_PROVINCES:
            if province in name:
                return province
        return None
    
    for code in all_codes:
        lat, lon = get_lat_lon_from_master(code)
        coord_cache[code] = (lat, lon)
        
        # Cache ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î - ‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏Ñ‡∏ß‡∏≤‡∏°‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç:
        # 1. Master Data
        # 2. ‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤ (‡∏î‡∏∂‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠) üÜï
        # 3. Province column
        # 4. branch_info
        
        found_province = False
        
        # 1. ‡∏•‡∏≠‡∏á Master Data ‡∏Å‡πà‡∏≠‡∏ô
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
            if len(master_row) > 0:
                prov = master_row.iloc[0].get('‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', '')
                if prov and str(prov).strip() and prov != 'UNKNOWN':
                    province_cache[code] = str(prov).strip()
                    found_province = True
        
        # 2. üÜï ‡∏î‡∏∂‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤ (‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏°‡∏µ‡πÉ‡∏ô Master)
        if not found_province and 'Name' in test_df.columns:
            code_data = test_df[test_df['Code'] == code]
            if len(code_data) > 0:
                branch_name = code_data['Name'].iloc[0]
                prov_from_name = extract_province_from_name(branch_name)
                if prov_from_name:
                    province_cache[code] = prov_from_name
                    found_province = True
        
        # 3. Province column
        if not found_province and 'Province' in test_df.columns:
            prov = test_df[test_df['Code'] == code]['Province'].iloc[0] if len(test_df[test_df['Code'] == code]) > 0 else None
            if prov and prov != 'UNKNOWN' and str(prov).strip():
                province_cache[code] = prov
                found_province = True
        
        # 4. branch_info
        if not found_province and code in branch_info:
            prov = branch_info[code].get('province', 'UNKNOWN')
            if prov and prov != 'UNKNOWN' and str(prov).strip():
                province_cache[code] = prov
                found_province = True
        
        # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡πÄ‡∏à‡∏≠‡πÄ‡∏•‡∏¢
        if not found_province:
            province_cache[code] = 'UNKNOWN'
    
    # üéØ ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ï‡∏≤‡∏°‡∏û‡∏¥‡∏Å‡∏±‡∏î‡∏Å‡πà‡∏≠‡∏ô (‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏£‡∏±‡∏®‡∏°‡∏µ‡∏™‡∏π‡∏á - ‡∏Ç‡∏≠‡∏ö‡πÄ‡∏Ç‡∏ï‡πÉ‡∏´‡∏ç‡πà‡∏Ç‡∏∂‡πâ‡∏ô)
    spatial_clusters = create_distance_based_clusters(all_codes, max_distance_km=25)
    
    # üîí ‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏´‡∏°‡πà‡πÑ‡∏õ‡∏¢‡∏±‡∏á‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ (FAST VERSION)
    def check_distance_to_all_trip_branches(new_code, trip_codes, max_dist=40):
        """
        ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏´‡∏°‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ö‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà (‡πÉ‡∏ä‡πâ sampling ‡∏ñ‡πâ‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏ç‡πà)
        ‡∏Ñ‡∏∑‡∏ô‡∏Ñ‡πà‡∏≤: (avg_distance, max_distance, all_within_limit)
        """
        if not trip_codes:
            return 0, 0, True
        
        new_lat, new_lon = coord_cache.get(new_code, (None, None))
        if not new_lat:
            return 9999, 9999, False
        
        # ‚ö° Speed: ‡∏ñ‡πâ‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏°‡∏µ‡∏´‡∏•‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤ ‡πÉ‡∏´‡πâ sample ‡πÅ‡∏Ñ‡πà 3 ‡∏™‡∏≤‡∏Ç‡∏≤ (‡∏•‡∏î‡∏à‡∏≤‡∏Å 5)
        sample_codes = trip_codes if len(trip_codes) <= 3 else trip_codes[:2] + trip_codes[-1:]
        distances = []
        for code in sample_codes:
            code_lat, code_lon = coord_cache.get(code, (None, None))
            if code_lat:
                dist = haversine_distance(new_lat, new_lon, code_lat, code_lon)
                distances.append(dist)
        
        if not distances:
            return 9999, 9999, False
        
        avg_dist = sum(distances) / len(distances)
        max_dist_found = max(distances)
        all_within = max_dist_found <= max_dist
        
        return avg_dist, max_dist_found, all_within
    
    # ‚ö° Speed: Pre-compute trip centroids for fast lookup
    trip_centroids = {}  # {trip_num: (lat, lon)}
    
    def update_trip_centroid(trip_num, codes):
        """‡∏≠‡∏±‡∏û‡πÄ‡∏î‡∏ï centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ"""
        if not codes:
            trip_centroids[trip_num] = (None, None)
            return
        lats, lons = [], []
        for code in codes:
            lat, lon = coord_cache.get(code, (None, None))
            if lat:
                lats.append(lat)
                lons.append(lon)
        if lats:
            trip_centroids[trip_num] = (sum(lats)/len(lats), sum(lons)/len(lons))
        else:
            trip_centroids[trip_num] = (None, None)
    
    def find_closest_trip_for_branch(branch_code, all_trip_codes_dict, exclude_trip=None):
        """
        ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤ - ‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ (‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πâ centroid)
        ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏õ‡πâ‡∏≠‡∏á‡∏Å‡∏±‡∏ô‡∏Å‡∏≤‡∏£‡∏Å‡∏£‡∏∞‡πÇ‡∏î‡∏î‡∏Ç‡πâ‡∏≤‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏ß‡πà‡∏≤
        """
        branch_lat, branch_lon = coord_cache.get(branch_code, (None, None))
        if not branch_lat:
            return None, 9999
        
        best_trip = None
        best_avg_dist = 9999
        best_max_dist = 9999
        
        for trip_num, codes in all_trip_codes_dict.items():
            if exclude_trip and trip_num == exclude_trip:
                continue
            if not codes:
                continue
            
            # üîí ‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ (‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πâ centroid)
            distances = []
            for code in codes:
                code_lat, code_lon = coord_cache.get(code, (None, None))
                if code_lat:
                    dist = haversine_distance(branch_lat, branch_lon, code_lat, code_lon)
                    distances.append(dist)
            
            if not distances:
                continue
            
            avg_dist = sum(distances) / len(distances)
            max_dist = max(distances)
            
            # ‚ö° ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏£‡∏∞‡∏¢‡∏∞‡πÄ‡∏â‡∏•‡∏µ‡πà‡∏¢‡∏ï‡πà‡∏≥‡∏™‡∏∏‡∏î ‡πÅ‡∏•‡∏∞‡∏£‡∏∞‡∏¢‡∏∞‡πÑ‡∏Å‡∏•‡∏™‡∏∏‡∏î‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 40km
            if avg_dist < best_avg_dist and max_dist <= 40:
                best_avg_dist = avg_dist
                best_max_dist = max_dist
                best_trip = trip_num
        
        return best_trip, best_avg_dist
    
    # üîÑ ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡∏à‡∏≤‡∏Å‡πÉ‡∏Å‡∏•‡πâ ‚Üí ‡πÑ‡∏Å‡∏• ‡∏à‡∏≤‡∏Å DC
    def sort_by_distance_from_dc(codes):
        """‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡∏à‡∏≤‡∏Å‡πÉ‡∏Å‡∏•‡πâ DC ‡πÑ‡∏õ‡πÑ‡∏Å‡∏• DC"""
        def get_distance_from_dc(code):
            lat, lon = coord_cache.get(code, (None, None))
            if lat and lon:
                return calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
            return 9999  # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏û‡∏¥‡∏Å‡∏±‡∏î ‡πÉ‡∏´‡πâ‡πÑ‡∏ß‡πâ‡∏ó‡πâ‡∏≤‡∏¢
        return sorted(codes, key=get_distance_from_dc)
    
    # üÜï Cache ‡∏ä‡∏∑‡πà‡∏≠‡πÅ‡∏•‡∏∞‡∏ï‡∏≥‡∏ö‡∏•/‡∏≠‡∏≥‡πÄ‡∏†‡∏≠ ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°
    name_cache = {}
    subdistrict_cache = {}
    district_cache = {}
    
    for code in test_df['Code'].unique():
        # Cache ‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤
        if 'Name' in test_df.columns:
            code_data = test_df[test_df['Code'] == code]
            if len(code_data) > 0:
                name_cache[code] = str(code_data['Name'].iloc[0]).strip()
        
        # Cache ‡∏ï‡∏≥‡∏ö‡∏•/‡∏≠‡∏≥‡πÄ‡∏†‡∏≠ ‡∏à‡∏≤‡∏Å Master
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
            if len(master_row) > 0:
                m = master_row.iloc[0]
                if '‡∏ï‡∏≥‡∏ö‡∏•' in m.index and pd.notna(m['‡∏ï‡∏≥‡∏ö‡∏•']):
                    subdistrict_cache[code] = str(m['‡∏ï‡∏≥‡∏ö‡∏•']).strip()
                if '‡∏≠‡∏≥‡πÄ‡∏†‡∏≠' in m.index and pd.notna(m['‡∏≠‡∏≥‡πÄ‡∏†‡∏≠']):
                    district_cache[code] = str(m['‡∏≠‡∏≥‡πÄ‡∏†‡∏≠']).strip()
    
    # üÜï ‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô‡∏´‡∏≤ base name (‡πÄ‡∏ä‡πà‡∏ô "‡πÇ‡∏•‡∏ï‡∏±‡∏™ ‡∏û‡∏£‡∏∞‡∏£‡∏≤‡∏° 2" -> "‡πÇ‡∏•‡∏ï‡∏±‡∏™ ‡∏û‡∏£‡∏∞‡∏£‡∏≤‡∏°")
    def get_base_name(name):
        import re
        if not name:
            return ""
        
        # ‡πÅ‡∏õ‡∏•‡∏á‡πÄ‡∏õ‡πá‡∏ô lowercase ‡πÅ‡∏•‡∏∞‡∏ï‡∏±‡∏î whitespace
        name_lower = str(name).strip().lower()
        
        # üÜï Normalize ‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏Ñ‡∏•‡πâ‡∏≤‡∏¢‡∏Å‡∏±‡∏ô
        # Future/‡∏ü‡∏¥‡∏ß‡πÄ‡∏à‡∏≠‡∏£‡πå
        if 'future' in name_lower or '‡∏ü‡∏¥‡∏ß‡πÄ‡∏à‡∏≠‡∏£‡πå' in name_lower or '‡∏ü‡∏¥‡∏ß‡πÄ‡∏à‡∏≠' in name_lower:
            if 'rangsit' in name_lower or '‡∏£‡∏±‡∏á‡∏™‡∏¥‡∏ï' in name_lower:
                return "‡∏ü‡∏¥‡∏ß‡πÄ‡∏à‡∏≠‡∏£‡πå‡∏£‡∏±‡∏á‡∏™‡∏¥‡∏ï"  # ‡∏£‡∏ß‡∏°‡πÄ‡∏õ‡πá‡∏ô‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß
        
        # Lotus/‡πÇ‡∏•‡∏ï‡∏±‡∏™
        if 'lotus' in name_lower or '‡πÇ‡∏•‡∏ï‡∏±‡∏™' in name_lower:
            # ‡∏ï‡∏±‡∏î‡πÄ‡∏•‡∏Ç‡∏ó‡πâ‡∏≤‡∏¢‡∏≠‡∏≠‡∏Å ‡πÄ‡∏ä‡πà‡∏ô ‡πÇ‡∏•‡∏ï‡∏±‡∏™ 1 -> ‡πÇ‡∏•‡∏ï‡∏±‡∏™
            base = re.sub(r'\s*\d+\s*$', '', name_lower)
            return base.strip()
        
        # Big C/‡∏ö‡∏¥‡πä‡∏Å‡∏ã‡∏µ
        if 'big c' in name_lower or 'bigc' in name_lower or '‡∏ö‡∏¥‡πä‡∏Å‡∏ã‡∏µ' in name_lower or '‡∏ö‡∏¥‡πä‡∏Å‡∏ã' in name_lower:
            base = re.sub(r'\s*\d+\s*$', '', name_lower)
            return base.strip()
        
        # Makro/‡πÅ‡∏°‡πá‡∏Ñ‡πÇ‡∏Ñ‡∏£
        if 'makro' in name_lower or '‡πÅ‡∏°‡πá‡∏Ñ‡πÇ‡∏Ñ‡∏£' in name_lower or '‡πÅ‡∏°‡∏Ñ‡πÇ‡∏Ñ‡∏£' in name_lower:
            base = re.sub(r'\s*\d+\s*$', '', name_lower)
            return base.strip()
        
        # ‡∏Ñ‡∏•‡∏≠‡∏á (‡∏Ñ‡∏•‡∏≠‡∏á 1, ‡∏Ñ‡∏•‡∏≠‡∏á 2, ‡∏Ñ‡∏•‡∏≠‡∏á 3, ...)
        if '‡∏Ñ‡∏•‡∏≠‡∏á' in name_lower:
            # ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏ï‡∏±‡∏ß‡πÄ‡∏•‡∏Ç ‡πÄ‡∏ä‡πà‡∏ô "‡∏Ñ‡∏•‡∏≠‡∏á 3" -> "‡∏Ñ‡∏•‡∏≠‡∏á"
            if re.search(r'‡∏Ñ‡∏•‡∏≠‡∏á\s*\d+', name_lower):
                return "‡∏Ñ‡∏•‡∏≠‡∏á"
        
        # ‡∏ï‡∏±‡∏î‡∏ï‡∏±‡∏ß‡πÄ‡∏•‡∏Ç‡∏ó‡πâ‡∏≤‡∏¢‡∏ä‡∏∑‡πà‡∏≠‡πÅ‡∏•‡∏∞ whitespace
        base = re.sub(r'\s*\d+\s*$', '', str(name).strip())
        # ‡∏ï‡∏±‡∏î "‡∏™‡∏≤‡∏Ç‡∏≤" ‡∏≠‡∏≠‡∏Å
        base = re.sub(r'^‡∏™‡∏≤‡∏Ç‡∏≤\s*', '', base)
        # ‡∏ï‡∏±‡∏î FC_, _FC ‡∏≠‡∏≠‡∏Å
        base = re.sub(r'_FC\d+$', '', base)
        base = re.sub(r'^FC\s*', '', base)
        return base.strip().lower()
    
    # üÜï ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡∏≤‡∏°‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô + ‡∏ï‡∏≥‡∏ö‡∏•‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô + ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏°‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á nearest neighbor
    def group_by_name_and_subdistrict(codes):
        """
        ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡∏≤‡∏°‡∏•‡∏≥‡∏î‡∏±‡∏ö ‡πÅ‡∏•‡πâ‡∏ß‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏°‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á nearest neighbor:
        1. ‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô‡∏Å‡∏±‡∏ô + ‡∏ï‡∏≥‡∏ö‡∏•‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô + ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô (‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î)
        2. ‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô‡∏Å‡∏±‡∏ô + ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
        3. ‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏´‡∏°‡∏∑‡∏≠‡∏ô‡∏Å‡∏±‡∏ô (‡∏ï‡πà‡∏≤‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î)
        4. ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î + ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
        5. ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
        6. ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏•‡∏∑‡∏≠
        
        üÜï ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ï‡∏≤‡∏°‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á: ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å‡πÉ‡∏Å‡∏•‡πâ DC ‚Üí nearest neighbor ‡πÑ‡∏õ‡πÄ‡∏£‡∏∑‡πà‡∏≠‡∏¢‡πÜ
        """
        # ‡∏™‡∏£‡πâ‡∏≤‡∏á key ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°
        groups = {}  # key: (priority, province, district, base_name, subdistrict) -> [codes]
        
        for code in codes:
            name = name_cache.get(code, '')
            base_name = get_base_name(name)
            subdistrict = subdistrict_cache.get(code, '')
            district = district_cache.get(code, '')
            province = province_cache.get(code, '')
            
            # ‡∏™‡∏£‡πâ‡∏≤‡∏á group key - ‡πÉ‡∏ä‡πâ‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏Ñ‡∏ß‡∏≤‡∏°‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç (‡πÄ‡∏•‡∏Ç‡∏ô‡πâ‡∏≠‡∏¢ = ‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏Å‡∏ß‡πà‡∏≤)
            if base_name and subdistrict and province:
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 1: ‡∏ä‡∏∑‡πà‡∏≠ + ‡∏ï‡∏≥‡∏ö‡∏• + ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î (‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î)
                key = (1, province, district, base_name, subdistrict)
            elif base_name and province:
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 2: ‡∏ä‡∏∑‡πà‡∏≠ + ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î (üî• ‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏°‡∏≤‡∏Å - ‡∏£‡∏ß‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô)
                key = (2, province, district, base_name, '')
            elif base_name:
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 3: ‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô (‡πÅ‡∏°‡πâ‡∏ï‡πà‡∏≤‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î - ‡πÄ‡∏ä‡πà‡∏ô ‡πÇ‡∏•‡∏ï‡∏±‡∏™ ‡∏Å‡∏ó‡∏° ‡∏Å‡∏±‡∏ö ‡πÇ‡∏•‡∏ï‡∏±‡∏™ ‡∏ä‡∏•‡∏ö‡∏∏‡∏£‡∏µ)
                key = (3, province, '', base_name, '')
            elif province and district:
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 4: ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î + ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠ (‡∏£‡∏ß‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô)
                key = (4, province, district, '', '')
            elif province:
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 5: ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
                key = (5, province, '', '', '')
            else:
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 6: ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏•‡∏∑‡∏≠
                key = (6, province if province else code, '', '', '', '')
            
            if key not in groups:
                groups[key] = []
            groups[key].append(code)
        
        # üÜï ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏î‡πâ‡∏ß‡∏¢ nearest neighbor approach
        # ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ DC ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î ‡πÅ‡∏•‡πâ‡∏ß‡∏´‡∏≤‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ñ‡∏±‡∏î‡πÑ‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
        result = []
        remaining_groups = list(groups.items())  # [(key, [codes]), ...]
        
        # ‡∏´‡∏≤‡∏ï‡∏≥‡πÅ‡∏´‡∏ô‡πà‡∏á‡πÄ‡∏â‡∏•‡∏µ‡πà‡∏¢‡∏Ç‡∏≠‡∏á‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏Å‡∏•‡∏∏‡πà‡∏°
        def get_group_center(group_codes):
            lats, lons = [], []
            for code in group_codes:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    lats.append(lat)
                    lons.append(lon)
            if lats and lons:
                return (sum(lats) / len(lats), sum(lons) / len(lons))
            return (None, None)
        
        # ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ DC ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
        if remaining_groups:
            # ‡∏´‡∏≤‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ DC ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
            def dist_from_dc(item):
                key, group_codes = item
                priority = key[0]
                center_lat, center_lon = get_group_center(group_codes)
                if center_lat and center_lon:
                    dist = calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, center_lat, center_lon)
                    return (priority, dist)  # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏° priority ‡∏Å‡πà‡∏≠‡∏ô ‡πÅ‡∏•‡πâ‡∏ß‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á
                return (priority, 9999)
            
            remaining_groups.sort(key=dist_from_dc)
            current_key, current_group = remaining_groups.pop(0)
            
            # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏™‡∏°‡∏≤‡∏ä‡∏¥‡∏Å‡πÉ‡∏ô‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÅ‡∏£‡∏Å‡∏ï‡∏≤‡∏° nearest neighbor
            group_sorted = sort_by_distance_from_dc(current_group)
            ordered = build_route_nearest_neighbor(group_sorted)
            result.extend(ordered)
            
            # ‡∏´‡∏≤‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ñ‡∏±‡∏î‡πÑ‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ö‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏õ‡∏±‡∏à‡∏à‡∏∏‡∏ö‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
            while remaining_groups:
                # ‡πÉ‡∏ä‡πâ‡∏ï‡∏≥‡πÅ‡∏´‡∏ô‡πà‡∏á‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢‡∏Ç‡∏≠‡∏á result ‡πÄ‡∏õ‡πá‡∏ô‡∏à‡∏∏‡∏î‡∏≠‡πâ‡∏≤‡∏á‡∏≠‡∏¥‡∏á
                last_code = result[-1] if result else None
                last_lat, last_lon = coord_cache.get(last_code, (None, None)) if last_code else (DC_WANG_NOI_LAT, DC_WANG_NOI_LON)
                
                if not last_lat:
                    last_lat, last_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
                
                # ‡∏´‡∏≤‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
                def dist_from_last(item):
                    key, group_codes = item
                    priority = key[0]
                    center_lat, center_lon = get_group_center(group_codes)
                    if center_lat and center_lon:
                        dist = calculate_distance(last_lat, last_lon, center_lat, center_lon)
                        return (priority, dist)  # priority ‡∏Å‡πà‡∏≠‡∏ô ‡πÅ‡∏•‡πâ‡∏ß‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á
                    return (priority, 9999)
                
                remaining_groups.sort(key=dist_from_last)
                next_key, next_group = remaining_groups.pop(0)
                
                # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏™‡∏°‡∏≤‡∏ä‡∏¥‡∏Å‡πÉ‡∏ô‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ï‡∏≤‡∏° nearest neighbor ‡∏à‡∏≤‡∏Å‡∏à‡∏∏‡∏î‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢
                ordered_group = build_route_nearest_neighbor_from_point(next_group, last_lat, last_lon)
                result.extend(ordered_group)
        
        return result
    
    # üÜï ‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô nearest neighbor ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å‡∏à‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏Å‡∏≥‡∏´‡∏ô‡∏î
    def build_route_nearest_neighbor_from_point(codes, start_lat, start_lon):
        if not codes:
            return []
        
        result = []
        remaining = codes[:]
        current_lat, current_lon = start_lat, start_lon
        
        while remaining:
            # ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏à‡∏≤‡∏Å‡∏à‡∏∏‡∏î‡∏õ‡∏±‡∏à‡∏à‡∏∏‡∏ö‡∏±‡∏ô
            nearest = None
            min_dist = float('inf')
            for code in remaining:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    dist = calculate_distance(current_lat, current_lon, lat, lon)
                    if dist < min_dist:
                        min_dist = dist
                        nearest = code
            
            if nearest:
                result.append(nearest)
                remaining.remove(nearest)
                current_lat, current_lon = coord_cache.get(nearest, (current_lat, current_lon))
            else:
                # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏û‡∏¥‡∏Å‡∏±‡∏î ‡πÉ‡∏´‡πâ‡πÄ‡∏û‡∏¥‡πà‡∏°‡πÑ‡∏õ‡πÄ‡∏•‡∏¢
                result.extend(remaining)
                break
        
        return result
    
    # üîí Define helper functions ‡∏Å‡πà‡∏≠‡∏ô loop
    def get_province(branch_code):
        return province_cache.get(branch_code, 'UNKNOWN')
    
    def get_distance_from_dc(code):
        """‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å DC"""
        lat, lon = coord_cache.get(code, (None, None))
        if lat and lon:
            return haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
        return 0  # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏û‡∏¥‡∏Å‡∏±‡∏î ‡∏ñ‡∏∑‡∏≠‡∏ß‡πà‡∏≤‡πÉ‡∏Å‡∏•‡πâ DC
    
    # üöÄ **NEW ALGORITHM: Farthest First + Nearest Neighbor**
    # 1. ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÑ‡∏Å‡∏•‡∏à‡∏≤‡∏Å DC ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
    # 2. ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏Å‡∏•‡πâ‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏ô‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô ‡πÅ‡∏•‡∏∞‡∏£‡∏∞‡∏¢‡∏∞‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô MAX_DISTANCE_IN_TRIP
    # 3. ‡πÄ‡∏ï‡∏¥‡∏°‡∏£‡∏ñ‡∏à‡∏ô‡πÄ‡∏ï‡πá‡∏° (Cube limit)
    # 4. ‡∏ï‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà‡πÄ‡∏°‡∏∑‡πà‡∏≠: ‡πÄ‡∏ï‡πá‡∏° / ‡∏£‡∏∞‡∏¢‡∏∞‡πÄ‡∏Å‡∏¥‡∏ô / ‡∏Ç‡πâ‡∏≤‡∏°‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î
    # 5. ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ‡∏à‡∏±‡∏î‡∏ó‡∏µ‡πà‡πÑ‡∏Å‡∏•‡∏™‡∏∏‡∏î‡πÄ‡∏õ‡πá‡∏ô seed ‡πÉ‡∏´‡∏°‡πà
    
    # ‡∏£‡∏ß‡∏° codes ‡∏à‡∏≤‡∏Å spatial_clusters
    all_codes_flat = []
    for cluster in spatial_clusters:
        all_codes_flat.extend(cluster)
    
    # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏°‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å DC (‡πÑ‡∏Å‡∏•‡∏™‡∏∏‡∏î‡∏≠‡∏¢‡∏π‡πà‡∏´‡∏ô‡πâ‡∏≤)
    all_codes = sorted(all_codes_flat, key=lambda c: get_distance_from_dc(c), reverse=True)
    
    def get_lat_lon(branch_code):
        return coord_cache.get(branch_code, (None, None))
    
    # **Main Loop: Farthest First + Nearest Neighbor**
    while all_codes:
        # ‚è±Ô∏è Early stopping - ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πâ‡πÄ‡∏ß‡∏•‡∏≤‡∏°‡∏≤‡∏Å‡∏Å‡∏ß‡πà‡∏≤ MAX_PROCESSING_TIME
        if time.time() - start_time > MAX_PROCESSING_TIME:
            # ‡∏à‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏•‡∏∑‡∏≠‡πÄ‡∏Ç‡πâ‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà‡∏Ñ‡∏ô‡∏•‡∏∞‡∏Ñ‡∏±‡∏ô
            for remaining_code in all_codes:
                assigned_trips[remaining_code] = trip_counter
                trip_counter += 1
            break
        
        # üéØ Pop ‡∏™‡∏≤‡∏Ç‡∏≤‡πÅ‡∏£‡∏Å (‡πÑ‡∏Å‡∏•‡∏™‡∏∏‡∏î‡∏à‡∏≤‡∏Å DC)
        seed_code = all_codes.pop(0)
        current_trip = [seed_code]
        assigned_trips[seed_code] = trip_counter
        
        seed_province = get_province(seed_code)
        seed_lat, seed_lon = coord_cache.get(seed_code, (None, None))
        
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì Weight/Cube ‡∏õ‡∏±‡∏à‡∏à‡∏∏‡∏ö‡∏±‡∏ô
        current_weight = test_df[test_df['Code'] == seed_code]['Weight'].sum()
        current_cube = test_df[test_df['Code'] == seed_code]['Cube'].sum()
        
        # ‡∏Å‡∏≥‡∏´‡∏ô‡∏î‡∏Ç‡∏µ‡∏î‡∏à‡∏≥‡∏Å‡∏±‡∏î (‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å 6W)
        max_cube = LIMITS['6W']['max_c'] * BUFFER  # 20 cube
        max_weight = LIMITS['6W']['max_w'] * BUFFER  # 6000 kg
        
        # üîÑ ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ñ‡∏±‡∏î‡πÑ‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏™‡∏∏‡∏î‡πÅ‡∏•‡∏∞‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏ô‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
        while all_codes:
            best_code = None
            best_dist = 9999
            
            # ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏à‡∏≤‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
            last_code = current_trip[-1]
            last_lat, last_lon = coord_cache.get(last_code, (None, None))
            
            for code in all_codes:
                code_province = get_province(code)
                code_lat, code_lon = coord_cache.get(code, (None, None))
                
                if not last_lat or not code_lat:
                    continue
                
                # ‡∏£‡∏∞‡∏¢‡∏∞‡∏à‡∏≤‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢
                dist_from_last = haversine_distance(last_lat, last_lon, code_lat, code_lon)
                
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏à‡∏≤‡∏Å‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
                max_dist_in_trip = 0
                for trip_code in current_trip:
                    trip_lat, trip_lon = coord_cache.get(trip_code, (None, None))
                    if trip_lat:
                        dist = haversine_distance(code_lat, code_lon, trip_lat, trip_lon)
                        max_dist_in_trip = max(max_dist_in_trip, dist)
                
                # ‡∏Å‡∏é: ‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏ï‡πâ‡∏≠‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô MAX_DISTANCE_IN_TRIP (50km)
                if max_dist_in_trip > MAX_DISTANCE_IN_TRIP:
                    continue
                
                # ‡πÑ‡∏°‡πà‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡∏≠‡∏µ‡∏Å‡∏ï‡πà‡∏≠‡πÑ‡∏õ - ‡πÉ‡∏ä‡πâ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡πÄ‡∏õ‡πá‡∏ô‡∏´‡∏•‡∏±‡∏Å
                
                # ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
                if dist_from_last < best_dist:
                    best_dist = dist_from_last
                    best_code = code
            
            if not best_code:
                break  # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏° ‡∏ï‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ capacity
            next_weight = test_df[test_df['Code'] == best_code]['Weight'].sum()
            next_cube = test_df[test_df['Code'] == best_code]['Cube'].sum()
            
            new_weight = current_weight + next_weight
            new_cube = current_cube + next_cube
            
            # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô capacity ‚Üí ‡∏ï‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
            if new_cube > max_cube or new_weight > max_weight:
                break
            
            # ‚úÖ ‡∏ú‡πà‡∏≤‡∏ô‡∏ó‡∏∏‡∏Å‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç ‚Üí ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ
            all_codes.remove(best_code)
            current_trip.append(best_code)
            assigned_trips[best_code] = trip_counter
            current_weight = new_weight
            current_cube = new_cube
        
        trip_counter += 1
    
    test_df['Trip'] = test_df['Code'].map(assigned_trips)
    
    # ===============================================
    # üÜï ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ó‡∏£‡∏¥‡∏õ ‡πÅ‡∏•‡∏∞‡∏à‡∏±‡∏î‡πÉ‡∏´‡πâ‡∏Ñ‡∏£‡∏ö‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤
    # ===============================================
    unassigned_codes = test_df[test_df['Trip'].isna()]['Code'].tolist()
    
    if unassigned_codes:
        # ‡∏à‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏Ç‡πâ‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î ‡∏´‡∏£‡∏∑‡∏≠‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
        for code in unassigned_codes:
            code_lat, code_lon = coord_cache.get(code, (None, None))
            code_province = province_cache.get(code, 'UNKNOWN')
            code_weight = test_df[test_df['Code'] == code]['Weight'].sum()
            code_cube = test_df[test_df['Code'] == code]['Cube'].sum()
            
            best_trip = None
            best_score = float('inf')
            
            # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡πÅ‡∏•‡∏∞‡∏£‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ‡πÑ‡∏î‡πâ
            for trip_num in test_df['Trip'].dropna().unique():
                trip_data = test_df[test_df['Trip'] == trip_num]
                trip_codes = trip_data['Code'].tolist()
                trip_weight = trip_data['Weight'].sum()
                trip_cube = trip_data['Cube'].sum()
                
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
                trip_provinces = set()
                for tc in trip_codes:
                    tp = province_cache.get(tc, 'UNKNOWN')
                    if tp != 'UNKNOWN':
                        trip_provinces.add(tp)
                
                # ‡∏ï‡πâ‡∏≠‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î
                if code_province != 'UNKNOWN' and trip_provinces and code_province not in trip_provinces:
                    continue
                
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡πÉ‡∏™‡πà‡∏£‡∏ñ‡πÑ‡∏î‡πâ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà (‡πÉ‡∏ä‡πâ 6W ‡πÄ‡∏õ‡πá‡∏ô limit)
                new_weight = trip_weight + code_weight
                new_cube = trip_cube + code_cube
                new_util = max((new_weight / LIMITS['6W']['max_w']) * 100,
                              (new_cube / LIMITS['6W']['max_c']) * 100)
                
                if new_util > 100:  # ‡πÄ‡∏Å‡∏¥‡∏ô 100% ‡πÑ‡∏°‡πà‡∏£‡∏±‡∏ö
                    continue
                
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤
                if len(trip_codes) >= MAX_BRANCHES_PER_TRIP:
                    continue
                
                # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡πÄ‡∏â‡∏•‡∏µ‡πà‡∏¢‡πÑ‡∏õ‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
                if code_lat:
                    distances = []
                    for tc in trip_codes:
                        tc_lat, tc_lon = coord_cache.get(tc, (None, None))
                        if tc_lat:
                            dist = haversine_distance(code_lat, code_lon, tc_lat, tc_lon)
                            distances.append(dist)
                    
                    if distances:
                        avg_dist = sum(distances) / len(distances)
                        # ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
                        if avg_dist < best_score:
                            best_score = avg_dist
                            best_trip = trip_num
            
            if best_trip is not None:
                # ‡∏à‡∏±‡∏î‡πÄ‡∏Ç‡πâ‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_trip
                assigned_trips[code] = best_trip
            else:
                # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏° ‚Üí ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
                test_df.loc[test_df['Code'] == code, 'Trip'] = trip_counter
                assigned_trips[code] = trip_counter
                trip_counter += 1
    
    # ===============================================
    # üîí Post-processing: ‡∏™‡∏•‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏´‡πâ‡∏≠‡∏¢‡∏π‡πà‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î (FAST)
    # ===============================================
    def optimize_branch_placement():
        """‡∏™‡∏•‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏£‡∏∞‡∏´‡∏ß‡πà‡∏≤‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡πâ‡∏≠‡∏¢‡∏π‡πà‡∏Å‡∏±‡∏ö‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ô‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î (‡πÄ‡∏ß‡∏≠‡∏£‡πå‡∏ä‡∏±‡∏ô‡πÄ‡∏£‡πá‡∏ß)"""
        # ‡∏™‡∏£‡πâ‡∏≤‡∏á dict ‡∏Ç‡∏≠‡∏á trip ‚Üí codes
        trip_codes_dict = {}
        for trip_num in test_df['Trip'].unique():
            codes = test_df[test_df['Trip'] == trip_num]['Code'].tolist()
            trip_codes_dict[trip_num] = codes
            update_trip_centroid(trip_num, codes)
        
        # ‚ö° Speed: ‡πÄ‡∏ä‡πá‡∏Ñ‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏≠‡∏¢‡∏π‡πà‡πÑ‡∏Å‡∏•‡∏à‡∏≤‡∏Å centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡∏ï‡∏±‡∏ß‡πÄ‡∏≠‡∏á
        outliers = []  # (code, trip_num, dist_from_centroid)
        
        for trip_num, codes in trip_codes_dict.items():
            if len(codes) <= 2:
                continue
            
            centroid = trip_centroids.get(trip_num)
            if not centroid or not centroid[0]:
                continue
            
            for code in codes:
                code_lat, code_lon = coord_cache.get(code, (None, None))
                if code_lat:
                    dist = haversine_distance(code_lat, code_lon, centroid[0], centroid[1])
                    # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏Å‡∏•‡∏à‡∏≤‡∏Å centroid ‡∏°‡∏≤‡∏Å‡∏Å‡∏ß‡πà‡∏≤ 20km ‚Üí ‡∏≠‡∏≤‡∏à‡πÄ‡∏õ‡πá‡∏ô outlier
                    if dist > 20:
                        outliers.append((code, trip_num, dist))
        
        # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á outlier ‡∏à‡∏≤‡∏Å‡πÑ‡∏Å‡∏•‡∏™‡∏∏‡∏î‡∏Å‡πà‡∏≠‡∏ô ‡πÅ‡∏•‡∏∞‡∏à‡∏≥‡∏Å‡∏±‡∏î‡πÅ‡∏Ñ‡πà 50 ‡∏ï‡∏±‡∏ß
        outliers.sort(key=lambda x: -x[2])
        outliers = outliers[:50]
        
        # ‡∏•‡∏≠‡∏á‡∏¢‡πâ‡∏≤‡∏¢ outliers ‡πÑ‡∏õ‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏ß‡πà‡∏≤
        for code, trip_num, dist_current in outliers:
            if code not in trip_codes_dict.get(trip_num, []):
                continue
            
            best_trip, best_dist = find_closest_trip_for_branch(code, trip_codes_dict, exclude_trip=trip_num)
            
            # ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏ß‡πà‡∏≤‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏°‡∏µ‡∏ô‡∏±‡∏¢‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç (> 15km ‡∏î‡∏µ‡∏Å‡∏ß‡πà‡∏≤)
            if best_trip and best_dist < dist_current - 15:
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ
                code_max_vehicle = get_max_vehicle_for_branch(code)
                target_trip_codes = trip_codes_dict.get(best_trip, [])
                target_max_vehicle = get_max_vehicle_for_trip(set(target_trip_codes + [code]))
                
                vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
                if vehicle_priority.get(code_max_vehicle, 3) >= vehicle_priority.get(target_max_vehicle, 3):
                    # ‡∏¢‡πâ‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤
                    trip_codes_dict[trip_num].remove(code)
                    trip_codes_dict[best_trip].append(code)
                    assigned_trips[code] = best_trip
                    # ‡∏≠‡∏±‡∏û‡πÄ‡∏î‡∏ï centroids
                    update_trip_centroid(trip_num, trip_codes_dict[trip_num])
                    update_trip_centroid(best_trip, trip_codes_dict[best_trip])
        
        # ‡∏≠‡∏±‡∏û‡πÄ‡∏î‡∏ï test_df
        test_df['Trip'] = test_df['Code'].map(assigned_trips)
    
    # ‡πÄ‡∏£‡∏µ‡∏¢‡∏Å‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô optimization
    optimize_branch_placement()
    
    # ===============================================
    # Post-processing: ‡∏£‡∏ß‡∏°‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏•‡πá‡∏Å‡πÅ‡∏•‡∏∞‡∏õ‡∏£‡∏±‡∏ö‡∏Ç‡∏ô‡∏≤‡∏î‡∏£‡∏ñ
    # ===============================================
    # ‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏õ‡∏£‡∏±‡∏ö‡∏õ‡∏£‡∏∏‡∏á‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ
    
    # ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î‡∏û‡∏£‡πâ‡∏≠‡∏°‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•
    all_trips = []
    for trip_num in test_df['Trip'].unique():
        trip_data = test_df[test_df['Trip'] == trip_num]
        branch_count = len(trip_data)
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        trip_codes = set(trip_data['Code'].values)
        
        # ‡∏´‡∏≤‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ
        provinces = set()
        for code in trip_codes:
            prov = get_province(code)
            if prov != 'UNKNOWN':
                provinces.add(prov)
        
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì % ‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏£‡∏ñ 4W
        w_util = (total_w / LIMITS['4W']['max_w']) * 100
        c_util = (total_c / LIMITS['4W']['max_c']) * 100
        max_util = max(w_util, c_util)
        
        all_trips.append({
            'trip': trip_num,
            'count': branch_count,
            'util': max_util,
            'weight': total_w,
            'cube': total_c,
            'codes': trip_codes,
            'provinces': provinces
        })
    
    # üéØ ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì centroid (‡∏à‡∏∏‡∏î‡∏Å‡∏∂‡πà‡∏á‡∏Å‡∏•‡∏≤‡∏á) ‡∏Ç‡∏≠‡∏á‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏ó‡∏£‡∏¥‡∏õ
    for trip in all_trips:
        lats, lons = [], []
        for code in trip['codes']:
            lat, lon = get_lat_lon(code)
            if lat and lon:
                lats.append(lat)
                lons.append(lon)
        
        if lats and lons:
            trip['centroid_lat'] = sum(lats) / len(lats)
            trip['centroid_lon'] = sum(lons) / len(lons)
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å DC
            trip['distance_from_dc'] = haversine_distance(
                DC_WANG_NOI_LAT, DC_WANG_NOI_LON,
                trip['centroid_lat'], trip['centroid_lon']
            )
        else:
            trip['centroid_lat'] = DC_WANG_NOI_LAT
            trip['centroid_lon'] = DC_WANG_NOI_LON
            trip['distance_from_dc'] = 0
    
    # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡∏ï‡∏≤‡∏°: 1) ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏´‡∏•‡∏±‡∏Å 2) ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å DC (‡πÉ‡∏Å‡∏•‡πâ‡πÑ‡∏õ‡πÑ‡∏Å‡∏•) 3) ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤ 4) utilization
    # ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏ô‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡∏Å‡πà‡∏≠‡∏ô ‡∏Ñ‡πà‡∏≠‡∏¢‡∏Ç‡πâ‡∏≤‡∏°‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î
    def get_primary_province(trip):
        """‡∏´‡∏≤‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏´‡∏•‡∏±‡∏Å‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ (‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡∏°‡∏≤‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î)"""
        if not trip['provinces']:
            return 'UNKNOWN'
        # ‡∏ô‡∏±‡∏ö‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î
        province_counts = {}
        for code in trip['codes']:
            prov = get_province(code)
            province_counts[prov] = province_counts.get(prov, 0) + 1
        # ‡∏Ñ‡∏∑‡∏ô‡∏Ñ‡πà‡∏≤‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡∏°‡∏≤‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
        return max(province_counts.items(), key=lambda x: x[1])[0] if province_counts else 'UNKNOWN'
    
    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏´‡∏•‡∏±‡∏Å‡πÉ‡∏´‡πâ‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏ó‡∏£‡∏¥‡∏õ
    for trip in all_trips:
        trip['primary_province'] = get_primary_province(trip)
    
    # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏°‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏´‡∏•‡∏±‡∏Å ‚Üí ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á ‚Üí ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤ ‚Üí utilization
    all_trips.sort(key=lambda x: (x['primary_province'], x['distance_from_dc'], x['count'], x['util']))
    
    # üéØ Phase 1: ‡∏£‡∏ß‡∏°‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏•‡πá‡∏Å (‚â§3 ‡∏™‡∏≤‡∏Ç‡∏≤) ‡∏Å‡∏±‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á (FAST VERSION)
    merged = True
    merge_count = 0
    iteration = 0
    max_iterations = 1  # ‚ö° ‡∏•‡∏î‡πÄ‡∏õ‡πá‡∏ô 1 ‡∏£‡∏≠‡∏ö‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏£‡πá‡∏ß
    
    while merged and len(all_trips) > 1 and iteration < max_iterations:
        merged = False
        iteration += 1
        
        # ‚ö° Speed: ‡∏™‡∏£‡πâ‡∏≤‡∏á index ‡∏ï‡∏≤‡∏°‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á‡πÑ‡∏î‡πâ‡πÄ‡∏£‡πá‡∏ß
        province_to_trips = {}
        for idx, trip in enumerate(all_trips):
            if trip is None:
                continue
            for prov in trip['provinces']:
                if prov not in province_to_trips:
                    province_to_trips[prov] = []
                province_to_trips[prov].append(idx)
        
        # ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏•‡πá‡∏Å‡∏Å‡πà‡∏≠‡∏ô (‡πÄ‡∏£‡πá‡∏ß‡∏Å‡∏ß‡πà‡∏≤)
        small_trips = [(idx, t) for idx, t in enumerate(all_trips) if t and t['count'] <= 3]
        small_trips.sort(key=lambda x: x[1]['count'])
        
        for i, trip1 in small_trips:
            if all_trips[i] is None:
                continue
            
            # ‚ö° ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏à‡∏∞‡∏£‡∏ß‡∏°‡πÑ‡∏î‡πâ‡∏à‡∏≤‡∏Å‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
            candidate_indices = set()
            for prov in trip1['provinces']:
                for idx in province_to_trips.get(prov, []):
                    if idx != i and all_trips[idx] is not None:
                        candidate_indices.add(idx)
            
            # ‚ö° ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡πÅ‡∏Ñ‡πà 10 candidates ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î (‡∏ï‡∏≤‡∏° centroid)
            if len(candidate_indices) > 10 and 'centroid_lat' in trip1:
                candidates_with_dist = []
                for idx in candidate_indices:
                    trip2 = all_trips[idx]
                    if 'centroid_lat' in trip2:
                        dist = haversine_distance(
                            trip1['centroid_lat'], trip1['centroid_lon'],
                            trip2['centroid_lat'], trip2['centroid_lon']
                        )
                        candidates_with_dist.append((idx, dist))
                candidates_with_dist.sort(key=lambda x: x[1])
                candidate_indices = {x[0] for x in candidates_with_dist[:10]}
            
            for j in candidate_indices:
                if all_trips[j] is None:
                    continue
                
                trip2 = all_trips[j]
                
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏£‡∏∞‡∏´‡∏ß‡πà‡∏≤‡∏á centroid
                if 'centroid_lat' in trip1 and 'centroid_lat' in trip2:
                    centroid_distance = haversine_distance(
                        trip1['centroid_lat'], trip1['centroid_lon'],
                        trip2['centroid_lat'], trip2['centroid_lon']
                    )
                    if centroid_distance > 80:  # ‡πÑ‡∏Å‡∏•‡πÄ‡∏Å‡∏¥‡∏ô 80km ‡πÑ‡∏°‡πà‡∏£‡∏ß‡∏°
                        continue
                
                # üö® ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡∏Å‡πà‡∏≠‡∏ô‡∏£‡∏ß‡∏°
                combined_codes = trip1['codes'] | trip2['codes']
                max_allowed_combined = get_max_vehicle_for_trip(combined_codes)
                
                # ‡∏•‡∏≠‡∏á‡∏£‡∏ß‡∏°‡∏Å‡∏±‡∏ô
                combined_w = trip1['weight'] + trip2['weight']
                combined_c = trip1['cube'] + trip2['cube']
                combined_count = trip1['count'] + trip2['count']
                
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏£‡∏ß‡∏°‡πÅ‡∏•‡πâ‡∏ß‡πÉ‡∏™‡πà‡∏£‡∏ñ‡πÑ‡∏î‡πâ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
                if combined_count > MAX_BRANCHES_PER_TRIP:
                    continue
                
                # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì % ‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏´‡∏•‡∏±‡∏á‡∏£‡∏ß‡∏°
                combined_6w_util = max(
                    (combined_w / LIMITS['6W']['max_w']) * 100,
                    (combined_c / LIMITS['6W']['max_c']) * 100
                )
                
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡πÉ‡∏™‡πà‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï‡πÑ‡∏î‡πâ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
                vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
                allowed_priority = vehicle_priority.get(max_allowed_combined, 3)
                
                can_fit = False
                if allowed_priority >= 3 and combined_6w_util <= 100:  # 6W
                    can_fit = True
                elif allowed_priority >= 2 and combined_c <= LIMITS['JB']['max_c'] * BUFFER:  # JB
                    can_fit = True
                elif allowed_priority >= 1 and combined_c <= LIMITS['4W']['max_c'] * BUFFER:  # 4W
                    can_fit = True
                
                if can_fit:
                    # ‡∏£‡∏ß‡∏°‡∏ó‡∏£‡∏¥‡∏õ
                    for code in trip2['codes']:
                        test_df.loc[test_df['Code'] == code, 'Trip'] = trip1['trip']
                    
                    # ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• trip1
                    trip1['weight'] = combined_w
                    trip1['cube'] = combined_c
                    trip1['count'] = combined_count
                    trip1['codes'] |= trip2['codes']
                    trip1['provinces'] |= trip2['provinces']
                    trip1['util'] = combined_6w_util
                    
                    # ‡∏≠‡∏±‡∏û‡πÄ‡∏î‡∏ï centroid
                    lats, lons = [], []
                    for code in trip1['codes']:
                        lat, lon = coord_cache.get(code, (None, None))
                        if lat:
                            lats.append(lat)
                            lons.append(lon)
                    if lats:
                        trip1['centroid_lat'] = sum(lats) / len(lats)
                        trip1['centroid_lon'] = sum(lons) / len(lons)
                    
                    # ‡∏•‡∏ö trip2 ‡∏≠‡∏≠‡∏Å
                    all_trips[j] = None
                    merged = True
                    merge_count += 1
                    break
            
            if merged:
                break
        
        # ‡∏•‡∏ö None ‡∏≠‡∏≠‡∏Å
        all_trips = [t for t in all_trips if t is not None]
    
    # üéØ Phase 1.25: ‡∏¢‡πâ‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å‡πÑ‡∏õ‡∏¢‡∏±‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á‡∏ó‡∏µ‡πà‡πÉ‡∏´‡∏ç‡πà‡∏Å‡∏ß‡πà‡∏≤
    # üÜï ‡∏û‡∏¥‡πÄ‡∏®‡∏©: ‡πÄ‡∏ä‡πá‡∏Ñ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î+‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡∏î‡πâ‡∏ß‡∏¢
    reassign_count = 0
    
    # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡πÄ‡∏û‡∏µ‡∏¢‡∏á 1 ‡∏™‡∏≤‡∏Ç‡∏≤ ‡πÅ‡∏•‡∏∞ utilization ‡∏ï‡πà‡∏≥ (<40%)
    single_branch_trips = []
    for trip_num in test_df['Trip'].unique():
        if trip_num == 0:
            continue
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 1:
            branch_code = trip_data['Code'].values[0]
            branch_w = trip_data['Weight'].values[0]
            branch_c = trip_data['Cube'].values[0]
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì utilization (‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏•‡πá‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏û‡∏≠‡∏î‡∏µ)
            util_4w = max((branch_w / LIMITS['4W']['max_w']) * 100, 
                         (branch_c / LIMITS['4W']['max_c']) * 100)
            util_jb = max((branch_w / LIMITS['JB']['max_w']) * 100,
                         (branch_c / LIMITS['JB']['max_c']) * 100)
            
            # üÜï ‡∏î‡∏∂‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î+‡∏≠‡∏≥‡πÄ‡∏†‡∏≠
            branch_province = get_province(branch_code)
            branch_district = ''
            if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == branch_code]
                if len(master_row) > 0:
                    branch_district = str(master_row.iloc[0].get('‡∏≠‡∏≥‡πÄ‡∏†‡∏≠', '')).strip()
            
            # ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏ô‡πâ‡∏≠‡∏¢‡∏Å‡∏ß‡πà‡∏≤ 40% ‚Üí ‡∏û‡∏¥‡∏à‡∏≤‡∏£‡∏ì‡∏≤‡∏¢‡πâ‡∏≤‡∏¢
            if util_4w < 40 or util_jb < 40:
                lat, lon = get_lat_lon(branch_code)
                if lat and lon:
                    single_branch_trips.append({
                        'trip': trip_num,
                        'code': branch_code,
                        'weight': branch_w,
                        'cube': branch_c,
                        'lat': lat,
                        'lon': lon,
                        'util': util_4w,
                        'province': branch_province,
                        'district': branch_district
                    })
    
    # ‡∏û‡∏¢‡∏≤‡∏¢‡∏≤‡∏°‡∏¢‡πâ‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡πÑ‡∏õ‡∏£‡∏ß‡∏°‡∏Å‡∏±‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á
    for single_trip in single_branch_trips:
        branch_code = single_trip['code']
        branch_w = single_trip['weight']
        branch_c = single_trip['cube']
        branch_lat = single_trip['lat']
        branch_lon = single_trip['lon']
        
        # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î (‡πÑ‡∏°‡πà‡∏£‡∏ß‡∏°‡∏ó‡∏£‡∏¥‡∏õ‡∏Ç‡∏≠‡∏á‡∏ï‡∏±‡∏ß‡πÄ‡∏≠‡∏á)
        best_trip = None
        min_distance = float('inf')
        
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0 or trip_num == single_trip['trip']:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            
            # ‡∏ñ‡πâ‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏°‡∏µ‡∏°‡∏≤‡∏Å‡∏Å‡∏ß‡πà‡∏≤ 2 ‡∏™‡∏≤‡∏Ç‡∏≤ ‚Üí ‡∏û‡∏¥‡∏à‡∏≤‡∏£‡∏ì‡∏≤
            if len(trip_data) < 2:
                continue
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ
            trip_lats = []
            trip_lons = []
            for code in trip_data['Code'].values:
                lat, lon = get_lat_lon(code)
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if not trip_lats:
                continue
            
            centroid_lat = sum(trip_lats) / len(trip_lats)
            centroid_lon = sum(trip_lons) / len(trip_lons)
            
            distance = haversine_distance(branch_lat, branch_lon, centroid_lat, centroid_lon)
            
            # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏Å‡∏•‡πÄ‡∏Å‡∏¥‡∏ô 50km ‚Üí ‡∏Ç‡πâ‡∏≤‡∏° (‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å 30km)
            if distance > 50:
                continue
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡πÄ‡∏û‡∏¥‡πà‡∏°‡πÅ‡∏•‡πâ‡∏ß‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏´‡∏°
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            new_w = trip_w + branch_w
            new_c = trip_c + branch_c
            new_count = len(trip_data) + 1
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡πÉ‡∏™‡πà‡πÑ‡∏î‡πâ‡πÑ‡∏´‡∏° (‡∏¢‡∏≠‡∏°‡πÉ‡∏´‡πâ‡πÄ‡∏Å‡∏¥‡∏ô 125% ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏î‡∏µ‡∏¢‡∏ß)
            new_util = max(
                (new_w / LIMITS['6W']['max_w']) * 100,
                (new_c / LIMITS['6W']['max_c']) * 100
            )
            
            if new_util <= 100 and new_count <= MAX_BRANCHES_PER_TRIP:
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤
                trip_codes = set(trip_data['Code'].values) | {branch_code}
                max_allowed = get_max_vehicle_for_trip(trip_codes)
                
                # ‡∏ö‡∏±‡∏ô‡∏ó‡∏∂‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
                if distance < min_distance:
                    min_distance = distance
                    best_trip = trip_num
        
        # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏à‡∏≠‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏° ‚Üí ‡∏¢‡πâ‡∏≤‡∏¢
        if best_trip is not None:
            test_df.loc[test_df['Code'] == branch_code, 'Trip'] = best_trip
            reassign_count += 1
    
    # üéØ Phase 1.75: ‡∏£‡∏ß‡∏°‡∏ó‡∏£‡∏¥‡∏õ utilization ‡∏ï‡πà‡∏≥ (<50%) ‡πÉ‡∏´‡πâ‡πÄ‡∏ï‡πá‡∏°‡∏Ç‡∏∂‡πâ‡∏ô
    rebalance_count = 0
    LOW_UTIL_THRESHOLD = 50  # ‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏ï‡πà‡∏≥‡∏Å‡∏ß‡πà‡∏≤ 50% ‡∏ñ‡∏∑‡∏≠‡∏ß‡πà‡∏≤‡πÑ‡∏°‡πà‡∏Ñ‡∏∏‡πâ‡∏°
    
    # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà utilization ‡∏ï‡πà‡∏≥
    low_util_trips = []
    for trip_num in sorted(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        trip_w = trip_data['Weight'].sum()
        trip_c = trip_data['Cube'].sum()
        trip_count = len(trip_data)
        
        trip_util = max(
            (trip_w / LIMITS['6W']['max_w']) * 100,
            (trip_c / LIMITS['6W']['max_c']) * 100
        )
        
        # ‡∏ñ‡πâ‡∏≤ util < 50% ‡πÅ‡∏•‡∏∞‡∏°‡∏µ ‚â§6 ‡∏™‡∏≤‡∏Ç‡∏≤ ‚Üí ‡∏û‡∏¥‡∏à‡∏≤‡∏£‡∏ì‡∏≤‡∏£‡∏ß‡∏°
        if trip_util < LOW_UTIL_THRESHOLD and trip_count <= 6:
            # ‡∏´‡∏≤ centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ
            trip_lats, trip_lons = [], []
            for code in trip_data['Code'].values:
                lat, lon = get_lat_lon(code)
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if trip_lats:
                low_util_trips.append({
                    'trip': trip_num,
                    'util': trip_util,
                    'count': trip_count,
                    'weight': trip_w,
                    'cube': trip_c,
                    'codes': set(trip_data['Code'].values),
                    'lat': sum(trip_lats) / len(trip_lats),
                    'lon': sum(trip_lons) / len(trip_lons)
                })
    
    # ‡∏û‡∏¢‡∏≤‡∏¢‡∏≤‡∏°‡∏£‡∏ß‡∏°‡∏ó‡∏£‡∏¥‡∏õ‡∏ï‡πà‡∏≥‡∏Å‡∏±‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á
    for low_trip in low_util_trips:
        best_merge = None
        min_distance = float('inf')
        
        # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0 or trip_num == low_trip['trip']:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            trip_count = len(trip_data)
            
            # ‡∏Ç‡πâ‡∏≤‡∏°‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏¢‡∏≠‡∏∞‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ
            if trip_count >= MAX_BRANCHES_PER_TRIP:
                continue
            
            # ‡∏´‡∏≤ centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡∏ô‡∏µ‡πâ
            trip_lats, trip_lons = [], []
            for code in trip_data['Code'].values:
                lat, lon = get_lat_lon(code)
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if not trip_lats:
                continue
            
            target_lat = sum(trip_lats) / len(trip_lats)
            target_lon = sum(trip_lons) / len(trip_lons)
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á
            distance = haversine_distance(low_trip['lat'], low_trip['lon'], target_lat, target_lon)
            
            # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏Å‡∏•‡πÄ‡∏Å‡∏¥‡∏ô 50km ‚Üí ‡∏Ç‡πâ‡∏≤‡∏°
            if distance > 50:
                continue
            
            # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏£‡∏ß‡∏°‡πÅ‡∏•‡πâ‡∏ß‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏´‡∏°
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            combined_w = trip_w + low_trip['weight']
            combined_c = trip_c + low_trip['cube']
            combined_count = trip_count + low_trip['count']
            
            combined_util = max(
                (combined_w / LIMITS['6W']['max_w']) * 100,
                (combined_c / LIMITS['6W']['max_c']) * 100
            )
            
            # ‡∏£‡∏ß‡∏°‡πÑ‡∏î‡πâ‡∏ñ‡πâ‡∏≤ ‚â§120% ‡πÅ‡∏•‡∏∞‡∏™‡∏≤‡∏Ç‡∏≤ ‚â§MAX
            if combined_util <= 100 and combined_count <= MAX_BRANCHES_PER_TRIP:
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤
                combined_codes = low_trip['codes'] | set(trip_data['Code'].values)
                max_allowed = get_max_vehicle_for_trip(combined_codes)
                
                if distance < min_distance:
                    min_distance = distance
                    best_merge = trip_num
        
        # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏à‡∏≠‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏° ‚Üí ‡∏£‡∏ß‡∏°
        if best_merge is not None:
            for code in low_trip['codes']:
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_merge
            rebalance_count += 1
    
    # üéØ Phase 1.5: ‡πÄ‡∏Å‡πá‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏ô‡πÄ‡∏™‡πâ‡∏ô‡∏ó‡∏≤‡∏á (Route Pickup Optimization) - ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡πÄ‡∏ß‡∏•‡∏≤
    pickup_count = 0
    MAX_DETOUR_KM_LOCAL = MAX_DETOUR_KM  # ‡πÉ‡∏ä‡πâ‡∏Ñ‡πà‡∏≤‡∏à‡∏≤‡∏Å config (12 ‡∏Å‡∏°.)
    
    # ‚ö° Skip ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏ó‡∏£‡∏¥‡∏õ‡∏°‡∏≤‡∏Å‡πÄ‡∏Å‡∏¥‡∏ô 20 ‡∏ó‡∏£‡∏¥‡∏õ (‡∏õ‡∏£‡∏∞‡∏´‡∏¢‡∏±‡∏î‡πÄ‡∏ß‡∏•‡∏≤)
    unique_trips = test_df['Trip'].unique()
    if len(unique_trips) > 20:
        pass  # Skip Phase 1.5 ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏£‡πá‡∏ß
    else:
        # ‡∏ß‡∏ô‡∏•‡∏π‡∏õ‡∏ó‡∏∏‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏ï‡πá‡∏° (‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢ 95%) - ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡πÅ‡∏Ñ‡πà 15 ‡∏ó‡∏£‡∏¥‡∏õ‡πÅ‡∏£‡∏Å
        for trip_num in sorted(unique_trips)[:15]:
            trip_data = test_df[test_df['Trip'] == trip_num]
            current_w = trip_data['Weight'].sum()
            current_c = trip_data['Cube'].sum()
            current_count = len(trip_data)
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì % ‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏õ‡∏±‡∏à‡∏à‡∏∏‡∏ö‡∏±‡∏ô (‡πÉ‡∏ä‡πâ 6W ‡πÄ‡∏õ‡πá‡∏ô‡∏°‡∏≤‡∏ï‡∏£‡∏ê‡∏≤‡∏ô)
            current_util = max(
                (current_w / LIMITS['6W']['max_w']) * 100,
                (current_c / LIMITS['6W']['max_c']) * 100
            )
            
            # üéØ ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢: ‡πÄ‡∏Å‡πá‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏à‡∏ô‡πÄ‡∏ï‡πá‡∏°‡πÄ‡∏Å‡∏∑‡∏≠‡∏ö 100% (‡∏Ñ‡∏¥‡∏ß‡πÄ‡∏ï‡πá‡∏°)
            TARGET_UTIL = 100  # ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢ utilization
            MAX_PICKUP_UTIL = 100  # ‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏¢‡∏≠‡∏°‡πÄ‡∏Å‡πá‡∏ö‡πÑ‡∏î‡πâ (‡∏´‡πâ‡∏≤‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 100%)
            
            # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 100% ‡∏´‡∏£‡∏∑‡∏≠‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏¢‡∏≠‡∏∞‡πÅ‡∏•‡πâ‡∏ß ‚Üí ‡∏Ç‡πâ‡∏≤‡∏°
            if current_util >= MAX_PICKUP_UTIL or current_count >= MAX_BRANCHES_PER_TRIP:
                continue
            
            # ‡∏´‡∏≤‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡∏õ‡∏±‡∏à‡∏à‡∏∏‡∏ö‡∏±‡∏ô
            trip_provinces = set()
            trip_coords = []
            for code in trip_data['Code'].values:
                prov = get_province(code)
                if prov != 'UNKNOWN':
                    trip_provinces.add(prov)
                
                # ‡πÄ‡∏Å‡πá‡∏ö‡∏û‡∏¥‡∏Å‡∏±‡∏î
                lat, lon = get_lat_lon(code)
                if lat and lon:
                    trip_coords.append((lat, lon))
            
            # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏û‡∏¥‡∏Å‡∏±‡∏î ‚Üí ‡∏Ç‡πâ‡∏≤‡∏°
            if not trip_coords:
                continue
            
            # ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ (Trip = 0 ‡∏´‡∏£‡∏∑‡∏≠ NaN)
            unassigned = test_df[(test_df['Trip'] == 0) | (test_df['Trip'].isna())]
            
            for idx, row in unassigned.iterrows():
                branch_code = row['Code']
                branch_w = row['Weight']
                branch_c = row['Cube']
                branch_prov = get_province(branch_code)
                branch_lat, branch_lon = get_lat_lon(branch_code)
                
                # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏û‡∏¥‡∏Å‡∏±‡∏î ‚Üí ‡∏Ç‡πâ‡∏≤‡∏°
                if not branch_lat or not branch_lon:
                    continue
                
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏ô‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô‡∏´‡∏£‡∏∑‡∏≠‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á
                if branch_prov not in trip_provinces:
                    # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
                    min_distance = float('inf')
                    for trip_lat, trip_lon in trip_coords:
                        dist = haversine_distance(trip_lat, trip_lon, branch_lat, branch_lon)
                        if dist < min_distance:
                            min_distance = dist
                    
                    # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏ô‡πÄ‡∏™‡πâ‡∏ô‡∏ó‡∏≤‡∏á (‡πÑ‡∏Å‡∏•‡πÄ‡∏Å‡∏¥‡∏ô‡∏à‡∏≤‡∏Å‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤) ‚Üí ‡∏Ç‡πâ‡∏≤‡∏°
                    if min_distance > MAX_DETOUR_KM_LOCAL:
                        continue
                
                # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏ß‡πà‡∏≤‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ‡πÅ‡∏•‡πâ‡∏ß‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏´‡∏°
                new_w = current_w + branch_w
                new_c = current_c + branch_c
                new_count = current_count + 1
                
                # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì % ‡πÉ‡∏´‡∏°‡πà (‡πÄ‡∏ô‡πâ‡∏ô Cube)
                new_cube_util = (new_c / LIMITS['6W']['max_c']) * 100
                new_weight_util = (new_w / LIMITS['6W']['max_w']) * 100
                new_util = max(new_cube_util, new_weight_util)
                
                # üéØ ‡∏ñ‡πâ‡∏≤‡∏£‡∏ñ‡πÑ‡∏°‡πà‡πÄ‡∏ï‡πá‡∏° (<95%) ‚Üí ‡∏¢‡∏≠‡∏°‡πÉ‡∏´‡πâ‡πÄ‡∏û‡∏¥‡πà‡∏°‡πÅ‡∏ï‡πà‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 100%
                # ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢: Cube 95-100%, ‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å ‚â§100%
                if current_util < 95:
                    # ‡∏£‡∏ñ‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏ï‡πá‡∏° ‚Üí ‡∏¢‡∏≠‡∏°‡πÉ‡∏´‡πâ‡πÄ‡∏û‡∏¥‡πà‡∏°‡πÅ‡∏ï‡πà‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 100%
                    can_add = new_cube_util <= 100 and new_weight_util <= 100 and new_count <= MAX_BRANCHES_PER_TRIP
                else:
                    # ‡∏£‡∏ñ‡πÄ‡∏ï‡πá‡∏°‡∏û‡∏≠‡∏™‡∏°‡∏Ñ‡∏ß‡∏£‡πÅ‡∏•‡πâ‡∏ß ‚Üí ‡πÄ‡∏Ç‡πâ‡∏°‡∏á‡∏ß‡∏î (‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 100%)
                    can_add = new_cube_util <= 100 and new_weight_util <= 100 and new_count <= MAX_BRANCHES_PER_TRIP
                
                if can_add:
                    # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤
                    test_trip_codes = set(trip_data['Code'].values) | {branch_code}
                    max_allowed = get_max_vehicle_for_trip(test_trip_codes)
                    
                    # ‡∏ñ‡πâ‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å‡∏Å‡∏ß‡πà‡∏≤‡∏£‡∏ñ‡∏õ‡∏±‡∏à‡∏à‡∏∏‡∏ö‡∏±‡∏ô ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡πÉ‡∏™‡πà‡πÑ‡∏î‡πâ‡πÑ‡∏´‡∏°
                    # (‡∏õ‡∏•‡πà‡∏≠‡∏¢‡πÉ‡∏´‡πâ Phase 2 ‡∏à‡∏±‡∏î‡∏Å‡∏≤‡∏£)
                    
                    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏Ç‡πâ‡∏≤‡∏ó‡∏£‡∏¥‡∏õ
                    test_df.loc[test_df['Code'] == branch_code, 'Trip'] = trip_num
                    
                    # ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏õ‡∏±‡∏à‡∏à‡∏∏‡∏ö‡∏±‡∏ô
                    current_w = new_w
                    current_c = new_c
                    current_count = new_count
                    current_util = new_util
                    
                    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏û‡∏¥‡∏Å‡∏±‡∏î‡πÉ‡∏´‡∏°‡πà
                    trip_coords.append((branch_lat, branch_lon))
                    if branch_prov != 'UNKNOWN':
                        trip_provinces.add(branch_prov)
                    
                    pickup_count += 1
                    
                    # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏ï‡πá‡∏°‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ‡πÅ‡∏•‡πâ‡∏ß (Cube >100% ‡∏´‡∏£‡∏∑‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô MAX) ‚Üí ‡∏´‡∏¢‡∏∏‡∏î‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤
                    current_cube_util = (current_c / LIMITS['6W']['max_c']) * 100
                    if current_cube_util >= 100 or current_count >= MAX_BRANCHES_PER_TRIP:
                        break
    
    # üö® Phase 1.75: ‡πÅ‡∏¢‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ (4W/JB) ‡∏≠‡∏≠‡∏Å‡∏à‡∏≤‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà
    restriction_split_count = 0
    
    # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡∏ú‡∏™‡∏°‡∏Å‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏°‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î
    for trip_num in sorted(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        trip_codes = set(trip_data['Code'].values)
        
        # ‡πÅ‡∏¢‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡∏≤‡∏°‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ
        codes_4w_only = set()  # ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ 4W ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
        codes_jb_or_less = set()  # ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ‡πÅ‡∏Ñ‡πà JB ‡∏´‡∏£‡∏∑‡∏≠‡πÄ‡∏•‡πá‡∏Å‡∏Å‡∏ß‡πà‡∏≤
        codes_no_limit = set()  # ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î (‡πÉ‡∏ä‡πâ 6W ‡πÑ‡∏î‡πâ)
        
        for code in trip_codes:
            max_vehicle = get_max_vehicle_for_trip({code})
            if max_vehicle == '4W':
                codes_4w_only.add(code)
            elif max_vehicle == 'JB':
                codes_jb_or_less.add(code)
            else:
                codes_no_limit.add(code)
        
        # üö® ‡∏ñ‡πâ‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏£‡∏ñ‡∏ú‡∏™‡∏°‡∏Å‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏°‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å
        has_restrictions = len(codes_4w_only) > 0 or len(codes_jb_or_less) > 0
        has_no_limits = len(codes_no_limit) > 0
        
        if has_restrictions and has_no_limits:
            # ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô 2 ‡∏Å‡∏•‡∏∏‡πà‡∏°: 1) ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î 2) ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î
            restricted_codes = codes_4w_only | codes_jb_or_less
            unrestricted_codes = codes_no_limit
            
            # ‡πÄ‡∏Å‡πá‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏î‡∏¥‡∏°‡πÉ‡∏´‡πâ‡∏Å‡∏±‡∏ö‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤‡∏°‡∏≤‡∏Å‡∏Å‡∏ß‡πà‡∏≤
            if len(restricted_codes) >= len(unrestricted_codes):
                # restricted ‡πÉ‡∏ä‡πâ‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏î‡∏¥‡∏°
                keep_trip = trip_num
                new_trip = test_df['Trip'].max() + 1
                
                # ‡∏¢‡πâ‡∏≤‡∏¢ unrestricted ‡πÑ‡∏õ‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
                for code in unrestricted_codes:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip
            else:
                # unrestricted ‡πÉ‡∏ä‡πâ‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏î‡∏¥‡∏°
                keep_trip = trip_num
                new_trip = test_df['Trip'].max() + 1
                
                # ‡∏¢‡πâ‡∏≤‡∏¢ restricted ‡πÑ‡∏õ‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
                for code in restricted_codes:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip
            
            restriction_split_count += 1
    
    # üéØ Phase 2: ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏° (‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å 4W ‚Üí JB ‚Üí 6W ‡∏´‡∏£‡∏∑‡∏≠ 2 ‡∏Ñ‡∏±‡∏ô) - Optimized
    vehicle_assignment_count = 0
    downsize_count = 0
    region_changes = {
        '4w': 0, 
        'jb': 0, 
        '6w': 0, 
        'split_2_vehicles': 0,
        'nearby_6w_to_jb': 0,
        'far_keep_6w': 0,
        'other': 0
    }
    
    # ‚ö° Early stopping - ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πâ‡πÄ‡∏ß‡∏•‡∏≤‡∏°‡∏≤‡∏Å‡∏Å‡∏ß‡πà‡∏≤ 55 ‡∏ß‡∏¥‡∏ô‡∏≤‡∏ó‡∏µ
    if time.time() - start_time > 55:
        # Skip Phase 2 complex logic, ‡πÉ‡∏ä‡πâ logic ‡πÄ‡∏£‡πá‡∏ß
        for trip_num in test_df['Trip'].unique():
            trip_data = test_df[test_df['Trip'] == trip_num]
            total_c = trip_data['Cube'].sum()
            
            # ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ‡πÅ‡∏ö‡∏ö‡πÄ‡∏£‡πá‡∏ß (‡πÑ‡∏°‡πà‡∏°‡∏µ optimization)
            if total_c <= 5:
                trip_recommended_vehicles[trip_num] = '4W'
            elif total_c <= 7:
                trip_recommended_vehicles[trip_num] = 'JB'
            else:
                trip_recommended_vehicles[trip_num] = '6W'
    else:
        # ‡πÄ‡∏Å‡πá‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏•‡∏∑‡∏≠
        for trip_num in test_df['Trip'].unique():
            trip_data = test_df[test_df['Trip'] == trip_num]
            branch_count = len(trip_data)
            total_w = trip_data['Weight'].sum()
            total_c = trip_data['Cube'].sum()
            trip_codes = set(trip_data['Code'].values)
            
            # üîí ‡πÄ‡∏ä‡πá‡∏Ñ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î - ‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç‡∏°‡∏≤‡∏Å! ‡∏´‡πâ‡∏≤‡∏° 6W ‡πÉ‡∏ô‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•
            provinces = set()
            for code in trip_codes:
                prov = get_province(code)
                if prov and prov != 'UNKNOWN':
                    provinces.add(prov)
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏ó‡∏∏‡∏Å‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏õ‡πá‡∏ô‡∏û‡∏∑‡πâ‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
            all_nearby = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
            has_north = any(get_region_type(p) == 'north' for p in provinces) if provinces else False
            has_south = any(get_region_type(p) == 'south' for p in provinces) if provinces else False
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á max ‡∏à‡∏≤‡∏Å DC
            max_distance_from_dc = 0
            for code in trip_codes:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    dist = haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
                    max_distance_from_dc = max(max_distance_from_dc, dist)
            
            # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤
            max_allowed = get_max_vehicle_for_trip(trip_codes)
            
            # üîí ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö Punthai - ‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç‡∏û‡∏¥‡πÄ‡∏®‡∏©
            punthai_limits = get_punthai_vehicle_limits(trip_data, total_c, branch_count)
            punthai_type = is_punthai_only(trip_data)
            
            # ‡∏ñ‡πâ‡∏≤ Punthai ‡∏•‡πâ‡∏ß‡∏ô ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏ï‡∏≤‡∏° Punthai limits
            if punthai_type == 'punthai_only':
                # Punthai ‡∏•‡πâ‡∏ß‡∏ô: JB ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 7 drop, 4W ‡∏ñ‡πâ‡∏≤ Cube > 5 ‚Üí ‡∏ï‡∏±‡∏î‡πÄ‡∏õ‡πá‡∏ô JB
                punthai_max_vehicle = punthai_limits['max_vehicle']
                punthai_max_drops = punthai_limits['max_drops']
                
                # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö max_allowed ‡∏ï‡∏≤‡∏° Punthai
                vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
                if vehicle_priority.get(punthai_max_vehicle, 3) < vehicle_priority.get(max_allowed, 3):
                    max_allowed = punthai_max_vehicle
                
            elif punthai_type == 'mixed':
                # ‡∏ú‡∏™‡∏° Punthai + ‡∏≠‡∏∑‡πà‡∏ô: ‡∏ñ‡πâ‡∏≤ Cube 3-4 ‚Üí 6W ‡πÑ‡∏î‡πâ, ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô ‚Üí 4W ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
                if total_c > 4.0:
                    max_allowed = '4W'  # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö 4W ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
                # ‡∏ñ‡πâ‡∏≤ Cube 3-4 ‚Üí ‡∏¢‡∏±‡∏á‡πÉ‡∏ä‡πâ max_allowed ‡∏õ‡∏Å‡∏ï‡∏¥ (6W ‡πÑ‡∏î‡πâ)
            
            # ‚ö†Ô∏è ‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç: ‡∏ñ‡πâ‡∏≤‡∏ó‡∏∏‡∏Å‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏õ‡πá‡∏ô nearby ‚Üí ‡∏´‡πâ‡∏≤‡∏° 6W ‡πÄ‡∏î‡πá‡∏î‡∏Ç‡∏≤‡∏î!
            if all_nearby:
                very_far = False  # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏´‡πâ‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πâ 6W
                if max_allowed == '6W':
                    max_allowed = 'JB'  # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏Ç‡∏≠‡∏ö‡πÄ‡∏Ç‡∏ï‡πÄ‡∏õ‡πá‡∏ô JB
            # ‚ö†Ô∏è ‡∏†‡∏≤‡∏Ñ‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡πÅ‡∏•‡∏∞‡∏†‡∏≤‡∏Ñ‡πÉ‡∏ï‡πâ ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏ä‡πâ 6W
            elif has_north or has_south:
                very_far = True  # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏´‡πâ‡πÉ‡∏ä‡πâ 6W
            else:
                # üöõ ‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á - ‡πÑ‡∏Å‡∏•‡∏°‡∏≤‡∏Å‡∏û‡∏¥‡πÄ‡∏®‡∏© (>300km) ‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏ä‡πâ 6W
                very_far_by_distance = max_distance_from_dc > 300
                very_far = very_far_by_distance
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì % ‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó
            util_4w = max((total_w / LIMITS['4W']['max_w']) * 100, 
                          (total_c / LIMITS['4W']['max_c']) * 100)
            util_jb = max((total_w / LIMITS['JB']['max_w']) * 100,
                          (total_c / LIMITS['JB']['max_c']) * 100)
            util_6w = max((total_w / LIMITS['6W']['max_w']) * 100,
                          (total_c / LIMITS['6W']['max_c']) * 100)
            
            # üîí ‡πÑ‡∏°‡πà‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏£‡∏µ‡∏¢‡∏Å get_max_vehicle_for_trip ‡∏≠‡∏µ‡∏Å - ‡πÉ‡∏ä‡πâ‡∏Ñ‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö all_nearby ‡πÅ‡∏•‡πâ‡∏ß
            
            # üéØ ‡∏Å‡∏•‡∏¢‡∏∏‡∏ó‡∏ò‡πå‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ (‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å 4W ‚Üí JB ‚Üí ‡πÅ‡∏¢‡∏Å 2 ‡∏Ñ‡∏±‡∏ô/6W)
            recommended = None
            cube_util_4w = (total_c / LIMITS['4W']['max_c']) * 100
            cube_util_jb = (total_c / LIMITS['JB']['max_c']) * 100
            cube_util_6w = (total_c / LIMITS['6W']['max_c']) * 100
            weight_util_4w = (total_w / LIMITS['4W']['max_w']) * 100
            weight_util_jb = (total_w / LIMITS['JB']['max_w']) * 100
            weight_util_6w = (total_w / LIMITS['6W']['max_w']) * 100
            
            # üö® ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤‡∏Å‡πà‡∏≠‡∏ô
            if max_allowed == '4W':
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 1: ‡∏•‡∏≠‡∏á 4W ‡∏Å‡πà‡∏≠‡∏ô (95-100%)
                if 95 <= cube_util_4w <= 100 and weight_util_4w <= 100 and branch_count <= 12:
                    recommended = '4W'
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 2: ‡∏ñ‡πâ‡∏≤ 4W ‡πÑ‡∏°‡πà‡∏û‡∏≠‡∏î‡∏µ ‚Üí ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô 4W + 4W (75-95% ‡∏ï‡πà‡∏≠‡∏Ñ‡∏±‡∏ô)
                elif cube_util_4w > 100:
                    # ‡∏à‡∏∞‡πÅ‡∏¢‡∏Å‡πÉ‡∏ô Phase 2.5
                    recommended = '4W+4W'
                else:
                    # ‡∏ï‡πà‡∏≥‡∏Å‡∏ß‡πà‡∏≤ 95% ‚Üí ‡πÉ‡∏ä‡πâ 4W (‡πÅ‡∏ï‡πà‡∏≠‡∏≤‡∏à‡∏£‡∏ß‡∏°‡∏Å‡∏±‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô‡∏†‡∏≤‡∏¢‡∏´‡∏•‡∏±‡∏á)
                    recommended = '4W'
            elif max_allowed == 'JB':
                # üîí Punthai ‡∏•‡πâ‡∏ß‡∏ô: JB ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 7 drop
                max_jb_drops = 7 if punthai_type == 'punthai_only' else 12
                
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 1: ‡∏•‡∏≠‡∏á 4W ‡∏Å‡πà‡∏≠‡∏ô (95-100%)
                if 95 <= cube_util_4w <= 100 and weight_util_4w <= 100 and branch_count <= 12:
                    recommended = '4W'
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 2: ‡∏•‡∏≠‡∏á JB (95-100%) - ‡πÄ‡∏ä‡πá‡∏Ñ drop limit ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö Punthai
                elif 95 <= cube_util_jb <= 100 and weight_util_jb <= 100 and branch_count <= max_jb_drops:
                    recommended = 'JB'
                # ‡∏•‡∏≥‡∏î‡∏±‡∏ö 3: ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB + 4W ‡∏´‡∏£‡∏∑‡∏≠ JB + JB (75-95% ‡∏ï‡πà‡∏≠‡∏Ñ‡∏±‡∏ô)
                elif cube_util_jb > 100 or branch_count > max_jb_drops:
                    # ‡∏•‡∏≠‡∏á‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB + 4W (13 cube max)
                    if total_c <= 13:
                        recommended = 'JB+4W'
                    else:
                        recommended = 'JB+JB'  # 16 cube max
                else:
                    # ‡∏ï‡πà‡∏≥‡∏Å‡∏ß‡πà‡∏≤ 95% ‚Üí ‡πÉ‡∏ä‡πâ JB ‡∏´‡∏£‡∏∑‡∏≠ 4W
                    if cube_util_jb >= 75 and branch_count <= max_jb_drops:
                        recommended = 'JB'
                    else:
                        recommended = '4W'
            # üöõ ‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û+‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏• (nearby) ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏´‡πâ‡∏≤‡∏° 6W (‡∏•‡∏≥‡∏î‡∏±‡∏ö‡πÅ‡∏£‡∏Å‡∏™‡∏∏‡∏î!)
            elif all_nearby:
                # üîí Punthai ‡∏•‡πâ‡∏ß‡∏ô: JB ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 7 drop
                max_jb_drops = 7 if punthai_type == 'punthai_only' else 12
                
                # ‡∏•‡∏≠‡∏á 4W ‡∏Å‡πà‡∏≠‡∏ô
                if cube_util_4w <= 100 and weight_util_4w <= 100:
                    recommended = '4W'
                # ‡∏ñ‡πâ‡∏≤ 4W ‡πÑ‡∏°‡πà‡∏û‡∏≠ ‚Üí ‡∏•‡∏≠‡∏á JB (‡πÄ‡∏ä‡πá‡∏Ñ drop limit ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö Punthai)
                elif cube_util_jb <= 100 and weight_util_jb <= 100 and branch_count <= max_jb_drops:
                    recommended = 'JB'
                    region_changes['nearby_6w_to_jb'] += 1
                # ‡∏ñ‡πâ‡∏≤ JB ‡∏Å‡πá‡πÑ‡∏°‡πà‡∏û‡∏≠ ‡∏´‡∏£‡∏∑‡∏≠ Punthai ‡πÄ‡∏Å‡∏¥‡∏ô 7 drop ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ
                else:
                    recommended = 'JB'  # ‡∏Å‡∏≥‡∏´‡∏ô‡∏î‡πÑ‡∏ß‡πâ‡∏Å‡πà‡∏≠‡∏ô ‡∏à‡∏∞‡πÅ‡∏¢‡∏Å‡∏†‡∏≤‡∏¢‡∏´‡∏•‡∏±‡∏á
                    region_changes['nearby_6w_to_jb'] += 1
            # üöõ ‡∏†‡∏≤‡∏Ñ‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏ä‡πâ 6W ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô (‡∏ñ‡πâ‡∏≤‡πÑ‡∏î‡πâ ‚â•18 cube)
            elif has_north:
                if total_c >= 18.0:
                    recommended = '6W'
                    region_changes['far_keep_6w'] += 1
                else:
                    # ‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡πÅ‡∏ï‡πà‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á 18 cube ‚Üí ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB
                    recommended = 'JB'
                    region_changes['other'] += 1
            # üöõ ‡∏†‡∏≤‡∏Ñ‡πÉ‡∏ï‡πâ‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏ä‡πâ 6W ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô (‡∏ñ‡πâ‡∏≤‡πÑ‡∏î‡πâ ‚â•18 cube)
            elif has_south:
                if total_c >= 18.0:
                    recommended = '6W'
                    region_changes['far_keep_6w'] += 1
                else:
                    # ‡πÉ‡∏ï‡πâ‡πÅ‡∏ï‡πà‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á 18 cube ‚Üí ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB
                    recommended = 'JB'
                    region_changes['other'] += 1
            else:
                # üéØ ‡∏û‡∏∑‡πâ‡∏ô‡∏ó‡∏µ‡πà‡πÑ‡∏Å‡∏• (far) - ‡∏¢‡∏∑‡∏î‡∏´‡∏¢‡∏∏‡πà‡∏ô ‡πÉ‡∏ä‡πâ JB ‡πÑ‡∏î‡πâ‡∏ñ‡πâ‡∏≤‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°
                # ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢: 6W ‡∏ï‡πâ‡∏≠‡∏á‡πÑ‡∏î‡πâ‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥ 18 cube (90%), ‡∏´‡πâ‡∏≤‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 20 cube
                
                MIN_6W_CUBE = 18.0  # 6W ‡∏ï‡πâ‡∏≠‡∏á‡πÑ‡∏î‡πâ‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥ 18 cube
                MAX_6W_CUBE = 20.0  # 6W ‡∏´‡πâ‡∏≤‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 20 cube
                MIN_UTIL_THRESHOLD = 75   # ‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥ - ‡∏´‡πâ‡∏≤‡∏°‡∏£‡∏ñ‡πÄ‡∏´‡∏•‡∏∑‡∏≠‡∏ï‡πà‡∏≥‡∏Å‡∏ß‡πà‡∏≤‡∏ô‡∏µ‡πâ
                TARGET_MIN = 95 # ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢‡∏Ç‡∏±‡πâ‡∏ô‡∏ï‡πà‡∏≥
                TARGET_MAX = 100 # ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î (‡∏´‡πâ‡∏≤‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 100%)
                
                # üîí 6W ‡πÄ‡∏Å‡∏¥‡∏ô 20 cube ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å ‡∏™‡πà‡∏ß‡∏ô‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ 4W
                if total_c > MAX_6W_CUBE:
                    recommended = '6W'  # ‡∏à‡∏∞‡πÅ‡∏¢‡∏Å‡∏™‡πà‡∏ß‡∏ô‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ 4W ‡πÉ‡∏ô Phase 2.5
                    region_changes['far_keep_6w'] += 1
                
                # üéØ 6W ‡πÑ‡∏î‡πâ 18-20 cube ‚Üí ‡πÉ‡∏ä‡πâ 6W ‚úÖ
                elif total_c >= MIN_6W_CUBE:
                    recommended = '6W'
                    region_changes['far_keep_6w'] += 1
                
                # üîí 6W ‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á 18 cube (7-18) ‚Üí ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB ‡πÅ‡∏ó‡∏ô
                elif total_c >= 7.0:
                    recommended = 'JB'  # ‡∏à‡∏∞‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB ‡∏´‡∏•‡∏≤‡∏¢‡∏Ñ‡∏±‡∏ô‡πÉ‡∏ô Phase 2.1
                    region_changes['other'] += 1
                
                # 2. ‡∏ñ‡πâ‡∏≤ JB ‡∏û‡∏≠‡∏î‡∏µ (95-100%) ‚Üí ‡πÉ‡∏ä‡πâ JB ‚úÖ
                elif TARGET_MIN <= cube_util_jb <= TARGET_MAX and weight_util_jb <= TARGET_MAX:
                    recommended = 'JB'
                    region_changes['other'] += 1
                
                # 3. 4W ‡∏û‡∏≠‡∏î‡∏µ ‚Üí ‡πÉ‡∏ä‡πâ 4W
                elif cube_util_4w <= TARGET_MAX and weight_util_4w <= TARGET_MAX:
                    recommended = '4W'
                    region_changes['other'] += 1
                
                # 4. ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ó‡∏≤‡∏á‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏≠‡∏∑‡πà‡∏ô ‚Üí ‡πÉ‡∏ä‡πâ JB
                else:
                    recommended = 'JB'
                    region_changes['other'] += 1
            
            # üö® ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏ä‡πâ max_allowed ‡∏ñ‡πâ‡∏≤‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÅ‡∏ô‡∏∞‡∏ô‡∏≥‡πÉ‡∏´‡∏ç‡πà‡∏Å‡∏ß‡πà‡∏≤‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î (‡∏´‡πâ‡∏≤‡∏°‡∏Ç‡πâ‡∏≤‡∏°!)
            # ‚úÖ ‡∏ó‡∏∏‡∏Å‡∏†‡∏≤‡∏Ñ‡∏ï‡πâ‡∏≠‡∏á‡∏î‡∏π Auto Plan
            vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
            recommended_priority = vehicle_priority.get(recommended, 3)
            allowed_priority = vehicle_priority.get(max_allowed, 3)
            
            if recommended_priority > allowed_priority:
                # ‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÅ‡∏ô‡∏∞‡∏ô‡∏≥‡πÉ‡∏´‡∏ç‡πà‡∏Å‡∏ß‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÉ‡∏ä‡πâ max_allowed (‡∏´‡πâ‡∏≤‡∏°‡∏Ç‡πâ‡∏≤‡∏°‡∏Ç‡∏±‡πâ‡∏ô!)
                recommended = max_allowed
            
            # üîí Double check: ‡∏´‡πâ‡∏≤‡∏°‡∏Ç‡πâ‡∏≤‡∏°‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏î‡πá‡∏î‡∏Ç‡∏≤‡∏î!
            if max_allowed == '4W' and recommended != '4W':
                recommended = '4W'
            elif max_allowed == 'JB' and recommended == '6W':
                recommended = 'JB'
            
            # üîí Triple check: ‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û+‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏• ‡∏´‡πâ‡∏≤‡∏° 6W ‡πÄ‡∏î‡πá‡∏î‡∏Ç‡∏≤‡∏î!
            if all_nearby and recommended == '6W':
                # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡πÄ‡∏õ‡πá‡∏ô JB
                recommended = 'JB'
                region_changes['nearby_6w_to_jb'] += 1
            
            # ‡∏ö‡∏±‡∏ô‡∏ó‡∏∂‡∏Å‡∏Å‡∏≤‡∏£‡∏õ‡∏£‡∏±‡∏ö‡∏Ç‡∏ô‡∏≤‡∏î
            original_vehicle = trip_recommended_vehicles.get(trip_num, '6W')
            trip_recommended_vehicles[trip_num] = recommended
            if recommended != original_vehicle:
                downsize_count += 1
    
    # Phase 2 completed
    
    # üö® ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏≠‡∏µ‡∏Å‡∏Ñ‡∏£‡∏±‡πâ‡∏á: ‡∏´‡πâ‡∏≤‡∏°‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û+‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•+‡∏†‡∏≤‡∏Ñ‡∏Å‡∏•‡∏≤‡∏á‡πÉ‡∏ä‡πâ 6W (‡πÄ‡∏Ç‡πâ‡∏°‡∏á‡∏ß‡∏î)
    bangkok_6w_count = 0
    bangkok_6w_splits = 0
    
    for trip_num in test_df['Trip'].unique():
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        trip_codes = list(trip_data['Code'].values)
        
        # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏ó‡∏∏‡∏Å‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏õ‡πá‡∏ô‡∏û‡∏∑‡πâ‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
        provinces = set()
        for code in trip_codes:
            prov = get_province(code)
            if prov != 'UNKNOWN':
                provinces.add(prov)
        
        all_nearby = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
        current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
        
        if all_nearby and current_vehicle == '6W':
            # ‡∏û‡∏¢‡∏≤‡∏¢‡∏≤‡∏°‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡πÄ‡∏õ‡πá‡∏ô JB ‡∏Å‡πà‡∏≠‡∏ô
            total_w = trip_data['Weight'].sum()
            total_c = trip_data['Cube'].sum()
            jb_util = max((total_w / LIMITS['JB']['max_w']) * 100, (total_c / LIMITS['JB']['max_c']) * 100)
            
            if jb_util <= 100:
                # JB ‡πÉ‡∏™‡πà‡πÑ‡∏î‡πâ ‚Üí ‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡πÄ‡∏õ‡πá‡∏ô JB
                trip_recommended_vehicles[trip_num] = 'JB'
                bangkok_6w_count += 1
            else:
                # JB ‡πÄ‡∏ï‡πá‡∏° ‚Üí ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB ‡∏´‡∏•‡∏≤‡∏¢‡∏Ñ‡∏±‡∏ô
                new_trips = []
                current_group = []
                current_group_w = 0
                current_group_c = 0
                
                sorted_data = trip_data.sort_values('Weight', ascending=False)
                
                for _, row in sorted_data.iterrows():
                    code = row['Code']
                    w = row['Weight']
                    c = row['Cube']
                    
                    test_w = current_group_w + w
                    test_c = current_group_c + c
                    test_util = max((test_w / LIMITS['JB']['max_w']) * 100, (test_c / LIMITS['JB']['max_c']) * 100)
                    
                    if test_util <= 100 or len(current_group) == 0:
                        current_group.append(code)
                        current_group_w += w
                        current_group_c += c
                    else:
                        new_trips.append(current_group.copy())
                        current_group = [code]
                        current_group_w = w
                        current_group_c = c
                
                if current_group:
                    new_trips.append(current_group)
                
                # ‡∏≠‡∏±‡∏û‡πÄ‡∏î‡∏ó‡∏ó‡∏£‡∏¥‡∏õ
                if len(new_trips) > 1:
                    for code in new_trips[0]:
                        test_df.loc[test_df['Code'] == code, 'Trip'] = trip_num
                    trip_recommended_vehicles[trip_num] = 'JB'
                    
                    for group in new_trips[1:]:
                        new_trip_num = test_df['Trip'].max() + 1
                        for code in group:
                            test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip_num
                        trip_recommended_vehicles[new_trip_num] = 'JB'
                        bangkok_6w_splits += 1
                else:
                    trip_recommended_vehicles[trip_num] = 'JB'
                    bangkok_6w_count += 1
    
    # üö® Phase 2.1: ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡πÅ‡∏•‡∏∞‡πÅ‡∏Å‡πâ‡πÑ‡∏Ç‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡πÄ‡∏Å‡∏¥‡∏ô‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î (‡∏•‡∏î‡∏Å‡∏≤‡∏£ loop)
    fix_count = 0
    split_count = 0
    
    for trip_num in test_df['Trip'].unique():
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        trip_codes = list(trip_data['Code'].values)
        current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
        max_allowed = get_max_vehicle_for_trip(set(trip_codes))
        
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        
        # üîí ‡πÄ‡∏ä‡πá‡∏Ñ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î - ‡∏´‡πâ‡∏≤‡∏° 6W ‡πÉ‡∏ô‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•!
        provinces = set()
        for code in trip_codes:
            prov = get_province(code)
            if prov and prov != 'UNKNOWN':
                provinces.add(prov)
        all_nearby = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
        
        # üîí ‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏• = ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö JB ‡∏´‡∏£‡∏∑‡∏≠ 4W (‡∏´‡πâ‡∏≤‡∏° 6W)
        if all_nearby and max_allowed == '6W':
            max_allowed = 'JB'
        
        # üîí ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡∏≠‡∏¢‡∏π‡πà‡πÉ‡∏´‡∏ç‡πà‡∏Å‡∏ß‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà (‡∏´‡πâ‡∏≤‡∏°‡∏Ç‡πâ‡∏≤‡∏°‡∏Ç‡∏±‡πâ‡∏ô!)
        vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
        current_priority = vehicle_priority.get(current_vehicle, 3)
        allowed_priority = vehicle_priority.get(max_allowed, 3)
        
        allowed_w = LIMITS[max_allowed]['max_w']
        allowed_c = LIMITS[max_allowed]['max_c']
        
        # ‡πÄ‡∏ä‡πá‡∏Ñ utilization ‡∏Ç‡∏≠‡∏á‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï
        util_allowed = max((total_w / allowed_w) * 100, (total_c / allowed_c) * 100)
        
        # üö® ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö: ‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏Ñ‡∏≤‡∏£‡∏û‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤ ‡πÑ‡∏°‡πà‡∏ß‡πà‡∏≤‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥‡∏à‡∏∞‡∏ö‡∏≠‡∏Å‡∏≠‡∏∞‡πÑ‡∏£!
        # ‡∏Å‡∏£‡∏ì‡∏µ‡∏ó‡∏µ‡πà 1: ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏Å‡∏ß‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡∏´‡∏£‡∏∑‡∏≠‡πÅ‡∏¢‡∏Å
        # ‡∏Å‡∏£‡∏ì‡∏µ‡∏ó‡∏µ‡πà 2: ‡∏£‡∏ñ‡∏ï‡∏≤‡∏°‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡πÅ‡∏ï‡πà‡πÉ‡∏™‡πà‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ (>100%) ‚Üí ‡∏ï‡∏±‡∏î‡πÅ‡∏¢‡∏Å‡∏ó‡∏±‡∏ô‡∏ó‡∏µ!
        if current_priority > allowed_priority or util_allowed > 100:
            if util_allowed <= 100:
                # ‡πÉ‡∏™‡πà‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡∏≠‡∏ô‡∏∏‡∏ç‡∏≤‡∏ï‡πÑ‡∏î‡πâ ‚Üí ‡∏õ‡∏£‡∏±‡∏ö‡∏£‡∏ñ
                trip_recommended_vehicles[trip_num] = max_allowed
                fix_count += 1
            else:
                # üö® ‡πÉ‡∏™‡πà‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ
                # üéØ ‡∏Å‡∏•‡∏¢‡∏∏‡∏ó‡∏ò‡πå‡πÉ‡∏´‡∏°‡πà: 
                #   - ‡∏ñ‡πâ‡∏≤‡∏à‡∏≥‡∏Å‡∏±‡∏î 4W ‚Üí ‡∏•‡∏≠‡∏á 4W ‡∏Å‡πà‡∏≠‡∏ô ‚Üí ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô‡∏Ñ‡πà‡∏≠‡∏¢‡∏ï‡∏±‡∏î‡πÄ‡∏õ‡πá‡∏ô JB ‡∏´‡∏•‡∏≤‡∏¢‡∏Ñ‡∏±‡∏ô
                #   - ‡∏ñ‡πâ‡∏≤‡∏à‡∏≥‡∏Å‡∏±‡∏î JB ‚Üí ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB ‡∏´‡∏•‡∏≤‡∏¢‡∏Ñ‡∏±‡∏ô
                #   - ‡∏´‡πâ‡∏≤‡∏° 6W ‡πÉ‡∏ô‡∏û‡∏∑‡πâ‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏î‡πÜ ‚Üí ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å (4W/JB) ‡πÄ‡∏•‡∏¢
                
                target_vehicle = max_allowed
                
                # üö® ‡∏Å‡∏£‡∏ì‡∏µ‡∏û‡∏¥‡πÄ‡∏®‡∏©: 4W ‡∏à‡∏≥‡∏Å‡∏±‡∏î ‚Üí ‡∏•‡∏≠‡∏á 4W ‡∏Å‡πà‡∏≠‡∏ô
                if max_allowed == '4W':
                    # ‡∏•‡∏≠‡∏á 4W ‡∏´‡∏•‡∏≤‡∏¢‡∏Ñ‡∏±‡∏ô‡∏Å‡πà‡∏≠‡∏ô
                    fourw_w = LIMITS['4W']['max_w']
                    fourw_c = LIMITS['4W']['max_c']
                    fourw_util = max((total_w / fourw_w) * 100, (total_c / fourw_c) * 100)
                    
                    # ‡∏ñ‡πâ‡∏≤ 4W ‡πÉ‡∏™‡πà‡πÑ‡∏î‡πâ (‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 100%) ‚Üí ‡πÉ‡∏ä‡πâ 4W
                    if fourw_util <= 100:
                        trip_recommended_vehicles[trip_num] = '4W'
                        fix_count += 1
                        continue
                    else:
                        # 4W ‡πÄ‡∏ï‡πá‡∏° ‚Üí ‡∏ï‡∏±‡∏î‡πÄ‡∏õ‡πá‡∏ô JB (‡πÄ‡∏û‡∏£‡∏≤‡∏∞‡∏´‡πâ‡∏≤‡∏° 6W ‡πÉ‡∏ô‡∏û‡∏∑‡πâ‡∏ô‡∏ó‡∏µ‡πà‡∏ô‡∏µ‡πâ)
                        target_vehicle = 'JB'
                
                target_w = LIMITS[target_vehicle]['max_w']
                target_c = LIMITS[target_vehicle]['max_c']
                split_needed = True
                
                # üéØ ‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡πÇ‡∏î‡∏¢‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ï‡∏≤‡∏°‡∏ï‡∏≥‡πÅ‡∏´‡∏ô‡πà‡∏á‡πÅ‡∏•‡∏∞‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°
                if split_needed:
                    # üìç Step 1: ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡∏≤‡∏°‡∏ï‡∏≥‡πÅ‡∏´‡∏ô‡πà‡∏á (‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ô‡∏≠‡∏¢‡∏π‡πà‡∏î‡πâ‡∏ß‡∏¢‡∏Å‡∏±‡∏ô)
                    branch_info = []
                    for _, row in trip_data.iterrows():
                        code = row['Code']
                        lat, lon = get_lat_lon(code)
                        branch_info.append({
                            'code': code,
                            'weight': row['Weight'],
                            'cube': row['Cube'],
                            'lat': lat if lat else 0,
                            'lon': lon if lon else 0
                        })
                    
                    # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏° lat, lon ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ô
                    branch_info.sort(key=lambda x: (x['lat'], x['lon']))
                    
                    # üìç Step 2: ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÇ‡∏î‡∏¢‡∏û‡∏¥‡∏à‡∏≤‡∏£‡∏ì‡∏≤‡∏ó‡∏±‡πâ‡∏á Cube ‡πÅ‡∏•‡∏∞‡∏ï‡∏≥‡πÅ‡∏´‡∏ô‡πà‡∏á
                    new_trips = []
                    current_group = []
                    current_group_w = 0
                    current_group_c = 0
                    current_centroid_lat = None
                    current_centroid_lon = None
                    
                    for branch in branch_info:
                        code = branch['code']
                        w = branch['weight']
                        c = branch['cube']
                        b_lat = branch['lat']
                        b_lon = branch['lon']
                        
                        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å centroid ‡∏Ç‡∏≠‡∏á‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏õ‡∏±‡∏à‡∏à‡∏∏‡∏ö‡∏±‡∏ô
                        if current_centroid_lat and b_lat:
                            distance_from_group = haversine_distance(current_centroid_lat, current_centroid_lon, b_lat, b_lon)
                        else:
                            distance_from_group = 0
                        
                        # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏ñ‡πâ‡∏≤‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ ‡∏à‡∏∞‡πÄ‡∏Å‡∏¥‡∏ô‡∏£‡∏ñ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢‡πÑ‡∏´‡∏°
                        test_w = current_group_w + w
                        test_c = current_group_c + c
                        test_util = max((test_w / target_w) * 100, (test_c / target_c) * 100)
                        
                        # üö® ‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç: ‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å (4W/JB) ‡∏´‡πâ‡∏≤‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 12 ‡∏™‡∏≤‡∏Ç‡∏≤, 6W ‡πÑ‡∏°‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î
                        max_branches = 12 if target_vehicle in ['4W', 'JB'] else float('inf')
                        
                        # ‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç‡πÄ‡∏û‡∏¥‡πà‡∏°: ‡∏ñ‡πâ‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏´‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏°‡∏≤‡∏Å‡πÄ‡∏Å‡∏¥‡∏ô 50km ‚Üí ‡πÅ‡∏¢‡∏Å‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÉ‡∏´‡∏°‡πà
                        too_far = distance_from_group > 50 and len(current_group) > 0
                        
                        # ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢: 95-100% ‡πÅ‡∏•‡∏∞‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤ ‡πÅ‡∏•‡∏∞‡πÑ‡∏°‡πà‡πÑ‡∏Å‡∏•‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ
                        if ((test_util <= 100 and len(current_group) < max_branches and not too_far) or 
                            len(current_group) == 0):
                            # ‡πÉ‡∏™‡πà‡πÑ‡∏î‡πâ
                            current_group.append(code)
                            current_group_w += w
                            current_group_c += c
                            # ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï centroid
                            if b_lat:
                                if current_centroid_lat is None:
                                    current_centroid_lat = b_lat
                                    current_centroid_lon = b_lon
                                else:
                                    n = len(current_group)
                                    current_centroid_lat = ((current_centroid_lat * (n-1)) + b_lat) / n
                                    current_centroid_lon = ((current_centroid_lon * (n-1)) + b_lon) / n
                        else:
                            # ‡πÄ‡∏ï‡πá‡∏°‡πÅ‡∏•‡πâ‡∏ß ‡∏´‡∏£‡∏∑‡∏≠ ‡πÑ‡∏Å‡∏•‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ ‚Üí ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÉ‡∏´‡∏°‡πà
                            current_util = max((current_group_w / target_w) * 100, (current_group_c / target_c) * 100)
                            
                            if current_util >= 95 or len(current_group) >= 12 or too_far:
                                new_trips.append({
                                    'codes': current_group.copy(),
                                    'weight': current_group_w,
                                    'cube': current_group_c
                                })
                                current_group = [code]
                                current_group_w = w
                                current_group_c = c
                                current_centroid_lat = b_lat if b_lat else None
                                current_centroid_lon = b_lon if b_lon else None
                            else:
                                # ‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏ï‡πá‡∏°‡∏û‡∏≠ ‚Üí ‡πÉ‡∏™‡πà‡∏ï‡πà‡∏≠
                                current_group.append(code)
                                current_group_w += w
                                current_group_c += c
                    
                    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢
                    if current_group:
                        new_trips.append({
                            'codes': current_group,
                            'weight': current_group_w,
                            'cube': current_group_c
                        })
                    
                    # üìç Step 3: ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏Å‡∏•‡∏∏‡πà‡∏° (‡∏≠‡∏≤‡∏à‡∏Ñ‡∏ô‡∏•‡∏∞‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó)
                    for trip_info in new_trips:
                        trip_w = trip_info['weight']
                        trip_c = trip_info['cube']
                        trip_branches = len(trip_info['codes'])
                        
                        # ‡∏•‡∏≠‡∏á 4W ‡∏Å‡πà‡∏≠‡∏ô (‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡πÅ‡∏•‡∏∞‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 12 ‡∏™‡∏≤‡∏Ç‡∏≤)
                        util_4w = max((trip_w / LIMITS['4W']['max_w']) * 100, 
                                     (trip_c / LIMITS['4W']['max_c']) * 100)
                        util_jb = max((trip_w / LIMITS['JB']['max_w']) * 100,
                                     (trip_c / LIMITS['JB']['max_c']) * 100)
                        util_6w = max((trip_w / LIMITS['6W']['max_w']) * 100,
                                     (trip_c / LIMITS['6W']['max_c']) * 100)
                        
                        # ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î (Cube 95-120%)
                        if trip_branches <= 12:
                            if 95 <= util_4w <= 100 and max_allowed != 'JB' and max_allowed != '6W':
                                trip_info['vehicle'] = '4W'
                            elif 95 <= util_jb <= 100 and max_allowed != '6W':
                                trip_info['vehicle'] = 'JB'
                            elif util_6w <= 200 and max_allowed == '6W':
                                trip_info['vehicle'] = '6W'
                            elif util_jb <= 100 and max_allowed != '6W':
                                trip_info['vehicle'] = 'JB'
                            elif util_4w <= 100 and max_allowed != 'JB' and max_allowed != '6W':
                                trip_info['vehicle'] = '4W'
                            else:
                                trip_info['vehicle'] = target_vehicle
                        else:
                            # ‡πÄ‡∏Å‡∏¥‡∏ô 12 ‡∏™‡∏≤‡∏Ç‡∏≤ ‚Üí ‡πÉ‡∏ä‡πâ 6W
                            trip_info['vehicle'] = '6W' if max_allowed == '6W' else 'JB'
                    
                    # üìç Step 4: ‡∏£‡∏ß‡∏°‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏ô‡πâ‡∏≠‡∏¢‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ (<95%) ‡∏Å‡∏±‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
                    final_trips = []
                    low_util_trips = []
                    
                    for trip_info in new_trips:
                        vehicle = trip_info.get('vehicle', target_vehicle)
                        v_w = LIMITS[vehicle]['max_w']
                        v_c = LIMITS[vehicle]['max_c']
                        trip_util = max((trip_info['weight'] / v_w) * 100, (trip_info['cube'] / v_c) * 100)
                        
                        if trip_util >= 95:
                            final_trips.append(trip_info)
                        else:
                            low_util_trips.append(trip_info)
                    
                    # ‡∏Å‡∏£‡∏∞‡∏à‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤‡∏à‡∏≤‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏ô‡πâ‡∏≠‡∏¢‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ
                    for low_trip in low_util_trips:
                        for code in low_trip['codes']:
                            branch_w = test_df[test_df['Code'] == code]['Weight'].sum()
                            branch_c = test_df[test_df['Code'] == code]['Cube'].sum()
                            branch_lat, branch_lon = get_lat_lon(code)
                            
                            best_trip_idx = -1
                            best_score = float('inf')
                            
                            for idx, trip_info in enumerate(final_trips):
                                # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ
                                trip_coords = []
                                for c in trip_info['codes']:
                                    lat, lon = get_lat_lon(c)
                                    if lat and lon:
                                        trip_coords.append((lat, lon))
                                
                                if trip_coords and branch_lat:
                                    centroid_lat = sum(c[0] for c in trip_coords) / len(trip_coords)
                                    centroid_lon = sum(c[1] for c in trip_coords) / len(trip_coords)
                                    distance = haversine_distance(branch_lat, branch_lon, centroid_lat, centroid_lon)
                                else:
                                    distance = 50
                                
                                vehicle = trip_info.get('vehicle', target_vehicle)
                                v_w = LIMITS[vehicle]['max_w']
                                v_c = LIMITS[vehicle]['max_c']
                                new_w = trip_info['weight'] + branch_w
                                new_c = trip_info['cube'] + branch_c
                                new_util = max((new_w / v_w) * 100, (new_c / v_c) * 100)
                                
                                if new_util <= 100 and len(trip_info['codes']) < 12:
                                    score = distance + (new_util - 100) * 0.5
                                    if score < best_score:
                                        best_score = score
                                        best_trip_idx = idx
                            
                            if best_trip_idx >= 0:
                                final_trips[best_trip_idx]['codes'].append(code)
                                final_trips[best_trip_idx]['weight'] += branch_w
                                final_trips[best_trip_idx]['cube'] += branch_c
                            else:
                                # ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
                                final_trips.append({
                                    'codes': [code],
                                    'weight': branch_w,
                                    'cube': branch_c,
                                    'vehicle': target_vehicle
                                })
                    
                    # üìç Step 5: ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï DataFrame ‡πÅ‡∏•‡∏∞‡∏Å‡∏≥‡∏´‡∏ô‡∏î‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°
                    if len(final_trips) >= 1:
                        for idx, trip_info in enumerate(final_trips):
                            codes = trip_info['codes']
                            vehicle = trip_info.get('vehicle', target_vehicle)
                            
                            # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤
                            group_max_allowed = get_max_vehicle_for_trip(set(codes))
                            vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
                            if vehicle_priority.get(vehicle, 3) > vehicle_priority.get(group_max_allowed, 3):
                                vehicle = group_max_allowed
                            
                            if idx == 0:
                                # ‡∏ó‡∏£‡∏¥‡∏õ‡πÅ‡∏£‡∏Å‡πÉ‡∏ä‡πâ‡πÄ‡∏•‡∏Ç‡πÄ‡∏î‡∏¥‡∏°
                                for code in codes:
                                    test_df.loc[test_df['Code'] == code, 'Trip'] = trip_num
                                trip_recommended_vehicles[trip_num] = vehicle
                            else:
                                # ‡∏ó‡∏£‡∏¥‡∏õ‡∏ñ‡∏±‡∏î‡πÑ‡∏õ‡∏™‡∏£‡πâ‡∏≤‡∏á‡πÉ‡∏´‡∏°‡πà
                                new_trip_num = test_df['Trip'].max() + 1
                                for code in codes:
                                    test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip_num
                                trip_recommended_vehicles[new_trip_num] = vehicle
                                split_count += 1
                    else:
                        # ‡πÑ‡∏°‡πà‡πÅ‡∏¢‡∏Å ‚Üí ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÄ‡∏î‡∏¥‡∏°
                        trip_recommended_vehicles[trip_num] = target_vehicle
                        fix_count += 1
    # üéØ Phase 2.5: ‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà Cube ‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ‡∏°‡∏≤‡∏Å (‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å‡πÄ‡∏ö‡∏≤ ‡πÅ‡∏ï‡πà‡πÄ‡∏ï‡πá‡∏° Cube)
    cube_split_count = 0
    next_trip_num = test_df['Trip'].max() + 1
    
    for trip_num in sorted(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
            
        trip_data = test_df[test_df['Trip'] == trip_num]
        current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
        
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì Cube utilization
        should_split = False
        target_vehicle = current_vehicle
        
        if current_vehicle == '4W':
            cube_util = (total_c / LIMITS['4W']['max_c']) * 100
            weight_util = (total_w / LIMITS['4W']['max_w']) * 100
            # 4W Cube ‡πÄ‡∏Å‡∏¥‡∏ô 100% ‚Üí ‡πÅ‡∏¢‡∏Å
            if cube_util > 100 and len(trip_data) >= 4:
                should_split = True
                target_vehicle = 'JB'
        elif current_vehicle == 'JB':
            cube_util = (total_c / LIMITS['JB']['max_c']) * 100
            weight_util = (total_w / LIMITS['JB']['max_w']) * 100
            # JB Cube ‡πÄ‡∏Å‡∏¥‡∏ô 100% ‚Üí ‡πÅ‡∏¢‡∏Å (‡πÇ‡∏î‡∏¢‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏ó‡∏µ‡πà‡∏´‡πâ‡∏≤‡∏°‡πÉ‡∏ä‡πâ 6W)
            if cube_util > 100 and len(trip_data) >= 4:
                should_split = True
                target_vehicle = 'JB'  # ‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô JB ‡∏≠‡∏µ‡∏Å‡∏Ñ‡∏±‡∏ô
        elif current_vehicle == '6W':
            # üöõ 6W ‡πÑ‡∏°‡πà‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏Ñ‡∏¥‡∏ß - ‡πÉ‡∏™‡πà‡πÑ‡∏î‡πâ‡πÄ‡∏ï‡πá‡∏°‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å
            should_split = False
        
        if should_split:
            # ‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ï‡∏≤‡∏° Cube (‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏° Cube ‡∏à‡∏≤‡∏Å‡∏°‡∏≤‡∏Å‡πÑ‡∏õ‡∏ô‡πâ‡∏≠‡∏¢ ‡πÅ‡∏•‡πâ‡∏ß‡πÅ‡∏ö‡πà‡∏á‡∏Ñ‡∏£‡∏∂‡πà‡∏á)
            trip_data_sorted = trip_data.sort_values('Cube', ascending=False)
            codes = list(trip_data_sorted['Code'].values)
            
            # ‡πÅ‡∏ö‡πà‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏õ‡πá‡∏ô 2 ‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÉ‡∏´‡πâ Cube ‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á‡∏Å‡∏±‡∏ô
            g1_codes, g2_codes = [], []
            g1_cube, g2_cube = 0, 0
            
            for code in codes:
                branch_cube = trip_data_sorted[trip_data_sorted['Code'] == code]['Cube'].sum()
                if g1_cube <= g2_cube:
                    g1_codes.append(code)
                    g1_cube += branch_cube
                else:
                    g2_codes.append(code)
                    g2_cube += branch_cube
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏û‡∏≠‡∏î‡∏µ‡∏Å‡∏±‡∏ö‡∏£‡∏ñ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
            g1_w = trip_data[trip_data['Code'].isin(g1_codes)]['Weight'].sum()
            g1_c = trip_data[trip_data['Code'].isin(g1_codes)]['Cube'].sum()
            g2_w = trip_data[trip_data['Code'].isin(g2_codes)]['Weight'].sum()
            g2_c = trip_data[trip_data['Code'].isin(g2_codes)]['Cube'].sum()
            
            g1_cube_util = (g1_c / LIMITS[target_vehicle]['max_c']) * 100
            g1_weight_util = (g1_w / LIMITS[target_vehicle]['max_w']) * 100
            g2_cube_util = (g2_c / LIMITS[target_vehicle]['max_c']) * 100
            g2_weight_util = (g2_w / LIMITS[target_vehicle]['max_w']) * 100
            
            # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏ó‡∏±‡πâ‡∏á 2 ‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢‡πÑ‡∏î‡πâ‡πÅ‡∏•‡∏∞‡∏°‡∏µ‡∏õ‡∏£‡∏∞‡∏™‡∏¥‡∏ó‡∏ò‡∏¥‡∏†‡∏≤‡∏û (Cube ‚â•50%, ‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å ‚â§100%)
            g1_ok = g1_cube_util <= 100 and g1_weight_util <= 100 and g1_cube_util >= 50
            g2_ok = g2_cube_util <= 100 and g2_weight_util <= 100 and g2_cube_util >= 50
            
            # üö® ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏ñ‡πâ‡∏≤‡πÅ‡∏¢‡∏Å‡πÅ‡∏•‡πâ‡∏ß‡∏£‡∏ñ‡πÉ‡∏´‡∏°‡πà‡πÑ‡∏°‡πà‡πÄ‡∏ï‡πá‡∏° ‚Üí ‡πÑ‡∏°‡πà‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å ‡πÉ‡∏´‡πâ‡∏¢‡∏±‡∏î‡πÉ‡∏™‡πà‡∏£‡∏ñ‡πÄ‡∏î‡∏¥‡∏°‡πÅ‡∏°‡πâ‡πÄ‡∏Å‡∏¥‡∏ô
            if not (g1_ok and g2_ok):
                # ‡∏ñ‡πâ‡∏≤‡πÅ‡∏¢‡∏Å‡πÅ‡∏•‡πâ‡∏ß‡∏£‡∏ñ‡πÉ‡∏î‡∏£‡∏ñ‡∏´‡∏ô‡∏∂‡πà‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏ï‡πá‡∏° (Cube <100%) ‚Üí ‡πÑ‡∏°‡πà‡πÅ‡∏¢‡∏Å
                # ‡∏¢‡∏≠‡∏°‡πÉ‡∏´‡πâ‡∏£‡∏ñ‡πÄ‡∏î‡∏¥‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 120% ‡πÑ‡∏î‡πâ ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÑ‡∏°‡πà‡πÉ‡∏´‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏°‡πà‡∏ß‡∏¥‡πà‡∏á‡πÑ‡∏°‡πà‡∏Ñ‡∏∏‡πâ‡∏°
                should_split = False
            
            if should_split and g1_ok and g2_ok and len(g1_codes) >= 2 and len(g2_codes) >= 2:
                # ‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ: ‡πÄ‡∏Å‡πá‡∏ö trip_num ‡πÄ‡∏î‡∏¥‡∏° ‡πÉ‡∏´‡πâ g1, ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà‡πÉ‡∏´‡πâ g2
                for code in g2_codes:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = next_trip_num
                
                # ‡∏ö‡∏±‡∏ô‡∏ó‡∏∂‡∏Å‡∏£‡∏ñ‡∏ó‡∏±‡πâ‡∏á 2 ‡∏ó‡∏£‡∏¥‡∏õ
                trip_recommended_vehicles[trip_num] = target_vehicle
                trip_recommended_vehicles[next_trip_num] = target_vehicle
                
                next_trip_num += 1
                cube_split_count += 1
    
    # üéØ Phase 3: ‡∏õ‡∏£‡∏±‡∏ö‡∏õ‡∏£‡∏∏‡∏á 6W ‡πÉ‡∏´‡πâ‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏™‡∏°
    # - 6W ‚â•200% Cube ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å (‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ‡∏°‡∏≤‡∏Å)
    # - 6W 150-199% Cube ‚Üí ‡∏û‡∏¥‡∏à‡∏≤‡∏£‡∏ì‡∏≤‡πÅ‡∏¢‡∏Å (‡∏ñ‡πâ‡∏≤‡∏ó‡∏≥‡πÑ‡∏î‡πâ)
    # - 6W <150% Cube ‚Üí ‡πÑ‡∏°‡πà‡πÅ‡∏¢‡∏Å (‡πÉ‡∏ä‡πâ 6W ‡∏Ñ‡∏∏‡πâ‡∏°‡∏Ñ‡πà‡∏≤)
    split_count = 0
    
    # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ 6W ‡πÅ‡∏•‡∏∞ Cube ‚â•150%
    trips_to_check = []
    for trip_num in test_df['Trip'].unique():
        if trip_num == 0:
            continue
            
        trip_data = test_df[test_df['Trip'] == trip_num]
        current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
        
        if current_vehicle != '6W':
            continue
        
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì Cube utilization
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        cube_util = (total_c / LIMITS['6W']['max_c']) * 100
        
        # üöõ 6W ‚â•100% ‚Üí ‡∏û‡∏¥‡∏à‡∏≤‡∏£‡∏ì‡∏≤‡πÅ‡∏¢‡∏Å (‚â•200% ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÅ‡∏¢‡∏Å)
        if cube_util >= 100 and len(trip_data) >= 4:
            trips_to_check.append({
                'trip': trip_num,
                'data': trip_data,
                'cube_util': cube_util,
                'total_w': total_w,
                'total_c': total_c,
                'force_split': cube_util >= 200  # ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÅ‡∏¢‡∏Å‡∏ñ‡πâ‡∏≤ ‚â•200%
            })
    
    # ‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏°‡∏µ Cube ‚â•150% (‚â•200% ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÅ‡∏¢‡∏Å)
    for trip_info in trips_to_check:
        trip_num = trip_info['trip']
        trip_data = trip_info['data']
        trip_codes = list(trip_data['Code'].values)
        
        # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤
        max_allowed = get_max_vehicle_for_trip(set(trip_codes))
        if max_allowed == '6W':
            # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤ ‚Üí ‡∏•‡∏≠‡∏á‡πÅ‡∏¢‡∏Å‡πÄ‡∏õ‡πá‡∏ô‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å
            
            # ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå spatial clusters
            clusters = create_distance_based_clusters(trip_codes, max_distance_km=25)
            
            # ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ ‚â•2 ‡∏Å‡∏•‡∏∏‡πà‡∏° ‚Üí ‡∏•‡∏≠‡∏á‡πÅ‡∏¢‡∏Å
            if len(clusters) >= 2:
                # ‡∏à‡∏±‡∏î‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ï‡∏≤‡∏°‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å/‡∏Ñ‡∏¥‡∏ß
                cluster_info = []
                for cluster_codes in clusters:
                    cluster_data = trip_data[trip_data['Code'].isin(cluster_codes)]
                    cluster_w = cluster_data['Weight'].sum()
                    cluster_c = cluster_data['Cube'].sum()
                    cluster_info.append({
                        'codes': cluster_codes,
                        'weight': cluster_w,
                        'cube': cluster_c,
                        'branches': len(cluster_codes)
                    })
                
                # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏°‡∏Ñ‡∏¥‡∏ß (‡∏°‡∏≤‡∏Å‚Üí‡∏ô‡πâ‡∏≠‡∏¢)
                cluster_info.sort(key=lambda x: x['cube'], reverse=True)
                
                # ‡∏•‡∏≠‡∏á‡∏à‡∏±‡∏ö‡∏Ñ‡∏π‡πà‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏£‡∏ñ‡πÉ‡∏´‡∏°‡πà
                new_trips = []
                used_clusters = set()
                
                for i, cluster in enumerate(cluster_info):
                    if i in used_clusters:
                        continue
                    
                    # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ô‡∏µ‡πâ‡∏û‡∏≠‡∏î‡∏µ JB ‡∏´‡∏£‡∏∑‡∏≠ 4W ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
                    util_jb = max((cluster['weight'] / LIMITS['JB']['max_w']) * 100,
                                 (cluster['cube'] / LIMITS['JB']['max_c']) * 100)
                    util_4w = max((cluster['weight'] / LIMITS['4W']['max_w']) * 100,
                                 (cluster['cube'] / LIMITS['4W']['max_c']) * 100)
                    
                    # ‡∏ñ‡πâ‡∏≤‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ô‡∏µ‡πâ‡∏°‡∏µ‡∏™‡∏≤‡∏Ç‡∏≤ ‚â§12 ‡πÅ‡∏•‡∏∞‡∏û‡∏≠‡∏î‡∏µ JB ‡∏´‡∏£‡∏∑‡∏≠ 4W
                    if cluster['branches'] <= 12:
                        if util_4w >= 90 and util_4w <= 100:
                            # ‡∏û‡∏≠‡∏î‡∏µ 4W
                            new_trips.append({
                                'codes': cluster['codes'],
                                'vehicle': '4W'
                            })
                            used_clusters.add(i)
                        elif util_jb >= 90 and util_jb <= 100:
                            # ‡∏û‡∏≠‡∏î‡∏µ JB
                            new_trips.append({
                                'codes': cluster['codes'],
                                'vehicle': 'JB'
                            })
                            used_clusters.add(i)
                        else:
                            # ‡∏•‡∏≠‡∏á‡∏£‡∏ß‡∏°‡∏Å‡∏±‡∏ö‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏≠‡∏∑‡πà‡∏ô
                            for j, other_cluster in enumerate(cluster_info):
                                if j <= i or j in used_clusters:
                                    continue
                                
                                combined_codes = cluster['codes'] + other_cluster['codes']
                                combined_w = cluster['weight'] + other_cluster['weight']
                                combined_c = cluster['cube'] + other_cluster['cube']
                                combined_branches = cluster['branches'] + other_cluster['branches']
                                
                                if combined_branches <= 12:
                                    combined_util_jb = max((combined_w / LIMITS['JB']['max_w']) * 100,
                                                          (combined_c / LIMITS['JB']['max_c']) * 100)
                                    combined_util_4w = max((combined_w / LIMITS['4W']['max_w']) * 100,
                                                          (combined_c / LIMITS['4W']['max_c']) * 100)
                                    
                                    if combined_util_4w >= 90 and combined_util_4w <= 100:
                                        new_trips.append({
                                            'codes': combined_codes,
                                            'vehicle': '4W'
                                        })
                                        used_clusters.add(i)
                                        used_clusters.add(j)
                                        break
                                    elif combined_util_jb >= 90 and combined_util_jb <= 100:
                                        new_trips.append({
                                            'codes': combined_codes,
                                            'vehicle': 'JB'
                                        })
                                        used_clusters.add(i)
                                        used_clusters.add(j)
                                        break
                
                # üö® ‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç‡∏Å‡∏≤‡∏£‡πÅ‡∏¢‡∏Å:
                # - ‡∏ñ‡πâ‡∏≤ force_split = True (‚â•200%) ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÅ‡∏¢‡∏Å‡πÄ‡∏™‡∏°‡∏≠
                # - ‡∏ñ‡πâ‡∏≤ 150-199% ‚Üí ‡πÅ‡∏¢‡∏Å‡∏ñ‡πâ‡∏≤‡πÑ‡∏î‡πâ‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏ô‡πâ‡∏≠‡∏¢ 2 ‡∏ó‡∏£‡∏¥‡∏õ‡πÅ‡∏•‡∏∞‡πÉ‡∏ä‡πâ‡∏£‡∏ñ ‚â•90%
                force_split = trip_info.get('force_split', False)
                should_split = force_split or len(new_trips) >= 2
                
                if should_split:
                    # ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
                    max_trip = test_df['Trip'].max()
                    
                    # ‡∏ñ‡πâ‡∏≤‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÅ‡∏¢‡∏Å‡πÅ‡∏ï‡πà‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡∏°‡∏µ new_trips ‚Üí ‡πÅ‡∏ö‡πà‡∏á‡∏Ñ‡∏£‡∏∂‡πà‡∏á
                    if force_split and len(new_trips) < 2:
                        # ‡πÅ‡∏ö‡πà‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡∏Ñ‡∏£‡∏∂‡πà‡∏á‡∏´‡∏ô‡∏∂‡πà‡∏á‡∏ï‡∏≤‡∏° Cube
                        sorted_data = trip_data.sort_values('Cube', ascending=False)
                        mid = len(sorted_data) // 2
                        g1_codes = list(sorted_data.iloc[:mid]['Code'].values)
                        g2_codes = list(sorted_data.iloc[mid:]['Code'].values)
                        
                        # ‡∏Å‡∏≥‡∏´‡∏ô‡∏î‡∏£‡∏ñ‡πÉ‡∏´‡πâ‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏Å‡∏•‡∏∏‡πà‡∏° (‡πÉ‡∏ä‡πâ 6W ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏£‡∏≠‡∏á‡∏£‡∏±‡∏ö Cube ‡∏™‡∏π‡∏á)
                        new_trips = [
                            {'codes': g1_codes, 'vehicle': '6W'},
                            {'codes': g2_codes, 'vehicle': '6W'}
                        ]
                    
                    for idx, new_trip_info in enumerate(new_trips):
                        if idx == 0:
                            # ‡∏ó‡∏£‡∏¥‡∏õ‡πÅ‡∏£‡∏Å‡πÉ‡∏ä‡πâ‡πÄ‡∏•‡∏Ç‡πÄ‡∏î‡∏¥‡∏°
                            for code in new_trip_info['codes']:
                                test_df.loc[test_df['Code'] == code, 'Trip'] = trip_num
                            trip_recommended_vehicles[trip_num] = new_trip_info['vehicle']
                        else:
                            # ‡∏ó‡∏£‡∏¥‡∏õ‡∏ñ‡∏±‡∏î‡πÑ‡∏õ‡∏™‡∏£‡πâ‡∏≤‡∏á‡πÉ‡∏´‡∏°‡πà
                            new_trip_num = max_trip + idx
                            for code in new_trip_info['codes']:
                                test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip_num
                            trip_recommended_vehicles[new_trip_num] = new_trip_info['vehicle']
                            split_count += 1
    
    # üîÑ Phase 4: ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô nearby ‡∏à‡∏≤‡∏Å 6W ‚Üí JB/4W ‡πÅ‡∏•‡∏∞‡∏Å‡∏£‡∏∞‡∏à‡∏≤‡∏¢‡∏ó‡∏£‡∏¥‡∏õ‡∏ô‡πâ‡∏≠‡∏¢ (‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏ó‡∏µ‡πà‡∏à‡∏≥‡πÄ‡∏õ‡πá‡∏ô) - Optimized
    low_util_trips = []
    
    # ‚ö° Skip ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πâ‡πÄ‡∏ß‡∏•‡∏≤‡∏°‡∏≤‡∏Å‡∏Å‡∏ß‡πà‡∏≤ 22 ‡∏ß‡∏¥‡∏ô‡∏≤‡∏ó‡∏µ
    if time.time() - start_time > 22:
        pass  # Skip Phase 4 ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏£‡πá‡∏ß
    else:
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            trip_codes = set(trip_data['Code'].values)
            current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
            total_w = trip_data['Weight'].sum()
            total_c = trip_data['Cube'].sum()
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡πÄ‡∏õ‡πá‡∏ô nearby ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
            provinces = set()
            for code in trip_codes:
                prov = get_province(code)
            if prov != 'UNKNOWN':
                provinces.add(prov)
        
        all_nearby = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
        
        # ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πâ 6W ‡πÅ‡∏•‡∏∞‡πÄ‡∏õ‡πá‡∏ô nearby ‚Üí ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡πÄ‡∏õ‡πá‡∏ô JB
        if current_vehicle == '6W' and all_nearby:
            jb_util = max((total_w / LIMITS['JB']['max_w']) * 100, 
                         (total_c / LIMITS['JB']['max_c']) * 100)
            if jb_util <= 100:
                trip_recommended_vehicles[trip_num] = 'JB'
                current_vehicle = 'JB'
            else:
                trip_recommended_vehicles[trip_num] = 'JB'
                current_vehicle = 'JB'
        
        # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏ô‡πâ‡∏≠‡∏¢‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ (<65% ‡πÅ‡∏•‡∏∞ ‚â§ 2 ‡∏™‡∏≤‡∏Ç‡∏≤ - ‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏ó‡∏µ‡πà‡∏à‡∏≥‡πÄ‡∏õ‡πá‡∏ô‡∏à‡∏£‡∏¥‡∏á‡πÜ)
        util = max((total_w / LIMITS[current_vehicle]['max_w']) * 100,
                   (total_c / LIMITS[current_vehicle]['max_c']) * 100)
        
        if util < 65 and len(trip_data) <= 2:
            low_util_trips.append({
                'trip_num': trip_num,
                'codes': list(trip_codes),
                'weight': total_w,
                'cube': total_c,
                'vehicle': current_vehicle
            })
    
    # ‡∏Å‡∏£‡∏∞‡∏à‡∏≤‡∏¢‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏ô‡πâ‡∏≠‡∏¢‡πÄ‡∏Å‡∏¥‡∏ô‡πÑ‡∏õ (Skip ‡∏ñ‡πâ‡∏≤‡∏°‡∏µ‡∏°‡∏≤‡∏Å - ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏£‡πá‡∏ß)
    if len(low_util_trips) > 15:
        low_util_trips = []  # Skip ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏£‡πá‡∏ß
    
    # ‡∏Å‡∏£‡∏∞‡∏à‡∏≤‡∏¢‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏ó‡∏µ‡πà‡∏à‡∏≥‡πÄ‡∏õ‡πá‡∏ô
    if len(low_util_trips) == 0:
        pass  # Skip ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ó‡∏£‡∏¥‡∏õ‡∏ô‡πâ‡∏≠‡∏¢
    else:
        for low_trip in low_util_trips:
            # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡πÅ‡∏•‡∏∞‡∏£‡∏±‡∏ö‡πÑ‡∏î‡πâ
            best_target_trip = None
            best_score = float('inf')
            
            for target_trip_num in test_df['Trip'].unique():
                if target_trip_num == 0 or target_trip_num == low_trip['trip_num']:
                    continue
                
                target_data = test_df[test_df['Trip'] == target_trip_num]
                target_vehicle = trip_recommended_vehicles.get(target_trip_num, '6W')
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏£‡∏ß‡∏°‡πÑ‡∏î‡πâ‡πÑ‡∏´‡∏°
            pass
            new_w = target_data['Weight'].sum() + low_trip['weight']
            new_c = target_data['Cube'].sum() + low_trip['cube']
            new_util = max((new_w / LIMITS[target_vehicle]['max_w']) * 100,
                          (new_c / LIMITS[target_vehicle]['max_c']) * 100)
            
            max_branches = 12 if target_vehicle in ['4W', 'JB'] else float('inf')
            
            if new_util <= 100 and len(target_data) + len(low_trip['codes']) <= max_branches:
                # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á (‡πÄ‡∏â‡∏•‡∏µ‡πà‡∏¢)
                score = new_util
                if score < best_score:
                    best_score = score
                    best_target_trip = target_trip_num
        
        # ‡∏¢‡πâ‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤
        if best_target_trip:
            for code in low_trip['codes']:
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_target_trip
                # ‡∏•‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏î‡∏¥‡∏°
                if low_trip['trip_num'] in trip_recommended_vehicles:
                    del trip_recommended_vehicles[low_trip['trip_num']]
    
    # üö® Phase 5: Distance Optimization - ‡∏™‡∏•‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏´‡πâ‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ô‡∏°‡∏≤‡∏Å‡∏Ç‡∏∂‡πâ‡∏ô (FAST)
    # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô MAX_DISTANCE_IN_TRIP ‚Üí ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏ß‡πà‡∏≤‡∏à‡∏≤‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô‡∏°‡∏≤‡∏™‡∏•‡∏±‡∏ö
    # ‚ö° Skip ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πâ‡πÄ‡∏ß‡∏•‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 25 ‡∏ß‡∏¥‡∏ô‡∏≤‡∏ó‡∏µ
    if time.time() - start_time > 25:
        distance_swaps = 999  # Skip Phase 5
    else:
        distance_swaps = 0
    max_distance_swaps = 30  # ‡∏•‡∏î‡∏à‡∏≤‡∏Å 100 ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏£‡πá‡∏ß
    
    # üîí ‡πÄ‡∏Å‡πá‡∏ö ‡∏ï‡∏≥‡∏ö‡∏•/‡∏≠‡∏≥‡πÄ‡∏†‡∏≠/‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î ‡∏Ç‡∏≠‡∏á‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏™‡∏≤‡∏Ç‡∏≤
    def get_location_for_code(code):
        """‡∏Ñ‡∏∑‡∏ô‡∏Ñ‡πà‡∏≤ (‡∏ï‡∏≥‡∏ö‡∏•, ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠, ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î) ‡∏Ç‡∏≠‡∏á‡∏™‡∏≤‡∏Ç‡∏≤"""
        if not MASTER_DATA.empty:
            master = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
            if len(master) > 0:
                row = master.iloc[0]
                return (
                    row.get('‡∏ï‡∏≥‡∏ö‡∏•', ''),
                    row.get('‡∏≠‡∏≥‡πÄ‡∏†‡∏≠', ''),
                    row.get('‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', '')
                )
        return ('', '', '')
    
    def get_province_for_code(code):
        return get_location_for_code(code)[2]
    
    def calculate_location_bonus(code1_loc, other_codes):
        """
        ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì bonus ‡∏ï‡∏≤‡∏°‡∏Ñ‡∏ß‡∏≤‡∏°‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á‡∏Ç‡∏≠‡∏á‡∏ï‡∏≥‡∏ö‡∏•/‡∏≠‡∏≥‡πÄ‡∏†‡∏≠/‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î
        - ‡∏ï‡∏≥‡∏ö‡∏•‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô: +15km bonus
        - ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô: +10km bonus  
        - ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô: +5km bonus
        """
        subdistrict1, district1, province1 = code1_loc
        best_bonus = 0
        
        for code in other_codes:
            subdistrict2, district2, province2 = get_location_for_code(code)
            
            # ‡∏ï‡∏≥‡∏ö‡∏•‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô + ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô = ‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î
            if subdistrict1 and subdistrict1 == subdistrict2 and province1 == province2:
                return 15  # Bonus ‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î
            
            # ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô + ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
            if district1 and district1 == district2 and province1 == province2:
                best_bonus = max(best_bonus, 10)
            
            # ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
            elif province1 and province1 == province2:
                best_bonus = max(best_bonus, 5)
        
        return best_bonus
    
    for iteration in range(2):  # ‚ö° ‡∏•‡∏î‡πÄ‡∏õ‡πá‡∏ô 2 ‡∏£‡∏≠‡∏ö‡πÄ‡∏û‡∏∑‡πà‡∏≠‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏£‡πá‡∏ß
        if distance_swaps >= max_distance_swaps or time.time() - start_time > 28:
            break
            
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0 or distance_swaps >= max_distance_swaps:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            trip_codes = list(trip_data['Code'].values)
            
            if len(trip_codes) < 2:
                continue
            
            # ‡∏´‡∏≤ centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ
            trip_lats, trip_lons = [], []
            for code in trip_codes:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if not trip_lats:
                continue
                
            centroid_lat = sum(trip_lats) / len(trip_lats)
            centroid_lon = sum(trip_lons) / len(trip_lons)
            
            # ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÑ‡∏Å‡∏•‡∏à‡∏≤‡∏Å centroid ‡∏°‡∏≤‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
            farthest_code = None
            farthest_dist = 0
            
            for code in trip_codes:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    dist = haversine_distance(centroid_lat, centroid_lon, lat, lon)
                    if dist > farthest_dist:
                        farthest_dist = dist
                        farthest_code = code
            
            # ‡∏ñ‡πâ‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏Å‡∏•‡πÄ‡∏Å‡∏¥‡∏ô 40km ‡∏à‡∏≤‡∏Å centroid ‚Üí ‡∏•‡∏≠‡∏á‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏°‡∏≤‡∏∞‡∏Å‡∏ß‡πà‡∏≤ (‡∏•‡∏î‡∏à‡∏≤‡∏Å 50km)
            if farthest_dist > 40 and farthest_code:
                far_lat, far_lon = coord_cache.get(farthest_code, (None, None))
                if not far_lat:
                    continue
                
                far_branch_data = test_df[test_df['Code'] == farthest_code].iloc[0]
                far_weight = far_branch_data['Weight']
                far_cube = far_branch_data['Cube']
                far_province = get_province_for_code(farthest_code)
                
                # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏ß‡πà‡∏≤‡πÅ‡∏•‡∏∞‡∏¢‡∏±‡∏á‡πÉ‡∏™‡πà‡πÑ‡∏î‡πâ
                best_new_trip = None
                best_new_dist = farthest_dist
                best_same_province = False
                
                for other_trip in test_df['Trip'].unique():
                    if other_trip == 0 or other_trip == trip_num:
                        continue
                    
                    other_data = test_df[test_df['Trip'] == other_trip]
                    other_codes = list(other_data['Code'].values)
                    
                    # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô
                    other_provinces = set()
                    for code in other_codes:
                        prov = get_province_for_code(code)
                        if prov:
                            other_provinces.add(prov)
                    
                    # ‡πÉ‡∏´‡πâ priority ‡∏Å‡∏±‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
                    same_province = far_province in other_provinces
                    
                    # ‡∏´‡∏≤ centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô
                    other_lats, other_lons = [], []
                    for code in other_codes:
                        lat, lon = coord_cache.get(code, (None, None))
                        if lat and lon:
                            other_lats.append(lat)
                            other_lons.append(lon)
                    
                    if not other_lats:
                        continue
                    
                    other_centroid_lat = sum(other_lats) / len(other_lats)
                    other_centroid_lon = sum(other_lons) / len(other_lons)
                    
                    # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏Å‡∏•‡πÑ‡∏õ‡∏¢‡∏±‡∏á centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô
                    dist_to_other = haversine_distance(far_lat, far_lon, other_centroid_lat, other_centroid_lon)
                    
                    # üîí ‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏£‡∏¥‡∏á + bonus ‡∏ï‡∏≤‡∏°‡∏ï‡∏≥‡∏ö‡∏•/‡∏≠‡∏≥‡πÄ‡∏†‡∏≠/‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î
                    # ‡∏ï‡∏≥‡∏ö‡∏•‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô: +15km, ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô: +10km, ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô: +5km
                    far_location = get_location_for_code(farthest_code)
                    location_bonus = calculate_location_bonus(far_location, other_codes)
                    effective_dist = dist_to_other - location_bonus
                    
                    # ‡∏ï‡πâ‡∏≠‡∏á‡∏î‡∏µ‡∏Ç‡∏∂‡πâ‡∏ô‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏ô‡πâ‡∏≠‡∏¢ 10km (‡∏´‡∏•‡∏±‡∏á‡∏´‡∏±‡∏Å bonus)
                    if effective_dist < best_new_dist - 10:
                        other_vehicle = trip_recommended_vehicles.get(other_trip, '4W')
                        other_total_w = other_data['Weight'].sum() + far_weight
                        other_total_c = other_data['Cube'].sum() + far_cube
                        
                        other_util = max(
                            (other_total_w / LIMITS[other_vehicle]['max_w']) * 100,
                            (other_total_c / LIMITS[other_vehicle]['max_c']) * 100
                        )
                        
                        max_branches = 12 if other_vehicle in ['4W', 'JB'] else float('inf')
                        
                        if other_util <= 100 and len(other_codes) < max_branches:
                            best_new_trip = other_trip
                            best_new_dist = effective_dist
                
                # ‡∏¢‡πâ‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏õ‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
                if best_new_trip is not None:
                    test_df.loc[test_df['Code'] == farthest_code, 'Trip'] = best_new_trip
                    distance_swaps += 1
    
    # üó∫Ô∏è ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡∏≤‡∏° Nearest Neighbor (‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏ç‡πà ‚â• 8 ‡∏™‡∏≤‡∏Ç‡∏≤ - ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏ß‡∏≤‡∏°‡πÄ‡∏£‡πá‡∏ß)
    # ‚ö° Skip ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πâ‡πÄ‡∏ß‡∏•‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 28 ‡∏ß‡∏¥‡∏ô‡∏≤‡∏ó‡∏µ
    if time.time() - start_time <= 28:
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0:
                continue
            
            trip_codes = list(test_df[test_df['Trip'] == trip_num]['Code'].values)
            if len(trip_codes) < 8:  # ‚ö° Skip ‡∏ñ‡πâ‡∏≤‡∏ô‡πâ‡∏≠‡∏¢‡∏Å‡∏ß‡πà‡∏≤ 8 ‡∏™‡∏≤‡∏Ç‡∏≤ (‡πÄ‡∏î‡∏¥‡∏° 6)
                continue
            
            # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏° Nearest Neighbor ‡πÅ‡∏ö‡∏ö‡πÄ‡∏£‡πá‡∏ß (‡πÉ‡∏ä‡πâ cache)
            ordered = []
            remaining = trip_codes.copy()
            current_lat, current_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
            
            while remaining and len(ordered) < len(trip_codes):
                nearest = None
                min_dist = float('inf')
                
                for code in remaining:
                    lat, lon = coord_cache.get(code, (None, None))
                    if lat:
                        dist = haversine_distance(current_lat, current_lon, lat, lon)
                        if dist < min_dist:
                            min_dist = dist
                            nearest = code
                
                if nearest:
                    ordered.append(nearest)
                    remaining.remove(nearest)
                    lat, lon = coord_cache.get(nearest, (None, None))
                    if lat:
                        current_lat, current_lon = lat, lon
                else:
                    ordered.extend(remaining)
                    break
            
            # ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï Sequence
            for seq, code in enumerate(ordered, start=1):
                test_df.loc[(test_df['Code'] == code) & (test_df['Trip'] == trip_num), 'Sequence'] = seq
    
    # ===============================================
    # üéØ Phase 6: Capacity Balancing - ‡∏Å‡∏£‡∏∞‡∏à‡∏≤‡∏¢ load ‡πÉ‡∏´‡πâ‡πÄ‡∏ó‡πà‡∏≤‡∏Å‡∏±‡∏ô
    # MIN_UTIL: 4W ‚â• 70%, JB ‚â• 80%, 6W ‚â• 90%
    # ===============================================
    balance_count = 0
    MAX_BALANCE_ITERATIONS = 3
    
    for balance_iter in range(MAX_BALANCE_ITERATIONS):
        if time.time() - start_time > 50:  # ‡∏ñ‡πâ‡∏≤‡πÉ‡∏ä‡πâ‡πÄ‡∏ß‡∏•‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 50 ‡∏ß‡∏¥‡∏ô‡∏≤‡∏ó‡∏µ ‚Üí ‡∏´‡∏¢‡∏∏‡∏î
            break
            
        # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà util ‡∏ï‡πà‡∏≥‡∏Å‡∏ß‡πà‡∏≤ MIN_UTIL
        low_util_trips_balance = []
        high_util_trips_balance = []
        
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            trip_count = len(trip_data)
            trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì util ‡∏ï‡∏≤‡∏°‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ
            limits = LIMITS.get(trip_vehicle, LIMITS['4W'])
            trip_util = max(
                (trip_w / limits['max_w']) * 100,
                (trip_c / limits['max_c']) * 100
            )
            
            min_util = MIN_UTIL.get(trip_vehicle, 70)
            
            if trip_util < min_util and trip_count <= 3:
                # ‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà util ‡∏ï‡πà‡∏≥ ‚Üí ‡∏ï‡πâ‡∏≠‡∏á‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏û‡∏¥‡πà‡∏° ‡∏´‡∏£‡∏∑‡∏≠‡∏¢‡πâ‡∏≤‡∏¢‡πÑ‡∏õ‡∏£‡∏ß‡∏°‡∏Å‡∏±‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô
                low_util_trips_balance.append({
                    'trip_num': trip_num,
                    'util': trip_util,
                    'count': trip_count,
                    'weight': trip_w,
                    'cube': trip_c,
                    'vehicle': trip_vehicle,
                    'codes': set(trip_data['Code'].values),
                    'min_util': min_util
                })
            elif trip_util > 95 and trip_count >= 3:
                # ‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÄ‡∏Å‡∏∑‡∏≠‡∏ö‡πÄ‡∏ï‡πá‡∏° ‚Üí ‡∏≠‡∏≤‡∏à‡∏Å‡∏£‡∏∞‡∏à‡∏≤‡∏¢‡πÉ‡∏´‡πâ‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡∏∑‡πà‡∏ô‡πÑ‡∏î‡πâ
                high_util_trips_balance.append({
                    'trip_num': trip_num,
                    'util': trip_util,
                    'count': trip_count,
                    'weight': trip_w,
                    'cube': trip_c,
                    'vehicle': trip_vehicle,
                    'codes': list(trip_data['Code'].values)
                })
        
        if not low_util_trips_balance:
            break  # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡∏ï‡πâ‡∏≠‡∏á balance
        
        # ‡∏û‡∏¢‡∏≤‡∏¢‡∏≤‡∏°‡∏¢‡πâ‡∏≤‡∏¢‡∏ó‡∏£‡∏¥‡∏õ util ‡∏ï‡πà‡∏≥‡πÑ‡∏õ‡∏£‡∏ß‡∏°‡∏Å‡∏±‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á
        for low_trip in low_util_trips_balance:
            best_merge_trip = None
            best_merge_util = float('inf')
            best_merge_dist = float('inf')
            
            # ‡∏´‡∏≤ centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ util ‡∏ï‡πà‡∏≥
            low_lats, low_lons = [], []
            for code in low_trip['codes']:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    low_lats.append(lat)
                    low_lons.append(lon)
            
            if not low_lats:
                continue
            
            low_centroid_lat = sum(low_lats) / len(low_lats)
            low_centroid_lon = sum(low_lons) / len(low_lons)
            
            for trip_num in test_df['Trip'].unique():
                if trip_num == 0 or trip_num == low_trip['trip_num']:
                    continue
                
                target_data = test_df[test_df['Trip'] == trip_num]
                target_w = target_data['Weight'].sum()
                target_c = target_data['Cube'].sum()
                target_count = len(target_data)
                target_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
                
                # üö® ‡πÄ‡∏ä‡πá‡∏Ñ‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á centroid ‡∏Ç‡∏≠‡∏á target
                target_lats, target_lons = [], []
                for code in target_data['Code'].values:
                    lat, lon = coord_cache.get(code, (None, None))
                    if lat and lon:
                        target_lats.append(lat)
                        target_lons.append(lon)
                
                if not target_lats:
                    continue
                
                target_centroid_lat = sum(target_lats) / len(target_lats)
                target_centroid_lon = sum(target_lons) / len(target_lons)
                
                # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏£‡∏∞‡∏´‡∏ß‡πà‡∏≤‡∏á centroids
                centroid_dist = haversine_distance(low_centroid_lat, low_centroid_lon,
                                                   target_centroid_lat, target_centroid_lon)
                
                # üîí ‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á‡∏ï‡πâ‡∏≠‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô MAX_DISTANCE_IN_TRIP (50km)
                if centroid_dist > MAX_DISTANCE_IN_TRIP:
                    continue
                
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏£‡∏ß‡∏°‡πÑ‡∏î‡πâ‡πÑ‡∏´‡∏°
                new_w = target_w + low_trip['weight']
                new_c = target_c + low_trip['cube']
                new_count = target_count + low_trip['count']
                
                # ‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 12 ‡∏™‡∏≤‡∏Ç‡∏≤ (‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö 4W/JB)
                max_branches = 12 if target_vehicle in ['JB'] else 20
                if new_count > max_branches:
                    continue
                
                # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì util ‡πÉ‡∏´‡∏°‡πà
                limits = LIMITS.get(target_vehicle, LIMITS['4W'])
                new_util = max(
                    (new_w / limits['max_w']) * 100,
                    (new_c / limits['max_c']) * 100
                )
                
                # ‡∏ï‡πâ‡∏≠‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 100%
                if new_util > 100:
                    continue
                
                # ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡πÅ‡∏•‡∏∞‡∏£‡∏ß‡∏°‡πÅ‡∏•‡πâ‡∏ß‡πÑ‡∏î‡πâ util ‡∏î‡∏µ
                min_util_target = MIN_UTIL.get(target_vehicle, 70)
                if new_util >= min_util_target and centroid_dist < best_merge_dist:
                    best_merge_util = new_util
                    best_merge_trip = trip_num
                    best_merge_dist = centroid_dist
            
            # ‡∏¢‡πâ‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏õ‡∏ó‡∏£‡∏¥‡∏õ‡πÉ‡∏´‡∏°‡πà
            if best_merge_trip is not None:
                for code in low_trip['codes']:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = best_merge_trip
                balance_count += 1
    
    # ===============================================
    # üéØ Phase 6.5: ‡∏ö‡∏±‡∏á‡∏Ñ‡∏±‡∏ö‡∏£‡∏ß‡∏° 4W ‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏ú‡πà‡∏≤‡∏ô MIN_UTIL
    # 1. ‡∏´‡∏≤ 6W ‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡πÄ‡∏ï‡πá‡∏°‡πÅ‡∏•‡∏∞‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á
    # 2. ‡∏´‡∏≤ JB ‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡πÄ‡∏ï‡πá‡∏°‡πÅ‡∏•‡∏∞‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á
    # 3. ‡∏£‡∏ß‡∏° 2 ‡∏Ñ‡∏±‡∏ô 4W ‡πÄ‡∏Ç‡πâ‡∏≤‡∏î‡πâ‡∏ß‡∏¢‡∏Å‡∏±‡∏ô‡πÄ‡∏õ‡πá‡∏ô JB
    # ===============================================
    merge_4w_count = 0
    MERGE_DISTANCE_LIMIT = 80  # ‡∏¢‡∏≠‡∏°‡πÉ‡∏´‡πâ‡∏£‡∏ß‡∏°‡πÑ‡∏î‡πâ‡πÑ‡∏Å‡∏•‡∏Ç‡∏∂‡πâ‡∏ô‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö merge
    
    # ‡∏´‡∏≤‡∏ó‡∏£‡∏¥‡∏õ 4W ‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î (‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏° util ‡∏à‡∏≤‡∏Å‡∏ô‡πâ‡∏≠‡∏¢‡πÑ‡∏õ‡∏°‡∏≤‡∏Å)
    trips_4w_to_merge = []
    for trip_num in list(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 0:
            continue
            
        trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
        
        # ‡πÄ‡∏â‡∏û‡∏≤‡∏∞ 4W ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô
        if trip_vehicle != '4W':
            continue
        
        trip_w = trip_data['Weight'].sum()
        trip_c = trip_data['Cube'].sum()
        trip_count = len(trip_data)
        trip_codes_list = list(trip_data['Code'].values)
        
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì util ‡∏Ç‡∏≠‡∏á 4W
        util_4w = max(
            (trip_w / LIMITS['4W']['max_w']) * 100,
            (trip_c / LIMITS['4W']['max_c']) * 100
        )
        
        # ‡πÄ‡∏Å‡πá‡∏ö‡∏ó‡∏∏‡∏Å‡∏ó‡∏£‡∏¥‡∏õ 4W ‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡∏ú‡πà‡∏≤‡∏ô MIN_UTIL (70%)
        if util_4w < MIN_UTIL.get('4W', 70):
            # ‡∏´‡∏≤ centroid
            trip_lats, trip_lons = [], []
            for code in trip_codes_list:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if trip_lats:
                trips_4w_to_merge.append({
                    'trip_num': trip_num,
                    'util': util_4w,
                    'weight': trip_w,
                    'cube': trip_c,
                    'count': trip_count,
                    'codes': trip_codes_list,
                    'lat': sum(trip_lats) / len(trip_lats),
                    'lon': sum(trip_lons) / len(trip_lons)
                })
    
    # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏à‡∏≤‡∏Å util ‡∏ô‡πâ‡∏≠‡∏¢‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏Å‡πà‡∏≠‡∏ô (‡∏£‡∏ß‡∏°‡∏á‡πà‡∏≤‡∏¢‡∏Å‡∏ß‡πà‡∏≤)
    trips_4w_to_merge.sort(key=lambda x: x['util'])
    
    # ‡∏û‡∏¢‡∏≤‡∏¢‡∏≤‡∏°‡∏£‡∏ß‡∏°‡∏ó‡∏µ‡∏•‡∏∞‡∏ó‡∏£‡∏¥‡∏õ
    for trip_info in trips_4w_to_merge:
        trip_num = trip_info['trip_num']
        
        # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏ó‡∏£‡∏¥‡∏õ‡∏ô‡∏µ‡πâ‡∏¢‡∏±‡∏á‡∏°‡∏µ‡∏≠‡∏¢‡∏π‡πà‡πÑ‡∏´‡∏° (‡∏≠‡∏≤‡∏à‡∏ñ‡∏π‡∏Å‡∏£‡∏ß‡∏°‡πÑ‡∏õ‡πÅ‡∏•‡πâ‡∏ß)
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 0:
            continue
        
        trip_w = trip_info['weight']
        trip_c = trip_info['cube']
        trip_count = trip_info['count']
        trip_codes = trip_info['codes']
        
        # ‡∏´‡∏≤ centroid ‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ 4W ‡∏ô‡∏µ‡πâ
        trip_lats, trip_lons = [], []
        for code in trip_codes:
            lat, lon = coord_cache.get(code, (None, None))
            if lat and lon:
                trip_lats.append(lat)
                trip_lons.append(lon)
        
        if not trip_lats:
            continue
        
        trip_centroid_lat = sum(trip_lats) / len(trip_lats)
        trip_centroid_lon = sum(trip_lons) / len(trip_lons)
        
        # ‡∏´‡∏≤ 6W ‡∏´‡∏£‡∏∑‡∏≠ JB ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡πÅ‡∏•‡∏∞‡∏£‡∏±‡∏ö‡πÑ‡∏î‡πâ
        best_target = None
        best_dist = float('inf')
        
        for target_num in test_df['Trip'].unique():
            if target_num == 0 or target_num == trip_num:
                continue
            
            target_data = test_df[test_df['Trip'] == target_num]
            if len(target_data) == 0:
                continue
                
            target_vehicle = trip_recommended_vehicles.get(target_num, '4W')
            
            # ‡πÄ‡∏â‡∏û‡∏≤‡∏∞ 6W ‡∏´‡∏£‡∏∑‡∏≠ JB
            if target_vehicle not in ['6W', 'JB']:
                continue
            
            target_w = target_data['Weight'].sum()
            target_c = target_data['Cube'].sum()
            target_count = len(target_data)
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏£‡∏ß‡∏°‡πÅ‡∏•‡πâ‡∏ß‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô capacity
            new_w = target_w + trip_w
            new_c = target_c + trip_c
            new_count = target_count + trip_count
            
            limits = LIMITS.get(target_vehicle, LIMITS['6W'])
            new_util = max(
                (new_w / limits['max_w']) * 100,
                (new_c / limits['max_c']) * 100
            )
            
            # ‡∏ï‡πâ‡∏≠‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô 100% ‡πÅ‡∏•‡∏∞‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô
            max_branches = 20 if target_vehicle == '6W' else 12
            if new_util > 100 or new_count > max_branches:
                continue
            
            # ‡∏´‡∏≤ centroid ‡∏Ç‡∏≠‡∏á target
            target_lats, target_lons = [], []
            for code in target_data['Code'].values:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    target_lats.append(lat)
                    target_lons.append(lon)
            
            if not target_lats:
                continue
            
            target_centroid_lat = sum(target_lats) / len(target_lats)
            target_centroid_lon = sum(target_lons) / len(target_lons)
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á centroid
            dist = haversine_distance(trip_centroid_lat, trip_centroid_lon,
                                      target_centroid_lat, target_centroid_lon)
            
            # ‡∏£‡∏∞‡∏¢‡∏∞‡∏ï‡πâ‡∏≠‡∏á‡πÑ‡∏°‡πà‡πÄ‡∏Å‡∏¥‡∏ô MERGE_DISTANCE_LIMIT (80km ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö merge)
            if dist > MERGE_DISTANCE_LIMIT:
                continue
            
            # ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å target ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î
            if dist < best_dist:
                best_dist = dist
                best_target = target_num
        
        # ‡∏¢‡πâ‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏õ target
        if best_target is not None:
            for code in trip_codes:
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_target
            merge_4w_count += 1
    
    # ===============================================
    # üéØ Phase 6.6: ‡∏£‡∏ß‡∏° 4W ‡∏´‡∏•‡∏≤‡∏¢‡∏Ñ‡∏±‡∏ô‡πÄ‡∏Ç‡πâ‡∏≤‡∏î‡πâ‡∏ß‡∏¢‡∏Å‡∏±‡∏ô‡πÉ‡∏´‡πâ‡∏Å‡∏•‡∏≤‡∏¢‡πÄ‡∏õ‡πá‡∏ô JB ‡∏´‡∏£‡∏∑‡∏≠ 6W
    # ===============================================
    merge_4w_to_larger_count = 0
    
    # ‡πÄ‡∏£‡∏µ‡∏¢‡∏Å trips_4w_to_merge ‡πÉ‡∏´‡∏°‡πà (‡πÄ‡∏û‡∏£‡∏≤‡∏∞‡∏≠‡∏≤‡∏à‡∏°‡∏µ‡∏ö‡∏≤‡∏á‡∏™‡πà‡∏ß‡∏ô‡∏ñ‡∏π‡∏Å‡∏£‡∏ß‡∏°‡πÑ‡∏õ‡πÅ‡∏•‡πâ‡∏ß)
    trips_4w_remaining = []
    for trip_num in list(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 0:
            continue
            
        trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
        if trip_vehicle != '4W':
            continue
        
        trip_w = trip_data['Weight'].sum()
        trip_c = trip_data['Cube'].sum()
        trip_count = len(trip_data)
        trip_codes_list = list(trip_data['Code'].values)
        
        # ‡∏´‡∏≤ centroid
        trip_lats, trip_lons = [], []
        for code in trip_codes_list:
            lat, lon = coord_cache.get(code, (None, None))
            if lat and lon:
                trip_lats.append(lat)
                trip_lons.append(lon)
        
        if trip_lats:
            trips_4w_remaining.append({
                'trip_num': trip_num,
                'weight': trip_w,
                'cube': trip_c,
                'count': trip_count,
                'codes': trip_codes_list,
                'lat': sum(trip_lats) / len(trip_lats),
                'lon': sum(trip_lons) / len(trip_lons)
            })
    
    # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏° cube ‡∏ô‡πâ‡∏≠‡∏¢‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡∏Å‡πà‡∏≠‡∏ô
    trips_4w_remaining.sort(key=lambda x: x['cube'])
    
    # ‡∏û‡∏¢‡∏≤‡∏¢‡∏≤‡∏°‡∏£‡∏ß‡∏° 4W ‡∏Å‡∏±‡∏ö 4W ‡∏≠‡∏∑‡πà‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏Å‡∏±‡∏ô
    merged_trips = set()
    for i, trip1 in enumerate(trips_4w_remaining):
        if trip1['trip_num'] in merged_trips:
            continue
        
        # ‡∏´‡∏≤ 4W ‡∏≠‡∏∑‡πà‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡πÅ‡∏•‡∏∞‡∏£‡∏ß‡∏°‡πÑ‡∏î‡πâ
        for j, trip2 in enumerate(trips_4w_remaining):
            if i >= j or trip2['trip_num'] in merged_trips:
                continue
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏´‡πà‡∏≤‡∏á centroid
            dist = haversine_distance(trip1['lat'], trip1['lon'], trip2['lat'], trip2['lon'])
            if dist > MERGE_DISTANCE_LIMIT:  # ‡πÉ‡∏ä‡πâ 80km ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö merge
                continue
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏£‡∏ß‡∏°‡πÅ‡∏•‡πâ‡∏ß‡πÉ‡∏™‡πà JB ‡πÑ‡∏î‡πâ‡πÑ‡∏´‡∏°
            combined_w = trip1['weight'] + trip2['weight']
            combined_c = trip1['cube'] + trip2['cube']
            combined_count = trip1['count'] + trip2['count']
            
            # ‡∏•‡∏≠‡∏á‡πÉ‡∏™‡πà JB (7 cube, 3500kg)
            if combined_c <= LIMITS['JB']['max_c'] and combined_w <= LIMITS['JB']['max_w'] and combined_count <= 12:
                # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏£‡∏±‡∏ö JB ‡πÑ‡∏î‡πâ‡πÑ‡∏´‡∏°
                all_codes = trip1['codes'] + trip2['codes']
                can_use_jb = True
                for code in all_codes:
                    branch_max = get_max_vehicle_for_branch(code)
                    if branch_max == '4W':  # ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ‡∏£‡∏±‡∏ö JB ‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ
                        can_use_jb = False
                        break
                
                if can_use_jb:
                    # ‡∏£‡∏ß‡∏°‡πÑ‡∏î‡πâ! ‡∏¢‡πâ‡∏≤‡∏¢‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏à‡∏≤‡∏Å trip2 ‡πÑ‡∏õ trip1
                    for code in trip2['codes']:
                        test_df.loc[test_df['Code'] == code, 'Trip'] = trip1['trip_num']
                    
                    # ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï trip_recommended_vehicles ‡πÄ‡∏õ‡πá‡∏ô JB
                    trip_recommended_vehicles[trip1['trip_num']] = 'JB'
                    merged_trips.add(trip2['trip_num'])
                    merge_4w_to_larger_count += 1
                    
                    # ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï trip1 info
                    trip1['weight'] = combined_w
                    trip1['cube'] = combined_c
                    trip1['count'] = combined_count
                    trip1['codes'].extend(trip2['codes'])
    
    # ===============================================
    # üéØ Phase 6.7: ‡∏£‡∏ß‡∏° 4W ‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏•‡∏∑‡∏≠‡πÄ‡∏Ç‡πâ‡∏≤‡∏Å‡∏±‡∏ö JB ‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡πÄ‡∏ï‡πá‡∏°
    # ===============================================
    # ‡∏´‡∏≤ 4W ‡∏ó‡∏µ‡πà‡∏¢‡∏±‡∏á‡πÄ‡∏´‡∏•‡∏∑‡∏≠‡∏≠‡∏¢‡∏π‡πà
    for trip_num in list(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 0:
            continue
        
        trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
        if trip_vehicle != '4W':
            continue
        
        trip_w = trip_data['Weight'].sum()
        trip_c = trip_data['Cube'].sum()
        trip_count = len(trip_data)
        trip_codes_list = list(trip_data['Code'].values)
        
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì util ‡∏Ç‡∏≠‡∏á 4W
        util_4w = max(
            (trip_w / LIMITS['4W']['max_w']) * 100,
            (trip_c / LIMITS['4W']['max_c']) * 100
        )
        
        # ‡∏ñ‡πâ‡∏≤‡∏ú‡πà‡∏≤‡∏ô MIN_UTIL ‡πÅ‡∏•‡πâ‡∏ß ‚Üí ‡∏Ç‡πâ‡∏≤‡∏°
        if util_4w >= MIN_UTIL.get('4W', 70):
            continue
        
        # ‡∏´‡∏≤ centroid
        trip_lats, trip_lons = [], []
        for code in trip_codes_list:
            lat, lon = coord_cache.get(code, (None, None))
            if lat and lon:
                trip_lats.append(lat)
                trip_lons.append(lon)
        
        if not trip_lats:
            continue
        
        trip_lat = sum(trip_lats) / len(trip_lats)
        trip_lon = sum(trip_lons) / len(trip_lons)
        
        # ‡∏´‡∏≤ JB ‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡∏ó‡∏µ‡πà‡∏™‡∏∏‡∏î‡πÅ‡∏•‡∏∞‡∏£‡∏±‡∏ö‡πÑ‡∏î‡πâ
        best_jb = None
        best_dist = float('inf')
        
        for target_num in test_df['Trip'].unique():
            if target_num == 0 or target_num == trip_num:
                continue
            
            target_data = test_df[test_df['Trip'] == target_num]
            if len(target_data) == 0:
                continue
            
            target_vehicle = trip_recommended_vehicles.get(target_num, '4W')
            if target_vehicle != 'JB':
                continue
            
            target_w = target_data['Weight'].sum()
            target_c = target_data['Cube'].sum()
            target_count = len(target_data)
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ capacity
            new_w = target_w + trip_w
            new_c = target_c + trip_c
            new_count = target_count + trip_count
            
            if new_c > LIMITS['JB']['max_c'] or new_w > LIMITS['JB']['max_w'] or new_count > 12:
                continue
            
            # ‡∏´‡∏≤ centroid ‡∏Ç‡∏≠‡∏á JB
            target_lats, target_lons = [], []
            for code in target_data['Code'].values:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    target_lats.append(lat)
                    target_lons.append(lon)
            
            if not target_lats:
                continue
            
            target_lat = sum(target_lats) / len(target_lats)
            target_lon = sum(target_lons) / len(target_lons)
            
            dist = haversine_distance(trip_lat, trip_lon, target_lat, target_lon)
            if dist > MERGE_DISTANCE_LIMIT:
                continue
            
            if dist < best_dist:
                best_dist = dist
                best_jb = target_num
        
        # ‡∏¢‡πâ‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤‡πÑ‡∏õ JB
        if best_jb is not None:
            for code in trip_codes_list:
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_jb
            merge_4w_count += 1
    
    # ‡∏™‡∏£‡∏∏‡∏õ‡∏ú‡∏•‡πÅ‡∏•‡∏∞‡πÅ‡∏ô‡∏∞‡∏ô‡∏≥‡∏£‡∏ñ
    summary_data = []
    for trip_num in sorted(test_df['Trip'].unique()):
        trip_data = test_df[test_df['Trip'] == trip_num]
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        
        # ‡∏´‡∏≤‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÉ‡∏´‡∏ç‡πà‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏ó‡∏∏‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ (‡πÉ‡∏ä‡πâ get_max_vehicle_for_branch ‡πÇ‡∏î‡∏¢‡∏ï‡∏£‡∏á)
        trip_codes = trip_data['Code'].unique()
        max_vehicles = []
        for c in trip_codes:
            # ‡πÉ‡∏ä‡πâ‡∏ü‡∏±‡∏á‡∏Å‡πå‡∏ä‡∏±‡∏ô‡∏´‡∏•‡∏±‡∏Å‡∏ó‡∏µ‡πà‡∏£‡∏ß‡∏° Booking + Punthai ‡πÅ‡∏•‡πâ‡∏ß
            branch_max = get_max_vehicle_for_branch(c)
            max_vehicles.append(branch_max)
        
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤ max_vehicles ‡πÑ‡∏°‡πà‡∏ß‡πà‡∏≤‡∏á
        if max_vehicles:
            min_max_size = min(vehicle_sizes.get(v, 3) for v in max_vehicles)
        else:
            min_max_size = 3  # default ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• ‡πÉ‡∏´‡πâ‡πÉ‡∏ä‡πâ 6W ‡πÑ‡∏î‡πâ
        
        max_allowed_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(min_max_size, '6W')
        
        # ‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡∏£‡∏ñ: ‡πÄ‡∏ä‡πá‡∏Ñ‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤‡∏Å‡πà‡∏≠‡∏ô ‡πÅ‡∏•‡πâ‡∏ß‡∏Ñ‡πà‡∏≠‡∏¢‡πÉ‡∏ä‡πâ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥/AI
        if trip_num in trip_recommended_vehicles:
            # ‡∏°‡∏µ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥
            suggested_from_history = trip_recommended_vehicles[trip_num]
            
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏£‡∏ñ‡∏à‡∏≤‡∏Å‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥‡∏Ç‡∏±‡∏î‡∏Å‡∏±‡∏ö‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
            if vehicle_sizes.get(suggested_from_history, 0) > min_max_size:
                # ‡∏ñ‡πâ‡∏≤‡∏Ç‡∏±‡∏î - ‡∏ï‡πâ‡∏≠‡∏á‡∏•‡∏î‡∏•‡∏á‡∏ï‡∏≤‡∏°‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤
                suggested = max_allowed_vehicle
                source = f"üìú ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥ ‚Üí {max_allowed_vehicle} (‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤)"
            else:
                # ‡πÑ‡∏°‡πà‡∏Ç‡∏±‡∏î - ‡πÉ‡∏ä‡πâ‡∏ï‡∏≤‡∏°‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥
                suggested = suggested_from_history
                source = "üìú ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥"
        else:
            # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥ - ‡πÉ‡∏ä‡πâ AI ‡∏û‡∏£‡πâ‡∏≠‡∏°‡πÄ‡∏Ñ‡∏≤‡∏£‡∏û‡∏Ç‡πâ‡∏≠‡∏à‡∏≥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤
            suggested = suggest_truck(total_w, total_c, max_allowed_vehicle, trip_codes)
            if min_max_size < 3:
                source = f"ü§ñ AI (‡∏à‡∏≥‡∏Å‡∏±‡∏î {max_allowed_vehicle})"
            else:
                source = "ü§ñ AI"
        
        # üîí ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡πÄ‡∏õ‡πá‡∏ô‡∏û‡∏∑‡πâ‡∏ô‡∏ó‡∏µ‡πà nearby (‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û+‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•) ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà ‚Üí ‡∏´‡πâ‡∏≤‡∏° 6W ‡πÄ‡∏î‡πá‡∏î‡∏Ç‡∏≤‡∏î!
        provinces = set()
        for code in trip_codes:
            prov = get_province(code)
            if prov and prov != 'UNKNOWN':
                provinces.add(prov)
        is_nearby_trip = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
        
        # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÉ‡∏™‡πà‡∏Ç‡∏≠‡∏á‡πÑ‡∏î‡πâ‡∏à‡∏£‡∏¥‡∏á‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà (‡∏´‡πâ‡∏≤‡∏°‡πÄ‡∏Å‡∏¥‡∏ô 100%)
        if suggested in LIMITS:
            w_util = (total_w / LIMITS[suggested]['max_w']) * 100
            c_util = (total_c / LIMITS[suggested]['max_c']) * 100
            max_util = max(w_util, c_util)
            
            # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Å‡∏¥‡∏ô 100% ‡∏ï‡πâ‡∏≠‡∏á‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ç‡∏ô‡∏≤‡∏î‡∏£‡∏ñ
            if max_util > 100:
                if suggested == '4W' and 'JB' in LIMITS:
                    # ‡∏•‡∏≠‡∏á‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡πÄ‡∏õ‡πá‡∏ô JB
                    jb_w_util = (total_w / LIMITS['JB']['max_w']) * 100
                    jb_c_util = (total_c / LIMITS['JB']['max_c']) * 100
                    if max(jb_w_util, jb_c_util) <= 100:
                        suggested = 'JB'
                        source = source + " ‚Üí JB"
                        w_util, c_util = jb_w_util, jb_c_util
                    else:
                        # üö´ ‡∏´‡πâ‡∏≤‡∏°‡πÉ‡∏ä‡πâ 6W ‡πÉ‡∏ô‡∏û‡∏∑‡πâ‡∏ô‡∏ó‡∏µ‡πà nearby!
                        if is_nearby_trip:
                            # ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ (‡∏à‡∏∞‡πÅ‡∏¢‡∏Å‡πÉ‡∏ô Phase 2.5) - ‡∏¢‡∏±‡∏á‡∏Ñ‡∏á‡πÉ‡∏ä‡πâ JB ‡πÑ‡∏ß‡πâ‡∏Å‡πà‡∏≠‡∏ô
                            suggested = 'JB'
                            source = source + " ‚Üí JB (‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ)"
                            w_util, c_util = jb_w_util, jb_c_util
                        else:
                            suggested = '6W'
                            source = source + " ‚Üí 6W"
                            w_util = (total_w / LIMITS['6W']['max_w']) * 100
                            c_util = (total_c / LIMITS['6W']['max_c']) * 100
                elif suggested == 'JB' or suggested == '4W':
                    # üö´ ‡∏´‡πâ‡∏≤‡∏°‡πÉ‡∏ä‡πâ 6W ‡πÉ‡∏ô‡∏û‡∏∑‡πâ‡∏ô‡∏ó‡∏µ‡πà nearby!
                    if is_nearby_trip:
                        # ‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ (‡∏à‡∏∞‡πÅ‡∏¢‡∏Å‡πÉ‡∏ô Phase 2.5) - ‡∏¢‡∏±‡∏á‡∏Ñ‡∏á‡πÉ‡∏ä‡πâ JB ‡πÑ‡∏ß‡πâ‡∏Å‡πà‡∏≠‡∏ô
                        suggested = 'JB'
                        source = source + " ‚Üí JB (‡∏ï‡πâ‡∏≠‡∏á‡πÅ‡∏¢‡∏Å‡∏ó‡∏£‡∏¥‡∏õ)"
                        jb_w_util = (total_w / LIMITS['JB']['max_w']) * 100
                        jb_c_util = (total_c / LIMITS['JB']['max_c']) * 100
                        w_util, c_util = jb_w_util, jb_c_util
                    else:
                        suggested = '6W'
                        source = source + " ‚Üí 6W"
                        w_util = (total_w / LIMITS['6W']['max_w']) * 100
                        c_util = (total_c / LIMITS['6W']['max_c']) * 100
        else:
            w_util = c_util = 0
        
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏£‡∏ß‡∏°‡∏Ç‡∏≠‡∏á‡∏ó‡∏£‡∏¥‡∏õ (‡πÄ‡∏™‡πâ‡∏ô‡∏ó‡∏≤‡∏á: DC ‚Üí ‡∏™‡∏≤‡∏Ç‡∏≤1 ‚Üí ‡∏™‡∏≤‡∏Ç‡∏≤2 ‚Üí ... ‚Üí DC)
        total_distance = 0
        if trip_codes is not None and len(trip_codes) > 0:
            # ‡∏î‡∏∂‡∏á‡∏û‡∏¥‡∏Å‡∏±‡∏î‡∏Ç‡∏≠‡∏á‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏™‡∏≤‡∏Ç‡∏≤‡∏à‡∏≤‡∏Å Master
            branch_coords = []
            for code in trip_codes:
                if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                    master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                    if len(master_row) > 0:
                        lat = master_row.iloc[0].get('‡∏•‡∏∞‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
                        lon = master_row.iloc[0].get('‡∏•‡∏≠‡∏á‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
                        if lat != 0 and lon != 0:
                            branch_coords.append((lat, lon))
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏ï‡∏≤‡∏°‡πÄ‡∏™‡πâ‡∏ô‡∏ó‡∏≤‡∏á
            if len(branch_coords) > 0:
                # DC ‚Üí ‡∏™‡∏≤‡∏Ç‡∏≤‡πÅ‡∏£‡∏Å
                total_distance += calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, 
                                                    branch_coords[0][0], branch_coords[0][1])
                # ‡∏™‡∏≤‡∏Ç‡∏≤ ‚Üí ‡∏™‡∏≤‡∏Ç‡∏≤
                for i in range(len(branch_coords) - 1):
                    total_distance += calculate_distance(branch_coords[i][0], branch_coords[i][1],
                                                        branch_coords[i+1][0], branch_coords[i+1][1])
                # ‡∏™‡∏≤‡∏Ç‡∏≤‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢ ‚Üí DC
                total_distance += calculate_distance(branch_coords[-1][0], branch_coords[-1][1],
                                                    DC_WANG_NOI_LAT, DC_WANG_NOI_LON)
        
        summary_data.append({
            'Trip': trip_num,
            'Branches': len(trip_data),
            'Weight': total_w,
            'Cube': total_c,
            'Truck': f"{suggested} {source}",
            'Weight_Use%': w_util,
            'Cube_Use%': c_util,
            'Total_Distance': total_distance
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡∏à‡∏±‡∏î‡∏™‡πà‡∏á‡πÉ‡∏ô‡∏£‡∏≤‡∏¢‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î‡∏£‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤
    trip_truck_map = {}
    trip_truck_type_map = {}  # ‡πÄ‡∏Å‡πá‡∏ö‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ (‡πÑ‡∏°‡πà‡∏£‡∏ß‡∏° source)
    for _, row in summary_df.iterrows():
        trip_truck_map[row['Trip']] = row['Truck']
        # ‡∏î‡∏∂‡∏á‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏ñ (‡∏ï‡∏±‡∏î emoji ‡πÅ‡∏•‡∏∞ source ‡∏≠‡∏≠‡∏Å)
        truck_type = row['Truck'].split()[0] if row['Truck'] else '6W'
        trip_truck_type_map[row['Trip']] = truck_type
    
    test_df['Truck'] = test_df['Trip'].map(trip_truck_map)
    
    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏†‡∏≤‡∏Ñ (Region) ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏™‡∏î‡∏á‡∏ú‡∏•
    def get_region_name(code):
        """‡∏î‡∏∂‡∏á‡∏ä‡∏∑‡πà‡∏≠‡∏†‡∏≤‡∏Ñ‡∏à‡∏≤‡∏Å‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤"""
        if code not in test_df['Code'].values:
            return '‡πÑ‡∏°‡πà‡∏£‡∏∞‡∏ö‡∏∏'
        
        prov = test_df[test_df['Code'] == code]['Province'].iloc[0] if 'Province' in test_df.columns else None
        if pd.isna(prov) or prov == 'UNKNOWN':
            return '‡πÑ‡∏°‡πà‡∏£‡∏∞‡∏ö‡∏∏'
        
        region_type = get_region_type(prov)
        
        # ‡πÅ‡∏õ‡∏•‡∏á‡πÄ‡∏õ‡πá‡∏ô‡∏ä‡∏∑‡πà‡∏≠‡∏†‡∏≤‡∏Ñ‡∏†‡∏≤‡∏©‡∏≤‡πÑ‡∏ó‡∏¢
        if region_type == 'nearby':
            # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡πÄ‡∏õ‡πá‡∏ô‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏´‡∏£‡∏∑‡∏≠‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•
            bangkok = ['‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏°‡∏´‡∏≤‡∏ô‡∏Ñ‡∏£', '‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û']
            if prov in bangkok:
                return '‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û'
            else:
                return '‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•'
        else:
            # ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ï‡∏≤‡∏°‡∏†‡∏π‡∏°‡∏¥‡∏†‡∏≤‡∏Ñ
            province_regions = {
                '‡∏Å‡∏•‡∏≤‡∏á‡∏ï‡∏≠‡∏ô‡∏ö‡∏ô': ['‡∏ä‡∏±‡∏¢‡∏ô‡∏≤‡∏ó', '‡∏û‡∏£‡∏∞‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤', '‡∏•‡∏û‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏£‡∏∞‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏¥‡∏á‡∏´‡πå‡∏ö‡∏∏‡∏£‡∏µ', '‡∏≠‡πà‡∏≤‡∏á‡∏ó‡∏≠‡∏á', '‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤'],
                '‡∏Å‡∏•‡∏≤‡∏á‡∏ï‡∏≠‡∏ô‡∏•‡πà‡∏≤‡∏á': ['‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏™‡∏á‡∏Ñ‡∏£‡∏≤‡∏°', '‡∏™‡∏∏‡∏û‡∏£‡∏£‡∏ì‡∏ö‡∏∏‡∏£‡∏µ'],
                '‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏ï‡∏Å': ['‡∏Å‡∏≤‡∏ç‡∏à‡∏ô‡∏ö‡∏∏‡∏£‡∏µ', '‡∏õ‡∏£‡∏∞‡∏à‡∏ß‡∏ö‡∏Ñ‡∏µ‡∏£‡∏µ‡∏Ç‡∏±‡∏ô‡∏ò‡πå', '‡∏£‡∏≤‡∏ä‡∏ö‡∏∏‡∏£‡∏µ', '‡πÄ‡∏û‡∏ä‡∏£‡∏ö‡∏∏‡∏£‡∏µ'],
                '‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏≠‡∏≠‡∏Å': ['‡∏à‡∏±‡∏ô‡∏ó‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ä‡∏•‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ï‡∏£‡∏≤‡∏î', '‡∏ô‡∏Ñ‡∏£‡∏ô‡∏≤‡∏¢‡∏Å', '‡∏õ‡∏£‡∏≤‡∏à‡∏µ‡∏ô‡∏ö‡∏∏‡∏£‡∏µ', '‡∏£‡∏∞‡∏¢‡∏≠‡∏á', '‡∏™‡∏£‡∏∞‡πÅ‡∏Å‡πâ‡∏ß', '‡∏â‡∏∞‡πÄ‡∏ä‡∏¥‡∏á‡πÄ‡∏ó‡∏£‡∏≤'],
                '‡∏≠‡∏µ‡∏™‡∏≤‡∏ô‡πÄ‡∏´‡∏ô‡∏∑‡∏≠': ['‡∏ô‡∏Ñ‡∏£‡∏û‡∏ô‡∏°', '‡∏ö‡∏∂‡∏á‡∏Å‡∏≤‡∏¨', '‡∏°‡∏∏‡∏Å‡∏î‡∏≤‡∏´‡∏≤‡∏£', '‡∏™‡∏Å‡∏•‡∏ô‡∏Ñ‡∏£', '‡∏´‡∏ô‡∏≠‡∏á‡∏Ñ‡∏≤‡∏¢', '‡∏´‡∏ô‡∏≠‡∏á‡∏ö‡∏±‡∏ß‡∏•‡∏≥‡∏†‡∏π', '‡∏≠‡∏∏‡∏î‡∏£‡∏ò‡∏≤‡∏ô‡∏µ', '‡πÄ‡∏•‡∏¢'],
                '‡∏≠‡∏µ‡∏™‡∏≤‡∏ô‡∏Å‡∏•‡∏≤‡∏á': ['‡∏Å‡∏≤‡∏¨‡∏™‡∏¥‡∏ô‡∏ò‡∏∏‡πå', '‡∏Ç‡∏≠‡∏ô‡πÅ‡∏Å‡πà‡∏ô', '‡∏ä‡∏±‡∏¢‡∏†‡∏π‡∏°‡∏¥', '‡∏°‡∏´‡∏≤‡∏™‡∏≤‡∏£‡∏Ñ‡∏≤‡∏°', '‡∏£‡πâ‡∏≠‡∏¢‡πÄ‡∏≠‡πá‡∏î'],
                '‡∏≠‡∏µ‡∏™‡∏≤‡∏ô‡πÉ‡∏ï‡πâ': ['‡∏ô‡∏Ñ‡∏£‡∏£‡∏≤‡∏ä‡∏™‡∏µ‡∏°‡∏≤', '‡πÇ‡∏Ñ‡∏£‡∏≤‡∏ä', '‡∏ö‡∏∏‡∏£‡∏µ‡∏£‡∏±‡∏°‡∏¢‡πå', '‡∏¢‡πÇ‡∏™‡∏ò‡∏£', '‡∏®‡∏£‡∏µ‡∏™‡∏∞‡πÄ‡∏Å‡∏©', '‡∏™‡∏∏‡∏£‡∏¥‡∏ô‡∏ó‡∏£‡πå', '‡∏≠‡∏≥‡∏ô‡∏≤‡∏à‡πÄ‡∏à‡∏£‡∏¥‡∏ç', '‡∏≠‡∏∏‡∏ö‡∏•‡∏£‡∏≤‡∏ä‡∏ò‡∏≤‡∏ô‡∏µ'],
                '‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ï‡∏≠‡∏ô‡∏ö‡∏ô': ['‡∏ô‡πà‡∏≤‡∏ô', '‡∏û‡∏∞‡πÄ‡∏¢‡∏≤', '‡∏•‡∏≥‡∏õ‡∏≤‡∏á', '‡∏•‡∏≥‡∏û‡∏π‡∏ô', '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡∏£‡∏≤‡∏¢', '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡πÉ‡∏´‡∏°‡πà', '‡πÅ‡∏û‡∏£‡πà', '‡πÅ‡∏°‡πà‡∏Æ‡πà‡∏≠‡∏á‡∏™‡∏≠‡∏ô'],
                '‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ï‡∏≠‡∏ô‡∏•‡πà‡∏≤‡∏á': ['‡∏Å‡∏≥‡πÅ‡∏û‡∏á‡πÄ‡∏û‡∏ä‡∏£', '‡∏ï‡∏≤‡∏Å', '‡∏ô‡∏Ñ‡∏£‡∏™‡∏ß‡∏£‡∏£‡∏Ñ‡πå', '‡∏û‡∏¥‡∏à‡∏¥‡∏ï‡∏£', '‡∏û‡∏¥‡∏©‡∏ì‡∏∏‡πÇ‡∏•‡∏Å', '‡∏™‡∏∏‡πÇ‡∏Ç‡∏ó‡∏±‡∏¢', '‡∏≠‡∏∏‡∏ï‡∏£‡∏î‡∏¥‡∏ï‡∏ñ‡πå', '‡∏≠‡∏∏‡∏ó‡∏±‡∏¢‡∏ò‡∏≤‡∏ô‡∏µ', '‡πÄ‡∏û‡∏ä‡∏£‡∏ö‡∏π‡∏£‡∏ì‡πå'],
                '‡πÉ‡∏ï‡πâ‡∏ù‡∏±‡πà‡∏á‡∏≠‡∏±‡∏ô‡∏î‡∏≤‡∏°‡∏±‡∏ô': ['‡∏Å‡∏£‡∏∞‡∏ö‡∏µ‡πà', '‡∏ï‡∏£‡∏±‡∏á', '‡∏û‡∏±‡∏á‡∏á‡∏≤', '‡∏†‡∏π‡πÄ‡∏Å‡πá‡∏ï', '‡∏£‡∏∞‡∏ô‡∏≠‡∏á', '‡∏™‡∏ï‡∏π‡∏•'],
                '‡πÉ‡∏ï‡πâ‡∏ù‡∏±‡πà‡∏á‡∏≠‡πà‡∏≤‡∏ß‡πÑ‡∏ó‡∏¢': ['‡∏ä‡∏∏‡∏°‡∏û‡∏£', '‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏ò‡∏£‡∏£‡∏°‡∏£‡∏≤‡∏ä', '‡∏û‡∏±‡∏ó‡∏•‡∏∏‡∏á', '‡∏¢‡∏∞‡∏•‡∏≤', '‡∏™‡∏á‡∏Ç‡∏•‡∏≤', '‡∏™‡∏∏‡∏£‡∏≤‡∏©‡∏é‡∏£‡πå‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏õ‡∏±‡∏ï‡∏ï‡∏≤‡∏ô‡∏µ', '‡∏ô‡∏£‡∏≤‡∏ò‡∏¥‡∏ß‡∏≤‡∏™']
            }
            
            for region_name, provinces in province_regions.items():
                if prov in provinces:
                    return region_name
            
            return '‡∏ï‡πà‡∏≤‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î'
    
    test_df['Region'] = test_df['Code'].apply(get_region_name)
    
    # üÜï ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ï‡∏≥‡∏ö‡∏•‡πÅ‡∏•‡∏∞‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡∏à‡∏≤‡∏Å Master Data
    def get_subdistrict(code):
        """‡∏î‡∏∂‡∏á‡∏ï‡∏≥‡∏ö‡∏•‡∏à‡∏≤‡∏Å Master Data"""
        if MASTER_DATA.empty or 'Plan Code' not in MASTER_DATA.columns:
            return ''
        master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
        if len(master_row) > 0:
            sub = master_row.iloc[0].get('‡∏ï‡∏≥‡∏ö‡∏•', '')
            return str(sub).strip() if pd.notna(sub) else ''
        return ''
    
    def get_district(code):
        """‡∏î‡∏∂‡∏á‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡∏à‡∏≤‡∏Å Master Data"""
        if MASTER_DATA.empty or 'Plan Code' not in MASTER_DATA.columns:
            return ''
        master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
        if len(master_row) > 0:
            dist = master_row.iloc[0].get('‡∏≠‡∏≥‡πÄ‡∏†‡∏≠', '')
            return str(dist).strip() if pd.notna(dist) else ''
        return ''
    
    test_df['Subdistrict'] = test_df['Code'].apply(get_subdistrict)
    test_df['District'] = test_df['Code'].apply(get_district)
    
    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏£‡∏∞‡∏´‡∏ß‡πà‡∏≤‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ ‡πÅ‡∏•‡∏∞‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏•‡∏≥‡∏î‡∏±‡∏ö
    def add_distance_and_sort(df):
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á max ‡∏†‡∏≤‡∏¢‡πÉ‡∏ô‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏ó‡∏£‡∏¥‡∏õ
        trip_distances = {}
        for trip_num in df['Trip'].unique():
            trip_codes = df[df['Trip'] == trip_num]['Code'].tolist()
            max_dist = 0
            
            # ‡∏´‡∏≤‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏™‡∏π‡∏á‡∏™‡∏∏‡∏î‡∏£‡∏∞‡∏´‡∏ß‡πà‡∏≤‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
            for i in range(len(trip_codes)):
                for j in range(i + 1, len(trip_codes)):
                    code1, code2 = trip_codes[i], trip_codes[j]
                    
                    # ‡∏î‡∏∂‡∏á‡∏û‡∏¥‡∏Å‡∏±‡∏î
                    if not MASTER_DATA.empty:
                        m1 = MASTER_DATA[MASTER_DATA['Plan Code'] == code1]
                        m2 = MASTER_DATA[MASTER_DATA['Plan Code'] == code2]
                        
                        if len(m1) > 0 and len(m2) > 0:
                            lat1 = m1.iloc[0].get('‡∏•‡∏∞‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
                            lon1 = m1.iloc[0].get('‡∏•‡∏≠‡∏á‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
                            lat2 = m2.iloc[0].get('‡∏•‡∏∞‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
                            lon2 = m2.iloc[0].get('‡∏•‡∏≠‡∏á‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
                            
                            if lat1 and lon1 and lat2 and lon2:
                                dist = haversine_distance(lat1, lon1, lat2, lon2)
                                if dist > max_dist:
                                    max_dist = dist
            
            trip_distances[trip_num] = round(max_dist, 2)
        
        # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á max ‡πÉ‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
        df['Max_Distance_in_Trip'] = df['Trip'].map(trip_distances)
        
        # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏†‡∏≤‡∏¢‡πÉ‡∏ô‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏ó‡∏£‡∏¥‡∏õ: Trip ‚Üí Sequence (‡∏ñ‡πâ‡∏≤‡∏°‡∏µ) ‡∏´‡∏£‡∏∑‡∏≠ Weight
        if 'Sequence' in df.columns:
            df = df.sort_values(['Trip', 'Sequence'], ascending=[True, True])
        else:
            df = df.sort_values(['Trip', 'Weight'], ascending=[True, False])
        return df
    
    test_df = add_distance_and_sort(test_df)
    
    # üó∫Ô∏è ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡πÅ‡∏ö‡∏ö‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î
    def calculate_detailed_distances(df):
        """‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì: DC‚Üí‡∏™‡∏≤‡∏Ç‡∏≤‡πÅ‡∏£‡∏Å, ‡∏£‡∏∞‡∏´‡∏ß‡πà‡∏≤‡∏á‡∏™‡∏≤‡∏Ç‡∏≤, ‡∏£‡∏ß‡∏°‡∏ó‡∏±‡πâ‡∏á‡∏ó‡∏£‡∏¥‡∏õ"""
        
        # ‡πÄ‡∏ï‡∏£‡∏µ‡∏¢‡∏° dict ‡πÄ‡∏Å‡πá‡∏ö‡∏ú‡∏•‡∏•‡∏±‡∏û‡∏ò‡πå
        distance_from_dc = {}
        distance_to_next = {}
        total_trip_distance = {}
        
        for trip_num in df['Trip'].unique():
            if trip_num == 0:
                continue
            
            trip_data = df[df['Trip'] == trip_num].copy()
            
            # ‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏° Sequence (‡∏ñ‡πâ‡∏≤‡∏°‡∏µ) ‡∏´‡∏£‡∏∑‡∏≠ Weight ‡πÄ‡∏û‡∏∑‡πà‡∏≠‡πÉ‡∏´‡πâ‡πÑ‡∏î‡πâ‡∏•‡∏≥‡∏î‡∏±‡∏ö‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ö‡∏Å‡∏≤‡∏£‡πÅ‡∏™‡∏î‡∏á‡∏ú‡∏•
            if 'Sequence' in trip_data.columns:
                trip_data = trip_data.sort_values('Sequence', ascending=True)
            else:
                trip_data = trip_data.sort_values('Weight', ascending=False)
            codes = trip_data['Code'].tolist()
            
            # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á
            trip_total_dist = 0
            prev_lat, prev_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
            
            for i, code in enumerate(codes):
                # ‡∏´‡∏≤‡∏û‡∏¥‡∏Å‡∏±‡∏î‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ
                m = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                
                if len(m) > 0:
                    lat = m.iloc[0].get('‡∏•‡∏∞‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
                    lon = m.iloc[0].get('‡∏•‡∏≠‡∏á‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
                    
                    if lat and lon:
                        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á‡∏à‡∏≤‡∏Å‡∏à‡∏∏‡∏î‡∏Å‡πà‡∏≠‡∏ô‡∏´‡∏ô‡πâ‡∏≤
                        dist = haversine_distance(prev_lat, prev_lon, lat, lon)
                        
                        if i == 0:
                            # ‡∏™‡∏≤‡∏Ç‡∏≤‡πÅ‡∏£‡∏Å: ‡∏£‡∏∞‡∏¢‡∏∞‡∏à‡∏≤‡∏Å DC
                            distance_from_dc[code] = round(dist, 2)
                            distance_to_next[code] = 0  # ‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏£‡∏∞‡∏¢‡∏∞ "‡∏Å‡πà‡∏≠‡∏ô‡∏´‡∏ô‡πâ‡∏≤"
                        else:
                            # ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ñ‡∏±‡∏î‡πÑ‡∏õ: ‡∏£‡∏∞‡∏¢‡∏∞‡∏à‡∏≤‡∏Å‡∏™‡∏≤‡∏Ç‡∏≤‡∏Å‡πà‡∏≠‡∏ô‡∏´‡∏ô‡πâ‡∏≤
                            distance_from_dc[code] = round(haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon), 2)
                            distance_to_next[codes[i-1]] = round(dist, 2)  # ‡∏ö‡∏±‡∏ô‡∏ó‡∏∂‡∏Å‡∏ó‡∏µ‡πà‡∏™‡∏≤‡∏Ç‡∏≤‡∏Å‡πà‡∏≠‡∏ô‡∏´‡∏ô‡πâ‡∏≤
                            
                            if i == len(codes) - 1:
                                # ‡∏™‡∏≤‡∏Ç‡∏≤‡∏™‡∏∏‡∏î‡∏ó‡πâ‡∏≤‡∏¢: ‡πÑ‡∏°‡πà‡∏°‡∏µ "‡∏ñ‡∏±‡∏î‡πÑ‡∏õ"
                                distance_to_next[code] = 0
                        
                        trip_total_dist += dist
                        prev_lat, prev_lon = lat, lon
                    else:
                        distance_from_dc[code] = 0
                        distance_to_next[code] = 0
                else:
                    distance_from_dc[code] = 0
                    distance_to_next[code] = 0
            
            # ‡∏ö‡∏±‡∏ô‡∏ó‡∏∂‡∏Å‡∏£‡∏∞‡∏¢‡∏∞‡∏£‡∏ß‡∏°‡∏ó‡∏±‡πâ‡∏á‡∏ó‡∏£‡∏¥‡∏õ
            for code in codes:
                total_trip_distance[code] = round(trip_total_dist, 2)
        
        # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏•‡∏á‡πÉ‡∏ô DataFrame
        df['Distance_from_DC'] = df['Code'].map(distance_from_dc).fillna(0)
        df['Distance_to_Next'] = df['Code'].map(distance_to_next).fillna(0)
        df['Total_Trip_Distance'] = df['Code'].map(total_trip_distance).fillna(0)
        
        return df
    
    test_df = calculate_detailed_distances(test_df)
    
    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏ô‡∏µ‡πâ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
    def check_vehicle_history(row):
        code = row['Code']
        trip = row['Trip']
        truck_type = trip_truck_type_map.get(trip, '6W')
        
        if code not in branch_vehicles:
            return "‚úÖ ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ (‡∏™‡∏≤‡∏Ç‡∏≤‡πÉ‡∏´‡∏°‡πà)"
        
        vehicle_history = branch_vehicles.get(code, {})
        if not vehicle_history:
            return "‚úÖ ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ (‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥)"
        
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        requested_size = vehicle_sizes.get(truck_type, 0)
        
        # ‡∏´‡∏≤‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏™‡∏∏‡∏î‡∏ó‡∏µ‡πà‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ
        max_used_size = max(vehicle_sizes.get(v, 0) for v in vehicle_history)
        max_used_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(max_used_size, '6W')
        
        # ‡∏ñ‡πâ‡∏≤‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏ô‡∏µ‡πâ
        if truck_type in vehicle_history:
            count = vehicle_history[truck_type]
            return f"‚úÖ ‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ ({count} ‡∏Ñ‡∏£‡∏±‡πâ‡∏á)"
        
        # ‡∏ñ‡πâ‡∏≤‡∏Ç‡∏≠‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å‡∏Å‡∏ß‡πà‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ = ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ
        if requested_size < max_used_size:
            return f"‚úÖ ‡πÉ‡∏ä‡πâ‡πÑ‡∏î‡πâ (‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ {max_used_vehicle})"
        
        # ‡∏ñ‡πâ‡∏≤‡∏Ç‡∏≠‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà‡∏Å‡∏ß‡πà‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏Ñ‡∏¢‡πÉ‡∏ä‡πâ = ‡∏≠‡∏≤‡∏à‡πÄ‡∏Ç‡πâ‡∏≤‡πÑ‡∏°‡πà‡πÑ‡∏î‡πâ
        history_str = ", ".join([f"{v}:{c}" for v, c in vehicle_history.items()])
        return f"üö´ ‡∏à‡∏≥‡∏Å‡∏±‡∏î {max_used_vehicle} ({history_str})"
    
    test_df['VehicleCheck'] = test_df.apply(check_vehicle_history, axis=1)
    
    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏ó‡∏£‡∏¥‡∏õ‡∏ó‡∏µ‡πà‡πÄ‡∏Å‡∏¥‡∏ô 12 ‡∏™‡∏≤‡∏Ç‡∏≤ (‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å)
    def check_branch_count(row):
        trip_num = row['Trip']
        if trip_num == 0:
            return ""
        
        trip_branch_count = len(test_df[test_df['Trip'] == trip_num])
        truck_type = trip_truck_type_map.get(trip_num, '6W')
        
        # ‡πÄ‡∏ä‡πá‡∏Ñ‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏£‡∏ñ‡πÄ‡∏•‡πá‡∏Å (4W, JB) - ‡∏£‡∏ñ‡πÉ‡∏´‡∏ç‡πà (6W) ‡πÑ‡∏°‡πà‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô
        if trip_branch_count > 12:
            if truck_type in ['4W', 'JB']:
                return f"‚ö†Ô∏è ‡πÄ‡∏Å‡∏¥‡∏ô 12 ‡∏™‡∏≤‡∏Ç‡∏≤ ({trip_branch_count} ‡∏™‡∏≤‡∏Ç‡∏≤) - {truck_type}"
            else:
                return f"‚úÖ {trip_branch_count} ‡∏™‡∏≤‡∏Ç‡∏≤ - {truck_type} (‡∏¢‡∏≠‡∏°‡πÑ‡∏î‡πâ)"
        else:
            return f"‚úÖ {trip_branch_count} ‡∏™‡∏≤‡∏Ç‡∏≤ - {truck_type}"
    
    test_df['BranchCount'] = test_df.apply(check_branch_count, axis=1)
    
    # ===============================================
    # üéØ Renumber trips: ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å 1, 2, 3, ... (‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏ä‡πà‡∏≠‡∏á‡∏ß‡πà‡∏≤‡∏á)
    # ===============================================
    unique_trips = sorted(test_df['Trip'].dropna().unique())
    trip_renumber_map = {old: new for new, old in enumerate(unique_trips, start=1)}
    test_df['Trip'] = test_df['Trip'].map(trip_renumber_map)
    
    # ‡∏≠‡∏±‡∏õ‡πÄ‡∏î‡∏ï summary_df ‡∏î‡πâ‡∏ß‡∏¢
    if 'Trip' in summary_df.columns:
        summary_df['Trip'] = summary_df['Trip'].map(trip_renumber_map)
    
    return test_df, summary_df

# ==========================================
# STREAMLIT UI
# ==========================================
def main():
    st.set_page_config(
        page_title="‡∏£‡∏∞‡∏ö‡∏ö‡∏à‡∏±‡∏î‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏ß",
        page_icon="üöö",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # üîÑ Auto-refresh ‡∏ó‡∏∏‡∏Å‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏á‡∏Ñ‡∏∑‡∏ô (‡∏•‡πâ‡∏≤‡∏á‡πÅ‡∏Ñ‡∏ä)
    if AUTOREFRESH_AVAILABLE:
        now = datetime.now()
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡πÄ‡∏ß‡∏•‡∏≤‡∏ñ‡∏∂‡∏á‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏á‡∏Ñ‡∏∑‡∏ô (00:00:00)
        midnight = datetime.combine(now.date(), time(0, 0, 0))
        
        # ‡∏ñ‡πâ‡∏≤‡∏¢‡∏±‡∏á‡πÑ‡∏°‡πà‡∏ñ‡∏∂‡∏á‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏á‡∏Ñ‡∏∑‡∏ô ‡πÄ‡∏≠‡∏≤‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏á‡∏Ñ‡∏∑‡∏ô‡∏ß‡∏±‡∏ô‡∏ñ‡∏±‡∏î‡πÑ‡∏õ
        if now < midnight:
            next_midnight = midnight
        else:
            from datetime import timedelta
            next_midnight = midnight + timedelta(days=1)
        
        # ‡∏Ñ‡∏≥‡∏ô‡∏ß‡∏ì‡πÄ‡∏ß‡∏•‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏´‡∏•‡∏∑‡∏≠ (‡∏ß‡∏¥‡∏ô‡∏≤‡∏ó‡∏µ)
        seconds_until_midnight = int((next_midnight - now).total_seconds())
        
        # Refresh ‡∏ó‡∏∏‡∏Å‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏á‡∏Ñ‡∏∑‡∏ô
        if seconds_until_midnight > 0:
            # ‡πÄ‡∏ä‡πá‡∏Ñ‡πÉ‡∏ô‡∏ä‡πà‡∏ß‡∏á 5 ‡∏ô‡∏≤‡∏ó‡∏µ‡∏Å‡πà‡∏≠‡∏ô‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏á‡∏Ñ‡∏∑‡∏ô (‡∏´‡∏•‡∏±‡∏á 23:55)
            if seconds_until_midnight <= 300:  # 5 minutes
                st.info(f"üîÑ ‡∏£‡∏∞‡∏ö‡∏ö‡∏à‡∏∞ Refresh ‡∏≠‡∏±‡∏ï‡πÇ‡∏ô‡∏°‡∏±‡∏ï‡∏¥‡πÉ‡∏ô {seconds_until_midnight // 60} ‡∏ô‡∏≤‡∏ó‡∏µ")
                st_autorefresh(interval=seconds_until_midnight * 1000, key="midnight_refresh")
            else:
                # ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ó‡∏∏‡∏Å 1 ‡∏ä‡∏±‡πà‡∏ß‡πÇ‡∏°‡∏á
                st_autorefresh(interval=3600000, limit=24, key="hourly_check")
    
    # Header
    col1, col2 = st.columns([3, 1])
    with col1:
        st.title("üöö ‡∏£‡∏∞‡∏ö‡∏ö‡∏à‡∏±‡∏î‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏ß")
    with col2:
        st.image("https://raw.githubusercontent.com/twitter/twemoji/master/assets/svg/1f69a.svg", width=100)
    
    # Show Punthai learning stats
    if PUNTHAI_PATTERNS and 'stats' in PUNTHAI_PATTERNS and PUNTHAI_PATTERNS['stats']:
        stats = PUNTHAI_PATTERNS['stats']
        with st.expander("üìä ‡∏™‡∏ñ‡∏¥‡∏ï‡∏¥‡∏ó‡∏µ‡πà‡πÄ‡∏£‡∏µ‡∏¢‡∏ô‡∏£‡∏π‡πâ‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå Punthai Maxmart", expanded=False):
            col_a, col_b, col_c = st.columns(3)
            with col_a:
                st.metric("‡πÄ‡∏â‡∏•‡∏µ‡πà‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤/‡∏ó‡∏£‡∏¥‡∏õ", f"{stats.get('avg_branches', 0):.1f}")
            with col_b:
                st.metric("‡∏ó‡∏£‡∏¥‡∏õ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß", f"{stats.get('same_province_pct', 0):.1f}%")
            with col_c:
                total_trips = stats.get('same_province', 0) + stats.get('mixed_province', 0)
                st.metric("‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏ó‡∏£‡∏¥‡∏õ‡∏≠‡πâ‡∏≤‡∏á‡∏≠‡∏¥‡∏á", total_trips)
    
    st.markdown("---")
    
    # ‡πÇ‡∏´‡∏•‡∏î‡πÇ‡∏°‡πÄ‡∏î‡∏•
    model_data = load_model()
    
    if not model_data:
        st.error("‚ùå ‡πÑ‡∏°‡πà‡∏û‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÇ‡∏°‡πÄ‡∏î‡∏• ‡∏Å‡∏£‡∏∏‡∏ì‡∏≤‡πÄ‡∏ó‡∏£‡∏ô‡πÇ‡∏°‡πÄ‡∏î‡∏•‡∏Å‡πà‡∏≠‡∏ô‡πÉ‡∏ä‡πâ‡∏á‡∏≤‡∏ô")
        st.info("üí° ‡∏£‡∏±‡∏ô‡∏Ñ‡∏≥‡∏™‡∏±‡πà‡∏á: `python test_model.py`")
        st.stop()
    
    # ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏Ñ‡∏£‡∏±‡πâ‡∏á‡πÄ‡∏î‡∏µ‡∏¢‡∏ß
    st.markdown("### üìÇ ‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏≠‡∏≠‡πÄ‡∏î‡∏≠‡∏£‡πå")
    uploaded_file = st.file_uploader(
        "‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Excel (.xlsx)", 
        type=['xlsx'],
        help="‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå Excel ‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏™‡∏≤‡∏Ç‡∏≤‡πÅ‡∏•‡∏∞‡∏≠‡∏≠‡πÄ‡∏î‡∏≠‡∏£‡πå"
    )
    
    if uploaded_file:
        # ‡πÄ‡∏Å‡πá‡∏ö file content ‡πÑ‡∏ß‡πâ‡πÉ‡∏ä‡πâ‡∏ï‡∏≠‡∏ô export
        original_file_content = uploaded_file.read()
        uploaded_file.seek(0)  # reset pointer
        
        with st.spinner("‚è≥ ‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏≠‡πà‡∏≤‡∏ô‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•..."):
            df = load_excel(original_file_content)
            df = process_dataframe(df)
            
            if df is not None and 'Code' in df.columns:
                st.success(f"‚úÖ ‡∏≠‡πà‡∏≤‡∏ô‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏™‡∏≥‡πÄ‡∏£‡πá‡∏à: **{len(df):,}** ‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£")
                
                # ‡πÅ‡∏™‡∏î‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏û‡∏∑‡πâ‡∏ô‡∏ê‡∏≤‡∏ô
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("üìç ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤", f"{df['Code'].nunique():,}")
                with col2:
                    st.metric("‚öñÔ∏è ‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å‡∏£‡∏ß‡∏°", f"{df['Weight'].sum():,.0f} kg")
                with col3:
                    st.metric("üì¶ ‡∏Ñ‡∏¥‡∏ß‡∏£‡∏ß‡∏°", f"{df['Cube'].sum():.1f} m¬≥")
                with col4:
                    provinces = df['Province'].nunique() if 'Province' in df.columns else 0
                    st.metric("üó∫Ô∏è ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î", f"{provinces}")
                
                # ‡πÅ‡∏™‡∏î‡∏á‡∏ï‡∏±‡∏ß‡∏≠‡∏¢‡πà‡∏≤‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•
                with st.expander("üîç ‡∏î‡∏π‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ï‡∏±‡∏ß‡∏≠‡∏¢‡πà‡∏≤‡∏á"):
                    st.dataframe(df.head(10), use_container_width=True)
                
                st.markdown("---")
                
                # ‡πÅ‡∏ó‡πá‡∏ö‡∏´‡∏•‡∏±‡∏Å
                tab1, tab2 = st.tabs(["üì¶ ‡∏à‡∏±‡∏î‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏ß (‡∏ï‡∏≤‡∏°‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å)", "üó∫Ô∏è ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ï‡∏≤‡∏°‡∏†‡∏≤‡∏Ñ (‡πÑ‡∏°‡πà‡∏™‡∏ô‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å)"])
                    
                # ==========================================
                # ‡πÅ‡∏ó‡πá‡∏ö 1: ‡∏à‡∏±‡∏î‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏ß (‡∏ï‡∏≤‡∏°‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å)
                # ==========================================
                with tab1:
                    # ‡∏õ‡∏∏‡πà‡∏°‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ
                    if st.button("üöÄ ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏±‡∏î‡πÄ‡∏ó‡∏µ‡πà‡∏¢‡∏ß", type="primary", use_container_width=True):
                        with st.spinner("‚è≥ ‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏õ‡∏£‡∏∞‡∏°‡∏ß‡∏•‡∏ú‡∏•..."):
                            result_df, summary = predict_trips(df.copy(), model_data)
                            
                            st.balloons()
                            st.success(f"‚úÖ **‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ‡πÄ‡∏™‡∏£‡πá‡∏à‡∏™‡∏°‡∏ö‡∏π‡∏£‡∏ì‡πå!** ‡∏£‡∏ß‡∏° **{len(summary)}** ‡∏ó‡∏£‡∏¥‡∏õ")
                            
                            st.markdown("---")
                            
                            # ‡∏™‡∏ñ‡∏¥‡∏ï‡∏¥‡πÇ‡∏î‡∏¢‡∏£‡∏ß‡∏°
                            st.markdown("### üìä ‡∏™‡∏£‡∏∏‡∏õ‡∏ú‡∏•‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ")
                            
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("üöö ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏ó‡∏£‡∏¥‡∏õ", len(summary))
                            with col2:
                                st.metric("üìç ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤", len(result_df))
                            with col3:
                                avg_branches = len(result_df) / result_df['Trip'].nunique()
                                st.metric("üìä ‡πÄ‡∏â‡∏•‡∏µ‡πà‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤/‡∏ó‡∏£‡∏¥‡∏õ", f"{avg_branches:.1f}")
                            with col4:
                                avg_util = summary['Cube_Use%'].mean()
                                st.metric("üìà ‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡πÄ‡∏â‡∏•‡∏µ‡πà‡∏¢", f"{avg_util:.0f}%")
                            
                            st.markdown("---")
                            
                            # ‡∏ï‡∏≤‡∏£‡∏≤‡∏á‡∏™‡∏£‡∏∏‡∏õ‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏ó‡∏£‡∏¥‡∏õ
                            st.markdown("### üöõ ‡∏£‡∏≤‡∏¢‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏ó‡∏£‡∏¥‡∏õ")
                            st.dataframe(
                                summary.style.format({
                                    'Weight': '{:.2f}',
                                    'Cube': '{:.2f}',
                                    'Weight_Use%': '{:.1f}%',
                                    'Cube_Use%': '{:.1f}%',
                                    'Total_Distance': '{:.1f} km'
                                }).background_gradient(
                                    subset=['Weight_Use%', 'Cube_Use%'],
                                    cmap='RdYlGn',
                                    vmin=0,
                                    vmax=100
                                ),
                                use_container_width=True,
                                height=400
                            )
                            
                            # ‡∏ï‡∏≤‡∏£‡∏≤‡∏á‡∏£‡∏≤‡∏¢‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î (‡∏°‡∏µ‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏£‡∏ñ‡πÅ‡∏•‡∏∞‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á)
                            with st.expander("üìã ‡∏î‡∏π‡∏£‡∏≤‡∏¢‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î‡∏£‡∏≤‡∏¢‡∏™‡∏≤‡∏Ç‡∏≤ (‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏ï‡∏≤‡∏°‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å)"):
                                # ‡∏à‡∏±‡∏î‡πÄ‡∏£‡∏µ‡∏¢‡∏á‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ó‡∏µ‡πà‡∏™‡∏≥‡∏Ñ‡∏±‡∏ç
                                display_cols = ['Trip', 'Code', 'Name']
                                if 'Province' in result_df.columns:
                                    display_cols.append('Province')
                                if 'District' in result_df.columns:
                                    display_cols.append('District')
                                if 'Subdistrict' in result_df.columns:
                                    display_cols.append('Subdistrict')
                                if 'Region' in result_df.columns:
                                    display_cols.append('Region')
                                display_cols.extend(['Max_Distance_in_Trip', 'Weight', 'Cube', 'Truck', 'VehicleCheck'])
                                
                                # ‡∏Å‡∏£‡∏≠‡∏á‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏≠‡∏¢‡∏π‡πà‡∏à‡∏£‡∏¥‡∏á
                                display_cols = [col for col in display_cols if col in result_df.columns]
                                display_df = result_df[display_cols].copy()
                                
                                # ‡∏ï‡∏±‡πâ‡∏á‡∏ä‡∏∑‡πà‡∏≠‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏†‡∏≤‡∏©‡∏≤‡πÑ‡∏ó‡∏¢
                                col_names = {'Trip': '‡∏ó‡∏£‡∏¥‡∏õ', 'Code': '‡∏£‡∏´‡∏±‡∏™', 'Name': '‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤', 'Province': '‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', 
                                           'District': '‡∏≠‡∏≥‡πÄ‡∏†‡∏≠', 'Subdistrict': '‡∏ï‡∏≥‡∏ö‡∏•',
                                           'Region': '‡∏†‡∏≤‡∏Ñ', 'Max_Distance_in_Trip': '‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á Max(km)', 
                                           'Weight': '‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å(kg)', 'Cube': '‡∏Ñ‡∏¥‡∏ß(m¬≥)', 'Truck': '‡∏£‡∏ñ', 'VehicleCheck': '‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏£‡∏ñ'}
                                display_df.columns = [col_names.get(c, c) for c in display_cols]
                                
                                # ‡∏à‡∏±‡∏î‡∏£‡∏π‡∏õ‡πÅ‡∏ö‡∏ö‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á
                                st.dataframe(
                                    display_df.style.format({
                                        '‡∏£‡∏∞‡∏¢‡∏∞‡∏ó‡∏≤‡∏á(km)': '{:.1f}',
                                        '‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å(kg)': '{:.2f}',
                                        '‡∏Ñ‡∏¥‡∏ß(m¬≥)': '{:.2f}'
                                    }),
                                    use_container_width=True, 
                                    height=400
                                )
                            
                            # ‡πÅ‡∏™‡∏î‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡∏°‡∏µ‡∏Ñ‡∏≥‡πÄ‡∏ï‡∏∑‡∏≠‡∏ô
                            warning_branches = result_df[result_df['VehicleCheck'].str.contains('‚ö†Ô∏è', na=False)]
                            if len(warning_branches) > 0:
                                with st.expander(f"‚ö†Ô∏è ‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏ï‡πà‡∏≤‡∏á‡∏à‡∏≤‡∏Å‡∏õ‡∏Å‡∏ï‡∏¥ ({len(warning_branches)} ‡∏™‡∏≤‡∏Ç‡∏≤)"):
                                    st.warning("‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏´‡∏•‡πà‡∏≤‡∏ô‡∏µ‡πâ‡∏õ‡∏Å‡∏ï‡∏¥‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏≠‡∏∑‡πà‡∏ô ‡πÅ‡∏ï‡πà‡∏ñ‡∏π‡∏Å‡∏à‡∏±‡∏î‡πÉ‡∏´‡πâ‡πÉ‡∏ä‡πâ‡∏£‡∏ñ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏ó‡∏µ‡πà‡∏ï‡πà‡∏≤‡∏á‡∏≠‡∏≠‡∏Å‡πÑ‡∏õ")
                                    display_cols_warn = ['Trip', 'Code', 'Name', 'Truck', 'VehicleCheck']
                                    display_warn_df = warning_branches[display_cols_warn].copy()
                                    display_warn_df.columns = ['‡∏ó‡∏£‡∏¥‡∏õ', '‡∏£‡∏´‡∏±‡∏™', '‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤', '‡∏£‡∏ñ‡∏ó‡∏µ‡πà‡∏à‡∏±‡∏î', '‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥‡∏Å‡∏≤‡∏£‡πÉ‡∏ä‡πâ‡∏£‡∏ñ']
                                    st.dataframe(display_warn_df, use_container_width=True)
                            
                            st.markdown("---")
                            
                            # ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î - ‡πÄ‡∏Ç‡∏µ‡∏¢‡∏ô‡∏ó‡∏±‡∏ö‡∏ä‡∏µ‡∏ï 2.Punthai ‡πÉ‡∏ô‡πÑ‡∏ü‡∏•‡πå‡∏ï‡πâ‡∏ô‡∏â‡∏ö‡∏±‡∏ö ‡∏û‡∏£‡πâ‡∏≠‡∏°‡∏™‡∏•‡∏±‡∏ö‡∏™‡∏µ‡πÄ‡∏´‡∏•‡∏∑‡∏≠‡∏á‡πÇ‡∏ó‡∏ô‡∏™‡πâ‡∏°-‡∏Ç‡∏≤‡∏ß
                            output = io.BytesIO()
                            
                            # ‡∏™‡∏£‡πâ‡∏≤‡∏á Trip_No map (JB ‡πÉ‡∏ä‡πâ prefix 4WJ)
                            trip_no_map = {}
                            vehicle_counts = {'4W': 0, '4WJ': 0, '6W': 0}
                            
                            for trip_num in sorted(result_df['Trip'].unique()):
                                if trip_num == 0:
                                    continue
                                trip_summary = summary[summary['Trip'] == trip_num]
                                if len(trip_summary) > 0:
                                    truck_info = trip_summary.iloc[0]['Truck']
                                    vehicle_type = truck_info.split()[0] if truck_info else '6W'
                                    # JB ‡πÉ‡∏ä‡πâ prefix 4WJ ‡πÅ‡∏ó‡∏ô
                                    if vehicle_type == 'JB':
                                        vehicle_type = '4WJ'
                                    vehicle_counts[vehicle_type] = vehicle_counts.get(vehicle_type, 0) + 1
                                    trip_no = f"{vehicle_type}{vehicle_counts[vehicle_type]:03d}"
                                    trip_no_map[trip_num] = trip_no
                            
                            # ‡πÇ‡∏´‡∏•‡∏î‡πÑ‡∏ü‡∏•‡πå‡∏ï‡πâ‡∏ô‡∏â‡∏ö‡∏±‡∏ö‡πÄ‡∏û‡∏∑‡πà‡∏≠ copy ‡∏ó‡∏∏‡∏Å‡∏ä‡∏µ‡∏ï
                            from openpyxl import load_workbook
                            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                            from copy import copy
                            
                            try:
                                # ‡πÇ‡∏´‡∏•‡∏î workbook ‡∏ï‡πâ‡∏ô‡∏â‡∏ö‡∏±‡∏ö
                                wb = load_workbook(io.BytesIO(original_file_content))
                                
                                # ‡∏´‡∏≤‡∏ä‡∏µ‡∏ï‡πÄ‡∏õ‡πâ‡∏≤‡∏´‡∏°‡∏≤‡∏¢ (2.Punthai)
                                target_sheet = None
                                for sheet_name in wb.sheetnames:
                                    if 'punthai' in sheet_name.lower() or '2.' in sheet_name.lower():
                                        target_sheet = sheet_name
                                        break
                                
                                if not target_sheet:
                                    target_sheet = '2.Punthai'
                                    if target_sheet not in wb.sheetnames:
                                        wb.create_sheet(target_sheet)
                                
                                ws = wb[target_sheet]
                                
                                # ‡∏•‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÄ‡∏Å‡πà‡∏≤ (‡πÄ‡∏Å‡πá‡∏ö‡πÅ‡∏ñ‡∏ß‡πÅ‡∏£‡∏Å header)
                                # ‡∏´‡∏≤ header row (‡πÅ‡∏ñ‡∏ß‡∏ó‡∏µ‡πà‡∏°‡∏µ "‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤" ‡∏´‡∏£‡∏∑‡∏≠ "Trip")
                                header_row = 1
                                for row_idx in range(1, min(5, ws.max_row + 1)):
                                    for col_idx in range(1, min(15, ws.max_column + 1)):
                                        cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                                        if '‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤' in cell_val or 'Trip' in cell_val.upper():
                                            header_row = row_idx
                                            break
                                
                                # ‡∏•‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ï‡∏±‡πâ‡∏á‡πÅ‡∏ï‡πà‡πÅ‡∏ñ‡∏ß‡∏´‡∏•‡∏±‡∏á header
                                if ws.max_row > header_row:
                                    ws.delete_rows(header_row + 1, ws.max_row - header_row)
                                
                                # ‡∏™‡∏µ‡πÄ‡∏´‡∏•‡∏∑‡∏≠‡∏á‡πÇ‡∏ó‡∏ô‡∏™‡πâ‡∏°-‡∏Ç‡∏≤‡∏ß
                                yellow_orange = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                                white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                                thin_border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
                                
                                # ‡πÄ‡∏Ç‡∏µ‡∏¢‡∏ô‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÉ‡∏´‡∏°‡πà (‡πÅ‡∏ñ‡∏ß‡∏ó‡∏µ‡πà header_row + 1 ‡πÄ‡∏õ‡πá‡∏ô‡∏ï‡πâ‡∏ô‡πÑ‡∏õ)
                                current_trip = None
                                use_yellow = True
                                row_num = header_row + 1
                                sep_num = 1  # ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏ô‡∏±‡∏ö Sep. ‡∏à‡∏≤‡∏Å 1
                                
                                for trip_num in sorted(result_df['Trip'].unique()):
                                    if trip_num == 0:
                                        continue
                                    trip_data = result_df[result_df['Trip'] == trip_num].copy()
                                    trip_no = trip_no_map.get(trip_num, '')
                                    
                                    # ‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡∏™‡∏µ‡πÄ‡∏°‡∏∑‡πà‡∏≠‡πÄ‡∏õ‡∏•‡∏µ‡πà‡∏¢‡∏ô‡∏ó‡∏£‡∏¥‡∏õ
                                    if current_trip != trip_num:
                                        current_trip = trip_num
                                        use_yellow = not use_yellow
                                    
                                    fill = yellow_orange if use_yellow else white_fill
                                    
                                    for _, row in trip_data.iterrows():
                                        # ‡πÄ‡∏Ç‡∏µ‡∏¢‡∏ô‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ï‡∏≤‡∏°‡πÇ‡∏Ñ‡∏£‡∏á‡∏™‡∏£‡πâ‡∏≤‡∏á‡πÑ‡∏ü‡∏•‡πå‡∏ï‡πâ‡∏ô‡∏â‡∏ö‡∏±‡∏ö
                                        # ‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå: A=Sep, B=BU, C=‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤, D=‡∏£‡∏´‡∏±‡∏™ WMS, E=‡∏™‡∏≤‡∏Ç‡∏≤, F=Cube, G=Weight, H=Original QTY, I=Trip, J=Trip no
                                        data = [
                                            sep_num,  # A: Sep (‡∏•‡∏≥‡∏î‡∏±‡∏ö‡πÅ‡∏ñ‡∏ß)
                                            row.get('BU', 211),  # B: BU (‡∏à‡∏≤‡∏Å‡∏ï‡πâ‡∏ô‡∏â‡∏ö‡∏±‡∏ö)
                                            row.get('Code', ''),  # C: ‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤
                                            row.get('Code', ''),  # D: ‡∏£‡∏´‡∏±‡∏™ WMS
                                            row.get('Name', ''),  # E: ‡∏™‡∏≤‡∏Ç‡∏≤
                                            round(row.get('Cube', 0), 2) if pd.notna(row.get('Cube')) else 0,  # F: Cube
                                            round(row.get('Weight', 0), 2) if pd.notna(row.get('Weight')) else 0,  # G: Weight
                                            row.get('OriginalQty', 0) if pd.notna(row.get('OriginalQty')) else 0,  # H: Original QTY (‡∏à‡∏≤‡∏Å‡∏ï‡πâ‡∏ô‡∏â‡∏ö‡∏±‡∏ö)
                                            int(trip_num),  # I: Trip
                                            trip_no,  # J: Trip no
                                            '',  # K: ‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡πÇ‡∏´‡∏•‡∏î
                                            '',  # L: ‡πÄ‡∏ß‡∏•‡∏≤‡πÇ‡∏´‡∏•‡∏î
                                            '',  # M: ‡∏õ‡∏£‡∏∞‡∏ï‡∏π
                                            '',  # N: WAVE
                                            '',  # O: remark
                                            row.get('Latitude', 0) if pd.notna(row.get('Latitude')) else 0,  # P: lat
                                            row.get('Longitude', 0) if pd.notna(row.get('Longitude')) else 0,  # Q: lon
                                        ]
                                        
                                        for col_idx, value in enumerate(data, 1):
                                            cell = ws.cell(row=row_num, column=col_idx, value=value)
                                            cell.fill = fill
                                            cell.border = thin_border
                                        
                                        row_num += 1
                                        sep_num += 1  # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏•‡∏≥‡∏î‡∏±‡∏ö Sep
                                    
                                    # üöõ ‡πÄ‡∏û‡∏¥‡πà‡∏° DC011 ‡∏õ‡∏¥‡∏î‡∏ó‡πâ‡∏≤‡∏¢‡∏ó‡∏∏‡∏Å‡∏ó‡∏£‡∏¥‡∏õ
                                    dc_data = [
                                        sep_num,  # A: Sep
                                        'PROJECT',  # B: BU
                                        'DC011',  # C: ‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤
                                        'DC011',  # D: ‡∏£‡∏´‡∏±‡∏™ WMS
                                        '‡∏ö.‡∏û‡∏µ‡∏ó‡∏µ‡∏à‡∏µ ‡πÄ‡∏≠‡πá‡∏ô‡πÄ‡∏ô‡∏≠‡∏¢‡∏µ ‡∏à‡∏≥‡∏Å‡∏±‡∏î (‡∏°‡∏´‡∏≤‡∏ä‡∏ô) (DC‡∏ß‡∏±‡∏á‡∏ô‡πâ‡∏≠‡∏¢)',  # E: ‡∏™‡∏≤‡∏Ç‡∏≤
                                        0,  # F: Cube
                                        0,  # G: Weight
                                        0,  # H: Original QTY
                                        int(trip_num),  # I: Trip
                                        trip_no,  # J: Trip no
                                        '',  # K: ‡∏ß‡∏±‡∏ô‡∏ó‡∏µ‡πà‡πÇ‡∏´‡∏•‡∏î
                                        '',  # L: ‡πÄ‡∏ß‡∏•‡∏≤‡πÇ‡∏´‡∏•‡∏î
                                        '',  # M: ‡∏õ‡∏£‡∏∞‡∏ï‡∏π
                                        '',  # N: WAVE
                                        '',  # O: remark
                                        14.2167,  # P: lat (DC ‡∏ß‡∏±‡∏á‡∏ô‡πâ‡∏≠‡∏¢)
                                        100.6167,  # Q: lon (DC ‡∏ß‡∏±‡∏á‡∏ô‡πâ‡∏≠‡∏¢)
                                    ]
                                    
                                    for col_idx, value in enumerate(dc_data, 1):
                                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                                        cell.fill = fill
                                        cell.border = thin_border
                                    
                                    row_num += 1
                                    sep_num += 1
                                
                                # ‡∏ö‡∏±‡∏ô‡∏ó‡∏∂‡∏Å‡∏•‡∏á BytesIO
                                wb.save(output)
                                output.seek(0)
                                
                            except Exception as e:
                                st.warning(f"‚ö†Ô∏è ‡πÑ‡∏°‡πà‡∏™‡∏≤‡∏°‡∏≤‡∏£‡∏ñ‡πÄ‡∏Ç‡∏µ‡∏¢‡∏ô‡∏ó‡∏±‡∏ö‡πÑ‡∏ü‡∏•‡πå‡∏ï‡πâ‡∏ô‡∏â‡∏ö‡∏±‡∏ö‡πÑ‡∏î‡πâ: {e}")
                                # Fallback: ‡∏™‡∏£‡πâ‡∏≤‡∏á‡πÑ‡∏ü‡∏•‡πå‡πÉ‡∏´‡∏°‡πà
                                from openpyxl import Workbook
                                wb = Workbook()
                                ws = wb.active
                                ws.title = '2.Punthai'
                                
                                # ‡πÄ‡∏Ç‡∏µ‡∏¢‡∏ô header
                                headers = ['Sep.', 'BU', '‡∏£‡∏´‡∏±‡∏™‡∏™‡∏≤‡∏Ç‡∏≤', '‡∏£‡∏´‡∏±‡∏™ WMS', '‡∏™‡∏≤‡∏Ç‡∏≤', 'Total Cube', 'Total Wgt', 'Original QTY', 'Trip', 'Trip no']
                                for col_num, header in enumerate(headers, 1):
                                    ws.cell(row=1, column=col_num, value=header)
                                
                                row_num = 2
                                sep_num = 1  # ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏ô‡∏±‡∏ö Sep. ‡∏à‡∏≤‡∏Å 1
                                current_trip = None
                                use_yellow = True
                                yellow_orange = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                                white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                                
                                for trip_num in sorted(result_df['Trip'].unique()):
                                    if trip_num == 0:
                                        continue
                                    trip_data = result_df[result_df['Trip'] == trip_num]
                                    trip_no = trip_no_map.get(trip_num, '')
                                    
                                    if current_trip != trip_num:
                                        current_trip = trip_num
                                        use_yellow = not use_yellow
                                    fill = yellow_orange if use_yellow else white_fill
                                    
                                    for _, row in trip_data.iterrows():
                                        original_qty = row.get('OriginalQty', 0) if pd.notna(row.get('OriginalQty')) else 0
                                        data = [sep_num, row.get('BU', 211), row.get('Code', ''), row.get('Code', ''), row.get('Name', ''),
                                                round(row.get('Cube', 0), 2), round(row.get('Weight', 0), 2), original_qty, int(trip_num), trip_no]
                                        for col_idx, value in enumerate(data, 1):
                                            cell = ws.cell(row=row_num, column=col_idx, value=value)
                                            cell.fill = fill
                                        row_num += 1
                                        sep_num += 1
                                    
                                    # üöõ ‡πÄ‡∏û‡∏¥‡πà‡∏° DC011 ‡∏õ‡∏¥‡∏î‡∏ó‡πâ‡∏≤‡∏¢‡∏ó‡∏∏‡∏Å‡∏ó‡∏£‡∏¥‡∏õ
                                    dc_data = [sep_num, 'PROJECT', 'DC011', 'DC011', '‡∏ö.‡∏û‡∏µ‡∏ó‡∏µ‡∏à‡∏µ ‡πÄ‡∏≠‡πá‡∏ô‡πÄ‡∏ô‡∏≠‡∏¢‡∏µ ‡∏à‡∏≥‡∏Å‡∏±‡∏î (‡∏°‡∏´‡∏≤‡∏ä‡∏ô) (DC‡∏ß‡∏±‡∏á‡∏ô‡πâ‡∏≠‡∏¢)',
                                               0, 0, 0, int(trip_num), trip_no]
                                    for col_idx, value in enumerate(dc_data, 1):
                                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                                        cell.fill = fill
                                    row_num += 1
                                    sep_num += 1
                                
                                wb.save(output)
                                output.seek(0)
                            
                            col1, col2, col3 = st.columns([1, 2, 1])
                            with col2:
                                st.download_button(
                                    label="üì• ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î‡∏ú‡∏•‡∏•‡∏±‡∏û‡∏ò‡πå (Excel)",
                                    data=output.getvalue(),
                                    file_name=f"‡∏ú‡∏•‡∏à‡∏±‡∏î‡∏ó‡∏£‡∏¥‡∏õ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                
                # ==========================================
                # ‡πÅ‡∏ó‡πá‡∏ö 2: ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ï‡∏≤‡∏°‡∏†‡∏≤‡∏Ñ (‡πÑ‡∏°‡πà‡∏™‡∏ô‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å)
                # ==========================================
                with tab2:
                    df_region = df.copy()
                    
                    # ‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ï‡∏≤‡∏°‡∏†‡∏≤‡∏Ñ
                    branch_info = model_data.get('branch_info', {})
                    trip_pairs = model_data.get('trip_pairs', set())
                    
                    # ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏†‡∏≤‡∏Ñ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏™‡∏≤‡∏Ç‡∏≤ (‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥)
                    region_groups = {
                        '‡∏†‡∏≤‡∏Ñ‡∏Å‡∏•‡∏≤‡∏á-‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏ä‡∏±‡πâ‡∏ô‡πÉ‡∏ô': ['‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏°‡∏´‡∏≤‡∏ô‡∏Ñ‡∏£'],
                        '‡∏†‡∏≤‡∏Ñ‡∏Å‡∏•‡∏≤‡∏á-‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏ä‡∏±‡πâ‡∏ô‡∏Å‡∏•‡∏≤‡∏á': ['‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏°‡∏´‡∏≤‡∏ô‡∏Ñ‡∏£'],
                        '‡∏†‡∏≤‡∏Ñ‡∏Å‡∏•‡∏≤‡∏á-‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏ä‡∏±‡πâ‡∏ô‡∏ô‡∏≠‡∏Å': ['‡∏Å‡∏£‡∏∏‡∏á‡πÄ‡∏ó‡∏û‡∏°‡∏´‡∏≤‡∏ô‡∏Ñ‡∏£'],
                        '‡∏†‡∏≤‡∏Ñ‡∏Å‡∏•‡∏≤‡∏á-‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•': ['‡∏ô‡∏Ñ‡∏£‡∏õ‡∏ê‡∏°', '‡∏ô‡∏ô‡∏ó‡∏ö‡∏∏‡∏£‡∏µ', '‡∏õ‡∏ó‡∏∏‡∏°‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏õ‡∏£‡∏≤‡∏Å‡∏≤‡∏£', '‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏™‡∏≤‡∏Ñ‡∏£'],
                        '‡∏†‡∏≤‡∏Ñ‡∏Å‡∏•‡∏≤‡∏á-‡∏Å‡∏•‡∏≤‡∏á‡∏ï‡∏≠‡∏ô‡∏ö‡∏ô': ['‡∏ä‡∏±‡∏¢‡∏ô‡∏≤‡∏ó', '‡∏û‡∏£‡∏∞‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤', '‡∏•‡∏û‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏£‡∏∞‡∏ö‡∏∏‡∏£‡∏µ', '‡∏™‡∏¥‡∏á‡∏´‡πå‡∏ö‡∏∏‡∏£‡∏µ', '‡∏≠‡πà‡∏≤‡∏á‡∏ó‡∏≠‡∏á', '‡∏≠‡∏¢‡∏∏‡∏ò‡∏¢‡∏≤'],
                        '‡∏†‡∏≤‡∏Ñ‡∏Å‡∏•‡∏≤‡∏á-‡∏Å‡∏•‡∏≤‡∏á‡∏ï‡∏≠‡∏ô‡∏•‡πà‡∏≤‡∏á': ['‡∏™‡∏°‡∏∏‡∏ó‡∏£‡∏™‡∏á‡∏Ñ‡∏£‡∏≤‡∏°', '‡∏™‡∏∏‡∏û‡∏£‡∏£‡∏ì‡∏ö‡∏∏‡∏£‡∏µ'],
                        '‡∏†‡∏≤‡∏Ñ‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏ï‡∏Å': ['‡∏Å‡∏≤‡∏ç‡∏à‡∏ô‡∏ö‡∏∏‡∏£‡∏µ', '‡∏õ‡∏£‡∏∞‡∏à‡∏ß‡∏ö‡∏Ñ‡∏µ‡∏£‡∏µ‡∏Ç‡∏±‡∏ô‡∏ò‡πå', '‡∏£‡∏≤‡∏ä‡∏ö‡∏∏‡∏£‡∏µ', '‡πÄ‡∏û‡∏ä‡∏£‡∏ö‡∏∏‡∏£‡∏µ'],
                        '‡∏†‡∏≤‡∏Ñ‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏≠‡∏≠‡∏Å': ['‡∏à‡∏±‡∏ô‡∏ó‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ä‡∏•‡∏ö‡∏∏‡∏£‡∏µ', '‡∏ï‡∏£‡∏≤‡∏î', '‡∏ô‡∏Ñ‡∏£‡∏ô‡∏≤‡∏¢‡∏Å', '‡∏õ‡∏£‡∏≤‡∏à‡∏µ‡∏ô‡∏ö‡∏∏‡∏£‡∏µ', '‡∏£‡∏∞‡∏¢‡∏≠‡∏á', '‡∏™‡∏£‡∏∞‡πÅ‡∏Å‡πâ‡∏ß', '‡∏â‡∏∞‡πÄ‡∏ä‡∏¥‡∏á‡πÄ‡∏ó‡∏£‡∏≤'],
                        '‡∏†‡∏≤‡∏Ñ‡∏≠‡∏µ‡∏™‡∏≤‡∏ô-‡∏≠‡∏µ‡∏™‡∏≤‡∏ô‡πÄ‡∏´‡∏ô‡∏∑‡∏≠': ['‡∏ô‡∏Ñ‡∏£‡∏û‡∏ô‡∏°', '‡∏ö‡∏∂‡∏á‡∏Å‡∏≤‡∏¨', '‡∏°‡∏∏‡∏Å‡∏î‡∏≤‡∏´‡∏≤‡∏£', '‡∏™‡∏Å‡∏•‡∏ô‡∏Ñ‡∏£', '‡∏´‡∏ô‡∏≠‡∏á‡∏Ñ‡∏≤‡∏¢', '‡∏´‡∏ô‡∏≠‡∏á‡∏ö‡∏±‡∏ß‡∏•‡∏≥‡∏†‡∏π', '‡∏≠‡∏∏‡∏î‡∏£‡∏ò‡∏≤‡∏ô‡∏µ', '‡πÄ‡∏•‡∏¢'],
                        '‡∏†‡∏≤‡∏Ñ‡∏≠‡∏µ‡∏™‡∏≤‡∏ô-‡∏≠‡∏µ‡∏™‡∏≤‡∏ô‡∏Å‡∏•‡∏≤‡∏á': ['‡∏Å‡∏≤‡∏¨‡∏™‡∏¥‡∏ô‡∏ò‡∏∏‡πå', '‡∏Ç‡∏≠‡∏ô‡πÅ‡∏Å‡πà‡∏ô', '‡∏ä‡∏±‡∏¢‡∏†‡∏π‡∏°‡∏¥', '‡∏°‡∏´‡∏≤‡∏™‡∏≤‡∏£‡∏Ñ‡∏≤‡∏°', '‡∏£‡πâ‡∏≠‡∏¢‡πÄ‡∏≠‡πá‡∏î'],
                        '‡∏†‡∏≤‡∏Ñ‡∏≠‡∏µ‡∏™‡∏≤‡∏ô-‡∏≠‡∏µ‡∏™‡∏≤‡∏ô‡πÉ‡∏ï‡πâ': ['‡∏ô‡∏Ñ‡∏£‡∏£‡∏≤‡∏ä‡∏™‡∏µ‡∏°‡∏≤', '‡πÇ‡∏Ñ‡∏£‡∏≤‡∏ä', '‡∏ö‡∏∏‡∏£‡∏µ‡∏£‡∏±‡∏°‡∏¢‡πå', '‡∏¢‡πÇ‡∏™‡∏ò‡∏£', '‡∏®‡∏£‡∏µ‡∏™‡∏∞‡πÄ‡∏Å‡∏©', '‡∏™‡∏∏‡∏£‡∏¥‡∏ô‡∏ó‡∏£‡πå', '‡∏≠‡∏≥‡∏ô‡∏≤‡∏à‡πÄ‡∏à‡∏£‡∏¥‡∏ç', '‡∏≠‡∏∏‡∏ö‡∏•‡∏£‡∏≤‡∏ä‡∏ò‡∏≤‡∏ô‡∏µ'],
                        '‡∏†‡∏≤‡∏Ñ‡πÄ‡∏´‡∏ô‡∏∑‡∏≠-‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ï‡∏≠‡∏ô‡∏ö‡∏ô': ['‡∏ô‡πà‡∏≤‡∏ô', '‡∏û‡∏∞‡πÄ‡∏¢‡∏≤', '‡∏•‡∏≥‡∏õ‡∏≤‡∏á', '‡∏•‡∏≥‡∏û‡∏π‡∏ô', '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡∏£‡∏≤‡∏¢', '‡πÄ‡∏ä‡∏µ‡∏¢‡∏á‡πÉ‡∏´‡∏°‡πà', '‡πÅ‡∏û‡∏£‡πà', '‡πÅ‡∏°‡πà‡∏Æ‡πà‡∏≠‡∏á‡∏™‡∏≠‡∏ô'],
                        '‡∏†‡∏≤‡∏Ñ‡πÄ‡∏´‡∏ô‡∏∑‡∏≠-‡πÄ‡∏´‡∏ô‡∏∑‡∏≠‡∏ï‡∏≠‡∏ô‡∏•‡πà‡∏≤‡∏á': ['‡∏Å‡∏≥‡πÅ‡∏û‡∏á‡πÄ‡∏û‡∏ä‡∏£', '‡∏ï‡∏≤‡∏Å', '‡∏ô‡∏Ñ‡∏£‡∏™‡∏ß‡∏£‡∏£‡∏Ñ‡πå', '‡∏û‡∏¥‡∏à‡∏¥‡∏ï‡∏£', '‡∏û‡∏¥‡∏©‡∏ì‡∏∏‡πÇ‡∏•‡∏Å', '‡∏™‡∏∏‡πÇ‡∏Ç‡∏ó‡∏±‡∏¢', '‡∏≠‡∏∏‡∏ï‡∏£‡∏î‡∏¥‡∏ï‡∏ñ‡πå', '‡∏≠‡∏∏‡∏ó‡∏±‡∏¢‡∏ò‡∏≤‡∏ô‡∏µ', '‡πÄ‡∏û‡∏ä‡∏£‡∏ö‡∏π‡∏£‡∏ì‡πå'],
                        '‡∏†‡∏≤‡∏Ñ‡πÉ‡∏ï‡πâ-‡πÉ‡∏ï‡πâ‡∏ù‡∏±‡πà‡∏á‡∏≠‡∏±‡∏ô‡∏î‡∏≤‡∏°‡∏±‡∏ô': ['‡∏Å‡∏£‡∏∞‡∏ö‡∏µ‡πà', '‡∏ï‡∏£‡∏±‡∏á', '‡∏û‡∏±‡∏á‡∏á‡∏≤', '‡∏†‡∏π‡πÄ‡∏Å‡πá‡∏ï', '‡∏£‡∏∞‡∏ô‡∏≠‡∏á', '‡∏™‡∏ï‡∏π‡∏•'],
                        '‡∏†‡∏≤‡∏Ñ‡πÉ‡∏ï‡πâ-‡πÉ‡∏ï‡πâ‡∏ù‡∏±‡πà‡∏á‡∏≠‡πà‡∏≤‡∏ß‡πÑ‡∏ó‡∏¢': ['‡∏ä‡∏∏‡∏°‡∏û‡∏£', '‡∏ô‡∏Ñ‡∏£‡∏®‡∏£‡∏µ‡∏ò‡∏£‡∏£‡∏°‡∏£‡∏≤‡∏ä', '‡∏û‡∏±‡∏ó‡∏•‡∏∏‡∏á', '‡∏¢‡∏∞‡∏•‡∏≤', '‡∏™‡∏á‡∏Ç‡∏•‡∏≤', '‡∏™‡∏∏‡∏£‡∏≤‡∏©‡∏é‡∏£‡πå‡∏ò‡∏≤‡∏ô‡∏µ', '‡∏õ‡∏±‡∏ï‡∏ï‡∏≤‡∏ô‡∏µ', '‡∏ô‡∏£‡∏≤‡∏ò‡∏¥‡∏ß‡∏≤‡∏™']
                    }
                    
                    def get_region(province):
                        if pd.isna(province) or not province or str(province).strip() in ['', 'nan', 'UNKNOWN']:
                            return '‡πÑ‡∏°‡πà‡∏£‡∏∞‡∏ö‡∏∏'
                        
                        # üö® Override: ‡∏â‡∏∞‡πÄ‡∏ä‡∏¥‡∏á‡πÄ‡∏ó‡∏£‡∏≤ ‚Üí ‡∏†‡∏≤‡∏Ñ‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏≠‡∏≠‡∏Å (‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πà‡∏õ‡∏£‡∏¥‡∏°‡∏ì‡∏ë‡∏•)
                        if '‡∏â‡∏∞‡πÄ‡∏ä‡∏¥‡∏á‡πÄ‡∏ó‡∏£‡∏≤' in str(province):
                            return '‡∏†‡∏≤‡∏Ñ‡∏ï‡∏∞‡∏ß‡∏±‡∏ô‡∏≠‡∏≠‡∏Å'
                        
                        for region, provinces in region_groups.items():
                            if any(p in str(province) for p in provinces):
                                return region
                        return '‡∏≠‡∏∑‡πà‡∏ô‡πÜ'
                    
                    # ‡πÄ‡∏û‡∏¥‡πà‡∏°‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå‡∏†‡∏≤‡∏Ñ - ‡∏î‡∏∂‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏à‡∏≤‡∏Å Master ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ
                    if 'Province' not in df_region.columns or df_region['Province'].isna().any():
                        # ‡∏î‡∏∂‡∏á‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏à‡∏≤‡∏Å Master
                        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                            province_map = {}
                            for _, row in MASTER_DATA.iterrows():
                                code = row.get('Plan Code', '')
                                province = row.get('‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', '')
                                if code and province:
                                    province_map[code] = province
                            
                            # ‡πÉ‡∏™‡πà‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÉ‡∏´‡πâ‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏™‡∏≤‡∏Ç‡∏≤
                            if 'Province' not in df_region.columns:
                                df_region['Province'] = df_region['Code'].map(province_map)
                            else:
                                # ‡πÄ‡∏ï‡∏¥‡∏°‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏ó‡∏µ‡πà‡πÄ‡∏õ‡πá‡∏ô NaN
                                df_region['Province'] = df_region.apply(
                                    lambda row: province_map.get(row['Code'], row.get('Province', 'UNKNOWN')) 
                                    if pd.isna(row.get('Province')) else row['Province'],
                                    axis=1
                                )
                    
                    df_region['Region'] = df_region['Province'].apply(get_region)
                    
                    # ‡∏´‡∏≤‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤ (‡πÉ‡∏ä‡πâ Booking No. ‡πÄ‡∏õ‡πá‡∏ô‡∏´‡∏•‡∏±‡∏Å)
                    def find_paired_branches(code, code_province, df_data):
                        paired = set()
                        
                        # ‡∏´‡∏≤ Booking No. ‡∏Ç‡∏≠‡∏á‡∏™‡∏≤‡∏Ç‡∏≤‡∏ô‡∏µ‡πâ
                        code_rows = df_data[df_data['Code'] == code]
                        if len(code_rows) == 0:
                            return paired
                        
                        # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏°‡∏µ‡∏Ñ‡∏≠‡∏•‡∏±‡∏°‡∏ô‡πå Booking ‡∏´‡∏£‡∏∑‡∏≠‡πÑ‡∏°‡πà
                        if 'Booking' not in df_data.columns and 'Trip' not in df_data.columns:
                            return paired
                        
                        booking_col = 'Booking' if 'Booking' in df_data.columns else 'Trip'
                        code_bookings = set(code_rows[booking_col].dropna().astype(str))
                        
                        if not code_bookings:
                            return paired
                        
                        # ‡∏´‡∏≤‡∏™‡∏≤‡∏Ç‡∏≤‡∏≠‡∏∑‡πà‡∏ô‡∏ó‡∏µ‡πà‡∏≠‡∏¢‡∏π‡πà Booking ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô (‡πÑ‡∏°‡πà‡∏™‡∏ô‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î)
                        for booking in code_bookings:
                            if booking == 'nan' or not booking.strip():
                                continue
                            
                            same_booking = df_data[df_data[booking_col].astype(str) == booking]
                            for _, other_row in same_booking.iterrows():
                                other_code = other_row['Code']
                                
                                # ‡πÄ‡∏á‡∏∑‡πà‡∏≠‡∏ô‡πÑ‡∏Ç: Booking ‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô = ‡∏£‡∏ß‡∏°‡∏Å‡∏•‡∏∏‡πà‡∏° (‡πÑ‡∏°‡πà‡∏™‡∏ô‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î)
                                if other_code != code:
                                    paired.add(other_code)
                        
                        return paired
                    
                    all_codes_set = set(df_region['Code'].unique())
                    
                    # ‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡πÅ‡∏ö‡∏ö Union-Find (‡∏ï‡∏≤‡∏°‡∏•‡∏≥‡∏î‡∏±‡∏ö: ‡∏ï‡∏≥‡∏ö‡∏• ‚Üí ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠ ‚Üí ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î)
                    # Step 1: ‡πÄ‡∏£‡∏¥‡πà‡∏°‡∏à‡∏≤‡∏Å‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏™‡∏≤‡∏Ç‡∏≤‡πÄ‡∏õ‡πá‡∏ô‡∏Å‡∏•‡∏∏‡πà‡∏°‡πÜ ‡∏û‡∏£‡πâ‡∏≠‡∏°‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• Master
                    initial_groups = {}
                    for code in all_codes_set:
                        # ‡∏î‡∏∂‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏à‡∏≤‡∏Å Master
                        location = {}
                        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                            if len(master_row) > 0:
                                master_row = master_row.iloc[0]
                                location = {
                                    'subdistrict': master_row.get('‡∏ï‡∏≥‡∏ö‡∏•', ''),
                                    'district': master_row.get('‡∏≠‡∏≥‡πÄ‡∏†‡∏≠', ''),
                                    'province': master_row.get('‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', 'UNKNOWN'),
                                    'lat': master_row.get('‡∏•‡∏∞‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0),
                                    'lon': master_row.get('‡∏•‡∏≠‡∏á‡∏ï‡∏¥‡∏à‡∏π‡∏î', 0)
                                }
                        
                        # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡πÉ‡∏ô Master ‡∏•‡∏≠‡∏á‡∏î‡∏∂‡∏á‡∏à‡∏≤‡∏Å‡πÑ‡∏ü‡∏•‡πå‡∏≠‡∏±‡∏õ‡πÇ‡∏´‡∏•‡∏î
                        if not location or location.get('province', 'UNKNOWN') == 'UNKNOWN':
                            c_row = df_region[df_region['Code'] == code].iloc[0] if len(df_region[df_region['Code'] == code]) > 0 else None
                            if c_row is not None:
                                location = {
                                    'subdistrict': '',
                                    'district': '',
                                    'province': c_row.get('Province', 'UNKNOWN'),
                                    'lat': 0,
                                    'lon': 0
                                }
                        
                        if location:
                            initial_groups[(code,)] = {code: location}
                    
                    # ‡πÉ‡∏ä‡πâ initial_groups ‡πÅ‡∏ó‡∏ô booking_groups
                    booking_groups = initial_groups
                    
                    # Step 2: ‡∏£‡∏ß‡∏°‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏ï‡∏≤‡∏°‡∏•‡∏≥‡∏î‡∏±‡∏ö ‡∏ï‡∏≥‡∏ö‡∏• ‚Üí ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠ ‚Üí ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î
                    def groups_can_merge(locs1, locs2):
                        """‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤ 2 ‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏Ñ‡∏ß‡∏£‡∏£‡∏ß‡∏°‡∏Å‡∏±‡∏ô‡πÑ‡∏´‡∏° (‡∏ï‡∏≤‡∏°‡∏•‡∏≥‡∏î‡∏±‡∏ö‡∏Ñ‡∏ß‡∏≤‡∏°‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î)"""
                        # 1. ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ï‡∏≥‡∏ö‡∏•‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô (‡∏ï‡πâ‡∏≠‡∏á‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ï‡∏≥‡∏ö‡∏•)
                        subdistricts1 = set(loc.get('subdistrict', '') for loc in locs1.values() if loc.get('subdistrict', ''))
                        subdistricts2 = set(loc.get('subdistrict', '') for loc in locs2.values() if loc.get('subdistrict', ''))
                        if subdistricts1 and subdistricts2 and (subdistricts1 & subdistricts2):
                            return True, '‡∏ï‡∏≥‡∏ö‡∏•'
                        
                        # 2. ‡πÄ‡∏ä‡πá‡∏Ñ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô (‡∏ï‡πâ‡∏≠‡∏á‡∏°‡∏µ‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡πÅ‡∏•‡∏∞‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô)
                        districts1 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs1.values() if loc.get('district', '')}
                        districts2 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs2.values() if loc.get('district', '')}
                        if districts1 and districts2:
                            # ‡πÄ‡∏ä‡πá‡∏Ñ‡∏ß‡πà‡∏≤‡∏°‡∏µ‡∏≠‡∏≥‡πÄ‡∏†‡∏≠‡πÅ‡∏•‡∏∞‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏ï‡∏£‡∏á‡∏Å‡∏±‡∏ô
                            for d1, p1 in districts1:
                                for d2, p2 in districts2:
                                    if d1 == d2 and p1 == p2 and p1:
                                        return True, '‡∏≠‡∏≥‡πÄ‡∏†‡∏≠'
                        
                        # 3. ‡πÄ‡∏ä‡πá‡∏Ñ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏î‡∏µ‡∏¢‡∏ß‡∏Å‡∏±‡∏ô
                        provinces1 = set(loc.get('province', '') for loc in locs1.values() if loc.get('province', ''))
                        provinces2 = set(loc.get('province', '') for loc in locs2.values() if loc.get('province', ''))
                        if provinces1 & provinces2:
                            return True, '‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î'
                        
                        return False, None
                    
                    merged_groups = []
                    used_groups = set()
                    
                    for group1, locs1 in booking_groups.items():
                        if group1 in used_groups:
                            continue
                        
                        merged_codes = set(group1)
                        merged_locs = locs1.copy()
                        used_groups.add(group1)
                        
                        # ‡∏´‡∏≤‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏≠‡∏∑‡πà‡∏ô‡∏ó‡∏µ‡πà‡πÉ‡∏Å‡∏•‡πâ‡πÄ‡∏Ñ‡∏µ‡∏¢‡∏á
                        changed = True
                        while changed:
                            changed = False
                            for group2, locs2 in booking_groups.items():
                                if group2 in used_groups:
                                    continue
                                can_merge, level = groups_can_merge(merged_locs, locs2)
                                if can_merge:
                                    merged_codes |= set(group2)
                                    merged_locs.update(locs2)
                                    used_groups.add(group2)
                                    changed = True
                        
                        merged_groups.append({
                            'codes': merged_codes,
                            'locations': merged_locs
                        })
                    
                    # Step 3: ‡πÅ‡∏õ‡∏•‡∏á‡πÄ‡∏õ‡πá‡∏ô groups format
                    groups = []
                    for mg in merged_groups:
                        rep_code = list(mg['codes'])[0]
                        rep_row = df_region[df_region['Code'] == rep_code].iloc[0]
                        # ‡∏Å‡∏£‡∏≠‡∏á‡πÄ‡∏â‡∏û‡∏≤‡∏∞‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡∏ó‡∏µ‡πà‡πÑ‡∏°‡πà‡πÉ‡∏ä‡πà UNKNOWN ‡πÅ‡∏•‡∏∞‡πÑ‡∏°‡πà‡πÄ‡∏õ‡πá‡∏ô NaN
                        provinces = set(
                            str(loc.get('province', '')).strip() 
                            for loc in mg['locations'].values() 
                            if loc.get('province') and str(loc.get('province', '')).strip() not in ['UNKNOWN', 'nan', '']
                        )
                        
                        # ‡∏ñ‡πâ‡∏≤‡πÑ‡∏°‡πà‡∏°‡∏µ‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î‡πÄ‡∏•‡∏¢ ‡πÉ‡∏™‡πà "‡πÑ‡∏°‡πà‡∏£‡∏∞‡∏ö‡∏∏"
                        province_str = ', '.join(sorted(provinces)) if provinces else '‡πÑ‡∏°‡πà‡∏£‡∏∞‡∏ö‡∏∏'
                        
                        groups.append({
                            'codes': mg['codes'],
                            'region': rep_row.get('Region', '‡πÑ‡∏°‡πà‡∏£‡∏∞‡∏ö‡∏∏'),
                            'province': province_str
                        })
                    
                    # ‡πÅ‡∏™‡∏î‡∏á‡∏™‡∏ñ‡∏¥‡∏ï‡∏¥
                    st.markdown("---")
                    st.markdown("### üìä ‡∏™‡∏£‡∏∏‡∏õ‡∏Å‡∏≤‡∏£‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("üìç ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤", df_region['Code'].nunique())
                    with col2:
                        st.metric("üóÇÔ∏è ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏Å‡∏•‡∏∏‡πà‡∏°", len(groups))
                    with col3:
                        regions_count = df_region['Region'].nunique()
                        st.metric("üó∫Ô∏è ‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏†‡∏≤‡∏Ñ", regions_count)
                    
                    # ‡πÅ‡∏™‡∏î‡∏á‡∏ï‡∏≤‡∏°‡∏†‡∏≤‡∏Ñ
                    st.markdown("---")
                    st.markdown("### üó∫Ô∏è ‡∏™‡∏≤‡∏Ç‡∏≤‡πÅ‡∏¢‡∏Å‡∏ï‡∏≤‡∏°‡∏†‡∏≤‡∏Ñ")
                    
                    region_summary = df_region.groupby('Region').agg({
                        'Code': 'nunique',
                        'Weight': 'sum',
                        'Cube': 'sum'
                    }).reset_index()
                    region_summary.columns = ['‡∏†‡∏≤‡∏Ñ', '‡∏à‡∏≥‡∏ô‡∏ß‡∏ô‡∏™‡∏≤‡∏Ç‡∏≤', '‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å‡∏£‡∏ß‡∏°', '‡∏Ñ‡∏¥‡∏ß‡∏£‡∏ß‡∏°']
                    st.dataframe(region_summary, use_container_width=True)
                    
                    # ‡πÅ‡∏™‡∏î‡∏á‡∏£‡∏≤‡∏¢‡∏•‡∏∞‡πÄ‡∏≠‡∏µ‡∏¢‡∏î‡πÅ‡∏ï‡πà‡∏•‡∏∞‡∏†‡∏≤‡∏Ñ
                    for region in sorted(df_region['Region'].unique()):
                        region_data = df_region[df_region['Region'] == region]
                        with st.expander(f"üìç {region} ({region_data['Code'].nunique()} ‡∏™‡∏≤‡∏Ç‡∏≤)"):
                            display_cols = ['Code', 'Name', 'Province', 'Weight', 'Cube']
                            display_cols = [c for c in display_cols if c in region_data.columns]
                            
                            region_display = region_data[display_cols].drop_duplicates('Code')
                            col_names = {'Code': '‡∏£‡∏´‡∏±‡∏™', 'Name': '‡∏ä‡∏∑‡πà‡∏≠‡∏™‡∏≤‡∏Ç‡∏≤', 'Province': '‡∏à‡∏±‡∏á‡∏´‡∏ß‡∏±‡∏î', 'Weight': '‡∏ô‡πâ‡∏≥‡∏´‡∏ô‡∏±‡∏Å', 'Cube': '‡∏Ñ‡∏¥‡∏ß'}
                            region_display.columns = [col_names.get(c, c) for c in display_cols]
                            st.dataframe(region_display, use_container_width=True)
                    
                    # ‡πÅ‡∏™‡∏î‡∏á‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏Ñ‡∏¢‡πÑ‡∏õ‡∏î‡πâ‡∏ß‡∏¢‡∏Å‡∏±‡∏ô
                    st.markdown("---")
                    st.markdown("### üîó ‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏Ñ‡∏¢‡πÑ‡∏õ‡∏î‡πâ‡∏ß‡∏¢‡∏Å‡∏±‡∏ô (‡∏à‡∏≤‡∏Å‡∏õ‡∏£‡∏∞‡∏ß‡∏±‡∏ï‡∏¥)")
                    
                    paired_groups = [g for g in groups if len(g['codes']) > 1]
                    if paired_groups:
                        for i, group in enumerate(paired_groups, 1):
                            codes_list = list(group['codes'])
                            names = []
                            for c in codes_list:
                                name_row = df_region[df_region['Code'] == c]
                                if len(name_row) > 0 and 'Name' in name_row.columns:
                                    names.append(f"{c} ({name_row['Name'].iloc[0]})")
                                else:
                                    names.append(c)
                            
                            st.write(f"**‡∏Å‡∏•‡∏∏‡πà‡∏° {i}** - {group['region']}: {', '.join(names)}")
                    else:
                        st.info("‡πÑ‡∏°‡πà‡∏û‡∏ö‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏µ‡πà‡πÄ‡∏Ñ‡∏¢‡πÑ‡∏õ‡∏î‡πâ‡∏ß‡∏¢‡∏Å‡∏±‡∏ô‡πÉ‡∏ô‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£‡∏ô‡∏µ‡πâ")
                    
                    # ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î
                    st.markdown("---")
                    output_region = io.BytesIO()
                    with pd.ExcelWriter(output_region, engine='xlsxwriter') as writer:
                        df_region.to_excel(writer, sheet_name='‡∏™‡∏≤‡∏Ç‡∏≤‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î', index=False)
                        region_summary.to_excel(writer, sheet_name='‡∏™‡∏£‡∏∏‡∏õ‡∏ï‡∏≤‡∏°‡∏†‡∏≤‡∏Ñ', index=False)
                    
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        st.download_button(
                            label="üì• ‡∏î‡∏≤‡∏ß‡∏ô‡πå‡πÇ‡∏´‡∏•‡∏î‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏° (Excel)",
                            data=output_region.getvalue(),
                            file_name=f"‡∏à‡∏±‡∏î‡∏Å‡∏•‡∏∏‡πà‡∏°‡∏™‡∏≤‡∏Ç‡∏≤_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

if __name__ == "__main__":
    main()
