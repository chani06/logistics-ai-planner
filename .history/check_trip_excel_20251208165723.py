import openpyxl
import sys
import io
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# เปิดไฟล์ Excel
wb = openpyxl.load_workbook('Dc/แผนงาน Punthai Maxmart รอบสั่ง 24หยิบ 25พฤศจิกายน 2568 To.เฟิ(1) - สำเนา.xlsx')
ws = wb['2.Punthai']

print('=' * 80)
print('ตรวจสอบคอลัมน์ Trip (คอลัมน์ที่ 9)')
print('=' * 80)

# อ่านค่า Trip จากคอลัมน์ที่ 9 (เริ่มจากแถวที่ 3)
trips = []
codes = []
branches = []
provinces = []

for row_idx in range(3, ws.max_row + 1):
    trip_value = ws.cell(row=row_idx, column=9).value
    code_value = ws.cell(row=row_idx, column=3).value
    branch_value = ws.cell(row=row_idx, column=4).value
    province_value = ws.cell(row=row_idx, column=5).value
    
    if trip_value is not None:
        trips.append(int(trip_value))
        codes.append(str(code_value) if code_value else '')
        branches.append(str(branch_value) if branch_value else '')
        provinces.append(str(province_value) if province_value else '')

print(f'\n📊 จำนวนแถวทั้งหมด: {ws.max_row - 2}')
print(f'✅ จำนวนที่มี Trip: {len(trips)}')
print(f'❌ จำนวนที่ไม่มี Trip: {ws.max_row - 2 - len(trips)}')

# ดึง Trip ที่ไม่ซ้ำ
unique_trips = sorted(list(set(trips)))

print(f'\n🚛 จำนวน Trip ทั้งหมด: {len(unique_trips)}')
print(f'📝 Trip เลขที่: {unique_trips[:30]}')
if len(unique_trips) > 30:
    print(f'   ... และอีก {len(unique_trips) - 30} ทริป')

# ตรวจสอบความต่อเนื่อง
print(f'\n🔍 ตรวจสอบความต่อเนื่อง:')
print(f'   Trip เริ่มต้น: {min(unique_trips)}')
print(f'   Trip สุดท้าย: {max(unique_trips)}')

missing = [i for i in range(min(unique_trips), max(unique_trips)+1) if i not in unique_trips]
if missing:
    print(f'⚠️  Trip ที่ขาดหาย ({len(missing)} ทริป): {missing[:30]}')
    if len(missing) > 30:
        print(f'   ... และอีก {len(missing) - 30} ทริป')
else:
    print(f'✅ ไม่มี Trip ขาดหาย - ต่อเนื่องทุกหมายเลข')

# นับจำนวนสาขาในแต่ละ Trip
trip_counts = Counter(trips)

print(f'\n🔢 จำนวนสาขาในแต่ละ Trip (10 ทริปแรก):')
for trip in sorted(trip_counts.keys())[:10]:
    count = trip_counts[trip]
    print(f'   Trip {trip:3d}: {count:2d} สาขา')

# แสดงตัวอย่างข้อมูล
print(f'\n📋 ตัวอย่างข้อมูล (5 แถวแรก):')
for i in range(min(5, len(trips))):
    print(f'   Trip {trips[i]:3d} | {codes[i]:10s} | {branches[i][:30]:30s} | {provinces[i]}')

# เปรียบเทียบกับผลจาก test_planner_v2.py
print(f'\n📊 สรุปเปรียบเทียบ:')
print(f'   - ตามโปรแกรม: 78 trips (จาก output)')
print(f'   - ในไฟล์ Excel: {len(unique_trips)} trips')
print(f'   - Trip ขาดหาย: {len(missing)} trips')
if len(unique_trips) == 78:
    print(f'   ✅ ตรงกัน!')
else:
    print(f'   ⚠️  ไม่ตรงกัน!')
