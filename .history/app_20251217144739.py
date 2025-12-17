"""
Logistics Planner 
"""

import streamlit as st
import pandas as pd
import numpy as np
import pickle
import os
import glob
from datetime import datetime, time
import io

# Fuzzy String Matching
try:
    from rapidfuzz import fuzz, process
    FUZZY_AVAILABLE = True
except ImportError:
    FUZZY_AVAILABLE = False
    # Fallback: ใช้ difflib
    from difflib import SequenceMatcher

# Auto-refresh component
try:
    from streamlit_autorefresh import st_autorefresh
    AUTOREFRESH_AVAILABLE = True
except ImportError:
    AUTOREFRESH_AVAILABLE = False
    # Skip warning - streamlit-autorefresh is optional

# ==========================================
# CONFIG
# ==========================================
MODEL_PATH = 'models/decision_tree_model.pkl'

# ขีดจำกัดรถแต่ละประเภท
LIMITS = {
    '4W': {'max_w': 2500, 'max_c': 5.0},   # ไม่เกิน 12 จุด, Cube ≤ 5 (Punthai ล้วน)
    'JB': {'max_w': 3500, 'max_c': 7.0},   # ไม่เกิน 12 จุด, Cube ≤ 7
    '6W': {'max_w': 6000, 'max_c': 20.0}   # ไม่จำกัดจุด, Cube ต้องเต็ม, Weight ≤ 6000
}

# 🔒 ขีดจำกัดสำหรับ Punthai ล้วน
# - JB (Jumbo): ไม่เกิน 7 drop, Cube ≤ 7
# - 4W: ถ้า Cube > 5 → ตัดเป็น JB
PUNTHAI_LIMITS = {
    '4W': {'max_w': 2500, 'max_c': 5.0, 'max_drops': 12},  # Punthai ล้วน: ถ้าเกิน 5 cube → ใช้ JB
    'JB': {'max_w': 3500, 'max_c': 7.0, 'max_drops': 7},   # Punthai ล้วน: ไม่เกิน 7 drop, Cube ≤ 7
    '6W': {'max_w': 6000, 'max_c': 20.0, 'max_drops': 999}
}

# 🚨 ห้ามเกิน 100% - ไม่มี Buffer
BUFFER = 1.0

# 🚨 จำกัดระยะห่างสาขาในทริปเดียวกัน (consecutive distance)
MAX_DISTANCE_IN_TRIP = 50  # km - สาขาในทริปเดียวกันห้ามห่างกันเกิน 50km (สำหรับกรุงเทพ/ปริมณฑล)
MAX_DISTANCE_IN_TRIP_FAR = 100  # km - สำหรับต่างจังหวัดไกล (6W) ยอมให้เกินได้

# 🎯 Minimum utilization ต่อประเภทรถ (สำหรับ balancing)
MIN_UTIL = {
    '4W': 70,   # 4W ต้องใช้อย่างน้อย 70%
    'JB': 80,   # JB ต้องใช้อย่างน้อย 80%
    '6W': 90    # 6W ต้องใช้อย่างน้อย 90%
}

# จำนวนสาขาต่อทริป - ใช้กับ 4W/JB เท่านั้น (6W ไม่จำกัด)
MAX_BRANCHES_PER_TRIP = 12  # สูงสุด 12 สาขาต่อทริปสำหรับ 4W/JB (6W ไม่จำกัด)

# Performance Config - Optimized for < 1 minute
MAX_DETOUR_KM = 10  # ลดจาก 12km เป็น 10km
MAX_MERGE_ITERATIONS = 5  # ลดจาก 10 เป็น 5 เพื่อเร็วขึ้น
MAX_REBALANCE_ITERATIONS = 3  # จำกัดการ rebalance
MAX_PROCESSING_TIME = 55  # วินาที - เป้าหมาย < 1 minute
EARLY_STOP_UTIL = 95  # หยุดถ้าได้ utilization >= 95%
EARLY_STOP_THRESHOLD = 0.95  # หยุดถ้าได้ utilization >= 95%

# รายการสาขาที่ไม่ต้องการจัดส่ง (ตัดออก)
EXCLUDE_BRANCHES = ['DC011', 'PTDC', 'PTG DISTRIBUTION CENTER']

# รายชื่อที่ต้องตัดออก (ใช้ตรวจสอบชื่อ)
EXCLUDE_NAMES = ['Distribution Center', 'PTG Distribution', 'บ.พีทีจี เอ็นเนอยี']

# 🔒 ข้อจำกัดรถจากไฟล์ Auto Plan (ชีต info) - MaxTruckType
# จะถูก populate จากไฟล์ที่อัปโหลด
AUTO_PLAN_TRUCK_LIMITS = {}  # {branch_code: max_truck_type} เช่น {'11005514': '4W', 'G015': 'JB'}

# พิกัด DC วังน้อย (จุดกลาง)
DC_WANG_NOI_LAT = 14.179394
DC_WANG_NOI_LON = 100.648149

# ระยะทางที่ต้องใช้รถ 6W (กม.)
DISTANCE_REQUIRE_6W = 100  # ถ้าห่างจาก DC เกิน 100 กม. ต้องใช้ 6W

# ==========================================
# LOAD MASTER DATA
# ==========================================
@st.cache_data(ttl=7200)  # Cache 2 ชั่วโมง (เร็วขึ้น)
def load_master_data():
    """โหลดไฟล์ Master สถานที่ส่ง (Optimized)"""
    try:
        # โหลดเฉพาะคอลัมน์ที่จำเป็น
        usecols = ['Plan Code', 'ตำบล', 'อำเภอ', 'จังหวัด', 'ละติจูด', 'ลองติจูด']
        df_master = pd.read_excel('Dc/Master สถานที่ส่ง.xlsx', usecols=usecols)
        # ทำความสะอาด Plan Code (vectorized)
        if 'Plan Code' in df_master.columns:
            df_master['Plan Code'] = df_master['Plan Code'].astype(str).str.strip().str.upper()
        # สร้าง dict สำหรับค้นหาเร็ว
        df_master = df_master[df_master['Plan Code'] != '']
        return df_master
    except FileNotFoundError:
        return pd.DataFrame()
    except Exception as e:
        st.warning(f"ไม่สามารถโหลดไฟล์ Master: {e} (จะใช้ข้อมูลจากไฟล์อัปโหลดแทน)")
        return pd.DataFrame()

# โหลด Master Data
MASTER_DATA = load_master_data()

# ==========================================
# 🆕 โหลดไฟล์ สถานที่ส่ง.xlsx สำหรับจับกลุ่มสาขาที่อยู่ที่เดียวกัน (Reference)
# ==========================================
@st.cache_data(ttl=7200)
def load_location_reference():
    """โหลดไฟล์ สถานที่ส่ง.xlsx และสร้าง Reference mapping"""
    try:
        df = pd.read_excel('Dc/สถานที่ส่ง.xlsx')
        if 'Reference' in df.columns and 'Plan Code' in df.columns:
            # สร้าง mapping: branch_code -> reference
            code_to_ref = {}
            # สร้าง reverse mapping: reference -> [branch_codes]
            ref_to_codes = {}
            
            for _, row in df.iterrows():
                code = str(row['Plan Code']).strip().upper()
                ref = str(row['Reference']).strip()
                
                if code and ref and code != 'NAN' and ref != 'NAN':
                    code_to_ref[code] = ref
                    if ref not in ref_to_codes:
                        ref_to_codes[ref] = []
                    ref_to_codes[ref].append(code)
            
            return code_to_ref, ref_to_codes
        return {}, {}
    except Exception as e:
        return {}, {}

# โหลด Reference mapping
LOCATION_CODE_TO_REF, LOCATION_REF_TO_CODES = load_location_reference()

@st.cache_data(ttl=3600)  # Cache 1 ชั่วโมง
def load_booking_history_restrictions():
    """โหลดประวัติการจัดส่งจาก Booking History - ข้อมูลจริง 3,053 booking (Optimized)"""
    try:
        # ลองหาไฟล์ Booking History (อาจมีชื่อหลายแบบ)
        possible_files = [
            'Dc/ประวัติงานจัดส่ง DC วังน้อย(1).xlsx',
            'Dc/ประวัติงานจัดส่ง DC วังน้อย.xlsx',
            'branch_vehicle_restrictions_from_booking.xlsx'
        ]
        
        file_path = None
        for path in possible_files:
            if os.path.exists(path):
                file_path = path
                break
        
        if not file_path:
            # ถ้าไม่มีไฟล์ ใช้ข้อมูลที่เคยเรียนรู้ (fallback)
            return load_learned_restrictions_fallback()
        
        df = pd.read_excel(file_path)
        
        # แปลงประเภทรถ
        vehicle_mapping = {
            '4 ล้อ จัมโบ้ ตู้ทึบ': 'JB',
            '6 ล้อ ตู้ทึบ': '6W',
            '4 ล้อ ตู้ทึบ': '4W'
        }
        df['Vehicle_Type'] = df['ประเภทรถ'].map(vehicle_mapping)
        
        # วิเคราะห์ความสัมพันธ์สาขา-รถ
        branch_vehicle_history = {}
        booking_groups = df.groupby('Booking No')
        
        for booking_no, booking_data in booking_groups:
            vehicle_types = booking_data['Vehicle_Type'].dropna().unique()
            if len(vehicle_types) > 0:
                vehicle = booking_data['Vehicle_Type'].mode()[0] if len(booking_data['Vehicle_Type'].mode()) > 0 else vehicle_types[0]
                for branch_code in booking_data['รหัสสาขา'].dropna().unique():
                    if branch_code not in branch_vehicle_history:
                        branch_vehicle_history[branch_code] = []
                    branch_vehicle_history[branch_code].append(vehicle)
        
        # สร้าง restrictions
        branch_restrictions = {}
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for branch_code, vehicle_list in branch_vehicle_history.items():
            vehicles_used = set(vehicle_list)
            vehicle_counts = pd.Series(vehicle_list).value_counts().to_dict()
            
            if len(vehicles_used) == 1:
                # STRICT - ใช้รถเดียว
                vehicle = list(vehicles_used)[0]
                branch_restrictions[str(branch_code)] = {
                    'max_vehicle': vehicle,
                    'allowed': [vehicle],
                    'total_bookings': len(vehicle_list),
                    'restriction_type': 'STRICT'
                }
            else:
                # FLEXIBLE - ใช้ได้หลายประเภท
                max_vehicle = max(vehicles_used, key=lambda v: vehicle_sizes.get(v, 0))
                branch_restrictions[str(branch_code)] = {
                    'max_vehicle': max_vehicle,
                    'allowed': list(vehicles_used),
                    'total_bookings': len(vehicle_list),
                    'restriction_type': 'FLEXIBLE'
                }
        
        stats = {
            'total_branches': len(branch_restrictions),
            'strict': len([b for b, r in branch_restrictions.items() if r['restriction_type'] == 'STRICT']),
            'flexible': len([b for b, r in branch_restrictions.items() if r['restriction_type'] == 'FLEXIBLE']),
            'total_bookings': len(booking_groups)
        }
        
        return {
            'branch_restrictions': branch_restrictions,
            'stats': stats
        }
    except Exception as e:
        # ถ้าเกิด error ใช้ข้อมูลที่เคยเรียนรู้แทน
        return load_learned_restrictions_fallback()

def load_learned_restrictions_fallback():
    """
    ข้อมูลที่เรียนรู้จาก Booking History (backup)
    ใช้เมื่อไม่สามารถโหลดไฟล์ได้
    
    จากการวิเคราะห์ 3,053 bookings, 2,790 สาขา:
    - JB: รถกลาง (ใช้มากที่สุด 54.7%)
    - 6W: รถใหญ่ (30.1%)
    - 4W: รถเล็ก (0.2%)
    
    กลยุทธ์: ถ้าไม่มีข้อมูล default เป็น JB (รถกลาง ใช้ได้กับสาขาส่วนใหญ่)
    """
    return {
        'branch_restrictions': {},
        'stats': {
            'total_branches': 0,
            'strict': 0,
            'flexible': 0,
            'total_bookings': 0,
            'fallback': True,
            'message': 'ใช้ Punthai เป็นหลัก (ไม่พบไฟล์ Booking History)'
        }
    }

@st.cache_data(ttl=3600)  # Cache 1 ชั่วโมง
def load_punthai_reference():
    """โหลดไฟล์ Punthai Maxmart เพื่อเรียนรู้หลักการจัดทริป (Location patterns - Optimized)"""
    try:
        file_path = 'Dc/แผนงาน Punthai Maxmart รอบสั่ง 24หยิบ 25พฤศจิกายน 2568 To.เฟิ(1) - สำเนา.xlsx'
        df = pd.read_excel(file_path, sheet_name='2.Punthai', header=1)
        
        # กรองเฉพาะแถวที่มี Trip และไม่ใช่ DC/Distribution Center
        df_clean = df[df['Trip'].notna()].copy()
        df_clean = df_clean[~df_clean['BranchCode'].isin(['DC011', 'PTDC', 'PTG Distribution Center'])].copy()
        
        # Extract vehicle type from Trip no (เช่น 4W009 → 4W)
        df_clean['Vehicle_Type'] = df_clean['Trip no'].apply(
            lambda x: str(x)[:2] if pd.notna(x) else 'Unknown'
        )
        
        # Merge กับ Master เพื่อได้ข้อมูลตำบล/อำเภอ/จังหวัด
        try:
            df_master = pd.read_excel('Dc/Master สถานที่ส่ง.xlsx')
            df_clean = df_clean.merge(
                df_master[['Plan Code', 'ตำบล', 'อำเภอ', 'จังหวัด']],
                left_on='BranchCode',
                right_on='Plan Code',
                how='left'
            )
        except:
            pass
        
        # เรียนรู้ข้อจำกัดรถจาก Punthai (แผน) - สำหรับสาขาที่ไม่มีใน Booking
        punthai_restrictions = {}
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for branch_code in df_clean['BranchCode'].unique():
            branch_data = df_clean[df_clean['BranchCode'] == branch_code]
            vehicles_used = set(branch_data['Vehicle_Type'].dropna().tolist())
            vehicles_used = {v for v in vehicles_used if v in ['4W', 'JB', '6W']}
            
            if vehicles_used:
                if len(vehicles_used) == 1:
                    vehicle = list(vehicles_used)[0]
                    punthai_restrictions[str(branch_code)] = {
                        'max_vehicle': vehicle,
                        'allowed': [vehicle],
                        'source': 'PUNTHAI'
                    }
                else:
                    max_vehicle = max(vehicles_used, key=lambda v: vehicle_sizes.get(v, 0))
                    punthai_restrictions[str(branch_code)] = {
                        'max_vehicle': max_vehicle,
                        'allowed': list(vehicles_used),
                        'source': 'PUNTHAI'
                    }
        
        # สร้าง dictionary: Trip → ข้อมูล (location patterns)
        trip_patterns = {}
        location_stats = {
            'same_province': 0,
            'mixed_province': 0,
            'avg_branches': 0
        }
        
        for trip_num in df_clean['Trip'].unique():
            trip_data = df_clean[df_clean['Trip'] == trip_num]
            
            # Get location info
            provinces = set(trip_data['จังหวัด'].dropna().tolist()) if 'จังหวัด' in trip_data.columns else set()
            
            # Count same vs mixed province
            if len(provinces) == 1:
                location_stats['same_province'] += 1
            elif len(provinces) > 1:
                location_stats['mixed_province'] += 1
            
            trip_patterns[int(trip_num)] = {
                'branches': len(trip_data),
                'codes': trip_data['BranchCode'].tolist(),
                'weight': trip_data['TOTALWGT'].sum() if 'TOTALWGT' in trip_data.columns else 0,
                'cube': trip_data['TOTALCUBE'].sum() if 'TOTALCUBE' in trip_data.columns else 0,
                'provinces': list(provinces),
                'same_province': len(provinces) == 1
            }
        
        # Calculate stats
        if trip_patterns:
            location_stats['avg_branches'] = sum(t['branches'] for t in trip_patterns.values()) / len(trip_patterns)
            total = location_stats['same_province'] + location_stats['mixed_province']
            location_stats['same_province_pct'] = (location_stats['same_province'] / total * 100) if total > 0 else 0
        
        return {
            'patterns': trip_patterns, 
            'stats': location_stats,
            'punthai_restrictions': punthai_restrictions
        }
    except:
        return {'patterns': {}, 'stats': {}, 'punthai_restrictions': {}}

# โหลด Booking History (ข้อจำกัดรถ)
BOOKING_RESTRICTIONS = load_booking_history_restrictions()

# โหลด Punthai Reference (location patterns)
PUNTHAI_PATTERNS = load_punthai_reference()

# ==========================================
# HELPER FUNCTIONS
# ==========================================
def normalize(val):
    """ทำให้รหัสสาขาเป็นมาตรฐาน"""
    return str(val).strip().upper().replace(" ", "").replace(".0", "")

def calculate_distance(lat1, lon1, lat2, lon2):
    """คำนวณระยะทางระหว่างสองจุด (กม.) - Haversine formula"""
    # ตรวจสอบค่า None, NaN หรือ 0
    try:
        lat1 = float(lat1) if lat1 is not None else 0
        lon1 = float(lon1) if lon1 is not None else 0
        lat2 = float(lat2) if lat2 is not None else 0
        lon2 = float(lon2) if lon2 is not None else 0
    except (ValueError, TypeError):
        return 0
    
    if lat1 == 0 or lon1 == 0 or lat2 == 0 or lon2 == 0:
        return 0
    import math
    lat1_rad, lon1_rad = math.radians(lat1), math.radians(lon1)
    lat2_rad, lon2_rad = math.radians(lat2), math.radians(lon2)
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    return 6371 * c

def calculate_distance_from_dc(lat, lon):
    """คำนวณระยะทางจาก DC วังน้อย (กม.)"""
    return calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)

def check_branch_vehicle_compatibility(branch_code, vehicle_type):
    """ตรวจสอบว่าสาขานี้ใช้รถประเภทนี้ได้ไหม (รวม Booking + Punthai)"""
    branch_code_str = str(branch_code).strip()
    
    # 1. ลองหาจาก Booking History ก่อน (ข้อมูลจริง)
    booking_restrictions = BOOKING_RESTRICTIONS.get('branch_restrictions', {})
    if branch_code_str in booking_restrictions:
        allowed = booking_restrictions[branch_code_str].get('allowed', [])
        return vehicle_type in allowed
    
    # 2. ถ้าไม่มี ลองหาจาก Punthai (แผน)
    punthai_restrictions = PUNTHAI_PATTERNS.get('punthai_restrictions', {})
    if branch_code_str in punthai_restrictions:
        allowed = punthai_restrictions[branch_code_str].get('allowed', [])
        return vehicle_type in allowed
    
    # 3. ถ้าไม่มีข้อมูล = ยืดหยุ่น
    return True

def is_punthai_only(trip_data):
    """
    ตรวจสอบว่าทริปนี้เป็น Punthai ล้วนหรือไม่
    
    Returns:
        'punthai_only': ถ้าทั้งหมดเป็น Punthai (BU = 211 หรือชื่อมี PUNTHAI)
        'mixed': ถ้ามีทั้ง Punthai และอื่น
        'other': ถ้าไม่มี Punthai เลย
    """
    if trip_data is None or len(trip_data) == 0:
        return 'other'
    
    punthai_count = 0
    total_count = len(trip_data)
    
    for _, row in trip_data.iterrows():
        bu = row.get('BU', None)
        name = str(row.get('Name', '')).upper()
        
        # เช็ค BU = 211 หรือชื่อมี PUNTHAI
        if bu == 211 or bu == '211' or 'PUNTHAI' in name:
            punthai_count += 1
    
    if punthai_count == total_count:
        return 'punthai_only'
    elif punthai_count > 0:
        return 'mixed'
    else:
        return 'other'

def get_punthai_vehicle_limits(trip_data, total_cube, branch_count):
    """
    ดึงข้อจำกัดรถสำหรับ Punthai
    
    กฎ:
    - Punthai ล้วน + JB: ไม่เกิน 7 drop
    - Punthai ล้วน + 4W: ถ้า Cube > 5 → ตัดเป็น JB
    - ผสม (Punthai + อื่น): ถ้า Cube 3-4 → 6W ได้, ถ้าเกิน → ตัดเป็น 4W เท่านั้น
    
    Returns:
        dict: {'max_vehicle': '4W'/'JB'/'6W', 'max_drops': int, 'should_split': bool}
    """
    punthai_type = is_punthai_only(trip_data)
    
    if punthai_type == 'punthai_only':
        # Punthai ล้วน
        if total_cube > 5.0:
            # Cube เกิน 5 → ใช้ JB (ไม่ใช้ 4W)
            if branch_count > 7:
                # เกิน 7 drop → ต้องแยก
                return {'max_vehicle': 'JB', 'max_drops': 7, 'should_split': True}
            else:
                return {'max_vehicle': 'JB', 'max_drops': 7, 'should_split': False}
        else:
            # Cube ≤ 5 → ใช้ 4W ได้
            return {'max_vehicle': '4W', 'max_drops': 12, 'should_split': False}
    
    elif punthai_type == 'mixed':
        # ผสม (Punthai + อื่น): 4W max_c = 4.0
        if total_cube <= 4.0:
            # Cube ≤ 4 → ใช้ 4W ได้
            return {'max_vehicle': '4W', 'max_drops': 12, 'should_split': False}
        elif total_cube <= 7.0:
            # Cube 4-7 → ใช้ JB
            return {'max_vehicle': 'JB', 'max_drops': 12, 'should_split': False}
        else:
            # Cube > 7 → ต้องแยก
            return {'max_vehicle': 'JB', 'max_drops': 12, 'should_split': True}
    
    else:
        # ไม่มี Punthai → ไม่มีข้อจำกัดพิเศษ
        return {'max_vehicle': '6W', 'max_drops': 999, 'should_split': False}

def get_max_vehicle_for_branch(branch_code):
    """ดึงรถใหญ่สุดที่สาขานี้รองรับ - ใช้จาก Auto Plan เท่านั้น!"""
    branch_code_str = str(branch_code).strip()
    
    # 🔒 ใช้จาก Auto Plan (ชีต info - MaxTruckType) เท่านั้น!
    if branch_code_str in AUTO_PLAN_TRUCK_LIMITS:
        return AUTO_PLAN_TRUCK_LIMITS[branch_code_str]
    
    # ถ้าไม่มีในไฟล์ = ใช้รถใหญ่ได้ (6W)
    return '6W'

def get_max_vehicle_for_trip(trip_codes):
    """
    หารถใหญ่สุดที่ทริปนี้ใช้ได้ (เช็คข้อจำกัดของทุกสาขาในทริป)
    
    🔒 อ้างอิงจาก Auto Plan เท่านั้น (ไม่บังคับห้าม 6W ตามพื้นที่)
    
    Args:
        trip_codes: set ของ branch codes ในทริป
    
    Returns:
        str: '4W', 'JB', หรือ '6W'
    """
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    max_allowed = '6W'  # เริ่มจากใหญ่สุด แล้วจำกัดตามข้อจำกัดสาขา
    min_priority = 3  # ค่าใหญ่สุดคือไม่มีข้อจำกัด
    
    for code in trip_codes:
        # 🔒 ใช้ Auto Plan เท่านั้น
        branch_max = get_max_vehicle_for_branch(code)
        priority = vehicle_priority.get(branch_max, 3)
        
        # 🔒 เลือกรถที่เล็กที่สุด (ข้อจำกัดมากที่สุด) จากทุกสาขาในทริป
        if priority < min_priority:
            min_priority = priority
            max_allowed = branch_max
    
    return max_allowed

def get_required_vehicle_by_distance(branch_code):
    """ตรวจสอบว่าสาขาต้องใช้รถอะไรตามระยะทางจาก DC"""
    # ดึงพิกัดจาก Master
    if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
        master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == branch_code]
        if len(master_row) > 0:
            lat = master_row.iloc[0].get('ละติจูด', 0)
            lon = master_row.iloc[0].get('ลองติจูด', 0)
            distance = calculate_distance_from_dc(lat, lon)
            
            # ถ้าห่างจาก DC เกินกำหนด → ต้องใช้ 6W
            if distance > DISTANCE_REQUIRE_6W:
                return '6W', distance
    
    return None, 0

def can_fit_truck(total_weight, total_cube, truck_type):
    """เช็คว่าน้ำหนัก/คิวใส่รถได้หรือไม่"""
    limits = LIMITS[truck_type]
    max_w = limits['max_w'] * BUFFER
    max_c = limits['max_c'] * BUFFER
    return total_weight <= max_w and total_cube <= max_c

def suggest_truck(total_weight, total_cube, max_allowed='6W', trip_codes=None, prefer_jb_for_nearby=False):
    """
    แนะนำรถที่เหมาะสม โดยเลือกรถที่:
    1. ใส่ของได้พอดี (ไม่เกินขีดจำกัด 105%)
    2. ใช้งานได้ใกล้ 100% มากที่สุด (เป้าหมาย: 90-100%)
    3. เคารพข้อจำกัดของสาขา (ถ้าสาขาใช้แค่ 4W = ต้องใช้ 4W เท่านั้น)
    4. 🆕 พื้นที่ใกล้ DC → ใช้รถเล็กก่อน (4W → JB → 6W)
    """
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    max_size = vehicle_sizes.get(max_allowed, 3)
    
    # ตรวจสอบข้อจำกัดของสาขาทั้งหมดในกลุ่ม
    branch_max_vehicle = '6W'  # เริ่มจากใหญ่สุด
    is_nearby_area = False  # เช็คว่าเป็นพื้นที่ใกล้ DC หรือไม่
    avg_distance_from_dc = 0  # 🆕 ระยะเฉลี่ยจาก DC
    
    if trip_codes is not None and len(trip_codes) > 0:
        for code in trip_codes:
            branch_max = get_max_vehicle_for_branch(code)
            # หารถที่เล็กที่สุดที่อนุญาต
            if vehicle_sizes.get(branch_max, 3) < vehicle_sizes.get(branch_max_vehicle, 3):
                branch_max_vehicle = branch_max
        
        # จำกัด max_allowed ตามข้อจำกัดของสาขา
        if vehicle_sizes.get(branch_max_vehicle, 3) < max_size:
            max_allowed = branch_max_vehicle
            max_size = vehicle_sizes.get(max_allowed, 3)
        
        # 🆕 คำนวณระยะเฉลี่ยจาก DC และเช็คพื้นที่
        total_dist = 0
        dist_count = 0
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
            for code in trip_codes:
                master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                if len(master_row) > 0:
                    prov = master_row.iloc[0].get('จังหวัด', '')
                    if pd.notna(prov) and get_region_type(prov) == 'nearby':
                        is_nearby_area = True
                    
                    # คำนวณระยะจาก DC
                    lat = master_row.iloc[0].get('ละติจูด', None)
                    lon = master_row.iloc[0].get('ลองติจูด', None)
                    if lat and lon and pd.notna(lat) and pd.notna(lon):
                        dist = haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, float(lat), float(lon))
                        total_dist += dist
                        dist_count += 1
        
        if dist_count > 0:
            avg_distance_from_dc = total_dist / dist_count
    
    # 🆕 พื้นที่ใกล้ DC (nearby หรือ ระยะ < 100km) → ใช้รถเล็กก่อน (4W → JB)
    if is_nearby_area or avg_distance_from_dc < 100:
        truck_order = ['4W', 'JB', '6W']  # รถเล็กก่อน
    else:
        truck_order = ['4W', 'JB', '6W']  # ปกติก็ใช้รถเล็กก่อน
    
    best_truck = None
    best_utilization = 0
    best_distance_from_100 = 999  # ระยะห่างจาก 100%
    
    for truck in truck_order:
        truck_size = vehicle_sizes.get(truck, 0)
        # ถ้ารถใหญ่กว่าที่อนุญาต ข้ามไป
        if truck_size > max_size:
            continue
        if can_fit_truck(total_weight, total_cube, truck):
            # คำนวณ % การใช้รถ
            limits = LIMITS[truck]
            w_util = (total_weight / limits['max_w']) * 100
            c_util = (total_cube / limits['max_c']) * 100
            utilization = max(w_util, c_util)
            
            # คำนวณระยะห่างจาก 100%
            distance_from_100 = abs(100 - utilization)
            
            # เลือกรถที่ใกล้ 100% ที่สุด (90-105% เป็นเป้าหมาย)
            # ถ้าใช้งานใกล้เคียงกัน เลือกรถที่ใช้งานสูงกว่า
            if best_truck is None:
                best_truck = truck
                best_utilization = utilization
                best_distance_from_100 = distance_from_100
            else:
                # ถ้าอยู่ในช่วง 90-100% เลือกที่ใกล้ 100% ที่สุด
                if 90 <= utilization <= 100:
                    if distance_from_100 < best_distance_from_100 or best_utilization < 90:
                        best_truck = truck
                        best_utilization = utilization
                        best_distance_from_100 = distance_from_100
                # ถ้าทั้งคู่ไม่อยู่ในช่วง เลือกที่ใช้งานสูงกว่า
                elif utilization > best_utilization:
                    best_truck = truck
                    best_utilization = utilization
                    best_distance_from_100 = distance_from_100
    
    if best_truck:
        return best_truck
    
    # ถ้าไม่มีรถที่เหมาะสม ใช้รถใหญ่สุดที่อนุญาต
    return max_allowed if max_allowed in LIMITS else '6W+'

def calculate_optimal_vehicle_split(total_weight, total_cube, max_allowed='6W', branch_count=0):
    """
    🚛 คำนวณการแบ่งรถที่เหมาะสม
    
    เงื่อนไข:
    - 4W: ≤12 จุด, Cube ≤ 5
    - JB: ≤12 จุด, Cube ≤ 8  
    - 6W: ไม่จำกัดจุด, Cube ต้องเต็ม ≥100%
    
    ลำดับการเลือก:
    1. 4W (ถ้า cube ≤ 5)
    2. JB (ถ้า cube ≤ 8)
    3. JB + 4W (แยก 2 คัน, 75%-95% ต่อคัน)
    4. JB + JB (แยก 2 คัน, 75%-95% ต่อคัน)
    5. 6W + JB (แยก 2 คัน, 75%-95% ต่อคัน)
    6. 4W + 4W (แยก 2 คัน, 75%-95% ต่อคัน)
    7. 6W (cube ต้อง ≥100%)
    
    Returns: (vehicle_type, split_needed, split_config)
    """
    vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
    max_priority = vehicle_priority.get(max_allowed, 3)
    
    # คำนวณ utilization สำหรับแต่ละรถ (ใช้ Cube เป็นหลัก)
    cube_util_4w = (total_cube / LIMITS['4W']['max_c']) * 100  # max 5 cube
    cube_util_jb = (total_cube / LIMITS['JB']['max_c']) * 100  # max 7 cube
    cube_util_6w = (total_cube / LIMITS['6W']['max_c']) * 100  # max 20 cube
    
    weight_util_4w = (total_weight / LIMITS['4W']['max_w']) * 100
    weight_util_jb = (total_weight / LIMITS['JB']['max_w']) * 100
    weight_util_6w = (total_weight / LIMITS['6W']['max_w']) * 100
    
    # 🎯 เป้าหมาย: Utilization 75%-95% สำหรับการแยก, 95%-100% สำหรับคันเดียว
    SPLIT_MIN = 75   # ขั้นต่ำสำหรับแต่ละคันเมื่อแยก
    SPLIT_MAX = 95   # สูงสุดสำหรับแต่ละคันเมื่อแยก
    SINGLE_MIN = 95  # ขั้นต่ำสำหรับคันเดียว
    SINGLE_MAX = 100 # สูงสุดสำหรับคันเดียว (ห้ามเกิน 100%)
    
    # ตรวจสอบจำนวนสาขา (4W/JB ไม่เกิน 12 จุด)
    branch_ok_for_small = branch_count <= 12 or branch_count == 0
    
    # 1. ลอง 4W ก่อน (ถ้า cube ≤ 5 และ ≤12 จุด)
    if max_priority >= 1 and total_cube <= 5.0 and branch_ok_for_small:
        if cube_util_4w <= 100 and weight_util_4w <= 100:
            return ('4W', False, None)
    
    # 2. ลอง JB (ถ้า cube ≤ 7 และ ≤12 จุด)
    if max_priority >= 2 and total_cube <= 7.0 and branch_ok_for_small:
        if cube_util_jb <= 100 and weight_util_jb <= 100:
            return ('JB', False, None)
    
    # 3. ถ้ารถเดียวไม่พอ ต้องแยก (cube > 7 หรือ จุด > 12)
    need_split = total_cube > 7.0 or not branch_ok_for_small
    
    if need_split:
        # 🔄 ลองแบบต่างๆ ตามลำดับ
        
        # 🔒 กฎใหม่: cube 7-14 → บังคับแยกเป็น JB+JB หรือ JB+4W
        if total_cube > 7.0 and total_cube <= 14.0:
            if total_cube <= 12.0:
                # JB + 4W (JB 7 + 4W 5 = 12 max)
                return ('JB', True, {'split': ['JB', '4W'], 'reason': f'Cube {total_cube:.1f} แยกเป็น JB+4W'})
            else:
                # JB + JB (7 + 7 = 14 max)
                return ('JB', True, {'split': ['JB', 'JB'], 'reason': f'Cube {total_cube:.1f} แยกเป็น JB+JB'})
        
        # 🔒 กฎใหม่: cube 14-18 → บังคับแยกเป็น JB+JB+4W หรือ 6W ถ้าจำเป็น
        if total_cube > 14.0 and total_cube < 18.0:
            # ยังไม่ถึง 6W (ต้อง ≥18) แต่เกิน JB+JB → แยกเป็น JB+JB+4W หรือยอมใช้ 6W
            if max_priority >= 3:
                # ยอมใช้ 6W แม้ว่าจะไม่ถึง 18 cube (ดีกว่าแยก 3 คัน)
                return ('6W', False, {'reason': f'Cube {total_cube:.1f} ใช้ 6W (รอรวมเพิ่ม)'})
            else:
                return ('JB', True, {'split': ['JB', 'JB'], 'reason': f'Cube {total_cube:.1f} แยกเป็น JB+JB'})
        
        # 6W + JB (สำหรับ cube > 20)
        if max_priority >= 3 and total_cube > 20.0 and total_cube <= 27.0:
            return ('6W', True, {'split': ['6W', 'JB'], 'reason': f'Cube {total_cube:.1f} แยกเป็น 6W+JB'})
        
        # 4W + 4W (4W 5 + 4W 5 = 10 cube max) - สำหรับสาขาที่จำกัด 4W
        if max_priority == 1 and total_cube <= 10.0:
            four_w_util_half = (total_cube / 2 / LIMITS['4W']['max_c']) * 100
            if SPLIT_MIN <= four_w_util_half <= SPLIT_MAX:
                return ('4W', True, {'split': ['4W', '4W'], 'ratio': [0.5, 0.5]})
    
    # 4. 6W (ไม่จำกัดจุด แต่ต้องได้ขั้นต่ำ 18 cube (90%) และไม่เกิน 20 cube)
    if max_priority >= 3:
        # 🔒 6W เกิน 20 cube → ตัดส่วนเกินไป 4W
        if total_cube > 20.0:
            return ('6W', True, {'split': ['6W', '4W'], 'reason': 'เกิน 20 cube ตัดส่วนเกินไป 4W'})
        
        # 🔒 6W ต้องได้ขั้นต่ำ 18 cube (90%) ถ้าต่ำกว่า → แยกเป็น JB
        if total_cube >= 18.0:
            return ('6W', False, None)
        elif total_cube >= 7.0 and total_cube < 18.0:
            # 6W ไม่ถึง 18 cube → แยกเป็น JB แทน (ถ้าได้)
            if max_priority >= 2 and branch_ok_for_small:
                # ลองแยกเป็น JB + JB หรือ JB + 4W
                if total_cube <= 14.0:  # JB + JB = 14 cube max
                    return ('JB', True, {'split': ['JB', 'JB'], 'reason': '6W ไม่ถึง 18 cube แยกเป็น JB'})
                else:
                    return ('JB', True, {'split': ['JB', '4W'], 'reason': '6W ไม่ถึง 18 cube แยกเป็น JB+4W'})
            else:
                # ถ้าแยก JB ไม่ได้ → ยอมใช้ 6W แม้ไม่ถึง 18 cube
                return ('6W', False, None)
        else:
            # 6W ว่างมาก (<7 cube) → ลดเป็น JB หรือ 4W
            if total_cube <= 7.0 and branch_ok_for_small and max_priority >= 2:
                return ('JB', False, None)
            if total_cube <= 5.0 and branch_ok_for_small:
                return ('4W', False, None)
    
    # Default: ใช้ max_allowed
    return (max_allowed, False, None)

def can_branch_use_vehicle(code, vehicle_type, branch_vehicles):
    """
    เช็คว่าสาขาสามารถใช้รถประเภทนี้ได้หรือไม่
    - ถ้าไม่มีประวัติ = ใช้ได้ทุกประเภท
    - ถ้ามีประวัติใช้รถใหญ่ = ใช้รถเล็กกว่าได้
    - ถ้ามีประวัติใช้แค่รถเล็ก (เช่น 4W) = ใช้รถใหญ่ไม่ได้ (รถใหญ่เข้าไม่ได้)
    """
    if not branch_vehicles or code not in branch_vehicles:
        return True  # ไม่มีประวัติ = ใช้ได้ทุกประเภท
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return True  # ไม่มีข้อมูลรถ = ใช้ได้ทุกประเภท
    
    # ถ้าเคยใช้รถประเภทนี้ = ใช้ได้
    if vehicle_type in vehicle_history:
        return True
    
    # เช็คขนาดรถ (6W > JB > 4W)
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    requested_size = vehicle_sizes.get(vehicle_type, 0)
    
    # หารถที่ใหญ่ที่สุดที่สาขาเคยใช้
    max_used_size = max(vehicle_sizes.get(v, 0) for v in vehicle_history)
    
    # ถ้าขอใช้รถเล็กกว่าหรือเท่ากับที่เคยใช้ = ใช้ได้
    # ถ้าขอใช้รถใหญ่กว่าที่เคยใช้ = ใช้ไม่ได้ (รถใหญ่อาจเข้าไม่ได้)
    return requested_size <= max_used_size

def get_max_vehicle_for_branch_old(code, branch_vehicles):
    """[OLD] ดึงประเภทรถที่ใหญ่ที่สุดที่สาขาเคยใช้ (จำกัดไม่ให้ใช้รถใหญ่กว่านี้)"""
    if not branch_vehicles or code not in branch_vehicles:
        return '6W'  # ไม่มีประวัติ = ใช้ได้ถึง 6W
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return '6W'
    
    vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
    max_vehicle = max(vehicle_history.keys(), key=lambda v: vehicle_sizes.get(v, 0))
    return max_vehicle

def get_most_used_vehicle_for_branch(code, branch_vehicles):
    """ดึงประเภทรถที่สาขาใช้บ่อยที่สุด"""
    if not branch_vehicles or code not in branch_vehicles:
        return None
    
    vehicle_history = branch_vehicles[code]
    if not vehicle_history:
        return None
    
    return max(vehicle_history, key=vehicle_history.get)

def is_similar_name(name1, name2, similarity_threshold=85):
    """เช็คว่าชื่อสาขาคล้ายกันหรือไม่ - ใช้ Fuzzy Matching + คำสำคัญ
    
    Args:
        name1: ชื่อสาขาที่ 1
        name2: ชื่อสาขาที่ 2
        similarity_threshold: ความคล้ายขั้นต่ำ (0-100, default=85)
    
    Returns:
        True ถ้าชื่อคล้ายกัน
    """
    def extract_keywords(name):
        """ดึงคำสำคัญจากชื่อสาขา"""
        if pd.isna(name) or name is None:
            return set(), "", "", ""
        s = str(name).strip().upper()
        
        # คำสำคัญที่ต้องการจับคู่ (ไทย + อังกฤษ) - เพิ่มเติมคำเฉพาะของสาขา
        keywords = set()
        
        # เช็คคำสำคัญแบบ exact match
        important_words = [
            # สถานที่สำคัญ
            'ฟิวเจอร์', 'FUTURE', 'รังสิต', 'RANGSIT', 'คลองหลวง', 'KHLONGLUANG',
            'เซ็นทรัล', 'CENTRAL', 'เทสโก้', 'TESCO', 'โลตัส', 'LOTUS',
            'บิ๊กซี', 'BIGC', 'แม็คโคร', 'MAKRO', 'โฮมโปร', 'HOMEPRO',
            'ซีคอน', 'SEACON', 'เมกา', 'MEGA', 'พาราไดซ์', 'PARADISE',
            'เทอร์มินอล', 'TERMINAL', 'สยามพารากอน', 'SIAM', 'PARAGON',
            # หมายเลขสาขา (คลองหลวง 3, 4, 8, 10 ฯลฯ)
            'คลองหลวง3', 'คลองหลวง4', 'คลองหลวง8', 'คลองหลวง10',
        ]
        
        for word in important_words:
            if word in s:
                keywords.add(word)
        
        # ดึงชื่อพื้นฐาน (เช่น "คลองหลวง" จาก "คลองหลวง3")
        import re
        base_match = re.search(r'([ก-๙A-Z]+)\s*\d+', s)
        if base_match:
            base_name = base_match.group(1).strip()
            if len(base_name) >= 3:
                keywords.add(base_name)
        
        # Pattern 3: ดึงชื่อในวงเล็บ (เช่น "คลองหลวง4(ถ.พหลโยธิน กม.34)" → "คลองหลวง4")
        paren_match = re.search(r'^([^(]+)', s)
        if paren_match:
            main_name = paren_match.group(1).strip()
            if len(main_name) >= 3 and main_name != s:
                keywords.add(main_name)  # "คลองหลวง4"
        
        # ลบ prefix/suffix ที่พบบ่อย
        prefixes = ['PTC-MRT-', 'FC PTF ', 'PTC-', 'PTC ', 'PUN-', 'PTF ', 'FC ', 
                   'MAXMART', 'CW', 'NW', 'MI', 'PI', 'MH', 'ME', 'SE', 'SG', 'SH', 'MG']
        clean_s = s
        for prefix in prefixes:
            if clean_s.startswith(prefix):
                clean_s = clean_s[len(prefix):].strip()
                break
        
        # ลบตัวอักษรเดี่ยวที่ขึ้นต้น (M, P, N, S) ถ้าตามด้วยตัวเลข
        if re.match(r'^[MPNS]\d', clean_s):
            clean_s = clean_s[1:]
        
        # แยกภาษาไทยและอังกฤษ
        thai_chars = ''.join([c for c in s if '\u0e01' <= c <= '\u0e5b'])
        eng_chars = ''.join([c for c in s if c.isalpha() and c.isascii()])
        
        return keywords, thai_chars, eng_chars, clean_s
    
    keywords1, thai1, eng1, clean1 = extract_keywords(name1)
    keywords2, thai2, eng2, clean2 = extract_keywords(name2)
    
    # 🔥 ลำดับแรก: เช็คคำสำคัญก่อน (เช่น ฟิวเจอร์+รังสิต, คลองหลวง)
    if keywords1 and keywords2:
        common_keywords = keywords1 & keywords2
        
        # ✅ Case 1: ชื่อพื้นฐานเดียวกัน (เช่น "คลองหลวง" ใน คลองหลวง3, คลองหลวง4, คลองหลวง6)
        base_names = {k for k in common_keywords if len(k) >= 3 and not k.isdigit()}
        if base_names:
            # ถือว่าเป็นสาขาเดียวกัน (เช่น คลองหลวง 3, 4, 6, 8, 10)
            return True
        
        # ✅ Case 1.5: เช็ค partial match ของชื่อพื้นฐาน (เช่น "KHLONG" ใน keywords1, "LUANG" ใน keywords1 + "KHLONG" ใน keywords2)
        # → ถือว่าเป็น "KHLONG LUANG" เดียวกัน (รองรับภาษาอังกฤษ)
        for k1 in keywords1:
            for k2 in keywords2:
                # ถ้า k1 เป็น substring ของ k2 หรือตรงกันข้าม
                if len(k1) >= 4 and len(k2) >= 4:
                    if k1 in k2 or k2 in k1:
                        return True
        
        # ถ้ามีคำสำคัญร่วมกัน >= 2 คำ → ถือว่าเหมือนกัน
        if len(common_keywords) >= 2:
            return True
        
        # ถ้ามีคำสำคัญร่วมกัน 1 คำ แต่เป็นคำเฉพาะ → ถือว่าเหมือนกัน
        if len(common_keywords) >= 1:
            specific_places = {'รังสิต', 'RANGSIT', 'เซ็นทรัล', 'CENTRAL', 'ซีคอน', 'SEACON', 'คลองหลวง', 'KHLONGLUANG', 'ตลาดไท', 'TALADTHAI'}
            if common_keywords & specific_places:
                if len(common_keywords) >= 2 or (thai1 and thai2 and len(thai1) >= 4 and thai1[:4] in thai2):
                    return True
    
    # 🎯 Fuzzy Matching - ใช้ rapidfuzz ถ้ามี หรือ difflib ถ้าไม่มี
    if FUZZY_AVAILABLE:
        # ใช้ rapidfuzz (เร็วและแม่นยำกว่า)
        ratio = fuzz.token_sort_ratio(clean1, clean2)
        if ratio >= similarity_threshold:
            return True
        
        # เช็ค partial ratio สำหรับชื่อที่เป็น substring
        partial_ratio = fuzz.partial_ratio(clean1, clean2)
        if partial_ratio >= 90:  # ความคล้ายสูงมาก
            return True
    else:
        # Fallback: ใช้ difflib
        ratio = SequenceMatcher(None, clean1, clean2).ratio() * 100
        if ratio >= similarity_threshold:
            return True
    
    # ต้องมีความยาวพอสมควร
    if len(thai1) < 3 and len(eng1) < 3:
        return False
    if len(thai2) < 3 and len(eng2) < 3:
        return False
    
    # เช็คภาษาไทย
    if thai1 and thai2:
        shorter_thai = min(thai1, thai2, key=len)
        longer_thai = max(thai1, thai2, key=len)
        if len(shorter_thai) >= 3 and shorter_thai in longer_thai:
            return True
        # ความคล้าย 80%+
        if len(shorter_thai) >= 5:
            common = sum(1 for c in shorter_thai if c in longer_thai)
            if common / len(shorter_thai) >= 0.8:
                return True
    
    # เช็คภาษาอังกฤษ
    if eng1 and eng2:
        shorter_eng = min(eng1, eng2, key=len)
        longer_eng = max(eng1, eng2, key=len)
        if len(shorter_eng) >= 3 and shorter_eng in longer_eng:
            return True
        # ความคล้าย 80%+
        if len(shorter_eng) >= 5:
            common = sum(1 for c in shorter_eng if c in longer_eng)
            if common / len(shorter_eng) >= 0.8:
                return True
    
    # เช็คคำสำคัญระหว่างไทย-อังกฤษ (เช่น Future = ฟิวเจอร์, Rangsit = รังสิต)
    thai_eng_map = {
        'RANGSIT': 'รังสิต',
        'FUTURE': 'ฟิวเจอร',
        'PARK': 'ปารค',
        'TRIANGLE': 'ไตรแองเกิล',
    }
    
    for eng_word, thai_word in thai_eng_map.items():
        # ตรวจสอบว่ามีคำนี้ในชื่อทั้งสองฝั่ง (ไทย-อังกฤษ หรือ อังกฤษ-อังกฤษ หรือ ไทย-ไทย)
        has_eng_in_1 = eng_word in eng1
        has_eng_in_2 = eng_word in eng2
        has_thai_in_1 = thai_word in thai1
        has_thai_in_2 = thai_word in thai2
        
        # ถ้าทั้งสองมีคำเดียวกัน (ไม่ว่าจะไทยหรืออังกฤษ) = คล้ายกัน
        if (has_eng_in_1 and has_eng_in_2) or (has_thai_in_1 and has_thai_in_2):
            return True
        # ถ้าข้ามภาษา (อังกฤษ-ไทย)
        if (has_eng_in_1 and has_thai_in_2) or (has_eng_in_2 and has_thai_in_1):
            return True
    
    return False

def get_branch_base_code(code):
    """
    ดึง base code ของสาขา เพื่อจับกลุ่มสาขาที่เป็น location เดียวกัน
    
    ตัวอย่าง:
    - M862, P862, S862, ZS862 → 862 (ตลิ่งชัน2)
    - MF40, PF40, SF40, ZSF40 → F40 (ถ.พระเทพ1)
    - M036, P036, P723 → 036, 723 (มุกดาหาร)
    - MD65, PD65 → D65 (อำนาจเจริญ3)
    
    Prefix patterns:
    - M = MAX MART
    - P = PUNTHAI
    - N = PUNTHAI (old)
    - S = SUPPLY USE
    - ZS, ZF, Z = LUBE
    - O = MAX MART (outlet)
    """
    import re
    
    if not code or pd.isna(code):
        return None
    
    code = str(code).strip().upper()
    
    # Skip FC codes (11005xxx, 9100002xxx)
    if re.match(r'^\d{7,}$', code):
        return None
    
    # Remove prefix patterns
    # ZS, ZF, ZC first (2 chars)
    if code.startswith(('ZS', 'ZF', 'ZC')):
        base = code[2:]
    # Single letter prefix: M, P, N, S, O, F, C, E
    elif len(code) > 1 and code[0] in 'MPNSOFC' and not code[0:2].isalpha():
        base = code[1:]
    # Two letter prefix: MF, PF, SF, MD, PD, MI, PI, MH, PH, etc.
    elif len(code) > 2 and code[0] in 'MPS' and code[1].isalpha():
        base = code[1:]  # Keep the second letter as part of base
    else:
        base = code
    
    # Clean up: remove trailing spaces/special chars
    base = re.sub(r'[^A-Z0-9]', '', base)
    
    return base if len(base) >= 2 else None

def is_same_location(code1, code2):
    """
    เช็คว่า 2 codes เป็น location เดียวกันหรือไม่
    เช่น M862 และ P862 = ตลิ่งชัน2 (เป็น location เดียวกัน)
    """
    base1 = get_branch_base_code(code1)
    base2 = get_branch_base_code(code2)
    
    if base1 and base2:
        return base1 == base2
    return False

def haversine_distance(lat1, lon1, lat2, lon2):
    """
    คำนวณระยะทางระหว่างจุดสองจุดบนพื้นโลก (km)
    ใช้สูตร Haversine
    """
    from math import radians, sin, cos, sqrt, atan2
    
    # ตรวจสอบค่า None, NaN หรือ 0
    try:
        lat1 = float(lat1) if lat1 is not None else 0
        lon1 = float(lon1) if lon1 is not None else 0
        lat2 = float(lat2) if lat2 is not None else 0
        lon2 = float(lon2) if lon2 is not None else 0
    except (ValueError, TypeError):
        return 0
    
    if lat1 == 0 or lon1 == 0 or lat2 == 0 or lon2 == 0:
        return 0
    
    # แปลงองศาเป็น radians
    lat1_rad = radians(lat1)
    lon1_rad = radians(lon1)
    lat2_rad = radians(lat2)
    lon2_rad = radians(lon2)
    
    # ความต่างของพิกัด
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    
    # สูตร Haversine
    a = sin(dlat/2)**2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    
    # รัศมีโลก (km)
    R = 6371.0
    distance = R * c
    
    return distance

def calculate_max_consecutive_distance(codes, coord_cache, dc_lat=14.179394, dc_lon=100.648149):
    """
    คำนวณ max consecutive distance ของสาขาในทริป
    โดยเรียงลำดับสาขาตามเส้นทาง (Nearest Neighbor จาก DC)
    ไม่นับระยะจาก DC ไปสาขาแรก
    
    Returns:
        float: max consecutive distance (km), -1 ถ้าไม่มีข้อมูล
    """
    if len(codes) < 2:
        return 0.0
    
    # สร้างลิสต์พิกัด
    points = []
    for code in codes:
        lat, lon = coord_cache.get(code, (None, None))
        if lat and lon:
            points.append((code, lat, lon))
    
    if len(points) < 2:
        return 0.0
    
    # เรียงลำดับด้วย Nearest Neighbor จาก DC
    sorted_points = []
    remaining = points.copy()
    current_lat, current_lon = dc_lat, dc_lon
    
    while remaining:
        best_idx = 0
        best_dist = haversine_distance(current_lat, current_lon, remaining[0][1], remaining[0][2])
        
        for i, (_, lat, lon) in enumerate(remaining[1:], 1):
            dist = haversine_distance(current_lat, current_lon, lat, lon)
            if dist < best_dist:
                best_dist = dist
                best_idx = i
        
        best_point = remaining.pop(best_idx)
        sorted_points.append(best_point)
        current_lat, current_lon = best_point[1], best_point[2]
    
    # คำนวณ max consecutive distance (ไม่นับ DC ไปสาขาแรก)
    max_dist = 0.0
    for i in range(len(sorted_points) - 1):
        lat1, lon1 = sorted_points[i][1], sorted_points[i][2]
        lat2, lon2 = sorted_points[i + 1][1], sorted_points[i + 1][2]
        dist = haversine_distance(lat1, lon1, lat2, lon2)
        if dist > max_dist:
            max_dist = dist
    
    return max_dist

def get_region_type(province):
    """
    กำหนดประเภทพื้นที่ (ใช้เป็นข้อมูลประกอบ ไม่บังคับประเภทรถ)
    
    Returns:
        str: 'nearby' (กรุงเทพ+ปริมณฑล+ภาคกลาง),
             'north' (ภาคเหนือ),
             'south' (ภาคใต้),
             'far' (ภูมิภาคอื่น),
             'unknown'
    """
    if pd.isna(province):
        return 'unknown'
    
    prov = str(province).strip()
    
    # กรุงเทพ + ปริมณฑล + ภาคกลาง (ใช้เป็นข้อมูลประกอบ ไม่บังคับรถ)
    nearby_provinces = [
        'กรุงเทพมหานคร', 'กรุงเทพ',
        'นครปฐม', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ', 'สมุทรสาคร',
        'ชัยนาท', 'พระนครศรีอยุธยา', 'ลพบุรี', 'สระบุรี', 'สิงห์บุรี', 'อ่างทอง', 'อยุธยา',
        'สมุทรสงคราม', 'สุพรรณบุรี', 'นครนายก'
    ]
    
    for nearby in nearby_provinces:
        if nearby in prov:
            return 'nearby'
    
    # 🚛 ภาคเหนือทั้งหมด (18 จังหวัด) → บังคับใช้ 6W
    north_provinces = [
        # เหนือตอนบน
        'เชียงใหม่', 'เชียงราย', 'แม่ฮ่องสอน', 'น่าน', 'พะเยา', 'ลำปาง', 'ลำพูน', 'แพร่',
        # เหนือตอนล่าง
        'กำแพงเพชร', 'ตาก', 'นครสวรรค์', 'พิจิตร', 'พิษณุโลก', 'สุโขทัย', 
        'อุตรดิตถ์', 'อุทัยธานี', 'เพชรบูรณ์', 'ชัยภูมิ'
    ]
    
    for north in north_provinces:
        if north in prov:
            return 'north'
    
    # 🚛 ภาคใต้ทั้งหมด (14 จังหวัด) → บังคับใช้ 6W
    south_provinces = [
        # ฝั่งอันดามัน
        'กระบี่', 'ตรัง', 'พังงา', 'ภูเก็ต', 'ระนอง', 'สตูล',
        # ฝั่งอ่าวไทย
        'ชุมพร', 'นครศรีธรรมราช', 'พัทลุง', 'ยะลา', 'สงขลา', 
        'สุราษฎร์ธานี', 'ปัตตานี', 'นราธิวาส'
    ]
    
    for south in south_provinces:
        if south in prov:
            return 'south'
    
    # จังหวัดอื่นๆ (ตะวันออก, ตะวันตก, อีสาน) = ยืดหยุ่นตาม utilization
    return 'far'

def is_nearby_province(prov1, prov2):
    """เช็คว่าจังหวัดใกล้เคียงกันหรือไม่ (จากไฟล์ประวัติ)"""
    if pd.isna(prov1) or pd.isna(prov2):
        return False
    
    if prov1 == prov2:
        return True
    
    # จัดกลุ่มจังหวัดตามภาคย่อย (จากไฟล์ประวัติ)
    province_groups = {
        'กรุงเทพ': ['กรุงเทพมหานคร', 'กรุงเทพ'],
        'ปริมณฑล': ['นครปฐม', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ', 'สมุทรสาคร'],
        'กลางตอนบน': ['ชัยนาท', 'พระนครศรีอยุธยา', 'ลพบุรี', 'สระบุรี', 'สิงห์บุรี', 'อ่างทอง', 'อยุธยา'],
        'กลางตอนล่าง': ['สมุทรสงคราม', 'สุพรรณบุรี'],
        'ภาคตะวันตก': ['กาญจนบุรี', 'ประจวบคีรีขันธ์', 'ราชบุรี', 'เพชรบุรี'],
        'ภาคตะวันออก': ['จันทบุรี', 'ชลบุรี', 'ตราด', 'นครนายก', 'ปราจีนบุรี', 'ระยอง', 'สระแก้ว', 'ฉะเชิงเทรา'],
        'อีสานเหนือ': ['นครพนม', 'บึงกาฬ', 'มุกดาหาร', 'สกลนคร', 'หนองคาย', 'หนองบัวลำภู', 'อุดรธานี', 'เลย'],
        'อีสานกลาง': ['กาฬสินธุ์', 'ขอนแก่น', 'ชัยภูมิ', 'มหาสารคาม', 'ร้อยเอ็ด'],
        'อีสานใต้': ['นครราชสีมา', 'โคราช', 'บุรีรัมย์', 'ยโสธร', 'ศรีสะเกษ', 'สุรินทร์', 'อำนาจเจริญ', 'อุบลราชธานี'],
        'เหนือตอนบน': ['น่าน', 'พะเยา', 'ลำปาง', 'ลำพูน', 'เชียงราย', 'เชียงใหม่', 'แพร่', 'แม่ฮ่องสอน'],
        'เหนือตอนล่าง': ['กำแพงเพชร', 'ตาก', 'นครสวรรค์', 'พิจิตร', 'พิษณุโลก', 'สุโขทัย', 'อุตรดิตถ์', 'อุทัยธานี', 'เพชรบูรณ์'],
        'ใต้ฝั่งอันดามัน': ['กระบี่', 'ตรัง', 'พังงา', 'ภูเก็ต', 'ระนอง', 'สตูล'],
        'ใต้ฝั่งอ่าวไทย': ['ชุมพร', 'นครศรีธรรมราช', 'พัทลุง', 'ยะลา', 'สงขลา', 'สุราษฎร์ธานี', 'ปัตตานี', 'นราธิวาส']
    }
    
    # หาว่าจังหวัดทั้ง 2 อยู่กลุ่มเดียวกันหรือไม่
    for group, provinces in province_groups.items():
        in_group_1 = any(p in str(prov1) for p in provinces)
        in_group_2 = any(p in str(prov2) for p in provinces)
        
        if in_group_1 and in_group_2:
            return True
    
    return False

def load_model():
    """โหลดโมเดลที่เทรนไว้"""
    if not os.path.exists(MODEL_PATH):
        return None
    
    try:
        with open(MODEL_PATH, 'rb') as f:
            model_data = pickle.load(f)
        return model_data
    except Exception as e:
        st.error(f"❌ Error loading model: {e}")
        return None

def create_pair_features(code1, code2, branch_info):
    """สร้าง features สำหรับคู่สาขา"""
    import math
    
    info1 = branch_info[code1]
    info2 = branch_info[code2]
    
    # คำนวณความต่างของน้ำหนักและคิว
    weight_diff = abs(info1['avg_weight'] - info2['avg_weight'])
    cube_diff = abs(info1['avg_cube'] - info2['avg_cube'])
    weight_sum = info1['avg_weight'] + info2['avg_weight']
    cube_sum = info1['avg_cube'] + info2['avg_cube']
    
    # จังหวัดเดียวกันหรือไม่
    same_province = 1 if info1['province'] == info2['province'] else 0
    
    # คำนวณระยะทางจากพิกัด
    distance_km = 0.0
    if info1['latitude'] != 0 and info2['latitude'] != 0:
        lat1, lon1 = math.radians(info1['latitude']), math.radians(info1['longitude'])
        lat2, lon2 = math.radians(info2['latitude']), math.radians(info2['longitude'])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
        c = 2 * math.asin(math.sqrt(a))
        distance_km = 6371 * c
    
    # ความถี่
    freq_product = info1['total_trips'] * info2['total_trips']
    freq_diff = abs(info1['total_trips'] - info2['total_trips'])
    
    # Ratio
    weight_ratio = (info1['avg_weight'] / info2['avg_weight']) if info2['avg_weight'] > 0 else 0
    cube_ratio = (info1['avg_cube'] / info2['avg_cube']) if info2['avg_cube'] > 0 else 0
    
    # ข้อจำกัดรถ
    over_4w = 1 if (weight_sum > 2500 or cube_sum > 5.0) else 0
    over_jb = 1 if (weight_sum > 3500 or cube_sum > 7.0) else 0
    over_6w = 1 if (weight_sum > 5800 or cube_sum > 22.0) else 0
    
    return {
        'weight_sum': weight_sum,
        'cube_sum': cube_sum,
        'weight_diff': weight_diff,
        'cube_diff': cube_diff,
        'same_province': same_province,
        'distance_km': distance_km,
        'avg_weight_1': info1['avg_weight'],
        'avg_weight_2': info2['avg_weight'],
        'avg_cube_1': info1['avg_cube'],
        'avg_cube_2': info2['avg_cube'],
        'freq_product': freq_product,
        'freq_diff': freq_diff,
        'weight_ratio': weight_ratio,
        'cube_ratio': cube_ratio,
        'over_4w': over_4w,
        'over_jb': over_jb,
        'over_6w': over_6w
    }

def load_info_sheet_truck_limits(xls):
    """
    🔒 อ่านชีต info จากไฟล์ Auto Plan เพื่อดึง MaxTruckType
    
    คอลัมน์ที่ต้องการ:
    - Location Code: รหัสสาขา
    - MaxTruckType: ประเภทรถบรรทุกใหญ่สุด (เช่น 4W, 6W, 10W)
    """
    global AUTO_PLAN_TRUCK_LIMITS
    
    try:
        # หาชีต info
        info_sheet = None
        for s in xls.sheet_names:
            if 'info' in s.lower():
                info_sheet = s
                break
        
        if not info_sheet:
            return  # ไม่มีชีต info
        
        # อ่านชีต info
        df_info = pd.read_excel(xls, sheet_name=info_sheet)
        
        # หาคอลัมน์ Location Code และ MaxTruckType
        location_col = None
        truck_col = None
        
        for col in df_info.columns:
            col_str = str(col).lower()
            if 'location' in col_str and 'code' in col_str:
                location_col = col
            elif 'maxtruck' in col_str or 'max truck' in col_str:
                truck_col = col
        
        if location_col and truck_col:
            # สร้าง mapping
            AUTO_PLAN_TRUCK_LIMITS.clear()
            
            for _, row in df_info.iterrows():
                code = str(row[location_col]).strip()
                truck_type = str(row[truck_col]).strip().upper()
                
                if code and truck_type and code != 'nan' and truck_type != 'NAN':
                    # 🔒 ตรวจสอบว่ามี 6W ในเซลหรือไม่ (ถ้าไม่มี = ห้าม 6W)
                    # ตัวอย่าง: "6W" = ใช้ได้ทุกประเภท, "4WJB" = ห้าม 6W (ใช้ได้แค่ 4W, JB)
                    has_6w = '6W' in truck_type or '10W' in truck_type
                    has_jb = 'JB' in truck_type or 'JUMBO' in truck_type
                    has_4w = '4W' in truck_type
                    
                    # กำหนด max vehicle ตามที่ระบุในไฟล์
                    if has_6w:
                        # มี 6W หรือ 10W = ใช้ได้ทุกประเภท
                        normalized_truck = '6W'
                    elif has_jb or truck_type == '4WJB':
                        # มี JB หรือ "4WJB" แต่ไม่มี 6W = ห้าม 6W (ใช้ได้แค่ 4W, JB)
                        normalized_truck = 'JB'
                    elif has_4w:
                        # มีแค่ 4W = ห้าม 6W และ JB
                        normalized_truck = '4W'
                    else:
                        # ไม่ระบุ = ห้าม 6W (default เป็น JB)
                        normalized_truck = 'JB'
                    
                    AUTO_PLAN_TRUCK_LIMITS[code] = normalized_truck
            
            if AUTO_PLAN_TRUCK_LIMITS:
                # นับจำนวนแต่ละประเภท
                count_4w = sum(1 for v in AUTO_PLAN_TRUCK_LIMITS.values() if v == '4W')
                count_jb = sum(1 for v in AUTO_PLAN_TRUCK_LIMITS.values() if v == 'JB')
                count_6w = sum(1 for v in AUTO_PLAN_TRUCK_LIMITS.values() if v == '6W')
                st.info(f"📋 โหลดข้อจำกัดรถจากชีต info: {len(AUTO_PLAN_TRUCK_LIMITS)} สาขา (4W: {count_4w}, JB: {count_jb}, 6W: {count_6w})")
                
    except Exception as e:
        # ไม่มีชีต info หรืออ่านไม่ได้ → ไม่เป็นไร ใช้ default
        pass

def load_excel(file_content, sheet_name=None):
    """โหลด Excel และอ่านข้อจำกัดรถจากชีต info"""
    global AUTO_PLAN_TRUCK_LIMITS
    
    try:
        xls = pd.ExcelFile(io.BytesIO(file_content))
        
        # 🔒 ลองอ่านชีต info เพื่อดึง MaxTruckType
        load_info_sheet_truck_limits(xls)
        
        target_sheet = None
        if sheet_name and sheet_name in xls.sheet_names:
            target_sheet = sheet_name
        else:
            for s in xls.sheet_names:
                if 'punthai' in s.lower() or '2.' in s.lower():
                    target_sheet = s
                    break
        
        if not target_sheet:
            target_sheet = xls.sheet_names[0]
        
        # หา header row
        df_temp = pd.read_excel(xls, sheet_name=target_sheet, header=None)
        header_row = 0
        
        for i in range(min(10, len(df_temp))):
            row_values = df_temp.iloc[i].astype(str).str.upper()
            match_count = sum([
                'BRANCH' in ' '.join(row_values),
                'TRIP' in ' '.join(row_values),
                'รหัสสาขา' in ' '.join(df_temp.iloc[i].astype(str))
            ])
            if match_count >= 2:
                header_row = i
                break
        
        df = pd.read_excel(xls, sheet_name=target_sheet, header=header_row)
        df = df.loc[:, ~df.columns.duplicated()]
        
        return df
    except Exception as e:
        st.error(f"❌ Error: {e}")
        return None

def process_dataframe(df):
    """แปลงคอลัมน์เป็นรูปแบบมาตรฐาน"""
    if df is None:
        return None
    
    rename_map = {}
    
    # ถ้ามีคอลัมน์น้อยกว่า 15 = ใช้ลำดับคอลัมน์
    # ลำดับมาตรฐาน: Sep, BU, รหัสสาขา, รหัส WMS, สาขา, Total Cube, Total Wgt, จำนวนชิ้น, Trip, Trip no, ...
    if len(df.columns) >= 8:
        col_list = list(df.columns)
        # ลำดับ 1 = BU
        if len(col_list) > 1:
            rename_map[col_list[1]] = 'BU'
        # ลำดับ 2 = รหัสสาขา
        if len(col_list) > 2:
            rename_map[col_list[2]] = 'Code'
        # ลำดับ 4 = สาขา/ชื่อ
        if len(col_list) > 4:
            rename_map[col_list[4]] = 'Name'
        # ลำดับ 5 = Total Cube
        if len(col_list) > 5:
            rename_map[col_list[5]] = 'Cube'
        # ลำดับ 6 = Total Wgt
        if len(col_list) > 6:
            rename_map[col_list[6]] = 'Weight'
        # ลำดับ 7 = จำนวนชิ้น (Original QTY)
        if len(col_list) > 7:
            rename_map[col_list[7]] = 'OriginalQty'
        # ลำดับ 8 = Trip
        if len(col_list) > 8:
            rename_map[col_list[8]] = 'Trip'
        # ลำดับ 9 = Trip no
        if len(col_list) > 9:
            rename_map[col_list[9]] = 'TripNo'
    
    # ตรวจสอบเพิ่มเติมจากชื่อคอลัมน์
    for col in df.columns:
        if col in rename_map:
            continue
        col_clean = str(col).strip()
        col_upper = col_clean.upper().replace(' ', '').replace('_', '')
        
        if col_clean == 'BranchCode' or 'รหัสสาขา' in col_clean or col_clean == 'รหัส WMS' or 'BRANCH_CODE' in col_upper:
            rename_map[col] = 'Code'
        elif col_clean == 'Branch' or 'ชื่อสาขา' in col_clean or col_clean == 'สาขา' or 'BRANCH' in col_upper:
            rename_map[col] = 'Name'
        elif 'TOTALWGT' in col_upper or 'น้ำหนัก' in col_clean or 'WGT' in col_upper or 'WEIGHT' in col_upper:
            rename_map[col] = 'Weight'
        elif 'TOTALCUBE' in col_upper or 'คิว' in col_clean or 'CUBE' in col_upper:
            rename_map[col] = 'Cube'
        elif 'latitude' in col_clean.lower() or col_clean == 'ละติจูด' or 'LAT' in col_upper:
            rename_map[col] = 'Latitude'
        elif 'longitude' in col_clean.lower() or col_clean == 'ลองติจูด' or 'LONG' in col_upper or 'LNG' in col_upper:
            rename_map[col] = 'Longitude'
        elif 'จังหวัด' in col_clean or 'PROVINCE' in col_upper:
            rename_map[col] = 'Province'
        elif col_upper in ['TRIPNO', 'TRIP_NO'] or col_clean == 'Trip no':
            rename_map[col] = 'TripNo'
        elif col_upper == 'TRIP' or 'ทริป' in col_clean or 'เที่ยว' in col_clean:
            rename_map[col] = 'Trip'
        elif 'BOOKING' in col_upper:
            rename_map[col] = 'Booking'
    
    df = df.rename(columns=rename_map)
    
    # ลบคอลัมน์ซ้ำ
    df = df.loc[:, ~df.columns.duplicated()]
    
    if 'Code' in df.columns:
        df['Code'] = df['Code'].apply(normalize)
        
        # 🔒 กรองข้อมูลที่ไม่ถูกต้องออก (NaN, nan, NAN, ว่าง)
        df = df[df['Code'].notna()]  # กรอง NaN
        df = df[df['Code'].astype(str).str.upper() != 'NAN']  # กรอง "NAN", "nan"
        df = df[df['Code'].astype(str).str.strip() != '']  # กรองค่าว่าง
        
        # ตัดสาขาที่ไม่ต้องการออก (รหัส)
        df = df[~df['Code'].isin(EXCLUDE_BRANCHES)]
        
        # ตัดสาขาที่ชื่อมี keyword ที่ไม่ต้องการ
        if 'Name' in df.columns:
            exclude_pattern = '|'.join(EXCLUDE_NAMES)
            df = df[~df['Name'].str.contains(exclude_pattern, case=False, na=False)]
            # 🔒 กรองชื่อสาขาที่เป็น nan ออกด้วย
            df = df[df['Name'].notna()]
            df = df[df['Name'].astype(str).str.lower() != 'nan']
    
    for col in ['Weight', 'Cube']:
        if col not in df.columns:
            df[col] = 0.0
        else:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    
    # เพิ่มจังหวัดจาก Master ถ้ายังไม่มี
    if 'Province' not in df.columns or df['Province'].isna().all():
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns and 'Code' in df.columns:
            # สร้าง mapping จาก Master
            province_map = {}
            for _, row in MASTER_DATA.iterrows():
                code = row.get('Plan Code', '')
                province = row.get('จังหวัด', '')
                if code and province:
                    province_map[code] = province
            
            # ฟังก์ชันค้นหาจังหวัดจากชื่อสาขา
            def find_province_by_name(code, name):
                # ลองหาจาก code ก่อน
                if code in province_map:
                    return province_map[code]
                
                # ถ้าไม่เจอ ลองค้นหาจากชื่อสาขา
                if not name or pd.isna(name):
                    return None
                
                # แยกคำสำคัญจากชื่อ (เอาคำแรกที่ไม่ใช่ prefix)
                keywords = str(name).replace('MAX MART-', '').replace('PUNTHAI-', '').replace('LUBE', '').strip()
                if not keywords:
                    return None
                
                # ค้นหาในชื่อสาขาของ Master
                for _, master_row in MASTER_DATA.iterrows():
                    master_name = str(master_row.get('สาขา', ''))
                    # ถ้าชื่อคล้ายกัน (มีคำสำคัญเหมือนกัน)
                    if keywords[:10] in master_name or master_name[:10] in keywords:
                        province = master_row.get('จังหวัด', '')
                        if province:
                            return province
                
                return None
            
            # ใส่จังหวัดให้แต่ละสาขา
            if 'Name' in df.columns:
                df['Province'] = df.apply(lambda row: find_province_by_name(row['Code'], row.get('Name', '')), axis=1)
            else:
                df['Province'] = df['Code'].map(province_map)
    
    return df.reset_index(drop=True)

def predict_trips(test_df, model_data):
    """
    จัดทริปด้วยระบบอัจฉริยะ โดยใช้กฎลำดับความสำคัญ:
    
    เงื่อนไขบังคับ (ต้องผ่านก่อน):
    0. ✅ จำกัดจำนวนสาขาต่อทริป (สูงสุด 12 สาขา)
    0. ✅ เช็คจังหวัดใกล้เคียงกัน (ต้องอยู่กลุ่มเดียวกัน)
    0. ✅ ถ้าเกิน 10 สาขา ต้องเป็นจังหวัดเดียวกันเท่านั้น
    0. ✅ ตรวจสอบว่าสาขาสามารถใช้รถประเภทนี้ได้หรือไม่จากประวัติ
    
    กฎการจับคู่ (หลังผ่านเงื่อนไขบังคับ):
    1. ✅ เคยไปด้วยกันในประวัติ (trip_pairs) + ใช้รถแบบเดิม
    2. ✅ ชื่อสาขาคล้ายกัน (เช่น นครราชสีมา1, นครราชสีมา2)
    3. ✅ AI ทำนายจาก Decision Tree Model
    4. ✅ เช็คน้ำหนัก/คิว ไม่เกินขีดจำกัดรถ
    """
    # ตรวจสอบว่า model_data มีข้อมูลครบถ้วน
    if not model_data or not isinstance(model_data, dict):
        st.error("❌ ข้อมูลโมเดลไม่ถูกต้อง กรุณาเทรนโมเดลก่อน")
        return test_df, []
    
    model = model_data.get('model')
    trip_pairs = model_data.get('trip_pairs', set()).copy()  # คัดลอกเพื่อไม่ให้กระทบต้นฉบับ
    branch_info = model_data.get('branch_info', {})
    trip_vehicles = model_data.get('trip_vehicles', {}).copy()
    branch_vehicles = model_data.get('branch_vehicles', {})
    
    # ⚡ อนุญาตให้ทำงานโดยไม่มี model (ใช้กฎและประวัติเท่านั้น)
    # if model is None:
    #     st.error("❌ ไม่พบโมเดล กรุณาเทรนโมเดลก่อน")
    #     return test_df, []
    
    # ★ ถ้าไฟล์อัปโหลดมีคอลัมน์ Trip ให้ใช้เป็นข้อมูลอ้างอิงหลัก
    # เพราะเป็นแผนงานที่ใช้จริงมาแล้ว
    file_trip_vehicles = {}  # เก็บประเภทรถจากไฟล์แผนงาน
    use_file_trips = False  # ใช้ทริปจากไฟล์โดยตรง
    
    if 'Trip' in test_df.columns and test_df['Trip'].notna().any():
        use_file_trips = True
        st.info(f"📋 พบคอลัมน์ทริปในไฟล์ - ใช้การจัดทริปจากไฟล์โดยตรง")
        
        # สร้าง trip_pairs จากไฟล์แผนงาน
        for trip_id, group in test_df.groupby('Trip'):
            if pd.isna(trip_id):
                continue
            codes = group['Code'].unique().tolist()
            
            # ดึงประเภทรถจาก TripNo (เช่น 4W009 -> 4W, JB014 -> JB)
            if 'TripNo' in group.columns:
                trip_no = group['TripNo'].iloc[0]
                if pd.notna(trip_no):
                    trip_no_str = str(trip_no).strip()
                    if trip_no_str.startswith('4W'):
                        vehicle_type = '4W'
                    elif trip_no_str.startswith('JB'):
                        vehicle_type = 'JB'
                    elif trip_no_str.startswith('6W'):
                        vehicle_type = '6W'
                    else:
                        vehicle_type = None
                    
                    # เก็บประเภทรถสำหรับแต่ละคู่สาขา
                    if vehicle_type:
                        for i in range(len(codes)):
                            for j in range(i+1, len(codes)):
                                pair = tuple(sorted([codes[i], codes[j]]))
                                file_trip_vehicles[pair] = vehicle_type
            
            # สร้างคู่สาขาในทริปเดียวกัน
            for i in range(len(codes)):
                for j in range(i+1, len(codes)):
                    pair = tuple(sorted([codes[i], codes[j]]))
                    trip_pairs.add(pair)  # เพิ่มเข้า trip_pairs
    
    # รวม file_trip_vehicles เข้ากับ trip_vehicles (ให้ไฟล์มีความสำคัญกว่า)
    for pair, vehicle in file_trip_vehicles.items():
        trip_vehicles[pair] = {'most_used': vehicle, 'vehicle': vehicle}
    
    # เพิ่มสาขาใหม่
    for code in test_df['Code'].unique():
        if code not in branch_info:
            code_data = test_df[test_df['Code'] == code]
            branch_info[code] = {
                'avg_weight': code_data['Weight'].mean(),
                'avg_cube': code_data['Cube'].mean(),
                'total_trips': 1,
                'province': code_data['Province'].iloc[0] if 'Province' in code_data.columns else 'UNKNOWN',
                'latitude': 0.0,
                'longitude': 0.0
            }
    
    # 🔒 Final enforcement of vehicle constraints (ต้องนิยามก่อนเรียกใช้)
    def enforce_vehicle_constraints(test_df_input):
        """บังคับข้อจำกัดรถขั้นสุดท้าย - ไม่อนุญาต 6W หากสาขาจำกัด 4W/JB หรืออยู่ในกรุงเทพ/ปริมณฑล"""
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for trip_num in test_df_input['Trip'].unique():
            if pd.isna(trip_num):
                continue
            trip_data = test_df_input[test_df_input['Trip'] == trip_num]
            trip_codes = trip_data['Code'].unique()
            
            # 🔒 เช็คจังหวัด - ห้าม 6W ถ้ามีแม้แค่สาขาเดียวอยู่ในกรุงเทพ/ปริมณฑล!
            has_any_nearby = False  # เปลี่ยนจาก all_nearby เป็น has_any_nearby
            for code in trip_codes:
                prov = None
                # หาจังหวัดจาก test_df_input
                code_data = test_df_input[test_df_input['Code'] == code]
                if 'Province' in code_data.columns and len(code_data) > 0:
                    prov = code_data['Province'].iloc[0]
                # หาจาก MASTER_DATA ถ้าไม่มี
                if not prov and not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                    master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                    if len(master_row) > 0:
                        prov = master_row.iloc[0].get('จังหวัด', '')
                if prov and prov != 'UNKNOWN' and get_region_type(str(prov)) == 'nearby':
                    has_any_nearby = True
                    break  # พบแม้แค่สาขาเดียวก็พอ
            
            # ตรวจสอบข้อจำกัดที่เข้มงวดที่สุดในทริป
            max_vehicles = []
            for code in trip_codes:
                max_vehicle = get_max_vehicle_for_branch(code)
                max_vehicles.append(max_vehicle)
            
            min_max_size = min(vehicle_sizes.get(v, 3) for v in max_vehicles) if max_vehicles else 3
            
            # 🔒 กรุงเทพ/ปริมณฑล = บังคับ JB หรือเล็กกว่า (ห้าม 6W เด็ดขาด!)
            if has_any_nearby and min_max_size == 3:
                min_max_size = 2  # บังคับลงมาเป็น JB
            
            # หากมีสาขาใดจำกัด 4W/JB หรืออยู่ในกรุงเทพ/ปริมณฑล → ห้าม 6W
            if min_max_size < 3:
                # บังคับเปลี่ยนเป็น JB หรือ 4W
                allowed_vehicle = 'JB' if min_max_size >= 2 else '4W'
                current_truck = test_df_input.loc[test_df_input['Trip'] == trip_num, 'Truck'].iloc[0] if len(test_df_input[test_df_input['Trip'] == trip_num]) > 0 else ''
                if '6W' in str(current_truck):
                    reason = 'กทม/ปริมณฑล' if has_any_nearby else 'บังคับสาขา'
                    test_df_input.loc[test_df_input['Trip'] == trip_num, 'Truck'] = f'{allowed_vehicle} 🔒 {reason}'
        
        return test_df_input
    
    # 🔒 ฟังก์ชันแยกสาขาที่มีข้อจำกัดรถออกจากทริปที่ใช้รถใหญ่เกินไป
    def split_restricted_branches(df):
        """แยกสาขาที่จำกัด 4W/JB ออกจากทริปที่มี Cube/Weight เกินความจุรถที่อนุญาต"""
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        for trip_num in df['Trip'].dropna().unique():
            trip_data = df[df['Trip'] == trip_num]
            total_w = trip_data['Weight'].sum()
            total_c = trip_data['Cube'].sum()
            trip_codes = list(trip_data['Code'].unique())
            
            # หาสาขาที่มีข้อจำกัดและไม่มีข้อจำกัด
            codes_4w_only = []  # จำกัด 4W
            codes_jb_only = []  # จำกัด JB
            codes_no_limit = []  # ไม่จำกัด
            
            for code in trip_codes:
                max_v = get_max_vehicle_for_branch(code)
                if max_v == '4W':
                    codes_4w_only.append(code)
                elif max_v == 'JB':
                    codes_jb_only.append(code)
                else:
                    codes_no_limit.append(code)
            
            # 🔴 ถ้ามีสาขาจำกัด 4W แต่ Cube รวมเกิน 4W capacity (5.0) → แยก
            if codes_4w_only and total_c > LIMITS['4W']['max_c']:
                # สาขาที่จำกัด 4W ต้องแยกออกไปทริปใหม่
                new_trip_num = df['Trip'].max() + 1
                for code in codes_4w_only:
                    df.loc[df['Code'] == code, 'Trip'] = new_trip_num
            
            # 🔴 ถ้ามีสาขาจำกัด JB แต่ Cube รวมเกิน JB capacity (7.0) → แยก
            elif codes_jb_only and total_c > LIMITS['JB']['max_c']:
                # ถ้ามีทั้งสาขาจำกัด JB และไม่จำกัด → แยกสาขาที่จำกัด JB ออก
                if codes_no_limit:
                    new_trip_num = df['Trip'].max() + 1
                    for code in codes_jb_only:
                        df.loc[df['Code'] == code, 'Trip'] = new_trip_num
        
        return df
    
    # ★★★ ถ้ามีคอลัมน์ Trip ในไฟล์ ใช้โดยตรงเลย ★★★
    if use_file_trips:
        # ใช้ Trip จากไฟล์โดยตรง
        test_df_result = test_df.copy()
        
        # 🔒 แยกสาขาที่มีข้อจำกัดรถออกก่อน
        test_df_result = split_restricted_branches(test_df_result)
        
        # ดึงประเภทรถจาก TripNo
        trip_truck_map_file = {}
        if 'TripNo' in test_df.columns:
            for trip_id in test_df['Trip'].dropna().unique():
                trip_data = test_df[test_df['Trip'] == trip_id]
                if 'TripNo' in trip_data.columns and len(trip_data) > 0:
                    trip_no = trip_data['TripNo'].iloc[0]
                    if pd.notna(trip_no):
                        trip_no_str = str(trip_no).strip()
                        if trip_no_str.startswith('4W'):
                            trip_truck_map_file[trip_id] = '4W'
                        elif trip_no_str.startswith('JB'):
                            trip_truck_map_file[trip_id] = 'JB'
                        elif trip_no_str.startswith('6W'):
                            trip_truck_map_file[trip_id] = '6W'
        
        # สร้าง summary
        summary_data = []
        for trip_num in sorted(test_df_result['Trip'].dropna().unique()):
            trip_data = test_df_result[test_df_result['Trip'] == trip_num]
            total_w = trip_data['Weight'].sum()
            total_c = trip_data['Cube'].sum()
            
            # ตรวจสอบข้อจำกัดของสาขาในทริป
            trip_codes = trip_data['Code'].unique()
            max_vehicles = []
            has_any_nearby_branch = False  # เช็คกรุงเทพ/ปริมณฑล
            for c in trip_codes:
                max_vehicles.append(get_max_vehicle_for_branch(c))
                # เช็คว่าสาขาอยู่ในกรุงเทพ/ปริมณฑลไหม
                code_data = trip_data[trip_data['Code'] == c]
                if 'Province' in code_data.columns and len(code_data) > 0:
                    prov = code_data['Province'].iloc[0]
                    if prov and pd.notna(prov) and get_region_type(str(prov)) == 'nearby':
                        has_any_nearby_branch = True
            
            vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
            min_max_size = min(vehicle_sizes.get(v, 3) for v in max_vehicles) if max_vehicles else 3
            
            # 🔒 กรุงเทพ/ปริมณฑล → บังคับ JB หรือ 4W (ห้าม 6W เด็ดขาด!)
            if has_any_nearby_branch and min_max_size == 3:
                min_max_size = 2  # บังคับลงมาเป็น JB
            
            max_allowed_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(min_max_size, '6W')
            

            # 🚨 STRICT: Branch constraint (🔒) > History (📜) > AI (🤖)
            # 1. Branch constraint (never allow 6W if any branch restricts to 4W/JB)
            if min_max_size < 3:  # 1=4W, 2=JB
                # Only allow 4W/JB, never 6W
                allowed = ['JB', '4W'] if min_max_size == 2 else ['4W']
            else:
                allowed = ['JB', '4W', '6W']

            if trip_num in trip_truck_map_file:
                suggested = trip_truck_map_file[trip_num]
                # If suggested vehicle is not allowed, override to strictest allowed
                if suggested not in allowed:
                    suggested = allowed[0]
                    reason = 'กทม/ปริมณฑล' if has_any_nearby_branch else 'จำกัดสาขา'
                    source = f"📋 ไฟล์ → {suggested} (🔒 {reason})"
                else:
                    source = "📋 ไฟล์"
            else:
                # AI suggestion, but must respect allowed
                ai_suggested = suggest_truck(total_w, total_c, max_allowed_vehicle, trip_codes)
                if ai_suggested not in allowed:
                    suggested = allowed[0]
                    reason = 'กทม/ปริมณฑล' if has_any_nearby_branch else 'จำกัดสาขา'
                    source = f"🤖 AI → {suggested} (🔒 {reason})"
                else:
                    suggested = ai_suggested
                    source = "🤖 AI"

            # Double check: If strict constraint, never allow 6W even if utilization >105%
            if min_max_size < 3:
                # Only JB or 4W allowed, never 6W
                if suggested == '6W':
                    # fallback to JB if possible, else 4W
                    if 'JB' in allowed:
                        suggested = 'JB'
                        reason = 'กทม/ปริมณฑล' if has_any_nearby_branch else 'จำกัดสาขา'
                        source = source + f" (🔒 {reason})"
                    else:
                        suggested = '4W'
                        source = source + " (🔒 จำกัดสาขา)"

            # ตรวจสอบว่ารถที่เลือกใส่ของได้จริงหรือไม่ (ห้ามเกิน 100%)
            if suggested in LIMITS:
                w_util = (total_w / LIMITS[suggested]['max_w']) * 100
                c_util = (total_c / LIMITS[suggested]['max_c']) * 100
                max_util = max(w_util, c_util)

                # ถ้าเกิน 100% ต้องเพิ่มขนาดรถ
                if max_util > 100:
                    # ถ้ามีข้อจำกัดสาขา ห้ามขยายเป็น 6W
                    if min_max_size < 3:
                        # บังคับ JB หรือ 4W เท่านั้น
                        if 'JB' in allowed and suggested == '4W':
                            jb_w_util = (total_w / LIMITS['JB']['max_w']) * 100
                            jb_c_util = (total_c / LIMITS['JB']['max_c']) * 100
                            if max(jb_w_util, jb_c_util) <= 100:
                                suggested = 'JB'
                                source = source + " → JB"
                                w_util, c_util = jb_w_util, jb_c_util
                            else:
                                # JB ก็ยังเกิน → ให้เตือน
                                suggested = 'JB'
                                source = source + " → JB (🚫 เกินแต่ห้าม 6W)"
                                w_util, c_util = jb_w_util, jb_c_util
                        # ถ้า JB ก็ยังเกิน ให้เตือนว่าเกิน ไม่ขยายเป็น 6W
                        elif suggested == 'JB':
                            source = source + " (🚫 เกินขนาดแต่ห้ามใช้ 6W - ต้องแยกทริป)"
                        elif suggested == '4W' and '4W' in allowed and 'JB' not in allowed:
                            # สาขาจำกัด 4W เท่านั้น แต่เกิน → ต้องแยกทริป
                            source = source + " (🚫 4W เกินขนาด - ต้องแยกทริป)"
                    else:
                        # ไม่มีข้อจำกัดสาขา สามารถขยายเป็น 6W ได้
                        if suggested == '4W' and 'JB' in LIMITS:
                            jb_w_util = (total_w / LIMITS['JB']['max_w']) * 100
                            jb_c_util = (total_c / LIMITS['JB']['max_c']) * 100
                            if max(jb_w_util, jb_c_util) <= 100:
                                suggested = 'JB'
                                source = source + " → JB"
                                w_util, c_util = jb_w_util, jb_c_util
                            else:
                                suggested = '6W'
                                source = source + " → 6W"
                                w_util = (total_w / LIMITS['6W']['max_w']) * 100
                                c_util = (total_c / LIMITS['6W']['max_c']) * 100
                        elif suggested == 'JB' or suggested == '4W':
                            suggested = '6W'
                            source = source + " → 6W"
                            w_util = (total_w / LIMITS['6W']['max_w']) * 100
                            c_util = (total_c / LIMITS['6W']['max_c']) * 100
            else:
                w_util = c_util = 0
            
            # ⚡ Skip distance calculation completely for speed optimization
            trip_codes = trip_data['Code'].unique()
            total_distance = 0  # Skip all distance calculations
            
            # 🔴 ตรวจสอบว่าทริปผ่านเกณฑ์หรือไม่
            max_util_check = max(w_util, c_util)
            trip_issues = []
            
            # ปัญหา 1: เกิน 100%
            if max_util_check > 100:
                trip_issues.append(f'⛔ เกิน {max_util_check:.0f}%')
                # ถ้าเกินแล้วห้ามใช้รถใหญ่กว่า → ต้องแยกทริป
                if min_max_size < 3:
                    trip_issues.append('🔧 ต้องแยกทริป')
            
            # ปัญหา 2: ใช้รถผิดประเภท (6W ในเขตต้องห้าม)
            if suggested == '6W' and has_any_nearby_branch:
                trip_issues.append('⛔ 6W ในกทม/ปริมณฑล')
            
            # ปัญหา 3: ใช้รถใหญ่กว่าที่สาขาอนุญาต
            if suggested == '6W' and min_max_size < 3:
                trip_issues.append('⛔ 6W ในสาขาจำกัด')
            elif suggested == 'JB' and min_max_size < 2:
                trip_issues.append('⛔ JB ในสาขาจำกัด 4W')
            
            trip_status = '❌ ไม่ผ่าน: ' + ', '.join(trip_issues) if trip_issues else '✅ ผ่าน'
            
            summary_data.append({
                'Trip': int(trip_num),
                'Branches': len(trip_data['Code'].unique()),
                'Weight': total_w,
                'Cube': total_c,
                'Truck': f"{suggested} {source}",
                'Weight_Use%': w_util,
                'Cube_Use%': c_util,
                'Total_Distance': total_distance,
                'TripStatus': trip_status
            })
        

        summary_df = pd.DataFrame(summary_data)

        # 🚨 Double Check: No trip uses a vehicle larger than allowed by any branch
        for idx, row in summary_df.iterrows():
            trip_num = row['Trip']
            trip_codes = test_df_result[test_df_result['Trip'] == trip_num]['Code'].unique()
            max_allowed = get_max_vehicle_for_trip(trip_codes)
            vehicle_type = row['Truck'].split()[0]
            vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
            if vehicle_sizes.get(vehicle_type, 3) > vehicle_sizes.get(max_allowed, 3):
                # Override to strictest allowed
                summary_df.at[idx, 'Truck'] = f"{max_allowed} 🔒 จำกัดสาขา"

        # เพิ่มคอลัมน์รถ
        trip_truck_display = {}
        for _, row in summary_df.iterrows():
            trip_truck_display[row['Trip']] = row['Truck']

        test_df_result['Truck'] = test_df_result['Trip'].map(trip_truck_display)
        # 🔒 Final enforcement: Never allow 6W if any branch restricts to 4W/JB
        test_df_result = enforce_vehicle_constraints(test_df_result)
        
        # 🔍 Validate trip grouping - เพิ่มคอลัมน์ TripValidation
        def validate_trip_grouping(row):
            """ตรวจสอบว่าสาขาในทริปนี้ควรอยู่ด้วยกันหรือไม่"""
            trip_num = row['Trip']
            code = row['Code']
            if pd.isna(trip_num):
                return '❓ ไม่มีทริป'
            
            trip_data = test_df_result[test_df_result['Trip'] == trip_num]
            trip_codes = [c for c in trip_data['Code'].unique() if c != code]
            
            if len(trip_codes) == 0:
                return '✅ สาขาเดี่ยว'
            
            issues = []
            valid_reasons = []
            
            # เช็ค 1: เคยไปด้วยกันในประวัติ?
            paired_with_history = False
            for other_code in trip_codes:
                pair = tuple(sorted([code, other_code]))
                if pair in trip_pairs:
                    paired_with_history = True
                    break
            
            if paired_with_history:
                valid_reasons.append('📜 ประวัติ')
            
            # เช็ค 2: Reference เดียวกัน?
            same_reference = False
            code_ref = LOCATION_CODE_TO_REF.get(code)
            if code_ref:
                for other_code in trip_codes:
                    other_ref = LOCATION_CODE_TO_REF.get(other_code)
                    if other_ref and code_ref == other_ref:
                        same_reference = True
                        break
            
            if same_reference:
                valid_reasons.append('🏠 Reference')
            
            # เช็ค 3: ตำบลเดียวกัน?
            same_subdistrict = False
            code_data = test_df_result[test_df_result['Code'] == code]
            my_subdistrict = code_data['Subdistrict'].iloc[0] if 'Subdistrict' in code_data.columns and len(code_data) > 0 else None
            if my_subdistrict and pd.notna(my_subdistrict):
                for other_code in trip_codes:
                    other_data = test_df_result[test_df_result['Code'] == other_code]
                    other_subdist = other_data['Subdistrict'].iloc[0] if 'Subdistrict' in other_data.columns and len(other_data) > 0 else None
                    if other_subdist and my_subdistrict == other_subdist:
                        same_subdistrict = True
                        break
            
            if same_subdistrict:
                valid_reasons.append('📍 ตำบล')
            
            # เช็ค 4: ระยะทางใกล้กันพอ?
            close_distance = False
            # ดึงพิกัดจาก MASTER_DATA
            code_lat, code_lon = None, None
            if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                if len(master_row) > 0:
                    code_lat = master_row.iloc[0].get('ละติจูด', None)
                    code_lon = master_row.iloc[0].get('ลองติจูด', None)
            
            if code_lat and code_lon and pd.notna(code_lat) and pd.notna(code_lon):
                for other_code in trip_codes:
                    other_lat, other_lon = None, None
                    if not MASTER_DATA.empty:
                        other_row = MASTER_DATA[MASTER_DATA['Plan Code'] == other_code]
                        if len(other_row) > 0:
                            other_lat = other_row.iloc[0].get('ละติจูด', None)
                            other_lon = other_row.iloc[0].get('ลองติจูด', None)
                    if other_lat and other_lon and pd.notna(other_lat) and pd.notna(other_lon):
                        try:
                            dist = haversine_distance(float(code_lat), float(code_lon), float(other_lat), float(other_lon))
                            if dist <= MAX_DISTANCE_IN_TRIP:
                                close_distance = True
                                break
                        except:
                            pass
            
            if close_distance:
                valid_reasons.append('📏 ใกล้')
            
            # 🔴 เช็ค 5: จังหวัดต่างกันมาก? (ภาคเหนือ + ภาคใต้ = ไม่ควรรวม)
            different_regions = False
            my_province = code_data['Province'].iloc[0] if 'Province' in code_data.columns and len(code_data) > 0 else None
            my_region = get_region_type(str(my_province)) if my_province and pd.notna(my_province) else None
            
            for other_code in trip_codes:
                other_data = test_df_result[test_df_result['Code'] == other_code]
                other_province = other_data['Province'].iloc[0] if 'Province' in other_data.columns and len(other_data) > 0 else None
                other_region = get_region_type(str(other_province)) if other_province and pd.notna(other_province) else None
                
                # เช็คว่าภาคต่างกันมากไหม (north + south = ไม่ควรรวม)
                if my_region and other_region:
                    if (my_region == 'north' and other_region == 'south') or (my_region == 'south' and other_region == 'north'):
                        different_regions = True
                        issues.append(f'🚫 {my_province} + {other_province}')
                        break
            
            # สรุปผล
            if issues:
                return '❌ ' + ', '.join(issues)
            elif valid_reasons:
                return '✅ ' + ', '.join(valid_reasons)
            else:
                return '⚠️ ไม่พบเหตุผลจับคู่'
        
        test_df_result['TripValidation'] = test_df_result.apply(validate_trip_grouping, axis=1)
        
        # 🆕 เพิ่มคอลัมน์แสดงสาขาที่อยู่ในทริปเดียวกัน
        def get_trip_partners(row):
            """แสดงสาขาอื่นที่อยู่ในทริปเดียวกัน"""
            trip_num = row['Trip']
            code = row['Code']
            if pd.isna(trip_num):
                return ''
            
            trip_data = test_df_result[test_df_result['Trip'] == trip_num]
            trip_codes = [c for c in trip_data['Code'].unique() if c != code]
            
            if len(trip_codes) == 0:
                return '(สาขาเดี่ยว)'
            
            # แสดงรหัสสาขาที่อยู่ด้วยกัน
            return ', '.join(trip_codes[:5]) + ('...' if len(trip_codes) > 5 else '')
        
        test_df_result['TripPartners'] = test_df_result.apply(get_trip_partners, axis=1)
        
        # 🔴 เพิ่มคอลัมน์ TripStatus จาก summary_df
        trip_status_map = {}
        for _, row in summary_df.iterrows():
            trip_status_map[row['Trip']] = row['TripStatus']
        test_df_result['TripStatus'] = test_df_result['Trip'].map(trip_status_map)
        
        # Mark VehicleCheck if strict constraint enforced
        def vehicle_check_str(row):
            truck = str(row.get('Truck', '')) if pd.notna(row.get('Truck')) else ''
            if '🔒' in truck or 'บังคับสาขา' in truck:
                return '🔒 จำกัดสาขา'
            return '✅ ใช้ตามไฟล์'
        test_df_result['VehicleCheck'] = test_df_result.apply(vehicle_check_str, axis=1)

        return test_df_result, summary_df
    
    # ถ้าไม่มีคอลัมน์ Trip ให้จัดทริปใหม่
    
    # 🗺️ จัดกลุ่มสาขาตามพิกัดก่อน (Spatial Clustering) + จับคู่สาขาชื่อคล้ายกัน
    def create_distance_based_clusters(codes, max_distance_km=25):
        """จัดกลุ่มสาขาที่อยู่ใกล้กัน (ไม่เกิน max_distance_km) + บังคับรวมสาขาชื่อคล้ายกัน + บังคับรวม location เดียวกัน"""
        # ⚡ Speed: Skip clustering if too few codes
        if len(codes) < 10:
            return [codes]  # Return all as one cluster
        
        # 🔥 Phase 0: จับคู่สาขาที่มี base code เดียวกัน (M862, P862, S862, ZS862 → ต้องอยู่ด้วยกัน)
        location_groups = {}  # base_code -> [codes]
        for code in codes:
            base = get_branch_base_code(code)
            if base:
                if base not in location_groups:
                    location_groups[base] = []
                location_groups[base].append(code)
        
        # สร้างกลุ่มสาขาที่เป็น location เดียวกัน (มี 2+ codes)
        same_location_groups = [group for group in location_groups.values() if len(group) > 1]
        grouped_by_location = set()
        for group in same_location_groups:
            for code in group:
                grouped_by_location.add(code)
        
        # 🔥 Phase 1: จับคู่สาขาที่มีชื่อคล้ายกัน (เช่น คลองหลวง 3,4,8,10) ให้อยู่กลุ่มเดียวกันเสมอ
        similar_groups = []  # เก็บกลุ่มสาขาที่ชื่อคล้ายกัน
        grouped_codes = set(grouped_by_location)  # เริ่มจากสาขาที่ถูกจัดกลุ่มด้วย location แล้ว
        
        # ตรวจสอบทุกคู่สาขาที่ยังไม่ถูกจัดกลุ่ม
        for i, code1 in enumerate(codes):
            if code1 in grouped_codes:
                continue
            
            # หาชื่อสาขา
            name1 = test_df[test_df['Code'] == code1]['Name'].iloc[0] if 'Name' in test_df.columns and len(test_df[test_df['Code'] == code1]) > 0 else ''
            
            # หาสาขาที่ชื่อคล้ายกัน
            similar_group = [code1]
            for j, code2 in enumerate(codes):
                if i >= j or code2 in grouped_codes:
                    continue
                
                name2 = test_df[test_df['Code'] == code2]['Name'].iloc[0] if 'Name' in test_df.columns and len(test_df[test_df['Code'] == code2]) > 0 else ''
                
                # ถ้าชื่อคล้ายกัน (เช่น "คลองหลวง" ใน "คลองหลวง 3", "คลองหลวง 4")
                if is_similar_name(name1, name2, similarity_threshold=75):  # ลดเกณฑ์เหลือ 75% เพื่อจับได้มากขึ้น
                    # เช็คระยะทาง - ยอมให้ไกลได้ถึง 80km (เพราะสาขาชื่อเดียวกันอาจกระจาย)
                    lat1, lon1 = get_lat_lon_from_master(code1)
                    lat2, lon2 = get_lat_lon_from_master(code2)
                    
                    if lat1 and lat2:
                        dist = haversine_distance(lat1, lon1, lat2, lon2)
                        if dist < 80:  # เพิ่มจาก 50km → 80km
                            similar_group.append(code2)
                            grouped_codes.add(code2)
                    else:
                        # ถ้าไม่มีพิกัด → รวมเข้ากลุ่มเลย (ถือว่าชื่อเดียวกัน)
                        similar_group.append(code2)
                        grouped_codes.add(code2)
            
            if len(similar_group) > 1:
                # มีสาขาชื่อคล้ายกัน → สร้างกลุ่ม
                similar_groups.append(similar_group)
                grouped_codes.add(code1)
        
        # สาขาที่เหลือ (ไม่มีชื่อคล้ายกัน และไม่ถูกจัดกลุ่มด้วย location)
        remaining_codes = [c for c in codes if c not in grouped_codes]
        
        clusters = []
        remaining = remaining_codes.copy()
        
        while remaining:
            # เริ่มกลุ่มใหม่
            seed = remaining.pop(0)
            cluster = [seed]
            seed_lat, seed_lon = get_lat_lon_from_master(seed)
            
            if seed_lat is None:
                # ไม่มีพิกัด → ใส่คลัสเตอร์เดี่ยว
                clusters.append(cluster)
                continue
            
            # หาสาขาที่ใกล้กับ seed
            to_remove = []
            for code in remaining[:]:
                lat, lon = get_lat_lon_from_master(code)
                if lat and lon:
                    dist = haversine_distance(seed_lat, seed_lon, lat, lon)
                    if dist <= max_distance_km:
                        cluster.append(code)
                        to_remove.append(code)
            
            # ลบสาขาที่เพิ่มไปแล้ว
            for code in to_remove:
                if code in remaining:
                    remaining.remove(code)
            
            clusters.append(cluster)
        
        # 🔥 เพิ่มกลุ่มสาขา location เดียวกัน + ชื่อคล้ายกันเข้าไป (จะอยู่ข้างหน้าสุด - ส่งก่อน)
        all_clusters = same_location_groups + similar_groups + clusters
        
        return all_clusters
    
    def get_lat_lon_from_master(code):
        """ดึงพิกัดจาก Master Data"""
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
            if len(master_row) > 0:
                lat = master_row.iloc[0].get('ละติจูด', None)
                lon = master_row.iloc[0].get('ลองติจูด', None)
                if pd.notna(lat) and pd.notna(lon) and lat != 0 and lon != 0:
                    try:
                        return float(lat), float(lon)
                    except:
                        pass
        return None, None
    
    def build_route_nearest_neighbor(codes):
        """สร้างเส้นทางโดยเลือกสาขาที่ใกล้ที่สุดถัดไป (Nearest Neighbor)"""
        if len(codes) <= 1:
            return codes
        
        # เริ่มจาก DC
        route = []
        remaining = codes.copy()
        current_lat, current_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
        
        while remaining:
            # หาสาขาที่ใกล้ที่สุดจากตำแหน่งปัจจุบัน
            min_dist = float('inf')
            nearest_code = None
            
            for code in remaining:
                lat, lon = get_lat_lon_from_master(code)
                if lat and lon:
                    dist = haversine_distance(current_lat, current_lon, lat, lon)
                    if dist < min_dist:
                        min_dist = dist
                        nearest_code = code
            
            if nearest_code:
                route.append(nearest_code)
                remaining.remove(nearest_code)
                current_lat, current_lon = get_lat_lon_from_master(nearest_code)
                if current_lat is None:
                    current_lat, current_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
            else:
                # ไม่มีพิกัด → ใส่ที่เหลือตามลำดับ
                route.extend(remaining)
                break
        
        return route
    
    all_codes = test_df['Code'].unique().tolist()
    assigned_trips = {}
    trip_counter = 1
    trip_recommended_vehicles = {}  # เก็บรถที่แนะนำสำหรับแต่ละทริป
    
    total_codes = len(all_codes)
    processed = 0
    
    # ⏱️ Timer สำหรับ early stopping
    import time
    start_time = time.time()
    MAX_PROCESSING_TIME = 20  # วินาที - ลดเพื่อให้เร็วขึ้น (target: 30 วินาที รวม)
    
    # 🚀 Cache พิกัดและจังหวัดล่วงหน้า (ประหยัดเวลา 70%)
    coord_cache = {}
    province_cache = {}
    
    # 🆕 รายชื่อ 77 จังหวัดของไทยเพื่อดึงจากชื่อสาขา
    THAI_PROVINCES = [
        'กรุงเทพ', 'กรุงเทพมหานคร', 'สมุทรปราการ', 'นนทบุรี', 'ปทุมธานี', 'นครปฐม', 'สมุทรสาคร',
        'พระนครศรีอยุธยา', 'อยุธยา', 'อ่างทอง', 'ลพบุรี', 'สิงห์บุรี', 'ชัยนาท', 'สระบุรี',
        'ชลบุรี', 'ระยอง', 'จันทบุรี', 'ตราด', 'ฉะเชิงเทรา', 'ปราจีนบุรี', 'นครนายก', 'สระแก้ว',
        'นครราชสีมา', 'โคราช', 'บุรีรัมย์', 'สุรินทร์', 'ศรีสะเกษ', 'อุบลราชธานี', 'ยโสธร', 'ชัยภูมิ',
        'อำนาจเจริญ', 'หนองบัวลำภู', 'ขอนแก่น', 'อุดรธานี', 'เลย', 'หนองคาย', 'มหาสารคาม',
        'ร้อยเอ็ด', 'กาฬสินธุ์', 'สกลนคร', 'นครพนม', 'มุกดาหาร', 'บึงกาฬ',
        'เชียงใหม่', 'ลำพูน', 'ลำปาง', 'อุตรดิตถ์', 'แพร่', 'น่าน', 'พะเยา', 'เชียงราย', 'แม่ฮ่องสอน',
        'นครสวรรค์', 'อุทัยธานี', 'กำแพงเพชร', 'ตาก', 'สุโขทัย', 'พิษณุโลก', 'พิจิตร', 'เพชรบูรณ์',
        'ราชบุรี', 'กาญจนบุรี', 'สุพรรณบุรี', 'นครปฐม', 'สมุทรสงคราม', 'เพชรบุรี', 'ประจวบคีรีขันธ์',
        'ชุมพร', 'ระนอง', 'สุราษฎร์ธานี', 'พังงา', 'กระบี่', 'ภูเก็ต', 'นครศรีธรรมราช', 'ตรัง',
        'พัทลุง', 'สงขลา', 'สตูล', 'ปัตตานี', 'ยะลา', 'นราธิวาส'
    ]
    
    def extract_province_from_name(branch_name):
        """ดึงจังหวัดจากชื่อสาขา เช่น 'พิษณุโลก5' -> 'พิษณุโลก'"""
        if not branch_name:
            return None
        name = str(branch_name).strip()
        for province in THAI_PROVINCES:
            if province in name:
                return province
        return None
    
    for code in all_codes:
        lat, lon = get_lat_lon_from_master(code)
        coord_cache[code] = (lat, lon)
        
        # Cache จังหวัด - ลำดับความสำคัญ:
        # 1. Master Data
        # 2. ชื่อสาขา (ดึงจังหวัดจากชื่อ) 🆕
        # 3. Province column
        # 4. branch_info
        
        found_province = False
        
        # 1. ลอง Master Data ก่อน
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
            if len(master_row) > 0:
                prov = master_row.iloc[0].get('จังหวัด', '')
                if prov and str(prov).strip() and prov != 'UNKNOWN':
                    province_cache[code] = str(prov).strip()
                    found_province = True
        
        # 2. 🆕 ดึงจังหวัดจากชื่อสาขา (สำคัญสำหรับสาขาที่ไม่มีใน Master)
        if not found_province and 'Name' in test_df.columns:
            code_data = test_df[test_df['Code'] == code]
            if len(code_data) > 0:
                branch_name = code_data['Name'].iloc[0]
                prov_from_name = extract_province_from_name(branch_name)
                if prov_from_name:
                    province_cache[code] = prov_from_name
                    found_province = True
        
        # 3. Province column
        if not found_province and 'Province' in test_df.columns:
            prov = test_df[test_df['Code'] == code]['Province'].iloc[0] if len(test_df[test_df['Code'] == code]) > 0 else None
            if prov and prov != 'UNKNOWN' and str(prov).strip():
                province_cache[code] = prov
                found_province = True
        
        # 4. branch_info
        if not found_province and code in branch_info:
            prov = branch_info[code].get('province', 'UNKNOWN')
            if prov and prov != 'UNKNOWN' and str(prov).strip():
                province_cache[code] = prov
                found_province = True
        
        # ถ้าไม่เจอเลย
        if not found_province:
            province_cache[code] = 'UNKNOWN'
    
    # 🎯 จัดกลุ่มตามพิกัดก่อน (เพิ่มรัศมีสูง - ขอบเขตใหญ่ขึ้น)
    spatial_clusters = create_distance_based_clusters(all_codes, max_distance_km=25)
    
    # 🔒 ฟังก์ชันเช็คระยะทางจากสาขาใหม่ไปยังทุกสาขาในทริป (FAST VERSION)
    def check_distance_to_all_trip_branches(new_code, trip_codes, max_dist=40):
        """
        เช็คว่าสาขาใหม่ใกล้กับทุกสาขาในทริปหรือไม่ (ใช้ sampling ถ้าทริปใหญ่)
        คืนค่า: (avg_distance, max_distance, all_within_limit)
        """
        if not trip_codes:
            return 0, 0, True
        
        new_lat, new_lon = coord_cache.get(new_code, (None, None))
        if not new_lat:
            return 9999, 9999, False
        
        # ⚡ Speed: ถ้าทริปมีหลายสาขา ให้ sample แค่ 3 สาขา (ลดจาก 5)
        sample_codes = trip_codes if len(trip_codes) <= 3 else trip_codes[:2] + trip_codes[-1:]
        distances = []
        for code in sample_codes:
            code_lat, code_lon = coord_cache.get(code, (None, None))
            if code_lat:
                dist = haversine_distance(new_lat, new_lon, code_lat, code_lon)
                distances.append(dist)
        
        if not distances:
            return 9999, 9999, False
        
        avg_dist = sum(distances) / len(distances)
        max_dist_found = max(distances)
        all_within = max_dist_found <= max_dist
        
        return avg_dist, max_dist_found, all_within
    
    # ⚡ Speed: Pre-compute trip centroids for fast lookup
    trip_centroids = {}  # {trip_num: (lat, lon)}
    
    def update_trip_centroid(trip_num, codes):
        """อัพเดต centroid ของทริป"""
        if not codes:
            trip_centroids[trip_num] = (None, None)
            return
        lats, lons = [], []
        for code in codes:
            lat, lon = coord_cache.get(code, (None, None))
            if lat:
                lats.append(lat)
                lons.append(lon)
        if lats:
            trip_centroids[trip_num] = (sum(lats)/len(lats), sum(lons)/len(lons))
        else:
            trip_centroids[trip_num] = (None, None)
    
    def find_closest_trip_for_branch(branch_code, all_trip_codes_dict, exclude_trip=None):
        """
        หาทริปที่เหมาะสมที่สุดสำหรับสาขา - เช็คระยะห่างจากทุกสาขาในทริป (ไม่ใช้ centroid)
        เพื่อป้องกันการกระโดดข้ามสาขาที่ใกล้กว่า
        """
        branch_lat, branch_lon = coord_cache.get(branch_code, (None, None))
        if not branch_lat:
            return None, 9999
        
        best_trip = None
        best_avg_dist = 9999
        best_max_dist = 9999
        
        for trip_num, codes in all_trip_codes_dict.items():
            if exclude_trip and trip_num == exclude_trip:
                continue
            if not codes:
                continue
            
            # 🔒 เช็คระยะห่างจากทุกสาขาในทริป (ไม่ใช้ centroid)
            distances = []
            for code in codes:
                code_lat, code_lon = coord_cache.get(code, (None, None))
                if code_lat:
                    dist = haversine_distance(branch_lat, branch_lon, code_lat, code_lon)
                    distances.append(dist)
            
            if not distances:
                continue
            
            avg_dist = sum(distances) / len(distances)
            max_dist = max(distances)
            
            # ⚡ เลือกทริปที่มีระยะเฉลี่ยต่ำสุด และระยะไกลสุดไม่เกิน 40km
            if avg_dist < best_avg_dist and max_dist <= 40:
                best_avg_dist = avg_dist
                best_max_dist = max_dist
                best_trip = trip_num
        
        return best_trip, best_avg_dist
    
    # 🔄 เรียงสาขาจากใกล้ → ไกล จาก DC
    def sort_by_distance_from_dc(codes):
        """เรียงสาขาจากใกล้ DC ไปไกล DC"""
        def get_distance_from_dc(code):
            lat, lon = coord_cache.get(code, (None, None))
            if lat and lon:
                return calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
            return 9999  # ไม่มีพิกัด ให้ไว้ท้าย
        return sorted(codes, key=get_distance_from_dc)
    
    # 🆕 Cache ชื่อและตำบล/อำเภอ เพื่อจัดกลุ่ม
    name_cache = {}
    subdistrict_cache = {}
    district_cache = {}
    
    for code in test_df['Code'].unique():
        # Cache ชื่อสาขา
        if 'Name' in test_df.columns:
            code_data = test_df[test_df['Code'] == code]
            if len(code_data) > 0:
                name_cache[code] = str(code_data['Name'].iloc[0]).strip()
        
        # Cache ตำบล/อำเภอ จาก Master
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
            if len(master_row) > 0:
                m = master_row.iloc[0]
                if 'ตำบล' in m.index and pd.notna(m['ตำบล']):
                    subdistrict_cache[code] = str(m['ตำบล']).strip()
                if 'อำเภอ' in m.index and pd.notna(m['อำเภอ']):
                    district_cache[code] = str(m['อำเภอ']).strip()
    
    # 🆕 ฟังก์ชันหา base name (เช่น "โลตัส พระราม 2" -> "โลตัส พระราม")
    def get_base_name(name):
        import re
        if not name:
            return ""
        
        # แปลงเป็น lowercase และตัด whitespace
        name_lower = str(name).strip().lower()
        
        # 🆕 Normalize ชื่อสาขาที่คล้ายกัน
        # Future/ฟิวเจอร์
        if 'future' in name_lower or 'ฟิวเจอร์' in name_lower or 'ฟิวเจอ' in name_lower:
            if 'rangsit' in name_lower or 'รังสิต' in name_lower:
                return "ฟิวเจอร์รังสิต"  # รวมเป็นชื่อเดียว
        
        # Lotus/โลตัส
        if 'lotus' in name_lower or 'โลตัส' in name_lower:
            # ตัดเลขท้ายออก เช่น โลตัส 1 -> โลตัส
            base = re.sub(r'\s*\d+\s*$', '', name_lower)
            return base.strip()
        
        # Big C/บิ๊กซี
        if 'big c' in name_lower or 'bigc' in name_lower or 'บิ๊กซี' in name_lower or 'บิ๊กซ' in name_lower:
            base = re.sub(r'\s*\d+\s*$', '', name_lower)
            return base.strip()
        
        # Makro/แม็คโคร
        if 'makro' in name_lower or 'แม็คโคร' in name_lower or 'แมคโคร' in name_lower:
            base = re.sub(r'\s*\d+\s*$', '', name_lower)
            return base.strip()
        
        # คลอง (คลอง 1, คลอง 2, คลอง 3, ...)
        if 'คลอง' in name_lower:
            # ถ้ามีตัวเลข เช่น "คลอง 3" -> "คลอง"
            if re.search(r'คลอง\s*\d+', name_lower):
                return "คลอง"
        
        # ตัดตัวเลขท้ายชื่อและ whitespace
        base = re.sub(r'\s*\d+\s*$', '', str(name).strip())
        # ตัด "สาขา" ออก
        base = re.sub(r'^สาขา\s*', '', base)
        # ตัด FC_, _FC ออก
        base = re.sub(r'_FC\d+$', '', base)
        base = re.sub(r'^FC\s*', '', base)
        return base.strip().lower()
    
    # 🆕 จัดกลุ่มสาขาตามชื่อเดียวกัน + ตำบลเดียวกัน + เรียงตามระยะทาง nearest neighbor
    def group_by_name_and_subdistrict(codes):
        """
        จัดกลุ่มสาขาตามลำดับ แล้วเรียงตามระยะทาง nearest neighbor:
        🔥 ลำดับใหม่: ตำบล > ชื่อ > อำเภอ > จังหวัด
        0. 🆕 ตำบลเดียวกัน + จังหวัดเดียวกัน (สำคัญที่สุด - บังคับรวม)
        1. ชื่อเหมือนกัน + ตำบลเดียวกัน + จังหวัดเดียวกัน
        2. ชื่อเหมือนกัน + อำเภอเดียวกัน + จังหวัดเดียวกัน
        3. ตำบลเดียวกัน (แม้ชื่อต่าง) → รวมก่อน
        4. อำเภอเดียวกัน + จังหวัดเดียวกัน
        5. จังหวัดเดียวกัน
        6. ที่เหลือ
        
        🆕 เรียงกลุ่มตามระยะทาง: เริ่มจากใกล้ DC → nearest neighbor ไปเรื่อยๆ
        """
        # สร้าง key สำหรับจัดกลุ่ม
        groups = {}  # key: (priority, province, district, base_name, subdistrict) -> [codes]
        
        for code in codes:
            name = name_cache.get(code, '')
            base_name = get_base_name(name)
            subdistrict = subdistrict_cache.get(code, '')
            district = district_cache.get(code, '')
            province = province_cache.get(code, '')
            
            # 🔥 สร้าง group key - ใช้ลำดับความสำคัญ (เลขน้อย = สำคัญกว่า)
            # ให้ความสำคัญกับตำบลเดียวกันก่อน!
            if subdistrict and province:
                # ลำดับ 0: ตำบลเดียวกัน + จังหวัดเดียวกัน (🔥 สำคัญที่สุด - บังคับรวม)
                key = (0, province, district, '', subdistrict)
            elif base_name and subdistrict and province:
                # ลำดับ 1: ชื่อ + ตำบล + จังหวัด
                key = (1, province, district, base_name, subdistrict)
            elif base_name and district and province:
                # ลำดับ 2: ชื่อ + อำเภอ + จังหวัด
                key = (2, province, district, base_name, '')
            elif base_name and province:
                # ลำดับ 3: ชื่อ + จังหวัด
                key = (3, province, district, base_name, '')
            elif base_name:
                # ลำดับ 4: ชื่อเดียวกัน (แม้ต่างจังหวัด - เช่น โลตัส กทม กับ โลตัส ชลบุรี)
                key = (4, province, '', base_name, '')
            elif province and district:
                # ลำดับ 5: จังหวัด + อำเภอ (รวมสาขาในอำเภอเดียวกัน)
                key = (5, province, district, '', '')
            elif province:
                # ลำดับ 6: จังหวัดเดียวกัน
                key = (6, province, '', '', '')
            else:
                # ลำดับ 7: ที่เหลือ
                key = (7, province if province else code, '', '', '', '')
            
            if key not in groups:
                groups[key] = []
            groups[key].append(code)
        
        # 🆕 เรียงกลุ่มด้วย nearest neighbor approach
        # เริ่มจากกลุ่มที่ใกล้ DC ที่สุด แล้วหากลุ่มถัดไปที่ใกล้ที่สุด
        result = []
        remaining_groups = list(groups.items())  # [(key, [codes]), ...]
        
        # หาตำแหน่งเฉลี่ยของแต่ละกลุ่ม
        def get_group_center(group_codes):
            lats, lons = [], []
            for code in group_codes:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    lats.append(lat)
                    lons.append(lon)
            if lats and lons:
                return (sum(lats) / len(lats), sum(lons) / len(lons))
            return (None, None)
        
        # เริ่มจากกลุ่มที่ใกล้ DC ที่สุด
        if remaining_groups:
            # หากลุ่มที่ใกล้ DC ที่สุด
            def dist_from_dc(item):
                key, group_codes = item
                priority = key[0]
                center_lat, center_lon = get_group_center(group_codes)
                if center_lat and center_lon:
                    dist = calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, center_lat, center_lon)
                    return (priority, dist)  # เรียงตาม priority ก่อน แล้วระยะทาง
                return (priority, 9999)
            
            remaining_groups.sort(key=dist_from_dc)
            current_key, current_group = remaining_groups.pop(0)
            
            # เรียงสมาชิกในกลุ่มแรกตาม nearest neighbor
            group_sorted = sort_by_distance_from_dc(current_group)
            ordered = build_route_nearest_neighbor(group_sorted)
            result.extend(ordered)
            
            # หากลุ่มถัดไปที่ใกล้กับกลุ่มปัจจุบันที่สุด
            while remaining_groups:
                # ใช้ตำแหน่งสุดท้ายของ result เป็นจุดอ้างอิง
                last_code = result[-1] if result else None
                last_lat, last_lon = coord_cache.get(last_code, (None, None)) if last_code else (DC_WANG_NOI_LAT, DC_WANG_NOI_LON)
                
                if not last_lat:
                    last_lat, last_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
                
                # หากลุ่มที่ใกล้ที่สุด
                def dist_from_last(item):
                    key, group_codes = item
                    priority = key[0]
                    center_lat, center_lon = get_group_center(group_codes)
                    if center_lat and center_lon:
                        dist = calculate_distance(last_lat, last_lon, center_lat, center_lon)
                        return (priority, dist)  # priority ก่อน แล้วระยะทาง
                    return (priority, 9999)
                
                remaining_groups.sort(key=dist_from_last)
                next_key, next_group = remaining_groups.pop(0)
                
                # เรียงสมาชิกในกลุ่มตาม nearest neighbor จากจุดสุดท้าย
                ordered_group = build_route_nearest_neighbor_from_point(next_group, last_lat, last_lon)
                result.extend(ordered_group)
        
        return result
    
    # 🆕 ฟังก์ชัน nearest neighbor เริ่มจากจุดที่กำหนด
    def build_route_nearest_neighbor_from_point(codes, start_lat, start_lon):
        if not codes:
            return []
        
        result = []
        remaining = codes[:]
        current_lat, current_lon = start_lat, start_lon
        
        while remaining:
            # หาสาขาที่ใกล้ที่สุดจากจุดปัจจุบัน
            nearest = None
            min_dist = float('inf')
            for code in remaining:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    dist = calculate_distance(current_lat, current_lon, lat, lon)
                    if dist < min_dist:
                        min_dist = dist
                        nearest = code
            
            if nearest:
                result.append(nearest)
                remaining.remove(nearest)
                current_lat, current_lon = coord_cache.get(nearest, (current_lat, current_lon))
            else:
                # ไม่มีพิกัด ให้เพิ่มไปเลย
                result.extend(remaining)
                break
        
        return result
    
    # 🔒 Define helper functions ก่อน loop
    def get_province(branch_code):
        return province_cache.get(branch_code, 'UNKNOWN')
    
    def get_distance_from_dc(code):
        """คำนวณระยะทางจาก DC"""
        lat, lon = coord_cache.get(code, (None, None))
        if lat and lon:
            return haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
        return 0  # ไม่มีพิกัด ถือว่าใกล้ DC
    
    # 🚀 **NEW ALGORITHM: ใช้ลำดับไฟล์ต้นฉบับ + เช็คข้อห้ามรถ**
    # ไฟล์ต้นฉบับเรียงสาขาที่ใกล้กันไว้แล้ว → ใช้ลำดับนั้นเลย
    # 1. วนทีละสาขาตามลำดับไฟล์
    # 2. เช็คว่าเพิ่มเข้าทริปปัจจุบันได้ไหม (capacity + ระยะทาง + ข้อห้ามรถ)
    # 3. ถ้าได้ → เพิ่มเข้าทริป
    # 4. ถ้าไม่ได้ → ตัดทริปใหม่
    
    # ใช้ลำดับดั้งเดิมจากไฟล์ต้นฉบับ (ไม่ sort)
    all_codes_ordered = test_df['Code'].unique().tolist()
    all_codes = all_codes_ordered.copy()
    
    def get_lat_lon(branch_code):
        return coord_cache.get(branch_code, (None, None))
    
    # **Main Loop: Farthest First + Nearest Neighbor**
    while all_codes:
        # ⏱️ Early stopping - ถ้าใช้เวลามากกว่า MAX_PROCESSING_TIME
        if time.time() - start_time > MAX_PROCESSING_TIME:
            # จัดสาขาที่เหลือเข้าทริปใหม่คนละคัน
            for remaining_code in all_codes:
                assigned_trips[remaining_code] = trip_counter
                trip_counter += 1
            break
        
        # 🎯 Pop สาขาแรก (ไกลสุดจาก DC)
        seed_code = all_codes.pop(0)
        current_trip = [seed_code]
        assigned_trips[seed_code] = trip_counter
        
        seed_province = get_province(seed_code)
        seed_lat, seed_lon = coord_cache.get(seed_code, (None, None))
        
        # คำนวณ Weight/Cube ปัจจุบัน
        current_weight = test_df[test_df['Code'] == seed_code]['Weight'].sum()
        current_cube = test_df[test_df['Code'] == seed_code]['Cube'].sum()
        
        # กำหนดขีดจำกัด (เริ่มจาก 6W)
        max_cube = LIMITS['6W']['max_c'] * BUFFER  # 20 cube
        max_weight = LIMITS['6W']['max_w'] * BUFFER  # 6000 kg
        
        seed_subdistrict = subdistrict_cache.get(seed_code, '')
        seed_district = district_cache.get(seed_code, '')
        
        # 🔥🔥🔥 ขั้นตอนที่ 0: หาสาขาที่ต้องไปด้วยกัน (Reference เดียวกัน / เคยไปด้วยกัน / ตำบลเดียวกัน)
        # 🔒 เช็คข้อห้ามรถของ seed ก่อน
        seed_max_vehicle = get_max_vehicle_for_branch(seed_code)
        if seed_max_vehicle == '4W':
            max_cube = LIMITS['4W']['max_c'] * BUFFER
            max_weight = LIMITS['4W']['max_w'] * BUFFER
        elif seed_max_vehicle == 'JB':
            max_cube = LIMITS['JB']['max_c'] * BUFFER
            max_weight = LIMITS['JB']['max_w'] * BUFFER
        
        # 🆕 ลำดับที่ 0.1: หาสาขาที่มี Reference เดียวกัน (อยู่ที่เดียวกัน)
        seed_ref = LOCATION_CODE_TO_REF.get(seed_code, '')
        if seed_ref:
            same_ref_codes = [c for c in all_codes if LOCATION_CODE_TO_REF.get(c, '') == seed_ref]
            for same_code in same_ref_codes:
                next_weight = test_df[test_df['Code'] == same_code]['Weight'].sum()
                next_cube = test_df[test_df['Code'] == same_code]['Cube'].sum()
                
                # เช็คข้อห้ามรถ
                branch_max = get_max_vehicle_for_branch(same_code)
                temp_max_cube = max_cube
                temp_max_weight = max_weight
                if branch_max == '4W':
                    temp_max_cube = min(max_cube, LIMITS['4W']['max_c'] * BUFFER)
                    temp_max_weight = min(max_weight, LIMITS['4W']['max_w'] * BUFFER)
                elif branch_max == 'JB' and seed_max_vehicle == '6W':
                    temp_max_cube = min(max_cube, LIMITS['JB']['max_c'] * BUFFER)
                    temp_max_weight = min(max_weight, LIMITS['JB']['max_w'] * BUFFER)
                
                if current_cube + next_cube <= temp_max_cube and current_weight + next_weight <= temp_max_weight:
                    all_codes.remove(same_code)
                    current_trip.append(same_code)
                    assigned_trips[same_code] = trip_counter
                    current_weight += next_weight
                    current_cube += next_cube
                    max_cube = temp_max_cube
                    max_weight = temp_max_weight
        
        # 🆕 ลำดับที่ 0.2: หาสาขาที่เคยไปด้วยกันในประวัติ (trip_pairs)
        for pair_code in list(all_codes):
            pair_key = tuple(sorted([seed_code, pair_code]))
            if pair_key in trip_pairs:
                next_weight = test_df[test_df['Code'] == pair_code]['Weight'].sum()
                next_cube = test_df[test_df['Code'] == pair_code]['Cube'].sum()
                
                # เช็คข้อห้ามรถ
                branch_max = get_max_vehicle_for_branch(pair_code)
                temp_max_cube = max_cube
                temp_max_weight = max_weight
                if branch_max == '4W':
                    temp_max_cube = min(max_cube, LIMITS['4W']['max_c'] * BUFFER)
                    temp_max_weight = min(max_weight, LIMITS['4W']['max_w'] * BUFFER)
                elif branch_max == 'JB' and seed_max_vehicle == '6W':
                    temp_max_cube = min(max_cube, LIMITS['JB']['max_c'] * BUFFER)
                    temp_max_weight = min(max_weight, LIMITS['JB']['max_w'] * BUFFER)
                
                if current_cube + next_cube <= temp_max_cube and current_weight + next_weight <= temp_max_weight:
                    all_codes.remove(pair_code)
                    current_trip.append(pair_code)
                    assigned_trips[pair_code] = trip_counter
                    current_weight += next_weight
                    current_cube += next_cube
                    max_cube = temp_max_cube
                    max_weight = temp_max_weight
        
        # 🆕 ลำดับที่ 0.3: หาสาขาตำบลเดียวกัน
        if seed_subdistrict:
            # หาสาขาทั้งหมดที่อยู่ตำบลเดียวกัน
            same_sd_codes = [c for c in all_codes if subdistrict_cache.get(c, '') == seed_subdistrict]
            # เรียงตามระยะจาก seed
            if same_sd_codes and seed_lat and seed_lon:
                same_sd_codes.sort(key=lambda c: haversine_distance(
                    seed_lat, seed_lon, 
                    *coord_cache.get(c, (seed_lat, seed_lon))
                ))
            
            # เพิ่มสาขาตำบลเดียวกันเข้าทริป (ถ้า capacity พอ และ ข้อห้ามรถตรงกัน)
            for same_code in same_sd_codes:
                next_weight = test_df[test_df['Code'] == same_code]['Weight'].sum()
                next_cube = test_df[test_df['Code'] == same_code]['Cube'].sum()
                
                # 🔒 เช็คข้อห้ามรถของสาขาใหม่
                branch_max = get_max_vehicle_for_branch(same_code)
                if branch_max == '4W':
                    # ถ้าสาขาใหม่ใช้ได้แค่ 4W → ต้องปรับ limit ลง
                    if current_cube + next_cube > LIMITS['4W']['max_c'] * BUFFER:
                        continue
                    if current_weight + next_weight > LIMITS['4W']['max_w'] * BUFFER:
                        continue
                    max_cube = LIMITS['4W']['max_c'] * BUFFER
                    max_weight = LIMITS['4W']['max_w'] * BUFFER
                elif branch_max == 'JB' and seed_max_vehicle == '6W':
                    # ถ้าสาขาใหม่ใช้ได้ถึง JB แต่ seed ใช้ 6W ได้ → ปรับ limit เป็น JB
                    if current_cube + next_cube > LIMITS['JB']['max_c'] * BUFFER:
                        continue
                    max_cube = LIMITS['JB']['max_c'] * BUFFER
                    max_weight = LIMITS['JB']['max_w'] * BUFFER
                
                if current_cube + next_cube <= max_cube and current_weight + next_weight <= max_weight:
                    all_codes.remove(same_code)
                    current_trip.append(same_code)
                    assigned_trips[same_code] = trip_counter
                    current_weight += next_weight
                    current_cube += next_cube
        
        # 🔄 หาสาขาถัดไปที่ใกล้สุด (ระยะจากสาขาสุดท้าย ≤ MAX_DISTANCE_IN_TRIP)
        # 🆕 ให้ความสำคัญกับตำบลเดียวกันก่อน
        
        while all_codes:
            best_code = None
            best_dist = 9999
            best_same_subdistrict = False
            best_same_district = False
            
            # หาสาขาที่ใกล้ที่สุดจากสาขาสุดท้ายในทริป
            last_code = current_trip[-1]
            last_lat, last_lon = coord_cache.get(last_code, (None, None))
            last_subdistrict = subdistrict_cache.get(last_code, '')
            last_district = district_cache.get(last_code, '')
            
            for code in all_codes:
                code_province = get_province(code)
                code_lat, code_lon = coord_cache.get(code, (None, None))
                code_subdistrict = subdistrict_cache.get(code, '')
                code_district = district_cache.get(code, '')
                
                if not last_lat or not code_lat:
                    continue
                
                # ระยะจากสาขาสุดท้าย (สาขาติดกัน)
                dist_from_last = haversine_distance(last_lat, last_lon, code_lat, code_lon)
                
                # 🆕 เช็คว่าเป็นตำบล/อำเภอเดียวกันหรือไม่
                same_subdistrict = (code_subdistrict and code_subdistrict == last_subdistrict)
                same_district = (code_district and code_district == last_district)
                
                # 🔒 กฎใหม่: 
                # - ตำบลเดียวกัน → ไม่จำกัดระยะ (บังคับรวม)
                # - อำเภอเดียวกัน → ยืดหยุ่นระยะเป็น 80km
                # - อื่นๆ → ระยะจากสาขาก่อนหน้าต้องไม่เกิน MAX_DISTANCE_IN_TRIP (50km)
                if same_subdistrict:
                    # ตำบลเดียวกัน → ไม่จำกัดระยะ
                    pass
                elif same_district:
                    # อำเภอเดียวกัน → ยืดหยุ่นเป็น 80km
                    if dist_from_last > 80:
                        continue
                else:
                    # อื่นๆ → ไม่เกิน MAX_DISTANCE_IN_TRIP
                    if dist_from_last > MAX_DISTANCE_IN_TRIP:
                        continue
                
                # 🔥 เลือกสาขา: 1) ตำบลเดียวกัน > 2) อำเภอเดียวกัน > 3) ใกล้ที่สุด
                if best_code is None:
                    best_code = code
                    best_dist = dist_from_last
                    best_same_subdistrict = same_subdistrict
                    best_same_district = same_district
                elif same_subdistrict and not best_same_subdistrict:
                    # ตำบลเดียวกัน ดีกว่าที่เลือกไว้
                    best_code = code
                    best_dist = dist_from_last
                    best_same_subdistrict = same_subdistrict
                    best_same_district = same_district
                elif same_subdistrict and best_same_subdistrict and dist_from_last < best_dist:
                    # ทั้งคู่ตำบลเดียวกัน เลือกที่ใกล้กว่า
                    best_code = code
                    best_dist = dist_from_last
                    best_same_subdistrict = same_subdistrict
                    best_same_district = same_district
                elif not best_same_subdistrict and same_district and not best_same_district:
                    # อำเภอเดียวกัน ดีกว่าที่เลือกไว้ (ถ้าไม่มีตำบลเดียวกัน)
                    best_code = code
                    best_dist = dist_from_last
                    best_same_subdistrict = same_subdistrict
                    best_same_district = same_district
                elif not best_same_subdistrict and same_district and best_same_district and dist_from_last < best_dist:
                    # ทั้งคู่อำเภอเดียวกัน เลือกที่ใกล้กว่า
                    best_code = code
                    best_dist = dist_from_last
                    best_same_subdistrict = same_subdistrict
                    best_same_district = same_district
                elif not best_same_subdistrict and not best_same_district and dist_from_last < best_dist:
                    # ทั้งคู่ไม่ใช่ตำบล/อำเภอเดียวกัน เลือกที่ใกล้กว่า
                    best_code = code
                    best_dist = dist_from_last
                    best_same_subdistrict = same_subdistrict
                    best_same_district = same_district
            
            if not best_code:
                break  # ไม่มีสาขาที่เหมาะสม ตัดทริปใหม่
            
            # เช็ค capacity
            next_weight = test_df[test_df['Code'] == best_code]['Weight'].sum()
            next_cube = test_df[test_df['Code'] == best_code]['Cube'].sum()
            
            new_weight = current_weight + next_weight
            new_cube = current_cube + next_cube
            
            # ถ้าเกิน capacity → ตัดทริปใหม่
            if new_cube > max_cube or new_weight > max_weight:
                break
            
            # 🔒 เช็คข้อห้ามรถ: สาขาใหม่ต้องใช้รถร่วมกับสาขาเดิมได้
            # หารถที่ใหญ่ที่สุดที่ทุกสาขาในทริปใช้ได้
            trip_codes_with_new = current_trip + [best_code]
            max_vehicle_allowed = '6W'
            for trip_code in trip_codes_with_new:
                branch_max = get_max_vehicle_for_branch(trip_code)
                if branch_max == '4W':
                    max_vehicle_allowed = '4W'
                    break
                elif branch_max == 'JB' and max_vehicle_allowed == '6W':
                    max_vehicle_allowed = 'JB'
            
            # ถ้ารถที่อนุญาตไม่พอใส่ของ → ตัดทริปใหม่
            if max_vehicle_allowed == '4W' and (new_cube > LIMITS['4W']['max_c'] * BUFFER or new_weight > LIMITS['4W']['max_w'] * BUFFER):
                break
            if max_vehicle_allowed == 'JB' and (new_cube > LIMITS['JB']['max_c'] * BUFFER or new_weight > LIMITS['JB']['max_w'] * BUFFER):
                break
            
            # 🆕 เช็คระยะทางรวมทั้งทริป (DC → สาขาแรก → ... → สาขาสุดท้าย → DC)
            # คำนวณระยะทางแบบ consecutive (DC → สาขา1 → สาขา2 → ... → สาขาสุดท้าย → DC)
            trip_codes_for_dist = current_trip + [best_code]
            total_trip_distance = 0
            
            # DC → สาขาแรก
            first_code = trip_codes_for_dist[0]
            first_lat, first_lon = coord_cache.get(first_code, (None, None))
            if first_lat and first_lon:
                total_trip_distance += haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, first_lat, first_lon)
            
            # สาขา → สาขาถัดไป
            for i in range(len(trip_codes_for_dist) - 1):
                c1 = trip_codes_for_dist[i]
                c2 = trip_codes_for_dist[i + 1]
                lat1, lon1 = coord_cache.get(c1, (None, None))
                lat2, lon2 = coord_cache.get(c2, (None, None))
                if lat1 and lon1 and lat2 and lon2:
                    total_trip_distance += haversine_distance(lat1, lon1, lat2, lon2)
            
            # สาขาสุดท้าย → DC
            last_code_dist = trip_codes_for_dist[-1]
            last_lat_dist, last_lon_dist = coord_cache.get(last_code_dist, (None, None))
            if last_lat_dist and last_lon_dist:
                total_trip_distance += haversine_distance(last_lat_dist, last_lon_dist, DC_WANG_NOI_LAT, DC_WANG_NOI_LON)
            
            # 🔒 จำกัดระยะทางรวมทั้งทริป: 4W/JB = 400km, 6W = 1000km
            max_trip_distance = 1000 if max_vehicle_allowed == '6W' else 400
            if total_trip_distance > max_trip_distance:
                break
            
            # ✅ ผ่านทุกเงื่อนไข → เพิ่มสาขานี้
            all_codes.remove(best_code)
            current_trip.append(best_code)
            assigned_trips[best_code] = trip_counter
            current_weight = new_weight
            current_cube = new_cube
        
        trip_counter += 1
    
    test_df['Trip'] = test_df['Code'].map(assigned_trips)
    
    # ===============================================
    # 🆕 ตรวจสอบสาขาที่ยังไม่มีทริป และจัดให้ครบทุกสาขา
    # ===============================================
    unassigned_codes = test_df[test_df['Trip'].isna()]['Code'].tolist()
    
    if unassigned_codes:
        # จัดสาขาที่ยังไม่มีทริปเข้าทริปที่ใกล้ที่สุด หรือสร้างทริปใหม่
        for code in unassigned_codes:
            code_lat, code_lon = coord_cache.get(code, (None, None))
            code_province = province_cache.get(code, 'UNKNOWN')
            code_weight = test_df[test_df['Code'] == code]['Weight'].sum()
            code_cube = test_df[test_df['Code'] == code]['Cube'].sum()
            
            best_trip = None
            best_score = float('inf')
            
            # หาทริปที่ใกล้ที่สุดและรับสาขานี้ได้
            for trip_num in test_df['Trip'].dropna().unique():
                trip_data = test_df[test_df['Trip'] == trip_num]
                trip_codes = trip_data['Code'].tolist()
                trip_weight = trip_data['Weight'].sum()
                trip_cube = trip_data['Cube'].sum()
                
                # เช็คจังหวัดเดียวกัน
                trip_provinces = set()
                for tc in trip_codes:
                    tp = province_cache.get(tc, 'UNKNOWN')
                    if tp != 'UNKNOWN':
                        trip_provinces.add(tp)
                
                # ต้องจังหวัดเดียวกัน หรือไม่มีข้อมูลจังหวัด
                if code_province != 'UNKNOWN' and trip_provinces and code_province not in trip_provinces:
                    continue
                
                # เช็คว่าใส่รถได้หรือไม่ (ใช้ 6W เป็น limit)
                new_weight = trip_weight + code_weight
                new_cube = trip_cube + code_cube
                new_util = max((new_weight / LIMITS['6W']['max_w']) * 100,
                              (new_cube / LIMITS['6W']['max_c']) * 100)
                
                if new_util > 100:  # เกิน 100% ไม่รับ
                    continue
                
                # เช็คจำนวนสาขา
                if len(trip_codes) >= MAX_BRANCHES_PER_TRIP:
                    continue
                
                # 🚨 เช็ค consecutive distance หลังรวม
                combined_codes = trip_codes + [code]
                max_consec = calculate_max_consecutive_distance(combined_codes, coord_cache)
                if max_consec > MAX_DISTANCE_IN_TRIP:
                    continue
                
                # คำนวณระยะทางเฉลี่ยไปสาขาในทริป
                if code_lat:
                    distances = []
                    for tc in trip_codes:
                        tc_lat, tc_lon = coord_cache.get(tc, (None, None))
                        if tc_lat:
                            dist = haversine_distance(code_lat, code_lon, tc_lat, tc_lon)
                            distances.append(dist)
                    
                    if distances:
                        avg_dist = sum(distances) / len(distances)
                        # เลือกทริปที่ใกล้ที่สุด
                        if avg_dist < best_score:
                            best_score = avg_dist
                            best_trip = trip_num
            
            if best_trip is not None:
                # จัดเข้าทริปที่ใกล้ที่สุด
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_trip
                assigned_trips[code] = best_trip
            else:
                # ไม่มีทริปที่เหมาะสม → สร้างทริปใหม่
                test_df.loc[test_df['Code'] == code, 'Trip'] = trip_counter
                assigned_trips[code] = trip_counter
                trip_counter += 1
    
    # ===============================================
    # 🔒 Post-processing: สลับสาขาให้อยู่ทริปที่ใกล้กันที่สุด (FAST)
    # ===============================================
    def optimize_branch_placement():
        """สลับสาขาระหว่างทริปให้อยู่กับกลุ่มที่ใกล้กันที่สุด (เวอร์ชันเร็ว)"""
        # สร้าง dict ของ trip → codes
        trip_codes_dict = {}
        for trip_num in test_df['Trip'].unique():
            codes = test_df[test_df['Trip'] == trip_num]['Code'].tolist()
            trip_codes_dict[trip_num] = codes
            update_trip_centroid(trip_num, codes)
        
        # ⚡ Speed: เช็คเฉพาะสาขาที่อยู่ไกลจาก centroid ของทริปตัวเอง
        outliers = []  # (code, trip_num, dist_from_centroid)
        
        for trip_num, codes in trip_codes_dict.items():
            if len(codes) <= 2:
                continue
            
            centroid = trip_centroids.get(trip_num)
            if not centroid or not centroid[0]:
                continue
            
            for code in codes:
                code_lat, code_lon = coord_cache.get(code, (None, None))
                if code_lat:
                    dist = haversine_distance(code_lat, code_lon, centroid[0], centroid[1])
                    # ถ้าไกลจาก centroid มากกว่า 20km → อาจเป็น outlier
                    if dist > 20:
                        outliers.append((code, trip_num, dist))
        
        # เรียง outlier จากไกลสุดก่อน และจำกัดแค่ 50 ตัว
        outliers.sort(key=lambda x: -x[2])
        outliers = outliers[:50]
        
        # ลองย้าย outliers ไปทริปที่ใกล้กว่า
        for code, trip_num, dist_current in outliers:
            if code not in trip_codes_dict.get(trip_num, []):
                continue
            
            best_trip, best_dist = find_closest_trip_for_branch(code, trip_codes_dict, exclude_trip=trip_num)
            
            # ถ้ามีทริปอื่นที่ใกล้กว่าอย่างมีนัยสำคัญ (> 15km ดีกว่า)
            if best_trip and best_dist < dist_current - 15:
                # เช็คข้อจำกัดรถ
                code_max_vehicle = get_max_vehicle_for_branch(code)
                target_trip_codes = trip_codes_dict.get(best_trip, [])
                target_max_vehicle = get_max_vehicle_for_trip(set(target_trip_codes + [code]))
                
                # 🚨 เช็ค consecutive distance หลังรวม
                combined_codes = target_trip_codes + [code]
                max_consec = calculate_max_consecutive_distance(combined_codes, coord_cache)
                if max_consec > MAX_DISTANCE_IN_TRIP:
                    continue
                
                vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
                if vehicle_priority.get(code_max_vehicle, 3) >= vehicle_priority.get(target_max_vehicle, 3):
                    # ย้ายสาขา
                    trip_codes_dict[trip_num].remove(code)
                    trip_codes_dict[best_trip].append(code)
                    assigned_trips[code] = best_trip
                    # อัพเดต centroids
                    update_trip_centroid(trip_num, trip_codes_dict[trip_num])
                    update_trip_centroid(best_trip, trip_codes_dict[best_trip])
        
        # อัพเดต test_df
        test_df['Trip'] = test_df['Code'].map(assigned_trips)
    
    # เรียกใช้งาน optimization
    optimize_branch_placement()
    
    # ===============================================
    # Post-processing: รวมทริปเล็กและปรับขนาดรถ
    # ===============================================
    # กำลังปรับปรุงการจัดทริป
    
    # สร้างรายการทริปทั้งหมดพร้อมข้อมูล
    all_trips = []
    for trip_num in test_df['Trip'].unique():
        trip_data = test_df[test_df['Trip'] == trip_num]
        branch_count = len(trip_data)
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        trip_codes = set(trip_data['Code'].values)
        
        # หาจังหวัดของทริป
        provinces = set()
        for code in trip_codes:
            prov = get_province(code)
            if prov != 'UNKNOWN':
                provinces.add(prov)
        
        # คำนวณ % การใช้รถ 4W
        w_util = (total_w / LIMITS['4W']['max_w']) * 100
        c_util = (total_c / LIMITS['4W']['max_c']) * 100
        max_util = max(w_util, c_util)
        
        all_trips.append({
            'trip': trip_num,
            'count': branch_count,
            'util': max_util,
            'weight': total_w,
            'cube': total_c,
            'codes': trip_codes,
            'provinces': provinces
        })
    
    # 🎯 คำนวณ centroid (จุดกึ่งกลาง) ของแต่ละทริป
    for trip in all_trips:
        lats, lons = [], []
        for code in trip['codes']:
            lat, lon = get_lat_lon(code)
            if lat and lon:
                lats.append(lat)
                lons.append(lon)
        
        if lats and lons:
            trip['centroid_lat'] = sum(lats) / len(lats)
            trip['centroid_lon'] = sum(lons) / len(lons)
            # คำนวณระยะทางจาก DC
            trip['distance_from_dc'] = haversine_distance(
                DC_WANG_NOI_LAT, DC_WANG_NOI_LON,
                trip['centroid_lat'], trip['centroid_lon']
            )
        else:
            trip['centroid_lat'] = DC_WANG_NOI_LAT
            trip['centroid_lon'] = DC_WANG_NOI_LON
            trip['distance_from_dc'] = 0
    
    # เรียงทริปตาม: 1) จังหวัดหลัก 2) ระยะทางจาก DC (ใกล้ไปไกล) 3) จำนวนสาขา 4) utilization
    # เพื่อจัดกลุ่มทริปในจังหวัดเดียวกันก่อน ค่อยข้ามจังหวัด
    def get_primary_province(trip):
        """หาจังหวัดหลักของทริป (จังหวัดที่มีสาขามากที่สุด)"""
        if not trip['provinces']:
            return 'UNKNOWN'
        # นับจำนวนสาขาในแต่ละจังหวัด
        province_counts = {}
        for code in trip['codes']:
            prov = get_province(code)
            province_counts[prov] = province_counts.get(prov, 0) + 1
        # คืนค่าจังหวัดที่มีสาขามากที่สุด
        return max(province_counts.items(), key=lambda x: x[1])[0] if province_counts else 'UNKNOWN'
    
    # เพิ่มข้อมูลจังหวัดหลักให้แต่ละทริป
    for trip in all_trips:
        trip['primary_province'] = get_primary_province(trip)
    
    # 🔄 เรียงตามจังหวัดหลัก → ระยะทางจาก DC (ไกลไปใกล้ เพื่อให้รถไกลออกก่อน)
    all_trips.sort(key=lambda x: (x['primary_province'], -x['distance_from_dc'], x['count'], x['util']))
    
    # ===============================================
    # 🎯 Phase 0.3: บังคับรวมสาขาตำบลเดียวกันที่ถูกแยกทริป
    # ถ้ามีสาขาตำบลเดียวกันอยู่คนละทริป → ย้ายไปรวมกัน
    # ===============================================
    def get_subdistrict_for_trip(trip_num):
        """หาตำบลหลักของทริป"""
        trip_data = test_df[test_df['Trip'] == trip_num]
        subdistricts = {}
        for code in trip_data['Code'].values:
            sd = subdistrict_cache.get(code, '')
            if sd:
                subdistricts[sd] = subdistricts.get(sd, 0) + 1
        if subdistricts:
            return max(subdistricts.items(), key=lambda x: x[1])[0]
        return ''
    
    # สร้าง mapping: ตำบล → ทริปที่มีสาขาในตำบลนี้
    subdistrict_to_trips = {}
    for trip in all_trips:
        if trip is None:
            continue
        for code in trip['codes']:
            sd = subdistrict_cache.get(code, '')
            prov = province_cache.get(code, '')
            if sd and prov:
                key = (sd, prov)  # ตำบล + จังหวัด
                if key not in subdistrict_to_trips:
                    subdistrict_to_trips[key] = []
                subdistrict_to_trips[key].append(trip)
    
    # หาตำบลที่มีสาขาอยู่หลายทริป
    merge_same_subdistrict_count = 0
    for (sd, prov), trips in subdistrict_to_trips.items():
        if len(trips) <= 1:
            continue
        
        # หาทริปหลัก (มีสาขาในตำบลนี้มากที่สุด)
        main_trip = None
        max_count = 0
        for trip in trips:
            if trip is None:
                continue
            count = sum(1 for c in trip['codes'] if subdistrict_cache.get(c, '') == sd)
            if count > max_count:
                max_count = count
                main_trip = trip
        
        if main_trip is None:
            continue
        
        # ย้ายสาขาตำบลนี้จากทริปอื่นมาทริปหลัก
        for trip in trips:
            if trip is None or trip == main_trip:
                continue
            
            # หาสาขาในตำบลนี้
            codes_to_move = [c for c in trip['codes'] if subdistrict_cache.get(c, '') == sd]
            
            if not codes_to_move:
                continue
            
            # เช็คว่ารวมแล้วไม่เกิน capacity
            codes_weight = sum(test_df[test_df['Code'] == c]['Weight'].sum() for c in codes_to_move)
            codes_cube = sum(test_df[test_df['Code'] == c]['Cube'].sum() for c in codes_to_move)
            
            new_weight = main_trip['weight'] + codes_weight
            new_cube = main_trip['cube'] + codes_cube
            
            # ใช้ 6W capacity เป็น limit (เพราะจะเลือกรถทีหลัง)
            if new_cube > LIMITS['6W']['max_c'] * BUFFER:
                continue
            if new_weight > LIMITS['6W']['max_w'] * BUFFER:
                continue
            
            # ย้ายสาขา
            for code in codes_to_move:
                test_df.loc[test_df['Code'] == code, 'Trip'] = main_trip['trip']
                trip['codes'].discard(code)  # ลบจากทริปเดิม
                main_trip['codes'].add(code)  # เพิ่มเข้าทริปหลัก
            
            # อัปเดต weight/cube
            main_trip['weight'] = new_weight
            main_trip['cube'] = new_cube
            main_trip['count'] = len(main_trip['codes'])
            trip['weight'] -= codes_weight
            trip['cube'] -= codes_cube
            trip['count'] = len(trip['codes'])
            
            merge_same_subdistrict_count += 1
    
    # ลบทริปที่ไม่มีสาขาแล้ว
    all_trips = [t for t in all_trips if t is not None and len(t['codes']) > 0]
    
    # ===============================================
    # 🎯 Phase 0.5: รวมทริปภาคเหนือ/ใต้ ที่ยังไม่เต็ม 6W
    # จังหวัดในภาคเดียวกัน (เช่น น่าน + พะเยา) ควรรวมกันเป็น 6W
    # ===============================================
    def get_region_group(province):
        """คืนค่ากลุ่มภาค (เหนือตอนบน, เหนือตอนล่าง, ใต้ฝั่งอันดามัน, ใต้ฝั่งอ่าวไทย)"""
        region_groups = {
            'เหนือตอนบน': ['น่าน', 'พะเยา', 'ลำปาง', 'ลำพูน', 'เชียงราย', 'เชียงใหม่', 'แพร่', 'แม่ฮ่องสอน'],
            'เหนือตอนล่าง': ['กำแพงเพชร', 'ตาก', 'นครสวรรค์', 'พิจิตร', 'พิษณุโลก', 'สุโขทัย', 'อุตรดิตถ์', 'อุทัยธานี', 'เพชรบูรณ์'],
            'ใต้ฝั่งอันดามัน': ['กระบี่', 'ตรัง', 'พังงา', 'ภูเก็ต', 'ระนอง', 'สตูล'],
            'ใต้ฝั่งอ่าวไทย': ['ชุมพร', 'นครศรีธรรมราช', 'พัทลุง', 'ยะลา', 'สงขลา', 'สุราษฎร์ธานี', 'ปัตตานี', 'นราธิวาส']
        }
        prov_str = str(province).strip()
        for group, provinces in region_groups.items():
            for p in provinces:
                if p in prov_str:
                    return group
        return None
    
    # หาทริปภาคเหนือ/ใต้ที่ยังไม่เต็ม 6W
    north_south_trips = []
    for idx, trip in enumerate(all_trips):
        if trip is None:
            continue
        primary_prov = trip['primary_province']
        region = get_region_type(primary_prov)
        
        # เฉพาะภาคเหนือ/ใต้
        if region in ['north', 'south']:
            region_group = get_region_group(primary_prov)
            cube_6w = (trip['cube'] / LIMITS['6W']['max_c']) * 100
            
            # ยังไม่เต็ม 6W (< 95%)
            if cube_6w < 95:
                north_south_trips.append({
                    'idx': idx,
                    'trip': trip,
                    'region': region,
                    'region_group': region_group,
                    'cube': trip['cube'],
                    'weight': trip['weight'],
                    'count': trip['count'],
                    'cube_6w': cube_6w
                })
    
    # รวมทริปในภาคเดียวกัน
    merge_north_south_count = 0
    merged_indices = set()
    
    for i, t1 in enumerate(north_south_trips):
        if t1['idx'] in merged_indices:
            continue
        
        for j, t2 in enumerate(north_south_trips):
            if i >= j or t2['idx'] in merged_indices:
                continue
            
            # ต้องอยู่ในกลุ่มภาคเดียวกัน (เช่น เหนือตอนบน หรือ ใต้ฝั่งอ่าวไทย)
            if t1['region_group'] != t2['region_group'] or t1['region_group'] is None:
                continue
            
            # เช็ค capacity รวมกัน (ต้องไม่เกิน 6W)
            combined_cube = t1['cube'] + t2['cube']
            combined_weight = t1['weight'] + t2['weight']
            combined_count = t1['count'] + t2['count']
            
            if combined_cube > LIMITS['6W']['max_c'] * BUFFER:
                continue
            if combined_weight > LIMITS['6W']['max_w'] * BUFFER:
                continue
            
            # รวมได้! ย้ายสาขาจาก trip2 ไป trip1
            trip1 = all_trips[t1['idx']]
            trip2 = all_trips[t2['idx']]
            
            if trip1 is None or trip2 is None:
                continue
            
            # ย้ายสาขา
            for code in trip2['codes']:
                test_df.loc[test_df['Code'] == code, 'Trip'] = trip1['trip']
            
            # อัปเดต trip1
            trip1['codes'].update(trip2['codes'])  # codes เป็น set
            trip1['cube'] = combined_cube
            trip1['weight'] = combined_weight
            trip1['count'] = combined_count
            trip1['provinces'].update(trip2['provinces'])
            
            # ลบ trip2
            all_trips[t2['idx']] = None
            merged_indices.add(t2['idx'])
            merge_north_south_count += 1
            
            # อัปเดต t1 สำหรับ iteration ถัดไป
            t1['cube'] = combined_cube
            t1['weight'] = combined_weight
            t1['count'] = combined_count
    
    # ลบทริปที่ถูกรวมแล้ว
    all_trips = [t for t in all_trips if t is not None]
    
    # 🎯 Phase 1: รวมทริปเล็ก (≤3 สาขา) กับทริปใกล้เคียง (FAST VERSION)
    merged = True
    merge_count = 0
    iteration = 0
    max_iterations = 1  # ⚡ ลดเป็น 1 รอบเพื่อความเร็ว
    
    while merged and len(all_trips) > 1 and iteration < max_iterations:
        merged = False
        iteration += 1
        
        # ⚡ Speed: สร้าง index ตามจังหวัดเพื่อหาทริปใกล้เคียงได้เร็ว
        province_to_trips = {}
        for idx, trip in enumerate(all_trips):
            if trip is None:
                continue
            for prov in trip['provinces']:
                if prov not in province_to_trips:
                    province_to_trips[prov] = []
                province_to_trips[prov].append(idx)
        
        # เริ่มจากทริปเล็กก่อน (เร็วกว่า)
        small_trips = [(idx, t) for idx, t in enumerate(all_trips) if t and t['count'] <= 3]
        small_trips.sort(key=lambda x: x[1]['count'])
        
        for i, trip1 in small_trips:
            if all_trips[i] is None:
                continue
            
            # ⚡ หาทริปที่จะรวมได้จากจังหวัดเดียวกันเท่านั้น
            candidate_indices = set()
            for prov in trip1['provinces']:
                for idx in province_to_trips.get(prov, []):
                    if idx != i and all_trips[idx] is not None:
                        candidate_indices.add(idx)
            
            # ⚡ จำกัดแค่ 10 candidates ที่ใกล้ที่สุด (ตาม centroid)
            if len(candidate_indices) > 10 and 'centroid_lat' in trip1:
                candidates_with_dist = []
                for idx in candidate_indices:
                    trip2 = all_trips[idx]
                    if 'centroid_lat' in trip2:
                        dist = haversine_distance(
                            trip1['centroid_lat'], trip1['centroid_lon'],
                            trip2['centroid_lat'], trip2['centroid_lon']
                        )
                        candidates_with_dist.append((idx, dist))
                candidates_with_dist.sort(key=lambda x: x[1])
                candidate_indices = {x[0] for x in candidates_with_dist[:10]}
            
            for j in candidate_indices:
                if all_trips[j] is None:
                    continue
                
                trip2 = all_trips[j]
                
                # เช็คระยะทางระหว่าง centroid
                if 'centroid_lat' in trip1 and 'centroid_lat' in trip2:
                    centroid_distance = haversine_distance(
                        trip1['centroid_lat'], trip1['centroid_lon'],
                        trip2['centroid_lat'], trip2['centroid_lon']
                    )
                    if centroid_distance > 80:  # ไกลเกิน 80km ไม่รวม
                        continue
                
                # 🚨 เช็คข้อจำกัดรถก่อนรวม
                combined_codes = trip1['codes'] | trip2['codes']
                max_allowed_combined = get_max_vehicle_for_trip(combined_codes)
                
                # ลองรวมกัน
                combined_w = trip1['weight'] + trip2['weight']
                combined_c = trip1['cube'] + trip2['cube']
                combined_count = trip1['count'] + trip2['count']
                
                # เช็คว่ารวมแล้วใส่รถได้หรือไม่
                if combined_count > MAX_BRANCHES_PER_TRIP:
                    continue
                
                # คำนวณ % การใช้รถหลังรวม
                combined_6w_util = max(
                    (combined_w / LIMITS['6W']['max_w']) * 100,
                    (combined_c / LIMITS['6W']['max_c']) * 100
                )
                
                # เช็คว่าใส่รถที่อนุญาตได้หรือไม่
                vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
                allowed_priority = vehicle_priority.get(max_allowed_combined, 3)
                
                can_fit = False
                if allowed_priority >= 3 and combined_6w_util <= 100:  # 6W
                    can_fit = True
                elif allowed_priority >= 2 and combined_c <= LIMITS['JB']['max_c'] * BUFFER:  # JB
                    can_fit = True
                elif allowed_priority >= 1 and combined_c <= LIMITS['4W']['max_c'] * BUFFER:  # 4W
                    can_fit = True
                
                if can_fit:
                    # รวมทริป
                    for code in trip2['codes']:
                        test_df.loc[test_df['Code'] == code, 'Trip'] = trip1['trip']
                    
                    # อัปเดตข้อมูล trip1
                    trip1['weight'] = combined_w
                    trip1['cube'] = combined_c
                    trip1['count'] = combined_count
                    trip1['codes'] |= trip2['codes']
                    trip1['provinces'] |= trip2['provinces']
                    trip1['util'] = combined_6w_util
                    
                    # อัพเดต centroid
                    lats, lons = [], []
                    for code in trip1['codes']:
                        lat, lon = coord_cache.get(code, (None, None))
                        if lat:
                            lats.append(lat)
                            lons.append(lon)
                    if lats:
                        trip1['centroid_lat'] = sum(lats) / len(lats)
                        trip1['centroid_lon'] = sum(lons) / len(lons)
                    
                    # ลบ trip2 ออก
                    all_trips[j] = None
                    merged = True
                    merge_count += 1
                    break
            
            if merged:
                break
        
        # ลบ None ออก
        all_trips = [t for t in all_trips if t is not None]
    
    # 🎯 Phase 1.25: ย้ายสาขาเดียวที่ใช้รถเล็กไปยังทริปใกล้เคียงที่ใหญ่กว่า
    # 🆕 พิเศษ: เช็คจังหวัด+อำเภอเดียวกันด้วย
    reassign_count = 0
    
    # หาทริปที่มีเพียง 1 สาขา และ utilization ต่ำ (<40%)
    single_branch_trips = []
    for trip_num in test_df['Trip'].unique():
        if trip_num == 0:
            continue
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 1:
            branch_code = trip_data['Code'].values[0]
            branch_w = trip_data['Weight'].values[0]
            branch_c = trip_data['Cube'].values[0]
            
            # คำนวณ utilization (ใช้รถที่เล็กที่สุดที่พอดี)
            util_4w = max((branch_w / LIMITS['4W']['max_w']) * 100, 
                         (branch_c / LIMITS['4W']['max_c']) * 100)
            util_jb = max((branch_w / LIMITS['JB']['max_w']) * 100,
                         (branch_c / LIMITS['JB']['max_c']) * 100)
            
            # 🆕 ดึงจังหวัด+อำเภอ
            branch_province = get_province(branch_code)
            branch_district = ''
            if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == branch_code]
                if len(master_row) > 0:
                    branch_district = str(master_row.iloc[0].get('อำเภอ', '')).strip()
            
            # ถ้าใช้รถน้อยกว่า 40% → พิจารณาย้าย
            if util_4w < 40 or util_jb < 40:
                lat, lon = get_lat_lon(branch_code)
                if lat and lon:
                    single_branch_trips.append({
                        'trip': trip_num,
                        'code': branch_code,
                        'weight': branch_w,
                        'cube': branch_c,
                        'lat': lat,
                        'lon': lon,
                        'util': util_4w,
                        'province': branch_province,
                        'district': branch_district
                    })
    
    # พยายามย้ายสาขาเดียวไปรวมกับทริปอื่นที่ใกล้เคียง
    for single_trip in single_branch_trips:
        branch_code = single_trip['code']
        branch_w = single_trip['weight']
        branch_c = single_trip['cube']
        branch_lat = single_trip['lat']
        branch_lon = single_trip['lon']
        
        # หาทริปที่ใกล้ที่สุด (ไม่รวมทริปของตัวเอง)
        best_trip = None
        min_distance = float('inf')
        
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0 or trip_num == single_trip['trip']:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            
            # ถ้าทริปมีมากกว่า 2 สาขา → พิจารณา
            if len(trip_data) < 2:
                continue
            
            # คำนวณระยะทางจาก centroid ของทริป
            trip_lats = []
            trip_lons = []
            for code in trip_data['Code'].values:
                lat, lon = get_lat_lon(code)
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if not trip_lats:
                continue
            
            centroid_lat = sum(trip_lats) / len(trip_lats)
            centroid_lon = sum(trip_lons) / len(trip_lons)
            
            distance = haversine_distance(branch_lat, branch_lon, centroid_lat, centroid_lon)
            
            # ถ้าไกลเกิน 50km → ข้าม (เพิ่มจาก 30km)
            if distance > 50:
                continue
            
            # เช็คว่าเพิ่มแล้วเกินไหม
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            new_w = trip_w + branch_w
            new_c = trip_c + branch_c
            new_count = len(trip_data) + 1
            
            # เช็คว่าใส่ได้ไหม (ยอมให้เกิน 125% สำหรับสาขาเดียว)
            new_util = max(
                (new_w / LIMITS['6W']['max_w']) * 100,
                (new_c / LIMITS['6W']['max_c']) * 100
            )
            
            if new_util <= 100 and new_count <= MAX_BRANCHES_PER_TRIP:
                # เช็คข้อจำกัดสาขา
                trip_codes = set(trip_data['Code'].values) | {branch_code}
                max_allowed = get_max_vehicle_for_trip(trip_codes)
                
                # บันทึกทริปที่ใกล้ที่สุด
                if distance < min_distance:
                    min_distance = distance
                    best_trip = trip_num
        
        # ถ้าเจอทริปที่เหมาะสม → ย้าย
        if best_trip is not None:
            test_df.loc[test_df['Code'] == branch_code, 'Trip'] = best_trip
            reassign_count += 1
    
    # 🎯 Phase 1.75: รวมทริป utilization ต่ำ (<50%) ให้เต็มขึ้น
    rebalance_count = 0
    LOW_UTIL_THRESHOLD = 50  # ทริปที่ต่ำกว่า 50% ถือว่าไม่คุ้ม
    
    # หาทริปที่ utilization ต่ำ
    low_util_trips = []
    for trip_num in sorted(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        trip_w = trip_data['Weight'].sum()
        trip_c = trip_data['Cube'].sum()
        trip_count = len(trip_data)
        
        trip_util = max(
            (trip_w / LIMITS['6W']['max_w']) * 100,
            (trip_c / LIMITS['6W']['max_c']) * 100
        )
        
        # ถ้า util < 50% และมี ≤6 สาขา → พิจารณารวม
        if trip_util < LOW_UTIL_THRESHOLD and trip_count <= 6:
            # หา centroid ของทริป
            trip_lats, trip_lons = [], []
            for code in trip_data['Code'].values:
                lat, lon = get_lat_lon(code)
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if trip_lats:
                low_util_trips.append({
                    'trip': trip_num,
                    'util': trip_util,
                    'count': trip_count,
                    'weight': trip_w,
                    'cube': trip_c,
                    'codes': set(trip_data['Code'].values),
                    'lat': sum(trip_lats) / len(trip_lats),
                    'lon': sum(trip_lons) / len(trip_lons)
                })
    
    # พยายามรวมทริปต่ำกับทริปใกล้เคียง
    for low_trip in low_util_trips:
        best_merge = None
        min_distance = float('inf')
        
        # หาทริปที่ใกล้ที่สุด
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0 or trip_num == low_trip['trip']:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            trip_count = len(trip_data)
            
            # ข้ามทริปที่มีสาขาเยอะเกินไป
            if trip_count >= MAX_BRANCHES_PER_TRIP:
                continue
            
            # หา centroid ของทริปนี้
            trip_lats, trip_lons = [], []
            for code in trip_data['Code'].values:
                lat, lon = get_lat_lon(code)
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if not trip_lats:
                continue
            
            target_lat = sum(trip_lats) / len(trip_lats)
            target_lon = sum(trip_lons) / len(trip_lons)
            
            # คำนวณระยะทาง
            distance = haversine_distance(low_trip['lat'], low_trip['lon'], target_lat, target_lon)
            
            # ถ้าไกลเกิน 50km → ข้าม
            if distance > 50:
                continue
            
            # ตรวจสอบว่ารวมแล้วเกินไหม
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            combined_w = trip_w + low_trip['weight']
            combined_c = trip_c + low_trip['cube']
            combined_count = trip_count + low_trip['count']
            
            combined_util = max(
                (combined_w / LIMITS['6W']['max_w']) * 100,
                (combined_c / LIMITS['6W']['max_c']) * 100
            )
            
            # รวมได้ถ้า ≤120% และสาขา ≤MAX
            if combined_util <= 100 and combined_count <= MAX_BRANCHES_PER_TRIP:
                # เช็คข้อจำกัดสาขา
                combined_codes = low_trip['codes'] | set(trip_data['Code'].values)
                max_allowed = get_max_vehicle_for_trip(combined_codes)
                
                if distance < min_distance:
                    min_distance = distance
                    best_merge = trip_num
        
        # ถ้าเจอทริปที่เหมาะสม → รวม
        if best_merge is not None:
            for code in low_trip['codes']:
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_merge
            rebalance_count += 1
    
    # 🎯 Phase 1.5: เก็บสาขาที่อยู่ในเส้นทาง (Route Pickup Optimization) - จำกัดเวลา
    pickup_count = 0
    MAX_DETOUR_KM_LOCAL = MAX_DETOUR_KM  # ใช้ค่าจาก config (12 กม.)
    
    # ⚡ Skip ถ้ามีทริปมากเกิน 20 ทริป (ประหยัดเวลา)
    unique_trips = test_df['Trip'].unique()
    if len(unique_trips) > 20:
        pass  # Skip Phase 1.5 เพื่อความเร็ว
    else:
        # วนลูปทุกทริปที่ยังไม่เต็ม (เป้าหมาย 95%) - จำกัดแค่ 15 ทริปแรก
        for trip_num in sorted(unique_trips)[:15]:
            trip_data = test_df[test_df['Trip'] == trip_num]
            current_w = trip_data['Weight'].sum()
            current_c = trip_data['Cube'].sum()
            current_count = len(trip_data)
            
            # คำนวณ % การใช้รถปัจจุบัน (ใช้ 6W เป็นมาตรฐาน)
            current_util = max(
                (current_w / LIMITS['6W']['max_w']) * 100,
                (current_c / LIMITS['6W']['max_c']) * 100
            )
            
            # 🎯 เป้าหมาย: เก็บสาขาจนเต็มเกือบ 100% (คิวเต็ม)
            TARGET_UTIL = 100  # เป้าหมาย utilization
            MAX_PICKUP_UTIL = 100  # สูงสุดที่ยอมเก็บได้ (ห้ามเกิน 100%)
            
            # ถ้าเกิน 100% หรือมีสาขาเยอะแล้ว → ข้าม
            if current_util >= MAX_PICKUP_UTIL or current_count >= MAX_BRANCHES_PER_TRIP:
                continue
            
            # หาจังหวัดของทริปปัจจุบัน
            trip_provinces = set()
            trip_coords = []
            for code in trip_data['Code'].values:
                prov = get_province(code)
                if prov != 'UNKNOWN':
                    trip_provinces.add(prov)
                
                # เก็บพิกัด
                lat, lon = get_lat_lon(code)
                if lat and lon:
                    trip_coords.append((lat, lon))
            
            # ถ้าไม่มีพิกัด → ข้าม
            if not trip_coords:
                continue
            
            # หาสาขาที่ยังไม่ได้จัดทริป (Trip = 0 หรือ NaN)
            unassigned = test_df[(test_df['Trip'] == 0) | (test_df['Trip'].isna())]
            
            for idx, row in unassigned.iterrows():
                branch_code = row['Code']
                branch_w = row['Weight']
                branch_c = row['Cube']
                branch_prov = get_province(branch_code)
                branch_lat, branch_lon = get_lat_lon(branch_code)
                
                # ถ้าไม่มีพิกัด → ข้าม
                if not branch_lat or not branch_lon:
                    continue
                
                # เช็คว่าอยู่ในจังหวัดเดียวกันหรือใกล้เคียง
                if branch_prov not in trip_provinces:
                    # เช็คระยะทางจากทุกสาขาในทริป
                    min_distance = float('inf')
                    for trip_lat, trip_lon in trip_coords:
                        dist = haversine_distance(trip_lat, trip_lon, branch_lat, branch_lon)
                        if dist < min_distance:
                            min_distance = dist
                    
                    # ถ้าไม่ได้อยู่ในเส้นทาง (ไกลเกินจากทุกสาขา) → ข้าม
                    if min_distance > MAX_DETOUR_KM_LOCAL:
                        continue
                
                # คำนวณว่าเพิ่มสาขานี้แล้วเกินไหม
                new_w = current_w + branch_w
                new_c = current_c + branch_c
                new_count = current_count + 1
                
                # คำนวณ % ใหม่ (เน้น Cube)
                new_cube_util = (new_c / LIMITS['6W']['max_c']) * 100
                new_weight_util = (new_w / LIMITS['6W']['max_w']) * 100
                new_util = max(new_cube_util, new_weight_util)
                
                # 🎯 ถ้ารถไม่เต็ม (<95%) → ยอมให้เพิ่มแต่ไม่เกิน 100%
                # เป้าหมาย: Cube 95-100%, น้ำหนัก ≤100%
                if current_util < 95:
                    # รถยังไม่เต็ม → ยอมให้เพิ่มแต่ไม่เกิน 100%
                    can_add = new_cube_util <= 100 and new_weight_util <= 100 and new_count <= MAX_BRANCHES_PER_TRIP
                else:
                    # รถเต็มพอสมควรแล้ว → เข้มงวด (ไม่เกิน 100%)
                    can_add = new_cube_util <= 100 and new_weight_util <= 100 and new_count <= MAX_BRANCHES_PER_TRIP
                
                if can_add:
                    # เช็คข้อจำกัดสาขา
                    test_trip_codes = set(trip_data['Code'].values) | {branch_code}
                    max_allowed = get_max_vehicle_for_trip(test_trip_codes)
                    
                    # ถ้าสาขานี้จำกัดรถเล็กกว่ารถปัจจุบัน → ต้องเช็คว่าใส่ได้ไหม
                    # (ปล่อยให้ Phase 2 จัดการ)
                    
                    # เพิ่มสาขาเข้าทริป
                    test_df.loc[test_df['Code'] == branch_code, 'Trip'] = trip_num
                    
                    # อัปเดตข้อมูลปัจจุบัน
                    current_w = new_w
                    current_c = new_c
                    current_count = new_count
                    current_util = new_util
                    
                    # เพิ่มพิกัดใหม่
                    trip_coords.append((branch_lat, branch_lon))
                    if branch_prov != 'UNKNOWN':
                        trip_provinces.add(branch_prov)
                    
                    pickup_count += 1
                    
                    # ถ้าเต็มเกินไปแล้ว (Cube >100% หรือสาขาเกิน MAX) → หยุดเพิ่มสาขา
                    current_cube_util = (current_c / LIMITS['6W']['max_c']) * 100
                    if current_cube_util >= 100 or current_count >= MAX_BRANCHES_PER_TRIP:
                        break
    
    # 🚨 Phase 1.75: แยกสาขาที่มีข้อจำกัดรถ (4W/JB) ออกจากทริปที่ใช้รถใหญ่
    restriction_split_count = 0
    
    # หาทริปที่มีสาขาที่มีข้อจำกัดรถผสมกับสาขาไม่จำกัด
    for trip_num in sorted(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        trip_codes = set(trip_data['Code'].values)
        
        # แยกสาขาตามข้อจำกัดรถ
        codes_4w_only = set()  # สาขาที่ต้องใช้ 4W เท่านั้น
        codes_jb_or_less = set()  # สาขาที่ใช้ได้แค่ JB หรือเล็กกว่า
        codes_no_limit = set()  # สาขาที่ไม่มีข้อจำกัด (ใช้ 6W ได้)
        
        for code in trip_codes:
            max_vehicle = get_max_vehicle_for_trip({code})
            if max_vehicle == '4W':
                codes_4w_only.add(code)
            elif max_vehicle == 'JB':
                codes_jb_or_less.add(code)
            else:
                codes_no_limit.add(code)
        
        # 🚨 ถ้าทริปมีสาขาที่จำกัดรถผสมกับสาขาไม่จำกัด → ต้องแยก
        has_restrictions = len(codes_4w_only) > 0 or len(codes_jb_or_less) > 0
        has_no_limits = len(codes_no_limit) > 0
        
        if has_restrictions and has_no_limits:
            # แยกเป็น 2 กลุ่ม: 1) สาขาที่มีข้อจำกัด 2) สาขาที่ไม่มีข้อจำกัด
            restricted_codes = codes_4w_only | codes_jb_or_less
            unrestricted_codes = codes_no_limit
            
            # เก็บทริปเดิมให้กับกลุ่มที่มีสาขามากกว่า
            if len(restricted_codes) >= len(unrestricted_codes):
                # restricted ใช้ทริปเดิม
                keep_trip = trip_num
                new_trip = test_df['Trip'].max() + 1
                
                # ย้าย unrestricted ไปทริปใหม่
                for code in unrestricted_codes:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip
            else:
                # unrestricted ใช้ทริปเดิม
                keep_trip = trip_num
                new_trip = test_df['Trip'].max() + 1
                
                # ย้าย restricted ไปทริปใหม่
                for code in restricted_codes:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip
            
            restriction_split_count += 1
    
    # 🎯 Phase 2: เลือกรถที่เหมาะสม (เริ่มจาก 4W → JB → 6W หรือ 2 คัน) - Optimized
    vehicle_assignment_count = 0
    downsize_count = 0
    region_changes = {
        '4w': 0, 
        'jb': 0, 
        '6w': 0, 
        'split_2_vehicles': 0,
        'nearby_6w_to_jb': 0,
        'far_keep_6w': 0,
        'other': 0
    }
    
    # ⚡ Early stopping - ถ้าใช้เวลามากกว่า 55 วินาที
    if time.time() - start_time > 55:
        # Skip Phase 2 complex logic, ใช้ logic เร็ว
        for trip_num in test_df['Trip'].unique():
            trip_data = test_df[test_df['Trip'] == trip_num]
            total_c = trip_data['Cube'].sum()
            
            # เลือกรถแบบเร็ว (ไม่มี optimization)
            if total_c <= 5:
                trip_recommended_vehicles[trip_num] = '4W'
            elif total_c <= 7:
                trip_recommended_vehicles[trip_num] = 'JB'
            else:
                trip_recommended_vehicles[trip_num] = '6W'
    else:
        # เก็บข้อมูลทริปที่เหลือ
        for trip_num in test_df['Trip'].unique():
            trip_data = test_df[test_df['Trip'] == trip_num]
            branch_count = len(trip_data)
            total_w = trip_data['Weight'].sum()
            total_c = trip_data['Cube'].sum()
            trip_codes = set(trip_data['Code'].values)
            
            # 🔒 เช็คจังหวัด - สำคัญมาก! ห้าม 6W ในปริมณฑล
            provinces = set()
            for code in trip_codes:
                prov = get_province(code)
                if prov and prov != 'UNKNOWN':
                    provinces.add(prov)
            
            # เช็คว่าทุกจังหวัดเป็นพื้นที่ใกล้หรือไม่
            all_nearby = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
            has_north = any(get_region_type(p) == 'north' for p in provinces) if provinces else False
            has_south = any(get_region_type(p) == 'south' for p in provinces) if provinces else False
            
            # คำนวณระยะทาง max จาก DC
            max_distance_from_dc = 0
            for code in trip_codes:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    dist = haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
                    max_distance_from_dc = max(max_distance_from_dc, dist)
            
            # ตรวจสอบข้อจำกัดสาขา
            max_allowed = get_max_vehicle_for_trip(trip_codes)
            
            # 🔒 ตรวจสอบ Punthai - เงื่อนไขพิเศษ
            punthai_limits = get_punthai_vehicle_limits(trip_data, total_c, branch_count)
            punthai_type = is_punthai_only(trip_data)
            
            # ถ้า Punthai ล้วน → บังคับตาม Punthai limits
            if punthai_type == 'punthai_only':
                # Punthai ล้วน: JB ไม่เกิน 7 drop, 4W ถ้า Cube > 5 → ตัดเป็น JB
                punthai_max_vehicle = punthai_limits['max_vehicle']
                punthai_max_drops = punthai_limits['max_drops']
                
                # บังคับ max_allowed ตาม Punthai
                vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
                if vehicle_priority.get(punthai_max_vehicle, 3) < vehicle_priority.get(max_allowed, 3):
                    max_allowed = punthai_max_vehicle
                
            elif punthai_type == 'mixed':
                # ผสม Punthai + อื่น: ถ้า Cube 3-4 → 6W ได้, ถ้าเกิน → 4W เท่านั้น
                if total_c > 4.0:
                    max_allowed = '4W'  # บังคับ 4W เท่านั้น
                # ถ้า Cube 3-4 → ยังใช้ max_allowed ปกติ (6W ได้)
            
            # ⚠️ สำคัญ: ถ้าทุกจังหวัดเป็น nearby → ห้าม 6W เด็ดขาด!
            if all_nearby:
                very_far = False  # บังคับให้ไม่ใช้ 6W
                if max_allowed == '6W':
                    max_allowed = 'JB'  # บังคับขอบเขตเป็น JB
            # ⚠️ ภาคเหนือและภาคใต้ → บังคับใช้ 6W
            elif has_north or has_south:
                very_far = True  # บังคับให้ใช้ 6W
            else:
                # 🚛 เช็คระยะทาง - ไกลมากพิเศษ (>300km) ต้องใช้ 6W
                very_far_by_distance = max_distance_from_dc > 300
                very_far = very_far_by_distance
            
            # คำนวณ % การใช้รถแต่ละประเภท
            util_4w = max((total_w / LIMITS['4W']['max_w']) * 100, 
                          (total_c / LIMITS['4W']['max_c']) * 100)
            util_jb = max((total_w / LIMITS['JB']['max_w']) * 100,
                          (total_c / LIMITS['JB']['max_c']) * 100)
            util_6w = max((total_w / LIMITS['6W']['max_w']) * 100,
                          (total_c / LIMITS['6W']['max_c']) * 100)
            
            # 🔒 ไม่ต้องเรียก get_max_vehicle_for_trip อีก - ใช้ค่าที่บังคับ all_nearby แล้ว
            
            # 🎯 กลยุทธ์เลือกรถ (เริ่มจาก 4W → JB → แยก 2 คัน/6W)
            recommended = None
            cube_util_4w = (total_c / LIMITS['4W']['max_c']) * 100
            cube_util_jb = (total_c / LIMITS['JB']['max_c']) * 100
            cube_util_6w = (total_c / LIMITS['6W']['max_c']) * 100
            weight_util_4w = (total_w / LIMITS['4W']['max_w']) * 100
            weight_util_jb = (total_w / LIMITS['JB']['max_w']) * 100
            weight_util_6w = (total_w / LIMITS['6W']['max_w']) * 100
            
            # 🚨 ตรวจสอบข้อจำกัดสาขาก่อน
            if max_allowed == '4W':
                # ลำดับ 1: ลอง 4W ก่อน (95-100%)
                if 95 <= cube_util_4w <= 100 and weight_util_4w <= 100 and branch_count <= 12:
                    recommended = '4W'
                # ลำดับ 2: ถ้า 4W ไม่พอดี → แยกเป็น 4W + 4W (75-95% ต่อคัน)
                elif cube_util_4w > 100:
                    # จะแยกใน Phase 2.5
                    recommended = '4W+4W'
                else:
                    # ต่ำกว่า 95% → ใช้ 4W (แต่อาจรวมกับทริปอื่นภายหลัง)
                    recommended = '4W'
            elif max_allowed == 'JB':
                # 🔒 Punthai ล้วน: JB ไม่เกิน 7 drop
                max_jb_drops = 7 if punthai_type == 'punthai_only' else 12
                
                # ลำดับ 1: ลอง 4W ก่อน (95-100%)
                if 95 <= cube_util_4w <= 100 and weight_util_4w <= 100 and branch_count <= 12:
                    recommended = '4W'
                # ลำดับ 2: ลอง JB (95-100%) - เช็ค drop limit สำหรับ Punthai
                elif 95 <= cube_util_jb <= 100 and weight_util_jb <= 100 and branch_count <= max_jb_drops:
                    recommended = 'JB'
                # ลำดับ 3: แยกเป็น JB + 4W หรือ JB + JB (75-95% ต่อคัน)
                elif cube_util_jb > 100 or branch_count > max_jb_drops:
                    # ลองแยกเป็น JB + 4W (13 cube max)
                    if total_c <= 13:
                        recommended = 'JB+4W'
                    else:
                        recommended = 'JB+JB'  # 16 cube max
                else:
                    # ต่ำกว่า 95% → ใช้ JB หรือ 4W
                    if cube_util_jb >= 75 and branch_count <= max_jb_drops:
                        recommended = 'JB'
                    else:
                        recommended = '4W'
            # 🚛 กรุงเทพ+ปริมณฑล (nearby) → บังคับห้าม 6W (ลำดับแรกสุด!)
            elif all_nearby:
                # 🔒 Punthai ล้วน: JB ไม่เกิน 7 drop
                max_jb_drops = 7 if punthai_type == 'punthai_only' else 12
                
                # ลอง 4W ก่อน
                if cube_util_4w <= 100 and weight_util_4w <= 100:
                    recommended = '4W'
                # ถ้า 4W ไม่พอ → ลอง JB (เช็ค drop limit สำหรับ Punthai)
                elif cube_util_jb <= 100 and weight_util_jb <= 100 and branch_count <= max_jb_drops:
                    recommended = 'JB'
                    region_changes['nearby_6w_to_jb'] += 1
                # ถ้า JB ก็ไม่พอ หรือ Punthai เกิน 7 drop → ต้องแยกทริป
                else:
                    recommended = 'JB'  # กำหนดไว้ก่อน จะแยกภายหลัง
                    region_changes['nearby_6w_to_jb'] += 1
            # 🚛 ภาคเหนือทั้งหมด → บังคับใช้ 6W เท่านั้น (ถ้าได้ ≥18 cube)
            elif has_north:
                if total_c >= 18.0:
                    recommended = '6W'
                    region_changes['far_keep_6w'] += 1
                else:
                    # เหนือแต่ไม่ถึง 18 cube → แยกเป็น JB
                    recommended = 'JB'
                    region_changes['other'] += 1
            # 🚛 ภาคใต้ทั้งหมด → บังคับใช้ 6W เท่านั้น (ถ้าได้ ≥18 cube)
            elif has_south:
                if total_c >= 18.0:
                    recommended = '6W'
                    region_changes['far_keep_6w'] += 1
                else:
                    # ใต้แต่ไม่ถึง 18 cube → แยกเป็น JB
                    recommended = 'JB'
                    region_changes['other'] += 1
            else:
                # 🎯 พื้นที่ไกล (far) - ยืดหยุ่น ใช้ JB ได้ถ้าเหมาะสม
                # เป้าหมาย: 6W ต้องได้ขั้นต่ำ 18 cube (90%), ห้ามเกิน 20 cube
                
                MIN_6W_CUBE = 18.0  # 6W ต้องได้ขั้นต่ำ 18 cube
                MAX_6W_CUBE = 20.0  # 6W ห้ามเกิน 20 cube
                MIN_UTIL_THRESHOLD = 75   # ขั้นต่ำ - ห้ามรถเหลือต่ำกว่านี้
                TARGET_MIN = 95 # เป้าหมายขั้นต่ำ
                TARGET_MAX = 100 # เป้าหมายสูงสุด (ห้ามเกิน 100%)
                
                # 🔒 6W เกิน 20 cube → ต้องแยก ส่วนเกินไป 4W
                if total_c > MAX_6W_CUBE:
                    recommended = '6W'  # จะแยกส่วนเกินไป 4W ใน Phase 2.5
                    region_changes['far_keep_6w'] += 1
                
                # 🎯 6W ได้ 18-20 cube → ใช้ 6W ✅
                elif total_c >= MIN_6W_CUBE:
                    recommended = '6W'
                    region_changes['far_keep_6w'] += 1
                
                # 🔒 6W ไม่ถึง 18 cube (7-18) → แยกเป็น JB แทน
                elif total_c >= 7.0:
                    recommended = 'JB'  # จะแยกเป็น JB หลายคันใน Phase 2.1
                    region_changes['other'] += 1
                
                # 2. ถ้า JB พอดี (95-100%) → ใช้ JB ✅
                elif TARGET_MIN <= cube_util_jb <= TARGET_MAX and weight_util_jb <= TARGET_MAX:
                    recommended = 'JB'
                    region_changes['other'] += 1
                
                # 3. 4W พอดี → ใช้ 4W
                elif cube_util_4w <= TARGET_MAX and weight_util_4w <= TARGET_MAX:
                    recommended = '4W'
                    region_changes['other'] += 1
                
                # 4. ไม่มีทางเลือกอื่น → ใช้ JB
                else:
                    recommended = 'JB'
                    region_changes['other'] += 1
            
            # 🚨 บังคับใช้ max_allowed ถ้ารถที่แนะนำใหญ่กว่าข้อจำกัด (ห้ามข้าม!)
            # ✅ ทุกภาคต้องดู Auto Plan
            vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
            recommended_priority = vehicle_priority.get(recommended, 3)
            allowed_priority = vehicle_priority.get(max_allowed, 3)
            
            if recommended_priority > allowed_priority:
                # รถที่แนะนำใหญ่กว่าที่อนุญาต → บังคับใช้ max_allowed (ห้ามข้ามขั้น!)
                recommended = max_allowed
            
            # 🔒 Double check: ห้ามข้ามข้อจำกัดสาขาเด็ดขาด!
            if max_allowed == '4W' and recommended != '4W':
                recommended = '4W'
            elif max_allowed == 'JB' and recommended == '6W':
                recommended = 'JB'
            
            # 🔒 Triple check: กรุงเทพ+ปริมณฑล ห้าม 6W เด็ดขาด!
            if all_nearby and recommended == '6W':
                # บังคับเปลี่ยนเป็น JB
                recommended = 'JB'
                region_changes['nearby_6w_to_jb'] += 1
            
            # บันทึกการปรับขนาด
            original_vehicle = trip_recommended_vehicles.get(trip_num, '6W')
            trip_recommended_vehicles[trip_num] = recommended
            if recommended != original_vehicle:
                downsize_count += 1
    
    # Phase 2 completed
    
    # 🚨 ตรวจสอบอีกครั้ง: ห้ามกรุงเทพ+ปริมณฑล+ภาคกลางใช้ 6W (เข้มงวด)
    bangkok_6w_count = 0
    bangkok_6w_splits = 0
    
    for trip_num in test_df['Trip'].unique():
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        trip_codes = list(trip_data['Code'].values)
        
        # เช็คว่าทุกจังหวัดเป็นพื้นที่ใกล้หรือไม่
        provinces = set()
        for code in trip_codes:
            prov = get_province(code)
            if prov != 'UNKNOWN':
                provinces.add(prov)
        
        all_nearby = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
        current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
        
        if all_nearby and current_vehicle == '6W':
            # พยายามเปลี่ยนเป็น JB ก่อน
            total_w = trip_data['Weight'].sum()
            total_c = trip_data['Cube'].sum()
            jb_util = max((total_w / LIMITS['JB']['max_w']) * 100, (total_c / LIMITS['JB']['max_c']) * 100)
            
            if jb_util <= 100:
                # JB ใส่ได้ → เปลี่ยนเป็น JB
                trip_recommended_vehicles[trip_num] = 'JB'
                bangkok_6w_count += 1
            else:
                # JB เต็ม → แยกเป็น JB หลายคัน
                new_trips = []
                current_group = []
                current_group_w = 0
                current_group_c = 0
                
                sorted_data = trip_data.sort_values('Weight', ascending=False)
                
                for _, row in sorted_data.iterrows():
                    code = row['Code']
                    w = row['Weight']
                    c = row['Cube']
                    
                    test_w = current_group_w + w
                    test_c = current_group_c + c
                    test_util = max((test_w / LIMITS['JB']['max_w']) * 100, (test_c / LIMITS['JB']['max_c']) * 100)
                    
                    if test_util <= 100 or len(current_group) == 0:
                        current_group.append(code)
                        current_group_w += w
                        current_group_c += c
                    else:
                        new_trips.append(current_group.copy())
                        current_group = [code]
                        current_group_w = w
                        current_group_c = c
                
                if current_group:
                    new_trips.append(current_group)
                
                # อัพเดททริป
                if len(new_trips) > 1:
                    for code in new_trips[0]:
                        test_df.loc[test_df['Code'] == code, 'Trip'] = trip_num
                    trip_recommended_vehicles[trip_num] = 'JB'
                    
                    for group in new_trips[1:]:
                        new_trip_num = test_df['Trip'].max() + 1
                        for code in group:
                            test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip_num
                        trip_recommended_vehicles[new_trip_num] = 'JB'
                        bangkok_6w_splits += 1
                else:
                    trip_recommended_vehicles[trip_num] = 'JB'
                    bangkok_6w_count += 1
    
    # 🚨 Phase 2.1: ตรวจสอบและแก้ไขทริปที่ใช้รถใหญ่เกินข้อจำกัด (ลดการ loop)
    fix_count = 0
    split_count = 0
    
    for trip_num in test_df['Trip'].unique():
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        trip_codes = list(trip_data['Code'].values)
        current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
        max_allowed = get_max_vehicle_for_trip(set(trip_codes))
        
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        
        # 🔒 เช็คจังหวัด - ห้าม 6W ในปริมณฑล!
        provinces = set()
        for code in trip_codes:
            prov = get_province(code)
            if prov and prov != 'UNKNOWN':
                provinces.add(prov)
        all_nearby = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
        
        # 🔒 ปริมณฑล = บังคับ JB หรือ 4W (ห้าม 6W)
        if all_nearby and max_allowed == '6W':
            max_allowed = 'JB'
        
        # 🔒 เช็คว่ารถที่ใช้อยู่ใหญ่กว่าที่อนุญาตหรือไม่ (ห้ามข้ามขั้น!)
        vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
        current_priority = vehicle_priority.get(current_vehicle, 3)
        allowed_priority = vehicle_priority.get(max_allowed, 3)
        
        allowed_w = LIMITS[max_allowed]['max_w']
        allowed_c = LIMITS[max_allowed]['max_c']
        
        # เช็ค utilization ของรถที่อนุญาต
        util_allowed = max((total_w / allowed_w) * 100, (total_c / allowed_c) * 100)
        
        # 🚨 บังคับ: ต้องเคารพข้อจำกัดสาขา ไม่ว่าประวัติจะบอกอะไร!
        # กรณีที่ 1: รถใหญ่กว่าที่อนุญาต → บังคับเปลี่ยนหรือแยก
        # กรณีที่ 2: รถตามข้อจำกัดแต่ใส่ไม่ได้ (>100%) → ตัดแยกทันที!
        if current_priority > allowed_priority or util_allowed > 100:
            if util_allowed <= 100:
                # ใส่รถที่อนุญาตได้ → ปรับรถ
                trip_recommended_vehicles[trip_num] = max_allowed
                fix_count += 1
            else:
                # 🚨 ใส่ไม่ได้ → ต้องแยกทริป
                # 🎯 กลยุทธ์ใหม่: 
                #   - ถ้าจำกัด 4W → ลอง 4W ก่อน → ถ้าเกินค่อยตัดเป็น JB หลายคัน
                #   - ถ้าจำกัด JB → แยกเป็น JB หลายคัน
                #   - ห้าม 6W ในพื้นที่ใดๆ → ใช้รถเล็ก (4W/JB) เลย
                
                target_vehicle = max_allowed
                
                # 🚨 กรณีพิเศษ: 4W จำกัด → ลอง 4W ก่อน
                if max_allowed == '4W':
                    # ลอง 4W หลายคันก่อน
                    fourw_w = LIMITS['4W']['max_w']
                    fourw_c = LIMITS['4W']['max_c']
                    fourw_util = max((total_w / fourw_w) * 100, (total_c / fourw_c) * 100)
                    
                    # ถ้า 4W ใส่ได้ (ไม่เกิน 100%) → ใช้ 4W
                    if fourw_util <= 100:
                        trip_recommended_vehicles[trip_num] = '4W'
                        fix_count += 1
                        continue
                    else:
                        # 4W เต็ม → ตัดเป็น JB (เพราะห้าม 6W ในพื้นที่นี้)
                        target_vehicle = 'JB'
                
                target_w = LIMITS[target_vehicle]['max_w']
                target_c = LIMITS[target_vehicle]['max_c']
                split_needed = True
                
                # 🎯 แยกทริปโดยจัดกลุ่มตามตำแหน่งและเลือกรถที่เหมาะสม
                if split_needed:
                    # 📍 Step 1: จัดกลุ่มสาขาตามตำแหน่ง (ใกล้กันอยู่ด้วยกัน)
                    branch_info = []
                    for _, row in trip_data.iterrows():
                        code = row['Code']
                        lat, lon = get_lat_lon(code)
                        branch_info.append({
                            'code': code,
                            'weight': row['Weight'],
                            'cube': row['Cube'],
                            'lat': lat if lat else 0,
                            'lon': lon if lon else 0
                        })
                    
                    # เรียงตาม lat, lon เพื่อจัดกลุ่มใกล้กัน
                    branch_info.sort(key=lambda x: (x['lat'], x['lon']))
                    
                    # 📍 Step 2: สร้างกลุ่มโดยพิจารณาทั้ง Cube และตำแหน่ง
                    new_trips = []
                    current_group = []
                    current_group_w = 0
                    current_group_c = 0
                    current_centroid_lat = None
                    current_centroid_lon = None
                    
                    for branch in branch_info:
                        code = branch['code']
                        w = branch['weight']
                        c = branch['cube']
                        b_lat = branch['lat']
                        b_lon = branch['lon']
                        
                        # คำนวณระยะห่างจาก centroid ของกลุ่มปัจจุบัน
                        if current_centroid_lat and b_lat:
                            distance_from_group = haversine_distance(current_centroid_lat, current_centroid_lon, b_lat, b_lon)
                        else:
                            distance_from_group = 0
                        
                        # เช็คว่าถ้าเพิ่มสาขานี้ จะเกินรถเป้าหมายไหม
                        test_w = current_group_w + w
                        test_c = current_group_c + c
                        test_util = max((test_w / target_w) * 100, (test_c / target_c) * 100)
                        
                        # 🚨 เงื่อนไขสำคัญ: รถเล็ก (4W/JB) ห้ามเกิน 12 สาขา, 6W ไม่จำกัด
                        max_branches = 12 if target_vehicle in ['4W', 'JB'] else float('inf')
                        
                        # เงื่อนไขเพิ่ม: ถ้าสาขาห่างจากกลุ่มมากเกิน 50km → แยกกลุ่มใหม่
                        too_far = distance_from_group > 50 and len(current_group) > 0
                        
                        # เป้าหมาย: 95-100% และไม่เกินจำนวนสาขา และไม่ไกลเกินไป
                        if ((test_util <= 100 and len(current_group) < max_branches and not too_far) or 
                            len(current_group) == 0):
                            # ใส่ได้
                            current_group.append(code)
                            current_group_w += w
                            current_group_c += c
                            # อัปเดต centroid
                            if b_lat:
                                if current_centroid_lat is None:
                                    current_centroid_lat = b_lat
                                    current_centroid_lon = b_lon
                                else:
                                    n = len(current_group)
                                    current_centroid_lat = ((current_centroid_lat * (n-1)) + b_lat) / n
                                    current_centroid_lon = ((current_centroid_lon * (n-1)) + b_lon) / n
                        else:
                            # เต็มแล้ว หรือ ไกลเกินไป → สร้างกลุ่มใหม่
                            current_util = max((current_group_w / target_w) * 100, (current_group_c / target_c) * 100)
                            
                            if current_util >= 95 or len(current_group) >= 12 or too_far:
                                new_trips.append({
                                    'codes': current_group.copy(),
                                    'weight': current_group_w,
                                    'cube': current_group_c
                                })
                                current_group = [code]
                                current_group_w = w
                                current_group_c = c
                                current_centroid_lat = b_lat if b_lat else None
                                current_centroid_lon = b_lon if b_lon else None
                            else:
                                # ยังไม่เต็มพอ → ใส่ต่อ
                                current_group.append(code)
                                current_group_w += w
                                current_group_c += c
                    
                    # เพิ่มกลุ่มสุดท้าย
                    if current_group:
                        new_trips.append({
                            'codes': current_group,
                            'weight': current_group_w,
                            'cube': current_group_c
                        })
                    
                    # 📍 Step 3: เลือกรถที่เหมาะสมสำหรับแต่ละกลุ่ม (อาจคนละประเภท)
                    for trip_info in new_trips:
                        trip_w = trip_info['weight']
                        trip_c = trip_info['cube']
                        trip_branches = len(trip_info['codes'])
                        
                        # ลอง 4W ก่อน (ถ้าไม่มีข้อจำกัดและไม่เกิน 12 สาขา)
                        util_4w = max((trip_w / LIMITS['4W']['max_w']) * 100, 
                                     (trip_c / LIMITS['4W']['max_c']) * 100)
                        util_jb = max((trip_w / LIMITS['JB']['max_w']) * 100,
                                     (trip_c / LIMITS['JB']['max_c']) * 100)
                        util_6w = max((trip_w / LIMITS['6W']['max_w']) * 100,
                                     (trip_c / LIMITS['6W']['max_c']) * 100)
                        
                        # เลือกรถที่เหมาะสมที่สุด (Cube 95-120%)
                        if trip_branches <= 12:
                            if 95 <= util_4w <= 100 and max_allowed != 'JB' and max_allowed != '6W':
                                trip_info['vehicle'] = '4W'
                            elif 95 <= util_jb <= 100 and max_allowed != '6W':
                                trip_info['vehicle'] = 'JB'
                            elif util_6w <= 200 and max_allowed == '6W':
                                trip_info['vehicle'] = '6W'
                            elif util_jb <= 100 and max_allowed != '6W':
                                trip_info['vehicle'] = 'JB'
                            elif util_4w <= 100 and max_allowed != 'JB' and max_allowed != '6W':
                                trip_info['vehicle'] = '4W'
                            else:
                                trip_info['vehicle'] = target_vehicle
                        else:
                            # เกิน 12 สาขา → ใช้ 6W
                            trip_info['vehicle'] = '6W' if max_allowed == '6W' else 'JB'
                    
                    # 📍 Step 4: รวมทริปที่น้อยเกินไป (<95%) กับทริปที่ใกล้ที่สุด
                    final_trips = []
                    low_util_trips = []
                    
                    for trip_info in new_trips:
                        vehicle = trip_info.get('vehicle', target_vehicle)
                        v_w = LIMITS[vehicle]['max_w']
                        v_c = LIMITS[vehicle]['max_c']
                        trip_util = max((trip_info['weight'] / v_w) * 100, (trip_info['cube'] / v_c) * 100)
                        
                        if trip_util >= 95:
                            final_trips.append(trip_info)
                        else:
                            low_util_trips.append(trip_info)
                    
                    # กระจายสาขาจากทริปที่น้อยเกินไป
                    for low_trip in low_util_trips:
                        for code in low_trip['codes']:
                            branch_w = test_df[test_df['Code'] == code]['Weight'].sum()
                            branch_c = test_df[test_df['Code'] == code]['Cube'].sum()
                            branch_lat, branch_lon = get_lat_lon(code)
                            
                            best_trip_idx = -1
                            best_score = float('inf')
                            
                            for idx, trip_info in enumerate(final_trips):
                                # คำนวณ centroid ของทริป
                                trip_coords = []
                                for c in trip_info['codes']:
                                    lat, lon = get_lat_lon(c)
                                    if lat and lon:
                                        trip_coords.append((lat, lon))
                                
                                if trip_coords and branch_lat:
                                    centroid_lat = sum(c[0] for c in trip_coords) / len(trip_coords)
                                    centroid_lon = sum(c[1] for c in trip_coords) / len(trip_coords)
                                    distance = haversine_distance(branch_lat, branch_lon, centroid_lat, centroid_lon)
                                else:
                                    distance = 50
                                
                                vehicle = trip_info.get('vehicle', target_vehicle)
                                v_w = LIMITS[vehicle]['max_w']
                                v_c = LIMITS[vehicle]['max_c']
                                new_w = trip_info['weight'] + branch_w
                                new_c = trip_info['cube'] + branch_c
                                new_util = max((new_w / v_w) * 100, (new_c / v_c) * 100)
                                
                                if new_util <= 100 and len(trip_info['codes']) < 12:
                                    score = distance + (new_util - 100) * 0.5
                                    if score < best_score:
                                        best_score = score
                                        best_trip_idx = idx
                            
                            if best_trip_idx >= 0:
                                final_trips[best_trip_idx]['codes'].append(code)
                                final_trips[best_trip_idx]['weight'] += branch_w
                                final_trips[best_trip_idx]['cube'] += branch_c
                            else:
                                # สร้างทริปใหม่
                                final_trips.append({
                                    'codes': [code],
                                    'weight': branch_w,
                                    'cube': branch_c,
                                    'vehicle': target_vehicle
                                })
                    
                    # 📍 Step 5: อัปเดต DataFrame และกำหนดรถที่เหมาะสม
                    if len(final_trips) >= 1:
                        for idx, trip_info in enumerate(final_trips):
                            codes = trip_info['codes']
                            vehicle = trip_info.get('vehicle', target_vehicle)
                            
                            # ตรวจสอบข้อจำกัดสาขา
                            group_max_allowed = get_max_vehicle_for_trip(set(codes))
                            vehicle_priority = {'4W': 1, 'JB': 2, '6W': 3}
                            if vehicle_priority.get(vehicle, 3) > vehicle_priority.get(group_max_allowed, 3):
                                vehicle = group_max_allowed
                            
                            if idx == 0:
                                # ทริปแรกใช้เลขเดิม
                                for code in codes:
                                    test_df.loc[test_df['Code'] == code, 'Trip'] = trip_num
                                trip_recommended_vehicles[trip_num] = vehicle
                            else:
                                # ทริปถัดไปสร้างใหม่
                                new_trip_num = test_df['Trip'].max() + 1
                                for code in codes:
                                    test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip_num
                                trip_recommended_vehicles[new_trip_num] = vehicle
                                split_count += 1
                    else:
                        # ไม่แยก → ใช้รถเดิม
                        trip_recommended_vehicles[trip_num] = target_vehicle
                        fix_count += 1
    # 🎯 Phase 2.5: แยกทริปที่ Cube เกินไปมาก (น้ำหนักเบา แต่เต็ม Cube)
    cube_split_count = 0
    next_trip_num = test_df['Trip'].max() + 1
    
    for trip_num in sorted(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
            
        trip_data = test_df[test_df['Trip'] == trip_num]
        current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
        
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        
        # คำนวณ Cube utilization
        should_split = False
        target_vehicle = current_vehicle
        
        if current_vehicle == '4W':
            cube_util = (total_c / LIMITS['4W']['max_c']) * 100
            weight_util = (total_w / LIMITS['4W']['max_w']) * 100
            # 4W Cube เกิน 100% → แยก
            if cube_util > 100 and len(trip_data) >= 4:
                should_split = True
                target_vehicle = 'JB'
        elif current_vehicle == 'JB':
            cube_util = (total_c / LIMITS['JB']['max_c']) * 100
            weight_util = (total_w / LIMITS['JB']['max_w']) * 100
            # JB Cube เกิน 100% → แยก (โดยเฉพาะกรุงเทพที่ห้ามใช้ 6W)
            if cube_util > 100 and len(trip_data) >= 4:
                should_split = True
                target_vehicle = 'JB'  # แยกเป็น JB อีกคัน
        elif current_vehicle == '6W':
            # 🚛 6W ไม่จำกัดคิว - ใส่ได้เต็มที่ไม่ต้องแยก
            should_split = False
        
        if should_split:
            # แยกทริปตาม Cube (เรียงตาม Cube จากมากไปน้อย แล้วแบ่งครึ่ง)
            trip_data_sorted = trip_data.sort_values('Cube', ascending=False)
            codes = list(trip_data_sorted['Code'].values)
            
            # แบ่งสาขาเป็น 2 กลุ่มให้ Cube ใกล้เคียงกัน
            g1_codes, g2_codes = [], []
            g1_cube, g2_cube = 0, 0
            
            for code in codes:
                branch_cube = trip_data_sorted[trip_data_sorted['Code'] == code]['Cube'].sum()
                if g1_cube <= g2_cube:
                    g1_codes.append(code)
                    g1_cube += branch_cube
                else:
                    g2_codes.append(code)
                    g2_cube += branch_cube
            
            # เช็คว่าแต่ละกลุ่มพอดีกับรถเป้าหมายหรือไม่
            g1_w = trip_data[trip_data['Code'].isin(g1_codes)]['Weight'].sum()
            g1_c = trip_data[trip_data['Code'].isin(g1_codes)]['Cube'].sum()
            g2_w = trip_data[trip_data['Code'].isin(g2_codes)]['Weight'].sum()
            g2_c = trip_data[trip_data['Code'].isin(g2_codes)]['Cube'].sum()
            
            g1_cube_util = (g1_c / LIMITS[target_vehicle]['max_c']) * 100
            g1_weight_util = (g1_w / LIMITS[target_vehicle]['max_w']) * 100
            g2_cube_util = (g2_c / LIMITS[target_vehicle]['max_c']) * 100
            g2_weight_util = (g2_w / LIMITS[target_vehicle]['max_w']) * 100
            
            # ตรวจสอบว่าทั้ง 2 กลุ่มใช้รถเป้าหมายได้และมีประสิทธิภาพ (Cube ≥50%, น้ำหนัก ≤100%)
            g1_ok = g1_cube_util <= 100 and g1_weight_util <= 100 and g1_cube_util >= 50
            g2_ok = g2_cube_util <= 100 and g2_weight_util <= 100 and g2_cube_util >= 50
            
            # 🚨 เช็คว่าถ้าแยกแล้วรถใหม่ไม่เต็ม → ไม่ต้องแยก ให้ยัดใส่รถเดิมแม้เกิน
            if not (g1_ok and g2_ok):
                # ถ้าแยกแล้วรถใดรถหนึ่งไม่เต็ม (Cube <100%) → ไม่แยก
                # ยอมให้รถเดิมเกิน 120% ได้ เพื่อไม่ให้รถใหม่วิ่งไม่คุ้ม
                should_split = False
            
            if should_split and g1_ok and g2_ok and len(g1_codes) >= 2 and len(g2_codes) >= 2:
                # แยกทริป: เก็บ trip_num เดิม ให้ g1, สร้างทริปใหม่ให้ g2
                for code in g2_codes:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = next_trip_num
                
                # บันทึกรถทั้ง 2 ทริป
                trip_recommended_vehicles[trip_num] = target_vehicle
                trip_recommended_vehicles[next_trip_num] = target_vehicle
                
                next_trip_num += 1
                cube_split_count += 1
    
    # 🎯 Phase 3: ปรับปรุง 6W ให้เหมาะสม
    # - 6W ≥200% Cube → ต้องแยก (เกินไปมาก)
    # - 6W 150-199% Cube → พิจารณาแยก (ถ้าทำได้)
    # - 6W <150% Cube → ไม่แยก (ใช้ 6W คุ้มค่า)
    split_count = 0
    
    # หาทริปที่ใช้ 6W และ Cube ≥150%
    trips_to_check = []
    for trip_num in test_df['Trip'].unique():
        if trip_num == 0:
            continue
            
        trip_data = test_df[test_df['Trip'] == trip_num]
        current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
        
        if current_vehicle != '6W':
            continue
        
        # คำนวณ Cube utilization
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        cube_util = (total_c / LIMITS['6W']['max_c']) * 100
        
        # 🚛 6W ≥100% → พิจารณาแยก (≥200% บังคับแยก)
        if cube_util >= 100 and len(trip_data) >= 4:
            trips_to_check.append({
                'trip': trip_num,
                'data': trip_data,
                'cube_util': cube_util,
                'total_w': total_w,
                'total_c': total_c,
                'force_split': cube_util >= 200  # บังคับแยกถ้า ≥200%
            })
    
    # แยกทริปที่มี Cube ≥150% (≥200% บังคับแยก)
    for trip_info in trips_to_check:
        trip_num = trip_info['trip']
        trip_data = trip_info['data']
        trip_codes = list(trip_data['Code'].values)
        
        # เช็คข้อจำกัดสาขา
        max_allowed = get_max_vehicle_for_trip(set(trip_codes))
        if max_allowed == '6W':
            # ไม่มีข้อจำกัดสาขา → ลองแยกเป็นรถเล็ก
            
            # วิเคราะห์ spatial clusters
            clusters = create_distance_based_clusters(trip_codes, max_distance_km=25)
            
            # ถ้ามี ≥2 กลุ่ม → ลองแยก
            if len(clusters) >= 2:
                # จัดเรียงกลุ่มตามน้ำหนัก/คิว
                cluster_info = []
                for cluster_codes in clusters:
                    cluster_data = trip_data[trip_data['Code'].isin(cluster_codes)]
                    cluster_w = cluster_data['Weight'].sum()
                    cluster_c = cluster_data['Cube'].sum()
                    cluster_info.append({
                        'codes': cluster_codes,
                        'weight': cluster_w,
                        'cube': cluster_c,
                        'branches': len(cluster_codes)
                    })
                
                # เรียงตามคิว (มาก→น้อย)
                cluster_info.sort(key=lambda x: x['cube'], reverse=True)
                
                # ลองจับคู่กลุ่มเพื่อสร้างรถใหม่
                new_trips = []
                used_clusters = set()
                
                for i, cluster in enumerate(cluster_info):
                    if i in used_clusters:
                        continue
                    
                    # เช็คว่ากลุ่มนี้พอดี JB หรือ 4W หรือไม่
                    util_jb = max((cluster['weight'] / LIMITS['JB']['max_w']) * 100,
                                 (cluster['cube'] / LIMITS['JB']['max_c']) * 100)
                    util_4w = max((cluster['weight'] / LIMITS['4W']['max_w']) * 100,
                                 (cluster['cube'] / LIMITS['4W']['max_c']) * 100)
                    
                    # ถ้ากลุ่มนี้มีสาขา ≤12 และพอดี JB หรือ 4W
                    if cluster['branches'] <= 12:
                        if util_4w >= 90 and util_4w <= 100:
                            # พอดี 4W
                            new_trips.append({
                                'codes': cluster['codes'],
                                'vehicle': '4W'
                            })
                            used_clusters.add(i)
                        elif util_jb >= 90 and util_jb <= 100:
                            # พอดี JB
                            new_trips.append({
                                'codes': cluster['codes'],
                                'vehicle': 'JB'
                            })
                            used_clusters.add(i)
                        else:
                            # ลองรวมกับกลุ่มอื่น
                            for j, other_cluster in enumerate(cluster_info):
                                if j <= i or j in used_clusters:
                                    continue
                                
                                combined_codes = cluster['codes'] + other_cluster['codes']
                                combined_w = cluster['weight'] + other_cluster['weight']
                                combined_c = cluster['cube'] + other_cluster['cube']
                                combined_branches = cluster['branches'] + other_cluster['branches']
                                
                                if combined_branches <= 12:
                                    combined_util_jb = max((combined_w / LIMITS['JB']['max_w']) * 100,
                                                          (combined_c / LIMITS['JB']['max_c']) * 100)
                                    combined_util_4w = max((combined_w / LIMITS['4W']['max_w']) * 100,
                                                          (combined_c / LIMITS['4W']['max_c']) * 100)
                                    
                                    if combined_util_4w >= 90 and combined_util_4w <= 100:
                                        new_trips.append({
                                            'codes': combined_codes,
                                            'vehicle': '4W'
                                        })
                                        used_clusters.add(i)
                                        used_clusters.add(j)
                                        break
                                    elif combined_util_jb >= 90 and combined_util_jb <= 100:
                                        new_trips.append({
                                            'codes': combined_codes,
                                            'vehicle': 'JB'
                                        })
                                        used_clusters.add(i)
                                        used_clusters.add(j)
                                        break
                
                # 🚨 เงื่อนไขการแยก:
                # - ถ้า force_split = True (≥200%) → บังคับแยกเสมอ
                # - ถ้า 150-199% → แยกถ้าได้อย่างน้อย 2 ทริปและใช้รถ ≥90%
                force_split = trip_info.get('force_split', False)
                should_split = force_split or len(new_trips) >= 2
                
                if should_split:
                    # สร้างทริปใหม่
                    max_trip = test_df['Trip'].max()
                    
                    # ถ้าบังคับแยกแต่ยังไม่มี new_trips → แบ่งครึ่ง
                    if force_split and len(new_trips) < 2:
                        # แบ่งทริปครึ่งหนึ่งตาม Cube
                        sorted_data = trip_data.sort_values('Cube', ascending=False)
                        mid = len(sorted_data) // 2
                        g1_codes = list(sorted_data.iloc[:mid]['Code'].values)
                        g2_codes = list(sorted_data.iloc[mid:]['Code'].values)
                        
                        # กำหนดรถให้แต่ละกลุ่ม (ใช้ 6W เพื่อรองรับ Cube สูง)
                        new_trips = [
                            {'codes': g1_codes, 'vehicle': '6W'},
                            {'codes': g2_codes, 'vehicle': '6W'}
                        ]
                    
                    for idx, new_trip_info in enumerate(new_trips):
                        if idx == 0:
                            # ทริปแรกใช้เลขเดิม
                            for code in new_trip_info['codes']:
                                test_df.loc[test_df['Code'] == code, 'Trip'] = trip_num
                            trip_recommended_vehicles[trip_num] = new_trip_info['vehicle']
                        else:
                            # ทริปถัดไปสร้างใหม่
                            new_trip_num = max_trip + idx
                            for code in new_trip_info['codes']:
                                test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip_num
                            trip_recommended_vehicles[new_trip_num] = new_trip_info['vehicle']
                            split_count += 1
    
    # 🔄 Phase 4: บังคับเปลี่ยน nearby จาก 6W → JB/4W และกระจายทริปน้อย (เฉพาะที่จำเป็น) - Optimized
    low_util_trips = []
    
    # ⚡ Skip ถ้าใช้เวลามากกว่า 22 วินาที
    if time.time() - start_time > 22:
        pass  # Skip Phase 4 เพื่อความเร็ว
    else:
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            trip_codes = set(trip_data['Code'].values)
            current_vehicle = trip_recommended_vehicles.get(trip_num, '4W')  # Start with 4W
            total_w = trip_data['Weight'].sum()
            total_c = trip_data['Cube'].sum()
            
            # เช็คว่าเป็น nearby หรือไม่
            provinces = set()
            for code in trip_codes:
                prov = get_province(code)
            if prov != 'UNKNOWN':
                provinces.add(prov)
        
        all_nearby = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
        
        # ถ้าใช้ 6W และเป็น nearby → บังคับเปลี่ยนเป็น JB
        if current_vehicle == '6W' and all_nearby:
            jb_util = max((total_w / LIMITS['JB']['max_w']) * 100, 
                         (total_c / LIMITS['JB']['max_c']) * 100)
            if jb_util <= 100:
                trip_recommended_vehicles[trip_num] = 'JB'
                current_vehicle = 'JB'
            else:
                trip_recommended_vehicles[trip_num] = 'JB'
                current_vehicle = 'JB'
        
        # หาทริปที่ใช้รถน้อยเกินไป (<65% และ ≤ 2 สาขา - เฉพาะที่จำเป็นจริงๆ)
        util = max((total_w / LIMITS[current_vehicle]['max_w']) * 100,
                   (total_c / LIMITS[current_vehicle]['max_c']) * 100)
        
        if util < 65 and len(trip_data) <= 2:
            low_util_trips.append({
                'trip_num': trip_num,
                'codes': list(trip_codes),
                'weight': total_w,
                'cube': total_c,
                'vehicle': current_vehicle
            })
    
    # กระจายทริปที่น้อยเกินไป (Skip ถ้ามีมาก - เพื่อความเร็ว)
    if len(low_util_trips) > 15:
        low_util_trips = []  # Skip เพื่อความเร็ว
    
    # กระจายเฉพาะที่จำเป็น
    if len(low_util_trips) == 0:
        pass  # Skip ถ้าไม่มีทริปน้อย
    else:
        for low_trip in low_util_trips:
            # หาทริปที่ใกล้ที่สุดและรับได้
            best_target_trip = None
            best_score = float('inf')
            
            for target_trip_num in test_df['Trip'].unique():
                if target_trip_num == 0 or target_trip_num == low_trip['trip_num']:
                    continue
                
                target_data = test_df[test_df['Trip'] == target_trip_num]
                target_vehicle = trip_recommended_vehicles.get(target_trip_num, '6W')
            
            # เช็คว่ารวมได้ไหม
            pass
            new_w = target_data['Weight'].sum() + low_trip['weight']
            new_c = target_data['Cube'].sum() + low_trip['cube']
            new_util = max((new_w / LIMITS[target_vehicle]['max_w']) * 100,
                          (new_c / LIMITS[target_vehicle]['max_c']) * 100)
            
            max_branches = 12 if target_vehicle in ['4W', 'JB'] else float('inf')
            
            if new_util <= 100 and len(target_data) + len(low_trip['codes']) <= max_branches:
                # คำนวณระยะห่าง (เฉลี่ย)
                score = new_util
                if score < best_score:
                    best_score = score
                    best_target_trip = target_trip_num
        
        # ย้ายสาขา
        if best_target_trip:
            for code in low_trip['codes']:
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_target_trip
                # ลบทริปเดิม
                if low_trip['trip_num'] in trip_recommended_vehicles:
                    del trip_recommended_vehicles[low_trip['trip_num']]
    
    # 🚨 Phase 5: Distance Optimization - สลับสาขาให้ใกล้กันมากขึ้น (FAST)
    # ตรวจสอบระยะห่างในทริป ถ้าเกิน MAX_DISTANCE_IN_TRIP → หาสาขาที่ใกล้กว่าจากทริปอื่นมาสลับ
    # ⚡ Skip ถ้าใช้เวลาเกิน 25 วินาที
    if time.time() - start_time > 25:
        distance_swaps = 999  # Skip Phase 5
    else:
        distance_swaps = 0
    max_distance_swaps = 30  # ลดจาก 100 เพื่อความเร็ว
    
    # 🔒 เก็บ ตำบล/อำเภอ/จังหวัด ของแต่ละสาขา
    def get_location_for_code(code):
        """คืนค่า (ตำบล, อำเภอ, จังหวัด) ของสาขา"""
        if not MASTER_DATA.empty:
            master = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
            if len(master) > 0:
                row = master.iloc[0]
                return (
                    row.get('ตำบล', ''),
                    row.get('อำเภอ', ''),
                    row.get('จังหวัด', '')
                )
        return ('', '', '')
    
    def get_province_for_code(code):
        return get_location_for_code(code)[2]
    
    def calculate_location_bonus(code1_loc, other_codes):
        """
        คำนวณ bonus ตามความใกล้เคียงของตำบล/อำเภอ/จังหวัด
        - ตำบลเดียวกัน: +15km bonus
        - อำเภอเดียวกัน: +10km bonus  
        - จังหวัดเดียวกัน: +5km bonus
        """
        subdistrict1, district1, province1 = code1_loc
        best_bonus = 0
        
        for code in other_codes:
            subdistrict2, district2, province2 = get_location_for_code(code)
            
            # ตำบลเดียวกัน + จังหวัดเดียวกัน = สูงสุด
            if subdistrict1 and subdistrict1 == subdistrict2 and province1 == province2:
                return 15  # Bonus สูงสุด
            
            # อำเภอเดียวกัน + จังหวัดเดียวกัน
            if district1 and district1 == district2 and province1 == province2:
                best_bonus = max(best_bonus, 10)
            
            # จังหวัดเดียวกัน
            elif province1 and province1 == province2:
                best_bonus = max(best_bonus, 5)
        
        return best_bonus
    
    for iteration in range(2):  # ⚡ ลดเป็น 2 รอบเพื่อความเร็ว
        if distance_swaps >= max_distance_swaps or time.time() - start_time > 28:
            break
            
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0 or distance_swaps >= max_distance_swaps:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            trip_codes = list(trip_data['Code'].values)
            
            if len(trip_codes) < 2:
                continue
            
            # หา centroid ของทริป
            trip_lats, trip_lons = [], []
            for code in trip_codes:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if not trip_lats:
                continue
                
            centroid_lat = sum(trip_lats) / len(trip_lats)
            centroid_lon = sum(trip_lons) / len(trip_lons)
            
            # หาสาขาที่ไกลจาก centroid มากที่สุด
            farthest_code = None
            farthest_dist = 0
            
            for code in trip_codes:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    dist = haversine_distance(centroid_lat, centroid_lon, lat, lon)
                    if dist > farthest_dist:
                        farthest_dist = dist
                        farthest_code = code
            
            # ถ้าสาขาไกลเกิน 40km จาก centroid → ลองหาทริปที่เหมาะกว่า (ลดจาก 50km)
            if farthest_dist > 40 and farthest_code:
                far_lat, far_lon = coord_cache.get(farthest_code, (None, None))
                if not far_lat:
                    continue
                
                far_branch_data = test_df[test_df['Code'] == farthest_code].iloc[0]
                far_weight = far_branch_data['Weight']
                far_cube = far_branch_data['Cube']
                far_province = get_province_for_code(farthest_code)
                
                # หาทริปอื่นที่ใกล้กว่าและยังใส่ได้
                best_new_trip = None
                best_new_dist = farthest_dist
                best_same_province = False
                
                for other_trip in test_df['Trip'].unique():
                    if other_trip == 0 or other_trip == trip_num:
                        continue
                    
                    other_data = test_df[test_df['Trip'] == other_trip]
                    other_codes = list(other_data['Code'].values)
                    
                    # เช็คจังหวัดในทริปอื่น
                    other_provinces = set()
                    for code in other_codes:
                        prov = get_province_for_code(code)
                        if prov:
                            other_provinces.add(prov)
                    
                    # ให้ priority กับทริปที่มีจังหวัดเดียวกัน
                    same_province = far_province in other_provinces
                    
                    # หา centroid ของทริปอื่น
                    other_lats, other_lons = [], []
                    for code in other_codes:
                        lat, lon = coord_cache.get(code, (None, None))
                        if lat and lon:
                            other_lats.append(lat)
                            other_lons.append(lon)
                    
                    if not other_lats:
                        continue
                    
                    other_centroid_lat = sum(other_lats) / len(other_lats)
                    other_centroid_lon = sum(other_lons) / len(other_lons)
                    
                    # คำนวณระยะห่างจากสาขาไกลไปยัง centroid ของทริปอื่น
                    dist_to_other = haversine_distance(far_lat, far_lon, other_centroid_lat, other_centroid_lon)
                    
                    # 🔒 เช็คระยะทางจริง + bonus ตามตำบล/อำเภอ/จังหวัด
                    # ตำบลเดียวกัน: +15km, อำเภอเดียวกัน: +10km, จังหวัดเดียวกัน: +5km
                    far_location = get_location_for_code(farthest_code)
                    location_bonus = calculate_location_bonus(far_location, other_codes)
                    effective_dist = dist_to_other - location_bonus
                    
                    # ต้องดีขึ้นอย่างน้อย 10km (หลังหัก bonus)
                    if effective_dist < best_new_dist - 10:
                        other_vehicle = trip_recommended_vehicles.get(other_trip, '4W')
                        other_total_w = other_data['Weight'].sum() + far_weight
                        other_total_c = other_data['Cube'].sum() + far_cube
                        
                        other_util = max(
                            (other_total_w / LIMITS[other_vehicle]['max_w']) * 100,
                            (other_total_c / LIMITS[other_vehicle]['max_c']) * 100
                        )
                        
                        max_branches = 12 if other_vehicle in ['4W', 'JB'] else float('inf')
                        
                        if other_util <= 100 and len(other_codes) < max_branches:
                            # 🚨 เช็ค consecutive distance หลังรวม
                            combined_codes = other_codes + [farthest_code]
                            max_consec = calculate_max_consecutive_distance(combined_codes, coord_cache)
                            if max_consec <= MAX_DISTANCE_IN_TRIP:
                                best_new_trip = other_trip
                                best_new_dist = effective_dist
                
                # ย้ายสาขาไปทริปใหม่
                if best_new_trip is not None:
                    test_df.loc[test_df['Code'] == farthest_code, 'Trip'] = best_new_trip
                    distance_swaps += 1
    
    # 🗺️ เรียงลำดับสาขา: ไกลสุดจาก DC ก่อน → ใกล้สุด (เพื่อให้รถวิ่งกลับมา DC)
    # ⚡ Skip ถ้าใช้เวลาเกิน 28 วินาที
    if time.time() - start_time <= 28:
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0:
                continue
            
            trip_codes = list(test_df[test_df['Trip'] == trip_num]['Code'].values)
            if len(trip_codes) < 2:  # ทำทุกทริปที่มี 2+ สาขา
                continue
            
            # 🆕 เรียงจากไกลสุดมาใกล้สุด: หาระยะทางจาก DC แล้ว sort
            distances_from_dc = []
            for code in trip_codes:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    dist = haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
                    distances_from_dc.append((code, dist))
                else:
                    distances_from_dc.append((code, 0))
            
            # เรียงจากไกลสุด (dist มากสุด) มาใกล้สุด (dist น้อยสุด)
            distances_from_dc.sort(key=lambda x: x[1], reverse=True)
            ordered = [x[0] for x in distances_from_dc]
            
            # อัปเดต Sequence
            for seq, code in enumerate(ordered, start=1):
                test_df.loc[(test_df['Code'] == code) & (test_df['Trip'] == trip_num), 'Sequence'] = seq
    
    # ===============================================
    # 🎯 Phase 6: Capacity Balancing - กระจาย load ให้เท่ากัน
    # MIN_UTIL: 4W ≥ 70%, JB ≥ 80%, 6W ≥ 90%
    # ===============================================
    balance_count = 0
    MAX_BALANCE_ITERATIONS = 3
    
    for balance_iter in range(MAX_BALANCE_ITERATIONS):
        if time.time() - start_time > 50:  # ถ้าใช้เวลาเกิน 50 วินาที → หยุด
            break
            
        # หาทริปที่ util ต่ำกว่า MIN_UTIL
        low_util_trips_balance = []
        high_util_trips_balance = []
        
        for trip_num in test_df['Trip'].unique():
            if trip_num == 0:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            trip_count = len(trip_data)
            trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
            
            # คำนวณ util ตามประเภทรถ
            limits = LIMITS.get(trip_vehicle, LIMITS['4W'])
            trip_util = max(
                (trip_w / limits['max_w']) * 100,
                (trip_c / limits['max_c']) * 100
            )
            
            min_util = MIN_UTIL.get(trip_vehicle, 70)
            
            if trip_util < min_util and trip_count <= 3:
                # ทริปที่ util ต่ำ → ต้องหาสาขาเพิ่ม หรือย้ายไปรวมกับทริปอื่น
                low_util_trips_balance.append({
                    'trip_num': trip_num,
                    'util': trip_util,
                    'count': trip_count,
                    'weight': trip_w,
                    'cube': trip_c,
                    'vehicle': trip_vehicle,
                    'codes': set(trip_data['Code'].values),
                    'min_util': min_util
                })
            elif trip_util > 95 and trip_count >= 3:
                # ทริปที่เกือบเต็ม → อาจกระจายให้ทริปอื่นได้
                high_util_trips_balance.append({
                    'trip_num': trip_num,
                    'util': trip_util,
                    'count': trip_count,
                    'weight': trip_w,
                    'cube': trip_c,
                    'vehicle': trip_vehicle,
                    'codes': list(trip_data['Code'].values)
                })
        
        if not low_util_trips_balance:
            break  # ไม่มีทริปที่ต้อง balance
        
        # พยายามย้ายทริป util ต่ำไปรวมกับทริปใกล้เคียง
        for low_trip in low_util_trips_balance:
            best_merge_trip = None
            best_merge_util = float('inf')
            best_merge_dist = float('inf')
            
            # หา centroid ของทริป util ต่ำ
            low_lats, low_lons = [], []
            for code in low_trip['codes']:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    low_lats.append(lat)
                    low_lons.append(lon)
            
            if not low_lats:
                continue
            
            low_centroid_lat = sum(low_lats) / len(low_lats)
            low_centroid_lon = sum(low_lons) / len(low_lons)
            
            for trip_num in test_df['Trip'].unique():
                if trip_num == 0 or trip_num == low_trip['trip_num']:
                    continue
                
                target_data = test_df[test_df['Trip'] == trip_num]
                target_w = target_data['Weight'].sum()
                target_c = target_data['Cube'].sum()
                target_count = len(target_data)
                target_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
                
                # 🚨 เช็คระยะทาง centroid ของ target
                target_lats, target_lons = [], []
                for code in target_data['Code'].values:
                    lat, lon = coord_cache.get(code, (None, None))
                    if lat and lon:
                        target_lats.append(lat)
                        target_lons.append(lon)
                
                if not target_lats:
                    continue
                
                target_centroid_lat = sum(target_lats) / len(target_lats)
                target_centroid_lon = sum(target_lons) / len(target_lons)
                
                # คำนวณระยะห่างระหว่าง centroids
                centroid_dist = haversine_distance(low_centroid_lat, low_centroid_lon,
                                                   target_centroid_lat, target_centroid_lon)
                
                # 🔒 ระยะห่างต้องไม่เกิน MAX_DISTANCE_IN_TRIP (50km)
                if centroid_dist > MAX_DISTANCE_IN_TRIP:
                    continue
                
                # เช็คว่ารวมได้ไหม
                new_w = target_w + low_trip['weight']
                new_c = target_c + low_trip['cube']
                new_count = target_count + low_trip['count']
                
                # ไม่เกิน 12 สาขา (สำหรับ 4W/JB)
                max_branches = 12 if target_vehicle in ['JB'] else 20
                if new_count > max_branches:
                    continue
                
                # 🚨 เช็ค consecutive distance หลังรวม
                combined_codes = list(target_data['Code'].values) + list(low_trip['codes'])
                max_consec = calculate_max_consecutive_distance(combined_codes, coord_cache)
                if max_consec > MAX_DISTANCE_IN_TRIP:
                    continue
                
                # คำนวณ util ใหม่
                limits = LIMITS.get(target_vehicle, LIMITS['4W'])
                new_util = max(
                    (new_w / limits['max_w']) * 100,
                    (new_c / limits['max_c']) * 100
                )
                
                # ต้องไม่เกิน 100%
                if new_util > 100:
                    continue
                
                # เลือกทริปที่ใกล้ที่สุดและรวมแล้วได้ util ดี
                min_util_target = MIN_UTIL.get(target_vehicle, 70)
                if new_util >= min_util_target and centroid_dist < best_merge_dist:
                    best_merge_util = new_util
                    best_merge_trip = trip_num
                    best_merge_dist = centroid_dist
            
            # ย้ายสาขาไปทริปใหม่
            if best_merge_trip is not None:
                for code in low_trip['codes']:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = best_merge_trip
                balance_count += 1
    
    # ===============================================
    # 🎯 Phase 6.5: บังคับรวม 4W ที่ไม่ผ่าน MIN_UTIL
    # 1. หา 6W ที่ไม่เต็มและใกล้เคียง
    # 2. หา JB ที่ไม่เต็มและใกล้เคียง
    # 3. รวม 2 คัน 4W เข้าด้วยกันเป็น JB
    # ===============================================
    merge_4w_count = 0
    MERGE_DISTANCE_LIMIT = 80  # ยอมให้รวมได้ไกลขึ้นสำหรับ merge
    
    # หาทริป 4W ทั้งหมด (เรียงตาม util จากน้อยไปมาก)
    trips_4w_to_merge = []
    for trip_num in list(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 0:
            continue
            
        trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
        
        # เฉพาะ 4W เท่านั้น
        if trip_vehicle != '4W':
            continue
        
        trip_w = trip_data['Weight'].sum()
        trip_c = trip_data['Cube'].sum()
        trip_count = len(trip_data)
        trip_codes_list = list(trip_data['Code'].values)
        
        # คำนวณ util ของ 4W
        util_4w = max(
            (trip_w / LIMITS['4W']['max_w']) * 100,
            (trip_c / LIMITS['4W']['max_c']) * 100
        )
        
        # เก็บทุกทริป 4W ที่ไม่ผ่าน MIN_UTIL (70%)
        if util_4w < MIN_UTIL.get('4W', 70):
            # หา centroid
            trip_lats, trip_lons = [], []
            for code in trip_codes_list:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    trip_lats.append(lat)
                    trip_lons.append(lon)
            
            if trip_lats:
                trips_4w_to_merge.append({
                    'trip_num': trip_num,
                    'util': util_4w,
                    'weight': trip_w,
                    'cube': trip_c,
                    'count': trip_count,
                    'codes': trip_codes_list,
                    'lat': sum(trip_lats) / len(trip_lats),
                    'lon': sum(trip_lons) / len(trip_lons)
                })
    
    # เรียงจาก util น้อยที่สุดก่อน (รวมง่ายกว่า)
    trips_4w_to_merge.sort(key=lambda x: x['util'])
    
    # พยายามรวมทีละทริป
    for trip_info in trips_4w_to_merge:
        trip_num = trip_info['trip_num']
        
        # เช็คว่าทริปนี้ยังมีอยู่ไหม (อาจถูกรวมไปแล้ว)
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 0:
            continue
        
        trip_w = trip_info['weight']
        trip_c = trip_info['cube']
        trip_count = trip_info['count']
        trip_codes = trip_info['codes']
        
        # หา centroid ของทริป 4W นี้
        trip_lats, trip_lons = [], []
        for code in trip_codes:
            lat, lon = coord_cache.get(code, (None, None))
            if lat and lon:
                trip_lats.append(lat)
                trip_lons.append(lon)
        
        if not trip_lats:
            continue
        
        trip_centroid_lat = sum(trip_lats) / len(trip_lats)
        trip_centroid_lon = sum(trip_lons) / len(trip_lons)
        
        # หา 6W หรือ JB ที่ใกล้ที่สุดและรับได้
        best_target = None
        best_dist = float('inf')
        
        for target_num in test_df['Trip'].unique():
            if target_num == 0 or target_num == trip_num:
                continue
            
            target_data = test_df[test_df['Trip'] == target_num]
            if len(target_data) == 0:
                continue
                
            target_vehicle = trip_recommended_vehicles.get(target_num, '4W')
            
            # เฉพาะ 6W หรือ JB
            if target_vehicle not in ['6W', 'JB']:
                continue
            
            target_w = target_data['Weight'].sum()
            target_c = target_data['Cube'].sum()
            target_count = len(target_data)
            
            # เช็คว่ารวมแล้วไม่เกิน capacity
            new_w = target_w + trip_w
            new_c = target_c + trip_c
            new_count = target_count + trip_count
            
            limits = LIMITS.get(target_vehicle, LIMITS['6W'])
            new_util = max(
                (new_w / limits['max_w']) * 100,
                (new_c / limits['max_c']) * 100
            )
            
            # ต้องไม่เกิน 100% และจำนวนสาขาไม่เกิน
            max_branches = 20 if target_vehicle == '6W' else 12
            if new_util > 100 or new_count > max_branches:
                continue
            
            # หา centroid ของ target
            target_lats, target_lons = [], []
            for code in target_data['Code'].values:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    target_lats.append(lat)
                    target_lons.append(lon)
            
            if not target_lats:
                continue
            
            target_centroid_lat = sum(target_lats) / len(target_lats)
            target_centroid_lon = sum(target_lons) / len(target_lons)
            
            # คำนวณระยะห่าง centroid
            dist = haversine_distance(trip_centroid_lat, trip_centroid_lon,
                                      target_centroid_lat, target_centroid_lon)
            
            # ระยะต้องไม่เกิน MERGE_DISTANCE_LIMIT (80km สำหรับ merge)
            if dist > MERGE_DISTANCE_LIMIT:
                continue
            
            # 🚨 เช็ค consecutive distance หลังรวม
            combined_codes = list(target_data['Code'].values) + trip_codes
            max_consec = calculate_max_consecutive_distance(combined_codes, coord_cache)
            if max_consec > MAX_DISTANCE_IN_TRIP:
                continue
            
            # เลือก target ที่ใกล้ที่สุด
            if dist < best_dist:
                best_dist = dist
                best_target = target_num
        
        # ย้ายสาขาไป target
        if best_target is not None:
            for code in trip_codes:
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_target
            merge_4w_count += 1
    
    # ===============================================
    # 🎯 Phase 6.6: รวม 4W หลายคันเข้าด้วยกันให้กลายเป็น JB หรือ 6W
    # ===============================================
    merge_4w_to_larger_count = 0
    
    # เรียก trips_4w_to_merge ใหม่ (เพราะอาจมีบางส่วนถูกรวมไปแล้ว)
    trips_4w_remaining = []
    for trip_num in list(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 0:
            continue
            
        trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
        if trip_vehicle != '4W':
            continue
        
        trip_w = trip_data['Weight'].sum()
        trip_c = trip_data['Cube'].sum()
        trip_count = len(trip_data)
        trip_codes_list = list(trip_data['Code'].values)
        
        # หา centroid
        trip_lats, trip_lons = [], []
        for code in trip_codes_list:
            lat, lon = coord_cache.get(code, (None, None))
            if lat and lon:
                trip_lats.append(lat)
                trip_lons.append(lon)
        
        if trip_lats:
            trips_4w_remaining.append({
                'trip_num': trip_num,
                'weight': trip_w,
                'cube': trip_c,
                'count': trip_count,
                'codes': trip_codes_list,
                'lat': sum(trip_lats) / len(trip_lats),
                'lon': sum(trip_lons) / len(trip_lons)
            })
    
    # เรียงตาม cube น้อยที่สุดก่อน
    trips_4w_remaining.sort(key=lambda x: x['cube'])
    
    # พยายามรวม 4W กับ 4W อื่นที่ใกล้กัน
    merged_trips = set()
    for i, trip1 in enumerate(trips_4w_remaining):
        if trip1['trip_num'] in merged_trips:
            continue
        
        # หา 4W อื่นที่ใกล้ที่สุดและรวมได้
        for j, trip2 in enumerate(trips_4w_remaining):
            if i >= j or trip2['trip_num'] in merged_trips:
                continue
            
            # คำนวณระยะห่าง centroid
            dist = haversine_distance(trip1['lat'], trip1['lon'], trip2['lat'], trip2['lon'])
            if dist > MERGE_DISTANCE_LIMIT:  # ใช้ 80km สำหรับ merge
                continue
            
            # เช็คว่ารวมแล้วใส่ JB ได้ไหม
            combined_w = trip1['weight'] + trip2['weight']
            combined_c = trip1['cube'] + trip2['cube']
            combined_count = trip1['count'] + trip2['count']
            
            # ลองใส่ JB (7 cube, 3500kg)
            if combined_c <= LIMITS['JB']['max_c'] and combined_w <= LIMITS['JB']['max_w'] and combined_count <= 12:
                # เช็คว่าทุกสาขารับ JB ได้ไหม
                all_codes = trip1['codes'] + trip2['codes']
                can_use_jb = True
                for code in all_codes:
                    branch_max = get_max_vehicle_for_branch(code)
                    if branch_max == '4W':  # สาขานี้รับ JB ไม่ได้
                        can_use_jb = False
                        break
                
                if not can_use_jb:
                    continue
                
                # 🚨 เช็ค consecutive distance หลังรวม
                max_consec = calculate_max_consecutive_distance(all_codes, coord_cache)
                if max_consec > MAX_DISTANCE_IN_TRIP:
                    continue
                
                # รวมได้! ย้ายทุกสาขาจาก trip2 ไป trip1
                for code in trip2['codes']:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = trip1['trip_num']
                
                # อัปเดต trip_recommended_vehicles เป็น JB
                trip_recommended_vehicles[trip1['trip_num']] = 'JB'
                merged_trips.add(trip2['trip_num'])
                merge_4w_to_larger_count += 1
                
                # อัปเดต trip1 info
                trip1['weight'] = combined_w
                trip1['cube'] = combined_c
                trip1['count'] = combined_count
                trip1['codes'].extend(trip2['codes'])
    
    # ===============================================
    # 🎯 Phase 6.7: รวม 4W ที่เหลือเข้ากับ JB ที่ไม่เต็ม
    # ===============================================
    # หา 4W ที่ยังเหลืออยู่
    for trip_num in list(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 0:
            continue
        
        trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
        if trip_vehicle != '4W':
            continue
        
        trip_w = trip_data['Weight'].sum()
        trip_c = trip_data['Cube'].sum()
        trip_count = len(trip_data)
        trip_codes_list = list(trip_data['Code'].values)
        
        # คำนวณ util ของ 4W
        util_4w = max(
            (trip_w / LIMITS['4W']['max_w']) * 100,
            (trip_c / LIMITS['4W']['max_c']) * 100
        )
        
        # ถ้าผ่าน MIN_UTIL แล้ว → ข้าม
        if util_4w >= MIN_UTIL.get('4W', 70):
            continue
        
        # หา centroid
        trip_lats, trip_lons = [], []
        for code in trip_codes_list:
            lat, lon = coord_cache.get(code, (None, None))
            if lat and lon:
                trip_lats.append(lat)
                trip_lons.append(lon)
        
        if not trip_lats:
            continue
        
        trip_lat = sum(trip_lats) / len(trip_lats)
        trip_lon = sum(trip_lons) / len(trip_lons)
        
        # หา JB ที่ใกล้ที่สุดและรับได้
        best_jb = None
        best_dist = float('inf')
        
        for target_num in test_df['Trip'].unique():
            if target_num == 0 or target_num == trip_num:
                continue
            
            target_data = test_df[test_df['Trip'] == target_num]
            if len(target_data) == 0:
                continue
            
            target_vehicle = trip_recommended_vehicles.get(target_num, '4W')
            if target_vehicle != 'JB':
                continue
            
            target_w = target_data['Weight'].sum()
            target_c = target_data['Cube'].sum()
            target_count = len(target_data)
            
            # เช็ค capacity
            new_w = target_w + trip_w
            new_c = target_c + trip_c
            new_count = target_count + trip_count
            
            if new_c > LIMITS['JB']['max_c'] or new_w > LIMITS['JB']['max_w'] or new_count > 12:
                continue
            
            # หา centroid ของ JB
            target_lats, target_lons = [], []
            for code in target_data['Code'].values:
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    target_lats.append(lat)
                    target_lons.append(lon)
            
            if not target_lats:
                continue
            
            target_lat = sum(target_lats) / len(target_lats)
            target_lon = sum(target_lons) / len(target_lons)
            
            dist = haversine_distance(trip_lat, trip_lon, target_lat, target_lon)
            if dist > MERGE_DISTANCE_LIMIT:
                continue
            
            # 🚨 เช็ค consecutive distance หลังรวม
            combined_codes = list(target_data['Code'].values) + trip_codes_list
            max_consec = calculate_max_consecutive_distance(combined_codes, coord_cache)
            if max_consec > MAX_DISTANCE_IN_TRIP:
                continue
            
            if dist < best_dist:
                best_dist = dist
                best_jb = target_num
        
        # ย้ายสาขาไป JB
        if best_jb is not None:
            for code in trip_codes_list:
                test_df.loc[test_df['Code'] == code, 'Trip'] = best_jb
            merge_4w_count += 1
    
    # ===============================================
    # 🎯 Phase 6.8: Force Upgrade Low Util 4W to 6W
    # ถ้า 4W ไม่ผ่าน MIN_UTIL และรวมไม่ได้ → เปลี่ยนเป็น 6W แล้วดึงสาขาใกล้มาเพิ่ม
    # ===============================================
    force_upgrade_count = 0
    
    # หา 4W ที่ยังไม่ผ่าน MIN_UTIL
    for trip_num in list(test_df['Trip'].unique()):
        if trip_num == 0:
            continue
        
        trip_data = test_df[test_df['Trip'] == trip_num]
        if len(trip_data) == 0:
            continue
        
        trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
        if trip_vehicle != '4W':
            continue
        
        trip_w = trip_data['Weight'].sum()
        trip_c = trip_data['Cube'].sum()
        trip_codes_list = list(trip_data['Code'].values)
        
        # คำนวณ util ของ 4W
        util_4w = max(
            (trip_w / LIMITS['4W']['max_w']) * 100,
            (trip_c / LIMITS['4W']['max_c']) * 100
        )
        
        # ถ้าผ่าน MIN_UTIL แล้ว → ข้าม
        if util_4w >= MIN_UTIL.get('4W', 70):
            continue
        
        # เช็คว่าสาขาทั้งหมดรับ 6W ได้ไหม
        can_use_6w = True
        for code in trip_codes_list:
            branch_max = get_max_vehicle_for_branch(code)
            if branch_max in ['4W', 'JB']:  # ไม่รับ 6W
                can_use_6w = False
                break
        
        if not can_use_6w:
            continue
        
        # หา centroid ของทริปนี้
        trip_lats, trip_lons = [], []
        for code in trip_codes_list:
            lat, lon = coord_cache.get(code, (None, None))
            if lat and lon:
                trip_lats.append(lat)
                trip_lons.append(lon)
        
        if not trip_lats:
            continue
        
        trip_lat = sum(trip_lats) / len(trip_lats)
        trip_lon = sum(trip_lons) / len(trip_lons)
        
        # เปลี่ยนเป็น 6W
        trip_recommended_vehicles[trip_num] = '6W'
        
        # ดึงสาขาจากทริปอื่นที่ใกล้เคียงมาเพิ่ม จนกว่าจะผ่าน MIN_UTIL 6W (90%)
        current_w = trip_w
        current_c = trip_c
        current_codes = trip_codes_list.copy()
        
        # หาสาขาจากทริปอื่นที่อยู่ใกล้
        nearby_branches = []
        for other_trip in test_df['Trip'].unique():
            if other_trip == 0 or other_trip == trip_num:
                continue
            
            other_data = test_df[test_df['Trip'] == other_trip]
            for _, row in other_data.iterrows():
                code = row['Code']
                lat, lon = coord_cache.get(code, (None, None))
                if lat and lon:
                    # เช็คว่าสาขานี้รับ 6W ได้ไหม
                    branch_max = get_max_vehicle_for_branch(code)
                    if branch_max not in ['6W']:
                        continue
                    
                    dist_to_centroid = haversine_distance(lat, lon, trip_lat, trip_lon)
                    if dist_to_centroid <= MAX_DISTANCE_IN_TRIP:  # ใกล้พอที่จะรวมได้
                        nearby_branches.append({
                            'code': code,
                            'weight': row['Weight'],
                            'cube': row['Cube'],
                            'dist': dist_to_centroid,
                            'from_trip': other_trip,
                            'lat': lat,
                            'lon': lon
                        })
        
        # เรียงตามระยะใกล้สุดก่อน
        nearby_branches.sort(key=lambda x: x['dist'])
        
        # ดึงสาขามาเพิ่มจนผ่าน MIN_UTIL 6W (90%)
        for branch in nearby_branches:
            # คำนวณ util ปัจจุบัน
            util_6w = max(
                (current_w / LIMITS['6W']['max_w']) * 100,
                (current_c / LIMITS['6W']['max_c']) * 100
            )
            
            # ถ้าผ่านแล้ว → หยุด
            if util_6w >= MIN_UTIL.get('6W', 90):
                break
            
            # เช็ค capacity ว่าใส่ได้ไหม
            new_w = current_w + branch['weight']
            new_c = current_c + branch['cube']
            
            if new_w > LIMITS['6W']['max_w'] or new_c > LIMITS['6W']['max_c']:
                continue
            
            # เช็ค consecutive distance หลังเพิ่ม
            test_codes = current_codes + [branch['code']]
            max_consec = calculate_max_consecutive_distance(test_codes, coord_cache)
            if max_consec > MAX_DISTANCE_IN_TRIP:
                continue
            
            # ย้ายสาขามา
            test_df.loc[test_df['Code'] == branch['code'], 'Trip'] = trip_num
            current_w = new_w
            current_c = new_c
            current_codes.append(branch['code'])
        
        force_upgrade_count += 1
    
    # ===============================================
    # 🎯 Phase 6.9: FORCE Merge Low Util Trips
    # ทริป 4W ที่ไม่ผ่าน MIN_UTIL จะถูกบังคับรวมเข้ากับทริปอื่น
    # ถ้ารวมไม่ได้เลย → กระจายสาขาไปทริปอื่น
    # ===============================================
    force_merge_count = 0
    max_iterations = 10  # ป้องกัน infinite loop
    
    for iteration in range(max_iterations):
        # หา 4W ที่ไม่ผ่าน MIN_UTIL
        low_util_4w = []
        for trip_num in list(test_df['Trip'].unique()):
            if trip_num == 0:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            if len(trip_data) == 0:
                continue
            
            trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
            if trip_vehicle != '4W':
                continue
            
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            trip_codes_list = list(trip_data['Code'].values)
            
            util_4w = max(
                (trip_w / LIMITS['4W']['max_w']) * 100,
                (trip_c / LIMITS['4W']['max_c']) * 100
            )
            
            if util_4w < MIN_UTIL.get('4W', 70):
                # หา centroid
                trip_lats, trip_lons = [], []
                for code in trip_codes_list:
                    lat, lon = coord_cache.get(code, (None, None))
                    if lat and lon:
                        trip_lats.append(lat)
                        trip_lons.append(lon)
                
                if trip_lats:
                    low_util_4w.append({
                        'trip_num': trip_num,
                        'util': util_4w,
                        'weight': trip_w,
                        'cube': trip_c,
                        'codes': trip_codes_list,
                        'lat': sum(trip_lats) / len(trip_lats),
                        'lon': sum(trip_lons) / len(trip_lons)
                    })
        
        if not low_util_4w:
            break  # ไม่มี 4W ที่ไม่ผ่านแล้ว
        
        # เรียงจาก util น้อยสุด (รวมง่ายกว่า)
        low_util_4w.sort(key=lambda x: x['util'])
        
        merged_any = False
        for trip_info in low_util_4w:
            trip_num = trip_info['trip_num']
            
            # เช็คว่าทริปยังมีอยู่ไหม
            trip_data = test_df[test_df['Trip'] == trip_num]
            if len(trip_data) == 0:
                continue
            
            trip_codes = trip_info['codes']
            trip_lat = trip_info['lat']
            trip_lon = trip_info['lon']
            
            # หาทริปที่รวมได้ (ไม่จำกัดประเภทรถ)
            best_target = None
            best_score = float('inf')
            
            for target_num in test_df['Trip'].unique():
                if target_num == 0 or target_num == trip_num:
                    continue
                
                target_data = test_df[test_df['Trip'] == target_num]
                if len(target_data) == 0:
                    continue
                
                target_vehicle = trip_recommended_vehicles.get(target_num, '4W')
                target_w = target_data['Weight'].sum()
                target_c = target_data['Cube'].sum()
                target_count = len(target_data)
                target_codes = list(target_data['Code'].values)
                
                # คำนวณ combined values
                new_w = target_w + trip_info['weight']
                new_c = target_c + trip_info['cube']
                new_count = target_count + len(trip_codes)
                
                # เลือกรถที่เหมาะสมสำหรับ combined
                combined_codes = target_codes + trip_codes
                
                # หา max vehicle ที่ทุกสาขารับได้
                min_max_vehicle = '6W'
                for code in combined_codes:
                    branch_max = get_max_vehicle_for_branch(code)
                    if branch_max == '4W':
                        min_max_vehicle = '4W'
                        break
                    elif branch_max == 'JB' and min_max_vehicle == '6W':
                        min_max_vehicle = 'JB'
                
                # เลือกรถที่เหมาะสม
                if new_c <= LIMITS['4W']['max_c'] and new_w <= LIMITS['4W']['max_w'] and min_max_vehicle == '4W':
                    new_vehicle = '4W'
                    new_util = max((new_w / LIMITS['4W']['max_w']) * 100, (new_c / LIMITS['4W']['max_c']) * 100)
                    max_branches = 12
                elif new_c <= LIMITS['JB']['max_c'] and new_w <= LIMITS['JB']['max_w'] and min_max_vehicle in ['JB', '6W']:
                    new_vehicle = 'JB'
                    new_util = max((new_w / LIMITS['JB']['max_w']) * 100, (new_c / LIMITS['JB']['max_c']) * 100)
                    max_branches = 12
                elif new_c <= LIMITS['6W']['max_c'] and new_w <= LIMITS['6W']['max_w'] and min_max_vehicle == '6W':
                    new_vehicle = '6W'
                    new_util = max((new_w / LIMITS['6W']['max_w']) * 100, (new_c / LIMITS['6W']['max_c']) * 100)
                    max_branches = 25
                else:
                    continue  # เกิน capacity
                
                # เช็คจำนวนสาขา
                if new_count > max_branches:
                    continue
                
                # เช็ค consecutive distance
                max_consec = calculate_max_consecutive_distance(combined_codes, coord_cache)
                if max_consec > MAX_DISTANCE_IN_TRIP:
                    continue
                
                # หา centroid ของ target
                target_lats = [coord_cache.get(c, (None, None))[0] for c in target_codes if coord_cache.get(c, (None, None))[0]]
                target_lons = [coord_cache.get(c, (None, None))[1] for c in target_codes if coord_cache.get(c, (None, None))[1]]
                
                if not target_lats:
                    continue
                
                target_lat = sum(target_lats) / len(target_lats)
                target_lon = sum(target_lons) / len(target_lons)
                
                centroid_dist = haversine_distance(trip_lat, trip_lon, target_lat, target_lon)
                
                # Score: ระยะใกล้ + util สูง = ดี
                score = centroid_dist - (new_util * 0.5)  # ยิ่ง util สูง ยิ่งดี
                
                if score < best_score:
                    best_score = score
                    best_target = target_num
            
            # รวมเข้า target
            if best_target is not None:
                target_vehicle = trip_recommended_vehicles.get(best_target, '4W')
                target_data = test_df[test_df['Trip'] == best_target]
                target_codes = list(target_data['Code'].values)
                combined_codes = target_codes + trip_codes
                
                # คำนวณรถใหม่
                new_w = target_data['Weight'].sum() + trip_info['weight']
                new_c = target_data['Cube'].sum() + trip_info['cube']
                
                # หา max vehicle ที่ทุกสาขารับได้
                min_max_vehicle = '6W'
                for code in combined_codes:
                    branch_max = get_max_vehicle_for_branch(code)
                    if branch_max == '4W':
                        min_max_vehicle = '4W'
                        break
                    elif branch_max == 'JB' and min_max_vehicle == '6W':
                        min_max_vehicle = 'JB'
                
                # เลือกรถที่เหมาะสม
                if new_c <= LIMITS['4W']['max_c'] and new_w <= LIMITS['4W']['max_w'] and min_max_vehicle == '4W':
                    new_vehicle = '4W'
                elif new_c <= LIMITS['JB']['max_c'] and new_w <= LIMITS['JB']['max_w'] and min_max_vehicle in ['JB', '6W']:
                    new_vehicle = 'JB'
                else:
                    new_vehicle = '6W'
                
                # ย้ายสาขา
                for code in trip_codes:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = best_target
                
                trip_recommended_vehicles[best_target] = new_vehicle
                force_merge_count += 1
                merged_any = True
        
        if not merged_any:
            break  # ไม่มี merge เพิ่มแล้ว
    
    # ===============================================
    # 🎯 Phase 6.10: Force Merge Low Util JB
    # JB ที่ไม่ผ่าน MIN_UTIL (80%) → รวมเข้ากับ 6W หรือ JB อื่น
    # ===============================================
    for iteration in range(5):
        low_util_jb = []
        for trip_num in list(test_df['Trip'].unique()):
            if trip_num == 0:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            if len(trip_data) == 0:
                continue
            
            trip_vehicle = trip_recommended_vehicles.get(trip_num, '4W')
            if trip_vehicle != 'JB':
                continue
            
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            trip_codes_list = list(trip_data['Code'].values)
            
            util_jb = max(
                (trip_w / LIMITS['JB']['max_w']) * 100,
                (trip_c / LIMITS['JB']['max_c']) * 100
            )
            
            if util_jb < MIN_UTIL.get('JB', 80):
                trip_lats, trip_lons = [], []
                for code in trip_codes_list:
                    lat, lon = coord_cache.get(code, (None, None))
                    if lat and lon:
                        trip_lats.append(lat)
                        trip_lons.append(lon)
                
                if trip_lats:
                    low_util_jb.append({
                        'trip_num': trip_num,
                        'util': util_jb,
                        'weight': trip_w,
                        'cube': trip_c,
                        'codes': trip_codes_list,
                        'lat': sum(trip_lats) / len(trip_lats),
                        'lon': sum(trip_lons) / len(trip_lons)
                    })
        
        if not low_util_jb:
            break
        
        low_util_jb.sort(key=lambda x: x['util'])
        merged_any = False
        
        for trip_info in low_util_jb:
            trip_num = trip_info['trip_num']
            trip_data = test_df[test_df['Trip'] == trip_num]
            if len(trip_data) == 0:
                continue
            
            trip_codes = trip_info['codes']
            
            # หา 6W หรือ JB ที่รวมได้
            best_target = None
            best_dist = float('inf')
            
            for target_num in test_df['Trip'].unique():
                if target_num == 0 or target_num == trip_num:
                    continue
                
                target_data = test_df[test_df['Trip'] == target_num]
                if len(target_data) == 0:
                    continue
                
                target_vehicle = trip_recommended_vehicles.get(target_num, '4W')
                if target_vehicle not in ['6W', 'JB']:
                    continue
                
                target_codes = list(target_data['Code'].values)
                combined_codes = target_codes + trip_codes
                
                # เช็ค capacity
                new_w = target_data['Weight'].sum() + trip_info['weight']
                new_c = target_data['Cube'].sum() + trip_info['cube']
                new_count = len(combined_codes)
                
                # ลองใส่ 6W
                if target_vehicle == '6W':
                    if new_c > LIMITS['6W']['max_c'] or new_w > LIMITS['6W']['max_w'] or new_count > 25:
                        continue
                else:  # JB
                    if new_c > LIMITS['JB']['max_c'] or new_w > LIMITS['JB']['max_w'] or new_count > 12:
                        # ลองเปลี่ยนเป็น 6W
                        if new_c <= LIMITS['6W']['max_c'] and new_w <= LIMITS['6W']['max_w'] and new_count <= 25:
                            # เช็คว่าทุกสาขารับ 6W ได้
                            can_6w = True
                            for code in combined_codes:
                                if get_max_vehicle_for_branch(code) != '6W':
                                    can_6w = False
                                    break
                            if not can_6w:
                                continue
                        else:
                            continue
                
                # เช็ค consecutive distance
                max_consec = calculate_max_consecutive_distance(combined_codes, coord_cache)
                if max_consec > MAX_DISTANCE_IN_TRIP:
                    continue
                
                # คำนวณระยะ centroid
                target_lats = [coord_cache.get(c, (None, None))[0] for c in target_codes if coord_cache.get(c, (None, None))[0]]
                target_lons = [coord_cache.get(c, (None, None))[1] for c in target_codes if coord_cache.get(c, (None, None))[1]]
                if not target_lats:
                    continue
                
                dist = haversine_distance(trip_info['lat'], trip_info['lon'], 
                                          sum(target_lats)/len(target_lats), sum(target_lons)/len(target_lons))
                
                if dist < best_dist:
                    best_dist = dist
                    best_target = target_num
            
            if best_target is not None:
                target_vehicle = trip_recommended_vehicles.get(best_target, '4W')
                target_data = test_df[test_df['Trip'] == best_target]
                new_w = target_data['Weight'].sum() + trip_info['weight']
                new_c = target_data['Cube'].sum() + trip_info['cube']
                
                # เลือกรถ
                if target_vehicle == '6W' or new_c > LIMITS['JB']['max_c'] or new_w > LIMITS['JB']['max_w']:
                    trip_recommended_vehicles[best_target] = '6W'
                
                for code in trip_codes:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = best_target
                
                force_merge_count += 1
                merged_any = True
        
        if not merged_any:
            break
    
    # ===============================================
    # 🎯 Phase 7: Final Validation & Auto-Fix
    # 1. แยกทริปที่เกิน 100% ออกเป็น 2 คัน
    # 2. รวมทริปที่ต่ำกว่า MIN_UTIL เข้ากับทริปอื่น
    # ===============================================
    
    # 7.1 แยกทริปที่เกิน 100%
    for iteration in range(5):
        over_capacity_trips = []
        
        for trip_num in list(test_df['Trip'].unique()):
            if trip_num == 0:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            if len(trip_data) == 0:
                continue
            
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            
            # หารถที่ใหญ่ที่สุดที่ใช้ได้ (ตามข้อจำกัดสาขา)
            trip_codes = list(trip_data['Code'].values)
            max_allowed = get_max_vehicle_for_trip(trip_codes)
            vehicle = trip_recommended_vehicles.get(trip_num, max_allowed)
            
            # ถ้ารถที่แนะนำใหญ่กว่าที่อนุญาต ให้ใช้รถที่อนุญาต
            vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
            if vehicle_sizes.get(vehicle, 3) > vehicle_sizes.get(max_allowed, 3):
                vehicle = max_allowed
                trip_recommended_vehicles[trip_num] = max_allowed
            
            limits = LIMITS.get(vehicle, LIMITS['6W'])
            
            util = max((trip_w / limits['max_w']) * 100, (trip_c / limits['max_c']) * 100)
            
            # 🔒 ถ้าเกิน 100% และห้ามใช้รถใหญ่กว่า → ต้องแยกทริป
            should_split = False
            if util > 100:
                should_split = True
            
            # 🔒 ถ้าใช้ JB เกิน แต่ห้าม 6W → ต้องแยกทริป
            if vehicle == 'JB' and max_allowed in ['JB', '4W']:
                jb_limits = LIMITS['JB']
                jb_util = max((trip_w / jb_limits['max_w']) * 100, (trip_c / jb_limits['max_c']) * 100)
                if jb_util > 100:
                    should_split = True
            
            if should_split:
                over_capacity_trips.append({
                    'trip_num': trip_num,
                    'util': util,
                    'weight': trip_w,
                    'cube': trip_c,
                    'codes': list(trip_data['Code'].values),
                    'vehicle': vehicle
                })
        
        if not over_capacity_trips:
            break
        
        for trip_info in over_capacity_trips:
            trip_num = trip_info['trip_num']
            trip_codes = trip_info['codes']
            
            if len(trip_codes) < 2:
                continue  # ไม่สามารถแยกได้
            
            # แยกครึ่ง
            mid = len(trip_codes) // 2
            codes_stay = trip_codes[:mid]
            codes_move = trip_codes[mid:]
            
            # สร้างทริปใหม่
            new_trip_num = max(test_df['Trip'].max(), 0) + 1
            
            for code in codes_move:
                test_df.loc[test_df['Code'] == code, 'Trip'] = new_trip_num
            
            # กำหนดรถสำหรับทริปใหม่
            trip_recommended_vehicles[new_trip_num] = trip_info['vehicle']
    
    # 7.2 รวมทริปที่ต่ำกว่า MIN_UTIL หรือเปลี่ยนรถให้เหมาะสม
    for iteration in range(15):
        low_util_trips = []
        
        for trip_num in list(test_df['Trip'].unique()):
            if trip_num == 0:
                continue
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            if len(trip_data) == 0:
                continue
            
            trip_w = trip_data['Weight'].sum()
            trip_c = trip_data['Cube'].sum()
            trip_codes = list(trip_data['Code'].values)
            
            vehicle = trip_recommended_vehicles.get(trip_num, '4W')
            limits = LIMITS.get(vehicle, LIMITS['4W'])
            min_util_required = MIN_UTIL.get(vehicle, 70)
            
            util = max((trip_w / limits['max_w']) * 100, (trip_c / limits['max_c']) * 100)
            
            if util < min_util_required:
                # หา centroid
                trip_lats, trip_lons = [], []
                for code in trip_codes:
                    lat, lon = coord_cache.get(code, (None, None))
                    if lat and lon:
                        trip_lats.append(lat)
                        trip_lons.append(lon)
                
                if trip_lats:
                    low_util_trips.append({
                        'trip_num': trip_num,
                        'util': util,
                        'weight': trip_w,
                        'cube': trip_c,
                        'codes': trip_codes,
                        'vehicle': vehicle,
                        'lat': sum(trip_lats) / len(trip_lats),
                        'lon': sum(trip_lons) / len(trip_lons)
                    })
        
        if not low_util_trips:
            break
        
        # เรียงจาก util น้อยสุด
        low_util_trips.sort(key=lambda x: x['util'])
        
        fixed_any = False
        for trip_info in low_util_trips:
            trip_num = trip_info['trip_num']
            
            trip_data = test_df[test_df['Trip'] == trip_num]
            if len(trip_data) == 0:
                continue
            
            trip_codes = trip_info['codes']
            trip_w = trip_info['weight']
            trip_c = trip_info['cube']
            current_vehicle = trip_info['vehicle']
            
            # 🔧 วิธี 1: ลองเปลี่ยนเป็นรถเล็กลง (util จะสูงขึ้น)
            vehicle_order = ['4W', 'JB', '6W']
            current_idx = vehicle_order.index(current_vehicle) if current_vehicle in vehicle_order else 2
            
            changed_vehicle = False
            for smaller_idx in range(current_idx - 1, -1, -1):
                smaller_vehicle = vehicle_order[smaller_idx]
                smaller_limits = LIMITS[smaller_vehicle]
                
                # เช็คว่าใส่รถเล็กได้หรือไม่ (ห้ามเกิน 100%)
                w_util = (trip_w / smaller_limits['max_w']) * 100
                c_util = (trip_c / smaller_limits['max_c']) * 100
                new_util = max(w_util, c_util)
                
                # ต้องไม่เกิน 100% และผ่าน MIN_UTIL ของรถเล็ก
                if new_util <= 100 and new_util >= MIN_UTIL.get(smaller_vehicle, 70):
                    # เช็คจำนวนสาขา
                    if smaller_vehicle in ['4W', 'JB'] and len(trip_codes) > 12:
                        continue
                    
                    trip_recommended_vehicles[trip_num] = smaller_vehicle
                    changed_vehicle = True
                    fixed_any = True
                    break
            
            if changed_vehicle:
                continue
            
            # 🔧 วิธี 2: ลองเปลี่ยนเป็นรถใหญ่ขึ้น (ต้องผ่าน MIN_UTIL ของรถใหญ่)
            for larger_idx in range(current_idx + 1, len(vehicle_order)):
                larger_vehicle = vehicle_order[larger_idx]
                larger_limits = LIMITS[larger_vehicle]
                
                w_util = (trip_w / larger_limits['max_w']) * 100
                c_util = (trip_c / larger_limits['max_c']) * 100
                new_util = max(w_util, c_util)
                
                # ต้องผ่าน MIN_UTIL ของรถใหญ่
                if new_util >= MIN_UTIL.get(larger_vehicle, 70) and new_util <= 100:
                    trip_recommended_vehicles[trip_num] = larger_vehicle
                    changed_vehicle = True
                    fixed_any = True
                    break
            
            if changed_vehicle:
                continue
            
            # 🔧 วิธี 3: รวมกับทริปอื่นที่ใกล้เคียง
            best_target = None
            best_new_util = 0
            best_dist = float('inf')
            
            for target_num in test_df['Trip'].unique():
                if target_num == 0 or target_num == trip_num:
                    continue
                
                target_data = test_df[test_df['Trip'] == target_num]
                if len(target_data) == 0:
                    continue
                
                target_vehicle = trip_recommended_vehicles.get(target_num, '4W')
                target_codes = list(target_data['Code'].values)
                combined_codes = target_codes + trip_codes
                
                # คำนวณ capacity รวม
                new_w = target_data['Weight'].sum() + trip_w
                new_c = target_data['Cube'].sum() + trip_c
                new_count = len(combined_codes)
                
                # เลือกรถที่เหมาะสม (เล็กสุดที่ใส่ได้)
                new_vehicle = None
                for v in ['4W', 'JB', '6W']:
                    v_limits = LIMITS[v]
                    if new_w <= v_limits['max_w'] and new_c <= v_limits['max_c']:
                        if v in ['4W', 'JB'] and new_count > 12:
                            continue
                        new_vehicle = v
                        break
                
                if not new_vehicle:
                    continue  # เกิน capacity
                
                new_limits = LIMITS[new_vehicle]
                new_util = max((new_w / new_limits['max_w']) * 100, (new_c / new_limits['max_c']) * 100)
                
                # ต้องไม่เกิน 100% และผ่าน MIN_UTIL
                if new_util > 100:
                    continue
                
                new_min_util = MIN_UTIL.get(new_vehicle, 70)
                if new_util < new_min_util:
                    continue
                
                # เช็ค consecutive distance
                max_consec = calculate_max_consecutive_distance(combined_codes, coord_cache)
                if max_consec > MAX_DISTANCE_IN_TRIP:
                    continue
                
                # คำนวณระยะ centroid
                target_lats = [coord_cache.get(c, (None, None))[0] for c in target_codes if coord_cache.get(c, (None, None))[0]]
                target_lons = [coord_cache.get(c, (None, None))[1] for c in target_codes if coord_cache.get(c, (None, None))[1]]
                if not target_lats:
                    continue
                
                dist = haversine_distance(trip_info['lat'], trip_info['lon'],
                                          sum(target_lats)/len(target_lats), sum(target_lons)/len(target_lons))
                
                # เลือกทริปที่รวมแล้วได้ util สูงสุดและใกล้ที่สุด
                if new_util > best_new_util or (new_util == best_new_util and dist < best_dist):
                    best_new_util = new_util
                    best_target = target_num
                    best_dist = dist
                    best_new_vehicle = new_vehicle
            
            if best_target is not None:
                # รวมเข้า target
                trip_recommended_vehicles[best_target] = best_new_vehicle
                
                for code in trip_codes:
                    test_df.loc[test_df['Code'] == code, 'Trip'] = best_target
                
                fixed_any = True
        
        if not fixed_any:
            break
    
    # สรุปผลและแนะนำรถ
    summary_data = []
    for trip_num in sorted(test_df['Trip'].unique()):
        trip_data = test_df[test_df['Trip'] == trip_num]
        total_w = trip_data['Weight'].sum()
        total_c = trip_data['Cube'].sum()
        
        # หารถที่ใหญ่สุดที่ทุกสาขาในทริปสามารถใช้ได้ (ใช้ get_max_vehicle_for_branch โดยตรง)
        trip_codes = trip_data['Code'].unique()
        max_vehicles = []
        for c in trip_codes:
            # ใช้ฟังก์ชันหลักที่รวม Booking + Punthai แล้ว
            branch_max = get_max_vehicle_for_branch(c)
            max_vehicles.append(branch_max)
        
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        
        # ตรวจสอบว่า max_vehicles ไม่ว่าง
        if max_vehicles:
            min_max_size = min(vehicle_sizes.get(v, 3) for v in max_vehicles)
        else:
            min_max_size = 3  # default ถ้าไม่มีข้อมูล ให้ใช้ 6W ได้
        
        max_allowed_vehicle = {1: '4W', 2: 'JB', 3: '6W'}.get(min_max_size, '6W')
        
        # เลือกรถ: เช็คข้อจำกัดสาขาก่อน แล้วค่อยใช้ประวัติ/AI
        if trip_num in trip_recommended_vehicles:
            # มีประวัติ
            suggested_from_history = trip_recommended_vehicles[trip_num]
            
            # เช็คว่ารถจากประวัติขัดกับข้อจำกัดสาขาหรือไม่
            if vehicle_sizes.get(suggested_from_history, 0) > min_max_size:
                # ถ้าขัด - ต้องลดลงตามข้อจำกัดสาขา
                suggested = max_allowed_vehicle
                source = f"📜 ประวัติ → {max_allowed_vehicle} (จำกัดสาขา)"
            else:
                # ไม่ขัด - ใช้ตามประวัติ
                suggested = suggested_from_history
                source = "📜 ประวัติ"
        else:
            # ไม่มีประวัติ - ใช้ AI พร้อมเคารพข้อจำกัดสาขา
            suggested = suggest_truck(total_w, total_c, max_allowed_vehicle, trip_codes)
            if min_max_size < 3:
                source = f"🤖 AI (จำกัด {max_allowed_vehicle})"
            else:
                source = "🤖 AI"
        
        # 🔒 เช็คว่าเป็นพื้นที่ nearby (กรุงเทพ+ปริมณฑล) หรือไม่ → ห้าม 6W เด็ดขาด!
        provinces = set()
        for code in trip_codes:
            prov = get_province(code)
            if prov and prov != 'UNKNOWN':
                provinces.add(prov)
        is_nearby_trip = all(get_region_type(p) == 'nearby' for p in provinces) if provinces else False
        
        # ตรวจสอบว่ารถที่เลือกใส่ของได้จริงหรือไม่ (ห้ามเกิน 100%)
        if suggested in LIMITS:
            w_util = (total_w / LIMITS[suggested]['max_w']) * 100
            c_util = (total_c / LIMITS[suggested]['max_c']) * 100
            max_util = max(w_util, c_util)
            
            # ถ้าเกิน 100% ต้องเพิ่มขนาดรถ
            if max_util > 100:
                if suggested == '4W' and 'JB' in LIMITS:
                    # ลองเปลี่ยนเป็น JB
                    jb_w_util = (total_w / LIMITS['JB']['max_w']) * 100
                    jb_c_util = (total_c / LIMITS['JB']['max_c']) * 100
                    if max(jb_w_util, jb_c_util) <= 100:
                        suggested = 'JB'
                        source = source + " → JB"
                        w_util, c_util = jb_w_util, jb_c_util
                    else:
                        # 🚫 ห้ามใช้ 6W ในพื้นที่ nearby!
                        if is_nearby_trip:
                            # ต้องแยกทริป (จะแยกใน Phase 2.5) - ยังคงใช้ JB ไว้ก่อน
                            suggested = 'JB'
                            source = source + " → JB (ต้องแยกทริป)"
                            w_util, c_util = jb_w_util, jb_c_util
                        else:
                            suggested = '6W'
                            source = source + " → 6W"
                            w_util = (total_w / LIMITS['6W']['max_w']) * 100
                            c_util = (total_c / LIMITS['6W']['max_c']) * 100
                elif suggested == 'JB' or suggested == '4W':
                    # 🚫 ห้ามใช้ 6W ในพื้นที่ nearby!
                    if is_nearby_trip:
                        # ต้องแยกทริป (จะแยกใน Phase 2.5) - ยังคงใช้ JB ไว้ก่อน
                        suggested = 'JB'
                        source = source + " → JB (ต้องแยกทริป)"
                        jb_w_util = (total_w / LIMITS['JB']['max_w']) * 100
                        jb_c_util = (total_c / LIMITS['JB']['max_c']) * 100
                        w_util, c_util = jb_w_util, jb_c_util
                    else:
                        suggested = '6W'
                        source = source + " → 6W"
                        w_util = (total_w / LIMITS['6W']['max_w']) * 100
                        c_util = (total_c / LIMITS['6W']['max_c']) * 100
        else:
            w_util = c_util = 0
        
        # คำนวณระยะทางรวมของทริป (เส้นทาง: DC → สาขา1 → สาขา2 → ... → DC)
        total_distance = 0
        if trip_codes is not None and len(trip_codes) > 0:
            # ดึงพิกัดของแต่ละสาขาจาก Master
            branch_coords = []
            for code in trip_codes:
                if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                    master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                    if len(master_row) > 0:
                        lat = master_row.iloc[0].get('ละติจูด', 0)
                        lon = master_row.iloc[0].get('ลองติจูด', 0)
                        if lat != 0 and lon != 0:
                            branch_coords.append((lat, lon))
            
            # คำนวณระยะทางตามเส้นทาง
            if len(branch_coords) > 0:
                # DC → สาขาแรก
                total_distance += calculate_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, 
                                                    branch_coords[0][0], branch_coords[0][1])
                # สาขา → สาขา
                for i in range(len(branch_coords) - 1):
                    total_distance += calculate_distance(branch_coords[i][0], branch_coords[i][1],
                                                        branch_coords[i+1][0], branch_coords[i+1][1])
                # สาขาสุดท้าย → DC
                total_distance += calculate_distance(branch_coords[-1][0], branch_coords[-1][1],
                                                    DC_WANG_NOI_LAT, DC_WANG_NOI_LON)
        
        # 🔒 ตรวจสอบ MIN_UTIL และ MAX_UTIL (100%)
        max_util = max(w_util, c_util)
        min_util_required = MIN_UTIL.get(suggested, 70)
        
        # สร้าง status
        if max_util > 100:
            status = '🚫 เกิน100%'
        elif max_util < min_util_required:
            status = f'⚠️ ต่ำ{min_util_required}%'
        else:
            status = '✅ ผ่าน'
        
        summary_data.append({
            'Trip': trip_num,
            'Branches': len(trip_data),
            'Weight': total_w,
            'Cube': total_c,
            'Truck': f"{suggested} {source}",
            'Weight_Use%': w_util,
            'Cube_Use%': c_util,
            'Max_Util%': max_util,
            'Status': status,
            'Total_Distance': total_distance
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    # เพิ่มคอลัมน์รถที่จัดส่งในรายละเอียดรายสาขา
    trip_truck_map = {}
    trip_truck_type_map = {}  # เก็บเฉพาะประเภทรถ (ไม่รวม source)
    for _, row in summary_df.iterrows():
        trip_truck_map[row['Trip']] = row['Truck']
        # ดึงประเภทรถ (ตัด emoji และ source ออก)
        truck_type = row['Truck'].split()[0] if row['Truck'] else '6W'
        trip_truck_type_map[row['Trip']] = truck_type
    
    test_df['Truck'] = test_df['Trip'].map(trip_truck_map)
    
    # เพิ่มคอลัมน์ภาค (Region) สำหรับแสดงผล
    def get_region_name(code):
        """ดึงชื่อภาคจากรหัสสาขา"""
        if code not in test_df['Code'].values:
            return 'ไม่ระบุ'
        
        prov = test_df[test_df['Code'] == code]['Province'].iloc[0] if 'Province' in test_df.columns else None
        if pd.isna(prov) or prov == 'UNKNOWN':
            return 'ไม่ระบุ'
        
        region_type = get_region_type(prov)
        
        # แปลงเป็นชื่อภาคภาษาไทย
        if region_type == 'nearby':
            # ตรวจสอบว่าเป็นกรุงเทพหรือปริมณฑล
            bangkok = ['กรุงเทพมหานคร', 'กรุงเทพ']
            if prov in bangkok:
                return 'กรุงเทพ'
            else:
                return 'ปริมณฑล'
        else:
            # จัดกลุ่มตามภูมิภาค
            province_regions = {
                'กลางตอนบน': ['ชัยนาท', 'พระนครศรีอยุธยา', 'ลพบุรี', 'สระบุรี', 'สิงห์บุรี', 'อ่างทอง', 'อยุธยา'],
                'กลางตอนล่าง': ['สมุทรสงคราม', 'สุพรรณบุรี'],
                'ตะวันตก': ['กาญจนบุรี', 'ประจวบคีรีขันธ์', 'ราชบุรี', 'เพชรบุรี'],
                'ตะวันออก': ['จันทบุรี', 'ชลบุรี', 'ตราด', 'นครนายก', 'ปราจีนบุรี', 'ระยอง', 'สระแก้ว', 'ฉะเชิงเทรา'],
                'อีสานเหนือ': ['นครพนม', 'บึงกาฬ', 'มุกดาหาร', 'สกลนคร', 'หนองคาย', 'หนองบัวลำภู', 'อุดรธานี', 'เลย'],
                'อีสานกลาง': ['กาฬสินธุ์', 'ขอนแก่น', 'ชัยภูมิ', 'มหาสารคาม', 'ร้อยเอ็ด'],
                'อีสานใต้': ['นครราชสีมา', 'โคราช', 'บุรีรัมย์', 'ยโสธร', 'ศรีสะเกษ', 'สุรินทร์', 'อำนาจเจริญ', 'อุบลราชธานี'],
                'เหนือตอนบน': ['น่าน', 'พะเยา', 'ลำปาง', 'ลำพูน', 'เชียงราย', 'เชียงใหม่', 'แพร่', 'แม่ฮ่องสอน'],
                'เหนือตอนล่าง': ['กำแพงเพชร', 'ตาก', 'นครสวรรค์', 'พิจิตร', 'พิษณุโลก', 'สุโขทัย', 'อุตรดิตถ์', 'อุทัยธานี', 'เพชรบูรณ์'],
                'ใต้ฝั่งอันดามัน': ['กระบี่', 'ตรัง', 'พังงา', 'ภูเก็ต', 'ระนอง', 'สตูล'],
                'ใต้ฝั่งอ่าวไทย': ['ชุมพร', 'นครศรีธรรมราช', 'พัทลุง', 'ยะลา', 'สงขลา', 'สุราษฎร์ธานี', 'ปัตตานี', 'นราธิวาส']
            }
            
            for region_name, provinces in province_regions.items():
                if prov in provinces:
                    return region_name
            
            return 'ต่างจังหวัด'
    
    test_df['Region'] = test_df['Code'].apply(get_region_name)
    
    # 🆕 เพิ่มคอลัมน์ตำบลและอำเภอจาก Master Data
    def get_subdistrict(code):
        """ดึงตำบลจาก Master Data"""
        if MASTER_DATA.empty or 'Plan Code' not in MASTER_DATA.columns:
            return ''
        master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
        if len(master_row) > 0:
            sub = master_row.iloc[0].get('ตำบล', '')
            return str(sub).strip() if pd.notna(sub) else ''
        return ''
    
    def get_district(code):
        """ดึงอำเภอจาก Master Data"""
        if MASTER_DATA.empty or 'Plan Code' not in MASTER_DATA.columns:
            return ''
        master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
        if len(master_row) > 0:
            dist = master_row.iloc[0].get('อำเภอ', '')
            return str(dist).strip() if pd.notna(dist) else ''
        return ''
    
    test_df['Subdistrict'] = test_df['Code'].apply(get_subdistrict)
    test_df['District'] = test_df['Code'].apply(get_district)
    
    # เพิ่มคอลัมน์ระยะทางระหว่างสาขาในทริป และเรียงลำดับด้วย Nearest Neighbor
    def add_distance_and_sort(df):
        """
        🔄 เรียงสาขาภายในแต่ละทริปด้วย Nearest Neighbor Algorithm
        - เริ่มจาก DC → หาสาขาที่ใกล้ที่สุด → หาถัดไปที่ใกล้ที่สุด → ...
        - ป้องกันการกระโดดไปมา
        """
        # สร้าง coord cache จาก MASTER_DATA
        coord_cache_local = {}
        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
            for _, row in MASTER_DATA.iterrows():
                code = row['Plan Code']
                lat = row.get('ละติจูด', 0)
                lon = row.get('ลองติจูด', 0)
                if lat and lon and pd.notna(lat) and pd.notna(lon):
                    coord_cache_local[code] = (float(lat), float(lon))
        
        # เก็บลำดับ Sequence ใหม่สำหรับแต่ละ code
        new_sequences = {}
        trip_distances = {}
        
        for trip_num in df['Trip'].unique():
            if trip_num == 0:
                continue
                
            trip_data = df[df['Trip'] == trip_num]
            trip_codes = trip_data['Code'].tolist()
            
            if len(trip_codes) <= 1:
                # สาขาเดี่ยว
                if trip_codes:
                    new_sequences[trip_codes[0]] = 1
                trip_distances[trip_num] = 0
                continue
            
            # 🔄 Nearest Neighbor Algorithm
            # สร้างลิสต์พิกัด
            points = []
            for code in trip_codes:
                lat, lon = coord_cache_local.get(code, (None, None))
                if lat and lon:
                    points.append((code, lat, lon))
                else:
                    # ไม่มีพิกัด ให้ไว้ท้าย
                    points.append((code, DC_WANG_NOI_LAT, DC_WANG_NOI_LON))
            
            # เรียงลำดับด้วย Nearest Neighbor จาก DC
            sorted_codes = []
            remaining = points.copy()
            current_lat, current_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
            
            while remaining:
                best_idx = 0
                best_dist = haversine_distance(current_lat, current_lon, remaining[0][1], remaining[0][2])
                
                for i, (_, lat, lon) in enumerate(remaining[1:], 1):
                    dist = haversine_distance(current_lat, current_lon, lat, lon)
                    if dist < best_dist:
                        best_dist = dist
                        best_idx = i
                
                best_point = remaining.pop(best_idx)
                sorted_codes.append(best_point[0])
                current_lat, current_lon = best_point[1], best_point[2]
            
            # กำหนด Sequence ใหม่
            for seq, code in enumerate(sorted_codes, 1):
                new_sequences[code] = seq
            
            # คำนวณ max consecutive distance
            max_consecutive_dist = 0
            for i in range(len(sorted_codes) - 1):
                code1, code2 = sorted_codes[i], sorted_codes[i + 1]
                lat1, lon1 = coord_cache_local.get(code1, (0, 0))
                lat2, lon2 = coord_cache_local.get(code2, (0, 0))
                if lat1 and lon1 and lat2 and lon2:
                    dist = haversine_distance(lat1, lon1, lat2, lon2)
                    if dist > max_consecutive_dist:
                        max_consecutive_dist = dist
            
            trip_distances[trip_num] = round(max_consecutive_dist, 2)
        
        # เพิ่มคอลัมน์ Sequence ใหม่
        df['Sequence'] = df['Code'].map(new_sequences).fillna(999)
        
        # เพิ่มคอลัมน์ระยะทาง max ระหว่างสาขาติดกันในทริป
        df['Max_Distance_in_Trip'] = df['Trip'].map(trip_distances)
        
        # เรียงลำดับภายในแต่ละทริป: Trip → Sequence (Nearest Neighbor order)
        df = df.sort_values(['Trip', 'Sequence'], ascending=[True, True])
        return df
    
    test_df = add_distance_and_sort(test_df)
    
    # 🗺️ คำนวณระยะทางแบบละเอียด
    def calculate_detailed_distances(df):
        """คำนวณ: DC→สาขาแรก, ระหว่างสาขา, รวมทั้งทริป"""
        
        # เตรียม dict เก็บผลลัพธ์
        distance_from_dc = {}
        distance_to_next = {}
        total_trip_distance = {}
        
        for trip_num in df['Trip'].unique():
            if trip_num == 0:
                continue
            
            trip_data = df[df['Trip'] == trip_num].copy()
            
            # เรียงตาม Sequence (ถ้ามี) หรือ Weight เพื่อให้ได้ลำดับเดียวกับการแสดงผล
            if 'Sequence' in trip_data.columns:
                trip_data = trip_data.sort_values('Sequence', ascending=True)
            else:
                trip_data = trip_data.sort_values('Weight', ascending=False)
            codes = trip_data['Code'].tolist()
            
            # คำนวณระยะทาง
            trip_total_dist = 0
            prev_lat, prev_lon = DC_WANG_NOI_LAT, DC_WANG_NOI_LON
            
            for i, code in enumerate(codes):
                # หาพิกัดสาขานี้
                m = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                
                if len(m) > 0:
                    lat = m.iloc[0].get('ละติจูด', 0)
                    lon = m.iloc[0].get('ลองติจูด', 0)
                    
                    if lat and lon:
                        # คำนวณระยะทางจากจุดก่อนหน้า
                        dist = haversine_distance(prev_lat, prev_lon, lat, lon)
                        
                        if i == 0:
                            # สาขาแรก: ระยะจาก DC
                            distance_from_dc[code] = round(dist, 2)
                            distance_to_next[code] = 0  # ไม่มีระยะ "ก่อนหน้า"
                        else:
                            # สาขาถัดไป: ระยะจากสาขาก่อนหน้า
                            distance_from_dc[code] = round(haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon), 2)
                            distance_to_next[codes[i-1]] = round(dist, 2)  # บันทึกที่สาขาก่อนหน้า
                            
                            if i == len(codes) - 1:
                                # สาขาสุดท้าย: ไม่มี "ถัดไป"
                                distance_to_next[code] = 0
                        
                        trip_total_dist += dist
                        prev_lat, prev_lon = lat, lon
                    else:
                        distance_from_dc[code] = 0
                        distance_to_next[code] = 0
                else:
                    distance_from_dc[code] = 0
                    distance_to_next[code] = 0
            
            # บันทึกระยะรวมทั้งทริป
            for code in codes:
                total_trip_distance[code] = round(trip_total_dist, 2)
        
        # เพิ่มคอลัมน์ลงใน DataFrame
        df['Distance_from_DC'] = df['Code'].map(distance_from_dc).fillna(0)
        df['Distance_to_Next'] = df['Code'].map(distance_to_next).fillna(0)
        df['Total_Trip_Distance'] = df['Code'].map(total_trip_distance).fillna(0)
        
        return df
    
    test_df = calculate_detailed_distances(test_df)
    
    # เพิ่มคอลัมน์เช็คว่าสาขาใช้รถประเภทนี้ได้หรือไม่ (อ้างอิงจาก Auto Plan เท่านั้น!)
    def check_vehicle_history(row):
        code = row['Code']
        trip = row['Trip']
        truck_type = trip_truck_type_map.get(trip, '6W')
        
        vehicle_sizes = {'4W': 1, 'JB': 2, '6W': 3}
        requested_size = vehicle_sizes.get(truck_type, 3)
        
        # 🔒 อ้างอิงจาก AUTO_PLAN_TRUCK_LIMITS เท่านั้น (ไม่ใช้ Booking History)
        if code in AUTO_PLAN_TRUCK_LIMITS:
            max_allowed = AUTO_PLAN_TRUCK_LIMITS[code]
            max_allowed_size = vehicle_sizes.get(max_allowed, 3)
            
            if requested_size <= max_allowed_size:
                # รถที่ขอใช้ เล็กกว่าหรือเท่ากับที่อนุญาต = ใช้ได้
                return f"✅ ใช้ได้ (อนุญาต {max_allowed})"
            else:
                # รถที่ขอใช้ ใหญ่กว่าที่อนุญาต = ห้าม
                return f"🚫 จำกัด {max_allowed} (Auto Plan)"
        
        # ถ้าไม่มีใน Auto Plan = ใช้ได้ทุกประเภท (ไม่มีข้อจำกัด)
        return "✅ ใช้ได้ (ไม่มีข้อจำกัด)"
    
    test_df['VehicleCheck'] = test_df.apply(check_vehicle_history, axis=1)
    
    # เพิ่มคอลัมน์เตือนสำหรับทริปที่เกิน 12 สาขา (เฉพาะรถเล็ก)
    def check_branch_count(row):
        trip_num = row['Trip']
        if trip_num == 0:
            return ""
        
        trip_branch_count = len(test_df[test_df['Trip'] == trip_num])
        truck_type = trip_truck_type_map.get(trip_num, '6W')
        
        # เช็คเฉพาะรถเล็ก (4W, JB) - รถใหญ่ (6W) ไม่เตือน
        if trip_branch_count > 12:
            if truck_type in ['4W', 'JB']:
                return f"⚠️ เกิน 12 สาขา ({trip_branch_count} สาขา) - {truck_type}"
            else:
                return f"✅ {trip_branch_count} สาขา - {truck_type} (ยอมได้)"
        else:
            return f"✅ {trip_branch_count} สาขา - {truck_type}"
    
    test_df['BranchCount'] = test_df.apply(check_branch_count, axis=1)
    
    # ===============================================
    # 🎯 Renumber trips: เริ่มจาก 1, 2, 3, ... (ไม่มีช่องว่าง)
    # ===============================================
    unique_trips = sorted(test_df['Trip'].dropna().unique())
    trip_renumber_map = {old: new for new, old in enumerate(unique_trips, start=1)}
    test_df['Trip'] = test_df['Trip'].map(trip_renumber_map)
    
    # อัปเดต summary_df ด้วย
    if 'Trip' in summary_df.columns:
        summary_df['Trip'] = summary_df['Trip'].map(trip_renumber_map)
    
    return test_df, summary_df

# ==========================================
# STREAMLIT UI
# ==========================================
def main():
    st.set_page_config(
        page_title="ระบบจัดเที่ยว",
        page_icon="🚚",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # 🔧 Initialize session state สำหรับเก็บผลลัพธ์
    if 'result_df' not in st.session_state:
        st.session_state.result_df = None
    if 'summary' not in st.session_state:
        st.session_state.summary = None
    if 'processed' not in st.session_state:
        st.session_state.processed = False
    if 'original_file_content' not in st.session_state:
        st.session_state.original_file_content = None
    
    # 🔄 Auto-refresh ทุกเที่ยงคืน (ล้างแคช)
    if AUTOREFRESH_AVAILABLE:
        now = datetime.now()
        # คำนวณเวลาถึงเที่ยงคืน (00:00:00)
        midnight = datetime.combine(now.date(), time(0, 0, 0))
        
        # ถ้ายังไม่ถึงเที่ยงคืน เอาเที่ยงคืนวันถัดไป
        if now < midnight:
            next_midnight = midnight
        else:
            from datetime import timedelta
            next_midnight = midnight + timedelta(days=1)
        
        # คำนวณเวลาที่เหลือ (วินาที)
        seconds_until_midnight = int((next_midnight - now).total_seconds())
        
        # Refresh ทุกเที่ยงคืน (เฉพาะเมื่อมี autorefresh component)
        if AUTOREFRESH_AVAILABLE and seconds_until_midnight > 0:
            # เช็คในช่วง 5 นาทีก่อนเที่ยงคืน (หลัง 23:55)
            if seconds_until_midnight <= 300:  # 5 minutes
                st.info(f"🔄 ระบบจะ Refresh อัตโนมัติใน {seconds_until_midnight // 60} นาที")
                st_autorefresh(interval=seconds_until_midnight * 1000, key="midnight_refresh")
            else:
                # ตรวจสอบทุก 1 ชั่วโมง
                st_autorefresh(interval=3600000, limit=24, key="hourly_check")
    
    # Header
    col1, col2 = st.columns([3, 1])
    with col1:
        st.title("🚚 ระบบจัดเที่ยว")
    with col2:
        st.markdown("# 🚚")
    
    # Show Punthai learning stats
    if PUNTHAI_PATTERNS and 'stats' in PUNTHAI_PATTERNS and PUNTHAI_PATTERNS['stats']:
        stats = PUNTHAI_PATTERNS['stats']
        with st.expander("📊 สถิติที่เรียนรู้จากไฟล์ Punthai Maxmart", expanded=False):
            col_a, col_b, col_c = st.columns(3)
            with col_a:
                st.metric("เฉลี่ยสาขา/ทริป", f"{stats.get('avg_branches', 0):.1f}")
            with col_b:
                st.metric("ทริปจังหวัดเดียว", f"{stats.get('same_province_pct', 0):.1f}%")
            with col_c:
                total_trips = stats.get('same_province', 0) + stats.get('mixed_province', 0)
                st.metric("จำนวนทริปอ้างอิง", total_trips)
    
    st.markdown("---")
    
    # โหลดโมเดล
    model_data = load_model()
    
    if not model_data:
        st.error("❌ ไม่พบข้อมูลโมเดล กรุณาเทรนโมเดลก่อนใช้งาน")
        st.info("💡 รันคำสั่ง: `python test_model.py`")
        st.stop()
    
    # อัปโหลดไฟล์ครั้งเดียว
    st.markdown("### 📂 อัปโหลดไฟล์รายการออเดอร์")
    uploaded_file = st.file_uploader(
        "เลือกไฟล์ Excel (.xlsx)", 
        type=['xlsx'],
        help="อัปโหลดไฟล์ Excel ที่มีรายการสาขาและออเดอร์"
    )
    
    if uploaded_file:
        # เก็บ file content ไว้ใช้ตอน export - เก็บใน session_state
        file_content = uploaded_file.read()
        uploaded_file.seek(0)  # reset pointer
        
        # เก็บไว้ใน session_state
        st.session_state.original_file_content = file_content
        
        with st.spinner("⏳ กำลังอ่านข้อมูล..."):
            df = load_excel(file_content)
            df = process_dataframe(df)
            
            if df is not None and 'Code' in df.columns:
                st.success(f"✅ อ่านข้อมูลสำเร็จ: **{len(df):,}** รายการ")
                
                # แสดงข้อมูลพื้นฐาน
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("📍 จำนวนสาขา", f"{df['Code'].nunique():,}")
                with col2:
                    st.metric("⚖️ น้ำหนักรวม", f"{df['Weight'].sum():,.0f} kg")
                with col3:
                    st.metric("📦 คิวรวม", f"{df['Cube'].sum():.1f} m³")
                with col4:
                    provinces = df['Province'].nunique() if 'Province' in df.columns else 0
                    st.metric("🗺️ จังหวัด", f"{provinces}")
                
                # แสดงตัวอย่างข้อมูล
                with st.expander("🔍 ดูข้อมูลตัวอย่าง"):
                    st.dataframe(df.head(10), use_container_width=True)
                
                st.markdown("---")
                
                # แท็บหลัก
                tab1, tab2 = st.tabs(["📦 จัดเที่ยว (ตามน้ำหนัก)", "🗺️ จัดกลุ่มตามภาค (ไม่สนน้ำหนัก)"])
                    
                # ==========================================
                # แท็บ 1: จัดเที่ยว (ตามน้ำหนัก)
                # ==========================================
                with tab1:
                    # ปุ่มจัดทริป
                    if st.button("🚀 เริ่มจัดเที่ยว", type="primary", use_container_width=True):
                        with st.spinner("⏳ กำลังประมวลผล..."):
                            result_df, summary = predict_trips(df.copy(), model_data)
                            
                            # 🔧 เก็บผลลัพธ์ใน session_state เพื่อไม่ให้หายเมื่อกด download
                            st.session_state.result_df = result_df
                            st.session_state.summary = summary
                            st.session_state.processed = True
                            
                            st.balloons()
                    
                    # 🔧 แสดงผลลัพธ์จาก session_state (ถ้ามี)
                    if st.session_state.processed and st.session_state.result_df is not None:
                        result_df = st.session_state.result_df
                        summary = st.session_state.summary
                        
                        st.success(f"✅ **จัดทริปเสร็จสมบูรณ์!** รวม **{len(summary)}** ทริป")
                        
                        st.markdown("---")
                        
                        # สถิติโดยรวม
                        st.markdown("### 📊 สรุปผลการจัดทริป")
                        
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("🚚 จำนวนทริป", len(summary))
                        with col2:
                            st.metric("📍 จำนวนสาขา", len(result_df))
                        with col3:
                            avg_branches = len(result_df) / result_df['Trip'].nunique()
                            st.metric("📊 เฉลี่ยสาขา/ทริป", f"{avg_branches:.1f}")
                        with col4:
                            avg_util = summary['Cube_Use%'].mean()
                            st.metric("📈 การใช้รถเฉลี่ย", f"{avg_util:.0f}%")
                        
                        st.markdown("---")
                        
                        # ตารางสรุปแต่ละทริป
                        st.markdown("### 🚛 รายละเอียดแต่ละทริป")
                        
                        # แสดงสรุปทริปที่ไม่ผ่าน (ถ้ามี)
                        failed_trips = summary[summary['Status'] != '✅ ผ่าน']
                        if len(failed_trips) > 0:
                            st.warning(f"⚠️ พบ **{len(failed_trips)}** ทริปที่ไม่ผ่านเกณฑ์")
                        
                        st.dataframe(
                            summary.style.format({
                                'Weight': '{:.2f}',
                                'Cube': '{:.2f}',
                                'Weight_Use%': '{:.1f}%',
                                'Cube_Use%': '{:.1f}%',
                                'Max_Util%': '{:.1f}%',
                                'Total_Distance': '{:.1f} km'
                            }).background_gradient(
                                subset=['Max_Util%'],
                                cmap='RdYlGn',
                                vmin=0,
                                vmax=100
                            ),
                            use_container_width=True,
                            height=400
                        )
                        
                        # ตารางรายละเอียดทั้งหมด (มีคอลัมน์รถและระยะทาง)
                        with st.expander("📋 ดูรายละเอียดรายสาขา (เรียงตามน้ำหนัก)"):
                            # จัดเรียงคอลัมน์ที่สำคัญ
                            display_cols = ['Trip', 'Code', 'Name']
                            if 'Province' in result_df.columns:
                                display_cols.append('Province')
                            if 'District' in result_df.columns:
                                display_cols.append('District')
                            if 'Subdistrict' in result_df.columns:
                                display_cols.append('Subdistrict')
                            if 'Region' in result_df.columns:
                                display_cols.append('Region')
                            display_cols.extend(['Max_Distance_in_Trip', 'Weight', 'Cube', 'Truck', 'VehicleCheck'])
                            
                            # กรองคอลัมน์ที่มีอยู่จริง
                            display_cols = [col for col in display_cols if col in result_df.columns]
                            display_df = result_df[display_cols].copy()
                            
                            # ตั้งชื่อคอลัมน์ภาษาไทย
                            col_names = {'Trip': 'ทริป', 'Code': 'รหัส', 'Name': 'ชื่อสาขา', 'Province': 'จังหวัด', 
                                       'District': 'อำเภอ', 'Subdistrict': 'ตำบล',
                                       'Region': 'ภาค', 'Max_Distance_in_Trip': 'ระยะทาง Max(km)', 
                                       'Weight': 'น้ำหนัก(kg)', 'Cube': 'คิว(m³)', 'Truck': 'รถ', 'VehicleCheck': 'ตรวจสอบรถ'}
                            display_df.columns = [col_names.get(c, c) for c in display_cols]
                            
                            # จัดรูปแบบคอลัมน์ระยะทาง
                            st.dataframe(
                                display_df.style.format({
                                    'ระยะทาง(km)': '{:.1f}',
                                    'น้ำหนัก(kg)': '{:.2f}',
                                    'คิว(m³)': '{:.2f}'
                                }),
                                use_container_width=True, 
                                height=400
                            )
                        
                        # แสดงสาขาที่มีคำเตือน
                        warning_branches = result_df[result_df['VehicleCheck'].str.contains('⚠️', na=False)]
                        if len(warning_branches) > 0:
                            with st.expander(f"⚠️ สาขาที่ใช้รถต่างจากปกติ ({len(warning_branches)} สาขา)"):
                                st.warning("สาขาเหล่านี้ปกติใช้รถประเภทอื่น แต่ถูกจัดให้ใช้รถประเภทที่ต่างออกไป")
                                display_cols_warn = ['Trip', 'Code', 'Name', 'Truck', 'VehicleCheck']
                                display_warn_df = warning_branches[display_cols_warn].copy()
                                display_warn_df.columns = ['ทริป', 'รหัส', 'ชื่อสาขา', 'รถที่จัด', 'ประวัติการใช้รถ']
                                st.dataframe(display_warn_df, use_container_width=True)
                        
                        st.markdown("---")
                        
                        # ดาวน์โหลด - เขียนทับชีต 2.Punthai ในไฟล์ต้นฉบับ พร้อมสลับสีเหลืองโทนส้ม-ขาว
                        output = io.BytesIO()
                        
                        # สร้าง Trip_No map (JB ใช้ prefix 4WJ)
                        trip_no_map = {}
                        vehicle_counts = {'4W': 0, '4WJ': 0, '6W': 0}
                        
                        for trip_num in sorted(result_df['Trip'].unique()):
                            if trip_num == 0:
                                continue
                            trip_summary = summary[summary['Trip'] == trip_num]
                            if len(trip_summary) > 0:
                                truck_info = trip_summary.iloc[0]['Truck']
                                vehicle_type = truck_info.split()[0] if truck_info else '6W'
                                # JB ใช้ prefix 4WJ แทน
                                if vehicle_type == 'JB':
                                    vehicle_type = '4WJ'
                                vehicle_counts[vehicle_type] = vehicle_counts.get(vehicle_type, 0) + 1
                                trip_no = f"{vehicle_type}{vehicle_counts[vehicle_type]:03d}"
                                trip_no_map[trip_num] = trip_no
                        
                        # โหลดไฟล์ต้นฉบับเพื่อ copy ทุกชีต
                        from openpyxl import load_workbook
                        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                        from copy import copy
                        
                        try:
                            # โหลด workbook ต้นฉบับ จาก session_state
                            wb = load_workbook(io.BytesIO(st.session_state.original_file_content))
                            
                            # หาชีตเป้าหมาย (2.Punthai)
                            target_sheet = None
                            for sheet_name in wb.sheetnames:
                                if 'punthai' in sheet_name.lower() or '2.' in sheet_name.lower():
                                    target_sheet = sheet_name
                                    break
                            
                            if not target_sheet:
                                target_sheet = '2.Punthai'
                                if target_sheet not in wb.sheetnames:
                                    wb.create_sheet(target_sheet)
                            
                            ws = wb[target_sheet]
                            
                            # ลบข้อมูลเก่า (เก็บแถวแรก header)
                            # หา header row (แถวที่มี "รหัสสาขา" หรือ "Trip")
                            header_row = 1
                            for row_idx in range(1, min(5, ws.max_row + 1)):
                                for col_idx in range(1, min(15, ws.max_column + 1)):
                                    cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                                    if 'รหัสสาขา' in cell_val or 'Trip' in cell_val.upper():
                                        header_row = row_idx
                                        break
                            
                            # ลบข้อมูลตั้งแต่แถวหลัง header
                            if ws.max_row > header_row:
                                ws.delete_rows(header_row + 1, ws.max_row - header_row)
                            
                            # สีเหลืองโทนส้ม-ขาว
                            yellow_orange = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            # 🔴 สีแดงสำหรับทริปที่ไม่ผ่านเกณฑ์
                            red_font = Font(color='FF0000', bold=True)
                            
                            # สร้าง map ของทริปที่ไม่ผ่านเกณฑ์
                            failed_trips = set()
                            if 'TripStatus' in result_df.columns:
                                for t in result_df['Trip'].unique():
                                    trip_status = result_df[result_df['Trip'] == t]['TripStatus'].iloc[0] if len(result_df[result_df['Trip'] == t]) > 0 else ''
                                    if '❌' in str(trip_status) or '⛔' in str(trip_status):
                                        failed_trips.add(t)
                            
                            # เขียนข้อมูลใหม่ (แถวที่ header_row + 1 เป็นต้นไป)
                            current_trip = None
                            use_yellow = True
                            row_num = header_row + 1
                            sep_num = 1  # เริ่มนับ Sep. จาก 1
                            
                            for trip_num in sorted(result_df['Trip'].unique()):
                                if trip_num == 0:
                                    continue
                                trip_data = result_df[result_df['Trip'] == trip_num].copy()
                                trip_no = trip_no_map.get(trip_num, '')
                                
                                # เปลี่ยนสีเมื่อเปลี่ยนทริป
                                if current_trip != trip_num:
                                    current_trip = trip_num
                                    use_yellow = not use_yellow
                                
                                fill = yellow_orange if use_yellow else white_fill
                                
                                for _, row in trip_data.iterrows():
                                    # เขียนข้อมูลตามโครงสร้างไฟล์ต้นฉบับ
                                    # คอลัมน์: A=Sep, B=BU, C=รหัสสาขา, D=รหัส WMS, E=สาขา, F=Cube, G=Weight, H=Original QTY, I=Trip, J=Trip no
                                    data = [
                                        sep_num,  # A: Sep (ลำดับแถว)
                                        row.get('BU', 211),  # B: BU (จากต้นฉบับ)
                                        row.get('Code', ''),  # C: รหัสสาขา
                                        row.get('Code', ''),  # D: รหัส WMS
                                        row.get('Name', ''),  # E: สาขา
                                        round(row.get('Cube', 0), 2) if pd.notna(row.get('Cube')) else 0,  # F: Cube
                                        round(row.get('Weight', 0), 2) if pd.notna(row.get('Weight')) else 0,  # G: Weight
                                        row.get('OriginalQty', 0) if pd.notna(row.get('OriginalQty')) else 0,  # H: Original QTY (จากต้นฉบับ)
                                        int(trip_num),  # I: Trip
                                        trip_no,  # J: Trip no
                                        '',  # K: วันที่โหลด
                                        '',  # L: เวลาโหลด
                                        '',  # M: ประตู
                                        '',  # N: WAVE
                                        '',  # O: remark
                                        '',  # P: lat (เว้นว่าง)
                                        '',  # Q: lon (เว้นว่าง)
                                    ]
                                    
                                    for col_idx, value in enumerate(data, 1):
                                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                                        cell.fill = fill
                                        cell.border = thin_border
                                        # 🔴 ถ้าทริปไม่ผ่านเกณฑ์ ใช้ตัวหนังสือสีแดง
                                        if trip_num in failed_trips:
                                            cell.font = red_font
                                    
                                    row_num += 1
                                    sep_num += 1  # เพิ่มลำดับ Sep
                                
                                # 🚛 เพิ่ม DC011 ปิดท้ายทุกทริป
                                dc_data = [
                                    sep_num,  # A: Sep
                                    'PROJECT',  # B: BU
                                    'DC011',  # C: รหัสสาขา
                                    'DC011',  # D: รหัส WMS
                                    'บ.พีทีจี เอ็นเนอยี จำกัด (มหาชน) (DCวังน้อย)',  # E: สาขา
                                    0,  # F: Cube
                                    0,  # G: Weight
                                    0,  # H: Original QTY
                                    int(trip_num),  # I: Trip
                                    trip_no,  # J: Trip no
                                    '',  # K: วันที่โหลด
                                    '',  # L: เวลาโหลด
                                '',  # M: ประตู
                                    '',  # N: WAVE
                                    '',  # O: remark
                                    '',  # P: lat (เว้นว่าง)
                                    '',  # Q: lon (เว้นว่าง)
                                ]
                                
                                for col_idx, value in enumerate(dc_data, 1):
                                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                                    cell.fill = fill
                                    cell.border = thin_border
                                    # 🔴 ถ้าทริปไม่ผ่านเกณฑ์ ใช้ตัวหนังสือสีแดง
                                    if trip_num in failed_trips:
                                        cell.font = red_font
                                
                                row_num += 1
                                sep_num += 1
                            
                            # บันทึกลง BytesIO
                            wb.save(output)
                            output.seek(0)
                            
                        except Exception as e:
                            st.warning(f"⚠️ ไม่สามารถเขียนทับไฟล์ต้นฉบับได้: {e}")
                            # Fallback: สร้างไฟล์ใหม่
                            from openpyxl import Workbook
                            wb = Workbook()
                            ws = wb.active
                            ws.title = '2.Punthai'
                            
                            # เขียน header
                            headers = ['Sep.', 'BU', 'รหัสสาขา', 'รหัส WMS', 'สาขา', 'Total Cube', 'Total Wgt', 'Original QTY', 'Trip', 'Trip no']
                            for col_num, header in enumerate(headers, 1):
                                ws.cell(row=1, column=col_num, value=header)
                            
                            row_num = 2
                            sep_num = 1  # เริ่มนับ Sep. จาก 1
                            current_trip = None
                            use_yellow = True
                            yellow_orange = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                            red_font_fallback = Font(color='FF0000', bold=True)
                            
                            # สร้าง map ของทริปที่ไม่ผ่านเกณฑ์
                            failed_trips_fallback = set()
                            if 'TripStatus' in result_df.columns:
                                for t in result_df['Trip'].unique():
                                    trip_status = result_df[result_df['Trip'] == t]['TripStatus'].iloc[0] if len(result_df[result_df['Trip'] == t]) > 0 else ''
                                    if '❌' in str(trip_status) or '⛔' in str(trip_status):
                                        failed_trips_fallback.add(t)
                            
                            for trip_num in sorted(result_df['Trip'].unique()):
                                if trip_num == 0:
                                    continue
                                trip_data = result_df[result_df['Trip'] == trip_num]
                                trip_no = trip_no_map.get(trip_num, '')
                                
                                if current_trip != trip_num:
                                    current_trip = trip_num
                                    use_yellow = not use_yellow
                                fill = yellow_orange if use_yellow else white_fill
                                
                                for _, row in trip_data.iterrows():
                                    original_qty = row.get('OriginalQty', 0) if pd.notna(row.get('OriginalQty')) else 0
                                    data = [sep_num, row.get('BU', 211), row.get('Code', ''), row.get('Code', ''), row.get('Name', ''),
                                            round(row.get('Cube', 0), 2), round(row.get('Weight', 0), 2), original_qty, int(trip_num), trip_no]
                                    for col_idx, value in enumerate(data, 1):
                                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                                        cell.fill = fill
                                        # 🔴 ถ้าทริปไม่ผ่านเกณฑ์ ใช้ตัวหนังสือสีแดง
                                        if trip_num in failed_trips_fallback:
                                            cell.font = red_font_fallback
                                    row_num += 1
                                    sep_num += 1
                                
                                # 🚛 เพิ่ม DC011 ปิดท้ายทุกทริป
                                dc_data = [sep_num, 'PROJECT', 'DC011', 'DC011', 'บ.พีทีจี เอ็นเนอยี จำกัด (มหาชน) (DCวังน้อย)',
                                           0, 0, 0, int(trip_num), trip_no]
                                for col_idx, value in enumerate(dc_data, 1):
                                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                                    cell.fill = fill
                                    # 🔴 ถ้าทริปไม่ผ่านเกณฑ์ ใช้ตัวหนังสือสีแดง
                                    if trip_num in failed_trips_fallback:
                                        cell.font = red_font_fallback
                                row_num += 1
                                sep_num += 1
                            
                            wb.save(output)
                            output.seek(0)
                        
                        col1, col2, col3 = st.columns([1, 2, 1])
                        with col2:
                            st.download_button(
                                label="📥 ดาวน์โหลดผลลัพธ์ (Excel)",
                                data=output.getvalue(),
                                file_name=f"ผลจัดทริป_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                
                # ==========================================
                # แท็บ 2: จัดกลุ่มสาขาตามภาค (ไม่สนน้ำหนัก)
                # ==========================================
                with tab2:
                    df_region = df.copy()
                    
                    # จัดกลุ่มตามภาค
                    branch_info = model_data.get('branch_info', {})
                    trip_pairs = model_data.get('trip_pairs', set())
                    
                    # สร้างข้อมูลภาคสำหรับแต่ละสาขา (จากไฟล์ประวัติ)
                    region_groups = {
                        'ภาคกลาง-กรุงเทพชั้นใน': ['กรุงเทพมหานคร'],
                        'ภาคกลาง-กรุงเทพชั้นกลาง': ['กรุงเทพมหานคร'],
                        'ภาคกลาง-กรุงเทพชั้นนอก': ['กรุงเทพมหานคร'],
                        'ภาคกลาง-ปริมณฑล': ['นครปฐม', 'นนทบุรี', 'ปทุมธานี', 'สมุทรปราการ', 'สมุทรสาคร'],
                        'ภาคกลาง-กลางตอนบน': ['ชัยนาท', 'พระนครศรีอยุธยา', 'ลพบุรี', 'สระบุรี', 'สิงห์บุรี', 'อ่างทอง', 'อยุธยา'],
                        'ภาคกลาง-กลางตอนล่าง': ['สมุทรสงคราม', 'สุพรรณบุรี'],
                        'ภาคตะวันตก': ['กาญจนบุรี', 'ประจวบคีรีขันธ์', 'ราชบุรี', 'เพชรบุรี'],
                        'ภาคตะวันออก': ['จันทบุรี', 'ชลบุรี', 'ตราด', 'นครนายก', 'ปราจีนบุรี', 'ระยอง', 'สระแก้ว', 'ฉะเชิงเทรา'],
                        'ภาคอีสาน-อีสานเหนือ': ['นครพนม', 'บึงกาฬ', 'มุกดาหาร', 'สกลนคร', 'หนองคาย', 'หนองบัวลำภู', 'อุดรธานี', 'เลย'],
                        'ภาคอีสาน-อีสานกลาง': ['กาฬสินธุ์', 'ขอนแก่น', 'ชัยภูมิ', 'มหาสารคาม', 'ร้อยเอ็ด'],
                        'ภาคอีสาน-อีสานใต้': ['นครราชสีมา', 'โคราช', 'บุรีรัมย์', 'ยโสธร', 'ศรีสะเกษ', 'สุรินทร์', 'อำนาจเจริญ', 'อุบลราชธานี'],
                        'ภาคเหนือ-เหนือตอนบน': ['น่าน', 'พะเยา', 'ลำปาง', 'ลำพูน', 'เชียงราย', 'เชียงใหม่', 'แพร่', 'แม่ฮ่องสอน'],
                        'ภาคเหนือ-เหนือตอนล่าง': ['กำแพงเพชร', 'ตาก', 'นครสวรรค์', 'พิจิตร', 'พิษณุโลก', 'สุโขทัย', 'อุตรดิตถ์', 'อุทัยธานี', 'เพชรบูรณ์'],
                        'ภาคใต้-ใต้ฝั่งอันดามัน': ['กระบี่', 'ตรัง', 'พังงา', 'ภูเก็ต', 'ระนอง', 'สตูล'],
                        'ภาคใต้-ใต้ฝั่งอ่าวไทย': ['ชุมพร', 'นครศรีธรรมราช', 'พัทลุง', 'ยะลา', 'สงขลา', 'สุราษฎร์ธานี', 'ปัตตานี', 'นราธิวาส']
                    }
                    
                    def get_region(province):
                        if pd.isna(province) or not province or str(province).strip() in ['', 'nan', 'UNKNOWN']:
                            return 'ไม่ระบุ'
                        
                        # 🚨 Override: ฉะเชิงเทรา → ภาคตะวันออก (ไม่ใช่ปริมณฑล)
                        if 'ฉะเชิงเทรา' in str(province):
                            return 'ภาคตะวันออก'
                        
                        for region, provinces in region_groups.items():
                            if any(p in str(province) for p in provinces):
                                return region
                        return 'อื่นๆ'
                    
                    # เพิ่มคอลัมน์ภาค - ดึงจังหวัดจาก Master ถ้าไม่มี
                    if 'Province' not in df_region.columns or df_region['Province'].isna().any():
                        # ดึงจังหวัดจาก Master
                        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                            province_map = {}
                            for _, row in MASTER_DATA.iterrows():
                                code = row.get('Plan Code', '')
                                province = row.get('จังหวัด', '')
                                if code and province:
                                    province_map[code] = province
                            
                            # ใส่จังหวัดให้แต่ละสาขา
                            if 'Province' not in df_region.columns:
                                df_region['Province'] = df_region['Code'].map(province_map)
                            else:
                                # เติมเฉพาะที่เป็น NaN
                                df_region['Province'] = df_region.apply(
                                    lambda row: province_map.get(row['Code'], row.get('Province', 'UNKNOWN')) 
                                    if pd.isna(row.get('Province')) else row['Province'],
                                    axis=1
                                )
                    
                    df_region['Region'] = df_region['Province'].apply(get_region)
                    
                    # หากลุ่มสาขา (ใช้ Booking No. เป็นหลัก)
                    def find_paired_branches(code, code_province, df_data):
                        paired = set()
                        
                        # หา Booking No. ของสาขานี้
                        code_rows = df_data[df_data['Code'] == code]
                        if len(code_rows) == 0:
                            return paired
                        
                        # เช็คว่ามีคอลัมน์ Booking หรือไม่
                        if 'Booking' not in df_data.columns and 'Trip' not in df_data.columns:
                            return paired
                        
                        booking_col = 'Booking' if 'Booking' in df_data.columns else 'Trip'
                        code_bookings = set(code_rows[booking_col].dropna().astype(str))
                        
                        if not code_bookings:
                            return paired
                        
                        # หาสาขาอื่นที่อยู่ Booking เดียวกัน (ไม่สนจังหวัด)
                        for booking in code_bookings:
                            if booking == 'nan' or not booking.strip():
                                continue
                            
                            same_booking = df_data[df_data[booking_col].astype(str) == booking]
                            for _, other_row in same_booking.iterrows():
                                other_code = other_row['Code']
                                
                                # เงื่อนไข: Booking เดียวกัน = รวมกลุ่ม (ไม่สนจังหวัด)
                                if other_code != code:
                                    paired.add(other_code)
                        
                        return paired
                    
                    all_codes_set = set(df_region['Code'].unique())
                    
                    # สร้างกลุ่มสาขาแบบ Union-Find (ตามลำดับ: ตำบล → อำเภอ → จังหวัด)
                    # Step 1: เริ่มจากแต่ละสาขาเป็นกลุ่มๆ พร้อมข้อมูล Master
                    initial_groups = {}
                    for code in all_codes_set:
                        # ดึงข้อมูลจาก Master
                        location = {}
                        if not MASTER_DATA.empty and 'Plan Code' in MASTER_DATA.columns:
                            master_row = MASTER_DATA[MASTER_DATA['Plan Code'] == code]
                            if len(master_row) > 0:
                                master_row = master_row.iloc[0]
                                location = {
                                    'subdistrict': master_row.get('ตำบล', ''),
                                    'district': master_row.get('อำเภอ', ''),
                                    'province': master_row.get('จังหวัด', 'UNKNOWN'),
                                    'lat': master_row.get('ละติจูด', 0),
                                    'lon': master_row.get('ลองติจูด', 0)
                                }
                        
                        # ถ้าไม่มีใน Master ลองดึงจากไฟล์อัปโหลด
                        if not location or location.get('province', 'UNKNOWN') == 'UNKNOWN':
                            c_row = df_region[df_region['Code'] == code].iloc[0] if len(df_region[df_region['Code'] == code]) > 0 else None
                            if c_row is not None:
                                location = {
                                    'subdistrict': '',
                                    'district': '',
                                    'province': c_row.get('Province', 'UNKNOWN'),
                                    'lat': 0,
                                    'lon': 0
                                }
                        
                        if location:
                            initial_groups[(code,)] = {code: location}
                    
                    # ใช้ initial_groups แทน booking_groups
                    booking_groups = initial_groups
                    
                    # Step 2: รวมกลุ่มตามลำดับ ตำบล → อำเภอ → จังหวัด
                    def groups_can_merge(locs1, locs2):
                        """เช็คว่า 2 กลุ่มควรรวมกันไหม (ตามลำดับความละเอียด)"""
                        # 1. เช็คตำบลเดียวกัน (ต้องมีข้อมูลตำบล)
                        subdistricts1 = set(loc.get('subdistrict', '') for loc in locs1.values() if loc.get('subdistrict', ''))
                        subdistricts2 = set(loc.get('subdistrict', '') for loc in locs2.values() if loc.get('subdistrict', ''))
                        if subdistricts1 and subdistricts2 and (subdistricts1 & subdistricts2):
                            return True, 'ตำบล'
                        
                        # 2. เช็คอำเภอเดียวกัน (ต้องมีข้อมูลอำเภอและจังหวัดเดียวกัน)
                        districts1 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs1.values() if loc.get('district', '')}
                        districts2 = {(loc.get('district', ''), loc.get('province', '')) for loc in locs2.values() if loc.get('district', '')}
                        if districts1 and districts2:
                            # เช็คว่ามีอำเภอและจังหวัดตรงกัน
                            for d1, p1 in districts1:
                                for d2, p2 in districts2:
                                    if d1 == d2 and p1 == p2 and p1:
                                        return True, 'อำเภอ'
                        
                        # 3. เช็คจังหวัดเดียวกัน
                        provinces1 = set(loc.get('province', '') for loc in locs1.values() if loc.get('province', ''))
                        provinces2 = set(loc.get('province', '') for loc in locs2.values() if loc.get('province', ''))
                        if provinces1 & provinces2:
                            return True, 'จังหวัด'
                        
                        return False, None
                    
                    merged_groups = []
                    used_groups = set()
                    
                    for group1, locs1 in booking_groups.items():
                        if group1 in used_groups:
                            continue
                        
                        merged_codes = set(group1)
                        merged_locs = locs1.copy()
                        used_groups.add(group1)
                        
                        # หากลุ่มอื่นที่ใกล้เคียง
                        changed = True
                        while changed:
                            changed = False
                            for group2, locs2 in booking_groups.items():
                                if group2 in used_groups:
                                    continue
                                can_merge, level = groups_can_merge(merged_locs, locs2)
                                if can_merge:
                                    merged_codes |= set(group2)
                                    merged_locs.update(locs2)
                                    used_groups.add(group2)
                                    changed = True
                        
                        merged_groups.append({
                            'codes': merged_codes,
                            'locations': merged_locs
                        })
                    
                    # Step 3: แปลงเป็น groups format
                    groups = []
                    for mg in merged_groups:
                        rep_code = list(mg['codes'])[0]
                        rep_row = df_region[df_region['Code'] == rep_code].iloc[0]
                        # กรองเฉพาะจังหวัดที่ไม่ใช่ UNKNOWN และไม่เป็น NaN
                        provinces = set(
                            str(loc.get('province', '')).strip() 
                            for loc in mg['locations'].values() 
                            if loc.get('province') and str(loc.get('province', '')).strip() not in ['UNKNOWN', 'nan', '']
                        )
                        
                        # ถ้าไม่มีจังหวัดเลย ใส่ "ไม่ระบุ"
                        province_str = ', '.join(sorted(provinces)) if provinces else 'ไม่ระบุ'
                        
                        groups.append({
                            'codes': mg['codes'],
                            'region': rep_row.get('Region', 'ไม่ระบุ'),
                            'province': province_str
                        })
                    
                    # แสดงสถิติ
                    st.markdown("---")
                    st.markdown("### 📊 สรุปการจัดกลุ่ม")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("📍 จำนวนสาขา", df_region['Code'].nunique())
                    with col2:
                        st.metric("🗂️ จำนวนกลุ่ม", len(groups))
                    with col3:
                        regions_count = df_region['Region'].nunique()
                        st.metric("🗺️ จำนวนภาค", regions_count)
                    
                    # แสดงตามภาค
                    st.markdown("---")
                    st.markdown("### 🗺️ สาขาแยกตามภาค")
                    
                    region_summary = df_region.groupby('Region').agg({
                        'Code': 'nunique',
                        'Weight': 'sum',
                        'Cube': 'sum'
                    }).reset_index()
                    region_summary.columns = ['ภาค', 'จำนวนสาขา', 'น้ำหนักรวม', 'คิวรวม']
                    st.dataframe(region_summary, use_container_width=True)
                    
                    # แสดงรายละเอียดแต่ละภาค
                    for region in sorted(df_region['Region'].unique()):
                        region_data = df_region[df_region['Region'] == region]
                        with st.expander(f"📍 {region} ({region_data['Code'].nunique()} สาขา)"):
                            display_cols = ['Code', 'Name', 'Province', 'Weight', 'Cube']
                            display_cols = [c for c in display_cols if c in region_data.columns]
                            
                            region_display = region_data[display_cols].drop_duplicates('Code')
                            col_names = {'Code': 'รหัส', 'Name': 'ชื่อสาขา', 'Province': 'จังหวัด', 'Weight': 'น้ำหนัก', 'Cube': 'คิว'}
                            region_display.columns = [col_names.get(c, c) for c in display_cols]
                            st.dataframe(region_display, use_container_width=True)
                    
                    # แสดงกลุ่มสาขาที่เคยไปด้วยกัน
                    st.markdown("---")
                    st.markdown("### 🔗 กลุ่มสาขาที่เคยไปด้วยกัน (จากประวัติ)")
                    
                    paired_groups = [g for g in groups if len(g['codes']) > 1]
                    if paired_groups:
                        for i, group in enumerate(paired_groups, 1):
                            codes_list = list(group['codes'])
                            names = []
                            for c in codes_list:
                                name_row = df_region[df_region['Code'] == c]
                                if len(name_row) > 0 and 'Name' in name_row.columns:
                                    names.append(f"{c} ({name_row['Name'].iloc[0]})")
                                else:
                                    names.append(c)
                            
                            st.write(f"**กลุ่ม {i}** - {group['region']}: {', '.join(names)}")
                    else:
                        st.info("ไม่พบกลุ่มสาขาที่เคยไปด้วยกันในรายการนี้")
                    
                    # ดาวน์โหลด
                    st.markdown("---")
                    output_region = io.BytesIO()
                    with pd.ExcelWriter(output_region, engine='xlsxwriter') as writer:
                        df_region.to_excel(writer, sheet_name='สาขาทั้งหมด', index=False)
                        region_summary.to_excel(writer, sheet_name='สรุปตามภาค', index=False)
                    
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        st.download_button(
                            label="📥 ดาวน์โหลดข้อมูลจัดกลุ่ม (Excel)",
                            data=output_region.getvalue(),
                            file_name=f"จัดกลุ่มสาขา_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

if __name__ == "__main__":
    main()
