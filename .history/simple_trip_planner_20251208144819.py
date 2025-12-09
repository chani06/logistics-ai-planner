"""
Simple Trip Planner - จัดทริปตามเงื่อนไขใหม่
เรียงตามระยะทาง → ชื่อสาขา → จังหวัด → อำเภอ → ตำบล
จัดรถตามคิว: 4W (5 คิว punthai, 3-4 คิว คละ), JB (7 drop punthai)
"""

import pandas as pd
import numpy as np
import math
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import streamlit as st

# พิกัด DC วังน้อย
DC_WANG_NOI_LAT = 14.2682524
DC_WANG_NOI_LON = 100.8434858

def haversine_distance(lat1, lon1, lat2, lon2):
    """คำนวณระยะทางระหว่างสองจุด (km)"""
    R = 6371
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    return R * c

def is_punthai(branch_code):
    """เช็คว่าเป็นสาขา Punthai หรือไม่"""
    code_str = str(branch_code).upper()
    # Punthai ขึ้นต้นด้วย: PH, PU, PT
    return code_str.startswith('PH') or code_str.startswith('PU') or code_str.startswith('PT')

def get_base_name(name):
    """ดึงชื่อพื้นฐานของสาขา (ตัดเลขท้ายออก)"""
    import re
    if not name:
        return ""
    
    name_lower = str(name).strip().lower()
    
    # Normalize ชื่อที่คล้ายกัน
    if 'future' in name_lower or 'ฟิวเจอร์' in name_lower:
        if 'rangsit' in name_lower or 'รังสิต' in name_lower:
            return "ฟิวเจอร์รังสิต"
    
    if 'lotus' in name_lower or 'โลตัส' in name_lower:
        base = re.sub(r'\s*\d+\s*$', '', name_lower)
        return base.strip()
    
    # ตัดเลขท้ายออก
    base = re.sub(r'\s*\d+\s*$', '', str(name).strip())
    return base.strip().lower()

def simple_plan_trips(df, master_data):
    """
    จัดทริปแบบง่าย ตามเงื่อนไข:
    1. เรียงตามระยะทาง → ชื่อ → จังหวัด → อำเภอ → ตำบล
    2. จัดรถตามคิว (4W: 5/3-4 คิว, JB: 7 drop)
    3. ถ้าเกิน → ตัดเป็นคันใหม่
    """
    
    # เตรียมข้อมูล
    result_df = df.copy()
    result_df['Trip'] = 0
    result_df['Distance_DC'] = 0.0
    result_df['Base_Name'] = ''
    result_df['Province'] = ''
    result_df['District'] = ''
    result_df['Subdistrict'] = ''
    result_df['Is_Punthai'] = False
    
    # ดึงข้อมูลจาก Master
    for idx, row in result_df.iterrows():
        code = row['Code']
        
        # ระยะทางจาก DC
        if not master_data.empty and 'Plan Code' in master_data.columns:
            master_row = master_data[master_data['Plan Code'] == code]
            if len(master_row) > 0:
                m = master_row.iloc[0]
                lat = m.get('ละติจูด', 0)
                lon = m.get('ลองติจูด', 0)
                if lat and lon:
                    dist = haversine_distance(DC_WANG_NOI_LAT, DC_WANG_NOI_LON, lat, lon)
                    result_df.at[idx, 'Distance_DC'] = dist
                
                # ดึงข้อมูลพื้นที่
                result_df.at[idx, 'Province'] = str(m.get('จังหวัด', '')).strip()
                result_df.at[idx, 'District'] = str(m.get('อำเภอ', '')).strip()
                result_df.at[idx, 'Subdistrict'] = str(m.get('ตำบล', '')).strip()
        
        # Base name
        if 'Name' in result_df.columns:
            result_df.at[idx, 'Base_Name'] = get_base_name(row['Name'])
        
        # Punthai check
        result_df.at[idx, 'Is_Punthai'] = is_punthai(code)
    
    # เรียงลำดับ: ระยะทาง → ชื่อ → จังหวัด → อำเภอ → ตำบล (ใกล้ไปไกล)
    result_df = result_df.sort_values(
        by=['Distance_DC', 'Base_Name', 'Province', 'District', 'Subdistrict'],
        ascending=[True, True, True, True, True]  # True = ใกล้ไปไกล
    ).reset_index(drop=True)
    
    # จัดทริป
    trip_num = 1
    current_trip_codes = []
    current_trip_cubes = 0
    all_punthai = True
    
    for idx, row in result_df.iterrows():
        code = row['Code']
        cube = row['Cube']
        is_punthai_branch = row['Is_Punthai']  # เปลี่ยนชื่อตัวแปร
        base_name = row['Base_Name']
        
        # เช็คว่าควรเริ่มทริปใหม่หรือไม่
        should_start_new_trip = False
        
        if len(current_trip_codes) == 0:
            # ทริปว่าง เริ่มใหม่
            should_start_new_trip = False
        else:
            # เช็คเงื่อนไขการตัดทริป
            # 1. เช็คว่าเป็น punthai ล้วนหรือคละ
            if all_punthai and not is_punthai_branch:
                all_punthai = False
            
            # 2. คำนวณขีดจำกัดคิว
            if all_punthai and not is_punthai_branch:
                # เปลี่ยนจาก punthai ล้วน → คละ
                max_cubes_4w = 4  # ลดลงเหลือ 4 คิว
            elif all_punthai:
                max_cubes_4w = 5  # punthai ล้วน 5 คิว
            else:
                max_cubes_4w = 4  # คละ 3-4 คิว
            
            # 3. เช็คว่าเกินคิวหรือไม่ (ยกเว้น 6W)
            # สมมติว่าถ้า cube > 5 คือใช้ 6W (ไม่ตัดทริป)
            is_6w = current_trip_cubes > 5 or cube > 5
            if not is_6w and current_trip_cubes + cube > max_cubes_4w:
                should_start_new_trip = True
            
            # 4. เช็คจำนวน drop สำหรับ JB (ถ้า punthai ล้วน ไม่เกิน 7)
            if not is_6w and all_punthai and len(current_trip_codes) >= 7:
                should_start_new_trip = True
        
        # ตัดสินใจ
        if should_start_new_trip:
            # เริ่มทริปใหม่
            trip_num += 1
            current_trip_codes = [code]
            current_trip_cubes = cube
            all_punthai = is_punthai_branch
        else:
            # เพิ่มในทริปปัจจุบัน
            current_trip_codes.append(code)
            current_trip_cubes += cube
        
        result_df.at[idx, 'Trip'] = trip_num
    
    # สร้าง summary
    summary_data = []
    for trip in result_df['Trip'].unique():
        trip_data = result_df[result_df['Trip'] == trip]
        branches = len(trip_data)
        total_cube = trip_data['Cube'].sum()
        total_weight = trip_data['Weight'].sum()
        all_punthai = trip_data['Is_Punthai'].all()
        
        # กำหนดรถ
        if all_punthai and branches <= 7:
            truck = 'JB (Punthai)'
        elif all_punthai and total_cube <= 5:
            truck = '4W (Punthai 5คิว)'
        else:
            truck = f'4W ({"3-4" if not all_punthai else "5"}คิว)'
        
        summary_data.append({
            'Trip': trip,
            'Branches': branches,
            'Cube': total_cube,
            'Weight': total_weight,
            'Truck': truck,
            'Punthai': 'ล้วน' if all_punthai else 'คละ'
        })
    
    summary_df = pd.DataFrame(summary_data)
    
    return result_df, summary_df

def export_to_excel_with_colors(result_df, filename, original_file, sheet_name="2.Punthai"):
    """Export ไฟล์ Excel พร้อมแยกสีตามทริป (เหลือง-ขาว สลับกัน) กลับไปที่ไฟล์เดิม"""
    
    # โหลดไฟล์ต้นฉบับ
    wb = load_workbook(original_file)
    
    # เข้าถึงชีตที่ต้องการ
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # สร้างสีเหลือง-ขาว
    yellow = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # หาคอลัมน์ที่มี Code (คอลัมน์ที่ 3 = index 2 จากภาพ: Sep, BU, BranchCode)
    code_col = 3  # BranchCode
    
    # หาคอลัมน์ Trip (คอลัมน์ที่ 9 จากภาพ)
    trip_col = 9  # Trip column
    
    # สร้าง dictionary: code -> trip
    code_to_trip = dict(zip(result_df['Code'].astype(str), result_df['Trip']))
    
    # ระบายสีตามทริป (เริ่มจากแถวที่ 3 เพราะแถวที่ 1-2 เป็น header)
    current_trip = None
    use_yellow = True  # เริ่มด้วยเหลือง
    
    for row_idx in range(3, ws.max_row + 1):
        code_cell = ws.cell(row=row_idx, column=code_col)
        code_value = str(code_cell.value).strip() if code_cell.value else None
        
        if code_value and code_value in code_to_trip:
            trip = code_to_trip[code_value]
            
            # ถ้าเปลี่ยนทริป ให้สลับสี
            if current_trip != trip:
                current_trip = trip
                use_yellow = not use_yellow
            
            # เลือกสีและระบายทั้งแถว
            fill = yellow if use_yellow else white
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = fill
            
            # อัพเดทเลขทริปในคอลัมน์ Trip
            ws.cell(row=row_idx, column=trip_col, value=int(trip))
    
    # บันทึกไฟล์
    wb.save(filename)
    print(f"✅ บันทึกไฟล์: {filename}")
